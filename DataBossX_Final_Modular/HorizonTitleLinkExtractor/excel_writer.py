"""
excel_writer.py
==============
Excel reading + review-workbook writing for HorizonTitleLinkExtractor.

Uses openpyxl and preserves formatting, formulas, hyperlinks, and hidden sheets.
The ORIGINAL workbook is never modified: we always work on a separate review copy.

Provides:
  * load_workbook_copy()         - copy source -> review file, open it.
  * detect_title_sheets()        - find worksheets that look like runsheets.
  * detect_header()              - find header row + map known columns.
  * iter_link_rows()             - yield rows with a usable document hyperlink.
  * ensure_review_columns()      - append the AI Status / AI Suggested * columns.
  * write_review()               - write AI results + color-code a row.
  * compare_fields()             - field-by-field comparison (normalized).
"""

from __future__ import annotations

import datetime as _dt
import glob
import logging
import os
import re
import shutil
from dataclasses import dataclass, field
from typing import Any, Iterator

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import matching

log = logging.getLogger("horizon.excel")

# --- color fills -------------------------------------------------------------
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")

# --- header keyword -> canonical field --------------------------------------
HEADER_MAP: dict[str, str] = {}


def _reg(field_name: str, *aliases: str) -> None:
    for a in aliases:
        HEADER_MAP[a.lower()] = field_name


_reg("instrument_number", "instrument #", "instr #", "instrument number",
     "instrument no", "instr no", "instrument")
_reg("document_type", "instrument type", "type", "document type", "doc type")
_reg("book_page", "book/page", "book page", "book/pg", "bk/pg", "book-page")
_reg("filed_date", "filed date", "file date", "date filed")
_reg("recorded_date", "recorded date", "record date", "date recorded")
_reg("effective_date", "effective date", "date effective")
_reg("grantor", "grantor", "grantor(s)", "grantors")
_reg("grantee", "grantee", "grantee(s)", "grantees")
_reg("legal_description", "legal description", "legal", "lands", "land description")
_reg("section", "section", "sec", "sec.")
_reg("township", "township", "twp", "twp.")
_reg("range", "range", "rng", "rge")
_reg("county", "county")
_reg("state", "state")
_reg("remarks", "comments", "remarks", "notes", "comment")
_reg("document_link", "document link", "doc link", "link", "document", "doc",
     "image", "image link", "url")

# Sheets named like these are not runsheets.
_NON_TITLE_SHEET = re.compile(r"(cover|instructions?|legend|summary|toc|index)$", re.I)

# Review columns (in order) appended to the right of each processed sheet.
REVIEW_COLUMNS = [
    "AI Status", "AI Confidence", "AI Consensus",
    "AI Suggested Instrument #", "AI Suggested Instrument Type",
    "AI Suggested Book/Page", "AI Suggested Filed Date",
    "AI Suggested Recorded Date", "AI Suggested Effective Date",
    "AI Suggested Grantor", "AI Suggested Grantee",
    "AI Suggested Legal Description", "AI Suggested Section",
    "AI Suggested Township", "AI Suggested Range",
    "AI Notes", "AI Changed Fields", "AI Error", "AI Review Needed",
    "AI Processed Timestamp", "AI Source Link", "AI Viewer Method",
    "AI Model Used", "AI Validator Models Used",
]

# Map AI-suggested review column -> canonical extraction field.
_SUGGEST_FIELD = {
    "AI Suggested Instrument #": "instrument_number",
    "AI Suggested Instrument Type": "document_type",
    "AI Suggested Book/Page": "book_page",
    "AI Suggested Filed Date": "filed_date",
    "AI Suggested Recorded Date": "recorded_date",
    "AI Suggested Effective Date": "effective_date",
    "AI Suggested Grantor": "grantor",
    "AI Suggested Grantee": "grantee",
    "AI Suggested Legal Description": "legal_description",
    "AI Suggested Section": "section",
    "AI Suggested Township": "township",
    "AI Suggested Range": "range",
}

_URL_RE = re.compile(r"https?://[^\s\"'<>)\]]+", re.I)
_HYPERLINK_FORMULA = re.compile(r'HYPERLINK\(\s*"([^"]+)"', re.I)


@dataclass
class SheetLayout:
    sheet: Worksheet
    header_row: int
    columns: dict[str, int]           # canonical field -> column index (1-based)
    review_start_col: int = 0
    review_cols: dict[str, int] = field(default_factory=dict)


# =============================================================================
# Loading
# =============================================================================
def find_latest_review(output_dir: str, source_path: str) -> str | None:
    """Return the newest existing review workbook for this source (used to
    resume from prior results), or None if there isn't one."""
    stem = os.path.splitext(os.path.basename(source_path))[0]
    pattern = os.path.join(output_dir, f"{stem}_AI_REVIEW_*.xlsx")
    matches = [p for p in glob.glob(pattern) if os.path.isfile(p)]
    if not matches:
        return None
    return max(matches, key=os.path.getmtime)


def load_workbook_copy(source_path: str, output_dir: str,
                       resume_from: str | None = None) -> tuple[Any, str]:
    """
    Create a new AI_REVIEW working copy and open it. The original source is never
    touched. The new file is always named from the SOURCE stem.

    When resume_from is given (a prior review workbook), it is copied instead of
    the source so previously-written AI review columns/results are preserved -
    otherwise resuming would leave blank review data for already-processed rows.
    Returns (workbook, review_path).
    """
    os.makedirs(output_dir, exist_ok=True)
    base = os.path.basename(source_path)
    stem, ext = os.path.splitext(base)
    now = _dt.datetime.now()
    stamp = now.strftime("%Y-%m-%d_%H%M%S") + f"_{now.microsecond // 1000:03d}"
    review_name = f"{stem}_AI_REVIEW_{stamp}{ext}"
    review_path = os.path.join(output_dir, review_name)
    copy_src = resume_from if (resume_from and os.path.exists(resume_from)) else source_path
    shutil.copy2(copy_src, review_path)
    wb = load_workbook(review_path)  # keep formulas + formatting + hyperlinks
    if copy_src != source_path:
        log.info("Resuming from prior review workbook: %s", os.path.basename(copy_src))
    log.info("Opened review copy: %s", review_path)
    return wb, review_path


def detect_title_sheets(wb: Any) -> list[Worksheet]:
    """Return worksheets that look like title/runsheet sheets (hidden included)."""
    out: list[Worksheet] = []
    for ws in wb.worksheets:
        if _NON_TITLE_SHEET.search(ws.title.strip()):
            continue
        if detect_header(ws) is not None:
            out.append(ws)
    return out


def detect_header(ws: Worksheet, scan_rows: int = 25) -> SheetLayout | None:
    """
    Scan the first scan_rows rows for a header row matching known column names.
    Returns a SheetLayout (without review columns yet) or None.
    """
    best: tuple[int, dict[str, int]] | None = None
    max_col = min(ws.max_column or 1, 60)
    for r in range(1, min(ws.max_row or 1, scan_rows) + 1):
        cols: dict[str, int] = {}
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if not isinstance(val, str):
                continue
            key = " ".join(val.split()).strip().lower()
            canonical = HEADER_MAP.get(key)
            if canonical and canonical not in cols:
                cols[canonical] = c
        if best is None or len(cols) > len(best[1]):
            best = (r, cols)
    if not best or len(best[1]) < 2:
        return None
    # Require either a link column or enough title fields to be meaningful.
    has_link = "document_link" in best[1]
    has_fields = len(set(best[1]) - {"document_link"}) >= 2
    if not (has_link or has_fields):
        return None
    return SheetLayout(sheet=ws, header_row=best[0], columns=best[1])


# =============================================================================
# Hyperlink detection
# =============================================================================
def extract_link(ws: Worksheet, row: int, col: int) -> str | None:
    """
    Find a document URL for a cell using, in order:
      1. cell.hyperlink.target
      2. HYPERLINK() formula
      3. plain-text URL inside the cell
      4. rich-text hyperlink (best effort)
    """
    if not col:
        return None
    cell = ws.cell(row=row, column=col)

    # 1. explicit hyperlink object
    hl = getattr(cell, "hyperlink", None)
    if hl is not None and getattr(hl, "target", None):
        return str(hl.target)

    val = cell.value
    if isinstance(val, str):
        # 2. HYPERLINK() formula
        m = _HYPERLINK_FORMULA.search(val)
        if m:
            return m.group(1)
        # 3. plain text URL
        m = _URL_RE.search(val)
        if m:
            return m.group(0)

    # 4. rich text (openpyxl may expose CellRichText)
    try:
        if val is not None and not isinstance(val, (str, int, float)):
            text = "".join(getattr(part, "text", str(part)) for part in val)
            m = _URL_RE.search(text)
            if m:
                return m.group(0)
    except Exception:
        pass
    return None


@dataclass
class LinkRow:
    sheet_title: str
    row: int
    link: str
    values: dict[str, Any]            # canonical field -> existing cell value


def iter_link_rows(layout: SheetLayout) -> Iterator[LinkRow]:
    """
    Yield every data row that has a usable document link. Rows without a link
    are skipped (and should be logged by the caller via skipped_no_link()).
    """
    ws = layout.sheet
    link_col = layout.columns.get("document_link")
    start = layout.header_row + 1
    for r in range(start, (ws.max_row or start) + 1):
        # ignore fully empty rows
        if all(ws.cell(row=r, column=c).value in (None, "")
               for c in layout.columns.values()):
            continue
        link = extract_link(ws, r, link_col) if link_col else None
        if not link:
            # Some sheets stash the link in any column; try every mapped col.
            for c in layout.columns.values():
                link = extract_link(ws, r, c)
                if link:
                    break
        if not link:
            continue
        values = {fld: ws.cell(row=r, column=col).value
                  for fld, col in layout.columns.items()}
        yield LinkRow(ws.title, r, link, values)


def rows_without_links(layout: SheetLayout) -> list[int]:
    """Return data row numbers that have content but no usable link (for logging)."""
    ws = layout.sheet
    link_col = layout.columns.get("document_link")
    start = layout.header_row + 1
    skipped: list[int] = []
    for r in range(start, (ws.max_row or start) + 1):
        if all(ws.cell(row=r, column=c).value in (None, "")
               for c in layout.columns.values()):
            continue
        link = extract_link(ws, r, link_col) if link_col else None
        if not link:
            for c in layout.columns.values():
                if extract_link(ws, r, c):
                    link = "x"
                    break
        if not link:
            skipped.append(r)
    return skipped


# =============================================================================
# Review columns
# =============================================================================
def ensure_review_columns(layout: SheetLayout) -> None:
    """Ensure the review columns exist on this sheet and map them.

    Idempotent: if the sheet already has the review columns (e.g. we resumed from
    a prior review workbook), reuse them instead of appending a duplicate set.
    """
    ws = layout.sheet
    # Detect any review columns already present in the header row.
    existing: dict[str, int] = {}
    for c in range(1, (ws.max_column or 1) + 1):
        val = ws.cell(row=layout.header_row, column=c).value
        if isinstance(val, str) and val in REVIEW_COLUMNS:
            existing.setdefault(val, c)
    if "AI Status" in existing:
        # Reuse the existing block; backfill any columns a prior version lacked.
        layout.review_cols = dict(existing)
        next_col = max(existing.values()) + 1
        for name in REVIEW_COLUMNS:
            if name not in layout.review_cols:
                ws.cell(row=layout.header_row, column=next_col, value=name)
                layout.review_cols[name] = next_col
                next_col += 1
        layout.review_start_col = min(layout.review_cols.values())
        return

    start_col = (ws.max_column or 1) + 1
    layout.review_start_col = start_col
    for i, name in enumerate(REVIEW_COLUMNS):
        col = start_col + i
        ws.cell(row=layout.header_row, column=col, value=name)
        layout.review_cols[name] = col


# --- normalization for comparison -------------------------------------------
def _norm_text(v: Any) -> str:
    if v is None:
        return ""
    return " ".join(str(v).split()).strip()


def _norm_date(v: Any) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%B %d, %Y",
                "%b %d, %Y", "%d-%b-%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


_DATE_FIELDS = {"filed_date", "recorded_date", "effective_date"}


def compare_fields(existing: dict[str, Any], extracted: dict[str, Any],
                   major_fields: list[str],
                   same_threshold: float = matching.SAME_THRESHOLD,
                   minor_threshold: float = matching.MINOR_THRESHOLD
                   ) -> tuple[list[str], list[str]]:
    """
    Compare existing row values vs AI-suggested values using field-aware fuzzy
    matching. Returns (major_changes, minor_changes):

      * major_changes - substantive differences worth human review.
      * minor_changes - formatting/punctuation-level differences (noted, not
        treated as real corrections).

    AI nulls never produce a change. Names and legal descriptions are matched
    leniently so trivial punctuation/abbreviation differences are not flagged
    as hard changes.
    """
    major: list[str] = []
    minor: list[str] = []
    for fld in major_fields:
        ai_val = extracted.get(fld)
        if ai_val in (None, ""):
            continue
        level = matching.classify_diff(fld, existing.get(fld), ai_val,
                                       same_threshold, minor_threshold)
        if level == "major":
            major.append(fld)
        elif level == "minor":
            minor.append(fld)
    return major, minor


def write_review(layout: SheetLayout, row: int, result: dict[str, Any]) -> None:
    """
    Write a full AI result into the review columns for `row` and color-code it.
    Original cells are NEVER overwritten here (apply_corrections is handled by
    the orchestrator separately and is false by default).

    `result` keys used:
      status, confidence, consensus_label, suggestions(dict field->val),
      notes, changed_fields(list), error, review_needed(bool), timestamp,
      source_link, viewer_method, model_used, validator_models(list),
      color ('green'|'yellow'|'red'), consensus_fields(list)
    """
    ws = layout.sheet
    rc = layout.review_cols

    def put(colname: str, value: Any) -> None:
        col = rc.get(colname)
        if col:
            ws.cell(row=row, column=col, value=value)

    put("AI Status", result.get("status", ""))
    conf = result.get("confidence")
    put("AI Confidence", round(conf, 3) if isinstance(conf, (int, float)) else conf)
    put("AI Consensus", result.get("consensus_label", ""))

    suggestions = result.get("suggestions", {}) or {}
    consensus_fields = set(result.get("consensus_fields", []) or [])
    for colname, fld in _SUGGEST_FIELD.items():
        val = suggestions.get(fld)
        put(colname, val if val not in (None, "") else None)
        # Mark fields where >=2 models agreed with a cell comment.
        if fld in consensus_fields:
            cell = ws.cell(row=row, column=rc[colname])
            cell.comment = Comment("AI consensus (>=2 models agree)", "HorizonAI")

    put("AI Notes", result.get("notes", ""))
    put("AI Changed Fields", ", ".join(result.get("changed_fields", []) or []))
    put("AI Error", result.get("error", "") or "")
    put("AI Review Needed", "YES" if result.get("review_needed") else "no")
    put("AI Processed Timestamp", result.get("timestamp", ""))
    put("AI Source Link", result.get("source_link", ""))
    put("AI Viewer Method", result.get("viewer_method", ""))
    put("AI Model Used", result.get("model_used", ""))
    put("AI Validator Models Used",
        ", ".join(result.get("validator_models", []) or []))

    _color_row(layout, row, result.get("color", "yellow"))


def _color_row(layout: SheetLayout, row: int, color: str) -> None:
    fill = {"green": FILL_GREEN, "yellow": FILL_YELLOW, "red": FILL_RED}.get(
        color, FILL_YELLOW)
    ws = layout.sheet
    start = layout.review_start_col
    end = start + len(REVIEW_COLUMNS) - 1
    for c in range(start, end + 1):
        ws.cell(row=row, column=c).fill = fill
    # Also tint the "AI Status" cell strongly so the row reads at a glance.
    status_col = layout.review_cols.get("AI Status")
    if status_col:
        ws.cell(row=row, column=status_col).fill = fill


def apply_corrections(layout: SheetLayout, row: int, suggestions: dict[str, Any],
                      changed_fields: list[str],
                      highlight: bool = True) -> None:
    """
    Only called when config apply_corrections=true. Overwrites original cells for
    consensus-supported changed fields and highlights them. Never invents data.
    """
    ws = layout.sheet
    for fld in changed_fields:
        col = layout.columns.get(fld)
        val = suggestions.get(fld)
        if not col or val in (None, ""):
            continue
        cell = ws.cell(row=row, column=col)
        cell.value = val
        if highlight:
            cell.fill = FILL_YELLOW
