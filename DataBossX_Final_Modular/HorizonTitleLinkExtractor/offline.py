"""
offline.py
==========
A fully self-contained OFFLINE mode. It replaces the live browser and live AI
with deterministic mocks so the ENTIRE pipeline (workbook select -> open link ->
'View' -> capture image in memory -> AI extract -> compare -> write review
columns -> color code -> report -> logs) runs with NO API keys, NO browser,
and NO network. This powers RUN_OFFLINE_DEMO.bat and the end-to-end test.

The mock is designed to exercise every code path: green matches, yellow
corrections, low-confidence reviews, and red failed links.
"""

from __future__ import annotations

import io
from typing import Any

from PIL import Image, ImageDraw

from ai_extractors import MAJOR_FIELDS, TitleExtraction
from cache import CostTracker

# A small, deterministic synthetic "title plant" the demo draws rows from.
_GRANTORS = ["John A. Smith", "ACME Minerals, LLC", "Mary Jo Carter, Trustee",
             "Robert E. Lee, Jr.", "First National Bank", "Estate of Ada Vance"]
_GRANTEES = ["Jane B. Doe", "Horizon Resources, L.P.", "United Oil & Gas Co.",
             "The Carter Family Trust", "County of Roger Mills", "Ada Vance, et ux"]
_DOC_TYPES = ["Warranty Deed", "Mineral Deed", "Oil and Gas Lease",
              "Quitclaim Deed", "Mortgage", "Assignment"]


def demo_truth(idx: int, tract: str = "31-12N-24W") -> dict[str, Any]:
    """The 'real document' values for demo row idx (deterministic)."""
    sec, twp, rng = tract.split("-")
    g = _GRANTORS[idx % len(_GRANTORS)]
    ge = _GRANTEES[idx % len(_GRANTEES)]
    dt = _DOC_TYPES[idx % len(_DOC_TYPES)]
    year = 2018 + (idx % 8)
    month = (idx % 12) + 1
    day = (idx % 27) + 1
    return {
        "instrument_number": f"2024-{1000 + idx}",
        "document_type": dt,
        "book_page": f"{800 + idx}/{(idx * 7) % 600}",
        "filed_date": f"{year:04d}-{month:02d}-{day:02d}",
        "recorded_date": f"{year:04d}-{month:02d}-{day:02d}",
        "effective_date": None,
        "grantor": g,
        "grantee": ge,
        "legal_description": f"NW/4 of Section {sec}, Township {twp}, Range {rng}",
        "section": sec,
        "township": twp,
        "range": rng,
        "county": "Roger Mills",
        "state": "OK",
        "remarks": None,
    }


def demo_plan(idx: int) -> dict[str, Any]:
    """Decide what this demo row should demonstrate (color path)."""
    mod = idx % 6
    if mod == 5:
        return {"link_fails": True, "confidence": 0.0}        # RED failed link
    if mod == 4:
        return {"link_fails": False, "confidence": 0.55,      # RED low confidence
                "low_conf": True}
    if mod in (2, 3):
        return {"link_fails": False, "confidence": 0.93,      # YELLOW correction
                "spreadsheet_wrong": True}
    return {"link_fails": False, "confidence": 0.985}         # GREEN clean match


# --------------------------------------------------------------------------- #
# Mock browser viewer
# --------------------------------------------------------------------------- #
def _synthetic_doc_image(idx: int) -> bytes:
    """Render a small, real PNG that looks vaguely like a recorded document."""
    img = Image.new("RGB", (1100, 850), (250, 249, 245))
    d = ImageDraw.Draw(img)
    truth = demo_truth(idx)
    d.rectangle([20, 20, 1080, 830], outline=(120, 120, 120), width=2)
    lines = [
        "RECORDED DOCUMENT (DEMO)",
        f"Instrument No: {truth['instrument_number']}",
        f"Type: {truth['document_type']}",
        f"Grantor: {truth['grantor']}",
        f"Grantee: {truth['grantee']}",
        f"Filed: {truth['filed_date']}   Book/Page: {truth['book_page']}",
        f"Legal: {truth['legal_description']}",
        f"County: {truth['county']}, {truth['state']}",
    ]
    y = 60
    for ln in lines:
        d.text((50, y), ln, fill=(20, 20, 20))
        y += 90
    # texture so the blur metric is non-trivial
    for x in range(40, 1060, 6):
        d.line([(x, 800), (x, 810)], fill=(0, 0, 0))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


class MockViewer:
    """Drop-in replacement for BrowserViewer in offline mode."""

    def __init__(self, config: dict[str, Any]):
        self.config = config

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False

    def start(self):  # parity with BrowserViewer
        pass

    def stop(self):
        pass

    def capture_document(self, url: str, row_id: Any = "?") -> dict[str, Any]:
        idx = _idx_from(url, row_id)
        plan = demo_plan(idx)
        if plan.get("link_fails"):
            return {"images": [], "viewer_method": "none", "ok": False,
                    "error": "DEMO: document link failed / blocked page"}
        return {"images": [_synthetic_doc_image(idx)], "viewer_method": "image",
                "ok": True, "error": None}


# --------------------------------------------------------------------------- #
# Mock model router
# --------------------------------------------------------------------------- #
class MockRouter:
    """Drop-in replacement for ModelRouter in offline mode."""

    def __init__(self, config: dict[str, Any], cost: CostTracker | None = None,
                 result_cache=None):
        self.config = config
        self.cost = cost or CostTracker()
        self.openai_model = "offline-mock"

    def extract(self, images: list[bytes], row_changed_hint: bool = False,
                row_id: Any = "") -> dict[str, Any]:
        idx = _idx_from("", row_id)
        truth = demo_truth(idx)
        plan = demo_plan(idx)
        conf = float(plan.get("confidence", 0.95))
        primary = TitleExtraction(
            confidence=conf,
            needs_human_review=bool(plan.get("low_conf")),
            uncertain_fields=(["legal_description"] if plan.get("low_conf") else []),
            evidence_summary="OFFLINE DEMO extraction (no real document read).",
            **{k: truth.get(k) for k in MAJOR_FIELDS + ["county", "state", "remarks"]},
        )
        self.cost.record("offline-mock", 0, 0, row=row_id, cached=True)
        return {
            "primary": primary,
            "validators": {"gemini": None, "claude": None},
            "consensus": {}, "consensus_fields": [], "disagree_fields": [],
            "models_used": "offline-mock", "validator_models_used": [],
            "confidence": conf,
            "needs_human_review": bool(plan.get("low_conf")),
            "image_quality": 0.95, "poor_image": False,
        }


def _idx_from(url: str, row_id: Any) -> int:
    try:
        return int(row_id)
    except (TypeError, ValueError):
        return abs(hash(str(url))) % 24


# --------------------------------------------------------------------------- #
# Demo workbook generator
# --------------------------------------------------------------------------- #
def make_demo_workbook(path: str, tract: str = "31-12N-24W", rows: int = 12) -> str:
    """Create a realistic demo runsheet whose rows are crafted so the offline
    run produces a mix of green / yellow / red review states."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Runsheet"
    headers = ["Instrument #", "Instrument Type", "Book/Page", "Filed Date",
               "Grantor", "Grantee", "Legal Description", "Section", "Township",
               "Range", "Comments", "Document Link"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    field_to_col = {
        "instrument_number": 1, "document_type": 2, "book_page": 3,
        "filed_date": 4, "grantor": 5, "grantee": 6, "legal_description": 7,
        "section": 8, "township": 9, "range": 10,
    }
    # Index each data row by its WORKSHEET row number so the mock viewer/router
    # (which key off row_id == worksheet row) read the same document the row was
    # generated from.
    for i in range(1, rows + 1):
        ws_row = i + 1
        truth = demo_truth(ws_row, tract)
        plan = demo_plan(ws_row)
        for fld, col in field_to_col.items():
            val = truth.get(fld)
            # Introduce a deliberate spreadsheet error so AI shows a correction.
            if plan.get("spreadsheet_wrong") and fld == "grantor":
                val = "Jon Smith"  # misspelled vs document "John A. Smith"
            ws.cell(row=ws_row, column=col, value=val)
        link_cell = ws.cell(row=ws_row, column=12, value="View Document")
        link_cell.hyperlink = f"https://demo.county.local/preview?row={ws_row}"

    # one row with content but no link -> exercises "skipped" logging
    ws.cell(row=rows + 2, column=1, value="2024-NOLINK")
    ws.cell(row=rows + 2, column=5, value="No Link Grantor")

    wb.save(path)
    return path
