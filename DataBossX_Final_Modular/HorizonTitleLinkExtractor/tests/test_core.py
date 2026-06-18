"""
Offline unit tests for HorizonTitleLinkExtractor.

These cover the pure logic that does NOT need a browser or live AI APIs:
matching, excel detection/comparison, consensus, image preprocessing, caching,
cost tracking, config validation, drive selection, and the HTML report.

Run:  python -m pytest tests/ -q     (from the project folder)
"""

import io
import os
import sys

import pytest
from openpyxl import Workbook
from PIL import Image

# Make project modules importable when pytest runs from the project root.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import cache
import config_schema
import excel_writer as xl
import image_utils
import matching
import report as report_mod
from ai_extractors import MAJOR_FIELDS, ModelRouter, TitleExtraction
from drive_sync import NullDriveSync


# --------------------------------------------------------------------------- #
# matching
# --------------------------------------------------------------------------- #
def test_normalize_date_formats():
    assert matching.normalize_date("06/15/2026") == "2026-06-15"
    assert matching.normalize_date("June 15, 2026") == "2026-06-15"
    assert matching.normalize_date("2026-06-15") == "2026-06-15"
    assert matching.normalize_date("") == ""


def test_name_punctuation_is_minor_not_major():
    # "John A Smith" vs "John A. Smith" should NOT be a hard change.
    assert matching.classify_diff("grantor", "John A Smith", "John A. Smith") \
        in ("same", "minor")


def test_company_suffix_variants_agree():
    assert matching.values_agree("grantee", "ACME L.L.C.", "ACME LLC")
    assert matching.similarity("grantee", "ACME L.L.C.", "ACME LLC") >= 0.85


def test_legal_description_abbrev_match():
    a = "The Northwest Quarter of Section 31"
    b = "NW/4 Sec 31"
    assert matching.similarity("legal_description", a, b) >= 0.6


def test_real_difference_is_major():
    assert matching.classify_diff("grantor", "John Smith", "Mary Jones") == "major"
    assert matching.classify_diff("legal_description", "NW/4 Sec 31",
                                  "NE/4 Sec 31") == "major"


def test_blank_existing_with_ai_value_is_major():
    assert matching.classify_diff("instrument_number", "", "12345") == "major"


def test_ai_null_never_flags():
    assert matching.classify_diff("grantor", "John Smith", None) == "same"


# --------------------------------------------------------------------------- #
# excel detection + hyperlinks + comparison
# --------------------------------------------------------------------------- #
def _sample_wb(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Runsheet"
    headers = ["Instrument #", "Instrument Type", "Book/Page", "Filed Date",
               "Grantor", "Grantee", "Legal Description", "Document Link"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value="12345")
    ws.cell(row=2, column=5, value="John A Smith")
    hc = ws.cell(row=2, column=8, value="open")
    hc.hyperlink = "https://county.example.com/preview?id=12345"
    ws.cell(row=3, column=1, value="999")
    ws.cell(row=3, column=8,
            value='=HYPERLINK("https://county.example.com/doc/999","view")')
    ws.cell(row=4, column=1, value="777")
    ws.cell(row=4, column=8, value="see https://county.example.com/x/777")
    ws.cell(row=5, column=1, value="555")  # no link
    wb.save(path)
    return path


def test_header_and_link_detection(tmp_path):
    src = _sample_wb(str(tmp_path / "src.xlsx"))
    wb, review = xl.load_workbook_copy(src, str(tmp_path / "out"))
    assert os.path.abspath(review) != os.path.abspath(src)
    sheets = xl.detect_title_sheets(wb)
    assert sheets
    layout = xl.detect_header(sheets[0])
    assert layout.columns["instrument_number"] == 1
    assert "document_link" in layout.columns

    links = list(xl.iter_link_rows(layout))
    assert [r.link for r in links] == [
        "https://county.example.com/preview?id=12345",
        "https://county.example.com/doc/999",
        "https://county.example.com/x/777",
    ]
    assert 5 in xl.rows_without_links(layout)


def test_compare_fields_major_minor(tmp_path):
    src = _sample_wb(str(tmp_path / "src.xlsx"))
    wb, _ = xl.load_workbook_copy(src, str(tmp_path / "out"))
    layout = xl.detect_header(xl.detect_title_sheets(wb)[0])
    existing = {"grantor": "John A Smith", "grantee": "Johnson",
                "instrument_number": "12345"}
    suggestions = {f: None for f in MAJOR_FIELDS}
    suggestions.update({"grantor": "John A. Smith",   # punctuation only -> same
                        "grantee": "Johnston",        # typo-level -> minor
                        "instrument_number": "67890"})  # real change -> major
    major, minor = xl.compare_fields(existing, suggestions, MAJOR_FIELDS)
    assert "instrument_number" in major
    assert "grantor" not in major and "grantor" not in minor  # ignored
    assert "grantee" in minor and "grantee" not in major


def test_ensure_review_columns_idempotent(tmp_path):
    # Resuming reopens a workbook that already has review columns; we must reuse
    # them, not append a duplicate set.
    src = _sample_wb(str(tmp_path / "src.xlsx"))
    wb, _ = xl.load_workbook_copy(src, str(tmp_path / "out"))
    sheet = xl.detect_title_sheets(wb)[0]
    layout = xl.detect_header(sheet)
    xl.ensure_review_columns(layout)
    first_max = layout.sheet.max_column

    layout2 = xl.detect_header(sheet)          # fresh layout over same sheet
    xl.ensure_review_columns(layout2)
    assert layout.sheet.max_column == first_max  # no new columns added
    hdr = [layout.sheet.cell(row=layout2.header_row, column=c).value
           for c in range(1, layout.sheet.max_column + 1)]
    assert hdr.count("AI Status") == 1
    assert len(layout2.review_cols) == len(xl.REVIEW_COLUMNS)


def test_find_latest_review(tmp_path):
    out = tmp_path / "out"
    out.mkdir()
    src = tmp_path / "Book_31.xlsx"
    src.write_bytes(b"x")
    a = out / "Book_31_AI_REVIEW_1.xlsx"
    b = out / "Book_31_AI_REVIEW_2.xlsx"
    a.write_bytes(b"a")
    b.write_bytes(b"b")
    os.utime(a, (1, 1))
    os.utime(b, (2, 2))
    assert xl.find_latest_review(str(out), str(src)) == str(b)
    assert xl.find_latest_review(str(out), str(tmp_path / "Other.xlsx")) is None


def test_review_columns_written(tmp_path):
    src = _sample_wb(str(tmp_path / "src.xlsx"))
    wb, review = xl.load_workbook_copy(src, str(tmp_path / "out"))
    layout = xl.detect_header(xl.detect_title_sheets(wb)[0])
    xl.ensure_review_columns(layout)
    assert "AI Status" in layout.review_cols
    assert len(layout.review_cols) == len(xl.REVIEW_COLUMNS)
    xl.write_review(layout, 2, {
        "status": "ok", "confidence": 0.95, "consensus_label": "consensus",
        "suggestions": {"grantor": "John A. Smith"},
        "consensus_fields": ["grantor"], "notes": "n", "changed_fields": [],
        "error": "", "review_needed": False, "timestamp": "t",
        "source_link": "u", "viewer_method": "image", "model_used": "gpt",
        "validator_models": [], "color": "green"})
    col = layout.review_cols["AI Status"]
    assert layout.sheet.cell(row=2, column=col).value == "ok"


# --------------------------------------------------------------------------- #
# consensus (fuzzy, no API calls)
# --------------------------------------------------------------------------- #
def _router():
    return ModelRouter({"openai_model": "gpt-5.2", "use_cache": False,
                        "match_minor_threshold": 0.85})


def test_consensus_two_of_three_agree_fuzzy():
    r = _router()
    p = TitleExtraction(grantee="ACME L.L.C.", instrument_number="100")
    g = TitleExtraction(grantee="ACME LLC", instrument_number="100")
    c = TitleExtraction(grantee="Totally Different Co", instrument_number="100")
    cons, cfields, dfields = r._consensus(p, {"gemini": g, "claude": c})
    assert "grantee" in cfields            # ACME variants cluster
    assert "instrument_number" in cfields  # all agree
    assert "grantee" not in dfields


def test_consensus_disagreement_flags():
    r = _router()
    p = TitleExtraction(legal_description="NW/4 Sec 31")
    g = TitleExtraction(legal_description="SE/4 Sec 31")
    c = TitleExtraction(legal_description="NE/4 Sec 31")
    cons, cfields, dfields = r._consensus(p, {"gemini": g, "claude": c})
    assert "legal_description" in dfields
    assert "legal_description" not in cons


def test_consensus_never_invents_missing():
    r = _router()
    p = TitleExtraction()  # all null
    cons, cfields, dfields = r._consensus(p, {"gemini": None, "claude": None})
    assert cons == {} and cfields == [] and dfields == []


# --------------------------------------------------------------------------- #
# image utils
# --------------------------------------------------------------------------- #
def _png_bytes(w=200, h=120, color=128):
    img = Image.new("RGB", (w, h), (color, color, color))
    # add some structure so sharpness != 0
    for x in range(0, w, 8):
        for y in range(h):
            img.putpixel((x, y), (0, 0, 0))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def test_enhance_returns_valid_png_and_upscales():
    data = _png_bytes(200, 120)
    out = image_utils.enhance(data)
    img = Image.open(io.BytesIO(out))
    assert img.format == "PNG"
    assert max(img.size) >= image_utils.MIN_GOOD_DIM  # small image upscaled


def test_quality_assessment_and_prepare():
    data = _png_bytes(1200, 900)
    q = image_utils.assess_quality(data)
    assert 0.0 <= q.score <= 1.0
    prepared, worst = image_utils.prepare_for_ai([data, _png_bytes(100, 60)])
    assert len(prepared) == 2 and worst is not None


# --------------------------------------------------------------------------- #
# cache + cost
# --------------------------------------------------------------------------- #
def test_result_cache_roundtrip(tmp_path):
    c = cache.ResultCache(str(tmp_path / "c"), enabled=True)
    key = cache.image_key("gpt-5.2", [b"abc", b"def"])
    assert c.get(key) is None
    c.set(key, {"instrument_number": "1", "confidence": 0.9})
    assert c.get(key)["instrument_number"] == "1"
    # key stability
    assert key == cache.image_key("gpt-5.2", [b"abc", b"def"])
    assert key != cache.image_key("gpt-5.2", [b"abc", b"xyz"])


def test_cost_tracker():
    ct = cache.CostTracker()
    ct.record("gpt-5.2", 1000, 500, row=1)
    ct.record("gpt-5.2", 0, 0, row=2, cached=True)
    assert ct.total_calls == 1            # cached call not billed
    assert ct.total_usd > 0
    assert len(ct.rows) == 2


# --------------------------------------------------------------------------- #
# config validation
# --------------------------------------------------------------------------- #
def test_config_valid(tmp_path):
    p = tmp_path / "config.yaml"
    p.write_text("openai_model: gpt-5.2\ntest_mode_rows: 5\n")
    cfg = config_schema.load_validated(str(p))
    assert cfg["openai_model"] == "gpt-5.2"
    assert cfg["match_minor_threshold"] == 0.85   # default applied


def test_config_invalid_threshold(tmp_path):
    p = tmp_path / "config.yaml"
    # auto_safe < review must fail validation
    p.write_text("ai_confidence_auto_safe_threshold: 0.5\n"
                 "ai_confidence_review_threshold: 0.9\n")
    with pytest.raises(ValueError):
        config_schema.load_validated(str(p))


# --------------------------------------------------------------------------- #
# drive selection
# --------------------------------------------------------------------------- #
def test_drive_selection_excludes_ai_and_prefers_cursory(tmp_path, monkeypatch):
    folder = tmp_path / "input"
    folder.mkdir()
    # newest is an AI review file (must be excluded)
    (folder / "31-12N-24W_Roger_Mills_Cursory_Title_Report.xlsx").write_text("x")
    (folder / "random_other.xlsx").write_text("x")
    (folder / "31-12N-24W_AI_REVIEW_latest.xlsx").write_text("x")
    monkeypatch.chdir(tmp_path)
    sync = NullDriveSync({"drive_folder_url": "", "input_dir": "input"})
    # point its candidate scan at our folder
    sync.config["input_dir"] = "input"
    cands = sync.list_candidates()
    names = [c.name for c in cands]
    assert any("Cursory" in n for n in names)
    chosen = sync.select_source()
    assert "AI_REVIEW" not in chosen.name
    assert "Cursory" in chosen.name


# --------------------------------------------------------------------------- #
# html report
# --------------------------------------------------------------------------- #
def test_build_report(tmp_path):
    path = str(tmp_path / "r.html")
    summary = {"linked_rows": 3, "processed_rows": 2, "changed_rows": 1,
               "review_needed_rows": 1, "failed_rows": 1,
               "high_confidence_rows": 1, "skipped_rows": 0}
    rows = [{"sheet": "Runsheet", "row": 2, "status": "ok", "confidence": 0.95,
             "consensus_label": "consensus", "changed_fields": [],
             "review_needed": False, "viewer_method": "image", "notes": "",
             "color": "green", "link": "https://x"},
            {"sheet": "Runsheet", "row": 3, "status": "failed", "confidence": 0.0,
             "consensus_label": "", "changed_fields": [], "review_needed": True,
             "viewer_method": "", "notes": "blocked", "color": "red", "link": ""}]
    out = report_mod.build_report(path, "wb.xlsx", "2026-06-18", summary, rows,
                                  total_cost=0.1234)
    html = open(out, encoding="utf-8").read()
    assert "Horizon Title" in html
    assert "Runsheet" in html
    assert "$0.1234" in html
