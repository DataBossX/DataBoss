"""Write chained results into the template workbook, preserving formatting.

The template (``data/template/template.xlsx``) is the write destination
(rule 1). We load it with openpyxl -- which preserves fonts, fills, borders,
number formats, column widths, row heights, merged cells, freeze panes, page
setup and sheet order -- and write only the cells we own, copying style from a
styled reference row so new rows look native (rule 2).

Layout contract (see ``make_template.py``):

* Sheet order: ``Overview`` first (rule 3), then ``Title Sheet``, then
  ``Runsheet Register``.
* ``Title Sheet`` columns: Tract | Owner | Net Mineral Acres | Working
  Interest | OGL No. | Conveyance | Remarks / Review.
* Only owners surviving the chain are written (rules 4 & 5).
* OGL column carries OGL numbers only (rule 7) -- an assignment book/page can
  never land here because the schema strips it (rule 8).
* Only assumption / review cells are filled yellow (rules 15 & 16).
"""

from __future__ import annotations

from copy import copy
from decimal import Decimal
from pathlib import Path
from typing import Dict, Optional, Sequence

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from schemas import RunsheetInstrument, TractChainResult

TITLE_SHEET = "Title Sheet"
OVERVIEW_SHEET = "Overview"
REGISTER_SHEET = "Runsheet Register"

# Title Sheet column layout (1-based).
COL_TRACT = 1
COL_OWNER = 2
COL_ACRES = 3
COL_WI = 4
COL_OGL = 5
COL_CONV = 6
COL_REMARKS = 7
N_COLS = 7

# The one and only highlight used in the whole workbook (rule 16: no others).
YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")

# Row in the template that carries the reference style for data rows.
STYLE_REF_ROW = 4
HEADER_ROW = 3
FIRST_DATA_ROW = 4


def _copy_style(src_cell, dst_cell) -> None:
    """Copy visual style (not value) from one cell to another."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def _dec_to_number(value: Decimal) -> float:
    """openpyxl stores numbers as float; keep six-decimal precision."""
    return float(value)


def _clear_data_rows(ws: Worksheet, first_row: int) -> None:
    """Remove any previously written owner rows, keeping the styled ref row."""
    if ws.max_row >= first_row:
        ws.delete_rows(first_row, ws.max_row - first_row + 1)


def _ensure_sheet_order(wb) -> None:
    """Overview must remain the first tab (rule 3)."""
    if OVERVIEW_SHEET in wb.sheetnames:
        idx = wb.sheetnames.index(OVERVIEW_SHEET)
        if idx != 0:
            wb.move_sheet(OVERVIEW_SHEET, -idx)


def write_corrected_workbook(
    template_path: str | Path,
    results: Dict[str, TractChainResult],
    output_path: str | Path,
    instruments: Optional[Sequence[RunsheetInstrument]] = None,
) -> str:
    """Write ``results`` into a copy of the template and save to ``output_path``.

    ``results`` maps tract id -> :class:`TractChainResult`. ``instruments`` (the
    raw ingested runsheet rows) are optionally written to the Runsheet Register
    sheet. Returns the output path as a string. Raises ``FileNotFoundError`` if
    the template is missing (never silently writes to a fabricated path).
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template workbook not found at {template_path}. Run make_template.py "
            "or place your template there before exporting."
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(template_path)
    _ensure_sheet_order(wb)

    if TITLE_SHEET not in wb.sheetnames:
        raise KeyError(
            f"Template is missing a '{TITLE_SHEET}' sheet. The writer needs it as "
            "the destination for final owners."
        )
    ws = wb[TITLE_SHEET]

    # Capture the reference style row before we clear anything.
    style_cells = [ws.cell(row=STYLE_REF_ROW, column=c) for c in range(1, N_COLS + 1)]
    ref_style = [(c.has_style, copy(c.font), copy(c.border), copy(c.fill),
                  c.number_format, copy(c.protection), copy(c.alignment))
                 for c in style_cells]

    _clear_data_rows(ws, FIRST_DATA_ROW)

    row = FIRST_DATA_ROW
    for tract in sorted(results.keys(), key=_tract_sort_key):
        res = results[tract]
        for owner in res.final_owners:
            _write_owner_row(ws, row, tract, owner, res.tract_acres, ref_style)
            row += 1

    _write_overview(wb, results)
    if instruments:
        _write_register(wb, instruments)

    wb.save(output_path)
    return str(output_path)


# Runsheet Register column layout (matches make_template.build_runsheet_register).
_REG_HEADER_ROW = 3
_REG_FIRST_ROW = 4
_REG_NCOLS = 15


def _split_book_page(bk_pg: str) -> tuple[str, str]:
    """Split a 'book/page' style reference into (book, page)."""
    if not bk_pg:
        return "", ""
    import re

    m = re.match(r"\s*(\d+)\s*[/\-]\s*(\d+)\s*$", str(bk_pg))
    if m:
        return m.group(1), m.group(2)
    return str(bk_pg), ""


def _write_register(wb, instruments: Sequence[RunsheetInstrument]) -> None:
    """Write the ingested instruments to the Runsheet Register sheet."""
    if REGISTER_SHEET not in wb.sheetnames:
        return
    ws = wb[REGISTER_SHEET]
    ref_cells = [ws.cell(row=_REG_FIRST_ROW, column=c) for c in range(1, _REG_NCOLS + 1)]
    ref = [(c.has_style, copy(c.font), copy(c.border), copy(c.fill),
            c.number_format, copy(c.protection), copy(c.alignment)) for c in ref_cells]
    _clear_data_rows(ws, _REG_FIRST_ROW)
    row = _REG_FIRST_ROW
    for entry, inst in enumerate(instruments, start=1):
        book, page = _split_book_page(inst.bk_pg)
        gross = float(inst.gross_acres) if inst.gross_acres is not None else ""
        values = [
            entry, inst.date, "", inst.conveyance, inst.grantor, inst.grantee,
            book, page, inst.instrument_no, inst.legal, gross, "",
            inst.conveyed_interest or "", inst.ogl_ref or "", inst.notes,
        ]
        for col in range(1, _REG_NCOLS + 1):
            cell = ws.cell(row=row, column=col)
            _apply_ref_style(cell, ref[col - 1])
            cell.value = values[col - 1]
        row += 1


def _apply_ref_style(cell, ref) -> None:
    has_style, font, border, fill, number_format, protection, alignment = ref
    if has_style:
        cell.font = copy(font)
        cell.border = copy(border)
        cell.fill = copy(fill)
        cell.number_format = number_format
        cell.protection = copy(protection)
        cell.alignment = copy(alignment)


def _write_owner_row(ws, row, tract, owner, tract_acres, ref_style) -> None:
    """Write one final-owner row, styling it like the template's data row."""
    values = {
        COL_TRACT: tract,
        COL_OWNER: owner.owner,
        COL_ACRES: _dec_to_number(owner.acres),
        COL_WI: _dec_to_number(owner.wi),
        COL_OGL: owner.ogl or "",
        COL_CONV: owner.conveyance or "",
        COL_REMARKS: owner.assumption_note or "",
    }
    for col in range(1, N_COLS + 1):
        cell = ws.cell(row=row, column=col)
        _apply_ref_style(cell, ref_style[col - 1])
        cell.value = values[col]

    # Yellow only on the review cell of an assumption owner (rules 14, 15, 16).
    if owner.is_assumption:
        ws.cell(row=row, column=COL_REMARKS).fill = YELLOW
        if owner.ogl is None and owner.leased:
            # A lease with an unknown OGL number is itself a review item.
            ws.cell(row=row, column=COL_OGL).fill = YELLOW


def _write_overview(wb, results: Dict[str, TractChainResult]) -> None:
    """Refresh the Overview summary block without disturbing its formatting."""
    if OVERVIEW_SHEET not in wb.sheetnames:
        return
    ws = wb[OVERVIEW_SHEET]
    # The template reserves a summary block starting at a marker row; we only
    # update the values in the reserved cells, never restyle the sheet.
    marker_row = _find_marker(ws, "TRACT SUMMARY")
    if marker_row is None:
        return
    r = marker_row + 2  # header is marker_row + 1
    # Clear old summary lines (up to 200 rows) then rewrite.
    for rr in range(r, min(r + 200, ws.max_row + 1)):
        for c in range(1, 6):
            ws.cell(row=rr, column=c).value = None
    total_tracts = len(results)
    balanced = sum(1 for x in results.values() if x.balanced)
    review = sum(1 for x in results.values() if x.needs_review)
    for tract in sorted(results.keys(), key=_tract_sort_key):
        res = results[tract]
        ws.cell(row=r, column=1).value = tract
        ws.cell(row=r, column=2).value = _dec_to_number(res.tract_acres)
        ws.cell(row=r, column=3).value = _dec_to_number(res.total_acres)
        ws.cell(row=r, column=4).value = "BALANCED" if res.balanced else "REVIEW"
        ws.cell(row=r, column=5).value = "; ".join(res.review_flags)[:500]
        if not res.balanced:
            ws.cell(row=r, column=4).fill = YELLOW
        r += 1
    totals_row = _find_marker(ws, "TOTALS")
    if totals_row is not None:
        ws.cell(row=totals_row, column=2).value = total_tracts
        ws.cell(row=totals_row, column=3).value = balanced
        ws.cell(row=totals_row, column=4).value = review


def _find_marker(ws, marker: str):
    """Return the row index whose column A equals ``marker`` (case-insensitive)."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.strip().upper() == marker.upper():
            return cell.row
    return None


def _tract_sort_key(t: str):
    """Sort tracts numerically when possible, else lexically."""
    try:
        return (0, int(str(t).strip()))
    except (TypeError, ValueError):
        return (1, str(t))
