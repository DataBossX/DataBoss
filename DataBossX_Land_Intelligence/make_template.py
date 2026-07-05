"""Generate ``data/template/template.xlsx`` -- the workbook write destination.

Run once (``python make_template.py``) to produce a real, styled template that
:mod:`excel_writer` writes into. The layout matches the writer's contract:

* ``Overview``          -- section banner + tract summary block (stays first).
* ``Title Sheet``       -- final owners after chaining (Owner | Acres | WI |
                           OGL | Conveyance | Remarks).
* ``Runsheet Register`` -- the canonical instrument columns (entry, dates,
                           doc type, grantor/grantee, book, page, instrument,
                           legal, acreage, NMA, interest, remarks) mirroring the
                           existing Roger Mills report layout.

The template ships with header styling, column widths and a styled reference
data row so written rows inherit native formatting.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SECTION = "SECTION 31-12N-24W  —  ROGER MILLS COUNTY, OKLAHOMA"

NAVY = "0B1F3A"
NEON = "00D8FF"
LIGHT = "F2F6FA"
THIN = Side(style="thin", color="9DB2C6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name="Consolas", bold=True, color="FFFFFF", size=11)
BANNER_FONT = Font(name="Consolas", bold=True, color=NEON, size=16)
DATA_FONT = Font(name="Consolas", color="1A2A3A", size=10)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
BANNER_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _banner(ws, text: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = BANNER_FONT
    c.fill = BANNER_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30
    for col in range(1, span + 1):
        ws.cell(row=1, column=col).fill = BANNER_FILL


def _header_row(ws, row: int, headers, widths=None) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 26


def _style_ref_row(ws, row: int, ncols: int, number_formats=None) -> None:
    """Lay down one styled (empty) data row so the writer can copy its style."""
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.font = DATA_FONT
        c.border = BORDER
        c.alignment = LEFT if i in (2, 7) else CENTER
        c.fill = PatternFill(fill_type="solid", fgColor=LIGHT)
        if number_formats and number_formats.get(i):
            c.number_format = number_formats[i]


def build_overview(wb) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    _banner(ws, SECTION, span=5)

    sub = ws.cell(row=2, column=1, value="FINAL OWNERSHIP TITLE REPORT  ·  Command Center Export")
    sub.font = Font(name="Consolas", italic=True, color=NAVY, size=11)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    legend = ws.cell(
        row=4,
        column=1,
        value="LEGEND:  Yellow cells = assumption / examiner-review items only. "
        "All other cells are verified. OGL column carries OGL numbers only.",
    )
    legend.font = Font(name="Consolas", color="7A5A00", size=10)
    legend.fill = PatternFill(fill_type="solid", fgColor="FFF7CC")
    legend.alignment = LEFT
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
    ws.row_dimensions[4].height = 30

    # Totals marker (writer updates B/C/D on this row).
    t = ws.cell(row=6, column=1, value="TOTALS")
    t.font = Font(name="Consolas", bold=True, color=NAVY, size=11)
    ws.cell(row=6, column=2, value=0).font = DATA_FONT
    ws.cell(row=6, column=3, value=0).font = DATA_FONT
    ws.cell(row=6, column=4, value=0).font = DATA_FONT
    labels = ws.cell(row=5, column=1, value="Tracts / Balanced / Review:")
    labels.font = Font(name="Consolas", italic=True, color=NAVY, size=9)

    # Tract summary block (writer fills rows under the header).
    m = ws.cell(row=8, column=1, value="TRACT SUMMARY")
    m.font = Font(name="Consolas", bold=True, color=NAVY, size=12)
    _header_row(
        ws,
        9,
        ["Tract", "Tract Acres", "Chained Total", "Status", "Review Notes"],
        widths=[10, 14, 16, 12, 60],
    )

    ws.freeze_panes = "A10"


def build_title_sheet(wb) -> None:
    ws = wb.create_sheet("Title Sheet")
    ws.sheet_view.showGridLines = False
    headers = [
        "Tract",
        "Owner (Final)",
        "Net Mineral Acres",
        "Working Interest",
        "OGL No.",
        "Conveyance",
        "Remarks / Review",
    ]
    widths = [10, 40, 18, 16, 16, 14, 50]
    _banner(ws, SECTION + "  —  TITLE SHEET (FINAL OWNERS)", span=len(headers))
    ws.cell(row=2, column=1)  # spacer row
    _header_row(ws, 3, headers, widths)
    _style_ref_row(
        ws,
        4,
        len(headers),
        number_formats={3: "0.000000", 4: "0.000000"},
    )
    ws.freeze_panes = "A4"


def build_runsheet_register(wb) -> None:
    ws = wb.create_sheet("Runsheet Register")
    ws.sheet_view.showGridLines = False
    headers = [
        "Entry",
        "Instrument Date",
        "Recorded Date",
        "Doc Type",
        "Grantor",
        "Grantee",
        "Book",
        "Page",
        "Instrument No.",
        "Legal Description",
        "Gross Acres",
        "NMA",
        "Interest",
        "OGL No.",
        "Remarks",
    ]
    widths = [7, 14, 14, 10, 26, 26, 8, 8, 16, 40, 12, 12, 12, 12, 40]
    _banner(ws, SECTION + "  —  RUNSHEET REGISTER", span=len(headers))
    ws.cell(row=2, column=1)
    _header_row(ws, 3, headers, widths)
    _style_ref_row(ws, 4, len(headers))
    ws.freeze_panes = "A4"


def main(out_path: str | Path = None) -> str:
    out_path = Path(out_path) if out_path else Path(__file__).parent / "data" / "template" / "template.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_overview(wb)
    build_title_sheet(wb)
    build_runsheet_register(wb)
    wb.save(out_path)
    return str(out_path)


if __name__ == "__main__":
    p = main()
    print(f"Wrote template: {p}")
