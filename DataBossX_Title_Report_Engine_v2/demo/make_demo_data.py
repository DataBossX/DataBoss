#!/usr/bin/env python3
"""Generate a CLEARLY-SYNTHETIC demo corpus to exercise Engine v2 end-to-end.

    !!!  EVERYTHING THIS SCRIPT WRITES IS FAKE / SYNTHETIC TEST DATA  !!!
    !!!  It is NOT real Section 31-12N-24W Roger Mills title.          !!!

It builds, under demo/horizon_sample/:
  * template_v30.xlsx  -- a formatting template with an Overview/map tab (merged
    title cell, colored header, freeze panes) so template preservation is tested.
  * Roger_Mills_Runsheet.xlsx -- a runsheet sheet (Tract column) + an OGL sheet.

The synthetic chain deliberately exercises: a root/patent seed, a mineral deed
with a reservation, an oil & gas lease with an OGL number carried down, an
assignment (ASSN, leasehold only), an assumed-fraction conveyance, an impossible
conveyance that must be rejected, and an unleased owner (missing-lease flag).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent
OUT = HERE / "horizon_sample"

BANNER = "SYNTHETIC DEMO DATA — NOT REAL TITLE — for engine testing only"


def build_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ov = wb.active
    ov.title = "Overview"
    ov.merge_cells("A1:F1")
    c = ov["A1"]
    c.value = "SECTION 31-12N-24W — ROGER MILLS COUNTY, OKLAHOMA"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.alignment = Alignment(horizontal="center")
    c.fill = PatternFill("solid", fgColor="1F4E78")
    ov["A2"] = "Ownership / Cursory Title Report — Overview & Plat"
    ov["A2"].font = Font(italic=True)
    ov["A4"] = BANNER
    ov["A4"].font = Font(color="FF0000", bold=True)
    ov["A6"] = "[ plat / map image would live here in the real template ]"
    ov.column_dimensions["A"].width = 40
    ov.freeze_panes = "A3"
    # A hidden named-range/setup sheet to prove hidden sheets survive.
    hidden = wb.create_sheet("Setup")
    hidden["A1"] = "internal template settings"
    hidden.sheet_state = "hidden"
    wb.save(path)


def build_runsheet(path: Path) -> None:
    wb = openpyxl.Workbook()
    rs = wb.active
    rs.title = "Runsheet"
    rs["A1"] = BANNER
    rs["A1"].font = Font(color="FF0000", bold=True)
    headers = ["Instrument Date", "Recording Date", "Book/Page", "Instrument Type",
               "Grantor", "Grantee", "Tract", "Gross Acres", "Conveyed Interest",
               "Reserved Interest", "OGL", "Legal Description", "Notes"]
    rs.append([])  # row 2 spacer
    rs.append(headers)  # row 3 = header row
    rows = [
        # --- Tract 1 (160 ac): root, mineral deed w/ reservation, two leases, ASSN
        ["1920-01-05", "1920-02-01", "10/1", "US Patent", "United States",
         "A. Founder", "Tract 1", "160", "", "", "", "SW/4 Sec 31-12N-24W", "root of title"],
        ["1948-06-10", "1948-06-20", "88/210", "Mineral Deed", "A. Founder",
         "B. Smith", "Tract 1", "", "0.5", "0.5", "", "SW/4", "reserves 1/2"],
        ["1970-03-01", "1970-03-05", "150/44", "Oil & Gas Lease", "A. Founder",
         "Acme Oil Co", "Tract 1", "", "", "", "OGL-101", "SW/4", "lease"],
        ["1970-03-02", "1970-03-06", "150/45", "Oil & Gas Lease", "B. Smith",
         "Acme Oil Co", "Tract 1", "", "", "", "OGL-102", "SW/4", "lease"],
        ["1975-09-09", "1975-09-15", "201/88", "Assignment", "Acme Oil Co",
         "Bigwell LLC", "Tract 1", "", "", "", "OGL-101", "SW/4", "assigns OGL-101"],
        # --- Tract 2 (160 ac): root, deed 1/4, assumed-fraction warranty deed, one lease
        ["1921-04-11", "1921-05-01", "11/9", "US Patent", "United States",
         "C. Jones", "Tract 2", "160", "", "", "", "SE/4 Sec 31-12N-24W", "root of title"],
        ["1955-01-20", "1955-02-01", "95/120", "Mineral Deed", "C. Jones",
         "D. Brown", "Tract 2", "", "0.25", "", "", "SE/4", ""],
        ["1962-07-07", "1962-07-12", "120/300", "Warranty Deed", "D. Brown",
         "E. Green", "Tract 2", "", "", "", "", "SE/4", "conveys all owned (no fraction)"],
        ["1980-11-01", "1980-11-04", "260/15", "Oil & Gas Lease", "C. Jones",
         "Acme Oil Co", "Tract 2", "", "", "", "OGL-201", "SE/4", "lease"],
        # --- Tract 3 (80 ac): root, deed 3/4, IMPOSSIBLE quitclaim (0.9 > 0.75)
        ["1922-02-02", "1922-03-01", "12/50", "US Patent", "United States",
         "F. White", "Tract 3", "80", "", "", "", "N/2 NE/4 Sec 31-12N-24W", "root"],
        ["1958-05-05", "1958-05-10", "99/1", "Mineral Deed", "F. White",
         "G. Black", "Tract 3", "", "0.75", "", "", "N/2 NE/4", ""],
        ["1965-08-08", "1965-08-14", "130/77", "Quit Claim Deed", "G. Black",
         "H. Gray", "Tract 3", "", "0.9", "", "", "N/2 NE/4", "conveys more than owned"],
    ]
    for r in rows:
        rs.append(r)
    rs.freeze_panes = "A4"

    ogl = wb.create_sheet("OGLs")
    ogl["A1"] = BANNER
    ogl["A1"].font = Font(color="FF0000", bold=True)
    ogl.append([])
    ogl.append(["OGL Number", "Lessor", "Lessee", "Instrument Date", "Book/Page",
                "Legal Description"])
    for row in [
        ["OGL-101", "A. Founder", "Acme Oil Co", "1970-03-01", "150/44", "SW/4"],
        ["OGL-102", "B. Smith", "Acme Oil Co", "1970-03-02", "150/45", "SW/4"],
        ["OGL-201", "C. Jones", "Acme Oil Co", "1980-11-01", "260/15", "SE/4"],
    ]:
        ogl.append(row)
    wb.save(path)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    build_template(OUT / "template_v30.xlsx")
    build_runsheet(OUT / "Roger_Mills_Runsheet.xlsx")
    print(f"Wrote synthetic demo corpus to {OUT}")
    for p in sorted(OUT.glob("*.xlsx")):
        print("  -", p.name)


if __name__ == "__main__":
    main()
