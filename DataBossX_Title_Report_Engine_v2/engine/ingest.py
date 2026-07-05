"""Stage 1 -- discover, rank, and classify source workbooks under a root folder.

The engine is pointed at a folder (e.g. ``D:\\Desktop\\Horizon``) and must decide,
without human help, which workbook is the runsheet, which is the template, which
is a prior report, and which sheets inside them are runsheet / OGL / tract tabs.

Ranking is transparent: every candidate gets a score built from named,
inspectable signals (does it contain a runsheet-looking sheet? tract tabs? an
overview/map tab? how recently was it modified?) and the reasons are recorded so
``SOURCE_FILE_RANKING.xlsx`` can show *why* a file won.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

# --- sheet-name signals -----------------------------------------------------

RUNSHEET_NAMES = ("runsheet", "run sheet", "run-sheet", "instruments", "index",
                  "chain", "abstract")
OGL_NAMES = ("ogl", "ogls", "lease", "leases", "oil gas lease", "oil and gas",
             "oil & gas")
TRACT_RE = re.compile(r"\btract\s*#?\s*\d+", re.IGNORECASE)
OVERVIEW_NAMES = ("overview", "map", "plat", "cover", "summary", "title sheet",
                  "title")
TEMPLATE_HINTS = ("template",)
REPORT_HINTS = ("report", "final", "ownership", "title")

SPREADSHEET_EXTS = {".xlsx", ".xlsm", ".xls"}
# Generic junk directories we never descend into (matched by name anywhere).
SKIP_DIRS = {".git", "__pycache__", "node_modules", "venv", ".venv"}


@dataclass
class SheetInfo:
    title: str
    max_row: int = 0
    max_col: int = 0
    is_runsheet: bool = False
    is_ogl: bool = False
    is_tract: bool = False
    is_overview: bool = False
    hidden: bool = False


@dataclass
class Candidate:
    path: Path
    kind: str = "other"          # runsheet | template | prior_report | ogl | other
    score: float = 0.0
    modified: float = 0.0
    sheets: List[SheetInfo] = field(default_factory=list)
    reasons: List[str] = field(default_factory=list)
    error: str = ""

    @property
    def name(self) -> str:
        return self.path.name

    def has_runsheet(self) -> bool:
        return any(s.is_runsheet for s in self.sheets)

    def has_overview(self) -> bool:
        return any(s.is_overview for s in self.sheets)

    def tract_sheets(self) -> List[SheetInfo]:
        return [s for s in self.sheets if s.is_tract]


def _classify_sheet(title: str, hidden: bool, max_row: int, max_col: int) -> SheetInfo:
    low = title.strip().lower()
    is_tract = bool(TRACT_RE.search(title))
    is_ogl = any(k in low for k in OGL_NAMES) and not is_tract
    is_runsheet = any(k in low for k in RUNSHEET_NAMES)
    # An overview/map/plat tab -- but not if it's really a tract or runsheet.
    is_overview = (any(low == k or low.startswith(k) for k in OVERVIEW_NAMES)
                   and not is_tract and not is_runsheet)
    return SheetInfo(title=title, max_row=max_row, max_col=max_col,
                     is_runsheet=is_runsheet, is_ogl=is_ogl, is_tract=is_tract,
                     is_overview=is_overview, hidden=hidden)


def inspect_workbook(path: Path) -> Candidate:
    """Open a workbook read-only and record its sheet structure + a score."""
    import openpyxl

    cand = Candidate(path=path)
    try:
        cand.modified = path.stat().st_mtime
    except OSError:
        cand.modified = 0.0

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 -- a bad file must not kill the scan
        cand.error = f"{type(exc).__name__}: {exc}"
        cand.reasons.append(f"unreadable ({cand.error})")
        return cand

    try:
        for ws in wb.worksheets:
            hidden = getattr(ws, "sheet_state", "visible") != "visible"
            info = _classify_sheet(ws.title, hidden, ws.max_row or 0, ws.max_column or 0)
            cand.sheets.append(info)
    finally:
        wb.close()

    cand.score, cand.reasons, cand.kind = _score_candidate(cand)
    return cand


def _score_candidate(cand: Candidate):
    """Score a candidate from named signals and infer its most likely kind."""
    reasons: List[str] = []
    score = 0.0
    low_name = cand.name.lower()

    n_runsheet = sum(1 for s in cand.sheets if s.is_runsheet)
    n_tract = sum(1 for s in cand.sheets if s.is_tract)
    n_ogl = sum(1 for s in cand.sheets if s.is_ogl)
    n_overview = sum(1 for s in cand.sheets if s.is_overview)
    total_rows = sum(s.max_row for s in cand.sheets)

    if n_runsheet:
        score += 40
        reasons.append(f"has {n_runsheet} runsheet-type sheet(s) (+40)")
    if n_tract:
        score += min(n_tract * 6, 36)
        reasons.append(f"has {n_tract} tract sheet(s) (+{min(n_tract * 6, 36)})")
    if n_ogl:
        score += 15
        reasons.append(f"has {n_ogl} OGL/lease sheet(s) (+15)")
    if n_overview:
        score += 10
        reasons.append("has overview/map/title sheet (+10)")
    if total_rows > 40:
        score += 10
        reasons.append(f"substantial data ({total_rows} rows) (+10)")

    # Filename hints.
    if any(h in low_name for h in TEMPLATE_HINTS):
        score += 8
        reasons.append("filename says 'template' (+8)")
    if any(h in low_name for h in REPORT_HINTS):
        score += 6
        reasons.append("filename says report/final/ownership (+6)")

    # Infer kind. Priority: runsheet (controlling) > prior_report > template > ogl.
    kind = "other"
    if any(h in low_name for h in TEMPLATE_HINTS) and n_runsheet == 0:
        kind = "template"
    if n_runsheet:
        kind = "runsheet"
    elif n_tract and any(h in low_name for h in REPORT_HINTS):
        kind = "prior_report"
    elif n_tract:
        kind = "prior_report"
    elif n_ogl and kind == "other":
        kind = "ogl"

    if cand.error:
        reasons.append("penalty: unreadable")
        score = -1.0
    return round(score, 2), reasons, kind


def discover(root: Path, exclude_dirs: Optional[List[Path]] = None) -> List[Candidate]:
    """Recursively find and inspect every spreadsheet under ``root``.

    ``exclude_dirs`` are absolute directories to skip entirely -- used to keep the
    engine from re-ingesting its own copied inputs / prior outputs when its
    working folder sits inside the scanned root (e.g. under D:\\Desktop\\Horizon).
    """
    root = Path(root)
    excludes = [Path(d).resolve() for d in (exclude_dirs or [])]
    candidates: List[Candidate] = []
    if not root.exists():
        return candidates
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if path.suffix.lower() not in SPREADSHEET_EXTS:
            continue
        if path.name.startswith("~$"):  # Excel lock/temp files
            continue
        if any(part in SKIP_DIRS for part in path.parts):
            continue
        rp = path.resolve()
        if any(_is_under(rp, ex) for ex in excludes):
            continue
        candidates.append(inspect_workbook(path))
    candidates.sort(key=lambda c: (c.score, c.modified), reverse=True)
    return candidates


def _is_under(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


@dataclass
class SourceSelection:
    runsheet: Optional[Candidate] = None
    template: Optional[Candidate] = None
    prior_report: Optional[Candidate] = None
    ogl: Optional[Candidate] = None
    all_candidates: List[Candidate] = field(default_factory=list)


def select_sources(candidates: List[Candidate]) -> SourceSelection:
    """Pick the single best runsheet / template / prior report / OGL workbook.

    A template that best preserves formatting is the one with an overview/map
    tab; the runsheet is the highest-scoring workbook that actually contains a
    runsheet sheet. These can be the same physical file.
    """
    sel = SourceSelection(all_candidates=candidates)

    runsheets = [c for c in candidates if c.has_runsheet()]
    sel.runsheet = runsheets[0] if runsheets else None

    templates = [c for c in candidates if c.kind == "template"]
    if not templates:
        # Fall back to the best workbook that has an overview/map tab to preserve.
        templates = [c for c in candidates if c.has_overview()]
    sel.template = templates[0] if templates else (candidates[0] if candidates else None)

    priors = [c for c in candidates if c.kind == "prior_report"]
    sel.prior_report = priors[0] if priors else None

    ogls = [c for c in candidates if any(s.is_ogl for s in c.sheets)]
    sel.ogl = ogls[0] if ogls else None

    return sel
