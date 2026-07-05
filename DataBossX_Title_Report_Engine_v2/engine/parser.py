"""Stage 2 -- parse workbook sheets into structured :class:`EvidenceRow` objects.

Real runsheets never share one header layout, so the parser maps a wide table of
header synonyms onto the canonical evidence fields, finds the header row
heuristically (the row that hits the most known synonyms), and reads every data
row beneath it. It records ``source_sheet`` + ``source_row`` (1-based, matching
what a human sees in Excel) on every row so the audit trail is exact.

Nothing is invented: a cell that is blank or unparseable becomes ``None``.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

from .ingest import Candidate, SheetInfo
from .models import EvidenceRow, classify_type

# Header synonym -> canonical evidence field.
HEADER_SYNONYMS: Dict[str, str] = {
    "instrument date": "instrument_date", "inst date": "instrument_date",
    "dated": "instrument_date", "date": "instrument_date",
    "execution date": "instrument_date", "effective date": "instrument_date",
    "recording date": "recording_date", "recorded": "recording_date",
    "rec date": "recording_date", "filed": "recording_date",
    "book/page": "book_page", "book page": "book_page", "bk/pg": "book_page",
    "book-page": "book_page", "vol/page": "book_page", "recording": "book_page",
    "reception": "book_page", "doc no": "book_page", "document number": "book_page",
    "instrument type": "instrument_type", "type": "instrument_type",
    "inst type": "instrument_type", "doc type": "instrument_type",
    "instrument": "instrument_type", "conveyance": "instrument_type",
    "grantor": "grantor", "from": "grantor", "lessor": "grantor",
    "assignor": "grantor", "seller": "grantor",
    "grantee": "grantee", "to": "grantee", "lessee": "grantee",
    "assignee": "grantee", "buyer": "grantee",
    "legal description": "legal_description", "legal": "legal_description",
    "description": "legal_description", "lands": "legal_description",
    "land description": "legal_description",
    "tract": "tract", "tract no": "tract", "tract #": "tract", "parcel": "tract",
    "notes": "notes", "remarks": "notes", "comments": "notes", "comment": "notes",
    "conveyed interest": "conveyed_interest", "conveyed": "conveyed_interest",
    "interest": "conveyed_interest", "interest conveyed": "conveyed_interest",
    "fraction": "conveyed_interest", "mineral interest": "conveyed_interest",
    "reserved interest": "reserved_interest", "reserved": "reserved_interest",
    "reservation": "reserved_interest", "retained": "reserved_interest",
    "gross acres": "gross_acres", "acres": "gross_acres", "acreage": "gross_acres",
    "gross": "gross_acres", "gross acreage": "gross_acres",
    "ogl": "lease_number", "ogl no": "lease_number", "ogl number": "lease_number",
    "lease no": "lease_number", "lease number": "lease_number",
    "lease": "lease_number", "ogl #": "lease_number",
}

# An OGL number looks like OGL-123, OGL 123, L-45, or a bare integer used as a
# lease id; it is NEVER a book/page pair (rule #8). We keep the raw token.
OGL_TOKEN_RE = re.compile(r"\b(?:OGL|L)[\s\-]?\d+\b", re.IGNORECASE)


def _canon_header(value) -> Optional[str]:
    if value is None:
        return None
    key = " ".join(str(value).split()).strip().lower().rstrip(":")
    if key in HEADER_SYNONYMS:
        return HEADER_SYNONYMS[key]
    # loosen: strip punctuation and retry
    key2 = re.sub(r"[^a-z0-9 /#-]", "", key)
    return HEADER_SYNONYMS.get(key2)


def _find_header(rows: List[list]) -> Optional[int]:
    """Return the index of the row that best matches known headers (>=2 hits)."""
    best_idx, best_hits = None, 1
    for idx, values in enumerate(rows[:15]):  # header is near the top
        hits = sum(1 for v in values if _canon_header(v))
        if hits > best_hits:
            best_idx, best_hits = idx, hits
    return best_idx


def parse_sheet(rows: List[list], sheet: SheetInfo,
                default_tract: Optional[str] = None) -> List[EvidenceRow]:
    """Parse the already-read grid of one sheet into evidence rows."""
    if not rows:
        return []
    header_idx = _find_header(rows)
    if header_idx is None:
        return []
    header_map: Dict[int, str] = {}
    for col, value in enumerate(rows[header_idx]):
        field_name = _canon_header(value)
        if field_name:
            header_map.setdefault(col, field_name)
    if len(header_map) < 2:
        return []

    out: List[EvidenceRow] = []
    for r_offset, values in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        data: Dict[str, object] = {}
        for col, field_name in header_map.items():
            if col < len(values):
                cell = values[col]
                if cell is not None and str(cell).strip() != "":
                    data[field_name] = cell
        if not data:
            continue
        # Skip rows that are clearly not instruments (no parties at all).
        if not str(data.get("grantor", "")).strip() and not str(data.get("grantee", "")).strip():
            # keep only if it carries an interest/acreage figure worth flagging
            if not any(k in data for k in ("conveyed_interest", "reserved_interest", "gross_acres")):
                continue

        tract = data.get("tract") or default_tract
        row = EvidenceRow(
            instrument_date=_s(data.get("instrument_date")),
            recording_date=_s(data.get("recording_date")),
            book_page=_s(data.get("book_page")),
            instrument_type=_s(data.get("instrument_type")),
            conveyance_type=classify_type(data.get("instrument_type")),
            grantor=data.get("grantor", ""),
            grantee=data.get("grantee", ""),
            legal_description=data.get("legal_description", ""),
            tract=_norm_tract(tract),
            notes=data.get("notes", ""),
            conveyed_interest=data.get("conveyed_interest"),
            reserved_interest=data.get("reserved_interest"),
            gross_acres=data.get("gross_acres"),
            lease_number=_clean_ogl(data.get("lease_number")),
            source_sheet=sheet.title,
            source_row=r_offset,
        )
        out.append(row)
    return out


def _s(v) -> Optional[str]:
    if v is None:
        return None
    text = " ".join(str(v).split())
    return text or None


def _norm_tract(v) -> Optional[str]:
    if v is None:
        return None
    text = " ".join(str(v).split())
    m = re.search(r"(\d+)", text)
    if m and re.search(r"tract", text, re.IGNORECASE):
        return f"Tract {int(m.group(1))}"
    if m and text.strip() == m.group(1):
        return f"Tract {int(m.group(1))}"
    return text or None


def _clean_ogl(v) -> Optional[str]:
    """Return only a genuine OGL-style token; reject book/page pairs (rule #8)."""
    if v is None:
        return None
    text = " ".join(str(v).split())
    if not text:
        return None
    if "/" in text or ":" in text:  # looks like book/page -> not an OGL number
        m = OGL_TOKEN_RE.search(text)
        return m.group(0).upper() if m else None
    m = OGL_TOKEN_RE.search(text)
    if m:
        return m.group(0).upper()
    if re.fullmatch(r"\d{1,5}", text):  # a bare lease id number
        return f"OGL-{int(text)}"
    return None


def read_grid(path: Path, sheet_title: str) -> List[list]:
    """Read one worksheet into a list-of-lists grid (values only)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_title not in wb.sheetnames:
            return []
        ws = wb[sheet_title]
        return [list(vals) for vals in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def parse_candidate(cand: Candidate) -> List[EvidenceRow]:
    """Parse every runsheet + tract + OGL sheet of a candidate into evidence."""
    out: List[EvidenceRow] = []
    for sheet in cand.sheets:
        if not (sheet.is_runsheet or sheet.is_tract or sheet.is_ogl):
            continue
        grid = read_grid(cand.path, sheet.title)
        default_tract = sheet.title if sheet.is_tract else None
        out.extend(parse_sheet(grid, sheet, default_tract=_norm_tract(default_tract)))
    return out
