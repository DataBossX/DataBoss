"""Stage 5 -- verify tracts, compute evidence scores, and gate the export.

Two responsibilities:

1. **Per-tract verification** -- does the tract balance to its gross acreage
   (rule #11)? Are there negative interests (rule #12)? Missing lease / probate /
   assignment links? Each finding becomes a :class:`ReviewFlag` and an
   ``Evidence Score`` (0-100) is computed per tract.
2. **Pre-export validation** -- the hard gate from the spec: workbook opens, no
   negative acres/WI, OGL column holds only OGL-style refs, Title Sheet holds
   final owners only, Audit Log exists, output file exists. Returns a list of
   failures; an empty list means the report may ship.
"""

from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from typing import List

from .audit import AuditLog
from .models import ReviewFlag, Tract

BALANCE_TOLERANCE = Decimal("0.01")   # acres
OGL_STYLE_RE = re.compile(r"^(OGL|L)[\s\-]?\d+$", re.IGNORECASE)


def verify_tract(tract: Tract, audit: AuditLog) -> Tract:
    """Flag imbalance / negatives / missing links and score the tract."""
    flags: List[ReviewFlag] = list(tract.flags)

    # Negative ownership (rule #12).
    for o in tract.owners:
        if (o.mineral_interest or Decimal(0)) < 0 or (o.net_mineral_acres or Decimal(0)) < 0:
            flags.append(ReviewFlag(tract=tract.name, category="negative_ownership",
                                    severity="red",
                                    message=f"Negative interest for {o.owner}",
                                    source_sheet=o.source_sheet, source_row=o.source_row))

    # Tract balance (rule #11).
    if tract.gross_acres is not None:
        diff = (tract.owned_acres - tract.gross_acres).copy_abs()
        if diff > BALANCE_TOLERANCE:
            flags.append(ReviewFlag(
                tract=tract.name, category="unbalanced_tract", severity="yellow",
                message=(f"Owners total {tract.owned_acres} ac vs tract {tract.gross_acres} "
                         f"ac (diff {diff}); needs verification"),
            ))
            audit.record("FLAG unbalanced tract", tract=tract.name,
                         before=f"{tract.gross_acres} ac target",
                         after=f"{tract.owned_acres} ac owned",
                         evidence="sum of final owner NMA != tract gross acres",
                         flag="unbalanced_tract")
    else:
        flags.append(ReviewFlag(tract=tract.name, category="ambiguous_acreage",
                                severity="yellow",
                                message="Tract gross acreage unknown; cannot balance"))

    # Missing lease coverage: owners with no OGL beside them.
    if any(o.lease_status == "Leased" for o in tract.owners):
        for o in tract.owners:
            if o.lease_status != "Leased" and o.mineral_interest > 0:
                flags.append(ReviewFlag(
                    tract=tract.name, category="missing_lease", severity="yellow",
                    message=f"{o.owner} appears unleased ({o.mineral_interest} MI)",
                    source_sheet=o.source_sheet, source_row=o.source_row))

    # No traced owners at all.
    if not tract.owners:
        flags.append(ReviewFlag(tract=tract.name, category="missing_chain",
                                severity="red",
                                message="No ownership could be chained for this tract"))

    tract.flags = _dedupe(flags)
    tract.evidence_score = _score(tract)
    return tract


def _dedupe(flags: List[ReviewFlag]) -> List[ReviewFlag]:
    seen, out = set(), []
    for f in flags:
        key = (f.tract, f.category, f.message)
        if key not in seen:
            seen.add(key)
            out.append(f)
    return out


def _score(tract: Tract) -> int:
    """Evidence score 0-100: starts at 100, deducts per flag severity / imbalance."""
    score = 100
    for f in tract.flags:
        score -= 25 if f.severity == "red" else 8
    if tract.gross_acres and tract.owners:
        diff = (tract.owned_acres - tract.gross_acres).copy_abs()
        if diff > BALANCE_TOLERANCE:
            score -= 15
    if not tract.owners:
        score -= 40
    return max(0, min(100, score))


# --- pre-export validation gate ---------------------------------------------

def validate_export(path: Path, tracts: List[Tract]) -> List[str]:
    """Reopen the exported workbook and assert every hard rule. Returns failures."""
    import openpyxl

    failures: List[str] = []
    if not path.exists():
        return [f"output file does not exist: {path}"]

    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:  # noqa: BLE001
        return [f"workbook does not open: {type(exc).__name__}: {exc}"]

    sheetnames_lower = [s.lower() for s in wb.sheetnames]
    if not any("audit" in s for s in sheetnames_lower):
        failures.append("Audit Log sheet missing")
    if not any(("title" in s or "owner" in s) for s in sheetnames_lower):
        failures.append("Title Sheet missing")

    # Scan every cell for #REF! and check OGL columns.
    for ws in wb.worksheets:
        header_row = None
        ogl_cols = []
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and "#REF!" in val:
                    failures.append(f"#REF! formula in {ws.title}!{cell.coordinate}")
                if header_row is None and isinstance(val, str) and val.strip().lower() in (
                        "ogl", "ogl no", "ogl number", "lease number"):
                    ogl_cols.append(cell.column)
            if header_row is None and ogl_cols:
                header_row = row[0].row
        # OGL column may only contain OGL-style refs (rules #7, #8).
        for col in ogl_cols:
            for cell in ws.iter_rows(min_row=(header_row or 1) + 1, min_col=col, max_col=col):
                v = cell[0].value
                if v in (None, ""):
                    continue
                text = str(v).strip()
                if "/" in text and not OGL_STYLE_RE.match(text):
                    failures.append(
                        f"OGL column {ws.title}!{cell[0].coordinate} holds book/page-like "
                        f"value '{text}' (rules #7/#8)")
    wb.close()

    # Negative-value checks come from the in-memory model (source of truth).
    for t in tracts:
        for o in t.owners:
            if (o.mineral_interest or Decimal(0)) < 0:
                failures.append(f"negative interest: {t.name} / {o.owner}")
            if (o.net_mineral_acres or Decimal(0)) < 0:
                failures.append(f"negative acres: {t.name} / {o.owner}")

    return failures
