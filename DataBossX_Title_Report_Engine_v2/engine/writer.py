"""Stage 6 -- export the final workbook, preserving the template exactly.

The template is the formatting authority. We copy it byte-for-byte and let
openpyxl load/save it, which preserves merged cells, borders, colors, column
widths, row heights, page setup, hidden sheets, named ranges, print settings,
freeze panes, and tab order/colors. We then:

* keep the Overview/map tab as the first sheet (rule: preserve it first);
* write a **Title Sheet** of final calculated owners only (rule #6) with the OGL
  number carried down beside each leased owner (rule #10);
* append an **Audit Log** sheet and a **Review Flags** sheet.

Highlighting discipline (rules #14-#16): the ONLY fill we ever apply is a single
yellow on a genuinely unresolved review cell. No decorative colors.
"""

from __future__ import annotations

import shutil
from decimal import Decimal
from pathlib import Path
from typing import List, Optional

from .audit import AuditLog
from .models import Tract

YELLOW = "FFFF00"

TITLE_SHEET = "Title Sheet"
AUDIT_SHEET = "Audit Log"
FLAGS_SHEET = "Review Flags"

TITLE_HEADERS = ["Tract", "Owner", "Mineral Interest", "Net Mineral Acres",
                 "Lease Status", "OGL", "Source", "Remarks"]
AUDIT_HEADERS = ["Timestamp", "Tract", "Source Sheet", "Source Row", "Action",
                 "Before", "After", "Evidence", "Flag"]
FLAG_HEADERS = ["Tract", "Severity", "Category", "Message", "Source Sheet", "Source Row"]


def _yellow_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")


def _bold():
    from openpyxl.styles import Font
    return Font(bold=True)


def _dec(v: Optional[Decimal]) -> str:
    """Fixed-point string (never scientific notation), trailing zeros trimmed."""
    if v is None:
        return ""
    q = v.quantize(Decimal("0.00000001")).normalize()
    return format(q, "f")


def _num(v: Optional[Decimal]):
    """A real numeric cell value (float) so Excel can sum it; blank if unknown."""
    if v is None:
        return ""
    return float(v.quantize(Decimal("0.00000001")))


def _new_workbook_with_overview(template: Optional[Path], out_path: Path):
    """Return an openpyxl workbook seeded from the template (formatting kept)."""
    import openpyxl

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if template is not None and Path(template).exists():
        shutil.copy2(template, out_path)
        wb = openpyxl.load_workbook(out_path)
        # Ensure an overview/map tab stays first if one exists.
        _move_overview_first(wb)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Overview"
        ws["A1"] = "SECTION 31-12N-24W, ROGER MILLS COUNTY, OKLAHOMA"
        ws["A1"].font = _bold()
        ws["A2"] = "Cursory Ownership / Title Report"
        ws["A4"] = ("No template workbook was supplied; a minimal Overview tab was "
                    "created. Provide data/template/template.xlsx to preserve full "
                    "formatting, plat/map, and page setup.")
    return wb


def _move_overview_first(wb) -> None:
    overview_idx = None
    for i, name in enumerate(wb.sheetnames):
        low = name.strip().lower()
        if any(low == k or low.startswith(k) for k in ("overview", "map", "plat", "cover")):
            overview_idx = i
            break
    if overview_idx is not None and overview_idx != 0:
        ws = wb[wb.sheetnames[overview_idx]]
        wb.move_sheet(ws, -overview_idx)


def _replace_sheet(wb, title: str, index: int):
    """Drop any existing sheet named ``title`` and create a fresh one at index."""
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title, index=min(index, len(wb.sheetnames)))


def write_report(tracts: List[Tract], audit: AuditLog, out_path: Path,
                 template: Optional[Path] = None, section: str = "") -> Path:
    wb = _new_workbook_with_overview(template, out_path)
    yellow = _yellow_fill()

    # --- Title Sheet: final owners only ------------------------------------
    # Placed right after the Overview tab (index 1) so the map stays first.
    ws = _replace_sheet(wb, TITLE_SHEET, index=1)
    ws.append([f"{section or 'SECTION 31-12N-24W'} — FINAL CALCULATED OWNERSHIP"])
    ws["A1"].font = _bold()
    ws.append([])
    ws.append(TITLE_HEADERS)
    for c in ws[3]:
        c.font = _bold()
    for tract in tracts:
        tacres = _dec(tract.gross_acres)
        ws.append([f"{tract.name}  (target {tacres} ac)" if tacres else tract.name])
        ws[ws.max_row][0].font = _bold()
        if not tract.owners:
            r = ws.max_row + 1
            ws.append(["", "NO OWNERSHIP CHAINED — NEEDS VERIFICATION"])
            ws.cell(row=r, column=2).fill = yellow
        unbalanced = any(f.category in ("unbalanced_tract", "ambiguous_acreage")
                         for f in tract.flags)
        for o in tract.owners:
            row = [tract.name, o.owner, _num(o.mineral_interest),
                   _num(o.net_mineral_acres), o.lease_status, o.lease_number or "",
                   f"{o.source_sheet} r{o.source_row}" if o.source_sheet else "",
                   o.remarks]
            ws.append(row)
        # Tract total line; highlight the acres cell yellow only if unbalanced.
        total_row = ["", "TRACT TOTAL", _num(tract.owned_interest),
                     _num(tract.owned_acres), "", "", "", ""]
        ws.append(total_row)
        r = ws.max_row
        ws.cell(row=r, column=2).font = _bold()
        if unbalanced:
            ws.cell(row=r, column=4).fill = yellow
            ws.cell(row=r, column=8, value="Needs Verification — does not balance").fill = yellow
        ws.append([])
    ws.freeze_panes = "A4"
    _autosize(ws)

    # --- Audit Log ----------------------------------------------------------
    aws = _replace_sheet(wb, AUDIT_SHEET, index=len(wb.sheetnames))
    aws.append(AUDIT_HEADERS)
    for c in aws[1]:
        c.font = _bold()
    for e in audit.entries:
        aws.append([e.timestamp, e.tract or "", e.source_sheet, e.source_row,
                    e.action, e.before, e.after, e.evidence, e.flag])
        if e.flag and e.flag not in ("assumption",):
            aws.cell(row=aws.max_row, column=9).fill = yellow
    aws.freeze_panes = "A2"
    _autosize(aws)

    # --- Review Flags -------------------------------------------------------
    fws = _replace_sheet(wb, FLAGS_SHEET, index=len(wb.sheetnames))
    fws.append(FLAG_HEADERS)
    for c in fws[1]:
        c.font = _bold()
    for tract in tracts:
        for f in tract.flags:
            fws.append([f.tract or "", f.severity, f.category, f.message,
                        f.source_sheet, f.source_row])
            # Highlight true unresolved review items yellow (rule #15).
            fws.cell(row=fws.max_row, column=4).fill = yellow
    fws.freeze_panes = "A2"
    _autosize(fws)

    # Land the reader on the Overview tab.
    wb.active = 0
    wb.save(out_path)
    wb.close()
    return out_path


def _autosize(ws, cap: int = 60) -> None:
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            widths[i] = max(widths.get(i, 10), min(len(str(v)) + 2, cap))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def write_audit_workbook(audit: AuditLog, out_path: Path) -> Path:
    """Standalone SECTION31_AUDIT_LOG.xlsx deliverable."""
    import openpyxl

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = AUDIT_SHEET
    ws.append(AUDIT_HEADERS)
    for c in ws[1]:
        c.font = _bold()
    for e in audit.entries:
        ws.append([e.timestamp, e.tract or "", e.source_sheet, e.source_row,
                   e.action, e.before, e.after, e.evidence, e.flag])
    ws.freeze_panes = "A2"
    _autosize(ws)
    wb.save(out_path)
    wb.close()
    return out_path


def write_ranking_workbook(candidates, selection, out_path: Path) -> Path:
    """SOURCE_FILE_RANKING.xlsx: every candidate, its score, and why it ranked."""
    import openpyxl

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Source Ranking"
    ws.append(["Rank", "File", "Inferred Kind", "Score", "Selected As",
               "Sheets", "Reasons", "Error"])
    for c in ws[1]:
        c.font = _bold()

    def role(c):
        roles = []
        if selection.runsheet is c:
            roles.append("RUNSHEET")
        if selection.template is c:
            roles.append("TEMPLATE")
        if selection.prior_report is c:
            roles.append("PRIOR REPORT")
        if selection.ogl is c:
            roles.append("OGL")
        return ", ".join(roles)

    for rank, c in enumerate(candidates, start=1):
        sheet_summary = "; ".join(
            s.title + ("*" if (s.is_runsheet or s.is_tract or s.is_ogl) else "")
            for s in c.sheets)
        ws.append([rank, str(c.path), c.kind, c.score, role(c),
                   sheet_summary, " | ".join(c.reasons), c.error])
    ws.freeze_panes = "A2"
    _autosize(ws, cap=80)
    wb.save(out_path)
    wb.close()
    return out_path
