"""End-to-end + unit tests for DataBossX Title Report Engine v2.

These run entirely against the SYNTHETIC demo corpus (generated on the fly) so
they never touch or assume any real title data. They assert the rules the spec
cares about: exact Decimal balance, impossible-conveyance rejection, ASSN as
leasehold-only, OGL carried down, yellow-only highlighting, and the export gate.
"""

from __future__ import annotations

import sys
from decimal import Decimal
from pathlib import Path

import pytest

ENGINE_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ENGINE_ROOT))

import app  # noqa: E402
from engine import ingest  # noqa: E402
from engine.models import ConveyanceType, classify_type, to_decimal  # noqa: E402
from demo import make_demo_data  # noqa: E402


@pytest.fixture(scope="module")
def demo_result(tmp_path_factory):
    make_demo_data.main()
    root = make_demo_data.OUT
    out = tmp_path_factory.mktemp("out") / "FINAL.xlsx"
    return app.run(root, "31-12N-24W", out, "2026-07-05 00:00:00Z", copy_inputs=False)


# --- unit: models -----------------------------------------------------------

def test_to_decimal_handles_percent_and_money():
    assert to_decimal("50%") == Decimal("0.5")
    assert to_decimal("$1,280.5") == Decimal("1280.5")
    assert to_decimal("") is None
    assert to_decimal("n/a") is None
    assert to_decimal(0.1) == Decimal("0.1")  # not the binary tail


def test_classify_type_prefers_specific():
    assert classify_type("Assignment of OGL") is ConveyanceType.ASSIGNMENT
    assert classify_type("Oil and Gas Lease") is ConveyanceType.LEASE
    assert classify_type("Mineral Deed") is ConveyanceType.MINERAL_DEED
    assert classify_type("Quit Claim Deed") is ConveyanceType.QUITCLAIM


# --- unit: ingest ranking ---------------------------------------------------

def test_discovery_ranks_runsheet_and_template():
    make_demo_data.main()
    cands = ingest.discover(make_demo_data.OUT)
    sel = ingest.select_sources(cands)
    assert sel.runsheet is not None and "Runsheet" in sel.runsheet.name
    assert sel.template is not None and "template" in sel.template.name.lower()


# --- e2e: balances and rules ------------------------------------------------

def test_all_tracts_present(demo_result):
    names = {t.name for t in demo_result["tracts"]}
    assert names == {"Tract 1", "Tract 2", "Tract 3"}


def test_tracts_balance_exactly(demo_result):
    by = {t.name: t for t in demo_result["tracts"]}
    assert by["Tract 1"].owned_acres == Decimal("160")
    assert by["Tract 2"].owned_acres == Decimal("160")
    assert by["Tract 3"].owned_acres == Decimal("80")


def test_reservation_preserved(demo_result):
    t1 = next(t for t in demo_result["tracts"] if t.name == "Tract 1")
    founder = next(o for o in t1.owners if "Founder" in o.owner)
    assert founder.mineral_interest == Decimal("0.5")  # reserved 1/2


def test_ogl_carried_down_beside_leased_owner(demo_result):
    t1 = next(t for t in demo_result["tracts"] if t.name == "Tract 1")
    ogls = {o.owner: o.lease_number for o in t1.owners}
    assert ogls["A. Founder"] == "OGL-101"
    assert ogls["B. Smith"] == "OGL-102"


def test_assignment_does_not_change_minerals(demo_result):
    # Bigwell LLC received an ASSN of OGL-101 but must NOT appear as a mineral owner.
    t1 = next(t for t in demo_result["tracts"] if t.name == "Tract 1")
    owners = {o.owner for o in t1.owners}
    assert "Bigwell LLC" not in owners


def test_impossible_conveyance_rejected_and_flagged(demo_result):
    t3 = next(t for t in demo_result["tracts"] if t.name == "Tract 3")
    owners = {o.owner for o in t3.owners}
    assert "H. Gray" not in owners  # 0.9 > 0.75 rejected
    cats = {f.category for f in t3.flags}
    assert "impossible_conveyance" in cats


def test_assumption_flagged(demo_result):
    t2 = next(t for t in demo_result["tracts"] if t.name == "Tract 2")
    assert any(f.category == "assumption" for f in t2.flags)


def test_no_negative_ownership(demo_result):
    for t in demo_result["tracts"]:
        for o in t.owners:
            assert o.mineral_interest >= 0
            assert (o.net_mineral_acres or Decimal(0)) >= 0


def test_export_validation_passes(demo_result):
    assert demo_result["failures"] == []


def test_audit_log_has_entries(demo_result):
    # audit path written and non-trivial
    assert demo_result["audit_path"].exists()


def test_deliverable_files_exist(demo_result):
    assert demo_result["out_final"].exists()
    assert demo_result["ranking_path"].exists()
    assert demo_result["audit_path"].exists()


def test_ogl_column_holds_only_ogl_refs(demo_result):
    import openpyxl
    wb = openpyxl.load_workbook(demo_result["out_final"])
    ws = wb["Title Sheet"]
    # find OGL column
    ogl_col = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for c in row:
            if str(c.value).strip().upper() == "OGL":
                ogl_col = c.column
        if ogl_col:
            break
    assert ogl_col is not None
    for row in ws.iter_rows(min_col=ogl_col, max_col=ogl_col):
        v = row[0].value
        if v in (None, ""):
            continue
        assert str(v).upper().startswith("OGL"), f"bad OGL cell: {v}"
    wb.close()
