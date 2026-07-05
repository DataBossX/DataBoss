"""Task 1 -- discover every project file under the Horizon root and classify it.

Non-destructive: nothing is moved, deleted, or overwritten. Produces a scored
inventory and a role ranking so the engine (and a human) can see exactly which
files were treated as the controlling runsheet / template / OGL schedule.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List

from .models import FileRecord
from .utils import get_logger, sha256

log = get_logger()

TITLE_EXTS = {".xlsx", ".xlsm", ".xls", ".csv", ".pdf", ".docx", ".txt", ".md", ".json", ".py"}

# Directories that never contain source title material.
SKIP_DIRS = {
    ".git", "node_modules", "venv", ".venv", "__pycache__", ".mypy_cache",
    ".pytest_cache", "dist", "build", ".idea", ".vscode", "site-packages",
    "DataBossX_Title_Report_Engine_v3",  # our own workspace -- don't ingest ourselves
}

# Filename fragments that identify the engine's OWN generated artifacts. These
# must never be re-ingested as a source (a prior run's output masquerading as a
# runsheet / prior report would let the engine "confirm" its own guesses).
SELF_OUTPUT_MARKERS = [
    "final_ownership_title_report",
    "section31_audit_log",
    "file_inventory", "source_file_ranking", "template_profile", "tract1_profile",
    "parsed_runsheet_evidence", "ogl_match_report", "chain_calculation_ledger",
    "assumptions_and_review_flags", "evidence_score_by_tract",
    "qa_validation_report", "run_summary",
]


def _is_self_output(name: str) -> bool:
    n = name.lower()
    return any(m in n for m in SELF_OUTPUT_MARKERS)


# Keyword weights used for role scoring.
SECTION_HINTS = ["31-12n-24w", "31 12n 24w", "3112n24w", "sec 31", "section 31",
                 "roger mills", "12n-24w", "12n24w"]


def _role_scores(name: str, ext: str, text_sample: str) -> Dict[str, float]:
    """Heuristic confidence (0..1) that a file plays each candidate role."""
    n = name.lower()
    t = (text_sample or "").lower()
    hay = n + " " + t
    scores: Dict[str, float] = {}

    def bump(role, amt):
        scores[role] = min(1.0, scores.get(role, 0.0) + amt)

    section_bonus = 0.15 if any(h in hay for h in SECTION_HINTS) else 0.0

    if ext in {".xlsx", ".xlsm", ".xls"}:
        if "runsheet" in n or "run sheet" in n or "run_sheet" in n or "instrument" in n or "index" in n:
            bump("runsheet", 0.85 + section_bonus)
        if "ogl" in n or "lease" in n:
            bump("ogl", 0.8 + section_bonus)
        if "template" in n or "blank" in n or "master" in n:
            bump("template", 0.8)
        if "report" in n or "cursory" in n or "nhe" in n or "title" in n:
            bump("prior_report", 0.7 + section_bonus)
            bump("template", 0.35)  # a prior report can donate its formatting
        # generic workbook: could be any of the above
        bump("runsheet", 0.25 + section_bonus)
        bump("ogl", 0.2)
        bump("prior_report", 0.2)
        if "runsheet" in t or "run sheet" in t:
            bump("runsheet", 0.4)
        if "ogl" in t or "oil and gas lease" in t or "lessor" in t:
            bump("ogl", 0.35)
    elif ext == ".csv":
        if "runsheet" in n or "instrument" in n:
            bump("runsheet", 0.7 + section_bonus)
        if "ogl" in n or "lease" in n:
            bump("ogl", 0.65)
        bump("runsheet", 0.2)
    elif ext == ".pdf":
        if "runsheet" in n or "run sheet" in n:
            bump("runsheet", 0.6 + section_bonus)
        if "report" in n or "cursory" in n or "title" in n:
            bump("prior_report", 0.55 + section_bonus)
        bump("prior_report", 0.15)
    elif ext in {".md", ".txt", ".json", ".docx", ".py"}:
        if any(k in hay for k in ("rule", "prompt", "instruction", "note", "chat", "readme")):
            bump("rule_reference", 0.6)
        else:
            bump("rule_reference", 0.2)
    return scores


def _text_sample(path: Path, ext: str, limit: int = 4000) -> str:
    try:
        if ext in {".txt", ".md", ".json", ".py", ".csv"}:
            return path.read_text(encoding="utf-8", errors="ignore")[:limit]
        if ext in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            parts = ["|".join(wb.sheetnames)]
            for ws in wb.worksheets[:3]:
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i > 8:
                        break
                    parts.append(" ".join("" if c is None else str(c) for c in row))
            wb.close()
            return " ".join(parts)[:limit]
    except Exception as exc:  # unreadable file -- record nothing, don't crash
        log.warning("could not sample %s: %s", path.name, exc)
    return ""


def discover(root: Path) -> List[FileRecord]:
    root = Path(root)
    records: List[FileRecord] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        # Only skip based on directory names BELOW the scan root -- checking the
        # full absolute path would wrongly skip everything when the root itself
        # lives inside a folder whose name is in SKIP_DIRS (e.g. a demo fixture
        # under the engine's own directory).
        try:
            rel_parts = path.relative_to(root).parts
        except ValueError:
            rel_parts = path.parts
        if any(part in SKIP_DIRS for part in rel_parts):
            continue
        ext = path.suffix.lower()
        if ext not in TITLE_EXTS:
            continue
        if _is_self_output(path.name):
            continue  # never ingest the engine's own generated artifacts
        try:
            stat = path.stat()
        except OSError:
            continue
        sample = _text_sample(path, ext)
        scores = _role_scores(path.name, ext, sample)
        best_role, best_conf = ("irrelevant", 0.0)
        if scores:
            best_role, best_conf = max(scores.items(), key=lambda kv: kv[1])
        note_bits = [f"{k}:{v:.2f}" for k, v in sorted(scores.items(), key=lambda kv: -kv[1])]
        records.append(FileRecord(
            path=str(path),
            name=path.name,
            ext=ext,
            size_bytes=stat.st_size,
            modified=datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            sha256=sha256(path)[:16],
            role=best_role,
            confidence=round(best_conf, 3),
            notes="; ".join(note_bits) if note_bits else "no role signal",
        ))
    log.info("discovered %d candidate files under %s", len(records), root)
    return records


def rank_sources(records: List[FileRecord]) -> Dict[str, List[FileRecord]]:
    """Bucket records by role and rank each bucket by confidence."""
    buckets: Dict[str, List[FileRecord]] = {
        "template": [], "runsheet": [], "ogl": [], "prior_report": [],
        "rule_reference": [], "irrelevant": [],
    }
    for rec in records:
        # Re-derive per-role scores from the stored notes so a file can appear in
        # multiple buckets (a prior report is both prior_report and template cand.)
        placed = False
        for token in rec.notes.split(";"):
            token = token.strip()
            if ":" not in token:
                continue
            role, val = token.rsplit(":", 1)
            role = role.strip()
            try:
                conf = float(val)
            except ValueError:
                continue
            if role in buckets and conf >= 0.4:
                clone = rec.model_copy(update={"confidence": round(conf, 3), "role": role})
                buckets[role].append(clone)
                placed = True
        if not placed:
            buckets["irrelevant"].append(rec)
    for role in buckets:
        buckets[role].sort(key=lambda r: r.confidence, reverse=True)
    return buckets
