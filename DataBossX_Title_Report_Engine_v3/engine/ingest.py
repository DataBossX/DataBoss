"""Load raw tabular data from workbooks / CSV / PDF into plain Python tables.

Every loader degrades gracefully: an unreadable file yields an empty table plus
a logged reason -- it never invents rows. PDF text extraction depends on a
backend (pdfplumber/pypdf); where that backend fails to import in this
environment we record that OCR / manual extraction is required.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from .utils import get_logger, pdf_support

log = get_logger()


@dataclass
class RawSheet:
    source_file: str
    sheet_name: str
    rows: List[List[object]] = field(default_factory=list)   # includes header row(s)
    note: str = ""

    @property
    def n_rows(self) -> int:
        return len(self.rows)


@dataclass
class RawBook:
    source_file: str
    sheets: List[RawSheet] = field(default_factory=list)
    note: str = ""


def load_workbook_rows(path: Path) -> RawBook:
    path = Path(path)
    ext = path.suffix.lower()
    book = RawBook(source_file=str(path))
    try:
        if ext in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                book.sheets.append(RawSheet(str(path), ws.title, rows))
            wb.close()
        elif ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(path)
            for ws in wb.sheets():
                rows = [[ws.cell_value(r, c) for c in range(ws.ncols)]
                        for r in range(ws.nrows)]
                book.sheets.append(RawSheet(str(path), ws.name, rows))
        elif ext == ".csv":
            import csv
            with open(path, newline="", encoding="utf-8", errors="ignore") as f:
                rows = [list(r) for r in csv.reader(f)]
            book.sheets.append(RawSheet(str(path), path.stem, rows))
        elif ext == ".pdf":
            book.sheets.extend(_load_pdf_tables(path))
        else:
            book.note = f"unsupported extension {ext}"
    except Exception as exc:
        book.note = f"load error: {type(exc).__name__}: {exc}"
        log.warning("ingest failed for %s: %s", path.name, exc)
    return book


def _load_pdf_tables(path: Path) -> List[RawSheet]:
    support = pdf_support()
    if not support.available:
        return [RawSheet(str(path), "PDF",
                         note=f"PDF backend unavailable ({support.reason}); "
                              "manual/OCR extraction required -- no data invented")]
    sheets: List[RawSheet] = []
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            for pno, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables() or []
                if tables:
                    for tno, table in enumerate(tables, start=1):
                        sheets.append(RawSheet(str(path), f"page{pno}_tbl{tno}",
                                               [list(r) for r in table]))
                else:
                    text = page.extract_text() or ""
                    if text.strip():
                        sheets.append(RawSheet(str(path), f"page{pno}_text",
                                               [[line] for line in text.splitlines()],
                                               note="text only -- no ruled table detected"))
        if not sheets:
            sheets.append(RawSheet(str(path), "PDF",
                                   note="no extractable text/tables; likely image-only "
                                        "PDF requiring OCR -- no data invented"))
    except Exception as exc:
        sheets.append(RawSheet(str(path), "PDF", note=f"pdf parse error: {exc}"))
    return sheets


def load_many(paths: List[Path]) -> Dict[str, RawBook]:
    return {str(p): load_workbook_rows(p) for p in paths}
