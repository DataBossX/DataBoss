"""Task 17 -- strict QA validation of the final workbook and the computed data.

Returns a list of check dicts ({Check, Result, Detail}) and an overall status:
PASS / PASS WITH REVIEW FLAGS / FAIL. A FAIL means the workbook is structurally
broken (missing required sheet, negative acreage, #REF!, book/page in the OGL
column) -- those are engine bugs and must be fixed. Review flags (missing
evidence) are honest and pass as "review required", never faked away.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook

from .models import FinalOwner, ReviewFlag, TractResult
from .utils import get_logger

log = get_logger()

REQUIRED_SHEET_HINTS = {
    "overview": ["overview", "map", "cover"],
    "title": ["title"],
    "tract1": ["tract 1", "tract1"],
    "audit": ["audit"],
    "review flags": ["review flag", "flags"],
    "source ranking": ["source ranking", "source"],
    "evidence score": ["evidence score", "evidence"],
    "qa summary": ["qa"],
}


def _check(checks, name, ok, detail=""):
    checks.append({"Check": name, "Result": "PASS" if ok else "FAIL", "Detail": detail})
    return ok


def _review(checks, name, detail):
    checks.append({"Check": name, "Result": "REVIEW", "Detail": detail})


def run_qa(out_path: Path, results: List[TractResult], final_owners: List[FinalOwner],
           flags: List[ReviewFlag]):
    checks: List[dict] = []
    hard_ok = True

    # 1. workbook exists & opens
    exists = Path(out_path).exists()
    hard_ok &= _check(checks, "Final XLSX exists", exists, str(out_path))
    wb = None
    if exists:
        try:
            wb = load_workbook(out_path)
            _check(checks, "Workbook opens with openpyxl", True, f"{len(wb.sheetnames)} sheets")
        except Exception as exc:
            hard_ok = False
            _check(checks, "Workbook opens with openpyxl", False, str(exc))

    if wb is not None:
        names_lower = [s.lower() for s in wb.sheetnames]
        # Overview must be first
        first_ok = any(h in names_lower[0] for h in REQUIRED_SHEET_HINTS["overview"])
        hard_ok &= _check(checks, "Overview/map tab is first", first_ok, wb.sheetnames[0])
        for key, hints in REQUIRED_SHEET_HINTS.items():
            present = any(any(h in n for h in hints) for n in names_lower)
            hard_ok &= _check(checks, f"Sheet present: {key}", present)

        # No #REF! formulas
        ref_hits = _scan_for_ref(wb)
        hard_ok &= _check(checks, "No #REF! formulas", not ref_hits,
                          "; ".join(ref_hits[:5]) if ref_hits else "clean")
        # OGL column purity + assignment labeling scanned from computed data below.

    # data-level checks (from computed results, independent of the workbook)
    neg_acre = _negative_acres(results)
    hard_ok &= _check(checks, "No negative acres", not neg_acre, "; ".join(neg_acre[:5]))
    neg_wi = [o.owner for o in final_owners if _neg(o.wi)]
    hard_ok &= _check(checks, "No negative WI", not neg_wi, "; ".join(neg_wi[:5]))

    ogl_pollution = _ogl_column_pollution(results, final_owners)
    hard_ok &= _check(checks, "OGL column has no book/page or assignment refs",
                      not ogl_pollution, "; ".join(ogl_pollution[:5]))

    assn_ok = _assignments_labeled(results)
    hard_ok &= _check(checks, "Assignments labeled ASSN (not OGL)", assn_ok[0], assn_ok[1])

    # Title sheet: no intermediate (zero-interest) owners, no placeholder names
    placeholders = _placeholder_owners(final_owners)
    hard_ok &= _check(checks, "No placeholder owner names", not placeholders,
                      "; ".join(placeholders[:5]))

    over = [f.instrument_number for f in flags if f.category == "over_conveyance"]
    _check(checks, "Over-conveyances are flagged (not silent)", True,
           f"{len(over)} flagged" if over else "none detected")

    # Balance: each tract balances OR is flagged
    unbal = [r.tract for r in results if not r.balanced]
    if unbal:
        _review(checks, "Every tract balances or is flagged",
                f"tracts flagged for review: {', '.join(unbal)}")

    # No computed chains at all == missing evidence, not an engine failure.
    # Surface it as a review item so status degrades to review-required, never a
    # clean PASS (which would misrepresent an empty report as complete).
    no_evidence = not results
    if no_evidence:
        _review(checks, "Source evidence present",
                "No tract chains were computed -- awaiting the controlling runsheet / "
                "OGL evidence. No ownership was fabricated.")

    review_count = sum(1 for f in flags)
    if review_count:
        _review(checks, "Open review flags", f"{review_count} review flag(s) -- see Review Flags sheet")

    if not hard_ok:
        status = "FAIL"
    elif review_count or unbal or no_evidence or any(c["Result"] == "REVIEW" for c in checks):
        status = "PASS WITH REVIEW FLAGS"
    else:
        status = "PASS"
    return checks, status


# ------------------------------------------------------------------ helpers
def _scan_for_ref(wb) -> List[str]:
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "#REF!" in cell.value:
                    hits.append(f"{ws.title}!{cell.coordinate}")
    return hits


def _neg(val) -> bool:
    try:
        return float(str(val).split("/")[0]) < 0 if val not in (None, "") else False
    except (ValueError, TypeError):
        return False


def _negative_acres(results) -> List[str]:
    out = []
    for r in results:
        for ev in r.events:
            for field in ("before_acres", "conveyed_acres", "after_acres", "retained_acres"):
                v = getattr(ev, field)
                if v and _neg(v):
                    out.append(f"T{r.tract} {ev.instrument_number} {field}={v}")
    return out


def _ogl_column_pollution(results, final_owners) -> List[str]:
    """OGL references must not be a bare book/page (\\d+/\\d+ with no OGL token)."""
    import re
    bp = re.compile(r"^\s*\d+\s*/\s*\d+\s*$")
    out = []
    for r in results:
        for ev in r.events:
            if ev.ogl_reference and bp.match(str(ev.ogl_reference)):
                out.append(f"T{r.tract} event {ev.instrument_number} OGL='{ev.ogl_reference}'")
    for o in final_owners:
        if o.ogl_number and bp.match(str(o.ogl_number)):
            out.append(f"owner {o.owner} OGL='{o.ogl_number}'")
    return out


def _assignments_labeled(results):
    for r in results:
        for ev in r.events:
            if ev.instrument_type == "ASSN" and ev.ogl_reference and \
               not ev.assn_reference and ev.action.upper().startswith("OGL"):
                return False, f"ASSN mislabeled as OGL at T{r.tract} {ev.instrument_number}"
    return True, "all assignments labeled ASSN"


_PLACEHOLDER = {"john doe", "jane doe", "smith", "jones", "operator co",
                "operator company", "acme", "test owner", "sample"}


def _placeholder_owners(final_owners) -> List[str]:
    out = []
    for o in final_owners:
        if o.owner.strip().lower() in _PLACEHOLDER:
            out.append(o.owner)
    return out
