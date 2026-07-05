"""Tasks 6 & 7 -- profile the template workbook and the Tract 1 master sheet.

The template is authority for formatting; Tract 1 is authority for column order,
OGL carry-down, note style, and total-row logic. We read (never mutate) the
template with openpyxl and capture enough structure that the writer can mirror
Tract 1 onto every tract sheet without flattening the template's formatting.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from .utils import get_logger

log = get_logger()

OVERVIEW_HINTS = ["overview", "map", "cover", "index", "summary map"]
TITLE_HINTS = ["title sheet", "title", "ownership", "final ownership"]
TRACT_HINTS = ["tract"]
OGL_HINTS = ["ogl", "lease schedule", "leases"]


@dataclass
class SheetProfile:
    name: str
    index: int
    hidden: bool = False
    max_row: int = 0
    max_col: int = 0
    merged_cells: List[str] = field(default_factory=list)
    freeze_panes: Optional[str] = None
    has_formulas: bool = False
    role: str = "other"


@dataclass
class Tract1Profile:
    sheet_name: str = ""
    header_row: int = 1
    headers: List[str] = field(default_factory=list)
    column_index: Dict[str, int] = field(default_factory=dict)  # canonical -> col idx (1-based)
    first_data_row: int = 2
    total_row: Optional[int] = None
    yellow_fill_seen: bool = False
    notes: str = ""


@dataclass
class TemplateProfile:
    path: str = ""
    sheet_names: List[str] = field(default_factory=list)
    sheets: List[SheetProfile] = field(default_factory=list)
    named_ranges: List[str] = field(default_factory=list)
    overview_sheet: Optional[str] = None
    title_sheet: Optional[str] = None
    tract1_sheet: Optional[str] = None
    ogl_sheet: Optional[str] = None
    tract_sheets: List[str] = field(default_factory=list)
    tract1: Tract1Profile = field(default_factory=Tract1Profile)
    available: bool = False
    note: str = ""


# Canonical header matching for the Tract 1 grid (reused from parser vocabulary).
_TRACT_HEADER_MAP = {
    "owner": ["owner", "grantee", "current owner", "mineral owner", "party"],
    "conveyance": ["conveyance", "conveyance type", "type", "instrument type", "doc type"],
    "instrument": ["instrument", "instrument no", "instrument number", "doc no"],
    "date": ["date", "instrument date", "recorded date", "recording date"],
    "book_page": ["book/page", "book page", "book", "vol/page"],
    "legal": ["legal", "legal description", "description", "tract description"],
    "acres": ["acres", "net acres", "net mineral acres", "nma", "acreage"],
    "wi": ["wi", "working interest"],
    "mineral_interest": ["mineral interest", "interest", "decimal interest", "mi"],
    "royalty": ["royalty", "nri", "orri", "npri"],
    "ogl": ["ogl", "ogl number", "ogl no", "lease number", "lease no"],
    "notes": ["notes", "remarks", "comments"],
}


def _match_tract_header(text: str) -> Optional[str]:
    t = str(text or "").strip().lower()
    if not t:
        return None
    for canon, syns in _TRACT_HEADER_MAP.items():
        if t in syns:
            return canon
    for canon, syns in _TRACT_HEADER_MAP.items():
        if any(s in t for s in syns):
            return canon
    return None


def _classify_sheet(name: str) -> str:
    n = name.lower()
    if any(h in n for h in OVERVIEW_HINTS):
        return "overview"
    if any(h in n for h in OGL_HINTS):
        return "ogl"
    if any(h in n for h in TRACT_HINTS):
        return "tract"
    if any(h in n for h in TITLE_HINTS):
        return "title"
    return "other"


def profile_template(path: Optional[Path]) -> TemplateProfile:
    prof = TemplateProfile()
    if not path or not Path(path).exists():
        prof.note = "no template workbook available in this checkout"
        return prof
    prof.path = str(path)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        prof.note = f"template load error: {exc}"
        return prof

    prof.available = True
    prof.sheet_names = list(wb.sheetnames)
    try:
        prof.named_ranges = list(wb.defined_names.keys())  # openpyxl >=3.1
    except Exception:
        prof.named_ranges = []

    for idx, ws in enumerate(wb.worksheets):
        role = _classify_sheet(ws.title)
        sp = SheetProfile(
            name=ws.title, index=idx,
            hidden=(ws.sheet_state != "visible"),
            max_row=ws.max_row, max_col=ws.max_column,
            merged_cells=[str(r) for r in ws.merged_cells.ranges],
            freeze_panes=ws.freeze_panes,
            role=role,
        )
        # cheap formula scan
        for row in ws.iter_rows(max_row=min(ws.max_row, 200)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    sp.has_formulas = True
                    break
            if sp.has_formulas:
                break
        prof.sheets.append(sp)

        if role == "overview" and not prof.overview_sheet:
            prof.overview_sheet = ws.title
        elif role == "title" and not prof.title_sheet:
            prof.title_sheet = ws.title
        elif role == "ogl" and not prof.ogl_sheet:
            prof.ogl_sheet = ws.title
        elif role == "tract":
            prof.tract_sheets.append(ws.title)

    # Overview must be first if present; if template's first sheet is not overview,
    # note it (writer will preserve existing order, never reorder silently).
    if prof.tract_sheets:
        # "Tract 1" preferred; else first tract sheet
        t1 = next((t for t in prof.tract_sheets if "1" in t), prof.tract_sheets[0])
        prof.tract1_sheet = t1
        prof.tract1 = _profile_tract1(wb[t1])

    wb.close()
    log.info("template profiled: %d sheets, tract1=%r", len(prof.sheets), prof.tract1_sheet)
    return prof


def _profile_tract1(ws) -> Tract1Profile:
    p = Tract1Profile(sheet_name=ws.title)
    # find header row: the row with the most matched tract headers in first 15 rows
    best_row, best_hits, best_map = 1, 0, {}
    for r in range(1, min(ws.max_row, 15) + 1):
        cmap, hits = {}, 0
        for c in range(1, ws.max_column + 1):
            canon = _match_tract_header(ws.cell(row=r, column=c).value)
            if canon and canon not in cmap:
                cmap[canon] = c
                hits += 1
        if hits > best_hits:
            best_row, best_hits, best_map = r, hits, cmap
    p.header_row = best_row
    p.first_data_row = best_row + 1
    p.column_index = best_map
    p.headers = [str(ws.cell(row=best_row, column=c).value or "")
                 for c in range(1, ws.max_column + 1)]
    # detect a total row (contains 'total') and any yellow fills
    for r in range(best_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            v = str(cell.value or "").lower()
            if "total" in v and p.total_row is None and r > best_row:
                p.total_row = r
            fill = cell.fill
            if fill and fill.fgColor and str(fill.fgColor.rgb or "").upper().endswith("FFF2CC"):
                p.yellow_fill_seen = True
    return p
