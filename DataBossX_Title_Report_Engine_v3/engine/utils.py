"""Shared helpers: exact math, logging, styling, and safe workbook I/O.

Every quantity that represents acreage or an ownership interest is carried as a
:class:`fractions.Fraction` or :class:`decimal.Decimal` -- never a Python float --
because a float cannot represent common mineral fractions (1/3, 1/16, 5/128)
exactly and the error compounds into a fabricated balance over a long chain.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from fractions import Fraction
from pathlib import Path
from typing import Optional, Union

# ---------------------------------------------------------------------------
# Styling constants (single source of truth so highlights stay disciplined).
# Yellow == a true unresolved assumption / review item. Nothing else is yellow.
# ---------------------------------------------------------------------------
YELLOW_HEX = "FFFFF2CC"      # soft yellow assumption fill
HEADER_HEX = "FF1F3864"      # dark navy header
HEADER_FONT_HEX = "FFFFFFFF"
TOTAL_HEX = "FFD9E1F2"       # light blue total row
FLAG_HEX = "FFFCE4D6"        # light orange for review-flag emphasis (non-yellow)
THIN = "thin"

ACRE_QUANT = Decimal("0.0001")   # acreage rounding precision (4 dp)


def get_logger(name: str = "databossx") -> logging.Logger:
    logger = logging.getLogger(name)
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        h = logging.StreamHandler()
        h.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-7s  %(message)s",
                                         datefmt="%H:%M:%S"))
        logger.addHandler(h)
    return logger


def now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def now_compact() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ---------------------------------------------------------------------------
# Exact interest / acreage parsing
# ---------------------------------------------------------------------------
Number = Union[str, int, float, Fraction, Decimal]

_FRACTION_RE = re.compile(r"^\s*(-?\d+)\s*/\s*(\d+)\s*$")
_MIXED_RE = re.compile(r"^\s*(-?\d+)\s+(\d+)\s*/\s*(\d+)\s*$")
_PERCENT_RE = re.compile(r"^\s*(-?[\d.]+)\s*%\s*$")
_DECOR_RE = re.compile(r"\b(an?|undivided|interest|of|the|mineral|royalty|all)\b", re.I)
_ORDINAL_RE = re.compile(r"(\d)(st|nd|rd|th)\b", re.I)

FULL = Fraction(1, 1)


class InterestError(ValueError):
    """Raised when a value cannot be parsed as an exact interest -- we never guess."""


def parse_interest(value: Optional[Number]) -> Fraction:
    """Parse ``value`` into an exact :class:`Fraction` or raise :class:`InterestError`.

    Accepts fractions ("1/2"), mixed numbers ("1 1/2"), decimals ("0.5" -> exact
    via Decimal), percentages ("12.5%"), and "x of y" ("1/2 of 8/8").
    """
    if value is None:
        raise InterestError("cannot parse interest from None")
    if isinstance(value, Fraction):
        return value
    if isinstance(value, int):
        return Fraction(value)
    if isinstance(value, Decimal):
        return Fraction(value)
    if isinstance(value, float):
        # A float sneaking in is a bug upstream; convert exactly via str() so we
        # at least do not silently introduce binary rounding error.
        return Fraction(Decimal(str(value)))

    s = str(value).strip()
    if not s:
        raise InterestError("cannot parse interest from empty string")

    low = s.lower()
    if " of " in low:
        left, right = low.split(" of ", 1)
        return parse_interest(left) * parse_interest(right)

    cleaned = _ORDINAL_RE.sub(r"\1", s)
    cleaned = _DECOR_RE.sub(" ", cleaned).strip()

    m = _PERCENT_RE.match(cleaned)
    if m:
        try:
            return Fraction(Decimal(m.group(1))) / 100
        except (InvalidOperation, ValueError) as exc:
            raise InterestError(f"unparseable percent {value!r}") from exc

    m = _MIXED_RE.match(cleaned)
    if m:
        whole, num, den = (int(g) for g in m.groups())
        if den == 0:
            raise InterestError(f"zero denominator in {value!r}")
        sign = -1 if whole < 0 else 1
        return Fraction(whole) + sign * Fraction(num, den)

    m = _FRACTION_RE.match(cleaned)
    if m:
        num, den = int(m.group(1)), int(m.group(2))
        if den == 0:
            raise InterestError(f"zero denominator in {value!r}")
        return Fraction(num, den)

    try:
        return Fraction(Decimal(cleaned))
    except (InvalidOperation, ValueError) as exc:
        raise InterestError(f"cannot parse interest from {value!r}") from exc


def try_parse_interest(value: Optional[Number]) -> Optional[Fraction]:
    try:
        return parse_interest(value)
    except InterestError:
        return None


def parse_acres(value: Optional[Number]) -> Optional[Decimal]:
    """Parse an acreage value into an exact Decimal, or None if unparseable."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, Fraction)):
        return Decimal(value.numerator) / Decimal(value.denominator) if isinstance(value, Fraction) else Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    s = str(value).strip().replace(",", "")
    s = re.sub(r"[^0-9./\-]", "", s)
    if not s:
        return None
    if "/" in s:
        frac = try_parse_interest(s)
        return None if frac is None else Decimal(frac.numerator) / Decimal(frac.denominator)
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def quantize_acres(value: Optional[Decimal]) -> Optional[Decimal]:
    if value is None:
        return None
    return value.quantize(ACRE_QUANT, rounding=ROUND_HALF_UP)


def fraction_to_decimal(frac: Optional[Fraction], places: int = 8) -> Optional[Decimal]:
    if frac is None:
        return None
    q = Decimal(1).scaleb(-places)
    return (Decimal(frac.numerator) / Decimal(frac.denominator)).quantize(q, rounding=ROUND_HALF_UP)


def format_fraction(frac: Optional[Fraction]) -> str:
    if frac is None:
        return ""
    if frac.denominator == 1:
        return str(frac.numerator)
    return f"{frac.numerator}/{frac.denominator}"


def fmt_acres(value: Optional[Decimal]) -> str:
    q = quantize_acres(value)
    return "" if q is None else f"{q.normalize():f}" if q == q.to_integral() else f"{q:.4f}"


# ---------------------------------------------------------------------------
# openpyxl styling helpers (used by the log writers and the report writer)
# ---------------------------------------------------------------------------
def style_header_row(ws, row: int, ncols: int):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    fill = PatternFill("solid", fgColor=HEADER_HEX)
    font = Font(bold=True, color=HEADER_FONT_HEX, size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side = Side(style=THIN, color="FFBFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    # Use a string coordinate: accessing ws.cell(row+1, 1) would materialize an
    # empty row and push subsequent ws.append() calls down by one.
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = f"{get_column_letter(1)}{row + 1}"


def yellow_cell(ws, row: int, col: int):
    """Highlight exactly one cell yellow -- reserved for true assumptions/review items."""
    from openpyxl.styles import PatternFill
    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=YELLOW_HEX)


def autosize(ws, max_width: int = 60):
    from openpyxl.utils import get_column_letter
    widths: dict = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            widths[col] = max(widths.get(col, 10), min(max_width, len(str(cell.value)) + 2))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def write_table_xlsx(path: Path, headers, rows, title: str = "Sheet1",
                     yellow_predicate=None):
    """Write a simple styled table to a standalone .xlsx log file.

    ``yellow_predicate(row_dict, col_name) -> bool`` optionally marks specific
    cells yellow (assumption/review). Used sparingly and only for real flags.
    """
    from openpyxl import Workbook
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(list(headers))
    style_header_row(ws, 1, len(headers))
    for r in rows:
        if isinstance(r, dict):
            values = [r.get(h, "") for h in headers]
        else:
            values = list(r)
        ws.append(["" if v is None else v for v in values])
        if yellow_predicate and isinstance(r, dict):
            excel_row = ws.max_row
            for ci, h in enumerate(headers, start=1):
                if yellow_predicate(r, h):
                    yellow_cell(ws, excel_row, ci)
    autosize(ws)
    wb.save(path)
    return path


def sha256(path: Path) -> str:
    import hashlib
    h = hashlib.sha256()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1 << 16), b""):
                h.update(chunk)
    except OSError:
        return ""
    return h.hexdigest()


@dataclass
class PdfSupport:
    available: bool
    reason: str = ""


def pdf_support() -> PdfSupport:
    """Report whether a PDF text backend imports cleanly in this environment."""
    try:
        import pdfplumber  # noqa: F401
        return PdfSupport(True)
    except Exception as exc:  # cryptography rust panic, missing wheel, etc.
        return PdfSupport(False, f"{type(exc).__name__}: {exc}")
