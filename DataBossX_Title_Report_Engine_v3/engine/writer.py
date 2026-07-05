"""Task 14 -- write the final workbook.

Two modes:
  * template mode -- load a COPY of the selected template and write verified
    output into its data areas, preserving every existing sheet, formula, merge,
    width, fill, and print setting (the template is never rebuilt from scratch);
  * clean mode -- when no template exists in the checkout, build a disciplined
    workbook from scratch that still honors the Tract 1 master style, keeps the
    Overview tab first, and adds the required audit/review/score/QA sheets.

In both modes yellow fill marks only true assumptions / unresolved review items.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .audit import AUDIT_COLUMNS, AuditEntry
from .models import Assumption, FinalOwner, ReviewFlag, TractResult
from .tract1_profile import TemplateProfile
from .utils import (YELLOW_HEX, autosize, get_logger, now_stamp, style_header_row,
                    yellow_cell)

log = get_logger()

TRACT_COLUMNS = ["Owner", "Conveyance Type", "Instrument No.", "Date", "Book/Page",
                 "Legal Description", "Net Acres", "Mineral Interest", "WI",
                 "Royalty/NPRI", "OGL", "Notes"]

TITLE_COLUMNS = ["Tract", "Final Owner", "Conveyance", "Instrument No.", "Net Acres",
                 "Mineral Interest", "WI", "Royalty/NPRI", "OGL", "Status", "Notes"]

_THIN = Side(style="thin", color="FFBFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_YELLOW = PatternFill("solid", fgColor=YELLOW_HEX)


def write_report(out_path: Path, results: List[TractResult],
                 final_owners: List[FinalOwner], assumptions: List[Assumption],
                 flags: List[ReviewFlag], audit: List[AuditEntry],
                 source_ranking: List[dict], evidence_scores: List[dict],
                 qa_summary: List[dict], template: Optional[TemplateProfile],
                 banner: str = "") -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if template and template.available and Path(template.path).exists():
        wb = _open_template_copy(template, out_path)
        _fill_into_template(wb, template, results, final_owners, banner)
    else:
        wb = _build_clean(results, final_owners, banner)
    # Added sheets (both modes).
    _add_audit_sheet(wb, audit)
    _add_flags_sheet(wb, flags)
    _add_assumptions_sheet(wb, assumptions)
    _add_source_ranking_sheet(wb, source_ranking)
    _add_evidence_score_sheet(wb, evidence_scores)
    _add_qa_sheet(wb, qa_summary)
    wb.save(out_path)
    log.info("final workbook written: %s", out_path)
    return out_path


# --------------------------------------------------------------------------- template mode
def _open_template_copy(template: TemplateProfile, out_path: Path) -> Workbook:
    shutil.copyfile(template.path, out_path)     # preserve original byte-for-byte
    return load_workbook(out_path)               # then edit the copy in place


def _fill_into_template(wb, template, results, final_owners, banner):
    # Only write into recognized data areas; never reorder or delete template sheets.
    by_tract = {r.tract: r for r in results}
    for tname in template.tract_sheets:
        ws = wb[tname]
        # best-effort: append verified rows below the template's first data row
        res = _tract_for_sheet(by_tract, tname)
        if res is None:
            continue
        start = (template.tract1.first_data_row
                 if tname == template.tract1_sheet else _guess_first_data_row(ws))
        _write_tract_rows(ws, res, start, template)
    if template.title_sheet:
        _write_title_rows(wb[template.title_sheet], final_owners, _guess_first_data_row(wb[template.title_sheet]))
    if banner and template.overview_sheet:
        _banner_cell(wb[template.overview_sheet], banner)


def _tract_for_sheet(by_tract, tname):
    digits = "".join(ch for ch in tname if ch.isdigit())
    return by_tract.get(digits) or by_tract.get(tname) or None


def _guess_first_data_row(ws) -> int:
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 12) + 1)]
        if sum(1 for v in vals if v not in (None, "")) >= 3:
            return r + 1
    return 2


# --------------------------------------------------------------------------- clean mode
def _build_clean(results, final_owners, banner) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _clean_overview(wb, results, banner)          # first tab
    _clean_title(wb, final_owners)
    if results:
        for res in results:
            _clean_tract(wb, res)
    else:
        _clean_tract_stub(wb, banner)             # honest placeholder, no fake owners
    return wb


def _clean_tract_stub(wb, banner):
    """Emit a Tract 1 placeholder when no runsheet evidence exists.

    This is NOT a fabricated tract -- it computes nothing and lists no owners. It
    exists so the workbook is structurally complete (Tract 1 present) while the
    banner makes the missing-evidence condition unmistakable.
    """
    ws = wb.create_sheet("Tract 1")
    ws["A1"] = "Tract 1 — awaiting controlling runsheet"
    ws["A1"].font = Font(bold=True, size=13, color="FF9C0006")
    msg = banner or ("No runsheet evidence was found in the Horizon root, so no "
                     "ownership was computed. No owners, OGLs, acreage, or book/page "
                     "were invented. Drop the controlling runsheet into the Horizon "
                     "folder and rerun.")
    ws["A2"] = msg
    ws["A2"].fill = PatternFill("solid", fgColor=YELLOW_HEX)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:L6")
    hr = 8
    for c, h in enumerate(TRACT_COLUMNS, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header_row(ws, hr, len(TRACT_COLUMNS))
    ws.cell(row=hr + 1, column=1, value="(no evidence — nothing computed)")
    autosize(ws, max_width=55)


def _clean_overview(wb, results, banner):
    ws = wb.create_sheet("Overview")
    ws["A1"] = "DataBossX Cursory Title Report"
    ws["A1"].font = Font(bold=True, size=16, color="FF1F3864")
    ws["A2"] = "Section 31, Township 12 North, Range 24 West, Roger Mills County, Oklahoma"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = f"Generated: {now_stamp()}   |   Engine: DataBossX Title Report Engine v3"
    row = 5
    if banner:
        ws.cell(row=row, column=1, value=banner).font = Font(bold=True, color="FF9C0006")
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="FFFFC7CE")
        row += 2
    headers = ["Tract", "Control Acres", "Control Source", "Chain Events",
               "Final Owners", "Balanced", "Evidence Score", "Note"]
    hr = row
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header_row(ws, hr, len(headers))
    r = hr
    for res in results:
        r += 1
        ws.cell(row=r, column=1, value=res.tract)
        ws.cell(row=r, column=2, value=res.control_acres or "")
        ws.cell(row=r, column=3, value=res.control_source)
        ws.cell(row=r, column=4, value=len(res.events))
        ws.cell(row=r, column=5, value=len(res.final_owners))
        ws.cell(row=r, column=6, value="Yes" if res.balanced else "NO")
        ws.cell(row=r, column=7, value=res.evidence_score)
        ws.cell(row=r, column=8, value=res.balance_note)
        if not res.balanced:
            ws.cell(row=r, column=6).font = Font(bold=True, color="FF9C0006")
    autosize(ws, max_width=70)
    ws.freeze_panes = "A" + str(hr + 1)


def _clean_title(wb, final_owners):
    ws = wb.create_sheet("Title Sheet")
    ws["A1"] = "Final Calculated Ownership (Title Sheet)"
    ws["A1"].font = Font(bold=True, size=13, color="FF1F3864")
    hr = 3
    for c, h in enumerate(TITLE_COLUMNS, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header_row(ws, hr, len(TITLE_COLUMNS))
    r = hr
    for o in final_owners:
        r += 1
        row_vals = [o.tract, o.owner, o.conveyance_type, o.instrument_number,
                    o.net_acres, o.mineral_interest, o.wi, o.royalty, o.ogl_number,
                    o.status, o.notes]
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = _BORDER
        if o.is_assumption:
            yellow_cell(ws, r, 2)  # only the owner cell, and only when assumed
        if o.status != "ok":
            ws.cell(row=r, column=10).font = Font(bold=True, color="FF9C5700")
    autosize(ws, max_width=60)
    ws.freeze_panes = "A" + str(hr + 1)


def _clean_tract(wb, res: TractResult):
    title = f"Tract {res.tract}"[:31]
    ws = wb.create_sheet(title)
    ws["A1"] = f"Tract {res.tract} — Ownership Chain"
    ws["A1"].font = Font(bold=True, size=13, color="FF1F3864")
    ws["A2"] = (f"Control acreage: {res.control_acres or 'UNDETERMINED'} "
                f"({res.control_source})   |   Evidence score: {res.evidence_score}")
    ws["A2"].font = Font(italic=True, size=10)
    hr = 4
    for c, h in enumerate(TRACT_COLUMNS, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header_row(ws, hr, len(TRACT_COLUMNS))
    _write_tract_rows(ws, res, hr + 1, template=None)
    autosize(ws, max_width=55)
    ws.freeze_panes = "A" + str(hr + 1)


def _write_tract_rows(ws, res: TractResult, start: int, template):
    """Write chain events, then final owners, then a total row (Tract 1 style)."""
    r = start - 1
    # Section: chain events
    for ev in res.events:
        r += 1
        vals = [ev.grantee or ev.grantor, ev.instrument_type, ev.instrument_number,
                ev.instrument_date, "", "", ev.conveyed_acres,
                "", "", "", ev.ogl_reference, ev.notes]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = _BORDER
        if ev.assumption:
            yellow_cell(ws, r, TRACT_COLUMNS.index("Notes") + 1)
        if ev.flag:
            ws.cell(row=r, column=TRACT_COLUMNS.index("Notes") + 1).font = \
                Font(bold=True, color="FF9C0006")
    # Spacer + final owners header
    r += 2
    ws.cell(row=r, column=1, value="FINAL OWNERS (calculated)").font = Font(bold=True)
    for o in res.final_owners:
        r += 1
        vals = [o.owner, o.conveyance_type, o.instrument_number, "", o.book_page,
                "", o.net_acres, o.mineral_interest, o.wi, o.royalty, o.ogl_number, o.notes]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = _BORDER
        if o.is_assumption:
            yellow_cell(ws, r, 1)
    # Total row
    r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=7, value=res.control_acres or "")
    tfill = PatternFill("solid", fgColor="FFD9E1F2")
    for c in range(1, len(TRACT_COLUMNS) + 1):
        ws.cell(row=r, column=c).fill = tfill
        ws.cell(row=r, column=c).font = Font(bold=True)
    ws.cell(row=r, column=len(TRACT_COLUMNS),
            value=res.balance_note)
    if not res.balanced:
        yellow_cell(ws, r, len(TRACT_COLUMNS))


def _write_title_rows(ws, final_owners, start):
    r = start - 1
    for o in final_owners:
        r += 1
        vals = [o.tract, o.owner, o.conveyance_type, o.instrument_number, o.net_acres,
                o.mineral_interest, o.wi, o.royalty, o.ogl_number, o.status, o.notes]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        if o.is_assumption:
            yellow_cell(ws, r, 2)


def _banner_cell(ws, banner):
    ws.cell(row=max(ws.max_row + 2, 2), column=1, value=banner).font = \
        Font(bold=True, color="FF9C0006")


# --------------------------------------------------------------------------- added sheets
def _simple_sheet(wb, title, headers, rows, yellow_cols=None):
    ws = wb.create_sheet(title[:31])
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    yellow_cols = yellow_cols or {}
    for r_i, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            v = row.get(h, "") if isinstance(row, dict) else (row[c - 1] if c - 1 < len(row) else "")
            cell = ws.cell(row=r_i, column=c, value=v)
            cell.border = _BORDER
        # yellow specific cells when the predicate says so
        if isinstance(row, dict):
            for col_name, pred in yellow_cols.items():
                if col_name in headers and pred(row):
                    yellow_cell(ws, r_i, headers.index(col_name) + 1)
    autosize(ws)
    return ws


def _add_audit_sheet(wb, audit: List[AuditEntry]):
    headers = [c.replace("_", " ").title() for c in AUDIT_COLUMNS]
    rows = []
    for e in audit:
        d = e.model_dump()
        rows.append({c.replace("_", " ").title(): d[c] for c in AUDIT_COLUMNS})
    _simple_sheet(wb, "Audit Log", headers, rows,
                  yellow_cols={"Assumption": lambda r: bool(r.get("Assumption"))})


def _add_flags_sheet(wb, flags: List[ReviewFlag]):
    headers = ["Tract", "Severity", "Category", "Message", "Instrument No.", "Recommended Action"]
    rows = [{"Tract": f.tract, "Severity": f.severity, "Category": f.category,
             "Message": f.message, "Instrument No.": f.instrument_number,
             "Recommended Action": f.recommended_action} for f in flags]
    _simple_sheet(wb, "Review Flags", headers, rows,
                  yellow_cols={"Severity": lambda r: r.get("Severity") == "yellow"})


def _add_assumptions_sheet(wb, assumptions: List[Assumption]):
    headers = ["Tract", "Kind", "Assumption Statement", "Basis", "Confidence Penalty", "Location"]
    rows = [{"Tract": a.tract, "Kind": a.kind, "Assumption Statement": a.statement,
             "Basis": a.basis, "Confidence Penalty": a.confidence_penalty,
             "Location": a.location} for a in assumptions]
    _simple_sheet(wb, "Assumptions", headers, rows,
                  yellow_cols={"Assumption Statement": lambda r: True})


def _add_source_ranking_sheet(wb, ranking: List[dict]):
    headers = ["Role", "Selected", "Path", "Confidence", "Reason"]
    _simple_sheet(wb, "Source Ranking", headers, ranking)


def _add_evidence_score_sheet(wb, scores: List[dict]):
    headers = ["Tract", "Evidence Score", "Balanced", "Control Acres", "Note"]
    _simple_sheet(wb, "Evidence Score", headers, scores,
                  yellow_cols={"Balanced": lambda r: str(r.get("Balanced")) == "NO"})


def _add_qa_sheet(wb, qa: List[dict]):
    headers = ["Check", "Result", "Detail"]
    _simple_sheet(wb, "QA Summary", headers, qa,
                  yellow_cols={"Result": lambda r: str(r.get("Result")).upper() not in ("PASS", "OK", "PASS ")})
