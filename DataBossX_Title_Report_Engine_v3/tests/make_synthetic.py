"""Build a SYNTHETIC, clearly-labeled Horizon folder to exercise the engine.

Every party name here is FICTIONAL (Greek-alphabet placeholders that are NOT in
the QA placeholder blacklist). This corpus contains no real ownership, lease, or
recording data -- it exists only to prove the chain math, OGL carry-down,
over-conveyance flagging, and balance checks run end-to-end.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

RUNSHEET_HEADERS = ["Tract", "Instrument Date", "Recording Date", "Instrument Type",
                    "Grantor", "Grantee", "Lessor", "Lessee", "Assignor", "Assignee",
                    "Legal Description", "Acres", "Conveyed Interest", "Reserved Interest",
                    "Book/Page", "OGL Number", "Notes"]

# Tract, date, recdate, type, grantor, grantee, lessor, lessee, assignor, assignee,
# legal, acres, conveyed, reserved, book/page, ogl, notes
RUNSHEET_ROWS = [
    # Tract 1 -- balances to 8/8, exercises opening assumption + OGL + ASSN
    ["1", "2001-01-15", "2001-01-20", "Warranty Deed", "ALPHA MINERALS LLC",
     "BRAVO HOLDINGS LLC", "", "", "", "", "N/2 Sec 31-12N-24W", "160", "1/2", "",
     "500/100", "", "SYNTHETIC fixture row"],
    ["1", "2005-06-01", "2005-06-10", "Oil and Gas Lease", "", "", "BRAVO HOLDINGS LLC",
     "CHARLIE OPERATING CO", "", "", "N/2 Sec 31-12N-24W", "80", "", "", "610/220",
     "OGL-1001", "lessor Bravo retains minerals"],
    ["1", "2010-03-03", "2010-03-09", "Assignment of Oil and Gas Lease", "", "", "", "",
     "CHARLIE OPERATING CO", "DELTA ENERGY PARTNERS", "N/2 Sec 31-12N-24W", "", "", "",
     "700/330", "", "assignment of OGL-1001 leasehold"],
    # Tract 2 -- exercises over-conveyance flag; still nets to 8/8
    ["2", "2002-02-02", "2002-02-08", "Mineral Deed", "ECHO FAMILY TRUST",
     "FOXTROT LP", "", "", "", "", "SW/4 Sec 31-12N-24W", "160", "1/2", "",
     "520/140", "", "SYNTHETIC fixture row"],
    ["2", "2008-08-08", "2008-08-15", "Mineral Deed", "FOXTROT LP", "GOLF RESOURCES INC",
     "", "", "", "", "SW/4 Sec 31-12N-24W", "160", "3/4", "", "660/260", "",
     "conveys 3/4 but grantor holds only 1/2 -> over-conveyance expected"],
]

OGL_HEADERS = ["OGL Number", "Lessor", "Lessee", "Effective Date", "Recording Date",
               "Book/Page", "Legal Description", "Royalty", "Tract", "Acres"]
OGL_ROWS = [
    ["OGL-1001", "BRAVO HOLDINGS LLC", "CHARLIE OPERATING CO", "2005-06-01",
     "2005-06-10", "610/220", "N/2 Sec 31-12N-24W", "3/16", "1", "80"],
    ["OGL-2002", "GOLF RESOURCES INC", "HOTEL DRILLING LLC", "2009-09-09",
     "2009-09-20", "680/280", "SW/4 Sec 31-12N-24W", "1/8", "2", "40"],
]


def _write_runsheet(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Runsheet"
    ws.append(RUNSHEET_HEADERS)
    for row in RUNSHEET_ROWS:
        ws.append(row)
    ogl = wb.create_sheet("OGL")
    ogl.append(OGL_HEADERS)
    for row in OGL_ROWS:
        ogl.append(row)
    wb.save(path)


def _write_template(path: Path):
    wb = Workbook()
    ov = wb.active
    ov.title = "Overview Map"
    ov["A1"] = "SYNTHETIC TEMPLATE — Overview / Map"
    ov["A2"] = "Section 31-12N-24W (fictional test fixture)"
    title = wb.create_sheet("Title Sheet")
    title["A1"] = "Final Ownership"
    title.append(["Tract", "Owner", "Net Acres", "Mineral Interest", "OGL", "Notes"])
    for tno in ("1", "2"):
        ws = wb.create_sheet(f"Tract {tno}")
        ws["A1"] = f"Tract {tno}"
        ws.append(["Owner", "Conveyance Type", "Instrument No.", "Date",
                   "Book/Page", "Legal Description", "Net Acres", "Mineral Interest",
                   "WI", "Royalty/NPRI", "OGL", "Notes"])
    wb.save(path)


def build_synthetic_horizon(dest: Path) -> Path:
    dest = Path(dest)
    dest.mkdir(parents=True, exist_ok=True)
    _write_runsheet(dest / "SYNTHETIC_31-12N-24W_Runsheet.xlsx")
    _write_template(dest / "SYNTHETIC_Title_Template.xlsx")
    (dest / "README_SYNTHETIC.txt").write_text(
        "These files are FICTIONAL test fixtures for the DataBossX engine. "
        "They contain no real ownership, lease, or recording data.\n",
        encoding="utf-8")
    return dest


if __name__ == "__main__":
    d = build_synthetic_horizon(Path(__file__).resolve().parent / "_synthetic_horizon")
    print("Wrote synthetic Horizon fixture to", d)
