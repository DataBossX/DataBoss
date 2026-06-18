"""
excel_writer.py
===============
openpyxl-based reading and writing.

Responsibilities:
  - Open a workbook while preserving formulas, hyperlinks, hidden sheets,
    and as much formatting as possible.
  - Detect the header row on each title/runsheet sheet.
  - Map known column headers to canonical field names.
  - Extract the document URL from each row (hyperlink target, HYPERLINK()
    formula, plain-text URL, or rich-text link).
  - Append AI review columns to the right of each processed sheet.
  - Write AI suggestions, color-code rows, and (only if apply_corrections)
    overwrite original cells.

The original workbook is NEVER modified in place - the caller always works
on a separate output copy.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger("horizon.excel")

URL_RE = re.compile(r"https?://[^\s\"'<>)\]]+", re.IGNORECASE)

# ---- Header detection ------------------------------------------------------
# Map many possible spreadsheet headers -> canonical field name.
HEADER_ALIASES: Dict[str, List[str]] = {
    "instrument_number": [
        "instrument #", "instr #", "instrument number", "instrument no",
        "instrument", "doc #", "document #", "document number",
    ],
    "document_type": ["instrument type", "type", "document type", "doc type"],
    "book_page": ["book/page", "book page", "book/pg", "bk/pg", "book and page"],
    "filed_date": ["filed date", "file date", "date filed"],
    "recorded_date": ["recorded date", "record date", "date recorded"],
    "effective_date": ["effective date", "date effective", "dated", "instrument date"],
    "grantor": ["grantor", "grantor(s)", "from"],
    "grantee": ["grantee", "grantee(s)", "to"],
    "legal_description": ["legal description", "legal", "lands", "description"],
    "section": ["section", "sec", "sec."],
    "township": ["township", "twp", "twp."],
    "range": ["range", "rng", "rge"],
    "county": ["county"],
    "state": ["state"],
    "remarks": ["comments", "remarks", "notes"],
    "doc_link": [
        "document link", "doc link", "link", "document", "url", "hyperlink",
        "image link", "image",
    ],
}

# Headers that, if any are present, mark a sheet as a title/runsheet sheet.
TITLE_SHEET_SIGNAL_FIELDS = {
    "instrument_number", "document_type", "book_page", "grantor", "grantee",
    "legal_description", "doc_link",
}

REVIEW_COLUMNS = [
    "AI Status",
    "AI Confidence",
    "AI Consensus",
    "AI Suggested Instrument #",
    "AI Suggested Instrument Type",
    "AI Suggested Book/Page",
    "AI Suggested Filed Date",
    "AI Suggested Recorded Date",
    "AI Suggested Effective Date",
    "AI Suggested Grantor",
    "AI Suggested Grantee",
    "AI Suggested Legal Description",
    "AI Suggested Section",
    "AI Suggested Township",
    "AI Suggested Range",
    "AI Notes",
    "AI Changed Fields",
    "AI Error",
    "AI Review Needed",
    "AI Processed Timestamp",
    "AI Source Link",
    "AI Viewer Method",
    "AI Model Used",
    "AI Validator Models Used",
]

# Map review "AI Suggested X" columns to canonical fields, in order.
SUGGEST_FIELD_ORDER = [
    ("AI Suggested Instrument #", "instrument_number"),
    ("AI Suggested Instrument Type", "document_type"),
    ("AI Suggested Book/Page", "book_page"),
    ("AI Suggested Filed Date", "filed_date"),
    ("AI Suggested Recorded Date", "recorded_date"),
    ("AI Suggested Effective Date", "effective_date"),
    ("AI Suggested Grantor", "grantor"),
    ("AI Suggested Grantee", "grantee"),
    ("AI Suggested Legal Description", "legal_description"),
    ("AI Suggested Section", "section"),
    ("AI Suggested Township", "township"),
    ("AI Suggested Range", "range"),
]

# Original-cell fields eligible for apply_corrections overwrite.
APPLY_FIELDS = [f for _, f in SUGGEST_FIELD_ORDER]

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ---------------------------------------------------------------------------
# Sheet model
# ---------------------------------------------------------------------------


class SheetMap:
    """Header layout for one worksheet."""

    def __init__(self, header_row: int):
        self.header_row = header_row
        self.field_to_col: Dict[str, int] = {}  # canonical field -> col index (1-based)
        self.review_col_start: Optional[int] = None
        self.review_col_index: Dict[str, int] = {}  # review header -> col index

    @property
    def is_title_sheet(self) -> bool:
        return bool(set(self.field_to_col) & TITLE_SHEET_SIGNAL_FIELDS)

    @property
    def link_col(self) -> Optional[int]:
        return self.field_to_col.get("doc_link")


def _normalize_header(text) -> str:
    return " ".join(str(text).strip().lower().split()) if text is not None else ""


def detect_header(ws, scan_rows: int = 15) -> Optional[SheetMap]:
    """Scan the first `scan_rows` rows to find the header row + column map."""
    best: Optional[SheetMap] = None
    best_hits = 0
    max_col = min(ws.max_column, 60)
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        field_to_col: Dict[str, int] = {}
        for c in range(1, max_col + 1):
            val = _normalize_header(ws.cell(row=r, column=c).value)
            if not val:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if field in field_to_col:
                    continue
                if val in aliases:
                    field_to_col[field] = c
                    break
        hits = len(field_to_col)
        if hits > best_hits:
            sm = SheetMap(r)
            sm.field_to_col = field_to_col
            best, best_hits = sm, hits
    # Require at least 2 recognizable headers to treat as a data sheet.
    if best and best_hits >= 2:
        return best
    return None


# ---------------------------------------------------------------------------
# Hyperlink extraction
# ---------------------------------------------------------------------------


def extract_link(ws, row: int, col: int) -> Optional[str]:
    """Find a document URL in a cell using the required priority order."""
    if not col:
        return None
    cell = ws.cell(row=row, column=col)

    # 1. cell.hyperlink.target
    if cell.hyperlink is not None and cell.hyperlink.target:
        return cell.hyperlink.target.strip()

    val = cell.value

    # 2. HYPERLINK() formula
    if isinstance(val, str) and val.upper().startswith("=HYPERLINK("):
        m = re.search(r'=HYPERLINK\(\s*"([^"]+)"', val, re.IGNORECASE)
        if m:
            return m.group(1).strip()

    # 3. plain-text URL inside the cell
    if isinstance(val, str):
        m = URL_RE.search(val)
        if m:
            return m.group(0).strip()

    # 4. rich-text hyperlink, if available
    try:
        if hasattr(val, "__iter__") and not isinstance(val, str):
            for part in val:
                target = getattr(part, "hyperlink", None)
                if target:
                    return str(target).strip()
    except Exception:
        pass

    return None


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------


def _norm_cmp(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _norm_date(value) -> str:
    """Normalize a date-ish value to YYYY-MM-DD for comparison only."""
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return _norm_cmp(s)


DATE_FIELDS = {"filed_date", "recorded_date", "effective_date"}


def compare_row(sheet: SheetMap, ws, row: int, merged: Dict[str, Optional[str]]) -> List[str]:
    """Return the list of canonical fields where AI differs from the sheet."""
    changed: List[str] = []
    for _, field in SUGGEST_FIELD_ORDER:
        col = sheet.field_to_col.get(field)
        ai_val = merged.get(field)
        if ai_val is None or str(ai_val).strip() == "":
            continue  # never flag where AI has nothing - never invent
        existing = ws.cell(row=row, column=col).value if col else None
        if field in DATE_FIELDS:
            if _norm_date(existing) != _norm_date(ai_val):
                changed.append(field)
        else:
            if _norm_cmp(existing) != _norm_cmp(ai_val):
                changed.append(field)
    return changed


# ---------------------------------------------------------------------------
# Review-column writing
# ---------------------------------------------------------------------------


def ensure_review_columns(ws, sheet: SheetMap) -> None:
    """Append the AI review headers to the right of the existing headers."""
    if sheet.review_col_start is not None:
        return
    start = ws.max_column + 2  # leave a one-column gap
    sheet.review_col_start = start
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for i, header in enumerate(REVIEW_COLUMNS):
        col = start + i
        cell = ws.cell(row=sheet.header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Readable, generous widths; legal description gets the most room.
        width = 40 if "Legal" in header else (22 if "Suggested" in header or "Notes" in header else 16)
        ws.column_dimensions[get_column_letter(col)].width = width
        sheet.review_col_index[header] = col
    # Freeze the header row so it stays visible while scrolling long sheets.
    try:
        ws.freeze_panes = ws.cell(row=sheet.header_row + 1, column=1)
    except Exception:
        pass


def _fmt(value) -> str:
    return "" if value is None else str(value)


def classify_status(status: str, confidence: float, changed: List[str], config: dict) -> str:
    """Return one of: green / yellow / red."""
    red_states = {"failed", "blocked", "login_expired", "unreadable", "error"}
    if status in red_states:
        return "red"
    review_thr = float(config.get("ai_confidence_review_threshold", 0.85))
    safe_thr = float(config.get("ai_confidence_auto_safe_threshold", 0.97))
    if confidence < review_thr:
        return "red"
    if changed or confidence < safe_thr:
        return "yellow"
    return "green"


def write_review(ws, sheet: SheetMap, row: int, payload: dict, config: dict) -> str:
    """Write one row's AI results into the review columns. Returns color."""
    ensure_review_columns(ws, sheet)
    ci = sheet.review_col_index
    merged: Dict = payload.get("merged", {})
    status = payload.get("status", "processed")
    confidence = float(payload.get("confidence", 0.0))
    changed = payload.get("changed_fields", [])
    color = classify_status(status, confidence, changed, config)

    def put(header: str, value) -> None:
        ws.cell(row=row, column=ci[header], value=value)

    put("AI Status", status)
    put("AI Confidence", round(confidence, 3) if confidence else 0)
    put("AI Consensus", ", ".join(payload.get("consensus_fields", [])))
    for header, field in SUGGEST_FIELD_ORDER:
        put(header, _fmt(merged.get(field)))
    put("AI Notes", payload.get("evidence_summary", ""))
    put("AI Changed Fields", ", ".join(changed))
    put("AI Error", payload.get("error", ""))
    put("AI Review Needed", "YES" if payload.get("needs_human_review") else "no")
    put("AI Processed Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    put("AI Source Link", payload.get("link", ""))
    put("AI Viewer Method", payload.get("viewer_method", ""))
    put("AI Model Used", payload.get("model_used", ""))
    put("AI Validator Models Used", ", ".join(payload.get("validators_used", [])))

    # Color-code the AI Status cell and the changed original cells.
    fill = {"green": FILL_GREEN, "yellow": FILL_YELLOW, "red": FILL_RED}[color]
    ws.cell(row=row, column=ci["AI Status"]).fill = fill
    if config.get("highlight_changed_cells", True):
        for header, field in SUGGEST_FIELD_ORDER:
            if field in changed:
                ws.cell(row=row, column=ci[header]).fill = FILL_YELLOW

    # apply_corrections: only overwrite original cells when explicitly enabled.
    if config.get("apply_corrections", False):
        for field in APPLY_FIELDS:
            if field in changed and merged.get(field):
                col = sheet.field_to_col.get(field)
                if col:
                    ws.cell(row=row, column=col, value=merged[field])
                    ws.cell(row=row, column=col).fill = FILL_YELLOW

    return color


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------


def load_workbook(path: str):
    """Load a workbook preserving formulas, links, and hidden sheets."""
    from openpyxl import load_workbook as _load

    # data_only=False keeps formulas (incl. HYPERLINK); keep_links preserves links.
    return _load(path, data_only=False, keep_links=True, rich_text=True)


def iter_data_rows(ws, sheet: SheetMap) -> List[int]:
    """Yield 1-based data row indices below the header row."""
    return list(range(sheet.header_row + 1, ws.max_row + 1))


def add_summary_sheet(wb, stats: dict, workbook_name: str) -> None:
    """Prepend an '_AI_Review_Summary' sheet with run stats and a color legend.

    Inserted at the front so the reviewer sees it first. Never touches data sheets.
    """
    title = "_AI_Review_Summary"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    big = Font(bold=True, size=14)
    bold = Font(bold=True)
    ws["A1"] = "Horizon Title Link Extractor - AI Review Summary"
    ws["A1"].font = big
    ws["A2"] = "Source workbook"
    ws["A2"].font = bold
    ws["B2"] = workbook_name
    ws["A3"] = "Generated"
    ws["A3"].font = bold
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    r = 5
    ws.cell(row=r, column=1, value="Statistic").font = bold
    ws.cell(row=r, column=2, value="Value").font = bold
    r += 1
    for k, v in stats.items():
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Color legend").font = bold
    r += 1
    legend = [
        ("GREEN", "High confidence, no major mismatch", FILL_GREEN),
        ("YELLOW", "Possible correction / medium confidence / disagreement", FILL_YELLOW),
        ("RED", "Failed / blocked / login expired / unreadable / low confidence", FILL_RED),
    ]
    for label, desc, fill in legend:
        c = ws.cell(row=r, column=1, value=label)
        c.fill = fill
        c.font = bold
        ws.cell(row=r, column=2, value=desc)
        r += 1

    r += 1
    note = ws.cell(
        row=r, column=1,
        value="REVIEW ONLY: the original workbook was not modified. Suggested "
              "corrections are in the AI columns on each data sheet.",
    )
    note.font = Font(italic=True, color="888888")
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass
