"""
test_logic.py
=============
Automated tests for the pure (non-network) logic of HorizonTitleLinkExtractor:
config validation, Excel header/hyperlink/comparison/writing, Drive workbook
selection, the AI consensus router, caching, and the HTML report.

These run without Playwright, Google Drive, or any AI API key.
Run:  python -m pytest   (or RUN_TESTS.bat)
"""

import os

import pytest
from openpyxl import Workbook

import ai_extractors as ai
import cache as cache_mod
import config_schema
import drive_sync as ds
import excel_writer as ew
import report as report_mod


# ----------------------------- config_schema ------------------------------


def test_config_defaults_and_validation():
    cfg = config_schema.validate_config({})
    assert cfg["apply_corrections"] is False  # review-only default
    assert cfg["test_mode_rows"] == 5
    assert cfg["enable_ai_cache"] is True


def test_config_rejects_bad_threshold():
    with pytest.raises(Exception):
        config_schema.validate_config({"ai_confidence_review_threshold": 1.5})


def test_config_preserves_unknown_keys():
    cfg = config_schema.validate_config({"some_future_key": 123})
    assert cfg["some_future_key"] == 123


# ------------------------------ excel_writer -------------------------------


def _make_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Runsheet"
    ws.append(["Some Title Banner"])
    ws.append([
        "Instrument #", "Instrument Type", "Book/Page", "Filed Date",
        "Grantor", "Grantee", "Legal Description", "Section", "Township",
        "Range", "Document Link",
    ])
    ws.append([
        "2021-001", "Warranty Deed", "BK 100 PG 5", "2021-01-15",
        "John Doe", "Jane Roe", "SW/4 SEC 31", "31", "12N", "24W",
        "https://county.example/preview?doc=1",
    ])
    return wb, ws


def test_detect_header_and_link_columns():
    wb, ws = _make_sheet()
    sm = ew.detect_header(ws)
    assert sm is not None
    assert sm.header_row == 2
    assert sm.is_title_sheet
    assert sm.link_col == 11
    assert sm.field_to_col["grantor"] == 5


def test_extract_link_all_paths():
    wb, ws = _make_sheet()
    sm = ew.detect_header(ws)
    # plain text URL
    assert ew.extract_link(ws, 3, sm.link_col) == "https://county.example/preview?doc=1"
    # HYPERLINK formula
    ws.cell(row=4, column=11, value='=HYPERLINK("https://county.example/d2","view")')
    assert ew.extract_link(ws, 4, sm.link_col) == "https://county.example/d2"
    # cell.hyperlink.target
    c = ws.cell(row=5, column=11, value="Open")
    c.hyperlink = "https://county.example/d3"
    assert ew.extract_link(ws, 5, sm.link_col) == "https://county.example/d3"
    # no link
    assert ew.extract_link(ws, 6, sm.link_col) is None


def test_compare_row_date_and_whitespace_normalization():
    wb, ws = _make_sheet()
    sm = ew.detect_header(ws)
    # Same date in a different format -> NOT changed.
    merged = {
        "filed_date": "01/15/2021",        # equals 2021-01-15
        "grantor": "  John   Doe ",        # whitespace-only diff -> not changed
        "legal_description": "NW/4 SEC 31",  # genuinely different -> changed
        "instrument_number": None,          # AI empty -> never flagged
    }
    changed = ew.compare_row(sm, ws, 3, merged)
    assert "filed_date" not in changed
    assert "grantor" not in changed
    assert "legal_description" in changed
    assert "instrument_number" not in changed


def test_classify_status_colors():
    cfg = {"ai_confidence_review_threshold": 0.85, "ai_confidence_auto_safe_threshold": 0.97}
    assert ew.classify_status("failed", 0.99, [], cfg) == "red"
    assert ew.classify_status("processed", 0.5, [], cfg) == "red"
    assert ew.classify_status("processed", 0.9, ["grantor"], cfg) == "yellow"
    assert ew.classify_status("processed", 0.99, [], cfg) == "green"


def test_write_review_never_touches_original_when_apply_off():
    wb, ws = _make_sheet()
    sm = ew.detect_header(ws)
    ew.ensure_review_columns(ws, sm)
    original_grantor = ws.cell(row=3, column=sm.field_to_col["grantor"]).value
    payload = {
        "merged": {"grantor": "DIFFERENT NAME"}, "status": "processed",
        "confidence": 0.9, "changed_fields": ["grantor"], "consensus_fields": [],
        "evidence_summary": "x", "needs_human_review": False, "error": "",
        "link": "u", "viewer_method": "img", "model_used": "m", "validators_used": [],
    }
    ew.write_review(ws, sm, 3, payload, {"apply_corrections": False,
                                         "highlight_changed_cells": True,
                                         "ai_confidence_review_threshold": 0.85,
                                         "ai_confidence_auto_safe_threshold": 0.97})
    # Original cell unchanged.
    assert ws.cell(row=3, column=sm.field_to_col["grantor"]).value == original_grantor
    # Suggestion written to the review column.
    sug_col = sm.review_col_index["AI Suggested Grantor"]
    assert ws.cell(row=3, column=sug_col).value == "DIFFERENT NAME"


def test_write_review_applies_correction_when_enabled():
    wb, ws = _make_sheet()
    sm = ew.detect_header(ws)
    ew.ensure_review_columns(ws, sm)
    payload = {
        "merged": {"grantor": "CORRECTED NAME"}, "status": "processed",
        "confidence": 0.99, "changed_fields": ["grantor"], "consensus_fields": [],
        "evidence_summary": "x", "needs_human_review": False, "error": "",
        "link": "u", "viewer_method": "img", "model_used": "m", "validators_used": [],
    }
    ew.write_review(ws, sm, 3, payload, {"apply_corrections": True,
                                         "highlight_changed_cells": True,
                                         "ai_confidence_review_threshold": 0.85,
                                         "ai_confidence_auto_safe_threshold": 0.97})
    assert ws.cell(row=3, column=sm.field_to_col["grantor"]).value == "CORRECTED NAME"


def test_summary_sheet_added_first():
    wb, ws = _make_sheet()
    ew.add_summary_sheet(wb, {"processed_rows": 3, "failed_rows": 0}, "book.xlsx")
    assert wb.sheetnames[0] == "_AI_Review_Summary"


# ------------------------------- drive_sync --------------------------------


def test_extract_folder_id():
    url = "https://drive.google.com/drive/folders/1cmfCYsFkUoBXgo5fanDTOZsprtFZrGhc"
    assert ds.extract_folder_id(url) == "1cmfCYsFkUoBXgo5fanDTOZsprtFZrGhc"


def test_select_workbook_excludes_and_prefers_newest():
    cands = [
        ds.WorkbookCandidate("31-12N-24W_Roger_Mills_Cursory_Title_Report_(6-10-2026).xlsx", modified=100),
        ds.WorkbookCandidate("31-12N-24W_Roger_Mills_Cursory_Title_Report_(6-15-2026).xlsx", modified=200),
        ds.WorkbookCandidate("something_AI_REVIEW.xlsx", modified=999),
        ds.WorkbookCandidate("backup_copy.xlsx", modified=999),
        ds.WorkbookCandidate("_AI_Updated_thing.xlsx", modified=999),
    ]
    chosen = ds.select_workbook(cands, None)
    assert chosen.name == "31-12N-24W_Roger_Mills_Cursory_Title_Report_(6-15-2026).xlsx"


def test_select_workbook_override():
    cands = [
        ds.WorkbookCandidate("a.xlsx", modified=100),
        ds.WorkbookCandidate("target.xlsx", modified=1),
    ]
    chosen = ds.select_workbook(cands, "target.xlsx")
    assert chosen.name == "target.xlsx"


def test_timestamped_folder_name_format():
    name = ds.timestamped_folder_name("_AI_Updated_Reports")
    assert name.startswith("_AI_Updated_Reports_")
    # _AI_Updated_Reports_YYYY-MM-DD_HHMM
    tail = name.rsplit("_", 2)[-2:]
    assert len(tail[0]) == 10 and len(tail[1]) == 4


# ------------------------------ ai_extractors ------------------------------


def test_safe_parse_strips_code_fences():
    raw = '```json\n{"instrument_number":"5","grantor":"A","confidence":0.9}\n```'
    te = ai._safe_parse(raw)
    assert te.instrument_number == "5"
    assert te.confidence == 0.9


def test_safe_parse_garbage_returns_empty_low_confidence():
    te = ai._safe_parse("not json at all")
    assert te.confidence == 0.0
    assert te.needs_human_review is True


def test_consensus_two_of_three_and_empty_not_counted():
    r = ai.ModelRouter({"openai_model": "gpt-5.2", "ai_confidence_review_threshold": 0.85,
                        "enable_ai_cache": False})
    res = ai.RouterResult()
    a = ai.TitleExtraction(instrument_number="100", grantor="John Doe",
                           legal_description="SW/4", confidence=0.9, needs_human_review=False)
    b = ai.TitleExtraction(instrument_number="100", grantor="Jon Doe",
                           legal_description="SW/4", confidence=0.88, needs_human_review=False)
    c = ai.TitleExtraction(instrument_number="100", grantor="John Doe",
                           legal_description="SE/4", confidence=0.8, needs_human_review=False)
    r._merge([("openai", a), ("gemini", b), ("claude", c)], res)
    # instrument: all agree; grantor: 2 of 3 -> John Doe; legal: SW/4 has 2.
    assert res.merged["instrument_number"] == "100"
    assert res.merged["grantor"] == "John Doe"
    # document_type is null in all -> must NOT be marked consensus.
    assert "document_type" not in res.consensus_fields
    assert set(res.consensus_fields) >= {"instrument_number", "grantor", "legal_description"}


def test_should_validate_rules():
    r = ai.ModelRouter({"openai_model": "x", "ai_confidence_review_threshold": 0.85,
                        "enable_ai_cache": False})
    low = ai.TitleExtraction(confidence=0.5, needs_human_review=False)
    high = ai.TitleExtraction(confidence=0.99, needs_human_review=False, uncertain_fields=[])
    assert r._should_validate(low, row_changed=False) is True
    assert r._should_validate(high, row_changed=False) is False
    assert r._should_validate(high, row_changed=True) is True


# --------------------------------- cache -----------------------------------


def test_cache_roundtrip(tmp_path):
    c = cache_mod.AIResultCache(str(tmp_path / "c"), enabled=True)
    c.put("k1", {"a": 1})
    assert c.get("k1") == {"a": 1}
    assert c.get("missing") is None


def test_cache_disabled_is_noop(tmp_path):
    c = cache_mod.AIResultCache(str(tmp_path / "c"), enabled=False)
    c.put("k1", {"a": 1})
    assert c.get("k1") is None


def test_image_signature_stable_and_model_sensitive():
    imgs = [b"abc", b"def"]
    s1 = cache_mod.image_signature(imgs, "gpt")
    s2 = cache_mod.image_signature(imgs, "gpt")
    s3 = cache_mod.image_signature(imgs, "claude")
    assert s1 == s2
    assert s1 != s3


def test_usage_tracker_counts():
    t = cache_mod.UsageTracker()
    t.record_call("gpt", 10, 5)
    t.record_call("gpt", 20, 5)
    t.record_cache_hit("gpt")
    s = t.summary()
    assert s["ai_calls_total"] == 2
    assert s["ai_cache_hits_total"] == 1
    assert s["ai_input_tokens_total"] == 30
    assert s["ai_output_tokens_total"] == 10


# --------------------------------- report ----------------------------------


def test_generate_report(tmp_path):
    logs = tmp_path / "logs"
    logs.mkdir()
    (logs / "run_log.csv").write_text(
        "timestamp,workbook,sheet,row,link,status,confidence,changed_fields,review_needed,error_message\n"
        "2026-06-18,book.xlsx,Runsheet,3,http://x,green,0.99,,no,\n"
        "2026-06-18,book.xlsx,Runsheet,4,http://y,red,0,,yes,login expired\n",
        encoding="utf-8",
    )
    (logs / "summary.csv").write_text("processed_rows,2\nfailed_rows,1\n", encoding="utf-8")
    out = report_mod.generate_report(str(logs), "book.xlsx")
    assert os.path.exists(out)
    html = open(out, encoding="utf-8").read()
    assert "Review Report" in html
    assert "login expired" in html
    assert "green" in html and "red" in html
