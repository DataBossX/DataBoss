# -*- coding: utf-8 -*-
"""Additive CSV exports + a portable pull-list. Reads only verified data;
invents nothing. Writes to RogerMills_Title_Finisher/exports/."""
import sys, os, csv, re, datetime
sys.path.insert(0, 'report')
import data as D
import openpyxl

os.makedirs('exports', exist_ok=True)
TRACT_LEGAL = {t['n']: t['legal'] for t in D.TRACTS}
TRACT_GROSS = {t['n']: t['gross'] for t in D.TRACTS}

# ---------- 1) ownership_by_tract.csv (with decimal interest in tract) ----------
with open('exports/ownership_by_tract.csv','w',newline='') as f:
    w=csv.writer(f)
    w.writerow(['Tract','Tract legal','Tract gross ac','Owner','Net mineral acres',
                'Decimal interest in tract','Base OGL','Silver Oak top OGL','Royalty','Note'])
    for n in range(1,11):
        legal=TRACT_LEGAL[n]; gross=TRACT_GROSS[n]
        rows=list(D.OWNERS.get(n,[])); ident=sum(na for _,na,*_ in rows); rem=round(gross-ident,2)
        if rem>0.05 and n in D.OPEN_NOTES: rows=rows+[(D.OPEN_NOTES[n],rem,'HBP base','—','3/16')]
        for name,na,base,top,roy in rows:
            is_open = name.lower().startswith('open') or 'retained balance' in name.lower()
            dec = '' if gross in (0,None) else round(na/gross,8)
            w.writerow([n, legal, f'{gross:.2f}', name, f'{na:.4f}', dec, base, top, roy,
                        'OPEN / unreconciled balance' if is_open else ''])
print("wrote exports/ownership_by_tract.csv")

# ---------- 2) ogl_register.csv (full 69-lease register, verbatim from workbook) ----------
wb=openpyxl.load_workbook('source2/NHE_7-3-26.xlsx', data_only=True)
ws=wb['OGL']
def dt(v): return v.strftime('%Y-%m-%d') if isinstance(v,datetime.datetime) else (v if v is not None else '')
def trf(v):
    if isinstance(v,datetime.datetime): return str((v-datetime.datetime(1900,1,1)).days+1)
    return v if v is not None else ''
def royf(v):
    try: return {0.1875:'3/16',0.25:'1/4',0.125:'1/8'}.get(round(float(v),4), f'{float(v):.6f}')
    except: return v if v is not None else ''
with open('exports/ogl_register.csv','w',newline='') as f:
    w=csv.writer(f)
    w.writerow(['OGL#','Status','Instrument','Bk/Pg','Execution','Filing','Lessor','Lessee',
                'Legal description','Term','Gross ac','Tracts','Expiration','Net mineral ac','Royalty',
                'Depth clause','Option to extend','Top-lease clause'])
    for r in range(2, ws.max_row+1):
        num=ws.cell(r,1).value
        if num is None: continue
        g=lambda c: ws.cell(r,c).value
        w.writerow([num, g(2), g(4), g(5), dt(g(6)), dt(g(7)), g(8), g(9), g(10), g(12),
                    g(13), trf(g(14)), dt(g(15)), g(16), royf(g(17)),
                    g(22), g(23), g(25)])
print("wrote exports/ogl_register.csv")

# ---------- 3) open_items_pull_list.csv (66 source-in + 39 fraction) ----------
def parse_instr(s):
    m=re.search(r'(\d{4}-\d{6})', s);  inst=m.group(1) if m else ''
    b=re.search(r'Bk\s*([D0-9]+/\d+)', s); bp=b.group(1) if b else ''
    typ=s.split()[0] if s else ''
    return inst,bp,typ
with open('exports/open_items_pull_list.csv','w',newline='') as f:
    w=csv.writer(f)
    w.writerow(['Priority','Category','County','Party / subject','Doc type','Instrument #','Bk/Pg',
                'What to pull / read','okcountyrecords query hint'])
    for party,instr,why in D.GAPS:
        inst,bp,typ=parse_instr(instr)
        w.writerow(['A','Source-in gap','Roger Mills',party,typ,inst,bp,
                    f'Grantor vesting instrument ({why})',
                    f'county=Roger+Mills&number={inst}' if inst else ''])
    for tract, items in D.FRACTION_GAPS.items():
        for inst in re.findall(r'\d{4}-\d{6}', items):
            w.writerow(['B','Conveyed-fraction gap','Roger Mills',tract,'(see instrument)',inst,'',
                        'Read granting language for conveyed fraction / net acres',
                        f'county=Roger+Mills&number={inst}'])
    # regulatory pulls
    for pr,subj,doc,hint in [
        ('A','Alexander 1-31 (API 35-129-22925)','OTC gross production by PUN','HBP proof — continuous production'),
        ('A','Alexander 1-31 (API 35-129-22925)','OCC Form 1002A','spud / perforations / TD-TVD'),
        ('A','Alexander 1-31 (API 35-129-22925)','OCC Form 1073','operator of record (Martin\'s Resources vs Empire)'),
        ('B','Section 31-12N-24W','OCC Case Processing','spacing / pooling orders by legal'),
        ('B','Martin\'s Empire LLC / Martin\'s Resources LLC','OK SOS + OCC operator list','confirm entity + operator number'),
    ]:
        w.writerow([pr,'Regulatory','—',subj,doc,'','',hint,''])
print("wrote exports/open_items_pull_list.csv")

# ---------- 4) section_net_acre_summary.csv ----------
with open('exports/section_net_acre_summary.csv','w',newline='') as f:
    w=csv.writer(f)
    w.writerow(['Tract','Legal','Gross ac','Identified NMA','Open/unreconciled NMA','Status'])
    tot_g=tot_id=tot_open=0.0
    for n in range(1,11):
        g=TRACT_GROSS[n]
        idn=sum(na for nm,na,*_ in D.OWNERS.get(n,[]))
        opn=round(g-idn,2) if (round(g-idn,2)>0.05 and n in D.OPEN_NOTES) else 0.0
        st={'reconciled':'Reconciled','partial':'Open balance','estate':'Estate/heirship','succession':'Succession verify'}[[t for t in D.TRACTS if t['n']==n][0]['status']]
        w.writerow([n, TRACT_LEGAL[n], f'{g:.2f}', f'{idn:.2f}', f'{opn:.2f}', st])
        tot_g+=g; tot_id+=idn; tot_open+=opn
    w.writerow(['TOTAL','Section 31','%.2f'%tot_g,'%.2f'%tot_id,'%.2f'%tot_open,''])
print("wrote exports/section_net_acre_summary.csv")
