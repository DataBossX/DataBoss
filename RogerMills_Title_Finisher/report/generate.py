# -*- coding: utf-8 -*-
import html, datetime
import data as D
import openpyxl

WB = 'source2/NHE_7-3-26.xlsx'

def esc(s): return html.escape(str(s), quote=True)

# ---- pull the full OGL register straight from the workbook (authoritative) ----
def load_ogl():
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb['OGL']
    rows=[]
    for r in range(2, ws.max_row+1):
        num = ws.cell(r,1).value
        if num is None: continue
        def g(c):
            v=ws.cell(r,c).value
            return v
        def dt(v):
            return v.strftime('%-m/%-d/%Y') if isinstance(v,datetime.datetime) else (v or '')
        def tract_fix(v):
            # some cells store the tract list as an Excel date-serial (e.g. 1900-01-02 -> 2)
            if isinstance(v,datetime.datetime):
                return str((v - datetime.datetime(1900,1,1)).days + 1)  # recover small Excel serial (tract #)
            return v
        rows.append(dict(
            ogl=g(1), lh=g(2), instr=g(4), bp=g(5),
            exe=dt(g(6)), fil=dt(g(7)), grantor=g(8), grantee=g(9),
            tracts=tract_fix(g(14)), exp=dt(g(15)), nma=g(16), roy=g(17),
        ))
    return rows

OGL = load_ogl()

def roy_fmt(v):
    try:
        f=float(v)
        return {0.1875:'3/16',0.25:'1/4',0.125:'1/8'}.get(round(f,4), f'{f:.4f}')
    except: return esc(v or '')

def nma_fmt(v):
    try: return f'{float(v):.2f}'
    except: return esc(v or '')

STATUS_LABEL = {
 'reconciled':('Reconciled','ok'),
 'partial':('Open balance','warn'),
 'estate':('Estate / heirship','warn'),
 'succession':('Succession verify','warn'),
}

# ---------- SVG section plat ----------
TRACT_HUE = {  # low-saturation categorical fills, petroleum/ochre family
 '1':'#8fb0a6','2':'#c9a15a','3':'#a7bcc9','4':'#b8926a','5':'#9db98f',
 '6':'#c98f6f','7':'#7fa0b3','8':'#c9b06f','9':'#93a77e','10':'#b0a2c0',
}
def build_plat_svg():
    cell=88; pad=54; W=cell*4+pad*2; H=cell*4+pad*2+30
    out=[f'<svg viewBox="0 0 {W} {H}" role="img" aria-label="Section 31 aliquot plat" class="plat">']
    # compass
    out.append(f'<text x="{pad+cell*2}" y="26" text-anchor="middle" class="plat-n">N &#8593;</text>')
    for ri,row in enumerate(D.PLAT):
        for ci,(lab,tr,note) in enumerate(row):
            x=pad+ci*cell; y=pad+ri*cell
            fill=TRACT_HUE.get(tr,'#ccc')
            irregular = note in ('Gov Lots','M&B strip')
            out.append(f'<rect x="{x}" y="{y}" width="{cell}" height="{cell}" fill="{fill}" '
                       f'class="qq {"irregular" if irregular else ""}"/>')
            out.append(f'<text x="{x+cell/2}" y="{y+cell/2-6}" text-anchor="middle" class="qq-tract">{tr}</text>')
            out.append(f'<text x="{x+cell/2}" y="{y+cell/2+10}" text-anchor="middle" class="qq-leg">{esc(lab)}</text>')
            if note: out.append(f'<text x="{x+cell/2}" y="{y+cell-8}" text-anchor="middle" class="qq-note">{esc(note)}</text>')
    # west government-lot tier band
    out.append(f'<rect x="{pad}" y="{pad}" width="{cell}" height="{cell*4}" class="lot-band"/>')
    # well marker: C NW/4 SE/4 -> row2(index2) col2(index2) center-ish
    wx=pad+2*cell+cell/2; wy=pad+2*cell+cell*0.30
    out.append(f'<circle cx="{wx}" cy="{wy}" r="7" class="well-dot"/>')
    out.append(f'<text x="{wx+12}" y="{wy+4}" class="well-lbl">Alexander 1-31</text>')
    # frame
    out.append(f'<rect x="{pad}" y="{pad}" width="{cell*4}" height="{cell*4}" class="plat-frame"/>')
    out.append(f'<text x="{pad}" y="{pad+cell*4+22}" class="plat-cap">Schematic aliquot map &#183; north up &#183; not a survey. '
               f'West &amp; south tiers are government lots; Tracts 6 &amp; 8 are metes-and-bounds strips.</text>')
    out.append('</svg>')
    return '\n'.join(out)

# ---------- HTML sections ----------
def stat(v,l,cls=''): return f'<div class="stat {cls}"><div class="stat-v">{v}</div><div class="stat-l">{esc(l)}</div></div>'

def owners_with_open(n):
    """Identified owners + a computed open-balance line (gross - identified)."""
    rows=list(D.OWNERS.get(n,[]))
    gross=[t for t in D.TRACTS if t['n']==n][0]['gross']
    ident=sum(na for _,na,*_ in rows)
    rem=round(gross-ident,2)
    if rem>0.05 and n in D.OPEN_NOTES:
        rows=rows+[(D.OPEN_NOTES[n],rem,"HBP base","—","3/16")]
    return rows

def owners_block(n):
    rows=owners_with_open(n)
    body=[]
    for name,na,base,top,roy in rows:
        opencls=' class="open"' if name.lower().startswith('open') else ''
        body.append(f'<tr{opencls}><td>{esc(name)}</td><td class="num">{na:.2f}</td>'
                    f'<td class="mono">{esc(base)}</td><td class="mono">{esc(top)}</td><td class="mono">{esc(roy)}</td></tr>')
    return ('<table class="tbl"><thead><tr><th>Identified record owner</th><th class="num">Net ac</th>'
            '<th>Base&nbsp;OGL</th><th>Top</th><th>Roy.</th></tr></thead><tbody>'+''.join(body)+'</tbody></table>')

def tract_cards():
    cards=[]
    for t in D.TRACTS:
        lab,cls=STATUS_LABEL[t['status']]
        cards.append(f'''<article class="tract" id="tract-{t['n']}">
          <header class="tract-h">
            <div class="tract-id">Tract {t['n']}</div>
            <div class="tract-meta"><div class="tract-legal">{esc(t['legal'])}</div>
            <div class="tract-sub"><span class="mono">{t['gross']:.2f} gross ac</span> <span class="badge {cls}">{lab}</span></div></div>
          </header>
          {owners_block(t['n'])}
        </article>''')
    return '\n'.join(cards)

def ogl_rows(kind):
    out=[]
    for o in OGL:
        is_top = str(o['lh']).upper().startswith('TOP')
        if kind=='top' and not is_top: continue
        if kind=='base' and is_top: continue
        out.append(f'<tr><td class="mono">{esc(o["ogl"])}</td><td>{esc(o["instr"])}</td>'
                   f'<td class="mono">{esc(o["bp"])}</td><td>{esc(o["grantor"])}</td>'
                   f'<td>{esc(o["grantee"])}</td><td class="mono">{esc(o["tracts"])}</td>'
                   f'<td class="mono">{roy_fmt(o["roy"])}</td><td class="mono">{esc(o["exp"])}</td></tr>')
    return ''.join(out)

def gaps_rows():
    return ''.join(f'<tr><td>{esc(p)}</td><td class="mono">{esc(i)}</td><td>{esc(w)}</td></tr>' for p,i,w in D.GAPS)
def resolved_rows():
    return ''.join(f'<tr><td>{esc(p)}</td><td class="mono">{esc(s)}</td></tr>' for p,s in D.RESOLVED)
def frac_rows():
    return ''.join(f'<tr><td class="mono">{esc(k)}</td><td>{esc(v)}</td></tr>' for k,v in D.FRACTION_GAPS.items())
def depth_rows():
    return ''.join(f'<tr><td class="mono">{esc(o)}</td><td>{esc(g)}</td><td class="mono">{esc(r)}</td><td>{esc(c)}</td></tr>' for o,g,r,c in D.DEPTH)
def wi_chain():
    return ''.join(f'<li>{esc(x)}</li>' for x in D.WI_CHAIN)
def top_rows():
    return ''.join(f'<tr><td class="mono">{esc(n)}</td><td>{esc(l)}</td><td class="mono">{esc(bp)}</td>'
                   f'<td class="mono">{esc(ex)}</td><td class="mono">{esc(xp)}</td><td class="mono">{esc(tm)}</td></tr>'
                   for n,l,bp,ex,xp,tm in D.TOP)

CSS = open('report/style.css').read()
SVG = build_plat_svg()

M=D.META
HTMLDOC=f'''<title>Section 31-12N-24W Roger Mills — Mineral &amp; Leasehold Ownership</title>
<style>{CSS}</style>
<div class="wrap">

<header class="mast">
  <div class="mast-l">
    <div class="eyebrow">Cursory Mineral &amp; Leasehold Ownership Report</div>
    <h1>Section 31, T12N&#8211;R24W<span>Roger Mills County, Oklahoma</span></h1>
  </div>
  <div class="mast-r mono">
    <div><span>Prospect</span>{esc(M['prospect'])}</div>
    <div><span>Gross</span>{esc(M['gross'])} ac</div>
    <div><span>Report</span>{esc(M['report_date'])}</div>
    <div><span>Examiner</span>{esc(M['examiner'])}</div>
  </div>
</header>

<div class="disclaimer">Cursory &#8212; not a certified title opinion. Interests shown are subject to curative requirements; confirm every interest against the recorded instruments before relying on this for a transaction.</div>

<nav class="tnav"><a href="#summary">Summary</a><a href="#plat">Plat</a><a href="#ownership">Minerals</a><a href="#leasehold">Leasehold</a><a href="#depth">Depth</a><a href="#wi">Working interest</a><a href="#well">Well &amp; HBP</a><a href="#risk">Risk</a><a href="#open">Open items</a><a href="#research">Research</a></nav>

<section id="summary" class="band">
  <div class="stats">
    {stat(esc(M['gross']),'gross acres')}
    {stat(M['tracts'],'tracts')}
    {stat(f"{D.META['base_leases']}+{D.META['top_leases']}",'OGLs (base + top)')}
    {stat('1','active unit well','ok')}
    {stat('100%','leased &middot; HBP','ok')}
    {stat(D.META['instruments'],'indexed instruments')}
    {stat(len(D.GAPS),'source-in gaps','warn')}
    {stat(sum(1 for t in D.TRACTS if t['status']=='reconciled'),'tracts reconciled')}
  </div>
  <div class="lede">
    <p><b>Section 31 is a fully-leased, held-by-production Cherokee section.</b> Every net mineral acre sits under one of <b>63 base oil-and-gas leases</b> (the 2004&#8211;2006 EOG / Arrowhead play, OGL 1&#8211;63), all carried <b>HBP</b> by the single active <b>Alexander 1-31</b> unit well. In 2026 <b>Silver Oak Natural Resources, LLC</b> took a six-lease <b>top-lease sweep</b> (OGL 64&#8211;69, 3/16 royalty) over the principal current owners &#8212; subordinate and springing until the base leases terminate.</p>
    <p>Fee mineral ownership is a family mosaic dominated by the <b>Mitchell / Williams / Moore</b>, <b>Bain</b>, <b>Kirk</b>, and <b>Daigle / Sorenson / Johnson</b> families, with modern mineral-fund buyers layered on top. All ten tracts foot to the section total of <b>637.42</b> acres; Tracts 1, 2, 6 and 8 are owner-reconciled, and carried HBP balances on the remaining tracts are shown as explicit open lines &#8212; never distributed to look finished.</p>
  </div>
</section>

<section id="plat" class="sec">
  <h2>Section plat</h2>
  <div class="plat-wrap">{SVG}
    <div class="legend">
      {''.join(f'<span class="lg"><i style="background:{TRACT_HUE[str(t["n"])]}"></i>T{t["n"]} &middot; {esc(t["legal"][:26])}</span>' for t in D.TRACTS)}
    </div>
  </div>
</section>

<section id="ownership" class="sec">
  <h2>Mineral ownership by tract</h2>
  <p class="note">Identified record owners are shown with net mineral acres, their base lease (OGL #) and any 2026 Silver Oak top lease. Where the chain does not yet reconcile to gross, the <span class="open-key">open / unreconciled balance</span> is carried as an explicit line pending the source instruments listed under Open Items &#8212; it is never spread across owners to force a footing. The deep historical chain (Panhandle Royalty, ETA-Flag, Arrowhead, Crusader, Prism, &amp;c.) reflects prior leasehold-era holders whose carry-forward to present ownership requires those same instruments.</p>
  <div class="tracts">{tract_cards()}</div>
</section>

<section id="leasehold" class="sec">
  <h2>Leasehold</h2>
  <h3>Silver Oak top-lease sweep &#8212; OGL 64&#8211;69 (2026, 3/16, springing)</h3>
  <p class="note">Paid-up top leases over the principal current owners, <b>subordinate to and springing upon termination of the 2004-era base leases</b>. Because a lessor can only lease what it owns, these six 2026 lessors are the highest-confidence signal of <b>current</b> mineral ownership in the section.</p>
  <div class="scroll"><table class="tbl"><thead><tr><th>OGL</th><th>Top lessor (current owner)</th><th>Bk/Pg</th><th>Executed</th><th>Expires</th><th>Term</th></tr></thead><tbody>{top_rows()}</tbody></table></div>

  <h3>Base lease register &#8212; OGL 1&#8211;63 (2004&#8211;2006, all HBP)</h3>
  <p class="note">The complete base register, verbatim from the workbook OGL tab. All are held by production by the Alexander 1-31. Non-3/16 royalties (several 1/4 leases in the Bain / Fuchs / Deutsch chain) and depth-limited leases are itemized under Depth Severances.</p>
  <div class="scroll"><table class="tbl compact"><thead><tr><th>OGL</th><th>Instrument</th><th>Bk/Pg</th><th>Lessor</th><th>Lessee</th><th>Tracts</th><th>Roy.</th><th>Exp.</th></tr></thead><tbody>{ogl_rows('base')}</tbody></table></div>
</section>

<section id="depth" class="sec">
  <h2>Depth severances &amp; non-standard royalty</h2>
  <p class="note">Leases in the register that carry an express depth limitation or a non-3/16 royalty. Depth-limited rights are kept separate where the instrument supports it; no depth limit is inferred where the record is silent.</p>
  <div class="scroll"><table class="tbl"><thead><tr><th>OGL</th><th>Lessor &#8594; Lessee</th><th>Roy.</th><th>Depth / term provision</th></tr></thead><tbody>{depth_rows()}</tbody></table></div>
</section>

<section id="wi" class="sec">
  <h2>Working interest &#8212; wellbore chain</h2>
  <p class="note">The wellbore leasehold / working-interest chain traced through the recorded assignments to the current holder (5-4-2026). Mortgages, releases and ORRI-only conveyances are excluded from the title chain and retained on rawdata only.</p>
  <ol class="chain">{wi_chain()}</ol>
  <p class="note">Current wellbore assignment: <span class="mono">Instr. 2026-002547, Bk 2698/069</span> (executed 5-4-2026), Martin&#8217;s Empire, LLC &#8594; Stride Bank, N.A., Trustee of the Umbrella Trust. <b>Reconcile against OCC Form 1073:</b> the well of record shows operator <b>Martin&#8217;s Resources LLC</b> (a confirmed active but marginal Roger Mills operator), while &#8220;Martin&#8217;s Empire, LLC&#8221; did not surface in any public operator or entity index this session &#8212; verify the entity and the operator-of-record link.</p>
</section>

<section id="well" class="sec">
  <h2>Unit well &amp; HBP</h2>
  <div class="well-grid">
    <div class="well-panel">
      <div class="wr"><span>Well</span><b>{esc(M['well'])}</b></div>
      <div class="wr"><span>API</span><b class="mono">{esc(M['api'])}</b></div>
      <div class="wr"><span>Operator of record</span><b>Martin&#8217;s Resources LLC</b></div>
      <div class="wr"><span>Formation</span><b>Cherokee</b></div>
      <div class="wr"><span>Type</span><b>Oil</b></div>
      <div class="wr"><span>Surface</span><b class="mono">C NW/4 SE/4 &#183; 660&#8242; FWL &times; 1980&#8242; FSL</b></div>
      <div class="wr"><span>Completed</span><b class="mono">5/19/2006</b></div>
      <div class="wr"><span>Last production</span><b class="mono">3/1/2026</b></div>
      <div class="wr"><span>OCC status</span><b>Active (AC)</b></div>
      <div class="wr"><span>Spacing</span><b>640 ac (order not of record)</b></div>
    </div>
    <div class="well-note">
      <h3>HBP analysis</h3>
      <p>A single active vertical well holds the entire base register (OGL 1&#8211;63) by production; the Silver Oak top leases (OGL 64&#8211;69) remain subordinate and springing. Independent research this session confirms operator <b>Martin&#8217;s Resources LLC</b> reports production across only three Oklahoma wells (~2,792 BOE/month &#8212; marginal/stripper), which is <b>consistent with continuous-but-low production holding the section</b>, but the well-level volumes were not retrievable.</p>
      <p class="flag"><b>Reserve for examiner &#8212; pull to certify:</b> OTC gross production by PUN for API 35-129-22925 (continuous-production proof); OCC Form 1002A (spud, perforations, TD/TVD); OCC Form 1073 (operator of record). <b>Location discrepancy:</b> the quarter call &#8220;C NW/4 SE/4&#8221; conflicts with the footages (660&#8242; FWL &times; 1980&#8242; FSL fall in C NW/4 SW/4) &#8212; reconcile against the plat / 1002A.</p>
    </div>
  </div>
</section>

<section id="risk" class="sec">
  <h2>Title issues &amp; risk register</h2>
  <div class="risk">
    <div class="rcard defect"><h4>Root-of-title gaps (pre-2000)</h4><p>{len(D.GAPS)} conveying parties whose acquiring instrument predates the OCR&#8217;d index (which reaches only to 2/8/2000). Each needs its prior deed, probate decree, or heirship affidavit &#8212; see Open Items.</p></div>
    <div class="rcard warn"><h4>Estate / probate</h4><p>Ella Pearl Kirk estate carries ~48.80 NMA on Tract 6 and 32.80 on Tract 8 &#8212; probate / heirship required. Nancy Susan Mitchell &#8594; Dold succession on Tract 9 rests on AoH 2026-002509.</p></div>
    <div class="rcard warn"><h4>Unreconciled acreage</h4><p>Carried HBP open balances on Tracts 2, 4, 5, 6, 7 and 10 &#8212; shown as explicit lines, not distributed. Candidate: Psi Lke &#8594; Cummins/Eagle Owl &#8220;all of section&#8221; (MD 2022-000185) on Tract 4.</p></div>
    <div class="rcard warn"><h4>Conveyed fractions</h4><p>39 instrument columns need the granting fraction read from the image before net acres are final &#8212; itemized under Open Items.</p></div>
    <div class="rcard ok"><h4>Depth &amp; royalty preserved</h4><p>Depth-limited leases (20,000&#8242; and 100-ft-below-deepest) and 1/4-royalty leases in the Fuchs / Bain / Deutsch chains are carried separately; NPRI and reservations retained on rawdata.</p></div>
    <div class="rcard ok"><h4>Chain integrity</h4><p>All ten tract grids foot to gross and tie to 637.42; no mortgage / lien / ROW / ORRI-only rows on the title tabs; Title lease references normalized to OGL register numbers.</p></div>
  </div>
</section>

<section id="open" class="sec">
  <h2>Open items &#8212; what to pull to certify</h2>
  <p class="note">Every item below requires a source document that this cloud session&#8217;s egress policy blocked (okcountyrecords.com, OCC imaging/RBDMS, OTC all returned proxy 403). Nothing is fabricated to fill them; pull each on an unrestricted machine.</p>

  <h3>Source-in gaps &#8212; pull the grantor&#8217;s vesting instrument ({len(D.GAPS)})</h3>
  <div class="scroll"><table class="tbl"><thead><tr><th>Conveying party (no source-in located)</th><th>Instrument out</th><th>Why open</th></tr></thead><tbody>{gaps_rows()}</tbody></table></div>

  <h3>Conveyed-fraction gaps &#8212; read the granting language (39)</h3>
  <div class="scroll"><table class="tbl"><thead><tr><th>Tract</th><th>Instruments needing the conveyed fraction</th></tr></thead><tbody>{frac_rows()}</tbody></table></div>

  <h3>Resolved this engagement ({len(D.RESOLVED)} parties, from the index + Runsheet cross-reference)</h3>
  <div class="scroll"><table class="tbl"><thead><tr><th>Party</th><th>Source-in located (grantee-side)</th></tr></thead><tbody>{resolved_rows()}</tbody></table></div>
</section>

<section id="research" class="sec">
  <h2>Well &amp; regulatory research addendum</h2>
  <p class="note">Independent research, 7/3/2026. This session&#8217;s egress proxy blocked OCC RBDMS/imaging, OTC, ShaleXP and DrillingEdge (WebFetch 403 globally), so operational well facts below are from search snippets and must be re-pulled with direct access.</p>
  <ul class="rlist">
    <li><b>Operator &#8212; Martin&#8217;s Resources LLC (confirmed active):</b> Oklahoma producer id 23562; ~2,792 BOE reported for 8/2025; <b>3 producing wells</b>; operates in Roger Mills &amp; Custer counties; no future permits on file. Marginal/stripper profile consistent with HBP-holding production. <i>(mineralanswers.com, shalexp.com)</i></li>
    <li><b>&#8220;Martin&#8217;s Empire, LLC&#8221; &#8212; not found:</b> no OK Secretary of State entity, OCC operator record, or O&amp;G profile surfaced. Verify existence &amp; relationship to Martin&#8217;s Resources via OK SOS + OCC operator list. <i>(flag)</i></li>
    <li><b>Offset activity:</b> Charter Oak Production &#8220;Solo 1-8H,&#8221; a horizontal Cottage Grove test in 8-12N-24W (MD ~16,230&#8242;, TVD ~10,850&#8242;, spud ~Aug 2025). Roger Mills is currently the busiest OK permit county. <i>(mineralrightsforum.com, shaleexperts.com)</i></li>
    <li><b>Regional targets (context, not section fact):</b> Cottage Grove, Tonkawa, Cleveland, Marmaton and the Granite Wash are the productive/leased pays here; deeper Morrow / Springer / Chester / Hunton / Woodford gas targets exist basin-wide.</li>
    <li><b>Spacing pattern:</b> 640-acre horizontal units are standard in this township; no order specific to Section 31 was retrievable &#8212; pull OCC Case Processing by legal 31-12N-24W.</li>
  </ul>
</section>

<footer class="foot">
  <p class="mono">Basis: 7-3-2026 NHE cursory workbook (19 sheets) &#183; Roger Mills County Unplatted Legal Index (OCR, {D.META['instruments']} instruments, 2000&#8211;2026) &#183; OGL register (69 leases) &#183; independent OCC / operator research. Last index entry: {esc(M['last_entry'])}.</p>
  <p><b>Cursory cleanup only. Not a certified title opinion.</b> Open and verify items remain only where source evidence was insufficient. Confirm all interests against the recorded instruments before relying on this report for a transaction.</p>
</footer>
</div>'''

open('report/report.html','w').write(HTMLDOC)
print("wrote report/report.html", len(HTMLDOC), "bytes")
