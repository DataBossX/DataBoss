"""
Build the best merged Diversified Cursory Title Report for Section 32-11N-25W.
Strategy:
  - Base: NHE report (legal narrative, Overview, Title, OGL, Runsheet, Well 1)
  - From Report1: Tract 1-5 matrices, WI 1/2, PLAT (template-format chain grids)
  - Enhancements from evidence register
"""
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(r"D:\Desktop\Horizon\32-11N-25W, Beckham County\FINAL")
NHE = BASE / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_7-10-26 - NHE.xlsx"
R1 = BASE / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_2026-07-10.xlsx"
OUT = BASE / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_MERGED_BEST_2026-07-10.xlsx"

# Sheets to copy entirely from Report1 (values + styles where possible)
COPY_FROM_R1 = ["Tract 1", "Tract 2", "Tract 3", "Tract 4", "Tract 5", "WI 1", "WI 2", "PLAT"]

# Sheets to keep from NHE
KEEP_FROM_NHE = ["Overview", "Title ", "OGL", "Runsheet", "Well 1"]


def copy_sheet_values_and_styles(src_ws, dst_ws):
    """Copy all cell values and styles from source to destination worksheet."""
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws[cell.coordinate]
            dst_cell.value = cell.value
            if cell.has_style:
                dst_cell.font = copy(cell.font)
                dst_cell.border = copy(cell.border)
                dst_cell.fill = copy(cell.fill)
                dst_cell.number_format = cell.number_format
                dst_cell.protection = copy(cell.protection)
                dst_cell.alignment = copy(cell.alignment)
    # Copy column dimensions
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
    # Copy row dimensions
    for row_num, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dim.height
    # Copy merged cells
    for merged in list(dst_ws.merged_cells.ranges):
        dst_ws.unmerge_cells(str(merged))
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines


def clear_sheet(ws):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    if ws.max_row and ws.max_column:
        for row in ws.iter_rows():
            for cell in row:
                if type(cell).__name__ != "MergedCell":
                    cell.value = None


def enhance_overview(ws):
    """Align Overview with template field layout while preserving NHE legal content."""
    updates = {
        "Q3": "32-11N-25W / Diversified",
        "H5": 32,
        "R5": "11N",
        "AC5": "25W",
        "H7": "Beckham",
        "B9": "Containing 640.00 acres, more or less (nominal section acreage; not a title decimal)",
        "AB15": "Tract 1 — NE/4 / Crook wellbore",
        "U22": "Tract 2 — E/2 SE/4 / Pearl",
        "M26": "Tract 3 — N/2 below top Hunton",
        "E26": "Tract 4 — Deep leasehold refs",
        "X34": "Tract 5 — All Sec. 32 / Diversified branch",
        "S48": datetime(2026, 7, 10),
        "AC48": "Bk 2490/P471 (WD) — last indexed entry",
        "S50": datetime(2026, 7, 1),
        "B53": (
            "CONTROLLING CONCLUSION: Diversified-affiliated entities acquired a material interest "
            "indexed against Section 32, but missing post-1988 operative instruments and schedules "
            "prevent determination of exact estate, current vested entity, WI, NRI, leasehold/ORRI "
            "decimal, mineral/royalty interest, or net acres. No unsupported decimal is stated."
        ),
        "B54": (
            "SCOPE/CUTOFFS: 4,893 direct images end at Bk 1014/P75, recorded 04/26/1988. "
            "Section index certified through 07/01/2026; last entry Bk 2490/P471. "
            "Last principal Diversified-chain item reviewed: Bk 2476/P121 (substantive effect OPEN)."
        ),
        "B55": (
            "PRIORITY COPY PULL: 2340/403, 2340/490, 2371/470, /495, /514, /533, 2389/500, /581, "
            "2395/415, 2400/551, 2451/4 and 2476/121 with every exhibit/schedule; "
            "then bridge post-04/26/1988 predecessor chain."
        ),
        "F56": "New Horizon Energy (NHE) — merged best cursory / Cursor assist",
    }
    for coord, val in updates.items():
        ws[coord] = val


def enhance_well1(ws):
    """Add status column context from evidence register where missing."""
    status_map = {
        "35-009-00075": "PA — Plugged/abandoned",
        "35-009-20312": "AC — Crook wellbore; Granite Wash Lower 13,348-13,595 ft",
        "35-009-20605": "PA — Historical Pearl well",
        "35-009-20738": "PA — Plugged/abandoned",
        "35-009-21072": "AC — Active; formation OPEN",
        "35-009-21085": "AC — Active; formation OPEN",
        "35-009-21166": "PA — Plugged/abandoned",
        "35-009-21236": "AC — Active; formation OPEN",
        "35-009-21245": "EX — Expired/nonactive",
        "35-009-21258": "AC — Active; formation OPEN",
        "35-009-21391": "AC — Granite Wash Low; spud 2004-07-15",
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        api = str(row[1].value or "").strip()
        if api in status_map and row[3].value in (None, "", "OPEN"):
            row[3].value = status_map[api].split(" — ")[0]
        if api in status_map:
            comments_cell = row[7] if len(row) > 7 else None
            if comments_cell and (comments_cell.value in (None, "", "OPEN")):
                comments_cell.value = status_map[api]


def add_assumptions_sheet(wb):
    """Add Assumptions sheet documenting reasonable assumptions per user request."""
    if "Assumptions" in wb.sheetnames:
        del wb["Assumptions"]
    ws = wb.create_sheet("Assumptions", 0)
    rows = [
        ("ASSUMPTIONS & LIMITATIONS — Section 32-11N-25W Diversified Cursory", ""),
        ("", ""),
        ("#", "Assumption / Limitation"),
        ("A1", "Full-section Q-Q index ticks do NOT prove 100% leasehold, all formations, or fee minerals."),
        ("A2", "Historical Leede/Crook/Pearl interests are wellbore/depth-limited; not carried forward as current WI without post-1988 bridge."),
        ("A3", "Bk 2395/P967 is financing termination only; Diversified ABS conveyance is Bk 2395/P415."),
        ("A4", "SE/NW/SW sovereign inception based on receiver receipts (Images 0004-0006); actual patents not reviewed — obtain for fee chain."),
        ("A5", "No Diversified decimal, NRI, or net acres stated where operative schedules absent (C4)."),
        ("A6", "Section-level well activity does not prove any particular lease is HBP (C8)."),
        ("A7", "FourPoint/Unbridled/MNR branch (2415/328 et al.) may overlap Tapstone/KL CHK branch — reconcile before final opinion."),
        ("A8", "Countywide DP instruments (2434/751, 2464/200, 2480/824, 2480/903) screened but not confirmed Section 32 applicability."),
        ("A9", "Surface ownership per assessor roll is context only; does not establish mineral title."),
        ("A10", "Report date 2026-07-10; continuation from 2026-07-01 required for final effective date."),
        ("", ""),
        ("METHODOLOGY", ""),
        ("", "Evidence tiers: (1) Direct image, (2) Section index, (3) Public metadata, (4) Regulatory context only."),
        ("", "Merged from NHE legal narrative + Report1 tract/WI matrices on 2026-07-10."),
    ]
    for i, (a, b) in enumerate(rows, 1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 100


def main():
    shutil.copy2(NHE, OUT)
    wb_out = openpyxl.load_workbook(OUT)
    wb_r1 = openpyxl.load_workbook(R1, data_only=False)

    for sheet_name in COPY_FROM_R1:
        if sheet_name not in wb_r1.sheetnames:
            continue
        src = wb_r1[sheet_name]
        dst = wb_out[sheet_name]
        clear_sheet(dst)
        copy_sheet_values_and_styles(src, dst)
        print(f"Copied {sheet_name} from Report1")

    enhance_overview(wb_out["Overview"])
    enhance_well1(wb_out["Well 1"])
    add_assumptions_sheet(wb_out)

    wb_out.save(OUT)
    print(f"Saved merged report: {OUT}")


if __name__ == "__main__":
    main()
