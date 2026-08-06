import openpyxl
import re
from pathlib import Path

base = Path(r"D:\Desktop\Horizon\32-11N-25W, Beckham County\FINAL")
files = {
    "template": base / "Template.xlsx",
    "report1": base / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_2026-07-10.xlsx",
    "report2": base / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_7-10-26 - NHE.xlsx",
}


def cell_val(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def get_sheet_data(wb, sheet):
    ws = wb[sheet]
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                data[cell.coordinate] = cell_val(cell.value)
    return data


def sort_key(coord):
    col = re.sub(r"\d+", "", coord)
    row = int(re.sub(r"[A-Z]+", "", coord))
    return (openpyxl.utils.column_index_from_string(col), row)


def compare_sheets(d1, d2):
    all_keys = set(d1.keys()) | set(d2.keys())
    only1, only2, diff = [], [], []
    for k in sorted(all_keys, key=sort_key):
        v1, v2 = d1.get(k, ""), d2.get(k, "")
        if v1 and not v2:
            only1.append((k, v1))
        elif v2 and not v1:
            only2.append((k, v2))
        elif v1 != v2:
            diff.append((k, v1, v2))
    return only1, only2, diff


def count_open_items(ws):
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "OPEN" in str(cell.value).upper():
                n += 1
    return n


def count_placeholder(ws):
    n = 0
    markers = ["[", "TBD", "INSERT", "XXX", "PLACEHOLDER", "TODO", "???"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                s = str(cell.value).upper()
                if any(m in s for m in markers):
                    n += 1
    return n


wbs = {k: openpyxl.load_workbook(p, data_only=True) for k, p in files.items()}

# Report1 vs template
out = base / "comparison_report1_vs_template.txt"
with open(out, "w", encoding="utf-8") as f:
    for sheet in wbs["template"].sheetnames:
        dt = get_sheet_data(wbs["template"], sheet)
        dr = get_sheet_data(wbs["report1"], sheet)
        o_t, o_r, diff = compare_sheets(dt, dr)
        f.write(f"\n{'='*60}\nSHEET: {sheet}\n")
        f.write(f"Template-only: {len(o_t)} | Report-only: {len(o_r)} | Different: {len(diff)}\n")
        if diff:
            f.write("DIFFS:\n")
            for k, v1, v2 in diff[:30]:
                f.write(f"  {k}: T='{v1[:100]}' | R='{v2[:100]}'\n")
            if len(diff) > 30:
                f.write(f"  ... and {len(diff)-30} more\n")

print("Report1 vs Template summary:")
for sheet in wbs["template"].sheetnames:
    dt = get_sheet_data(wbs["template"], sheet)
    dr = get_sheet_data(wbs["report1"], sheet)
    o_t, o_r, diff = compare_sheets(dt, dr)
    print(f"  {sheet}: t_only={len(o_t)}, r_only={len(o_r)}, diff={len(diff)}")

print("\nOPEN ITEM counts:")
for name in ["report1", "report2"]:
    total = 0
    for sheet in wbs["template"].sheetnames:
        c = count_open_items(wbs[name][sheet])
        total += c
        if c:
            print(f"  {name} {sheet}: {c}")
    print(f"  {name} TOTAL OPEN: {total}")

print("\nPlaceholder/TBD counts:")
for name in ["report1", "report2", "template"]:
    total = 0
    for sheet in wbs["template"].sheetnames:
        c = count_placeholder(wbs[name][sheet])
        total += c
    print(f"  {name}: {total}")

# Runsheet row counts with data
print("\nRunsheet populated rows:")
for name in ["template", "report1", "report2"]:
    ws = wbs[name]["Runsheet"]
    rows = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if any(c.value for c in row[:8]):
            rows.add(row[0].row)
    print(f"  {name}: {len(rows)} data rows")

# Title sheet structure
print("\nTitle sheet - tract sections:")
for name in ["template", "report1", "report2"]:
    ws = wbs[name]["Title "]
    tracts = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        v = cell_val(row[0].value) if row[0].value else ""
        if v.startswith("TRACT"):
            tracts.append(v)
    print(f"  {name}: {tracts}")

# Sample tract 1 header area report1
print("\n--- TRACT 1 report1 first 15 rows with data ---")
ws = wbs["report1"]["Tract 1"]
for r in range(1, 30):
    vals = []
    for c in range(1, 10):
        v = ws.cell(r, c).value
        if v:
            vals.append(f"{openpyxl.utils.get_column_letter(c)}={cell_val(v)}")
    if vals:
        print(f"Row{r}: " + " | ".join(vals))

print("\n--- TRACT 1 report2 first 15 rows with data ---")
ws = wbs["report2"]["Tract 1"]
for r in range(1, 30):
    vals = []
    for c in range(1, 10):
        v = ws.cell(r, c).value
        if v:
            vals.append(f"{openpyxl.utils.get_column_letter(c)}={cell_val(v)}")
    if vals:
        print(f"Row{r}: " + " | ".join(vals))
