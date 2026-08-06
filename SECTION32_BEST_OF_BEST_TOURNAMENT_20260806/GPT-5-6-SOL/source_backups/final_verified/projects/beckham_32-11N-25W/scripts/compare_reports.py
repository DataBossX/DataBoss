import openpyxl
import re
from pathlib import Path

base = Path(r"D:\Desktop\Horizon\32-11N-25W, Beckham County\FINAL")
files = {
    "template": base / "Template.xlsx",
    "report1": base / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_2026-07-10.xlsx",
    "report2": base / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_7-10-26 - NHE.xlsx",
}


def cell_val(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def get_sheet_data(wb, sheet):
    ws = wb[sheet]
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                data[cell.coordinate] = cell_val(cell.value)
    return data


def sort_key(coord):
    col = re.sub(r"\d+", "", coord)
    row = int(re.sub(r"[A-Z]+", "", coord))
    return (openpyxl.utils.column_index_from_string(col), row)


def compare_sheets(d1, d2):
    all_keys = set(d1.keys()) | set(d2.keys())
    only1, only2, diff = [], [], []
    for k in sorted(all_keys, key=sort_key):
        v1, v2 = d1.get(k, ""), d2.get(k, "")
        if v1 and not v2:
            only1.append((k, v1))
        elif v2 and not v1:
            only2.append((k, v2))
        elif v1 != v2:
            diff.append((k, v1, v2))
    return only1, only2, diff


def count_filled(ws):
    n = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                n += 1
    return n


wbs = {k: openpyxl.load_workbook(p, data_only=True) for k, p in files.items()}

print("=" * 80)
print("FILLED CELL COUNTS BY SHEET")
print("=" * 80)
for sheet in wbs["template"].sheetnames:
    counts = {name: count_filled(wbs[name][sheet]) for name in wbs}
    print(f"{sheet}: template={counts['template']}, report1={counts['report1']}, report2={counts['report2']}")

print("\n" + "=" * 80)
print("REPORT1 vs REPORT2")
print("=" * 80)
total_r1_only = total_r2_only = total_diff = 0
for sheet in wbs["template"].sheetnames:
    d1 = get_sheet_data(wbs["report1"], sheet)
    d2 = get_sheet_data(wbs["report2"], sheet)
    o1, o2, diff = compare_sheets(d1, d2)
    total_r1_only += len(o1)
    total_r2_only += len(o2)
    total_diff += len(diff)
    print(f"{sheet}: only_r1={len(o1)}, only_r2={len(o2)}, different={len(diff)}")
print(f"TOTAL report1 vs report2: only_r1={total_r1_only}, only_r2={total_r2_only}, diff={total_diff}")

print("\n" + "=" * 80)
print("REPORT2 (NHE) vs TEMPLATE - FULL DETAIL")
print("=" * 80)
out_path = base / "comparison_report2_vs_template.txt"
with open(out_path, "w", encoding="utf-8") as out:
    grand_t_only = grand_r_only = grand_diff = 0
    for sheet in wbs["template"].sheetnames:
        dt = get_sheet_data(wbs["template"], sheet)
        dr = get_sheet_data(wbs["report2"], sheet)
        o_t, o_r, diff = compare_sheets(dt, dr)
        grand_t_only += len(o_t)
        grand_r_only += len(o_r)
        grand_diff += len(diff)
        out.write(f"\n{'='*60}\nSHEET: {sheet}\n{'='*60}\n")
        out.write(f"Template-only: {len(o_t)}\nReport-only: {len(o_r)}\nDifferent: {len(diff)}\n\n")
        if o_t:
            out.write("--- TEMPLATE ONLY (placeholder/instruction text) ---\n")
            for k, v in o_t:
                out.write(f"  {k}: {v}\n")
        if o_r:
            out.write("\n--- REPORT ONLY (new data) ---\n")
            for k, v in o_r:
                out.write(f"  {k}: {v}\n")
        if diff:
            out.write("\n--- DIFFERENT VALUES ---\n")
            for k, v1, v2 in diff:
                out.write(f"  {k}: T='{v1}' | R='{v2}'\n")
        print(f"{sheet}: t_only={len(o_t)}, r_only={len(o_r)}, diff={len(diff)}")
    out.write(f"\nGRAND TOTAL: template_only={grand_t_only}, report_only={grand_r_only}, different={grand_diff}\n")
print(f"\nFull detail written to {out_path}")

# Key sheets content dump
for sheet in ["Overview", "Title ", "Runsheet", "Well 1"]:
    print(f"\n{'#'*60}\nKEY CONTENT: {sheet} (report2)\n{'#'*60}")
    ws = wbs["report2"][sheet]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), max_col=min(ws.max_column, 12)):
        vals = [cell_val(c.value) for c in row if c.value is not None and str(c.value).strip()]
        if vals:
            print(" | ".join(vals[:8]))
