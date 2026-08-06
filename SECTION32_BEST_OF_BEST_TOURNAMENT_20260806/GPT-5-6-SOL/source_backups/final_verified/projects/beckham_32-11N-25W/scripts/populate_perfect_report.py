"""
Populate the perfect Section 32-11N-25W Diversified Cursory Title Report.
Sources: evidence register, draft workbook, NHE merged report, Report1 tract structure.
"""
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

BASE = Path(r"D:\Desktop\Horizon\32-11N-25W, Beckham County\FINAL")
SRC = BASE / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_MERGED_BEST_2026-07-10.xlsx"
OUT = BASE / "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_PERFECT_2026-07-10.xlsx"

# --- COMPLETE RUNSHEET DATA (one dict per instrument, chronological) ---
RUNSHEET = [
    # SOVEREIGN INCEPTION
    dict(seq=1, ogl=None, doc="PAT", bk="Patent Bk 3/469", exec_eff="1907-03-22", rec="1907-06-08",
         grantor="United States", grantee="Cornelius B. Cornelius", legal="NE/4 Section 32 (160 ac nominal)",
         notes="Homestead patent; no mineral reservation apparent on reviewed image.", tier="DIRECT IMAGE P-001 — 98%"),
    dict(seq=2, ogl=None, doc="FR", bk="Image 0004", exec_eff="1906-06-28 receipt", rec="1906-08-08 image stamp",
         grantor="U.S. Land Office", grantee="Velmore Biggs", legal="SE/4 Section 32",
         notes="Earliest supplied land-office evidence; actual patent remains OPEN.", tier="DIRECT IMAGE P-002 — 90%"),
    dict(seq=3, ogl=None, doc="FR", bk="Image 0005", exec_eff="1905-05-04 receipt", rec="1905-05-12",
         grantor="U.S. Land Office", grantee="Herman Stevens", legal="NW/4 Section 32",
         notes="Final Receiver Receipt No. 3894; actual patent remains OPEN.", tier="DIRECT IMAGE P-003 — 90%"),
    dict(seq=4, ogl=None, doc="FR", bk="Image 0006", exec_eff="1905-05-04 receipt", rec="1905-05-12",
         grantor="U.S. Land Office", grantee="Augustus P. Stevens", legal="SW/4 Section 32",
         notes="Final Receiver Receipt No. 3895; actual patent remains OPEN.", tier="DIRECT IMAGE P-004 — 90%"),
    # PRE-1988 DIRECT IMAGES
    dict(seq=5, ogl=None, doc="Correction Assignment", bk="Bk 502/P267 (supersedes Bk 471/P330)",
         exec_eff="1981-05-28; corrected 1981-09-16", rec="1981-09-23 image stamp",
         grantor="Union Oil Co. of California", grantee="Leede Exploration",
         legal="Crook #1-32 wellbore only; Morrow-Springer 16,210-20,451 ft; effective first runs",
         notes="Union retained 7/64 gas/condensate/distillate ORRI and 1/16 oil/casinghead-gas ORRI; subject to unrecorded 1979 agreement.",
         tier="DIRECT IMAGE L-001 — 98%"),
    dict(seq=6, ogl=None, doc="Corrected ORRI Assignment", bk="Bk 509/P176",
         exec_eff="1981-09-17", rec="1981-10-13 image stamp",
         grantor="Leede Exploration",
         grantee="Campbell 1%; Park 1%; Cone .10%; Popham .10%; Hickman .05%",
         legal="Same Crook wellbore/depth interest",
         notes="Aggregate 2.25% ORRI; cost-free except taxes; pooling/free-fuel provisions.",
         tier="DIRECT IMAGE L-003 — 99%"),
    dict(seq=7, ogl=None, doc="Correction Assignment", bk="Bk 509/P220 (supersedes Bk 471/P337)",
         exec_eff="1981-09-17", rec="1981-10-13 image stamp",
         grantor="Leede Exploration",
         grantee="Mesa 64%; American Petrofina 25%; ENI 5%; McCall 4%; Midland National Bank, Trustee 1%; Sterling Petroleum-1978-A 1%",
         legal="Same Crook wellbore/depth interest; effective first runs",
         notes="Participants total 100%; Leede reserved ORRI for 75% aggregate NRI plus 1/4 after-payout reversion.",
         tier="DIRECT IMAGE L-002 — 98%"),
    dict(seq=8, ogl=None, doc="Assignment of Interest", bk="Bk 502/P265",
         exec_eff="1981-09-17", rec="1981-09-23 image stamp",
         grantor="Leede Exploration",
         grantee="Edward H. Leede 21.053%; Midland National Bank, Trustee 52.632%; Pine/Hudgeons/Frazier/Ray/Morgan 5.263% each",
         legal="All Leede royalty/ORRI in Section 32",
         notes="Allocations total 100%; expressly excludes Leede after-payout Crook WI.",
         tier="DIRECT IMAGE L-004 — 99%"),
    dict(seq=9, ogl=None, doc="Assignment", bk="Bk 839/P188",
         exec_eff="1985-09-12; effective 1984-07-01", rec="1985-09-27",
         grantor="First City National Bank of Midland, Trustee", grantee="Consolidation of Leede Trusts",
         legal="Only royalty/ORRI acquired under Bk 502/P265",
         notes="No warranty; does not transfer Bank's separate 1% Crook participant WI.",
         tier="DIRECT IMAGE L-005 — 99%"),
    dict(seq=10, ogl=None, doc="Assignment of Borehole Rights", bk="Bk 845/P47",
         exec_eff="1985-07-29; effective 1985-02-04", rec="1985-10-23",
         grantor="Cities Service Oil & Gas Corp.", grantee="Leede Exploration",
         legal="E/2 SE/4; Pearl #1-32 wellbore below 11,100 ft to but not below 19,480 ft",
         notes="Reserves 1/8 of 8/8 ORRI; five-year renewal/extension reach; subject to unrecorded 1985 agreement.",
         tier="DIRECT IMAGE L-006 — 98%"),
    dict(seq=11, ogl=None, doc="Assignment", bk="Bk 872/P279",
         exec_eff="1985-10-29; effective first production", rec="1986-03-06",
         grantor="Union Oil Co. of California", grantee="Leede Exploration",
         legal="Listed Section 32 leases top Morrow to 19,480 ft; E/2 SE/4 limited 11,100-19,480 ft; excludes Crook rights",
         notes="Reserves 1/8 gas/condensate/distillate ORRI and 1/16 oil/casinghead-gas ORRI.",
         tier="DIRECT IMAGE L-007 — 98%"),
    dict(seq=12, ogl=None, doc="PRF Agreement", bk="Bk 903/P50",
         exec_eff="1986-02-19; effective 1983-10-01", rec="See record image",
         grantor="Clyde Pine, Charles Ray, Robert Frazier / Leede Exploration", grantee="Mutual parties",
         legal="McCall #1-29 Hunton spacing: S/2 Sec. 29 + N/2 Sec. 32; Pearl #1-32 Lower Morrow: all Sec. 32",
         notes="Before payout: each PRF .526300% and Leede 98.421100% of LGWI; after payout: each PRF 1.315750% and Leede 96.052750%.",
         tier="DIRECT IMAGE L-008 — 98%"),
    dict(seq=13, ogl=None, doc="Assignment and Reservation", bk="Bk 987/P159 et seq.",
         exec_eff="1987-12-02; effective 1987-08-31", rec="1987-12-04",
         grantor="TCW Royalty Co., Trust Company of the West, Richard Grantham, Trustee, Arthur Carlson",
         grantee="Leede Exploration",
         legal="Returns rights under 1984 term-loan/royalty documents",
         notes="Reserves ORRI equal to .5% of Leede NRI: McCall .043146%, Pearl .052310%, Crook 0%.",
         tier="DIRECT IMAGE L-009 — 99%"),
    dict(seq=14, ogl=None, doc="Release", bk="Bk 995/P52",
         exec_eff="1988-01-13", rec="See record image",
         grantor="First City National Bank of Midland", grantee="Leede parties",
         legal="Enumerated Pearl/Crook/McCall mortgage interests",
         notes="Full/final release of enumerated 1984/1987 TCW, Vinson and First City liens; unlisted liens unaffected.",
         tier="DIRECT IMAGE L-010 — 97%"),
    dict(seq=15, ogl=None, doc="Assignment", bk="Bk 1014/P75",
         exec_eff="1988-02-18; effective 1988-01-01", rec="1988-04-26",
         grantor="Consolidation of Leede Trusts", grantee="Leede Exploration",
         legal="N/2 Sec. 32 all depths below top Hunton; plus S/2 Sec. 29 all depths",
         notes="Purports to assign all OGLs and working/cost-bearing interests without warranty. SOURCE-TITLE DEFECT: Consolidation source proves only royalty/ORRI. Last supplied direct image.",
         tier="DIRECT IMAGE L-011 — 99% text / 20% ownership effect"),
    # IMMEDIATE POST-CUTOFF (index only)
    dict(seq=16, ogl=None, doc="Assignment", bk="Bk 1014/P76", exec_eff="OPEN", rec="May 1988 index entry",
         grantor="Union Oil & Chemical et al.", grantee="Leede Exploration", legal="Section 32 index marks",
         notes="Early post-cutoff transfer; operative copy absent. PRIORITY PULL.", tier="INDEX ONLY — 85% metadata"),
    dict(seq=17, ogl=None, doc="Assignment", bk="Bk 1014/P228", exec_eff="OPEN", rec="May 1988 index entry",
         grantor="Mesa Operating Ltd. Partnership", grantee="Leede Exploration", legal="Section 32 index marks",
         notes="Early post-cutoff transfer; operative copy absent. PRIORITY PULL.", tier="INDEX ONLY — 85% metadata"),
    dict(seq=18, ogl=None, doc="Assignment", bk="Bk 1016/P38", exec_eff="OPEN", rec="1988-05-05",
         grantor="Leede Exploration et al.", grantee="Exxon Corp.", legal="Section 32 index entry",
         notes="Early post-cutoff transfer; underlying copy absent.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=19, ogl=None, doc="Assignment", bk="Bk 1137/P65", exec_eff="OPEN", rec="1990-01-05",
         grantor="Leede Exploration et al.", grantee="TCW Royalty Co. et al.", legal="Section 32 index entry",
         notes="Possible royalty/ORRI burden.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=20, ogl=None, doc="Stipulation", bk="Bk 1137/P340; Bk 1140/P1", exec_eff="OPEN", rec="1990-03-21",
         grantor="Atom Land & Exploration et al.", grantee="Leede Exploration et al.", legal="Section 32 index entry",
         notes="Potential title clarification.", tier="INDEX ONLY — 85% metadata"),
    dict(seq=21, ogl=None, doc="Assignments", bk="Bk 1217/P304; Bk 1217/P318", exec_eff="OPEN", rec="1991-07-25",
         grantor="Leede Exploration", grantee="NICOR Exploration Co.", legal="Section 32 index entry",
         notes="Possible outstanding leasehold slices.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=22, ogl=None, doc="Conveyance", bk="Bk 1257/P197", exec_eff="OPEN", rec="1992-04-21",
         grantor="Leede Exploration et al.", grantee="TCW Royalty Co. et al.", legal="Section 32 index entry",
         notes="Potential full-section royalty/ORRI conveyance.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=23, ogl=None, doc="Assignment", bk="Bk 1303/P202", exec_eff="OPEN", rec="1993-03-31",
         grantor="Leede Exploration", grantee="Exxon Corp.", legal="Full-section-indexed Exxon link",
         notes="Full-section-indexed Exxon link.", tier="INDEX ONLY — 92% metadata"),
    dict(seq=24, ogl=None, doc="Assignment / Conveyance", bk="Bk 1346/P34; Bk 1442/P270", exec_eff="OPEN", rec="1994-02-03 / 1995-10-24",
         grantor="Leede Exploration", grantee="Leede Operating Co. / Leede Operating Co. LLC", legal="Leede operating branch",
         notes="Leede operating branch.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=25, ogl=None, doc="Assignments", bk="Bk 1522/P84; Bk 1522/P87", exec_eff="OPEN", rec="1997-06-19",
         grantor="Leede Operating Co.", grantee="St. Mary Land & Exploration", legal="Section 32 index entry",
         notes="Potential outstanding/divested branch.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=26, ogl=None, doc="Assignment / Conveyance", bk="Bk 1524/P36; Bk 1533/P362", exec_eff="OPEN", rec="1997-07-07 / 1997-09-15",
         grantor="Mid-Continent Energy / Leede Exploration", grantee="Sanguine Ltd. / Exxon Corp.", legal="Separate/partial branches",
         notes="Exact tracts open.", tier="INDEX ONLY — 88% metadata"),
    dict(seq=27, ogl=None, doc="Assignment / Stipulations", bk="Bk 1577/P450; Bk 1581/P303; Bk 1581/P319", exec_eff="OPEN", rec="1998-09-24 / 1998-11-06",
         grantor="Leede Exploration / Leede Operating", grantee="Kabala Oil & Gas LLC / self-stipulations", legal="Section 32 index entry",
         notes="Potential outstanding branch/title clarification.", tier="INDEX ONLY — 88% metadata"),
    dict(seq=28, ogl=None, doc="Assignment", bk="Bk 1697/P574", exec_eff="OPEN", rec="2001-06-14",
         grantor="Staghorn Resources", grantee="Chesapeake Exploration", legal="Full-section-indexed Chesapeake acquisition",
         notes="Principal Chesapeake entry point.", tier="INDEX ONLY — 93% metadata"),
    dict(seq=29, ogl=None, doc="Assignment-type entries", bk="Bk 1783/P118; Bk 1784/P959", exec_eff="OPEN", rec="2003-07-28 / 2003-08-18",
         grantor="Mobil Exploration Co. / Exxon Mobil Corp.", grantee="Chesapeake Exploration", legal="Partial-tract predecessor links",
         notes="Partial-tract predecessor links.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=30, ogl=None, doc="Wellbore sale", bk="Bk 2183/P174", exec_eff="OPEN", rec="2015-02-18",
         grantor="Chesapeake Exploration et al.", grantee="Linn Operating", legal="Section 32 index entry",
         notes="Potential wellbore-specific divestiture.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=31, ogl=None, doc="Conveyance / Assignment / Memorandum", bk="Bk 2185/P639; Bk 2225/P1; Bk 2225/P710", exec_eff="OPEN", rec="2015-03-31 / 2016-06-07",
         grantor="Chesapeake Exploration et al.", grantee="FourPoint Energy", legal="FourPoint branch",
         notes="Wellbore/depth/excluded assets open; reconcile to Tapstone/KL CHK.", tier="INDEX ONLY — 95% metadata"),
    dict(seq=32, ogl=None, doc="Name Change", bk="Bk 2340/P218", exec_eff="OPEN", rec="2020-12-10",
         grantor="Unbridled Resources et al.", grantee="Public", legal="FourPoint/Unbridled identity evidence",
         notes="Corporate identity notice; not a conveyance.", tier="INDEX ONLY — 95% metadata"),
    dict(seq=33, ogl=None, doc="Assignment", bk="Bk 2340/P403", exec_eff="OPEN", rec="2020-12-14",
         grantor="Chesapeake Exploration et al.", grantee="Tapstone Energy LLC", legal="Full-section-indexed Tapstone branch",
         notes="Parallel co-acquirer branch; operative terms absent. PRIORITY PULL.", tier="INDEX ONLY — 97% metadata"),
    dict(seq=34, ogl=None, doc="Assignment", bk="Bk 2340/P490", exec_eff="OPEN", rec="2020-12-14",
         grantor="Chesapeake Exploration et al.", grantee="KL CHK SPV LLC", legal="Full-section-indexed KL CHK branch",
         notes="Principal predecessor to Diversified. PRIORITY PULL.", tier="INDEX ONLY — 98% metadata"),
    dict(seq=35, ogl=None, doc="Assignment", bk="Bk 2371/P470-494; I-2022-000152", exec_eff="2021-12-02 instrument date", rec="2022-01-12",
         grantor="KL CHK SPV LLC", grantee="Diversified Production LLC and OCM Denali Holdings LLC",
         legal="All 16 Section 32 Q-Q boxes indexed; long description",
         notes="Principal Diversified acquisition; fractions/schedules absent. PRIORITY PULL.", tier="PUBLIC METADATA + INDEX C-001 — 99% metadata / 0% decimal"),
    dict(seq=36, ogl=None, doc="Assignment / Agreement", bk="Bk 2371/P495; Bk 2371/P514", exec_eff="OPEN", rec="2022-01-12",
         grantor="Tapstone Energy LLC et al.", grantee="OCM Denali Holdings LLC", legal="Parallel co-acquirer branch",
         notes="Operative terms absent.", tier="INDEX ONLY — 97% metadata"),
    dict(seq=37, ogl=None, doc="Memorandum", bk="Bk 2371/P533; I-2022-000155", exec_eff="OPEN", rec="2022-01-12",
         grantor="Diversified Production LLC et al.", grantee="OCM Denali Holdings LLC et al.", legal="All 16 Q-Q ticks",
         notes="Notice of relationship; NOT a conveyance.", tier="INDEX ONLY — 97% metadata"),
    dict(seq=38, ogl=None, doc="Merger", bk="Bk 2389/P500-580; I-2022-003506", exec_eff="OPEN", rec="2022-08-24",
         grantor="Diversified Production LLC; DP Legacy Central LLC; DP KL CHK Merger Co LLC; DP Legacy Merger Co LLC",
         grantee="DP Legacy Tapstone LLC", legal="All 16 Q-Q ticks; corporate succession",
         notes="Certificate/terms absent. PRIORITY PULL.", tier="PUBLIC METADATA + INDEX — 99% metadata / 35% vesting"),
    dict(seq=39, ogl=None, doc="Amended Affidavit", bk="Bk 2389/P581; I-2022-003507", exec_eff="OPEN", rec="2022-08-24",
         grantor="Diversified Oil & Gas Corp. et al.", grantee="Public", legal="Corporate identity notice",
         notes="Not a conveyance.", tier="INDEX ONLY — 95% metadata"),
    dict(seq=40, ogl=None, doc="Assignment", bk="Bk 2393/P1", exec_eff="OPEN", rec="2022-10-17",
         grantor="Burlington Resources Oil & Gas", grantee="DP Legacy Central LLC", legal="Full-section index marks",
         notes="Separate DP Legacy Central branch; schedule absent.", tier="INDEX ONLY — 90% metadata"),
    dict(seq=41, ogl=None, doc="Conveyance", bk="Bk 2395/P415-464; I-2022-004838", exec_eff="OPEN", rec="2022-11-18",
         grantor="DP Legacy Tapstone LLC", grantee="Diversified ABS VI Upstream LLC and DP Sooner HoldCo LLC",
         legal="All 16 Q-Q ticks; long description",
         notes="Principal ABS conveyance; estates/fractions/schedules absent. PRIORITY PULL.", tier="PUBLIC METADATA C-002 — 99% metadata / 0% decimal"),
    dict(seq=42, ogl=None, doc="Termination", bk="Bk 2395/P967-972; I-2022-004938", exec_eff="OPEN", rec="2022-11-28",
         grantor="KeyBank National Association", grantee="Tapstone Energy LLC; Diversified Production LLC",
         legal="Financing-statement termination only",
         notes="NOT a title conveyance. Diversified ABS conveyance is Bk 2395/P415.", tier="INDEX ONLY — 99% (RESOLVED)"),
    dict(seq=43, ogl=None, doc="Conveyance", bk="Bk 2400/P551-567; I-2023-000796", exec_eff="OPEN", rec="2023-03-02",
         grantor="DP Sooner HoldCo LLC", grantee="Diversified Production LLC", legal="Broad/all-section ticks; long description",
         notes="Relation to prior branch and subject asset fraction open. PRIORITY PULL.", tier="PUBLIC METADATA C-004 — 99% metadata / 25% effect"),
    dict(seq=44, ogl=None, doc="Conveyance", bk="Bk 2415/P328", exec_eff="OPEN", rec="2023-11-14",
         grantor="Unbridled Resources", grantee="MNR ABS Issuer I WB", legal="FourPoint/Unbridled securitization branch",
         notes="Reconcile to Tapstone/KL CHK/Diversified.", tier="INDEX ONLY — 92% metadata"),
    dict(seq=45, ogl=None, doc="Merger", bk="Bk 2451/P4-23; I-2025-001808", exec_eff="OPEN", rec="2025-05-13",
         grantor="Diversified Energy Company PLC", grantee="Public", legal="Corporate merger metadata",
         notes="Merged/surviving entities and title effect absent.", tier="PUBLIC METADATA C-005 — 99% metadata / 20% effect"),
    dict(seq=46, ogl=None, doc="Merger", bk="Bk 2475/P943", exec_eff="OPEN", rec="2026-01-23",
         grantor="Canvas Energy LLC", grantee="Canvas Energy II LLC", legal="All 16 Q-Q ticks",
         notes="Canvas corporate succession; separate from principal KL CHK branch.", tier="INDEX ONLY — 95% metadata"),
    dict(seq=47, ogl=None, doc="Conveyance", bk="Bk 2476/P1", exec_eff="OPEN", rec="2026-01-23",
         grantor="Canvas Energy II LLC", grantee="DP Ponies LLC", legal="All 16 Q-Q ticks",
         notes="Separate DP-affiliated acquisition; relation to principal Diversified branch open.", tier="INDEX ONLY — 95% metadata"),
    dict(seq=48, ogl=None, doc="Agreement", bk="Bk 2476/P121-130; I-2026-000215", exec_eff="OPEN", rec="2026-01-23",
         grantor="Diversified Production LLC; DP Ponies LLC; DP Secretariat LLC; DP American Pharoah LLC; DP Seabiscuit LLC; DP Sovereignty LLC",
         grantee="Same group (DE American Pharoah LLC name variance on grantee side)", legal="All 16 Q-Q ticks; long description",
         notes="Agreement only; not facially a conveyance; effect absent. PRIORITY PULL.", tier="PUBLIC METADATA C-006 — 99% metadata / 10% effect"),
    # COUNTYWIDE SCREEN (not confirmed Section 32)
    dict(seq=49, ogl=None, doc="Wellbore Assignment (SCREENED)", bk="Bk 2434/P751-760; I-2024-002930", exec_eff="OPEN", rec="OPEN",
         grantor="Diversified Production LLC and DP Legacy Central LLC", grantee="Teocalli Exploration LLC",
         legal="Long description; NOT indexed to subject section in supplied index",
         notes="Pull schedule before excluding.", tier="COUNTYWIDE SCREEN — applicability OPEN"),
    dict(seq=50, ogl=None, doc="Wellbore Assignment (SCREENED)", bk="Bk 2464/P200-217; I-2025-003599", exec_eff="OPEN", rec="OPEN",
         grantor="Stephens Production Co. / Continental Properties affiliates", grantee="Diversified Production LLC",
         legal="Long description; Section 32 applicability not established",
         notes="Potential acquisition; pull schedule.", tier="COUNTYWIDE SCREEN — applicability OPEN"),
    dict(seq=51, ogl=None, doc="Affidavit (SCREENED)", bk="Bk 2480/P824-883; I-2026-001031", exec_eff="OPEN", rec="OPEN",
         grantor="DP Legacy Central LLC; Diversified Production LLC", grantee="Public / affidavit filing",
         legal="60 pages; long description",
         notes="Subject applicability and substantive effect OPEN.", tier="COUNTYWIDE SCREEN — applicability OPEN"),
    dict(seq=52, ogl=None, doc="Merger (SCREENED)", bk="Bk 2480/P903-909; I-2026-001039", exec_eff="OPEN", rec="OPEN",
         grantor="Multiple DP/Diversified/Canvas affiliates", grantee="Public",
         legal="No indexed legal description",
         notes="Corporate effect may be global; pull certificate/plan.", tier="COUNTYWIDE SCREEN — applicability OPEN"),
]

# OGL LEASE DATA
OGL_LEASES = [
    (1, "1947-11-13", "Garretts", "C.E. McKinley", "63/33", "E/2 SE/4", "80 ac nominal", "Below 11,100 ft in later assignment", "OPEN", "OPEN", "Tract 2 / Pearl"),
    (2, "1972-02-05", "Charlie B. and Ina Pearl Crook", "Union", "264/345", "NE/4", "160 ac nominal", "Crook/deep assignment subject", "OPEN", "OPEN", "Tract 1"),
    (3, "1976-04-29", "Cora B. LeGrand", "Union", "307/31", "W/2 SE/4 + S/2 SW/4", "120 ac nominal", "Deep assignment subject", "OPEN", "OPEN", "Tract 4"),
    (4, "1977-10-14", "Gertrude L. LaFortune et al.", "Union", "340/302", "N/2 SW/4", "40 ac nominal", "Deep assignment subject", "OPEN", "OPEN", "Tract 4"),
    (5, "1977-10-14", "Gertrude LaFortune, widow", "Union", "340/304", "N/2 SW/4", "40 ac nominal", "Overlap — read with 340/302", "OPEN", "OPEN", "Tract 4"),
    (6, "1977-09-29", "George T. Harrison Jr. and Renata Harrison", "Union", "340/323", "N/2 NW/4 + N/2 SW/4", "80 ac nominal", "Deep assignment subject", "OPEN", "OPEN", "Tract 4"),
    (7, "1977-09-29", "Carol Seidenbach Leach", "Union", "342/564", "N/2 SW/4", "40 ac nominal", "Deep assignment subject", "OPEN", "OPEN", "Tract 4"),
    (8, "1977-11-01", "Cynthia Seidenbach Kruse", "Union", "342/566", "N/2 SW/4", "40 ac nominal", "Deep assignment subject", "OPEN", "OPEN", "Tract 4"),
    (9, "1978-05-16", "F.L. Randel", "Union", "352/719", "NE/4", "160 ac nominal", "Crook/deep assignment subject", "OPEN", "OPEN", "Tract 1"),
    (10, "1978-05-16", "J.L. Randel", "Union", "352/721", "NE/4", "160 ac nominal", "Crook/deep assignment subject", "OPEN", "OPEN", "Tract 1"),
    (11, "1979-03-21", "Great Western Oil & Gas", "Union", "376/369", "NE/4", "160 ac nominal", "Crook/deep assignment subject", "OPEN", "OPEN", "Tract 1"),
]

# TRACT INSTRUMENT COLUMN DEFINITIONS: col_letter -> {type, img, bk, date, conveyance, tier, grantor_path}
TRACT_INSTRUMENTS = {
    "Tract 1": {
        "desc": "NE/4 (160 ac nominal); Crook #1-32 wellbore; Morrow-Springer 16,210-20,451 ft",
        "acres": 160,
        "cols": [
            ("H", "PAT", "0003", "Patent Bk 3/469", "1907-06-08", "NE/4 patent — ARTI to patentee only", "DIRECT IMAGE REVIEWED", "United States → Cornelius B. Cornelius"),
            ("I", "CORR ASSN", "1940-1943", "502/267", "1981-09-23", "Crook wellbore WI; Union ORRI reserved", "DIRECT IMAGE REVIEWED", "Union Oil → Leede Exploration"),
            ("J", "ORRI ASSN", "1946-1948", "509/176", "1981-10-13", "2.25% aggregate employee ORRI", "DIRECT IMAGE REVIEWED", "Leede → Campbell/Park/Cone/Popham/Hickman"),
            ("K", "CORR ASSN", "1949-1952", "509/220", "1981-10-13", "Participant WI; 75% agg NRI; payout reversion", "DIRECT IMAGE REVIEWED", "Leede → Mesa/Petrofina/ENI/McCall/Midland/Sterling"),
            ("L", "ASSN INT", "1938-1939", "502/265", "1981-09-23", "Section 32 royalty/ORRI pool allocation", "DIRECT IMAGE REVIEWED", "Leede → Leede beneficiaries"),
            ("M", "ASSN", "4103", "839/188", "1985-09-27", "Royalty/ORRI only from 502/265", "DIRECT IMAGE REVIEWED", "First City Bank → Consolidation"),
            ("N", "RELEASE", "4868-4870", "995/52", "1988-01-13", "Enumerated financing release", "DIRECT IMAGE REVIEWED", "First City Bank → Leede parties"),
            ("O", "ASSN", "INDEX", "1016/38", "1988-05-05", "Leede → Exxon; INDEX ONLY", "INDEX ONLY", "Leede et al. → Exxon Corp."),
            ("P", "ASSN", "INDEX", "1303/202", "1993-03-31", "Leede → Exxon; INDEX ONLY", "INDEX ONLY", "Leede → Exxon Corp."),
            ("Q", "ASSN", "INDEX", "1697/574", "2001-06-14", "Staghorn → Chesapeake; INDEX ONLY", "INDEX ONLY", "Staghorn → Chesapeake"),
            ("R", "ASSN", "INDEX", "2340/490", "2020-12-14", "Chesapeake → KL CHK; INDEX ONLY", "INDEX ONLY", "Chesapeake → KL CHK SPV LLC"),
            ("S", "ASSN", "INDEX", "2371/470", "2022-01-12", "KL CHK → Diversified/OCM Denali", "INDEX + METADATA", "KL CHK → Diversified Production LLC"),
            ("T", "CONV", "INDEX", "2395/415", "2022-11-18", "DP Legacy Tapstone → ABS VI / Sooner HoldCo", "INDEX + METADATA", "DP Legacy Tapstone → Diversified ABS VI"),
            ("U", "CONV", "INDEX", "2400/551", "2023-03-02", "Sooner HoldCo → Diversified Production", "INDEX + METADATA", "DP Sooner HoldCo → Diversified Production LLC"),
            ("V", "AGMT", "INDEX", "2476/121", "2026-01-23", "DP affiliate agreement; NOT conveyance", "INDEX + METADATA", "DP affiliates (agreement)"),
        ],
    },
    "Tract 2": {
        "desc": "E/2 SE/4 (80 ac nominal); Pearl #1-32 wellbore 11,100-19,480 ft",
        "acres": 80,
        "cols": [
            ("H", "FR", "0004", "Image 0004", "1906-08-08", "SE/4 sovereign inception (receipt)", "DIRECT IMAGE REVIEWED", "U.S. Land Office → Velmore Biggs"),
            ("I", "OGL", "INDEX", "63/33", "1947-11-13", "Garretts → McKinley; E/2 SE/4; terms OPEN", "LEASE REF ONLY", "Garretts → C.E. McKinley"),
            ("J", "BORRH ASSN", "4118-4119", "845/47", "1985-10-23", "Pearl wellbore; Cities Service 1/8 ORRI reserved", "DIRECT IMAGE REVIEWED", "Cities Service → Leede Exploration"),
            ("K", "ASSN", "4388-4393", "872/279", "1986-03-06", "Deep leasehold top Morrow-19,480 ft; E/2 SE/4 11,100-19,480", "DIRECT IMAGE REVIEWED", "Union Oil → Leede Exploration"),
            ("L", "PRF", "4438-4463", "903/50", "See image", "Pearl Lower Morrow spacing; payout LGWI", "DIRECT IMAGE REVIEWED", "PRF parties / Leede"),
            ("M", "ASSN RES", "4840-4852", "987/159", "1987-12-04", "TCW .5% of Leede NRI ORRI on Pearl", "DIRECT IMAGE REVIEWED", "TCW et al. → Leede Exploration"),
            ("N", "RELEASE", "4868-4870", "995/52", "1988-01-13", "Enumerated financing release", "DIRECT IMAGE REVIEWED", "First City Bank → Leede parties"),
            ("O", "ASSN", "INDEX", "2340/403", "2020-12-14", "Chesapeake → Tapstone; INDEX ONLY", "INDEX ONLY", "Chesapeake → Tapstone Energy LLC"),
            ("P", "ASSN", "INDEX", "2371/470", "2022-01-12", "KL CHK → Diversified/OCM Denali", "INDEX + METADATA", "KL CHK → Diversified Production LLC"),
            ("Q", "CONV", "INDEX", "2395/415", "2022-11-18", "DP Legacy Tapstone → ABS VI / Sooner HoldCo", "INDEX + METADATA", "DP Legacy Tapstone → Diversified ABS VI"),
        ],
    },
    "Tract 3": {
        "desc": "N/2 Section 32 below top Hunton (320 ac nominal)",
        "acres": 320,
        "cols": [
            ("H", "FR", "0005", "Image 0005", "1905-05-12", "NW/4 receipt (N/2 chain component)", "DIRECT IMAGE REVIEWED", "U.S. Land Office → Herman Stevens"),
            ("I", "FR", "0006", "Image 0006", "1905-05-12", "SW/4 receipt (N/2 chain component)", "DIRECT IMAGE REVIEWED", "U.S. Land Office → Augustus P. Stevens"),
            ("J", "PRF", "4438-4463", "903/50", "See image", "McCall Hunton N/2 Sec. 32 spacing; payout LGWI", "DIRECT IMAGE REVIEWED", "PRF parties / Leede"),
            ("K", "ASSN", "4893", "1014/75", "1988-04-26", "N/2 below top Hunton; SOURCE-TITLE DEFECT", "DIRECT IMAGE REVIEWED", "Consolidation → Leede Exploration"),
            ("L", "ASSN", "INDEX", "1014/76", "May 1988", "Post-cutoff; INDEX ONLY", "INDEX ONLY", "Union Oil et al. → Leede"),
            ("M", "ASSN", "INDEX", "1014/228", "May 1988", "Post-cutoff; INDEX ONLY", "INDEX ONLY", "Mesa Operating → Leede"),
            ("N", "ASSN", "INDEX", "2371/470", "2022-01-12", "KL CHK → Diversified/OCM Denali", "INDEX + METADATA", "KL CHK → Diversified Production LLC"),
            ("O", "CONV", "INDEX", "2395/415", "2022-11-18", "DP Legacy Tapstone → ABS VI / Sooner HoldCo", "INDEX + METADATA", "DP Legacy Tapstone → Diversified ABS VI"),
        ],
    },
    "Tract 4": {
        "desc": "Deep leasehold references — listed OGLs top Morrow to 19,480 ft (multiple aliquots)",
        "acres": "Multiple",
        "cols": [
            ("H", "OGL", "INDEX", "63/33", "1947-11-13", "E/2 SE/4", "LEASE REF ONLY", "Garretts → McKinley"),
            ("I", "OGL", "INDEX", "264/345", "1972-02-05", "NE/4", "LEASE REF ONLY", "Crook → Union"),
            ("J", "OGL", "INDEX", "307/31", "1976-04-29", "W/2 SE/4 + S/2 SW/4", "LEASE REF ONLY", "LeGrand → Union"),
            ("K", "OGL", "INDEX", "340/302", "1977-10-14", "N/2 SW/4", "LEASE REF ONLY", "LaFortune → Union"),
            ("L", "OGL", "INDEX", "340/323", "1977-09-29", "N/2 NW/4 + N/2 SW/4", "LEASE REF ONLY", "Harrison → Union"),
            ("M", "OGL", "INDEX", "342/564", "1977-09-29", "N/2 SW/4", "LEASE REF ONLY", "Leach → Union"),
            ("N", "OGL", "INDEX", "352/719", "1978-05-16", "NE/4", "LEASE REF ONLY", "F.L. Randel → Union"),
            ("O", "OGL", "INDEX", "376/369", "1979-03-21", "NE/4", "LEASE REF ONLY", "Great Western → Union"),
            ("P", "ASSN", "4388-4393", "872/279", "1986-03-06", "Union → Leede deep leasehold package", "DIRECT IMAGE REVIEWED", "Union Oil → Leede Exploration"),
            ("Q", "ASSN RES", "4840-4852", "987/159", "1987-12-04", "TCW ORRI schedule", "DIRECT IMAGE REVIEWED", "TCW et al. → Leede"),
            ("R", "ASSN", "INDEX", "2371/470", "2022-01-12", "KL CHK → Diversified (full-section index)", "INDEX + METADATA", "KL CHK → Diversified Production LLC"),
        ],
    },
    "Tract 5": {
        "desc": "All Section 32 — Diversified-affiliated branch (index/metadata only post-1988)",
        "acres": 640,
        "cols": [
            ("H", "ASSN", "INDEX", "2340/403", "2020-12-14", "Tapstone branch from Chesapeake", "INDEX ONLY", "Chesapeake → Tapstone Energy LLC"),
            ("I", "ASSN", "INDEX", "2340/490", "2020-12-14", "KL CHK branch from Chesapeake", "INDEX ONLY", "Chesapeake → KL CHK SPV LLC"),
            ("J", "ASSN", "INDEX", "2371/470", "2022-01-12", "KL CHK → Diversified Production / OCM Denali", "INDEX + METADATA", "KL CHK → Diversified Production LLC"),
            ("K", "ASSN", "INDEX", "2371/495", "2022-01-12", "Tapstone → OCM Denali parallel branch", "INDEX ONLY", "Tapstone → OCM Denali Holdings LLC"),
            ("L", "MEMO", "INDEX", "2371/533", "2022-01-12", "Relationship notice; NOT conveyance", "INDEX ONLY", "Diversified Production et al."),
            ("M", "MERGER", "INDEX", "2389/500", "2022-08-24", "DP entities → DP Legacy Tapstone LLC", "INDEX + METADATA", "DP merger group → DP Legacy Tapstone"),
            ("N", "AFF", "INDEX", "2389/581", "2022-08-24", "Corporate identity notice", "INDEX ONLY", "Diversified Oil & Gas Corp."),
            ("O", "ASSN", "INDEX", "2393/1", "2022-10-17", "Burlington → DP Legacy Central (separate branch)", "INDEX ONLY", "Burlington → DP Legacy Central LLC"),
            ("P", "CONV", "INDEX", "2395/415", "2022-11-18", "DP Legacy Tapstone → ABS VI / Sooner HoldCo", "INDEX + METADATA", "DP Legacy Tapstone → Diversified ABS VI"),
            ("Q", "TERM", "INDEX", "2395/967", "2022-11-28", "KeyBank UCC termination — NOT conveyance", "INDEX ONLY (RESOLVED)", "KeyBank → Tapstone/Diversified"),
            ("R", "CONV", "INDEX", "2400/551", "2023-03-02", "Sooner HoldCo → Diversified Production", "INDEX + METADATA", "DP Sooner HoldCo → Diversified Production LLC"),
            ("S", "CONV", "INDEX", "2415/328", "2023-11-14", "Unbridled → MNR ABS (FourPoint branch)", "INDEX ONLY", "Unbridled → MNR ABS Issuer I WB"),
            ("T", "MERGER", "INDEX", "2451/4", "2025-05-13", "Diversified Energy Company PLC merger notice", "INDEX + METADATA", "Diversified Energy Company PLC"),
            ("U", "MERGER", "INDEX", "2475/943", "2026-01-23", "Canvas Energy succession", "INDEX ONLY", "Canvas Energy LLC → Canvas Energy II LLC"),
            ("V", "CONV", "INDEX", "2476/1", "2026-01-23", "Canvas → DP Ponies (separate branch)", "INDEX ONLY", "Canvas Energy II → DP Ponies LLC"),
            ("W", "AGMT", "INDEX", "2476/121", "2026-01-23", "DP affiliate agreement; NOT conveyance", "INDEX + METADATA", "Diversified Production + DP affiliates"),
        ],
    },
}

WELLS = [
    ("1", "35-009-00075", "BP America", "PA", "Cora Garrett C-310", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Plugged/abandoned; no Diversified title inference"),
    ("2", "35-009-20312", "Complete Oil & Gas Services LLC", "AC", "Crook 1-32", "Granite Wash Lower", "Section 32, T11N-R25W", "OPEN", "1980-01-26", "Crook wellbore; Granite Wash Lower 13,348-13,595 ft; Tract 1 context only"),
    ("3", "35-009-20605", "SM Energy", "PA", "Pearl 1-32", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Historical Pearl well; Tract 2 context only"),
    ("4", "35-009-20738", "Leede Oil & Gas", "PA", "Legrand 1-32", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Plugged/abandoned"),
    ("5", "35-009-21072", "BCE-Mach II LLC", "AC", "JGL 1-32", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Active; formation OPEN; no HBP inference"),
    ("6", "35-009-21085", "BCE-Mach II LLC", "AC", "Hanks Crook 1-32", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Active; formation OPEN"),
    ("7", "35-009-21166", "Vastar", "PA", "Crook-Legrand 1-32", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Plugged/abandoned"),
    ("8", "35-009-21236", "BCE-Mach II LLC", "AC", "Legrand 1-32", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Active; formation OPEN"),
    ("9", "35-009-21245", "BP America", "EX", "Katie 1-32", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Expired/nonactive"),
    ("10", "35-009-21258", "BCE-Mach II LLC", "AC", "Legrand 2-32", "OPEN", "Section 32, T11N-R25W", "OPEN", "OPEN", "Active; formation OPEN"),
    ("11", "35-009-21391", "BCE-Mach II LLC", "AC", "Twin 1-32", "Granite Wash Low", "Section 32, T11N-R25W", "OPEN", "2004-07-15", "Granite Wash Low 12,548-12,604 ft; no HBP inference"),
]


def set_cell(ws, coord, value):
    ws[coord] = value


def populate_runsheet(wb):
    ws = wb["Runsheet"]
    # Clear existing data rows
    for r in range(2, ws.max_row + 1):
        for c in range(1, 12):
            cell = ws.cell(r, c)
            if type(cell).__name__ != "MergedCell":
                cell.value = None
    headers = ["Seq", "OGL", "Doc Type", "Book / Page", "Execution / Effective Date",
               "Recording Date", "Grantor", "Grantee", "Legal / Estate", "Notes", "Evidence Tier / Confidence"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for i, row in enumerate(RUNSHEET, 2):
        ws.cell(i, 1, row["seq"])
        ws.cell(i, 2, row.get("ogl"))
        ws.cell(i, 3, row["doc"])
        ws.cell(i, 4, row["bk"])
        ws.cell(i, 5, row["exec_eff"])
        ws.cell(i, 6, row["rec"])
        ws.cell(i, 7, row["grantor"])
        ws.cell(i, 8, row["grantee"])
        ws.cell(i, 9, row["legal"])
        ws.cell(i, 10, row["notes"])
        ws.cell(i, 11, row["tier"])
    print(f"Runsheet: {len(RUNSHEET)} rows populated")


def populate_tract(wb, sheet_name, tract_data):
    ws = wb[sheet_name]
    desc = tract_data["desc"]
    acres = tract_data["acres"]
    cols = tract_data["cols"]

    # Row 1: Instrument types
    ws["E1"] = "Instrument"
    for col_letter, inst_type, *_ in cols:
        ws[f"{col_letter}1"] = inst_type

    # Row 2: Tract header
    ws["A2"] = 1
    ws["B2"] = "TRACT"
    ws["D2"] = desc
    ws["E2"] = "Image Number"
    for col_letter, _, img, *_ in cols:
        ws[f"{col_letter}2"] = img

    # Row 3: Book/Page
    ws["B3"] = "DESCRIPTION"
    ws["E3"] = "Book / Page"
    for col_letter, _, _, bk, *_ in cols:
        ws[f"{col_letter}3"] = bk

    # Row 4: Date label
    ws["E4"] = "Date"

    # Row 5: Dates + acres
    ws["B5"] = "TRACT ACRES"
    ws["D5"] = acres
    ws["E5"] = "Conveyance"
    for col_letter, _, _, _, date, *_ in cols:
        ws[f"{col_letter}5"] = date

    # Row 6: Conveyance descriptions
    ws["B6"] = "BASE OGL"
    ws["C6"] = acres if isinstance(acres, (int, float)) else None
    ws["D6"] = sheet_name
    ws["F6"] = "SURFACE"
    for col_letter, _, _, _, _, conv, *_ in cols:
        ws[f"{col_letter}6"] = conv

    # Row 8: Present owner status
    ws["A8"] = "Historical chain through direct images and index; present Diversified owner OPEN ITEM"
    for col_letter, *_ in cols:
        tier = cols[[c[0] for c in cols].index(col_letter)][6]
        ws[f"{col_letter}8"] = tier

    # Row 9: Evidence tier label
    ws["A9"] = "Evidence Tier (row 8 per instrument column)"

    # Row 10+: Party paths
    ws["B10"] = "Grantor → Grantee (summary)"
    for idx, (col_letter, *rest) in enumerate(cols):
        party = rest[6] if len(rest) > 6 else ""
        ws.cell(10 + idx, 4, party)  # column D for readability
        ws[f"{col_letter}10"] = party

    # Row 20: Current owner
    ws["A20"] = "CURRENT RECORD OWNER"
    ws["B20"] = "OPEN ITEM"
    ws["D20"] = "Diversified-affiliated entity not established — missing post-1988 schedules"
    ws["H20"] = "Pull priority copies per Overview B55"

    print(f"{sheet_name}: {len(cols)} instrument columns populated")


def populate_ogl(wb):
    ws = wb["OGL"]
    # Find or create data area - read first row headers
    headers = ["OGL #", "Effective Date", "Lessor", "Lessee", "Book/Page", "Legal Description",
               "Gross Acres", "Depth/Later Treatment", "Royalty", "Primary Term", "HBP Status", "Tract", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r, lease in enumerate(OGL_LEASES, 2):
        num, eff, lessor, lessee, bk, legal, acres, depth, royalty, term, tract = lease
        ws.cell(r, 1, num)
        ws.cell(r, 2, eff)
        ws.cell(r, 3, lessor)
        ws.cell(r, 4, lessee)
        ws.cell(r, 5, bk)
        ws.cell(r, 6, legal)
        ws.cell(r, 7, acres)
        ws.cell(r, 8, depth)
        ws.cell(r, 9, royalty)
        ws.cell(r, 10, term)
        ws.cell(r, 11, "OPEN — no lease-to-well proof")
        ws.cell(r, 12, tract)
        ws.cell(r, 13, "OPEN — original OGL not abstracted")
    print(f"OGL: {len(OGL_LEASES)} leases populated")


def populate_wells(wb):
    ws = wb["Well 1"]
    headers = ["Well #", "API #", "Current Listed Operator", "Status", "Well Name", "Formation",
               "Surface Location", "Bottom Hole Location", "Spud Date", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for r, w in enumerate(WELLS, 2):
        for c, val in enumerate(w, 1):
            ws.cell(r, c, val)
    print(f"Well 1: {len(WELLS)} wells populated")


def populate_wi_sheets(wb):
    # WI 1 - Crook participants (historical)
    ws = wb["WI 1"]
    participants = [
        ("Mesa Operating Ltd. Partnership", "64%", "509/220", "Crook wellbore WI", "HISTORICAL", "Payout/reassignments OPEN"),
        ("American Petrofina Exploration", "25%", "509/220", "Crook wellbore WI", "HISTORICAL", "Payout/reassignments OPEN"),
        ("ENI Exploration", "5%", "509/220", "Crook wellbore WI", "HISTORICAL", "Payout/reassignments OPEN"),
        ("Jack O. McCall", "4%", "509/220", "Crook wellbore WI", "HISTORICAL", "Payout/reassignments OPEN"),
        ("Midland National Bank, Trustee", "1%", "509/220", "Crook participant WI", "HISTORICAL", "Separate 1% WI source OPEN"),
        ("Sterling Petroleum-1978-A", "1%", "509/220", "Crook wellbore WI", "HISTORICAL", "Payout/reassignments OPEN"),
        ("Leede Exploration (after-payout reversion)", "1/4 reversion", "509/220", "Crook ORRI/reversion", "OPEN", "Payout status OPEN"),
        ("Diversified Production LLC (or successor)", "OPEN ITEM", "2371/470 et seq.", "Full-section Diversified branch", "OPEN ITEM", "No supported decimal"),
    ]
    ws["A1"] = "WI 1 — CROOK #1-32 WELLBORE PARTICIPANTS (HISTORICAL + DIVERSIFIED STATUS)"
    ws["A3"] = "Party"
    ws["B3"] = "Historical WI %"
    ws["C3"] = "Source"
    ws["D3"] = "Estate"
    ws["E3"] = "Status"
    ws["F3"] = "Open Item"
    for r, p in enumerate(participants, 4):
        for c, v in enumerate(p, 1):
            ws.cell(r, c, v)

    ws2 = wb["WI 2"]
    burdens = [
        ("Union Oil ORRI (Crook)", "7/64 gas + 1/16 oil", "502/267", "Crook 16,210-20,451 ft", "OPEN survival"),
        ("Employee ORRI aggregate", "2.25% of 8/8", "509/176", "Crook wellbore", "OPEN survival"),
        ("Cities Service ORRI (Pearl)", "1/8 of 8/8", "845/47", "Pearl 11,100-19,480 ft", "OPEN survival"),
        ("Union deep ORRI", "1/8 gas + 1/16 oil", "872/279", "Top Morrow-19,480 ft", "OPEN survival"),
        ("PRF LGWI (3 parties)", ".5263%/.5263%/.5263% pre-payout", "903/50", "McCall Hunton / Pearl Morrow", "Payout OPEN"),
        ("TCW ORRI", ".5% of Leede NRI", "987/159", "McCall/Pearl/Crook table", "OPEN survival"),
        ("Modern blanket financing", "Schedules absent", "2341/127; 2371/554; 2389/590; 2415/410; 2476/67", "Tapstone/DP/MNR/Ponies", "Lien status OPEN"),
    ]
    ws2["A1"] = "WI 2 — BURDENS / ORRI / FINANCING"
    ws2["A3"] = "Burden"
    ws2["B3"] = "Quantum"
    ws2["C3"] = "Source"
    ws2["D3"] = "Scope"
    ws2["E3"] = "Present Effect"
    for r, b in enumerate(burdens, 4):
        for c, v in enumerate(b, 1):
            ws2.cell(r, c, v)
    print("WI 1 and WI 2 populated")


def populate_overview(wb):
    ws = wb["Overview"]
    updates = {
        "Q3": "32-11N-25W / Diversified",
        "H5": 32, "R5": "11N", "AC5": "25W", "H7": "Beckham",
        "B9": "Containing 640.00 acres, more or less (nominal section acreage; not a title decimal)",
        "AB15": "Tract 1 — NE/4 / Crook wellbore",
        "U22": "Tract 2 — E/2 SE/4 / Pearl",
        "M26": "Tract 3 — N/2 below top Hunton",
        "E26": "Tract 4 — Deep leasehold refs",
        "X34": "Tract 5 — All Sec. 32 / Diversified branch",
        "S48": datetime(2026, 7, 10),
        "AC48": "Bk 2490/P471 (WD) — last indexed entry",
        "S50": datetime(2026, 7, 1),
        "B53": (
            "CONTROLLING CONCLUSION: Diversified-affiliated entities acquired a material interest indexed "
            "against Section 32 through the 2021/2022 KL CHK/Tapstone structure, but missing post-1988 "
            "operative instruments and schedules prevent determination of exact estate, current vested entity, "
            "WI, NRI, leasehold/ORRI decimal, mineral/royalty interest, or net acres."
        ),
        "B54": (
            "SCOPE/CUTOFFS: 4,893 direct images end at Bk 1014/P75 (rec. 1988-04-26). Section index certified "
            "through 07/01/2026; last entry Bk 2490/P471. Runsheet contains 52 instruments (15 direct-image, "
            "37 index/metadata). Last principal Diversified item: Bk 2476/P121 (effect OPEN)."
        ),
        "B55": (
            "PRIORITY COPY PULL: 2340/403, 2340/490, 2371/470, /495, /514, /533, 2389/500, /581, 2393/1, "
            "2395/415, 2400/551, 2451/4, 2475/943, 2476/1 and 2476/121 with every exhibit/schedule; "
            "bridge post-04/26/1988 chain from 1014/76 forward."
        ),
        "F56": "New Horizon Energy (NHE) / Cursor — perfected cursory 2026-07-10",
    }
    for coord, val in updates.items():
        ws[coord] = val
    print("Overview updated")


def populate_assumptions(wb):
    if "Assumptions" not in wb.sheetnames:
        ws = wb.create_sheet("Assumptions", 0)
    else:
        ws = wb["Assumptions"]
    rows = [
        ("ASSUMPTIONS & LIMITATIONS — Section 32-11N-25W", ""),
        ("", ""),
        ("ID", "Assumption / Limitation"),
        ("A1", "Full-section Q-Q index ticks do NOT prove 100% leasehold, all formations, or fee minerals."),
        ("A2", "Historical Leede/Crook/Pearl interests are wellbore/depth-limited; not carried forward without post-1988 bridge."),
        ("A3", "Bk 2395/P967 is financing termination only; Diversified ABS conveyance is Bk 2395/P415."),
        ("A4", "SE/NW/SW sovereign inception based on receiver receipts; actual patents not reviewed."),
        ("A5", "No Diversified decimal, NRI, or net acres stated where operative schedules absent."),
        ("A6", "Section-level well activity does not prove any lease is HBP."),
        ("A7", "FourPoint/Unbridled/MNR branch may overlap Tapstone/KL CHK — reconcile before final opinion."),
        ("A8", "Countywide DP instruments (2434/751, 2464/200, 2480/824, 2480/903) screened; Section 32 applicability OPEN."),
        ("A9", "Surface assessor data is context only; does not establish mineral title."),
        ("A10", "Report date 2026-07-10; continuation from 2026-07-01 required for final effective date."),
        ("A11", "Separate branches: principal KL CHK/Tapstone-DP, DP Legacy Central/Burlington, Canvas/DP Ponies, FourPoint/MNR."),
        ("", ""),
        ("QA VERIFIED CALCULATIONS", ""),
        ("", "Crook participants: 64+25+5+4+1+1 = 100%"),
        ("", "Employee ORRI: 1+1+.10+.10+.05 = 2.25%"),
        ("", "Leede royalty pool: 21.053+52.632+5x5.263 = 100%"),
        ("", "PRF before payout: 3x.526300+98.421100 = 100% LGWI"),
        ("", "PRF after payout: 3x1.315750+96.052750 = 100% LGWI"),
    ]
    for i, (a, b) in enumerate(rows, 1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 100
    print("Assumptions updated")


def populate_title(wb):
    ws = wb["Title "]
    # Clear and rebuild with tract summaries
    tract_sections = [
        ("TRACT 1:", "NE/4; Crook #1-32 wellbore; Morrow-Springer 16,210-20,451 ft", "160 ac nominal"),
        ("TRACT 2:", "E/2 SE/4; Pearl #1-32 wellbore 11,100-19,480 ft", "80 ac nominal"),
        ("TRACT 3:", "N/2 Section 32 below top Hunton", "320 ac nominal"),
        ("TRACT 4:", "Deep leasehold — listed OGLs top Morrow to 19,480 ft", "Multiple aliquots"),
        ("TRACT 5:", "All Section 32 — Diversified-affiliated branch", "640 ac nominal (index extent)"),
    ]
    interests = [
        ("Historical Leede/participant WI (Crook)", "Formation/wellbore limited", "502/267; 509/220", "HISTORICAL ONLY", "98%", "Payout/reassignments OPEN", "Do not carry forward"),
        ("Historical Pearl borehole WI", "E/2 SE/4 wellbore limited", "845/47; 872/279", "HISTORICAL ONLY", "98%", "Successor chain OPEN", "Do not carry forward"),
        ("Historical N/2 Hunton leasehold", "Below top Hunton", "1014/75", "HISTORICAL ONLY", "99% text/20% effect", "Source-title defect OPEN", "Consolidation WI source unproved"),
        ("Union ORRI burdens (Crook/Pearl/deep)", "Wellbore/depth specific", "502/267; 845/47; 872/279", "BURDEN — OPEN", "95% existence", "Survival OPEN", "Carry as open burden"),
        ("Employee ORRI (Crook)", "2.25% aggregate historical", "509/176", "BURDEN — OPEN", "99%", "Survival OPEN", "Carry as open burden"),
        ("PRF / TCW burdens", "LGWI and .5% Leede NRI ORRI", "903/50; 987/159", "BURDEN — OPEN", "98%", "Payout OPEN", "Carry as open burden"),
        ("Diversified Production LLC (principal branch)", "OPEN ITEM", "2371/470; 2395/415; 2400/551", "OPEN ITEM", "90% acquisition/0% decimal", "Schedules absent", "No decimal stated"),
        ("Diversified ABS VI Upstream LLC", "OPEN ITEM", "2395/415", "OPEN ITEM", "99% metadata/0% decimal", "Estate/fraction OPEN", "Provisional branch holder"),
        ("DP Sooner HoldCo LLC / Diversified Production LLC", "OPEN ITEM", "2395/415; 2400/551", "OPEN ITEM", "99% metadata/25% effect", "2400/551 schedule absent", "Branch reconciliation OPEN"),
        ("DP Legacy Central LLC (Burlington branch)", "OPEN ITEM", "2393/1", "OPEN ITEM", "90% metadata", "Separate branch", "Overlap with principal OPEN"),
        ("DP Ponies LLC (Canvas branch)", "OPEN ITEM", "2476/1; 2476/121", "OPEN ITEM", "95% metadata", "Separate branch", "Overlap with principal OPEN"),
        ("MNR ABS Issuer I WB (FourPoint branch)", "OPEN ITEM", "2415/328", "OPEN ITEM", "92% metadata", "Diversified attribution OPEN", "Reconcile to Tapstone/KL CHK"),
        ("Fee mineral / lessor royalty / NPRI", "NOT ESTABLISHED", "Reviewed packet", "NOT ESTABLISHED", "85% reviewed/0% negative", "Complete fee chain required", "No negative opinion"),
        ("Surface estate (assessor context)", "NOT IN SCOPE", "Plat Map only", "CONTEXT ONLY", "N/A", "Not mineral title", "Ancillary only"),
    ]
    # Write tract headers starting row 4
    row = 4
    for tract_hdr, containing, acres in tract_sections:
        ws.cell(row, 2, tract_hdr)
        ws.cell(row, 3, containing)
        row += 1
        ws.cell(row, 2, "Containing:")
        ws.cell(row, 3, f"{acres}; depth/wellbore limits apply")
        row += 2
        ws.cell(row, 2, "Interest Holder / Category")
        ws.cell(row, 3, "Net Acres")
        ws.cell(row, 4, "OGL No.")
        ws.cell(row, 5, "Royalty / Burden")
        ws.cell(row, 6, "Expiration")
        ws.cell(row, 7, "Evidence / Status")
        row += 1
        # Add 2-3 relevant interests per tract
        if "TRACT 1" in tract_hdr:
            items = [interests[0], interests[3], interests[4], interests[6]]
        elif "TRACT 2" in tract_hdr:
            items = [interests[1], interests[3], interests[6]]
        elif "TRACT 3" in tract_hdr:
            items = [interests[2], interests[5], interests[6]]
        elif "TRACT 4" in tract_hdr:
            items = [interests[1], interests[3], interests[5]]
        else:
            items = [interests[6], interests[7], interests[8], interests[9], interests[10], interests[11]]
        for item in items:
            for c, v in enumerate(item, 2):
                ws.cell(row, c, v)
            row += 1
        ws.cell(row, 2, "REPORT TOTAL")
        ws.cell(row, 3, "Totals not forced — chain incomplete")
        ws.cell(row, 7, "OPEN ITEM")
        row += 3
    print("Title sheet populated with 5 tract sections")


def main():
    import shutil
    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)

    populate_runsheet(wb)
    for sheet_name, tract_data in TRACT_INSTRUMENTS.items():
        populate_tract(wb, sheet_name, tract_data)
    populate_ogl(wb)
    populate_wells(wb)
    populate_wi_sheets(wb)
    populate_title(wb)
    populate_overview(wb)
    populate_assumptions(wb)

    wb.save(OUT)
    print(f"\nSaved perfect report: {OUT}")


if __name__ == "__main__":
    main()
