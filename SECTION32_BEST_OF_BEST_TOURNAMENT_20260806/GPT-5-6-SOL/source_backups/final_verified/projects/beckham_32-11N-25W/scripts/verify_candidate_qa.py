#!/usr/bin/env python3
"""Independent QA verification for the 32-11N-25W Diversified cursory report workbook.

Reproduces the QA suite described in FINAL_READINESS_STATEMENT_2026-07-11 and the
BEST_AVAILABLE QA JSONs, from scratch, against a local workbook copy:

  - sheet name/order check against the 13-sheet template layout
  - stale-string contamination scan (prior-project residue terms)
  - negative numeric constants
  - formula census + formula error tokens
  - external links / broken defined names
  - duplicate Runsheet book/page detection
  - Runsheet hyperlink count
  - evidence-tier census of Runsheet rows

Every check reports observed values; nothing is asserted that is not measured.
Usage: python3 verify_candidate_qa.py <workbook.xlsx> <output.json>
"""
import json
import re
import sys
import zipfile
from datetime import date

import openpyxl

EXPECTED_SHEETS = [
    "Overview", "Title ", "OGL", "PLAT", "Runsheet",
    "Tract 1", "Tract 2", "Tract 3", "Tract 4", "Tract 5",
    "WI 2", "WI 1", "Well 1",
]

# Stale terms from the readiness statement scan plus prior-project residue
# terms named in READ_ME_Diversified_Report_Audit_and_Completion_2026-07-10.
STALE_TERMS = [
    "448.333333", "377.53845154", "2237/381", "2572/490", "2610/574",
    "Section 27", "15N-24W", "27-15N-24W", "Roger Mills", "Presidio WAB",
    "White Sail", "Greenhead",
]

FORMULA_ERROR_TOKENS = ["#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!"]


def cell_text(c):
    v = c.value
    return v if isinstance(v, str) else (str(v) if v is not None else "")


def main(path, out_path):
    wb = openpyxl.load_workbook(path, data_only=False)
    results = {"workbook": path.split("/")[-1], "run_date": date.today().isoformat()}

    results["sheet_names"] = wb.sheetnames
    results["sheet_order_ok"] = wb.sheetnames == EXPECTED_SHEETS
    results["sheet_states"] = {ws.title: ws.sheet_state for ws in wb.worksheets}

    stale_hits, negatives, formula_cells, error_hits = [], [], 0, []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                if isinstance(v, str):
                    if v.startswith("="):
                        formula_cells += 1
                        for tok in FORMULA_ERROR_TOKENS:
                            if tok in v:
                                error_hits.append(f"{ws.title}!{c.coordinate}:{tok}")
                    for term in STALE_TERMS:
                        if term.lower() in v.lower():
                            stale_hits.append(f"{ws.title}!{c.coordinate}:{term}")
                    for tok in FORMULA_ERROR_TOKENS:
                        if not v.startswith("=") and tok in v:
                            error_hits.append(f"{ws.title}!{c.coordinate}:{tok}(literal)")
                elif isinstance(v, (int, float)) and v < 0:
                    negatives.append(f"{ws.title}!{c.coordinate}={v}")

    results["stale_term_hits"] = stale_hits
    results["negative_constants"] = negatives
    results["formula_cells"] = formula_cells
    results["formula_error_tokens"] = error_hits

    dn = wb.defined_names
    names = list(dn.keys()) if hasattr(dn, "keys") else [n.name for n in dn.definedName]
    broken = []
    for n in names:
        obj = dn[n] if hasattr(dn, "__getitem__") else None
        val = getattr(obj, "value", "") or ""
        if "#REF!" in val:
            broken.append(n)
    results["defined_names_total"] = len(names)
    results["defined_names_broken"] = broken

    with zipfile.ZipFile(path) as z:
        ext = [n for n in z.namelist() if "externalLink" in n]
    results["external_link_parts"] = ext

    rs = wb["Runsheet"] if "Runsheet" in wb.sheetnames else None
    if rs is not None:
        hyperlinks = sum(1 for row in rs.iter_rows() for c in row if c.hyperlink is not None)
        results["runsheet_hyperlinks"] = hyperlinks
        results["runsheet_dimensions"] = rs.dimensions

        bookpage = {}
        bp_re = re.compile(r"^\s*(\d{2,4})\s*/\s*(\d{1,4})")
        for row in rs.iter_rows():
            joined = " | ".join(cell_text(c) for c in row[:6] if cell_text(c))
            m = bp_re.search(joined)
            if m:
                key = f"{m.group(1)}/{m.group(2)}"
                bookpage.setdefault(key, []).append(row[0].row)
        dupes = {k: v for k, v in bookpage.items() if len(v) > 1}
        results["runsheet_bookpage_rows"] = len(bookpage)
        results["runsheet_duplicate_bookpage"] = dupes

        tiers = {"DIRECT": 0, "INDEX": 0, "PUBLIC METADATA": 0, "OPEN": 0}
        for row in rs.iter_rows():
            rowtext = " ".join(cell_text(c).upper() for c in row)
            for t in tiers:
                if t in rowtext:
                    tiers[t] += 1
        results["runsheet_tier_keyword_row_counts"] = tiers

    results["verdict_notes"] = [
        "This scan measures workbook mechanics only. It does not and cannot",
        "validate title conclusions; ownership remains OPEN per the readiness statement.",
    ]

    with open(out_path, "w") as f:
        json.dump(results, f, indent=2)
    print(json.dumps(results, indent=2))


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
