#!/usr/bin/env python3
"""Build Section 32-11N-25W NE/4 (Tract 1) ownership reconstruction workbook.
Clones the authoritative Template.xlsx (Section 27-15N-24W example) to inherit
all formatting, styles, merges and formulas, then rewrites the sheets with
independently-derived Section 32 NE/4 content. No data is copied from any
competitor workbook; all facts trace to the county index (SECTION_32_INDEX_
MASTER), the direct-image runsheet, and the Horizon findings brief.
"""
import copy, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

S="/tmp/claude-0/-home-user-DataBoss/14de7032-a11e-5ecb-9b29-220574151d44/scratchpad"
wb = load_workbook(f"{S}/drive/Template.xlsx")

thin = Side(style="thin", color="999999")
box = Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
sub_fill = PatternFill("solid", fgColor="D6DCE4")
open_fill = PatternFill("solid", fgColor="FCE4D6")
wrap = Alignment(wrap_text=True, vertical="top")
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

def clear_sheet(ws, keep_dims=True):
    for row in ws.iter_rows():
        for c in row:
            c.value=None
# ---------------------------------------------------------------- helper
def setrow(ws, r, values, start=1):
    for i,v in enumerate(values):
        ws.cell(row=r, column=start+i, value=v)

D = datetime.datetime
# ======================================================================
# 1. OVERVIEW
# ======================================================================
ws = wb["Overview"]
# preserve template layout cells; overwrite the identifying values
ws["Q3"]="32-11N-25W / Diversified — TRACT 1 (NE/4)"
ws["H5"]=32; ws["R5"]="11N"; ws["AC5"]="25W"
ws["H7"]="Beckham"; ws["X7"]="Oklahoma"
ws["B9"]="Tract 1 = NE/4, containing 160.00 acres, more or less (of the 640-acre Section 32)"
ws["AB15"]="TRACT 1 — NE/4 (160 ac) [THIS REPORT]"
ws["U22"]="T2: E/2 SE/4 etc."; ws["E26"]="T4: SW/4"; ws["M26"]="T3: NW/4"; ws["X34"]="T5: W/2 SE/4"
# status block (reuse lower area)
ws["I48"]="CURSORY REPORT DATED: "; ws["S48"]=D(2026,7,18); ws["Y48"]="LAST ENTRY:"; ws["AC48"]="2485/685"
ws["G48"]="x"
ws["B52"]="REMARKS:"
remark=("SOURCE-LIMITED CURSORY — HOLD_NO_RELEASE. Scope: NE/4 (Tract 1), 160 ac. "
 "FEE MINERALS OPEN: the NE/4 mineral estate fragmented before 1930 through The Heller Co., "
 "La Fortune/Seidenbach, Sanditen, Garrett (1/80 deeds Bk 89/190-195) and Crook branches and cannot be "
 "reduced to a proven current decimal from the supplied Patent-1988 image run plus post-1988 index metadata. "
 "Current assessor/surface owners (Hanks; Crook, Ola Mae) are SURFACE/CURRENT-OWNER LEADS ONLY. "
 "LEASEHOLD/WI: Diversified Production LLC is the record-derived leasehold claimant on lease OK48147.001.1 "
 "(Schedule E 2395/462; Schedule B 2400/566), but exact WI/NRI/net acres are NOT DETERMINED. Base OGL behind "
 "OK48147 is UNPROVEN — 1697/236 is an Assignment (Staghorn->CHK), not the base lease; the historic NE/4 base "
 "leases are the Union Oil leases (Cook/Crook 264/345; Randel 352/719 & 352/721; Great Western 376/369). "
 "Do NOT treat OCM Denali 48.75% (2371/495) as Diversified WI. Carlson 32-11-25 7H wellbore carve "
 "(2183/174; 2185/639) unresolved. Index stamped Page 48 missing; Page 54 duplicate.")
ws["F52"]=remark
ws["F52"].alignment=wrap
ws["B56"]="EXAMINER:"; ws["F56"]="AI Tournament Competitor — Tract 1 (NE/4)"

# ======================================================================
# 2. RUNSHEET  (NE/4-affecting instruments, chronological)
# ======================================================================
ws = wb["Runsheet"]
clear_sheet(ws)
setrow(ws,1,["","OGL","Doc Type","Bk/Pg.","Effective Date","Recorded Date","Grantor","Grantee","Legal Description (NE/4 scope)","Notes / Evidence Class"])
for c in range(1,11):
    cell=ws.cell(row=1,column=c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=ctr; cell.border=box

# (ogl#, type, bk/pg, eff, rec, grantor, grantee, legal, notes)
RS = [
 ("","PAT","6/275", D(1910,5,23), D(1910,12,16),"United States of America","(patentees of record)","NE/4 (see note)","Patent era. Supplied image run earliest NE/4 instruments run through the Martin Land & Inv./Benton/Farrar and Green A. Crooks branches (Bk 3-33). ARTI. INDEX/IMAGE."),
 ("","M.D.","89/190-195", None, None,"Cora B. Garrett & S.A. Garrett","Fowler; Andrews; Kuykendall; Gertrude Garrett; Pauline Roberts; Carl Garrett","NE/4 (all QQ) — 1/80 int each","Six mineral deeds, undivided 1/80 each, seed the persistent Garrett fractional-mineral cluster under the NE/4. INDEX."),
 ("","M.D.","89/85-98 series", None, None,"The Heller Company","Heller heirs; La Fortune; Seidenbach; May/Langer; Sanditen; Jacobs; Lockwood; Murray; Holbrook; Smith; Harrison","NE/4 & NW/4 fractional QQ","Heller Co. broke the NE/4 mineral estate into many small fractional (1/8, 1/10, 1/80) mineral deeds pre-1930. Chain fragmented; OPEN. INDEX."),
 ("","WD","90/545 / 89/306,311,536", None, None,"Crook (Everett/John/others)","Charlie B. Crook & wife","NE/4 / ALL (240 ac)","Charlie B. Crook consolidated a large surface + partial mineral position across the NE/4 by the 1940s. INDEX."),
 (1,"OL","264/271", None, None,"Shearer, Jean Jenkins","Union Oil Company of California","NE/4 (all)","Union Oil leasing program, NE/4. Historic base-lease candidate. INDEX."),
 (2,"OL","264/345", D(1972,2,5), None,"Charles R. Cook (Crook) et ux","Union Oil Company of California","NE/NE, NW, SE (Crook lease)","THE CROOK LEASE. Covers SW/4 NW/4 & NW/4 SW/4 Sec 28 and NE/4 Sec 32; held by production from No. 1 Bruner Unit. Released as to Leede at 395/618 (1979). DIRECT IMAGE 1575."),
 (3,"OL","270/72", None, None,"L.S. & J.L. Randell","Union Oil Company of California","NE/4 (all); SE/SE","Union Oil lease. Later ROL to Randell at 270/553. INDEX."),
 (4,"OL","341/714", None, None,"Shearer, Jean Watkins","Leede Exploration","NE/4 (all)","Leede leasing, NE/4. INDEX."),
 (5,"OL","352/719", D(1978,5,16), None,"F.L. Randel","Union Oil Company of California","NE/4 (all)","Union Oil base-lease candidate (NE/4). Referenced on 872/279 deep-rights exhibit. INDEX/IMAGE."),
 (6,"OL","352/721", D(1978,5,16), None,"J.L. Randel","Union Oil Company of California","NE/4 (all)","Union Oil base-lease candidate (NE/4). INDEX/IMAGE."),
 (7,"OL","376/369", D(1979,3,21), None,"Great Western Oil & Gas, Inc.","Union Oil Company of California","NE/4 (all)","Union Oil lease; corrected/re-recorded at 1039/20-23. INDEX; OKCOUNTY metadata."),
 ("","REL","395/618", D(1979,11,21), None,"Charlie B. & Ima Pearl Crook","Leede Exploration (released)","NE/4 (Crook lease 264/345)","Crooks own surface + portion of NE/4 minerals; recites Crook lease 264/345 held by production. DIRECT IMAGE 1575."),
 ("","ASG","872/279", D(1985,10,29), D(1986,3,6),"Union Oil Company of California","Leede Exploration","Listed leases incl. NE/4; top Morrow-19,480  ft","Historic deep-rights assignment; the exhibit list is the basis for the modern lease-schedule references. DIRECT IMAGE L-007 (4388-4393)."),
 ("","PA","845/149-150", D(1985,10,24), None,"Estate of Charlie B. Crook (Final Decree)","Devisees per decree","NE/4 Section 32","Probate of Charlie B. Crook; NE/4 devised (undivided fractions). Companion Estate of Ima Pearl Crook at 845/154. DIRECT IMAGE 4121."),
 ("","MD","839/329 / 842/93-94", D(1985,9,30), None,"Ralph W. Viersen Jr. et al; Keathley; Brown","Asher Land & Exploration Co.","NE","Viersen/Keathley NE/4 mineral interests conveyed to Asher Land & Exploration; feeds Walne/Asher exhibit chain (623/149; 506/205). INDEX/IMAGE."),
 (50,"ASG","1697/574 (also 1697/236)", D(2001,6,14), None,"Staghorn Resources","Chesapeake Exploration","Section 32 (long desc.)","Staghorn->CHK assignment. NOTE: 1697/236 is the ASSIGNMENT cited by lease OK48147.001.1 — NOT the base OGL. INDEX ONLY."),
 (58,"ASG","2340/403-433", D(2020,12,14), None,"Chesapeake affiliates","Tapstone Energy LLC","Section 32 (long desc.)","CHK->Tapstone. Schedules absent. PUBLIC METADATA + INDEX."),
 (64,"ASG","2371/470-494", D(2022,1,12), None,"KL CHK SPV LLC","Diversified Production LLC; OCM Denali Holdings LLC","All 16 Q-Q ticks (incl. NE/4)","CHK SPV -> Diversified + OCM Denali (co-assignees). Fractions/schedules absent; do NOT split 51.25/48.75 as WI. PUBLIC METADATA + INDEX."),
 (65,"ASG","2371/495-513", D(2022,1,12), None,"Tapstone Energy; Tapstone Midstream","OCM Denali Holdings LLC","Section 32","Parallel Tapstone->OCM Denali; schedule lists OK48147.001.1 Sec 32 (quantum OPEN). Separate from Diversified WI. PUBLIC METADATA."),
 (75,"MERGER","2389/500-580", D(2022,8,24), None,"Diversified Production; DP Legacy Central; merger entities","DP Legacy Tapstone LLC","All Section 32 ticks","Corporate merger into DP Legacy Tapstone. PUBLIC METADATA + INDEX."),
 (81,"CONV","2395/415-464", D(2022,11,18), None,"DP Legacy Tapstone LLC","Diversified ABS VI Upstream LLC; DP Sooner HoldCo LLC","All 16 Q-Q ticks (incl. NE/4)","I-2022-004838 Confirmation of Division (ER-01 IMAGE). OK48147.001.1 on Exhibit E at 2395/462; interest columns BLANK. DIRECT IMAGE."),
 (86,"CONV","2400/551-567", D(2023,3,2), None,"DP Sooner HoldCo LLC","Diversified Production LLC","Long description (incl. NE/4)","I-2023-000796 Confirmation of Merger (IMAGE). Schedule B includes OK48147.001.1 at 2400/566; legal/interest BLANK. Quantum NOT CALCULABLE. DIRECT IMAGE."),
 (53,"ASG","2183/174", D(2015,2,18), None,"Chesapeake affiliates","Linn Operating","Carlson 32-11-25 7H wellbore (Sec 32)","CHK->Linn Carlson 7H (API 35-009-2192300) wellbore carve; intersection with OK48147/NE/4 leasehold OPEN. PUBLIC METADATA."),
 (54,"CONV","2185/639", D(2015,3,31), None,"Chesapeake affiliates","FourPoint Energy","Carlson 7H wellbore (Sec 32)","Companion Carlson 7H carve. Wellbore/depth-limited. PUBLIC METADATA."),
 ("","TODD","2290/494; 2291/474", D(2018,9,4), None,"Hanks, Darrell Lee","Public (transfer-on-death)","NE","Transfer-on-death affidavits — current N/2 NE/4 surface/current-owner lead (Assessor 050005625; Bk 2290/494). SURFACE LEAD ONLY. INDEX."),
 ("","WD","2111/585", None, None,"(grantor of record)","Crook, Ola Mae","SE/4 NE/4 area","Current surface/current-owner lead for the SE/4 NE/4 portion (Assessor 050005631/050005630). SURFACE LEAD ONLY. INDEX."),
 ("","MD","2485/400; 2485/685", D(2026,5,7), D(2026,5,12),"Gillespie/Jones parties","Marsh, Kelcey H. et al.","NE & SW QQ marks","Most recent NE/4-indexed mineral conveyances (last index entries). Schedules unreviewed. INDEX."),
]
r=2
for row in RS:
    ogl,typ,bp,eff,rec,gtor,gtee,legal,notes=row
    setrow(ws,r,["",ogl,typ,bp,eff,rec,gtor,gtee,legal,notes])
    for c in range(2,11):
        cell=ws.cell(row=r,column=c); cell.border=box; cell.alignment=wrap
        cell.font=Font(size=9)
    if "OPEN" in notes or "NOT" in notes or "BLANK" in notes:
        ws.cell(row=r,column=10).fill=open_fill
    r+=1
ws.column_dimensions['B'].width=6
ws.column_dimensions['C'].width=12
ws.column_dimensions['D'].width=16
ws.column_dimensions['E'].width=12
ws.column_dimensions['F'].width=12
ws.column_dimensions['G'].width=26
ws.column_dimensions['H'].width=28
ws.column_dimensions['I'].width=24
ws.column_dimensions['J'].width=60

print("Overview + Runsheet done; rows:", r-2)
wb.save(f"{S}/NE4_build_stage1.xlsx")
print("saved stage1")
