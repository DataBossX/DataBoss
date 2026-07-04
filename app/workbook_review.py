"""Golden-Law-compliant workbook writer.

Policy rule #1: original Excel/PDF files are NEVER modified. This module writes
verified results to a NEW `<stem>_REVIEW_<ts>.xlsx` copy and never touches the
source workbook.

NOTE (surfaced for approval): automation/writer.py currently uses
`pd.ExcelWriter(path, mode="w")`, which overwrites the source workbook in place
— a violation of policy rule #1. This module is the compliant alternative; call
sites should migrate to it. That swap is a behavior change to the scraper, so it
is left for Rodney to approve rather than applied silently.
"""
from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any, Dict, List

from tools import guardrails


def review_path(source_path: Path | str, out_dir: Path | str | None = None, ts: str | None = None) -> Path:
    """Compute the _REVIEW_<ts> destination for *source_path* (pure, no I/O)."""
    p = Path(source_path)
    ts = ts or datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out = Path(out_dir) if out_dir else p.parent
    return out / f"{p.stem}_REVIEW_{ts}{p.suffix}"


def write_review_workbook(
    source_path: Path | str,
    updates: List[Dict[str, Any]],
    sheet: str | None = None,
    out_dir: Path | str | None = None,
    ts: str | None = None,
) -> Path:
    """Copy *source_path* to a _REVIEW_ sibling and apply *updates*, never
    overwriting the original. Returns the review path written.

    Requires openpyxl (for .xlsx). Raises PermissionError if asked to target a
    protected root, and refuses to write onto the source.
    """
    src = Path(source_path)
    dest = review_path(src, out_dir=out_dir, ts=ts)
    if dest.resolve() == src.resolve():  # pragma: no cover - defensive
        raise PermissionError("Refusing to overwrite the source workbook")
    if guardrails.is_protected_path(dest):
        raise PermissionError(f"Refusing to write under protected root: {dest}")

    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dep
        raise RuntimeError(
            "openpyxl is required to write .xlsx review copies but is not installed."
        ) from exc

    wb = openpyxl.load_workbook(src)
    ws = wb[sheet] if sheet else wb.active
    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header)}
    for upd in updates:
        row_idx = upd.get("_row")
        if not row_idx:
            continue
        for key, value in upd.items():
            if key == "_row" or key not in col:
                continue
            ws.cell(row=row_idx, column=col[key], value=value)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)  # NEW file — source workbook untouched
    return dest
