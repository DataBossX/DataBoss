"""Tests for the footing repair + title logic core (no network, no fixtures)."""
from openpyxl import Workbook

from title_report_builder.report import footing as F
from title_report_builder.report.title_logic import (
    normalize_aliquot, normalize_bookpage, Instrument, assign_tracts, dedupe,
)


def _make_tract_wb(d5=80.0):
    """Synthetic tract grid: D5 acres, owners in D10.., E = net interest,
    C/F carry the (wrong) old formula we expect to standardize."""
    wb = Workbook()
    ws = wb.active
    ws["D5"] = d5
    owners = [("The Public", -1.0), ("Owner A", 1.0), ("Owner B", 3.0), ("Owner C", 2.0)]
    for i, (nm, e) in enumerate(owners):
        r = 10 + i
        ws[f"D{r}"] = nm
        ws[f"E{r}"] = e
        ws[f"C{r}"] = "=OLD()"   # placeholder formula to be overwritten
        ws[f"F{r}"] = "=OLD()"
    return wb


def test_footing_foots_to_d5_excluding_public():
    wb = _make_tract_wb(80.0)
    ws = wb.active
    ft = F.repair_tract(ws, ws)  # same sheet provides values here
    # positive non-Public interest = 1+3+2 = 6 -> allocation foots to D5 exactly
    assert ft.foots is True
    assert abs(ft.exact_foot - 80.0) < 1e-9
    assert ft.positive_owners == 3
    # the rewritten formula must exclude The Public and normalize to D5
    assert 'The Public' in ws["C11"].value and 'SUMIFS' in ws["C11"].value


def test_exact_fractions_replaces_rounded_decimals():
    wb = Workbook()
    ws = wb.active
    ws["I9"] = "=0.0769230769230769"   # == 1/13
    ws["J9"] = "=0.5"                   # already exact -> skipped (==1/2? len check)
    ws["K9"] = "=1"                     # whole -> skipped
    n = F.exact_fractions(ws)
    assert ws["I9"].value == "=1/13"
    assert n >= 1


def test_normalize_aliquot():
    assert normalize_aliquot("SW/4 of NE/4") == "SW NE"
    assert normalize_aliquot("ne ne") == "NE NE"
    assert normalize_aliquot("lot 3") is None


def test_normalize_bookpage():
    assert normalize_bookpage("001608 / 0038") == "1608/38"
    assert normalize_bookpage("D22/248") == "22/248"
    assert normalize_bookpage(None) == ""


def test_assign_tracts_and_dedupe():
    a2t = {"SW NE": 2, "NW SE": 2, "NE NE": 1}
    recs = [
        {"grantor": "X", "grantee": "Y", "book_page": "100/1", "aliquots": ["SW NE", "NW SE"]},
        {"grantor": "X", "grantee": "Y", "book_page": "100/1", "aliquots": ["NE NE"]},  # dupe
    ]
    insts = [Instrument.from_record(r) for r in recs]
    assign_tracts(insts, a2t)
    out = dedupe(insts)
    assert len(out) == 1
    assert set(out[0].tracts) == {1, 2}
