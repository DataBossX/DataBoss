"""Excel extraction via openpyxl (formulas + cached values)."""
from __future__ import annotations
from openpyxl import load_workbook


def read_workbook_values(path: str) -> dict[str, list[list]]:
    """Return {sheet_name: 2D list of cached cell values}."""
    wb = load_workbook(path, data_only=True)
    out = {}
    for ws in wb.worksheets:
        out[ws.title] = [[c.value for c in row] for row in ws.iter_rows()]
    return out


def sheet_dimensions(path: str) -> dict[str, tuple[int, int]]:
    wb = load_workbook(path, read_only=True)
    return {ws.title: (ws.max_row, ws.max_column) for ws in wb.worksheets}
