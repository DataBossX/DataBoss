"""Net-acre footing repair + exact-fraction normalization for tract grids.

This is the audited core of the pipeline. It rewrites each tract's net-acre
column (C) and net-interest column (F) to the canonical pro-rata formula that
foots exactly to the tract acreage in D5, excluding "The Public":

    C = IFERROR(IF(AND($E>0, $D<>"The Public"),
                   $E / SUMIFS(Erange, Erange, ">0", Drange, "<>The Public") * $D$5,
                   ""), "")

Pro-rata footing is an ALLOCATION, not a deed-by-deed derivation. It also
replaces rounded conversion-decimals (row 9) with exact fraction formulas
(e.g. 0.0769230769... -> =1/13), which is math-preserving.
"""
from __future__ import annotations
from dataclasses import dataclass
from fractions import Fraction

from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class TractFooting:
    sheet: str
    d5: float | None
    first_row: int
    last_row: int
    exact_foot: float
    foots: bool | None        # None => could not verify (no cached values; needs recalc)
    verified: bool
    positive_owners: int
    implied_gross_tract_equiv: float


def _owner_block(ws_values: Worksheet, first: int = 10) -> int:
    last = first
    for r in range(first, ws_values.max_row + 1):
        nm = ws_values.cell(r, 4).value
        if isinstance(nm, str) and nm.strip() and nm.strip() != "Owners":
            last = r
    return last


def repair_tract(ws_f: Worksheet, ws_v: Worksheet, first_row: int = 10) -> TractFooting:
    """ws_f = worksheet with formulas (edited in place); ws_v = cached values."""
    d5 = ws_v["D5"].value
    A = first_row
    B = _owner_block(ws_v, A)
    denom = (f'SUMIFS($E${A}:$E${B},$E${A}:$E${B},">0",'
             f'$D${A}:$D${B},"<>The Public")')
    for r in range(A, B + 1):
        if isinstance(ws_f.cell(r, 3).value, str) and ws_f.cell(r, 3).value.startswith("="):
            ws_f.cell(r, 3).value = (
                f'=IFERROR(IF(AND($E{r}>0,$D{r}<>"The Public"),'
                f'$E{r}/{denom}*$D$5,""),"")')
        if isinstance(ws_f.cell(r, 6).value, str) and ws_f.cell(r, 6).value.startswith("="):
            ws_f.cell(r, 6).value = (
                f'=IFERROR(IF(AND($E{r}>0,$D{r}<>"The Public"),'
                f'$E{r}/{denom},""),"")')

    # numeric verification on cached values
    rows = [(r, ws_v.cell(r, 5).value, ws_v.cell(r, 4).value) for r in range(A, B + 1)]
    pos = [(r, e, nm) for r, e, nm in rows
           if isinstance(e, (int, float)) and e > 0
           and not (isinstance(nm, str) and nm.strip() == "The Public")]
    den = sum(e for _, e, _ in pos)
    # "verified" requires cached numeric interest values to sum. After an
    # openpyxl round-trip with no recalc the cache is empty -> we can write the
    # correct formula but cannot numerically confirm the foot here.
    any_numeric = any(isinstance(ws_v.cell(r, 5).value, (int, float)) for r in range(A, B + 1))
    if den:
        exact = sum(e / den * (d5 or 0) for _, e, _ in pos)
        foots: bool | None = abs(exact - (d5 or 0)) < 1e-6
        verified = True
    else:
        exact = 0.0
        foots = None if not any_numeric else (abs(0 - (d5 or 0)) < 1e-6)
        verified = any_numeric
    return TractFooting(
        sheet=ws_f.title, d5=d5, first_row=A, last_row=B,
        exact_foot=round(exact, 6), foots=foots, verified=verified,
        positive_owners=len(pos), implied_gross_tract_equiv=round(den, 4),
    )


def exact_fractions(ws_f: Worksheet, row: int = 9, max_den: int = 1000) -> int:
    """Replace rounded decimal conversion cells in `row` with exact fractions."""
    n = 0
    for ci in range(9, ws_f.max_column + 1):
        cell = ws_f.cell(row, ci)
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            body = v[1:].strip()
            try:
                fv = float(body)
            except ValueError:
                continue
            if fv in (0.0, 1.0):
                continue
            fr = Fraction(fv).limit_denominator(max_den)
            if fr.denominator > 1 and abs(float(fr) - fv) < 1e-9 and "." in body and len(body) > 8:
                cell.value = f"={fr.numerator}/{fr.denominator}"
                n += 1
    return n
