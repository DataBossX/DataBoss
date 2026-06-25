"""Formatting: extract a style map from a format-only template, and provide a
clean professional fallback theme. Template DATA is never copied — only visual
attributes (fonts, fills, borders, widths, number formats, freeze panes)."""
from __future__ import annotations
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---- professional fallback theme (used when no template is supplied) ----
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=10, bold=True)
NORMAL = Font(name="Calibri", size=10)
LINK = Font(name="Calibri", size=10, color="0563C1", underline="single")
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")
FILL_SUB = PatternFill("solid", fgColor="D6E4F0")
FILL_FLAG = PatternFill("solid", fgColor="FFF2CC")
_THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


@dataclass
class StyleMap:
    """Visual-only fingerprint of a template (no business data)."""
    column_widths: dict = field(default_factory=dict)   # sheet -> {col_letter: width}
    row_heights: dict = field(default_factory=dict)      # sheet -> {row: height}
    freeze_panes: dict = field(default_factory=dict)     # sheet -> "A2"
    number_formats: dict = field(default_factory=dict)   # sheet -> {coord: fmt}
    sheet_order: list = field(default_factory=list)
    tab_colors: dict = field(default_factory=dict)


def extract_style_map(template_path: str) -> StyleMap:
    wb = load_workbook(template_path)
    sm = StyleMap(sheet_order=list(wb.sheetnames))
    for ws in wb.worksheets:
        sm.column_widths[ws.title] = {
            k: v.width for k, v in ws.column_dimensions.items() if v.width}
        sm.row_heights[ws.title] = {
            k: v.height for k, v in ws.row_dimensions.items() if v.height}
        if ws.freeze_panes:
            sm.freeze_panes[ws.title] = ws.freeze_panes
        if ws.sheet_properties.tabColor:
            sm.tab_colors[ws.title] = str(ws.sheet_properties.tabColor.rgb)
        fmts = {}
        for row in ws.iter_rows():
            for c in row:
                if c.number_format and c.number_format != "General":
                    fmts[c.coordinate] = c.number_format
        sm.number_formats[ws.title] = fmts
    return sm


def style_header_row(ws, row: int, ncols: int) -> None:
    for j in range(1, ncols + 1):
        c = ws.cell(row, j)
        c.font = HEADER_FONT
        c.fill = FILL_HEADER
        c.border = BORDER
        c.alignment = CENTER


def banner(ws, text: str, ncols: int, row: int = 1) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = TITLE_FONT
    c.fill = FILL_TITLE
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 26
