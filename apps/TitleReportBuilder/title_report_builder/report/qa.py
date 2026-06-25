"""Quality gates: formula-error scan, footing checks, link-target validation."""
from __future__ import annotations
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
_ERR_RE = re.compile(r"#(REF|DIV/0|VALUE|NAME|N/A|NULL|NUM)!")


@dataclass
class QAReport:
    cached_errors: int = 0
    error_samples: list = field(default_factory=list)
    formula_error_literals: int = 0
    broken_link_targets: list = field(default_factory=list)
    foots: dict = field(default_factory=dict)

    @property
    def passed(self) -> bool:
        return (self.cached_errors == 0
                and self.formula_error_literals == 0
                and not self.broken_link_targets
                and all(self.foots.values()) if self.foots else True)


def scan(path: str) -> QAReport:
    rep = QAReport()
    wb_v = load_workbook(path, data_only=True)
    for ws in wb_v.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip() in ERROR_TOKENS:
                    rep.cached_errors += 1
                    if len(rep.error_samples) < 10:
                        rep.error_samples.append(f"{ws.title}!{c.coordinate}={c.value}")
    wb_f = load_workbook(path, data_only=False)
    names = set(wb_f.sheetnames)
    for ws in wb_f.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("=") and _ERR_RE.search(v):
                    rep.formula_error_literals += 1
                if c.hyperlink and c.hyperlink.target and str(c.hyperlink.target).startswith("#"):
                    m = re.match(r"#'?([^'!]+)'?!", c.hyperlink.target)
                    if m and m.group(1) not in names:
                        rep.broken_link_targets.append(f"{ws.title}!{c.coordinate}->{c.hyperlink.target}")
    return rep
