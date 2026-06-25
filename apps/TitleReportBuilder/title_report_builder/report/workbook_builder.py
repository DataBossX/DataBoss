"""Assemble report sheets: Index Evidence, Chain of Title, Summary punch-list.
Applies footing repair to existing tract grids and wires working links."""
from __future__ import annotations
from openpyxl import load_workbook

from . import footing as F
from . import formatting as S
from .title_logic import Instrument, assign_tracts, dedupe, chains_by_tract


def okc_link(doc: str, text: str) -> str:
    doc = (doc or "").strip()
    if doc:
        return f'=HYPERLINK("https://okcountyrecords.com/detail/roger+mills/{doc}","{text}")'
    return text


def repair_all_tracts(wb_path: str, tract_sheets: list[str], out_path: str) -> dict:
    """Standardize footing + exact fractions on each tract sheet; save."""
    wb_f = load_workbook(wb_path)
    wb_v = load_workbook(wb_path, data_only=True)
    results = {}
    frac = 0
    for sh in tract_sheets:
        ft = F.repair_tract(wb_f[sh], wb_v[sh])
        frac += F.exact_fractions(wb_f[sh])
        results[sh] = ft.__dict__
    wb_f.save(out_path)
    return {"tracts": results, "fractions_replaced": frac}


def add_index_evidence(wb_path: str, instruments: list[Instrument], out_path: str,
                       after_sheet: str = "Runsheet") -> int:
    wb = load_workbook(wb_path)
    if "Index Evidence" in wb.sheetnames:
        del wb["Index Evidence"]
    idx = wb.sheetnames.index(after_sheet) + 1 if after_sheet in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet("Index Evidence", idx)
    S.banner(ws, "INDEX EVIDENCE — vision-verified recorded instruments", 10)
    hdr = ["Date", "Doc #", "Type", "Grantor", "Grantee", "Book/Page",
           "Tract(s)", "Aliquots", "Source", "Go to tract"]
    for j, h in enumerate(hdr, 1):
        ws.cell(3, j, h)
    S.style_header_row(ws, 3, len(hdr))
    r = 4
    for it in instruments:
        sw = len(it.tracts) >= 8
        ws.cell(r, 1, it.date); ws.cell(r, 2, it.doc_number); ws.cell(r, 3, it.doc_type)
        ws.cell(r, 4, it.grantor).alignment = S.WRAP
        ws.cell(r, 5, it.grantee).alignment = S.WRAP
        ws.cell(r, 6, it.book_page)
        ws.cell(r, 7, "ENTIRE SECTION" if sw else ", ".join(map(str, it.tracts)))
        ws.cell(r, 8, "All 16" if sw else "; ".join(it.aliquots[:8])).alignment = S.WRAP
        src = ws.cell(r, 9, okc_link(it.doc_number, f"OKCR {it.doc_number or it.book_page}"))
        if str(src.value).startswith("=HYPERLINK"):
            src.font = S.LINK
        if it.tracts:
            nav = ws.cell(r, 10, f"Tract {it.tracts[0]}")
            nav.hyperlink = f"#'Tract {it.tracts[0]}'!A1"
            nav.font = S.LINK
        for j in range(1, 11):
            ws.cell(r, j).border = S.BORDER
        r += 1
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J{r - 1}"
    for col, w in zip("ABCDEFGHIJ", [11, 9, 9, 30, 30, 12, 13, 22, 20, 11]):
        ws.column_dimensions[col].width = w
    wb.save(out_path)
    return r - 4
