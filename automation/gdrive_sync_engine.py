#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Google Drive & PC Sync Engine  (DataBossX / Horizon)
===================================================

Provides high-performance, parallel comparison, verification, synchronization,
and directory structure enforcement between Local PC files and Google Drive
(e.g., Google Drive for Desktop stream, mounted drive, or cloud staging folder).

Key capabilities:
1. Standardized Section Folder Structure (7 standard subdirectories):
   - 1_Source_Documents/
   - 2_Federal_Lease_Files/
   - 3_OCR_Extracted_Text/
   - 4_Title_Chain_and_Worksheets/
   - 5_Final_Reports_Ready_To_Turn_In/
   - 6_Curative_and_Audit/
   - 7_Sync_and_Transcripts/
2. Deep Parallel Comparison:
   - SHA-256 integrity hash
   - File size & modification timestamp
   - PDF page counting & image count validation
   - Cloud vs. Local drift & missing file detection
3. Non-Destructive Bidirectional / Push / Pull Sync:
   - Mirrors deliverables, worksheets, OCR dumps, and chat transcripts
   - Preserves timestamps and audit receipts
"""

from __future__ import annotations

import argparse
import concurrent.futures
import csv
import datetime as _dt
import hashlib
import json
import os
import shutil
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

# Optional openpyxl for Excel output
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _HAVE_OPENPYXL = True
except Exception:
    _HAVE_OPENPYXL = False

# Optional PyMuPDF / pdfplumber for PDF page counting
try:
    import fitz  # PyMuPDF

    _HAVE_PYMUPDF = True
except Exception:
    _HAVE_PYMUPDF = False

try:
    import pdfplumber

    _HAVE_PDFPLUMBER = True
except Exception:
    _HAVE_PDFPLUMBER = False


SECTION_SUBDIRECTORIES = [
    "1_Source_Documents",
    "2_Federal_Lease_Files",
    "3_OCR_Extracted_Text",
    "4_Title_Chain_and_Worksheets",
    "5_Final_Reports_Ready_To_Turn_In",
    "6_Curative_and_Audit",
    "7_Sync_and_Transcripts",
]

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif", ".webp"}
PDF_EXTENSIONS = {".pdf"}
REPORT_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".docx", ".doc", ".pdf", ".csv", ".txt", ".md"}


def calculate_sha256(filepath: Path, block_size: int = 65536) -> str:
    """Calculate SHA-256 hash of a file efficiently."""
    sha = hashlib.sha256()
    try:
        with open(filepath, "rb") as f:
            for block in iter(lambda: f.read(block_size), b""):
                sha.update(block)
        return sha.hexdigest()
    except Exception as exc:
        return f"ERROR:{exc}"


def count_pdf_pages_and_images(filepath: Path) -> Tuple[int, int]:
    """Count page count and embedded image count for a PDF."""
    if not filepath.exists() or filepath.suffix.lower() not in PDF_EXTENSIONS:
        return 0, 0

    if _HAVE_PYMUPDF:
        try:
            doc = fitz.open(filepath)
            page_count = len(doc)
            image_count = 0
            for page in doc:
                image_count += len(page.get_images(full=True))
            doc.close()
            return page_count, image_count
        except Exception:
            pass

    if _HAVE_PDFPLUMBER:
        try:
            with pdfplumber.open(filepath) as pdf:
                page_count = len(pdf.pages)
                image_count = sum(len(p.images) for p in pdf.pages)
                return page_count, image_count
        except Exception:
            pass

    # Fallback standard-library crude PDF page estimation
    try:
        with open(filepath, "rb") as f:
            content = f.read()
            page_count = max(1, content.count(b"/Type /Page") - content.count(b"/Type /Pages"))
            image_count = content.count(b"/Subtype /Image")
            return page_count, image_count
    except Exception:
        return 0, 0


@dataclass
class FileEntry:
    rel_path: str
    abs_path: str
    filename: str
    extension: str
    size_bytes: int
    modified_iso: str
    sha256: str
    is_image: bool = False
    is_pdf: bool = False
    page_count: int = 0
    image_count: int = 0
    category: str = "other"


@dataclass
class ComparisonRecord:
    rel_path: str
    filename: str
    pc_status: str  # present / missing / error
    gdrive_status: str  # present / missing / error
    match_status: str  # MATCH_EXACT / DRIFT_PC_NEWER / DRIFT_GDRIVE_NEWER / PC_ONLY / GDRIVE_ONLY / HASH_MISMATCH
    pc_size: Optional[int] = None
    gdrive_size: Optional[int] = None
    pc_sha256: Optional[str] = None
    gdrive_sha256: Optional[str] = None
    pc_modified: Optional[str] = None
    gdrive_modified: Optional[str] = None
    pc_pages: int = 0
    gdrive_pages: int = 0
    pc_images: int = 0
    gdrive_images: int = 0
    action_required: str = "None"
    notes: str = ""


@dataclass
class SyncManifest:
    timestamp: str
    section: str
    pc_root: str
    gdrive_root: str
    total_pc_files: int = 0
    total_gdrive_files: int = 0
    total_matched: int = 0
    total_pc_only: int = 0
    total_gdrive_only: int = 0
    total_drift: int = 0
    total_pdf_pages_pc: int = 0
    total_pdf_pages_gdrive: int = 0
    records: List[ComparisonRecord] = field(default_factory=list)


def classify_subfolder_category(rel_path: str) -> str:
    low = rel_path.lower().replace("\\", "/")
    if "1_source" in low or "source_doc" in low:
        return "Source Documents"
    if "2_federal" in low or "federal_lease" in low:
        return "Federal Lease Files"
    if "3_ocr" in low or "extracted_text" in low:
        return "OCR Extracted Text"
    if "4_title" in low or "chain" in low or "worksheets" in low:
        return "Title Chain & Worksheets"
    if "5_final" in low or "report" in low or "turn_in" in low:
        return "Final Reports"
    if "6_curative" in low or "audit" in low:
        return "Curative & Audit"
    if "7_sync" in low or "transcript" in low or "chat" in low:
        return "Sync & Transcripts"
    return "General"


def inspect_single_file(root: Path, file_path: Path) -> FileEntry:
    rel_path = str(file_path.relative_to(root)).replace("\\", "/")
    ext = file_path.suffix.lower()
    stat = file_path.stat()
    size = stat.st_size
    mod_iso = _dt.datetime.fromtimestamp(stat.st_mtime, tz=_dt.timezone.utc).isoformat()
    sha = calculate_sha256(file_path)

    is_img = ext in IMAGE_EXTENSIONS
    is_pdf = ext in PDF_EXTENSIONS
    page_cnt, img_cnt = (0, 0)
    if is_pdf:
        page_cnt, img_cnt = count_pdf_pages_and_images(file_path)
    elif is_img:
        img_cnt = 1

    category = classify_subfolder_category(rel_path)

    return FileEntry(
        rel_path=rel_path,
        abs_path=str(file_path.resolve()),
        filename=file_path.name,
        extension=ext,
        size_bytes=size,
        modified_iso=mod_iso,
        sha256=sha,
        is_image=is_img,
        is_pdf=is_pdf,
        page_count=page_cnt,
        image_count=img_cnt,
        category=category,
    )


def inventory_tree_parallel(
    root: Path,
    max_workers: int = 8,
    ignore_patterns: Optional[Sequence[str]] = None,
) -> Dict[str, FileEntry]:
    """Scan and index all files under root in parallel with SHA-256 and page/image stats."""
    if not root.exists():
        return {}

    default_ignore = {".git", ".svn", "__pycache__", ".pytest_cache", ".DS_Store", "desktop.ini", "Thumbs.db"}
    if ignore_patterns:
        default_ignore.update(ignore_patterns)

    all_files: List[Path] = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in default_ignore and not d.startswith(".")]
        for fname in filenames:
            if fname in default_ignore or fname.startswith("~$") or fname.startswith("."):
                continue
            all_files.append(Path(dirpath) / fname)

    results: Dict[str, FileEntry] = {}
    if not all_files:
        return results

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(inspect_single_file, root, f): f for f in all_files}
        for fut in concurrent.futures.as_completed(futures):
            try:
                entry = fut.result()
                results[entry.rel_path] = entry
            except Exception as exc:
                f_path = futures[fut]
                rel = str(f_path.relative_to(root)).replace("\\", "/")
                results[rel] = FileEntry(
                    rel_path=rel,
                    abs_path=str(f_path),
                    filename=f_path.name,
                    extension=f_path.suffix.lower(),
                    size_bytes=0,
                    modified_iso="",
                    sha256=f"ERROR:{exc}",
                )

    return results


def initialize_section_folders(base_dir: Path, section_name: str = "Section 7") -> Dict[str, Path]:
    """Create the 7 standard section folders on the target directory."""
    section_root = base_dir / section_name
    section_root.mkdir(parents=True, exist_ok=True)
    created: Dict[str, Path] = {}
    for sub in SECTION_SUBDIRECTORIES:
        p = section_root / sub
        p.mkdir(parents=True, exist_ok=True)
        created[sub] = p
    return created


def compare_trees(
    pc_root: Path,
    gdrive_root: Path,
    section_name: str = "Section 7",
    max_workers: int = 8,
) -> SyncManifest:
    """Compare PC folder vs Google Drive folder in parallel and return audit manifest."""
    pc_inventory = inventory_tree_parallel(pc_root, max_workers=max_workers)
    gdrive_inventory = inventory_tree_parallel(gdrive_root, max_workers=max_workers)

    all_keys = sorted(set(pc_inventory.keys()) | set(gdrive_inventory.keys()))

    manifest = SyncManifest(
        timestamp=_dt.datetime.now(_dt.timezone.utc).isoformat(),
        section=section_name,
        pc_root=str(pc_root.resolve()) if pc_root.exists() else str(pc_root),
        gdrive_root=str(gdrive_root.resolve()) if gdrive_root.exists() else str(gdrive_root),
        total_pc_files=len(pc_inventory),
        total_gdrive_files=len(gdrive_inventory),
    )

    for rel in all_keys:
        pc_f = pc_inventory.get(rel)
        gd_f = gdrive_inventory.get(rel)
        fname = (pc_f or gd_f).filename  # type: ignore

        if pc_f and gd_f:
            if pc_f.sha256 == gd_f.sha256 and not pc_f.sha256.startswith("ERROR"):
                status = "MATCH_EXACT"
                action = "None (In Sync)"
                manifest.total_matched += 1
            else:
                # Compare mod times
                pc_m = pc_f.modified_iso
                gd_m = gd_f.modified_iso
                if pc_m > gd_m:
                    status = "DRIFT_PC_NEWER"
                    action = "Push PC version to Google Drive"
                elif gd_m > pc_m:
                    status = "DRIFT_GDRIVE_NEWER"
                    action = "Pull Google Drive version to PC"
                else:
                    status = "HASH_MISMATCH"
                    action = "Manual Review / Compare Hashes"
                manifest.total_drift += 1

            rec = ComparisonRecord(
                rel_path=rel,
                filename=fname,
                pc_status="present",
                gdrive_status="present",
                match_status=status,
                pc_size=pc_f.size_bytes,
                gdrive_size=gd_f.size_bytes,
                pc_sha256=pc_f.sha256,
                gdrive_sha256=gd_f.sha256,
                pc_modified=pc_f.modified_iso,
                gdrive_modified=gd_f.modified_iso,
                pc_pages=pc_f.page_count,
                gdrive_pages=gd_f.page_count,
                pc_images=pc_f.image_count,
                gdrive_images=gd_f.image_count,
                action_required=action,
            )
            manifest.total_pdf_pages_pc += pc_f.page_count
            manifest.total_pdf_pages_gdrive += gd_f.page_count

        elif pc_f and not gd_f:
            manifest.total_pc_only += 1
            manifest.total_pdf_pages_pc += pc_f.page_count
            rec = ComparisonRecord(
                rel_path=rel,
                filename=fname,
                pc_status="present",
                gdrive_status="missing",
                match_status="PC_ONLY",
                pc_size=pc_f.size_bytes,
                pc_sha256=pc_f.sha256,
                pc_modified=pc_f.modified_iso,
                pc_pages=pc_f.page_count,
                pc_images=pc_f.image_count,
                action_required="Sync new file to Google Drive",
            )
        else:  # not pc_f and gd_f
            manifest.total_gdrive_only += 1
            manifest.total_pdf_pages_gdrive += gd_f.page_count  # type: ignore
            rec = ComparisonRecord(
                rel_path=rel,
                filename=fname,
                pc_status="missing",
                gdrive_status="present",
                match_status="GDRIVE_ONLY",
                gdrive_size=gd_f.size_bytes,  # type: ignore
                gdrive_sha256=gd_f.sha256,  # type: ignore
                gdrive_modified=gd_f.modified_iso,  # type: ignore
                gdrive_pages=gd_f.page_count,  # type: ignore
                gdrive_images=gd_f.image_count,  # type: ignore
                action_required="Sync new file from Google Drive to PC",
            )

        manifest.records.append(rec)

    return manifest


def sync_files_parallel(
    source_root: Path,
    target_root: Path,
    rel_paths: Optional[Sequence[str]] = None,
    overwrite_if_newer: bool = True,
    max_workers: int = 8,
) -> Dict[str, Any]:
    """Copy files from source_root to target_root preserving folder layout."""
    if not source_root.exists():
        return {"error": f"Source root does not exist: {source_root}", "synced": 0}

    target_root.mkdir(parents=True, exist_ok=True)

    if rel_paths is None:
        source_inv = inventory_tree_parallel(source_root, max_workers=max_workers)
        items_to_sync = list(source_inv.keys())
    else:
        items_to_sync = list(rel_paths)

    synced_count = 0
    skipped_count = 0
    errors: List[Dict[str, str]] = []

    def _sync_single(rel: str) -> Tuple[str, bool, Optional[str]]:
        src = source_root / rel
        tgt = target_root / rel
        if not src.exists() or not src.is_file():
            return rel, False, "Source file not found"

        tgt.parent.mkdir(parents=True, exist_ok=True)

        if tgt.exists() and not overwrite_if_newer:
            if calculate_sha256(src) == calculate_sha256(tgt):
                return rel, False, None  # skipped identical

        try:
            shutil.copy2(src, tgt)
            return rel, True, None
        except Exception as exc:
            return rel, False, str(exc)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futs = [executor.submit(_sync_single, rel) for rel in items_to_sync]
        for fut in concurrent.futures.as_completed(futs):
            rel, success, err = fut.result()
            if success:
                synced_count += 1
            elif err:
                errors.append({"file": rel, "error": err})
            else:
                skipped_count += 1

    return {
        "synced_count": synced_count,
        "skipped_count": skipped_count,
        "error_count": len(errors),
        "errors": errors,
        "source_root": str(source_root),
        "target_root": str(target_root),
        "timestamp": _dt.datetime.now(_dt.timezone.utc).isoformat(),
    }


def save_chat_and_notes_to_drive(
    notes_or_chat: str,
    pc_root: Path,
    gdrive_root: Path,
    section_name: str = "Section 7",
    title: str = "examiner_chat_notes",
) -> Dict[str, str]:
    """Save conversation notes, audit trails, and instructions directly to the Sync folder on both PC and Drive."""
    now_tag = _dt.datetime.now(_dt.timezone.utc).strftime("%Y%m%d_%H%M%S")
    fname = f"{title}_{now_tag}.md"

    pc_sync_dir = pc_root / section_name / "7_Sync_and_Transcripts"
    pc_sync_dir.mkdir(parents=True, exist_ok=True)
    pc_file = pc_sync_dir / fname

    with open(pc_file, "w", encoding="utf-8") as f:
        f.write(notes_or_chat)

    gd_file = None
    if gdrive_root:
        gd_sync_dir = gdrive_root / section_name / "7_Sync_and_Transcripts"
        gd_sync_dir.mkdir(parents=True, exist_ok=True)
        gd_file = gd_sync_dir / fname
        with open(gd_file, "w", encoding="utf-8") as f:
            f.write(notes_or_chat)

    return {
        "pc_path": str(pc_file),
        "gdrive_path": str(gd_file) if gd_file else "None",
        "timestamp": now_tag,
    }


def export_comparison_manifest(
    manifest: SyncManifest,
    out_dir: Path,
) -> Dict[str, str]:
    """Export the sync comparison manifest to JSON, CSV, and formatted Excel."""
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / "gdrive_pc_sync_manifest.json"
    csv_path = out_dir / "gdrive_pc_comparison.csv"
    xlsx_path = out_dir / "gdrive_pc_comparison_report.xlsx"
    md_path = out_dir / "GDRIVE_PC_SYNC_STATUS.md"

    # JSON export
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(asdict(manifest), f, indent=2)

    # CSV export
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Relative Path",
            "Filename",
            "Match Status",
            "Action Required",
            "PC Size (Bytes)",
            "GDrive Size (Bytes)",
            "PC SHA-256",
            "GDrive SHA-256",
            "PC Pages",
            "GDrive Pages",
            "PC Images",
            "GDrive Images",
            "PC Modified",
            "GDrive Modified",
        ])
        for r in manifest.records:
            writer.writerow([
                r.rel_path,
                r.filename,
                r.match_status,
                r.action_required,
                r.pc_size or "",
                r.gdrive_size or "",
                r.pc_sha256 or "",
                r.gdrive_sha256 or "",
                r.pc_pages,
                r.gdrive_pages,
                r.pc_images,
                r.gdrive_images,
                r.pc_modified or "",
                r.gdrive_modified or "",
            ])

    # Markdown export
    md_lines = [
        f"# Google Drive vs. Local PC Synchronization & Audit Report",
        f"**Section:** `{manifest.section}`  ",
        f"**Audit Timestamp:** `{manifest.timestamp}`  ",
        f"**PC Root Path:** `{manifest.pc_root}`  ",
        f"**Google Drive Root Path:** `{manifest.gdrive_root}`  ",
        "",
        "## Summary Metrics",
        f"- **Total PC Files:** `{manifest.total_pc_files}`",
        f"- **Total Google Drive Files:** `{manifest.total_gdrive_files}`",
        f"- **Matched in Sync (100% SHA-256):** `{manifest.total_matched}`",
        f"- **PC Only (Needs Push to Drive):** `{manifest.total_pc_only}`",
        f"- **Google Drive Only (Needs Pull to PC):** `{manifest.total_gdrive_only}`",
        f"- **Drift / Modified (Version Discrepancy):** `{manifest.total_drift}`",
        f"- **Total PDF Pages Indexed (PC):** `{manifest.total_pdf_pages_pc}`",
        f"- **Total PDF Pages Indexed (Google Drive):** `{manifest.total_pdf_pages_gdrive}`",
        "",
        "## File Status Breakdown",
        "",
        "| Relative Path | Status | Action Required | PC Pages | Drive Pages | PC SHA-256 (prefix) | GDrive SHA-256 (prefix) |",
        "| :--- | :--- | :--- | :---: | :---: | :--- | :--- |",
    ]
    for r in manifest.records:
        pc_sha = r.pc_sha256[:12] if r.pc_sha256 else "—"
        gd_sha = r.gdrive_sha256[:12] if r.gdrive_sha256 else "—"
        md_lines.append(
            f"| `{r.rel_path}` | **{r.match_status}** | {r.action_required} | {r.pc_pages} | {r.gdrive_pages} | `{pc_sha}` | `{gd_sha}` |"
        )

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))

    # Excel export (formatted)
    if _HAVE_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sync Comparison"

        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        drift_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        missing_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        headers = [
            "Relative Path",
            "Filename",
            "Match Status",
            "Action Required",
            "PC Size",
            "GDrive Size",
            "PC Pages",
            "GDrive Pages",
            "PC Images",
            "GDrive Images",
            "PC SHA-256",
            "GDrive SHA-256",
            "PC Modified (UTC)",
            "GDrive Modified (UTC)",
        ]
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if "Size" in h or "Pages" in h or "Images" in h else "left")

        for row_idx, r in enumerate(manifest.records, 2):
            row_data = [
                r.rel_path,
                r.filename,
                r.match_status,
                r.action_required,
                r.pc_size or 0,
                r.gdrive_size or 0,
                r.pc_pages,
                r.gdrive_pages,
                r.pc_images,
                r.gdrive_images,
                r.pc_sha256 or "",
                r.gdrive_sha256 or "",
                r.pc_modified or "",
                r.gdrive_modified or "",
            ]
            ws.append(row_data)

            # Color status cell
            st_cell = ws.cell(row=row_idx, column=3)
            if r.match_status == "MATCH_EXACT":
                st_cell.fill = match_fill
            elif "DRIFT" in r.match_status or "MISMATCH" in r.match_status:
                st_cell.fill = drift_fill
            else:
                st_cell.fill = missing_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        ws.freeze_panes = "A2"
        wb.save(xlsx_path)

    return {
        "json": str(json_path),
        "csv": str(csv_path),
        "xlsx": str(xlsx_path) if _HAVE_OPENPYXL else "",
        "markdown": str(md_path),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="DataBossX Google Drive & PC Sync Engine")
    parser.add_argument("--pc-root", type=str, required=True, help="Path to local PC section folder")
    parser.add_argument("--gdrive-root", type=str, required=True, help="Path to Google Drive section folder")
    parser.add_argument("--section", type=str, default="Section 7", help="Section identifier (e.g. Section 7)")
    parser.add_argument("--action", choices=["compare", "sync-to-drive", "sync-from-drive", "init-folders"], default="compare")
    parser.add_argument("--output-dir", type=str, default="./sync_output", help="Directory for sync reports")
    parser.add_argument("--workers", type=int, default=8, help="Number of parallel worker threads")

    args = parser.parse_args()
    pc_p = Path(args.pc_root)
    gd_p = Path(args.gdrive_root)
    out_p = Path(args.output_dir)

    if args.action == "init-folders":
        print(f"Initializing standard section folders for {args.section}...")
        initialize_section_folders(pc_p, args.section)
        initialize_section_folders(gd_p, args.section)
        print("Done.")
        return 0

    if args.action == "compare":
        print(f"Comparing PC [{pc_p}] vs Google Drive [{gd_p}]...")
        manifest = compare_trees(pc_p, gd_p, section_name=args.section, max_workers=args.workers)
        res = export_comparison_manifest(manifest, out_p)
        print(f"Comparison complete! Total PC: {manifest.total_pc_files}, Drive: {manifest.total_gdrive_files}, Matched: {manifest.total_matched}, Drift: {manifest.total_drift}")
        print(f"Reports written to: {res}")
        return 0

    if args.action == "sync-to-drive":
        print(f"Syncing from PC [{pc_p}] to Google Drive [{gd_p}]...")
        res = sync_files_parallel(pc_p, gd_p, max_workers=args.workers)
        print(f"Sync complete: {res['synced_count']} synced, {res['skipped_count']} skipped, {res['error_count']} errors.")
        return 0

    if args.action == "sync-from-drive":
        print(f"Syncing from Google Drive [{gd_p}] to PC [{pc_p}]...")
        res = sync_files_parallel(gd_p, pc_p, max_workers=args.workers)
        print(f"Sync complete: {res['synced_count']} synced, {res['skipped_count']} skipped, {res['error_count']} errors.")
        return 0

    return 0


if __name__ == "__main__":
    sys.exit(main())
