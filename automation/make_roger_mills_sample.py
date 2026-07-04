#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Synthetic Roger Mills corpus generator (for TESTING the v2 builder only).

Creates a clearly-labeled FAKE folder tree that mimics the real layout:

    <out>/Horizon/
        Template(30).xlsx
        Roger Mills/   <report>.xlsx  <OGL sheet>.xlsx  <index>.pdf-not-real
        Roger Mills 2/ <report>.xlsx  ...
        Roger Mills 3/ <report>.xlsx  ...
        .env

Every value here is SYNTHETIC and marked as such. This exists so we can prove
the real tool runs end to end; it is never shipped as data.
"""
from __future__ import annotations
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CANON_HEADERS = [
    "Entry", "Instrument Date", "Recorded Date", "Type",
    "Grantor", "Grantee", "Book", "Page", "Instrument No",
    "Legal Description", "Acres", "NMA", "Interest", "Remarks",
]

# (entry, inst_date, rec_date, type, grantor, grantee, book, page, instno, legal, acres, nma, interest, remarks)
BASE_ROWS = [
    (1, "1/5/1961", "1/9/1961", "Patent", "United States of America", "Roy T. Mills", "12", "301", "", "NE/4 Sec 31-12N-24W", "160", "160", "1.0", ""),
    (2, "3/2/1974", "3/6/1974", "Warranty Deed", "Roy T. Mills and Ada Mills", "John Q. Rancher", "88", "112", "", "NE/4 Sec 31-12N-24W", "160", "160", "1.0", ""),
    (3, "6/14/1998", "6/20/1998", "Oil and Gas Lease", "John Q. Rancher", "Apache Corporation", "241", "455", "", "NE/4 Sec 31-12N-24W", "160", "160", "0.125 RI", "3 yr primary term"),
    (4, "8/1/2005", "8/9/2005", "Mineral Deed", "John Q. Rancher", "Jane R. Heir", "402", "77", "", "N/2 NE/4 Sec 31-12N-24W", "80", "80", "0.5", "undivided 1/2"),
    (5, "2/11/2019", "2/15/2019", "Oil and Gas Lease", "Jane R. Heir", "Continental Resources", "915", "12", "2019-000812", "N/2 NE/4 Sec 31-12N-24W", "80", "80", "0.1875 RI", "5 yr term"),
]

OGL_ROWS = [
    # ogl_no, lessor, lessee, date, recorded, book, page, tract, acres, royalty
    ("OGL-001", "John Q. Rancher", "Apache Corporation", "6/14/1998", "6/20/1998", "241", "455", "NE/4 Sec 31-12N-24W", "160", "1/8"),
    ("OGL-002", "Jane R. Heir", "Continental Resources", "2/11/2019", "2/15/2019", "915", "12", "N/2 NE/4 Sec 31-12N-24W", "80", "3/16"),
]

TRACT_ROWS = [
    # tract_no, description, gross_acres
    ("Tract 1", "NE/4 of Section 31, Township 12 North, Range 24 West", "160"),
]


def _style_header(ws, ncols):
    fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="BBBBBB")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18
    ws.freeze_panes = "A2"


def make_template(path: Path):
    wb = openpyxl.Workbook()
    # Title sheet
    ts = wb.active
    ts.title = "Title Sheet"
    ts["A1"] = "CURSORY TITLE REPORT"
    ts["A1"].font = Font(bold=True, size=16)
    ts["A3"] = "Section:"; ts["B3"] = "31-12N-24W"
    ts["A4"] = "County:"; ts["B4"] = "Roger Mills"
    ts["A5"] = "State:"; ts["B5"] = "Oklahoma"
    ts["A6"] = "Effective Date:"; ts["B6"] = ""
    ts["A7"] = "Prepared For:"; ts["B7"] = ""
    ts["A8"] = "Gross Acres:"; ts["B8"] = ""
    for r in range(3, 9):
        ts.cell(row=r, column=1).font = Font(bold=True)
    ts.column_dimensions["A"].width = 18
    ts.column_dimensions["B"].width = 40
    # Runsheet (the data table)
    rs = wb.create_sheet("Runsheet")
    rs.append(CANON_HEADERS)
    _style_header(rs, len(CANON_HEADERS))
    # Tract sheet
    tr = wb.create_sheet("Tract Sheet")
    tr.append(["Tract No", "Description", "Gross Acres"])
    _style_header(tr, 3)
    # OGL sheet
    og = wb.create_sheet("OGL")
    og.append(["OGL No", "Lessor", "Lessee", "Lease Date", "Recorded", "Book", "Page", "Tract", "Acres", "Royalty"])
    _style_header(og, 10)
    # Notes
    nt = wb.create_sheet("Notes")
    nt["A1"] = "TITLE NOTES"; nt["A1"].font = Font(bold=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_report(path: Path, rows, title_date="", prepared="", note="", drop_ogl_no=True):
    wb = openpyxl.Workbook()
    ts = wb.active
    ts.title = "Title Sheet"
    ts["A1"] = "CURSORY TITLE REPORT"
    ts["A1"].font = Font(bold=True, size=16)
    ts["A3"] = "Section:"; ts["B3"] = "31-12N-24W"
    ts["A4"] = "County:"; ts["B4"] = "Roger Mills"
    ts["A5"] = "State:"; ts["B5"] = "Oklahoma"
    ts["A6"] = "Effective Date:"; ts["B6"] = title_date
    ts["A7"] = "Prepared For:"; ts["B7"] = prepared
    ts["A8"] = "Gross Acres:"; ts["B8"] = "160"
    rs = wb.create_sheet("Runsheet")
    rs.append(CANON_HEADERS)
    _style_header(rs, len(CANON_HEADERS))
    for row in rows:
        row = list(row)
        if drop_ogl_no:
            row[8] = ""  # simulate missing/blank OGL instrument numbers to be filled from OGL sheet
        rs.append(row)
    tr = wb.create_sheet("Tract Sheet")
    tr.append(["Tract No", "Description", "Gross Acres"])
    for t in TRACT_ROWS:
        tr.append(list(t))
    og = wb.create_sheet("OGL")
    og.append(["OGL No", "Lessor", "Lessee", "Lease Date", "Recorded", "Book", "Page", "Tract", "Acres", "Royalty"])
    for o in OGL_ROWS:
        og.append(list(o))
    nt = wb.create_sheet("Notes")
    nt["A1"] = "TITLE NOTES"; nt["A1"].font = Font(bold=True)
    if note:
        nt["A2"] = note
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    out = Path(sys.argv[1] if len(sys.argv) > 1 else "./_sample_horizon")
    horizon = out / "Horizon"
    make_template(horizon / "Template(30).xlsx")
    # Folder 1: fewest columns filled, older
    make_report(horizon / "Roger Mills" / "31-12N-24W_Cursory_Title_Report_v1.xlsx",
                BASE_ROWS[:4], title_date="", prepared="", note="Need recording info for Tract 1 lease.")
    # Folder 2: more complete, has a conflict (different book/page on row 2)
    r2 = [list(r) for r in BASE_ROWS]
    r2[1][6] = "89"  # conflicting book
    make_report(horizon / "Roger Mills 2" / "31-12N-24W_Cursory_Title_Report_FINAL.xlsx",
                r2, title_date="6/27/2026", prepared="Horizon Minerals", note="Confirm 1974 deed book/page.")
    # Folder 3: most complete + 'updated' filename (should win tournament)
    make_report(horizon / "Roger Mills 3" / "31-12N-24W_Cursory_Title_Report_updated_best.xlsx",
                BASE_ROWS, title_date="6/27/2026", prepared="Horizon Minerals",
                note="Curative: obtain lessor affidavits for OGL-002.")
    (horizon / ".env").write_text("ANTHROPIC_API_KEY=\nHORIZON_NOTE=synthetic\n", encoding="utf-8")
    print(f"Wrote synthetic Horizon corpus under: {horizon}")


if __name__ == "__main__":
    main()
