#!/usr/bin/env python3
"""
Roger Mills 2 — Cursory Title Report: format alignment + data-correction pass.

Target report : 31-12N-24W Roger Mills Co. Cursory Title Report (7-6-26) - NHE.xlsx
Style source  : Template.xlsx  (the QA'd Section 31 rebuild — its visible sheets
                define the approved visual language)
Root folder   : D:\\Desktop\\Horizon\\Roger Mills 2   (override with --root)

What this script does, in order:
  1. Backs up the untouched target into  _Backups\\  (timestamped) before any edit.
  2. Applies the audited DATA CORRECTIONS (each one is guarded: the cell must
     still contain the expected "before" value or the fix is skipped and logged,
     so re-running on an already-fixed file is safe and idempotent).
  3. Applies the FORMATTING ALIGNMENT to the template's visual language:
       - OGL + Runsheet: exact per-cell style transfer from Template.xlsx
         (both sheets are row-aligned to the same source records — verified
         566/566 rows on Runsheet, 69/69 lease rows on OGL).
       - Title / WI 1 / WI 2 / Well 1: rule-based restyle (these sheets have a
         different row structure than the template, so styles are mapped by
         structural role — banner rows, header rows, total rows, data rows —
         using the exact fills/fonts/borders harvested from the template).
       - Overview: template column widths + scaffolding row heights; the plat
         map rows (10-42) are left untouched so the merged-cell map keeps its
         proportions and the medium-gray unit shading is preserved.
       - PLAT: hidden (it is empty in the target; the template hides it too).
       - Tract 1-10 sheets: untouched — they already match the template's
         Tract-sheet format cell-for-cell (verified against template Tract 1).
  4. Saves the finished report under its original name in the root folder.
  5. Sweeps every other file in the root into newly created subfolders
     (_Backups / _Reference / _Working_Files) so only the deliverable and the
     subfolders remain at the top level.

Usage (from any directory):
    python roger_mills_2_format_fix.py                 # real run on D:\\...
    python roger_mills_2_format_fix.py --dry-run       # report, change nothing
    python roger_mills_2_format_fix.py --root "C:\\somewhere\\else"
    python roger_mills_2_format_fix.py --no-sweep      # skip the folder sweep

Requires: Python 3.9+ and openpyxl (pip install openpyxl). Nothing else.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import re
import shutil
import sys
from copy import copy
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

DEFAULT_ROOT = r"D:\Desktop\Horizon\Roger Mills 2"
TARGET_NAME = "31-12N-24W Roger Mills Co. Cursory Title Report (7-6-26) - NHE.xlsx"
TEMPLATE_NAME = "Template.xlsx"

# Fallback globs in case the local filenames drift slightly (underscores, etc.)
TARGET_GLOB = "*31*12N*24W*Cursory*Title*Report*NHE*.xlsx"
TEMPLATE_GLOB = "*Template*.xlsx"

# Template palette (harvested from Template.xlsx)
NAVY = "FF1D2330"      # section banner / header band
STEEL = "FF1F4E78"     # Runsheet header band, Title top banner
GRAY_HDR = "FFE2E3E5"  # column-header rows
TOTAL_BG = "FFFFF2CC"  # totals band
TOTAL_BRD = "FFBF9000" # totals band border accent
GRID = "FFBFBFBF"      # thin data-grid borders
WHITE = "FFFFFFFF"

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_HDR = PatternFill("solid", fgColor=GRAY_HDR)
FILL_TOTAL = PatternFill("solid", fgColor=TOTAL_BG)

THIN = Side(style="thin", color=GRID)
TOTAL_EDGE = Side(style="medium", color=TOTAL_BRD)

LOG: list[str] = []


def log(msg: str) -> None:
    LOG.append(msg)
    print(msg)


# --------------------------------------------------------------------------- #
# Guarded data corrections
# --------------------------------------------------------------------------- #

def set_cell(ws, addr: str, expect, new, why: str, dry: bool) -> None:
    """Replace a cell value only if it still holds the expected 'before' value."""
    cur = ws[addr].value
    ok = cur == expect if not callable(expect) else expect(cur)
    if not ok:
        log(f"  SKIP  {ws.title}!{addr}: value changed since audit "
            f"(found {str(cur)[:60]!r}) — {why}")
        return
    if not dry:
        ws[addr] = new
    log(f"  FIX   {ws.title}!{addr}: {why}")


def sub_cell(ws, addr: str, old: str, new: str, why: str, dry: bool) -> None:
    """Substring replacement, guarded on the substring being present."""
    cur = ws[addr].value
    if not isinstance(cur, str) or old not in cur:
        log(f"  SKIP  {ws.title}!{addr}: {old!r} not found — {why}")
        return
    if not dry:
        ws[addr] = cur.replace(old, new)
    log(f"  FIX   {ws.title}!{addr}: {why}")


def apply_data_corrections(wb, dry: bool) -> None:
    log("== Data corrections ==")
    title, ogl, rs = wb["Title"], wb["OGL"], wb["Runsheet"]
    t2, t5, t8, t9 = wb["Tract 2"], wb["Tract 5"], wb["Tract 8"], wb["Tract 9"]
    well = wb["Well 1"]

    # -- 1. Unclosed parenthesis after "William Cobal Bain, Jr." ------------- #
    bain = "William C. Bain, Jr. (presumed to be a/k/a William Cobal Bain, Jr."
    set_cell(title, "B37", bain, bain + ")", "close parenthesis", dry)
    set_cell(t2, "D21", bain, bain + ")", "close parenthesis", dry)

    # -- 2. "W. 1820 chains" is a dropped decimal: same sentence reads
    #       "ADA W 18.20 chains"; Title!C158 carries the correct figure. ----- #
    for ws, addr in ((title, "C128"), (t5, "D2"), (t8, "D2")):
        sub_cell(ws, addr, "W. 1820 chains", "W. 18.20 chains",
                 '"1820 chains" -> "18.20 chains"', dry)

    # -- 3. Tract 9 legal-description punctuation ---------------------------- #
    for ws, addr in ((title, "C167"), (t9, "D2")):
        sub_cell(ws, addr, "SE/corner", "SE corner", "SE/corner -> SE corner", dry)
        sub_cell(ws, addr, "West 2547.60 feet:", "West 2547.60 feet;",
                 "colon -> semicolon in metes-and-bounds call", dry)
        sub_cell(ws, addr, "point of beginning.of 31-12N-24W",
                 "point of beginning, of 31-12N-24W",
                 "missing separator before section reference", dry)

    # -- 4. Stray trailing token in OGL 27's legal description --------------- #
    sub_cell(ogl, "J28", " aol", "", 'drop stray trailing "aol"', dry)

    # -- 5. Scrambled grantor name order on top lease OGL 66 ----------------- #
    set_cell(ogl, "H67",
             "a/k/a Gena Williams, a/k/a Gena Mitchell Williams, Gena W. Moore",
             "Gena W. Moore, a/k/a Gena Mitchell Williams, a/k/a Gena Williams",
             "primary name first, a/k/a's after", dry)

    # -- 6. Duplicated word in grantor on top lease OGL 69 ------------------- #
    set_cell(ogl, "H70",
             "Kevin Johnson Johnson, Trustee of The Kevin Warner Johnson Trust",
             "Kevin Warner Johnson, Trustee of The Kevin Warner Johnson Trust",
             '"Kevin Johnson Johnson" -> "Kevin Warner Johnson"', dry)

    # -- 7. "Hazel Trust Hamilton" is a surname-first clerk-index entry
    #       ("Hamilton, Hazel Trust") flipped mechanically; the grantee entity
    #       of the 1123/100 mineral deed is the Hazel Hamilton Trust. --------- #
    for ws, addr in ((title, "B62"), (t2, "D62"), (rs, "H173")):
        set_cell(ws, addr, "Hazel Trust Hamilton", "Hazel Hamilton Trust",
                 "un-flip clerk-index name order", dry)
    sub_cell(rs, "AA173", "to Hazel Trust Hamilton", "to Hazel Hamilton Trust",
             "un-flip clerk-index name order in summary", dry)

    # -- 8. Normalize the trust styling used at Title!B65 ("Rive Lind McCaul") #
    for ws, addr in ((title, "B68"), (t2, "D78")):
        set_cell(ws, addr, "RIVE Lind McCaul Trust", "Rive Lind McCaul Trust",
                 "normalize record-styled all-caps", dry)

    # -- 9. Well 1: stated spud 6/5/2006 postdates stated completion
    #       5/19/2006. The completion date is corroborated by both workbooks;
    #       the spud date is corroborated by nothing in the source data, so it
    #       takes the report's own "Not confirmed of record" convention. ------ #
    set_cell(well, "H3",
             lambda v: isinstance(v, _dt.datetime) and v.date() == _dt.date(2006, 6, 5),
             "Not confirmed of record",
             "spud 6/5/2006 contradicts completion 5/19/2006 — flag for OCC 1002A check",
             dry)
    if not dry and isinstance(well["H3"].value, str):
        well["H3"].number_format = "General"

    # -- 10. OGL "Tracts" column: rebuilt on the FINAL 10-tract layout.
    #        The column still carried an earlier 8-tract draft numbering
    #        (e.g. "All (1-8)" in a 10-tract report). Each entry below was
    #        re-derived from that lease's own legal description. -------------- #
    tract_fixes = {
        # OGL# : (expected current value, corrected value)
        1:  ("All (1-8)", "1, 3, 4, 5, 7, 9, 10"),
        2:  ("All (1-8)", "1, 3, 4, 5, 7, 9, 10"),
        3:  ("All (1-8)", "1, 3, 4, 5, 7, 9, 10"),
        4:  ("All (1-8)", "1, 3, 4, 5, 7, 9, 10"),
        8:  ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
        9:  ("2, 3, 5, 6, 7", "5, 6, 9"),
        10: ("5, 6", "5, 6, 8"),
        12: ("All (1-8)", "5, 6, 8"),
        15: (lambda v: isinstance(v, _dt.datetime), 2),   # "2" saved as a date
        20: ("2, 5", 2),
        23: ("2, 3, 5, 6, 7", "5, 9"),
        24: ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
        25: ("2, 3, 5, 7", "5, 9"),
        40: ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
        46: ("All (1-8)", "5, 8"),
        48: ("5, 6, 7", "5, 9"),
        49: ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
        50: ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
        61: ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
        62: (5, 8),
        64: ("1, 3, 4, 5, 6, 7, 8", "1, 3, 4, 5, 7, 9, 10"),
        65: ("All (1-8)", "All (1-10)"),
        66: ("All (1-8)", "1, 3, 4, 5, 6, 7, 8, 9, 10"),
        67: ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
        68: ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
        69: ("1, 3, 4, 7, 8", "3, 4, 7, 10"),
    }
    for ogl_no, (expect, new) in tract_fixes.items():
        row = ogl_no + 1                       # OGL 1 lives on sheet row 2
        assert ogl["A%d" % row].value == ogl_no, f"OGL row misalignment at {ogl_no}"
        addr = f"N{row}"
        set_cell(ogl, addr, expect, new,
                 f"OGL {ogl_no}: Tracts column onto final 10-tract layout", dry)
        if not dry and ogl_no == 15:
            ogl[addr].number_format = "General"  # cell previously formatted as a date


# --------------------------------------------------------------------------- #
# Formatting alignment
# --------------------------------------------------------------------------- #

def clone_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def copy_block_styles(tpl_ws, tgt_ws, max_row: int, max_col: int,
                      skip_widths: set[str] = frozenset()) -> None:
    """Exact style transfer for row-aligned sheets, plus widths and heights."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            clone_style(tpl_ws.cell(r, c), tgt_ws.cell(r, c))
        h = tpl_ws.row_dimensions[r].height if r in tpl_ws.row_dimensions else None
        if h is not None:
            tgt_ws.row_dimensions[r].height = h
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        if letter in skip_widths:
            continue
        if letter in tpl_ws.column_dimensions:
            w = tpl_ws.column_dimensions[letter].width
            if w is not None:
                tgt_ws.column_dimensions[letter].width = w


def restyled_font(cell, bold=None, color=None) -> Font:
    """Calibri-10 version of a cell's font, optionally forcing bold/color."""
    f = cell.font
    return Font(name="Calibri", size=10,
                bold=f.bold if bold is None else bold,
                italic=f.italic, underline=f.underline, strike=f.strike,
                vertAlign=f.vertAlign,
                color=f.color if color is None else color)


def _style_band(ws, r: int, styler) -> None:
    """Style B..G of row r. Cells hidden under a merge are skipped: Excel
    paints a merged range with its anchor cell's formatting, and the anchor
    (e.g. C4 of the merged legal-description cell C4:G4) is a real cell that
    does take the style — so the band renders solid across the full width."""
    from openpyxl.cell.cell import MergedCell
    for c in range(2, 8):
        cell = ws.cell(r, c)
        if isinstance(cell, MergedCell):
            continue
        styler(cell)


def restyle_title(ws) -> None:
    """Map the template's visual roles onto the target's tract-block layout."""
    band_pat = re.compile(r"^(TRACT \d+:|Working Interest\b.*)", re.I)
    max_row = ws.max_row
    r = 1
    while r <= max_row:
        b = ws.cell(r, 2).value
        btxt = b.strip() if isinstance(b, str) else ""
        if band_pat.match(btxt):
            # Section banner: navy band, white bold, across B:G
            def navy(cell):
                cell.fill = FILL_NAVY
                cell.font = restyled_font(cell, bold=True, color=WHITE)
                cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                           vertical="top", wrap_text=True)
            _style_band(ws, r, navy)
        elif btxt in ("Mineral Owner", "Owner"):
            # Column-header row: light gray band with grid borders
            for c in range(2, 8):
                cell = ws.cell(r, c)
                cell.fill = FILL_HDR
                cell.font = restyled_font(cell, bold=True)
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        elif btxt in ("REPORT TOTAL", "TOTAL"):
            # Totals band: template's soft-gold treatment
            for c in range(2, 8):
                cell = ws.cell(r, c)
                cell.fill = FILL_TOTAL
                cell.font = restyled_font(cell, bold=True)
                cell.border = Border(left=THIN, right=THIN,
                                     top=TOTAL_EDGE, bottom=TOTAL_EDGE)
            ws.cell(r, 3).number_format = "0.00"
        else:
            in_data = _row_is_owner_data(ws, r)
            for c in range(1, 8):
                cell = ws.cell(r, c)
                cell.font = restyled_font(cell)
                if in_data and c >= 2:
                    cell.border = Border(left=THIN, right=THIN,
                                         top=THIN, bottom=THIN)
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="top", wrap_text=True)
            if in_data:
                ws.cell(r, 3).number_format = "0.00000000"  # net acres
                ws.cell(r, 5).number_format = "0.0000"      # royalty decimal
        r += 1


def _row_is_owner_data(ws, r: int) -> bool:
    """True when row r sits between a column-header row and its REPORT TOTAL."""
    up, down = r, r
    while up >= 1:
        v = ws.cell(up, 2).value
        v = v.strip() if isinstance(v, str) else ""
        if v in ("Mineral Owner", "Owner"):
            break
        if v in ("REPORT TOTAL", "TOTAL") or v.upper().startswith("TRACT "):
            return False
        up -= 1
    else:
        return False
    if up < 1:
        return False
    while down <= ws.max_row:
        v = ws.cell(down, 2).value
        v = v.strip() if isinstance(v, str) else ""
        if v == "REPORT TOTAL":
            return True
        if v in ("Mineral Owner", "Owner") and down != up:
            return False
        down += 1
    return False


def restyle_wi(ws) -> None:
    """WI sheets keep their structure; the body font moves to Calibri 10."""
    for row in ws.iter_rows():
        for cell in row:
            f = cell.font
            if f.name and f.name.startswith("Arial") and (f.size or 10) in (9.0, 10.0):
                cell.font = Font(name="Calibri", size=10, bold=f.bold,
                                 italic=f.italic, underline=f.underline,
                                 strike=f.strike, vertAlign=f.vertAlign,
                                 color=f.color)


def restyle_well(tgt, tpl) -> None:
    """Well 1: template's navy header band + Calibri 10 body + template widths."""
    white_side = Side(style="thin", color=WHITE)
    for r in (1, 2):
        for c in range(1, 18):
            cell = tgt.cell(r, c)
            cell.fill = FILL_NAVY
            cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="top",
                                       wrap_text=True)
            cell.border = Border(left=white_side, right=white_side,
                                 top=white_side, bottom=white_side)
    for c in range(1, 18):
        cell = tgt.cell(3, c)
        cell.font = Font(name="Calibri", size=10, color=cell.font.color)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in (tgt["N3"], tgt["O3"]):
        if isinstance(cell.value, _dt.datetime):
            cell.number_format = "m/d/yy"
    tgt.row_dimensions[1].height = 26.25
    for c in range(1, 18):
        letter = get_column_letter(c)
        if letter in tpl.column_dimensions and tpl.column_dimensions[letter].width:
            tgt.column_dimensions[letter].width = tpl.column_dimensions[letter].width


def align_overview(tgt, tpl) -> None:
    """Adopt template scaffolding without disturbing the plat-map grid (10-42)."""
    for letter in ("A", "B", "AH", "AI"):
        if letter in tpl.column_dimensions and tpl.column_dimensions[letter].width:
            tgt.column_dimensions[letter].width = tpl.column_dimensions[letter].width
    for r in list(range(1, 10)) + list(range(43, 57)):
        h = tpl.row_dimensions[r].height if r in tpl.row_dimensions else None
        if h is not None:
            tgt.row_dimensions[r].height = h
    tgt.page_setup.orientation = tpl.page_setup.orientation
    tgt.page_margins = copy(tpl.page_margins)


def apply_formatting(wb, tpl, dry: bool) -> None:
    log("== Formatting alignment ==")
    if dry:
        log("  (dry-run: formatting steps listed but not applied)")
        for step in ("OGL exact style transfer", "Runsheet exact style transfer",
                     "Title restyle", "WI 1 / WI 2 font alignment",
                     "Well 1 header band", "Overview scaffolding", "hide PLAT"):
            log(f"  PLAN  {step}")
        return

    # OGL & Runsheet are row-aligned with the template's copies of the same
    # source records, so the template's exact cell styles transfer 1:1.
    copy_block_styles(tpl["OGL"], wb["OGL"], max_row=70, max_col=28)
    # The Tracts column holds plain tract lists; scrub any numeric/date
    # formats the template carried over (e.g. its N16 was "#,##0.0000").
    for r in range(2, 71):
        wb["OGL"].cell(r, 14).number_format = "General"
    log("  DONE  OGL: exact per-cell style transfer (A1:AB70, widths, heights)")

    # Template Runsheet col C carries a 129-char autofit artifact; keep the
    # target's readable width there — everything else transfers verbatim.
    copy_block_styles(tpl["Runsheet"], wb["Runsheet"], max_row=566, max_col=28,
                      skip_widths={"C"})
    log("  DONE  Runsheet: exact per-cell style transfer (A1:AB566, widths "
        "except C, heights)")

    restyle_title(wb["Title"])
    log("  DONE  Title: Calibri 10 body, navy tract banners, gray headers, "
        "gold totals, gridded owner rows")

    restyle_wi(wb["WI 1"])
    restyle_wi(wb["WI 2"])
    log("  DONE  WI 1 / WI 2: body font aligned to Calibri 10")

    restyle_well(wb["Well 1"], tpl["Well 1"])
    log("  DONE  Well 1: navy header band, Calibri 10 body, template widths")

    align_overview(wb["Overview"], tpl["Overview"])
    log("  DONE  Overview: template widths/heights outside the map grid, "
        "portrait print setup")

    # Title print setup follows the template (portrait, 80% like the template)
    wb["Title"].page_setup.orientation = tpl["Title "].page_setup.orientation

    wb["PLAT"].sheet_state = "hidden"
    log("  DONE  PLAT: hidden (sheet is empty; template hides it as well)")


# --------------------------------------------------------------------------- #
# Folder sweep
# --------------------------------------------------------------------------- #

def unique_dest(folder: Path, name: str) -> Path:
    """Never overwrite during the sweep — append (2), (3)... when needed."""
    dest = folder / name
    stem, suffix = dest.stem, dest.suffix
    n = 2
    while dest.exists():
        dest = folder / f"{stem} ({n}){suffix}"
        n += 1
    return dest


def sweep_root(root: Path, keep: Path, dry: bool) -> None:
    """Move every loose file except the deliverable into tidy subfolders."""
    log("== Root-folder sweep ==")
    backups = root / "_Backups"
    reference = root / "_Reference"
    working = root / "_Working_Files"

    backup_pat = re.compile(r"(backup|\.bak$|~\$|copy of|\s-\s?copy|\(\d+\)\.xlsx$)", re.I)
    this_script = Path(__file__).resolve()

    for item in sorted(root.iterdir()):
        if item.is_dir():
            continue                          # existing folders stay put
        if item.resolve() == keep.resolve() or item.resolve() == this_script:
            continue                          # the deliverable / this script
        if backup_pat.search(item.name):
            folder = backups
        elif item.suffix.lower() in (".xlsx", ".xlsm", ".xls", ".csv"):
            folder = reference if "template" in item.name.lower() else working
        else:
            folder = working
        if dry:
            log(f"  PLAN  {item.name}  ->  {folder.name}\\")
            continue
        folder.mkdir(exist_ok=True)
        try:
            shutil.move(str(item), str(unique_dest(folder, item.name)))
            log(f"  MOVE  {item.name}  ->  {folder.name}\\")
        except PermissionError:
            log(f"  SKIP  {item.name}: locked/in use — close it and re-run")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def find_file(root: Path, exact: str, pattern: str, label: str) -> Path:
    # Search the root first, then the sweep subfolders (so a re-run still
    # finds Template.xlsx after a previous run tidied it into _Reference).
    search_dirs = [root, root / "_Reference", root / "_Working_Files",
                   root / "_Backups"]
    for d in search_dirs:
        if d.is_dir() and (d / exact).exists():
            if d != root:
                log(f"  NOTE  {label} found in {d.name}\\")
            return d / exact
    for d in search_dirs:
        if not d.is_dir():
            continue
        hits = sorted(d.glob(pattern))
        if len(hits) == 1:
            log(f"  NOTE  {label} matched by pattern: {d.name}\\{hits[0].name}")
            return hits[0]
    sys.exit(f"ERROR: cannot locate the {label} under {root}\n"
             f"  looked for {exact!r} then pattern {pattern!r}.")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--root", default=DEFAULT_ROOT,
                    help=f"working folder (default: {DEFAULT_ROOT})")
    ap.add_argument("--dry-run", action="store_true",
                    help="log every planned change without touching any file")
    ap.add_argument("--no-sweep", action="store_true",
                    help="skip the root-folder cleanup step")
    args = ap.parse_args()

    root = Path(args.root)
    if not root.is_dir():
        sys.exit(f"ERROR: root folder not found: {root}")

    log(f"Root      : {root}")
    target = find_file(root, TARGET_NAME, TARGET_GLOB, "target report")
    template = find_file(root, TEMPLATE_NAME, TEMPLATE_GLOB, "template")
    log(f"Target    : {target.name}")
    log(f"Template  : {template.name}")
    log(f"Mode      : {'DRY RUN — no files will be modified' if args.dry_run else 'live'}")

    # 1. Backup before any modification
    stamp = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_dir = root / "_Backups"
    if not args.dry_run:
        backup_dir.mkdir(exist_ok=True)
        backup = backup_dir / f"{target.stem}.pre-fix.{stamp}{target.suffix}"
        shutil.copy2(target, backup)
        log(f"Backup    : _Backups\\{backup.name}")

    # 2. + 3. Open once, correct data, align formatting, save in place.
    #    keep_vba=False is fine (plain .xlsx); formulas are preserved because
    #    the workbook is loaded with formulas intact (data_only=False).
    wb = openpyxl.load_workbook(target)
    tpl = openpyxl.load_workbook(template)

    apply_data_corrections(wb, args.dry_run)
    apply_formatting(wb, tpl, args.dry_run)

    if not args.dry_run:
        wb.save(target)
        log(f"Saved     : {target}")

    # 4. Sweep the root so only the deliverable + subfolders remain.
    if not args.no_sweep:
        sweep_root(root, keep=target, dry=args.dry_run)

    # 5. Write a run log next to the backups for the project file.
    if not args.dry_run:
        logfile = backup_dir / f"fix-log.{stamp}.txt"
        logfile.write_text("\n".join(LOG), encoding="utf-8")
        log(f"Run log   : _Backups\\{logfile.name}")
    log("All done.")


if __name__ == "__main__":
    main()
