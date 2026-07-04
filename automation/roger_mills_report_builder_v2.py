#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Roger Mills Cursory Title Report Builder  (codexv2)
===================================================

A single-command, defensive, *run-it-locally* tool that takes the BEST report
from each Roger Mills folder, runs loops + a tournament to pick and merge the
best data, fixes the tract / title / OGL sheets, matches your template, and
writes finished workbooks into  ...\Horizon\rogermillsfinalreports.

WHY THIS IS A LOCAL TOOL
------------------------
The cloud assistant that wrote this CANNOT see your D:\ drive. A cloud
container has no path to files on your PC, so instead of inventing data
(worthless for a title report) it produced this tool for you to run *where the
files actually live*.

WHAT'S NEW vs codexv1
---------------------
1.  MULTI-FOLDER: scans "Roger Mills", "Roger Mills 2", "Roger Mills 3" (any
    "Roger Mills*" folder) under --horizon, or explicit --roots.
2.  PER-FOLDER + CROSS-FOLDER TOURNAMENT ("loops and tournaments"): every report
    workbook is scored; the best base in each folder is reported, then the best
    overall wins as the merge base. Data from ALL folders is still merged.
3.  OGL AUTHORITY: finds the OGL sheet (oil & gas lease schedule) and uses its
    OGL numbers as the authoritative lease identifiers, filling/repairing the
    instrument column on matching lease rows ("use the OGL numbers").
4.  TRACT SPINE: finds the tract sheet, keeps its tract set authoritative
    ("don't go off the tract sheets") while normalizing/formatting it
    ("fix the tract sheets"). Every data row is tagged to its tract.
5.  TITLE SHEET FIX: finds the title/cover sheet and normalizes its metadata
    (section, county, state, effective date, prepared-for, gross acres) to match
    the template ("fix... title sheet data").
6.  NOTES & LEGALS: preserves the Notes sheet and every row's remarks; legal
    descriptions are standardized (STR formatting) without inventing land.
7.  TEMPLATE = FORMATTING AUTHORITY: output is a copy of Template(30).xlsx with
    all styling preserved; data written in.
8.  FIX THE FINAL COPY: the existing "FINAL"/newest report is used as the merge
    base but then de-duplicated, OGL-filled, date/name-normalized, and every
    unverifiable row flagged [REVIEW] instead of silently trusted.
9.  .ENV + OPTIONAL AI POLISH: loads Horizon\.env; with --use-llm and an
    ANTHROPIC_API_KEY it will only *reformat* legal descriptions/notes it was
    given (never adds facts) and writes a full before/after audit.
10. OUTPUT: finished workbook(s) + all support files go to
    ...\Horizon\rogermillsfinalreports.

IT NEVER INVENTS DATA. Unverifiable rows are flagged for human review.

USAGE (Windows PowerShell or CMD) -- one line:
    py -m pip install --upgrade openpyxl pandas pdfplumber PyMuPDF pytesseract Pillow python-dateutil rapidfuzz
    py automation\roger_mills_report_builder_v2.py --horizon "D:\Desktop\Horizon"

Run --dry-run first to see the plan without writing anything.
Only openpyxl is strictly required; everything else degrades gracefully.
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
import os
import re
import sys
import traceback
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

# --- reuse the proven codexv1 engine (same folder) --------------------------
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    import roger_mills_title_report_builder as v1  # noqa: E402
except Exception as exc:  # pragma: no cover
    print("FATAL: could not import roger_mills_title_report_builder.py "
          "(must sit in the same folder as this file).")
    raise

import openpyxl  # noqa: E402  (v1 already proved this import works)
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402

LOG = v1.LOG
norm_text = v1.norm_text
norm_name = v1.norm_name
norm_doc_ref = v1.norm_doc_ref
norm_date = v1.norm_date
similar = v1.similar
CANON_FIELDS = v1.CANON_FIELDS
match_header = v1.match_header
_detect_header_row = v1._detect_header_row


# ============================================================================
# .env loader (stdlib only)
# ============================================================================
def load_env(horizon: Path) -> Dict[str, str]:
    """Load KEY=VALUE pairs from Horizon\\.env (also checks CWD). Never overrides
    a value already present in the real environment."""
    env: Dict[str, str] = {}
    for candidate in (horizon / ".env", Path.cwd() / ".env"):
        try:
            if candidate.exists():
                for line in candidate.read_text(encoding="utf-8", errors="ignore").splitlines():
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    k, _, v = line.partition("=")
                    k, v = k.strip(), v.strip().strip('"').strip("'")
                    if k and k not in env:
                        env[k] = v
                LOG(f"Loaded .env: {candidate}  ({len(env)} keys)")
        except Exception as exc:
            LOG(f"  .env read failed for {candidate}: {exc}", "WARN")
    # real environment wins
    for k in list(env):
        if os.environ.get(k):
            env[k] = os.environ[k]
    return env


# ============================================================================
# Folder discovery + per-folder / cross-folder tournament
# ============================================================================
def discover_roots(horizon: Path, explicit: Sequence[str]) -> List[Path]:
    if explicit:
        roots = [Path(p) for p in explicit]
    else:
        roots = []
        if horizon.exists():
            for child in sorted(horizon.iterdir()):
                if child.is_dir() and re.match(r"roger\s*mills", child.name, re.I) \
                        and "finalreport" not in child.name.lower():
                    roots.append(child)
    return [r for r in roots if r.exists()]


@dataclass
class FolderResult:
    root: Path
    recs: List["v1.FileRec"]
    analyses: Dict[str, "v1.WorkbookAnalysis"]
    scored: List[Tuple[float, "v1.WorkbookAnalysis"]]
    best: Optional["v1.WorkbookAnalysis"]


def analyze_folder(root: Path) -> FolderResult:
    recs = v1.inventory(root)
    analyses: Dict[str, "v1.WorkbookAnalysis"] = {}
    for rec in recs:
        if rec.ext in v1.WORKBOOK_EXT:
            analyses[str(rec.path)] = v1.analyze_workbook(rec)
    report_analyses = [analyses[str(r.path)] for r in recs
                       if r.kind == "report" and str(r.path) in analyses]
    if recs:
        newest = max(r.mtime.timestamp() for r in recs)
        oldest = min(r.mtime.timestamp() for r in recs)
    else:
        newest = oldest = _dt.datetime.now().timestamp()
    scored = sorted(((v1.score_report(wa, newest, oldest), wa) for wa in report_analyses),
                    key=lambda t: t[0], reverse=True)
    best = scored[0][1] if scored and scored[0][0] >= 0 else None
    return FolderResult(root=root, recs=recs, analyses=analyses, scored=scored, best=best)


# ============================================================================
# Special-sheet detection: OGL / tract / title / notes
# ============================================================================
def _sheet_kind(sheet_name: str) -> str:
    n = sheet_name.lower()
    if "ogl" in n or ("oil" in n and "gas" in n) or "lease sched" in n:
        return "ogl"
    if "tract" in n:
        return "tract"
    if "title" in n or "cover" in n:
        return "title"
    if "note" in n or "requirement" in n or "curative" in n:
        return "notes"
    if "run" in n and "sheet" in n:
        return "runsheet"
    return "other"


# OGL sheet header recognition
OGL_SYNS: Dict[str, Sequence[str]] = {
    "ogl_no": ("ogl", "ogl no", "ogl #", "lease no", "lease number", "lease #", "no", "#"),
    "lessor": ("lessor", "grantor", "owner", "from"),
    "lessee": ("lessee", "grantee", "operator", "to"),
    "lease_date": ("date", "lease date", "dated", "instrument date"),
    "recorded": ("recorded", "filed", "recording date"),
    "book": ("book", "vol", "bk"),
    "page": ("page", "pg"),
    "tract": ("tract", "legal", "description", "land"),
    "acres": ("acres", "acreage", "nma"),
    "royalty": ("royalty", "ri", "rate"),
}


def _map_headers(ws, synonyms: Dict[str, Sequence[str]], scan_rows: int = 20) -> Tuple[int, Dict[int, str]]:
    best_row, best_map = 0, {}
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        hmap: Dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            t = norm_text(val).lower()
            if not t:
                continue
            for field_name, syns in synonyms.items():
                if field_name in hmap.values():
                    continue
                if any(s == t or s in t for s in syns):
                    hmap[c] = field_name
                    break
        if len(hmap) > len(best_map):
            best_row, best_map = r, hmap
    return best_row, best_map


@dataclass
class OGLRecord:
    ogl_no: str
    lessor: str
    lessee: str
    lease_date: str
    recorded: str
    book: str
    page: str
    tract: str
    acres: str
    royalty: str
    source: str


def extract_ogl(rec: "v1.FileRec") -> List[OGLRecord]:
    """Read every OGL-looking sheet in a workbook and return lease records."""
    out: List[OGLRecord] = []
    try:
        wb = openpyxl.load_workbook(rec.path, data_only=True, read_only=True)
    except Exception as exc:
        LOG(f"  OGL read failed for {rec.path.name}: {exc}", "WARN")
        return out
    for ws in wb.worksheets:
        if _sheet_kind(ws.title) != "ogl":
            continue
        hr, hmap = _map_headers(ws, OGL_SYNS)
        if not hmap:
            continue
        for r in range(hr + 1, ws.max_row + 1):
            row = {f: norm_text(ws.cell(row=r, column=c).value) for c, f in hmap.items()}
            if not any(row.values()):
                continue
            out.append(OGLRecord(
                ogl_no=row.get("ogl_no", ""), lessor=row.get("lessor", ""),
                lessee=row.get("lessee", ""), lease_date=row.get("lease_date", ""),
                recorded=row.get("recorded", ""), book=row.get("book", ""),
                page=row.get("page", ""), tract=row.get("tract", ""),
                acres=row.get("acres", ""), royalty=row.get("royalty", ""),
                source=f"{rec.path.name}:{ws.title}",
            ))
    wb.close()
    return out


@dataclass
class TractRecord:
    tract_no: str
    description: str
    acres: str
    source: str


def extract_tracts(rec: "v1.FileRec") -> List[TractRecord]:
    out: List[TractRecord] = []
    try:
        wb = openpyxl.load_workbook(rec.path, data_only=True, read_only=True)
    except Exception:
        return out
    for ws in wb.worksheets:
        if _sheet_kind(ws.title) != "tract":
            continue
        hr, hmap = _map_headers(ws, {
            "tract_no": ("tract", "tract no", "no", "#"),
            "description": ("description", "legal", "land", "lands"),
            "acres": ("acres", "gross acres", "acreage"),
        })
        if not hmap:
            continue
        for r in range(hr + 1, ws.max_row + 1):
            row = {f: norm_text(ws.cell(row=r, column=c).value) for c, f in hmap.items()}
            if not any(row.values()):
                continue
            out.append(TractRecord(
                tract_no=row.get("tract_no", ""), description=row.get("description", ""),
                acres=row.get("acres", ""), source=f"{rec.path.name}:{ws.title}"))
    wb.close()
    return out


def extract_title_meta(rec: "v1.FileRec") -> Dict[str, str]:
    """Pull labelled metadata (Section:, County:, ...) from a title/cover sheet."""
    meta: Dict[str, str] = {}
    # ordered so more specific labels win before generic ones
    label_map = [
        ("effective", "effective_date"),
        ("prepared for", "prepared_for"), ("prepared", "prepared_for"),
        ("gross acres", "gross_acres"), ("acres", "gross_acres"),
        ("section", "section"), ("county", "county"), ("state", "state"),
        ("date", "effective_date"),
    ]
    try:
        wb = openpyxl.load_workbook(rec.path, data_only=True, read_only=True)
    except Exception:
        return meta
    for ws in wb.worksheets:
        if _sheet_kind(ws.title) != "title":
            continue
        for r in range(1, min(ws.max_row, 40) + 1):
            for c in range(1, min(ws.max_column, 8) + 1):
                label = norm_text(ws.cell(row=r, column=c).value).lower().rstrip(":")
                if not label:
                    continue
                key = next((k for lbl, k in label_map if label == lbl or label.startswith(lbl)), None)
                if key:
                    val = norm_text(ws.cell(row=r, column=c + 1).value)
                    if val and not meta.get(key):
                        meta[key] = val
    wb.close()
    return meta


def extract_notes(rec: "v1.FileRec") -> List[str]:
    out: List[str] = []
    try:
        wb = openpyxl.load_workbook(rec.path, data_only=True, read_only=True)
    except Exception:
        return out
    for ws in wb.worksheets:
        if _sheet_kind(ws.title) != "notes":
            continue
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column, 4) + 1):
                v = norm_text(ws.cell(row=r, column=c).value)
                if v and v.upper() not in ("TITLE NOTES", "NOTES"):
                    out.append(f"{rec.path.name}: {v}")
    wb.close()
    return out


# ============================================================================
# Legal-description standardization (formatting only; invents nothing)
# ============================================================================
_QTR = {
    "northeast": "NE", "ne": "NE", "northwest": "NW", "nw": "NW",
    "southeast": "SE", "se": "SE", "southwest": "SW", "sw": "SW",
}


def standardize_legal(text: str) -> str:
    """Normalize section/township/range phrasing. Purely cosmetic; if nothing is
    recognized the original text is returned unchanged."""
    s = norm_text(text)
    if not s:
        return s
    # "Section 31, Township 12 North, Range 24 West" -> "Sec 31-12N-24W"
    m = re.search(r"section\s*(\d+)[, ]+township\s*(\d+)\s*(north|south|[ns])[, ]+range\s*(\d+)\s*(east|west|[ew])",
                  s, re.I)
    if m:
        sec, twp, twpd, rng, rngd = m.groups()
        twpd = twpd[0].upper()
        rngd = rngd[0].upper()
        str_std = f"Sec {sec}-{twp}{twpd}-{rng}{rngd}"
        s = s[:m.start()].rstrip(" ,") + (" " if m.start() else "") + str_std + s[m.end():]
    # collapse "N/2 NE/4" style spacing
    s = re.sub(r"\s*/\s*", "/", s)
    return norm_text(s)


# ============================================================================
# Optional AI polish (stdlib urllib; opt-in; reformat-only; fully audited)
# ============================================================================
def llm_polish(items: List[str], api_key: str, model: str = "claude-opus-4-8") -> Dict[str, str]:
    """Ask the model to CLEAN UP (not change facts of) each legal description.
    Returns {original: cleaned}. Any failure -> identity (no change). Every
    change is returned to the caller for audit; nothing is silently trusted."""
    import urllib.request
    result: Dict[str, str] = {o: o for o in items}
    uniq = [o for o in dict.fromkeys(items) if o.strip()]
    if not uniq or not api_key:
        return result
    system = ("You standardize U.S. oil-and-gas legal descriptions for FORMATTING "
              "ONLY. Never add, remove, or change any section, township, range, "
              "quarter-call, lot, or acreage. Only fix capitalization, spacing, and "
              "abbreviations (Section->Sec, Township 12 North->12N, Range 24 West->24W, "
              "Northeast Quarter->NE/4). If unsure, return the input unchanged. "
              "Return STRICT JSON: {\"cleaned\":[...]} same length/order as input.")
    payload = {
        "model": model, "max_tokens": 2000,
        "system": system,
        "messages": [{"role": "user", "content": json.dumps({"legals": uniq})}],
    }
    try:
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=json.dumps(payload).encode("utf-8"),
            headers={"content-type": "application/json", "x-api-key": api_key,
                     "anthropic-version": "2023-06-01"},
            method="POST")
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = json.loads(resp.read().decode("utf-8"))
        text = "".join(b.get("text", "") for b in body.get("content", []))
        cleaned = json.loads(re.search(r"\{.*\}", text, re.S).group(0)).get("cleaned", [])
        for orig, new in zip(uniq, cleaned):
            new = norm_text(new)
            if new:
                result[orig] = new
        LOG(f"  AI polish applied to {len(uniq)} legal descriptions.")
    except Exception as exc:
        LOG(f"  AI polish unavailable ({exc}); using deterministic formatting only.", "WARN")
    return result


# ============================================================================
# Tract tagging + OGL fill
# ============================================================================
def tag_tract(legal: str, tracts: List[TractRecord]) -> str:
    """Return the tract_no whose description best matches this row's legal."""
    if not tracts:
        return ""
    lg = standardize_legal(legal)
    best_no, best_score = "", 0.0
    for t in tracts:
        sc = similar(standardize_legal(t.description), lg)
        if sc > best_score:
            best_no, best_score = t.tract_no, sc
    return best_no if best_score >= 0.45 else ""


def collapse_near_duplicates(merged: List["v1.TitleRow"]
                             ) -> Tuple[List["v1.TitleRow"], List[Dict[str, str]]]:
    """Collapse rows that describe the SAME instrument but differ only in
    book/page/instrument_no (typical of a data-entry typo across folders).

    Two rows collapse ONLY when grantor, grantee, instrument date, doc type, and
    standardized legal all match. Distinct instruments (differing parties/date)
    are always preserved. Every collapse is recorded as a conflict for review."""
    def sig(row: "v1.TitleRow") -> str:
        return "|".join((
            norm_name(row.data.get("grantor", "")),
            norm_name(row.data.get("grantee", "")),
            str(norm_date(row.data.get("instrument_date", ""))
                or norm_date(row.data.get("recorded_date", "")) or ""),
            norm_text(row.data.get("doc_type", "")).upper(),
            standardize_legal(row.data.get("legal_description", "")).upper(),
        ))
    groups: Dict[str, List["v1.TitleRow"]] = {}
    order: List[str] = []
    for row in merged:
        s = sig(row)
        if s not in groups:
            groups[s] = []
            order.append(s)
        groups[s].append(row)

    out: List["v1.TitleRow"] = []
    conflicts: List[Dict[str, str]] = []
    for s in order:
        group = groups[s]
        if len(group) == 1:
            out.append(group[0])
            continue
        # never collapse when we cannot even identify the parties (too risky)
        if not norm_name(group[0].data.get("grantor", "")) and not norm_name(group[0].data.get("grantee", "")):
            out.extend(group)
            continue
        # collapse: keep the most complete row (most non-empty canonical fields)
        keep = max(group, key=lambda r: sum(1 for f in CANON_FIELDS if norm_text(r.data.get(f, ""))))
        for fld in ("book", "page", "instrument_no"):
            chosen = norm_text(keep.data.get(fld, ""))
            vals = sorted({norm_text(r.data.get(fld, "")) for r in group if norm_text(r.data.get(fld, ""))})
            if len(vals) > 1:
                conflicts.append({
                    "key": keep.key(), "field": fld,
                    "chosen": chosen or "(blank)",
                    "alternatives": " | ".join(v for v in vals if v != chosen),
                    "sources": " ; ".join(sorted({r.source for r in group})),
                })
        keep.source = ";".join(sorted({r.source for r in group}))
        out.append(keep)
    return out, conflicts


def apply_ogl_numbers(merged: List["v1.TitleRow"], ogls: List[OGLRecord]) -> Tuple[int, List[Dict[str, str]]]:
    """Fill/repair instrument_no on lease rows using the OGL sheet.
    Match by (book,page) first, then by lessor+date. Records every change."""
    changes: List[Dict[str, str]] = []
    if not ogls:
        return 0, changes
    by_bp: Dict[str, OGLRecord] = {}
    for o in ogls:
        bp = norm_doc_ref(o.book + o.page)
        if bp:
            by_bp[bp] = o
    filled = 0
    for row in merged:
        dtype = norm_text(row.data.get("doc_type", "")).lower()
        is_lease = "lease" in dtype or "ogl" in dtype
        if not is_lease:
            continue
        match: Optional[OGLRecord] = None
        bp = norm_doc_ref(row.data.get("book", "") + row.data.get("page", ""))
        if bp and bp in by_bp:
            match = by_bp[bp]
        else:
            for o in ogls:
                if (similar(norm_name(o.lessor), norm_name(row.data.get("grantor", ""))) >= 0.8
                        and norm_date(o.lease_date) == norm_date(row.data.get("instrument_date", ""))
                        and norm_date(o.lease_date) is not None):
                    match = o
                    break
        if match and match.ogl_no:
            old = norm_text(row.data.get("instrument_no", ""))
            if old != match.ogl_no:
                row.data["instrument_no"] = match.ogl_no
                filled += 1
                changes.append({"grantor": row.data.get("grantor", ""),
                                "grantee": row.data.get("grantee", ""),
                                "old_instrument": old or "(blank)",
                                "ogl_no": match.ogl_no, "ogl_source": match.source})
    return filled, changes


# ============================================================================
# Support writers
# ============================================================================
def write_csv(path: Path, header: List[str], rows: List[List[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)


def write_tract_sheet(path: Path, tracts: List[TractRecord]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tract Sheet (fixed)"
    hdr = ["Tract No", "Description (standardized)", "Original Description", "Gross Acres", "Source"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for t in tracts:
        ws.append([t.tract_no, standardize_legal(t.description), t.description, t.acres, t.source])
    ws.freeze_panes = "A2"
    for i, h in enumerate(hdr, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(16, min(60, len(h) + 20))
    wb.save(path)
    wb.close()


# ============================================================================
# Main
# ============================================================================
def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Roger Mills report builder v2 (codexv2)")
    ap.add_argument("--horizon", default=r"D:\Desktop\Horizon",
                    help="Parent folder holding the Roger Mills* folders")
    ap.add_argument("--roots", nargs="*", default=[],
                    help="Explicit folders to scan (overrides auto-discovery)")
    ap.add_argument("--out-dir", default="",
                    help="Output folder (default: <horizon>\\rogermillsfinalreports)")
    ap.add_argument("--template", default="", help="Explicit template .xlsx (else auto-detect)")
    ap.add_argument("--section", default="31-12N-24W")
    ap.add_argument("--use-llm", action="store_true",
                    help="Reformat legal descriptions via ANTHROPIC_API_KEY (formatting only)")
    ap.add_argument("--dry-run", action="store_true", help="Plan only; write nothing final")
    args = ap.parse_args(argv)

    horizon = Path(args.horizon)
    out_dir = Path(args.out_dir) if args.out_dir else horizon / "rogermillsfinalreports"

    LOG.section(f"Roger Mills Report Builder v2 - section {args.section}")
    LOG(f"Horizon:  {horizon}")
    LOG(f"Out dir:  {out_dir}")
    LOG(f"Caps: pandas={v1._HAVE_PANDAS} dateutil={v1._HAVE_DATEUTIL} "
        f"rapidfuzz={v1._HAVE_RAPIDFUZZ} pdfplumber={v1._HAVE_PDFPLUMBER} "
        f"pymupdf={v1._HAVE_PYMUPDF} ocr={v1._HAVE_OCR}")

    env = load_env(horizon)

    roots = discover_roots(horizon, args.roots)
    if not roots:
        LOG("FATAL: no Roger Mills* folders found under --horizon and no --roots given.", "ERROR")
        LOG("Are you running this on the machine where the files live?", "ERROR")
        return 2
    LOG.section("1. Folder discovery")
    for r in roots:
        LOG(f"  root: {r}")

    # ---- per-folder tournament (the "loops") ----
    LOG.section("2. Per-folder tournaments")
    folder_results: List[FolderResult] = []
    for r in roots:
        fr = analyze_folder(r)
        folder_results.append(fr)
        LOG(f"[{r.name}] files={len(fr.recs)} report-workbooks={len(fr.scored)}")
        for sc, wa in fr.scored:
            LOG(f"    score={sc:6.3f}  {wa.rec.path.name}")
        if fr.best:
            LOG(f"    -> folder winner: {fr.best.rec.path.name}")

    all_recs = [rec for fr in folder_results for rec in fr.recs]

    # ---- cross-folder tournament (the "tournament") ----
    LOG.section("3. Cross-folder tournament (overall best base)")
    newest = max((r.mtime.timestamp() for r in all_recs), default=_dt.datetime.now().timestamp())
    oldest = min((r.mtime.timestamp() for r in all_recs), default=newest)
    global_scored: List[Tuple[float, "v1.WorkbookAnalysis", Path]] = []
    for fr in folder_results:
        for wa in (fr.analyses[str(r.path)] for r in fr.recs
                   if r.kind == "report" and str(r.path) in fr.analyses):
            global_scored.append((v1.score_report(wa, newest, oldest), wa, fr.root))
    global_scored.sort(key=lambda t: t[0], reverse=True)
    for sc, wa, root in global_scored:
        LOG(f"  score={sc:6.3f}  [{root.name}] {wa.rec.path.name}")
    best_base = global_scored[0][1] if global_scored and global_scored[0][0] >= 0 else None

    # ---- template = formatting authority ----
    template_rec = None
    if args.template:
        tp = Path(args.template)
        template_rec = next((r for r in all_recs if r.path.name.lower() == tp.name.lower()), None)
        if template_rec is None and tp.exists():
            template_rec = v1.FileRec(path=tp, rel=tp.name, ext=tp.suffix.lower(),
                                      size=tp.stat().st_size,
                                      mtime=_dt.datetime.fromtimestamp(tp.stat().st_mtime),
                                      kind="template")
    if template_rec is None:
        # look in horizon root + all folders
        for search in [horizon, *roots]:
            for p in sorted(Path(search).glob("*.xls*")):
                if "template" in p.name.lower():
                    template_rec = v1.FileRec(path=p, rel=p.name, ext=p.suffix.lower(),
                                              size=p.stat().st_size,
                                              mtime=_dt.datetime.fromtimestamp(p.stat().st_mtime),
                                              kind="template")
                    break
            if template_rec:
                break
    if template_rec is None and best_base is not None:
        LOG("No Template(*) found; using overall best base as formatting authority.", "WARN")
        template_rec = best_base.rec
    if template_rec is None:
        LOG("FATAL: no template and no usable report workbook. Cannot build.", "ERROR")
        out_dir.mkdir(parents=True, exist_ok=True)
        LOG.dump(out_dir / "build_log_codexv2.txt")
        return 3
    LOG(f"Formatting authority: {template_rec.path.name}")
    if best_base:
        LOG(f"Best base data workbook: {best_base.rec.path.name}")

    # ---- extract + merge rows from ALL report workbooks in ALL folders ----
    LOG.section("4. Extract & merge runsheet data (all folders)")
    all_rows: List[List["v1.TitleRow"]] = []
    for fr in folder_results:
        for wa in (fr.analyses[str(r.path)] for r in fr.recs
                   if r.kind == "report" and str(r.path) in fr.analyses):
            rows = v1.extract_rows(wa)
            if rows:
                LOG(f"  [{fr.root.name}] {wa.rec.path.name}: {len(rows)} rows")
                all_rows.append(rows)
    merged, audit, conflicts = v1.merge_rows(all_rows)
    before = len(merged)
    merged, near_conflicts = collapse_near_duplicates(merged)
    conflicts += near_conflicts
    LOG(f"Merged unique records: {len(merged)} "
        f"(collapsed {before - len(merged)} near-duplicate typo rows) | "
        f"field conflicts: {len(conflicts)}")

    # ---- OGL / tract / title / notes from every workbook ----
    LOG.section("5. OGL / Tract / Title / Notes")
    ogls: List[OGLRecord] = []
    tracts: List[TractRecord] = []
    title_meta: Dict[str, str] = {}
    notes: List[str] = []
    for rec in all_recs:
        if rec.ext in v1.WORKBOOK_EXT:
            ogls += extract_ogl(rec)
            tracts += extract_tracts(rec)
            for k, v in extract_title_meta(rec).items():
                title_meta.setdefault(k, v)
            notes += extract_notes(rec)
    # de-dup tracts by (tract_no, standardized description)
    seen = set()
    uniq_tracts: List[TractRecord] = []
    for t in tracts:
        key = (t.tract_no.lower(), standardize_legal(t.description).lower())
        if key not in seen:
            seen.add(key)
            uniq_tracts.append(t)
    tracts = uniq_tracts
    LOG(f"OGL leases: {len(ogls)} | tracts: {len(tracts)} | "
        f"title fields: {len(title_meta)} | notes: {len(notes)}")

    # ---- fix legals (deterministic + optional AI) ----
    LOG.section("6. Standardize legal descriptions")
    all_legals = [norm_text(r.data.get("legal_description", "")) for r in merged]
    all_legals += [t.description for t in tracts]
    polish = {}
    if args.use_llm:
        key = env.get("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_API_KEY", "")
        if key:
            polish = llm_polish([l for l in all_legals if l], key)
        else:
            LOG("--use-llm set but no ANTHROPIC_API_KEY in .env/env; skipping AI polish.", "WARN")
    legal_changes: List[List[str]] = []
    for row in merged:
        orig = norm_text(row.data.get("legal_description", ""))
        new = polish.get(orig, standardize_legal(orig))
        if new and new != orig:
            legal_changes.append([orig, new, "; ".join(sorted(set(row.source.split(";"))))])
        row.data["legal_description"] = new or orig

    # ---- OGL number authority ----
    LOG.section("7. Apply OGL numbers")
    n_ogl_filled, ogl_changes = apply_ogl_numbers(merged, ogls)
    LOG(f"OGL numbers filled/repaired on {n_ogl_filled} lease rows.")

    # ---- tract tagging (put every row on the tract spine) ----
    for row in merged:
        tno = tag_tract(row.data.get("legal_description", ""), tracts)
        if tno:
            rem = norm_text(row.data.get("remarks", ""))
            tag = f"[{tno}]"
            if tag not in rem:
                row.data["remarks"] = (tag + " " + rem).strip()

    # ---- index verification ----
    LOG.section("8. Index PDF verification")
    index_text, method = "", "none"
    index_pdf = next((r for r in all_recs if r.kind == "index_pdf"), None)
    if index_pdf:
        LOG(f"Index PDF: {index_pdf.path.name}")
        index_text, method = v1.extract_pdf_text(index_pdf.path)
        LOG(f"Extraction method: {method}; chars={len(index_text)}")
    else:
        LOG("No index PDF found.", "WARN")
    verified = v1.verify_against_index(merged, index_text)
    n_verified = sum(1 for x in verified.values() if x)
    LOG(f"Rows verified vs index: {n_verified}/{len(merged)}")

    if args.dry_run:
        LOG.section("DRY-RUN complete (nothing written).")
        out_dir.mkdir(parents=True, exist_ok=True)
        LOG.dump(out_dir / "build_log_codexv2.txt")
        _final_console(out_dir, all_recs, merged, conflicts, n_verified, method,
                       n_ogl_filled, len(tracts), dry=True)
        return 0

    # ---- backups (of every original, once, before writing anything) ----
    LOG.section("9. Timestamped backups")
    backup_root = out_dir / "_backups"
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    import shutil
    copied = 0
    for rec in all_recs:
        try:
            dest = backup_root / stamp / rec.path.parent.name / rec.path.name
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(rec.path, dest)
            copied += 1
        except Exception as exc:
            LOG(f"  backup failed {rec.path.name}: {exc}", "WARN")
    LOG(f"Backed up {copied}/{len(all_recs)} originals -> {backup_root / stamp}")

    # ---- build final workbook from template ----
    LOG.section("10. Build final workbook (template formatting)")
    out_name = f"{args.section}_Roger_Mills_Cursory_Title_Report_codexv2.xlsx"
    output_path = out_dir / out_name
    data_sheet, rows_written = v1.build_output(
        template_rec.path, output_path, merged, verified, args.section)
    LOG(f"Wrote {rows_written} rows to '{data_sheet}' -> {output_path}")

    # ---- fix the title sheet + tract sheet inside the output ----
    _fix_output_extras(output_path, title_meta, tracts, args.section, notes)

    # ---- support files ----
    LOG.section("11. Support files")
    support = out_dir / "files"
    support.mkdir(parents=True, exist_ok=True)
    v1.write_inventory_csv(support / "source_inventory_codexv2.csv", all_recs,
                           {str(rec.path): fr.analyses.get(str(rec.path))
                            for fr in folder_results for rec in fr.recs
                            if str(rec.path) in fr.analyses})
    v1.write_audit_csv(support / "merge_audit_codexv2.csv", audit, verified)
    v1.write_conflicts_xlsx(support / "conflicts_review_codexv2.xlsx", conflicts)
    write_tract_sheet(support / "tract_sheet_fixed_codexv2.xlsx", tracts)
    write_csv(support / "ogl_number_changes_codexv2.csv",
              ["grantor", "grantee", "old_instrument", "ogl_no", "ogl_source"],
              [[c["grantor"], c["grantee"], c["old_instrument"], c["ogl_no"], c["ogl_source"]]
               for c in ogl_changes])
    write_csv(support / "legal_description_changes_codexv2.csv",
              ["original", "standardized", "sources"], legal_changes)
    write_csv(support / "notes_collected_codexv2.csv", ["note"], [[n] for n in notes])
    write_csv(support / "tournament_scores_codexv2.csv",
              ["scope", "folder", "workbook", "score"],
              [["overall", root.name, wa.rec.path.name, f"{sc:.4f}"]
               for sc, wa, root in global_scored])
    LOG("Wrote inventory, merge audit, conflicts, fixed tract sheet, OGL changes, "
        "legal changes, notes, tournament scores.")

    # ---- validation ----
    LOG.section("12. Validation")
    ok, val_notes = v1.validate_output(output_path, data_sheet, rows_written)
    for n in val_notes:
        LOG(f"  {n}")
    LOG(f"VALIDATION: {'PASS' if ok else 'FAIL - see notes'}")

    _write_summary(out_dir, args.section, all_recs, folder_results, global_scored,
                   template_rec, best_base, merged, conflicts, method, n_verified,
                   n_ogl_filled, len(tracts), title_meta, notes, val_notes, ok, output_path)
    LOG.dump(out_dir / "build_log_codexv2.txt")
    _final_console(out_dir, all_recs, merged, conflicts, n_verified, method,
                   n_ogl_filled, len(tracts), dry=False, output_path=output_path)
    return 0


def _fix_output_extras(output_path: Path, title_meta: Dict[str, str],
                       tracts: List[TractRecord], section: str, notes: List[str]) -> None:
    """Repair the Title Sheet metadata and rewrite the Tract Sheet in the output
    workbook, preserving the template's own styling."""
    try:
        wb = openpyxl.load_workbook(output_path)
    except Exception as exc:
        LOG(f"  could not reopen output to fix extras: {exc}", "WARN")
        return
    # Title sheet
    for ws in wb.worksheets:
        if _sheet_kind(ws.title) != "title":
            continue
        # longest labels first so "effective date" wins before "date", etc.
        want = [
            ("effective date", title_meta.get("effective_date", "")),
            ("effective", title_meta.get("effective_date", "")),
            ("prepared for", title_meta.get("prepared_for", "")),
            ("prepared", title_meta.get("prepared_for", "")),
            ("gross acres", title_meta.get("gross_acres", "")),
            ("section", title_meta.get("section") or section),
            ("county", title_meta.get("county") or "Roger Mills"),
            ("state", title_meta.get("state") or "Oklahoma"),
        ]
        for r in range(1, min(ws.max_row, 40) + 1):
            for c in range(1, min(ws.max_column, 6) + 1):
                label = norm_text(ws.cell(row=r, column=c).value).lower().rstrip(":")
                if not label:
                    continue
                val = next((v for lbl, v in want if (label == lbl or label.startswith(lbl)) and v), None)
                if val:
                    tgt = ws.cell(row=r, column=c + 1)
                    if not norm_text(tgt.value):
                        tgt.value = val
        LOG(f"  fixed title sheet '{ws.title}'")
    # Tract sheet (authoritative set preserved; formatting standardized)
    for ws in wb.worksheets:
        if _sheet_kind(ws.title) != "tract":
            continue
        hr, hmap = _map_headers(ws, {
            "tract_no": ("tract", "tract no", "no", "#"),
            "description": ("description", "legal", "land"),
            "acres": ("acres", "gross acres"),
        })
        col_desc = next((c for c, f in hmap.items() if f == "description"), None)
        if col_desc:
            for r in range(hr + 1, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_desc)
                if norm_text(cell.value):
                    cell.value = standardize_legal(cell.value)
            LOG(f"  standardized tract descriptions in '{ws.title}'")
        # if the template tract sheet is empty, seed it from collected tracts
        if col_desc and ws.max_row <= hr and tracts:
            for t in tracts:
                ws.append([t.tract_no, standardize_legal(t.description), t.acres])
    # Notes sheet
    for ws in wb.worksheets:
        if _sheet_kind(ws.title) != "notes":
            continue
        start = ws.max_row + 1 if norm_text(ws.cell(row=1, column=1).value) else 1
        for i, n in enumerate(notes):
            ws.cell(row=start + i, column=1, value=n)
        if notes:
            LOG(f"  wrote {len(notes)} collected notes into '{ws.title}'")
    wb.save(output_path)
    wb.close()


def _write_summary(out_dir, section, all_recs, folder_results, global_scored,
                   template_rec, best_base, merged, conflicts, method, n_verified,
                   n_ogl_filled, n_tracts, title_meta, notes, val_notes, ok, output_path):
    lines = [
        "Roger Mills Cursory Title Report - v2 final summary (codexv2)",
        f"Generated: {_dt.datetime.now():%Y-%m-%d %H:%M:%S}",
        f"Section: {section}",
        "",
        f"Folders scanned:        {len(folder_results)}",
        *[f"  - {fr.root}  (winner: "
          f"{fr.best.rec.path.name if fr.best else 'n/a'})" for fr in folder_results],
        f"Total source files:     {len(all_recs)}",
        f"Formatting authority:   {template_rec.path.name if template_rec else '(none)'}",
        f"Overall best base:      {best_base.rec.path.name if best_base else '(none)'}",
        f"Unique records merged:  {len(merged)}",
        f"Field conflicts:        {len(conflicts)}",
        f"OGL numbers applied:    {n_ogl_filled}",
        f"Tracts (spine):         {n_tracts}",
        f"Index method:           {method}",
        f"Rows verified vs index: {n_verified}/{len(merged)}",
        f"Notes collected:        {len(notes)}",
        f"Title metadata fields:  {len(title_meta)}",
        f"Output workbook:        {output_path}",
        "",
        "Validation:",
        *[f"  {l}" for l in val_notes],
        f"  -> {'PASS' if ok else 'FAIL'}",
        "",
        "HUMAN REVIEW NEEDED:",
        "  - Rows flagged [REVIEW: not found in index].",
        "  - All field conflicts in files\\conflicts_review_codexv2.xlsx.",
        "  - OGL number changes in files\\ogl_number_changes_codexv2.csv.",
        "  - Legal-description edits in files\\legal_description_changes_codexv2.csv.",
        "  - Spot-check acreage / legals against the source index.",
    ]
    (out_dir / "final_summary_codexv2.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    LOG("Wrote final_summary_codexv2.txt")


def _final_console(out_dir, all_recs, merged, conflicts, n_verified, method,
                   n_ogl_filled, n_tracts, dry=False, output_path=None):
    print("\n" + "=" * 70)
    print("FINAL REPORT (codexv2)" + ("  [DRY-RUN]" if dry else ""))
    print("=" * 70)
    print(f"Output folder:      {out_dir}")
    if output_path and not dry:
        print(f"Final workbook:     {output_path}")
    print(f"Source files:       {len(all_recs)}")
    print(f"Records merged:     {len(merged)}")
    print(f"OGL numbers filled: {n_ogl_filled}")
    print(f"Tracts (spine):     {n_tracts}")
    print(f"Conflicts:          {len(conflicts)}")
    print(f"Index verified:     {n_verified}/{len(merged)}  (method: {method})")
    print("Review:             see final_summary_codexv2.txt")


if __name__ == "__main__":
    raise SystemExit(main())
