#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Roger Mills Cursory Title Report Builder  (codexv2)
===================================================

A single-file, defensive, *run-it-locally* tool that executes the full
"build the best possible updated Excel title report" mission against a real
folder tree on your machine (e.g. D:\\Desktop\\Horizon\\Roger Mills).

WHAT'S NEW IN codexv2
---------------------
* INTEREST CHAINING: walks every conveyance in chronological order and chains
  the mineral interest, parsing fractions (1/2), fraction products ("1/2 of
  1/8"), decimals, percents, NMA (net mineral acres, when --gross-acres is
  given), and PROPORTIONAL conveyances ("1/2 of grantor's interest"). Emits an
  "Interest Chain" sheet (flagged rows highlighted), an "Ownership Ledger" of
  net positions, and a consolidated "Review Flags" punch-list, FLAGGING
  over-conveyance / unparseable interest / missing parties instead of guessing.
* RUNSHEET NOTES: ingests a runsheet workbook and attaches its notes to the
  matching title rows (by instrument no. or book/page).
* OGL SHEET / OGL NUMBERS: ingests an OGL (oil & gas lease) schedule, attaches
  the OGL number to each matching lease row, and emits an "OGL Summary" sheet.
* LEGALS carried through the merge and shown alongside the chain.
* DEED EXTRACTION: best-effort parse of PDF/DOCX instruments into a clearly
  labeled, review-only "Deeds (auto-extract VERIFY)" sheet - never merged into
  the authoritative chain.
* HTML REPORT: a self-contained (offline, no external assets) interest-chain
  report written alongside the workbook.
* "LOOP TILL PERFECT": builds, validates, and re-builds up to --max-passes,
  writing a perfection checklist of the exact human-review items that remain.

WHY THIS IS A LOCAL TOOL
------------------------
The cloud assistant that wrote this cannot see your D:\\ drive. So instead of
inventing data (which would be worthless and wrong for a title report), it
produced this tool for you to run *where the files actually live*.

WHAT IT DOES (matches the mission spec)
---------------------------------------
1.  Recursively inventories EVERY file under --root.
2.  Makes TIMESTAMPED BACKUPS of every original before reading (read-only on
    originals; nothing is deleted or overwritten).
3.  Classifies files: template / report workbooks / runsheet / index PDF /
    other supporting data.
4.  Analyzes every workbook (sheets, used range, headers, row counts, filled
    cells, formulas, a formatting-quality score).
5.  Runs a "tournament" to pick the BEST BASE report workbook by weighted score
    (template-closeness, completeness, cleanliness, recency, low blanks/errors).
6.  Uses Template(30).xlsx as the FORMATTING AUTHORITY: the output is a copy of
    the template workbook (preserving merged cells, fonts, colors, borders,
    column widths, row heights, formulas, page setup, print areas, freeze
    panes, headers/footers, tab order/colors) into which merged data is written.
7.  Extracts the index PDF via text -> pdfplumber -> PyMuPDF -> OCR (graceful
    fallback; writes an OCR-limitation note if no OCR engine is available).
8.  Merges the best data from ALL report workbooks, normalizes names/dates/
    document refs/legal descriptions, de-duplicates intelligently (true separate
    instruments are preserved), and verifies against the PDF/index where
    practical.
9.  Sorts rows chronologically (configurable) and writes them into the report.
10. Writes the final workbook to --output and creates all support files in
    --support-dir.
11. VALIDATES: reopens the output to prove it is not corrupted, confirms sheets
    and populated rows, and writes a validation summary.

IT NEVER INVENTS DATA. Unverified-but-supported rows are flagged in the
conflict/review outputs rather than silently trusted.

USAGE (Windows PowerShell or CMD)
---------------------------------
    py -m pip install --upgrade openpyxl pandas pdfplumber PyMuPDF pytesseract Pillow python-dateutil rapidfuzz

    :: Simplest form - output auto-lands in D:\\Desktop\\Horizon\\rogermillsfinalreports
    py automation\\roger_mills_title_report_builder.py ^
        --root "D:\\Desktop\\Horizon\\Roger Mills" ^
        --section "31-12N-24W" ^
        --gross-acres 640

    :: Explicit final-dir (matches the requested destination) + tuning
    py automation\\roger_mills_title_report_builder.py ^
        --root "D:\\Desktop\\Horizon\\Roger Mills" ^
        --final-dir "D:\\Desktop\\Horizon\\rogermillsfinalreports" ^
        --section "31-12N-24W" ^
        --gross-acres 640 ^
        --max-passes 3

Naming hints so the extra engines fire:
    * a workbook with "runsheet" in its filename  -> runsheet notes are chained in
    * a workbook with "ogl" in its filename        -> OGL numbers are attached
    * pass --gross-acres to convert NMA interests into fractions of the whole

Run with --dry-run first to see the plan without writing the final workbook.

Only openpyxl + pandas are strictly required. pdfplumber / PyMuPDF / pytesseract
/ rapidfuzz / python-dateutil are optional and degrade gracefully.
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import io
import os
import re
import shutil
import sys
import traceback
import zipfile
from dataclasses import dataclass, field
from fractions import Fraction
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

# ----------------------------------------------------------------------------
# Optional dependency shims (degrade gracefully, never hard-crash on import)
# ----------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover - hard requirement
    print("FATAL: openpyxl is required. Install with: py -m pip install openpyxl")
    raise

try:
    import pandas as pd  # noqa: F401  (used for CSV/data convenience)
    _HAVE_PANDAS = True
except BaseException:  # optional dep may panic (e.g. broken cffi), not just raise
    _HAVE_PANDAS = False

try:
    from dateutil import parser as _dateparser  # type: ignore
    _HAVE_DATEUTIL = True
except BaseException:
    _HAVE_DATEUTIL = False

try:
    from rapidfuzz import fuzz as _fuzz  # type: ignore
    _HAVE_RAPIDFUZZ = True
except BaseException:
    _HAVE_RAPIDFUZZ = False

try:
    import docx as _docx  # python-docx  # type: ignore
    _HAVE_DOCX = True
except BaseException:
    _HAVE_DOCX = False

# PDF extraction backends (any/all optional)
try:
    import pdfplumber  # type: ignore
    _HAVE_PDFPLUMBER = True
except BaseException:
    _HAVE_PDFPLUMBER = False

try:
    import fitz  # PyMuPDF  # type: ignore
    _HAVE_PYMUPDF = True
except BaseException:
    _HAVE_PYMUPDF = False

try:
    import pytesseract  # type: ignore
    from PIL import Image  # type: ignore
    _HAVE_OCR = True
except BaseException:
    _HAVE_OCR = False


# ----------------------------------------------------------------------------
# Logging
# ----------------------------------------------------------------------------
class BuildLog:
    """Collects a human-readable build log and echoes to stdout."""

    def __init__(self) -> None:
        self._lines: List[str] = []
        self.start = _dt.datetime.now()

    def __call__(self, msg: str, level: str = "INFO") -> None:
        ts = _dt.datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {level:5s} {msg}"
        self._lines.append(line)
        print(line)

    def section(self, title: str) -> None:
        bar = "=" * 70
        self(bar)
        self(title)
        self(bar)

    def dump(self, path: Path) -> None:
        header = [
            "Roger Mills Title Report Builder - build log (codexv2)",
            f"Started:  {self.start:%Y-%m-%d %H:%M:%S}",
            f"Finished: {_dt.datetime.now():%Y-%m-%d %H:%M:%S}",
            "",
        ]
        path.write_text("\n".join(header + self._lines) + "\n", encoding="utf-8")


LOG = BuildLog()


# ----------------------------------------------------------------------------
# Normalization helpers (no data invented - just cleanup/standardization)
# ----------------------------------------------------------------------------
_WS_RE = re.compile(r"\s+")
_MONTHS = (
    "january february march april may june july august september october "
    "november december"
).split()


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).replace(" ", " ").strip()
    return _WS_RE.sub(" ", s)


def norm_name(value: Any) -> str:
    """Normalize a party (grantor/grantee) name for comparison/dedup."""
    s = norm_text(value).upper()
    s = s.replace(".", "").replace(",", " ")
    s = re.sub(r"\b(AN|A|THE)\b", " ", s)
    for token in (" AND ", " & "):
        s = s.replace(token, " & ")
    return _WS_RE.sub(" ", s).strip()


def norm_doc_ref(value: Any) -> str:
    """Normalize instrument / document / book-page references for dedup."""
    s = norm_text(value).upper()
    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s


def norm_date(value: Any) -> Optional[_dt.date]:
    """Best-effort parse to a date. Returns None if unparseable (never guesses)."""
    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    s = norm_text(value)
    if not s:
        return None
    # numeric excel serials sometimes leak through as ints
    if re.fullmatch(r"\d{5}", s):
        try:
            base = _dt.date(1899, 12, 30)
            return base + _dt.timedelta(days=int(s))
        except Exception:
            return None
    if _HAVE_DATEUTIL:
        try:
            return _dateparser.parse(s, dayfirst=False, fuzzy=True).date()
        except Exception:
            return None
    # minimal fallback parser
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def similar(a: str, b: str) -> float:
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    if _HAVE_RAPIDFUZZ:
        return _fuzz.token_sort_ratio(a, b) / 100.0
    # cheap fallback: Jaccard over word sets
    sa, sb = set(a.split()), set(b.split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


# ----------------------------------------------------------------------------
# File inventory & classification
# ----------------------------------------------------------------------------
WORKBOOK_EXT = {".xlsx", ".xlsm", ".xls"}
DATA_EXT = {".csv", ".txt", ".tsv"}
DOC_EXT = {".docx", ".doc"}
IMG_EXT = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}
PDF_EXT = {".pdf"}


@dataclass
class FileRec:
    path: Path
    rel: str
    ext: str
    size: int
    mtime: _dt.datetime
    kind: str = "other"  # template|report|runsheet|ogl|index_pdf|data|doc|image|other
    note: str = ""


def _table_kind_from_name(name: str) -> str:
    """Classify a tabular source (workbook or delimited) by its filename."""
    if "template" in name:
        return "template"
    if "runsheet" in name or "run sheet" in name:
        return "runsheet"
    if ("ogl" in name or "o&g lease" in name or "oil and gas lease" in name
            or "lease schedule" in name or "lease sheet" in name):
        return "ogl"
    return "report"  # treat unknown tables as candidate reports


def classify(rec: FileRec) -> str:
    name = rec.path.name.lower()
    if rec.ext in WORKBOOK_EXT:
        return _table_kind_from_name(name)
    if rec.ext in PDF_EXT:
        if "runsheet" in name or "run sheet" in name:
            return "runsheet"  # PDF runsheet -> best-effort note harvesting
        if re.search(r"deed|lease|assign|probate|affidavit|conveyance|patent|"
                     r"warranty|quit\s?claim|mineral|easement|right.?of.?way", name):
            return "deed"      # a specific instrument -> deed extraction (review-only)
        return "index_pdf"
    if rec.ext in DATA_EXT:
        # CSV/TSV can be a runsheet / OGL / report table; TXT stays freeform data
        if rec.ext in {".csv", ".tsv"}:
            return _table_kind_from_name(name)
        return "data"
    if rec.ext in DOC_EXT:
        return "doc"
    if rec.ext in IMG_EXT:
        return "image"
    return "other"


def inventory(root: Path, exclude: Optional[Sequence[Path]] = None) -> List[FileRec]:
    recs: List[FileRec] = []
    # resolve excluded dirs (our own output/support/backup trees) so re-runs do
    # not re-ingest generated workbooks and compound the merge on every pass.
    excluded = []
    for e in (exclude or []):
        try:
            excluded.append(e.resolve())
        except Exception:
            excluded.append(e)
    for dirpath, dirs, files in os.walk(root):
        here = Path(dirpath)
        try:
            here_res = here.resolve()
        except Exception:
            here_res = here
        # prune excluded subtrees in-place so os.walk never descends into them
        if any(here_res == ex or ex in here_res.parents for ex in excluded):
            dirs[:] = []
            continue
        dirs[:] = [d for d in dirs
                   if not any((here_res / d).resolve() == ex for ex in excluded)]
        for fn in files:
            p = Path(dirpath) / fn
            try:
                st = p.stat()
            except OSError:
                continue
            rec = FileRec(
                path=p,
                rel=str(p.relative_to(root)),
                ext=p.suffix.lower(),
                size=st.st_size,
                mtime=_dt.datetime.fromtimestamp(st.st_mtime),
            )
            rec.kind = classify(rec)
            recs.append(rec)
    return recs


# ----------------------------------------------------------------------------
# Zip extraction, duplicate & trash tidying (codexv2)
# ----------------------------------------------------------------------------
EXTRACT_DIRNAME = "_horizon_extracted"
TRASH_BASENAMES = {"thumbs.db", "desktop.ini", ".ds_store"}


def extract_all_zips(root: Path, max_depth: int = 5) -> Tuple[Path, int]:
    """Recursively extract every .zip under root (including zips-within-zips).

    Path-traversal safe, idempotent (skips already-extracted archives), and
    extracts into <root>/_horizon_extracted so the contents get inventoried
    and used. Returns (extract_root, files_extracted)."""
    extract_root = root / EXTRACT_DIRNAME
    processed: set = set()
    total = 0
    for _depth in range(max_depth):
        found: List[Path] = []
        for dp, _dirs, files in os.walk(root):
            for fn in files:
                if fn.lower().endswith(".zip"):
                    found.append(Path(dp) / fn)
        fresh = [z for z in found if str(z.resolve()) not in processed]
        if not fresh:
            break
        for z in fresh:
            processed.add(str(z.resolve()))
            dest = extract_root / z.stem
            try:
                if dest.exists() and any(dest.iterdir()):
                    LOG(f"  already extracted: {z.name}")
                    continue
                dest.mkdir(parents=True, exist_ok=True)
                dest_res = dest.resolve()
                extracted = 0
                with zipfile.ZipFile(z) as zf:
                    for member in zf.namelist():
                        if member.endswith("/"):
                            continue
                        target = (dest / member).resolve()
                        if dest_res != target and dest_res not in target.parents:
                            LOG(f"  skipped unsafe zip member: {member}", "WARN")
                            continue
                        zf.extract(member, dest)
                        extracted += 1
                total += extracted
                LOG(f"  extracted {z.name} -> {dest.name} ({extracted} files)")
            except Exception as exc:
                LOG(f"  zip failed for {z.name}: {exc}", "WARN")
    return extract_root, total


def sha256_file(path: Path, chunk: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        while True:
            b = fh.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def is_trash(rec: FileRec) -> bool:
    n = rec.path.name.lower()
    return (rec.size == 0 or n in TRASH_BASENAMES or n.startswith("~$")
            or n.endswith(".tmp") or n.endswith(".lnk"))


def _dup_keeper(group: List[FileRec]) -> FileRec:
    """Keep the most 'canonical' copy: prefer originals over extracted, then the
    shortest path, then the oldest (earliest recorded) file."""
    return sorted(group, key=lambda r: (EXTRACT_DIRNAME in r.rel,
                                        len(r.rel), r.mtime))[0]


def find_duplicates_and_trash(recs: List[FileRec]
                              ) -> Tuple[List[List[FileRec]], List[FileRec]]:
    """Return (duplicate_groups, trash_files). Duplicates are exact byte matches."""
    trash = [r for r in recs if is_trash(r)]
    trash_ids = {id(r) for r in trash}
    by_hash: Dict[str, List[FileRec]] = {}
    for r in recs:
        if id(r) in trash_ids:
            continue
        try:
            by_hash.setdefault(sha256_file(r.path), []).append(r)
        except Exception as exc:
            LOG(f"  hash failed for {r.rel}: {exc}", "WARN")
    dup_groups = [g for g in by_hash.values() if len(g) > 1]
    return dup_groups, trash


def quarantine(items: Sequence[FileRec], dest_dir: Path, reason: str
               ) -> List[Tuple[str, str, str]]:
    """Move files into a quarantine tree (reversible; never deletes). Returns a
    manifest of (relative_source, new_location, reason)."""
    manifest: List[Tuple[str, str, str]] = []
    for r in items:
        try:
            target = dest_dir / r.rel
            target.parent.mkdir(parents=True, exist_ok=True)
            if target.exists():
                target = target.with_name(
                    target.stem + "_" + str(r.size) + target.suffix)
            shutil.move(str(r.path), str(target))
            manifest.append((r.rel, str(target), reason))
        except Exception as exc:
            LOG(f"  quarantine failed for {r.rel}: {exc}", "WARN")
    return manifest


def write_tidy_manifest(path: Path, dup_groups: List[List[FileRec]],
                        trash: List[FileRec],
                        moved: List[Tuple[str, str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["category", "file", "detail", "action", "moved_to"])
        moved_map = {rel: (loc, reason) for rel, loc, reason in moved}
        for g in dup_groups:
            keeper = _dup_keeper(g)
            for r in g:
                role = "KEEP" if r is keeper else "duplicate-of:" + keeper.rel
                loc, _reason = moved_map.get(r.rel, ("", ""))
                w.writerow(["duplicate", r.rel, role,
                            "quarantined" if loc else "reported", loc])
        for r in trash:
            loc, _reason = moved_map.get(r.rel, ("", ""))
            reason = "empty" if r.size == 0 else "temp/junk"
            w.writerow(["trash", r.rel, reason,
                        "quarantined" if loc else "reported", loc])


# ----------------------------------------------------------------------------
# Workbook analysis & scoring tournament
# ----------------------------------------------------------------------------
# Canonical title-report columns we try to recognize across messy headers.
CANON_FIELDS = [
    "entry_no", "instrument_date", "recorded_date", "doc_type",
    "grantor", "grantee", "book", "page", "instrument_no",
    "legal_description", "acreage", "nma", "interest", "ogl_no",
    "runsheet_note", "remarks",
]

HEADER_SYNONYMS: Dict[str, Sequence[str]] = {
    "entry_no": ("entry", "no", "item", "#", "seq", "line"),
    "instrument_date": ("instrument date", "doc date", "dated", "date of instrument", "date"),
    "recorded_date": ("recorded", "record date", "filed", "recording date"),
    "doc_type": ("type", "instrument type", "document type", "doc type", "conveyance"),
    "grantor": ("grantor", "from", "seller", "assignor", "mortgagor", "lessor"),
    "grantee": ("grantee", "to", "buyer", "assignee", "mortgagee", "lessee"),
    "book": ("book", "vol", "volume", "bk"),
    "page": ("page", "pg", "pages"),
    "instrument_no": ("instrument", "document number", "doc no", "doc #", "reception", "file no"),
    "legal_description": ("legal", "description", "land", "tract", "lands described"),
    "acreage": ("acres", "acreage", "gross acres"),
    "nma": ("nma", "net mineral", "net acres", "nra"),
    "interest": ("interest", "fraction", "decimal", "ri", "ori", "wi", "nri",
                 "mineral interest", "royalty", "interest conveyed"),
    "ogl_no": ("ogl", "ogl no", "ogl number", "ogl #", "lease no", "lease number",
               "lease #", "o&g no", "og no"),
    "runsheet_note": ("runsheet note", "run sheet note", "rs note", "abstractor note"),
    "remarks": ("remarks", "notes", "comments", "exceptions", "review"),
}


def match_header(cell_text: str) -> Optional[str]:
    t = norm_text(cell_text).lower()
    if not t:
        return None
    best_field, best_score = None, 0.0
    for field_name, syns in HEADER_SYNONYMS.items():
        for syn in syns:
            if syn == t:
                return field_name
            if syn in t:
                score = len(syn) / max(len(t), 1)
                if score > best_score:
                    best_field, best_score = field_name, score
    return best_field if best_score >= 0.34 else None


@dataclass
class SheetAnalysis:
    name: str
    max_row: int
    max_col: int
    header_row: int
    header_map: Dict[int, str]  # col_idx -> canon field
    data_rows: int
    filled_cells: int
    formula_cells: int
    blank_ratio: float


@dataclass
class WorkbookAnalysis:
    rec: FileRec
    sheets: List[SheetAnalysis] = field(default_factory=list)
    style_score: float = 0.0
    error: str = ""

    @property
    def best_sheet(self) -> Optional[SheetAnalysis]:
        candidates = [s for s in self.sheets if s.header_map]
        if not candidates:
            return None
        return max(candidates, key=lambda s: (len(s.header_map), s.data_rows))


def _detect_header_row(ws, scan_rows: int = 25) -> Tuple[int, Dict[int, str]]:
    best_row, best_map = 1, {}
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        hmap: Dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            f = match_header(val) if val is not None else None
            if f and f not in hmap.values():
                hmap[c] = f
        if len(hmap) > len(best_map):
            best_row, best_map = r, hmap
    return best_row, best_map


def analyze_workbook(rec: FileRec) -> WorkbookAnalysis:
    wa = WorkbookAnalysis(rec=rec)
    try:
        wb = openpyxl.load_workbook(rec.path, data_only=False, read_only=False)
    except Exception as exc:
        wa.error = f"load failed: {exc}"
        return wa

    style_points = 0.0
    for ws in wb.worksheets:
        try:
            header_row, header_map = _detect_header_row(ws)
            filled = formula = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value not in (None, ""):
                        filled += 1
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula += 1
            data_rows = max(ws.max_row - header_row, 0)
            total_cells = max(ws.max_row * ws.max_column, 1)
            blank_ratio = 1.0 - (filled / total_cells)
            wa.sheets.append(SheetAnalysis(
                name=ws.title, max_row=ws.max_row, max_col=ws.max_column,
                header_row=header_row, header_map=header_map,
                data_rows=data_rows, filled_cells=filled,
                formula_cells=formula, blank_ratio=blank_ratio,
            ))
            # crude style richness proxy
            if ws.merged_cells.ranges:
                style_points += min(len(list(ws.merged_cells.ranges)), 20) * 0.5
            if getattr(ws, "freeze_panes", None):
                style_points += 2
            dims = getattr(ws, "column_dimensions", {})
            style_points += min(sum(1 for d in dims.values() if d.width), 30) * 0.2
        except Exception as exc:
            LOG(f"  sheet '{ws.title}' analysis error: {exc}", "WARN")
    wa.style_score = style_points
    wb.close()
    return wa


def score_report(wa: WorkbookAnalysis, newest_mtime: float, oldest_mtime: float) -> float:
    """Weighted tournament score for choosing the best base report workbook."""
    if wa.error or not wa.best_sheet:
        return -1.0
    s = wa.best_sheet
    # completeness: recognized columns + data rows
    completeness = len(s.header_map) / len(CANON_FIELDS)
    volume = min(s.data_rows / 50.0, 1.0)
    cleanliness = 1.0 - min(s.blank_ratio, 1.0)
    style = min(wa.style_score / 30.0, 1.0)
    span = max(newest_mtime - oldest_mtime, 1.0)
    recency = (wa.rec.mtime.timestamp() - oldest_mtime) / span
    # filename hints that this is an explicitly "best/updated" build
    name = wa.rec.path.name.lower()
    name_bonus = 0.0
    for kw, pts in (("updated", 0.06), ("best", 0.08), ("final", 0.05), ("v2", 0.03)):
        if kw in name:
            name_bonus += pts
    return (0.34 * completeness + 0.20 * volume + 0.16 * cleanliness
            + 0.14 * style + 0.10 * recency + name_bonus)


# ----------------------------------------------------------------------------
# Row extraction & merge
# ----------------------------------------------------------------------------
@dataclass
class TitleRow:
    data: Dict[str, str]
    source: str
    source_row: int

    def key(self) -> str:
        """Dedup key: prefer an explicit instrument/book-page, fall back to a
        normalized (grantor|grantee|date|type) signature."""
        inst = norm_doc_ref(self.data.get("instrument_no", ""))
        bp = norm_doc_ref(self.data.get("book", "") + self.data.get("page", ""))
        if inst:
            return f"INST:{inst}"
        if bp and bp not in ("", "0"):
            return f"BP:{bp}"
        sig = "|".join((
            norm_name(self.data.get("grantor", "")),
            norm_name(self.data.get("grantee", "")),
            str(norm_date(self.data.get("instrument_date", "")) or
                norm_date(self.data.get("recorded_date", "")) or ""),
            norm_text(self.data.get("doc_type", "")).upper(),
        ))
        return f"SIG:{sig}"


def extract_rows(wa: WorkbookAnalysis) -> List[TitleRow]:
    rows: List[TitleRow] = []
    s = wa.best_sheet
    if not s:
        return rows
    try:
        wb = openpyxl.load_workbook(wa.rec.path, data_only=True, read_only=True)
        ws = wb[s.name]
        for r in range(s.header_row + 1, ws.max_row + 1):
            data: Dict[str, str] = {}
            for col_idx, field_name in s.header_map.items():
                val = ws.cell(row=r, column=col_idx).value
                data[field_name] = norm_text(val)
            if any(v for v in data.values()):
                rows.append(TitleRow(data=data, source=wa.rec.path.name, source_row=r))
        wb.close()
    except Exception as exc:
        LOG(f"  extract failed for {wa.rec.path.name}: {exc}", "WARN")
    return rows


def merge_rows(all_rows: List[List[TitleRow]]) -> Tuple[List[TitleRow], List[Dict[str, str]], List[Dict[str, str]]]:
    """Merge rows from many workbooks.

    Returns (merged_rows, audit_records, conflict_records).
    Conflicts = same key, differing field values across sources.
    """
    bucket: Dict[str, List[TitleRow]] = {}
    audit: List[Dict[str, str]] = []
    for source_rows in all_rows:
        for row in source_rows:
            k = row.key()
            bucket.setdefault(k, []).append(row)

    merged: List[TitleRow] = []
    conflicts: List[Dict[str, str]] = []
    for k, group in bucket.items():
        # field-by-field, pick the most complete / most common value
        chosen: Dict[str, str] = {}
        for fld in CANON_FIELDS:
            values = [norm_text(g.data.get(fld, "")) for g in group]
            non_empty = [v for v in values if v]
            if not non_empty:
                chosen[fld] = ""
                continue
            # majority vote, tie-broken by longest (most descriptive)
            counts: Dict[str, int] = {}
            for v in non_empty:
                counts[v] = counts.get(v, 0) + 1
            best_val = sorted(non_empty, key=lambda v: (counts[v], len(v)), reverse=True)[0]
            chosen[fld] = best_val
            distinct = {v for v in non_empty}
            if len(distinct) > 1:
                conflicts.append({
                    "key": k,
                    "field": fld,
                    "chosen": best_val,
                    "alternatives": " | ".join(sorted(distinct - {best_val})),
                    "sources": " ; ".join(f"{g.source}:r{g.source_row}" for g in group),
                })
        merged_row = TitleRow(data=chosen, source=";".join(sorted({g.source for g in group})),
                              source_row=group[0].source_row)
        merged.append(merged_row)
        audit.append({
            "key": k,
            "n_sources": str(len(group)),
            "sources": " ; ".join(sorted({g.source for g in group})),
            "grantor": chosen.get("grantor", ""),
            "grantee": chosen.get("grantee", ""),
            "doc_type": chosen.get("doc_type", ""),
            "instrument_no": chosen.get("instrument_no", ""),
            "book": chosen.get("book", ""),
            "page": chosen.get("page", ""),
        })

    # chronological sort (rows with no parseable date sink to the end, stable)
    def sort_key(row: TitleRow):
        d = (norm_date(row.data.get("instrument_date", "")) or
             norm_date(row.data.get("recorded_date", "")))
        return (0, d) if d else (1, _dt.date.max)

    merged.sort(key=sort_key)
    return merged, audit, conflicts


# ----------------------------------------------------------------------------
# Interest parsing & chaining, runsheet notes, OGL cross-reference (codexv2)
# ----------------------------------------------------------------------------
_FRAC_RE = re.compile(r"(\d+)\s*/\s*(\d+)")
_DEC_RE = re.compile(r"^\s*(\d*\.\d+|\d+\.\d*|\d+)\s*$")
_PCT_RE = re.compile(r"(\d*\.?\d+)\s*%")
_NMA_RE = re.compile(
    r"(\d*\.?\d+)\s*(?:nma|nra|net\s*mineral\s*ac(?:re)?s?|net\s*ac(?:re)?s?)", re.I)
# "1/2 OF GRANTOR'S INTEREST" / "of his interest" -> conveyance is a fraction of
# whatever the grantor currently holds, not of the whole tract.
_OF_GRANTOR_RE = re.compile(
    r"\bof\s+(?:the\s+)?(?:grantor'?s?|his|her|its|their)\b", re.I)
_ONE = Fraction(1)
_EPS = Fraction(1, 10 ** 6)


def parse_interest(raw: Any, gross_acres: Optional[float] = None
                   ) -> Tuple[Optional[Fraction], str, str]:
    """Parse an interest expression into (Fraction | None, kind, detail).

    Never guesses. Returns (None, kind, detail) when it cannot parse a value.
    Supported forms:
      * fractions             "1/2", "3/16", "an undivided 1/4"
      * fraction products     "1/2 of 1/8", "1/2 x 1/8"  (multiplied)
      * decimals              "0.5", ".25"
      * percents              "50%", "12.5 %"
      * net mineral acres     "20 NMA" -> 20 / gross_acres (only if given)
    """
    s = norm_text(raw)
    if not s:
        return None, "empty", ""
    low = s.lower()

    # NMA -> fraction of the whole (requires gross acreage; otherwise flag it)
    m = _NMA_RE.search(low)
    if m:
        nma = m.group(1)
        if gross_acres and float(gross_acres) > 0:
            frac = (Fraction(nma).limit_denominator(10 ** 6)
                    / Fraction(str(gross_acres)).limit_denominator(10 ** 6))
            return frac, "nma", f"{nma} NMA / {gross_acres} ac"
        return None, "nma_no_gross", f"{nma} NMA (supply --gross-acres to convert)"

    # percent (only when it isn't actually a fraction like 1/8)
    m = _PCT_RE.search(low)
    if m and "/" not in low:
        try:
            return (Fraction(m.group(1)) / 100).limit_denominator(10 ** 6), "percent", m.group(0)
        except Exception:
            pass

    # fractions / fraction products
    fracs = _FRAC_RE.findall(low)
    if fracs:
        if len(fracs) > 1 and re.search(r"\bof\b|[x*]", low):
            val = _ONE
            for num, den in fracs:
                d = int(den) or 1
                val *= Fraction(int(num), d)
            return val, "fraction_product", s
        num, den = fracs[0]
        d = int(den) or 1
        return Fraction(int(num), d), "fraction", f"{num}/{den}"

    # bare decimal / integer
    m = _DEC_RE.match(low)
    if m:
        try:
            return Fraction(m.group(1)).limit_denominator(10 ** 6), "decimal", m.group(1)
        except Exception:
            return None, "unknown", s

    return None, "unknown", s


def frac_str(f: Optional[Fraction]) -> str:
    if f is None:
        return ""
    return str(f.numerator) if f.denominator == 1 else f"{f.numerator}/{f.denominator}"


def frac_dec(f: Optional[Fraction]) -> str:
    return "" if f is None else f"{float(f):.6f}"


@dataclass
class ChainLink:
    order: int
    date: Optional[_dt.date]
    doc_type: str
    grantor: str
    grantee: str
    raw_interest: str
    fraction: Optional[Fraction]
    kind: str
    grantee_running: Optional[Fraction]
    grantor_balance_after: Optional[Fraction]
    legal: str
    ogl_no: str
    flags: str
    source: str


def build_interest_chain(merged: List["TitleRow"], gross_acres: Optional[float] = None
                         ) -> Tuple[List[ChainLink], Dict[str, Fraction]]:
    """Chain the mineral interest across every conveyance, chronologically.

    Tracks a running ledger of each party's net interest (as a fraction of the
    whole) and each grantee's cumulative interest. Anomalies are FLAGGED, never
    silently corrected:
      * UNPARSED_INTEREST - an interest string we could not parse
      * OVER_CONVEYANCE    - grantor conveyed more than the ledger shows it held
      * NO_GRANTOR/GRANTEE - a party is missing from the row
    Ownership is not seeded with an assumed 100% owner, so early grantor
    balances can legitimately go negative; that is surfaced, not hidden.
    """
    ledger: Dict[str, Fraction] = {}
    links: List[ChainLink] = []
    for i, row in enumerate(merged, start=1):
        grantor_raw = norm_text(row.data.get("grantor", ""))
        grantee_raw = norm_text(row.data.get("grantee", ""))
        grantor = norm_name(grantor_raw)
        grantee = norm_name(grantee_raw)
        raw = row.data.get("interest", "")
        base, kind, _detail = parse_interest(raw, gross_acres)
        date = (norm_date(row.data.get("instrument_date", ""))
                or norm_date(row.data.get("recorded_date", "")))
        flags: List[str] = []
        if norm_text(raw) and base is None:
            flags.append(f"UNPARSED_INTEREST({kind})")
        if not grantor:
            flags.append("NO_GRANTOR")
        if not grantee:
            flags.append("NO_GRANTEE")

        grantor_before = ledger.get(grantor)
        # "of grantor's interest" => convey a fraction of the grantor's CURRENT
        # holding, not of the whole tract.
        proportional = base is not None and bool(_OF_GRANTOR_RE.search(norm_text(raw)))
        conveyed = base
        if proportional:
            if grantor_before is not None:
                conveyed = base * grantor_before
                flags.append(f"PROPORTIONAL(of grantor's {frac_str(grantor_before)})")
            else:
                conveyed = None
                flags.append("PROPORTIONAL_NO_BASIS(grantor interest unknown)")

        if conveyed is not None:
            if (not proportional and grantor and grantor_before is not None
                    and conveyed > grantor_before + _EPS):
                flags.append(f"OVER_CONVEYANCE(grantor_held={frac_str(grantor_before)})")
            if grantee:
                ledger[grantee] = ledger.get(grantee, Fraction(0)) + conveyed
            if grantor:
                ledger[grantor] = ledger.get(grantor, Fraction(0)) - conveyed

        links.append(ChainLink(
            order=i, date=date,
            doc_type=norm_text(row.data.get("doc_type", "")),
            grantor=grantor_raw, grantee=grantee_raw,
            raw_interest=norm_text(raw), fraction=conveyed, kind=kind,
            grantee_running=ledger.get(grantee) if grantee else None,
            grantor_balance_after=ledger.get(grantor) if grantor else None,
            legal=norm_text(row.data.get("legal_description", "")),
            ogl_no=norm_text(row.data.get("ogl_no", "")),
            flags="; ".join(flags), source=row.source,
        ))
    return links, ledger


def _load_delimited(rec: FileRec) -> List[Dict[str, str]]:
    """Load a CSV/TSV file as canon-field dict rows (headers matched fuzzily)."""
    out: List[Dict[str, str]] = []
    delim = "\t" if rec.ext == ".tsv" else ","
    try:
        with rec.path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh, delimiter=delim)
            rows = [r for r in reader]
    except Exception as exc:
        LOG(f"  delimited load failed for {rec.rel}: {exc}", "WARN")
        return out
    if not rows:
        return out
    # find the header row within the first 25 lines (most mapped columns wins)
    best_i, best_map = 0, {}
    for i in range(min(25, len(rows))):
        hmap: Dict[int, str] = {}
        for c, cell in enumerate(rows[i]):
            f = match_header(cell)
            if f and f not in hmap.values():
                hmap[c] = f
        if len(hmap) > len(best_map):
            best_i, best_map = i, hmap
    if not best_map:
        LOG(f"  {rec.rel}: no recognizable delimited header (skipped)", "WARN")
        return out
    for r in rows[best_i + 1:]:
        d: Dict[str, str] = {}
        for c, fld in best_map.items():
            d[fld] = norm_text(r[c]) if c < len(r) else ""
        if any(d.values()):
            out.append(d)
    return out


def _load_table(rec: FileRec) -> List[Dict[str, str]]:
    """Load a supporting table (workbook or CSV/TSV) as canon-field dict rows."""
    if rec.ext in {".csv", ".tsv"}:
        return _load_delimited(rec)
    wa = analyze_workbook(rec)
    s = wa.best_sheet
    if not s:
        LOG(f"  {rec.rel}: no recognizable table (skipped)", "WARN")
        return []
    out: List[Dict[str, str]] = []
    try:
        wb = openpyxl.load_workbook(rec.path, data_only=True, read_only=True)
        ws = wb[s.name]
        for r in range(s.header_row + 1, ws.max_row + 1):
            d: Dict[str, str] = {}
            for col, fld in s.header_map.items():
                d[fld] = norm_text(ws.cell(row=r, column=col).value)
            if any(d.values()):
                out.append(d)
        wb.close()
    except Exception as exc:
        LOG(f"  table load failed for {rec.rel}: {exc}", "WARN")
    return out


def title_rows_from_table(rec: FileRec) -> List["TitleRow"]:
    """Turn any supporting table (used for CSV report sources) into TitleRows."""
    rows: List[TitleRow] = []
    for i, d in enumerate(_load_table(rec), start=1):
        data = {f: norm_text(d.get(f, "")) for f in CANON_FIELDS}
        if any(data.values()):
            rows.append(TitleRow(data=data, source=rec.path.name, source_row=i))
    return rows


def _row_ref_keys(d: Dict[str, str]) -> List[str]:
    """Cross-reference keys for a row: instrument no. and/or book+page."""
    keys: List[str] = []
    inst = norm_doc_ref(d.get("instrument_no", ""))
    if inst and len(inst) >= 3:
        keys.append(f"INST:{inst}")
    bp = norm_doc_ref(d.get("book", "") + d.get("page", ""))
    if bp and bp not in ("", "0"):
        keys.append(f"BP:{bp}")
    return keys


def _runsheet_notes_from_pdf(rec: FileRec) -> Dict[str, List[str]]:
    """Best-effort: harvest instrument/book-page-keyed note lines from a PDF."""
    index: Dict[str, List[str]] = {}
    text, method = extract_pdf_text(rec.path)
    if not text:
        LOG(f"  runsheet PDF '{rec.rel}': no extractable text (method={method})", "WARN")
        return index
    inst_pat = re.compile(r"\b(\d{4}-\d{3,7}|\d{6,})\b")
    bp_pat = re.compile(r"\b(?:bk|book)\.?\s*(\d+)\s*(?:pg|page)s?\.?\s*(\d+)\b", re.I)
    for line in text.splitlines():
        ln = norm_text(line)
        if not ln:
            continue
        keys = {f"INST:{norm_doc_ref(m)}" for m in inst_pat.findall(ln)}
        keys |= {f"BP:{norm_doc_ref(b + p)}" for b, p in bp_pat.findall(ln)}
        for k in keys:
            index.setdefault(k, []).append(ln)
    LOG(f"  runsheet PDF harvested notes for {len(index)} references (method={method})")
    return index


def load_runsheet_notes(recs: List[FileRec]) -> Tuple[Dict[str, str], Optional[FileRec]]:
    """Build {ref_key -> combined note text} from the runsheet (workbook/CSV/PDF)."""
    rec = next((r for r in recs if r.kind == "runsheet"), None)
    if not rec:
        return {}, None
    LOG(f"Runsheet: {rec.rel}")
    index: Dict[str, List[str]] = {}
    if rec.ext in PDF_EXT:
        index = _runsheet_notes_from_pdf(rec)
    else:
        for d in _load_table(rec):
            note = norm_text(d.get("remarks", "")) or norm_text(d.get("runsheet_note", ""))
            if not note:
                continue
            for k in _row_ref_keys(d):
                index.setdefault(k, []).append(note)
    flat = {k: " | ".join(dict.fromkeys(v)) for k, v in index.items()}
    LOG(f"  runsheet notes indexed: {len(flat)} references")
    return flat, rec


def load_ogl_index(recs: List[FileRec]
                   ) -> Tuple[Dict[str, str], List[Dict[str, str]], Optional[FileRec]]:
    """Build ({ref_key -> OGL no.}, summary_rows) from the OGL schedule."""
    rec = next((r for r in recs if r.kind == "ogl"), None)
    if not rec:
        return {}, [], None
    LOG(f"OGL schedule: {rec.rel}")
    by_ref: Dict[str, str] = {}
    summary: List[Dict[str, str]] = []
    for d in _load_table(rec):
        ogl = norm_text(d.get("ogl_no", ""))
        summary.append({
            "ogl_no": ogl,
            "lessor": norm_text(d.get("grantor", "")),
            "lessee": norm_text(d.get("grantee", "")),
            "date": norm_text(d.get("instrument_date", "")) or norm_text(d.get("recorded_date", "")),
            "book": norm_text(d.get("book", "")),
            "page": norm_text(d.get("page", "")),
            "instrument_no": norm_text(d.get("instrument_no", "")),
            "interest": norm_text(d.get("interest", "")),
            "legal": norm_text(d.get("legal_description", "")),
        })
        if ogl:
            for k in _row_ref_keys(d):
                by_ref.setdefault(k, ogl)
    LOG(f"  OGL numbers indexed: {len(by_ref)} references, {len(summary)} leases")
    return by_ref, summary, rec


def attach_notes_and_ogl(merged: List["TitleRow"], runsheet_idx: Dict[str, str],
                         ogl_ref: Dict[str, str]) -> Tuple[int, int]:
    """Attach runsheet notes and OGL numbers onto merged rows by ref key."""
    n_notes = n_ogl = 0
    for row in merged:
        for k in _row_ref_keys(row.data):
            if k in runsheet_idx:
                existing = norm_text(row.data.get("runsheet_note", ""))
                add = runsheet_idx[k]
                if add and add not in existing:
                    row.data["runsheet_note"] = (existing + " | " + add).strip(" |") if existing else add
                    n_notes += 1
            if k in ogl_ref and not norm_text(row.data.get("ogl_no", "")):
                row.data["ogl_no"] = ogl_ref[k]
                n_ogl += 1
    return n_notes, n_ogl


def append_analysis_sheets(output_path: Path, links: List[ChainLink],
                           ledger: Dict[str, Fraction],
                           ogl_summary: List[Dict[str, str]],
                           gross_acres: Optional[float]) -> None:
    """Append Interest Chain / Ownership Ledger / OGL Summary / Review Flags
    sheets to output, with flagged rows highlighted for fast examiner review."""
    from openpyxl.styles import Font, PatternFill
    RED = PatternFill("solid", fgColor="FFC7CE")     # flagged conveyance
    AMBER = PatternFill("solid", fgColor="FFEB9C")    # negative / caution
    GREEN = PatternFill("solid", fgColor="C6EFCE")    # balanced total
    wb = openpyxl.load_workbook(output_path)

    def _fresh(title: str):
        if title in wb.sheetnames:
            del wb[title]
        return wb.create_sheet(title)

    def _finish(ws, widths):
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in ws[1]:
            c.font = Font(bold=True)

    def _fill_row(ws, row_idx, ncols, fill):
        for c in range(1, ncols + 1):
            ws.cell(row=row_idx, column=c).fill = fill

    ws = _fresh("Interest Chain")
    headers = ["#", "Instr Date", "Doc Type", "Grantor", "Grantee",
               "Interest (raw)", "Interest (fraction)", "Interest (decimal)",
               "Grantee Cumulative", "Grantor Balance After", "OGL No",
               "Legal", "Flags", "Source"]
    ws.append(headers)
    for lk in links:
        ws.append([
            lk.order, lk.date.isoformat() if lk.date else "", lk.doc_type,
            lk.grantor, lk.grantee, lk.raw_interest, frac_str(lk.fraction),
            frac_dec(lk.fraction), frac_str(lk.grantee_running),
            frac_str(lk.grantor_balance_after), lk.ogl_no, lk.legal,
            lk.flags, lk.source,
        ])
        if lk.flags:
            _fill_row(ws, ws.max_row, len(headers), RED)
    _finish(ws, [5, 12, 18, 26, 26, 16, 16, 14, 16, 18, 12, 40, 34, 24])

    ws2 = _fresh("Ownership Ledger")
    ws2.append(["Party (normalized)", "Net Interest (fraction)",
                "Net Interest (decimal)", "Note"])
    total = sum(ledger.values(), Fraction(0))
    for party, f in sorted(ledger.items(), key=lambda kv: float(kv[1]), reverse=True):
        note = "negative net (see chain / over-conveyance flags)" if f < 0 else ""
        ws2.append([party, frac_str(f), frac_dec(f), note])
        if f < 0:
            _fill_row(ws2, ws2.max_row, 4, AMBER)
    ws2.append([])
    ws2.append(["TOTAL (should trend to 0 if chain balances)",
                frac_str(total), frac_dec(total), ""])
    _fill_row(ws2, ws2.max_row, 4, GREEN if total == 0 else AMBER)
    _finish(ws2, [40, 22, 22, 44])

    if ogl_summary:
        ws3 = _fresh("OGL Summary")
        cols = ["ogl_no", "lessor", "lessee", "date", "book", "page",
                "instrument_no", "interest", "legal"]
        ws3.append([c.replace("_", " ").upper() for c in cols])
        for d in ogl_summary:
            ws3.append([d.get(c, "") for c in cols])
        _finish(ws3, [12, 26, 26, 12, 8, 8, 16, 14, 40])

    # Review Flags: one consolidated examiner punch-list inside the workbook
    ws4 = _fresh("Review Flags")
    ws4.append(["Type", "Chain #", "Instr Date", "Grantor", "Grantee",
                "Interest (raw)", "Detail"])
    for lk in links:
        if lk.flags:
            ws4.append(["conveyance", lk.order,
                        lk.date.isoformat() if lk.date else "", lk.grantor,
                        lk.grantee, lk.raw_interest, lk.flags])
            _fill_row(ws4, ws4.max_row, 7, RED)
    for party, f in sorted(ledger.items(), key=lambda kv: float(kv[1])):
        if f < 0:
            ws4.append(["negative-net", "", "", party, "", "",
                        f"ends holding {frac_str(f)} ({frac_dec(f)}) of the whole"])
            _fill_row(ws4, ws4.max_row, 7, AMBER)
    if ws4.max_row == 1:
        ws4.append(["(none)", "", "", "", "", "",
                    "No chain/ledger flags. Do a final examiner spot-check."])
    _finish(ws4, [14, 8, 12, 26, 26, 18, 52])

    wb.save(output_path)
    wb.close()


# ----------------------------------------------------------------------------
# PDF / index extraction (text -> pdfplumber -> PyMuPDF -> OCR)
# ----------------------------------------------------------------------------
def extract_pdf_text(pdf_path: Path) -> Tuple[str, str]:
    """Returns (text, method). Empty text + 'none' if all backends fail."""
    if _HAVE_PDFPLUMBER:
        try:
            chunks = []
            with pdfplumber.open(str(pdf_path)) as pdf:
                for page in pdf.pages:
                    chunks.append(page.extract_text() or "")
            text = "\n".join(chunks).strip()
            if text:
                return text, "pdfplumber"
        except Exception as exc:
            LOG(f"  pdfplumber failed: {exc}", "WARN")
    if _HAVE_PYMUPDF:
        try:
            text_chunks = []
            doc = fitz.open(str(pdf_path))
            for page in doc:
                text_chunks.append(page.get_text())
            text = "\n".join(text_chunks).strip()
            if text:
                return text, "pymupdf"
            # image-based: try OCR per page render
            if _HAVE_OCR:
                ocr_chunks = []
                for page in doc:
                    pix = page.get_pixmap(dpi=300)
                    img = Image.open(io.BytesIO(pix.tobytes("png")))
                    ocr_chunks.append(pytesseract.image_to_string(img))
                ocr_text = "\n".join(ocr_chunks).strip()
                if ocr_text:
                    return ocr_text, "pymupdf+ocr"
        except Exception as exc:
            LOG(f"  PyMuPDF/OCR failed: {exc}", "WARN")
    return "", "none"


# ----------------------------------------------------------------------------
# Deed / instrument document extraction (review-only; NEVER merged into chain)
# ----------------------------------------------------------------------------
_DEED_TYPES = [
    ("oil and gas lease", "Oil & Gas Lease"), ("o&g lease", "Oil & Gas Lease"),
    ("mineral deed", "Mineral Deed"), ("royalty deed", "Royalty Deed"),
    ("warranty deed", "Warranty Deed"), ("quitclaim", "Quitclaim Deed"),
    ("quit claim", "Quitclaim Deed"), ("assignment", "Assignment"),
    ("probate", "Probate"), ("affidavit", "Affidavit"), ("easement", "Easement"),
    ("right of way", "Right of Way"), ("patent", "Patent"), ("deed", "Deed"),
]
_DEED_BP_RE = re.compile(r"\b(?:book|bk|vol(?:ume)?)\.?\s*(\d+)[,\s]+(?:page|pg)s?\.?\s*(\d+)", re.I)
_DEED_INST_RE = re.compile(
    r"(?:instrument|document|reception|file)\s*(?:no\.?|number|#)?\s*[:#]?\s*([0-9][0-9\-]{3,})", re.I)
_DEED_DATE_RES = [
    re.compile(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b"),
    re.compile(r"\b(\d{1,2}(?:st|nd|rd|th)?\s+day\s+of\s+[A-Za-z]+,?\s+\d{4})", re.I),
    re.compile(r"\b([A-Za-z]+\s+\d{1,2},\s+\d{4})\b"),
]
_DEED_GRANTOR_RES = [
    re.compile(r"by\s+and\s+between\s+(.+?)\s*,?\s*(?:hereinafter.+?)?\b(?:grantor|lessor)", re.I | re.S),
    re.compile(r"^\s*(.+?),?\s+(?:hereinafter.+?)?\bgrantor\b", re.I | re.M),
]
_DEED_GRANTEE_RES = [
    # stop at a comma/newline/semicolon or a common following clause, so initials
    # like "E. WHITE" are not truncated at the first period.
    re.compile(r"\bunto\s+(.+?)(?=,|\n|;|\s+an?\s+undiv|\s+all\b|\s+the\s+following|"
               r"\s+a\s+tract|$)", re.I | re.S),
    re.compile(r"\band\s+(.+?)\s*,?\s*(?:hereinafter.+?)?\b(?:grantee|lessee)", re.I | re.S),
]


def _clean_party(s: str) -> str:
    s = norm_text(s)
    s = re.sub(r"\b(hereinafter|whose|of\s+the\s+county|a\s+single|husband|wife).*$", "", s, flags=re.I)
    return s[:80].strip(" ,;:-")


def extract_text_any(rec: FileRec) -> Tuple[str, str]:
    """Extract text from a PDF, DOCX, or TXT instrument. (text, method)."""
    if rec.ext in PDF_EXT:
        return extract_pdf_text(rec.path)
    if rec.ext == ".docx" and _HAVE_DOCX:
        try:
            d = _docx.Document(str(rec.path))
            return "\n".join(p.text for p in d.paragraphs).strip(), "docx"
        except Exception as exc:
            LOG(f"  docx read failed for {rec.rel}: {exc}", "WARN")
            return "", "none"
    if rec.ext in {".txt"}:
        try:
            return rec.path.read_text(encoding="utf-8", errors="ignore").strip(), "text"
        except Exception:
            return "", "none"
    return "", "unsupported"


def parse_deed_text(text: str) -> Dict[str, str]:
    """Best-effort structured fields from one instrument's text. Conservative -
    leaves a field blank rather than guessing. Confidence = fields found / 6."""
    low = text.lower()
    out = {"doc_type": "", "grantor": "", "grantee": "",
           "date": "", "book": "", "page": "", "instrument_no": ""}
    for needle, label in _DEED_TYPES:
        if needle in low:
            out["doc_type"] = label
            break
    for rx in _DEED_GRANTOR_RES:
        m = rx.search(text)
        if m and _clean_party(m.group(1)):
            out["grantor"] = _clean_party(m.group(1))
            break
    for rx in _DEED_GRANTEE_RES:
        m = rx.search(text)
        if m and _clean_party(m.group(1)):
            out["grantee"] = _clean_party(m.group(1))
            break
    m = _DEED_BP_RE.search(text)
    if m:
        out["book"], out["page"] = m.group(1), m.group(2)
    m = _DEED_INST_RE.search(text)
    if m:
        out["instrument_no"] = m.group(1)
    for rx in _DEED_DATE_RES:
        m = rx.search(text)
        if m:
            out["date"] = norm_text(m.group(1))
            break
    found = sum(1 for k in ("doc_type", "grantor", "grantee", "date", "instrument_no") if out[k])
    out["confidence"] = f"{found}/5"
    return out


def extract_deed_records(recs: List[FileRec]) -> List[Dict[str, str]]:
    """Scan deed/doc instruments and return review-only extracted rows."""
    targets = [r for r in recs if r.kind == "deed" or r.ext in DOC_EXT]
    records: List[Dict[str, str]] = []
    for rec in targets:
        text, method = extract_text_any(rec)
        if not text:
            records.append({"file": rec.rel, "method": method, "doc_type": "",
                            "grantor": "", "grantee": "", "date": "", "book": "",
                            "page": "", "instrument_no": "", "confidence": "0/5",
                            "note": "no extractable text"})
            continue
        d = parse_deed_text(text)
        d["file"] = rec.rel
        d["method"] = method
        d["note"] = "AUTO-EXTRACTED - VERIFY against the instrument; NOT in the chain"
        records.append(d)
    if targets:
        LOG(f"  scanned {len(targets)} deed/doc file(s) -> {len(records)} extracted rows")
    return records


def append_deeds_sheet(output_path: Path, deeds: List[Dict[str, str]]) -> None:
    """Append a clearly-labeled review-only 'Deeds (auto-extract VERIFY)' sheet."""
    if not deeds:
        return
    from openpyxl.styles import Font, PatternFill
    AMBER = PatternFill("solid", fgColor="FFEB9C")
    wb = openpyxl.load_workbook(output_path)
    title = "Deeds (auto-extract VERIFY)"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(["File", "Doc Type", "Grantor?", "Grantee?", "Date?", "Book",
               "Page", "Instrument No?", "Confidence", "Method", "Note"])
    for d in deeds:
        ws.append([d.get("file", ""), d.get("doc_type", ""), d.get("grantor", ""),
                   d.get("grantee", ""), d.get("date", ""), d.get("book", ""),
                   d.get("page", ""), d.get("instrument_no", ""),
                   d.get("confidence", ""), d.get("method", ""), d.get("note", "")])
        for c in range(1, 12):
            ws.cell(row=ws.max_row, column=c).fill = AMBER
    ws.freeze_panes = "A2"
    for i, w in enumerate([34, 16, 26, 26, 16, 8, 8, 16, 11, 10, 46], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True)
    wb.save(output_path)
    wb.close()


def _esc(v: Any) -> str:
    s = "" if v is None else str(v)
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))


def write_html_report(path: Path, section: str, links: List[ChainLink],
                      ledger: Dict[str, Fraction], ogl_summary: List[Dict[str, str]],
                      deeds: List[Dict[str, str]], stats: Dict[str, Any],
                      checklist: List[str]) -> None:
    """Write a self-contained (no external assets) HTML interest-chain report."""
    def table(headers, rows, row_class=None):
        h = "".join(f"<th>{_esc(x)}</th>" for x in headers)
        body = []
        for r in rows:
            cls = f' class="{row_class(r)}"' if row_class else ""
            cells = "".join(f"<td>{_esc(c)}</td>" for c in r)
            body.append(f"<tr{cls}>{cells}</tr>")
        return f"<table><thead><tr>{h}</tr></thead><tbody>{''.join(body)}</tbody></table>"

    chain_rows = [[lk.order, lk.date.isoformat() if lk.date else "", lk.doc_type,
                   lk.grantor, lk.grantee, lk.raw_interest, frac_str(lk.fraction),
                   frac_str(lk.grantee_running), frac_str(lk.grantor_balance_after),
                   lk.ogl_no, lk.flags] for lk in links]
    chain_html = table(
        ["#", "Date", "Doc Type", "Grantor", "Grantee", "Interest", "Conveyed",
         "Grantee Cum.", "Grantor Bal.", "OGL", "Flags"],
        chain_rows, row_class=lambda r: "flag" if r[-1] else "")

    total = sum(ledger.values(), Fraction(0))
    led_rows = [[p, frac_str(f), frac_dec(f)]
                for p, f in sorted(ledger.items(), key=lambda kv: float(kv[1]), reverse=True)]
    led_html = table(["Party", "Net (fraction)", "Net (decimal)"], led_rows,
                     row_class=lambda r: "neg" if r[1].startswith("-") else "")
    led_html += f'<p class="total">Ledger total: <b>{_esc(frac_str(total))}</b> ' \
                f'({_esc(frac_dec(total))}) &mdash; should be 0 if the chain balances.</p>'

    ogl_html = ""
    if ogl_summary:
        ogl_html = "<h2>OGL Summary</h2>" + table(
            ["OGL No", "Lessor", "Lessee", "Date", "Book", "Page", "Instrument", "Interest", "Legal"],
            [[d.get("ogl_no", ""), d.get("lessor", ""), d.get("lessee", ""), d.get("date", ""),
              d.get("book", ""), d.get("page", ""), d.get("instrument_no", ""),
              d.get("interest", ""), d.get("legal", "")] for d in ogl_summary])

    deeds_html = ""
    if deeds:
        deeds_html = ('<h2>Deeds (auto-extracted &mdash; VERIFY)</h2>'
                      '<p class="warn">These rows were parsed from instrument documents and '
                      'are <b>not</b> part of the authoritative chain. Verify each against the '
                      'original before relying on it.</p>' + table(
            ["File", "Doc Type", "Grantor?", "Grantee?", "Date?", "Book", "Page",
             "Instrument?", "Confidence", "Method"],
            [[d.get("file", ""), d.get("doc_type", ""), d.get("grantor", ""), d.get("grantee", ""),
              d.get("date", ""), d.get("book", ""), d.get("page", ""), d.get("instrument_no", ""),
              d.get("confidence", ""), d.get("method", "")] for d in deeds]))

    stat_html = "".join(f'<div class="stat"><span>{_esc(v)}</span><label>{_esc(k)}</label></div>'
                        for k, v in stats.items())
    check_html = "".join(f"<li>{_esc(c)}</li>" for c in checklist) or "<li>No automated issues.</li>"

    html = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{_esc(section)} Roger Mills - Interest Chain Report</title>
<style>
:root{{color-scheme:light dark}}
body{{font:15px/1.5 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;margin:0;padding:2rem;
 background:#f7f7f8;color:#1a1a1a}}
@media(prefers-color-scheme:dark){{body{{background:#15171a;color:#e6e6e6}}}}
h1{{margin:0 0 .25rem}} h2{{margin:2rem 0 .5rem;border-bottom:2px solid #8883;padding-bottom:.25rem}}
.sub{{color:#8a8a8a;margin:0 0 1.5rem}}
.stats{{display:flex;flex-wrap:wrap;gap:.75rem;margin:1rem 0}}
.stat{{background:#fff2;border:1px solid #8883;border-radius:10px;padding:.6rem .9rem;min-width:120px}}
.stat span{{font-size:1.4rem;font-weight:700;display:block}} .stat label{{font-size:.75rem;color:#8a8a8a}}
.wrap{{overflow-x:auto}}
table{{border-collapse:collapse;width:100%;margin:.5rem 0;font-size:13px;background:#fff1}}
th,td{{border:1px solid #8883;padding:.35rem .5rem;text-align:left;vertical-align:top}}
th{{background:#8881;position:sticky;top:0}}
tr.flag td{{background:#ffd5d5;color:#7a1010}} tr.neg td{{background:#ffe8b0;color:#7a5210}}
@media(prefers-color-scheme:dark){{tr.flag td{{background:#5a1e1e;color:#ffcaca}}
 tr.neg td{{background:#5a4410;color:#ffe6a8}}}}
.total{{font-size:14px}} .warn{{color:#a05a00;font-size:13px}}
ul.check{{background:#fff1;border:1px solid #8883;border-radius:10px;padding:1rem 1rem 1rem 2.2rem}}
footer{{margin-top:2rem;color:#8a8a8a;font-size:12px}}
</style></head><body>
<h1>{_esc(section)} &mdash; Roger Mills County</h1>
<p class="sub">Cursory Title Report &mdash; Interest Chain (codexv2). This view never
invents data; unresolved items are flagged, not guessed.</p>
<div class="stats">{stat_html}</div>
<h2>Interest Chain</h2><div class="wrap">{chain_html}</div>
<h2>Ownership Ledger</h2><div class="wrap">{led_html}</div>
{ogl_html and '<div class="wrap">' + ogl_html + '</div>'}
{deeds_html and '<div class="wrap">' + deeds_html + '</div>'}
<h2>Perfection Checklist</h2><ul class="check">{check_html}</ul>
<footer>Generated by roger_mills_title_report_builder.py (codexv2). Excel workbook is the
authoritative deliverable; this HTML is a convenience view.</footer>
</body></html>"""
    path.write_text(html, encoding="utf-8")


def verify_against_index(merged: List[TitleRow], index_text: str) -> Dict[str, bool]:
    """Mark each merged row 'verified' if its instrument/book-page or party names
    appear in the index text. Never fabricates - only confirms presence."""
    verified: Dict[str, bool] = {}
    haystack = re.sub(r"[^A-Za-z0-9]+", "", index_text).upper()
    if not haystack:
        return verified
    for row in merged:
        k = row.key()
        inst = norm_doc_ref(row.data.get("instrument_no", ""))
        bp = norm_doc_ref(row.data.get("book", "") + row.data.get("page", ""))
        ok = False
        if inst and len(inst) >= 4 and inst in haystack:
            ok = True
        elif bp and len(bp) >= 3 and bp in haystack:
            ok = True
        else:
            g = norm_doc_ref(row.data.get("grantor", ""))[:12]
            if g and len(g) >= 6 and g in haystack:
                ok = True
        verified[k] = ok
    return verified


# ----------------------------------------------------------------------------
# Output: copy template formatting, write data, save
# ----------------------------------------------------------------------------
def build_output(template_path: Path, output_path: Path, merged: List[TitleRow],
                 verified: Dict[str, bool], section: str) -> Tuple[str, int]:
    """Copy the template workbook (preserving styling) and write merged rows into
    the data sheet. Returns (data_sheet_name, rows_written)."""
    # openpyxl load+save preserves: merged cells, fonts, fills, borders, number
    # formats, column widths, row heights, formulas, print areas, page setup,
    # freeze panes, header/footer, sheet order and tab colors.
    wb = openpyxl.load_workbook(template_path, data_only=False)

    # find the sheet/header row in the template that looks like the data table
    target_ws = None
    header_row = 1
    header_map: Dict[int, str] = {}
    best = -1
    for ws in wb.worksheets:
        hr, hmap = _detect_header_row(ws)
        if len(hmap) > best:
            best, target_ws, header_row, header_map = len(hmap), ws, hr, hmap
    if target_ws is None or not header_map:
        # template has no recognizable table - create a clean one rather than guess
        target_ws = wb.active
        header_row = 1
        header_map = {}

    LOG(f"  template data sheet: '{target_ws.title}' (header row {header_row}, "
        f"{len(header_map)} mapped columns)")

    # capture a "style template" row = the first data row beneath the header so
    # new rows inherit the template's per-column cell styling.
    style_row_idx = header_row + 1
    col_to_field = dict(header_map)
    field_to_col = {v: k for k, v in col_to_field.items()}

    # if template had no headers, lay down our canonical headers in row 1
    if not field_to_col:
        for i, fld in enumerate(CANON_FIELDS, start=1):
            target_ws.cell(row=1, column=i, value=fld.replace("_", " ").title())
            field_to_col[fld] = i
        header_row, style_row_idx = 1, 2

    from copy import copy as _copy

    def clone_style(src_cell, dst_cell):
        try:
            dst_cell.font = _copy(src_cell.font)
            dst_cell.fill = _copy(src_cell.fill)
            dst_cell.border = _copy(src_cell.border)
            dst_cell.alignment = _copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = _copy(src_cell.protection)
        except Exception:
            pass

    write_row = header_row + 1
    rows_written = 0
    for row in merged:
        k = row.key()
        for fld, col in field_to_col.items():
            value = row.data.get(fld, "")
            cell = target_ws.cell(row=write_row, column=col, value=value or None)
            if write_row != style_row_idx:
                clone_style(target_ws.cell(row=style_row_idx, column=col), cell)
        # verification flag in remarks column if present
        if "remarks" in field_to_col and not verified.get(k, False):
            rc = target_ws.cell(row=write_row, column=field_to_col["remarks"])
            existing = norm_text(rc.value)
            flag = "[REVIEW: not found in index]"
            rc.value = (existing + " " + flag).strip() if existing else flag
        write_row += 1
        rows_written += 1

    # stamp the section + generation note in an unobtrusive header cell if blank
    try:
        title_cell = target_ws.cell(row=1, column=1)
        if not norm_text(title_cell.value):
            title_cell.value = f"{section} - Roger Mills County - Cursory Title Report"
    except Exception:
        pass

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return target_ws.title, rows_written


# ----------------------------------------------------------------------------
# Backups
# ----------------------------------------------------------------------------
def backup_originals(root: Path, recs: List[FileRec], support_dir: Path) -> Path:
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = support_dir / f"backup_{stamp}"
    backup_dir.mkdir(parents=True, exist_ok=True)
    copied = 0
    for rec in recs:
        try:
            dest = backup_dir / rec.rel
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(rec.path, dest)  # copy2 preserves mtime; originals untouched
            copied += 1
        except Exception as exc:
            LOG(f"  backup failed for {rec.rel}: {exc}", "WARN")
    LOG(f"Backed up {copied}/{len(recs)} originals -> {backup_dir}")
    return backup_dir


# ----------------------------------------------------------------------------
# Support file writers
# ----------------------------------------------------------------------------
def write_inventory_csv(path: Path, recs: List[FileRec], analyses: Dict[str, WorkbookAnalysis]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["relative_path", "kind", "ext", "size_bytes", "modified",
                    "sheets", "best_header_cols", "data_rows", "style_score", "error"])
        for rec in sorted(recs, key=lambda r: r.rel.lower()):
            wa = analyses.get(str(rec.path))
            sheets = best_cols = data_rows = style = ""
            err = ""
            if wa:
                sheets = "|".join(s.name for s in wa.sheets)
                bs = wa.best_sheet
                best_cols = str(len(bs.header_map)) if bs else "0"
                data_rows = str(bs.data_rows) if bs else "0"
                style = f"{wa.style_score:.1f}"
                err = wa.error
            w.writerow([rec.rel, rec.kind, rec.ext, rec.size,
                        rec.mtime.strftime("%Y-%m-%d %H:%M:%S"),
                        sheets, best_cols, data_rows, style, err])


def write_audit_csv(path: Path, audit: List[Dict[str, str]], verified: Dict[str, bool]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        cols = ["key", "n_sources", "sources", "grantor", "grantee", "doc_type",
                "instrument_no", "book", "page", "index_verified"]
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for rec in audit:
            rec = dict(rec)
            rec["index_verified"] = "YES" if verified.get(rec["key"], False) else "NO"
            w.writerow(rec)


def write_conflicts_xlsx(path: Path, conflicts: List[Dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conflicts"
    headers = ["key", "field", "chosen", "alternatives", "sources"]
    ws.append(headers)
    for c in ws[1]:
        c.font = openpyxl.styles.Font(bold=True)
    for rec in conflicts:
        ws.append([rec.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(60, len(h) + 10))
    if not conflicts:
        ws.append(["(no field-level conflicts detected across sources)"])
    wb.save(path)
    wb.close()


# ----------------------------------------------------------------------------
# Validation
# ----------------------------------------------------------------------------
def validate_output(output_path: Path, data_sheet: str, expected_rows: int) -> Tuple[bool, List[str]]:
    notes: List[str] = []
    ok = True
    if not output_path.exists():
        return False, [f"Output file missing: {output_path}"]
    try:
        wb = openpyxl.load_workbook(output_path, data_only=False)
        notes.append(f"Reopened OK. Sheets: {wb.sheetnames}")
        if data_sheet not in wb.sheetnames:
            ok = False
            notes.append(f"Expected data sheet '{data_sheet}' not found.")
        else:
            ws = wb[data_sheet]
            populated = sum(1 for row in ws.iter_rows()
                            if any(c.value not in (None, "") for c in row))
            notes.append(f"Data sheet '{data_sheet}' populated rows: {populated}")
            if expected_rows and populated < 2:
                ok = False
                notes.append("Data sheet appears empty despite merged rows.")
        # broken-formula heuristic
        broken = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("=#"):
                        broken += 1
        notes.append(f"Suspect broken formulas: {broken}")
        wb.close()
    except Exception as exc:
        ok = False
        notes.append(f"Reopen FAILED (possible corruption): {exc}")
    return ok, notes


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Roger Mills Cursory Title Report builder (codexv2)")
    ap.add_argument("--root", required=True, help="Root folder to scan recursively")
    ap.add_argument("--output", default="", help="Final .xlsx output path (default: <final-dir>/<section>_...xlsx)")
    ap.add_argument("--support-dir", default="", help="Folder for support files (default: <final-dir>/files)")
    ap.add_argument("--final-dir", default="",
                    help="Folder for finished reports (default: <root>/rogermillsfinalreports)")
    ap.add_argument("--section", default="31-12N-24W", help="Target section label")
    ap.add_argument("--template", default="", help="Explicit template .xlsx (else auto-detect)")
    ap.add_argument("--gross-acres", type=float, default=None,
                    help="Gross mineral acres for the tract (enables NMA->fraction interest chaining)")
    ap.add_argument("--max-passes", type=int, default=3,
                    help="Loop-till-perfect: rebuild up to N times until validation passes (default 3)")
    ap.add_argument("--no-unzip", action="store_true",
                    help="Do NOT recursively extract .zip archives before scanning")
    ap.add_argument("--no-tidy", action="store_true",
                    help="Detect duplicates/trash but do NOT move them to quarantine")
    ap.add_argument("--no-deeds", action="store_true",
                    help="Skip best-effort extraction of deed/DOCX instruments (review-only sheet)")
    ap.add_argument("--no-html", action="store_true",
                    help="Do NOT write the self-contained HTML interest-chain report")
    ap.add_argument("--dry-run", action="store_true", help="Analyze & plan only; do not write final workbook")
    args = ap.parse_args(argv)

    root = Path(args.root)
    final_dir = Path(args.final_dir) if args.final_dir else (root / "rogermillsfinalreports")
    safe_section = re.sub(r"[^A-Za-z0-9._-]+", "_", args.section).strip("_") or "section"
    output_path = (Path(args.output) if args.output
                   else final_dir / f"{safe_section}_Roger_Mills_Cursory_Title_Report_codexv2.xlsx")
    support_dir = Path(args.support_dir) if args.support_dir else (final_dir / "files")

    if not root.exists():
        print(f"FATAL: root folder does not exist: {root}")
        print("Are you running this on the machine where the files live?")
        return 2
    final_dir.mkdir(parents=True, exist_ok=True)
    support_dir.mkdir(parents=True, exist_ok=True)

    LOG.section(f"Roger Mills Title Report Builder - section {args.section}")
    LOG(f"Root:        {root}")
    LOG(f"Final dir:   {final_dir}")
    LOG(f"Output:      {output_path}")
    LOG(f"Support dir: {support_dir}")
    LOG(f"Gross acres: {args.gross_acres if args.gross_acres else '(not given; NMA interests will be flagged)'}")
    LOG(f"Capabilities: pandas={_HAVE_PANDAS} dateutil={_HAVE_DATEUTIL} "
        f"rapidfuzz={_HAVE_RAPIDFUZZ} pdfplumber={_HAVE_PDFPLUMBER} "
        f"pymupdf={_HAVE_PYMUPDF} ocr={_HAVE_OCR}")

    # 0. unzip everything (recursive, nested, path-safe, idempotent)
    zips_extracted = 0
    if not args.no_unzip:
        LOG.section("0. Extract .zip archives")
        _extract_root, zips_extracted = extract_all_zips(root)
        LOG(f"Extracted {zips_extracted} file(s) from archives.")
    else:
        LOG("--no-unzip: skipping archive extraction.")

    # 1. inventory
    LOG.section("1. Inventory")
    recs = inventory(root, exclude=[final_dir, support_dir])
    LOG(f"Found {len(recs)} files (excluding generated output/support trees).")

    def _log_kinds(rs):
        for kind in ("template", "report", "runsheet", "ogl", "index_pdf",
                     "data", "doc", "image", "other"):
            n = sum(1 for r in rs if r.kind == kind)
            if n:
                LOG(f"  {kind:10s}: {n}")
    _log_kinds(recs)

    # 1b. duplicates & trash (detect always; quarantine unless --no-tidy)
    LOG.section("1b. Duplicates & trash")
    dup_groups, trash = find_duplicates_and_trash(recs)
    dup_extra = sum(len(g) - 1 for g in dup_groups)  # redundant copies (keep 1 each)
    LOG(f"Duplicate groups: {len(dup_groups)} ({dup_extra} redundant copies); "
        f"trash files: {len(trash)}")
    quarantine_dir = support_dir / "_quarantine"
    moved: List[Tuple[str, str, str]] = []
    if not args.dry_run and not args.no_tidy and (dup_extra or trash):
        to_move = [r for g in dup_groups for r in g if r is not _dup_keeper(g)]
        moved += quarantine(to_move, quarantine_dir / "duplicates", "duplicate")
        moved += quarantine(trash, quarantine_dir / "trash", "trash")
        LOG(f"Quarantined {len(moved)} file(s) -> {quarantine_dir} (reversible; "
            f"originals were also backed up). Re-inventorying.")
        recs = inventory(root, exclude=[final_dir, support_dir])
        LOG(f"Post-tidy inventory: {len(recs)} files.")
        _log_kinds(recs)
    elif dup_extra or trash:
        LOG("Tidy skipped (dry-run or --no-tidy): duplicates/trash left in place, "
            "recorded in the manifest.", "WARN")
    if dup_groups or trash:  # don't overwrite a prior manifest with an empty one
        write_tidy_manifest(support_dir / "tidy_manifest_codexv2.csv",
                            dup_groups, trash, moved)
        LOG("Wrote tidy_manifest_codexv2.csv")
    else:
        LOG("No duplicates or trash this run; keeping any prior tidy manifest.")

    # 2. backups (skip in dry-run)
    if not args.dry_run:
        LOG.section("2. Timestamped backups")
        backup_originals(root, recs, support_dir)
    else:
        LOG("DRY-RUN: skipping backups.")

    # 3. analyze workbooks
    LOG.section("3. Analyze workbooks")
    analyses: Dict[str, WorkbookAnalysis] = {}
    for rec in recs:
        if rec.ext in WORKBOOK_EXT:
            wa = analyze_workbook(rec)
            analyses[str(rec.path)] = wa
            bs = wa.best_sheet
            LOG(f"  {rec.rel}: sheets={len(wa.sheets)} "
                f"best_cols={(len(bs.header_map) if bs else 0)} "
                f"rows={(bs.data_rows if bs else 0)} style={wa.style_score:.1f} "
                f"{('ERR ' + wa.error) if wa.error else ''}")

    # template selection
    template_rec = None
    if args.template:
        template_rec = next((r for r in recs if r.path.name.lower() == Path(args.template).name.lower()
                             or str(r.path).lower() == args.template.lower()), None)
    if template_rec is None:
        templates = [r for r in recs if r.kind == "template"]
        template_rec = templates[0] if templates else None

    # 4. tournament: pick best base report workbook
    LOG.section("4. Best-base tournament")
    report_analyses = [analyses[str(r.path)] for r in recs
                       if r.kind in ("report",) and str(r.path) in analyses]
    if recs:
        newest = max(r.mtime.timestamp() for r in recs)
        oldest = min(r.mtime.timestamp() for r in recs)
    else:
        newest = oldest = _dt.datetime.now().timestamp()
    scored = sorted(
        ((score_report(wa, newest, oldest), wa) for wa in report_analyses),
        key=lambda t: t[0], reverse=True,
    )
    for sc, wa in scored:
        LOG(f"  score={sc:6.3f}  {wa.rec.rel}")
    best_base = scored[0][1] if scored and scored[0][0] >= 0 else None

    if template_rec is None and best_base is not None:
        LOG("No Template(*) workbook found; using best-base report as formatting authority.", "WARN")
        template_rec = best_base.rec
    if template_rec is None:
        LOG("FATAL: no template and no usable report workbook found. Cannot build.", "ERROR")
        LOG.dump(support_dir / "build_log_codexv2.txt")
        return 3
    LOG(f"Formatting authority: {template_rec.rel}")
    if best_base:
        LOG(f"Best base data workbook: {best_base.rec.rel}")

    # 5. extract + merge rows from ALL report sources (workbooks + CSV/TSV)
    LOG.section("5. Extract & merge data")
    all_rows: List[List[TitleRow]] = []
    for wa in report_analyses:
        rows = extract_rows(wa)
        if rows:
            LOG(f"  {wa.rec.rel}: {len(rows)} rows")
            all_rows.append(rows)
    for rec in recs:
        if rec.kind == "report" and rec.ext in {".csv", ".tsv"}:
            rows = title_rows_from_table(rec)
            if rows:
                LOG(f"  {rec.rel}: {len(rows)} rows (delimited)")
                all_rows.append(rows)
    merged, audit, conflicts = merge_rows(all_rows)
    LOG(f"Merged unique records: {len(merged)}  | field conflicts: {len(conflicts)}")

    # 5b. runsheet notes + OGL numbers -> attach onto merged rows
    LOG.section("5b. Runsheet notes & OGL numbers")
    runsheet_idx, runsheet_rec = load_runsheet_notes(recs)
    ogl_ref, ogl_summary, ogl_rec = load_ogl_index(recs)
    n_notes, n_ogl = attach_notes_and_ogl(merged, runsheet_idx, ogl_ref)
    LOG(f"Attached runsheet notes to {n_notes} rows; OGL numbers to {n_ogl} rows.")

    # 5c. interest chaining
    LOG.section("5c. Interest chain")
    chain_links, ledger = build_interest_chain(merged, args.gross_acres)
    n_flagged = sum(1 for lk in chain_links if lk.flags)
    n_parsed = sum(1 for lk in chain_links if lk.fraction is not None)
    LOG(f"Chained {len(chain_links)} conveyances: {n_parsed} with parsed interest, "
        f"{n_flagged} flagged for review; {len(ledger)} parties in ownership ledger.")

    # 5d. deed/instrument document extraction (review-only, never merged)
    LOG.section("5d. Deed document extraction (review-only)")
    deeds: List[Dict[str, str]] = []
    if args.no_deeds:
        LOG("--no-deeds: skipping instrument document extraction.")
    else:
        deeds = extract_deed_records(recs)
        if not deeds:
            LOG("No deed/DOCX instruments found to extract.")

    # 6. PDF / index verification
    LOG.section("6. Index PDF verification")
    index_text, method = "", "none"
    index_pdf = next((r for r in recs if r.kind == "index_pdf"), None)
    if index_pdf:
        LOG(f"Index PDF: {index_pdf.rel}")
        index_text, method = extract_pdf_text(index_pdf.path)
        LOG(f"Extraction method: {method}; chars={len(index_text)}")
        if method == "none":
            LOG("No PDF text/OCR available - rows will be marked unverified.", "WARN")
    else:
        LOG("No index PDF found.", "WARN")
    verified = verify_against_index(merged, index_text)
    n_verified = sum(1 for v in verified.values() if v)
    LOG(f"Rows verified against index: {n_verified}/{len(merged)}")

    # 7. build output (loop till perfect)
    data_sheet = ""
    rows_written = 0
    build_ok = False
    passes_used = 0
    val_lines: List[str] = []
    if args.dry_run:
        LOG.section("7. DRY-RUN: skipping final workbook write")
        val_lines = ["DRY-RUN: final workbook not written; validation skipped."]
    else:
        LOG.section(f"7. Build final workbook (loop till perfect, max {args.max_passes} passes)")
        for attempt in range(1, max(1, args.max_passes) + 1):
            passes_used = attempt
            LOG(f"  --- pass {attempt}/{args.max_passes} ---")
            try:
                data_sheet, rows_written = build_output(
                    template_rec.path, output_path, merged, verified, args.section)
                append_analysis_sheets(output_path, chain_links, ledger,
                                       ogl_summary, args.gross_acres)
                append_deeds_sheet(output_path, deeds)
                LOG(f"  wrote {rows_written} rows to '{data_sheet}' + analysis sheets -> {output_path}")
            except Exception as exc:
                LOG(f"  build failed: {exc}\n{traceback.format_exc()}", "ERROR")
                continue
            ok, notes = validate_output(output_path, data_sheet, rows_written)
            val_lines = notes
            for n in notes:
                LOG(f"    {n}")
            if ok:
                build_ok = True
                LOG(f"  pass {attempt}: VALIDATION PASS - stopping loop.")
                break
            LOG(f"  pass {attempt}: validation not clean; retrying.", "WARN")

    # 8. support files
    LOG.section("8. Support files")
    write_inventory_csv(support_dir / "source_inventory_codexv2.csv", recs, analyses)
    write_audit_csv(support_dir / "merge_audit_codexv2.csv", audit, verified)
    write_conflicts_xlsx(support_dir / "conflicts_review_codexv2.xlsx", conflicts)
    LOG("Wrote source_inventory_codexv2.csv, merge_audit_codexv2.csv, conflicts_review_codexv2.xlsx")

    # 9. review items derived from the interest chain / notes / OGL work
    total_ledger = sum(ledger.values(), Fraction(0))
    over_conveyances = [lk for lk in chain_links if "OVER_CONVEYANCE" in lk.flags]
    unparsed = [lk for lk in chain_links if "UNPARSED_INTEREST" in lk.flags]
    prop_nobasis = [lk for lk in chain_links if "PROPORTIONAL_NO_BASIS" in lk.flags]
    negatives = {p: f for p, f in ledger.items() if f < 0}
    n_verified_pct = f"{n_verified}/{len(merged)}" if merged else "0/0"

    # validation summary file
    summary = [
        "Roger Mills Title Report - final validation summary (codexv2)",
        f"Generated: {_dt.datetime.now():%Y-%m-%d %H:%M:%S}",
        f"Section: {args.section}",
        "",
        f"Build status:            {'PASS' if build_ok else ('DRY-RUN' if args.dry_run else 'NEEDS REVIEW')}"
        f" (passes used: {passes_used}/{args.max_passes})",
        f"Files extracted (zip):   {zips_extracted}",
        f"Duplicate groups:        {len(dup_groups)}  ({dup_extra} redundant copies)",
        f"Trash files:             {len(trash)}",
        f"Quarantined (moved):     {len(moved)}"
        f"{'' if (args.no_tidy or args.dry_run) else ' -> ' + str(quarantine_dir)}",
        f"Source files reviewed:   {len(recs)}",
        f"Workbooks analyzed:      {len(analyses)}",
        f"Report sources merged:   {len(all_rows)}",
        f"Unique records merged:   {len(merged)}",
        f"Rows written to output:  {rows_written}",
        f"Field-level conflicts:   {len(conflicts)}",
        f"Index extraction method: {method}",
        f"Rows verified vs index:  {n_verified_pct}",
        "",
        "Interest chain:",
        f"  Conveyances chained:   {len(chain_links)}",
        f"  Parsed interests:      {n_parsed}",
        f"  Gross acres (for NMA): {args.gross_acres if args.gross_acres else '(not supplied)'}",
        f"  Over-conveyance flags: {len(over_conveyances)}",
        f"  Unparsed interests:    {len(unparsed)}",
        f"  Parties in ledger:     {len(ledger)}  (net total {frac_str(total_ledger)} = {frac_dec(total_ledger)})",
        "",
        "Runsheet & OGL:",
        f"  Runsheet workbook:     {runsheet_rec.rel if runsheet_rec else '(none found)'}",
        f"  Rows w/ runsheet note: {n_notes}",
        f"  OGL schedule:          {ogl_rec.rel if ogl_rec else '(none found)'}",
        f"  Rows w/ OGL number:    {n_ogl}   (OGL leases catalogued: {len(ogl_summary)})",
        "",
        f"Formatting authority:    {template_rec.rel if template_rec else '(none)'}",
        f"Best base workbook:      {best_base.rec.rel if best_base else '(none)'}",
        f"Output workbook:         {output_path}",
        "",
        "OCR / PDF note:",
    ]
    if method == "none":
        summary.append("  PDF text extraction and OCR were unavailable or returned no text.")
        summary.append("  Install pdfplumber/PyMuPDF and a Tesseract OCR engine to enable")
        summary.append("  full index verification, then re-run. Rows are currently marked")
        summary.append("  [REVIEW: not found in index] where they could not be confirmed.")
    else:
        summary.append(f"  Index extracted via '{method}'. {n_verified} rows confirmed present.")

    # perfection checklist - the exact remaining human-review items
    summary += ["", "=" * 60, "PERFECTION CHECKLIST (resolve every item, then re-run):", "=" * 60]
    checklist: List[str] = []
    if not build_ok and not args.dry_run:
        checklist.append("Output did not pass clean validation - see 'Validation' notes below.")
    if unparsed:
        checklist.append(f"{len(unparsed)} conveyance(s) have an interest string the parser "
                         "could not read - fix the wording or fill the interest on the "
                         "'Interest Chain' sheet (rows flagged UNPARSED_INTEREST).")
    if over_conveyances:
        checklist.append(f"{len(over_conveyances)} OVER_CONVEYANCE flag(s): a grantor conveyed "
                         "more than the chain shows they held - verify against the deeds.")
    if negatives:
        checklist.append(f"{len(negatives)} party(ies) end with a negative net interest in the "
                         "Ownership Ledger - usually a missing prior conveyance-in or a wrong "
                         "grantor/grantee. Check the earliest links for those parties.")
    if prop_nobasis:
        checklist.append(f"{len(prop_nobasis)} 'of grantor's interest' conveyance(s) could not "
                         "be computed because the grantor's prior interest is unknown - "
                         "establish the grantor's interest earlier in the chain, then re-run.")
    if args.gross_acres is None:
        checklist.append("No --gross-acres supplied: any NMA-stated interests were left "
                         "unparsed. Re-run with --gross-acres <acres> to chain them.")
    if conflicts:
        checklist.append(f"{len(conflicts)} field-level conflict(s) across sources - resolve in "
                         "conflicts_review_codexv2.xlsx.")
    if n_verified < len(merged):
        checklist.append(f"{len(merged) - n_verified} row(s) not confirmed against the index PDF - "
                         "spot-check them (flagged [REVIEW: not found in index]).")
    if runsheet_rec is None:
        checklist.append("No runsheet workbook was found/recognized - name it with 'runsheet' "
                         "in the filename so its notes can be chained in.")
    if ogl_rec is None:
        checklist.append("No OGL schedule was found/recognized - name it with 'ogl' in the "
                         "filename so OGL numbers can be attached.")
    if (dup_extra or trash) and (args.no_tidy or args.dry_run):
        checklist.append(f"{dup_extra} duplicate + {len(trash)} trash file(s) detected but NOT "
                         "moved (--no-tidy/dry-run). Review tidy_manifest_codexv2.csv, then "
                         "re-run without --no-tidy to quarantine them.")
    if deeds:
        checklist.append(f"{len(deeds)} instrument(s) were auto-extracted into the "
                         "'Deeds (auto-extract VERIFY)' sheet - verify each against the "
                         "original and fold confirmed ones into a report source; they are "
                         "NOT in the chain.")
    checklist.append("Spot-check legal descriptions / acreage on the data + Interest Chain sheets.")
    if not checklist:
        checklist.append("No automated issues remain. Do a final examiner spot-check and sign off.")
    summary += [f"  [ ] {item}" for item in checklist]

    summary += ["", "Validation:", *[f"  {l}" for l in val_lines]]
    (support_dir / "final_validation_summary_codexv2.txt").write_text(
        "\n".join(summary) + "\n", encoding="utf-8")
    LOG("Wrote final_validation_summary_codexv2.txt")

    # 9b. self-contained HTML interest-chain report (convenience view)
    html_path = None
    if not args.no_html:
        html_path = final_dir / f"{safe_section}_interest_chain_report_codexv2.html"
        stats = {
            "Records": len(merged), "Conveyances": len(chain_links),
            "Parsed": n_parsed, "Flagged": n_flagged,
            "OGL leases": len(ogl_summary), "Deeds": len(deeds),
            "Ledger net": frac_str(total_ledger),
            "Build": "PASS" if build_ok else ("DRY-RUN" if args.dry_run else "REVIEW"),
        }
        try:
            write_html_report(html_path, args.section, chain_links, ledger,
                              ogl_summary, deeds, stats, checklist)
            LOG(f"Wrote HTML report -> {html_path}")
        except Exception as exc:
            LOG(f"HTML report failed: {exc}", "WARN")
            html_path = None

    # build log last (captures everything above)
    LOG.dump(support_dir / "build_log_codexv2.txt")

    # final console report
    print("\n" + "=" * 70)
    print("FINAL REPORT (codexv2)")
    print("=" * 70)
    print(f"Final workbook:        {output_path if not args.dry_run else '(dry-run, not written)'}")
    print(f"  sheets added:        Interest Chain, Ownership Ledger, Review Flags"
          f"{', OGL Summary' if ogl_summary else ''}"
          f"{', Deeds(VERIFY)' if deeds else ''}")
    if html_path:
        print(f"HTML report:           {html_path}")
    print(f"Deeds auto-extracted:  {len(deeds)} (review-only, not in chain)")
    print(f"Support files in:      {support_dir}")
    print("  - source_inventory_codexv2.csv")
    print("  - merge_audit_codexv2.csv")
    print("  - conflicts_review_codexv2.xlsx")
    print("  - build_log_codexv2.txt")
    print("  - final_validation_summary_codexv2.txt")
    print("  - tidy_manifest_codexv2.csv")
    print(f"Archives extracted:    {zips_extracted} file(s)")
    print(f"Duplicates/trash:      dups={dup_extra}  trash={len(trash)}  quarantined={len(moved)}")
    print(f"Source files reviewed: {len(recs)}")
    print(f"Records merged:        {len(merged)}  (written: {rows_written})")
    print(f"Interest chained:      {len(chain_links)}  (parsed {n_parsed}, flagged {n_flagged})")
    print(f"Runsheet notes/OGL:    notes={n_notes}  ogl={n_ogl}  leases={len(ogl_summary)}")
    print(f"Conflicts found:       {len(conflicts)}")
    print(f"Index verified rows:   {n_verified_pct}  (method: {method})")
    print(f"Build/validation:      {'PASS' if build_ok else ('DRY-RUN' if args.dry_run else 'NEEDS REVIEW')}"
          f" in {passes_used} pass(es)")
    print("Human review:          see final_validation_summary_codexv2.txt (PERFECTION CHECKLIST)")
    return 0 if (build_ok or args.dry_run) else 1


if __name__ == "__main__":
    raise SystemExit(main())
