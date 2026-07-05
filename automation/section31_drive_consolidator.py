#!/usr/bin/env python3
"""Build SECTION31 final ownership/title report from the best base workbook.

Non-destructive: copies the 2026-07-05 HBP turn-in (best base) exactly and
APPENDS three hidden analysis sheets (QA_Checks, Gap_Log, Source_Index).
No existing cell, style, merge, or formula is altered. Also emits a standalone
QA log workbook and a $0 image-spend log.
"""
import openpyxl, re, datetime, os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "candidates/HBP_UPDATE_2026-07-05_RAW_BACKUP.xlsx"
FINAL = "exports/SECTION31_12N_24W_ROGER_MILLS_FINAL_OWNERSHIP_TITLE_REPORT.xlsx"
QA_LOG = "exports/SECTION31_QA_LOG.xlsx"
SPEND = "exports/image_spend_log.xlsx"

ERR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")

def num(v):
    """Parse a numeric net-acre value; None if not numeric."""
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None

# ---------------------------------------------------------------- load
wb = openpyxl.load_workbook(BASE, data_only=False)          # keep formulas
wbv = openpyxl.load_workbook(BASE, data_only=True)          # cached values
orig_sheets = list(wb.sheetnames)

findings = []   # (severity, category, sheet, cell, detail)
def add(sev, cat, sheet, cell, detail):
    findings.append((sev, cat, sheet, cell, detail))

# ---------------------------------------------------------------- QA 1: formula errors (all sheets, cached values)
for name in wbv.sheetnames:
    ws = wbv[name]
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip() in ERR_TOKENS:
                add("HIGH", "formula-error", name, c.coordinate, f"cell evaluates to {c.value.strip()}")

# ---------------------------------------------------------------- QA 2: parse Title tract blocks
tw = wbv["Title"]
maxr = tw.max_row
def cell(r, col): return tw.cell(r, col).value

tracts = []   # dict per tract
cur = None
for r in range(1, maxr + 1):
    b = cell(r, 2)
    bs = "" if b is None else str(b).strip()
    m = re.match(r"TRACT\s+(\d+)\s*:", bs, re.I)
    if m:
        if cur: tracts.append(cur)
        cur = {"num": int(m.group(1)), "legal": cell(r, 3), "start": r,
               "gross": None, "report_total": None, "owner_rows": []}
        continue
    if cur is None:
        continue
    if bs.lower().startswith("containing"):
        g = cell(r, 3)
        if g is not None:
            mm = re.search(r"([\d.]+)", str(g))
            if mm: cur["gross"] = float(mm.group(1))
        continue
    if bs in ("REPORT TOTAL", "TOTAL"):
        v = num(cell(r, 3))
        if v is not None and cur["report_total"] is None:
            cur["report_total"] = v
        # a TRACT block ends at its TOTAL
        if bs == "TOTAL":
            tracts.append(cur); cur = None
        continue
    if bs.startswith("Working Interest"):
        if cur: tracts.append(cur); cur = None
        continue
    # owner row: has a name in col B and a value in col C
    nv = num(cell(r, 3))
    if bs and bs not in ("Mineral Owner",) and nv is not None:
        cur["owner_rows"].append((r, b, nv, cell(r, 4), cell(r, 5), cell(r, 6), cell(r, 7)))
if cur: tracts.append(cur)
# de-dup tracts by num keeping first full
seen = {}
for t in tracts:
    if t["num"] not in seen and t.get("gross"):
        seen[t["num"]] = t
tracts = [seen[k] for k in sorted(seen)]

# ---------------------------------------------------------------- QA 3: tract acre reconciliation + negatives + fp noise
for t in tracts:
    s = sum(v for (_r, _n, v, *_x) in t["owner_rows"])
    gross = t["gross"]
    rt = t["report_total"]
    # net vs gross
    if gross is not None and s - gross > 0.01:
        add("HIGH", "over-conveyance",
            "Title", f"Tract {t['num']}",
            f"owner NMA sum {s:.2f} exceeds gross tract acres {gross:.2f} by {s-gross:.2f}")
    # report total vs owner sum
    if rt is not None and abs(rt - s) > 0.02:
        add("MED", "total-mismatch", "Title", f"Tract {t['num']}",
            f"stated REPORT TOTAL {rt:.4f} != sum of owner rows {s:.4f}")
    # report total vs gross (open balances allowed but note when it foots to gross by an OPEN line)
    for (r, n, v, ogl, roy, exp, com) in t["owner_rows"]:
        nm = "" if n is None else str(n).replace("\n", " ")[:45]
        if v < 0:
            add("HIGH", "negative-acres", "Title", f"C{r}", f"{nm}: net acres {v}")
        elif 0 < v < 1e-3:
            add("LOW", "fp-noise", "Title", f"C{r}",
                f"{nm}: net acres {v:g} is floating-point dust (~0); consider rounding to 0")
    # total with long fp tail
    if rt is not None and abs(rt - round(rt, 2)) > 1e-6:
        add("LOW", "fp-noise", "Title", f"Tract {t['num']} TOTAL",
            f"REPORT TOTAL {rt} carries floating-point tail; display rounds fine but value is not clean")

# ---------------------------------------------------------------- QA 4: missing address / royalty / OGL on leased rows
for t in tracts:
    for (r, n, v, ogl, roy, exp, com) in t["owner_rows"]:
        nm = "" if n is None else str(n).replace("\n", " ")
        ogls = "" if ogl is None else str(ogl).strip()
        exps = "" if exp is None else str(exp).strip()
        if "address not located" in nm.lower() or ("/" not in nm and len(nm) < 40 and v and v > 0.5):
            if "address not located" in nm.lower():
                add("MED", "missing-address", "Title", f"B{r}", f"{nm[:55]}: address not located")
        # leased (HBP / has expiration date) but no OGL number tied
        leased = ("HBP" in exps.upper()) or bool(re.search(r"\d/\d", exps))
        if leased and (ogls in ("", "—", "-") ):
            add("MED", "missing-ogl", "Title", f"D{r}",
                f"{nm[:45]}: row appears leased/HBP but no OGL number tied")

# ---------------------------------------------------------------- QA 5: OGL sheet basic completeness + dup book/page
ogl = wbv["OGL"]
# find header row
ogl_hdr = None
for r in range(1, 6):
    rowvals = [str(ogl.cell(r, c).value or "").strip().lower() for c in range(1, 12)]
    if any("ogl" in x for x in rowvals):
        ogl_hdr = r; break
ogl_count = 0
if ogl_hdr:
    for r in range(ogl_hdr + 1, ogl.max_row + 1):
        a = ogl.cell(r, 1).value
        if a not in (None, ""):
            ogl_count += 1

# ---------------------------------------------------------------- QA 6: Runsheet duplicate instrument scan
run = wbv["Runsheet"]
# find an instrument-number-ish column
inst_seen = {}
dups = 0
# scan columns for values like 20xx-xxxxxx or book/page NNNN/NNNN
inst_re = re.compile(r"\b(19|20)\d{2}-\d{4,6}\b")
for row in run.iter_rows():
    for c in row:
        if isinstance(c.value, str):
            for mm in inst_re.findall(c.value):
                pass
# lightweight: count exact instrument tokens across runsheet col
tok_counts = {}
for row in run.iter_rows():
    for c in row:
        if isinstance(c.value, str):
            for tok in inst_re.findall_iter if False else []:
                pass
# simpler robust dup detection over all cells
allmatches = {}
for row in run.iter_rows():
    for c in row:
        v = c.value
        if isinstance(v, str):
            for tok in re.findall(r"\b(?:19|20)\d{2}-\d{4,6}\b", v):
                allmatches.setdefault(tok, 0)
                allmatches[tok] += 1
dup_inst = {k: n for k, n in allmatches.items() if n > 3}   # >3 mentions = worth a look
for k, n in sorted(dup_inst.items(), key=lambda x: -x[1])[:15]:
    add("LOW", "duplicate-instrument", "Runsheet", "-",
        f"instrument {k} appears {n} times across runsheet (verify not double-counted)")

# ---------------------------------------------------------------- QA 7: open / needs-verification balances
open_rows = 0
for t in tracts:
    for (r, n, v, ogl2, roy, exp, com) in t["owner_rows"]:
        blob = " ".join(str(x) for x in (n, exp, com) if x)
        if re.search(r"OPEN\s*/\s*VERIFY|undetermined|Needs? Verif|verify", blob, re.I):
            open_rows += 1
            add("MED", "open-balance", "Title", f"Tract {t['num']} row {r}",
                f"{str(n)[:40]}: open/undetermined balance carried ({v:.2f} NMA) — needs source verification")

# ---------------------------------------------------------------- summarize
sev_order = {"HIGH": 0, "MED": 1, "LOW": 2}
findings.sort(key=lambda f: (sev_order.get(f[0], 3), f[1], f[2]))

# ================================================================ write hidden sheets into final
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF")
def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cc = ws.cell(1, c)
        cc.fill = HDR_FILL; cc.font = HDR_FONT
        cc.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

def newsheet(name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)

# QA_Checks
qc = newsheet("QA_Checks")
qc_cols = ["Severity", "Category", "Sheet", "Cell/Ref", "Finding"]
qc.append(qc_cols)
for f in findings:
    qc.append(list(f))
if not findings:
    qc.append(["INFO", "none", "-", "-", "No automated QA exceptions detected."])
style_header(qc, len(qc_cols))
widths = [10, 22, 12, 16, 90]
for i, w in enumerate(widths, 1):
    qc.column_dimensions[get_column_letter(i)].width = w
qc.sheet_state = "hidden"

# Gap_Log
gl = newsheet("Gap_Log")
gl_cols = ["Area", "Gap / Missing Item", "Tract/Where", "Impact", "Suggested Source to Pull", "Status"]
gl.append(gl_cols)
gap_rows = [
    ["Tract acres", "Tract 2 owner NMA (221.70) exceeds 160 gross — over-conveyance in Bain-family chain", "Tract 2", "High", "Bain/Herbold/Jensen mineral-deed chain review; reconcile double-allocated fractions", "OPEN"],
    ["Open balances", "Undetermined mineral fee balances carried on cursory basis", "Tracts 2,4,5,7,9", "Medium", "Source deeds/probate/assignments per Curative_Requirements sheet", "OPEN — disclosed"],
    ["Addresses", "Owners marked 'Address not located'", "Tracts 2,4,5", "Medium", "OKCountyRecords grantee index / skip-trace before owner notice", "OPEN"],
    ["HBP proof", "Alexander #1-31 HBP carried per client direction; OCC/OTC production not independently pulled", "WI / Well 1", "Medium", "OCC well file + OTC production; FracFocus completion", "OPEN — client-directed, optional QC"],
    ["OGL ties", "Some HBP-base successors not tied to a specific OGL number", "Tract 2 successors", "Medium", "Base lease images + assignment chain to bind exact OGL #", "OPEN"],
    ["Source images", "Instrument images not pulled in this environment (no OKCountyRecords credentials present)", "All", "Low", "OKCountyRecords API/browser on operator machine; log in image_spend_log", "OPEN"],
]
for row in gap_rows:
    gl.append(row)
style_header(gl, len(gl_cols))
for i, w in enumerate([16, 60, 16, 10, 50, 22], 1):
    gl.column_dimensions[get_column_letter(i)].width = w
gl.sheet_state = "hidden"

# Source_Index
si = newsheet("Source_Index")
si_cols = ["Source Type", "Reference", "Where Used", "Obtained?", "Notes"]
si.append(si_cols)
si_rows = [
    ["County index", "12N 24W 31 Index PDF (through 05/05/2026, last entry 2704/142)", "Runsheet / whole report", "Referenced", "Controlling index for cursory turn-in"],
    ["Base OGLs", "OGL sheet OGL numbers (69 leases incl. 6 Silver Oak top leases)", "OGL / Title / Tract / WI", "In workbook", "OGL numbers controlling per 2026-07-05 client direction"],
    ["Well", "Alexander #1-31, API 35-129-22925, Martin's Resources LLC", "Well 1 / WI", "In workbook", "HBP basis per client direction"],
    ["WI assignments", "Wellbore assignment chain 1567/0006 → 2698/0069", "WI 1", "In workbook", "12-step leasehold successor chain"],
    ["Instrument images", "Not pulled this session", "-", "No", "No OKCountyRecords credentials in this environment; $0 spent (see image_spend_log.xlsx)"],
]
for row in si_rows:
    si.append(row)
style_header(si, len(si_cols))
for i, w in enumerate([16, 60, 26, 12, 60], 1):
    si.column_dimensions[get_column_letter(i)].width = w
si.sheet_state = "hidden"

os.makedirs("exports", exist_ok=True)
wb.save(FINAL)

# ================================================================ standalone QA log (visible)
qwb = openpyxl.Workbook()
s = qwb.active; s.title = "QA_Checks"
s.append(qc_cols)
for f in findings:
    s.append(list(f))
style_header(s, len(qc_cols))
for i, w in enumerate(widths, 1):
    s.column_dimensions[get_column_letter(i)].width = w
s2 = qwb.create_sheet("Gap_Log"); s2.append(gl_cols)
for row in gap_rows: s2.append(row)
style_header(s2, len(gl_cols))
for i, w in enumerate([16, 60, 16, 10, 50, 22], 1):
    s2.column_dimensions[get_column_letter(i)].width = w
s3 = qwb.create_sheet("Source_Index"); s3.append(si_cols)
for row in si_rows: s3.append(row)
style_header(s3, len(si_cols))
for i, w in enumerate([16, 60, 26, 12, 60], 1):
    s3.column_dimensions[get_column_letter(i)].width = w
qwb.save(QA_LOG)

# ================================================================ image spend log ($0)
swb = openpyxl.Workbook(); ss = swb.active; ss.title = "image_spend_log"
scols = ["Date", "Source", "County", "Book", "Page", "Instrument Number", "Image Number",
         "Party", "Type", "Cost", "Reason", "Used In Workbook", "Notes"]
ss.append(scols)
ss.append(["2026-07-05", "OKCountyRecords", "Roger Mills", "", "", "", "", "", "", 0.00,
           "No credentials present in cloud environment; no images purchased",
           "No", "Cursory turn-in built from existing compiled workbooks; $0 spend. Total spend: $0.00 of $100 cap."])
style_header(ss, len(scols))
for i in range(1, len(scols) + 1):
    ss.column_dimensions[get_column_letter(i)].width = 16
swb.save(SPEND)

# ================================================================ report
print("BASE sheets:", len(orig_sheets))
print("FINAL sheets:", len(wb.sheetnames), "(added hidden: QA_Checks, Gap_Log, Source_Index)")
print(f"OGL leases counted: {ogl_count}")
print(f"Tracts parsed: {[t['num'] for t in tracts]}")
print(f"QA findings: {len(findings)}  "
      f"(HIGH={sum(1 for f in findings if f[0]=='HIGH')}, "
      f"MED={sum(1 for f in findings if f[0]=='MED')}, "
      f"LOW={sum(1 for f in findings if f[0]=='LOW')})")
print("---- HIGH/MED findings ----")
for f in findings:
    if f[0] in ("HIGH", "MED"):
        print(f"  [{f[0]}] {f[1]} | {f[2]} {f[3]} | {f[4]}")
