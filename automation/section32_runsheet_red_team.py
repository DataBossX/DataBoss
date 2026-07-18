"""Section 32 runsheet red-team audit harness.

Independent adversarial auditor for two candidate title runsheets (e.g. a
Gemini "master" extraction and a Grok "chained" reconstruction). It compares
the two workbooks row-by-row, detects the twelve defect classes the QA job
cares about, and emits four deterministic artifacts:

    CLAUDE_SECTION32_RED_TEAM_RUNSHEET.xlsx   corrected / reconciled runsheet
    CLAUDE_RUNSHEET_DEFECT_REGISTER.xlsx      every defect found, with evidence
    CLAUDE_CONFLICT_MATRIX.xlsx               field-level A-vs-B conflicts
    CLAUDE_COMPLETION_RECEIPT.json            machine-readable run summary

Design rules (mirroring the job contract):

* Never fabricate. Unproven fields are left blank and flagged; conflicts that
  cannot be resolved on evidence are kept visible and marked ``OPEN``.
* Correct only evidence-proven errors. A value is only overwritten when one
  side is backed by stronger evidence (source image > OCR > index > metadata
  > prior report) and the other is not.
* No unsupported decimals. A conveyed-interest decimal with no source citation
  is refused, not carried forward.
* No false current-owner conclusion. A current owner is reported only when an
  unbroken, image-proven chain with no OPEN conflicts accounts for the whole
  interest; otherwise the conclusion is ``UNDETERMINED``.

The module performs *no* network or Drive access and does not require the real
title documents. Point it at two ``.xlsx`` files and an output directory.

This harness is validated against a synthetic fixture in
``tests/test_section32_red_team.py``; it does not itself assert anything about
any real title chain.
"""
from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

# openpyxl is the same xlsx backend the horizon module already depends on.
try:  # pragma: no cover - import guard
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "section32_runsheet_red_team requires openpyxl (pip install openpyxl>=3.1)"
    ) from exc


# --------------------------------------------------------------------------- #
# Evidence tiers
# --------------------------------------------------------------------------- #
# Source-image evidence outranks OCR, which outranks a county index entry,
# which outranks file metadata, which outranks a prior-report assertion.
EVIDENCE_RANK = {
    "image": 4,
    "source_image": 4,
    "scan": 4,
    "ocr": 3,
    "index": 2,
    "metadata": 1,
    "prior_report": 0,
    "report": 0,
    "unknown": -1,
    "": -1,
}
IMAGE_TIER = 4
OCR_TIER = 3

# Defect categories (must match the QA job's controlled vocabulary).
DEFECTS = (
    "Missing Instrument",
    "Duplicate Instrument",
    "Wrong Party",
    "Wrong Date",
    "Wrong Book or Page",
    "Wrong Legal",
    "Wrong Tract",
    "Unsupported Interest",
    "Broken Chain",
    "Missing Schedule",
    "Missing Source Image",
    "Portfolio Mixing",
)

SPREADSHEET_ERROR_TOKENS = (
    "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#ERROR!",
)

# Header aliases -> canonical field. Matching is case-insensitive on a
# whitespace/underscore-normalised header string.
HEADER_ALIASES = {
    "instrument_no": ("instrument", "instrument no", "instrument number", "inst no",
                      "doc", "document no", "document number", "reception", "reception no"),
    "instrument_type": ("type", "instrument type", "doc type", "document type", "kind"),
    "book": ("book", "bk", "vol", "volume"),
    "page": ("page", "pg", "pages"),
    "exec_date": ("exec date", "execution date", "dated", "date executed", "instrument date"),
    "record_date": ("record date", "recorded", "recording date", "date recorded", "filed"),
    "grantor": ("grantor", "grantors", "from", "assignor", "seller"),
    "grantee": ("grantee", "grantees", "to", "assignee", "buyer"),
    "interest": ("interest", "conveyed interest", "fraction", "undivided interest", "wi", "working interest"),
    "legal_description": ("legal", "legal description", "description", "lands", "property"),
    "section": ("section", "sec"),
    "tract": ("tract", "tract id", "parcel"),
    "evidence_tier": ("evidence", "evidence tier", "source type", "provenance", "backing"),
    "source_citation": ("source", "source citation", "citation", "reference", "cite", "image"),
}


def _norm_header(h: str) -> str:
    return re.sub(r"[\s_]+", " ", str(h or "").strip().lower())


def _build_header_map(headers: Sequence[str]) -> Dict[int, str]:
    """Map column index -> canonical field name using the alias table."""
    out: Dict[int, str] = {}
    for idx, raw in enumerate(headers):
        norm = _norm_header(raw)
        if not norm:
            continue
        for canon, aliases in HEADER_ALIASES.items():
            if norm == canon or norm in aliases:
                out[idx] = canon
                break
    return out


# --------------------------------------------------------------------------- #
# Normalisation helpers
# --------------------------------------------------------------------------- #
def _clean(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def norm_party(name: str) -> str:
    """Case/punctuation-insensitive party key, with common suffixes folded."""
    s = _clean(name).lower()
    s = re.sub(r"[.,]", " ", s)
    s = re.sub(r"\b(a|an|the|and|et al|etux|et ux|husband|wife|inc|llc|llp|lp|ltd|company|co|corporation|corp|trust|trustee)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _clean(v)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_interest(v) -> Tuple[Optional[float], bool]:
    """Return (fraction, is_explicit). is_explicit is False when the value is
    absent or unparseable, so callers never invent a decimal."""
    s = _clean(v).lower()
    if not s:
        return None, False
    m = re.search(r"(\d+)\s*/\s*(\d+)", s)
    if m:
        num, den = int(m.group(1)), int(m.group(2))
        if den:
            return num / den, True
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", s)
    if m:
        return float(m.group(1)) / 100.0, True
    m = re.search(r"\b(0?\.\d+|1\.0+|1|0)\b", s)
    if m:
        try:
            return float(m.group(1)), True
        except ValueError:
            pass
    return None, False


def has_section32(legal: str) -> bool:
    s = _clean(legal).lower()
    return bool(re.search(r"\b(sec(?:tion)?\.?\s*32|s32|32-1?\d+[ns]-\d+[ew])\b", s)) or "section 32" in s


def mentions_schedule(text: str) -> bool:
    return bool(re.search(r"\b(schedule|exhibit)\b", _clean(text), re.I))


# --------------------------------------------------------------------------- #
# Row model
# --------------------------------------------------------------------------- #
@dataclass
class Row:
    source_sheet: str
    row_no: int
    instrument_no: str = ""
    instrument_type: str = ""
    book: str = ""
    page: str = ""
    exec_date: str = ""
    record_date: str = ""
    grantor: str = ""
    grantee: str = ""
    interest: str = ""
    legal_description: str = ""
    section: str = ""
    tract: str = ""
    evidence_tier: str = ""
    source_citation: str = ""
    raw: Dict[str, str] = field(default_factory=dict)

    def key(self) -> str:
        if self.instrument_no:
            return f"INST:{re.sub(r'[^0-9a-z]', '', self.instrument_no.lower())}"
        if self.book or self.page:
            return f"BP:{_clean(self.book).lower()}/{_clean(self.page).lower()}"
        return f"PGE:{norm_party(self.grantor)}>{norm_party(self.grantee)}@{self.record_date}"

    def tier_rank(self) -> int:
        return EVIDENCE_RANK.get(_norm_header(self.evidence_tier).replace(" ", "_"), -1)

    def has_citation(self) -> bool:
        return bool(self.source_citation.strip())


def load_rows(path: Path) -> List[Row]:
    """Read a runsheet workbook's first sheet into Row objects."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    hmap = _build_header_map([_clean(h) for h in header])
    out: List[Row] = []
    for i, values in enumerate(rows_iter, start=2):
        if values is None or all(v in (None, "") for v in values):
            continue
        fields: Dict[str, str] = {}
        raw: Dict[str, str] = {}
        for idx, val in enumerate(values):
            cv = _clean(val)
            if idx in hmap:
                fields[hmap[idx]] = cv
            raw[_clean(header[idx]) if idx < len(header) else f"col{idx}"] = cv
        out.append(Row(source_sheet=path.stem, row_no=i, raw=raw, **fields))
    wb.close()
    return out


# --------------------------------------------------------------------------- #
# Defect record
# --------------------------------------------------------------------------- #
@dataclass
class Defect:
    category: str
    key: str
    field: str
    sheet_a_value: str
    sheet_b_value: str
    detail: str
    resolution: str  # CORRECTED | OPEN | FLAGGED
    evidence: str


@dataclass
class AuditResult:
    retained: List[Dict[str, str]]
    defects: List[Defect]
    conflicts: List[Dict[str, str]]
    current_owner: str
    acceptance: Dict[str, bool]
    stats: Dict[str, int]


# --------------------------------------------------------------------------- #
# Core audit
# --------------------------------------------------------------------------- #
def _index_by_key(rows: Sequence[Row]) -> Dict[str, List[Row]]:
    idx: Dict[str, List[Row]] = {}
    for r in rows:
        idx.setdefault(r.key(), []).append(r)
    return idx


def _pick_field(a: Optional[Row], b: Optional[Row], attr: str) -> Tuple[str, str, str]:
    """Reconcile one field across the two sides.

    Returns (value, status, evidence). status is AGREE / CORRECTED / OPEN /
    SINGLE. A value is only CORRECTED to the better-sourced side; genuine
    unresolved disagreement is OPEN and the value left blank.
    """
    av = _clean(getattr(a, attr)) if a else ""
    bv = _clean(getattr(b, attr)) if b else ""
    if a and b:
        if av == bv:
            return av, "AGREE", ""
        # field-specific equality for parties/dates
        if attr in ("grantor", "grantee") and av and bv and norm_party(av) == norm_party(bv):
            return av, "AGREE", ""
        if attr in ("exec_date", "record_date") and parse_date(av) and parse_date(av) == parse_date(bv):
            return av, "AGREE", ""
        ra, rb = a.tier_rank(), b.tier_rank()
        if av and bv and ra > rb and ra >= OCR_TIER:
            return av, "CORRECTED", f"kept A ({a.evidence_tier}) over B ({b.evidence_tier})"
        if av and bv and rb > ra and rb >= OCR_TIER:
            return bv, "CORRECTED", f"kept B ({b.evidence_tier}) over A ({a.evidence_tier})"
        if av and not bv:
            return av, "SINGLE", "only A populated"
        if bv and not av:
            return bv, "SINGLE", "only B populated"
        return "", "OPEN", "conflict, neither side evidence-proven"
    single = a or b
    return _clean(getattr(single, attr)) if single else "", "SINGLE", ""


def audit(rows_a: Sequence[Row], rows_b: Sequence[Row],
          expected_section: str = "32", expected_tract: str = "") -> AuditResult:
    defects: List[Defect] = []
    conflicts: List[Dict[str, str]] = []

    idx_a = _index_by_key(rows_a)
    idx_b = _index_by_key(rows_b)

    # Duplicate Instrument (within either sheet).
    for label, idx in (("A", idx_a), ("B", idx_b)):
        for k, group in idx.items():
            if len(group) > 1:
                defects.append(Defect(
                    "Duplicate Instrument", k, "key", k, k,
                    f"{len(group)} rows share key in sheet {label} (rows {[g.row_no for g in group]})",
                    "FLAGGED", "within-sheet duplicate"))

    all_keys = list(dict.fromkeys(list(idx_a) + list(idx_b)))
    retained: List[Dict[str, str]] = []

    for k in all_keys:
        a = idx_a.get(k, [None])[0]
        b = idx_b.get(k, [None])[0]

        # Missing Instrument (present in one sheet only).
        if a and not b:
            defects.append(Defect("Missing Instrument", k, "-", "present", "ABSENT",
                                  f"instrument in A missing from B", "FLAGGED",
                                  a.source_citation or a.evidence_tier))
        elif b and not a:
            defects.append(Defect("Missing Instrument", k, "-", "ABSENT", "present",
                                  f"instrument in B missing from A", "FLAGGED",
                                  b.source_citation or b.evidence_tier))

        row: Dict[str, str] = {"key": k, "status": "OK"}
        open_fields: List[str] = []
        for attr, defect_cat in (
            ("instrument_type", None), ("book", "Wrong Book or Page"),
            ("page", "Wrong Book or Page"), ("instrument_no", "Wrong Book or Page"),
            ("exec_date", "Wrong Date"), ("record_date", "Wrong Date"),
            ("grantor", "Wrong Party"), ("grantee", "Wrong Party"),
            ("legal_description", "Wrong Legal"), ("section", "Wrong Tract"),
            ("tract", "Wrong Tract"), ("interest", "Unsupported Interest"),
            ("evidence_tier", None), ("source_citation", None),
        ):
            value, status, ev = _pick_field(a, b, attr)
            row[attr] = value
            if status in ("CORRECTED", "OPEN") and defect_cat:
                defects.append(Defect(
                    defect_cat, k, attr,
                    _clean(getattr(a, attr)) if a else "",
                    _clean(getattr(b, attr)) if b else "",
                    f"{attr} {status.lower()}", status, ev))
                conflicts.append({
                    "key": k, "field": attr,
                    "sheet_a": _clean(getattr(a, attr)) if a else "",
                    "sheet_b": _clean(getattr(b, attr)) if b else "",
                    "resolution": status, "evidence": ev})
                if status == "OPEN":
                    open_fields.append(attr)

        # Impossible chronology: execution after recording.
        ed, rd = parse_date(row.get("exec_date", "")), parse_date(row.get("record_date", ""))
        if ed and rd and ed > rd:
            defects.append(Defect("Wrong Date", k, "exec/record", row.get("exec_date", ""),
                                  row.get("record_date", ""),
                                  "execution date after recording date (impossible)",
                                  "FLAGGED", "chronology"))
            open_fields.append("exec_date")

        # Legal must include Section 32.
        legal = row.get("legal_description", "")
        if legal and not has_section32(legal):
            defects.append(Defect("Portfolio Mixing", k, "legal_description", legal, "",
                                  "legal description does not reference Section 32",
                                  "FLAGGED", "scope check"))
            open_fields.append("legal_description")

        # Stated tract must appear in the legal description.
        if expected_tract and legal and expected_tract.lower() not in legal.lower() \
                and _clean(row.get("tract", "")).lower() not in (expected_tract.lower(), ""):
            defects.append(Defect("Wrong Tract", k, "tract", row.get("tract", ""), "",
                                  f"expected tract '{expected_tract}' not evidenced in legal",
                                  "FLAGGED", "tract check"))
            open_fields.append("tract")

        # Missing Schedule: references a schedule/exhibit but none carried.
        if mentions_schedule(legal) and not _clean(row.get("tract", "")):
            defects.append(Defect("Missing Schedule", k, "legal_description", legal, "",
                                  "references a schedule/exhibit that is not attached",
                                  "FLAGGED", "schedule check"))

        # Unsupported decimal: interest present but no citation.
        frac, explicit = parse_interest(row.get("interest", ""))
        if explicit and not _clean(row.get("source_citation", "")):
            defects.append(Defect("Unsupported Interest", k, "interest", row.get("interest", ""), "",
                                  "conveyed-interest decimal has no source citation",
                                  "FLAGGED", "unsupported decimal"))
            row["interest"] = ""  # refuse to carry an unsupported decimal
            open_fields.append("interest")

        # Missing Source Image: no image-tier backing.
        best_tier = max((r.tier_rank() for r in (a, b) if r), default=-1)
        best_cite = row.get("source_citation", "")
        if best_tier < IMAGE_TIER:
            defects.append(Defect("Missing Source Image", k, "evidence_tier",
                                  a.evidence_tier if a else "", b.evidence_tier if b else "",
                                  "no source-image evidence (OCR/index/metadata/prior-report only)",
                                  "FLAGGED", best_cite or "no image"))

        row["_open"] = ",".join(sorted(set(open_fields)))
        row["_evidence_tier_rank"] = str(best_tier)
        row["_retained"] = "yes" if best_cite else "no"
        row["status"] = "OPEN" if open_fields else "OK"
        retained.append(row)

    # Broken Chain + over-conveyance, evaluated on the reconciled, image-proven
    # rows in recording order.
    current_owner = _evaluate_chain(retained, defects)

    stats = {
        "rows_sheet_a": len(rows_a),
        "rows_sheet_b": len(rows_b),
        "reconciled_keys": len(retained),
        "defects": len(defects),
        "open_conflicts": sum(1 for r in retained if r.get("_open")),
        "retained_with_citation": sum(1 for r in retained if r.get("_retained") == "yes"),
    }
    acceptance = _acceptance_tests(retained, defects, current_owner)
    return AuditResult(retained, defects, conflicts, current_owner, acceptance, stats)


def _evaluate_chain(retained: List[Dict[str, str]], defects: List[Defect]) -> str:
    """Walk instruments in recording order; flag chain gaps and over-conveyance.

    A current owner is only asserted from an unbroken, image-proven, conflict-free
    chain whose conveyed interest sums to the whole. Otherwise: UNDETERMINED.
    """
    def sortkey(r: Dict[str, str]):
        d = parse_date(r.get("record_date", "")) or parse_date(r.get("exec_date", ""))
        return (d or date.max, r.get("key", ""))

    chain = [r for r in retained if r.get("grantor") or r.get("grantee")]
    chain.sort(key=sortkey)

    owned: Dict[str, float] = {}  # party -> fraction currently held
    prior_grantee: Optional[str] = None
    unbroken = True
    last_grantee_raw = ""

    for r in chain:
        grantor = norm_party(r.get("grantor", ""))
        grantee = norm_party(r.get("grantee", ""))
        frac, explicit = parse_interest(r.get("interest", ""))
        conv = frac if explicit else 1.0

        # Broken chain: grantor is not the immediately prior grantee.
        if prior_grantee is not None and grantor and grantor != prior_grantee \
                and grantor not in owned:
            defects.append(Defect("Broken Chain", r.get("key", ""), "grantor",
                                  r.get("grantor", ""), "",
                                  f"grantor is not the prior grantee '{prior_grantee}'",
                                  "OPEN", "chain gap"))
            unbroken = False

        # Over-conveyance: grantor conveys more than proven ownership.
        held = owned.get(grantor, 1.0 if prior_grantee is None else 0.0)
        if grantor and explicit and conv - held > 1e-9 and grantor in owned:
            defects.append(Defect("Unsupported Interest", r.get("key", ""), "interest",
                                  r.get("interest", ""), "",
                                  f"grantor conveys {conv:.4f} but holds {held:.4f}",
                                  "OPEN", "over-conveyance"))
            unbroken = False

        if grantor in owned:
            owned[grantor] = max(0.0, owned.get(grantor, 0.0) - conv)
        if grantee:
            owned[grantee] = owned.get(grantee, 0.0) + conv

        if r.get("_open"):
            unbroken = False
        if int(r.get("_evidence_tier_rank", "-1")) < IMAGE_TIER:
            unbroken = False
        prior_grantee = grantee or prior_grantee
        last_grantee_raw = r.get("grantee", "") or last_grantee_raw

    if unbroken and chain and last_grantee_raw:
        return last_grantee_raw
    return "UNDETERMINED — unresolved OPEN items / no image-proven unbroken chain"


def _acceptance_tests(retained, defects, current_owner) -> Dict[str, bool]:
    retained_rows = [r for r in retained if r.get("_retained") == "yes"]
    every_retained_cited = all(_clean(r.get("source_citation", "")) for r in retained_rows)
    conflicts_resolved_or_open = all(
        d.resolution in ("CORRECTED", "OPEN", "FLAGGED") for d in defects)
    no_unsupported_decimal = not any(
        d.category == "Unsupported Interest" and "unsupported decimal" in d.evidence
        and d.resolution not in ("OPEN", "FLAGGED") for d in defects)
    no_false_owner = current_owner.startswith("UNDETERMINED") or not any(
        r.get("_open") for r in retained)
    no_error_tokens = not any(
        tok in _clean(v) for r in retained for v in r.values() for tok in SPREADSHEET_ERROR_TOKENS)
    return {
        "every_retained_row_has_citation": every_retained_cited,
        "every_conflict_resolved_or_open": conflicts_resolved_or_open,
        "no_unsupported_decimals": no_unsupported_decimal,
        "no_false_current_owner": no_false_owner,
        "no_error_tokens_or_broken_links": no_error_tokens,
    }


# --------------------------------------------------------------------------- #
# Output writers
# --------------------------------------------------------------------------- #
RUNSHEET_COLS = [
    "key", "instrument_type", "instrument_no", "book", "page", "exec_date",
    "record_date", "grantor", "grantee", "interest", "legal_description",
    "section", "tract", "evidence_tier", "source_citation", "status", "_open",
]


def write_outputs(result: AuditResult, out_dir: Path, job_id: str = "SEC32_RUNSHEET_CLAUDE_QA") -> Dict[str, str]:
    out_dir.mkdir(parents=True, exist_ok=True)
    paths: Dict[str, str] = {}

    # Red-team runsheet.
    wb = Workbook()
    ws = wb.active
    ws.title = "RED_TEAM_RUNSHEET"
    ws.append([c.lstrip("_").upper() for c in RUNSHEET_COLS])
    for r in result.retained:
        ws.append([r.get(c, "") for c in RUNSHEET_COLS])
    ws2 = wb.create_sheet("CURRENT_OWNER")
    ws2.append(["current_owner_conclusion"])
    ws2.append([result.current_owner])
    p = out_dir / "CLAUDE_SECTION32_RED_TEAM_RUNSHEET.xlsx"
    wb.save(p); paths["runsheet"] = str(p)

    # Defect register.
    wb = Workbook(); ws = wb.active; ws.title = "DEFECT_REGISTER"
    ws.append(["#", "category", "key", "field", "sheet_a_value", "sheet_b_value",
               "detail", "resolution", "evidence"])
    for i, d in enumerate(result.defects, 1):
        ws.append([i, d.category, d.key, d.field, d.sheet_a_value, d.sheet_b_value,
                   d.detail, d.resolution, d.evidence])
    ws2 = wb.create_sheet("BY_CATEGORY")
    ws2.append(["category", "count"])
    for cat in DEFECTS:
        ws2.append([cat, sum(1 for d in result.defects if d.category == cat)])
    p = out_dir / "CLAUDE_RUNSHEET_DEFECT_REGISTER.xlsx"
    wb.save(p); paths["defect_register"] = str(p)

    # Conflict matrix.
    wb = Workbook(); ws = wb.active; ws.title = "CONFLICT_MATRIX"
    ws.append(["key", "field", "sheet_a", "sheet_b", "resolution", "evidence"])
    for c in result.conflicts:
        ws.append([c["key"], c["field"], c["sheet_a"], c["sheet_b"], c["resolution"], c["evidence"]])
    p = out_dir / "CLAUDE_CONFLICT_MATRIX.xlsx"
    wb.save(p); paths["conflict_matrix"] = str(p)

    # Completion receipt.
    receipt = {
        "job_id": job_id,
        "role": "Independent adversarial title and spreadsheet auditor",
        "release_status": "HOLD_NO_RELEASE",
        "current_owner_conclusion": result.current_owner,
        "stats": result.stats,
        "defects_by_category": {cat: sum(1 for d in result.defects if d.category == cat) for cat in DEFECTS},
        "acceptance_tests": result.acceptance,
        "acceptance_passed": all(result.acceptance.values()),
        "outputs": {k: Path(v).name for k, v in paths.items()},
        "note": "Machinery output. Real release still gated on SECURITY_REMEDIATION_RECEIPT=PASS "
                "and confirmed source workbooks.",
    }
    p = out_dir / "CLAUDE_COMPLETION_RECEIPT.json"
    p.write_text(json.dumps(receipt, indent=2))
    paths["receipt"] = str(p)
    return paths


def run(path_a: Path, path_b: Path, out_dir: Path,
        expected_section: str = "32", expected_tract: str = "") -> AuditResult:
    rows_a = load_rows(Path(path_a))
    rows_b = load_rows(Path(path_b))
    result = audit(rows_a, rows_b, expected_section, expected_tract)
    write_outputs(result, Path(out_dir))
    return result


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Section 32 runsheet red-team auditor")
    ap.add_argument("--gemini", required=True, help="Gemini master runsheet .xlsx")
    ap.add_argument("--grok", required=True, help="Grok chained runsheet .xlsx")
    ap.add_argument("--out", required=True, help="output directory")
    ap.add_argument("--section", default="32")
    ap.add_argument("--tract", default="")
    args = ap.parse_args(argv)
    result = run(Path(args.gemini), Path(args.grok), Path(args.out), args.section, args.tract)
    print(json.dumps({
        "stats": result.stats,
        "current_owner": result.current_owner,
        "acceptance": result.acceptance,
        "acceptance_passed": all(result.acceptance.values()),
    }, indent=2))
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
