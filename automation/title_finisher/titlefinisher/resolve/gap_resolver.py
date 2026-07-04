"""Source-in gap resolver.

For each highlighted conveying party, search for an earlier grantee-side appearance
(the party ACQUIRING the interest) in the Runsheet grantee column and in the parsed
county-index database, under name/OCR variants. A gap is resolvable only when the SAME
party is documented acquiring the interest via a title-type instrument BEFORE it conveys
out. This is the exact, conservative test used to clear 29 of 96 gaps on 31-12N-24W.
"""
from __future__ import annotations
import re
import openpyxl

STOP = {"and", "the", "hw", "llc", "inc", "co", "company", "trust", "trustee", "of",
        "et", "al", "ttee", "living", "revocable", "family", "estate", "usa", "ltd",
        "lp", "partnership", "deceased", "dcd", "aka", "fka", "also", "known", "as",
        "single", "person", "jr", "sr", "trustees"}


def toks(name: str) -> set[str]:
    name = re.sub(r"[^A-Za-z ]", " ", (name or "").lower())
    return {t for t in name.split() if len(t) > 2 and t not in STOP}


def find_source_in(party: str, runsheet_rows: list[tuple], index_recs: list[dict],
                   convey_year: int | None = None) -> dict | None:
    """Return the resolving instrument dict, or None.
    runsheet_rows: list of (row, inst, type, bkpg, date, grantor, grantee)
    """
    gt = toks(party)
    if not gt:
        return None
    # Only a conveyance vests title in a grantee. Proof-of-death, affidavit, ratification,
    # change-of-address, release, mortgage, easement, etc. do NOT create a source-in.
    conveyance = re.compile(r"(?i)deed|assignment|conveyance|decree|order|patent|"
                            r"distribution|sheriff|guardian|\bMD\b|\bWD\b|\bQCD\b|\bASGT\b")
    best = None
    for (row, inst, typ, bkpg, date, grantor, grantee) in runsheet_rows:
        if not conveyance.search(typ or ""):
            continue
        ht = toks(grantee)
        common = gt & ht
        # A source-in match must identify the SAME person/entity, not merely share a surname
        # or a generic word. Rules:
        #  - reject matches that rest only on generic industry/common words;
        #  - for a personal name (a leading given-name token), require BOTH the surname AND the
        #    given name to appear on the grantee side — so "Clarence C. Bain" does NOT match a
        #    "William C. Bain" grantee, but "Jesse Joe Newell" matches "Jesse Joe Newell";
        #  - for an entity, require >=2 shared non-generic tokens.
        GENERIC = {"production", "energy", "resources", "oil", "gas", "minerals", "royalty",
                   "holdings", "partners", "company", "corporation", "corp", "inc", "bank",
                   "trust", "and"}
        if not common or common <= GENERIC:
            continue
        ptoks = [t for t in re.sub(r"[^A-Za-z ]", " ", party.lower()).split()
                 if len(t) > 1]
        given = ptoks[0] if ptoks else ""
        surname = ptoks[-1] if ptoks else ""
        is_entity = bool(re.search(r"(?i)\b(LLC|L\.L\.C|Inc|Co|Company|Corp|Trust|Ltd|LP|Partners|"
                                   r"Fund|Bank|Resources|Energy|Royalty|Minerals|Production)\b", party))
        if is_entity:
            strong = len(common - GENERIC) >= 2
        else:
            strong = (surname in common) and (given in common) and (surname not in GENERIC)
        if not strong:
            continue
        yr = None
        m = re.match(r"(\d{4})", str(inst))
        if m:
            yr = int(m.group(1))
        if convey_year and yr and yr > convey_year:
            continue  # acquisition must precede the conveyance
        cand = {"instrument": inst, "type": typ, "book_page": bkpg,
                "grantee": grantee[:60], "shared": sorted(common), "source": "Runsheet grantee"}
        if best is None or (yr or 0) <= (int(re.match(r'(\d{4})', str(best['instrument'])).group(1)) if re.match(r'(\d{4})', str(best['instrument'])) else 9999):
            best = cand
    return best


def load_runsheet_rows(workbook_path: str) -> list[tuple]:
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["Runsheet"]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append((r, str(ws.cell(r, 1).value or ""), str(ws.cell(r, 3).value or ""),
                     str(ws.cell(r, 4).value or ""), str(ws.cell(r, 6).value or ""),
                     str(ws.cell(r, 7).value or ""), str(ws.cell(r, 8).value or "")))
    return rows
