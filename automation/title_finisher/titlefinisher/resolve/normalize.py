"""Deterministic, source-safe normalization transforms applied before gap resolution.

These encode conventions, not new facts:
  1. NET ACRES (tract col C) — the report's pro-rata SUMIFS footing, so each tract foots
     to its gross acres. Computes the cached value too, so the file previews correctly.
  2. Wellbore/leasehold TBD cells — replaced with the sheet's own N/A conventions
     (a wellbore-only assignment carries no NMA; a springing top lease is 1.00 WI / 0.8125
     NRI insofar as the lessor's interest). No fabricated ownership.

Returns a list of (sheet, coord, op) ready for SurgicalEditor.apply().
"""
from __future__ import annotations
import openpyxl
from openpyxl.utils import get_column_letter

C_FORMULA = ('=IFERROR(IF(AND($E{r}>0,$D{r}<>"The Public"),'
             '$E{r}/SUMIFS($E$10:$E$203,$E$10:$E$203,">0",$D$10:$D$203,"<>The Public")'
             '*$D$5,""),"")')

WI1_WELLBORE = "N/A - wellbore WI (no NMA)"
WI2_TOPLEASE = "1.00 WI / 0.8125 NRI insofar as lessor's interest (springing top lease, 3/16 RI)"


def net_acre_ops(workbook_path: str) -> list[tuple]:
    wbv = openpyxl.load_workbook(workbook_path, data_only=True)
    ops: list[tuple] = []
    for i in range(1, 11):
        t = f"Tract {i}"
        if t not in wbv.sheetnames:
            continue
        ws = wbv[t]
        d5 = ws["D5"].value
        if not isinstance(d5, (int, float)):
            continue
        Ev, Dv = {}, {}
        for r in range(10, 204):
            e = ws.cell(r, 5).value
            Ev[r] = e if isinstance(e, (int, float)) else None
            Dv[r] = ws.cell(r, 4).value
        denom = sum(e for r, e in Ev.items() if e and e > 0 and str(Dv[r]) != "The Public")
        for r in range(10, 204):
            e = Ev[r]
            if e and e > 0 and str(Dv[r]) != "The Public" and denom:
                cached = e / denom * d5
            else:
                cached = ""
            ops.append((t, f"C{r}", ("f", C_FORMULA.format(r=r), cached)))
    return ops


def tbd_convention_ops(workbook_path: str) -> list[tuple]:
    wb = openpyxl.load_workbook(workbook_path)
    ops: list[tuple] = []
    if "WI 1" in wb.sheetnames:
        ws = wb["WI 1"]
        for r in range(10, 42):
            if str(ws.cell(r, 3).value).strip() == "TBD":
                ops.append(("WI 1", f"C{r}", ("v", WI1_WELLBORE)))
    if "WI 2" in wb.sheetnames:
        ws = wb["WI 2"]
        for r in range(9, 17):
            v = ws.cell(r, 3).value
            if isinstance(v, str) and "TBD" in v.upper():
                ops.append(("WI 2", f"C{r}", ("v", WI2_TOPLEASE)))
    if "Title " in wb.sheetnames:
        ws = wb["Title "]
        conv = {
            "wellbore WI only": "N/A - wellbore WI only",
            "lien/collateral": "N/A - lien/collateral only",
            "release only": "N/A - release only",
            "mortgage/lien only": "N/A - mortgage/lien only",
        }
        for r in range(120, 150):
            v = ws.cell(r, 3).value
            if isinstance(v, str) and v.strip() == "TBD":
                # look at the row's comment (col G / D) to pick the right N/A convention
                ctx = " ".join(str(ws.cell(r, cc).value or "") for cc in (4, 7)).lower()
                repl = next((txt for key, txt in conv.items() if key.split()[0] in ctx),
                            "N/A - not applicable (see comment)")
                ops.append(("Title ", f"C{r}", ("v", repl)))
    return ops


def exclusion_removal_ops(workbook_path: str) -> list[tuple]:
    """Clear (values only, grid intact) any Runsheet row whose doc type is a
    removal-excluded instrument (mortgage/lien/UCC/ROW/easement/subordination/partial
    release). rawdata retains them for awareness; only the title tabs are scrubbed."""
    import re
    from ..config import REMOVAL_EXCLUDED_RX
    wb = openpyxl.load_workbook(workbook_path)
    ops: list[tuple] = []
    if "Runsheet" not in wb.sheetnames:
        return ops
    ws = wb["Runsheet"]
    for r in range(2, ws.max_row + 1):
        typ = ws.cell(r, 3).value
        if isinstance(typ, str) and re.search(REMOVAL_EXCLUDED_RX, typ):
            for c in range(1, 19):
                if ws.cell(r, c).value is not None:
                    ops.append(("Runsheet", f"{get_column_letter(c)}{r}", ("clear",)))
    return ops


def all_ops(workbook_path: str) -> list[tuple]:
    return (net_acre_ops(workbook_path)
            + tbd_convention_ops(workbook_path)
            + exclusion_removal_ops(workbook_path))
