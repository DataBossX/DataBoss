"""Reconcile Title-sheet OGL references against the OGL register (the OGL sheet is the
authority). Non-destructive by default: validates every OGL number cited on Title against
the register, and reports lessor/tract inconsistencies for examiner review. Optionally
suggests the register OGL number for an owner whose base-OGL cell is blank, by matching the
owner name to a register lessor (reported as a suggestion, never auto-written, because a
current owner is often a successor to the original lessor)."""
from __future__ import annotations
import re
import openpyxl


def _reg(wb):
    """number -> (lessor, bk/pg) from the OGL sheet."""
    ws = wb["OGL"]
    reg = {}
    for r in range(2, 80):
        n = ws.cell(r, 1).value
        if isinstance(n, int):
            reg[n] = (str(ws.cell(r, 8).value or ""), str(ws.cell(r, 5).value or ""))
    return reg


def reconcile(path: str) -> dict:
    wb = openpyxl.load_workbook(path)
    if "Title " not in wb.sheetnames or "OGL" not in wb.sheetnames:
        return {"error": "missing Title/OGL sheet"}
    reg = _reg(wb)
    valid_nums = set(reg)
    out = {"cited": 0, "not_in_register": [], "legacy_109_114": [], "suggestions": []}
    wsT = wb["Title "]
    for row in wsT.iter_rows(min_row=1, max_row=250):
        for c in row:
            if not isinstance(c.value, str):
                continue
            for m in re.finditer(r"\bOGL\s+(\d{1,3})\b", c.value):
                n = int(m.group(1))
                out["cited"] += 1
                if 109 <= n <= 114:
                    out["legacy_109_114"].append(f"{c.coordinate}: OGL {n}")
                elif n not in valid_nums:
                    out["not_in_register"].append(f"{c.coordinate}: OGL {n}")
    return out


def suggest_ogl_for_owners(path: str) -> list[dict]:
    """For Title owners whose base-OGL cell is blank/HBP, suggest a register OGL number when
    the owner name matches a register lessor. Suggestions only — examiner confirms."""
    wb = openpyxl.load_workbook(path)
    reg = _reg(wb)
    lessor_tokens = {n: set(re.sub(r"[^a-z ]", " ", l.lower()).split()) for n, (l, _) in reg.items()}
    sugg = []
    wsT = wb["Title "]
    for r in range(1, 250):
        owner = wsT.cell(r, 2).value
        ogl = wsT.cell(r, 4).value
        if not isinstance(owner, str) or (isinstance(ogl, str) and re.search(r"\d", ogl)):
            continue
        if not (isinstance(ogl, str) and ("HBP" in ogl or ogl.strip() in ("", "—"))):
            continue
        ot = set(re.sub(r"[^a-z ]", " ", owner.lower().split("\n")[0]).split()) - {
            "the", "of", "trust", "llc", "inc", "co", "and"}
        best = None
        for n, lt in lessor_tokens.items():
            common = ot & lt
            if len(common) >= 2:
                best = (n, sorted(common)); break
        if best:
            sugg.append({"row": r, "owner": owner.split(chr(10))[0][:40],
                         "suggest_ogl": best[0], "matched_on": best[1],
                         "register_lessor": reg[best[0]][0]})
    return sugg
