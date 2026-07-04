"""Tournament selection: score candidate NHE workbooks on a compliance + completeness
rubric and pick the strongest base to finish. Used to choose "the best report" out of a
folder (or several folders) of competing versions before the finishing pass runs.

Scoring is deterministic and explainable — every candidate gets a per-criterion breakdown
so the choice is auditable, not a black box.
"""
from __future__ import annotations
import os, glob, re
from dataclasses import dataclass, field
import openpyxl
from .config import REQUIRED_SHEETS, REMOVAL_EXCLUDED_RX, YELLOW

# criterion -> max points (higher total = better base)
WEIGHTS = {
    "structure_19": 20,      # exactly the 19 required sheets, in order
    "tracts_foot": 20,       # each tract net acres == gross
    "section_total": 8,      # ties to the section gross (637.42 for 31-12N-24W)
    "no_tbd": 8,             # no bare "TBD"
    "no_formula_err": 8,     # no #REF!/#VALUE!/… literals
    "owner_identified": 16,  # share of net acres credited to a named owner
    "few_gaps": 8,           # fewer unresolved source-in highlights
    "no_excluded": 6,        # no mortgage/lien/etc. on Runsheet
    "ogl_numeric": 6,        # Title OGL refs numeric & in register
}


@dataclass
class Score:
    path: str
    total: float = 0.0
    breakdown: dict = field(default_factory=dict)
    valid: bool = True
    notes: list = field(default_factory=list)


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def score_workbook(path: str, section_gross: float = 637.42) -> Score:
    s = Score(path=path)
    try:
        wb = openpyxl.load_workbook(path)
        wbv = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        s.valid = False; s.notes.append(f"unreadable: {e}"); return s
    b = s.breakdown

    b["structure_19"] = WEIGHTS["structure_19"] if wb.sheetnames == REQUIRED_SHEETS else 0
    if b["structure_19"] == 0:
        s.notes.append(f"{len(wb.sheetnames)} sheets (need the 19 NHE sheets in order)")

    # tract footing + section total + owner-identified share
    footed = 0; gross_sum = 0.0; ident_sum = 0.0
    for i in range(1, 11):
        t = f"Tract {i}"
        if t not in wbv.sheetnames:
            continue
        d5 = _num(wbv[t]["D5"].value)
        cs = sum(_num(c[0].value) or 0 for c in wbv[t].iter_rows(min_row=10, max_row=203, min_col=3, max_col=3)
                 if _num(c[0].value))
        if d5:
            gross_sum += d5
            if abs(cs - d5) <= 0.02:
                footed += 1
    b["tracts_foot"] = round(WEIGHTS["tracts_foot"] * footed / 10, 1)
    b["section_total"] = WEIGHTS["section_total"] if abs(gross_sum - section_gross) < 1.0 else 0

    # owner-identified share: sum of Title identified net acres vs section gross
    if "Title " in wbv.sheetnames:
        idv = 0.0
        for row in wbv["Title "].iter_rows(min_col=2, max_col=3):
            name, net = row[0].value, _num(row[1].value)
            if isinstance(name, str) and net and not re.search(r"(?i)open|undetermined|balance|total|owner", name):
                idv += net
        ident_sum = idv
    b["owner_identified"] = round(WEIGHTS["owner_identified"] * min(ident_sum / section_gross, 1.0), 1)
    s.notes.append(f"~{ident_sum:.0f}/{section_gross:.0f} net ac owner-identified; {footed}/10 tracts foot")

    tbd = any(isinstance(c.value, str) and c.value.strip().upper() == "TBD"
              for ws in wbv.worksheets for row in ws.iter_rows() for c in row)
    b["no_tbd"] = 0 if tbd else WEIGHTS["no_tbd"]

    err = any(isinstance(c.value, str) and re.fullmatch(r"#(REF|VALUE|NAME|DIV/0|NUM|NULL|N/A)!?\??", c.value)
              for ws in wbv.worksheets for row in ws.iter_rows() for c in row)
    b["no_formula_err"] = 0 if err else WEIGHTS["no_formula_err"]

    # gaps: fewer yellow highlights on tract/runsheet is better
    gaps = 0
    for ws in wb.worksheets:
        if not (ws.title.startswith("Tract") or ws.title == "Runsheet"):
            continue
        for row in ws.iter_rows():
            for c in row:
                f = c.fill
                if f and f.patternType == "solid" and f.fgColor and str(f.fgColor.rgb) == "FF" + YELLOW:
                    gaps += 1
    b["few_gaps"] = round(WEIGHTS["few_gaps"] / (1 + gaps / 20), 1)
    s.notes.append(f"{gaps} source-in gap highlights")

    excluded = 0
    if "Runsheet" in wb.sheetnames:
        for row in wb["Runsheet"].iter_rows(min_col=3, max_col=3):
            if isinstance(row[0].value, str) and re.search(REMOVAL_EXCLUDED_RX, row[0].value):
                excluded += 1
    b["no_excluded"] = WEIGHTS["no_excluded"] if excluded == 0 else 0
    if excluded:
        s.notes.append(f"{excluded} excluded-type rows on Runsheet")

    # Title OGL discipline: references numeric and within register
    ogl_ok = True
    if "Title " in wb.sheetnames and "OGL" in wb.sheetnames:
        reg = {wb["OGL"].cell(r, 1).value for r in range(2, 80)
               if isinstance(wb["OGL"].cell(r, 1).value, int)}
        for row in wb["Title "].iter_rows(min_row=1, max_row=200):
            for c in row:
                if isinstance(c.value, str) and re.search(r"\bOGL\s+(10[9]|11[0-4])\b", c.value):
                    ogl_ok = False
    b["ogl_numeric"] = WEIGHTS["ogl_numeric"] if ogl_ok else 0

    s.total = round(sum(b.values()), 1)
    s.valid = b["structure_19"] > 0
    return s


def find_candidates(folder: str) -> list[str]:
    """Candidate NHE cursory workbooks in a folder (recursively), excluding temp/backup."""
    hits = []
    for p in glob.glob(os.path.join(folder, "**", "*.xlsx"), recursive=True):
        base = os.path.basename(p).lower()
        if base.startswith("~$") or "backup" in base or base.startswith("bak_"):
            continue
        hits.append(p)
    return hits


def run_tournament(folders: list[str], section_gross: float = 637.42) -> list[Score]:
    """Score every candidate across the folders; return ranked best-first."""
    cands = []
    for f in folders:
        cands += find_candidates(f)
    scores = [score_workbook(p, section_gross) for p in dict.fromkeys(cands)]
    scores.sort(key=lambda s: (s.valid, s.total), reverse=True)
    return scores
