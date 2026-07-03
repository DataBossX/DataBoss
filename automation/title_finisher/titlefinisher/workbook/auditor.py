"""QC auditor — asserts the full lock-set against a baseline. Returns (ok, issues, notes)."""
from __future__ import annotations
import zipfile, re
import openpyxl
from openpyxl.utils import get_column_letter
from ..config import REQUIRED_SHEETS, REMOVAL_EXCLUDED_RX, YELLOW


def _yellows(wb):
    out = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = c.fill
                if f and f.patternType == "solid" and f.fgColor and str(f.fgColor.rgb) == "FF" + YELLOW:
                    out.add(f"{ws.title}!{c.coordinate}")
    return out


def audit(baseline_path: str, out_path: str, expected_unhighlight: set[str] | None = None):
    issues, notes = [], []
    expected_unhighlight = expected_unhighlight or set()

    za, zb = zipfile.ZipFile(baseline_path), zipfile.ZipFile(out_path)
    na, nb = set(za.namelist()), set(zb.namelist())
    if na != nb:
        issues.append(f"zip part set changed: {na ^ nb}")
    else:
        notes.append(f"zip parts identical ({len(na)})")
    changed = [n for n in na & nb if za.read(n) != zb.read(n)]
    allowed = re.compile(r"xl/(styles\.xml|workbook\.xml|worksheets/sheet\d+\.xml)$")
    bad = [n for n in changed if not allowed.match(n)]
    if bad:
        issues.append(f"unexpected changed parts: {bad}")
    else:
        notes.append(f"only sheet/styles/workbook XML changed ({len(changed)}); media/comments intact")

    wa = openpyxl.load_workbook(baseline_path)
    wo = openpyxl.load_workbook(out_path)
    if wo.sheetnames != REQUIRED_SHEETS:
        issues.append(f"sheet names/order != required 19: {wo.sheetnames}")
    else:
        notes.append("19-sheet structure identical")
    for s in wa.sheetnames:
        if s in wo.sheetnames and set(map(str, wa[s].merged_cells.ranges)) != set(map(str, wo[s].merged_cells.ranges)):
            issues.append(f"merged ranges changed on {s}")

    wov = openpyxl.load_workbook(out_path, data_only=True)
    tbd = [f"{ws.title}!{c.coordinate}" for ws in wov.worksheets for row in ws.iter_rows()
           for c in row if isinstance(c.value, str) and c.value.strip().upper() == "TBD"]
    if tbd:
        issues.append(f"bare TBD remains: {tbd[:10]}")
    else:
        notes.append("no bare-TBD cells")

    total = 0.0
    for i in range(1, 11):
        t = f"Tract {i}"
        if t not in wov.sheetnames:
            continue
        d5 = wov[t]["D5"].value
        cs = sum(c[0].value for c in wov[t].iter_rows(min_row=10, max_row=203, min_col=3, max_col=3)
                 if isinstance(c[0].value, (int, float)))
        if isinstance(d5, (int, float)):
            total += d5
            if abs(cs - d5) > 0.02:
                issues.append(f"{t}: net {cs:.3f} != gross {d5}")
    if abs(total - 637.42) > 0.5:
        notes.append(f"tract gross total = {total} (verify vs section)")

    errs = [f"{ws.title}!{c.coordinate}={c.value}" for ws in wov.worksheets for row in ws.iter_rows()
            for c in row if isinstance(c.value, str) and re.fullmatch(r"#(REF|VALUE|NAME|DIV/0|NUM|NULL|N/A)!?\??", c.value)]
    if errs:
        issues.append(f"formula error literals: {errs[:10]}")
    else:
        notes.append("no formula error literals")

    for sheet in ("Runsheet", *[f"Tract {i}" for i in range(1, 11)], "Title "):
        if sheet not in wo.sheetnames:
            continue
        ws = wo[sheet]
        col = 3 if sheet == "Runsheet" else None
        for row in ws.iter_rows():
            for c in row:
                if col and c.column != col:
                    continue
                if isinstance(c.value, str) and re.search(REMOVAL_EXCLUDED_RX, c.value) and sheet == "Runsheet":
                    issues.append(f"{sheet}!{c.coordinate} excluded-type: {c.value[:40]}")

    ya, yo = _yellows(wa), _yellows(wo)
    unexpected_removed = (ya - yo) - expected_unhighlight
    if unexpected_removed or (yo - ya):
        issues.append(f"yellow drift: unexpected removed={unexpected_removed} added={yo - ya}")
    else:
        notes.append(f"highlights: {len(expected_unhighlight)} resolved-removed; {len(yo)} preserved")

    return (len(issues) == 0), issues, notes
