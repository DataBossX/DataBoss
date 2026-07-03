"""Read a cursory-title workbook and auto-build the open-item worklist.

No hand-maintained list: every open item is discovered from the sheet itself —
the yellow source-in gaps, the `VERIFY — NEED int.` instrument columns, the
`Verify base HBP` cells, and the Well 1 TBDs.
"""
from __future__ import annotations
import re
from dataclasses import dataclass, asdict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from ..config import YELLOW


@dataclass
class GapItem:
    kind: str          # 'source_in_gap' | 'need_interest' | 'hbp' | 'well' | 'tbd'
    sheet: str
    coord: str
    party: str = ""
    instrument: str = ""
    doc_type: str = ""
    book_page: str = ""
    detail: str = ""
    populated: bool = False


def _is_yellow(cell) -> bool:
    if isinstance(cell, MergedCell):
        return False
    f = cell.fill
    return bool(f and f.patternType == "solid" and f.fgColor
               and str(f.fgColor.rgb) == "FF" + YELLOW)


def build_worklist(path: str) -> list[GapItem]:
    wb = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    items: list[GapItem] = []

    # 1) yellow source-in gaps (Runsheet col G grantor, Tract col D owner)
    if "Runsheet" in wb.sheetnames:
        ws = wb["Runsheet"]
        for row in ws.iter_rows():
            for c in row:
                if c.column == 7 and _is_yellow(c):
                    r = c.row
                    items.append(GapItem(
                        "source_in_gap", "Runsheet", c.coordinate,
                        party=str(ws.cell(r, 7).value or "").split("\n")[0],
                        instrument=str(ws.cell(r, 1).value or ""),
                        doc_type=str(ws.cell(r, 3).value or ""),
                        book_page=str(ws.cell(r, 4).value or "")))
    for i in range(1, 11):
        t = f"Tract {i}"
        if t not in wb.sheetnames:
            continue
        ws = wb[t]
        for row in ws.iter_rows():
            for c in row:
                if c.column == 4 and _is_yellow(c):
                    items.append(GapItem(
                        "source_in_gap", t, c.coordinate,
                        party=str(ws.cell(c.row, 4).value or "").split("\n")[0]))
        # 2) VERIFY — NEED int. instrument columns (row 5 header)
        wsv = wbv[t]
        for ci in range(7, ws.max_column + 1):
            h = ws.cell(5, ci).value
            if isinstance(h, str) and "NEED" in h.upper():
                col = get_column_letter(ci)
                nz = any(isinstance(wsv.cell(r, ci).value, (int, float))
                         and wsv.cell(r, ci).value for r in range(10, 204))
                items.append(GapItem(
                    "need_interest", t, f"{col}5",
                    instrument=str(ws.cell(2, ci).value or ""),
                    doc_type=str(ws.cell(1, ci).value or ""),
                    populated=nz,
                    detail="confirm fraction against image" if nz else "place conveyed interest"))

    # 3) HBP verify cells on Title
    if "Title " in wb.sheetnames:
        ws = wb["Title "]
        for row in ws.iter_rows(min_col=7, max_col=7):
            c = row[0]
            if isinstance(c.value, str) and re.search(r"(?i)verify base hbp|verify.*hbp", c.value):
                items.append(GapItem("hbp", "Title ", c.coordinate,
                                     party=str(ws.cell(c.row, 2).value or "").split("\n")[0],
                                     detail=c.value[:120]))

    # 4) Well 1 TBD / verify-w-OCC cells
    if "Well 1" in wb.sheetnames:
        ws = wb["Well 1"]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and re.search(r"(?i)\bTBD\b|verify w/ occ|form 1002a|not in occ gis", c.value):
                    items.append(GapItem("well", "Well 1", c.coordinate, detail=c.value[:120]))

    return items


def worklist_summary(items: list[GapItem]) -> dict:
    out: dict[str, int] = {}
    for it in items:
        out[it.kind] = out.get(it.kind, 0) + 1
    return out
