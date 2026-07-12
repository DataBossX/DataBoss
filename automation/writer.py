"""Safe, atomic, concurrency-guarded workbook writer.

Design guarantees:

* External or AI-derived cell values are written as literal text. A value that
  begins with a spreadsheet formula trigger (``= + - @``) or a control
  character (tab/CR/LF) is neutralized so a viewer never executes it, unless the
  exact value is explicitly allowlisted as an intended formula.
* Writes are serialized with an OS lock that survives process death (stale lock
  recovery) and replaced atomically via a temp file + ``os.replace``.
* ``.xlsm`` macro projects are preserved (``keep_vba``).
* Digitally signed OOXML packages and unsupported package types are refused
  (fail closed) because an openpyxl round-trip cannot preserve their signatures.
"""

from __future__ import annotations

import os
import shutil
import tempfile
import time
import zipfile
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

FIELDS = [
    "Verified Address",
    "Status",
    "Last Affecting Doc No",
    "Last Doc Type",
    "Last Doc Date",
    "Source URL",
    "Notes",
    "Confidence%",
]

SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
_FORMULA_TRIGGERS = ("=", "+", "-", "@")
_CONTROL_PREFIXES = ("\t", "\r", "\n")


class WorkbookSecurityError(ValueError):
    """Raised when a workbook cannot be edited safely (fail closed)."""


def sanitize_cell_value(value, *, allowlisted_formulas: Iterable[str] = ()):
    """Neutralize spreadsheet-injection payloads in a single value.

    Non-string values pass through unchanged. A string that begins with a
    formula trigger or control character is prefixed with an apostrophe so it is
    stored and displayed as literal text, unless it exactly matches an
    explicitly allowlisted intended formula.
    """
    if not isinstance(value, str):
        return value
    if value in set(allowlisted_formulas):
        return value
    if value and (value[0] in _FORMULA_TRIGGERS or value[0] in _CONTROL_PREFIXES):
        return "'" + value
    return value


def _assert_editable_package(path: Path) -> None:
    """Refuse signed or unsupported OOXML packages (fail closed)."""
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise WorkbookSecurityError(
            f"unsupported workbook type {path.suffix!r}; expected one of "
            f"{sorted(SUPPORTED_SUFFIXES)}"
        )
    try:
        with zipfile.ZipFile(path) as package:
            names = package.namelist()
    except zipfile.BadZipFile as exc:
        raise WorkbookSecurityError(f"not a valid OOXML package: {path}") from exc
    signed = [name for name in names if name.startswith("_xmlsignatures/")]
    if signed:
        raise WorkbookSecurityError(
            "refusing to edit a digitally signed workbook: an openpyxl round-trip "
            f"would strip its signature ({path})"
        )


def create_staging_copy(path: str | Path, staging_dir: str | Path = "data/staging") -> Path:
    """Create a timestamped working copy; the supplied workbook remains untouched."""
    source = Path(path)
    destination_dir = Path(staging_dir)
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / (
        f"{source.stem}_staging_{datetime.now():%Y%m%d_%H%M%S_%f}{source.suffix}"
    )
    shutil.copy2(source, destination)
    return destination


@contextmanager
def _workbook_lock(path: Path, timeout: float = 30.0):
    """Serialize writes with an OS lock that is released if the process dies."""
    lock_path = path.with_name(f".{path.name}.lock")
    deadline = time.monotonic() + timeout
    lock_file = lock_path.open("a+b")
    if os.name == "nt":
        lock_file.seek(0, os.SEEK_END)
        if lock_file.tell() == 0:
            lock_file.write(b"\0")
            lock_file.flush()
    acquired = False
    while not acquired:
        try:
            if os.name == "nt":
                import msvcrt

                lock_file.seek(0)
                msvcrt.locking(lock_file.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl

                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            acquired = True
        except OSError:
            if time.monotonic() >= deadline:
                lock_file.close()
                raise TimeoutError(f"timed out waiting for workbook lock: {lock_path}")
            time.sleep(0.1)
    try:
        yield
    finally:
        if os.name == "nt":
            lock_file.seek(0)
            msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
        else:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
        lock_file.close()


def update_workbook(
    path: str | Path,
    sheet: str,
    owner: str,
    row_data: dict,
    *,
    allowlisted_formulas: Iterable[str] = (),
) -> bool:
    """Update a complete workbook atomically without dropping concurrent edits."""
    path = Path(path)
    _assert_editable_package(path)
    with _workbook_lock(path):
        return _update_workbook_unlocked(
            path, sheet, owner, row_data, allowlisted_formulas
        )


def _update_workbook_unlocked(
    path: Path, sheet: str, owner: str, row_data: dict, allowlisted_formulas: Iterable[str]
) -> bool:
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(path, keep_vba=keep_vba)
    temporary = None
    try:
        if sheet not in workbook.sheetnames:
            raise KeyError(f"worksheet not found: {sheet}")
        worksheet = workbook[sheet]
        headers = {
            str(cell.value).strip(): cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }
        missing = {"Name (Owner)", *FIELDS} - headers.keys()
        if missing:
            raise ValueError(f"missing required columns in {sheet}: {sorted(missing)}")

        owner_key = owner.strip().upper()
        owner_column = headers["Name (Owner)"]
        matching_rows = []
        for row in range(2, worksheet.max_row + 1):
            workbook_owner = str(worksheet.cell(row, owner_column).value or "")
            if workbook_owner.strip().upper() == owner_key:
                matching_rows.append(row)
        if not matching_rows:
            return False

        values = {
            "Verified Address": row_data.get("verified_address"),
            "Status": row_data.get("status"),
            "Last Affecting Doc No": row_data.get("doc_no"),
            "Last Doc Type": row_data.get("instrument"),
            "Last Doc Date": row_data.get("recording_date"),
            "Source URL": row_data.get("source_url"),
            "Notes": row_data.get("notes"),
            "Confidence%": int(round((row_data.get("confidence") or 0) * 100)),
        }
        for row in matching_rows:
            for field, value in values.items():
                safe_value = sanitize_cell_value(
                    value, allowlisted_formulas=allowlisted_formulas
                )
                worksheet.cell(row, headers[field], safe_value)

        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{path.stem}-", suffix=path.suffix, dir=path.parent
        )
        os.close(descriptor)
        temporary = Path(temporary_name)
        workbook.save(temporary)
        workbook.close()
        load_workbook(temporary, read_only=True, keep_vba=keep_vba).close()
        os.replace(temporary, path)
        return True
    finally:
        workbook.close()
        if temporary is not None:
            temporary.unlink(missing_ok=True)
