"""Landman run sheet generator — JSON + Excel output."""

import json
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any

from loguru import logger

REPORTS_DIR = Path(os.getenv("REPORTS_OUTPUT_DIR", "./reports_output"))


def _safe_excel():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        return openpyxl, Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None, None, None, None, None, None


def generate_report(job: Any) -> tuple[dict, str]:
    """
    Generate a landman run sheet for the completed job.

    Returns:
        (report_json dict, path to saved Excel file)
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    instruments = job.instruments or []
    chain = job.chain

    # Sort instruments chronologically
    sorted_instr = sorted(
        instruments,
        key=lambda i: (i.recording_date or "0000-00-00"),
    )

    # Build run sheet rows
    run_sheet = []
    for idx, instr in enumerate(sorted_instr, 1):
        run_sheet.append({
            "no": idx,
            "doc_no": instr.doc_no or "—",
            "instrument_type": instr.instrument_type or "Unknown",
            "recording_date": instr.recording_date or "—",
            "instrument_date": instr.instrument_date or "—",
            "grantors": "; ".join(instr.grantors) if instr.grantors else "—",
            "grantees": "; ".join(instr.grantees) if instr.grantees else "—",
            "interest_conveyed": instr.interest_conveyed or "—",
            "legal_description": (instr.legal_description or "—")[:200],
            "section": instr.section or "—",
            "township": instr.township or "—",
            "range": instr.range_val or "—",
            "county": instr.county or job.county or "—",
            "state": instr.state or job.state or "—",
            "book": instr.book or "—",
            "page": instr.page or "—",
            "lease_terms": instr.lease_royalty_terms or "—",
            "reservations": instr.reservations or "—",
            "notes": instr.notes or "—",
            "confidence": instr.confidence_score,
            "needs_review": instr.needs_review,
            "source_file": Path(instr.source_file).name if instr.source_file else "—",
        })

    # Build report JSON
    report_json: dict = {
        "report_version": "1.0",
        "generated_at": datetime.utcnow().isoformat(),
        "generated_date": str(date.today()),
        "job_id": job.id,
        "source": job.source,
        "tract_description": job.tract_description or "Not specified",
        "owner_name": job.owner_name or "Not specified",
        "county": job.county or "Not specified",
        "state": job.state or "Not specified",
        "document_count": len(sorted_instr),
        "run_sheet": run_sheet,
        "chain_of_title": chain.to_dict() if chain else None,
        "summary": {
            "status": chain.status if chain else "Unknown—needs manual",
            "confidence": chain.confidence if chain else 0.0,
            "current_owners": chain.current_owners if chain else [],
            "encumbrances": chain.encumbrances if chain else [],
            "gaps": chain.gaps if chain else [],
            "curative_needed": chain.curative_needed if chain else [],
            "explanation": chain.explanation if chain else "",
            "title_opinion_notes": chain.title_opinion_notes if chain else None,
        },
        "cost_summary": {
            "total_cost": round(job.total_cost, 4),
            "stage_breakdown": [e.to_dict() for e in job.stage_log if e.cost > 0],
        },
        "needs_review_count": sum(1 for i in instruments if i.needs_review),
        "low_confidence_count": sum(1 for i in instruments if i.confidence_score < 0.7),
    }

    # Save JSON
    json_path = REPORTS_DIR / f"report_{job.id}.json"
    json_path.write_text(json.dumps(report_json, indent=2))

    # Save Excel
    excel_path = _write_excel(report_json, job)

    return report_json, str(excel_path)


def _write_excel(report: dict, job: Any) -> Path:
    openpyxl, Font, PatternFill, Alignment, Border, Side = _safe_excel()
    excel_path = REPORTS_DIR / f"report_{job.id}.xlsx"

    if not openpyxl:
        logger.warning("openpyxl not installed — skipping Excel export")
        # Write a placeholder JSON file with .xlsx name note
        excel_path = REPORTS_DIR / f"report_{job.id}_no_excel.json"
        excel_path.write_text(json.dumps({"note": "Install openpyxl for Excel export"}, indent=2))
        return excel_path

    wb = openpyxl.Workbook()

    # ── Sheet 1: Run Sheet ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Run Sheet"

    # Title block
    ws.merge_cells("A1:T1")
    ws["A1"] = "LANDMAN RUN SHEET"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:T2")
    ws["A2"] = (
        f"Tract: {report['tract_description']}  |  "
        f"County: {report['county']}, {report['state']}  |  "
        f"Prepared: {report['generated_date']}  |  "
        f"Source: {report['source'].upper()}"
    )
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:T3")
    summary = report["summary"]
    ws["A3"] = (
        f"Status: {summary['status']}  |  "
        f"Confidence: {summary['confidence']:.0%}  |  "
        f"Documents: {report['document_count']}"
    )
    ws["A3"].font = Font(bold=True)
    ws["A3"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = [
        "No.", "Doc No.", "Instrument Type", "Rec. Date", "Inst. Date",
        "Grantors", "Grantees", "Interest Conveyed", "Sec", "Twp", "Rng",
        "County", "State", "Book", "Page", "Lease Terms",
        "Reservations", "Notes", "Conf.", "Review?"
    ]
    header_row = 5
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    low_conf_fill = PatternFill("solid", fgColor="FCE4D6")

    for row_data in report["run_sheet"]:
        r = header_row + row_data["no"]
        needs_review = row_data["needs_review"]
        low_conf = row_data["confidence"] < 0.7

        row_vals = [
            row_data["no"], row_data["doc_no"], row_data["instrument_type"],
            row_data["recording_date"], row_data["instrument_date"],
            row_data["grantors"], row_data["grantees"], row_data["interest_conveyed"],
            row_data["section"], row_data["township"], row_data["range"],
            row_data["county"], row_data["state"], row_data["book"], row_data["page"],
            row_data["lease_terms"], row_data["reservations"], row_data["notes"],
            f"{row_data['confidence']:.0%}", "YES" if needs_review else "",
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if needs_review:
                cell.fill = review_fill
            elif low_conf:
                cell.fill = low_conf_fill

    # Column widths
    col_widths = [5, 12, 22, 12, 12, 30, 30, 22, 6, 8, 8, 12, 6, 8, 8, 25, 25, 25, 7, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    # ── Sheet 2: Chain Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Chain Summary")
    chain = report.get("chain_of_title") or {}
    summary = report["summary"]

    ws2["A1"] = "CHAIN OF TITLE SUMMARY"
    ws2["A1"].font = Font(bold=True, size=13)

    rows = [
        ("Status", summary["status"]),
        ("Confidence", f"{summary['confidence']:.0%}"),
        ("", ""),
        ("Current Owners", ""),
    ]
    for owner in summary.get("current_owners", []):
        rows.append(("", f"  {owner.get('name', '—')} — {owner.get('interest_type', '—')} {owner.get('fraction', '') or ''}"))

    rows += [
        ("", ""),
        ("Encumbrances", ""),
    ]
    for enc in summary.get("encumbrances", []):
        rows.append(("", f"  {enc}"))

    rows += [
        ("", ""),
        ("Gaps in Chain", ""),
    ]
    for gap in summary.get("gaps", []):
        rows.append(("", f"  {gap}"))

    rows += [
        ("", ""),
        ("Curative Needed", ""),
    ]
    for cur in summary.get("curative_needed", []):
        rows.append(("", f"  {cur}"))

    rows += [
        ("", ""),
        ("Explanation", summary.get("explanation", "—")),
        ("", ""),
        ("Title Opinion Notes", summary.get("title_opinion_notes") or "None"),
        ("", ""),
        ("Total Pipeline Cost", f"${report['cost_summary']['total_cost']:.4f}"),
    ]

    for r, (label, value) in enumerate(rows, 3):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=bool(label))
        cell = ws2.cell(row=r, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 70

    wb.save(excel_path)
    logger.info(f"Excel report saved: {excel_path}")
    return excel_path
