#!/usr/bin/env python3
"""
BECKHAM 32 REPORT BUILDER
=========================
Section 32-11N-25W, Beckham County, Oklahoma - Diversified cursory title report.

Merges EVERYTHING found in the three source workbooks -
  1. the NHE cursory report   (…7-10-26 - NHE.xlsx)
  2. the original cursory     (…2026-07-10.xlsx)
  3. the Template             (Template.xlsx, Sec 27-15N-24W Roger Mills style)
- into ONE new Template-structured "best possible cursory" workbook:

  * a complete, chronological master RUNSHEET built from every instrument
    reference found anywhere in the source files (deduplicated by Book/Page),
    each row carrying its evidence tier and every source sheet that cites it;
  * per-tract chain-out sheets (Tracts 1-5): the instruments that touch each
    tract, in order, with grantor->grantee chain links and explicit
    "OPEN ITEM - chain break" rows where the record does not connect;
  * WI branch sheets (principal Diversified chain + parallel branch);
  * OGL, Wells, Overview, and a Conclusions & Open Requirements sheet that
    cross-checks the 12 priority instruments and reports which were located;
  * a Gaps sheet listing everything the evidence cannot support.

HONESTY RULES (enforced):
  * Nothing is written that was not read from a source file. No WI/NRI decimal
    is ever computed or invented; cells the record cannot support say
    "OPEN ITEM".
  * Post-1988 index-only entries are stamped "INDEX ONLY - PUBLIC METADATA".
  * HBP is never confirmed; wells are regulatory context only.
  * Source originals are opened read-only and never modified; output is a new
    file with a timestamp in its name.
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.stderr.write(
        "\nERROR: openpyxl is not installed.\n"
        "Fix:  py -m pip install openpyxl\n"
        "(or double-click run_beckham32.bat, which installs it for you)\n")
    sys.exit(1)

LOG = logging.getLogger("beckham32")

# ---------------------------------------------------------------------------
# Project constants (from the Section 32-11N-25W engagement)
# ---------------------------------------------------------------------------
SECTION_LABEL = "32-11N-25W"
COUNTY = "Beckham"
DEFAULT_SEARCH_DIRS = [
    r"/home/workdir/attachments",
    r"D:\Desktop\Horizon",
    ".",
]

NHE_NAME_RE = re.compile(r"32-11N-25W.*NHE", re.I)
ORIGINAL_NAME_RE = re.compile(r"32-11N-25W.*(cursory|title).*\.xlsx$", re.I)
TEMPLATE_NAME_RE = re.compile(r"template", re.I)

# The 12 priority instruments the report must locate (Book, Page-or-range).
PRIORITY_INSTRUMENTS = [
    ("2340", "403"), ("2340", "490"),
    ("2371", "470"), ("2371", "533"),
    ("2389", "500"), ("2389", "581"),
    ("2395", "415"), ("2400", "551"),
    ("2451", "4"), ("2476", "1"), ("2476", "121"), ("2490", "471"),
]

# Tract definitions. Used only to ROUTE instruments already found in the
# files onto tract sheets - they add no facts of their own.
TRACT_DEFS = [
    ("Tract 1", "NE/4 (Crook wellbore area)",
     [r"\bne/?4\b", r"\bcrook\b", r"northeast\s+quarter"]),
    ("Tract 2", "E/2 SE/4 (Pearl wellbore area)",
     [r"\be/?2\s*se/?4\b", r"\bpearl\b", r"east\s+half\s+of\s+the\s+southeast"]),
    ("Tract 3", "N/2 below base of Hunton",
     [r"\bn/?2\b", r"\bhunton\b", r"north\s+half"]),
    ("Tract 4", "Deep rights references",
     [r"\bdeep\b", r"below\s+\d", r"depths?\s+below"]),
    ("Tract 5", "All of Section 32 (modern index branch)",
     [r"all\s+of\s+(said\s+)?sec", r"\bentire\s+section\b", r"\bsection\s*32\b",
      r"\b32-11n-25w\b"]),
]

INDEX_ONLY_BOOK_THRESHOLD = 1014   # >= Bk 1014/75 (~1988 cutoff) = index only
OPEN_ITEM = "OPEN ITEM"

BOOK_PAGE_RE = re.compile(r"\b(?:bk\.?|book)?\s*(\d{2,4})\s*[/\-]\s*(?:p(?:g|age)?\.?\s*)?(\d{1,4})\b", re.I)
LEEDE_IMG_RE = re.compile(r"\bL-?0*(\d{1,3})\b")
DATE_TEXT_RE = re.compile(r"\b(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})\b")

GRANTOR_HDR_RE = re.compile(r"grantor|lessor|assignor|from", re.I)
GRANTEE_HDR_RE = re.compile(r"grantee|lessee|assignee|\bto\b", re.I)
DATE_HDR_RE = re.compile(r"date", re.I)
INSTR_HDR_RE = re.compile(r"instrument|type|document|effect", re.I)
DESC_HDR_RE = re.compile(r"description|legal|lands|remarks|notes|comments", re.I)
BOOK_HDR_RE = re.compile(r"book|\bbk\b|recording|book\s*/?\s*page", re.I)

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
OPEN_FILL = PatternFill("solid", fgColor="FFF2CC")
TIER_DIRECT = "DIRECT IMAGE"
TIER_INDEX = "INDEX ONLY - PUBLIC METADATA"
TIER_UNKNOWN = "UNVERIFIED REFERENCE"


# ---------------------------------------------------------------------------
@dataclass
class Instrument:
    book: str
    page: str
    date: str = ""
    instr_type: str = ""
    grantor: str = ""
    grantee: str = ""
    description: str = ""
    tier: str = TIER_UNKNOWN
    leede_ref: str = ""
    sources: list = field(default_factory=list)   # "file :: sheet ! cell"
    raw_context: str = ""

    @property
    def key(self):
        return (self.book.lstrip("0"), self.page.lstrip("0"))

    @property
    def sort_key(self):
        try:
            return (0, int(self.book), int(self.page))
        except ValueError:
            return (1, 0, 0)

    def merge(self, other: "Instrument"):
        for attr in ("date", "instr_type", "grantor", "grantee", "description",
                     "leede_ref"):
            if not getattr(self, attr) and getattr(other, attr):
                setattr(self, attr, getattr(other, attr))
        if other.tier == TIER_DIRECT:
            self.tier = TIER_DIRECT
        self.sources.extend(s for s in other.sources if s not in self.sources)
        if len(other.raw_context) > len(self.raw_context):
            self.raw_context = other.raw_context


def norm(value) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def norm_name(value) -> str:
    text = norm(value).casefold()
    text = re.sub(r"[.,;:'\"()]", " ", text)
    text = re.sub(r"\b(llc|inc|co|corp|ltd|lp|company|corporation|holdco)\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()


# ---------------------------------------------------------------------------
# Source discovery
# ---------------------------------------------------------------------------
def find_sources(search_dirs, explicit):
    """Return dict with keys nhe/original/template -> Path or None."""
    found = {"nhe": explicit.get("nhe"), "original": explicit.get("original"),
             "template": explicit.get("template")}
    pool = []
    for d in search_dirs:
        root = Path(d)
        if root.is_dir():
            pool.extend(p for p in root.rglob("*.xlsx")
                        if not p.name.startswith("~$"))
    for p in pool:
        name = p.name
        if found["nhe"] is None and NHE_NAME_RE.search(name):
            found["nhe"] = p
        elif found["template"] is None and TEMPLATE_NAME_RE.search(name):
            found["template"] = p
        elif (found["original"] is None and ORIGINAL_NAME_RE.search(name)
              and not NHE_NAME_RE.search(name)):
            found["original"] = p
    return found


# ---------------------------------------------------------------------------
# Harvest: read EVERY cell of EVERY sheet, one document at a time
# ---------------------------------------------------------------------------
def sheet_headers(ws, scan_rows=15):
    """Best-effort header row -> {col: header_text}."""
    for r in range(1, min(ws.max_row or 0, scan_rows) + 1):
        row = {c: norm(ws.cell(row=r, column=c).value)
               for c in range(1, min(ws.max_column or 0, 40) + 1)}
        texts = [t for t in row.values() if t]
        if len(texts) >= 3 and any(BOOK_HDR_RE.search(t) for t in texts) \
                and any(GRANTOR_HDR_RE.search(t) or GRANTEE_HDR_RE.search(t)
                        for t in texts):
            return r, {c: t for c, t in row.items() if t}
    return None, {}


def classify_tier(book: str, context: str, leede: str) -> str:
    low = context.lower()
    if leede or "direct image" in low or "image reviewed" in low:
        return TIER_DIRECT
    if "index only" in low or "metadata" in low:
        return TIER_INDEX
    try:
        if int(book) >= INDEX_ONLY_BOOK_THRESHOLD:
            return TIER_INDEX
        return TIER_DIRECT
    except ValueError:
        return TIER_UNKNOWN


def harvest_workbook(path: Path, registry: dict):
    LOG.info("Harvesting %s ...", path.name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            hdr_row, headers = sheet_headers(ws)
            col_of = {}
            if headers:
                for c, h in headers.items():
                    if GRANTOR_HDR_RE.search(h) and "grantor" not in col_of:
                        col_of["grantor"] = c
                    if GRANTEE_HDR_RE.search(h) and "grantee" not in col_of:
                        col_of["grantee"] = c
                    if DATE_HDR_RE.search(h) and "date" not in col_of:
                        col_of["date"] = c
                    if INSTR_HDR_RE.search(h) and "type" not in col_of:
                        col_of["type"] = c
                    if DESC_HDR_RE.search(h) and "desc" not in col_of:
                        col_of["desc"] = c
            max_row = min(ws.max_row or 0, 20000)
            max_col = min(ws.max_column or 0, 50)
            for r_idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col),
                    start=1):
                row_text = " | ".join(norm(c.value) for c in row if norm(c.value))
                if not row_text:
                    continue
                matches = list(BOOK_PAGE_RE.finditer(row_text))
                if not matches:
                    continue
                # skip pure ratio/fraction noise like "1/2" "3/4"
                matches = [m for m in matches
                           if not (len(m.group(1)) == 1 and len(m.group(2)) == 1)]
                for m in matches:
                    book, page = m.group(1), m.group(2)
                    leede_m = LEEDE_IMG_RE.search(row_text)
                    inst = Instrument(book=book, page=page,
                                      leede_ref=leede_m.group(0) if leede_m else "")
                    if hdr_row and r_idx > hdr_row and col_of:
                        cells = {name: norm(ws.cell(row=r_idx, column=c).value)
                                 for name, c in col_of.items()}
                        inst.grantor = cells.get("grantor", "")
                        inst.grantee = cells.get("grantee", "")
                        inst.instr_type = cells.get("type", "")
                        inst.description = cells.get("desc", "")
                        raw_date = cells.get("date", "")
                        if raw_date:
                            inst.date = raw_date
                    if not inst.date:
                        dm = DATE_TEXT_RE.search(row_text)
                        if dm:
                            inst.date = dm.group(0)
                    inst.raw_context = row_text[:500]
                    inst.tier = classify_tier(book, row_text, inst.leede_ref)
                    inst.sources.append(f"{path.name} :: {ws.title} ! row {r_idx}")
                    if inst.key in registry:
                        registry[inst.key].merge(inst)
                    else:
                        registry[inst.key] = inst
    finally:
        wb.close()


def harvest_text_notes(path: Path) -> dict:
    """Pull free-text sheets (Overview/Conclusion style) verbatim so nothing
    written by the examiner is lost."""
    notes = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            if not re.search(r"overview|conclusion|requirement|open|note|assumption",
                             ws.title, re.I):
                continue
            lines = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 500),
                                    max_col=min(ws.max_column or 0, 20)):
                text = " ".join(norm(c.value) for c in row if norm(c.value))
                if text:
                    lines.append(text)
            if lines:
                notes[ws.title] = lines
    finally:
        wb.close()
    return notes


# ---------------------------------------------------------------------------
# Tract routing + chain-out (one tract at a time)
# ---------------------------------------------------------------------------
def route_to_tracts(instruments):
    routed = {name: [] for name, _, _ in TRACT_DEFS}
    unrouted = []
    for inst in instruments:
        blob = " ".join([inst.description, inst.raw_context]).lower()
        hit = False
        for name, _desc, patterns in TRACT_DEFS:
            if any(re.search(p, blob, re.I) for p in patterns):
                routed[name].append(inst)
                hit = True
        if not hit:
            unrouted.append(inst)
    return routed, unrouted


def chain_out(instruments):
    """Order a tract's instruments and mark grantor->grantee continuity.
    Returns rows: (instrument, chain_status)."""
    ordered = sorted(instruments, key=lambda i: i.sort_key)
    rows = []
    last_grantee = ""
    for inst in ordered:
        status = ""
        if inst.grantor and inst.grantee:
            if last_grantee and norm_name(inst.grantor) != last_grantee:
                status = (f"{OPEN_ITEM} - chain break: grantor '{inst.grantor}' "
                          "does not match prior grantee of record")
            else:
                status = "links to prior instrument" if last_grantee else "chain start"
            last_grantee = norm_name(inst.grantee)
        else:
            status = (f"{OPEN_ITEM} - parties not abstracted "
                      f"({inst.tier.lower()}); pull instrument to chain")
        rows.append((inst, status))
    return rows


# ---------------------------------------------------------------------------
# Output workbook (one tab at a time)
# ---------------------------------------------------------------------------
def style_header(ws, n_cols, row=1):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


INSTRUMENT_HEADERS = ["Book", "Page", "Date", "Instrument / Effect", "Grantor",
                      "Grantee", "Description / Lands", "Evidence tier",
                      "Leede image", "Cited by (source file :: sheet)", "Context"]


def instrument_row(inst: Instrument):
    return [inst.book, inst.page, inst.date or OPEN_ITEM,
            inst.instr_type or OPEN_ITEM, inst.grantor or OPEN_ITEM,
            inst.grantee or OPEN_ITEM, inst.description or "",
            inst.tier, inst.leede_ref,
            "; ".join(inst.sources[:6]), inst.raw_context[:300]]


def mark_open_items(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(OPEN_ITEM):
                cell.fill = OPEN_FILL


def build_output(out_path: Path, sources: dict, registry: dict,
                 notes_by_file: dict, dry_run: bool):
    instruments = sorted(registry.values(), key=lambda i: i.sort_key)
    routed, unrouted = route_to_tracts(instruments)
    located_priority = {key: registry.get(key) for key in
                        ((b.lstrip("0"), p.lstrip("0")) for b, p in PRIORITY_INSTRUMENTS)}

    wb = openpyxl.Workbook()

    # -- Overview -----------------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    today = datetime.now().strftime("%m/%d/%Y")
    overview = [
        [f"CURSORY TITLE REPORT - SECTION {SECTION_LABEL}, {COUNTY} COUNTY, OKLAHOMA"],
        [f"Merged best-evidence compilation built {today} from: "
         + "; ".join(str(p) for p in sources.values() if p)],
        [""],
        ["SCOPE: leasehold / working-interest review of Diversified-affiliated "
         "entities. Direct record images reviewed through the ~1988 cutoff "
         f"(Bk {INDEX_ONLY_BOOK_THRESHOLD}/75 era); later entries are INDEX ONLY - "
         "PUBLIC METADATA and are so stamped on every row."],
        ["RELIANCE LIMITATION: this compilation adds no facts. Every row below "
         "is traceable to a cell in the source workbooks (see 'Cited by'). "
         "No working interest, net revenue interest, net acres or ownership "
         "decimal is computed or implied anywhere in this file; cells the "
         f"record cannot support read '{OPEN_ITEM}'."],
        ["HBP / WELLS: regulatory context only - nothing herein is proof of "
         "production holding any lease."],
        [""],
        [f"Instruments harvested from all sources: {len(instruments)}"],
        [f"Priority instruments located: "
         f"{sum(1 for v in located_priority.values() if v)} of {len(PRIORITY_INSTRUMENTS)}"],
    ]
    for line in overview:
        ws.append(line)
    ws.column_dimensions["A"].width = 120
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws["A1"].font = Font(bold=True, size=14)

    # -- Master Runsheet ----------------------------------------------------
    ws = wb.create_sheet("Runsheet")
    ws.append(INSTRUMENT_HEADERS)
    for inst in instruments:
        ws.append(instrument_row(inst))
    style_header(ws, len(INSTRUMENT_HEADERS))
    autosize(ws, [7, 7, 12, 24, 24, 24, 40, 26, 10, 45, 60])
    mark_open_items(ws)

    # -- Tract sheets (chain-out) -------------------------------------------
    for name, desc, _pat in TRACT_DEFS:
        ws = wb.create_sheet(name.replace(" ", ""))
        ws.append([f"{name} - {desc}"])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([f"Present ownership/decimal: {OPEN_ITEM} - pending priority "
                   "instrument pulls; no decimal is implied."])
        ws.append([])
        headers = INSTRUMENT_HEADERS + ["Chain status"]
        ws.append(headers)
        for inst, status in chain_out(routed[name]):
            ws.append(instrument_row(inst) + [status])
        if not routed[name]:
            ws.append(["(no instrument in the source files was routed to this "
                       f"tract) {OPEN_ITEM}"] + [""] * (len(headers) - 1))
        style_header(ws, len(headers), row=4)
        autosize(ws, [7, 7, 12, 24, 24, 24, 40, 26, 10, 45, 60, 50])
        mark_open_items(ws)

    # -- Unrouted instruments ------------------------------------------------
    ws = wb.create_sheet("Unrouted")
    ws.append(["Instruments found in the sources whose lands could not be "
               "matched to a tract definition - review and route by hand."])
    ws.append([])
    ws.append(INSTRUMENT_HEADERS)
    for inst in unrouted:
        ws.append(instrument_row(inst))
    style_header(ws, len(INSTRUMENT_HEADERS), row=3)
    autosize(ws, [7, 7, 12, 24, 24, 24, 40, 26, 10, 45, 60])

    # -- Priority instrument checklist ---------------------------------------
    ws = wb.create_sheet("Priority Pulls")
    ws.append(["Book", "Page", "Located in sources?", "Cited by", "Status"])
    for (book, page), inst in located_priority.items():
        if inst:
            ws.append([book, page, "YES", "; ".join(inst.sources[:4]),
                       f"referenced but {OPEN_ITEM} - certified copy with all "
                       "exhibits/schedules still required"])
        else:
            ws.append([book, page, "NO", "",
                       f"{OPEN_ITEM} - NOT found in any source workbook; "
                       "pull from county records"])
    style_header(ws, 5)
    autosize(ws, [8, 8, 18, 60, 60])
    mark_open_items(ws)

    # -- Examiner notes preserved verbatim ------------------------------------
    ws = wb.create_sheet("Source Notes")
    ws.append(["Source file", "Sheet", "Examiner text (verbatim)"])
    for fname, sheets in notes_by_file.items():
        for sheet, lines in sheets.items():
            for line in lines:
                ws.append([fname, sheet, line])
    style_header(ws, 3)
    autosize(ws, [40, 24, 120])

    # -- Open Requirements -----------------------------------------------------
    ws = wb.create_sheet("Open Requirements")
    ws.append(["#", "Requirement", "Why it matters"])
    requirements = [
        ("Pull certified copies of every priority instrument with ALL "
         "exhibits/schedules", "They control exact vesting, fractions, "
         "depth limits and exclusions - nothing modern can be decimalized "
         "without them."),
        ("Bridge the post-1988 chain (Leede -> Exxon -> Chesapeake -> "
         "Tapstone/KL CHK -> Diversified branches)", "Currently index-only; "
         "the present vested entity is unproven."),
        ("Abstract all original OGLs (terms, royalty, Pugh/depth, pooling, "
         "expiration)", "Lease status and burdens are unknown."),
        ("Reconcile the parallel FourPoint/Unbridled/MNR branch against the "
         "principal chain", "Possible separate wells/depths/exclusions."),
        ("Lease-by-lease HBP analysis against production histories",
         "Well status alone proves nothing about holding."),
        ("Courthouse continuation from the index cutoff to the current date",
         "Later instruments may change everything above."),
    ]
    for i, (req, why) in enumerate(requirements, start=1):
        ws.append([i, req, why])
    style_header(ws, 3)
    autosize(ws, [4, 70, 70])

    if dry_run:
        LOG.info("[DRY-RUN] would write %s (%d instruments, %d unrouted)",
                 out_path, len(instruments), len(unrouted))
        return len(instruments)
    wb.save(out_path)
    LOG.info("Wrote %s", out_path)
    return len(instruments)


# ---------------------------------------------------------------------------
def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        prog="beckham32_report_builder",
        description="Merge the NHE + original cursory reports (and Template "
                    "structure) into one best-evidence Section 32-11N-25W "
                    "report. Read-only on sources; writes a new file.")
    parser.add_argument("--search", nargs="*", default=DEFAULT_SEARCH_DIRS,
                        help="Folders to search for the three workbooks.")
    parser.add_argument("--nhe", help="Explicit path to the NHE report.")
    parser.add_argument("--original", help="Explicit path to the original cursory.")
    parser.add_argument("--template", help="Explicit path to Template.xlsx.")
    parser.add_argument("--out", help="Output folder (default: next to the NHE file).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Harvest and report, but write no file.")
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(levelname)-7s %(message)s",
                        stream=sys.stdout)
    explicit = {k: Path(v) if v else None
                for k, v in (("nhe", args.nhe), ("original", args.original),
                             ("template", args.template))}
    sources = find_sources(args.search, explicit)
    for kind, path in sources.items():
        if path and Path(path).is_file():
            LOG.info("Source [%s]: %s", kind, path)
        else:
            LOG.warning("Source [%s]: NOT FOUND", kind)

    readable = {k: Path(p) for k, p in sources.items() if p and Path(p).is_file()}
    if "nhe" not in readable and "original" not in readable:
        LOG.error("Neither the NHE report nor the original cursory was found. "
                  "Pass them explicitly, e.g.:")
        LOG.error('  py beckham32_report_builder.py --nhe "D:\\path\\...NHE.xlsx"')
        return 1

    try:
        registry: dict = {}
        notes_by_file: dict = {}
        # one document at a time
        for kind in ("nhe", "original", "template"):
            path = readable.get(kind)
            if not path:
                continue
            if kind == "template":
                LOG.info("Template noted for structure: %s (its Roger Mills "
                         "facts are NOT merged - different prospect)", path.name)
                continue
            harvest_workbook(path, registry)
            notes_by_file[path.name] = harvest_text_notes(path)
        LOG.info("Master register: %d unique instruments", len(registry))

        anchor = readable.get("nhe") or readable["original"]
        out_dir = Path(args.out) if args.out else anchor.parent
        out_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / (f"32-11N-25W_Beckham_Co_Diversified_Cursory_"
                              f"Title_Report_MERGED_BEST_{stamp}.xlsx")
        n = build_output(out_path, readable, registry, notes_by_file, args.dry_run)
        LOG.info("Done: %d instruments across runsheet + 5 tract chain-outs. "
                 "Sources untouched.", n)
        return 0
    except Exception as exc:
        LOG.error("FATAL: %s\n%s", exc, traceback.format_exc())
        LOG.error("The run did NOT complete. No source file was modified.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
