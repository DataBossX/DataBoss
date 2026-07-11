#!/usr/bin/env python3
"""FINAL_VERIFIED v2 — builds on build_final.py output with:
   - remaining register instruments swept into the Runsheet (incl. separate
     Presidio WAB and BCE-Mach parallel chains, clearly flagged)
   - instrument dates/timestamps patched onto modern rows from the OKCR detail
   - WI 1 / WI 2 grids extended with their branches' lien/release/carveout rows
   - Tract 1 / Tract 4 grids chained out with the verified historic branch
   - Well 1 rebuilt against the full verified OCC order stack (dates + depths)
   Nothing is invented: every row carries its register/OKCR/OCC source."""
import subprocess, sys
subprocess.run([sys.executable, "build_final.py"], check=True, stdout=subprocess.DEVNULL)

import openpyxl
from datetime import date

FIN = f"s32/32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_FINAL_VERIFIED_{date.today().isoformat()}.xlsx"
wb = openpyxl.load_workbook(FIN)

# ------------------------------------------------------------- Runsheet v2 --
ws = wb["Runsheet"]
NC = 11
def last_data_row():
    return max(r for r in range(1, ws.max_row + 1)
               if any(ws.cell(row=r, column=c).value for c in range(3, NC + 1)))

# patch instrument dates / recording timestamps onto existing modern rows
PATCH = {  # book/page fragment -> (col5 execution/instrument date, col6 recording)
    "2371/470": ("Instrument 2021-12-02", "2022-01-12 12:16 PM"),
    "2389/500": ("Instrument 2022-07-27", "2022-08-24 10:24 AM"),
    "2395/415": ("Instrument 2022-10-27", "2022-11-18 8:22 AM"),
    "2400/551": (None, "2023-03-02 1:36 PM"),
    "2451/4": ("Executed 2025-05-07", "2025-05-13 9:20 AM"),
    "2476/1": ("Instrument 2025-11-24", "2026-01-23 12:20 PM"),
    "2476/121": (None, "2026-01-23 12:20 PM"),
    "2167/241": ("Instrument 2014-04-01; I-2014-004882", None),
}
patched = 0
for r in range(2, last_data_row() + 1):
    bp = str(ws.cell(row=r, column=4).value or "")
    for frag, (ex, rec) in PATCH.items():
        if bp.startswith(frag):
            if ex:
                cur = str(ws.cell(row=r, column=5).value or "")
                if ex.split()[-1] not in cur:
                    ws.cell(row=r, column=5).value = (cur + "; " + ex).strip("; ")
            if rec:
                ws.cell(row=r, column=6).value = rec
            patched += 1
print("Runsheet: patched instrument/recording detail on", patched, "rows")

# additional instruments from Diversified Candidates / Entity Map / portfolios
MORE = [
    ("Assignment; I-2014-005089", "2168/35", "Instrument 2014-04-01", "2014-09-02",
     "Fortune Natural Leasing Company", "FourPoint Energy LLC and named EnerVest funds",
     "Section 32 index entry",
     "Companion Fortune-to-FourPoint input. Some transcriptions show page 44 — page discrepancy carried; the record controls.",
     "INDEX ONLY — PUBLIC METADATA", "2014-09-02"),
    ("Mortgage (referenced)", "2341/127", "OPEN", "2020 (referenced)",
     "Tapstone Energy LLC", "Bank of America, N.A.",
     "Referenced collateral", "Mortgage family referenced by the 2371/439 release — pull to confirm released collateral scope.",
     "REFERENCED BY RELEASE — not independently reviewed", "2020-12-15"),
    ("Mortgage (referenced)", "2389/590", "OPEN", "2022 (referenced)",
     "DP Legacy Tapstone LLC and named parties", "KeyBank National Association",
     "Referenced collateral", "Mortgage referenced by the 2395/370 release — pull to confirm full release.",
     "REFERENCED BY RELEASE — not independently reviewed", "2022-08-25"),
    ("Financing Statement (referenced)", "2389/630", "OPEN", "2022 (referenced)",
     "Tapstone Energy LLC; Diversified Production LLC", "KeyBank National Association",
     "Referenced filing", "Financing statement referenced by terminations 2395/967-972 and 2395/973.",
     "REFERENCED BY TERMINATION — not independently reviewed", "2022-08-26"),
    ("Release continuation pages; I-2025-002299/A/B", "2453/389-2455/700", "OPEN", "2025-06-11",
     "JPMorgan Chase Bank, N.A.", "Unbridled Resources LLC and indexed parties",
     "Section 32 index entries",
     "Release family runs across continuation pages 2453/389 through 2455/700; companion termination runs 2455/705 through 2457/936. Separate from the MNR/UMB lien family.",
     "INDEX ONLY — PUBLIC METADATA", "2025-06-11"),
    # Parallel chains — flagged separate, never blended
    ("Separate-chain records (Presidio WAB)", "2307/566; 2307/583; 2359/304; 2360/336", "OPEN", "2019-2021",
     "Apache/Presidio WAB parties per tract index", "Presidio WAB LLC / BMO lien parties",
     "Section 32 tract index",
     "SEPARATE CHAIN: 2019 Apache acquisition and BMO mortgage/2021 terminations. No record ties Presidio WAB to the drilled Carlson wells or any Diversified branch — do not blend.",
     "INDEX ONLY — PUBLIC METADATA; separate portfolio", "2019-10-01"),
    ("Separate-chain records (BCE-Mach)", "2317/613; 2377/765; 2416/49; 2424/175-530; 2446/1-232", "OPEN", "2019-10-15 through 2025-03-05",
     "Everest / B&W / Eclipse / BCE-Mach parties and lenders", "BCE-Mach II LLC and lenders",
     "Section 32 index entries",
     "SEPARATE CHAIN: parallel BCE-Mach operator/title and financing cycle (operator of JGL, Hanks Crook, LeGrand, Twin wells). Do not blend with Diversified absent a recorded bridge.",
     "INDEX ONLY — PUBLIC METADATA; separate portfolio", "2019-10-15"),
]
r = last_data_row()
import re
def sort_date(row_idx):
    for col in (6, 5):
        v = str(ws.cell(row=row_idx, column=col).value or "")
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", v)
        if m:
            return m.group(0)
    return "9999"
# append then re-sort whole data block chronologically (stable)
start = 2
rows = []
for rr in range(start, r + 1):
    rows.append([ws.cell(row=rr, column=c).value for c in range(1, NC + 1)])
for m in MORE:
    rows.append([None, None, m[0], m[1], m[2], m[3], m[4], m[5], m[6], m[7], m[8]])
def keyof(row):
    for v in (row[5], row[4], row[10] if len(row) > 10 else None):
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(v or ""))
        if m:
            return m.group(0)
    return None
last_key = "1900-01-01"
keyed = []
for row in rows:
    k = keyof(row[:10]) or (row[10] if len(row) > 10 and isinstance(row[10], str) and re.match(r"\d{4}-", str(row[10])) else None) or last_key
    if k < last_key:
        k = last_key
    last_key = k
    keyed.append((k, row))
keyed.sort(key=lambda t: t[0])
for i, (_, row) in enumerate(keyed, start=2):
    ws.cell(row=i, column=1).value = i - 1
    for c in range(2, NC + 1):
        ws.cell(row=i, column=c).value = row[c - 1] if c - 1 < len(row) else None
print(f"Runsheet v2: {len(keyed)} data rows")

# --------------------------------------------------- WI / Tract grid helper --
def find_panel_row(sheet, label="Interest / Category"):
    for rr in range(1, sheet.max_row + 1):
        if str(sheet.cell(row=rr, column=1).value or "").strip().startswith(label):
            return rr
    return None

def insert_grid_rows(sheet, rows, before_label="Interest / Category"):
    pos = find_panel_row(sheet, before_label)
    if pos is None:
        pos = sheet.max_row + 1
    sheet.insert_rows(pos, amount=len(rows))
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            sheet.cell(row=pos + i, column=c).value = v
    return pos

# --------------------------------------------------------------- WI 1 -------
wi1 = wb["WI 1"]
wi1_rows = [
    ("— DIRECT-BRANCH LIEN / RELEASE CYCLE (verified index metadata) —", "", "", "", "", "", "", ""),
    ("Release of Mortgage", "INDEX ONLY", "2371/439", "rec. 2022-01-12",
     "Bank of America, N.A.", "Tapstone Energy, LLC", "Releases 2341/127 family; collateral scope OPEN", "95% metadata"),
    ("Release of Mortgage", "INDEX ONLY", "2395/370", "rec. 2022-11-18",
     "KeyBank N.A.", "DP Legacy Tapstone LLC", "Refers to 2389/590 mortgage; full-release proof OPEN", "95% metadata"),
    ("Terminations", "INDEX ONLY", "2395/967-972; 2395/973", "rec. 2022-11-28",
     "KeyBank N.A.", "Tapstone Energy LLC; Diversified Production LLC; DP Legacy Tapstone LLC",
     "Terminate 2389/630 financing statement; NOT conveyances", "95% metadata"),
    ("— CANVAS / DP PONIES BRANCH (indexed 2026; schedules OPEN) —", "", "", "", "", "", "", ""),
    ("Merger", "PUBLIC METADATA", "2475/943-959", "rec. 2026-01-23 12:20 PM",
     "Canvas Energy LLC", "Canvas Energy II LLC", "17-page filing controls", "95% metadata"),
    ("Conveyance", "PUBLIC METADATA", "2476/1-66", "inst. 2025-11-24 / rec. 2026-01-23",
     "Canvas Energy II LLC", "DP Ponies LLC", "66-page schedule/exclusions control; estate/fraction OPEN", "95% metadata / 0% decimal"),
    ("Mortgage", "PUBLIC METADATA", "2476/67-108", "inst. 2025-11-24 / rec. 2026-01-23",
     "DP Ponies LLC", "UMB Bank, N.A.", "No release found — apparently outstanding; verify", "95% metadata"),
    ("Financing Statement", "PUBLIC METADATA", "2476/109-120", "rec. 2026-01-23",
     "DP Ponies LLC", "UMB Bank, N.A.", "No termination found — apparently active; verify", "95% metadata"),
    ("Merger filing", "PUBLIC METADATA", "2480/903-909", "rec. 2026-03-11",
     "Diversified Gas and Oil Corp.; Canvas and DP entities", "Public", "No legal indexed; title effect requires filing", "90% metadata"),
    ("Affidavit", "PUBLIC METADATA", "2487/28-31", "rec. 2026-05-27",
     "Diversified Gas and Oil Corp.; Canvas/Corsair entities", "Public", "Post-cutoff continuation; not a conveyance", "90% metadata"),
]
pos = insert_grid_rows(wi1, wi1_rows)
print("WI 1: inserted", len(wi1_rows), "rows at", pos)

# --------------------------------------------------------------- WI 2 -------
wi2 = wb["WI 2"]
wi2_rows = [
    ("Assignments", "INDEX ONLY", "2167/241; 2168/35", "inst. 2014-04-01 / rec. 2014-08-22 & 09-02",
     "Fortune Natural Resources Corp.; Fortune Natural Leasing Co.", "FourPoint Energy LLC and named EnerVest funds",
     "Historic branch inputs; page-44 transcription discrepancy carried", "90% metadata"),
    ("Assignment (CARVEOUT)", "INDEX ONLY", "2401/214", "rec. 2023-03-14",
     "Unbridled Resources LLC", "Teocalli Exploration LLC",
     "CRITICAL potential carveout — leases/wells/depths/reservations unknown", "95% metadata / 0% scope"),
    ("Mortgage", "INDEX ONLY", "2415/410", "rec. 2023-11-14",
     "MNR ABS Issuer I LLC", "UMB Bank, N.A.", "No release found — apparently outstanding; verify", "95% metadata"),
    ("Companion Mortgage / Agreement", "INDEX ONLY", "2415/515; 2415/617", "rec. 2023-11-14",
     "Unbridled Resources LLC and indexed parties", "MNR ABS Issuer I LLC and indexed parties",
     "Party list / collateral schedule require image", "92% metadata"),
    ("Merger filing", "PUBLIC METADATA", "2451/4-23", "exec. 2025-05-07 / rec. 2025-05-13",
     "Diversified Energy Company PLC", "Public",
     "County filing tied to Diversified/MNR/Unbridled (Maverick closed 3/14/2025 per official evidence); corporate filing alone transfers no unscheduled title", "99% metadata / 20% effect"),
    ("Release + Termination", "INDEX ONLY", "2453/389-2455/700; 2455/705-2457/936", "rec. 2025-06-11",
     "JPMorgan Chase Bank, N.A.", "Unbridled Resources LLC and indexed parties",
     "Clears separate 2022 JPM family only — does NOT release the MNR/UMB liens", "95% metadata"),
]
pos = insert_grid_rows(wi2, wi2_rows)
print("WI 2: inserted", len(wi2_rows), "rows at", pos)

# ------------------------------------------------------------- Tract 1 ------
t1 = wb["Tract 1"]
t1_rows = [
    ("— CROOK UNIT POOLING + PARALLEL BRANCHES (verified this session) —", "", "", "", "", "", "", ""),
    ("Pooling Order (verified official PDF)", "OFFICIAL OCC", "Order 156126 / Cause CD 59995", "exec. 1979-08-06 (hearing 1979-05-22)",
     "Corporation Commission of Oklahoma", "All Section 32 owners",
     "Pools Lower Permian/Virgil/UGW/LGW/Atoka/Morrow-Springer/Mississippian (640 ac) + Hunton unit incl. Secs 29-31 (1,440 ac); elections/JOA No. 1580 OPEN", "100% order / 0% elections"),
    ("Assignment (Hunt branch; Crook schedule)", "DIRECT IMAGE 2275-2309", "678/777-811", "eff. 1983-09-01 / rec. 1983-10-14",
     "Hunt Petroleum Corporation", "Petro-Hunt Corporation 46%; Rosewood Resources (HPC), Inc. 12%",
     "Crook #1-32 schedule cites lease 66032-21-1, JOA 1580 and Order 156126; schedule-stated NRI .09996 / WI .12305 need second visual check", "High parties / Medium decimals"),
    ("Release of Security Instruments", "DIRECT IMAGE 3673-3680", "792/51-58", "1985-01-31 / rec. 1985-02-15",
     "Seattle-First National Bank; Ralph L. Hawkins, Jr., Trustee", "ENI Joint Venture 1979 XII (body) / XIII (heading)",
     "Releases 611/154 only as to scheduled ENI Crook-unit properties; does NOT reach lease 66032-21-1 burdens", "High; XII/XIII conflict carried"),
    ("Assignment, Bill of Sale and Conveyance", "DIRECT IMAGE 3682-3711", "792/72-101", "eff. 1984-05-01 / rec. 1985-02-15",
     "113 ENI Schedule I entities (roster in working workbook)", "Quinoco Petroleum, Inc.",
     "5% scheduled leasehold + Order 156126 rights + 4.0144 Garrett interest; vintage ENI 'Diversified' program names are a name collision only", "High"),
    ("Operator chain (regulatory only)", "OFFICIAL OCC FORMS", "API 35-009-20312", "1986-2025",
     "Mesa → Mesa Operating → Samson → [Vastar step unresolved] → BP", "Latigo → Complete Oil & Gas Services LLC (Form 1073MW appr. 12/3/2025)",
     "Current RBDMS AC/DRY; open wellbore is NOT production, HBP or county title", "High operator chain / 0% title"),
]
pos = insert_grid_rows(t1, t1_rows)
print("Tract 1: inserted", len(t1_rows), "rows at", pos)

# ------------------------------------------------------------- Tract 4 ------
t4 = wb["Tract 4"]
t4_rows = [
    ("— HISTORIC LEASE 66032-21-1 NEEDLE CHAIN (N/2 SW/4 + N/2 NW/4; direct images) —", "", "", "", "", "", "", ""),
    ("Oil and Gas Lease (root; executed copy absent)", "RECITED imgs 1568/1812/2031/2300", "341/538", "1977-11-29",
     "R. L. Minton and Ella Minton, his wife", "John E. Hartman",
     "Creates lease 66032-21-1; royalty/term/Pugh/depth/pooling clauses unreviewed", "High identity / 0% terms"),
    ("Assignment", "DIRECT IMAGE 1568", "376/287", "1979-04-05 / rec. 1979-04-20",
     "John E. Hartman", "Hunt Energy Corporation",
     "All right, title and interest; reserves ORRI printed '.0125% of 8/8s' (literal 0.000125 — do not normalize); no release found", "High"),
    ("Assignment", "DIRECT IMAGE 1681-1688", "407/676-683", "1980-03-25 / rec. 1980-04-14",
     "Hunt Energy Corporation", "Nine Hunt entities (22/2/2/21/1/19/12/9/11%); Hunt Energy retains 1%",
     "99% allocation; schedule prints 1/14/1978 lease date vs 11/29/1977 — conflict carried", "High"),
    ("Assignment", "DIRECT IMAGE 1799-1812", "430/416-429", "1980-08-18 (eff. 1979-04-01) / rec. 1980-08-28",
     "Ten Hunt-group owners", "Great American Exploration Corporation",
     "25% of each assignor's interest; conditional 50% recovery reservation — not a clean 25% WI. Duplicate re-record at 440/269-282 (not double-counted)", "High"),
    ("Assignment", "DIRECT IMAGE 1813-1819", "431/674-680", "1980-09-10 / rec. 1980-09-30",
     "Great American Exploration Corporation", "Great American Drilling Partnership III",
     "All GAEC interest; GAEC reserves formula ORRI (25% of 8/8 less royalty/burdens, proportionately reduced); no release found", "High"),
    ("Five Mortgages + amendments/restatements", "DIRECT IMAGE 1696-1793; 1871-2103", "423/238-335; 470/29-68; 518/5-44; 568/121-191", "1980-09-04 through 1982-03-20",
     "GAEC; GAE-III; DAE-II Partnership", "First National Bank in Dallas",
     "Tranches 73.9874/13.3531/8.6709/3.9886%; no final release in supplied set; Bk 425M amendments absent — historic lien requirement", "High"),
    ("Assignment of OGLs", "DIRECT IMAGE 2010-2016", "546/68-74", "1981-12-23 (eff. 1981-06-30)",
     "Great American Drilling Partnership III", "DAE-II Partnership",
     "15.7533% of the GAEC-derived branch (a fraction of a fraction, NOT of 8/8)", "High"),
    ("Assignment and Conveyance", "DIRECT IMAGE 2024-2032", "568/26-34", "1982-04-27 (GC eff. 1/1/1982)",
     "Hassie Hunt Trust, Margaret Hunt Hill, Trustee", "Hassie Hunt Production Company",
     "2% Hunt-allocation branch, subject to earlier 25%-of-branch GAEC transfer", "High"),
    ("Current status", "SEE RUNSHEET / OGL #12", "—", "—", "—", "—",
     "Present ownership/decimal OPEN ITEM — trace successors, releases, elections and executed lease before any math", "—"),
]
pos = insert_grid_rows(t4, t4_rows)
print("Tract 4: inserted", len(t4_rows), "rows at", pos)

# ------------------------------------------------------------- Well 1 v2 ----
w = wb["Well 1"]
hdrs = {c: str(w.cell(row=2, column=c).value or "") for c in range(1, 18)}
mu = next((c for c, h in hdrs.items() if "MU Order" in h), 12)
spacing = next((c for c, h in hdrs.items() if "Spacing" in h), 11)
wells = {str(w.cell(row=r, column=2).value): r for r in range(3, 17)}
ORDER_STACK = ("Verified formation-specific stack: 67733 (12/15/1967: Virgil 6,450-7,150 ft; "
               "UGW 10,380-12,295 ft; LGW 12,295-13,880 ft), 612937 (6/19/2013: UGW=Missourian, "
               "LGW=Des Moines), 145162 (9/19/1978: Mississippi + Morrow-Springer; Hunton relief denied), "
               "147147 (11/21/1978: Lower Permian incl. Chase/Brown Dolomite + Atoka), "
               "295220 (3/27/1986: Upper Permian)")
def setw(api, col, val):
    r = wells.get(api)
    if r:
        w.cell(row=r, column=col).value = val
for api, extra in {
    "35-009-20312": "Order 475430 (5/5/2003, additional Atoka well); pooling Order 156126 (exec. 8/6/1979); " + ORDER_STACK,
    "35-009-21085": "Order 410493 (3/24/1997, additional Morrow-Springer well; Crook member ~19,180-19,198 ft); " + ORDER_STACK,
    "35-009-21072": "Order 147147 (Lower Permian + Atoka; JGL ITD ref to Order 129891 REJECTED — covers T10N lands); " + ORDER_STACK,
    "35-009-21391": "Orders 486531 (2/24/2004, additional Atoka well) + 487644 (3/19/2004 location exception); " + ORDER_STACK,
    "35-009-20738": "LeGrand-family Orders 454332/454333 (7/19/2001, incr. density Atoka + Morrow-Springer excl. Crook member); " + ORDER_STACK,
    "35-009-21236": "LeGrand-family Orders 454332/454333; " + ORDER_STACK,
    "35-009-21258": "LeGrand #2-32 Orders 463052/463053 (4/26/2002, incr. density Morrow-Springer excl. Crook member); " + ORDER_STACK,
    "35-009-21893": "Carlson Orders 613064 (6/24/2013 interim horiz. location, UGW/Missourian) and related 624342/625127/636221/625525/625773; " + ORDER_STACK,
    "35-009-21923": "Orders 624342 (4/23/2014 incr. density Missourian) / 625127 / 636221 (7H passed within 72 ft of 12H, >200 ft vertical); " + ORDER_STACK,
    "35-009-21933": "Orders 625525 (5/16/2014 incr. density Missourian) / 625773 (5/21/2014 location exception; corrects erroneous 147147 refs to 67733); " + ORDER_STACK,
}.items():
    setw(api, mu, extra)
for api in wells:
    r = wells[api]
    if not w.cell(row=r, column=spacing).value or str(w.cell(row=r, column=spacing).value) == "OPEN":
        w.cell(row=r, column=spacing).value = "640-ac formation-specific vertical units (Hunton: 1,440-ac unit incl. parts of Secs 29-31)"
print("Well 1 v2: full dated order stack applied to", len(wells), "well rows")

wb.save(FIN)
print("SAVED v2:", FIN)
