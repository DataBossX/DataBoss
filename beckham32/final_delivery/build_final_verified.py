#!/usr/bin/env python3
"""Build 32-11N-25W FINAL_VERIFIED workbook from the clean NHE base.

Every added row is transcribed from the project's visually-verified evidence
register (Working Report Workbook) or the official OCC order PDF read this
session. No decimal, party, or date is invented; unsupported cells say OPEN.
"""
import re
import shutil
from datetime import date, datetime

import openpyxl

SRC = "s32/NHE.xlsx"
OUT = f"s32/32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_FINAL_VERIFIED_{date.today().isoformat()}.xlsx"
shutil.copy2(SRC, OUT)
wb = openpyxl.load_workbook(OUT)

# ---------------------------------------------------------------- Runsheet --
# (doc_type, book_page, exec_date, rec_date, grantor, grantee, legal, notes, tier, sortdate)
NEW_ROWS = [
    ("Assignment of OGL", "376/287", "1979-04-05", "1979-04-20",
     "John E. Hartman", "Hunt Energy Corporation",
     "Lease 66032-21-1: N/2 SW/4 + N/2 NW/4; 160 acres",
     "All right, title and interest; assignor reserves ORRI printed “.0125% of 8/8s” (literal decimal 0.000125 — do not normalize), free of development/operating costs except taxes. Underlying lease 341/538 dated 1977-11-29. No release found.",
     "DIRECT IMAGE 1568 — visually verified", "1979-04-20"),
    ("Pooling Order (OCC)", "OCC Order 156126; Cause CD 59995", "Executed 1979-08-06 (hearing 1979-05-22)", "Not county-recorded",
     "Regulatory — Corporation Commission of Oklahoma", "All owners, Section 32 units",
     "All Sec. 32 for Lower Permian (incl. Chase/Brown Dolomite), Virgil Lime, Upper & Lower Granite Wash, Atoka, Morrow-Springer, Mississippian; Hunton unit = S/2 29 + SE/4 30 + E/2 31 + all 32 (1,440 ac)",
     "Leede Exploration pooling for proposed Sec. 32 well. $400/ac bonus or 1/16-of-8/8 elections; 20/27/35-day mechanics; 180-day commencement. Later schedules recite 8/8/1979 — conflict carried; official order date controls. Owner elections/defaults and JOA No. 1580 remain OPEN.",
     "OFFICIAL OCC PDF — verified this run", "1979-08-06"),
    ("Assignment", "407/676-683", "1980-03-25", "1980-04-14",
     "Hunt Energy Corporation",
     "Hassie Hunt Exploration 22%; Hassie Hunt Trust 2%; H.L. Hunt Jr. Trust Estate 2%; Hunt Petroleum 21%; Hunt Industries 1%; Rosewood Corp. 19%; N.B. Hunt Trust Estate 12%; W.H. Hunt Trust Estate 9%; Lamar Hunt Trust Estate 11%; Hunt Energy retains 1%",
     "Lease 66032-21-1",
     "Transfers 99% of Hunt Energy's lease interest; as-is/no warranty. Schedule prints 1/14/1978 lease date, conflicting with 11/29/1977 recited elsewhere — discrepancy carried, not harmonized.",
     "DIRECT IMAGE 1681-1688 — visually verified", "1980-04-14"),
    ("Assignment", "430/416-429", "1980-08-18 (eff. 1979-04-01 7:00 a.m.)", "1980-08-28",
     "Ten Hunt-group owners (see 407/676 allocation)", "Great American Exploration Corporation",
     "Lease 66032-21-1 scheduled at image 1812",
     "Undivided 25% of each assignor's then-owned scheduled lease interest; as-is/no warranty; conditional 50% production/sale-proceeds recovery reservation — NOT a clean absolute 25% WI.",
     "DIRECT IMAGE 1799-1812 — visually verified", "1980-08-28"),
    ("Five Mortgages", "423/238-335", "1980-09-04", "1980-09-10",
     "Great American Exploration Corp.; Great American Drilling Partnership-III", "First National Bank in Dallas",
     "Schedules include lease 66032-21-1",
     "One GAEC mortgage ($2,750,000 note) + four GAE-III tranche mortgages (stated 73.9874%, 13.3531%, 8.6709%, 3.9886%). No release found in supplied set — carried as historic lien requirement, not a current-lien conclusion.",
     "DIRECT IMAGE 1696-1793 — visually verified", "1980-09-10"),
    ("Assignment", "431/674-680", "1980-09-10", "1980-09-30",
     "Great American Exploration Corporation", "Great American Drilling Partnership III",
     "GAEC interest in lease 66032-21-1",
     "All GAEC interest from prior assignment; assignor reserves formula ORRI = 25% of 8/8 less lessor royalty and other burdens, proportionately reduced; as-is/no warranty. No release found.",
     "DIRECT IMAGE 1813-1819 — visually verified", "1980-09-30"),
    ("Duplicate / re-recorded Assignment", "440/269-282", "1980-08-18 (eff. 1979-04-01)", "OPEN — read stamp",
     "Same ten Hunt-group owners", "Great American Exploration Corporation",
     "Same scheduled lease 66032-21-1",
     "Duplicate/re-record of 430/416-429 — corroboration only; MUST NOT be counted as a second 25% transfer.",
     "DIRECT IMAGE 1824-1837 — visually verified", "1980-10-01"),
    ("Five Second Amendments (mortgages)", "470/29-68", "1981-05-12", "OPEN — read stamp",
     "GAEC; GAE-III", "First National Bank in Dallas",
     "Schedules continue lease 66032-21-1",
     "Amend and continue the five 423/238-335 mortgage tranches. Amendments, not releases. Bk 425M/Pgs 672-680 amendments referenced but absent from supplied set — OPEN pull.",
     "DIRECT IMAGE 1871-1910 — visually verified", "1981-05-12"),
    ("Four Third Amendments (mortgages)", "518/5-44", "1981-09-04", "OPEN — read stamp",
     "GAE-III; DAE-II Partnership", "First National Bank in Dallas",
     "Schedules include lease 66032-21-1",
     "Adds DAE-II Partnership as co-mortgagor for its 15.7533% assumed share. Amendments, not releases.",
     "DIRECT IMAGE 1953-1992 — visually verified", "1981-09-05"),
    ("Assignment of OGLs", "546/68-74", "1981-12-23 (eff. 1981-06-30 12:01 a.m.)", "OPEN — read stamp",
     "Great American Drilling Partnership III", "DAE-II Partnership",
     "GAEC-derived Section 32 lease interest (lease 66032-21-1)",
     "Undivided 15.7533% of the interest GAEC conveyed to GAE-III — a fraction of a prior fraction, NOT 15.7533% of 8/8; subject to proportional burdens; no warranty.",
     "DIRECT IMAGE 2010-2016 — visually verified", "1981-12-23"),
    ("Consolidated & Amended Mortgage + Restatements", "568/121-191", "1982-01-11 through 1982-03-20", "OPEN — read stamp",
     "GAE-III; DAE-II Partnership; GAEC", "First National Bank in Dallas",
     "Schedules include lease 66032-21-1",
     "Consolidation, two restatements, and Third Amendment (GAEC note changed to $2,350,000 demand, eff. 2/26/1982). None is a release — lien family carried as requirement.",
     "DIRECT IMAGE 2033-2103 — visually verified", "1982-03-20"),
    ("Assignment and Conveyance", "568/26-34", "1982-04-27 (General Conveyance eff. 1/1/1982)", "OPEN — read stamp",
     "Hassie Hunt Trust, Margaret Hunt Hill, Trustee", "Hassie Hunt Production Company",
     "Lease 66032-21-1 schedule at image 2031",
     "All Trust right, title and interest — the stated 2% Hunt-allocation branch, subject to the earlier 25%-of-branch GAEC transfer and existing burdens.",
     "DIRECT IMAGE 2024-2032 — visually verified", "1982-04-27"),
    ("Assignment, Conveyance and Bill of Sale", "678/777-811", "eff. 1983-09-01 7:00 a.m.", "1983-10-14",
     "Hunt Petroleum Corporation", "Petro-Hunt Corporation (46%); Rosewood Resources (HPC), Inc. (12%)",
     "Crook #1-32 schedule includes lease 66032-21-1; recites Hartman assignment, JOA No. 1580 and OCC Order 156126",
     "Conveys 58% of Hunt Petroleum's then-owned interest (42% retained). Schedule states assignor-row NRI .0999600000 / WI .1230500000 — schedule-stated figures, second visual check required before ANY use in math.",
     "DIRECT IMAGE 2275-2309 — visually verified", "1983-10-14"),
    ("Release of Security Instruments", "792/51-58", "1985-01-31", "1985-02-15",
     "Seattle-First National Bank; Ralph L. Hawkins, Jr., Trustee",
     "ENI Joint Venture 1979 XII (body) / XIII (schedule heading) — conflict carried",
     "Crook #1-32 / NE Mayfield scheduled Leede leases and force-pooling rights; NOT lease 66032-21-1",
     "Releases Bk 611/154 only as to scheduled ENI Crook-unit properties. Schedule does not list the Minton-Hartman lease, so this release does NOT reach the Hartman or Great American burdens.",
     "DIRECT IMAGE 3673-3680 — visually verified", "1985-02-15"),
    ("Assignment, Bill of Sale and Conveyance", "792/72-101", "eff. 1984-05-01 7:00 a.m.; ack. 1985-02-01", "1985-02-15",
     "All 113 Schedule I ENI programs and joint ventures (roster preserved in working workbook)", "Quinoco Petroleum, Inc. (DE)",
     "Crook #1-32 / NE Mayfield: 33 Leede leases, 10 Union leases, Garrett lease interest, OCC Order 156126 rights",
     "5% scheduled leasehold/pooled-rights interest + stated 4.0144 Garrett-lease interest; exhibit states before-payout OI .0495703/NRI .0371778, after-payout OI .0371778/NRI .0278833. Distinct Crook-unit chain; three vintage ENI “Diversified Drilling Program” names are a name collision — NO bridge to modern Diversified.",
     "DIRECT IMAGE 3682-3711 — visually verified", "1985-02-15"),
    # ----------------------------- modern index/metadata rows ---------------
    ("Assignment", "2167/241", "OPEN", "2014-08-22",
     "Fortune Natural Resources Corporation", "FourPoint Energy, LLC",
     "Section 32 index entry", "Input to FourPoint branch; schedule required.",
     "INDEX ONLY — PUBLIC METADATA", "2014-08-22"),
    ("Assignment", "2168/44", "OPEN", "2014-09-02",
     "Fortune Natural Leasing Company", "FourPoint Energy, LLC",
     "Section 32 index entry", "Additional FourPoint input; some transcriptions show a page discrepancy — record controls.",
     "INDEX ONLY — PUBLIC METADATA", "2014-09-02"),
    ("County-record instrument (type pending image); I-2020-001026", "2327/1", "OPEN", "2020-04-03",
     "BP America Production Company", "Ellipse Resources, LLC (styling pending image)",
     "Section 32 tract index", "CRITICAL: county shows BP→Ellipse BEFORE OCC's 5/14/2020 BP-to-Latigo operator transfer; no county-title bridge to Latigo or Diversified is proved.",
     "INDEX ONLY — PUBLIC METADATA; critical disconnect", "2020-04-03"),
    ("Companion instrument; I-2020-001027", "2327/108", "OPEN", "2020-04-03",
     "BP America Production Company", "Ellipse Resources, LLC",
     "Section 32 tract index", "Companion to 2327/1; exact type, parties and schedule require image.",
     "INDEX ONLY — PUBLIC METADATA; critical disconnect", "2020-04-03"),
    ("Mortgage; I-2020-001596", "2328/269", "OPEN", "2020-05-11",
     "Latigo Petroleum, LLC (+ additional grantors pending image)", "Lone Star State Bank",
     "Long Description", "Latigo collateral mortgage in current Carlson producer chain; no full release located — potentially outstanding.",
     "INDEX ONLY — PUBLIC METADATA", "2020-05-11"),
    ("Financing Statement; I-2020-001597", "2328/379", "OPEN", "2020-05-11",
     "Latigo Petroleum, LLC", "Lone Star State Bank",
     "Long Description", "Companion financing statement to 2328/269.",
     "INDEX ONLY — PUBLIC METADATA", "2020-05-11"),
    ("Amended Mortgage; I-2021-004510", "2365/743", "OPEN", "2021-10-12",
     "Latigo Petroleum, LLC", "Lone Star State Bank",
     "Long Description", "Amends the 2020 Latigo mortgage family; no full release located.",
     "INDEX ONLY — PUBLIC METADATA", "2021-10-12"),
    ("Release of Mortgage; I-2022-000151", "2371/439", "OPEN", "2022-01-12",
     "Bank of America, N.A.", "Tapstone Energy, LLC",
     "Section 32 index entry", "Releases the Bk 2341/127 mortgage family; confirm complete released collateral.",
     "INDEX ONLY — PUBLIC METADATA", "2022-01-11"),
    ("Partial Release of Mortgage; I-2022-001728", "2377/635", "OPEN", "2022-04-28",
     "Lone Star State Bank", "Latigo Petroleum, LLC",
     "Long Description", "Partially releases 2328/269 and 2365/743 family; partial release does not establish full satisfaction.",
     "INDEX ONLY — PUBLIC METADATA", "2022-04-28"),
    ("Release of Mortgage; I-2022-004837", "2395/370", "OPEN", "2022-11-18",
     "KeyBank National Association", "DP Legacy Tapstone LLC",
     "Section 32 index entry", "Release in direct branch; confirm referenced mortgage and collateral.",
     "INDEX ONLY — PUBLIC METADATA", "2022-11-17"),
    ("Termination; I-2022-004939", "2395/973", "OPEN", "2022-11-28",
     "KeyBank National Association", "DP Legacy Tapstone LLC",
     "Section 32 index entry", "Companion termination to 2395/967-972; confirm referenced filing. Not a title conveyance.",
     "INDEX ONLY — PUBLIC METADATA", "2022-11-28"),
    ("Assignment; I-2023-000993", "2401/214", "OPEN", "2023-03-14",
     "Unbridled Resources LLC", "Teocalli Exploration LLC",
     "Section 32-indexed assignment",
     "CRITICAL potential carveout ahead of the Maverick/Diversified sequence — exact leases, wells, depths, reservations and retained interests unknown until the schedule is read.",
     "INDEX ONLY — PUBLIC METADATA; critical open", "2023-03-14"),
    ("Mortgage; I-2023-004424", "2415/1", "OPEN", "2023-11-13",
     "Latigo Petroleum, LLC", "Lone Star State Bank",
     "Long Description", "New Latigo mortgage; no final release located through the 5/12/2026 tract cutoff.",
     "INDEX ONLY — PUBLIC METADATA", "2023-11-13"),
    ("Mortgage; I-2023-004457", "2415/410", "OPEN", "2023-11-14",
     "MNR ABS Issuer I LLC", "UMB Bank, N.A.",
     "Section 32 index entry", "MNR ABS collateral lien; NO release found in current search — apparently outstanding; verify. The 2025 JPMorgan releases do NOT reach this family.",
     "INDEX ONLY — PUBLIC METADATA; apparently outstanding", "2023-11-14"),
    ("Companion Mortgage; I-2023-004458", "2415/515", "OPEN", "2023-11-14",
     "Unbridled Resources LLC and indexed parties", "MNR ABS Issuer I LLC and indexed parties",
     "Section 32 index entry", "Companion financing/collateral instrument; exact party list and collateral schedule require image.",
     "INDEX ONLY — PUBLIC METADATA; critical open", "2023-11-14"),
    ("Agreement / Amendment; I-2023-004459", "2415/617", "OPEN", "2023-11-14",
     "Unbridled Resources LLC and indexed parties", "MNR ABS Issuer I LLC and indexed parties",
     "Section 32 index entry", "Related MNR ABS agreement; operative relationship requires full text.",
     "INDEX ONLY — PUBLIC METADATA", "2023-11-14"),
    ("Release of Mortgage; I-2025-002299", "2453/389", "OPEN", "2025-06-11",
     "JPMorgan Chase Bank, N.A.", "Unbridled Resources LLC and indexed parties",
     "Section 32 index entry", "Appears to release the separate 2022 JPMorgan family; does NOT facially release the MNR/UMB family.",
     "INDEX ONLY — PUBLIC METADATA", "2025-06-11"),
    ("Termination; I-2025-002301", "2455/705", "OPEN", "2025-06-11",
     "JPMorgan Chase Bank, N.A.", "Unbridled Resources LLC and indexed parties",
     "Section 32 index entry", "Terminates the related 2022 financing statement; separate from the UMB lien family.",
     "INDEX ONLY — PUBLIC METADATA", "2025-06-11"),
    ("Amended Mortgage; I-2025-004873", "2470/153", "OPEN", "2025-11-21",
     "Latigo Petroleum, LLC", "Prosperity Bank",
     "Long Description", "Amended mortgage in current producer chain; relationship to the prior Lone Star lien requires full text.",
     "INDEX ONLY — PUBLIC METADATA", "2025-11-21"),
    ("Partial Release of Mortgage; I-2025-005515", "2474/292", "OPEN", "2025-12-22",
     "Prosperity Bank", "Latigo Petroleum, LLC",
     "Long Description", "Partial release in the Latigo lien family; no final/full release located through the 5/12/2026 tract cutoff.",
     "INDEX ONLY — PUBLIC METADATA", "2025-12-22"),
    ("Merger; I-2026-000211", "2475/943-959", "OPEN", "2026-01-23",
     "Canvas Energy LLC", "Canvas Energy II LLC",
     "Long Description", "Corporate succession into Canvas Energy II; full 17-page filing controls.",
     "INDEX ONLY — PUBLIC METADATA", "2026-01-23"),
    ("Mortgage; I-2026-000213", "2476/67-108", "2025-11-24", "2026-01-23",
     "DP Ponies LLC", "UMB Bank, N.A.",
     "Long Description", "Canvas-derived collateral lien; NO release found in current search — apparently outstanding; verify.",
     "INDEX ONLY — PUBLIC METADATA; apparently outstanding", "2026-01-23"),
    ("Financing Statement; I-2026-000214", "2476/109-120", "OPEN", "2026-01-23",
     "DP Ponies LLC", "UMB Bank, N.A.",
     "Long Description", "Companion financing statement; no termination found — apparently active; verify.",
     "INDEX ONLY — PUBLIC METADATA; apparently active", "2026-01-23"),
    ("Merger; I-2026-001039", "2480/903-909", "OPEN", "2026-03-11",
     "Diversified Gas and Oil Corp.; Diversified Production LLC; Canvas and DP entities", "Public",
     "No legal indexed", "Corporate filing; no Section 32 legal indexed — title effect requires the filing itself.",
     "INDEX ONLY — PUBLIC METADATA", "2026-03-11"),
    ("Affidavit; I-2026-002717", "2487/28-31", "OPEN", "2026-05-27",
     "Diversified Gas and Oil Corp.; Canvas/Corsair asset entities", "Public",
     "Long Description; Section 32 tract result", "Post-index-cutoff continuation record; an affidavit is not a conveyance.",
     "INDEX ONLY — PUBLIC METADATA (post-cutoff continuation)", "2026-05-27"),
    ("Release; I-2026-003623", "2490/469-470", "OPEN", "2026-07-07 13:32",
     "Jeffrey Harrell; Kevin Harrell", "Public",
     "Section 32", "First of two filings after the visible index cutoff (OKCR continuation closed through 7/9/2026).",
     "INDEX ONLY — PUBLIC METADATA (post-cutoff continuation)", "2026-07-07"),
    ("Warranty Deed; I-2026-003624", "2490/471-473", "OPEN", "2026-07-07 13:37",
     "Abby Colleen Johnson", "Jesse J. Newell, Custodian; Paige Newell; Wade Newell",
     "Section 32; $70,000 consideration", "Current terminal record of the continuation (surface/fee context; not a leasehold instrument on its face — image controls).",
     "INDEX ONLY — PUBLIC METADATA (post-cutoff continuation)", "2026-07-07"),
    # 2011-2012 index-gap fill records (surface/wind context)
    ("Easement; I-2011-001600", "2042/183-197", "OPEN", "2011-02-18",
     "Maddoux, Gerry L Living Trust", "Dempsey Ridge Wind Farm LLC",
     "Exact Section 32 hit", "Index-gap fill (missing logical p.48): surface/wind-energy context; operative terms require image.",
     "INDEX ONLY — gap-fill via OKCR exact-legal search", "2011-02-18"),
    ("Notice; I-2011-002123", "2043/917-919", "OPEN", "2011-03-07",
     "Elk City II Wind LLC", "Maddoux, Gerry L Living Trust",
     "Exact Section 32 hit", "Index-gap fill; party orientation per public row — verify scan if orientation matters.",
     "INDEX ONLY — gap-fill via OKCR exact-legal search", "2011-03-07"),
    ("Warranty Deed; I-2011-004744", "2052/466-469", "OPEN", "2011-06-02",
     "Dempsey Ridge Wind Farm LLC", "Acciona Wind Energy USA LLC",
     "Exact Section 32 hit", "Index-gap fill; deed estate/legal require image.",
     "INDEX ONLY — gap-fill via OKCR exact-legal search", "2011-06-02"),
    ("Terminations; I-2011-004747 / -004748", "2052/476-478; 2052/479-481", "OPEN", "2011-06-02",
     "Dempsey Ridge Wind Farm LLC", "Hanks, Carole Faye; Crook, C B Jr and Ola Mae",
     "Exact Section 32 hits", "Index-gap fill; terminated instruments/scope require images.",
     "INDEX ONLY — gap-fill via OKCR exact-legal search", "2011-06-02"),
    ("Easements; I-2012-000124 / -000125", "2077/190-196; 2077/197-201", "OPEN", "2012-01-06",
     "Maddoux, Gerry L Living Trust; LeGrand, Juanita", "Enbridge G&P LP",
     "Exact Section 32 hits", "Index-gap fill; operative easement terms require images. Long Description records in the 2/12/2011-2/8/2012 bracket remain OPEN.",
     "INDEX ONLY — gap-fill via OKCR exact-legal search", "2012-01-06"),
]

ws = wb["Runsheet"]
NCOLS = 11
existing = []
for r in range(2, 48):
    row = [ws.cell(row=r, column=c).value for c in range(1, NCOLS + 1)]
    existing.append(row)

def parse_date(txt):
    if txt is None:
        return None
    if isinstance(txt, datetime):
        return txt.date()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(txt))
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"(\d{4})", str(txt))
    if m and 1900 <= int(m.group(1)) <= 2026:
        return date(int(m.group(1)), 6, 30)
    return None

# assign monotone sort keys to the curated existing order
keyed = []
last = date(1900, 1, 1)
for row in existing:
    d = parse_date(row[5]) or parse_date(row[4]) or last
    if d < last:
        d = last
    last = d
    keyed.append((d, 0, row))
for nr in NEW_ROWS:
    d = parse_date(nr[9]) or date(2026, 1, 1)
    row = [None, None, nr[0], nr[1], nr[2], nr[3], nr[4], nr[5], nr[6], nr[7], nr[8]]
    keyed.append((d, 1, row))
keyed.sort(key=lambda t: (t[0], t[1]))

for i, (_, _, row) in enumerate(keyed, start=2):
    ws.cell(row=i, column=1).value = i - 1  # renumber Seq
    for c in range(2, NCOLS + 1):
        ws.cell(row=i, column=c).value = row[c - 1]
print(f"Runsheet: {len(existing)} existing + {len(NEW_ROWS)} verified additions = {len(keyed)} rows")

# ------------------------------------------------------------------- OGL ---
og = wb["OGL"]
mirror = [og.cell(row=7, column=c).value for c in range(1, 29)]  # 340/323 row (N/2 NW/4 + N/2 SW/4)
newr = 13
vals = {1: 12, 3: "Recited at imgs 1568 / 1812 / 2031 / 2300", 4: "OGL (executed copy absent)",
        5: "341/538", 6: "1977-11-29", 7: "OPEN",
        8: "R. L. Minton and Ella Minton, his wife", 9: "John E. Hartman",
        10: "N/2 SW/4 + N/2 NW/4",
        11: "Lease 66032-21-1 — root of the historic Crook/Hunt branch. Identity and legal expressly reproduced in later recorded instruments; the executed lease itself is not in the supplied image set.",
        12: "OPEN ITEM", 13: 160, 14: mirror[13] if mirror[13] else "3/4", 15: "OPEN ITEM",
        16: "OPEN ITEM", 17: "OPEN ITEM", 18: "OPEN ITEM", 19: "OPEN ITEM", 20: "OPEN ITEM",
        21: "RECITED — identity high; terms 0% (executed lease absent)"}
for c, v in vals.items():
    og.cell(row=newr, column=c).value = v
print("OGL: appended lease 66032-21-1 root row at 13; Tract col mirrored:", vals[14])

# -------------------------------------------------------------- Title ------
t = wb["Title "]
hdr = [t.cell(row=125, column=c).value for c in range(1, 8)]
print("C-matrix header:", hdr)
t.insert_rows(134, amount=2)
c9 = {2: "C9",
      3: "Official OCC order PDFs verify a formation-specific vertical unit stack for Section 32 (Orders 67733, 612937, 145162, 147147, 295220) and Leede pooling Order 156126 / Cause CD 59995 executed 8/6/1979 (recorded schedules recite 8/8/1979 — conflict carried). Owner elections, payments/defaults and JOA No. 1580 remain OPEN; production/HBP conclusions must stay lease-, formation-, depth- and order-specific.",
      4: "100% for the orders themselves", 5: "OPEN ITEM",
      6: "Verified official OCC PDFs", 7: "Pooling elections and JOA still required"}
c10 = {2: "C10",
       3: "The historic Crook/Hunt branch of lease 66032-21-1 (Minton-Hartman, Bk 341/538) is fully abstracted from direct images: Hartman ORRI at 376/287; Hunt 99%/1% allocation at 407/676; 25%-of-branch GAEC conveyance at 430/416 (duplicate 440/269 — not double-counted); GAEC formula ORRI at 431/674; DAE-II 15.7533%-of-branch at 546/68; Hassie Hunt branch at 568/26; Petro-Hunt/Rosewood 58% split at 678/777; and the unreleased First National Bank in Dallas mortgage family at 423/238-335 with amendments/restatements through 568/191.",
       4: "High (visually verified images)", 5: "Burdens carried — releases OPEN",
       6: "Direct images 1568-2309; evidence packs in project Drive", 7: "Trace releases/successors before any decimal"}
for c, v in c9.items():
    t.cell(row=134, column=c).value = v
for c, v in c10.items():
    t.cell(row=135, column=c).value = v
# find requirement 11 row (shifted +2 from 150 -> 152)
for r in range(148, 158):
    if t.cell(row=r, column=2).value == 11:
        cur = t.cell(row=r, column=4).value or ""
        t.cell(row=r, column=4).value = str(cur).rstrip(".") + \
            ". STATUS UPDATE: OKCR exact-legal continuation completed through 07/09/2026 (145 Section 32 hits; post-cutoff Release 2490/469-470 and WD 2490/471-473, both recorded 07/07/2026). Long Description records and the final-date rerun remain OPEN."
        print(f"Requirement 11 updated at row {r}")
        break

# -------------------------------------------------------------- Well 1 -----
w = wb["Well 1"]
wells = {str(w.cell(row=r, column=2).value): r for r in range(3, 14)}
def setw(api, col, val):
    r = wells.get(api)
    if r:
        w.cell(row=r, column=col).value = val
# col 12 = MU Order, col 15 = Last Production, col 17 = notes/allocation (verify by header)
hdrs = {c: str(w.cell(row=2, column=c).value or "") for c in range(1, 18)}
mu = next((c for c, h in hdrs.items() if "MU Order" in h), 12)
lastp = next((c for c, h in hdrs.items() if "Last Production" in h), 15)
setw("35-009-20312", mu, "OCC Order 475430; pooling Order 156126 (Cause CD 59995, executed 8/6/1979)")
setw("35-009-20312", lastp, "No rolling-12-month OTC rows (PUN 009-066620-0-0000, Active, 640 ac, still lists BP). AC/DRY under Complete Oil & Gas Services LLC; Latigo-to-Complete Form 1073MW approved 12/3/2025 (Crook first of 38 wells). AC proves an open wellbore — NOT production or HBP.")
setw("35-009-21072", mu, "OCC order per verified stack — formation-specific units only")
setw("35-009-21072", lastp, "Gas reported 7/2025-4/2026; 4/2026 = 2,367 MCF (OTC PUN 009-101800-0-0000, reported by NextEra Energy Marketing). Production fact only — HBP not concluded.")
setw("35-009-21085", mu, "OCC Order 410493")
setw("35-009-21085", lastp, "Active PUN returned no rolling-12-month rows on 7/9/2026.")
setw("35-009-21391", mu, "OCC Orders 486531 / 487644")
for api in ("35-009-20738", "35-009-21236", "35-009-21258"):
    setw(api, mu, "LeGrand-family OCC Orders 454332 / 454333 / 463052 / 463053 (well-specific tie required)")
# Append verified Carlson rows (drilled wells surface/BHL evidence; regulatory only)
base = 14
carlson = [
    ("12H", "35-009-21893", "Latigo Petroleum, LLC", "Carlson 32-11-25 12H", "Granite Wash Upper (Missourian nomenclature per Order 612937)",
     "Completed 9/3/2013; perforations 11,524-16,241 ft",
     "Gas through 5/2026; 5/2026 = 2,113 MCF (OTC PUN 009-211663-0-0000, Active, 640 ac). OTC displays API 009-21903 on this PUN — that API belongs to the permit-only Riviera 9H (approved 10/2/2013, expired 4/2/2014); conflict carried. Operator chain per official OCC forms: Linn → EnerVest → BP → Latigo."),
    ("7H", "35-009-21923", "Latigo Petroleum, LLC", "Carlson 32-11-25 7H", "Missourian",
     "First production 9/18/2014; perforations 11,648-16,133 ft",
     "Historic completion facts only — not proof of current production, lease HBP or title."),
    ("10H", "35-009-21933", "Latigo Petroleum, LLC", "Carlson 32-11-25 10H", "Missourian",
     "First production 9/13/2014; perforations 11,839-16,187 ft",
     "Historic completion facts only — not proof of current production, lease HBP or title."),
]
for i, (tag, api, op, name, formation, comp, note) in enumerate(carlson):
    r = base + i
    w.cell(row=r, column=1).value = 12 + i
    w.cell(row=r, column=2).value = api
    w.cell(row=r, column=3).value = op
    w.cell(row=r, column=4).value = name
    w.cell(row=r, column=5).value = formation
    w.cell(row=r, column=6).value = "Section 32, T11N-R25W, Beckham County, OK (drilled Carlson wells; PUN legal = all of Section 32)"
    w.cell(row=r, column=14).value = comp
    w.cell(row=r, column=lastp).value = note
    w.cell(row=r, column=16).value = "AC"
    w.cell(row=r, column=mu).value = "Carlson OCC Orders 613064 / 624342 / 625127 / 636221 / 625525 / 625773"
print("Well 1: verified Crook/JGL/Hanks/Twin/LeGrand facts set; Carlson 12H/7H/10H rows appended (regulatory context only)")

# ------------------------------------------------------------ Overview -----
ov = wb["Overview"]
b54 = ov["B54"].value or ""
ov["B54"].value = str(b54).rstrip() + (
    "  CONTINUATION & INDEX INTEGRITY (verified 07/09-07/11/2026): the OKCR exact-legal search returned 145 Section 32 hits and is closed through 07/09/2026; two post-cutoff filings recorded 07/07/2026 — Release I-2026-003623 (Bk 2490/469-470) and WD I-2026-003624 (Bk 2490/471-473). Both supplied index PDFs omit logical page 48 and triplicate logical page 54; seven exact-legal records recovered for the 2/12/2011-2/8/2012 bracket (Long Description records remain OPEN). The supplied Tax Roll PDF duplicates the assessor parcel map (identical SHA-256) — the actual tax roll is an OPEN pull. Official OCC order stack verified: Orders 67733, 612937, 145162, 147147 and 295220 (formation-specific 640-acre vertical units; separate 1,440-acre Hunton unit) and pooling Order 156126 / Cause CD 59995 executed 8/6/1979."
)
print("Overview B54 SCOPE/CUTOFFS extended")

wb.save(OUT)
print("SAVED:", OUT)
