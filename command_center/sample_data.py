"""Synthetic, clearly-labelled sample data so the command center runs offline.

There are **no real title documents in the cloud checkout** (see PROJECT_STATUS.md)
and Horizon's law is *never fabricate a real balance*. So this module builds an
obviously-synthetic demonstration workbook for ``31-12N-24W`` whose only purpose
is to exercise the swarm end-to-end. It contains an **OGL** sheet and a
**Runsheet** sheet — exactly the two registers :mod:`command_center.engine`
reads — plus a blank cursory-report **template** to write into.

The corpus is crafted to hit every branch of the self-healing loop:

* **Tract 1 — SW/4 (160 ac):** a clean A→B→C chain that ties to 8/8 with no help.
* **Tract 2 — NE/4 (160 ac):** an assignment with a **blank** conveyed interest
  whose runsheet note documents "all right, title and interest". The Mathematician
  flags it undetermined; the Auditor HEALS it (assumed 8/8, highlighted yellow).
* **Tract 3 — SE/4 (80 ac):** an **over-conveyance** (a party conveys 3/4 while
  holding only 1/2). Cannot be healed by assumption → ESCALATED to examiner review.
"""

from __future__ import annotations

from pathlib import Path

SECTION = "31-12N-24W"
COUNTY = "Roger Mills County, Oklahoma"

# (instrument, grantor, grantee, conveyed, legal, doc_type, date, gross_acres)
_OGL_ROWS = [
    # Tract 1 — clean chain, SW/4, 160 ac
    ("2018-0101", "United States of America", "Alice Mineral Co", "1/2",
     "SW/4 Sec 31-12N-24W", "Mineral Deed", "2018-01-05", "160"),
    ("2019-0210", "Alice Mineral Co", "Bravo Royalty LLC", "1/2",
     "SW/4 Sec 31-12N-24W", "Assignment", "2019-02-10", "160"),
    # Tract 2 — healable blank interest, NE/4, 160 ac
    ("2020-0330", "State of Oklahoma", "Delta Energy Partners", "3/4",
     "NE/4 Sec 31-12N-24W", "Oil & Gas Lease", "2020-03-30", "160"),
    ("2021-0415", "Delta Energy Partners", "Echo Operating LLC", "",  # blank → heal
     "NE/4 Sec 31-12N-24W", "Assignment", "2021-04-15", "160"),
    # Tract 3 — over-conveyance, SE/4, 80 ac
    ("2017-0512", "Foxtrot Holdings", "Golf Minerals Inc", "1/2",
     "SE/4 Sec 31-12N-24W", "Mineral Deed", "2017-05-12", "80"),
    ("2022-0620", "Golf Minerals Inc", "Hotel Resources LP", "3/4",  # > held 1/2
     "SE/4 Sec 31-12N-24W", "Assignment", "2022-06-20", "80"),
]

# (instrument, note, legal)
_RUNSHEET_ROWS = [
    ("2018-0101", "Federal mineral patent; grantee takes an undivided 1/2.",
     "SW/4 Sec 31-12N-24W"),
    ("2019-0210", "Assignment of the 1/2 mineral interest acquired under 2018-0101.",
     "SW/4 Sec 31-12N-24W"),
    ("2020-0330", "Lease covers an undivided 3/4 mineral interest.",
     "NE/4 Sec 31-12N-24W"),
    # This note is what lets the Auditor HEAL instrument 2021-0415.
    ("2021-0415", "Assignment conveys all right, title and interest of grantor "
                  "(assume 8/8 of grantor's holding).", "NE/4 Sec 31-12N-24W"),
    ("2017-0512", "Grantor conveys an undivided 1/2 mineral interest.",
     "SE/4 Sec 31-12N-24W"),
    ("2022-0620", "ASSN references a 3/4 interest — exceeds grantor's holding; "
                  "runsheet flags a probable scrivener error.", "SE/4 Sec 31-12N-24W"),
]

OGL_HEADERS = [
    "Instrument_Number", "Grantor", "Grantee", "Conveyed_Interest",
    "Legal_Description", "Doc_Type", "Instrument_Date", "Gross_Acres",
]
RUNSHEET_HEADERS = ["Instrument_Number", "Note", "Legal_Description"]


def write_sample_workbook(path: Path) -> Path:
    """Write the synthetic OGL + Runsheet workbook to ``path``."""
    import openpyxl

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ogl = wb.active
    ogl.title = "OGL"
    ogl.append([f"SYNTHETIC DEMO DATA — {SECTION} {COUNTY} — NOT A REAL TITLE RECORD"])
    ogl.append(OGL_HEADERS)
    for row in _OGL_ROWS:
        ogl.append(list(row))

    run = wb.create_sheet("Runsheet")
    run.append([f"SYNTHETIC DEMO RUNSHEET — {SECTION} — NOT A REAL TITLE RECORD"])
    run.append(RUNSHEET_HEADERS)
    for row in _RUNSHEET_ROWS:
        run.append(list(row))

    wb.save(path)
    wb.close()
    return path


def write_template(path: Path) -> Path:
    """Write a blank cursory-report template workbook (a ``Title Report`` sheet).

    The engine's writer can target this so the operator's real branded template
    can be dropped in later; here it is just a header shell.
    """
    import openpyxl

    from horizon.models import CANONICAL_COLUMNS
    from horizon.report_io import DATA_SHEET_TITLE

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = DATA_SHEET_TITLE
    ws.append([f"{SECTION} — {COUNTY} — Cursory Title Report (TEMPLATE)"])
    ws.append([c.replace("_", " ").title() for c in CANONICAL_COLUMNS])
    wb.save(path)
    wb.close()
    return path


def ensure_sample_workspace(root: Path) -> tuple[Path, Path]:
    """Create ``root`` with a sample workbook + template; return their paths."""
    root = Path(root)
    root.mkdir(parents=True, exist_ok=True)
    wb = write_sample_workbook(root / f"{SECTION}_SAMPLE.xlsx")
    tpl = write_template(root / "template.xlsx")
    return wb, tpl


if __name__ == "__main__":  # pragma: no cover
    import sys

    dest = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("command_center/sample")
    wb, tpl = ensure_sample_workspace(dest)
    print(f"Sample workbook: {wb}")
    print(f"Template:        {tpl}")
