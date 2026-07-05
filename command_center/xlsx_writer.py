"""Write the perfected report to Excel, highlighting ASSUMED data in yellow.

The Cursor meta-prompt requires: *use the local template.xlsx as the final write
destination; highlight assumed data points in yellow using openpyxl.* This module
does exactly that, on top of the canonical Horizon column layout so the output
stays format-aligned with the Roger Mills cursory report.

Colour legend written into the sheet:

* **Yellow**  — an ASSUMED value (the Auditor healed a gap from the runsheet; the
  number is a documented assumption, not a recorded fact). Verify before relying.
* **Amber**   — a row tagged ``Needs Examiner Review`` (a gap the loop could not
  heal). The row's ``remarks`` say why.

Zero-destruction: if ``template`` carries embedded plats (``xl/media``) they are
preserved — only the data sheet is rewritten (delegated to
:func:`horizon.report_io.write_report`).
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional, Tuple

from horizon.models import CANONICAL_COLUMNS, ReportModel

# Hex fills (openpyxl wants ARGB, so the leading FF is opacity).
ASSUMED_FILL = "FFFFF200"   # bright yellow  — assumed data point
REVIEW_FILL = "FFFFC000"    # amber          — needs examiner review
HEADER_FILL = "FF00161F"    # deep space navy (matches the UI)
LEGEND_ASSUMED = "Yellow = ASSUMED (documented assumption, verify before use)"
LEGEND_REVIEW = "Amber = Needs Examiner Review (gap the loop could not heal)"

# The data grid starts on this row in the written sheet (title, legend, header).
_TITLE_ROW = 1
_LEGEND_ROW = 2
_HEADER_ROW = 3
_FIRST_DATA_ROW = 4


def write_highlighted_report(
    report: ReportModel,
    path: Path,
    assumed_cells: Iterable[Tuple[int, str]] = (),
    template: Optional[Path] = None,
) -> Path:
    """Write ``report`` to ``path``; paint assumed cells yellow, review rows amber.

    ``assumed_cells`` is an iterable of ``(row_index, canonical_field_name)`` where
    ``row_index`` indexes ``report.rows`` (0-based). A fresh workbook is built so
    the title/legend/header banner and freeze panes are consistent; when
    ``template`` carries embedded media we fall back to the media-preserving
    writer and then re-open to apply highlighting.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    assumed = set(assumed_cells)
    field_to_col = {name: i + 1 for i, name in enumerate(CANONICAL_COLUMNS)}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Title Report"

    title = f"{report.section} — Roger Mills County — Cursory Title Report"
    ws.cell(row=_TITLE_ROW, column=1, value=title)
    ws.cell(row=_TITLE_ROW, column=1).font = Font(bold=True, size=13, color="FF00D8FF")

    ws.cell(row=_LEGEND_ROW, column=1, value=f"{LEGEND_ASSUMED}    |    {LEGEND_REVIEW}")
    ws.cell(row=_LEGEND_ROW, column=1).font = Font(italic=True, size=9, color="FF808A90")

    header_font = Font(bold=True, color="FF00D8FF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for name, col in field_to_col.items():
        c = ws.cell(row=_HEADER_ROW, column=col, value=name.replace("_", " ").title())
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    assumed_pat = PatternFill("solid", fgColor=ASSUMED_FILL)
    review_pat = PatternFill("solid", fgColor=REVIEW_FILL)

    for r_off, row in enumerate(report.rows):
        excel_row = _FIRST_DATA_ROW + r_off
        review_row = row.needs_review
        for name, col in field_to_col.items():
            value = getattr(row, name, "") or ""
            cell = ws.cell(row=excel_row, column=col, value=value)
            if (r_off, name) in assumed:
                cell.fill = assumed_pat  # ASSUMED wins — it is the sharper warning
            elif review_row:
                cell.fill = review_pat

    ws.freeze_panes = f"A{_FIRST_DATA_ROW}"
    _autosize(ws, field_to_col)
    wb.save(path)
    wb.close()

    # If a media-bearing template was supplied, preserve its plats: rewrite the
    # data through horizon's media-preserving writer, then re-apply highlights.
    if template is not None and _has_media(Path(template)):
        _rewrite_into_template(report, path, assumed, template, field_to_col)

    return path


def summarise_fills(path: Path) -> dict:
    """Return counts of highlighted cells — used by tests and the UI badge."""
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb["Title Report"] if "Title Report" in wb.sheetnames else wb.active
    counts = {"assumed": 0, "review": 0}
    for row in ws.iter_rows(min_row=_FIRST_DATA_ROW):
        for cell in row:
            rgb = getattr(cell.fill.fgColor, "rgb", None)
            if rgb == ASSUMED_FILL:
                counts["assumed"] += 1
            elif rgb == REVIEW_FILL:
                counts["review"] += 1
    wb.close()
    return counts


# --------------------------------------------------------------------------- #
# internals
# --------------------------------------------------------------------------- #
def _autosize(ws, field_to_col) -> None:
    from openpyxl.utils import get_column_letter

    widths = {
        "grantor": 24, "grantee": 24, "legal_description": 26, "remarks": 48,
        "instrument_number": 16, "status": 22,
    }
    for name, col in field_to_col.items():
        ws.column_dimensions[get_column_letter(col)].width = widths.get(name, 14)


def _has_media(path: Path) -> bool:
    import zipfile

    try:
        with zipfile.ZipFile(path) as zf:
            return any(n.startswith("xl/media/") for n in zf.namelist())
    except (zipfile.BadZipFile, OSError):
        return False


def _rewrite_into_template(report, path, assumed, template, field_to_col) -> None:
    """Preserve template plats, then re-apply the two fills to the data sheet."""
    import openpyxl
    from openpyxl.styles import PatternFill

    from horizon.report_io import DATA_SHEET_TITLE, write_report

    write_report(report, path, template=template)  # media-preserving, plain values
    wb = openpyxl.load_workbook(path)
    ws = wb[DATA_SHEET_TITLE] if DATA_SHEET_TITLE in wb.sheetnames else wb.active
    # horizon.report_io lays out: row1 title, row2 headers, row3+ data.
    first_data = 3
    assumed_pat = PatternFill("solid", fgColor=ASSUMED_FILL)
    review_pat = PatternFill("solid", fgColor=REVIEW_FILL)
    for r_off, row in enumerate(report.rows):
        excel_row = first_data + r_off
        review_row = row.needs_review
        for name, col in field_to_col.items():
            if (r_off, name) in assumed:
                ws.cell(row=excel_row, column=col).fill = assumed_pat
            elif review_row:
                ws.cell(row=excel_row, column=col).fill = review_pat
    wb.save(path)
    wb.close()
