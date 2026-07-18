"""Workbook comparison (Moves #23/#67) — diff two workbooks sheet-by-sheet,
cell-by-cell, so "the current workbook silently differs from the prior
version" becomes a report instead of a surprise.

Operates on the same plain shape as ``qa_engine``: sheets are dicts of
sheet name -> list of row dicts. ``load_workbook_sheets`` provides an
optional openpyxl loader for real .xlsx files (values + formulas).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path


@dataclass
class Difference:
    kind: str            # sheet_added | sheet_removed | row_added | row_removed | cell_changed
    sheet: str
    row: int | None = None       # 1-based data row
    column: str | None = None
    before: object = None
    after: object = None

    def describe(self) -> str:
        loc = self.sheet + (f"!r{self.row}" if self.row else "") + \
            (f".{self.column}" if self.column else "")
        if self.kind == "cell_changed":
            return f"{loc}: {self.before!r} -> {self.after!r}"
        return f"{self.kind}: {loc}"


def diff_sheets(before: dict[str, list[dict]], after: dict[str, list[dict]]) -> list[Difference]:
    diffs: list[Difference] = []
    for name in before:
        if name not in after:
            diffs.append(Difference("sheet_removed", name))
    for name in after:
        if name not in before:
            diffs.append(Difference("sheet_added", name))

    for name in sorted(set(before) & set(after)):
        b_rows, a_rows = before[name], after[name]
        for i in range(max(len(b_rows), len(a_rows))):
            if i >= len(b_rows):
                diffs.append(Difference("row_added", name, row=i + 1, after=a_rows[i]))
                continue
            if i >= len(a_rows):
                diffs.append(Difference("row_removed", name, row=i + 1, before=b_rows[i]))
                continue
            b_row, a_row = b_rows[i], a_rows[i]
            for col in sorted(set(b_row) | set(a_row)):
                bv, av = b_row.get(col), a_row.get(col)
                if bv != av:
                    diffs.append(Difference("cell_changed", name, row=i + 1,
                                            column=col, before=bv, after=av))
    return diffs


def summarize(diffs: list[Difference]) -> dict:
    counts: dict[str, int] = {}
    for d in diffs:
        counts[d.kind] = counts.get(d.kind, 0) + 1
    return {"total": len(diffs), **counts}


def load_workbook_sheets(path: str | Path, formulas: bool = False) -> dict[str, list[dict]]:
    """Load an .xlsx into the plain sheets shape. Requires openpyxl.

    Column keys are the row-1 header values (fallback: A, B, C…). With
    formulas=True, cells carry formula strings instead of cached values,
    so formula drift shows up in the diff.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path, data_only=not formulas, read_only=True)
    sheets: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets[ws.title] = []
            continue
        header = [str(h) if h is not None else get_column_letter(i + 1)
                  for i, h in enumerate(rows[0])]
        sheets[ws.title] = [dict(zip(header, r)) for r in rows[1:]]
    wb.close()
    return sheets
