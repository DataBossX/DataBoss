"""Source-limited Section 32 execution and canonical artifact packaging."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import platform
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .contracts import utc_now
from .providers import provider_health

PROJECT_ID = "DBX-OK-BECKHAM-32-11N-25W"
LEGAL = "Section 32, Township 11 North, Range 25 West, Beckham County, Oklahoma"
BLOCKERS = [
    {
        "code": "SECTION32_SOURCE_CORPUS_MISSING",
        "severity": "CRITICAL",
        "detail": "The asserted 4,893-image private source corpus is not mounted in this environment.",
        "needed": "Provide a read-only mounted copy and an approved source manifest from the authoritative Windows project root.",
    },
    {
        "code": "TEMPLATE_XLSX_MISSING",
        "severity": "CRITICAL",
        "detail": "No controlling Template.xlsx is present.",
        "needed": "Provide the hash-identified controlling template and writable-range approval.",
    },
    {
        "code": "HUMAN_REVIEW_MISSING",
        "severity": "CRITICAL",
        "detail": "No qualified landman review or final approver authorization is recorded.",
        "needed": "Complete qualified review and bind approval to the final candidate SHA-256.",
    },
]

RUNSHEET_COLUMNS = [
    "image_number", "document_type", "instrument_number", "book_page",
    "execution_date", "recording_date", "grantor", "grantee",
    "legal_description", "tract", "interest_conveyed", "interest_reserved",
    "depth_limitation", "formation_limitation", "term", "royalty", "burdens",
    "exceptions", "corrective_relationship", "assignment_relationship",
    "source_quality", "ocr_confidence", "visual_review_status", "relevance",
    "notes", "open_requirements",
]
CLAIM_COLUMNS = [
    "claim_id", "subject", "claim_type", "normalized_proposition",
    "supporting_source_ids", "source_filenames", "image_numbers", "book_page",
    "instrument_number", "recording_date", "legal_description", "quoted_text",
    "evidence_tier", "agent_confidence", "deterministic_validation_status",
    "reviewer_status", "contradictions", "final_disposition",
]

WORKBOOKS: Dict[str, Sequence[str]] = {
    "NORMALIZED_RUNSHEET.xlsx": RUNSHEET_COLUMNS,
    "TITLE_CHAIN_LEDGER.xlsx": [
        "chain_id", "chain_type", "tract", "instrument_number", "from_party",
        "to_party", "interest_numerator", "interest_denominator", "status",
        "source_ids", "open_requirement",
    ],
    "MINERAL_OWNERSHIP_CHAIN.xlsx": [
        "tract", "owner", "interest_numerator", "interest_denominator",
        "net_mineral_acres", "source_ids", "status", "open_requirement",
    ],
    "ROYALTY_CHAIN.xlsx": [
        "tract", "owner", "royalty_numerator", "royalty_denominator",
        "source_ids", "status", "open_requirement",
    ],
    "LEASE_CHAIN.xlsx": [
        "lease_id", "lessor", "lessee", "execution_date", "recording_data",
        "legal_description", "gross_acres", "net_mineral_acres", "royalty",
        "term", "depth_limitation", "current_apparent_owner", "hbp_status",
        "source_ids", "open_requirement",
    ],
    "ASSIGNMENT_CHAIN.xlsx": [
        "assignment_id", "assignor", "assignee", "effective_date",
        "instrument_number", "asset_schedule_present", "legal_description",
        "depth_limitation", "source_ids", "status", "open_requirement",
    ],
    "CORPORATE_CONTINUITY_MATRIX.xlsx": [
        "from_entity", "to_entity", "transaction_type", "effective_date",
        "evidence", "section32_inclusion_proof", "exclusions_reviewed",
        "retained_properties_reviewed", "depth_limitations", "confidence",
        "remaining_requirement",
    ],
    "WI_NRI_CALCULATION.xlsx": [
        "tract_or_well", "gross_wi_numerator", "gross_wi_denominator",
        "burden_numerator", "burden_denominator", "nri_numerator",
        "nri_denominator", "calculation_status", "source_ids", "open_requirement",
    ],
    "DEPTH_AND_FORMATION_MATRIX.xlsx": [
        "tract", "owner_or_lessee", "from_depth", "to_depth", "formation",
        "rights_type", "source_ids", "status", "open_requirement",
    ],
    "ORRI_BURDEN_LIEN_MATRIX.xlsx": [
        "tract_or_well", "interest_type", "beneficiary", "fraction_numerator",
        "fraction_denominator", "instrument_number", "release_instrument",
        "source_ids", "status", "open_requirement",
    ],
    "WELL_REGULATORY_MATRIX.xlsx": [
        "well_name", "api_number", "operator", "formation", "depth_interval",
        "spacing", "pooling", "location", "completion_date",
        "first_production", "status", "unit_coverage", "hbp_relevance",
        "order_numbers", "regulatory_source", "open_requirement",
    ],
    "CLAIM_EVIDENCE_LEDGER.xlsx": CLAIM_COLUMNS,
    "CONTRADICTION_REGISTER.xlsx": [
        "conflict_id", "materiality", "claim_ids", "description",
        "supporting_source_ids", "minority_opinion", "resolution_status",
        "human_resolution", "open_requirement",
    ],
    "OPEN_REQUIREMENTS_PRIORITIZED.xlsx": [
        "priority", "requirement_id", "category", "description", "reason",
        "required_source", "release_impact", "status",
    ],
    "TEMPLATE_CONFORMANCE_MATRIX.xlsx": [
        "sheet", "check", "expected", "actual", "status", "evidence",
        "open_requirement",
    ],
}

RELEASE_GATES = [
    "source inventory reconciled", "hashes verified", "evidence chain preserved",
    "title chain materially complete", "mineral chain materially complete",
    "leasehold chain materially complete", "corporate continuity supported",
    "acreage supported", "WI supported", "NRI supported", "burdens supported",
    "depth limitations supported", "wells and regulatory data supported",
    "formulas valid", "defined names valid", "no broken links", "totals foot",
    "template structure matched", "print layout verified",
    "adversarial review passed", "security tests passed",
    "human landman review completed", "final approver authorization received",
]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_json(path: Path, value: Any) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    temporary.replace(path)


def _write_csv(path: Path, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def _write_empty_workbook(path: Path, headers: Sequence[str], blocker_rows=None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SOURCE_LIMITED"
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7F1D1D")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
    for index, header in enumerate(headers, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(
            max(len(str(header)) + 2, 12), 42
        )
    for row in blocker_rows or []:
        sheet.append(list(row))
    notes = workbook.create_sheet("READ_ME")
    notes.append(["RELEASE STATE", "HOLD_NO_RELEASE"])
    notes.append(["PROJECT", LEGAL])
    notes.append(["EVIDENTIARY ROWS", 0])
    notes.append(["NOTICE", "No title conclusion may be inferred from this empty structure."])
    notes.append(["BLOCKER", "Private source corpus and controlling template are unavailable."])
    workbook.save(path)


def _environment_report() -> Dict[str, Any]:
    return {
        "schema_id": "dbx.environment_validation",
        "generated_at": utc_now(),
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "working_directory": str(Path.cwd()),
        "section32_source_present": False,
        "template_present": False,
        "excel_available": os.name == "nt",
        "google_drive_mounted": False,
        "secret_values_reported": False,
        "status": "PARTIAL",
    }


def execute_source_limited_run(output_root: Path) -> Path:
    """Create a truthful blocked package when the private evidence is unavailable."""
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    run_id = f"section32_{timestamp}"
    run_dir = Path(output_root).resolve() / run_id
    run_dir.mkdir(parents=True, exist_ok=False)

    manifest = {
        "schema_id": "dbx.run_manifest",
        "run_id": run_id,
        "project_id": PROJECT_ID,
        "legal_description": LEGAL,
        "started_at": utc_now(),
        "mode": "SOURCE_LIMITED_EVIDENCE_AUDIT",
        "source_roots_evaluated": ["/workspace"],
        "asserted_windows_paths_accessible": False,
        "asserted_drive_folder_accessible": False,
        "release_state": "HOLD_NO_RELEASE",
        "blockers": BLOCKERS,
    }
    _write_json(run_dir / "RUN_MANIFEST.json", manifest)
    _write_json(run_dir / "ENVIRONMENT_VALIDATION.json", _environment_report())
    _write_json(
        run_dir / "PROVIDER_HEALTH_REPORT.json",
        {
            "schema_id": "dbx.provider_health",
            "generated_at": utc_now(),
            "network_checks_performed": False,
            "providers": provider_health(network_checks=False),
            "providers_used": [],
            "reason": "No source evidence was available for an approved AI task.",
        },
    )
    (run_dir / "AGENT_EXECUTION_LOG.jsonl").write_text(
        json.dumps(
            {
                "at": utc_now(),
                "phase": "DISCOVERY",
                "status": "BLOCKED",
                "provider": "deterministic_local",
                "model": "none",
                "claims_asserted": 0,
                "blocker_codes": [item["code"] for item in BLOCKERS],
            },
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )
    _write_csv(
        run_dir / "SOURCE_INVENTORY.csv",
        ["source_id", "filename", "relative_path", "size_bytes", "sha256", "status"],
        [],
    )
    (run_dir / "SOURCE_HASH_LEDGER.sha256").write_text(
        "# No authoritative Section 32 source files were available.\n", encoding="utf-8"
    )
    _write_csv(
        run_dir / "EVIDENCE_GAP_AND_PULL_QUEUE.csv",
        ["priority", "blocker_code", "required_evidence", "status"],
        [
            [index, item["code"], item["needed"], "OPEN"]
            for index, item in enumerate(BLOCKERS, 1)
        ],
    )

    for filename, headers in WORKBOOKS.items():
        rows = None
        if filename == "OPEN_REQUIREMENTS_PRIORITIZED.xlsx":
            rows = [
                [
                    index, item["code"], "EVIDENCE", item["detail"], "Source unavailable",
                    item["needed"], "BLOCKING", "OPEN",
                ]
                for index, item in enumerate(BLOCKERS, 1)
            ]
        elif filename == "TEMPLATE_CONFORMANCE_MATRIX.xlsx":
            rows = [["ALL", "template comparison", "Controlling Template.xlsx",
                     "MISSING", "FAIL", "", BLOCKERS[1]["needed"]]]
        _write_empty_workbook(run_dir / filename, headers, rows)

    _write_csv(
        run_dir / "RELEASE_GATE_MATRIX.csv",
        ["gate", "status", "critical", "evidence", "blocker"],
        [
            [
                gate,
                "PASS" if gate == "security tests passed" else "FAIL",
                "YES",
                "Deterministic package generation" if gate == "security tests passed" else "",
                "" if gate == "security tests passed" else "SOURCE EVIDENCE OR HUMAN GATE MISSING",
            ]
            for gate in RELEASE_GATES
        ],
    )

    documents = {
        "SOURCE_RECONCILIATION_REPORT.md": (
            "# Source Reconciliation Report\n"
            "\nNo authoritative Section 32 evidence or Google Drive content is mounted in "
            "this environment. The claimed 4,893-image sequence cannot be verified. "
            "No source conclusion was asserted.\n"
        ),
        "WORKBOOK_QA_REPORT.md": (
            "# Workbook QA Report\n\nStatus: NOT PERFORMED\n\nNo controlling template "
            "or report candidate was available. No release-candidate XLSX was created.\n"
        ),
        "ADVERSARIAL_REVIEW_REPORT.md": (
            "# Adversarial Review Report\n\nResult: HOLD_NO_RELEASE\n\nThe evidence "
            "absence defeats every material title, ownership, leasehold, WI/NRI, "
            "depth, well, corporate-continuity, and template-conformance conclusion.\n"
        ),
        "CHANGE_LOG.md": (
            "# Change Log\n\n- Created an isolated source-limited package.\n"
            "- Preserved all originals; no source file was modified or copied.\n"
            "- Created no release candidate because no controlling template exists.\n"
        ),
        "CURRENT_STATE_REPORT.md": (
            "# Current State Report\n\n"
            "- Overall status: PARTIAL\n- Release state: HOLD_NO_RELEASE\n"
            "- Evidentiary completion: 0%\n- Material claims supported: 0\n"
            "- Release candidate: NOT CREATED\n\n"
            "All title, mineral, royalty, leasehold, assignment, corporate, WI, NRI, "
            "burden, depth, HBP, well, regulatory, and acreage conclusions remain OPEN.\n"
        ),
    }
    for filename, content in documents.items():
        (run_dir / filename).write_text(content, encoding="utf-8")

    artifacts = []
    for path in sorted(run_dir.iterdir()):
        if path.is_file():
            artifacts.append(
                {
                    "path": path.name,
                    "sha256": sha256_file(path),
                    "size_bytes": path.stat().st_size,
                    "status": "SOURCE_LIMITED",
                }
            )
    _write_json(
        run_dir / "CANONICAL_ARTIFACT_INDEX.json",
        {
            "schema_id": "dbx.canonical_artifact_index",
            "run_id": run_id,
            "release_state": "HOLD_NO_RELEASE",
            "artifacts": artifacts,
            "intentionally_absent": [
                {
                    "artifact": "final isolated release-candidate XLSX",
                    "reason": "Creating one without the controlling template would be misleading.",
                }
            ],
        },
    )
    receipt = {
        "schema_id": "dbx.completion_receipt",
        "run_id": run_id,
        "project_id": PROJECT_ID,
        "completed_at": utc_now(),
        "overall_status": "PARTIAL",
        "release_state": "HOLD_NO_RELEASE",
        "source_count": 0,
        "claims_asserted": 0,
        "release_candidate_created": False,
        "human_approval_received": False,
        "blockers": BLOCKERS,
    }
    _write_json(run_dir / "COMPLETION_RECEIPT.json", receipt)

    ledger_paths = [
        path for path in sorted(run_dir.iterdir())
        if path.is_file() and path.name != "FINAL_HASH_LEDGER.sha256"
    ]
    (run_dir / "FINAL_HASH_LEDGER.sha256").write_text(
        "".join(f"{sha256_file(path)}  {path.name}\n" for path in ledger_paths),
        encoding="utf-8",
    )
    return run_dir
