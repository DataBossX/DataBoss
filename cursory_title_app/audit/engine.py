"""Forensic QC / contradiction / math audit over the EXISTING workbook data.

This reads the workbook read-only and reports concrete, checkable findings. It
does NOT reconstruct chains from primary documents (that requires reading the
instruments themselves) and it never asserts ownership it cannot compute from
the data already in the book. Every finding carries a confidence tag.

Run:  python -m cursory_title_app.audit.engine <workbook.xlsx>
"""
from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter as L

from .. import config

TODAY = datetime.now()
ERROR_TOKENS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def _s(v):
    return "" if v is None else str(v).strip()


def _is_date(v):
    return isinstance(v, datetime)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


class Finding(dict):
    def __init__(self, code, severity, confidence, where, msg, **extra):
        super().__init__(code=code, severity=severity, confidence=confidence,
                         where=where, msg=msg, **extra)


# --------------------------------------------------------------------------- #
def audit_runsheet(wsF, wsV) -> list[Finding]:
    out: list[Finding] = []
    last = 1
    for r in range(2, wsF.max_row + 1):
        if any(_s(wsF.cell(r, c).value) for c in (1, 4, 7, 8, 9)):
            last = r

    bkpg_rows = defaultdict(list)
    for r in range(2, last + 1):
        a = _s(wsF.cell(r, 1).value)          # instrument #
        doc = _s(wsF.cell(r, 3).value)        # doc type
        d = _s(wsF.cell(r, 4).value)          # bk/pg
        eff = wsF.cell(r, 5).value
        rec = wsF.cell(r, 6).value
        gr = _s(wsF.cell(r, 7).value)
        ge = _s(wsF.cell(r, 8).value)
        legal = _s(wsF.cell(r, 9).value)
        tracts = _s(wsF.cell(r, 12).value)
        conv = _s(wsF.cell(r, 13).value)
        namt = _num(wsF.cell(r, 14).value)

        if d:
            bkpg_rows[d.upper().replace(" ", "")].append(r)

        # R-2 date sanity
        if _is_date(eff) and _is_date(rec) and eff > rec:
            out.append(Finding("R-DATE-ORDER", "med", "HIGH", f"Runsheet!E{r}/F{r}",
                       f"Effective date {eff.date()} is after recorded date {rec.date()}."))
        for lbl, dv in (("effective", eff), ("recorded", rec)):
            if _is_date(dv) and (dv.year < 1889 or dv > TODAY):
                out.append(Finding("R-DATE-RANGE", "med", "HIGH", f"Runsheet row {r}",
                           f"{lbl} date {dv.date()} is out of plausible range."))

        # R-3 completeness on substantive conveyances
        substantive = conv in ("Mineral", "Leasehold") or doc in (
            "Oil and Gas Lease", "Mineral Deed", "Warranty Deed", "Assignment")
        if substantive:
            if not d:
                out.append(Finding("R-MISSING-BKPG", "med", "HIGH", f"Runsheet row {r}",
                           f"{doc or conv} row has no Bk/Pg."))
            if not gr or not ge:
                out.append(Finding("R-MISSING-PARTY", "low", "MEDIUM", f"Runsheet row {r}",
                           f"{doc or conv} missing {'grantor' if not gr else ''}"
                           f"{' & ' if not gr and not ge else ''}{'grantee' if not ge else ''}."))
            if not tracts:
                out.append(Finding("R-MISSING-TRACT", "low", "MEDIUM", f"Runsheet row {r}",
                           f"{doc or conv} row has no Tract(s) assignment."))

        # R-4 STR / legal consistency
        if legal:
            up = legal.upper()
            if (config.TARGET_SECTION not in up and "ALL" not in up
                    and "AOL" not in up):
                out.append(Finding("R-LEGAL-STR", "low", "LOW", f"Runsheet!I{r}",
                           f"Legal description does not reference Sec "
                           f"{config.TARGET_SECTION}: {legal[:60]!r}"))

        # R-5 conveyance/doc type consistency
        if doc == "Oil and Gas Lease" and conv and conv != "Leasehold":
            out.append(Finding("R-TYPE-MISMATCH", "low", "MEDIUM", f"Runsheet row {r}",
                       f"Doc 'Oil and Gas Lease' but Conveyance type is '{conv}'."))
        if doc == "Mineral Deed" and conv and conv != "Mineral":
            out.append(Finding("R-TYPE-MISMATCH", "low", "MEDIUM", f"Runsheet row {r}",
                       f"Doc 'Mineral Deed' but Conveyance type is '{conv}'."))

        # R-7 conveyance amount sanity
        if conv == "Mineral" and namt is not None and (namt > 1 or namt < -1):
            out.append(Finding("R-CONV-AMT", "med", "HIGH", f"Runsheet!N{r}",
                       f"Mineral conveyance decimal {namt} is outside [-1, 1]."))

        # R-8 tract token validity
        if tracts:
            for tok in tracts.replace(";", ",").split(","):
                tok = tok.strip().lstrip("T").strip()
                if tok.isdigit() and not (1 <= int(tok) <= 8) and "all" not in tracts.lower():
                    out.append(Finding("R-TRACT-RANGE", "low", "MEDIUM", f"Runsheet!L{r}",
                               f"Tract token '{tok}' outside 1-8."))

    # R-1 duplicate bk/pg
    for d, rows in bkpg_rows.items():
        if len(rows) > 2:  # >2 is unusual (1 lease can have leasehold+mineral pair)
            out.append(Finding("R-DUP-BKPG", "low", "MEDIUM", f"Runsheet Bk/Pg {d}",
                       f"Bk/Pg appears on {len(rows)} rows: {rows[:12]}", rows=rows))

    # R-6 formula/error integrity on the live formula columns O..S
    error_cells = []
    for r in range(2, last + 1):
        for c in range(15, 20):
            v = wsV.cell(r, c).value
            if isinstance(v, str) and any(t in v for t in ERROR_TOKENS):
                error_cells.append(f"{L(c)}{r}={v}")
    if error_cells:
        out.append(Finding("R-FORMULA-ERROR", "high", "HIGH", "Runsheet O:S",
                   f"Formula error tokens present: {error_cells[:15]}", cells=error_cells))

    out.insert(0, Finding("R-INFO", "info", "HIGH", "Runsheet",
               f"{last - 1} populated data rows audited."))
    return out


# --------------------------------------------------------------------------- #
def audit_ogls(ws) -> list[Finding]:
    out: list[Finding] = []
    n = 0
    for r in range(2, ws.max_row + 1):
        if not _s(ws.cell(r, 4).value) and not _s(ws.cell(r, 8).value):
            continue
        n += 1
        bkpg = _s(ws.cell(r, 5).value)
        lessor = _s(ws.cell(r, 8).value)
        lessee = _s(ws.cell(r, 9).value)
        gross = ws.cell(r, 13).value
        exp = ws.cell(r, 15).value
        tracts = _s(ws.cell(r, 14).value)
        if not bkpg:
            out.append(Finding("O-MISSING-BKPG", "med", "HIGH", f"OGLs row {r}",
                       "Lease missing Bk/Pg."))
        if not lessor or not lessee:
            out.append(Finding("O-MISSING-PARTY", "med", "HIGH", f"OGLs row {r}",
                       f"Lease missing {'lessor' if not lessor else 'lessee'}."))
        if gross in (None, ""):
            out.append(Finding("O-MISSING-GROSS", "low", "MEDIUM", f"OGLs row {r}",
                       "Lease missing Gross Acres."))
        if exp in (None, ""):
            out.append(Finding("O-MISSING-EXP", "low", "MEDIUM", f"OGLs row {r}",
                       "Lease missing Expiration date (HBP/released status unknown)."))
        if not tracts:
            out.append(Finding("O-MISSING-TRACT", "low", "MEDIUM", f"OGLs row {r}",
                       "Lease not mapped to any tract."))
    out.insert(0, Finding("O-INFO", "info", "HIGH", "OGLs", f"{n} leases audited."))
    return out


# --------------------------------------------------------------------------- #
def _find_owner_block(ws):
    """Locate the 'OWNERS' header row in a tract tab."""
    for r in range(1, min(20, ws.max_row + 1)):
        for c in range(1, 10):
            if _s(ws.cell(r, c).value).upper() == "OWNERS":
                return r, c
    return None, None


def audit_tract(ws, name) -> list[Finding]:
    out: list[Finding] = []
    hdr_r, owners_c = _find_owner_block(ws)
    if hdr_r is None:
        out.append(Finding("T-PARSE", "info", "LOW", name,
                   "Could not locate OWNERS block; skipped tract math."))
        return out

    # NET ACRES is the column left of OWNERS; TOTAL is right; NET INT two right.
    net_ac_c = owners_c - 1
    total_c = owners_c + 1
    net_int_c = owners_c + 2

    tract_acres = None
    for r in range(1, hdr_r + 2):
        for c in range(1, 8):
            if _s(ws.cell(r, c).value).upper() == "TRACT ACRES":
                tract_acres = _num(ws.cell(r, c + 2).value) or _num(ws.cell(r, c + 1).value)

    sum_total = 0.0
    cur_owners = []
    neg_net = []
    seen = Counter()
    for r in range(hdr_r + 1, ws.max_row + 1):
        owner = _s(ws.cell(r, owners_c).value)
        if not owner:
            continue
        tot = _num(ws.cell(r, total_c).value)
        nac = _num(ws.cell(r, net_ac_c).value)
        nint = _num(ws.cell(r, net_int_c).value)
        if tot is not None:
            sum_total += tot
        if nint is not None and nint > 0:
            cur_owners.append((owner, nac, nint))
            seen[owner.upper()] += 1   # only count CURRENT owners for dup check
        if nac is not None and nac < -1e-9:
            neg_net.append((r, owner, nac))

    # T-1 balance: signed TOTAL column should net to ~0 (every interest in==out)
    out.append(Finding("T-BALANCE", "high" if abs(sum_total) > 1e-3 else "info",
               "HIGH", name,
               f"Conveyance ledger nets to {sum_total:.6g} "
               f"({'OUT OF BALANCE — investigate' if abs(sum_total) > 1e-3 else 'balanced'})."))

    # T-2 current ownership summary
    sum_cur_nac = sum(n for _, n, _ in cur_owners if n is not None)
    sum_cur_int = sum(i for _, _, i in cur_owners if i is not None)
    out.append(Finding("T-OWNERS", "info", "MEDIUM", name,
               f"{len(cur_owners)} current owners; net acres traced={sum_cur_nac:.4f}"
               f"{f' of {tract_acres} tract acres' if tract_acres else ''}; "
               f"net interest sum={sum_cur_int:.6f}.",
               tract_acres=tract_acres, current_owners=len(cur_owners),
               net_acres_traced=round(sum_cur_nac, 6),
               net_interest_sum=round(sum_cur_int, 8)))

    # T-3 negative current net acres = contradiction
    for r, owner, nac in neg_net:
        out.append(Finding("T-NEG-ACRES", "high", "HIGH", f"{name} row {r}",
                   f"Owner '{owner[:40]}' shows negative net acres {nac:.4f}."))

    # T-4 duplicate owner names (possible double counting / merge needed)
    for owner, cnt in seen.items():
        if cnt > 1:
            out.append(Finding("T-DUP-OWNER", "low", "MEDIUM", name,
                       f"Owner name appears {cnt}x: {owner[:40]} (verify same party/merge)."))
    return out


# --------------------------------------------------------------------------- #
def audit_title(ws) -> list[Finding]:
    out: list[Finding] = []
    verify = 0
    nonnum_acres = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, 8):
            v = _s(ws.cell(r, c).value)
            if "VERIFY" in v.upper() or "NEED" in v.upper():
                verify += 1
        # net acres column is typically C; flag non-numeric ones
        cval = ws.cell(r, 3).value
        if isinstance(cval, str) and ("VERIFY" in cval.upper() or "NEED" in cval.upper()):
            nonnum_acres += 1
    out.append(Finding("TI-VERIFY", "med", "HIGH", "Title",
               f"{verify} existing VERIFY/NEED flags in Title tab (human-review backlog)."))
    out.append(Finding("TI-NONNUM-ACRES", "med", "HIGH", "Title",
               f"{nonnum_acres} owner rows have non-numeric Net Acres (unresolved)."))
    return out


# --------------------------------------------------------------------------- #
# Codes that are completeness/quality gaps -> collapse to one counted finding.
_AGGREGATE = {
    "O-MISSING-GROSS", "O-MISSING-EXP", "O-MISSING-BKPG", "O-MISSING-PARTY",
    "O-MISSING-TRACT", "R-MISSING-BKPG", "R-MISSING-PARTY", "R-MISSING-TRACT",
    "R-LEGAL-STR", "R-TYPE-MISMATCH", "R-DUP-BKPG", "T-DUP-OWNER", "R-TRACT-RANGE",
}


def _aggregate(findings: list[Finding]) -> list[Finding]:
    keep, buckets = [], defaultdict(list)
    for f in findings:
        if f["code"] in _AGGREGATE:
            buckets[f["code"]].append(f)
        else:
            keep.append(f)
    for code, items in buckets.items():
        sev = items[0]["severity"]
        conf = items[0]["confidence"]
        examples = [f"{i['where']}: {i['msg']}" for i in items[:5]]
        keep.append(Finding(code, sev, conf, f"{len(items)} occurrences",
                            f"{len(items)} occurrences. e.g. " + " | ".join(examples),
                            count=len(items)))
    return keep


def run_audit(path: Path) -> dict:
    path = Path(path)
    wbF = openpyxl.load_workbook(path, data_only=False, keep_links=True)
    wbV = openpyxl.load_workbook(path, data_only=True, keep_links=True)

    findings: list[Finding] = []
    findings += audit_runsheet(wbF["Runsheet"], wbV["Runsheet"])
    if "OGLs" in wbF.sheetnames:
        findings += audit_ogls(wbV["OGLs"])
    for t in [s for s in wbF.sheetnames if s.startswith("Tract ")]:
        findings += audit_tract(wbV[t], t)
    if "Title " in wbF.sheetnames:
        findings += audit_title(wbV["Title "])

    findings = _aggregate(findings)
    by_sev = Counter(f["severity"] for f in findings)
    by_code = Counter(f["code"] for f in findings)
    return {
        "workbook": str(path),
        "generated": datetime.now().isoformat(timespec="seconds"),
        "totals": {"findings": len(findings), "by_severity": dict(by_sev),
                   "by_code": dict(by_code)},
        "findings": findings,
    }


def to_markdown(report: dict) -> str:
    lines = [f"# Section {config.TARGET_SECTION}-{config.TARGET_TOWNSHIP}-"
             f"{config.TARGET_RANGE} — Forensic QC Audit",
             f"_{report['generated']}_  ·  workbook: `{Path(report['workbook']).name}`",
             "",
             f"**{report['totals']['findings']} findings** — "
             + ", ".join(f"{k}:{v}" for k, v in report["totals"]["by_severity"].items()),
             ""]
    order = {"high": 0, "med": 1, "low": 2, "info": 3}
    for sev in ("high", "med", "low", "info"):
        sub = [f for f in report["findings"] if f["severity"] == sev]
        if not sub:
            continue
        lines.append(f"## {sev.upper()} ({len(sub)})")
        for f in sub:
            lines.append(f"- `{f['code']}` **[{f['confidence']}]** {f['where']}: {f['msg']}")
        lines.append("")
    return "\n".join(lines)


if __name__ == "__main__":
    wb = sys.argv[1] if len(sys.argv) > 1 else None
    if not wb:
        print("usage: python -m cursory_title_app.audit.engine <workbook.xlsx>")
        sys.exit(2)
    rep = run_audit(Path(wb))
    md = to_markdown(rep)
    out_json = config.DATA_DIR / "audit_report.json"
    out_md = config.DATA_DIR / "audit_report.md"
    config.ensure_dirs()
    out_json.write_text(json.dumps(rep, indent=2, default=str))
    out_md.write_text(md)
    print(md)
    print(f"\n[written] {out_json}\n[written] {out_md}")
