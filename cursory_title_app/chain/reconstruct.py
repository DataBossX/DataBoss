"""Independent chain-of-title reconstruction from the existing Runsheet.

For each tract it orders the MINERAL conveyances + root grants chronologically
and verifies grantor continuity: a grantor must have previously taken title
(been a grantee/patentee) in that tract, or be the sovereign. Grantors that
appear from nowhere are flagged as POSSIBLE MISSING LINKS (wild deeds / missing
instruments) — the classic title defect.

This is computed only from data already in the workbook. It does not prove
ownership; it surfaces breaks for a human to cure. Findings are MED confidence
because partial-interest conveyances and OCR name noise can mislead; entity
resolution (chain/entities.py) suppresses the obvious false positives.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from .entities import persons, any_match, is_sovereign

ROOT_DOCS = {"Patent", "Homestead", "Homestead Patent"}
# doc types that genuinely move mineral title (exclude leases/affidavits/releases)
MINERAL_MOVERS = {
    "Mineral Deed", "Warranty Deed", "Deed", "Quit Claim Deed",
    "Quit Claim Mineral Deed", "Quitclaim Deed", "Quitclaim Mineral Deed",
    "Joint Tenancy Warranty Deed", "Trustee's Mineral Deed", "Mineral and Royalty Deed",
    "Sheriff's Deed", "Trustee's Deed", "Correction Mineral Deed",
    "Final Decree", "Order and Final Decree", "Decree", "Patent", "Homestead",
    "Homestead Patent", "Non-Participating Royalty Deed",
}


def _s(v):
    return "" if v is None else str(v).strip()


def _tracts_of(cell) -> set[str]:
    s = _s(cell).lower()
    if "all" in s:
        return {str(i) for i in range(1, 9)}
    out = set()
    for tok in s.replace(";", ",").split(","):
        tok = tok.strip().lstrip("t").strip()
        if tok.isdigit():
            out.add(tok)
    return out


def _last_data_row(ws):
    last = 1
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in (1, 4, 7, 8, 9)):
            last = r
    return last


def _load_events(ws):
    """All mineral-moving conveyances with parsed parties + tract set."""
    last = _last_data_row(ws)
    evs = []
    for r in range(2, last + 1):
        doc = _s(ws.cell(r, 3).value)
        conv = _s(ws.cell(r, 13).value)
        if conv != "Mineral" and doc not in MINERAL_MOVERS:
            continue
        eff = ws.cell(r, 5).value
        rec = ws.cell(r, 6).value
        date = rec or eff
        evs.append({
            "row": r, "doc": doc, "conv": conv,
            "date": date,
            "grantor": _s(ws.cell(r, 7).value),
            "grantee": _s(ws.cell(r, 8).value),
            "bkpg": _s(ws.cell(r, 4).value),
            "tracts": _tracts_of(ws.cell(r, 12).value),
        })

    def _key(e):
        d = e["date"]
        return (getattr(d, "year", 0), getattr(d, "month", 0),
                getattr(d, "day", 0), e["row"])
    evs.sort(key=_key)
    return evs


def _date_str(d):
    return str(d.date()) if hasattr(d, "date") else str(d)


import re as _re
_ENTITY_RE = _re.compile(
    r"\b(inc|llc|l\.?p\.?|company|co\.?|corp|corporation|bank|trust|trustee|"
    r"n\.?a\.?|partners|holdings|resources|energy|operating|exploration)\b", _re.I)


def _categorize(doc: str, grantor: str):
    d = (doc or "").lower()
    if any(k in d for k in ("decree", "order", "court", "probate", "final account",
                            "return of sale", "sheriff")):
        return ("probate",
                "estate/probate/court transfer — verify the decree vests grantor & heirship",
                "LOW")
    if _ENTITY_RE.search(grantor or ""):
        return ("entity-succession",
                "corporate predecessor/merger likely not separately indexed — verify succession",
                "LOW")
    return ("wild-deed",
            "possible missing conveyance / wild deed / unindexed link / OCR name variant",
            "MEDIUM")


def reconstruct(workbook: Path) -> dict:
    """Section-wide single pass: a grantor is 'supported' if it is the sovereign,
    a root grant, or matches a grantee from any EARLIER instrument in the whole
    section. Otherwise it is a POSSIBLE MISSING LINK (deduped by row)."""
    wb = openpyxl.load_workbook(workbook, data_only=True, keep_links=True)
    events = _load_events(wb["Runsheet"])

    vested: list[list] = []          # person-lists who have taken title (section)
    for e in events:
        gtor_p = persons(e["grantor"])
        is_root = e["doc"] in ROOT_DOCS or is_sovereign(e["grantor"])
        # A grantor we cannot identify (empty/unparseable) is NOT support — it is
        # a break to verify, not silently "ok".
        supported = is_root or any(any_match(gtor_p, v) for v in vested)
        e["status"] = "root" if is_root else ("ok" if supported else "BREAK")
        gtee_p = persons(e["grantee"])
        if gtee_p:
            vested.append(gtee_p)

    defects = []
    for e in events:
        if e["status"] != "BREAK":
            continue
        cat, curative, conf = _categorize(e["doc"], e["grantor"])
        defects.append({
            "row": e["row"], "date": _date_str(e["date"]), "doc": e["doc"],
            "bkpg": e["bkpg"], "grantor": e["grantor"],
            "tracts": ",".join(sorted(e["tracts"])),
            "category": cat,
            "issue": "grantor not found vested earlier in the section chain",
            "curative": curative, "confidence": conf,
        })

    tracts = []
    for i in range(1, 9):
        ti = str(i)
        tev = [e for e in events if ti in e["tracts"]]
        tracts.append({
            "tract": ti,
            "events": len(tev),
            "breaks": sum(1 for e in tev if e["status"] == "BREAK"),
            "has_root": any(e["status"] == "root" for e in tev),
            "timeline": [{
                "row": e["row"], "date": _date_str(e["date"]), "doc": e["doc"],
                "bkpg": e["bkpg"], "grantor": e["grantor"][:40],
                "grantee": e["grantee"][:40], "status": e["status"],
            } for e in tev],
        })

    from collections import Counter
    return {
        "total_mineral_events": len(events),
        "tracts": tracts,
        "defects": defects,
        "total_defects": len(defects),
        "defects_by_category": dict(Counter(d["category"] for d in defects)),
        "tracts_without_root": [t["tract"] for t in tracts if not t["has_root"]],
    }
