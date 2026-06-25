"""Write the Chain-of-Title report (external file; never touches the workbook).

Two sheets:
  * Defects        — categorized missing-link findings + curative + search link
  * Chain Timeline — every mineral-moving instrument per tract, in order, with
                     status (root / ok / BREAK): the chained-out title narrative.
Plus a markdown summary.
"""
from __future__ import annotations

from pathlib import Path
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as L

from .. import config
from . import reconstruct

DATED = "(6-25-2026)"
SEARCH = "https://okcountyrecords.com/search/roger%20mills"
_CAT_FILL = {"wild-deed": "F4CCCC", "probate": "FFF2CC", "entity-succession": "D9E1F2"}
_STATUS_FILL = {"BREAK": "F4CCCC", "root": "E2EFDA", "ok": "FFFFFF"}


def _hdr(ws, headers, fill="1F4E78"):
    f = PatternFill("solid", fgColor=fill)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = f
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def build(source: Path, out_dir: Path = config.OUTPUT_DIR) -> dict:
    rep = reconstruct.reconstruct(source)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx = out_dir / f"Section31_Chain_of_Title_{DATED}.xlsx"
    md = out_dir / f"Section31_Chain_of_Title_{DATED}.md"

    wb = openpyxl.Workbook()

    # --- Defects sheet ---
    ws = wb.active
    ws.title = "Defects"
    cols = ["row", "date", "doc", "bkpg", "tracts", "category", "issue",
            "curative", "confidence", "grantor", "search"]
    _hdr(ws, cols, fill="7F1D1D")
    for i, d in enumerate(rep["defects"], 2):
        link = f"{SEARCH}?q={quote(d['grantor'][:60])}" if d["grantor"] else SEARCH
        vals = [d["row"], d["date"], d["doc"], d["bkpg"], d["tracts"], d["category"],
                d["issue"], d["curative"], d["confidence"], d["grantor"], link]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            if cols[c - 1] == "category":
                cell.fill = PatternFill("solid", fgColor=_CAT_FILL.get(v, "FFFFFF"))
            if cols[c - 1] == "search" and v:
                cell.hyperlink = v
                cell.font = Font(color="0563C1", underline="single")
    ws.auto_filter.ref = f"A1:{L(len(cols))}{len(rep['defects'])+1}"
    for c, w in {1: 6, 2: 11, 3: 20, 4: 10, 5: 11, 6: 16, 7: 30, 8: 38,
                 9: 11, 10: 34, 11: 40}.items():
        ws.column_dimensions[L(c)].width = w

    # --- Chain Timeline sheet ---
    ws2 = wb.create_sheet("Chain Timeline")
    cols2 = ["tract", "row", "date", "doc", "bkpg", "grantor", "grantee", "status"]
    _hdr(ws2, cols2)
    r = 2
    for t in rep["tracts"]:
        for e in t["timeline"]:
            vals = [t["tract"], e["row"], e["date"], e["doc"], e["bkpg"],
                    e["grantor"], e["grantee"], e["status"]]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(r, c, v)
                if cols2[c - 1] == "status":
                    cell.fill = PatternFill("solid", fgColor=_STATUS_FILL.get(v, "FFFFFF"))
            r += 1
    ws2.auto_filter.ref = f"A1:{L(len(cols2))}{r-1}"
    for c, w in {1: 6, 2: 6, 3: 11, 4: 20, 5: 10, 6: 34, 7: 34, 8: 8}.items():
        ws2.column_dimensions[L(c)].width = w
    wb.save(xlsx)

    # --- markdown ---
    lines = [f"# Section 31-12N-24W — Chain-of-Title Reconstruction {DATED}",
             "",
             f"- Mineral-moving instruments analyzed: **{rep['total_mineral_events']}**",
             f"- Possible missing links (deduped by row): **{rep['total_defects']}**",
             f"- By category: {rep['defects_by_category']}",
             f"- Tracts lacking a root grant: {rep['tracts_without_root'] or 'none'}",
             "",
             "_Computed from the existing Runsheet. A 'break' means the grantor was "
             "not found vested earlier in the section — it flags a link to verify, "
             "not a proven defect. Cursory report: gaps are expected._",
             "", "## Per-tract", ""]
    for t in rep["tracts"]:
        lines.append(f"- **Tract {t['tract']}**: {t['events']} events, "
                     f"root grant {'present' if t['has_root'] else 'MISSING'}, "
                     f"{t['breaks']} flagged links")
    lines += ["", "## Missing-link defects", ""]
    for d in rep["defects"]:
        lines.append(f"- `r{d['row']}` {d['date']} {d['doc']} [{d['category']}] "
                     f"**{d['grantor']}** — {d['curative']} ({d['confidence']})")
    md.write_text("\n".join(lines))

    return {"xlsx": str(xlsx), "md": str(md), "summary": {
        "events": rep["total_mineral_events"], "defects": rep["total_defects"],
        "by_category": rep["defects_by_category"],
        "tracts_without_root": rep["tracts_without_root"]}}


if __name__ == "__main__":
    import json
    import sys
    print(json.dumps(build(Path(sys.argv[1]))["summary"], indent=2))
