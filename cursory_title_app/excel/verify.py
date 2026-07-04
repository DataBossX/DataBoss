"""Post-write verification / QA on the produced workbook.

Confirms the hard guarantees the user requires:
  * the file opens cleanly (loadable, no repair needed)
  * sheet names are unchanged vs the source
  * the live-formula columns O..S still contain formulas (not clobbered)
  * no #REF! / #VALUE! / #NAME? / #DIV/0! error strings were introduced
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from .. import config
from ..runsheet.columns import FORMULA_COLUMNS, FIRST_DATA_ROW

ERROR_TOKENS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def verify_workbook(output_path: Path, source_path: Path) -> dict:
    result = {"ok": True, "checks": [], "errors": []}

    def check(name, passed, detail=""):
        result["checks"].append({"check": name, "passed": bool(passed), "detail": detail})
        if not passed:
            result["ok"] = False
            result["errors"].append(f"{name}: {detail}")

    # 1) opens cleanly
    try:
        out_wb = openpyxl.load_workbook(output_path, data_only=False, keep_links=True)
        check("output_opens", True)
    except Exception as e:  # pragma: no cover
        check("output_opens", False, repr(e))
        return result

    src_wb = openpyxl.load_workbook(source_path, data_only=False, keep_links=True)

    # 2) sheet names unchanged
    check("tabs_unchanged", src_wb.sheetnames == out_wb.sheetnames,
          f"source={src_wb.sheetnames} output={out_wb.sheetnames}")

    # 3) expected tabs present, none added
    check("no_unexpected_tabs",
          set(out_wb.sheetnames) == set(config.EXPECTED_TABS),
          f"output tabs={out_wb.sheetnames}")

    # 4) formula columns still hold formulas where the source had them
    ws_src = src_wb[config.RUNSHEET_TAB]
    ws_out = out_wb[config.RUNSHEET_TAB]
    clobbered = []
    for col in FORMULA_COLUMNS:
        for row in range(FIRST_DATA_ROW, ws_src.max_row + 1):
            sv = ws_src[f"{col}{row}"].value
            if isinstance(sv, str) and sv.startswith("="):
                ov = ws_out[f"{col}{row}"].value
                if not (isinstance(ov, str) and ov.startswith("=")):
                    clobbered.append(f"{col}{row}")
    # Appended rows (below the source range) that carry data in A–N MUST also
    # carry the O–S formulas, or their acreage/interest math is dead.
    missing_on_appended = []
    for row in range(ws_src.max_row + 1, ws_out.max_row + 1):
        has_data = any(ws_out[f"{c}{row}"].value not in (None, "")
                       for c in ("A", "D", "G", "H", "I"))
        if not has_data:
            continue
        for col in FORMULA_COLUMNS:
            ov = ws_out[f"{col}{row}"].value
            if not (isinstance(ov, str) and ov.startswith("=")):
                missing_on_appended.append(f"{col}{row}")
    check("formula_columns_intact", not clobbered and not missing_on_appended,
          f"clobbered={clobbered[:20]} missing_on_appended={missing_on_appended[:20]}")

    # 5) no error tokens introduced in the Runsheet
    introduced = []
    for row in ws_out.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and any(tok in v for tok in ERROR_TOKENS):
                introduced.append(f"{cell.coordinate}={v}")
    check("no_error_tokens", not introduced, f"tokens={introduced[:20]}")

    return result
