"""Format-preserving Runsheet writer.

Two backends, same contract:
  * COMWriter  (Windows + Excel installed)  -> PRIMARY. Opens the real .xlsx in
    Excel, sets only specific cells, saves. Nothing else in the file is touched,
    so every style/merge/filter/print-area survives exactly.
  * OpenpyxlWriter (cross-platform fallback) -> loads the workbook with
    keep_vba/rich formatting and writes only target cells. Never rebuilds from a
    blank workbook (that is what pandas.to_excel does — REJECTED).

Both refuse to write the live-formula columns O..S, both make a timestamped
backup, and both write to a NEW output file (never overwrite the source unless
the caller explicitly passes the same path AND approve_overwrite=True).
"""
from __future__ import annotations

import datetime as _dt
import shutil
from pathlib import Path
from typing import Optional

from .. import config
from ..runsheet.columns import (
    FIELD_TO_COLUMN,
    FORMULA_COLUMNS,
    assert_writable,
)


_FORMULA_TRIGGERS = ("=", "+", "-", "@")


def _is_injection(value) -> bool:
    """A plain (non-formula) value that Excel would interpret as a formula."""
    return isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS


def _set_openpyxl(ws, coord, value, is_formula: bool):
    """Set a cell; force string type for non-formula values that look like
    formulas (e.g. a party name '=HYPERLINK(...)') so they can never execute."""
    cell = ws[coord]
    cell.value = value
    if not is_formula and _is_injection(value):
        cell.data_type = "s"


def _set_com(ws, coord, value, is_formula: bool):
    if is_formula:
        ws.Range(coord).Formula = value
    elif _is_injection(value):
        # Leading apostrophe = Excel text indicator (hidden), neutralizes formula.
        ws.Range(coord).Value = "'" + value
    else:
        ws.Range(coord).Value = value


def _timestamp() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def make_backup(src: Path, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    dest = backup_dir / f"{src.stem}.BACKUP_{_timestamp()}{src.suffix}"
    shutil.copy2(src, dest)
    return dest


def hyperlink_formula(url: str, label: Optional[str] = None) -> str:
    """Build an Excel =HYPERLINK formula matching the workbook's existing style."""
    label = label or url
    safe_url = url.replace('"', '""')
    safe_label = str(label).replace('"', '""')
    return f'=HYPERLINK("{safe_url}","{safe_label}")'


def _resolve_cells(field_values: dict) -> dict[str, object]:
    """Map logical field names -> column letters, enforcing write guards."""
    cells: dict[str, object] = {}
    for field, value in field_values.items():
        col = FIELD_TO_COLUMN.get(field)
        if col is None:
            raise KeyError(f"Unknown field '{field}' (no Runsheet column).")
        assert_writable(col)          # hard stop on O..S or unknown columns
        cells[col] = value
    return cells


class BaseWriter:
    backend = "base"

    def write_rows(self, writes: list) -> dict:
        raise NotImplementedError


# --------------------------------------------------------------------------- #
# openpyxl fallback (also used for verification on any OS)
# --------------------------------------------------------------------------- #
class OpenpyxlWriter(BaseWriter):
    backend = "openpyxl"

    def __init__(self, source_path: Path):
        self.source_path = Path(source_path)

    def write(self, writes: list, output_path: Path) -> dict:
        import openpyxl

        wb = openpyxl.load_workbook(self.source_path, keep_links=True)
        before_tabs = list(wb.sheetnames)
        ws = wb[config.RUNSHEET_TAB]

        written = 0
        for w in writes:
            cells = _resolve_cells(w.values)
            for col, value in cells.items():
                _set_openpyxl(ws, f"{col}{w.row}", value, is_formula=False)
                written += 1
            if w.document_link_url:
                assert_writable("K")
                ws[f"K{w.row}"] = hyperlink_formula(
                    w.document_link_url, w.document_link_label
                )
                written += 1

        after_tabs = list(wb.sheetnames)
        if before_tabs != after_tabs:
            raise RuntimeError(
                f"Tab set changed during write! before={before_tabs} after={after_tabs}"
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return {"backend": self.backend, "cells_written": written,
                "output": str(output_path), "tabs": after_tabs}


# --------------------------------------------------------------------------- #
# Excel COM primary (Windows only)
# --------------------------------------------------------------------------- #
class COMWriter(BaseWriter):
    backend = "com"

    def __init__(self, source_path: Path):
        self.source_path = Path(source_path)

    def write(self, writes: list, output_path: Path) -> dict:
        import win32com.client as win32  # pywin32, Windows only

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = None
        try:
            wb = excel.Workbooks.Open(str(self.source_path.resolve()))
            before_tabs = [s.Name for s in wb.Worksheets]
            ws = wb.Worksheets(config.RUNSHEET_TAB)

            written = 0
            for w in writes:
                cells = _resolve_cells(w.values)
                for col, value in cells.items():
                    _set_com(ws, f"{col}{w.row}", value, is_formula=False)
                    written += 1
                if w.document_link_url:
                    assert_writable("K")
                    # Use the formula form so it matches existing K-column style.
                    ws.Range(f"K{w.row}").Formula = hyperlink_formula(
                        w.document_link_url, w.document_link_label
                    )
                    written += 1

            after_tabs = [s.Name for s in wb.Worksheets]
            if before_tabs != after_tabs:
                raise RuntimeError(
                    f"Tab set changed! before={before_tabs} after={after_tabs}"
                )

            output_path.parent.mkdir(parents=True, exist_ok=True)
            # 51 = xlOpenXMLWorkbook (.xlsx)
            wb.SaveAs(str(output_path.resolve()), FileFormat=51)
            return {"backend": self.backend, "cells_written": written,
                    "output": str(output_path), "tabs": after_tabs}
        finally:
            if wb is not None:
                wb.Close(SaveChanges=False)
            excel.Quit()


# --------------------------------------------------------------------------- #
# Generic edit applier (used by the report builder and the re-importer).
# An edit is {"sheet": "Runsheet", "cell": "K12", "value": ..., "is_formula": bool}.
# Unlike write_runsheet(), this trusts the caller's edit list — callers MUST
# guard formula columns themselves (the re-importer only writes O..S on rows it
# has verified are previously blank, i.e. appended rows).
# --------------------------------------------------------------------------- #
def _apply_edits_openpyxl(source: Path, edits: list[dict], out: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(source, keep_links=True)
    before = list(wb.sheetnames)
    for e in edits:
        ws = wb[e.get("sheet", "Runsheet")]
        _set_openpyxl(ws, e["cell"], e["value"], e.get("is_formula", False))
    if list(wb.sheetnames) != before:
        raise RuntimeError("Tab set changed during edit apply!")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def _apply_edits_com(source: Path, edits: list[dict], out: Path) -> None:
    import win32com.client as win32

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(Path(source).resolve()))
        before = [s.Name for s in wb.Worksheets]
        for e in edits:
            ws = wb.Worksheets(e.get("sheet", "Runsheet"))
            _set_com(ws, e["cell"], e["value"], e.get("is_formula", False))
        if [s.Name for s in wb.Worksheets] != before:
            raise RuntimeError("Tab set changed during edit apply!")
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.SaveAs(str(Path(out).resolve()), FileFormat=51)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()


def apply_edits(source: Path, edits: list[dict], out: Path, prefer: str = "auto") -> str:
    """Apply a list of cell edits to a NEW output file. Returns backend used."""
    if prefer in ("com", "auto"):
        try:
            import win32com.client  # noqa: F401
            _apply_edits_com(Path(source), edits, Path(out))
            return "com"
        except Exception:
            if prefer == "com":
                raise
    _apply_edits_openpyxl(Path(source), edits, Path(out))
    return "openpyxl"


def get_writer(source_path: Path, prefer: str = "auto") -> BaseWriter:
    """Pick the safest available backend. COM on Windows, else openpyxl."""
    if prefer in ("com", "auto"):
        try:
            import win32com.client  # noqa: F401
            return COMWriter(source_path)
        except Exception:
            if prefer == "com":
                raise RuntimeError(
                    "COM backend requested but pywin32/Excel unavailable."
                )
    return OpenpyxlWriter(source_path)


def write_runsheet(
    source_path: Path,
    writes: list,
    output_dir: Path = config.OUTPUT_DIR,
    backup_dir: Path = config.BACKUP_DIR,
    prefer: str = "auto",
) -> dict:
    """Top-level entry: backup, write to a NEW output file, return a report.

    Never overwrites the source. Caller verifies the result with excel.verify.
    """
    source_path = Path(source_path)
    backup = make_backup(source_path, backup_dir)

    output_dir.mkdir(parents=True, exist_ok=True)
    out = output_dir / f"{source_path.stem}.UPDATED_{_timestamp()}{source_path.suffix}"

    writer = get_writer(source_path, prefer=prefer)
    report = writer.write(writes, out)
    report["backup"] = str(backup)
    report["source"] = str(source_path)
    report["protected_columns"] = FORMULA_COLUMNS
    return report
