"""Build the best HONEST 6-25-2026 deliverable set from the existing workbook.

Produces THREE files in the output folder:
  1) 31-12N-24W_Roger_Mills_Cursory_Title_Report_(6-25-2026).xlsx
       - format-preserving copy of the source
       - repairs plain-text Document Link cells into working =HYPERLINK formulas
       - adds QC flags (effective>recorded) into BLANK Review cells only
       - never fabricates ownership; never adds/renames tabs; never touches O..S
  2) Section31_Ownership_Reconciliation_(6-25-2026).xlsx   (external)
       - per-tract current owners computed from the balanced Tract ledgers
       - reconciled against the curated Title-tab totals, gap highlighted
  3) Section31_Curative_Manifest_(6-25-2026).xlsx/.csv      (external)
       - every Title cell that says VERIFY/NEED + the document required to fill it
"""
from __future__ import annotations

import csv
import datetime as _dt
import shutil
from pathlib import Path
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as L

from .. import config
from ..excel import verify as xlverify
from ..excel.writer import hyperlink_formula, make_backup, apply_edits
from ..runsheet.columns import assert_writable, FIRST_DATA_ROW
from . import ownership

DATED = "(6-25-2026)"
DETAIL = "https://okcountyrecords.com/detail/roger+mills/"
SEARCH = "https://okcountyrecords.com/search/roger%20mills"


def _last_runsheet_row(ws) -> int:
    last = 1
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in (1, 4, 7, 8, 9)):
            last = r
    return last


# --------------------------------------------------------------------------- #
def _compute_dated_edits(source: Path) -> tuple[list[dict], list[int], list[int]]:
    """Read-only analysis. Returns (edits, link_fix_rows, date_flag_rows).
    Each edit: {cell, value, is_formula}. Backend-agnostic so it can be applied
    via openpyxl OR Excel COM identically."""
    wbF = openpyxl.load_workbook(source, keep_links=True)
    wbV = openpyxl.load_workbook(source, data_only=True, keep_links=True)
    ws, wv = wbF["Runsheet"], wbV["Runsheet"]
    last = _last_runsheet_row(ws)

    edits: list[dict] = []
    link_fixes, date_flags = [], []
    for r in range(FIRST_DATA_ROW, last + 1):
        # (1) repair Document Link cells that aren't already HYPERLINK formulas:
        #     a bare http(s) URL -> clickable HYPERLINK to that URL; other plain
        #     text (e.g. "OKCountyRecords: 1994-007065") -> detail link by instr#.
        k = ws.cell(r, 11).value
        if isinstance(k, str) and k.strip() and not k.strip().upper().startswith("=HYPERLINK"):
            ks = k.strip()
            assert_writable("K")
            if ks.lower().startswith("http"):
                edits.append({"cell": f"K{r}", "is_formula": True,
                              "value": hyperlink_formula(ks, ks)})
                link_fixes.append(r)
            else:
                instr = str(ws.cell(r, 1).value or "").strip()
                if instr:
                    edits.append({"cell": f"K{r}", "is_formula": True,
                                  "value": hyperlink_formula(DETAIL + quote(instr), ks)})
                    link_fixes.append(r)

        # (2) QC flag: effective date after recorded date (computed values)
        eff, rec = wv.cell(r, 5).value, wv.cell(r, 6).value
        if isinstance(eff, _dt.datetime) and isinstance(rec, _dt.datetime) and eff > rec:
            t = ws.cell(r, 20).value
            note = "VERIFY: effective date after recorded date"
            if t in (None, ""):
                newval = note
            elif note not in str(t):
                newval = f"{t}; {note}"
            else:
                newval = t
            if newval != t:
                assert_writable("T")
                edits.append({"cell": f"T{r}", "is_formula": False, "value": newval})
            date_flags.append(r)
    return edits, link_fixes, date_flags


def build_dated_workbook(source: Path, out_dir: Path, prefer: str = "auto") -> dict:
    source = Path(source)
    make_backup(source, config.BACKUP_DIR)
    out = out_dir / f"31-12N-24W_Roger_Mills_Cursory_Title_Report_{DATED}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)

    edits, link_fixes, date_flags = _compute_dated_edits(source)
    backend = apply_edits(source, edits, out, prefer=prefer)

    res = xlverify.verify_workbook(out, source)
    return {"output": str(out), "backend": backend, "link_fixes": link_fixes,
            "date_flags": date_flags, "verify_ok": res["ok"], "verify": res}


# --------------------------------------------------------------------------- #
def _style_header(ws, headers, fill="1F4E78"):
    hf = PatternFill("solid", fgColor=fill)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = hf
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{L(len(headers))}1"


def build_ownership_reconciliation(source: Path, out_dir: Path) -> dict:
    rec = ownership.reconcile(source)
    out = out_dir / f"Section31_Ownership_Reconciliation_{DATED}.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: Reconciliation summary
    ws = wb.active
    ws.title = "Reconciliation"
    hdr = ["Tract", "Gross Acres", "Ledger Owners (current)", "Ledger NMA (computed)",
           "Title Tab TOTAL", "Title Owners Listed", "Title VERIFY rows",
           "Gap (Ledger - Title)", "Status"]
    _style_header(ws, hdr)
    warn = PatternFill("solid", fgColor="FCE4D6")
    ok = PatternFill("solid", fgColor="E2EFDA")
    tot_led = tot_title = 0.0
    for i, s in enumerate(rec["summary"], start=2):
        gap = s["gap_acres"]
        status = "TIES" if (gap is not None and abs(gap) < 0.01) else "GAP — resolve"
        row = [s["tract"], s["tract_acres"], s["ledger_owners"], s["ledger_net_acres"],
               s["title_total"], s["title_owners_listed"], s["title_verify_rows"],
               gap, status]
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            if c == 9:
                cell.fill = ok if status == "TIES" else warn
        tot_led += s["ledger_net_acres"] or 0
        tot_title += s["title_total"] or 0
    n = len(rec["summary"]) + 2
    ws.cell(n, 1, "TOTAL").font = Font(bold=True)
    ws.cell(n, 4, round(tot_led, 2)).font = Font(bold=True)
    ws.cell(n, 5, round(tot_title, 2)).font = Font(bold=True)
    for c, w in {1: 7, 2: 11, 3: 18, 4: 18, 5: 14, 6: 16, 7: 14, 8: 16, 9: 16}.items():
        ws.column_dimensions[L(c)].width = w

    # Sheet 2: full ledger ownership (the chained-out current owners)
    ws2 = wb.create_sheet("Ledger Ownership")
    hdr2 = ["Tract", "Gross Acres", "Owner (computed current)", "Net Acres", "Net Interest"]
    _style_header(ws2, hdr2)
    r = 2
    for t in rec["tracts"]:
        for o in t["owners"]:
            ws2.cell(r, 1, t["tract"]); ws2.cell(r, 2, t.get("tract_acres"))
            ws2.cell(r, 3, o["owner"]); ws2.cell(r, 4, o["net_acres"])
            ws2.cell(r, 5, o["net_interest"]); r += 1
    for c, w in {1: 7, 2: 11, 3: 46, 4: 12, 5: 14}.items():
        ws2.column_dimensions[L(c)].width = w
    ws2.auto_filter.ref = f"A1:E{r-1}"

    out_dir.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return {"output": str(out), "summary": rec["summary"],
            "ledger_total_nma": round(tot_led, 2),
            "title_total_nma": round(tot_title, 2)}


# --------------------------------------------------------------------------- #
def curative_rows(source: Path) -> list[dict]:
    """The curative items as data (shared by the file export and the browser
    walker) — every Title cell that says VERIFY/NEED + the doc needed."""
    rec = ownership.reconcile(source)
    rows = []
    for sec in rec["title"]:
        for o in sec["owners"]:
            if not o["needs"]:
                continue
            q = o["owner"] or o["ogl"]
            rows.append({
                "tract": sec.get("tract_num"),
                "legal": sec["legal"],
                "owner": o["owner"],
                "stated_net_acres": o["net_acres_raw"],
                "ogl_or_bkpg": o["ogl"],
                "royalty": o["royalty"],
                "what_is_needed": o["net_acres_raw"],
                "okcountyrecords_search": f"{SEARCH}?q={quote(q)}" if q else SEARCH,
            })
    return rows


def build_curative_manifest(source: Path, out_dir: Path) -> dict:
    rows = curative_rows(source)
    cols = ["tract", "legal", "owner", "stated_net_acres", "ogl_or_bkpg",
            "royalty", "what_is_needed", "okcountyrecords_search"]
    out_csv = out_dir / f"Section31_Curative_Manifest_{DATED}.csv"
    out_xlsx = out_dir / f"Section31_Curative_Manifest_{DATED}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader(); w.writerows(rows)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Curative"
    _style_header(ws, cols, fill="7F6000")
    for i, row in enumerate(rows, 2):
        for c, name in enumerate(cols, 1):
            cell = ws.cell(i, c, row[name])
            if name == "okcountyrecords_search" and row[name]:
                cell.hyperlink = row[name]; cell.font = Font(color="0563C1", underline="single")
    for c, w in {1: 6, 2: 22, 3: 34, 4: 22, 5: 22, 6: 14, 7: 24, 8: 40}.items():
        ws.column_dimensions[L(c)].width = w
    ws.auto_filter.ref = f"A1:{L(len(cols))}{len(rows)+1}"
    wb.save(out_xlsx)
    return {"csv": str(out_csv), "xlsx": str(out_xlsx), "items": len(rows)}


def build_all(source: Path, out_dir: Path = config.OUTPUT_DIR,
              prefer: str = "auto") -> dict:
    out_dir = Path(out_dir)
    config.ensure_dirs()
    from ..chain import report as chain_report
    from . import title_report
    return {
        "dated_workbook": build_dated_workbook(source, out_dir, prefer=prefer),
        "ownership": build_ownership_reconciliation(source, out_dir),
        "curative": build_curative_manifest(source, out_dir),
        "chain": chain_report.build(source, out_dir),
        "title_report": title_report.build(source, out_dir),
    }


if __name__ == "__main__":
    import json, sys
    src = Path(sys.argv[1])
    res = build_all(src)
    print(json.dumps(res, indent=2, default=str))
