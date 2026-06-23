"""Export the missing/conflict instruments as an EXTERNAL pull-list.

Writes a CSV and a standalone .xlsx (a brand-new workbook) so you can work the
list directly. This NEVER touches the title report workbook or its tabs — it is
a separate file written to the output folder.
"""
from __future__ import annotations

import csv
import datetime as _dt
import re
from pathlib import Path
from urllib.parse import quote

from .. import config
from ..db import store

COLUMNS = [
    "queue_id", "diff_kind", "status", "index_page", "confidence",
    "doc_type (index)", "book_page", "grantor", "grantee",
    "runsheet_row", "okcountyrecords_search", "document_link", "notes",
]


def _idx_id_from_key(dedup_key: str | None) -> int | None:
    if not dedup_key:
        return None
    m = re.search(r"idx(\d+)$", dedup_key)
    return int(m.group(1)) if m else None


def _search_url(grantor: str | None, book_page: str | None) -> str:
    """A *hint* search URL for the user's own logged-in session. Not a deep link."""
    q = (grantor or "").strip() or (book_page or "").strip()
    base = "https://okcountyrecords.com/search/roger%20mills"
    return f"{base}?q={quote(q)}" if q else base


def _rows() -> list[dict]:
    items = store.fetch_all(
        "SELECT id,dedup_key,diff_kind,status,book_page,doc_type,grantor,grantee,"
        "runsheet_row,document_url FROM work_queue "
        "WHERE source='index' AND diff_kind IN ('missing','conflict') "
        "ORDER BY (diff_kind='missing') DESC, id")
    # confidence/page live on index_rows
    idx = {r["id"]: r for r in store.fetch_all(
        "SELECT id,page_number,confidence FROM index_rows")}
    out = []
    for it in items:
        ix = idx.get(_idx_id_from_key(it["dedup_key"]), {})
        out.append({
            "queue_id": it["id"],
            "diff_kind": it["diff_kind"],
            "status": it["status"],
            "index_page": ix.get("page_number"),
            "confidence": ix.get("confidence"),
            "doc_type (index)": it["doc_type"],
            "book_page": it["book_page"],
            "grantor": it["grantor"],
            "grantee": it["grantee"],
            "runsheet_row": it["runsheet_row"],
            "okcountyrecords_search": _search_url(it["grantor"], it["book_page"]),
            "document_link": it["document_url"] or "",
            "notes": "",
        })
    return out


def export(out_dir: Path = config.OUTPUT_DIR) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    rows = _rows()
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = out_dir / f"section31_pull_list_{ts}.csv"
    xlsx_path = out_dir / f"section31_pull_list_{ts}.xlsx"

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)

    _write_xlsx(rows, xlsx_path)
    store.audit("pull_list_exported",
                f"{len(rows)} rows -> {csv_path.name}, {xlsx_path.name}")
    return {"rows": len(rows), "csv": str(csv_path), "xlsx": str(xlsx_path)}


def _write_xlsx(rows: list[dict], path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter as L

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pull List"
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    miss_fill = PatternFill("solid", fgColor="FCE4D6")   # missing -> light orange
    conf_fill = PatternFill("solid", fgColor="FFF2CC")   # conflict -> light yellow

    for c, name in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{L(len(COLUMNS))}{len(rows) + 1}"

    for r, row in enumerate(rows, 2):
        for c, name in enumerate(COLUMNS, 1):
            v = row.get(name)
            cell = ws.cell(r, c, v)
            if name == "okcountyrecords_search" and v:
                cell.hyperlink = v
                cell.font = Font(color="0563C1", underline="single")
            if row["diff_kind"] == "missing":
                cell.fill = miss_fill
            elif row["diff_kind"] == "conflict":
                cell.fill = conf_fill

    widths = {"grantor": 26, "grantee": 26, "okcountyrecords_search": 40,
              "doc_type (index)": 16, "notes": 30, "document_link": 30}
    for c, name in enumerate(COLUMNS, 1):
        ws.column_dimensions[L(c)].width = widths.get(name, 13)
    wb.save(path)


if __name__ == "__main__":
    config.ensure_dirs()
    store.init()
    res = export()
    print(res)
