"""Extract ownership two ways and reconcile them:

  * LEDGER ownership  — the current owners computed inside each Tract tab's
    signed conveyance ledger (these already balance to 100% of the tract).
  * TITLE ownership    — the curated owners/NMA stated on the Title tab.

We do NOT decide which is "right" (that needs the source documents). We surface
both and the gap, with confidence tags, so a human resolves it.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _s(v):
    return "" if v is None else str(v).strip()


def _find_owner_block(ws):
    for r in range(1, min(20, ws.max_row + 1)):
        for c in range(1, 10):
            if _s(ws.cell(r, c).value).upper() == "OWNERS":
                return r, c
    return None, None


def ledger_ownership(ws) -> dict:
    """Current owners (net interest > 0) from a Tract tab ledger."""
    hdr_r, owners_c = _find_owner_block(ws)
    if hdr_r is None:
        return {"tract_acres": None, "owners": [], "balanced": None}
    net_ac_c, net_int_c = owners_c - 1, owners_c + 2
    total_c = owners_c + 1

    tract_acres = None
    for r in range(1, hdr_r + 2):
        for c in range(1, 8):
            if _s(ws.cell(r, c).value).upper() == "TRACT ACRES":
                tract_acres = _num(ws.cell(r, c + 2).value) or _num(ws.cell(r, c + 1).value)

    owners, sum_total = [], 0.0
    for r in range(hdr_r + 1, ws.max_row + 1):
        owner = _s(ws.cell(r, owners_c).value)
        if not owner:
            continue
        tot = _num(ws.cell(r, total_c).value)
        if tot is not None:
            sum_total += tot
        nint = _num(ws.cell(r, net_int_c).value)
        nac = _num(ws.cell(r, net_ac_c).value)
        if nint is not None and nint > 1e-9:
            owners.append({"owner": owner, "net_acres": nac, "net_interest": nint})
    owners.sort(key=lambda o: -(o["net_acres"] or 0))
    return {"tract_acres": tract_acres, "owners": owners,
            "net_acres_traced": round(sum(o["net_acres"] or 0 for o in owners), 6),
            "balanced": abs(sum_total) < 1e-3}


# Title tab tract sections (parsed by "TRACT n:" / "REPORT TOTAL").
def title_ownership(ws) -> list[dict]:
    sections, cur = [], None
    for r in range(1, ws.max_row + 1):
        b = _s(ws.cell(r, 2).value)
        c = ws.cell(r, 3).value
        if b.upper().startswith("TRACT ") and b.rstrip().endswith(":"):
            cur = {"tract": b, "legal": _s(c), "owners": [], "report_total": None,
                   "start_row": r}
            sections.append(cur)
        elif cur is not None and b.upper() in ("REPORT TOTAL", "TOTAL"):
            cur["report_total"] = _num(c)
        elif cur is not None and b and b not in ("Mineral Owner", "Containing:"):
            # an owner row: net acres in col C may be numeric or a VERIFY string
            owners = cur["owners"]
            owners.append({
                "row": r,
                "owner": b.replace("\n", " ")[:60],
                "net_acres": _num(c),
                "net_acres_raw": _s(c),
                "ogl": _s(ws.cell(r, 4).value),
                "royalty": _s(ws.cell(r, 5).value),
                "needs": ("VERIFY" in _s(c).upper() or "NEED" in _s(c).upper()),
            })
    return sections


def consolidate_owners(owner_list: list[dict]) -> list[dict]:
    """Cluster owners by resolved identity (M. G. Mitchell == Marvin G. Mitchell),
    summing net acres / interest. Representative name = the longest variant."""
    from ..chain.entities import persons, any_match
    clusters: list[dict] = []
    for o in owner_list:
        op = persons(o["owner"])
        hit = None
        for cl in clusters:
            if op and any_match(op, cl["_persons"]):
                hit = cl
                break
        if hit is None:
            clusters.append({"owner": o["owner"], "_persons": op,
                             "variants": {o["owner"]},
                             "net_acres": o.get("net_acres") or 0.0,
                             "net_interest": o.get("net_interest") or 0.0})
        else:
            hit["variants"].add(o["owner"])
            if len(o["owner"]) > len(hit["owner"]):
                hit["owner"], hit["_persons"] = o["owner"], op
            hit["net_acres"] += o.get("net_acres") or 0.0
            hit["net_interest"] += o.get("net_interest") or 0.0
    for cl in clusters:
        cl.pop("_persons", None)
        cl["variants"] = sorted(cl["variants"])
        cl["net_acres"] = round(cl["net_acres"], 6)
        cl["net_interest"] = round(cl["net_interest"], 8)
    clusters.sort(key=lambda c: -c["net_acres"])
    return clusters


def reconcile(workbook: Path) -> dict:
    wb = openpyxl.load_workbook(workbook, data_only=True, keep_links=True)
    tracts = []
    for i in range(1, 9):
        name = f"Tract {i}"
        if name not in wb.sheetnames:
            continue
        led = ledger_ownership(wb[name])
        tracts.append({"tract": i, **led})

    title = title_ownership(wb["Title "]) if "Title " in wb.sheetnames else []
    # Align title sections to tract numbers by the actual "TRACT n:" header,
    # not list position — gaps/reordering must not skew reconciliation.
    for idx, sec in enumerate(title):
        m = re.search(r"TRACT\s+(\d+)", str(sec.get("tract", "")), re.I)
        sec["tract_num"] = int(m.group(1)) if m else (idx + 1 if idx < 8 else None)

    from ..chain.entities import persons, any_match
    summary = []
    for t in tracts:
        tnum = t["tract"]
        tsec = next((s for s in title if s.get("tract_num") == tnum), None)
        title_total = tsec["report_total"] if tsec else None
        led_acres = t.get("net_acres_traced")
        gap = None
        if led_acres is not None and title_total is not None:
            gap = round(led_acres - title_total, 4)

        # Entity-resolved current owners + which are already in the Title summary.
        consolidated = consolidate_owners(t["owners"])
        title_persons = [persons(o["owner"]) for o in (tsec["owners"] if tsec else [])]
        missing_from_title = []
        for c in consolidated:
            cp = persons(c["owner"])
            in_title = any(any_match(cp, tp) for tp in title_persons if tp)
            c["in_title"] = in_title
            if not in_title:
                missing_from_title.append(c)
        t["consolidated_owners"] = consolidated
        t["missing_from_title"] = missing_from_title

        summary.append({
            "tract": tnum,
            "tract_acres": t.get("tract_acres"),
            "ledger_owners": len(t["owners"]),
            "ledger_owners_resolved": len(consolidated),
            "ledger_net_acres": led_acres,
            "title_total": title_total,
            "title_owners_listed": len(tsec["owners"]) if tsec else 0,
            "title_verify_rows": sum(1 for o in tsec["owners"] if o["needs"]) if tsec else 0,
            "owners_missing_from_title": len(missing_from_title),
            "acres_missing_from_title": round(
                sum(c["net_acres"] for c in missing_from_title), 4),
            "gap_acres": gap,
        })
    return {"tracts": tracts, "title": title, "summary": summary}
