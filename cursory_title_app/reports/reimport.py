"""Re-import reviewed picks back into the Runsheet (human-in-the-loop).

Flow:
  1) template_csv(workbook)  -> an editable CSV of:
        * action=add    : index 'missing' instruments to append (prefilled, low-conf)
        * action=update : existing Runsheet rows carrying NEED/VERIFY flags
  2) You open each document in your browser, paste the real document_link and any
     corrected fields, and set approve=yes on rows you've verified.
  3) apply_csv(workbook, edited.csv) -> writes ONLY approved rows to a NEW dated
     workbook via the format-preserving writer:
        * update rows  -> writable cols A-N/T/U only (never the O..S formulas)
        * add rows     -> appended at the next blank row, A-N/T/U filled AND the
                          O..S formulas copied down from the template row so the
                          acreage/interest math stays live.

Nothing is written unless approve is truthy. No fabrication: blanks stay blank.
"""
from __future__ import annotations

import csv
import datetime as _dt
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator

from .. import config
from ..db import store
from ..excel import verify as xlverify
from ..excel.writer import apply_edits, hyperlink_formula, make_backup
from ..runsheet.columns import FIELD_TO_COLUMN, FORMULA_COLUMNS, FIRST_DATA_ROW

DATED = "(6-25-2026)"
DETAIL = "https://okcountyrecords.com/detail/roger+mills/"

CSV_COLUMNS = [
    "action", "approve", "runsheet_row", "confidence",
    "instrument_number", "ogl", "doc_type", "book_page",
    "effective_date", "recorded_date", "grantor", "grantee",
    "legal_description", "tracts", "conveyance_type", "conveyance_amount_dec",
    "document_link", "notes", "review", "need_action",
]
# CSV field name -> logical runsheet field (writable cols only)
_CSV_TO_FIELD = {
    "instrument_number": "instrument_number", "ogl": "ogl", "doc_type": "doc_type",
    "book_page": "book_page", "effective_date": "effective_date",
    "recorded_date": "recorded_date", "grantor": "grantor", "grantee": "grantee",
    "legal_description": "legal_description", "tracts": "tracts",
    "conveyance_type": "conveyance_type",
    "conveyance_amount_dec": "conveyance_amount_dec",
    "notes": "notes", "review": "review", "need_action": "need_action",
}
_TRUTHY = {"y", "yes", "true", "1", "x", "approve", "approved"}


def _last_data_row(ws) -> int:
    last = 1
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in (1, 4, 7, 8, 9)):
            last = r
    return last


def _parse_date(s):
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d", "%m/%d/%y"):
        try:
            return _dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s  # leave as text; caller adds a VERIFY flag


# --------------------------------------------------------------------------- #
def template_csv(workbook: Path, out_csv: Path) -> dict:
    wb = openpyxl.load_workbook(workbook, data_only=False, keep_links=True)
    ws = wb["Runsheet"]
    last = _last_data_row(ws)
    rows: list[dict] = []

    # action=add : index 'missing' instruments (if any have been extracted)
    for q in store.fetch_all(
            "SELECT book_page, doc_type, grantor, grantee, document_url "
            "FROM work_queue WHERE diff_kind='missing' ORDER BY id"):
        rows.append({c: "" for c in CSV_COLUMNS} | {
            "action": "add", "approve": "no", "doc_type": q["doc_type"] or "",
            "book_page": q["book_page"] or "", "grantor": q["grantor"] or "",
            "grantee": q["grantee"] or "", "document_link": q["document_url"] or "",
            "confidence": "low",
        })

    # action=update : existing rows carrying NEED/VERIFY flags
    for r in range(FIRST_DATA_ROW, last + 1):
        t = str(ws.cell(r, 20).value or "")
        u = str(ws.cell(r, 21).value or "")
        if "NEED" in (t + u).upper() or "VERIFY" in (t + u).upper():
            rows.append({c: "" for c in CSV_COLUMNS} | {
                "action": "update", "approve": "no", "runsheet_row": r,
                "instrument_number": ws.cell(r, 1).value or "",
                "doc_type": ws.cell(r, 3).value or "",
                "book_page": ws.cell(r, 4).value or "",
                "grantor": ws.cell(r, 7).value or "",
                "grantee": ws.cell(r, 8).value or "",
                "review": t, "need_action": u, "document_link": "",
            })

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        w.writeheader()
        w.writerows(rows)
    return {"csv": str(out_csv), "add_rows": sum(1 for x in rows if x["action"] == "add"),
            "update_rows": sum(1 for x in rows if x["action"] == "update")}


# --------------------------------------------------------------------------- #
def _value_edits_for_fields(rowdict: dict, target_row: int) -> list[dict]:
    """Build edits for writable A-N/T/U from a CSV row (skips blanks)."""
    edits = []
    for csv_field, logical in _CSV_TO_FIELD.items():
        raw = (rowdict.get(csv_field) or "").strip()
        if raw == "":
            continue
        col = FIELD_TO_COLUMN[logical]
        if col in FORMULA_COLUMNS:                  # never via this path
            raise PermissionError(f"Refusing formula column {col}.")
        val = raw
        if logical in ("effective_date", "recorded_date"):
            val = _parse_date(raw)
        elif logical == "conveyance_amount_dec":
            try:
                val = float(raw)
            except ValueError:
                val = raw
        edits.append({"sheet": "Runsheet", "cell": f"{col}{target_row}",
                      "value": val, "is_formula": False})
    # document_link -> HYPERLINK in K
    link = (rowdict.get("document_link") or "").strip()
    if link:
        if link.lower().startswith("=hyperlink"):
            formula = link
        elif link.lower().startswith("http"):
            formula = hyperlink_formula(link, rowdict.get("book_page") or "Document")
        else:
            formula = hyperlink_formula(DETAIL + link, link)
        edits.append({"sheet": "Runsheet", "cell": f"K{target_row}",
                      "value": formula, "is_formula": True})
    return edits


def apply_csv(workbook: Path, csv_path: Path, out_dir: Path = config.OUTPUT_DIR,
              prefer: str = "auto") -> dict:
    workbook = Path(workbook)
    make_backup(workbook, config.BACKUP_DIR)

    wb = openpyxl.load_workbook(workbook, data_only=False, keep_links=True)
    ws = wb["Runsheet"]
    next_row = _last_data_row(ws) + 1
    tmpl_row = FIRST_DATA_ROW   # row 2 holds the O..S formula pattern

    # capture template formulas for O..S
    tmpl_formulas = {}
    for col in FORMULA_COLUMNS:
        v = ws[f"{col}{tmpl_row}"].value
        if isinstance(v, str) and v.startswith("="):
            tmpl_formulas[col] = v

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = list(csv.DictReader(f))

    max_row = ws.max_row
    edits: list[dict] = []
    added, updated, skipped, rejected = [], [], 0, []
    for row in reader:
        if (row.get("approve") or "").strip().lower() not in _TRUTHY:
            skipped += 1
            continue
        action = (row.get("action") or "update").strip().lower()

        if action == "update":
            try:
                tr = int(float(row["runsheet_row"]))
            except (KeyError, ValueError, TypeError):
                skipped += 1
                continue
            # Never write above the first data row (would corrupt the header) or
            # past the current data range.
            if tr < FIRST_DATA_ROW or tr > max_row:
                rejected.append({"runsheet_row": tr, "reason": "out of data range"})
                continue
            edits += _value_edits_for_fields(row, tr)
            updated.append(tr)

        elif action == "add":
            tr = next_row
            next_row += 1
            edits += _value_edits_for_fields(row, tr)
            # copy the O..S formulas down so the math computes on the new row
            for col, tf in tmpl_formulas.items():
                new_f = Translator(tf, origin=f"{col}{tmpl_row}") \
                    .translate_formula(f"{col}{tr}")
                edits.append({"sheet": "Runsheet", "cell": f"{col}{tr}",
                              "value": new_f, "is_formula": True})
            added.append(tr)
        else:
            skipped += 1

    out = out_dir / f"31-12N-24W_Roger_Mills_Cursory_Title_Report_{DATED}.REIMPORT.xlsx"
    backend = apply_edits(workbook, edits, out, prefer=prefer)
    res = xlverify.verify_workbook(out, workbook)
    store.audit("reimport_applied",
                f"added={len(added)} updated={len(updated)} skipped={skipped}",
                {"backend": backend, "verify_ok": res["ok"]})
    return {"output": str(out), "backend": backend,
            "rows_added": added, "rows_updated": updated, "rows_skipped": skipped,
            "rows_rejected": rejected,
            "cells_written": len(edits), "verify_ok": res["ok"], "verify": res}
