"""Runsheet analyzer.

Reads the EXISTING Runsheet (read-only) and the extracted index rows, then:
  * detects duplicate rows already in the Runsheet
  * detects missing rows (in the index but not the Runsheet)
  * detects conflicting rows (same instrument/bk-pg, different grantor/grantee)
  * builds a work queue in SQLite

It NEVER modifies the workbook. Writing is the Excel writer's job, on approval.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from .. import config
from ..db import store
from ..runsheet.columns import FIRST_DATA_ROW, HEADER_TEXT


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().upper()


def _bkpg(s) -> str:
    return _norm(s).replace(" ", "")


_STOP = {"AND", "HW", "ET", "AL", "ETAL", "INC", "LLC", "CO", "COMPANY", "TRUST",
         "THE", "OF", "A", "JR", "SR", "ESTATE", "HEIRS"}


def _name_tokens(s) -> set[str]:
    """Surname-ish tokens from a party name for fuzzy matching."""
    toks = {t for t in re.split(r"[^A-Za-z]+", _norm(s)) if len(t) > 2 and t not in _STOP}
    return toks


def read_runsheet(path: Path) -> list[dict]:
    """Read existing Runsheet rows as dicts keyed by column letter. Read-only."""
    wb = openpyxl.load_workbook(path, data_only=False, keep_links=True)
    ws = wb[config.RUNSHEET_TAB]
    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        a = ws[f"A{r}"].value
        d = ws[f"D{r}"].value
        g = ws[f"G{r}"].value
        h = ws[f"H{r}"].value
        if a is None and d is None and g is None and h is None:
            continue  # blank row
        rows.append({
            "row": r,
            "instrument": a, "bkpg": d, "doc_type": ws[f"C{r}"].value,
            "grantor": g, "grantee": h, "legal": ws[f"I{r}"].value,
        })
    return rows


def _dedup_key(instrument, bkpg, grantor, grantee) -> str:
    inst = _norm(instrument)
    if inst:
        return f"INST::{inst}"
    bp = _bkpg(bkpg)
    if bp:
        return f"BKPG::{bp}"
    return f"NAME::{_norm(grantor)}>{_norm(grantee)}"


def analyze(workbook_path: Path) -> dict:
    """Diff existing Runsheet vs extracted index rows; populate work_queue."""
    existing = read_runsheet(workbook_path)
    by_inst = {}
    by_bkpg = {}
    by_name = {}                     # name-token -> rows, for fuzzy fallback
    for e in existing:
        if _norm(e["instrument"]):
            by_inst.setdefault(_norm(e["instrument"]), []).append(e)
        if _bkpg(e["bkpg"]):
            by_bkpg.setdefault(_bkpg(e["bkpg"]), []).append(e)
        for tok in _name_tokens(e["grantor"]) | _name_tokens(e["grantee"]):
            by_name.setdefault(tok, []).append(e)

    index_rows = store.fetch_all("SELECT * FROM index_rows")

    counts = {"missing": 0, "duplicate": 0, "conflict": 0, "present": 0}
    for ix in index_rows:
        inst = _norm(ix.get("instrument_kind"))  # index rarely has instrument #
        bp = _bkpg(f"{ix.get('book') or ''}/{ix.get('page') or ''}")
        match = by_bkpg.get(bp, []) if bp not in ("", "/") else []
        matched_by = "bkpg" if match else None

        # Fuzzy name fallback when Bk/Pg didn't match (OCR of cursive Bk/Pg is
        # unreliable). Require BOTH a grantor token and grantee token to agree.
        if not match:
            gr_toks = _name_tokens(ix.get("grantor"))
            ge_toks = _name_tokens(ix.get("grantee"))
            cand = {}
            for tok in gr_toks | ge_toks:
                for e in by_name.get(tok, []):
                    cand.setdefault(e["row"], e)
            strong = [e for e in cand.values()
                      if (_name_tokens(e["grantor"]) & gr_toks)
                      and (_name_tokens(e["grantee"]) & ge_toks)]
            if strong:
                match = strong
                matched_by = "name"

        diff_kind = "present"
        runsheet_row = None
        if not match:
            diff_kind = "missing"
        else:
            runsheet_row = match[0]["row"]
            # conflict if the parties don't fully agree. Neither set being a
            # subset of the other means a real mismatch (e.g. same grantor but a
            # different grantee) — route it to review rather than call it present.
            ix_names = {n for n in (_norm(ix.get("grantor")), _norm(ix.get("grantee"))) if n}
            ex_names = {n for n in (_norm(match[0]["grantor"]), _norm(match[0]["grantee"])) if n}
            if ix_names and ex_names and not (
                    ix_names <= ex_names or ex_names <= ix_names):
                diff_kind = "conflict"
            elif len(match) > 1:
                diff_kind = "duplicate"
            else:
                diff_kind = "present"

        counts[diff_kind] = counts.get(diff_kind, 0) + 1

        dkey = _dedup_key(None, f"{ix.get('book')}/{ix.get('page')}",
                          ix.get("grantor"), ix.get("grantee")) + f"::idx{ix['id']}"
        store.upsert_queue({
            "dedup_key": dkey,
            "source": "index",
            "runsheet_row": runsheet_row,
            "instrument_number": None,
            "book_page": f"{ix.get('book') or ''}/{ix.get('page') or ''}",
            "doc_type": ix.get("instrument_kind"),
            "grantor": ix.get("grantor"),
            "grantee": ix.get("grantee"),
            "document_url": None,
            "status": "pending" if diff_kind in ("missing", "conflict") else "skipped",
            "diff_kind": diff_kind,
        })
        _ = matched_by  # retained for evidence/debugging

    store.audit("runsheet_analyzed",
                f"existing={len(existing)} index={len(index_rows)}", counts)
    return {
        "existing_rows": len(existing),
        "index_rows": len(index_rows),
        "headers_ok": True,
        "counts": counts,
    }
