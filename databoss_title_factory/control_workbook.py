"""Separate examiner/control workbook generation.

Audit, confidence, model, prompt, and review data never enter the client report.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable, Sequence


def _safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _rows(value: Any) -> list[dict[str, Any]]:
    return value if isinstance(value, list) else []


def _write_sheet(
    workbook: Any,
    name: str,
    rows: Iterable[dict[str, Any]],
    fields: Sequence[str],
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(name[:31])
    sheet.append([field.replace("_", " ").title() for field in fields])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="163A3D")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        sheet.append([_safe(row.get(field, "")) for field in fields])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, field in enumerate(fields, start=1):
        width = 52 if field in {
            "citation", "text", "field_provenance", "candidate_values",
            "differences", "reconciliation_notes", "missing_or_issue",
        } else 22
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_control_workbook(
    run_dir: str | Path,
    destination: str | Path,
    preservation_audit: dict[str, Any],
    project_metadata: dict[str, Any],
) -> Path:
    import openpyxl

    run_path = Path(run_dir)
    output = Path(destination)
    output.parent.mkdir(parents=True, exist_ok=True)

    def read_json(name: str, default: Any) -> Any:
        path = run_path / name
        return json.loads(path.read_text(encoding="utf-8")) if path.exists() else default

    def read_jsonl(name: str) -> list[dict[str, Any]]:
        path = run_path / name
        if not path.exists():
            return []
        return [
            json.loads(line)
            for line in path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]

    inventory = _rows(read_json("file_inventory.json", []))
    ocr = read_jsonl("ocr_citations.jsonl")
    instruments = _rows(read_json("instruments.json", []))
    bundle = read_json("runsheet_bundle.json", {})
    candidates = read_jsonl("candidates/candidate_archive.jsonl")
    conflicts = _rows(read_json("field_conflicts.json", []))
    reconciliation = read_jsonl("reconciliation_audit.jsonl")
    missing = _rows(bundle.get("missing_documents", [])) if isinstance(bundle, dict) else []
    review = [
        row for row in instruments
        if str(row.get("status", "")).upper() != "ACCEPTED"
    ]

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    _write_sheet(
        workbook,
        "Run Control",
        [{"key": key, "value": value} for key, value in project_metadata.items()],
        ("key", "value"),
    )
    _write_sheet(
        workbook,
        "File Manifest",
        inventory,
        (
            "source_path", "extension", "category", "probable_purpose",
            "size_bytes", "sha256", "duplicate_hash_group", "page_count",
            "processing_status", "ocr_status", "extraction_status", "review_status",
        ),
    )
    _write_sheet(
        workbook,
        "OCR Provenance",
        ocr,
        (
            "citation", "source_file_hash", "source_path", "source_locator",
            "derived_image_path", "preprocessing_recipe", "method", "engine_version",
            "ocr_config", "confidence", "status", "blocks", "text",
        ),
    )
    _write_sheet(
        workbook,
        "Candidate Archive",
        candidates,
        (
            "candidate_id", "engine", "status", "sha256", "citation",
            "instrument_number", "book", "page", "grantor", "grantee",
            "legal_description", "confidence", "field_provenance",
        ),
    )
    _write_sheet(
        workbook,
        "Reconciled Instruments",
        instruments,
        (
            "entry_no", "status", "confidence", "instrument_number",
            "instrument_type", "book", "page", "grantor", "grantee",
            "legal_description", "citation", "field_provenance",
            "reconciliation_notes",
        ),
    )
    _write_sheet(
        workbook,
        "Field Conflicts",
        conflicts,
        (
            "match_id", "field", "material", "decision", "candidate_values",
            "source_support", "review_reason",
        ),
    )
    _write_sheet(
        workbook,
        "Reconciliation Audit",
        reconciliation,
        ("match_id", "status", "candidates", "field_decisions"),
    )
    _write_sheet(
        workbook,
        "Review Queue",
        review,
        (
            "entry_no", "status", "instrument_number", "book", "page",
            "citation", "reconciliation_notes",
        ),
    )
    _write_sheet(
        workbook,
        "Missing Documents",
        missing,
        ("instrument_number", "missing_or_issue", "citation", "recommended_action"),
    )
    _write_sheet(
        workbook,
        "Workbook Audit",
        [
            {
                "preservation_passed": preservation_audit.get("preservation_passed"),
                "binary_identical": preservation_audit.get("binary_identical"),
                "high_risk_features_present": preservation_audit.get(
                    "high_risk_features_present"
                ),
                "claim": preservation_audit.get("claim"),
                "differences": preservation_audit.get("differences", []),
            }
        ],
        (
            "preservation_passed", "binary_identical", "high_risk_features_present",
            "claim", "differences",
        ),
    )
    workbook.save(output)
    workbook.close()
    return output
