"""Local, non-destructive processing pipeline for DataBoss Title Factory.

Every source is read-only. Generated files are written to a separate,
configurable workspace and weak results are copied into that run's quarantine
area for examiner review.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import importlib.metadata
import json
import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

OUTPUT_DIR_NAME = "DataBoss_Title_Factory_Output"
SKIP_DIRS = {
    ".git", ".venv", "venv", "__pycache__", "node_modules", OUTPUT_DIR_NAME,
    "_DATABOSS_BACKUPS",
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}
PDF_EXTENSIONS = {".pdf"}
TEXT_EXTENSIONS = {".txt", ".md", ".log", ".text"}
TABLE_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm"}
SUPPORTED_EXTENSIONS = IMAGE_EXTENSIONS | PDF_EXTENSIONS | TEXT_EXTENSIONS | TABLE_EXTENSIONS

INSTRUMENT_FIELDS = (
    "instrument_number", "instrument_type", "instrument_date", "recorded_date",
    "book", "page", "grantor", "grantee", "legal_description",
    "interest_conveyed", "lease_royalty_terms",
)
EXPORT_FIELDS = (
    "entry_no", "instrument_date", "recorded_date", "instrument_type",
    "grantor", "grantee", "instrument_number", "book", "page",
    "legal_description", "interest_conveyed", "lease_royalty_terms",
    "confidence", "status", "citation", "source_path", "source_locator",
    "reconciliation_notes",
)


@dataclass(frozen=True)
class RunContext:
    root: Path
    output_dir: Path
    run_dir: Path
    run_id: str

    @property
    def quarantine_dir(self) -> Path:
        return self.run_dir / "quarantine"

    @property
    def preprocessed_dir(self) -> Path:
        return self.run_dir / "preprocessed"


def _now_id() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")


def _output_dir(root_path: Path, output_root: str | Path | None = None) -> Path:
    """Return a stable generated workspace that is never inside the source."""
    configured = output_root or os.environ.get("DATABOSS_OUTPUT_ROOT")
    workspace = (
        Path(configured).expanduser().resolve()
        if configured
        else Path(__file__).resolve().parent.parent / "runtime" / "title-intelligence"
    )
    source_key = hashlib.sha256(str(root_path).encode("utf-8")).hexdigest()[:12]
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", root_path.name).strip("._") or "project"
    output = workspace / f"{safe_name}-{source_key}" / OUTPUT_DIR_NAME
    if output == root_path or output.is_relative_to(root_path):
        raise ValueError("Generated output workspace must be outside the source folder")
    return output


def start_run(root: str | Path, output_root: str | Path | None = None) -> RunContext:
    """Create a new immutable run folder and mark it as the latest run."""
    root_path = Path(root).expanduser().resolve()
    if not root_path.is_dir():
        raise ValueError(f"Project folder does not exist: {root_path}")
    output = _output_dir(root_path, output_root)
    run_id = _now_id()
    run_dir = output / "runs" / run_id
    run_dir.mkdir(parents=True, exist_ok=False)
    (run_dir / "quarantine").mkdir()
    output.mkdir(parents=True, exist_ok=True)
    _atomic_text(output / "latest_run.txt", run_id)
    _write_json(
        run_dir / "run_manifest.json",
        {
            "run_id": run_id,
            "project_root": str(root_path),
            "created_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "source_policy": "read-only; generated results are versioned; nothing is deleted",
        },
    )
    from .project_db import register_run

    register_run(output, run_id, str(root_path))
    return RunContext(root_path, output, run_dir, run_id)


def latest_run(root: str | Path, output_root: str | Path | None = None) -> RunContext:
    root_path = Path(root).expanduser().resolve()
    output = _output_dir(root_path, output_root)
    pointer = output / "latest_run.txt"
    if not pointer.exists():
        raise FileNotFoundError("No run exists. Run Inventory first.")
    run_id = pointer.read_text(encoding="utf-8").strip()
    run_dir = output / "runs" / run_id
    if not run_dir.is_dir():
        raise FileNotFoundError(f"Latest run folder is missing: {run_dir}")
    return RunContext(root_path, output, run_dir, run_id)


def resume_pipeline(
    ctx: RunContext,
    weak_threshold: float = 0.65,
    cursor_json: Optional[str | Path] = None,
    codex_json: Optional[str | Path] = None,
) -> dict[str, Any]:
    """Continue the same run from its last validated artifact checkpoint."""
    from .project_db import (
        artifact_checkpoint_valid,
        is_pause_requested,
        request_pause,
        stage_statuses,
    )

    request_pause(ctx.output_dir, ctx.run_id, False)
    statuses = {row["stage"]: row["status"] for row in stage_statuses(ctx.output_dir, ctx.run_id)}
    stages = (
        (
            "INVENTORY",
            ctx.run_dir / "file_inventory.json",
            lambda: build_inventory(ctx),
            ("file_inventory.json", "project_manifest.csv", "duplicate_map.csv"),
        ),
        (
            "OCR",
            ctx.run_dir / "ocr_citations.jsonl",
            lambda: run_ocr(ctx, weak_threshold=weak_threshold),
            ("ocr_citations.jsonl", "ocr_attempts.jsonl"),
        ),
        (
            "RECONCILE",
            ctx.run_dir / "instruments.json",
            lambda: extract_and_reconcile(
                ctx,
                weak_threshold=weak_threshold,
                cursor_json=cursor_json,
                codex_json=codex_json,
            ),
            (
                "instruments.json", "field_conflicts.json",
                "reconciliation_audit.jsonl",
            ),
        ),
        (
            "RUNSHEET",
            ctx.run_dir / "runsheet_bundle.json",
            lambda: build_runsheet(ctx),
            ("runsheet_bundle.json", "runsheet.csv", "missing_documents.csv"),
        ),
    )
    completed: list[str] = []
    skipped: list[str] = []
    rerun_downstream = False
    for stage, _required_artifact, action, artifacts in stages:
        if (
            not rerun_downstream
            and statuses.get(stage) in {"PASSED", "READY FOR REVIEW"}
            and all(
                artifact_checkpoint_valid(
                    ctx.output_dir,
                    ctx.run_id,
                    stage,
                    ctx.run_dir / relative_path,
                )
                for relative_path in artifacts
            )
        ):
            skipped.append(stage)
            continue
        if is_pause_requested(ctx.output_dir, ctx.run_id):
            return {
                "run_id": ctx.run_id,
                "status": "PAUSED",
                "completed": completed,
                "skipped": skipped,
            }
        _archive_retry_artifacts(ctx, stage, artifacts)
        action()
        completed.append(stage)
        rerun_downstream = True
    return {
        "run_id": ctx.run_id,
        "status": "READY FOR REVIEW",
        "completed": completed,
        "skipped": skipped,
    }


def _archive_retry_artifacts(
    ctx: RunContext,
    stage: str,
    relative_paths: Sequence[str],
) -> None:
    existing = [ctx.run_dir / relative for relative in relative_paths if (ctx.run_dir / relative).exists()]
    if not existing:
        return
    archive = ctx.run_dir / "retry_archive" / f"{_now_id()}_{stage.lower()}"
    archive.mkdir(parents=True, exist_ok=False)
    for source in existing:
        destination = archive / source.relative_to(ctx.run_dir)
        destination.parent.mkdir(parents=True, exist_ok=True)
        if source.is_file():
            shutil.copy2(source, destination)


def _record_stage_artifacts(
    ctx: RunContext,
    stage: str,
    paths: Sequence[Path],
) -> None:
    from .project_db import record_artifact

    for path in paths:
        if path.is_file():
            record_artifact(ctx.output_dir, ctx.run_id, stage, path)


def _atomic_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.{_now_id()}.tmp")
    temp.write_text(text, encoding="utf-8")
    temp.replace(path)


def _write_json(path: Path, value: Any) -> None:
    _atomic_text(path, json.dumps(value, indent=2, ensure_ascii=False, default=str))


def _read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def _write_csv(path: Path, rows: Sequence[dict[str, Any]], fields: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.{_now_id()}.tmp")
    with temp.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fields), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _safe_cell(row.get(field, "")) for field in fields})
    temp.replace(path)


def _sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(chunk_size), b""):
            digest.update(block)
    return digest.hexdigest()


def _assert_inventory_hash(path: Path, expected_hash: str) -> None:
    current_hash = _sha256(path)
    if not expected_hash or current_hash != expected_hash:
        raise RuntimeError(
            "SOURCE HASH MISMATCH: file changed after inventory; processing blocked "
            f"(expected={expected_hash or 'missing'}, current={current_hash})"
        )


def build_inventory(ctx: RunContext) -> list[dict[str, Any]]:
    """Inventory every source file without modifying the source tree."""
    from .project_db import update_stage

    update_stage(ctx.output_dir, ctx.run_id, "INVENTORY", "IN PROGRESS")
    records: list[dict[str, Any]] = []
    for dirpath, dirnames, filenames in os.walk(ctx.root):
        dirnames[:] = sorted(
            name for name in dirnames
            if name not in SKIP_DIRS
            and not name.startswith("FINAL_DATABOSS_OUTPUT_")
            and not (Path(dirpath) / name).is_symlink()
        )
        parent = Path(dirpath)
        for filename in sorted(filenames):
            path = parent / filename
            if path.is_symlink() or not path.is_file():
                continue
            try:
                stat = path.stat()
                rel = path.relative_to(ctx.root).as_posix()
                records.append(
                    {
                        "source_path": rel,
                        "absolute_path": str(path.resolve()),
                        "folder": path.parent.relative_to(ctx.root).as_posix(),
                        "filename": filename,
                        "extension": path.suffix.lower(),
                        "category": _file_category(path),
                        "probable_purpose": _probable_purpose(path),
                        "size_bytes": stat.st_size,
                        "modified_at": dt.datetime.fromtimestamp(
                            stat.st_mtime, tz=dt.timezone.utc
                        ).isoformat(),
                        "sha256": _sha256(path),
                        "duplicate_hash_group": "",
                        "page_count": _page_count(path),
                        "supported": path.suffix.lower() in SUPPORTED_EXTENSIONS,
                        "processing_status": "INVENTORIED",
                        "ocr_status": "NOT STARTED",
                        "extraction_status": "NOT STARTED",
                        "review_status": "NOT STARTED",
                        "status": "inventoried",
                    }
                )
            except OSError as exc:
                records.append(
                    {
                        "source_path": str(path),
                        "absolute_path": str(path),
                        "folder": "",
                        "filename": filename,
                        "extension": path.suffix.lower(),
                        "category": "scan_error",
                        "probable_purpose": "unavailable",
                        "size_bytes": "",
                        "modified_at": "",
                        "sha256": "",
                        "duplicate_hash_group": "",
                        "page_count": "",
                        "supported": False,
                        "processing_status": "FAILED",
                        "ocr_status": "BLOCKED",
                        "extraction_status": "BLOCKED",
                        "review_status": "READY FOR REVIEW",
                        "status": f"scan_error: {exc}",
                    }
                )
    hash_counts: dict[str, int] = {}
    for record in records:
        if record["sha256"]:
            hash_counts[record["sha256"]] = hash_counts.get(record["sha256"], 0) + 1
    group_number = 0
    duplicate_rows: list[dict[str, Any]] = []
    for digest in sorted(key for key, count in hash_counts.items() if count > 1):
        group_number += 1
        group_id = f"DUP-{group_number:04d}"
        for record in records:
            if record["sha256"] == digest:
                record["duplicate_hash_group"] = group_id
                duplicate_rows.append(record)
    fields = (
        "source_path", "absolute_path", "folder", "filename", "extension",
        "category", "probable_purpose", "size_bytes", "modified_at", "sha256",
        "duplicate_hash_group", "page_count", "supported", "processing_status",
        "ocr_status", "extraction_status", "review_status", "status",
    )
    _write_json(ctx.run_dir / "file_inventory.json", records)
    _write_csv(ctx.run_dir / "file_inventory.csv", records, fields)
    _write_csv(ctx.run_dir / "project_manifest.csv", records, fields)
    _write_csv(ctx.run_dir / "duplicate_map.csv", duplicate_rows, fields)
    _build_report_candidate_scorecard(ctx, records)
    _record_stage_artifacts(
        ctx,
        "INVENTORY",
        (
            ctx.run_dir / "file_inventory.json",
            ctx.run_dir / "project_manifest.csv",
            ctx.run_dir / "duplicate_map.csv",
        ),
    )
    update_stage(
        ctx.output_dir,
        ctx.run_id,
        "INVENTORY",
        "PASSED",
        {"file_count": len(records), "duplicate_file_count": len(duplicate_rows)},
    )
    return records


def _file_category(path: Path) -> str:
    name = path.name.lower()
    extension = path.suffix.lower()
    if extension in IMAGE_EXTENSIONS:
        return "index_image" if "index" in name else "instrument_image"
    if extension in PDF_EXTENSIONS:
        return "index_pdf" if "index" in name else "source_pdf"
    if extension in {".xlsx", ".xlsm"}:
        if "template" in name:
            return "client_template"
        if "runsheet" in name or "run sheet" in name:
            return "runsheet"
        if "ogl" in name or "lease" in name:
            return "ogl_sheet"
        if "wi" in name or "working interest" in name:
            return "wi_sheet"
        if "tract" in name:
            return "tract_sheet"
        return "report_candidate"
    if extension in {".csv", ".tsv"}:
        return "structured_data"
    if extension in TEXT_EXTENSIONS:
        return "ocr_or_research_text"
    return "supporting_or_unrelated"


def _probable_purpose(path: Path) -> str:
    category = _file_category(path)
    return {
        "client_template": "formatting authority; never edit in place",
        "report_candidate": "candidate working report; score before selection",
        "index_image": "indexed-document register evidence",
        "index_pdf": "indexed-document register evidence",
        "instrument_image": "controlling source image",
        "source_pdf": "controlling source document",
        "runsheet": "existing chronology candidate",
        "tract_sheet": "tract definition candidate",
        "ogl_sheet": "oil-and-gas lease candidate",
        "wi_sheet": "working-interest candidate",
    }.get(category, "supporting material requiring classification")


def _page_count(path: Path) -> int | str:
    try:
        if path.suffix.lower() in IMAGE_EXTENSIONS:
            from PIL import Image

            with Image.open(path) as image:
                return int(getattr(image, "n_frames", 1))
        if path.suffix.lower() in PDF_EXTENSIONS:
            import fitz

            with fitz.open(path) as document:
                return document.page_count
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            import openpyxl

            workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
            count = len(workbook.sheetnames)
            workbook.close()
            return count
    except Exception:
        return ""
    return 1


def _build_report_candidate_scorecard(
    ctx: RunContext,
    records: Sequence[dict[str, Any]],
) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    candidates = [
        record for record in records
        if record.get("category") == "report_candidate"
        and record.get("extension") in {".xlsx", ".xlsm"}
    ]
    rows: list[dict[str, Any]] = []
    risks: list[str] = []
    for record in candidates:
        path = ctx.root / record["source_path"]
        try:
            from .workbook_audit import audit_workbook

            audit = audit_workbook(path)
            workbook = openpyxl.load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_vba=path.suffix.lower() == ".xlsm",
            )
            populated = formulas = citations = 0
            title_tabs = ogl_tabs = wi_tabs = tract_tabs = 0
            for sheet in workbook.worksheets:
                tab = sheet.title.lower()
                title_tabs += int("title" in tab or "report" in tab)
                ogl_tabs += int("ogl" in tab or "lease" in tab)
                wi_tabs += int("wi" in tab or "working" in tab)
                tract_tabs += int("tract" in tab)
                for values in sheet.iter_rows(values_only=True):
                    for value in values:
                        if value not in (None, ""):
                            populated += 1
                            text = str(value)
                            formulas += int(text.startswith("="))
                            citations += int(
                                bool(re.search(r"\b(?:book|bk\.?)\s*\d+.*\b(?:page|pg\.?)\s*\d+", text, re.I))
                            )
            workbook.close()
            sheet_coverage = min(
                1.0, (title_tabs + ogl_tabs + wi_tabs + tract_tabs) / 4.0
            )
            completeness = min(1.0, populated / 1000.0)
            citation_score = min(1.0, citations / 25.0)
            formatting_integrity = min(
                1.0,
                (
                    sum(sheet["formula_count"] for sheet in audit["sheets"])
                    + sum(len(sheet["merged_cells"]) for sheet in audit["sheets"])
                    + len(audit["sheet_order"])
                ) / 100.0,
            )
            score = round(
                100
                * (
                    0.30 * sheet_coverage
                    + 0.30 * completeness
                    + 0.20 * citation_score
                    + 0.20 * formatting_integrity
                ),
                2,
            )
            if audit["high_risk_features_present"]:
                risks.append(
                    f"- HIGH: `{record['source_path']}` contains drawings, charts, media, "
                    "macros, links, or other high-risk OOXML parts; require COM or exact-copy audit."
                )
            rows.append(
                {
                    "relative_path": record["source_path"],
                    "sha256": record["sha256"],
                    "worksheet_count": len(audit["sheet_order"]),
                    "required_sheet_coverage": round(sheet_coverage, 3),
                    "populated_cells": populated,
                    "formula_cells": formulas,
                    "source_citation_hits": citations,
                    "high_risk_workbook_features": audit["high_risk_features_present"],
                    "score": score,
                    "selection_status": "CANDIDATE - PRESERVED",
                }
            )
        except Exception as exc:
            rows.append(
                {
                    "relative_path": record["source_path"],
                    "sha256": record["sha256"],
                    "score": 0,
                    "selection_status": f"REVIEW REQUIRED: {exc}",
                }
            )
    rows.sort(key=lambda row: float(row.get("score", 0)), reverse=True)
    if rows:
        rows[0]["selection_status"] = "PROVISIONAL WORKING MASTER - HUMAN CONFIRMATION REQUIRED"
        _atomic_text(ctx.run_dir / "selected_master_report.txt", rows[0]["relative_path"])
    else:
        _atomic_text(
            ctx.run_dir / "selected_master_report.txt",
            "NONE - no report candidate workbook found",
        )
        risks.append("- BLOCKED: No report candidate workbook found.")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Candidate Scorecard"
    headers = (
        "relative_path", "sha256", "worksheet_count", "required_sheet_coverage",
        "populated_cells", "formula_cells", "source_citation_hits",
        "high_risk_workbook_features", "score", "selection_status",
    )
    sheet.append([header.replace("_", " ").title() for header in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="163A3D")
    for row in rows:
        sheet.append([_safe_cell(row.get(header, "")) for header in headers])
    workbook.save(ctx.run_dir / "report_candidate_scorecard.xlsx")
    workbook.close()
    _atomic_text(
        ctx.run_dir / "project_risk_register.md",
        "# Project Risk Register\n\n"
        + ("\n".join(risks) if risks else "- No high-risk workbook feature detected during inventory.")
        + "\n\nThis register is preliminary until examiner review and the Section 32 corpus are complete.\n",
    )


def load_inventory(ctx: RunContext) -> list[dict[str, Any]]:
    records = _read_json(ctx.run_dir / "file_inventory.json")
    if records is None:
        return build_inventory(ctx)
    return records


def _safe_rel_name(rel: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", rel).strip("._")
    suffix = hashlib.sha1(rel.encode("utf-8")).hexdigest()[:10]
    return f"{clean[:100]}_{suffix}" if clean else suffix


def _prepare_image_variants(
    source: Path,
    destination_stem: Path,
    frame_index: int = 0,
) -> list[dict[str, str]]:
    from PIL import Image, ImageEnhance, ImageFilter, ImageOps

    destination_stem.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(source) as opened:
        opened.seek(frame_index)
        oriented = ImageOps.exif_transpose(opened)
        if oriented.width < 1800:
            scale = min(3.0, 1800 / max(oriented.width, 1))
            oriented = oriented.resize(
                (int(oriented.width * scale), int(oriented.height * scale)),
                Image.Resampling.LANCZOS,
            )
        grayscale = oriented.convert("L")
        deskewed = grayscale
        try:
            import cv2
            import numpy as np
            from PIL import Image

            pixels = np.array(grayscale)
            coordinates = np.column_stack(np.where(pixels < 245))
            if len(coordinates) > 50:
                angle = cv2.minAreaRect(coordinates[:, ::-1].astype("float32"))[-1]
                angle = -(90 + angle) if angle < -45 else -angle
                if 0.15 < abs(angle) < 15:
                    center = (pixels.shape[1] / 2, pixels.shape[0] / 2)
                    matrix = cv2.getRotationMatrix2D(center, angle, 1.0)
                    pixels = cv2.warpAffine(
                        pixels,
                        matrix,
                        (pixels.shape[1], pixels.shape[0]),
                        flags=cv2.INTER_CUBIC,
                        borderMode=cv2.BORDER_REPLICATE,
                    )
                    deskewed = Image.fromarray(pixels)
        except (ImportError, ValueError, cv2.error if "cv2" in locals() else ValueError):
            deskewed = grayscale
        contrast = ImageEnhance.Contrast(
            ImageOps.autocontrast(deskewed, cutoff=1)
        ).enhance(1.35)
        variants = {
            "orientation_corrected": oriented.convert("RGB"),
            "grayscale": grayscale,
            "deskewed": deskewed,
            "contrast_enhanced": contrast,
            "thresholded": contrast.point(lambda pixel: 255 if pixel > 165 else 0),
            "sharpened": contrast.filter(ImageFilter.SHARPEN),
        }
        outputs: list[dict[str, str]] = []
        for recipe, image in variants.items():
            destination = destination_stem.with_name(
                f"{destination_stem.name}_{recipe}.png"
            )
            image.save(destination, format="PNG", optimize=True)
            outputs.append({"recipe": recipe, "path": str(destination)})
        return outputs


def preprocess_images(ctx: RunContext, dpi: int = 240) -> list[dict[str, Any]]:
    """Create OCR-ready image copies; originals remain untouched."""
    from .project_db import update_stage

    update_stage(ctx.output_dir, ctx.run_id, "PREPROCESS", "IN PROGRESS")
    inventory = load_inventory(ctx)
    ctx.preprocessed_dir.mkdir(parents=True, exist_ok=True)
    pages: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for record in inventory:
        rel = str(record.get("source_path", ""))
        ext = str(record.get("extension", "")).lower()
        source = ctx.root / rel
        stem = _safe_rel_name(rel)
        if ext in IMAGE_EXTENSIONS:
            try:
                _assert_inventory_hash(source, str(record.get("sha256", "")))
                from PIL import Image

                with Image.open(source) as image:
                    frame_count = int(getattr(image, "n_frames", 1))
                for frame_index in range(frame_count):
                    page_number = frame_index + 1
                    destination_stem = (
                        ctx.preprocessed_dir / f"{stem}_page_{page_number:04d}"
                    )
                    variants = _prepare_image_variants(
                        source,
                        destination_stem,
                        frame_index=frame_index,
                    )
                    pages.append(
                        _page_map(
                            rel,
                            record["sha256"],
                            page_number,
                            variants,
                            ctx,
                        )
                    )
            except Exception as exc:
                errors.append({"source_path": rel, "error": str(exc)})
        elif ext in PDF_EXTENSIONS:
            try:
                _assert_inventory_hash(source, str(record.get("sha256", "")))
                import fitz

                with fitz.open(source) as document:
                    matrix = fitz.Matrix(dpi / 72, dpi / 72)
                    for index, page in enumerate(document, start=1):
                        raw = ctx.preprocessed_dir / f"{stem}_page_{index:04d}_rendered.png"
                        destination_stem = ctx.preprocessed_dir / f"{stem}_page_{index:04d}"
                        page.get_pixmap(matrix=matrix, colorspace=fitz.csGRAY, alpha=False).save(raw)
                        variants = [
                            {"recipe": "pdf_rendered", "path": str(raw)},
                            *_prepare_image_variants(raw, destination_stem),
                        ]
                        pages.append(
                            _page_map(rel, record["sha256"], index, variants, ctx)
                        )
            except Exception as exc:
                errors.append({"source_path": rel, "error": str(exc)})
    _write_json(ctx.run_dir / "preprocessed_pages.json", pages)
    _write_json(ctx.run_dir / "preprocess_errors.json", errors)
    _write_csv(
        ctx.quarantine_dir / "preprocess_failures.csv",
        errors,
        ("source_path", "error"),
    )
    _record_stage_artifacts(
        ctx,
        "PREPROCESS",
        (
            ctx.run_dir / "preprocessed_pages.json",
            ctx.run_dir / "preprocess_errors.json",
        ),
    )
    update_stage(
        ctx.output_dir,
        ctx.run_id,
        "PREPROCESS",
        "PASSED" if not errors else "READY FOR REVIEW",
        {"page_count": len(pages), "failure_count": len(errors)},
    )
    return pages


def _page_map(
    rel: str,
    source_hash: str,
    page: int,
    variants: list[dict[str, str]],
    ctx: RunContext,
) -> dict[str, Any]:
    normalized_variants = [
        {
            "recipe": item["recipe"],
            "path": Path(item["path"]).relative_to(ctx.run_dir).as_posix(),
        }
        for item in variants
    ]
    preferred = next(
        (
            item["path"] for item in normalized_variants
            if item["recipe"] == "contrast_enhanced"
        ),
        normalized_variants[0]["path"],
    )
    return {
        "source_path": rel,
        "source_file_hash": source_hash,
        "page": page,
        "preprocessed_path": preferred,
        "variants": normalized_variants,
        "citation": f"{rel}#page={page}",
    }


def _ocr_image(path: Path, config: str = "--oem 3 --psm 6") -> dict[str, Any]:
    import pytesseract
    from PIL import Image

    with Image.open(path) as image:
        data = pytesseract.image_to_data(
            image, config=config, output_type=pytesseract.Output.DICT
        )
    words: list[str] = []
    confidences: list[float] = []
    current_line = None
    chunks: list[str] = []
    blocks: list[dict[str, Any]] = []
    for index, word in enumerate(data.get("text", [])):
        word = str(word).strip()
        if not word:
            continue
        line_key = (
            data["block_num"][index],
            data["par_num"][index],
            data["line_num"][index],
        )
        if current_line is not None and line_key != current_line:
            chunks.append(" ".join(words))
            words = []
        current_line = line_key
        words.append(word)
        word_confidence = 0.0
        try:
            confidence = float(data["conf"][index])
            if confidence >= 0:
                word_confidence = confidence / 100.0
                confidences.append(word_confidence)
        except (TypeError, ValueError):
            pass
        left = int(data["left"][index])
        top = int(data["top"][index])
        width = int(data["width"][index])
        height = int(data["height"][index])
        blocks.append(
            {
                "text": word,
                "normalized_text": re.sub(r"\s+", " ", word).strip(),
                "confidence": round(word_confidence, 4),
                "bounding_box": {
                    "x1": left,
                    "y1": top,
                    "x2": left + width,
                    "y2": top + height,
                    "coordinate_space": "pixels",
                },
                "block_number": int(data["block_num"][index]),
                "paragraph_number": int(data["par_num"][index]),
                "line_number": int(data["line_num"][index]),
                "word_number": int(data["word_num"][index]),
            }
        )
    if words:
        chunks.append(" ".join(words))
    return {
        "text": "\n".join(chunks).strip(),
        "confidence": sum(confidences) / len(confidences) if confidences else 0.0,
        "blocks": blocks,
        "ocr_config": config,
        "engine_version": str(pytesseract.get_tesseract_version()).splitlines()[0],
    }


def _text_records(ctx: RunContext, record: dict[str, Any]) -> list[dict[str, Any]]:
    rel = record["source_path"]
    source = ctx.root / rel
    ext = source.suffix.lower()
    results: list[dict[str, Any]] = []
    if ext in TEXT_EXTENSIONS:
        text = source.read_text(encoding="utf-8", errors="replace")
        results.append(
            _citation(
                rel, "file", text, 1.0, "native_text",
                source_file_hash=record["sha256"],
                blocks=_native_text_blocks(text),
            )
        )
    elif ext in {".csv", ".tsv"}:
        delimiter = "\t" if ext == ".tsv" else ","
        with source.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            for row_number, row in enumerate(csv.reader(handle, delimiter=delimiter), start=1):
                text = " | ".join(str(value) for value in row if value)
                if text:
                    results.append(
                        _citation(
                            rel, f"row={row_number}", text, 1.0, "table_text",
                            source_file_hash=record["sha256"],
                            blocks=_native_text_blocks(text),
                        )
                    )
    elif ext in {".xlsx", ".xlsm"}:
        import openpyxl

        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                    text = " | ".join(str(value) for value in values if value not in (None, ""))
                    if text:
                        locator = f"sheet={sheet.title};row={row_number}"
                        results.append(
                            _citation(
                                rel, locator, text, 1.0, "workbook_text",
                                source_file_hash=record["sha256"],
                                blocks=_native_text_blocks(text),
                            )
                        )
        finally:
            workbook.close()
    return results


def _citation(
    source_path: str,
    locator: str,
    text: str,
    confidence: float,
    method: str,
    *,
    source_file_hash: str = "",
    derived_image_path: str = "",
    preprocessing_recipe: str = "",
    blocks: Optional[list[dict[str, Any]]] = None,
    engine_version: str = "",
    ocr_config: str = "",
) -> dict[str, Any]:
    return {
        "citation": f"{source_path}#{locator}",
        "source_path": source_path,
        "source_locator": locator,
        "source_file_hash": source_file_hash,
        "derived_image_path": derived_image_path,
        "preprocessing_recipe": preprocessing_recipe,
        "text": text,
        "confidence": round(float(confidence), 4),
        "method": method,
        "engine_version": engine_version,
        "ocr_config": ocr_config,
        "blocks": blocks or [],
        "status": "ok" if text.strip() else "weak",
    }


def _native_text_blocks(text: str) -> list[dict[str, Any]]:
    return [
        {
            "text": word,
            "normalized_text": word,
            "confidence": 1.0,
            "bounding_box": None,
            "block_number": None,
            "paragraph_number": None,
            "line_number": line_number,
            "word_number": word_number,
        }
        for line_number, line in enumerate(text.splitlines(), start=1)
        for word_number, word in enumerate(line.split(), start=1)
    ]


def _best_page_ocr(
    ctx: RunContext,
    mapped: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    attempts: list[dict[str, Any]] = []
    for variant in mapped.get("variants", []):
        result = _ocr_image(ctx.run_dir / variant["path"])
        quality = float(result["confidence"]) + min(len(result["text"]) / 5000.0, 0.15)
        attempts.append(
            {
                **result,
                "quality_score": round(quality, 4),
                "derived_image_path": variant["path"],
                "preprocessing_recipe": variant["recipe"],
            }
        )
    if not attempts:
        raise RuntimeError(f"No preprocessing variants for {mapped.get('citation', 'page')}")
    return max(attempts, key=lambda item: item["quality_score"]), attempts


def run_ocr(ctx: RunContext, weak_threshold: float = 0.55) -> list[dict[str, Any]]:
    """Extract text from native documents and OCR page images with citations."""
    from .project_db import update_stage

    update_stage(ctx.output_dir, ctx.run_id, "OCR", "IN PROGRESS")
    inventory = load_inventory(ctx)
    pages = preprocess_images(ctx)
    page_index = {(p["source_path"], p["page"]): p for p in pages}
    records: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    attempts: list[dict[str, Any]] = []
    for item in inventory:
        rel = str(item.get("source_path", ""))
        ext = str(item.get("extension", "")).lower()
        try:
            if ext in SUPPORTED_EXTENSIONS:
                _assert_inventory_hash(ctx.root / rel, str(item.get("sha256", "")))
            if ext in TEXT_EXTENSIONS | TABLE_EXTENSIONS:
                records.extend(_text_records(ctx, item))
            elif ext in PDF_EXTENSIONS:
                records.extend(_ocr_pdf(ctx, rel, page_index, attempts))
            elif ext in IMAGE_EXTENSIONS:
                source_pages = sorted(
                    (
                        mapped for (source_path, _page), mapped in page_index.items()
                        if source_path == rel
                    ),
                    key=lambda mapped: int(mapped["page"]),
                )
                for mapped in source_pages:
                    page_number = int(mapped["page"])
                    best, page_attempts = _best_page_ocr(ctx, mapped)
                    attempts.extend(
                        {
                            **attempt,
                            "source_path": rel,
                            "source_file_hash": mapped["source_file_hash"],
                            "page": page_number,
                        }
                        for attempt in page_attempts
                    )
                    records.append(
                        _citation(
                            rel, f"page={page_number}", best["text"],
                            best["confidence"], "tesseract",
                            source_file_hash=mapped["source_file_hash"],
                            derived_image_path=best["derived_image_path"],
                            preprocessing_recipe=best["preprocessing_recipe"],
                            blocks=best["blocks"],
                            engine_version=best["engine_version"],
                            ocr_config=best["ocr_config"],
                        )
                    )
        except Exception as exc:
            failures.append({"source_path": rel, "error": str(exc)})
    for record in records:
        record["project_id"] = hashlib.sha256(str(ctx.root).encode("utf-8")).hexdigest()[:16]
        if not record["text"].strip() or float(record["confidence"]) < weak_threshold:
            record["status"] = "weak"
    fields = (
        "citation", "source_file_hash", "source_path", "source_locator",
        "derived_image_path", "preprocessing_recipe", "confidence", "method",
        "engine_version", "ocr_config", "status", "blocks", "text",
    )
    _write_jsonl(ctx.run_dir / "ocr_citations.jsonl", records)
    _write_jsonl(ctx.run_dir / "ocr_attempts.jsonl", attempts)
    _write_csv(ctx.run_dir / "ocr_citations.csv", records, fields)
    weak = [record for record in records if record["status"] == "weak"]
    _write_json(ctx.quarantine_dir / "weak_ocr_results.json", weak)
    _write_csv(ctx.quarantine_dir / "weak_ocr_results.csv", weak, fields)
    _write_json(ctx.quarantine_dir / "ocr_failures.json", failures)
    _write_csv(ctx.quarantine_dir / "ocr_failures.csv", failures, ("source_path", "error"))
    _record_stage_artifacts(
        ctx,
        "OCR",
        (
            ctx.run_dir / "ocr_citations.jsonl",
            ctx.run_dir / "ocr_attempts.jsonl",
        ),
    )
    update_stage(
        ctx.output_dir,
        ctx.run_id,
        "OCR",
        "PASSED" if not failures else "READY FOR REVIEW",
        {
            "citation_count": len(records),
            "weak_count": len(weak),
            "failure_count": len(failures),
        },
    )
    return records


def _ocr_pdf(
    ctx: RunContext,
    rel: str,
    page_index: dict[tuple[str, int], dict[str, Any]],
    attempts: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    import fitz

    output: list[dict[str, Any]] = []
    with fitz.open(ctx.root / rel) as document:
        for page_number, page in enumerate(document, start=1):
            embedded = page.get_text("text").strip()
            if len(re.sub(r"\s+", "", embedded)) >= 40:
                words = page.get_text("words")
                blocks = [
                    {
                        "text": str(word[4]),
                        "normalized_text": str(word[4]).strip(),
                        "confidence": 1.0,
                        "bounding_box": {
                            "x1": float(word[0]),
                            "y1": float(word[1]),
                            "x2": float(word[2]),
                            "y2": float(word[3]),
                            "coordinate_space": "pdf_points",
                        },
                        "block_number": int(word[5]),
                        "paragraph_number": int(word[6]),
                        "line_number": int(word[7]),
                        "word_number": None,
                    }
                    for word in words
                ]
                mapped = page_index.get((rel, page_number), {})
                output.append(
                    _citation(
                        rel, f"page={page_number}", embedded, 0.99, "pdf_text",
                        source_file_hash=mapped.get(
                            "source_file_hash", _sha256(ctx.root / rel)
                        ),
                        derived_image_path=mapped.get("preprocessed_path", ""),
                        preprocessing_recipe="native_pdf_text_with_render_reference",
                        blocks=blocks,
                        engine_version=importlib.metadata.version("PyMuPDF"),
                        ocr_config="page.get_text('words')",
                    )
                )
                continue
            mapped = page_index.get((rel, page_number))
            if mapped:
                best, page_attempts = _best_page_ocr(ctx, mapped)
                attempts.extend(
                    {
                        **attempt,
                        "source_path": rel,
                        "source_file_hash": mapped["source_file_hash"],
                        "page": page_number,
                    }
                    for attempt in page_attempts
                )
                output.append(
                    _citation(
                        rel, f"page={page_number}", best["text"], best["confidence"],
                        "tesseract",
                        source_file_hash=mapped["source_file_hash"],
                        derived_image_path=best["derived_image_path"],
                        preprocessing_recipe=best["preprocessing_recipe"],
                        blocks=best["blocks"],
                        engine_version=best["engine_version"],
                        ocr_config=best["ocr_config"],
                    )
                )
    return output


def _write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    text = "".join(json.dumps(row, ensure_ascii=False, default=str) + "\n" for row in rows)
    _atomic_text(path, text)


def load_ocr(ctx: RunContext) -> list[dict[str, Any]]:
    path = ctx.run_dir / "ocr_citations.jsonl"
    if not path.exists():
        raise FileNotFoundError("OCR output is missing. Run OCR first.")
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


_LABEL_PATTERNS: dict[str, tuple[str, ...]] = {
    "instrument_number": (
        r"(?:instrument|document|doc|reception)\s*(?:number|no\.?|#)\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-/.]+)",
        r"\b(?:INST|DOC)\s*#\s*([A-Z0-9][A-Z0-9\-/.]+)",
    ),
    "book": (r"\b(?:book|bk\.?)\s*[:#]?\s*([A-Z0-9-]+)",),
    "page": (r"\b(?:page|pg\.?)\s*[:#]?\s*([A-Z0-9-]+)",),
    "grantor": (
        r"\b(?:grantor|lessor|assignor)\s*[:\-]\s*([^\n|;]{2,100})",
        r"\bfrom\s*[:\-]\s*([^\n|;]{2,100})",
    ),
    "grantee": (
        r"\b(?:grantee|lessee|assignee)\s*[:\-]\s*([^\n|;]{2,100})",
        r"\bto\s*[:\-]\s*([^\n|;]{2,100})",
    ),
    "legal_description": (
        r"\b(?:legal description|lands?|tract)\s*[:\-]\s*([^\n]{4,250})",
        r"\b((?:NE|NW|SE|SW|N/2|S/2|E/2|W/2)[^\n]{0,180}\b(?:section|sec\.?)\s*\d+[^\n]{0,100})",
    ),
    "interest_conveyed": (
        r"\b(?:interest conveyed|conveyed interest|interest)\s*[:\-]\s*([^\n|;]{1,80})",
    ),
    "lease_royalty_terms": (
        r"\b(?:royalty|lease royalty|royalty terms?)\s*[:\-]\s*([^\n|;]{1,100})",
    ),
}
_DATE_RX = re.compile(
    r"\b(?:0?[1-9]|1[0-2])[/.-](?:0?[1-9]|[12]\d|3[01])[/.-](?:19|20)\d{2}\b"
)
_TYPE_RX = re.compile(
    r"\b(warranty deed|quitclaim deed|mineral deed|oil and gas lease|"
    r"assignment(?: of [^\n,;]+)?|affidavit|memorandum|probate|court order|deed)\b",
    re.IGNORECASE,
)


def _first_match(text: str, patterns: Sequence[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return re.sub(r"\s+", " ", match.group(1)).strip(" ,.;:-")
    return ""


def _extract_cursor_candidate(citation: dict[str, Any]) -> Optional[dict[str, Any]]:
    text = str(citation.get("text", ""))
    fields = {name: _first_match(text, patterns) for name, patterns in _LABEL_PATTERNS.items()}
    dates = _DATE_RX.findall(text)
    fields["instrument_date"] = dates[0] if dates else ""
    fields["recorded_date"] = dates[1] if len(dates) > 1 else ""
    type_match = _TYPE_RX.search(text)
    fields["instrument_type"] = type_match.group(1).title() if type_match else ""
    if not any(fields.values()):
        return None
    return _candidate(citation, fields, "cursor")


def _extract_codex_candidate(citation: dict[str, Any]) -> Optional[dict[str, Any]]:
    text = str(citation.get("text", ""))
    fields = {name: _first_match(text, patterns) for name, patterns in _LABEL_PATTERNS.items()}
    if not fields["instrument_number"]:
        fallback = re.search(
            r"\b(?:recorded|filed)\s+(?:as\s+)?(?:instrument\s+)?([A-Z]?\d[\d-]{4,})\b",
            text,
            flags=re.IGNORECASE,
        )
        fields["instrument_number"] = fallback.group(1) if fallback else ""
    dates = _DATE_RX.findall(text)
    recorded = re.search(
        r"\b(?:recorded|filed)\s*(?:on)?\s*[:\-]?\s*(" + _DATE_RX.pattern[2:-2] + r")",
        text,
        flags=re.IGNORECASE,
    )
    fields["recorded_date"] = recorded.group(1) if recorded else (dates[-1] if dates else "")
    fields["instrument_date"] = dates[0] if dates else ""
    type_match = _TYPE_RX.search(text)
    fields["instrument_type"] = type_match.group(1).title() if type_match else ""
    if not any(fields.values()):
        return None
    return _candidate(citation, fields, "codex")


def _candidate(
    citation: dict[str, Any], fields: dict[str, Any], engine: str
) -> dict[str, Any]:
    completeness = sum(bool(fields.get(field)) for field in INSTRUMENT_FIELDS) / len(
        INSTRUMENT_FIELDS
    )
    source_confidence = float(citation.get("confidence", 0.0) or 0.0)
    confidence = min(0.98, 0.35 * source_confidence + 0.65 * completeness)
    field_provenance = {
        field: _field_provenance(citation, field, str(value))
        for field, value in fields.items()
        if value
    }
    return {
        **{field: str(fields.get(field, "") or "").strip() for field in INSTRUMENT_FIELDS},
        "engine": engine,
        "candidate_id": hashlib.sha256(
            (
                engine
                + str(citation.get("citation", ""))
                + json.dumps(fields, sort_keys=True, default=str)
            ).encode("utf-8")
        ).hexdigest()[:20],
        "confidence": round(confidence, 4),
        "citation": citation.get("citation", ""),
        "source_path": citation.get("source_path", ""),
        "source_locator": citation.get("source_locator", ""),
        "source_file_hash": citation.get("source_file_hash", ""),
        "derived_image_path": citation.get("derived_image_path", ""),
        "extraction_method": f"{citation.get('method', 'unknown')}+deterministic_parser",
        "field_provenance": field_provenance,
    }


def _field_provenance(
    citation: dict[str, Any],
    field: str,
    value: str,
) -> dict[str, Any]:
    blocks = citation.get("blocks", []) or []
    value_tokens = [
        _norm(token) for token in re.findall(r"[A-Za-z0-9/.-]+", value) if _norm(token)
    ]
    block_tokens = [_norm(block.get("text", "")) for block in blocks]
    matched: list[dict[str, Any]] = []
    match_start: Optional[int] = None
    if value_tokens:
        for start in range(max(0, len(block_tokens) - len(value_tokens) + 1)):
            segment = block_tokens[start:start + len(value_tokens)]
            if segment == value_tokens:
                matched = blocks[start:start + len(value_tokens)]
                match_start = start
                break
    semantic_support = _field_span_is_semantic(
        field,
        value,
        value_tokens,
        blocks,
        block_tokens,
        match_start,
    )
    boxes = [block["bounding_box"] for block in matched if block.get("bounding_box")]
    bounding_box = None
    if boxes:
        bounding_box = {
            "x1": min(box["x1"] for box in boxes),
            "y1": min(box["y1"] for box in boxes),
            "x2": max(box["x2"] for box in boxes),
            "y2": max(box["y2"] for box in boxes),
            "coordinate_space": boxes[0].get("coordinate_space", "pixels"),
        }
    direct_text = (
        _norm(value) in _norm(citation.get("text", ""))
        and semantic_support
    )
    confidence = (
        sum(float(block.get("confidence", 0.0)) for block in matched) / len(matched)
        if matched else float(citation.get("confidence", 0.0))
    )
    if bounding_box and semantic_support:
        status = "DIRECT_SOURCE_SUPPORT"
        support = min(1.0, 0.65 + 0.35 * confidence)
    elif direct_text:
        status = "PAGE_LEVEL_SUPPORT"
        support = min(0.89, 0.55 + 0.34 * confidence)
    else:
        status = "UNSUPPORTED"
        support = 0.0
    page_match = re.search(r"page=(\d+)", str(citation.get("source_locator", "")))
    return {
        "field": field,
        "project_id": citation.get("project_id", ""),
        "source_file": citation.get("source_path", ""),
        "source_file_hash": citation.get("source_file_hash", ""),
        "source_page": int(page_match.group(1)) if page_match else None,
        "source_image_number": None,
        "source_bounding_box": bounding_box,
        "derived_image_path": citation.get("derived_image_path", ""),
        "preprocessing_recipe": citation.get("preprocessing_recipe", ""),
        "extraction_method": citation.get("method", ""),
        "extraction_model": citation.get("engine_version", "deterministic"),
        "raw_text_snippet": value,
        "confidence": round(confidence, 4),
        "source_support_score": round(support, 4),
        "review_status": status,
    }


_FIELD_LABEL_SEQUENCES: dict[str, tuple[tuple[str, ...], ...]] = {
    "instrument_number": (
        ("INSTRUMENT", "NUMBER"), ("INSTRUMENT", "NO"), ("DOCUMENT", "NUMBER"),
        ("DOC", "NO"), ("RECEPTION", "NUMBER"),
    ),
    "instrument_date": (("INSTRUMENT", "DATE"), ("EXECUTION", "DATE"), ("DATED",)),
    "instrument_type": (
        ("INSTRUMENT", "TYPE"), ("DOCUMENT", "TYPE"), ("DOC", "TYPE"),
    ),
    "recorded_date": (("RECORDED", "DATE"), ("RECORDING", "DATE"), ("FILED",)),
    "book": (("BOOK",), ("BK",)),
    "page": (("PAGE",), ("PG",)),
    "grantor": (("GRANTOR",), ("LESSOR",), ("ASSIGNOR",), ("FROM",)),
    "grantee": (("GRANTEE",), ("LESSEE",), ("ASSIGNEE",), ("TO",)),
    "legal_description": (
        ("LEGAL", "DESCRIPTION"), ("LAND",), ("LANDS",), ("TRACT",),
    ),
    "interest_conveyed": (
        ("INTEREST", "CONVEYED"), ("CONVEYED", "INTEREST"), ("INTEREST",),
    ),
    "lease_royalty_terms": (
        ("LEASE", "ROYALTY"), ("ROYALTY", "TERMS"), ("ROYALTY",),
    ),
}


def _field_span_is_semantic(
    field: str,
    value: str,
    value_tokens: list[str],
    blocks: list[dict[str, Any]],
    block_tokens: list[str],
    match_start: Optional[int],
) -> bool:
    if not value_tokens or match_start is None:
        return False
    label_words = {
        token
        for sequences in _FIELD_LABEL_SEQUENCES.values()
        for sequence in sequences
        for token in sequence
    }
    if len(value_tokens) == 1 and value_tokens[0] in label_words:
        return False
    match_end = match_start + len(value_tokens)
    line_key = _ocr_line_key(blocks[match_start])
    if any(_ocr_line_key(block) != line_key for block in blocks[match_start:match_end]):
        return False
    line_indices = [
        index for index, block in enumerate(blocks)
        if _ocr_line_key(block) == line_key
    ]
    if not line_indices:
        return False
    line_start = min(line_indices)
    line_end = max(line_indices) + 1
    if field == "instrument_type" and _TYPE_RX.fullmatch(value.strip()):
        ordered_lines: list[tuple[Any, Any, Any]] = []
        for block in blocks:
            key = _ocr_line_key(block)
            if key not in ordered_lines:
                ordered_lines.append(key)
        recognized_type_lines: list[tuple[Any, Any, Any]] = []
        for candidate_line in ordered_lines:
            candidate_text = " ".join(
                str(block.get("text", "")).strip()
                for block in blocks
                if _ocr_line_key(block) == candidate_line
            ).strip(" \t:;-")
            if _TYPE_RX.fullmatch(candidate_text):
                recognized_type_lines.append(candidate_line)
        if (
            match_start == line_start
            and match_end == line_end
            and recognized_type_lines
            and line_key == recognized_type_lines[0]
            and ordered_lines.index(line_key) < 5
        ):
            return True
    sequences = _FIELD_LABEL_SEQUENCES.get(field, ())
    all_labels: list[tuple[int, int, str]] = []
    for label_field, field_sequences in _FIELD_LABEL_SEQUENCES.items():
        for sequence in field_sequences:
            length = len(sequence)
            for start in range(line_start, max(line_start, line_end - length + 1)):
                if tuple(block_tokens[start:start + length]) == sequence:
                    all_labels.append((start, start + length, label_field))
    for sequence in sequences:
        length = len(sequence)
        for label_start in range(line_start, match_start):
            if tuple(block_tokens[label_start:label_start + length]) == sequence:
                label_end = label_start + length
                next_label_start = min(
                    (
                        start for start, _end, label_field in all_labels
                        if start >= label_end
                        and start != label_start
                        and label_field != field
                    ),
                    default=line_end,
                )
                if label_end == match_start and match_end == next_label_start:
                    return True
    return False


def _ocr_line_key(block: dict[str, Any]) -> tuple[Any, Any, Any]:
    return (
        block.get("block_number"),
        block.get("paragraph_number"),
        block.get("line_number"),
    )


def _load_external_candidates(path: Optional[str | Path], engine: str) -> Optional[list[dict[str, Any]]]:
    if not path:
        return None
    candidate_path = Path(path).expanduser()
    if not candidate_path.exists():
        raise FileNotFoundError(f"{engine.title()} output does not exist: {candidate_path}")
    data = json.loads(candidate_path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = data.get("instruments") or data.get("results") or data.get("rows") or [data]
    if not isinstance(data, list):
        raise ValueError(f"{engine.title()} output must be a JSON object or array.")
    normalized: list[dict[str, Any]] = []
    for index, row in enumerate(data):
        if not isinstance(row, dict):
            normalized.append(
                {
                    **{field: "" for field in INSTRUMENT_FIELDS},
                    "engine": engine,
                    "confidence": 0.0,
                    "citation": "",
                    "source_path": "",
                    "source_locator": "",
                    "source_file_hash": "",
                    "derived_image_path": "",
                    "extraction_method": engine,
                    "field_provenance": {},
                    "raw_candidate": row,
                    "raw_candidate_index": index,
                    "validation_errors": ["Candidate entry is not a JSON object."],
                }
            )
            continue
        normalized.append(
            {
                **{field: str(row.get(field, "") or "").strip() for field in INSTRUMENT_FIELDS},
                "engine": engine,
                "confidence": float(row.get("confidence", row.get("confidence_score", 0.5)) or 0.0),
                "citation": str(row.get("citation", "") or ""),
                "source_path": str(row.get("source_path", "") or ""),
                "source_locator": str(row.get("source_locator", row.get("page", "")) or ""),
                "source_file_hash": str(row.get("source_file_hash", "") or ""),
                "derived_image_path": str(row.get("derived_image_path", "") or ""),
                "extraction_method": str(row.get("extraction_method", engine) or engine),
                "field_provenance": (
                    row.get("field_provenance")
                    if isinstance(row.get("field_provenance"), dict)
                    else {}
                ),
                "raw_candidate": row,
                "raw_candidate_index": index,
                "validation_errors": (
                    [] if isinstance(row.get("field_provenance", {}), dict)
                    else ["field_provenance must be an object."]
                ),
            }
        )
    return normalized


def _validate_candidate_evidence(
    candidates: list[dict[str, Any]],
    evidence_by_citation: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    for candidate in candidates:
        citation = str(candidate.get("citation", "")).strip()
        evidence = evidence_by_citation.get(citation)
        if not candidate.get("candidate_id"):
            candidate["candidate_id"] = hashlib.sha256(
                json.dumps(candidate, sort_keys=True, default=str).encode("utf-8")
            ).hexdigest()[:20]
        if evidence and not candidate.get("source_file_hash"):
            candidate["source_file_hash"] = evidence.get("source_file_hash", "")
        hash_matches = bool(
            evidence
            and candidate.get("source_file_hash")
            and candidate.get("source_file_hash") == evidence.get("source_file_hash")
        )
        # Never trust model-supplied support scores or geometry. Rebuild support
        # exclusively from the current run's source evidence ledger.
        provenance = {
            field: _field_provenance(evidence, field, str(candidate.get(field, "")))
            for field in INSTRUMENT_FIELDS
            if evidence and str(candidate.get(field, "")).strip()
        }
        candidate["field_provenance"] = provenance
        populated = [
            field for field in INSTRUMENT_FIELDS if str(candidate.get(field, "")).strip()
        ]
        supported_fields = [
            field for field in populated
            if isinstance(provenance.get(field), dict)
            and float(provenance[field].get("source_support_score", 0.0)) > 0
        ]
        candidate["evidence_valid"] = bool(
            citation and evidence and hash_matches and len(supported_fields) == len(populated)
        )
        candidate["evidence_validation"] = {
            "citation_in_current_run": bool(evidence),
            "source_hash_matches": hash_matches,
            "populated_field_count": len(populated),
            "supported_field_count": len(supported_fields),
        }
    return candidates


def _archive_external_candidate_file(
    ctx: RunContext,
    path: Optional[str | Path],
    engine: str,
) -> Optional[dict[str, Any]]:
    if not path:
        return None
    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"{engine.title()} output does not exist: {source}")
    digest = _sha256(source)
    destination = (
        ctx.run_dir / "candidates"
        / f"{engine}_raw_input_{digest[:16]}_{_now_id()}{source.suffix.lower()}"
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    manifest = {
        "engine": engine,
        "original_path": str(source),
        "archived_path": str(destination),
        "sha256": digest,
        "size_bytes": destination.stat().st_size,
        "status": "RAW INPUT - RETAINED BEFORE PARSING",
    }
    _write_json(
        destination.with_suffix(destination.suffix + ".manifest.json"),
        manifest,
    )
    return manifest


def extract_and_reconcile(
    ctx: RunContext,
    weak_threshold: float = 0.65,
    cursor_json: Optional[str | Path] = None,
    codex_json: Optional[str | Path] = None,
) -> list[dict[str, Any]]:
    """Create/load Cursor and Codex candidates, then run a field-level tournament."""
    from .project_db import update_stage

    update_stage(ctx.output_dir, ctx.run_id, "RECONCILE", "IN PROGRESS")
    _archive_current_reconciliation(ctx)
    ocr = load_ocr(ctx)
    evidence_by_citation = {
        str(item.get("citation", "")).strip(): item
        for item in ocr if str(item.get("citation", "")).strip()
    }
    _archive_external_candidate_file(ctx, cursor_json, "cursor")
    _archive_external_candidate_file(ctx, codex_json, "codex")
    cursor = _load_external_candidates(cursor_json, "cursor")
    codex = _load_external_candidates(codex_json, "codex")
    if cursor is None:
        cursor = [
            candidate for item in ocr
            if (candidate := _extract_cursor_candidate(item)) is not None
        ]
    if codex is None:
        codex = [
            candidate for item in ocr
            if (candidate := _extract_codex_candidate(item)) is not None
        ]
    cursor = _validate_candidate_evidence(cursor, evidence_by_citation)
    codex = _validate_candidate_evidence(codex, evidence_by_citation)
    candidates_dir = ctx.run_dir / "candidates"
    fields = (
        *INSTRUMENT_FIELDS, "engine", "confidence", "citation", "source_path",
        "source_locator", "source_file_hash", "derived_image_path",
        "extraction_method", "field_provenance", "evidence_valid",
        "evidence_validation", "candidate_id",
    )
    archive: list[dict[str, Any]] = []
    for engine, rows in (("cursor", cursor), ("codex", codex)):
        _write_json(candidates_dir / f"{engine}_output.json", rows)
        _write_csv(candidates_dir / f"{engine}_output.csv", rows, fields)
        for row in rows:
            immutable = dict(row)
            immutable["sha256"] = hashlib.sha256(
                json.dumps(row, sort_keys=True, default=str).encode("utf-8")
            ).hexdigest()
            immutable["status"] = "CANDIDATE - RETAINED"
            archive.append(immutable)
    _write_jsonl(candidates_dir / "candidate_archive.jsonl", archive)
    reconciled, reconciliation_audit, conflicts = _tournament_with_audit(
        cursor, codex, weak_threshold
    )
    _write_jsonl(ctx.run_dir / "reconciliation_audit.jsonl", reconciliation_audit)
    _write_json(ctx.run_dir / "field_conflicts.json", conflicts)
    _write_csv(
        ctx.run_dir / "field_conflicts.csv",
        conflicts,
        (
            "match_id", "field", "material", "decision", "candidate_values",
            "source_support", "review_reason",
        ),
    )
    for number, row in enumerate(reconciled, start=1):
        row["entry_no"] = number
    _write_json(ctx.run_dir / "instruments.json", reconciled)
    _write_csv(ctx.run_dir / "instruments.csv", reconciled, EXPORT_FIELDS)
    weak = [row for row in reconciled if row["status"] == "QUARANTINED - REVIEW REQUIRED"]
    review_rows = [row for row in reconciled if row["status"] != "ACCEPTED"]
    if weak:
        _write_json(ctx.quarantine_dir / "weak_instruments.json", weak)
        _write_csv(ctx.quarantine_dir / "weak_instruments.csv", weak, EXPORT_FIELDS)
    _write_json(
        ctx.quarantine_dir / "README.json",
        {
            "policy": "Review copies only. No source file was moved, overwritten, or deleted.",
            "weak_threshold": weak_threshold,
            "weak_instrument_count": len(weak),
        },
    )
    _record_stage_artifacts(
        ctx,
        "RECONCILE",
        (
            ctx.run_dir / "instruments.json",
            ctx.run_dir / "field_conflicts.json",
            ctx.run_dir / "reconciliation_audit.jsonl",
            ctx.run_dir / "candidates" / "candidate_archive.jsonl",
        ),
    )
    update_stage(
        ctx.output_dir,
        ctx.run_id,
        "RECONCILE",
        "READY FOR REVIEW" if review_rows or conflicts else "PASSED",
        {
            "candidate_count": len(archive),
            "instrument_count": len(reconciled),
            "conflict_count": len(conflicts),
            "quarantine_count": len(weak),
            "review_count": len(review_rows),
        },
    )
    return reconciled


def _archive_current_reconciliation(ctx: RunContext) -> None:
    sources = [
        ctx.run_dir / "instruments.json",
        ctx.run_dir / "instruments.csv",
        ctx.run_dir / "field_conflicts.json",
        ctx.run_dir / "field_conflicts.csv",
        ctx.run_dir / "reconciliation_audit.jsonl",
        ctx.run_dir / "candidates" / "candidate_archive.jsonl",
        ctx.run_dir / "candidates" / "cursor_output.json",
        ctx.run_dir / "candidates" / "cursor_output.csv",
        ctx.run_dir / "candidates" / "codex_output.json",
        ctx.run_dir / "candidates" / "codex_output.csv",
    ]
    existing = [path for path in sources if path.is_file()]
    if not existing:
        return
    archive = ctx.run_dir / "reconciliation_versions" / _now_id()
    archive.mkdir(parents=True, exist_ok=False)
    for source in existing:
        destination = archive / source.relative_to(ctx.run_dir)
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)


def _norm(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def _candidate_pairs(
    cursor_rows: Sequence[dict[str, Any]],
    codex_rows: Sequence[dict[str, Any]],
) -> list[list[dict[str, Any]]]:
    """Pair cross-engine candidates by instrument, then by source identity.

    Instrument number cannot be the only grouping key because a disagreement in
    that field is exactly what the tournament must surface. Source citations are
    the fallback identity; each Codex row is consumed at most once.
    """
    available = [dict(row) for row in codex_rows]
    groups: list[list[dict[str, Any]]] = []
    for cursor in cursor_rows:
        cursor = dict(cursor)
        cursor_instrument = _norm(cursor.get("instrument_number", ""))
        cursor_citation = str(cursor.get("citation", "")).strip()
        match_index: Optional[int] = None
        if cursor_instrument:
            match_index = next(
                (
                    index for index, row in enumerate(available)
                    if _norm(row.get("instrument_number", "")) == cursor_instrument
                ),
                None,
            )
        if match_index is None and cursor_citation:
            citation_matches = [
                index for index, row in enumerate(available)
                if str(row.get("citation", "")).strip() == cursor_citation
            ]
            if citation_matches:
                def evidence_score(index: int) -> int:
                    row = available[index]
                    return sum(
                        bool(_norm(cursor.get(field, "")))
                        and _norm(cursor.get(field, "")) == _norm(row.get(field, ""))
                        for field in ("grantor", "grantee", "instrument_date", "recorded_date")
                    )

                match_index = max(citation_matches, key=evidence_score)
        group = [cursor]
        if match_index is not None:
            group.append(available.pop(match_index))
        groups.append(group)
    groups.extend([[row] for row in available])
    return groups


def tournament_reconcile(
    cursor_rows: Sequence[dict[str, Any]],
    codex_rows: Sequence[dict[str, Any]],
    weak_threshold: float = 0.65,
) -> list[dict[str, Any]]:
    """Compatibility wrapper for the source-controlled tournament."""
    output, _audit, _conflicts = _tournament_with_audit(
        cursor_rows, codex_rows, weak_threshold
    )
    return output


MATERIAL_FIELDS = {
    "instrument_number", "instrument_type", "instrument_date", "recorded_date",
    "book", "page", "grantor", "grantee", "legal_description",
    "interest_conveyed", "lease_royalty_terms",
}


def _field_support(candidate: dict[str, Any], field: str) -> float:
    provenance = candidate.get("field_provenance", {})
    if not isinstance(provenance, dict) or not isinstance(provenance.get(field), dict):
        return 0.0
    return float(provenance[field].get("source_support_score", 0.0) or 0.0)


def _tournament_with_audit(
    cursor_rows: Sequence[dict[str, Any]],
    codex_rows: Sequence[dict[str, Any]],
    weak_threshold: float,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Reconcile against source support; model agreement never establishes truth."""
    output: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    conflict_rows: list[dict[str, Any]] = []
    for group_number, candidates in enumerate(
        _candidate_pairs(cursor_rows, codex_rows), start=1
    ):
        match_id = f"MATCH-{group_number:06d}"
        winner: dict[str, Any] = {}
        winner_provenance: dict[str, Any] = {}
        field_decisions: dict[str, Any] = {}
        material_conflicts: list[str] = []
        unresolved: list[str] = []
        selected_support: list[float] = []
        for field in INSTRUMENT_FIELDS:
            populated = [row for row in candidates if str(row.get(field, "")).strip()]
            if not populated:
                winner[field] = ""
                field_decisions[field] = {
                    "decision": "NOT_PRESENT",
                    "final_value": None,
                    "candidate_values": {},
                    "source_support": {},
                    "human_review_required": False,
                }
                continue
            by_value: dict[str, list[dict[str, Any]]] = {}
            for row in populated:
                by_value.setdefault(_norm(row[field]), []).append(row)
            ranked = sorted(
                populated,
                key=lambda row: (_field_support(row, field), str(row.get("engine", ""))),
                reverse=True,
            )
            best = ranked[0]
            best_support = _field_support(best, field)
            values = {
                str(row.get("candidate_id", row.get("engine", "candidate"))): row[field]
                for row in populated
            }
            supports = {
                str(row.get("candidate_id", row.get("engine", "candidate"))): _field_support(
                    row, field
                )
                for row in populated
            }
            conflict = len(by_value) > 1
            material = field in MATERIAL_FIELDS
            if conflict and material:
                winner[field] = ""
                decision = "UNRESOLVED"
                material_conflicts.append(field)
                unresolved.append(field)
                review_reason = "Material candidate values conflict; source review required."
            elif best_support <= 0:
                winner[field] = ""
                decision = "UNRESOLVED"
                unresolved.append(field)
                review_reason = "No candidate has direct source support."
            else:
                winner[field] = best[field]
                winner_provenance[field] = best.get("field_provenance", {}).get(field, {})
                selected_support.append(best_support)
                decision = (
                    f"ACCEPT_{str(best.get('engine', 'SOURCE')).upper()}"
                    if not conflict else "ACCEPT_SOURCE_SUPPORTED"
                )
                review_reason = (
                    "Source-supported value selected."
                    if not conflict else "Non-material conflict resolved by stronger source support."
                )
            field_decisions[field] = {
                "decision": decision,
                "final_value": winner[field] or None,
                "candidate_values": values,
                "source_support": supports,
                "human_review_required": decision == "UNRESOLVED",
                "review_reason": review_reason,
            }
            if conflict:
                conflict_rows.append(
                    {
                        "match_id": match_id,
                        "field": field,
                        "material": material,
                        "decision": decision,
                        "candidate_values": values,
                        "source_support": supports,
                        "review_reason": review_reason,
                    }
                )
        score = sum(selected_support) / len(selected_support) if selected_support else 0.0
        citations = sorted({str(row.get("citation", "")) for row in candidates if row.get("citation")})
        verified_citations = sorted(
            {
                str(row.get("citation", ""))
                for row in candidates
                if row.get("citation") and row.get("evidence_valid") is not False
            }
        )
        source_paths = sorted(
            {str(row.get("source_path", "")) for row in candidates if row.get("source_path")}
        )
        locators = sorted(
            {str(row.get("source_locator", "")) for row in candidates if row.get("source_locator")}
        )
        reasons = []
        if material_conflicts:
            reasons.append("material source review: " + ", ".join(material_conflicts))
        if unresolved:
            reasons.append("unresolved fields: " + ", ".join(sorted(set(unresolved))))
        if not winner["instrument_number"]:
            reasons.append("missing instrument number")
        if not citations:
            reasons.append("missing source citation")
        elif not verified_citations:
            reasons.append("citation not found in current OCR evidence ledger")
        if (
            score < weak_threshold
            or not winner["instrument_number"]
            or not verified_citations
        ):
            status = "QUARANTINED - REVIEW REQUIRED"
        elif material_conflicts or unresolved or score < 0.80:
            status = "HUMAN REVIEW REQUIRED"
        elif score < 0.95:
            status = "PROVISIONAL - SAMPLE QA"
        else:
            status = "ACCEPTED"
        output.append(
            {
                **winner,
                "entry_no": "",
                "confidence": round(max(0.0, min(1.0, score)), 4),
                "status": status,
                "citation": " | ".join(citations),
                "source_path": " | ".join(source_paths),
                "source_locator": " | ".join(locators),
                "source_file_hash": " | ".join(
                    sorted(
                        {
                            str(row.get("source_file_hash", ""))
                            for row in candidates if row.get("source_file_hash")
                        }
                    )
                ),
                "field_provenance": winner_provenance,
                "match_id": match_id,
                "reconciliation_notes": "; ".join(reasons) or "candidate evidence reconciled",
            }
        )
        audit_rows.append(
            {
                "match_id": match_id,
                "status": status,
                "candidates": candidates,
                "field_decisions": field_decisions,
            }
        )
    return output, audit_rows, conflict_rows


def build_runsheet(ctx: RunContext) -> dict[str, list[dict[str, Any]]]:
    from .project_db import update_stage

    update_stage(ctx.output_dir, ctx.run_id, "RUNSHEET", "IN PROGRESS")
    instruments = _read_json(ctx.run_dir / "instruments.json")
    if instruments is None:
        raise FileNotFoundError("Instrument output is missing. Run Extract first.")
    runsheet = sorted(
        instruments,
        key=lambda row: (
            _date_sort_value(row.get("recorded_date", "")),
            _date_sort_value(row.get("instrument_date", "")),
            str(row.get("instrument_number", "")),
        ),
    )
    missing: list[dict[str, Any]] = []
    for row in runsheet:
        missing_fields = [
            field for field in ("instrument_number", "instrument_type", "grantor", "grantee", "legal_description")
            if not str(row.get(field, "")).strip()
        ]
        if missing_fields or row.get("status") != "ACCEPTED":
            missing.append(
                {
                    "instrument_number": row.get("instrument_number", ""),
                    "missing_or_issue": ", ".join(missing_fields) or row.get("status", ""),
                    "citation": row.get("citation", ""),
                    "recommended_action": "Locate source document and obtain examiner verification.",
                }
            )
    failures = _read_json(ctx.quarantine_dir / "ocr_failures.json", []) or []
    for failure in failures:
        missing.append(
            {
                "instrument_number": "",
                "missing_or_issue": f"OCR failed: {failure.get('error', '')}",
                "citation": failure.get("source_path", ""),
                "recommended_action": "Inspect the source manually and rerun OCR.",
            }
        )
    preprocess_failures = _read_json(ctx.run_dir / "preprocess_errors.json", []) or []
    for failure in preprocess_failures:
        missing.append(
            {
                "instrument_number": "",
                "missing_or_issue": f"Image preprocessing failed: {failure.get('error', '')}",
                "citation": failure.get("source_path", ""),
                "recommended_action": "Inspect image/PDF integrity and retry preprocessing.",
            }
        )
    represented_sources: set[str] = set()
    for row in runsheet:
        represented_sources.update(
            source.strip()
            for source in str(row.get("source_path", "")).split(" | ")
            if source.strip()
        )
    inventory = load_inventory(ctx)
    already_missing = {str(row.get("citation", "")) for row in missing}
    for item in inventory:
        source_path = str(item.get("source_path", ""))
        if (
            item.get("supported")
            and source_path
            and source_path not in represented_sources
            and source_path not in already_missing
        ):
            missing.append(
                {
                    "instrument_number": "",
                    "missing_or_issue": "No instrument candidate extracted from inventoried source.",
                    "citation": source_path,
                    "recommended_action": "Confirm this source contains no title instrument or review it manually.",
                }
            )
    ogl = [
        row for row in runsheet
        if any(term in str(row.get("instrument_type", "")).lower() for term in ("lease", "memorandum"))
        or str(row.get("lease_royalty_terms", "")).strip()
    ]
    tracts: list[dict[str, Any]] = []
    for row in runsheet:
        if str(row.get("legal_description", "")).strip():
            tracts.append(
                {
                    "tract_id": hashlib.sha1(
                        str(row["legal_description"]).encode("utf-8")
                    ).hexdigest()[:10].upper(),
                    **row,
                }
            )
    result = {
        "runsheet": runsheet,
        "missing_documents": missing,
        "ogl_draft": ogl,
        "tract_drafts": tracts,
    }
    _write_json(ctx.run_dir / "runsheet_bundle.json", result)
    _write_csv(ctx.run_dir / "runsheet.csv", runsheet, EXPORT_FIELDS)
    _write_csv(
        ctx.run_dir / "missing_documents.csv",
        missing,
        ("instrument_number", "missing_or_issue", "citation", "recommended_action"),
    )
    _write_csv(ctx.run_dir / "ogl_draft.csv", ogl, EXPORT_FIELDS)
    _write_csv(ctx.run_dir / "tract_drafts.csv", tracts, ("tract_id", *EXPORT_FIELDS))
    _record_stage_artifacts(
        ctx,
        "RUNSHEET",
        (
            ctx.run_dir / "runsheet_bundle.json",
            ctx.run_dir / "runsheet.csv",
            ctx.run_dir / "missing_documents.csv",
        ),
    )
    update_stage(
        ctx.output_dir,
        ctx.run_id,
        "RUNSHEET",
        "READY FOR REVIEW",
        {
            "runsheet_count": len(runsheet),
            "missing_document_count": len(missing),
        },
    )
    return result


def _date_sort_value(value: Any) -> tuple[dt.date, str]:
    text = str(value or "").strip()
    for date_format in (
        "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m.%d.%Y",
        "%m/%d/%y", "%m-%d-%y", "%B %d, %Y", "%b %d, %Y",
    ):
        try:
            return dt.datetime.strptime(text, date_format).date(), text
        except ValueError:
            continue
    return dt.date.max, text


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _unique_export_path(output_dir: Path, stem: str, suffix: str = ".xlsx") -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    candidate = output_dir / f"{stem}_{_now_id()}{suffix}"
    counter = 1
    while candidate.exists():
        candidate = output_dir / f"{stem}_{_now_id()}_{counter}{suffix}"
        counter += 1
    return candidate


def _unique_sheet_title(workbook: Any, base: str) -> str:
    if base not in workbook.sheetnames:
        return base
    number = 2
    while True:
        suffix = f" ({number})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        number += 1


def export_safe_xlsx(
    ctx: RunContext,
    template: str | Path,
    section: str = "32-11N-25W",
    selected_master: Optional[str | Path] = None,
) -> Path:
    """Create a review package without adding audit sheets to the client file.

    Until an approved writable-range mapping exists, the client candidate is an
    exact copy. All runsheet, provenance, reconciliation, and audit material is
    written to a separate control workbook.
    """
    from .control_workbook import create_control_workbook
    from .project_db import update_stage
    from .workbook_audit import audit_workbook, compare_workbooks

    update_stage(ctx.output_dir, ctx.run_id, "EXPORT_REVIEW_PACKAGE", "IN PROGRESS")

    template_path = Path(template).expanduser().resolve()
    if not template_path.is_file() or template_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Choose an existing .xlsx or .xlsm template.")
    bundle = _read_json(ctx.run_dir / "runsheet_bundle.json")
    if bundle is None:
        build_runsheet(ctx)
    suffix = template_path.suffix.lower()
    package = ctx.root / f"FINAL_DATABOSS_OUTPUT_{_now_id()}"
    package.mkdir(parents=True, exist_ok=False)
    backup_record = _backup_for_operation(
        ctx,
        template_path,
        action="read-only template copy for preservation-audited review package",
    )
    untouched_template = package / f"UNTOUCHED_TEMPLATE_COPY{suffix}"
    destination = package / f"CLIENT_REPORT_CANDIDATE_{section.replace(' ', '_')}{suffix}"
    shutil.copy2(template_path, untouched_template)
    shutil.copy2(template_path, destination)
    selected_master_copy = ""
    if selected_master:
        master_path = Path(selected_master).expanduser().resolve()
        if not master_path.is_file():
            raise FileNotFoundError(f"Selected master does not exist: {master_path}")
        _backup_for_operation(
            ctx,
            master_path,
            action="read-only selected-master copy for review package",
        )
        selected_master_destination = package / (
            "UNTOUCHED_SELECTED_MASTER_COPY" + master_path.suffix.lower()
        )
        shutil.copy2(master_path, selected_master_destination)
        selected_master_copy = str(selected_master_destination)

    before = audit_workbook(template_path)
    after = audit_workbook(destination)
    preservation = compare_workbooks(before, after)
    _write_json(package / "workbook_before_inventory.json", before)
    _write_json(package / "workbook_after_inventory.json", after)
    _write_json(package / "workbook_preservation_audit.json", preservation)

    control_workbook = package / "DATABOSS_CONTROL_WORKBOOK.xlsx"
    project_metadata = {
        "project": "Section 32, Township 11 North, Range 25 West",
        "county": "Beckham County",
        "state": "Oklahoma",
        "work_type": "Horizon Oklahoma Title Report / Diversified research",
        "run_id": ctx.run_id,
        "source_root": str(ctx.root),
        "template": str(template_path),
        "selected_master_copy": selected_master_copy,
        "client_candidate": str(destination),
        "preservation_status": preservation["claim"],
        "submission_status": "NOT READY TO SUBMIT",
    }
    create_control_workbook(
        ctx.run_dir,
        control_workbook,
        preservation,
        project_metadata,
    )
    control_data = package / "CONTROL_DATA"
    control_data.mkdir()
    for name in (
        "project_manifest.csv", "duplicate_map.csv", "file_inventory.json",
        "ocr_citations.jsonl", "instruments.json", "instruments.csv",
        "field_conflicts.csv", "reconciliation_audit.jsonl", "runsheet.csv",
        "missing_documents.csv", "ogl_draft.csv", "tract_drafts.csv",
    ):
        source = ctx.run_dir / name
        if source.exists():
            shutil.copy2(source, control_data / source.name)
    candidate_archive = ctx.run_dir / "candidates" / "candidate_archive.jsonl"
    if candidate_archive.exists():
        shutil.copy2(candidate_archive, control_data / candidate_archive.name)

    review_count = sum(
        str(row.get("status", "")).upper() != "ACCEPTED"
        for row in (_read_json(ctx.run_dir / "instruments.json", []) or [])
    )
    readiness = {
        "report_candidate_path": str(destination),
        "template_path": str(template_path),
        "selected_master_path": selected_master_copy,
        "source_cutoff": dt.datetime.now(dt.timezone.utc).isoformat(),
        "number_of_source_files": len(load_inventory(ctx)),
        "number_of_instruments_extracted": len(
            _read_json(ctx.run_dir / "instruments.json", []) or []
        ),
        "number_requiring_review": review_count,
        "workbook_preservation_status": preservation["claim"],
        "gates_passed": ["source template unchanged", "workbook preservation audit"]
        if preservation["preservation_passed"] else ["source template unchanged"],
        "gates_failed": [
            "approved client workbook mapping not configured",
            "Section 32 production source corpus unavailable in this environment",
            "human title review incomplete",
        ],
        "readiness_score": 20 if preservation["preservation_passed"] else 0,
        "ready_to_submit": False,
        "conclusion": "NOT READY TO SUBMIT",
    }
    _write_json(package / "FINAL_READINESS_STATEMENT.json", readiness)
    _atomic_text(
        package / "FINAL_READINESS_STATEMENT.md",
        "# Final Readiness Statement\n\n"
        "**NOT READY TO SUBMIT**\n\n"
        "The workbook preservation audit covers an exact, unpopulated template copy. "
        "No client cells were populated because no approved writable-range mapping "
        "or Section 32 production corpus is available in this environment.\n",
    )
    export_manifest = {
        "package_path": str(package),
        "export_path": str(destination),
        "control_workbook_path": str(control_workbook),
        "untouched_template_copy": str(untouched_template),
        "template_path": str(template_path),
        "template_sha256": _sha256(template_path),
        "export_sha256": _sha256(destination),
        "section": section,
        "client_worksheets_added": [],
        "preservation_audit_passed": preservation["preservation_passed"],
        "backup_record": backup_record,
    }
    _write_json(ctx.run_dir / "export_manifest.json", export_manifest)
    _write_json(package / "package_manifest.json", export_manifest)
    _record_stage_artifacts(
        ctx,
        "EXPORT_REVIEW_PACKAGE",
        (
            destination,
            control_workbook,
            package / "workbook_preservation_audit.json",
            package / "package_manifest.json",
        ),
    )
    update_stage(
        ctx.output_dir,
        ctx.run_id,
        "EXPORT_REVIEW_PACKAGE",
        "READY FOR REVIEW",
        {
            "package_path": str(package),
            "preservation_audit_passed": preservation["preservation_passed"],
            "ready_to_submit": False,
        },
    )
    return destination


def _backup_for_operation(
    ctx: RunContext,
    source: Path,
    action: str,
) -> dict[str, Any]:
    stamp = "stamp_" + _now_id()
    backup_dir = ctx.root / "_DATABOSS_BACKUPS" / stamp
    backup_dir.mkdir(parents=True, exist_ok=False)
    before_hash = _sha256(source)
    backup_path = backup_dir / f"{before_hash[:12]}_{source.name}"
    shutil.copy2(source, backup_path)
    after_hash = _sha256(source)
    record = {
        "original_path": str(source),
        "backup_path": str(backup_path),
        "sha256_before": before_hash,
        "sha256_after": after_hash,
        "backup_sha256": _sha256(backup_path),
        "action": action,
        "timestamp": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source_unchanged": before_hash == after_hash,
    }
    _write_json(backup_dir / "backup_manifest.json", record)
    return record


def run_summary(ctx: RunContext) -> dict[str, Any]:
    inventory = _read_json(ctx.run_dir / "file_inventory.json", []) or []
    ocr = load_ocr(ctx) if (ctx.run_dir / "ocr_citations.jsonl").exists() else []
    instruments = _read_json(ctx.run_dir / "instruments.json", []) or []
    bundle = _read_json(ctx.run_dir / "runsheet_bundle.json", {}) or {}
    return {
        "run_id": ctx.run_id,
        "files": len(inventory),
        "ocr_citations": len(ocr),
        "instruments": len(instruments),
        "quarantined": sum(
            row.get("status") != "ACCEPTED" for row in instruments
        ),
        "runsheet_rows": len(bundle.get("runsheet", [])),
        "run_dir": str(ctx.run_dir),
    }
