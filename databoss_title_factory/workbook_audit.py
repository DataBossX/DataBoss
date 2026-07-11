"""Excel feature inventory and before/after preservation auditing.

OpenPyXL is not treated as a complete workbook-preservation engine. This module
uses both OOXML package inspection and OpenPyXL's object model, then compares
the inventories before any output can be described as preservation-verified.
"""

from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path
from typing import Any, Iterable, Optional

HIGH_RISK_PREFIXES = (
    "xl/charts/",
    "xl/drawings/",
    "xl/media/",
    "xl/externalLinks/",
    "xl/slicers/",
    "xl/timelines/",
    "xl/ctrlProps/",
    "xl/activeX/",
    "xl/embeddings/",
)
HIGH_RISK_NAMES = {
    "xl/vbaProject.bin",
    "xl/connections.xml",
    "xl/pivotCache/",
}


def _digest_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _json_digest(value: Any) -> str:
    encoded = json.dumps(value, sort_keys=True, ensure_ascii=False, default=str).encode("utf-8")
    return _digest_bytes(encoded)


def _package_inventory(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        names = sorted(archive.namelist())
        high_risk = {
            name: _digest_bytes(archive.read(name))
            for name in names
            if name.startswith(HIGH_RISK_PREFIXES)
            or name in HIGH_RISK_NAMES
            or any(name.startswith(prefix) for prefix in HIGH_RISK_NAMES if prefix.endswith("/"))
        }
        return {
            "member_count": len(names),
            "members": names,
            "high_risk_part_hashes": high_risk,
            "vba_project_present": "xl/vbaProject.bin" in names,
            "chart_parts": [name for name in names if name.startswith("xl/charts/")],
            "drawing_parts": [name for name in names if name.startswith("xl/drawings/")],
            "image_parts": [name for name in names if name.startswith("xl/media/")],
            "slicer_parts": [name for name in names if name.startswith("xl/slicers/")],
            "external_link_parts": [
                name for name in names if name.startswith("xl/externalLinks/")
            ],
            "embedded_object_parts": [
                name for name in names if name.startswith(("xl/embeddings/", "xl/activeX/"))
            ],
        }


def _dimension_inventory(dimensions: Iterable[tuple[Any, Any]]) -> dict[str, Any]:
    return {
        str(key): {
            "hidden": bool(getattr(dimension, "hidden", False)),
            "width": getattr(dimension, "width", None),
            "height": getattr(dimension, "height", None),
            "outline_level": getattr(dimension, "outlineLevel", 0),
        }
        for key, dimension in dimensions
        if (
            getattr(dimension, "hidden", False)
            or getattr(dimension, "width", None) is not None
            or getattr(dimension, "height", None) is not None
            or getattr(dimension, "outlineLevel", 0)
        )
    }


def _sheet_inventory(sheet: Any) -> dict[str, Any]:
    formulas: dict[str, str] = {}
    comments: list[str] = []
    cells: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None or cell.has_style or cell.hyperlink is not None:
                cells[cell.coordinate] = {
                    "value": cell.value,
                    "data_type": cell.data_type,
                    "style_id": cell.style_id,
                    "number_format": cell.number_format,
                    "hyperlink": (
                        cell.hyperlink.target if cell.hyperlink is not None else None
                    ),
                }
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas[cell.coordinate] = cell.value
            if cell.comment is not None:
                comments.append(cell.coordinate)
    merged = sorted(str(item) for item in sheet.merged_cells.ranges)
    validations = []
    if getattr(sheet, "data_validations", None):
        validations = sorted(
            str(item.sqref) for item in sheet.data_validations.dataValidation
        )
    conditional_formatting = sorted(
        str(key) for key in getattr(sheet.conditional_formatting, "_cf_rules", {})
    )
    return {
        "title": sheet.title,
        "state": sheet.sheet_state,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_cells": merged,
        "formula_count": len(formulas),
        "formulas": formulas,
        "formula_digest": _json_digest(formulas),
        "cell_inventory": cells,
        "cell_digest": _json_digest(cells),
        "comments": sorted(comments),
        "image_count": len(getattr(sheet, "_images", [])),
        "chart_count": len(getattr(sheet, "_charts", [])),
        "conditional_formatting_ranges": conditional_formatting,
        "data_validation_ranges": validations,
        "freeze_panes": str(sheet.freeze_panes or ""),
        "print_area": str(sheet.print_area or ""),
        "print_title_rows": str(sheet.print_title_rows or ""),
        "print_title_cols": str(sheet.print_title_cols or ""),
        "auto_filter": str(sheet.auto_filter.ref or ""),
        "sheet_protected": bool(sheet.protection.sheet),
        "row_dimensions": _dimension_inventory(sheet.row_dimensions.items()),
        "column_dimensions": _dimension_inventory(sheet.column_dimensions.items()),
        "page_setup": {
            "orientation": sheet.page_setup.orientation,
            "paper_size": sheet.page_setup.paperSize,
            "fit_to_width": sheet.page_setup.fitToWidth,
            "fit_to_height": sheet.page_setup.fitToHeight,
            "scale": sheet.page_setup.scale,
        },
        "page_margins": {
            "left": sheet.page_margins.left,
            "right": sheet.page_margins.right,
            "top": sheet.page_margins.top,
            "bottom": sheet.page_margins.bottom,
            "header": sheet.page_margins.header,
            "footer": sheet.page_margins.footer,
        },
    }


def audit_workbook(path: str | Path) -> dict[str, Any]:
    """Return a deterministic workbook feature inventory."""
    import openpyxl

    workbook_path = Path(path).expanduser().resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Workbook audit supports .xlsx and .xlsm files.")
    package = _package_inventory(workbook_path)
    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=False,
        data_only=False,
        keep_vba=workbook_path.suffix.lower() == ".xlsm",
        keep_links=True,
    )
    try:
        sheets = [_sheet_inventory(sheet) for sheet in workbook.worksheets]
        defined_names = sorted(
            {
                str(name): str(value)
                for name, value in workbook.defined_names.items()
            }.items()
        )
        result = {
            "path": str(workbook_path),
            "sha256": _digest_bytes(workbook_path.read_bytes()),
            "file_size": workbook_path.stat().st_size,
            "sheet_order": workbook.sheetnames,
            "sheets": sheets,
            "defined_names": defined_names,
            "calculation_mode": getattr(workbook.calculation, "calcMode", None),
            "package": package,
            "high_risk_features_present": bool(
                package["high_risk_part_hashes"]
                or any(sheet["chart_count"] or sheet["image_count"] for sheet in sheets)
            ),
        }
    finally:
        workbook.close()
    result["feature_digest"] = _json_digest(
        {key: value for key, value in result.items() if key not in {"path", "sha256", "file_size"}}
    )
    return result


def compare_workbooks(
    before: dict[str, Any],
    after: dict[str, Any],
    approved_cell_changes: Optional[dict[str, list[str]]] = None,
) -> dict[str, Any]:
    """Compare workbook inventories and identify preservation failures."""
    approved = approved_cell_changes or {}
    differences: list[dict[str, Any]] = []
    if not approved and before.get("sha256") != after.get("sha256"):
        differences.append(
            {
                "feature": "binary_identity",
                "severity": "HIGH",
                "before": before.get("sha256"),
                "after": after.get("sha256"),
            }
        )
    checks = (
        "sheet_order",
        "defined_names",
        "calculation_mode",
    )
    for key in checks:
        if before.get(key) != after.get(key):
            differences.append(
                {"feature": key, "severity": "HIGH", "before": before.get(key), "after": after.get(key)}
            )
    before_parts = before["package"]["high_risk_part_hashes"]
    after_parts = after["package"]["high_risk_part_hashes"]
    if before_parts != after_parts:
        differences.append(
            {
                "feature": "high_risk_ooxml_parts",
                "severity": "HIGH",
                "before": before_parts,
                "after": after_parts,
            }
        )
    before_sheets = {sheet["title"]: sheet for sheet in before["sheets"]}
    after_sheets = {sheet["title"]: sheet for sheet in after["sheets"]}
    if set(before_sheets) != set(after_sheets):
        differences.append(
            {
                "feature": "worksheet_set",
                "severity": "HIGH",
                "before": sorted(before_sheets),
                "after": sorted(after_sheets),
            }
        )
    structural_keys = (
        "state", "merged_cells", "comments", "image_count", "chart_count",
        "conditional_formatting_ranges", "data_validation_ranges", "freeze_panes",
        "print_area", "print_title_rows", "print_title_cols", "sheet_protected",
        "row_dimensions", "column_dimensions", "page_setup", "page_margins",
    )
    for title in sorted(set(before_sheets) & set(after_sheets)):
        old = before_sheets[title]
        new = after_sheets[title]
        for key in structural_keys:
            if old.get(key) != new.get(key):
                differences.append(
                    {
                        "feature": f"{title}.{key}",
                        "severity": "HIGH",
                        "before": old.get(key),
                        "after": new.get(key),
                    }
                )
        old_formulas = old["formulas"]
        new_formulas = new["formulas"]
        allowed = set(approved.get(title, []))
        old_cells = old["cell_inventory"]
        new_cells = new["cell_inventory"]
        changed_cells = sorted(
            coordinate
            for coordinate in set(old_cells) | set(new_cells)
            if old_cells.get(coordinate) != new_cells.get(coordinate)
            and coordinate not in allowed
        )
        if changed_cells:
            differences.append(
                {
                    "feature": f"{title}.cells_or_styles",
                    "severity": "HIGH",
                    "changed_cells": changed_cells,
                }
            )
        changed_formula_cells = sorted(
            coordinate
            for coordinate in set(old_formulas) | set(new_formulas)
            if old_formulas.get(coordinate) != new_formulas.get(coordinate)
            and coordinate not in allowed
        )
        if changed_formula_cells:
            differences.append(
                {
                    "feature": f"{title}.formulas",
                    "severity": "HIGH",
                    "changed_cells": changed_formula_cells,
                }
            )
    passed = not any(item["severity"] in {"HIGH", "CRITICAL"} for item in differences)
    return {
        "preservation_passed": passed,
        "template_sha256": before["sha256"],
        "candidate_sha256": after["sha256"],
        "binary_identical": before["sha256"] == after["sha256"],
        "high_risk_features_present": before["high_risk_features_present"],
        "difference_count": len(differences),
        "differences": differences,
        "claim": (
            "PRESERVATION AUDIT PASSED"
            if passed
            else "PRESERVATION NOT PROVEN - CLIENT EXPORT BLOCKED"
        ),
    }
