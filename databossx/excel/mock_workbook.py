"""Mock title/runsheet workbook generator for DataBossX.

Creates a realistic-but-fake runsheet (hyperlinks, a formula, merged header)
so the pipeline can be proven end-to-end with NO real client data.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from ..core import paths

HEADERS = [
    "Row", "Instrument_Type", "Document_Date", "Recorded_Date",
    "Book", "Page", "Grantor", "Grantee", "Legal_Description", "Doc_Link",
]

MOCK_ROWS = [
    [1, "Warranty Deed", "1998-04-12", "1998-04-15", "412", "221",
     "John A. Sample", "Jane B. Example", "NW/4 Sec 12-3N-4W", "https://example.gov/records/doc/1001"],
    [2, "Oil & Gas Lease", "2005-09-30", "2005-10-02", "905", "118",
     "Jane B. Example", "Acme Minerals LLC", "SE/4 Sec 12-3N-4W", "https://example.gov/records/doc/1002"],
    [3, "Mineral Deed", "2011-01-22", "2011-02-01", "1102", "55",
     "Acme Minerals LLC", "Beta Royalty Trust", "S/2 Sec 12-3N-4W", "https://example.gov/records/doc/1003"],
    [4, "Assignment", "2018-07-08", "2018-07-10", "1807", "300",
     "Beta Royalty Trust", "Gamma Energy Inc", "ALL Sec 12-3N-4W", "https://example.gov/records/doc/1004"],
    [5, "Release", "2021-03-15", "2021-03-18", "2103", "9",
     "Gamma Energy Inc", "Jane B. Example", "NW/4 Sec 12-3N-4W", "https://example.gov/records/doc/1005"],
]


def create_mock_workbook(out_path: Path | None = None, ts: str | None = None) -> Path:
    paths.ensure_dirs()
    ts = ts or paths.timestamp()
    out_path = Path(out_path) if out_path else paths.REVIEW_OUTPUTS / f"mock_runsheet_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Runsheet"

    # Merged title banner across the header columns.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.cell(row=1, column=1, value="MOCK COUNTY RUNSHEET — TEST DATA ONLY")
    ws.cell(row=1, column=1).font = Font(bold=True)

    # Header row.
    for col, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=header)
        c.font = Font(bold=True)

    # Data rows with a clickable hyperlink in the Doc_Link column.
    link_col = HEADERS.index("Doc_Link") + 1
    for i, row in enumerate(MOCK_ROWS, start=3):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            if col == link_col:
                cell.hyperlink = value
                cell.style = "Hyperlink"

    # A formula cell to prove formula preservation: count of data rows.
    summary_row = len(MOCK_ROWS) + 4
    ws.cell(row=summary_row, column=1, value="Total rows:")
    ws.cell(row=summary_row, column=2, value=f"=COUNTA(A3:A{len(MOCK_ROWS) + 2})")

    wb.save(out_path)
    return out_path
