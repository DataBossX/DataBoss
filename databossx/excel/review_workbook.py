"""Safe review-copy builder for DataBossX.

Copies a source workbook to a NEW review file and appends the AI review
columns to the copy only. The source is never opened for writing.

Satisfies: "workbook review copy does not alter source hash".
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from ..core import file_guard, paths

REVIEW_COLUMNS = [
    "AI_Status", "AI_Confidence", "Suggested_Document_Date", "Suggested_Recorded_Date",
    "Suggested_Type", "Suggested_Book", "Suggested_Page", "Suggested_Grantor",
    "Suggested_Grantee", "Suggested_Legal", "Suggested_Notes", "Review_Reason",
]


def create_review_copy(source: Path, ts: str | None = None, header_row: int = 2) -> Path:
    """Create <source>_REVIEW_<ts>.xlsx with empty AI review columns appended.

    The copy is made via the guarded copier (never overwrites), then the AI
    columns are written into the COPY only.
    """
    source = Path(source)
    paths.ensure_dirs()
    ts = ts or paths.timestamp()

    dest = paths.REVIEW_OUTPUTS / f"{source.stem}_REVIEW_{ts}.xlsx"
    # Guarded copy: refuses overwrite / protected / excluded destinations.
    file_guard.guarded_copy(source, dest, dry_run=False, overwrite=False)

    # Now edit only the copy.
    wb = load_workbook(dest)
    ws = wb.active
    start_col = ws.max_column + 1
    for offset, name in enumerate(REVIEW_COLUMNS):
        cell = ws.cell(row=header_row, column=start_col + offset, value=name)
        cell.font = Font(bold=True)
    wb.save(dest)
    wb.close()
    return dest
