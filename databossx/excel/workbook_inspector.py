"""Read-only workbook inspector for DataBossX.

Opens a workbook WITHOUT saving it, and reports sheets, dimensions, headers,
formulas and hyperlinks. Also exports hyperlinks to CSV.

Satisfies: "hyperlink export works".
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from ..core import paths


@dataclass
class SheetInfo:
    name: str
    max_row: int
    max_col: int
    headers: list[str]
    formula_count: int
    hyperlinks: list[tuple[str, str]] = field(default_factory=list)  # (cell, target)


@dataclass
class InspectionResult:
    path: str
    sheets: list[SheetInfo]

    @property
    def total_hyperlinks(self) -> int:
        return sum(len(s.hyperlinks) for s in self.sheets)


def inspect(path: Path) -> InspectionResult:
    path = Path(path)
    # data_only=False keeps formulas visible; read_only would drop hyperlinks,
    # so we use a normal load but never call save().
    wb = load_workbook(path, data_only=False)
    sheets: list[SheetInfo] = []
    for ws in wb.worksheets:
        headers: list[str] = []
        formula_count = 0
        hyperlinks: list[tuple[str, str]] = []
        # Heuristic: first non-empty row with >1 string cells is the header.
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5)):
            strvals = [c.value for c in row if isinstance(c.value, str)]
            if len(strvals) >= 2:
                headers = [str(c.value) if c.value is not None else "" for c in row]
                break
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                if cell.hyperlink is not None:
                    target = cell.hyperlink.target or (cell.value if isinstance(cell.value, str) else "")
                    hyperlinks.append((cell.coordinate, target or ""))
        sheets.append(SheetInfo(ws.title, ws.max_row, ws.max_column, headers, formula_count, hyperlinks))
    wb.close()
    return InspectionResult(path.as_posix(), sheets)


def export_hyperlinks_csv(result: InspectionResult, ts: str | None = None) -> Path:
    paths.ensure_dirs()
    ts = ts or paths.timestamp()
    out = paths.REPORTS / f"workbook_links_{ts}.csv"
    with out.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["sheet", "cell", "target"])
        for sheet in result.sheets:
            for cell, target in sheet.hyperlinks:
                writer.writerow([sheet.name, cell, target])
    return out


def render_report(result: InspectionResult, ts: str) -> str:
    lines = [
        f"# Workbook Inspection — {ts}",
        "",
        f"Workbook: `{result.path}`",
        f"Total hyperlinks: **{result.total_hyperlinks}**",
        "",
    ]
    for s in result.sheets:
        lines += [
            f"## Sheet: `{s.name}`",
            "",
            f"- Dimensions: {s.max_row} rows × {s.max_col} cols",
            f"- Headers: {', '.join(f'`{h}`' for h in s.headers if h) or 'none detected'}",
            f"- Formula cells: {s.formula_count}",
            f"- Hyperlinks: {len(s.hyperlinks)}",
            "",
        ]
        if s.hyperlinks:
            lines.append("| Cell | Target |")
            lines.append("|------|--------|")
            for cell, target in s.hyperlinks[:25]:
                lines.append(f"| {cell} | {target} |")
            lines.append("")
    return "\n".join(lines)


def run(path: Path, ts: str | None = None) -> tuple[Path, Path, InspectionResult]:
    paths.ensure_dirs()
    ts = ts or paths.timestamp()
    result = inspect(path)
    report_path = paths.REPORTS / f"workbook_inspection_{ts}.md"
    report_path.write_text(render_report(result, ts))
    csv_path = export_hyperlinks_csv(result, ts)
    return report_path, csv_path, result
