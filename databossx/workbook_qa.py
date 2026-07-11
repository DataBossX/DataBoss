#!/usr/bin/env python3
"""Examiner-level workbook QA engine (CAP_QA / CAP_WORKBOOK).

Promotes the QA checks used to certify the Section 32-11N-25W BEST_AVAILABLE
workbook (see READ_ME_Diversified_Report_Audit_and_Completion_2026-07-10.md
section 7) into a reusable, spec-driven engine any project can run:

    python -m databossx.workbook_qa report.xlsx [--spec spec.json] [--json out.json]

Checks are declarative: pass a ``spec`` dict (or JSON file) describing what
the workbook must look like; the engine reports PASS/FAIL/WARN/INFO per check
and an overall verdict. ``DIVERSIFIED_32_SPEC`` is the certified spec for the
Section 32 deliverable and doubles as the template for future projects.

The engine enforces the DataBossX evidence rules mechanically where possible:
blank ownership fields are legitimate (blank never means zero), so the
fabricated-decimal check flags any interest-like decimal that appears on
tract/WI sheets that are declared evidence-only.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Optional

try:
    import openpyxl
except ImportError as _exc:  # pragma: no cover
    raise SystemExit("workbook_qa requires openpyxl (pip install openpyxl)") from _exc

ERROR_TOKENS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")

# Certified spec for the 32-11N-25W Diversified cursory report.
DIVERSIFIED_32_SPEC: Dict = {
    "name": "32-11N-25W Beckham Co Diversified Cursory",
    "expected_sheets": ["Overview", "Title ", "OGL", "PLAT", "Runsheet",
                        "Tract 1", "Tract 2", "Tract 3", "Tract 4", "Tract 5",
                        "WI 2", "WI 1", "Well 1"],
    "max_defined_names": 0,
    "expected_formula_cells": 12,
    "runsheet_sheet": "Runsheet",
    "min_runsheet_hyperlinks": 34,
    "contamination_tokens": ["Roger Mills", "27-15N-24W", "Presidio", "WAB",
                             "White Sail", "Greenhead"],
    "required_tokens": [
        # corrected direct-image citations (AUDIT_3 corrections)
        "4118", "4388", "4852",
        # material instruments that must appear in the runsheet
        "1014", "2371", "2393", "2415", "2475", "2476", "2480",
    ],
    "evidence_only_sheets": ["Tract 1", "Tract 2", "Tract 3", "Tract 4",
                             "Tract 5", "WI 1", "WI 2"],
}


def _iter_string_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                yield cell


def validate_workbook(path: Path, spec: Optional[Dict] = None) -> Dict:
    """Run every check; returns {workbook, spec, checks: [...], verdict}."""
    spec = spec or DIVERSIFIED_32_SPEC
    wb = openpyxl.load_workbook(path)
    checks: List[Dict] = []

    def check(check_id: str, status: str, detail: str) -> None:
        checks.append({"id": check_id, "status": status, "detail": detail})

    # 1. Sheet names and order
    expected = spec.get("expected_sheets")
    if expected is not None:
        ok = wb.sheetnames == expected
        check("sheet_names_order", "PASS" if ok else "FAIL",
              f"found {wb.sheetnames}" if not ok else f"{len(expected)} sheets in template order")

    # 2. Defined names
    max_names = spec.get("max_defined_names")
    if max_names is not None:
        n = len(wb.defined_names)
        check("defined_names", "PASS" if n <= max_names else "FAIL",
              f"{n} defined names (max {max_names})")

    # 3. External links
    ext = len(getattr(wb, "_external_links", []) or [])
    check("external_links", "PASS" if ext == 0 else "FAIL", f"{ext} external links")

    # Collect cell statistics in one pass
    formula_cells = 0
    error_cells: List[str] = []
    all_text: List[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    if v.startswith("="):
                        formula_cells += 1
                    if any(tok in v for tok in ERROR_TOKENS):
                        error_cells.append(f"{ws.title}!{cell.coordinate}")
                    all_text.append(v)
    corpus = "\n".join(all_text)

    # 4. Formula error literals
    check("formula_error_cells", "PASS" if not error_cells else "FAIL",
          f"{len(error_cells)} error cells" + (f": {error_cells[:5]}" if error_cells else ""))

    # 5. Formula cell count
    exp_f = spec.get("expected_formula_cells")
    if exp_f is not None:
        check("formula_cell_count", "PASS" if formula_cells == exp_f else "WARN",
              f"{formula_cells} formula cells (expected {exp_f})")

    # 6. Runsheet hyperlinks (traceability)
    rs_name = spec.get("runsheet_sheet")
    if rs_name and rs_name in wb.sheetnames:
        n_links = sum(1 for row in wb[rs_name].iter_rows() for c in row
                      if c.hyperlink is not None)
        min_links = spec.get("min_runsheet_hyperlinks", 1)
        check("runsheet_hyperlinks", "PASS" if n_links >= min_links else "FAIL",
              f"{n_links} hyperlinks on {rs_name} (min {min_links})")

    # 7. Prior-project contamination scan
    hits = [tok for tok in spec.get("contamination_tokens", []) if tok in corpus]
    check("contamination_scan", "PASS" if not hits else "FAIL",
          "no stale-project tokens" if not hits else f"contamination: {hits}")

    # 8. Required instrument/citation tokens
    missing = [tok for tok in spec.get("required_tokens", []) if tok not in corpus]
    check("required_tokens", "PASS" if not missing else "FAIL",
          "all required instrument/citation tokens present" if not missing
          else f"missing tokens: {missing}")

    # 9. Fabricated-decimal guard on evidence-only sheets
    suspicious: List[str] = []
    for name in spec.get("evidence_only_sheets", []):
        if name not in wb.sheetnames:
            continue
        for row in wb[name].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, (int, float)) and 0 < v < 1:
                    suspicious.append(f"{name}!{cell.coordinate}={v}")
    check("fabricated_decimal_guard", "PASS" if not suspicious else "FAIL",
          "no interest-like decimals on evidence-only sheets" if not suspicious
          else f"unsupported decimals: {suspicious[:10]}")

    # 10. Date-as-text density (informational)
    date_texts = len(re.findall(r"\b\d{4}-\d{2}-\d{2}\b", corpus))
    check("date_text_density", "INFO", f"{date_texts} ISO-date strings in text cells")

    failed = [c["id"] for c in checks if c["status"] == "FAIL"]
    return {
        "workbook": str(path),
        "spec": spec.get("name", "custom"),
        "checks": checks,
        "verdict": "PASS" if not failed else "FAIL",
        "failed_checks": failed,
    }


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--spec", type=Path, default=None,
                    help="JSON spec file (default: certified 32-11N-25W spec)")
    ap.add_argument("--json", type=Path, default=None, help="write report JSON here")
    args = ap.parse_args(argv)
    spec = None
    if args.spec:
        with open(args.spec, "r", encoding="utf-8") as fh:
            spec = json.load(fh)
    report = validate_workbook(args.workbook, spec)
    out = json.dumps(report, indent=2, ensure_ascii=False)
    if args.json:
        args.json.write_text(out + "\n", encoding="utf-8")
    print(out)
    return 0 if report["verdict"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
