"""
excel_manager.py
----------------
Append classification results to an Excel workbook.

Behaviour
~~~~~~~~~
* If the workbook does not exist it is created with a formatted header row.
* Each call appends exactly one data row and saves the file atomically.
* Rows are colour-coded by confidence level (green / amber / red) for
  instant visual scanning.
* Column widths are set on creation and respected on subsequent writes.
"""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import Config

logger = logging.getLogger(__name__)

# ── Styling constants ──────────────────────────────────────────────────────────

_HEADER_BG = "1B2A4A"       # deep navy
_HEADER_FG = "FFFFFF"       # white text

# Confidence → row background colour
_CONFIDENCE_BG: dict[str, str] = {
    "High":   "D6EAD7",   # soft green
    "Medium": "FFF3CD",   # soft amber
    "Low":    "FADADD",   # soft red
}

# Column widths (characters) — must match Config.EXCEL_COLUMNS order
_COLUMN_WIDTHS: list[int] = [20, 35, 11, 24, 12, 58, 46, 62]


# ── Public API ─────────────────────────────────────────────────────────────────


def append_result(
    filename: str,
    file_ext: str,
    classification: dict,
) -> Path:
    """
    Write one classification result to the Excel log, creating the file if
    it does not already exist.

    Args:
        filename:       Original uploaded filename.
        file_ext:       File extension without leading dot (e.g. ``"pdf"``).
        classification: ``ClassificationResult`` dict from classifier.py.

    Returns:
        Resolved path to the Excel file.

    Raises:
        RuntimeError: If the file cannot be written (permissions, disk full …).
    """
    excel_path = Config.EXCEL_OUTPUT_PATH

    try:
        if excel_path.exists():
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            logger.debug("Opened existing workbook: %s", excel_path)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Classifications"
            _write_header(ws)
            logger.info("Created new workbook: %s", excel_path)

        row_values = _build_row(filename, file_ext, classification)
        ws.append(row_values)
        _style_data_row(ws, ws.max_row, classification.get("confidence", ""))

        wb.save(excel_path)
        logger.info(
            "Saved classification to row %d of %s", ws.max_row, excel_path
        )
        return excel_path.resolve()

    except PermissionError as exc:
        raise RuntimeError(
            f"Cannot write to '{excel_path}' — the file may be open in Excel."
        ) from exc
    except Exception as exc:
        logger.error("Excel write failed: %s", exc, exc_info=True)
        raise RuntimeError(f"Excel write failed: {exc}") from exc


# ── Private helpers ────────────────────────────────────────────────────────────


def _build_row(filename: str, file_ext: str, classification: dict) -> list:
    """Assemble the flat list of values that maps to Config.EXCEL_COLUMNS."""
    topics = classification.get("key_topics", [])
    topics_str = ", ".join(str(t) for t in topics)

    return [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        filename,
        file_ext.upper(),
        classification.get("category", "Unknown"),
        classification.get("confidence", "Unknown"),
        classification.get("summary", ""),
        topics_str,
        classification.get("reasoning", ""),
    ]


def _write_header(ws) -> None:
    """Create and style the header row (only called once on file creation)."""
    ws.append(Config.EXCEL_COLUMNS)

    header_font = Font(bold=True, color=_HEADER_FG, size=11)
    header_fill = PatternFill(
        start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid"
    )
    center_wrap = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_wrap
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _COLUMN_WIDTHS[col_idx - 1]

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"    # header always visible when scrolling


def _style_data_row(ws, row_num: int, confidence: str) -> None:
    """Apply confidence-based row colouring and wrap long cells."""
    bg_hex = _CONFIDENCE_BG.get(confidence, "FFFFFF")
    fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    for cell in ws[row_num]:
        cell.fill = fill
        cell.alignment = wrap_top
