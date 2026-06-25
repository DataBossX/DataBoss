"""
Cursory Title Report generator — Section 31, T12N, R24W, I.M., Roger Mills County, OK.

Produces a single multi-sheet Excel workbook styled like the DOTO Image Commander
reports (navy header fill, white bold headers, thin borders, review/error/ok
highlight fills) and structured as a professional cursory (limited) title report:

    Cover  ·  Tract & Legal  ·  Chain of Title (runsheet)  ·
    Mineral Ownership  ·  Encumbrances & Leases  ·  Requirements  ·  Sources

DATA HONESTY
------------
This environment has no live access to Roger Mills County land records and the
repository contains no chain-of-title data for this tract. The chain, parties,
book/page, and dates below are an INTERNALLY CONSISTENT ILLUSTRATIVE EXAMPLE,
flagged "ILLUSTRATIVE — VERIFY OF RECORD" on every constructed row. The PLSS
math, mineral-acre arithmetic (decimals sum to 1.000000), and all source links
are real. Replace the rows in TRACT["chain"] / ["minerals"] / ["encumbrances"]
with instruments actually pulled and verified from the sources on the Sources
sheet before relying on this document.

Run:  python3 cursory_title_report.py [output.xlsx]
"""

from __future__ import annotations

import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Styling (matches doto_image_commander/reports/generator.py) ────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="1F4E79", bold=True, size=16)
SUB_FONT = Font(color="1F4E79", bold=True, size=12)
LABEL_FONT = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single")
WARN_FILL = PatternFill("solid", fgColor="FFE699")   # review / verify
ERROR_FILL = PatternFill("solid", fgColor="FF9999")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
BAND_FILL = PatternFill("solid", fgColor="DCE6F1")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

REPORT_DATE = "2026-06-25"
PREPARED = "DataBoss Cursory Title — automated draft (examiner review required)"


# ── The tract (edit these structures with verified record data) ────────────────
TRACT = {
    "caption": "Section 31, Township 12 North, Range 24 West, Indian Meridian",
    "county": "Roger Mills",
    "state": "Oklahoma",
    "meridian": "Indian Meridian (I.M.)",
    "section": "31",
    "township": "12N",
    "range": "24W",
    "nominal_acres": 640.0,
    "report_type": "Cursory (Limited) Title Report",
    "effective_date": REPORT_DATE,
    # Aliquot acreage table (nominal; confirm against GLO plat — Sec 31 may carry
    # government lots along the west tier from PLSS convergence).
    "aliquots": [
        ("NE/4", "NE/4 of Sec 31-12N-24W", 160.0),
        ("NW/4", "NW/4 of Sec 31-12N-24W (may include Gov't Lots 1-2)", 160.0),
        ("SE/4", "SE/4 of Sec 31-12N-24W", 160.0),
        ("SW/4", "SW/4 of Sec 31-12N-24W (may include Gov't Lots 3-4)", 160.0),
    ],
    # Chain of title, sovereignty → present. STATUS flags each row.
    # cols: entry, itype, doc_no, book, page, rec_date, inst_date,
    #       grantor, grantee, conveyance, aliquot, nma, note, source_key, status
    "chain": [
        (1, "U.S. Patent", "PAT-IM-0000", "GLO", "Plat", "1902-xx-xx", "1902-xx-xx",
         "The United States of America", "[Original Patentee — verify GLO]",
         "Surface + 8/8 minerals", "All of Sec 31", 640.0,
         "Sovereignty anchor. Verify patentee, date & any mineral reservation on the GLO patent + plat.",
         "glo", "VERIFY OF RECORD"),
        (2, "Warranty Deed", "WD-0001", "0042", "0117", "1928-03-11", "1928-03-02",
         "[Patentee heirs]", "[Grantee A]",
         "Surface + 8/8 minerals", "All of Sec 31", 640.0,
         "First private conveyance. Confirm full legal & marital/heirship recitals.",
         "okcr", "ILLUSTRATIVE — VERIFY"),
        (3, "Mineral Deed", "MD-0044", "0061", "0233", "1947-09-08", "1947-09-01",
         "[Grantee A]", "[Mineral Grantee B]",
         "Undivided 1/2 of 8/8 minerals", "All of Sec 31", 320.0,
         "Severs 1/2 minerals from surface. Surface stays with Grantee A.",
         "okcr", "ILLUSTRATIVE — VERIFY"),
        (4, "Warranty Deed", "WD-0210", "0098", "0410", "1962-05-20", "1962-05-15",
         "[Grantee A]", "[Surface Owner C]",
         "Surface + remaining 1/2 minerals", "All of Sec 31", 320.0,
         "Surface + the 1/2 minerals retained at entry 3.",
         "okcr", "ILLUSTRATIVE — VERIFY"),
        (5, "Mineral Deed", "MD-0455", "0140", "0066", "1979-11-02", "1979-10-28",
         "[Surface Owner C]", "[Mineral Grantee D]",
         "Undivided 1/4 of 8/8 minerals", "All of Sec 31", 160.0,
         "Further severance; Surface Owner C retains surface + 1/4 minerals.",
         "okcr", "ILLUSTRATIVE — VERIFY"),
        (6, "Oil & Gas Lease", "OGL-1201", "0233", "0512", "2014-02-18", "2014-02-10",
         "[All mineral owners of record]", "[Operator X Energy LLC]",
         "OGL, 3/16 royalty, 3-yr primary", "All of Sec 31", 640.0,
         "Lease of record. Held by production (HBP) — see OCC well data. Verify ratification by each owner.",
         "occ", "ILLUSTRATIVE — VERIFY"),
        (7, "Probate / Decree", "PB-2021-31", "0301", "0044", "2021-07-30", "2021-07-22",
         "Estate of [Mineral Grantee B]", "[Heirs of B — B1, B2]",
         "1/2 minerals to heirs (1/4 each)", "All of Sec 31", 320.0,
         "Distributes entry-3 minerals. Confirm certified decree & no pending appeal.",
         "okcr", "ILLUSTRATIVE — VERIFY"),
        (8, "Affidavit of Heirship", "AFF-3309", "0312", "0190", "2023-04-05", "2023-03-30",
         "Affiant re: [Mineral Grantee D]", "[Heirs of D — D1]",
         "1/4 minerals to sole heir", "All of Sec 31", 160.0,
         "Heirship affidavit; weaker than decree. Requirement raised — see Requirements.",
         "okcr", "ILLUSTRATIVE — VERIFY"),
    ],
    # Mineral ownership AS OF effective date. Must sum to 640 nma / 1.000000.
    # cols: owner, capacity, aliquot, gross_ac, nma, fraction_8_8, decimal, note
    "minerals": [
        ("[Surface Owner C]", "Surface + Minerals", "All of Sec 31", 640.0, 160.0,
         "1/4", "Surface fee; 1/4 of 8/8 minerals retained after entries 4-5."),
        ("[Heir B1]", "Minerals", "All of Sec 31", 640.0, 160.0,
         "1/4", "Via probate decree entry 7."),
        ("[Heir B2]", "Minerals", "All of Sec 31", 640.0, 160.0,
         "1/4", "Via probate decree entry 7."),
        ("[Heir D1]", "Minerals", "All of Sec 31", 640.0, 160.0,
         "1/4", "Via affidavit of heirship entry 8 (curative requirement open)."),
    ],
    # Encumbrances & active leases.
    # cols: etype, ref, date, parties, affects, status, source_key
    "encumbrances": [
        ("Oil & Gas Lease", "Bk 0233 / Pg 0512 (OGL-1201)", "2014-02-18",
         "Lessor: mineral owners → Lessee: [Operator X Energy LLC]",
         "All minerals, Sec 31", "ACTIVE — held by production (verify HBP via OCC)", "occ"),
        ("OCC Spacing / Pooling Order", "[Order No. — verify OCC OGCD]", "[date]",
         "Applicant: [Operator X] — Sec 31-12N-24W unit", "Drilling & spacing unit",
         "VERIFY — pull order from OCC imaging", "occ"),
        ("Right-of-Way Easement", "[Bk/Pg — verify]", "[date]",
         "Grantor: surface owner → Grantee: [pipeline/utility]", "Surface, Sec 31",
         "VERIFY — confirm of record", "okcr"),
    ],
    # Curative requirements / examiner comments.
    # cols: num, requirement, priority, reference
    "requirements": [
        (1, "Confirm U.S. patent date, patentee, and ANY mineral/coal reservation "
            "in favor of the United States (GLO patent + plat).", "HIGH", "Entry 1 / GLO"),
        (2, "Verify Sec 31 government lots & exact acreage on the GLO township plat; "
            "reconcile against 640-acre nominal used in the mineral math.", "HIGH", "Tract & Legal"),
        (3, "Obtain certified probate decree distributing the [B] minerals; confirm "
            "finality and heir identities.", "HIGH", "Entry 7"),
        (4, "Affidavit of heirship (entry 8) is insufficient alone — obtain probate "
            "or quiet-title decree for the [D] 1/4 minerals.", "HIGH", "Entry 8"),
        (5, "Pull OCC spacing/pooling order(s) and current well status to confirm the "
            "OGL is held by production.", "MEDIUM", "Encumbrances"),
        (6, "Run names through judgment/lien, UCC, and federal/state tax-lien indices; "
            "check pending litigation and bankruptcy.", "MEDIUM", "All owners"),
        (7, "Confirm ad valorem taxes paid (county treasurer) and no resale/tax deed.",
            "MEDIUM", "Roger Mills Treasurer"),
        (8, "Verify marital status / homestead joinder on each conveyance.", "MEDIUM", "Chain"),
    ],
}

# Real, public source systems (links are live and clickable).
SOURCES = {
    "okcr": ("Roger Mills County land records (OKCountyRecords)",
             "Deeds, leases, mortgages, probate, affidavits — book/page & instrument images.",
             "https://okcountyrecords.com/"),
    "glo": ("BLM GLO Records — federal land patents & survey plats",
            "Original U.S. patent, township plat, government lots, federal reservations.",
            "https://glorecords.blm.gov/"),
    "occ": ("Oklahoma Corporation Commission (OGCD)",
            "Spacing/pooling orders, well records, production (HBP determination).",
            "https://oklahoma.gov/occ.html"),
    "occimg": ("OCC Well & Imaging system",
               "Scanned well files, completion reports, OGCD case images.",
               "https://imaging.occ.ok.gov/"),
    "blmlr": ("BLM Land Records / survey services",
              "Survey plats, field notes, master title plats.",
              "https://www.blm.gov/services/land-records"),
    "rmco": ("Roger Mills County (Clerk / Assessor / Treasurer)",
             "County clerk recording, assessor parcels/GIS, treasurer tax status.",
             "https://www.rogermillscounty.org/"),
}


# ── Helpers ────────────────────────────────────────────────────────────────────
def style_header(ws: Worksheet, row: int = 1) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def grid(ws: Worksheet, first_row: int, last_row: int, ncols: int) -> None:
    for r in range(first_row, last_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.alignment.wrap_text is False:
                cell.alignment = WRAP


def autofit(ws: Worksheet, max_w: int = 60, min_w: int = 10) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = max(min(longest + 3, max_w), min_w)


def link_cell(ws: Worksheet, row: int, col: int, url: str, text: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=text or url)
    cell.hyperlink = url
    cell.font = LINK_FONT
    cell.border = BORDER
    cell.alignment = WRAP


def status_fill(value: str) -> PatternFill | None:
    v = (value or "").upper()
    if "VERIFY" in v or "ILLUSTRATIVE" in v:
        return WARN_FILL
    if v in ("ACTIVE", "OK", "CONFIRMED", "VESTED"):
        return OK_FILL
    return None


# ── Sheets ─────────────────────────────────────────────────────────────────────
def sheet_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "CURSORY (LIMITED) TITLE REPORT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:F2")
    ws["A2"] = TRACT["caption"]
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A3:F3")
    ws["A3"] = f'{TRACT["county"]} County, {TRACT["state"]}'
    ws["A3"].font = SUB_FONT

    meta = [
        ("Tract", TRACT["caption"]),
        ("County / State", f'{TRACT["county"]}, {TRACT["state"]}'),
        ("Meridian", TRACT["meridian"]),
        ("Sec-Twp-Rng", f'{TRACT["section"]}-{TRACT["township"]}-{TRACT["range"]}'),
        ("Nominal gross acres", f'{TRACT["nominal_acres"]:.2f} (verify GLO plat)'),
        ("Report type", TRACT["report_type"]),
        ("Effective date", TRACT["effective_date"]),
        ("Prepared by", PREPARED),
    ]
    r = 5
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    banner = ws.cell(row=r, column=1,
                     value="⚠  DATA STATUS: The chain of title, parties, book/page, and dates in this "
                           "workbook are an INTERNALLY CONSISTENT ILLUSTRATIVE EXAMPLE — every "
                           "constructed instrument is flagged “ILLUSTRATIVE — VERIFY OF RECORD.” "
                           "PLSS math, mineral-acre arithmetic, and all source links are real. "
                           "Pull and verify each instrument from the Sources sheet before relying on this report.")
    banner.fill = WARN_FILL
    banner.font = Font(bold=True, color="7F6000")
    banner.alignment = WRAP
    ws.row_dimensions[r].height = 70

    r += 2
    ws.cell(row=r, column=1, value="Scope & limitations").font = SUB_FONT
    r += 1
    for line in [
        "• Cursory (limited) examination — NOT a title opinion and not a commitment to insure.",
        "• Covers instruments of record affecting the captioned tract found in the listed indices.",
        "• Subject to: matters not of record, survey/boundary, unrecorded leases & assignments,",
        "  rights of parties in possession, taxes & assessments, and the requirements on the",
        "  Requirements sheet. Examiner review and curative required before reliance.",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=line)
        r += 1

    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18


def sheet_tract(wb: Workbook) -> None:
    ws = wb.create_sheet("Tract & Legal")
    ws["A1"] = "TRACT & LEGAL DESCRIPTION"
    ws["A1"].font = SUB_FONT
    ws.append([])
    ws.append(["Full legal", TRACT["caption"] + f', {TRACT["county"]} County, {TRACT["state"]}'])
    ws.append(["PLSS", f'Section {TRACT["section"]}, T{TRACT["township"]}, R{TRACT["range"]}, {TRACT["meridian"]}'])
    ws.append(["Nominal gross acres", f'{TRACT["nominal_acres"]:.2f}'])
    for r in range(3, 6):
        ws.cell(row=r, column=1).font = LABEL_FONT
    ws.append([])

    hr = ws.max_row + 1
    ws.append(["Aliquot", "Description", "Nominal Acres"])
    style_header(ws, hr)
    for code, desc, ac in TRACT["aliquots"]:
        ws.append([code, desc, ac])
    total_row = ws.max_row + 1
    ws.append(["TOTAL", "Section 31 (nominal)", sum(a[2] for a in TRACT["aliquots"])])
    for cell in ws[total_row]:
        cell.font = LABEL_FONT
        cell.fill = BAND_FILL
    grid(ws, hr, total_row, 3)
    ws.append([])
    note = ws.max_row + 1
    ws.cell(row=note, column=1,
            value="Note: Section 31 lies on the west tier of the township; PLSS convergence "
                  "typically creates government lots along the west boundary, so actual acreage "
                  "may differ from 640. Confirm exact lots/acreage on the GLO township plat "
                  "(see Sources → BLM GLO Records).")
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=3)
    ws.cell(row=note, column=1).fill = WARN_FILL
    ws.cell(row=note, column=1).alignment = WRAP
    ws.row_dimensions[note].height = 50
    autofit(ws)


def sheet_chain(wb: Workbook) -> None:
    ws = wb.create_sheet("Chain of Title")
    headers = ["#", "Instrument Type", "Doc /\nInstr #", "Book", "Page",
               "Recording Date", "Instrument Date", "Grantor", "Grantee",
               "Conveyance / Interest", "Aliquot / Legal", "Net Min. Ac.",
               "Note", "Source", "Status"]
    ws.append(headers)
    style_header(ws, 1)
    src_col = headers.index("Source") + 1
    status_col = headers.index("Status") + 1
    for row in TRACT["chain"]:
        (entry, itype, doc, book, page, rec, inst, gtor, gtee, conv,
         aliquot, nma, note, src_key, status) = row
        ws.append([entry, itype, doc, book, page, rec, inst, gtor, gtee, conv,
                   aliquot, nma, note, "", status])
        ridx = ws.max_row
        name, _desc, url = SOURCES[src_key]
        link_cell(ws, ridx, src_col, url, name)
        fill = status_fill(status)
        if fill:
            ws.cell(row=ridx, column=status_col).fill = fill
    grid(ws, 1, ws.max_row, len(headers))
    ws.freeze_panes = "A2"
    autofit(ws, max_w=42)
    ws.column_dimensions[get_column_letter(headers.index("Note") + 1)].width = 46
    ws.column_dimensions["A"].width = 5


def sheet_minerals(wb: Workbook) -> None:
    ws = wb.create_sheet("Mineral Ownership")
    ws["A1"] = f'MINERAL OWNERSHIP — as of {TRACT["effective_date"]}'
    ws["A1"].font = SUB_FONT
    ws.append([])
    hr = ws.max_row + 1
    headers = ["Owner", "Capacity", "Aliquot", "Gross Ac.", "Net Min. Ac.",
               "Fraction of 8/8", "Decimal Interest", "Note"]
    ws.append(headers)
    style_header(ws, hr)
    total_nma = 0.0
    total_dec = 0.0
    for owner, cap, aliquot, gross, nma, frac, note in TRACT["minerals"]:
        decimal = nma / TRACT["nominal_acres"]
        total_nma += nma
        total_dec += decimal
        ws.append([owner, cap, aliquot, gross, nma, frac, round(decimal, 6), note])
        ws.cell(row=ws.max_row, column=7).number_format = "0.000000"
    tr = ws.max_row + 1
    ws.append(["TOTAL", "", "", "", round(total_nma, 2), "8/8", round(total_dec, 6), ""])
    ws.cell(row=tr, column=7).number_format = "0.000000"
    for cell in ws[tr]:
        cell.font = LABEL_FONT
        cell.fill = OK_FILL if abs(total_dec - 1.0) < 1e-6 else ERROR_FILL
    grid(ws, hr, tr, len(headers))

    chk = tr + 2
    ok = abs(total_dec - 1.0) < 1e-6 and abs(total_nma - TRACT["nominal_acres"]) < 1e-6
    ws.cell(row=chk, column=1,
            value=("✔ Reconciliation OK: net mineral acres = %.2f and decimals sum to %.6f (8/8)."
                   % (total_nma, total_dec)) if ok else
                  ("✘ Reconciliation FAILED: nma=%.2f, decimals=%.6f — fix the mineral table."
                   % (total_nma, total_dec)))
    ws.merge_cells(start_row=chk, start_column=1, end_row=chk, end_column=8)
    ws.cell(row=chk, column=1).fill = OK_FILL if ok else ERROR_FILL
    ws.cell(row=chk, column=1).font = LABEL_FONT
    autofit(ws, max_w=46)


def sheet_encumbrances(wb: Workbook) -> None:
    ws = wb.create_sheet("Encumbrances & Leases")
    headers = ["Type", "Reference (Bk/Pg/Order)", "Date", "Parties",
               "Affects", "Status", "Source"]
    ws.append(headers)
    style_header(ws, 1)
    src_col = len(headers)
    status_col = headers.index("Status") + 1
    for etype, ref, dt, parties, affects, status, src_key in TRACT["encumbrances"]:
        ws.append([etype, ref, dt, parties, affects, status, ""])
        ridx = ws.max_row
        name, _d, url = SOURCES[src_key]
        link_cell(ws, ridx, src_col, url, name)
        fill = status_fill(status)
        if fill:
            ws.cell(row=ridx, column=status_col).fill = fill
    grid(ws, 1, ws.max_row, len(headers))
    autofit(ws, max_w=46)


def sheet_requirements(wb: Workbook) -> None:
    ws = wb.create_sheet("Requirements")
    headers = ["#", "Requirement / Examiner Comment", "Priority", "Reference"]
    ws.append(headers)
    style_header(ws, 1)
    for num, req, prio, ref in TRACT["requirements"]:
        ws.append([num, req, prio, ref])
        if prio.upper() == "HIGH":
            ws.cell(row=ws.max_row, column=3).fill = ERROR_FILL
        elif prio.upper() == "MEDIUM":
            ws.cell(row=ws.max_row, column=3).fill = WARN_FILL
    grid(ws, 1, ws.max_row, len(headers))
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22


def sheet_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    headers = ["Source", "What it provides", "Link (click)"]
    ws.append(headers)
    style_header(ws, 1)
    for _key, (name, what, url) in SOURCES.items():
        ws.append([name, what, ""])
        link_cell(ws, ws.max_row, 3, url)
    grid(ws, 1, ws.max_row, len(headers))
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 42
    note = ws.max_row + 2
    ws.cell(row=note, column=1,
            value="Per-instrument deep links: once a document is located in OKCountyRecords, "
                  "paste its image/permalink into the Source cell of the matching Chain of Title "
                  "row so the runsheet links straight to the recorded image.")
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=3)
    ws.cell(row=note, column=1).alignment = WRAP
    ws.row_dimensions[note].height = 40


def build(output_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_cover(wb)
    sheet_tract(wb)
    sheet_chain(wb)
    sheet_minerals(wb)
    sheet_encumbrances(wb)
    sheet_requirements(wb)
    sheet_sources(wb)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else \
        "31-12N-24W_Roger_Mills_Cursory_Title_Report_(6-25-2026).xlsx"
    print("Wrote:", build(out))
