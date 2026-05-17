"""Generate all DOTO Image Commander reports to Excel workbooks."""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.config import config
from core.database import fetch_all
from core.audit import record


# ── Style helpers ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REVIEW_FILL = PatternFill("solid", fgColor="FFE699")
ERROR_FILL = PatternFill("solid", fgColor="FF9999")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ── Individual sheets ─────────────────────────────────────────────────────────

def _sheet_image_index(wb: Workbook) -> None:
    ws = wb.create_sheet("Image Index")
    headers = [
        "Queue ID", "County", "Book", "Page", "Instrument #",
        "Image #", "Type", "Section", "Township", "Range",
        "Recording Date", "Status", "PDF Path",
    ]
    ws.append(headers)
    rows = fetch_all(
        """SELECT q.id, q.county, q.book, q.page, q.instrument_number, q.image_number,
                  q.instrument_type, q.section, q.township, q.range_val,
                  q.recording_date, q.status, d.file_path
           FROM image_queue q
           LEFT JOIN downloads d ON d.queue_item_id = q.id AND d.status='success'
           ORDER BY q.county, q.book, q.page"""
    )
    for r in rows:
        ws.append([
            r["id"], r["county"], r["book"], r["page"],
            r["instrument_number"], r["image_number"], r["instrument_type"],
            r["section"], r["township"], r["range_val"],
            r["recording_date"], r["status"], r["file_path"] or "",
        ])
    _style_header_row(ws)
    _autofit(ws)


def _sheet_analysis_results(wb: Workbook) -> None:
    ws = wb.create_sheet("Analysis Results")
    headers = [
        "Queue ID", "County", "Instrument Type", "Instrument #",
        "Book", "Page", "Recording Date", "Instrument Date",
        "Grantors", "Grantees", "Legal Description",
        "Section", "Township", "Range", "Interest Conveyed",
        "Lease/Royalty Terms", "Title Req Affected",
        "Confidence", "Needs Review",
    ]
    ws.append(headers)
    rows = fetch_all(
        """SELECT a.*, q.county FROM analyses a
           JOIN image_queue q ON q.id = a.queue_item_id
           ORDER BY q.county, a.book, a.page"""
    )
    for r in rows:
        def _parse_list(v):
            if not v:
                return ""
            try:
                lst = json.loads(v)
                return "; ".join(lst) if isinstance(lst, list) else str(v)
            except Exception:
                return str(v)

        row_data = [
            r["queue_item_id"], r["county"], r["instrument_type"], r["instrument_number"],
            r["book"], r["page"], r["recording_date"], r["instrument_date"],
            _parse_list(r["grantors"]), _parse_list(r["grantees"]),
            r["legal_description"], r["section"], r["township"], r["range_val"],
            r["interest_conveyed"], r["lease_royalty_terms"], r["title_requirement_affected"],
            r["confidence_score"], "YES" if r["needs_review"] else "no",
        ]
        ws.append(row_data)
        # Highlight review rows
        if r["needs_review"]:
            for cell in ws[ws.max_row]:
                cell.fill = REVIEW_FILL
    _style_header_row(ws)
    _autofit(ws)


def _sheet_missing_images(wb: Workbook) -> None:
    ws = wb.create_sheet("Missing Images")
    ws.append(["Queue ID", "County", "Book", "Page", "Instrument #",
                "Image #", "Type", "Status", "Note"])
    rows = fetch_all(
        """SELECT q.id, q.county, q.book, q.page, q.instrument_number,
                  q.image_number, q.instrument_type, q.status
           FROM image_queue q
           WHERE q.status NOT IN ('downloaded','analyzed')"""
    )
    for r in rows:
        note = "Not yet approved" if r["status"] == "pending" else \
               "Approved but not downloaded" if r["status"] == "approved" else \
               "Download failed" if r["status"] == "failed" else r["status"]
        ws.append([r["id"], r["county"], r["book"], r["page"],
                   r["instrument_number"], r["image_number"],
                   r["instrument_type"], r["status"], note])
        if r["status"] == "failed":
            for cell in ws[ws.max_row]:
                cell.fill = ERROR_FILL
    _style_header_row(ws)
    _autofit(ws)


def _sheet_failed_downloads(wb: Workbook) -> None:
    ws = wb.create_sheet("Failed Downloads")
    ws.append(["Download ID", "Queue ID", "County", "Book", "Page",
                "Instrument #", "Error", "Attempted At"])
    rows = fetch_all(
        """SELECT d.id, d.queue_item_id, d.error_message, d.downloaded_at,
                  q.county, q.book, q.page, q.instrument_number
           FROM downloads d
           JOIN image_queue q ON q.id = d.queue_item_id
           WHERE d.status='failed'
           ORDER BY d.downloaded_at DESC"""
    )
    for r in rows:
        ws.append([r["id"], r["queue_item_id"], r["county"], r["book"],
                   r["page"], r["instrument_number"], r["error_message"],
                   r["downloaded_at"]])
        for cell in ws[ws.max_row]:
            cell.fill = ERROR_FILL
    _style_header_row(ws)
    _autofit(ws)


def _sheet_cost_report(wb: Workbook) -> None:
    ws = wb.create_sheet("Cost Report")
    ws.append(["Date", "API", "Amount ($)", "Description", "Queue Item"])
    rows = fetch_all(
        "SELECT created_at, api_type, amount, description, queue_item_id "
        "FROM cost_tracking ORDER BY created_at"
    )
    total_ok = 0.0
    total_ai = 0.0
    for r in rows:
        ws.append([r["created_at"], r["api_type"], r["amount"],
                   r["description"], r["queue_item_id"]])
        if r["api_type"] == "okcounty":
            total_ok += r["amount"]
        else:
            total_ai += r["amount"]
    # Summary rows
    ws.append([])
    ws.append(["", "OKCountyRecords Total", total_ok])
    ws.append(["", "OpenAI Total", total_ai])
    ws.append(["", "GRAND TOTAL", total_ok + total_ai])
    for row in ws.iter_rows(min_row=ws.max_row - 2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(bold=True)
    _style_header_row(ws)
    _autofit(ws)


def _sheet_doto_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("DOTO Update Summary")
    ws.append([
        "County", "Instrument Type", "Book", "Page", "Instrument #",
        "Recording Date", "Grantors", "Grantees", "Legal Description",
        "STR", "Interest Conveyed", "Confidence",
    ])
    rows = fetch_all(
        """SELECT a.*, q.county FROM analyses a
           JOIN image_queue q ON q.id = a.queue_item_id
           WHERE a.needs_review = 0
           ORDER BY q.county, a.recording_date, a.book, a.page"""
    )
    for r in rows:
        def pl(v):
            if not v:
                return ""
            try:
                lst = json.loads(v)
                return "; ".join(lst) if isinstance(lst, list) else str(v)
            except Exception:
                return str(v)

        str_label = "/".join(filter(None, [r["section"], r["township"], r["range_val"]]))
        ws.append([
            r["county"], r["instrument_type"], r["book"], r["page"],
            r["instrument_number"], r["recording_date"],
            pl(r["grantors"]), pl(r["grantees"]), r["legal_description"],
            str_label, r["interest_conveyed"], r["confidence_score"],
        ])
    _style_header_row(ws)
    _autofit(ws)


def _sheet_needs_review(wb: Workbook) -> None:
    ws = wb.create_sheet("Needs Human Review")
    ws.append([
        "Analysis ID", "Queue ID", "County", "Book", "Page", "Instrument #",
        "Type", "Confidence", "PDF Path", "Analyzed At",
    ])
    rows = fetch_all(
        """SELECT a.id, a.queue_item_id, a.instrument_type, a.instrument_number,
                  a.book, a.page, a.confidence_score, a.analyzed_at,
                  q.county, d.file_path
           FROM analyses a
           JOIN image_queue q ON q.id = a.queue_item_id
           LEFT JOIN downloads d ON d.id = a.download_id
           WHERE a.needs_review = 1
           ORDER BY a.confidence_score ASC"""
    )
    for r in rows:
        ws.append([r["id"], r["queue_item_id"], r["county"], r["book"],
                   r["page"], r["instrument_number"], r["instrument_type"],
                   r["confidence_score"], r["file_path"] or "", r["analyzed_at"]])
        for cell in ws[ws.max_row]:
            cell.fill = REVIEW_FILL
    _style_header_row(ws)
    _autofit(ws)


# ── Public entry points ───────────────────────────────────────────────────────

def generate_full_report(output_path: Optional[str] = None) -> str:
    """Generate all report sheets into one Excel workbook. Returns file path."""
    config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    if not output_path:
        output_path = str(config.REPORTS_DIR / f"DOTO_Report_{_ts()}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _sheet_image_index(wb)
    _sheet_analysis_results(wb)
    _sheet_missing_images(wb)
    _sheet_failed_downloads(wb)
    _sheet_cost_report(wb)
    _sheet_doto_summary(wb)
    _sheet_needs_review(wb)

    wb.save(output_path)
    record("report_generated", f"Full report saved to {output_path}")
    return output_path


def write_results_to_input_excel(source_xlsx: str, output_path: Optional[str] = None) -> str:
    """
    Read original pull-list XLSX, append analysis columns, save to new file.
    Useful for returning enriched data in the original spreadsheet's format.
    """
    df_orig = pd.read_excel(source_xlsx, dtype=str)

    analyses = fetch_all(
        """SELECT a.instrument_number, a.book, a.page, a.recording_date,
                  a.instrument_date, a.grantors, a.grantees, a.legal_description,
                  a.section, a.township, a.range_val, a.interest_conveyed,
                  a.lease_royalty_terms, a.title_requirement_affected,
                  a.confidence_score, a.needs_review
           FROM analyses a"""
    )
    df_analysis = pd.DataFrame(analyses)
    if not df_analysis.empty:
        df_analysis["grantors"] = df_analysis["grantors"].apply(
            lambda v: "; ".join(json.loads(v)) if v and v.startswith("[") else (v or ""))
        df_analysis["grantees"] = df_analysis["grantees"].apply(
            lambda v: "; ".join(json.loads(v)) if v and v.startswith("[") else (v or ""))

    if not output_path:
        config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        output_path = str(config.REPORTS_DIR / f"DOTO_Results_{_ts()}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_orig.to_excel(writer, sheet_name="Original Pull List", index=False)
        if not df_analysis.empty:
            df_analysis.to_excel(writer, sheet_name="Analysis Results", index=False)

    record("report_results_excel", f"Results written to {output_path}")
    return output_path
