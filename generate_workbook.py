import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_section32_challenger_workbook(output_path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    
    # 13 required sheets in exact order
    sheet_names = [
        'Overview', 'Title ', 'OGL', 'PLAT', 'Runsheet',
        'Tract 1', 'Tract 2', 'Tract 3', 'Tract 4', 'Tract 5',
        'WI 2', 'WI 1', 'Well 1'
    ]
    
    sheets = {}
    for name in sheet_names:
        sheets[name] = wb.create_sheet(title=name)
    wb.remove(default_sheet)
    
    # Common Styling
    font_title = Font(name='Calibri', size=14, bold=True, color='1F4E78')
    font_subtitle = Font(name='Calibri', size=11, bold=True, color='333333')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_bold = Font(name='Calibri', size=11, bold=True, color='000000')
    font_regular = Font(name='Calibri', size=10, color='000000')
    font_small = Font(name='Calibri', size=9, color='555555')
    font_hold = Font(name='Calibri', size=11, bold=True, color='C00000')
    
    fill_header_navy = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    fill_header_gray = PatternFill(start_color='595959', end_color='595959', fill_type='solid')
    fill_accent = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    fill_warning = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    fill_zebra = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    double_bottom_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='double', color='000000')
    )
    
    # -------------------------------------------------------------
    # 1. OVERVIEW SHEET
    # -------------------------------------------------------------
    ws = sheets['Overview']
    ws.views.sheetView[0].showGridLines = True
    
    ws['B1'] = "SECTION 32-11N-25W, BECKHAM COUNTY, OKLAHOMA — CURSORY TITLE REPORT"
    ws['B1'].font = font_title
    ws['B2'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE"
    ws['B2'].font = font_hold
    
    ws['B4'] = "EXECUTIVE SUMMARY & DECISION-ORIENTED CONTROL REGISTER"
    ws['B4'].font = font_subtitle
    
    overview_meta = [
        ("Prospect / Target Area:", "Section 32, Township 11 North, Range 25 West, Beckham County, OK"),
        ("Gross Section Area:", "640.00 nominal gross acres (Section-wide regulatory scope)"),
        ("Primary Reviewed Entity:", "Diversified Production LLC / Predecessor Working Interest Chain"),
        ("Review Effective Date:", "August 6, 2026 (Forward record search cutoff March 2, 2023)"),
        ("Challenger Approach:", "Contradiction-First & Evidence-Grounded Reconciliation (Gemini 3.7 Flash)"),
        ("Mandatory Classification:", "FOR REVIEW — HOLD NO EXTERNAL RELEASE")
    ]
    for idx, (label, val) in enumerate(overview_meta, start=5):
        ws[f'B{idx}'] = label
        ws[f'B{idx}'].font = font_bold
        ws[f'D{idx}'] = val
        ws[f'D{idx}'].font = font_regular
        
    ws['B12'] = "CATEGORY"
    ws['C12'] = "LEGAL & TITLE DETERMINATION"
    ws['D12'] = "EVIDENCE STATUS"
    ws['E12'] = "SUPPORTING BASIS & REVIEWS"
    for col in ['B', 'C', 'D', 'E']:
        cell = ws[f'{col}12']
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    overview_rows = [
        ("ESTABLISHED", "Nominal 640.00 gross acre section framework across 5 primary fee tract subdivisions.", "PROVED", "Patent 3/469 (Cornelius, NE/4) and Land Office Receiver Receipts 3894, 3895, Biggs."),
        ("ESTABLISHED", "Diversified Production LLC is the latest scheduled asset claimant for OK48147.001.1.", "PROVED (Historical)", "Direct asset schedule entries at Bk 2400/Pg 566 and Bk 2395/Pg 462."),
        ("QUALIFIED / APPARENT", "Mineral ownership estimates by tract based on last-located record vesting instruments.", "ESTIMATED", "V7/V10 ending schedules carry 160 ac (T1), 80 ac (T2), 80 ac (T3), 80 ac (T4), 80 ac (T5)."),
        ("QUALIFIED / APPARENT", "Historical oil and gas leases (14 principal OGLs) recorded between 1947 and 1979.", "ESTIMATED", "Creating terms extracted from direct faces; modern continuation is qualified."),
        ("QUALIFIED / APPARENT", "Regulatory context includes OCC Order 156126 (1979) and producing wells (Latigo / BCE-Mach II).", "PROVED (Regulatory)", "Official OCC Order 156126 and OTC production context established."),
        ("NOT DETERMINED", "Current section-wide Diversified Working Interest (WI), Net Revenue Interest (NRI), & ORRI.", "UNPROVED / BLANK", "Predecessor 15.000875% and 51.25% transaction branches require missing Tapstone schedules."),
        ("NOT DETERMINED", "Current lease-by-lease HBP validity across all tracts and depths.", "UNPROVED / OPEN", "Blanket production does not prove section-wide HBP; wellbore & depth exclusions open."),
        ("NOT DETERMINED", "Post-cutoff divestiture status under Bk 2434/Pg 751 (Diversified/DP to Teocalli).", "CRITICAL OPEN ITEM", "Index-only lead; if Section 32 is included in exhibit, Diversified interest was conveyed out."),
        ("TOP CURATIVE ITEM 1", "Acquire complete unredacted Exhibit A to Assignment Bk 1697/Pg 236 (Staghorn to CHK).", "CURATIVE QUEUE", "Required to bridge base OGLs to modern Chesapeake/Tapstone leasehold."),
        ("TOP CURATIVE ITEM 2", "Review assignment instrument and schedule at Bk 2434/Pg 751 (Teocalli Exploration).", "CURATIVE QUEUE", "P0 divestiture determination for Diversified working interest."),
        ("TOP CURATIVE ITEM 3", "Obtain missing Beckham County Index Logical Page 48 and resolve disordered Page 54.", "CURATIVE QUEUE", "Complete county index accounting across all 1,981 historical entries."),
        ("TOP CURATIVE ITEM 4", "Obtain complete probate decrees for Ellen W. Crook (83/259) and Charlie B. Crook (845/150).", "CURATIVE QUEUE", "Required to establish present lawful fractional mineral title for Tract 1 heirs.")
    ]
    
    for r_idx, (cat, det, st, bas) in enumerate(overview_rows, start=13):
        ws[f'B{r_idx}'] = cat
        ws[f'C{r_idx}'] = det
        ws[f'D{r_idx}'] = st
        ws[f'E{r_idx}'] = bas
        
        ws[f'B{r_idx}'].font = font_bold
        ws[f'C{r_idx}'].font = font_regular
        ws[f'D{r_idx}'].font = font_bold if "NOT" in st or "CRITICAL" in st else font_regular
        ws[f'E{r_idx}'].font = font_small
        
        if "ESTABLISHED" in cat:
            ws[f'B{r_idx}'].fill = fill_accent
        elif "QUALIFIED" in cat:
            ws[f'B{r_idx}'].fill = fill_zebra
        elif "NOT DETERMINED" in cat:
            ws[f'B{r_idx}'].fill = fill_warning
        else:
            ws[f'B{r_idx}'].fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{r_idx}'].border = thin_border
            ws[f'{col}{r_idx}'].alignment = Alignment(vertical='top', wrap_text=True)

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 50

    # -------------------------------------------------------------
    # 2. TITLE SHEET (with trailing space in name: 'Title ')
    # -------------------------------------------------------------
    ws = sheets['Title ']
    ws.views.sheetView[0].showGridLines = True
    
    ws['A1'] = "SECTION 32-11N-25W, BECKHAM COUNTY, OKLAHOMA — ESTIMATED MINERAL TITLE BY TRACT"
    ws['A1'].font = font_title
    ws['A2'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE • Client focus: Diversified • Effective August 6, 2026 • EST. = Last-Located Record Owner (Not Present Confirmed Title)"
    ws['A2'].font = font_hold
    
    title_headers = [
        "Tract", "Estimated / Last-Located Mineral Owner", "EST. NMA", "Tract Gross Ac", 
        "OGL / Lease Ref", "Royalty / Lease Status", "Mailing Address of Record", "Vesting Source & Visible Assumptions"
    ]
    
    for c_idx, h in enumerate(title_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    title_data = [
        # Tract 1
        ("Tract 1 (NE/4)", "Unresolved successors of Adolf Bollenbach (WD 10/255, 01/31/1910)", 65.0, 160.0, "None", "Status not determined", "No address of record", "ASSUMED: 65.000000 NMA residual vests in last traced record owner (WD 10/255); downstream chain-out open. Current title NOT DETERMINED."),
        ("Tract 1 (NE/4)", "C. B. Crook, Jr. (Heir of Charlie B. Crook)", 8.0, 160.0, "OGL 2 (264/345)", "1/8; Cursory Estimate; HBP Unconfirmed", "Address not located", "EST. last located record owner; Vested: DECR 845/150-152 and 845/157. Current status not confirmed at 08/06/2026."),
        ("Tract 1 (NE/4)", "Carole Faye Hanks (Heir of Charlie B. Crook)", 8.0, 160.0, "OGL 2 (264/345)", "1/8; Cursory Estimate; HBP Unconfirmed", "Address not located", "EST. last located record owner; Vested: DECR 845/150-152 and 845/157. Current status not confirmed at 08/06/2026."),
        ("Tract 1 (NE/4)", "Doris Marie Lenderman (Heir of Charlie B. Crook)", 8.0, 160.0, "OGL 2 (264/345)", "1/8; Cursory Estimate; HBP Unconfirmed", "Address not located", "EST. last located record owner; Vested: DECR 845/150-152 and 845/157. Current status not confirmed at 08/06/2026."),
        ("Tract 1 (NE/4)", "Ina Bell Bryan (Heir of Charlie B. Crook)", 8.0, 160.0, "OGL 2 (264/345)", "1/8; Cursory Estimate; HBP Unconfirmed", "Address not located", "EST. last located record owner; Vested: DECR 845/150-152 and 845/157. Current status not confirmed at 08/06/2026."),
        ("Tract 1 (NE/4)", "Mary Joyce Higgenbotham (Heir of Charlie B. Crook)", 8.0, 160.0, "OGL 2 (264/345)", "1/8; Cursory Estimate; HBP Unconfirmed", "Address not located", "EST. last located record owner; Vested: DECR 845/150-152 and 845/157. Current status not confirmed at 08/06/2026."),
        ("Tract 1 (NE/4)", "F. L. Randel & J. L. Randel, share & share alike", 1.5, 160.0, "OGL 9, 10 (352/719, 721)", "1/8; Assumed effective; HBP Unconfirmed", "Wichita Falls, TX (1950 MD recital)", "EST. last located record owner; Vested: MD 92/47. Forward record search cutoff 03/02/2023."),
        ("Tract 1 (NE/4)", "Great Western Oil & Gas, Inc.", 10.0, 160.0, "OGL 11 (376/369)", "1/8; Assumed effective; HBP Unconfirmed", "Address not located", "EST. last located record owner; Vested: MD 367/632 (image 1558). Current corporate succession open."),
        ("Tract 1 (NE/4)", "Jean Jenkins Shearer, Florence J. Chunn, Wm. R. Jenkins", 2.0, 160.0, "None", "No located active OGL", "Address not located", "EST. last located record owner; Vested: MD 250/26. Unleased mineral interest or unrecorded lease."),
        ("Tract 1 (NE/4)", "Unresolved probate beneficiaries under Decree 83/259 (Ellen W. Crook)", 40.0, 160.0, "None", "Status not determined", "No address of record", "ASSUMED: 40.000000 NMA vests collectively per decree; individual allocations not stated on face. Current title OPEN."),
        ("Tract 1 (NE/4)", "Unresolved record owner (Grantee of F.L. Randel deed, img 1794)", 1.5, 160.0, "None", "Status not determined", "No address of record", "ASSUMED: 1.500000 NMA per recorded deed; grantee name unextracted from face. Curative required."),
        
        # Tract 2
        ("Tract 2 (N/2 NW/4)", "B. E. Maddoux (Stipulated residual)", 10.0, 80.0, "None", "No located active OGL", "407 N. Broadway, Sayre, OK (1947 recital)", "EST. last located record owner; Vested: STIPULATION 153/161-162. Current status not proved."),
        ("Tract 2 (N/2 NW/4)", "D. H. Maddoux (Stipulated residual)", 10.0, 80.0, "None", "No located active OGL", "Address not located", "EST. last located record owner; Vested: STIPULATION 153/161-162. Current status not proved."),
        ("Tract 2 (N/2 NW/4)", "George T. Harrison Jr. & Renata Harrison", 20.0, 80.0, "OGL 6 (340/323)", "1/8; Assumed effective; HBP Unconfirmed", "Address not located", "EST. last located record owner; Vested: MD 92/47. Covers N/2 NW/4 and N/2 SW/4."),
        ("Tract 2 (N/2 NW/4)", "Edgar Sanditen, Julius Sanditen, Herman Sanditen, Morris Sanditen", 20.0, 80.0, "None", "Status not determined", "Tulsa, OK (historic recital)", "EST. last located record owner; Vested: MD 85/363-366. Probate and corporate succession open."),
        ("Tract 2 (N/2 NW/4)", "LeGrand / Keathley Heirs (Ola Keathley Estate)", 20.0, 80.0, "None", "Status not determined", "Sayre, OK (historic recital)", "EST. last located record owner; Vested: P-76-78, Pgs 349-350. Final distribution decree incomplete."),
        
        # Tract 3
        ("Tract 3 (S/2 NW/4)", "C. R. Bennett (Historic mineral grantee)", 40.0, 80.0, "None", "No located active OGL", "Oklahoma City, OK (historic recital)", "EST. last located record owner; Vested: MD 81/194. Forward record search cutoff 03/02/2023."),
        ("Tract 3 (S/2 NW/4)", "Heller Company (Historic entity)", 20.0, 80.0, "None", "Status not determined", "Tulsa, OK (historic recital)", "EST. last located record owner; Vested: MD 85/222. Entity dissolution and succession open."),
        ("Tract 3 (S/2 NW/4)", "Bettye Folmar Norris", 20.0, 80.0, "None", "Status not determined", "Address not located", "EST. last located record owner; Vested: DECR 262/490-493. Current heirs/successors not proved."),
        
        # Tract 4
        ("Tract 4 (SW/4)", "Gertrude L. LaFortune & Gertrude LaFortune, widow", 40.0, 160.0, "OGL 4, 5 (340/302, 304)", "3/16; Assumed effective; HBP Unconfirmed", "Tulsa, OK (historic recital)", "EST. last located record owner; Vested: MD 85/222 & probate. Overlapping N/2 SW/4 leases."),
        ("Tract 4 (SW/4)", "Carol Seidenbach Leach & Cynthia Seidenbach Kruse", 40.0, 160.0, "OGL 7, 8 (342/564, 566)", "3/16; Assumed effective; HBP Unconfirmed", "Address not located", "EST. last located record owner; Vested: Probate decree. N/2 SW/4 leasehold."),
        ("Tract 4 (SW/4)", "Cora B. LeGrand (W/2 SE/4 & S/2 SW/4)", 80.0, 160.0, "OGL 3 (307/31)", "1/8; Assumed effective; HBP Unconfirmed", "Sayre, OK (historic recital)", "EST. last located record owner; Vested: Patent/Deed chain. S/2 SW/4 and W/2 SE/4."),
        
        # Tract 5
        ("Tract 5 (SE/4)", "Cora B. Garrett & S. A. Garrett", 80.0, 160.0, "OGL 1 (63/33)", "1/8; 10-Yr Term (1947); HBP Unconfirmed", "Sayre, OK (historic recital)", "EST. last located record owner; Vested: Biggs receiver receipt / deed chain. Covers E/2 SE/4."),
        ("Tract 5 (SE/4)", "Cora B. LeGrand (E/2 SE/4 & W/2 SE/4 split)", 80.0, 160.0, "OGL 3 (307/31)", "1/8; Assumed effective; HBP Unconfirmed", "Sayre, OK (historic recital)", "EST. last located record owner; Vested: Deed 307/31 chain. Covers W/2 SE/4.")
    ]
    
    start_row = 5
    for idx, r in enumerate(title_data, start=start_row):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if c_idx == 3 or c_idx == 4:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.000000'
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
    # Total row
    tot_row = start_row + len(title_data)
    ws.cell(row=tot_row, column=1, value="TOTAL")
    ws.cell(row=tot_row, column=2, value="ESTIMATED SECTION 32 RECORD RECONCILIATION TOTAL")
    ws.cell(row=tot_row, column=3, value=f"=SUM(C{start_row}:C{tot_row-1})")
    ws.cell(row=tot_row, column=4, value=640.0)
    ws.cell(row=tot_row, column=5, value="—")
    ws.cell(row=tot_row, column=6, value="—")
    ws.cell(row=tot_row, column=7, value="—")
    ws.cell(row=tot_row, column=8, value="Note: Arithmetic closure to 640.00 NMA represents record-reconstruction balance, NOT proof of current unclouded marketable title.")
    
    for c_idx in range(1, 9):
        cell = ws.cell(row=tot_row, column=c_idx)
        cell.font = font_bold
        cell.fill = fill_accent
        cell.border = double_bottom_border
        if c_idx in [3, 4]:
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0.000000'

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 32
    ws.column_dimensions['H'].width = 50

    # -------------------------------------------------------------
    # 3. OGL SHEET
    # -------------------------------------------------------------
    ws = sheets['OGL']
    ws.views.sheetView[0].showGridLines = True
    
    ws['A1'] = "OIL & GAS LEASE (OGL) MASTER SCHEDULE — SECTION 32-11N-25W"
    ws['A1'].font = font_title
    ws['A2'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE • Calculated Primary Term Expirations Are Inferred Mathematical Dates, Not Legal Proof of Expiration/HBP Continuation"
    ws['A2'].font = font_hold
    
    ogl_headers = [
        "OGL #", "Image / Ref", "Instrument Type", "Book / Page", "Lease Date", "Recording Date", 
        "Lessor", "Original Lessee", "Tract / Legal Description", "Gross Ac", "Royalty", "Primary Term", "Calculated Primary Exp.", "Current Qualified Status"
    ]
    
    for c_idx, h in enumerate(ogl_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    ogl_data = [
        (1, "Img 0353", "Oil and Gas Lease", "63/33", "1947-11-13", "1947-11-26", "Cora B. Garrett; S.A. Garrett", "C.E. McKinley", "E/2 SE/4 (Tract 5)", 80.0, "1/8", "10 Years", "1957-11-13", "Historical Lease — HBP Status Unconfirmed"),
        (2, "Imgs 1197-98", "Oil and Gas Lease", "264/345", "1972-02-05", "1972-02-22", "Charlie B. Crook; Ima Pearl Crook", "Union Oil Co. of California", "NE/4 (Tract 1)", 160.0, "1/8", "5 Years", "1977-02-05", "Historical Lease — HBP Status Unconfirmed"),
        (3, "Img 1372", "Oil and Gas Lease", "307/31", "1976-04-29", "1976-06-09", "Cora B. LeGrand", "Union Oil Co. of California", "W/2 SE/4 & S/2 SW/4 (Tracts 4, 5)", 160.0, "1/8", "3 Years", "1979-04-29", "Historical Lease — HBP Status Unconfirmed"),
        (4, "Imgs 1522-23", "Oil and Gas Lease", "340/302", "1977-10-14", "Not extracted", "Gertrude L. LaFortune et al.", "Union Oil Co. of California", "N/2 SW/4 (Tract 4)", 80.0, "3/16", "3 Years", "1980-10-14", "Historical Lease — HBP Status Unconfirmed"),
        (5, "Imgs 1524-25", "Oil and Gas Lease", "340/304", "1977-10-14", "Not extracted", "Gertrude LaFortune, widow", "Union Oil Co. of California", "N/2 SW/4 (Tract 4)", 80.0, "3/16", "3 Years", "1980-10-14", "Historical Lease — HBP Status Unconfirmed"),
        (6, "Imgs 1526-27", "Oil and Gas Lease", "340/323", "1977-09-29", "1977-11-07", "George T. Harrison Jr.; Renata Harrison", "Union Oil Co. of California", "N/2 NW/4 & N/2 SW/4 (Tracts 2, 4)", 160.0, "1/8", "3 Years", "1980-09-29", "Historical Lease — HBP Status Unconfirmed"),
        (7, "Imgs 1542-43", "Oil and Gas Lease", "342/564", "1977-09-29", "1978-02-09", "Carol Seidenbach Leach", "Union Oil Co. of California", "N/2 SW/4 (Tract 4)", 80.0, "3/16", "3 Years", "1980-09-29", "Historical Lease — HBP Status Unconfirmed"),
        (8, "Imgs 1544-45", "Oil and Gas Lease", "342/566", "1977-11-01", "1978-02-09", "Cynthia Seidenbach Kruse", "Union Oil Co. of California", "N/2 SW/4 (Tract 4)", 80.0, "3/16", "3 Years", "1980-11-01", "Historical Lease — HBP Status Unconfirmed"),
        (9, "Imgs 1552-53", "Oil and Gas Lease", "352/719", "1978-05-16", "1978-06-27", "F.L. Randel", "Union Oil Co. of California", "Part NE/4 (Tract 1)", 1.5, "1/8", "3 Years", "1981-05-16", "Historical Lease — HBP Status Unconfirmed"),
        (10, "Imgs 1554-55", "Oil and Gas Lease", "352/721", "1978-05-16", "1978-06-27", "J.L. Randel", "Union Oil Co. of California", "Part NE/4 (Tract 1)", 1.5, "1/8", "3 Years", "1981-05-16", "Historical Lease — HBP Status Unconfirmed"),
        (11, "Imgs 1569-70", "Oil and Gas Lease", "376/369", "1979-03-21", "1979-04-30", "Great Western Oil & Gas, Inc.", "Union Oil Co. of California", "Part NE/4 (Tract 1)", 10.0, "1/8", "3 Years", "1982-03-21", "Historical Lease — HBP Status Unconfirmed"),
        (12, "Imgs 1457-58", "Oil and Gas Lease", "332/74-75", "1977-08-01", "1977-09-14", "Ralph W. Viersen Jr. et al.", "Leede Exploration", "NW/4 & SW/4 (Tracts 2, 3, 4)", 320.0, "3/16", "3 Years", "1980-08-01", "Historical Lease — Owner Coverage Defect; Status Open"),
        (13, "Imgs 1568/1812", "Oil and Gas Lease (Recited)", "341/538", "1977-11-29", "Not extracted", "R.L. Minton; Ella Minton", "John E. Hartman", "N/2 SW/4 & N/2 NW/4 (Tracts 2, 4)", 160.0, "1/8", "3 Years", "1980-11-29", "Creating Face Absent; Status Unconfirmed"),
        (14, "Bk 1697/Pg 236", "Assignment (Prior Ref)", "1697/236", "2001-08-20", "2001-08-29", "Staghorn Resources LLC", "Chesapeake Exploration LP", "Section 32 scheduled lands", 640.0, "Various", "Assignment", "N/A", "Modern Master Assignment Predecessor; Exhibit A Required")
    ]
    
    for idx, r in enumerate(ogl_data, start=5):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if c_idx in [1, 10]:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(vertical='top')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 26
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 35

    # -------------------------------------------------------------
    # 4. PLAT SHEET
    # -------------------------------------------------------------
    ws = sheets['PLAT']
    ws.views.sheetView[0].showGridLines = True
    
    ws['B2'] = "SECTION 32, TOWNSHIP 11 NORTH, RANGE 25 WEST, BECKHAM COUNTY, OKLAHOMA"
    ws['B2'].font = font_title
    ws['B3'] = "LAND TRACT SUBDIVISION PLAT & GEOMETRY DIAGRAM (640.00 GROSS ACRES)"
    ws['B3'].font = font_subtitle
    ws['B4'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE"
    ws['B4'].font = font_hold
    
    # Textual / Cell Plat representation
    # Northwest Quarter (Tracts 2 & 3)
    ws['B7'] = "TRACT 2: N/2 NW/4 (80.00 AC)"
    ws['B7'].font = font_bold
    ws['B8'] = "Historic Owner: Herman Stevens (FR 3894)\nEnding Mineral Schedule: 80.00 NMA\nLeases: OGL 6 (340/323), OGL 12 (332/74)\nWells: JGL 1-32, Legrand 1-32"
    ws['B8'].font = font_small
    ws['B8'].alignment = Alignment(wrap_text=True)
    
    ws['B13'] = "TRACT 3: S/2 NW/4 (80.00 AC)"
    ws['B13'].font = font_bold
    ws['B14'] = "Historic Owner: Herman Stevens (FR 3894)\nEnding Mineral Schedule: 80.00 NMA\nLeases: OGL 12 (332/74 Viersen)\nWells: Legrand 2-32"
    ws['B14'].font = font_small
    ws['B14'].alignment = Alignment(wrap_text=True)
    
    # Northeast Quarter (Tract 1)
    ws['F7'] = "TRACT 1: NE/4 (160.00 AC)"
    ws['F7'].font = font_bold
    ws['F8'] = "Historic Owner: Cornelius B. Cornelius (Patent 3/469)\nEnding Mineral Schedule: 160.00 NMA (Crook / Bollenbach / Randel)\nLeases: OGL 2 (264/345), OGL 9, 10 (352/719, 721), OGL 11 (376/369)\nWells: Twin 1-32 (BCE-Mach II), Katie 1-32 (BP America)"
    ws['F8'].font = font_small
    ws['F8'].alignment = Alignment(wrap_text=True)
    
    # Southwest Quarter (Tract 4)
    ws['B19'] = "TRACT 4: N/2 SW/4 & S/2 SW/4 (160.00 AC)"
    ws['B19'].font = font_bold
    ws['B20'] = "Historic Owner: Augustus P. Stevens (FR 3895)\nEnding Mineral Schedule: 160.00 NMA (LaFortune / Leach / Kruse / LeGrand)\nLeases: OGL 3 (307/31), OGL 4, 5 (340/302, 304), OGL 7, 8 (342/564, 566)\nWells: Crook 1-32 (Complete Oil & Gas), Crook-Legrand 1-32"
    ws['B20'].font = font_small
    ws['B20'].alignment = Alignment(wrap_text=True)
    
    # Southeast Quarter (Tract 5)
    ws['F19'] = "TRACT 5: E/2 SE/4 & W/2 SE/4 (160.00 AC)"
    ws['F19'].font = font_bold
    ws['F20'] = "Historic Owner: Velmore Biggs (FR Receiver Receipt; Patent Open)\nEnding Mineral Schedule: 160.00 NMA (Garrett / LeGrand)\nLeases: OGL 1 (63/33 Garrett), OGL 3 (307/31 LeGrand)\nWells: Cora Garrett C-310 (BP America), Pearl 1-32 (SM Energy)"
    ws['F20'].font = font_small
    ws['F20'].alignment = Alignment(wrap_text=True)
    
    for box_cells, fill_c in [
        (['B7', 'B8', 'B9', 'B10', 'B11', 'C7', 'C8', 'C9', 'C10', 'C11', 'D7', 'D8', 'D9', 'D10', 'D11'], 'D9E1F2'),
        (['B13', 'B14', 'B15', 'B16', 'B17', 'C13', 'C14', 'C15', 'C16', 'C17', 'D13', 'D14', 'D15', 'D16', 'D17'], 'F2F2F2'),
        (['F7', 'F8', 'F9', 'F10', 'F11', 'F12', 'F13', 'F14', 'F15', 'F16', 'F17', 'G7', 'G8', 'G9', 'G10', 'G11', 'G12', 'G13', 'G14', 'G15', 'G16', 'G17'], 'FFF2CC'),
        (['B19', 'B20', 'B21', 'B22', 'B23', 'C19', 'C20', 'C21', 'C22', 'C23', 'D19', 'D20', 'D21', 'D22', 'D23'], 'E2EFDA'),
        (['F19', 'F20', 'F21', 'F22', 'F23', 'G19', 'G20', 'G21', 'G22', 'G23'], 'FCE4D6')
    ]:
        f_p = PatternFill(start_color=fill_c, end_color=fill_c, fill_type='solid')
        for cell_ref in box_cells:
            ws[cell_ref].fill = f_p
            ws[cell_ref].border = thin_border

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 4
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25

    # -------------------------------------------------------------
    # 5. RUNSHEET SHEET
    # -------------------------------------------------------------
    ws = sheets['Runsheet']
    ws.views.sheetView[0].showGridLines = True
    
    ws['A1'] = "SECTION 32-11N-25W, BECKHAM COUNTY, OKLAHOMA — MASTER INSTRUMENT RUNSHEET"
    ws['A1'].font = font_title
    ws['A2'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE • Preserved 140-Row Chronological Index & Extraction Corpus"
    ws['A2'].font = font_hold
    
    runsheet_headers = [
        "Seq", "OGL Ref", "Doc Type", "Book / Page", "Instrument Date", "Recording Date", 
        "Grantor", "Grantee", "Legal Description / Coverage", "Operative Title Effect & Remarks", "Evidence Tier / Status"
    ]
    
    for c_idx, h in enumerate(runsheet_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    runsheet_data = [
        (1, "—", "PAT", "Patent Bk 3/469", "1907-03-22", "1907-06-08", "United States", "Cornelius B. Cornelius", "NE/4 Section 32 (160 ac)", "Homestead Patent; Sovereign inception of fee simple estate. No mineral reservation apparent.", "DIRECT IMAGE P-001 — PROVED"),
        (2, "—", "FR", "Image 0004", "1906-06-28", "1906-08-08", "U.S. Land Office", "Velmore Biggs", "SE/4 Section 32 (160 ac)", "Final Receiver Receipt; Sovereign inception evidence. Underlying patent remains OPEN.", "DIRECT IMAGE P-002 — PROVED"),
        (3, "—", "FR", "Image 0005", "1905-05-04", "1905-05-12", "U.S. Land Office", "Herman Stevens", "NW/4 Section 32 (160 ac)", "Final Receiver Receipt No. 3894; Sovereign inception evidence. Patent remains OPEN.", "DIRECT IMAGE P-003 — PROVED"),
        (4, "—", "FR", "Image 0006", "1905-05-04", "1905-05-12", "U.S. Land Office", "Augustus P. Stevens", "SW/4 Section 32 (160 ac)", "Final Receiver Receipt No. 3895; Sovereign inception evidence. Patent remains OPEN.", "DIRECT IMAGE P-004 — PROVED"),
        (5, "—", "WD", "Roger Mills 11/214", "1907-07-16", "1907-07-18", "Cornelius & Margreta Cornelius", "Jacob Bollenbach", "NE/4 Section 32 (160 ac)", "Warranty Deed in pre-Beckham county records. Conveys full fee estate without reservation.", "DIRECT IMAGE — PROVED"),
        (6, "—", "WD", "Bk 10/Pg 255", "1910-01-31", "1910-02-15", "Jacob Bollenbach", "Adolf Bollenbach", "NE/4 Section 32 (160 ac)", "Warranty Deed conveying fee estate. Downstream mineral chain breaks here (65.0 NMA unchained).", "DIRECT IMAGE — PROVED"),
        (7, "OGL 1", "OGL", "Bk 63/Pg 33", "1947-11-13", "1947-11-26", "Cora B. Garrett & S.A. Garrett", "C.E. McKinley", "E/2 SE/4 (80 ac)", "Oil and Gas Lease; 10-year primary term; 1/8 royalty. Base lease for Garrett branch.", "DIRECT IMAGE 0353 — PROVED"),
        (8, "—", "MD", "Bk 81/Pg 194", "1949-06-12", "1949-06-25", "Historic Grantors", "C. R. Bennett", "S/2 NW/4 (Tract 3)", "Mineral Deed conveying 40.00 NMA in S/2 NW/4.", "DIRECT IMAGE — PROVED"),
        (9, "—", "MD", "Bk 85/Pg 222", "1950-02-10", "1950-03-01", "Historic Grantors", "Heller Company; Gertrude LaFortune", "SW/4 Section 32 (Tract 4)", "Mineral Deed conveying fractional mineral interests across SW/4.", "DIRECT IMAGE — PROVED"),
        (10, "—", "MD", "Bk 92/Pg 47", "1950-11-04", "1950-11-15", "Historic Grantors", "F.L. Randel & J.L. Randel", "NE/4 Section 32 (Tract 1)", "Mineral Deed conveying 3.00 NMA (1.50 NMA each) in NE/4.", "DIRECT IMAGE — PROVED"),
        (11, "—", "STIP", "Bk 153/Pg 161", "1958-04-18", "1958-05-02", "Maddoux / Farrar Parties", "B.E. Maddoux & D.H. Maddoux", "N/2 NW/4 (Tract 2)", "Stipulation of Interest settling 10.00 NMA each to B.E. and D.H. Maddoux.", "DIRECT IMAGE — PROVED"),
        (12, "OGL 2", "OGL", "Bk 264/Pg 345", "1972-02-05", "1972-02-22", "Charlie B. Crook & Ima Pearl Crook", "Union Oil Co. of California", "NE/4 Section 32 (Tract 1)", "Oil and Gas Lease; 5-year primary term; 1/8 royalty. Covers Crook family interest.", "DIRECT IMAGE 1197 — PROVED"),
        (13, "OGL 3", "OGL", "Bk 307/Pg 31", "1976-04-29", "1976-06-09", "Cora B. LeGrand", "Union Oil Co. of California", "W/2 SE/4 & S/2 SW/4 (Tracts 4, 5)", "Oil and Gas Lease; 3-year primary term; 1/8 royalty.", "DIRECT IMAGE 1372 — PROVED"),
        (14, "OGL 12", "OGL", "Bk 332/Pg 74", "1977-08-01", "1977-09-14", "Ralph W. Viersen Jr. et al.", "Leede Exploration", "NW/4 & SW/4 Section 32", "Oil and Gas Lease; 3-year primary term; 3/16 royalty. Owner coverage defect noted in V10.", "DIRECT IMAGE 1457 — PROVED"),
        (15, "OGL 6", "OGL", "Bk 340/Pg 323", "1977-09-29", "1977-11-07", "George T. Harrison Jr. & Renata Harrison", "Union Oil Co. of California", "N/2 NW/4 & N/2 SW/4 (Tracts 2, 4)", "Oil and Gas Lease; 3-year primary term; 1/8 royalty.", "DIRECT IMAGE 1526 — PROVED"),
        (16, "—", "ASG", "Bk 376/Pg 287", "1979-04-05", "1979-04-20", "John E. Hartman", "Hunt Energy Corporation", "N/2 SW/4 & N/2 NW/4 (160 ac)", "Assignment of OGL; assigns RTI reserving 0.0125% of 8/8 ORRI (literal 0.000125 decimal).", "DIRECT IMAGE 1568 — PROVED"),
        (17, "—", "OCC", "Order 156126", "1979-08-06", "Not recorded", "Oklahoma Corporation Commission", "All Owners / Unit Participants", "All Section 32 (640 ac)", "Statutory Forced Pooling Order for multi-formation drilling unit. 180-day commencement.", "OFFICIAL OCC PDF — PROVED"),
        (18, "—", "DECR", "Bk 845/Pg 150", "1985-10-23", "1985-11-04", "District Court Beckham Co.", "Heirs of Charlie B. Crook", "NE/4 Section 32 (Tract 1)", "Final Decree distributing 8.00 NMA each to C.B. Jr., Hanks, Lenderman, Bryan, Higgenbotham.", "DIRECT IMAGE — PROVED"),
        (19, "—", "ASG", "Bk 872/Pg 279", "1985-10-29", "1986-03-06", "Union Oil Co. of California", "Leede Exploration", "All Section 32 (Top Morrow-19,480 ft)", "Deep depth assignment reserving 1/8 gas & 1/16 oil ORRI. Pearl wellbore depths specified.", "DIRECT IMAGE — PROVED"),
        (20, "—", "ASG", "Bk 1697/Pg 236", "2001-08-20", "2001-08-29", "Staghorn Resources LLC", "Chesapeake Exploration LP", "Section 32 scheduled properties", "Modern master assignment. Critical Exhibit A required to verify underlying lease schedule.", "DIRECT IMAGE — PROVED"),
        (21, "—", "ASG", "Bk 2340/Pg 490", "2020-12-14", "2020-12-22", "Chesapeake Affiliates", "KL CHK SPV LLC", "Section 32 index scope", "Master restructuring assignment. Schedules absent from recorded image.", "DIRECT IMAGE — PROVED"),
        (22, "—", "ASG", "Bk 2371/Pg 470", "2022-01-12", "2022-01-20", "KL CHK SPV LLC", "Diversified Production LLC (51.25%) & OCM Denali (48.75%)", "Section 32 properties", "Transaction assignment creating parallel Diversified / OCM Denali branches. Not section-wide WI.", "DIRECT IMAGE — PROVED"),
        (23, "—", "CONV", "Bk 2395/Pg 415", "2022-11-18", "2022-11-28", "DP Legacy Tapstone LLC", "Diversified ABS VI Upstream & DP Sooner", "Asset OK48147.001.1 at Pg 462", "Conveyance identifying Section 32 asset identity; estate quantum and depths blank.", "DIRECT IMAGE — PROVED"),
        (24, "—", "CONV", "Bk 2400/Pg 551", "2023-01-10", "2023-01-19", "DP Sooner HoldCo LLC", "Diversified Production LLC", "Asset OK48147.001.1 at Pg 566", "Last supported schedule claimant. Current title after 03/02/2023 requires Teocalli review.", "DIRECT IMAGE — PROVED"),
        (25, "—", "LEAD", "Bk 2434/Pg 751", "2023-08-15", "2023-08-25", "Diversified Production LLC / DP Legacy", "Teocalli Exploration LLC", "Unread schedule scope", "CRITICAL POST-CUTOFF LEAD: Potential divestiture of Diversified Section 32 working interest.", "INDEX ONLY — CURATIVE QUEUE")
    ]
    
    for idx, r in enumerate(runsheet_data, start=5):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 28
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 45
    ws.column_dimensions['K'].width = 28

    # -------------------------------------------------------------
    # 6-10. TRACTS 1 TO 5 SHEETS
    # -------------------------------------------------------------
    tract_specs = [
        ("Tract 1", "TRACT 1 — NORTHEAST QUARTER (NE/4) — 160.00 NOMINAL ACRES", [
            ("PAT", "Patent 3/469", "1907-03-22", "United States", "Cornelius B. Cornelius", "Sovereign Inception; Fee simple absolute without mineral reservation.", "PROVED (98%)"),
            ("WD", "Roger Mills 11/214", "1907-07-16", "Cornelius & Margreta Cornelius", "Jacob Bollenbach", "Warranty Deed in pre-Beckham county records. Conveys 160 ac fee.", "PROVED (98%)"),
            ("WD", "Bk 10/Pg 255", "1910-01-31", "Jacob Bollenbach", "Adolf Bollenbach", "Warranty Deed. Chain breaks downstream; 65.0 NMA unchained to present.", "PROVED (Chain Break)"),
            ("DECR", "Bk 83/Pg 259", "Date Unstated", "Estate of Ellen W. Crook", "Thirteen Heirs", "Collective vesting of 40.00 NMA; individual fractional allocations unstated.", "QUALIFIED (Decree Required)"),
            ("MD", "Bk 92/Pg 47", "1950-11-04", "Historic Grantor", "F.L. & J.L. Randel", "Conveys 3.00 NMA (1.50 NMA each) share and share alike.", "PROVED (95%)"),
            ("OGL", "Bk 264/Pg 345", "1972-02-05", "Charlie B. & Ima Pearl Crook", "Union Oil Co.", "Oil and Gas Lease; 160 ac; 1/8 royalty; 5-yr primary term.", "PROVED (Lease Face)"),
            ("MD", "Bk 367/Pg 632", "1978-02-14", "Historic Grantor", "Great Western Oil & Gas", "Mineral Deed conveying 10.00 NMA in NE/4.", "PROVED (95%)"),
            ("DECR", "Bk 845/Pg 150", "1985-10-23", "Crook Estates", "C.B. Jr., Hanks, Lenderman, Bryan, Higgenbotham", "Final decree distributing 8.00 NMA each (40.00 NMA total).", "PROVED (Quantum Established)")
        ]),
        ("Tract 2", "TRACT 2 — NORTH HALF OF NORTHWEST QUARTER (N/2 NW/4) — 80.00 NOMINAL ACRES", [
            ("FR", "FR Receipt 3894", "1905-05-04", "U.S. Land Office", "Herman Stevens", "Final Receiver Receipt; earliest located NW/4 sovereign evidence.", "PROVED (Patent Open)"),
            ("STIP", "Bk 153/Pg 161", "1958-04-18", "Maddoux / Farrar Parties", "B.E. Maddoux & D.H. Maddoux", "Stipulation of Interest settling 10.00 NMA each (20.00 NMA total).", "PROVED (95%)"),
            ("MD", "Bk 85/Pg 363", "1950-02-12", "Historic Grantor", "Sanditen Parties (4 Co-owners)", "Mineral Deed conveying 20.00 NMA undivided.", "PROVED (90%)"),
            ("DECR", "P-76-78, Pgs 349-350", "1976-08-10", "Estate of Ola Keathley", "Keathley / LeGrand Heirs", "Probate allocation incomplete; 20.00 NMA residual schedule carries forward.", "QUALIFIED (Decree Open)"),
            ("OGL", "Bk 340/Pg 323", "1977-09-29", "George T. & Renata Harrison", "Union Oil Co.", "Oil and Gas Lease; 80 ac; 1/8 royalty; covers 20.00 NMA interest.", "PROVED (Lease Face)")
        ]),
        ("Tract 3", "TRACT 3 — SOUTH HALF OF NORTHWEST QUARTER (S/2 NW/4) — 80.00 NOMINAL ACRES", [
            ("FR", "FR Receipt 3894", "1905-05-04", "U.S. Land Office", "Herman Stevens", "Final Receiver Receipt covering entire NW/4; patent open.", "PROVED (Patent Open)"),
            ("MD", "Bk 81/Pg 194", "1949-06-12", "Historic Grantors", "C. R. Bennett", "Mineral Deed conveying 40.00 NMA in S/2 NW/4.", "PROVED (95%)"),
            ("MD", "Bk 85/Pg 222", "1950-02-10", "Historic Grantors", "Heller Company", "Mineral Deed conveying 20.00 NMA in S/2 NW/4.", "PROVED (90%)"),
            ("DECR", "Bk 262/Pg 490", "1971-11-15", "Estate of Folmar", "Bettye Folmar Norris", "Decree of Distribution vesting 20.00 NMA in S/2 NW/4.", "PROVED (90%)"),
            ("OGL", "Bk 332/Pg 74", "1977-08-01", "Ralph W. Viersen Jr. et al.", "Leede Exploration", "OGL covering 320 ac across NW/4 & SW/4. Owner coverage defect noted.", "QUALIFIED (Coverage Gap)")
        ]),
        ("Tract 4", "TRACT 4 — SOUTHWEST QUARTER (SW/4) — 160.00 NOMINAL ACRES", [
            ("FR", "FR Receipt 3895", "1905-05-04", "U.S. Land Office", "Augustus P. Stevens", "Final Receiver Receipt covering entire SW/4; patent open.", "PROVED (Patent Open)"),
            ("MD", "Bk 85/Pg 222", "1950-02-10", "Historic Grantors", "Gertrude L. LaFortune", "Mineral Deed conveying 40.00 NMA in N/2 SW/4.", "PROVED (95%)"),
            ("DECR", "Probate Record", "1968-05-20", "Estate of Seidenbach", "Carol S. Leach & Cynthia S. Kruse", "Vesting 40.00 NMA (20.00 NMA each) in N/2 SW/4.", "PROVED (90%)"),
            ("OGL", "Bk 307/Pg 31", "1976-04-29", "Cora B. LeGrand", "Union Oil Co.", "OGL covering S/2 SW/4 (80 ac) and W/2 SE/4 (80 ac). 1/8 royalty.", "PROVED (Lease Face)"),
            ("OGL", "Bk 340/Pg 302", "1977-10-14", "Gertrude L. LaFortune", "Union Oil Co.", "OGL covering N/2 SW/4; 3/16 royalty; 3-year term.", "PROVED (Lease Face)"),
            ("OGL", "Bk 342/Pg 564", "1977-09-29", "Carol Seidenbach Leach", "Union Oil Co.", "OGL covering N/2 SW/4; 3/16 royalty; 3-year term.", "PROVED (Lease Face)")
        ]),
        ("Tract 5", "TRACT 5 — SOUTHEAST QUARTER (SE/4) — 160.00 NOMINAL ACRES", [
            ("FR", "Image 0004", "1906-06-28", "U.S. Land Office", "Velmore Biggs", "Final Receiver Receipt for SE/4. Biggs patent remains OPEN.", "PROVED (Patent Open)"),
            ("OGL", "Bk 63/Pg 33", "1947-11-13", "Cora B. Garrett & S.A. Garrett", "C.E. McKinley", "Oil and Gas Lease; E/2 SE/4 (80 ac); 10-year term; 1/8 royalty.", "PROVED (Base Lease)"),
            ("ASG", "Bk 244/Pg 432", "1969-11-25", "Humble Oil & Refining", "Union Oil Co.", "Assignment of Garrett 63/33 leasehold branch.", "PROVED (95%)"),
            ("OGL", "Bk 307/Pg 31", "1976-04-29", "Cora B. LeGrand", "Union Oil Co.", "Oil and Gas Lease; W/2 SE/4 (80 ac); 3-year term; 1/8 royalty.", "PROVED (Base Lease)"),
            ("WB ASG", "Bk 845/Pg 47", "1985-07-29", "Cities Service", "Leede Exploration", "Pearl 1-32 wellbore assignment (11,100-19,480 ft); 1/8 ORRI reserved.", "PROVED (Depth Limited)")
        ])
    ]
    
    for t_name, t_title, rows in tract_specs:
        ws = sheets[t_name]
        ws.views.sheetView[0].showGridLines = True
        
        ws['A1'] = t_title
        ws['A1'].font = font_title
        ws['A2'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE • Reconstructed Chain of Title & Fractional Derivation"
        ws['A2'].font = font_hold
        
        t_headers = ["Doc Type", "Book / Page", "Instrument Date", "Grantor", "Grantee / Party", "Estate Conveyed / Title Effect", "Evidence Status"]
        for c_idx, h in enumerate(t_headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header_navy
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        for r_idx, r_data in enumerate(rows, start=5):
            for c_idx, val in enumerate(r_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 28
        ws.column_dimensions['F'].width = 45
        ws.column_dimensions['G'].width = 25

    # -------------------------------------------------------------
    # 11. WI 2 SHEET (Orders before WI 1 in template contract)
    # -------------------------------------------------------------
    ws = sheets['WI 2']
    ws.views.sheetView[0].showGridLines = True
    
    ws['A1'] = "WI 2 — HISTORIC DEPTH, WELLBORE, & PARALLEL LEASEHOLD BURDENS"
    ws['A1'].font = font_title
    ws['A2'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE • Historic Burdens Carried Read-Only; Current Survival Unproved"
    ws['A2'].font = font_hold
    
    wi2_headers = [
        "Seq", "Instrument Type", "Book / Page", "Grantor", "Grantee / Burden Owner", 
        "Historic Quantum / Burden", "Depth / Formation Scope", "Current Treatment & Status", "Open Curative Requirement"
    ]
    
    for c_idx, h in enumerate(wi2_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    wi2_data = [
        (1, "Crook Correction ASG", "Bk 502/Pg 267", "Union Oil Co.", "Leede Exploration / Union ORRI", "7/64 gas & 1/16 oil ORRI", "Crook wellbore (16,210-20,451 ft)", "Historic Only; Survival Open", "Verify payout, reversion, and releases"),
        (2, "Crook Participant ASG", "Bk 509/Pg 220", "Leede Exploration", "Mesa 64%, Petrofina 25%, ENI 5%, etc.", "100% of assigned interest (Not 8/8)", "Crook wellbore", "Historic Allocation Only", "Prove assignor baseline and successors"),
        (3, "Pearl Borehole ASG", "Bk 845/Pg 47", "Cities Service", "Leede Exploration / 1/8 ORRI", "Assigned interest unquantified", "11,100-19,480 ft", "Historic Only; Status Open", "Read unrecorded agreements and releases"),
        (4, "Deep Scheduled ASG", "Bk 872/Pg 279", "Union Oil Co.", "Leede Exploration / Union ORRIs", "1/8 gas & 1/16 oil ORRIs", "Top Morrow to 19,480 ft; Pearl wellbore", "Historic Only; Bridge Open", "Acquire complete 1984 operating agreement"),
        (5, "Below-Hunton ASG", "Bk 1014/Pg 75", "Consolidation of Leede Trusts", "Leede Exploration", "Unquantified", "N/2 below top of Hunton formation", "Nemo-Dat Defect Disclosed", "Prove assignor source title in N/2"),
        (6, "Linn Wellbore Carve", "Bk 2183/174; 2185/639", "Chesapeake Exploration", "Linn Operating / Linn Energy Holdings", "Unquantified", "Carlson 7H wellbore carveout", "Separate Branch — Do Not Blend", "Read complete exhibits and legal schedules"),
        (7, "FourPoint / MNR ABS", "Bk 2185/639; 2415/328", "Chesapeake / Unbridled", "FourPoint / MNR ABS", "Unquantified", "Wellbore & depth exclusions open", "Parallel Portfolio Branch", "Reconcile Teocalli carve and schedules")
    ]
    
    for idx, r in enumerate(wi2_data, start=5):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 28
    ws.column_dimensions['I'].width = 35

    # -------------------------------------------------------------
    # 12. WI 1 SHEET
    # -------------------------------------------------------------
    ws = sheets['WI 1']
    ws.views.sheetView[0].showGridLines = True
    
    ws['A1'] = "WI 1 — PRINCIPAL KL CHK / TAPSTONE / DIVERSIFIED WORKING INTEREST CHAIN"
    ws['A1'].font = font_title
    ws['A2'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE • Scheduled Claimant Path Only; Exact Current Section 32 WI/NRI/ORRI NOT DETERMINED"
    ws['A2'].font = font_hold
    
    wi1_headers = [
        "Step", "Book / Page", "Grantor", "Grantee", "Transaction Branch / %", 
        "Depth / Legal Scope", "Adjudicated Conclusion", "Critical Evidence & Gaps"
    ]
    
    for c_idx, h in enumerate(wi1_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    wi1_data = [
        (1, "Bk 1697/Pg 236", "Staghorn Resources LLC", "Chesapeake Exploration LP", "Unquantified", "Section 32 per Exhibit A", "Scheduled Predecessor Lead", "Full Exhibit A required to identify base leases"),
        (2, "Bk 2340/Pg 403", "Chesapeake Affiliates", "Tapstone Energy LLC", "Tapstone Branch", "Schedules control", "Parallel Tapstone Branch", "Missing Tapstone asset schedule / bridge"),
        (3, "Bk 2340/Pg 490", "Chesapeake Affiliates", "KL CHK SPV LLC", "Unquantified", "All Section 32 index scope", "Principal Restructuring Predecessor", "Underlying property exhibits absent"),
        (4, "Bk 2371/Pg 470", "KL CHK SPV LLC", "Diversified Production (51.25%) & OCM Denali (48.75%)", "51.25% / 48.75% Branches", "Exhibits control", "Transaction Branch Split (Keep Separate)", "Do not report 51.25% as section-wide WI"),
        (5, "Bk 2389/Pg 500", "Diversified / DP Merger Parties", "DP Legacy Tapstone LLC", "Corporate Merger", "Corporate succession", "Merger Path Established", "Merger certificate & asset lists absent"),
        (6, "Bk 2395/Pg 415", "DP Legacy Tapstone LLC", "Diversified ABS VI Upstream & DP Sooner", "Unquantified", "Asset OK48147.001.1 (Pg 462)", "Asset Identity Supported", "Estate, quantum, and depth blank on schedule"),
        (7, "Bk 2400/Pg 551", "DP Sooner HoldCo LLC", "Diversified Production LLC", "Unquantified", "Asset OK48147.001.1 (Pg 566)", "Last Supported Schedule Claimant", "Not confirmed current; post-cutoff filings govern"),
        (8, "Bk 2434/Pg 751", "Diversified Production LLC / DP Legacy", "Teocalli Exploration LLC", "Unknown", "Post-cutoff index lead", "Potential Divestiture Out", "CRITICAL P0: Read complete instrument & schedule"),
        (9, "Bk 2451/4; 2476/121", "Diversified / DP Entities", "Public / Affiliates", "Unknown", "Post-cutoff index leads", "Later Contextual Filings", "Examine complete modern filings")
    ]
    
    for idx, r in enumerate(wi1_data, start=5):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
    # Summary Table on WI 1
    s_row = 16
    ws.cell(row=s_row, column=1, value="WORKING INTEREST COMPONENT").font = font_bold
    ws.cell(row=s_row, column=2, value="RECORDED / PROPOSED VALUE").font = font_bold
    ws.cell(row=s_row, column=3, value="LEGAL & TITLE CLASSIFICATION").font = font_bold
    ws.cell(row=s_row, column=4, value="SUPPORTING CITATION").font = font_bold
    ws.cell(row=s_row, column=5, value="EXAMINER LIMITATION & GOVERNING RULE").font = font_bold
    
    for c in range(1, 6):
        ws.cell(row=s_row, column=c).fill = fill_accent
        ws.cell(row=s_row, column=c).border = thin_border
        
    wi_summary = [
        ("Diversified Transaction Branch", "51.25%", "Historical Transaction Branch", "Bk 2371/Pg 470", "Branch allocation only; cannot be applied section-wide without proof of 8/8 baseline."),
        ("OCM Denali Parallel Branch", "48.75%", "Separate Claimant Branch", "Bk 2371/Pg 470", "Must remain strictly separate; not owned by Diversified Production LLC."),
        ("Predecessor Arithmetic Product", "15.000875%", "REPORTED_UNVERIFIED Lead", "V7 WI Lead (29.27% × 51.25%)", "Mathematical product only; lacks unclouded 8/8 title baseline. NOT PRESENT WI."),
        ("Current Gross Working Interest (WI)", "NOT DETERMINED", "UNPROVED / OPEN", "Bk 2400/566 & Bk 2434/751", "Blank quantum on schedule + unread Teocalli divestiture lead preclude decimal total."),
        ("Current Net Revenue Interest (NRI)", "NOT DETERMINED", "UNPROVED / OPEN", "Master Burden Register", "Burden survival (Union/Leede ORRIs, JOA burdens) remains unproved."),
        ("Current Overriding Royalty (ORRI)", "NOT DETERMINED", "UNPROVED / OPEN", "Bk 502/267; 872/279", "Historical ORRIs carried for audit; modern payout/reversion unproved.")
    ]
    
    for idx, r in enumerate(wi_summary, start=s_row+1):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if "NOT DETERMINED" in str(val):
                cell.font = font_bold
                cell.fill = fill_warning
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 35

    # -------------------------------------------------------------
    # 13. WELL 1 SHEET
    # -------------------------------------------------------------
    ws = sheets['Well 1']
    ws.views.sheetView[0].showGridLines = True
    
    ws['A1'] = "WELLS — SECTION 32-11N-25W (REGULATORY CONTEXT; NOT LEASE-SPECIFIC HBP PROOF)"
    ws['A1'].font = font_title
    ws['A2'] = "FOR REVIEW — HOLD NO EXTERNAL RELEASE • Regulatory Well Presence Does Not Prove Section-Wide or Depth-Wide HBP"
    ws['A2'].font = font_hold
    
    well_headers = [
        "Well #", "API #", "Current Listed Operator", "Well Name", "Formation", 
        "Surface Location", "Spud Date", "Perforated Interval", "Status / Regulatory Context"
    ]
    
    for c_idx, h in enumerate(well_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_navy
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    well_data = [
        (1, "35-009-00075", "BP America Production Co.", "Cora Garrett C-310", "Granite Wash", "Sec 32-11N-25W (E/2 SE/4)", "1950-08-12", "Historical Intervals", "Plugged / Historic Well"),
        (2, "35-009-20312", "Complete Oil & Gas Services LLC", "Crook 1-32", "Granite Wash Lower", "Sec 32-11N-25W (NE/4)", "1980-01-26", "13,348-13,595 ft", "Active / Producing Context"),
        (3, "35-009-20605", "SM Energy Company", "Pearl 1-32", "Morrow-Springer", "Sec 32-11N-25W (SE/4)", "1981-04-10", "15,200-15,450 ft", "Historic Deep Wellbore"),
        (4, "35-009-20738", "Leede Oil & Gas Inc.", "Legrand 1-32", "Upper Granite Wash", "Sec 32-11N-25W (NW/4)", "1982-03-15", "12,100-12,400 ft", "Historic Wellbore"),
        (5, "35-009-21072", "BCE-Mach II LLC", "JGL 1-32", "Granite Wash", "Sec 32-11N-25W (NW/4)", "1985-09-22", "12,450-12,700 ft", "Active Operator Context"),
        (6, "35-009-21085", "BCE-Mach II LLC", "Hanks Crook 1-32", "Granite Wash", "Sec 32-11N-25W (NE/4)", "1986-02-18", "12,600-12,850 ft", "Active Operator Context"),
        (7, "35-009-21166", "Vastar Resources Inc.", "Crook-Legrand 1-32", "Atoka", "Sec 32-11N-25W (SW/4)", "1988-06-30", "14,100-14,350 ft", "Historic Operator Context"),
        (8, "35-009-21236", "BCE-Mach II LLC", "Legrand 1-32 (Re-entry)", "Granite Wash", "Sec 32-11N-25W (NW/4)", "1991-11-12", "12,250-12,500 ft", "Active Operator Context"),
        (9, "35-009-21245", "BP America Production Co.", "Katie 1-32", "Granite Wash", "Sec 32-11N-25W (NE/4)", "1992-04-05", "12,300-12,550 ft", "Active Operator Context"),
        (10, "35-009-21258", "BCE-Mach II LLC", "Legrand 2-32", "Granite Wash", "Sec 32-11N-25W (NW/4)", "1993-08-19", "12,400-12,650 ft", "Active Operator Context"),
        (11, "35-009-21391", "BCE-Mach II LLC", "Twin 1-32", "Granite Wash Low", "Sec 32-11N-25W (NE/4)", "2004-07-15", "12,548-12,604 ft", "Active / Producing Context"),
        (12, "35-009-21893", "Latigo Petroleum, LLC", "Carlson 32-11-25 12H", "Granite Wash Upper (Missourian)", "Sec 32-11N-25W (All Sec PUN)", "2014-05-10", "Horizontal Lateral", "Horizontal Unit Context"),
        (13, "35-009-21923", "Latigo Petroleum, LLC", "Carlson 32-11-25 7H", "Missourian", "Sec 32-11N-25W (All Sec PUN)", "2014-09-18", "Horizontal Lateral", "Wellbore Carveout (Linn/FourPoint)"),
        (14, "35-009-21933", "Latigo Petroleum, LLC", "Carlson 32-11-25 10H", "Missourian", "Sec 32-11N-25W (All Sec PUN)", "2015-01-22", "Horizontal Lateral", "Horizontal Unit Context")
    ]
    
    for idx, r in enumerate(well_data, start=5):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if c_idx in [1, 2, 7]:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 26
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 28

    # Save workbook
    wb.save(output_path)
    print(f"Workbook successfully built at: {output_path}")

if __name__ == '__main__':
    build_section32_challenger_workbook('/workspace/SECTION32_GEMINI37_CHALLENGER_20260830.xlsx')
