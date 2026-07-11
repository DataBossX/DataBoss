"""State-driven Section 32 workbook QA and repair loop.

Each invocation performs bounded work in an immutable run directory. The source
workbook is hash-verified, copied to staging, checked deterministically, repaired
one defect at a time, and never promoted automatically.
"""

from __future__ import annotations

import argparse
import json
import shutil
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

from .project_manifest import (
    ControlFileError,
    ProjectManifest,
    WorkOrder,
    load_project_manifest,
    load_work_order,
    sha256_file,
)
from .workbook_qa import QAReport, inspect_workbook, load_workbook_profile


@dataclass(frozen=True)
class RunResult:
    run_id: str
    status: str
    run_directory: Path
    staged_workbook: Optional[Path]
    receipt_path: Path
    promotion_package: Optional[Path]
    attempts: int


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _write_json(path: Path, value: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(value, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def _is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def _assert_safe_paths(work_order: WorkOrder) -> None:
    if work_order.candidate_path.resolve() == work_order.staging_root.resolve():
        raise ControlFileError("Candidate path cannot be the staging root")
    for prohibited in work_order.prohibited_paths:
        if _is_relative_to(work_order.staging_root, prohibited):
            raise ControlFileError(
                f"Staging root {work_order.staging_root} is inside prohibited "
                f"path {prohibited}"
            )


def _report_has_regression(before: QAReport, after: QAReport) -> bool:
    prior = {check.check_id: check for check in before.checks}
    current = {check.check_id: check for check in after.checks}
    return any(
        result.passed and not current[check_id].passed
        for check_id, result in prior.items()
        if check_id in current
    )


def _blocking_count(report: QAReport) -> int:
    return sum(
        finding.severity == "blocking"
        for finding in report.findings
        if finding.check_id != "human_landman_release"
    )


def _is_improvement(before: QAReport, after: QAReport) -> bool:
    return (
        _blocking_count(after) < _blocking_count(before)
        or after.score.overall_score > before.score.overall_score
    )


def _next_formula_repair(report: QAReport, attempts: Dict[str, int], limit: int):
    for finding in report.findings:
        key = f"{finding.sheet}!{finding.cell}"
        if (
            finding.code == "broken_formula"
            and finding.repairable
            and attempts.get(key, 0) < limit
        ):
            return finding, key
    return None, ""


def _restore_formula_from_template(
    staged_path: Path,
    template_path: Path,
    sheet_name: str,
    cell_reference: str,
) -> Dict[str, str]:
    import openpyxl

    keep_vba = staged_path.suffix.lower() == ".xlsm"
    staged = openpyxl.load_workbook(staged_path, data_only=False, keep_vba=keep_vba)
    template = openpyxl.load_workbook(
        template_path,
        data_only=False,
        keep_vba=template_path.suffix.lower() == ".xlsm",
    )
    try:
        if sheet_name not in staged.sheetnames or sheet_name not in template.sheetnames:
            raise ControlFileError(
                f"Cannot restore formula: sheet {sheet_name!r} is not in both workbooks"
            )
        old_formula = staged[sheet_name][cell_reference].value
        new_formula = template[sheet_name][cell_reference].value
        if not isinstance(new_formula, str) or not new_formula.startswith("="):
            raise ControlFileError(
                f"Template {sheet_name}!{cell_reference} has no formula authority"
            )
        staged[sheet_name][cell_reference] = new_formula
        staged.save(staged_path)
        return {
            "repair_type": "restore_formula_from_template",
            "sheet": sheet_name,
            "cell": cell_reference,
            "before": str(old_formula),
            "after": new_formula,
        }
    finally:
        staged.close()
        template.close()


class ControlledWorkbookLoop:
    """Run one hash-bound, bounded QA work order."""

    def __init__(self, manifest: ProjectManifest, work_order: WorkOrder) -> None:
        self.manifest = manifest
        self.work_order = work_order

    @classmethod
    def from_files(
        cls, manifest_path: Path, work_order_path: Path
    ) -> "ControlledWorkbookLoop":
        manifest = load_project_manifest(manifest_path)
        return cls(manifest, load_work_order(work_order_path, manifest))

    def run(self) -> RunResult:
        run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        run_id = f"{run_id}-{uuid4().hex[:8]}"
        run_directory = self.work_order.staging_root / run_id
        receipt_path = run_directory / "run_receipt.json"
        run_directory.mkdir(parents=True, exist_ok=False)
        started_at = _utc_now()
        staged_path: Optional[Path] = None
        promotion_path: Optional[Path] = None
        iterations: List[Dict[str, Any]] = []
        status = "failed"

        try:
            _assert_safe_paths(self.work_order)
            candidate = self.work_order.candidate_path
            if not candidate.is_file():
                raise ControlFileError(f"Candidate workbook not found: {candidate}")
            actual_hash = sha256_file(candidate)
            if actual_hash != self.work_order.expected_sha256:
                raise ControlFileError(
                    "Candidate hash mismatch: expected "
                    f"{self.work_order.expected_sha256}, found {actual_hash}"
                )
            if self.work_order.template_path and not self.work_order.template_path.is_file():
                raise ControlFileError(
                    f"Template workbook not found: {self.work_order.template_path}"
                )

            input_backup = run_directory / f"input{candidate.suffix.lower()}"
            staged_path = run_directory / f"staged{candidate.suffix.lower()}"
            shutil.copy2(candidate, input_backup)
            shutil.copy2(candidate, staged_path)
            if sha256_file(input_backup) != actual_hash or sha256_file(staged_path) != actual_hash:
                raise ControlFileError("Staging copy hash verification failed")

            profile = load_workbook_profile(self.work_order.profile_path)
            report = inspect_workbook(
                staged_path,
                self.work_order.acceptance_tests,
                template_path=self.work_order.template_path,
                profile=profile,
            )
            _write_json(run_directory / "qa_before.json", report.to_dict())
            attempts_by_defect: Dict[str, int] = {}

            while not report.score.technical_pass:
                allowed = "restore_formula_from_template" in self.work_order.allowed_repairs
                finding, defect_key = _next_formula_repair(
                    report,
                    attempts_by_defect,
                    self.work_order.max_attempts_per_defect,
                )
                if not allowed or finding is None or self.work_order.template_path is None:
                    status = "blocked"
                    break

                attempts_by_defect[defect_key] = attempts_by_defect.get(defect_key, 0) + 1
                snapshot = run_directory / (
                    f"before_attempt_{len(iterations) + 1:03d}{staged_path.suffix}"
                )
                shutil.copy2(staged_path, snapshot)
                repair = _restore_formula_from_template(
                    staged_path,
                    self.work_order.template_path,
                    finding.sheet,
                    finding.cell,
                )
                after = inspect_workbook(
                    staged_path,
                    self.work_order.acceptance_tests,
                    template_path=self.work_order.template_path,
                    profile=profile,
                )
                regression = _report_has_regression(report, after)
                improved = _is_improvement(report, after)
                kept = improved and not regression
                iteration = {
                    "index": len(iterations) + 1,
                    "repair": repair,
                    "before_sha256": report.sha256,
                    "after_sha256": after.sha256,
                    "before_score": asdict(report.score),
                    "after_score": asdict(after.score),
                    "regression": regression,
                    "kept": kept,
                }
                iterations.append(iteration)
                _write_json(
                    run_directory / f"repair_receipt_{len(iterations):03d}.json",
                    iteration,
                )
                if not kept:
                    shutil.copy2(snapshot, staged_path)
                    status = "regression_stopped" if regression else "no_improvement"
                    break
                report = after

            if report.score.technical_pass:
                status = "technically_verified_pending_approval"
                promotion_path = run_directory / "promotion_package.json"
                _write_json(
                    promotion_path,
                    {
                        "schema_id": "dbx.promotion_package",
                        "schema_version": "1.0",
                        "project_id": self.manifest.project_id,
                        "work_order_id": self.work_order.work_order_id,
                        "run_id": run_id,
                        "candidate_sha256": report.sha256,
                        "technical_checks_passed": True,
                        "promotion_executed": False,
                        "required_human_gate": self.manifest.human_gate,
                        "approval_must_name_sha256": report.sha256,
                        "backup_sha256": actual_hash,
                        "status": "PENDING_HUMAN_APPROVAL",
                    },
                )

            _write_json(run_directory / "qa_after.json", report.to_dict())
            receipt = {
                "schema_id": "dbx.run_receipt",
                "schema_version": "1.0",
                "run_id": run_id,
                "project_id": self.manifest.project_id,
                "work_order_id": self.work_order.work_order_id,
                "objective": self.work_order.objective,
                "started_at": started_at,
                "completed_at": _utc_now(),
                "status": status,
                "input": {
                    "path": str(candidate),
                    "expected_sha256": self.work_order.expected_sha256,
                    "actual_sha256": actual_hash,
                    "backup": str(input_backup),
                },
                "output": {
                    "path": str(staged_path),
                    "sha256": report.sha256,
                },
                "score": asdict(report.score),
                "iterations": iterations,
                "human_approval_required": True,
                "promotion_executed": False,
            }
            _write_json(receipt_path, receipt)
        except Exception as exc:
            status = "failed"
            _write_json(
                receipt_path,
                {
                    "schema_id": "dbx.run_receipt",
                    "schema_version": "1.0",
                    "run_id": run_id,
                    "project_id": self.manifest.project_id,
                    "work_order_id": self.work_order.work_order_id,
                    "started_at": started_at,
                    "completed_at": _utc_now(),
                    "status": status,
                    "error_type": type(exc).__name__,
                    "error": str(exc),
                    "human_approval_required": True,
                    "promotion_executed": False,
                },
            )

        return RunResult(
            run_id=run_id,
            status=status,
            run_directory=run_directory,
            staged_workbook=staged_path,
            receipt_path=receipt_path,
            promotion_package=promotion_path,
            attempts=len(iterations),
        )


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Run a controlled workbook QA loop")
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--work-order", type=Path, required=True)
    args = parser.parse_args(argv)
    result = ControlledWorkbookLoop.from_files(args.manifest, args.work_order).run()
    print(result.receipt_path)
    return 0 if result.status == "technically_verified_pending_approval" else 1


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
