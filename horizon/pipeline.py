"""The Intelligence Layer, operationalized.

This is where the isolated engines are wired together into an actual title
build: read the OGL register and the runsheet from a workbook, cross-reference
them on ``Instrument_Number``, reconcile interest down each tract's chain with
exact math, and emit a :class:`ReportModel` whose rows carry the reconciled
retained interest / net acres -- or a ``Needs Examiner Review`` /
``ESCALATED`` tag when the files don't support a determination.

Nothing here fabricates a balance. A chain break, an over-conveyance, an
unknown starting interest, or a legal description that doesn't tie all produce
tags, not invented numbers.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from .chaining import (
    CrossRef,
    OGLRecord,
    RunsheetNote,
    build_cross_reference,
    find_chain_breaks,
    normalize_instrument,
    reconcile_chain,
)
from .interest import FULL, format_acres, format_fraction, net_acres, try_parse_acres
from .models import ESCALATED_TAG, REVIEW_TAG, ReportModel, TitleRow

# Sheet-name hints for auto-detecting the two registers inside one workbook.
# NB: no "runs"/"run" here -- those are substrings of "Runsheet" and would make
# a runsheet tab (which precedes the OGL tab) get parsed as the OGL register.
_OGL_SHEET_HINTS = ("ogl", "o&g", "lease", "conveyance", "instrument", "register")
_RUNSHEET_HINTS = ("runsheet", "run sheet", "notes")

# Header synonyms for the reference sheets (superset of report_io's map).
_OGL_HEADERS: Dict[str, str] = {
    "instrument": "instrument_number", "instrument_number": "instrument_number",
    "instrument_no": "instrument_number", "doc_no": "instrument_number",
    "document_number": "instrument_number", "reception": "instrument_number",
    "grantor": "grantor", "from": "grantor", "assignor": "grantor", "lessor": "grantor",
    "grantee": "grantee", "to": "grantee", "assignee": "grantee", "lessee": "grantee",
    "conveyed": "conveyed_interest", "conveyed_interest": "conveyed_interest",
    "interest": "conveyed_interest", "fraction": "conveyed_interest",
    "legal": "legal_description", "legal_description": "legal_description",
    "description": "legal_description", "tract": "legal_description",
    "type": "doc_type", "doc_type": "doc_type", "instrument_type": "doc_type",
    "date": "instrument_date", "instrument_date": "instrument_date", "dated": "instrument_date",
    "acres": "gross_acres", "gross_acres": "gross_acres", "gross": "gross_acres",
    "acreage": "gross_acres", "gross_acreage": "gross_acres",
}
_RUNSHEET_HEADERS: Dict[str, str] = {
    "instrument": "instrument_number", "instrument_number": "instrument_number",
    "instrument_no": "instrument_number", "doc_no": "instrument_number",
    "note": "note", "notes": "note", "comment": "note", "remarks": "note",
    "legal": "legal_description", "legal_description": "legal_description",
    "description": "legal_description",
}


def _canon(header, table) -> Optional[str]:
    key = str(header or "").strip().lower().replace(" ", "_")
    return table.get(key)


def _detect_header(ws, table) -> Tuple[int, Dict[int, str]]:
    """Find the header row and column->field map for a reference sheet."""
    best_row, best_map = -1, {}
    for r_idx, values in enumerate(ws.iter_rows(values_only=True)):
        mapped = {i: _canon(v, table) for i, v in enumerate(values)}
        mapped = {i: f for i, f in mapped.items() if f}
        if len(mapped) > len(best_map):
            best_row, best_map = r_idx, mapped
        if r_idx > 30:
            break
    return best_row, best_map


def read_ogl_records(path: Path, sheet: Optional[str] = None) -> List[OGLRecord]:
    """Read OGL / conveyance rows from a workbook sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = _pick_ogl_sheet(wb, sheet)
    header_row, header_map = _detect_header(ws, _OGL_HEADERS)
    records: List[OGLRecord] = []
    if header_row >= 0:
        for r_idx, values in enumerate(ws.iter_rows(values_only=True)):
            if r_idx <= header_row:
                continue
            data = {f: (str(values[i]).strip() if i < len(values) and values[i] is not None else "")
                    for i, f in header_map.items()}
            if data.get("instrument_number") or data.get("grantor") or data.get("grantee"):
                records.append(OGLRecord(
                    instrument_number=data.get("instrument_number", ""),
                    grantor=data.get("grantor", ""),
                    grantee=data.get("grantee", ""),
                    conveyed_interest=data.get("conveyed_interest", ""),
                    legal_description=data.get("legal_description", ""),
                    doc_type=data.get("doc_type", ""),
                    instrument_date=data.get("instrument_date", ""),
                    gross_acres=data.get("gross_acres", ""),
                ))
    wb.close()
    return records


def read_runsheet_notes(path: Path, sheet: Optional[str] = None) -> List[RunsheetNote]:
    """Read runsheet notes from a workbook sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = _pick_runsheet_sheet(wb, sheet)
    if ws is None:
        # No dedicated runsheet tab -> no notes (never read the OGL as a runsheet).
        wb.close()
        return []
    header_row, header_map = _detect_header(ws, _RUNSHEET_HEADERS)
    notes: List[RunsheetNote] = []
    if header_row >= 0:
        for r_idx, values in enumerate(ws.iter_rows(values_only=True)):
            if r_idx <= header_row:
                continue
            data = {f: (str(values[i]).strip() if i < len(values) and values[i] is not None else "")
                    for i, f in header_map.items()}
            if data.get("instrument_number"):
                notes.append(RunsheetNote(
                    instrument_number=data.get("instrument_number", ""),
                    note=data.get("note", ""),
                    legal_description=data.get("legal_description", ""),
                ))
    wb.close()
    return notes


def _matches(name: str, hints) -> bool:
    low = str(name).lower()
    return any(h in low for h in hints)


def _is_ogl_sheet(name: str) -> bool:
    """OGL-like *and not* runsheet-like -- so an ambiguous tab (e.g. "OGL Notes")
    is claimed by neither register rather than by both."""
    return _matches(name, _OGL_SHEET_HINTS) and not _matches(name, _RUNSHEET_HINTS)


def _is_runsheet_sheet(name: str) -> bool:
    return _matches(name, _RUNSHEET_HINTS) and not _matches(name, _OGL_SHEET_HINTS)


def _pick_ogl_sheet(wb, name: Optional[str]):
    """The OGL register: explicit name > unambiguous OGL tab > first sheet."""
    if name and name in wb.sheetnames:
        return wb[name]
    for sn in wb.sheetnames:
        if _is_ogl_sheet(sn):
            return wb[sn]
    # The first sheet is the conventional home of the register; safe default.
    return wb.worksheets[0]


def _pick_runsheet_sheet(wb, name: Optional[str]):
    """The runsheet: explicit name > unambiguous runsheet tab > None.

    Unlike the OGL, there is NO fall-back to sheet 0 -- doing so would read the
    OGL register as synthetic runsheet notes and hide real chain breaks.
    """
    if name and name in wb.sheetnames:
        return wb[name]
    for sn in wb.sheetnames:
        if _is_runsheet_sheet(sn):
            return wb[sn]
    return None


def score_reference_workbook(path: Path) -> int:
    """Score how well a workbook works as an OGL+runsheet reference source.

    Higher is better. A workbook that has *both* an unambiguous OGL sheet and a
    *distinct* unambiguous runsheet sheet scores highest; the canonical NHE
    report name gets a bonus. Returns 0 (unusable) if it can't be opened or has
    no OGL register.
    """
    try:
        import openpyxl
    except ImportError:
        return 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return 0
    names = list(wb.sheetnames)
    wb.close()

    ogl_sheets = [n for n in names if _is_ogl_sheet(n)]
    run_sheets = [n for n in names if _is_runsheet_sheet(n)]
    has_ogl = bool(ogl_sheets)
    # A usable reference workbook MUST have an OGL register. Without one, no
    # filename bonus can rescue it -- otherwise a file merely *named* like the
    # report but containing no register would be auto-selected and its first
    # sheet processed as OGL data. Return 0 so it is never chosen.
    if not has_ogl:
        return 0
    # Require the runsheet to be a *different* sheet than the chosen OGL one.
    has_run = any(n not in ogl_sheets for n in run_sheets)
    score = 3  # has_ogl
    if has_run:
        score += 3
        score += 4  # a single workbook carrying both registers is ideal
    low_name = path.name.lower()
    if "cursory" in low_name or "nhe" in low_name or "title report" in low_name:
        score += 2
    if "roger" in low_name and "mills" in low_name:
        score += 1
    return score


def find_reference_workbook(paths) -> Optional[Path]:
    """Pick the best OGL+runsheet reference workbook from candidate paths."""
    best: Optional[Path] = None
    best_score = 0
    for p in paths:
        p = Path(p)
        if p.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        s = score_reference_workbook(p)
        if s > best_score:
            best, best_score = p, s
    return best


@dataclass
class ChainedBuild:
    report: ReportModel
    cross_reference: Dict[str, CrossRef] = field(default_factory=dict)
    chain_breaks: List[CrossRef] = field(default_factory=list)
    tracts_reviewed: List[str] = field(default_factory=list)


def chain_to_report(
    ogl_records: List[OGLRecord],
    runsheet_notes: List[RunsheetNote],
    section: str = "31-12N-24W",
    tract_legals: Optional[Dict[str, str]] = None,
    starting_interests: Optional[Dict[str, object]] = None,
) -> ChainedBuild:
    """Build a reconciled report from the OGL register + runsheet.

    Records are grouped into chains by their (normalized) legal description --
    that is the "tract". Within a tract they are walked in file order and the
    interest is reconciled exactly. Rows that cannot be tied out carry a
    ``Needs Examiner Review`` status; instruments that break the OGL<->runsheet
    link are marked in remarks. Nothing is balanced by fabrication.
    """
    tract_legals = tract_legals or {}
    starting_interests = starting_interests or {}

    xref = build_cross_reference(ogl_records, runsheet_notes)
    breaks = find_chain_breaks(xref)
    break_keys = {c.key for c in breaks}

    # group OGL records into tract chains keyed by normalized legal description
    tracts: Dict[str, List[OGLRecord]] = {}
    for rec in ogl_records:
        tract_key = _tract_key(rec.legal_description) or "UNSPECIFIED"
        tracts.setdefault(tract_key, []).append(rec)

    # An instrument that appears under more than one tract is reconciled from a
    # full starting interest in each -- correct only if those really are separate
    # estates. Because we cannot prove that from the files, flag such instruments
    # for examiner review rather than silently trusting the per-tract math.
    tracts_by_instrument: Dict[str, set] = {}
    for tract_key, records in tracts.items():
        for rec in records:
            k = normalize_instrument(rec.instrument_number)
            if k:
                tracts_by_instrument.setdefault(k, set()).add(tract_key)
    cross_tract_keys = {k for k, ts in tracts_by_instrument.items() if len(ts) > 1}

    rows: List[TitleRow] = []
    reviewed: List[str] = []

    for tract_key, records in tracts.items():
        tract_legal = tract_legals.get(tract_key, records[0].legal_description if records else "")
        start = starting_interests.get(tract_key, FULL)

        # Only the first row per instrument number participates in the chain
        # math; further rows sharing an instrument number are duplicates and are
        # emitted as flagged rows WITHOUT advancing the interest ledger (so the
        # same instrument is never reconciled twice). Rows without an instrument
        # number are all treated as primaries (nothing to dedupe on).
        primaries: List[OGLRecord] = []
        duplicate_recs: List[OGLRecord] = []
        seen_keys = set()
        for rec in records:
            k = normalize_instrument(rec.instrument_number)
            if k and k in seen_keys:
                duplicate_recs.append(rec)
            else:
                if k:
                    seen_keys.add(k)
                primaries.append(rec)

        result = reconcile_chain(tract_key, primaries, starting_interest=start,
                                 tract_legal=tract_legal)
        has_cross_tract = any(normalize_instrument(r.instrument_number) in cross_tract_keys
                              for r in records)
        if result.needs_examiner_review or duplicate_recs or has_cross_tract:
            reviewed.append(tract_key)

        for link, rec in zip(result.links, primaries):
            remarks_bits: List[str] = []
            status = "ok"

            key = normalize_instrument(rec.instrument_number)
            if key and key in break_keys:
                xstatus = xref.get(key, CrossRef(key=key)).status
                detail = {
                    "orphan_note": "runsheet note references an instrument with no OGL record",
                    "unreferenced_ogl": "OGL instrument never referenced in the runsheet",
                    "duplicate_ogl": "instrument number appears on multiple OGL rows",
                }.get(xstatus, "instrument not matched across OGL/runsheet")
                remarks_bits.append(f"Chain break: {detail}")
                status = REVIEW_TAG
            if key and key in cross_tract_keys:
                remarks_bits.append(
                    "Instrument spans multiple tracts; per-tract interest not "
                    "assumed to be independent")
                status = REVIEW_TAG
            if not str(rec.legal_description).strip():
                # No legal to tie the conveyance to a tract; the legal-tie check
                # can't verify it, so flag rather than silently accept.
                remarks_bits.append("No legal description to tie to a tract")
                status = REVIEW_TAG
            elif not link.tied_to_legal:
                remarks_bits.append("Legal description does not tie to tract")
                status = REVIEW_TAG
            if not link.grantor_continuity:
                remarks_bits.append(
                    f"Chain-of-title break: grantor does not match prior grantee")
                status = REVIEW_TAG
            rc = link.reconciliation
            if rc.status != "balanced":
                remarks_bits.append(rc.note)
                status = REVIEW_TAG
            if not key and (rec.grantor or rec.grantee):
                remarks_bits.append("Missing instrument number")
                status = REVIEW_TAG

            # Only emit a retained interest when it is a real, non-negative value.
            # An over-conveyance yields a negative `retained`; writing it would
            # feed a nonsensical figure back into downstream validation (the row
            # is already tagged for examiner review via its status).
            retained_txt = (format_fraction(rc.retained)
                            if rc.retained is not None and rc.retained >= 0 else "")

            # Net mineral acres = conveyed interest x gross acres, exact. Blank
            # (never guessed) when gross acres or the conveyed interest is unknown.
            # Gross acres is a decimal/whole number -- a fraction form is rejected.
            nma_txt = ""
            gross = try_parse_acres(rec.gross_acres) \
                if str(rec.gross_acres).strip() else None
            if gross is not None and link.conveyed is not None:
                nma_txt = format_acres(net_acres(link.conveyed, gross))
            elif str(rec.gross_acres).strip() and link.conveyed is not None:
                remarks_bits.append("Gross acres unparseable; net acres not computed")
                status = REVIEW_TAG

            # attach runsheet notes into remarks (verification trail)
            for note in xref.get(key, CrossRef(key=key)).notes if key else []:
                if note.note:
                    remarks_bits.append(f"Runsheet: {note.note}")

            rows.append(TitleRow(
                doc_type=rec.doc_type or None,
                instrument_date=rec.instrument_date or None,
                grantor=rec.grantor,
                grantee=rec.grantee,
                instrument_number=rec.instrument_number,
                legal_description=rec.legal_description,
                conveyed_interest=(format_fraction(link.conveyed)
                                   if link.conveyed is not None else ""),
                retained_interest=retained_txt,
                net_mineral_acres=nma_txt,
                status=status,
                remarks="; ".join(b for b in remarks_bits if b),
            ))

        # Emit duplicate-instrument rows as review items only -- no interest math,
        # so a single instrument's interest is never double-counted in the chain.
        for rec in duplicate_recs:
            rows.append(TitleRow(
                doc_type=rec.doc_type or None,
                instrument_date=rec.instrument_date or None,
                grantor=rec.grantor,
                grantee=rec.grantee,
                instrument_number=rec.instrument_number,
                legal_description=rec.legal_description,
                conveyed_interest="",
                retained_interest="",
                net_mineral_acres="",
                status=REVIEW_TAG,
                remarks="Chain break: duplicate instrument number "
                        "(not reconciled to avoid double-counting)",
            ))

    report = ReportModel(section=section, source="chain_to_report", rows=rows)
    return ChainedBuild(report=report, cross_reference=xref,
                        chain_breaks=breaks, tracts_reviewed=reviewed)


def _tract_key(legal: str) -> str:
    import re
    return re.sub(r"[^A-Za-z0-9]+", "", str(legal or "")).upper()


def build_from_workbook(path: Path, section: str = "31-12N-24W",
                        ogl_sheet: Optional[str] = None,
                        runsheet_sheet: Optional[str] = None) -> ChainedBuild:
    """Convenience: read both registers from one workbook and chain them."""
    ogl = read_ogl_records(path, ogl_sheet)
    notes = read_runsheet_notes(path, runsheet_sheet)
    return chain_to_report(ogl, notes, section=section)
