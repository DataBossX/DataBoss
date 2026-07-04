"""Read/write cursory title reports as .xlsx, aligned to the canonical columns.

Writing goes through openpyxl; the output always uses :data:`CANONICAL_COLUMNS`
so every iteration is format-aligned with the Roger Mills template logic. Reading
is tolerant of messy real-world headers (it maps synonyms onto canonical fields)
so an existing report can be ingested and re-perfected.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from .models import CANONICAL_COLUMNS, ReportModel, TitleRow

_HEADER_SYNONYMS: Dict[str, str] = {
    "entry": "entry_no", "entry_no": "entry_no", "no": "entry_no", "item": "entry_no",
    "instrument_date": "instrument_date", "date": "instrument_date", "dated": "instrument_date",
    "recorded_date": "recorded_date", "recorded": "recorded_date", "filed": "recorded_date",
    "doc_type": "doc_type", "type": "doc_type", "instrument_type": "doc_type",
    "grantor": "grantor", "from": "grantor",
    "grantee": "grantee", "to": "grantee",
    "instrument_number": "instrument_number", "instrument": "instrument_number",
    "instrument_no": "instrument_number", "doc_no": "instrument_number",
    "document_number": "instrument_number", "reception": "instrument_number",
    "book": "book", "vol": "book", "volume": "book",
    "page": "page", "pg": "page",
    "legal_description": "legal_description", "legal": "legal_description",
    "description": "legal_description", "tract": "legal_description",
    "gross_acres": "gross_acres", "acres": "gross_acres", "acreage": "gross_acres",
    "conveyed_interest": "conveyed_interest", "conveyed": "conveyed_interest",
    "retained_interest": "retained_interest", "retained": "retained_interest",
    "net_mineral_acres": "net_mineral_acres", "nma": "net_mineral_acres",
    "net_acres": "net_mineral_acres",
    "status": "status",
    "remarks": "remarks", "notes": "remarks", "comments": "remarks",
}


def _canon(header: str) -> Optional[str]:
    key = str(header or "").strip().lower().replace(" ", "_")
    return _HEADER_SYNONYMS.get(key)


DATA_SHEET_TITLE = "Title Report"


def _has_media(path: Path) -> bool:
    import zipfile
    try:
        with zipfile.ZipFile(path) as zf:
            return any(n.startswith("xl/media/") for n in zf.namelist())
    except (zipfile.BadZipFile, OSError):
        return False


def write_report(report: ReportModel, path: Path, template: Optional[Path] = None) -> Path:
    """Write ``report`` to ``path`` using the canonical column layout.

    If ``template`` is an existing workbook that carries embedded media (plats),
    the output is written *into a copy of that template* so ``xl/media`` parts
    and other sheets survive (openpyxl preserves images across load/save); only
    the data sheet is rewritten. Otherwise a fresh workbook is created.
    """
    import openpyxl

    path.parent.mkdir(parents=True, exist_ok=True)

    if template is not None and Path(template).exists() and _has_media(Path(template)):
        import shutil
        shutil.copy2(template, path)
        wb = openpyxl.load_workbook(path)
        # find the data sheet (preferred title, else the active sheet) and
        # rewrite only it -- every other sheet (e.g. embedded plats) is preserved.
        ws = wb[DATA_SHEET_TITLE] if DATA_SHEET_TITLE in wb.sheetnames else wb.active
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
        ws.title = DATA_SHEET_TITLE
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = DATA_SHEET_TITLE

    headers = [c.replace("_", " ").title() for c in CANONICAL_COLUMNS]
    ws.append([f"{report.section} - Roger Mills County - Cursory Title Report"])
    ws.append(headers)
    for row in report.rows:
        ws.append([getattr(row, c, "") or "" for c in CANONICAL_COLUMNS])
    ws.freeze_panes = "A3"
    wb.save(path)
    wb.close()
    return path


def read_report(path: Path, section: str = "") -> ReportModel:
    """Read an existing .xlsx into a :class:`ReportModel`, mapping headers."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = None
    header_map: Dict[int, str] = {}
    rows: List[TitleRow] = []
    for r_idx, values in enumerate(ws.iter_rows(values_only=True)):
        if header_row is None:
            mapped = {i: _canon(v) for i, v in enumerate(values) if _canon(v)}
            if len(mapped) >= 3:
                header_row = r_idx
                header_map = {i: f for i, f in mapped.items() if f}
            continue
        data = {}
        for i, field_name in header_map.items():
            if i < len(values) and values[i] is not None:
                data[field_name] = str(values[i]).strip()
        if any(data.values()):
            rows.append(TitleRow(**data))
    wb.close()
    return ReportModel(section=section or "unknown", source=str(path), rows=rows)
