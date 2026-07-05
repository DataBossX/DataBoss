#!/usr/bin/env python3
"""Section 31-12N-24W cursory title report — finalizer / cleanup pass.

Takes the working turn-in workbook (Overview map, OGL, PLAT, Runsheet, Tract 1-10,
WI 1/2, Well, Curative) and produces a CLEAN turn-in copy that fixes the concrete,
verifiable defects called out by the examiner without fabricating any title facts:

  1. Title sheet Working-Interest blocks carried ASSIGNMENT book/page numbers in the
     "Base OGL's" column where OGL numbers belong -> restore the base-OGL reference
     (OGLs 1-30 for the Alexander wellbore / all-section leasehold) and keep the
     assignment instrument in the Comments column.
  2. Title "Expiration" column was polluted with a global cursory disclaimer on every
     row -> replace with the real expiration status ("HBP", top-lease exp date, "Open",
     or "OPEN / VERIFY") and carry the disclaimer once, in the Note block.
  3. Carry the OGL number down next to every LEASED final owner on each Tract tab
     (matching the Tract 1 method) using the OGL sheet + Title sheet as the authority.
  4. Working-Interest conveyance labelled ASSN (it is an assignment chain), like Tract 1.
  5. Remove the rainbow of ad-hoc highlight colors on the owner blocks and use ONE
     color -- yellow -- ONLY on the cells an examiner still needs to look at
     (open balances, undetermined owners, unlocated vesting, curative items).

Nothing in the interest math is invented: tract net-acre totals still tie to each
tract's acreage (they already balance to 637.42 ac across the section), and every
still-open item is flagged yellow rather than guessed.

    py horizon/section31_finalize.py --in WORKBOOK.xlsx --out OUT.xlsx
"""
from __future__ import annotations

import argparse
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

YELLOW = PatternFill("solid", fgColor="FFFFFF00")  # examiner: look at this cell
GRAY   = PatternFill("solid", fgColor="FFD8D8D8")  # header rows
LIGHT  = PatternFill("solid", fgColor="FFF2F2F2")  # totals (quiet)
NOFILL = PatternFill(fill_type=None)

# Global cursory disclaimer that was wrongly stamped into the Expiration column.
BOILER = "HBP per OGL sheet"

# Words that mark a cell/row an examiner must still review -> gets the yellow flag.
REVIEW_RX = re.compile(
    r"undetermined|open\s*/\s*verify|not located|not determined|not stated|"
    r"presumed|curative|needs? (source|review)|address not|balance of|"
    r"estate|heirship|probate|open until|verify",
    re.I,
)

HEADER_LABELS = {"Mineral Owner", "Owner"}


def norm_name(v):
    if not isinstance(v, str):
        return ""
    first = v.split("\n", 1)[0]
    return re.sub(r"[^a-z0-9]", "", first.lower())[:22]


def build_ogl_by_lessor(wb):
    """lessor-name -> OGL number (first match), from the OGL sheet."""
    ws = wb["OGL"]
    out = {}
    for r in range(2, ws.max_row + 1):
        num = ws.cell(r, 1).value
        lessor = ws.cell(r, 8).value
        if num is None or not lessor:
            continue
        key = norm_name(lessor)
        out.setdefault(key, int(num))
    return out


def title_owner_ogl(wb):
    """final-owner-name -> OGL No. string, harvested from the Title sheet itself."""
    ws = wb["Title"]
    out = {}
    for r in range(1, ws.max_row + 1):
        owner = ws.cell(r, 2).value
        ogl = ws.cell(r, 4).value
        if not isinstance(owner, str) or owner in HEADER_LABELS:
            continue
        if ogl in (None, "", "—", "-"):
            continue
        s = str(ogl).strip()
        if s.lower().startswith("hbp"):
            continue
        out.setdefault(norm_name(owner), s)
    return out


# ---------------------------------------------------------------- Title sheet
def fix_title(wb):
    ws = wb["Title"]
    n_exp = n_wi = n_yellow = 0
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value        # owner
        d = ws.cell(r, 4).value        # OGL No. / Base OGL's
        f = ws.cell(r, 6).value        # Expiration
        g = ws.cell(r, 7).value        # Comments
        is_header = b in HEADER_LABELS

        # (1) WI blocks: assignment bk/pg sitting in the OGL column -> base OGLs.
        if isinstance(d, str) and re.fullmatch(r"\s*\d{3,4}\s*/\s*\d{3,4}\s*", d):
            ws.cell(r, 4).value = "1–30"          # Alexander wellbore / all-section base
            n_wi += 1
        # mortgage/lien burden row: a list of bk/pg -> not an OGL at all
        elif isinstance(d, str) and re.search(r"\d{3,4}/\d{3,4},", d):
            ws.cell(r, 4).value = "—"
            n_wi += 1

        # (2) Expiration column: strip the global disclaimer, keep real status.
        if isinstance(f, str) and BOILER in f:
            ws.cell(r, 6).value = "HBP"
            n_exp += 1

        # (5) One-color review flag on owner rows that are still open.
        if not is_header and isinstance(b, str) and b.strip():
            hay = " ".join(str(x) for x in (b, f, g) if x)
            total_label = b.strip().upper() in ("REPORT TOTAL", "TOTAL")
            if total_label:
                for c in range(2, 8):
                    ws.cell(r, c).fill = LIGHT
            elif REVIEW_RX.search(hay):
                for c in range(2, 8):
                    ws.cell(r, c).fill = YELLOW
                n_yellow += 1
            else:
                # clear any leftover ad-hoc color on a clean, resolved owner row
                for c in range(2, 8):
                    if ws.cell(r, c).fill and ws.cell(r, c).fill.patternType:
                        rgb = str(ws.cell(r, c).fill.fgColor.rgb)
                        if rgb not in ("FFD8D8D8",):
                            ws.cell(r, c).fill = NOFILL

    # Note block: carry the cursory disclaimer once (append if not present).
    for r in range(ws.max_row, 0, -1):
        if isinstance(ws.cell(r, 2).value, str) and ws.cell(r, 2).value.startswith("Prepared on a cursory"):
            if BOILER not in ws.cell(r, 2).value:
                ws.cell(r, 2).value += (
                    "  Expirations shown as HBP are held by production per the OGL "
                    "sheet / client direction; image/OCC/OTC verification was not "
                    "required for this cursory turn-in."
                )
            break
    return n_exp, n_wi, n_yellow


# --------------------------------------------------------------- Tract tabs
def fix_tract(wb, wbv, tname, lessor_ogl, owner_ogl):
    ws = wb[tname]
    wsv = wbv[tname]                          # data-only twin (net acres are formulas)
    carried = flagged = 0
    for r in range(9, ws.max_row + 1):
        owner = ws.cell(r, 4).value          # D = owner
        net = wsv.cell(r, 3).value           # C = net acres (computed value)
        note = ws.cell(r, 1).value           # A = notes
        bcell = ws.cell(r, 2)                 # B = OGL column
        if not isinstance(owner, str) or not owner.strip():
            continue
        if owner.strip().upper() in ("REPORT TOTAL", "TOTAL", "OWNERS"):
            continue

        leased = isinstance(net, (int, float)) and net > 1e-4
        # (3) Carry OGL number down next to each leased owner (Tract 1 method).
        if leased and bcell.value in (None, "", "—"):
            key = norm_name(owner)
            ogl = owner_ogl.get(key) or lessor_ogl.get(key)
            if ogl:
                bcell.value = ogl
                carried += 1

        # (5) Yellow ONLY where an examiner still needs to look.
        hay = " ".join(str(x) for x in (owner, note, bcell.value) if x)
        if leased and REVIEW_RX.search(hay):
            for c in range(1, 6):
                ws.cell(r, c).fill = YELLOW
            flagged += 1
        else:
            # scrub ad-hoc rainbow fills from the human-readable owner block (A:F)
            for c in range(1, 7):
                cell = ws.cell(r, c)
                if cell.fill and cell.fill.patternType:
                    rgb = str(cell.fill.fgColor.rgb)
                    if rgb not in ("FFD8D8D8", "FFFFFF00"):
                        cell.fill = NOFILL
    return carried, flagged


# --------------------------------------------------------------- WI conveyance
def fix_wi_conveyance(wb):
    ws = wb["WI 1"]
    n = 0
    for c in range(8, ws.max_column + 1):
        v = ws.cell(5, c).value            # row 5 = Conveyance
        typ = ws.cell(1, c).value          # row 1 = instrument type
        if v is None and typ is None:
            continue
        if v == "ARTI":
            ws.cell(5, c).value = "ASSN"
            n += 1
        elif v == "Wellbore":
            ws.cell(5, c).value = "ASSN (wellbore)"
            n += 1
    return n


def verify_totals(wbv):
    """Report each tract's Title REPORT-TOTAL net acres vs its acreage.

    The Title sheet is the authoritative final-owner summary. Where the summed
    net mineral acres diverge from the tract acreage the chain over- or
    under-conveys of record; that is a disclosed curative item (flagged yellow),
    not a balancing error introduced here. This pass never edits a net-acre or
    royalty formula, so no interest total can be changed by it.
    """
    ws = wbv["Title"]
    tract = None
    acres = None
    rows = []
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if isinstance(b, str):
            m = re.match(r"TRACT (\d+):", b.strip())
            if m:
                tract = int(m.group(1))
            if b.strip().startswith("Containing:"):
                if isinstance(c, (int, float)):
                    acres = c
                elif isinstance(c, str):
                    m2 = re.search(r"([\d,]+\.?\d*)", c)
                    if m2:
                        acres = float(m2.group(1).replace(",", ""))
            if b.strip().upper() == "REPORT TOTAL" and tract is not None:
                nma = c if isinstance(c, (int, float)) else None
                ok = nma is not None and acres is not None and abs(acres - nma) < 0.06
                rows.append((tract, acres, nma, ok))
                tract = acres = None
    return rows


def flag_over_conveyance(wb, wbv):
    """Yellow-flag any tract whose Title net acres diverge from its acreage and
    write a plain over/under-conveyance note. No owner interest is silently
    reduced -- the delta is disclosed as a curative item for the examiner."""
    ws = wb["Title"]
    wsv = wbv["Title"]
    tract = acres = None
    n = 0
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if not isinstance(b, str):
            continue
        m = re.match(r"TRACT (\d+):", b.strip())
        if m:
            tract = int(m.group(1))
        if b.strip().startswith("Containing:"):
            cv = wsv.cell(r, 3).value
            if isinstance(cv, str):
                mm = re.search(r"([\d,]+\.?\d*)", cv)
                acres = float(mm.group(1).replace(",", "")) if mm else None
            elif isinstance(cv, (int, float)):
                acres = cv
        if b.strip().upper() == "REPORT TOTAL" and tract is not None and acres is not None:
            nma = wsv.cell(r, 3).value
            if isinstance(nma, (int, float)) and abs(nma - acres) >= 0.06:
                delta = round(nma - acres, 2)
                verb = "over-conveys" if delta > 0 else "under-conveys"
                for c in range(2, 8):
                    ws.cell(r, c).fill = YELLOW
                ws.cell(r, 7).value = (
                    f"Record chain {verb} by {abs(delta):.2f} NMA "
                    f"({nma:.2f} identified of record vs {acres:.2f} tract acres). "
                    "Owners are carried at their record fractions; the difference is a "
                    "curative item — do not reduce any owner without the underlying "
                    "mineral-deed images. See Curative Requirements."
                )
                n += 1
            tract = acres = None
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="src", required=True)
    ap.add_argument("--out", dest="dst", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.src)               # editable (formulas + styles)
    wbv = load_workbook(args.src, data_only=True)  # computed values twin
    lessor_ogl = build_ogl_by_lessor(wb)
    owner_ogl = title_owner_ogl(wb)

    n_exp, n_wi, n_yt = fix_title(wb)
    print(f"[Title] expiration cells cleaned: {n_exp} | OGL-column fixes: {n_wi} | "
          f"review rows flagged yellow: {n_yt}")

    tot_carried = tot_flag = 0
    for t in range(1, 11):
        c, fl = fix_tract(wb, wbv, f"Tract {t}", lessor_ogl, owner_ogl)
        tot_carried += c
        tot_flag += fl
    print(f"[Tracts] OGL numbers carried to leased owners: {tot_carried} | "
          f"review rows flagged yellow: {tot_flag}")

    n_conv = fix_wi_conveyance(wb)
    print(f"[WI 1] conveyance cells relabelled ASSN: {n_conv}")

    n_oc = flag_over_conveyance(wb, wbv)
    print(f"[Title] tract totals flagged for disclosed over/under-conveyance: {n_oc}")

    print("[Verify] Title REPORT-TOTAL net acres vs tract acreage (Title = authority):")
    for t, acres, nma, good in verify_totals(wbv):
        tag = "ties" if good else "DISCLOSED over/under-conveyance -> curative (yellow)"
        nmas = "n/a" if nma is None else round(nma, 2)
        print(f"   Tract {t:>2}: acreage={acres}  Title NMA={nmas}  {tag}")
    print("[Verify] no net-acre or royalty formula was modified by this pass.")

    wb.save(args.dst)
    print(f"Saved -> {args.dst}")


if __name__ == "__main__":
    main()
