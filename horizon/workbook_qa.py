"""Deterministic workbook inspection, validation, and rule-derived scoring."""

from __future__ import annotations

import json
import math
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import unquote, urlparse

from .project_manifest import ControlFileError, sha256_file


EXCEL_ERROR_VALUES = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
}
MATH_CHECKS = {
    "negative_current_ownership",
    "ownership_totals",
    "tract_acreage_totals",
    "lease_totals",
    "working_interest_totals",
    "no_duplicate_owners",
}
EVIDENCE_CHECKS = {
    "source_manifest_complete",
    "evidence_crosswalk_complete",
    "evidence_links_resolve",
}
TEMPLATE_CHECKS = {"no_broken_formulas", "template_compliance", "print_rendering"}
TOTAL_CHECKS = {
    "ownership_totals",
    "tract_acreage_totals",
    "lease_totals",
    "working_interest_totals",
}


@dataclass(frozen=True)
class QAFinding:
    check_id: str
    severity: str
    code: str
    message: str
    sheet: str = ""
    cell: str = ""
    repairable: bool = False


@dataclass
class CheckResult:
    check_id: str
    status: str
    score: float
    findings: List[QAFinding] = field(default_factory=list)
    metrics: Dict[str, Any] = field(default_factory=dict)

    @property
    def passed(self) -> bool:
        return self.status == "passed"


@dataclass(frozen=True)
class QAScore:
    completeness: float
    mathematical_accuracy: float
    evidence_support: float
    template_compliance: float
    unresolved_conflicts: int
    overall_score: float
    technical_pass: bool
    promotion_ready: bool = False


@dataclass
class QAReport:
    workbook: str
    sha256: str
    checks: List[CheckResult]
    score: QAScore
    inventory: Dict[str, Any]

    @property
    def findings(self) -> List[QAFinding]:
        return [finding for check in self.checks for finding in check.findings]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "workbook": self.workbook,
            "sha256": self.sha256,
            "checks": [asdict(check) for check in self.checks],
            "score": asdict(self.score),
            "inventory": self.inventory,
        }


def load_workbook_profile(path: Optional[Path]) -> Dict[str, Any]:
    if path is None:
        return {}
    try:
        profile = json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ControlFileError(f"Cannot read workbook profile {path}: {exc}") from exc
    if not isinstance(profile, dict):
        raise ControlFileError(f"Workbook profile {path} must be a JSON object")
    return profile


def _open_workbook(path: Path, *, data_only: bool = False):
    import openpyxl

    return openpyxl.load_workbook(
        path,
        data_only=data_only,
        read_only=False,
        keep_links=True,
        keep_vba=path.suffix.lower() == ".xlsm",
    )


def _formula_cells(workbook) -> Iterable[Tuple[str, str, str]]:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    yield sheet.title, cell.coordinate, str(cell.value)


def _is_broken_formula(formula: str) -> bool:
    upper = formula.upper()
    return upper.startswith("=#") or any(token in upper for token in EXCEL_ERROR_VALUES)


def _is_excel_error(value: Any) -> bool:
    return isinstance(value, str) and value.upper() in EXCEL_ERROR_VALUES


def _workbook_inventory(workbook) -> Dict[str, Any]:
    formula_count = sum(1 for _ in _formula_cells(workbook))
    return {
        "sheet_names": list(workbook.sheetnames),
        "sheet_states": {
            sheet.title: sheet.sheet_state for sheet in workbook.worksheets
        },
        "formula_count": formula_count,
        "defined_name_count": len(workbook.defined_names),
        "external_link_count": len(getattr(workbook, "_external_links", [])),
        "has_vba": getattr(workbook, "vba_archive", None) is not None,
    }


def _failed(
    check_id: str,
    code: str,
    message: str,
    *,
    status: str = "failed",
    sheet: str = "",
    cell: str = "",
    repairable: bool = False,
) -> CheckResult:
    return CheckResult(
        check_id=check_id,
        status=status,
        score=0.0,
        findings=[
            QAFinding(
                check_id=check_id,
                severity="blocking",
                code=code,
                message=message,
                sheet=sheet,
                cell=cell,
                repairable=repairable,
            )
        ],
    )


def _check_result(
    check_id: str,
    findings: List[QAFinding],
    metrics: Dict[str, Any],
) -> CheckResult:
    passed = not findings
    return CheckResult(
        check_id=check_id,
        status="passed" if passed else "failed",
        score=100.0 if passed else 0.0,
        findings=findings,
        metrics=metrics,
    )


def _check_formulas(workbook, values_workbook, template) -> CheckResult:
    findings: List[QAFinding] = []
    workbook_formulas = {
        (sheet, cell): formula for sheet, cell, formula in _formula_cells(workbook)
    }
    template_formulas = (
        {(sheet, cell): formula for sheet, cell, formula in _formula_cells(template)}
        if template
        else {}
    )
    formula_coordinates = sorted(set(workbook_formulas) | set(template_formulas))
    for sheet, cell in formula_coordinates:
        formula = workbook_formulas.get((sheet, cell), "")
        authoritative_formula = template_formulas.get((sheet, cell), "")
        if authoritative_formula and formula != authoritative_formula:
            findings.append(
                QAFinding(
                    check_id="no_broken_formulas",
                    severity="blocking",
                    code="authoritative_formula_mismatch",
                    message=(
                        f"Formula {formula!r} differs from template authority "
                        f"{authoritative_formula!r}"
                    ),
                    sheet=sheet,
                    cell=cell,
                    repairable=True,
                )
            )
            continue
        cached = values_workbook[sheet][cell].value
        formula_broken = _is_broken_formula(formula)
        cached_error = _is_excel_error(cached)
        cached_missing = cached is None
        if not formula_broken and not cached_error and not cached_missing:
            continue
        replacement = template_formulas.get((sheet, cell), "")
        repairable = bool(
            replacement
            and not _is_broken_formula(replacement)
            and (formula_broken or cached_error)
        )
        if formula_broken or cached_error:
            code = "broken_formula"
            message = (
                f"Formula {formula!r} has an invalid expression or cached "
                f"error value {cached!r}"
            )
        else:
            code = "formula_not_calculated"
            message = (
                f"Formula {formula!r} has no cached result; approved "
                "recalculation is required"
            )
        findings.append(
            QAFinding(
                check_id="no_broken_formulas",
                severity="blocking",
                code=code,
                message=message,
                sheet=sheet,
                cell=cell,
                repairable=repairable,
            )
        )
    return _check_result(
        "no_broken_formulas",
        findings,
        {
            "formula_cells_checked": len(formula_coordinates),
            "validation_mode": "formula_text_and_cached_results",
        },
    )


def _check_template(workbook, template, profile: Dict[str, Any]) -> CheckResult:
    expected_order = profile.get("required_sheet_order")
    if expected_order is None and template is not None:
        expected_order = list(template.sheetnames)
    if not expected_order:
        return _failed(
            "template_compliance",
            "template_authority_missing",
            "No template workbook or required_sheet_order was provided",
            status="not_evaluated",
        )

    actual = list(workbook.sheetnames)
    preserve_order = profile.get("preserve_sheet_order", True)
    allow_new = profile.get("allow_new_sheets", False)
    findings: List[QAFinding] = []
    missing = [name for name in expected_order if name not in actual]
    extras = [name for name in actual if name not in expected_order]
    if missing:
        findings.append(
            QAFinding(
                "template_compliance",
                "blocking",
                "missing_sheets",
                f"Missing sheets: {missing}",
            )
        )
    if extras and not allow_new:
        findings.append(
            QAFinding(
                "template_compliance",
                "blocking",
                "unexpected_sheets",
                f"Unexpected sheets: {extras}",
            )
        )
    if preserve_order and [name for name in actual if name in expected_order] != [
        name for name in expected_order if name in actual
    ]:
        findings.append(
            QAFinding(
                "template_compliance",
                "blocking",
                "sheet_order_changed",
                "Workbook sheet order differs from the template authority",
            )
        )
    if template is not None:
        for name in set(actual) & set(template.sheetnames):
            if workbook[name].sheet_state != template[name].sheet_state:
                findings.append(
                    QAFinding(
                        "template_compliance",
                        "blocking",
                        "sheet_visibility_changed",
                        "Sheet visibility differs from template",
                        sheet=name,
                    )
                )
    return _check_result(
        "template_compliance",
        findings,
        {"expected_sheet_count": len(expected_order), "actual_sheet_count": len(actual)},
    )


def _profile_cells(profile: Dict[str, Any], key: str) -> List[Dict[str, Any]]:
    value = profile.get(key, [])
    return value if isinstance(value, list) else []


def _check_total_assertions(workbook, profile: Dict[str, Any], check_id: str) -> CheckResult:
    assertions = [
        item
        for item in _profile_cells(profile, "total_assertions")
        if item.get("check_id") == check_id
    ]
    if not assertions:
        return _failed(
            check_id,
            "rule_not_configured",
            f"No deterministic assertions configured for {check_id}",
            status="not_evaluated",
        )
    findings: List[QAFinding] = []
    for rule in assertions:
        sheet, cell = str(rule.get("sheet", "")), str(rule.get("cell", ""))
        if sheet not in workbook.sheetnames or not cell:
            findings.append(
                QAFinding(
                    check_id,
                    "blocking",
                    "assertion_target_missing",
                    f"Assertion target {sheet}!{cell} does not exist",
                    sheet,
                    cell,
                )
            )
            continue
        value = workbook[sheet][cell].value
        expected = rule.get("expected")
        try:
            actual_number = float(value)
            expected_number = float(expected)
            tolerance = float(rule.get("tolerance", 0))
            difference = abs(actual_number - expected_number)
            valid = (
                math.isfinite(actual_number)
                and math.isfinite(expected_number)
                and math.isfinite(tolerance)
                and tolerance >= 0
                and math.isfinite(difference)
            )
        except (TypeError, ValueError):
            difference = float("inf")
            tolerance = float("nan")
            valid = False
        if not valid or difference > tolerance:
            findings.append(
                QAFinding(
                    check_id,
                    "blocking",
                    "total_out_of_tolerance",
                    f"Expected {expected!r} ± {tolerance}, found {value!r}",
                    sheet,
                    cell,
                )
            )
    return _check_result(
        check_id,
        findings,
        {"assertions_checked": len(assertions)},
    )


def _check_negative_ownership(workbook, profile: Dict[str, Any]) -> CheckResult:
    ranges = _profile_cells(profile, "current_ownership_ranges")
    if not ranges:
        return _failed(
            "negative_current_ownership",
            "rule_not_configured",
            "No current_ownership_ranges are configured",
            status="not_evaluated",
        )
    findings: List[QAFinding] = []
    checked = 0
    for rule in ranges:
        sheet_name, cell_range = str(rule.get("sheet", "")), str(rule.get("range", ""))
        if sheet_name not in workbook.sheetnames or not cell_range:
            findings.append(
                QAFinding(
                    "negative_current_ownership",
                    "blocking",
                    "ownership_range_missing",
                    f"Ownership range {sheet_name}!{cell_range} does not exist",
                )
            )
            continue
        for row in workbook[sheet_name][cell_range]:
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    checked += 1
                    if not math.isfinite(float(cell.value)) or cell.value < 0:
                        findings.append(
                            QAFinding(
                                "negative_current_ownership",
                                "blocking",
                                "negative_interest",
                                f"Invalid current ownership value {cell.value}",
                                sheet_name,
                                cell.coordinate,
                            )
                        )
    return _check_result(
        "negative_current_ownership",
        findings,
        {"numeric_cells_checked": checked},
    )


def _check_duplicate_owners(workbook, profile: Dict[str, Any]) -> CheckResult:
    tables = _profile_cells(profile, "current_owner_columns")
    if not tables:
        return _failed(
            "no_duplicate_owners",
            "rule_not_configured",
            "No current_owner_columns are configured",
            status="not_evaluated",
        )
    findings: List[QAFinding] = []
    checked = 0
    for rule in tables:
        sheet_name = str(rule.get("sheet", ""))
        column = str(rule.get("column", "A"))
        start_row = int(rule.get("start_row", 2))
        default_end_row = (
            workbook[sheet_name].max_row if sheet_name in workbook.sheetnames else 1
        )
        end_row = int(rule.get("end_row", default_end_row))
        if sheet_name not in workbook.sheetnames:
            findings.append(
                QAFinding(
                    "no_duplicate_owners",
                    "blocking",
                    "owner_sheet_missing",
                    f"Owner sheet {sheet_name!r} does not exist",
                )
            )
            continue
        seen: Dict[str, str] = {}
        for row_number in range(start_row, end_row + 1):
            cell = workbook[sheet_name][f"{column}{row_number}"]
            name = " ".join(str(cell.value or "").casefold().split())
            if not name:
                continue
            checked += 1
            if name in seen:
                findings.append(
                    QAFinding(
                        "no_duplicate_owners",
                        "blocking",
                        "duplicate_owner",
                        f"Owner duplicates {seen[name]}",
                        sheet_name,
                        cell.coordinate,
                    )
                )
            else:
                seen[name] = cell.coordinate
    return _check_result(
        "no_duplicate_owners",
        findings,
        {"owners_checked": checked},
    )


def _check_evidence_links(workbook, evidence_root: Path) -> CheckResult:
    findings: List[QAFinding] = []
    checked = 0
    resolved_root = evidence_root.resolve()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not cell.hyperlink:
                    continue
                checked += 1
                target = str(cell.hyperlink.target or "")
                if not target:
                    findings.append(
                        QAFinding(
                            "evidence_links_resolve",
                            "blocking",
                            "empty_or_internal_evidence_link",
                            "Evidence hyperlink has no external target",
                            sheet.title,
                            cell.coordinate,
                        )
                    )
                    continue
                parsed = urlparse(target)
                if parsed.scheme in ("http", "https") and parsed.netloc:
                    findings.append(
                        QAFinding(
                            "evidence_links_resolve",
                            "blocking",
                            "remote_link_unverified",
                            f"Remote evidence link requires an acquisition receipt: {target!r}",
                            sheet.title,
                            cell.coordinate,
                        )
                    )
                    continue
                if parsed.scheme == "file":
                    local = Path(unquote(parsed.path))
                elif not parsed.scheme:
                    local = resolved_root / unquote(target)
                else:
                    local = None
                contained = False
                if local is not None:
                    resolved_local = local.resolve()
                    try:
                        resolved_local.relative_to(resolved_root)
                        contained = True
                    except ValueError:
                        contained = False
                if (
                    local is None
                    or not contained
                    or not resolved_local.is_file()
                ):
                    findings.append(
                        QAFinding(
                            "evidence_links_resolve",
                            "blocking",
                            "unresolved_evidence_link",
                            f"Evidence link does not resolve: {target!r}",
                            sheet.title,
                            cell.coordinate,
                        )
                    )
    if not checked:
        return _failed(
            "evidence_links_resolve",
            "no_evidence_links",
            "No workbook evidence links were available to validate",
            status="not_evaluated",
        )
    return _check_result(
        "evidence_links_resolve",
        findings,
        {"links_checked": checked},
    )


def _unsupported(check_id: str) -> CheckResult:
    return _failed(
        check_id,
        "validator_unavailable",
        f"No deterministic validator is registered for {check_id}",
        status="not_evaluated",
    )


def _component_score(results: Dict[str, CheckResult], check_ids: set) -> float:
    selected = [results[item].score for item in check_ids if item in results]
    return round(sum(selected) / len(selected), 2) if selected else 0.0


def inspect_workbook(
    workbook_path: Path,
    acceptance_tests: List[str],
    *,
    template_path: Optional[Path] = None,
    profile: Optional[Dict[str, Any]] = None,
) -> QAReport:
    """Run configured checks against one immutable workbook snapshot."""
    workbook_path = Path(workbook_path)
    profile = profile or {}
    workbook = _open_workbook(workbook_path)
    values_workbook = _open_workbook(workbook_path, data_only=True)
    template = _open_workbook(template_path) if template_path else None
    inventory = _workbook_inventory(workbook)
    results: List[CheckResult] = []

    try:
        for check_id in acceptance_tests:
            if check_id == "no_broken_formulas":
                result = _check_formulas(workbook, values_workbook, template)
            elif check_id == "template_compliance":
                result = _check_template(workbook, template, profile)
            elif check_id == "negative_current_ownership":
                result = _check_negative_ownership(values_workbook, profile)
            elif check_id == "no_duplicate_owners":
                result = _check_duplicate_owners(workbook, profile)
            elif check_id in TOTAL_CHECKS:
                result = _check_total_assertions(
                    values_workbook, profile, check_id
                )
            elif check_id == "evidence_links_resolve":
                evidence_root = Path(
                    profile.get("evidence_root") or workbook_path.parent
                )
                result = _check_evidence_links(workbook, evidence_root)
            elif check_id == "human_landman_release":
                result = _failed(
                    check_id,
                    "human_approval_pending",
                    "Technical verification cannot satisfy the human release gate",
                    status="pending_approval",
                )
            else:
                result = _unsupported(check_id)
            results.append(result)
    finally:
        workbook.close()
        values_workbook.close()
        if template is not None:
            template.close()

    by_id = {result.check_id: result for result in results}
    technical = [
        result
        for result in results
        if result.check_id != "human_landman_release"
    ]
    technical_pass = bool(technical) and all(result.passed for result in technical)
    unresolved = sum(len(result.findings) for result in results if not result.passed)
    completeness = round(
        100.0 * sum(result.passed for result in results) / len(results), 2
    ) if results else 0.0
    math_score = _component_score(by_id, MATH_CHECKS)
    evidence_score = _component_score(by_id, EVIDENCE_CHECKS)
    template_score = _component_score(by_id, TEMPLATE_CHECKS)
    overall = round(
        0.25 * completeness
        + 0.35 * math_score
        + 0.20 * evidence_score
        + 0.20 * template_score,
        2,
    )
    score = QAScore(
        completeness=completeness,
        mathematical_accuracy=math_score,
        evidence_support=evidence_score,
        template_compliance=template_score,
        unresolved_conflicts=unresolved,
        overall_score=overall,
        technical_pass=technical_pass,
        promotion_ready=False,
    )
    return QAReport(
        workbook=str(workbook_path),
        sha256=sha256_file(workbook_path),
        checks=results,
        score=score,
        inventory=inventory,
    )
