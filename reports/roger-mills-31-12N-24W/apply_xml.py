#!/usr/bin/env python3
"""Apply changeset to NHE2 base via zip/XML surgery, preserving all other parts byte-identically.
Computes fresh cached values for every formula written or whose precedents changed.
"""
import zipfile, shutil, re, pickle, json, datetime, copy
from lxml import etree
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.datetime import to_excel

SRC = 'extracted/files10/31-12N-24W Roger Mills Co. Cursory Title Report (7-1-26) - NHE2.xlsx'
OUT = 'out/31-12N-24W Roger Mills Co. Cursory Title Report (7-1-26) - NHE - Claude FINAL.xlsx'
NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
def q(tag): return f'{{{NS}}}{tag}'

CS = pickle.load(open('qc/changeset.pkl', 'rb'))

wbF = openpyxl.load_workbook(SRC)                 # formulas
wbV = openpyxl.load_workbook(SRC, data_only=True) # cached values

# ---------------- compute new cached values ----------------
computed = {}   # (sheet, coord) -> cached value for formulas
TRACTS = [f'Tract {i}' for i in range(1, 11)]

def cellv(sheet, coord):
    """current cached value, honoring changeset value ops"""
    op = CS.get(sheet, {}).get(coord)
    if op:
        if op[0] == 'v':
            return op[1]
        if op[0] == 'clear':
            return None
    if (sheet, coord) in computed:
        return computed[(sheet, coord)]
    return wbV[sheet][coord].value

def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None

# Tract 2 E-column adjustments from new instrument values AT128 / CN129
E_ADJ = {('Tract 2', 128): 0.00379041530317005, ('Tract 2', 129): -0.00180499576736958}

tractC = {}   # sheet -> {row: cached C}
for t in TRACTS:
    wsV = wbV[t]
    d5 = wsV['D5'].value
    Ev, Dv = {}, {}
    for r in range(10, 204):
        e = num(wsV.cell(r, 5).value)
        if (t, r) in E_ADJ:
            e = (e or 0) + E_ADJ[(t, r)]
        Ev[r] = e
        Dv[r] = wsV.cell(r, 4).value
    denom = sum(e for r, e in Ev.items() if e and e > 0 and str(Dv[r]) != 'The Public')
    Cv = {}
    for r in range(10, 204):
        e = Ev[r]
        if e and e > 0 and str(Dv[r]) != 'The Public' and denom:
            Cv[r] = e / denom * d5
        else:
            Cv[r] = ''
    tractC[t] = Cv
    for r, v in Cv.items():
        coord = f'C{r}'
        if coord in CS.get(t, {}) and CS[t][coord][0] == 'f':
            computed[(t, coord)] = v

# Tract 2 E128/E129/E6 cache refresh (formulas unchanged, inputs changed)
for (t, r), adj in E_ADJ.items():
    old = num(wbV[t].cell(r, 5).value) or 0
    computed[(t, f'E{r}')] = old + adj
computed[('Tract 2', 'E6')] = None  # recompute below via SUBTOTAL parse if formula present

# SUBTOTAL formulas in changeset: compute sum over range from cached values (with adjustments)
SUBRE = re.compile(r'=SUBTOTAL\(9,\s*([A-Z]+)(\d+):([A-Z]+)(\d+)\)')
for sh, ops in CS.items():
    for coord, op in ops.items():
        if op[0] == 'f' and op[1].startswith('=SUBTOTAL'):
            m = SUBRE.match(op[1])
            if not m:
                continue
            col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(4))
            ci = column_index_from_string(col)
            s = 0.0
            for r in range(r1, r2 + 1):
                # honor changeset values in same column (e.g. WI new cols, T2 balancing)
                cc = f'{col}{r}'
                opv = CS.get(sh, {}).get(cc)
                if opv and opv[0] == 'v':
                    v = num(opv[1])
                elif opv and opv[0] == 'clear':
                    v = None
                else:
                    v = num(wbV[sh][cc].value) if r <= wbV[sh].max_row else None
                if v:
                    s += v
            computed[(sh, coord)] = s

# refresh Tract 2 E6 =SUBTOTAL(9,E9:E205) cache
f_e6 = wbF['Tract 2']['E6'].value
if isinstance(f_e6, str) and f_e6.startswith('=SUBTOTAL'):
    m = SUBRE.match(f_e6)
    col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(4))
    s = 0.0
    for r in range(r1, r2 + 1):
        e = num(wbV['Tract 2'].cell(r, column_index_from_string(col)).value)
        if (('Tract 2', r) in E_ADJ):
            e = (e or 0) + E_ADJ[('Tract 2', r)]
        if e:
            s += e
    computed[('Tract 2', 'E6')] = s
else:
    computed.pop(('Tract 2', 'E6'), None)

# Title C cells referencing tract C columns: recompute all formula cells w/ 'Tract'
SUMIFRE = re.compile(r"^=?(?:ROUND\()?SUMIF\('(Tract \d+)'!\$E\$(\d+):\$E\$(\d+),\">0\",'(Tract \d+)'!\$C\$(\d+):\$C\$(\d+)\)(?:,2\))?$")
title_refresh = {}
wsTF = wbF['Title ']
for row in wsTF.iter_rows():
    for c in row:
        f = CS.get('Title ', {}).get(c.coordinate, (None,))
        ftext = f[1] if f[0] == 'f' else (c.value if isinstance(c.value, str) and c.value.startswith('=') else None)
        if not ftext or 'Tract' not in ftext:
            continue
        m = SUMIFRE.match(ftext.replace(' ', '').replace("'Tract", "'Tract ").replace('Tract  ', 'Tract '))
        # simpler: direct parse
        m2 = re.match(r"=(ROUND\()?SUMIF\('(Tract \d+)'!\$E\$(\d+):\$E\$(\d+),\">0\",'(Tract \d+)'!\$C\$(\d+):\$C\$(\d+)\)(,2\))?$", ftext)
        if not m2:
            continue
        t = m2.group(2); r1, r2 = int(m2.group(3)), int(m2.group(4))
        wsVt = wbV[t]
        s = 0.0
        for r in range(r1, min(r2, 203) + 1):
            e = num(wsVt.cell(r, 5).value)
            if (t, r) in E_ADJ:
                e = (e or 0) + E_ADJ[(t, r)]
            if e and e > 0:
                cv = tractC[t].get(r, '')
                if isinstance(cv, (int, float)):
                    s += cv
        if m2.group(1):
            s = round(s, 2)
        title_refresh[c.coordinate] = s
        computed[('Title ', c.coordinate)] = s

# Title C135 / C147 SUM ranges (mostly text now)
for coord, rng in [('C135', ('C', 128, 134)), ('C147', ('C', 142, 146))]:
    col, r1, r2 = rng
    s = 0.0
    for r in range(r1, r2 + 1):
        v = cellv('Title ', f'{col}{r}')
        if num(v) is not None:
            s += num(v)
    computed[('Title ', coord)] = s

# OGL P-column formulas referencing Title C cells
REFRE = re.compile(r"'Title '!C(\d+)")
for coord, op in CS.get('OGL', {}).items():
    if isinstance(op, tuple) and op[0] == 'f' and "'Title '" in op[1]:
        refs = REFRE.findall(op[1])
        body = re.sub(r"'Title '!C\d+", 'X', op[1])
        if refs and re.fullmatch(r'=X(\+X)*', body):
            s = 0.0
            ok = True
            for rr in refs:
                v = cellv('Title ', f'C{rr}')
                if computed.get(('Title ', f'C{rr}')) is not None:
                    v = computed[('Title ', f'C{rr}')]
                if num(v) is None:
                    ok = False
                    break
                s += num(v)
            if ok:
                computed[('OGL', coord)] = s

# attach computed caches into changeset formula ops
for sh, ops in CS.items():
    for coord, op in list(ops.items()):
        if op[0] == 'f':
            ops[coord] = ('f', op[1], computed.get((sh, coord)))

# refresh ops for unchanged-formula cells with stale caches
REFRESH = {}
for (sh, coord), v in computed.items():
    if coord not in CS.get(sh, {}):
        REFRESH.setdefault(sh, {})[coord] = v

json.dump({f'{sh}!{c}': (repr(v)[:40]) for (sh, c), v in computed.items()},
          open('qc/computed_values.json', 'w'), indent=1)
print('computed caches:', len(computed), '| refresh-only cells:', sum(len(v) for v in REFRESH.values()))

# ---------------- XML surgery ----------------
shutil.copyfile(SRC, OUT)
zin = zipfile.ZipFile(SRC)

# map sheet name -> sheet xml path
wbxml = etree.fromstring(zin.read('xl/workbook.xml'))
rels = etree.fromstring(zin.read('xl/_rels/workbook.xml.rels'))
RNS = 'http://schemas.openxmlformats.org/package/2006/relationships'
relmap = {r.get('Id'): r.get('Target') for r in rels.iter(f'{{{RNS}}}Relationship')}
ONS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
sheetpath = {}
for s in wbxml.iter(q('sheet')):
    t = relmap[s.get(f'{{{ONS}}}id')]
    sheetpath[s.get('name')] = 'xl/' + t.lstrip('/')

# styles: prepare no-yellow xf clones
styles = etree.fromstring(zin.read('xl/styles.xml'))
cellXfs = styles.find(q('cellXfs'))
xfs = list(cellXfs)
fills = list(styles.find(q('fills')))
def fill_is_yellow(fid):
    f = fills[fid]
    pf = f.find(q('patternFill'))
    if pf is None: return False
    fg = pf.find(q('fgColor'))
    return pf.get('patternType') == 'solid' and fg is not None and (fg.get('rgb') == 'FFFFFF00')
noyellow_cache = {}
styles_changed = False
def noyellow_xf(sidx, target_fill=0):
    global styles_changed
    key = (sidx, target_fill)
    if key in noyellow_cache: return noyellow_cache[key]
    src = xfs[sidx]
    clone = copy.deepcopy(src)
    clone.set('fillId', str(target_fill))
    cellXfs.append(clone)
    xfs.append(clone)
    newidx = len(xfs) - 1
    cellXfs.set('count', str(len(xfs)))
    noyellow_cache[key] = newidx
    styles_changed = True
    return newidx

# collect O1 fill on OGL for header restyle
def get_cell_el(root, coord):
    r = int(re.match(r'[A-Z]+(\d+)', coord).group(1))
    sd = root.find(q('sheetData'))
    for rowel in sd.iterfind(q('row')):
        if int(rowel.get('r')) == r:
            for cel in rowel.iterfind(q('c')):
                if cel.get('r') == coord:
                    return cel
    return None

def set_cell(root, sheet, coord, op, shared_masters):
    col = re.match(r'([A-Z]+)', coord).group(1)
    rnum = int(re.match(r'[A-Z]+(\d+)', coord).group(1))
    ci = column_index_from_string(col)
    sd = root.find(q('sheetData'))
    # find or create row
    rowel = None; prev = None
    for rl in sd.iterfind(q('row')):
        rn = int(rl.get('r'))
        if rn == rnum: rowel = rl; break
        if rn > rnum: break
        prev = rl
    if rowel is None:
        rowel = etree.SubElement(sd, q('row')); rowel.set('r', str(rnum))
        sd.remove(rowel)
        idx = list(sd).index(prev) + 1 if prev is not None else 0
        sd.insert(idx, rowel)
    # find or create cell
    cel = None; prevc = None
    for cl in rowel.iterfind(q('c')):
        cc = column_index_from_string(re.match(r'([A-Z]+)', cl.get('r')).group(1))
        if cl.get('r') == coord: cel = cl; break
        if cc > ci: break
        prevc = cl
    if cel is None:
        cel = etree.Element(q('c')); cel.set('r', coord)
        # style: prefer same column previous rows, else left neighbor
        sref = None
        for rr in range(rnum - 1, max(rnum - 6, 0), -1):
            e = get_cell_el(root, f'{col}{rr}')
            if e is not None and e.get('s'):
                sref = e.get('s'); break
        if sref is None and prevc is not None and prevc.get('s'):
            sref = prevc.get('s')
        if sref: cel.set('s', sref)
        idx = list(rowel).index(prevc) + 1 if prevc is not None else 0
        rowel.insert(idx, cel)
    # if replacing a shared-formula master, expand group first
    fel = cel.find(q('f'))
    if fel is not None and fel.get('t') == 'shared' and fel.get('ref'):
        shared_masters.append((fel.get('si'), coord))
    # clear content
    for child in list(cel):
        cel.remove(child)
    if 't' in cel.attrib: del cel.attrib['t']
    kind = op[0]
    if kind == 'clear':
        return
    if kind == 'v':
        v = op[1]
        if isinstance(v, (datetime.datetime, datetime.date)):
            etree.SubElement(cel, q('v')).text = repr(to_excel(v))
        elif isinstance(v, bool):
            cel.set('t', 'b'); etree.SubElement(cel, q('v')).text = '1' if v else '0'
        elif isinstance(v, (int, float)):
            etree.SubElement(cel, q('v')).text = repr(v)
        else:
            cel.set('t', 'inlineStr')
            is_el = etree.SubElement(cel, q('is'))
            t_el = etree.SubElement(is_el, q('t'))
            t_el.text = str(v)
            t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
    elif kind == 'f':
        f_el = etree.SubElement(cel, q('f')); f_el.text = op[1][1:]
        cached = op[2]
        if cached is None:
            pass
        elif isinstance(cached, str):
            cel.set('t', 'str')
            etree.SubElement(cel, q('v')).text = cached
        else:
            etree.SubElement(cel, q('v')).text = repr(cached)

def refresh_v(root, coord, val):
    cel = get_cell_el(root, coord)
    if cel is None: return
    fel = cel.find(q('f'))
    if fel is None: return
    vel = cel.find(q('v'))
    if isinstance(val, str):
        cel.set('t', 'str')
    else:
        if 't' in cel.attrib: del cel.attrib['t']
    if vel is None:
        vel = etree.SubElement(cel, q('v'))
    vel.text = str(val) if isinstance(val, str) else repr(val)

# process each sheet
newparts = {}
for sheet, ops in CS.items():
    path = sheetpath[sheet]
    root = etree.fromstring(zin.read(path))
    shared_masters = []
    style_ops = [c for c, o in ops.items() if o[0] == 'style_noyellow']
    cell_ops = {c: o for c, o in ops.items() if o[0] != 'style_noyellow'}
    # expansion prep: collect openpyxl formulas for shared expansion
    wsF = wbF[sheet]
    # order ops by row for determinism
    for coord in sorted(cell_ops, key=lambda k: (int(re.match(r'[A-Z]+(\d+)', k).group(1)), column_index_from_string(re.match(r'([A-Z]+)', k).group(1)))):
        set_cell(root, sheet, coord, cell_ops[coord], shared_masters)
    # expand any shared groups whose master we replaced
    if shared_masters:
        sis = {si for si, _ in shared_masters}
        sd = root.find(q('sheetData'))
        for rowel in sd.iterfind(q('row')):
            for cel in rowel.iterfind(q('c')):
                fel = cel.find(q('f'))
                if fel is not None and fel.get('t') == 'shared' and fel.get('si') in sis and fel.get('ref') is None:
                    cc = cel.get('r')
                    ftext = wsF[cc].value
                    if isinstance(ftext, str) and ftext.startswith('='):
                        fel.attrib.pop('t', None); fel.attrib.pop('si', None); fel.attrib.pop('ref', None)
                        fel.text = ftext[1:]
                    else:
                        cel.remove(fel)
    # refresh stale caches
    for coord, val in REFRESH.get(sheet, {}).items():
        if val is not None:
            refresh_v(root, coord, val)
    # style ops (OGL header highlight removal)
    if style_ops:
        o1 = get_cell_el(root, 'O1')
        target_fill = int(xfs[int(o1.get('s', '0'))].get('fillId', '0')) if o1 is not None else 0
        for sop in style_ops:
            coord = sop.replace('_style', '')
            cel = get_cell_el(root, coord)
            if cel is not None:
                sidx = int(cel.get('s', '0'))
                if fill_is_yellow(int(xfs[sidx].get('fillId', '0'))):
                    cel.set('s', str(noyellow_xf(sidx, target_fill)))
    newparts[path] = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)

# clear yellow fills on cleared Runsheet rows (if any)
# handled generically above only for OGL; check Runsheet cleared-row fills via openpyxl styles
wsR = wbF['Runsheet']
runs_yellow = []
for r in list(range(566, 574)) + [399, 440]:
    for c in range(1, 19):
        cell = wsR.cell(r, c)
        f = cell.fill
        if f and f.patternType == 'solid' and f.fgColor and str(f.fgColor.rgb) == 'FFFFFF00':
            runs_yellow.append(cell.coordinate)
if runs_yellow:
    path = sheetpath['Runsheet']
    root = etree.fromstring(newparts[path])
    for coord in runs_yellow:
        cel = get_cell_el(root, coord)
        if cel is not None:
            sidx = int(cel.get('s', '0'))
            cel.set('s', str(noyellow_xf(sidx, 0)))
    newparts[path] = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)
print('cleared-row yellow cells restyled:', runs_yellow)

# ---- resolved source-in gap highlights: remove yellow fill (Phase 2 / HIGHLIGHT RULE) ----
import os
res_cleared = []
if os.path.exists('qc/gap_resolutions.json'):
    resolutions = json.load(open('qc/gap_resolutions.json'))
    bysheet = {}
    for r in resolutions:
        bysheet.setdefault(r['sheet'], []).append(r['coord'])
    for sheet, coords in bysheet.items():
        path = sheetpath[sheet]
        root = etree.fromstring(newparts.get(path) or zin.read(path))
        for coord in coords:
            cel = get_cell_el(root, coord)
            if cel is not None:
                sidx = int(cel.get('s', '0'))
                fid = int(xfs[sidx].get('fillId', '0'))
                if fill_is_yellow(fid):
                    cel.set('s', str(noyellow_xf(sidx, 0)))
                    res_cleared.append(f'{sheet}!{coord}')
        newparts[path] = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)
print('resolved gap highlights removed:', len(res_cleared), res_cleared)

# workbook.xml: force full recalc on load
wbroot = etree.fromstring(zin.read('xl/workbook.xml'))
calc = wbroot.find(q('calcPr'))
if calc is None:
    calc = etree.SubElement(wbroot, q('calcPr'))
calc.set('fullCalcOnLoad', '1')
newparts['xl/workbook.xml'] = etree.tostring(wbroot, xml_declaration=True, encoding='UTF-8', standalone=True)

if styles_changed:
    newparts['xl/styles.xml'] = etree.tostring(styles, xml_declaration=True, encoding='UTF-8', standalone=True)

# write output zip
zout = zipfile.ZipFile(OUT, 'w', zipfile.ZIP_DEFLATED)
for item in zin.infolist():
    data = newparts.get(item.filename, None)
    if data is None:
        data = zin.read(item.filename)
    zi = zipfile.ZipInfo(item.filename, date_time=item.date_time)
    zi.compress_type = zipfile.ZIP_DEFLATED
    zi.external_attr = item.external_attr
    zout.writestr(zi, data)
zout.close()
print('WROTE', OUT)
print('modified parts:', sorted(newparts))
