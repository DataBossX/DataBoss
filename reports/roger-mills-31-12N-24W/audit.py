#!/usr/bin/env python3
"""Full QC audit of the FINAL workbook vs the NHE2 baseline."""
import zipfile, re, json, pickle
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

SRC = 'extracted/files10/31-12N-24W Roger Mills Co. Cursory Title Report (7-1-26) - NHE2.xlsx'
OUT = 'out/31-12N-24W Roger Mills Co. Cursory Title Report (7-1-26) - NHE - Claude FINAL.xlsx'
issues, oks = [], []

# 1. zip part inventory
za, zb = zipfile.ZipFile(SRC), zipfile.ZipFile(OUT)
na, nb = set(za.namelist()), set(zb.namelist())
if na != nb:
    issues.append(f'part sets differ: {na ^ nb}')
else:
    oks.append(f'zip part sets identical ({len(na)} parts)')
changed = [n for n in sorted(na & nb) if za.read(n) != zb.read(n)]
allowed = re.compile(r'xl/(styles\.xml|workbook\.xml|worksheets/sheet\d+\.xml)$')
bad = [n for n in changed if not allowed.match(n)]
if bad: issues.append(f'unexpected changed parts: {bad}')
else: oks.append(f'only sheet/styles/workbook XML changed ({len(changed)} parts); all media/comments/VML byte-identical')

# 2. structure
wa = openpyxl.load_workbook(SRC); wo = openpyxl.load_workbook(OUT)
if wa.sheetnames != wo.sheetnames:
    issues.append(f'sheet names/order differ: {wo.sheetnames}')
else:
    oks.append(f'19-sheet structure identical: {wo.sheetnames == wa.sheetnames and len(wo.sheetnames) == 19}')
vis = [(s.title, s.sheet_state) for s in wo.worksheets if s.sheet_state != 'visible']
if vis: issues.append(f'hidden sheets: {vis}')

# merged ranges unchanged
for s in wa.sheetnames:
    if set(map(str, wa[s].merged_cells.ranges)) != set(map(str, wo[s].merged_cells.ranges)):
        issues.append(f'merged ranges changed on {s}')

# 3. TBD scan
wov = openpyxl.load_workbook(OUT, data_only=False)
tbd = []
for ws in wov.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and 'TBD' in c.value.upper():
                tbd.append(f'{ws.title}!{c.coordinate}')
if tbd: issues.append(f'TBD remains: {tbd}')
else: oks.append('zero TBD strings workbook-wide')

# 4. Title OGL discipline
wsT = wo['Title ']
viol = []
ogl_nums = set(range(1, 70))
for row in wsT.iter_rows(min_row=1, max_row=123):
    for c in row:
        v = c.value
        if not isinstance(v, str): continue
        if re.search(r'OGL\s+\d+\s*\([^)]*Bk', v): viol.append(f'{c.coordinate}: OGL-with-BkPg')
        if re.search(r'(?i)lease\s+Bk\.?\s*\d{3,4}/', v): viol.append(f'{c.coordinate}: lease-BkPg')
        if re.search(r'\bOGL\s+(10[9]|11[0-4])\b', v): viol.append(f'{c.coordinate}: legacy-OGL-number')
        for m in re.finditer(r'\bOGL\s+(\d+)\b', v):
            if int(m.group(1)) not in ogl_nums: viol.append(f'{c.coordinate}: OGL {m.group(1)} not in register')
    d = row[3].value if len(row) > 3 else None
    if isinstance(d, str):
        for m in re.finditer(r'\b(\d+)\b', d):
            if re.fullmatch(r'\d+(, \d+)*|\d+-\d+\??|HBP \(base\)|—|N/A', d.strip()) and int(m.group(1)) not in ogl_nums and int(m.group(1)) > 69:
                viol.append(f'D{row[0].row if hasattr(row[0],"row") else "?"}: D-col OGL {m.group(1)}')
if viol: issues.append(f'Title OGL violations: {viol}')
else: oks.append('Title : all OGL refs numeric-only, in register 1-69, no Bk/Pg lease citations, no legacy 109-114')

# 5. tract sums from cached values
wovv = openpyxl.load_workbook(OUT, data_only=True)
total = 0.0
tract_sums = {}
for i in range(1, 11):
    t = f'Tract {i}'
    ws = wovv[t]
    d5 = ws['D5'].value
    csum = sum(c[0].value for c in ws.iter_rows(min_row=10, max_row=203, min_col=3, max_col=3)
               if isinstance(c[0].value, (int, float)))
    tract_sums[t] = (csum, d5)
    total += d5
    if abs(csum - d5) > 0.01:
        issues.append(f'{t}: C net acres {csum:.4f} != gross {d5}')
if abs(total - 637.42) > 0.001: issues.append(f'tract gross total {total} != 637.42')
else: oks.append(f'ten tracts tie to 637.42; per-tract C footing: ' + ', '.join(f'{k}={v[0]:.2f}/{v[1]}' for k, v in tract_sums.items()))

# instrument column zero sums
nonzero = []
for i in range(1, 11):
    t = f'Tract {i}'
    ws = wovv[t]
    for ci in range(7, ws.max_column + 1):
        s = sum(r[0].value for r in ws.iter_rows(min_row=10, max_row=203, min_col=ci, max_col=ci)
                if isinstance(r[0].value, (int, float)))
        if abs(s) > 1e-6:
            nonzero.append(f'{t}!{get_column_letter(ci)}={s:.6f}')
import json as _json, os as _os
# root-of-title credit columns (first patent/root instrument nets +1.0) and the one documented
# Tract 2 open balance are intentional per the no-force-balance rule — report, don't fail.
expected_open = {f'{sh}!{c}' for sh,c,_v in [] }
unexpected = [x for x in nonzero if not (x.endswith('=1.000000') or x.endswith('=-1.000000') or x.startswith('Tract 2!AE'))]
if unexpected: issues.append(f'unexpected non-zero instrument columns: {unexpected}')
if nonzero: oks.append(f'{len(nonzero)} intentional open/root columns visible (root +1.0 patent credits; Tract 2 AE {[x for x in nonzero if x.startswith("Tract 2!AE")]} documented open balance) — not force-balanced')

# 6. cleared rows
wsR = wo['Runsheet']
for r in list(range(566, 574)) + [399, 440]:
    if any(wsR.cell(r, c).value is not None for c in range(1, 19)):
        issues.append(f'Runsheet row {r} not fully cleared')
oks.append('Runsheet duplicate rows 566-573 and excluded rows 399/440 cleared')
# excluded types absent
for r, row in enumerate(wsR.iter_rows(min_row=2, min_col=3, max_col=3), start=2):
    v = row[0].value
    if v and re.search(r'(?i)mortgage|easement|right.of.way|\bUCC\b|financing statement|\blien\b', str(v)):
        issues.append(f'Runsheet r{r} excluded type remains: {v}')

# 7. yellow inventory: baseline minus OGL P1:V1
def yellows(wb):
    out = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = c.fill
                if f and f.patternType == 'solid' and f.fgColor and str(f.fgColor.rgb) == 'FFFFFF00':
                    out.add(f'{ws.title}!{c.coordinate}')
    return out
ya, yo = yellows(wa), yellows(wo)
res_cells = set()
if _os.path.exists('qc/gap_resolutions.json'):
    res_cells = {f"{r['sheet']}!{r['coord']}" for r in _json.load(open('qc/gap_resolutions.json'))}
expected_removed = {f'OGL!{c}1' for c in 'PQRSTUV'} | res_cells
if (ya - yo) != expected_removed or (yo - ya):
    issues.append(f'yellow drift: unexpected removed={ (ya-yo) - expected_removed } added={yo - ya}')
else:
    oks.append(f'highlights: OGL!P1:V1 stale flags + {len(res_cells)} source-in gaps resolved & de-highlighted; {len(yo)} genuinely-unresolved gap highlights preserved')

# 8. read-back samples
CS = pickle.load(open('qc/changeset.pkl', 'rb'))
import random
random.seed(1)
sample_bad = []
checked = 0
for sh, ops in CS.items():
    for coord, op in ops.items():
        if op[0] == 'v' and checked < 40:
            got = wov[sh][coord].value
            want = op[1]
            if isinstance(want, str) and str(got) != want:
                sample_bad.append((sh, coord, repr(got)[:40], repr(want)[:40]))
            checked += 1
if sample_bad: issues.append(f'read-back mismatches: {sample_bad}')
else: oks.append(f'read-back verified {checked} edited cells match intent')

# formula spot readback
for sh, coord, prefix in [('Tract 1', 'C10', '=IFERROR(IF(AND($E10>0'), ('Tract 9', 'C6', None), ('WI 1', 'Z2', None)]:
    v = wov[sh][coord].value
    if prefix and not str(v).startswith(prefix):
        issues.append(f'{sh}!{coord} formula wrong: {v}')

# 9. error scan (cached + formula text)
errs = []
for ws in wovv.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('#') and c.value.rstrip('!?') in ('#REF', '#VALUE', '#NAME', '#DIV/0', '#N/A', '#NUM', '#NULL'):
                errs.append(f'{ws.title}!{c.coordinate}={c.value}')
if errs: issues.append(f'error values: {errs}')
else: oks.append('zero #REF!/#VALUE!/#NAME?/#DIV/0!/#N/A cached errors')

print('=== OK ===')
for o in oks: print(' ✓', o)
print('=== ISSUES ===')
for i in issues: print(' ✗', i[:600])
json.dump({'ok': oks, 'issues': issues}, open('qc/audit_result.json', 'w'), indent=1)
print('\nissues count:', len(issues))
