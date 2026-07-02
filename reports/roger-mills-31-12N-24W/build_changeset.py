#!/usr/bin/env python3
"""Build the verified change set: adopt vetted byfable improvements + own finishing edits.
Outputs qc/changeset.pkl : {sheet: {coord: ('v', value) | ('f', formula, cached) | ('clear',) | ('style_noyellow',)}}
Also outputs qc/changeset_report.json summarizing counts per sheet/category for the audit log.
"""
import openpyxl, re, pickle, json, datetime, collections

fN2 = 'extracted/files10/31-12N-24W Roger Mills Co. Cursory Title Report (7-1-26) - NHE2.xlsx'
fBF = 'src/4258af61-3112N24W_Roger_Mills_Co__Cursory_Title_Report__7126___NHEbyfable.xlsx'

wbA  = openpyxl.load_workbook(fN2)                       # formulas
wbAv = openpyxl.load_workbook(fN2, data_only=True)       # cached values
wbB  = openpyxl.load_workbook(fBF)

CS = collections.defaultdict(dict)   # changeset
report = collections.defaultdict(lambda: collections.defaultdict(int))

def isnoise(a, b):
    if a is not None and b is None and str(a) == '':
        return True
    try:
        fa, fb = float(a), float(b)
        return abs(fa - fb) < 1e-9
    except (TypeError, ValueError):
        return False

def colof(coord):
    return re.match(r'([A-Z]+)', coord).group(1)

def rowof(coord):
    return int(re.match(r'[A-Z]+(\d+)', coord).group(1))

def put(sheet, coord, val, cat):
    if isinstance(val, str) and val.startswith('='):
        CS[sheet][coord] = ('f', val, None)     # cached computed later
    elif val is None:
        CS[sheet][coord] = ('clear',)
    else:
        CS[sheet][coord] = ('v', val)
    report[sheet][cat] += 1

# ---------- 1. adopt byfable diffs per policy ----------
ADOPT_ALL = ['OGL', 'PLAT', 'Well 1', 'WI 1', 'WI 2']
for sh in ADOPT_ALL:
    wsA, wsB = wbA[sh], wbB[sh]
    maxr = max(wsA.max_row, wsB.max_row); maxc = max(wsA.max_column, wsB.max_column)
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            a, b = wsA.cell(r, c).value, wsB.cell(r, c).value
            if str(a) == str(b) or isnoise(a, b):
                continue
            coord = wsB.cell(r, c).coordinate
            put(sh, coord, b, 'adopted-byfable')

# Title: adopt all except B151 (own note later)
wsA, wsB = wbA['Title '], wbB['Title ']
maxr = max(wsA.max_row, wsB.max_row); maxc = max(wsA.max_column, wsB.max_column)
for r in range(1, maxr + 1):
    for c in range(1, maxc + 1):
        a, b = wsA.cell(r, c).value, wsB.cell(r, c).value
        if str(a) == str(b) or isnoise(a, b):
            continue
        coord = wsB.cell(r, c).coordinate
        if coord == 'B151':
            continue
        put('Title ', coord, b, 'adopted-byfable')

# Tracts: only C-column IFERROR rewrite, row-7 SUBTOTAL restorations, header fixes, Tract 2 balancing rows
TRACT2_BAL = {'AT128', 'D128', 'CN129', 'D129'}
for t in [f'Tract {i}' for i in range(1, 11)]:
    wsA, wsB = wbA[t], wbB[t]
    maxr = max(wsA.max_row, wsB.max_row); maxc = max(wsA.max_column, wsB.max_column)
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            a, b = wsA.cell(r, c).value, wsB.cell(r, c).value
            if str(a) == str(b) or isnoise(a, b):
                continue
            coord = wsB.cell(r, c).coordinate
            ok = False
            if isinstance(b, str) and b.startswith('=IFERROR(IF(AND($E') and colof(coord) == 'C':
                ok = True
            elif isinstance(b, str) and b.startswith('=SUBTOTAL') and rowof(coord) in (6, 7):
                ok = True
            elif coord in ('C6', 'D6'):
                ok = True
            elif t == 'Tract 2' and coord in TRACT2_BAL:
                ok = True
            if ok:
                put(t, coord, b, 'adopted-byfable')
            else:
                report[t]['rejected-byfable'] += 1

# Runsheet: adopt only phantom OGL cross-ref removals B526/B527
put('Runsheet', 'B526', None, 'phantom-OGL-ref-cleared')
put('Runsheet', 'B527', None, 'phantom-OGL-ref-cleared')

# ---------- 2. own edits ----------
# 2a. Runsheet duplicate block 566-573 (exact dup of 558-565) + excluded rows 399, 440
wsA = wbA['Runsheet']
for r in list(range(566, 574)) + [399, 440]:
    for c in range(1, 19):
        if wsA.cell(r, c).value is not None:
            CS['Runsheet'][wsA.cell(r, c).coordinate] = ('clear',)
    cat = 'dup-row-cleared' if r >= 566 else 'excluded-doc-row-cleared'
    report['Runsheet'][cat] += 1

# 2b. Title G/C/D Bk/Pg strip on OGL references (rows 1-123)
STRIP1 = re.compile(r'(OGL\s+\d+)\s*\(Bk\.?\s*\d{3,4}/\d{1,4}\)')
STRIP2 = re.compile(r'(OGL\s+\d+\s*\()([^)]*?),?\s*Bk\.?\s*\d{3,4}/\d{1,4}\)')
STRIP3 = re.compile(r'[Bb]ase lease Bk\.?\s*1787/0?047')
wsT = wbB['Title ']   # start from byfable text (already adopted)
for r in range(1, 124):
    for c in (2, 3, 4, 7):   # B, C, D, G
        cell = wsT.cell(r, c)
        v = cell.value
        # if we already adopted a change for this coord, base text = adopted value
        coord = cell.coordinate
        if coord in CS['Title '] and CS['Title '][coord][0] == 'v':
            v = CS['Title '][coord][1]
        if not isinstance(v, str):
            continue
        new = STRIP1.sub(r'\1', v)
        new = STRIP2.sub(lambda m: m.group(1) + m.group(2) + ')', new)
        new = STRIP3.sub('base OGL 24', new)
        new = new.replace('Base OGL 24 (Sorenson', 'Base OGL 24 (Sorenson')  # no-op guard
        if new != v:
            put('Title ', coord, new, 'bkpg-stripped')

# 2c. TBD resolutions
TITLE_TBD = {
    'B78':  lambda s: s.replace('successors TBD)', 'successors not established of record — probate/heirship documentation needed)'),
    'C125': lambda s: s.replace('exact WI/NRI TBD', 'exact WI/NRI not of record — assignment exhibits required'),
    'D139': lambda s: s.replace('exact WI/NRI TBD', 'exact WI/NRI not of record — assignment exhibits required'),
    'C129': lambda s: 'N/A - wellbore WI only',
    'C130': lambda s: 'N/A - wellbore WI only',
    'C131': lambda s: 'N/A - wellbore WI only',
    'C132': lambda s: 'N/A - lien/collateral only',
    'C133': lambda s: 'N/A - release only',
    'C142': lambda s: 'N/A - leasehold summary; HBP not concluded',
    'C144': lambda s: 'N/A - wellbore WI only',
    'C145': lambda s: 'N/A - mortgage/lien only',
}
for coord, fn in TITLE_TBD.items():
    base = wsT[coord].value
    if coord in CS['Title '] and CS['Title '][coord][0] == 'v':
        base = CS['Title '][coord][1]
    newv = fn(base if isinstance(base, str) else '')
    put('Title ', coord, newv, 'tbd-resolved')

for r in range(10, 42):
    coord = f'C{r}'
    cur = wbB['WI 1'][coord].value
    if coord in CS['WI 1']:
        cur = CS['WI 1'][coord][1] if CS['WI 1'][coord][0] == 'v' else cur
    if isinstance(cur, str) and cur.strip() == 'TBD':
        put('WI 1', coord, 'N/A - wellbore WI (no NMA)', 'tbd-resolved')

WI2_TBD = {
    'C9':  "1.00 WI / 0.8125 NRI insofar as lessor's interest (springing top lease, 3/16 RI)",
    'C10': "1.00 WI / 0.8125 NRI insofar as lessor's interest (springing top lease, 3/16 RI)",
    'C11': "1.00 WI / 0.8125 NRI insofar as lessor's interest (springing top lease, 3/16 RI)",
    'C12': "1.00 WI / 0.8125 NRI insofar as lessor's interest (springing top lease, 3/16 RI)",
    'C13': "1.00 WI / 0.8125 NRI insofar as lessor's interest (springing top lease, 3/16 RI)",
    'C14': "1.00 WI / 0.8125 NRI insofar as lessor's interest (springing top lease, 3/16 RI)",
    'C15': "Package: 1.00 WI / 0.8125 NRI insofar as lessors' interests (springing top leases, 3/16 RI)",
    'C16': "Open — base-lease HBP/production evidence required (OTC gross production by PUN)",
}
for coord, txt in WI2_TBD.items():
    put('WI 2', coord, txt, 'tbd-resolved')

# 2d. Title B151 own QA note
put('Title ', 'B151',
    "QA 7-1-26 (final assembly, Claude): base workbook = examiner's 7-1-26 original (format truth; all images, comments and styles preserved). "
    "Merged verified prior-pass work: completed OGL register data columns (live title-carried NMA links), WI 1 wellbore chain extended through "
    "Wellbore Assignment 2698/0069 (Martin's Empire, LLC -> Stride Bank, TTEE, filed 5/4/2026), WI 2 Silver Oak top-lease package (OGL 64-69), "
    "OCC-sourced Well 1 data, corrected PLAT labels, Tract 9/10 header corrections, and Tract 2 instrument-conservation suspense rows (flagged VERIFY). "
    "Rejected prior-pass Tract 3 owner-name splits as OCR/parse corruption (kept examiner's originals). Title OGL references normalized to register "
    "numbers (Bk/Pg lives on OGL tab); base-OGL numbers reconciled to the register. Runsheet: phantom OGL 70/71 cross-refs cleared; exact-duplicate "
    "tail rows 566-573 and excluded easement-option / mortgage partial-release rows cleared per report rules (rawdata retains awareness). "
    "No legal facts fabricated; open items remain flagged in-sheet with the specific missing document named. Not a certified title opinion.",
    'qa-note')

# 2e. stale highlight removals: OGL P1:V1 (columns completed)
for c in 'PQRSTUV':
    CS['OGL'][f'{c}1_style'] = ('style_noyellow',)
    report['OGL']['stale-highlight-removed'] += 1

pickle.dump(dict(CS), open('qc/changeset.pkl', 'wb'))
json.dump({k: dict(v) for k, v in report.items()}, open('qc/changeset_report.json', 'w'), indent=1)
tot = sum(sum(v.values()) for v in report.values())
print(json.dumps({k: dict(v) for k, v in report.items()}, indent=1))
print('total ops:', tot)
