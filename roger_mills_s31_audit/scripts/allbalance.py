import openpyxl
wb = openpyxl.load_workbook("sources/CURRENT_BROKEN.xlsx", data_only=True)
for t in range(1,11):
    ws=wb[f"Tract {t}"]
    d5=ws.cell(5,4).value            # tract acres
    c6=ws.cell(6,3).value            # structural header net
    tot=0.0; nz=0
    for r in range(7, ws.max_row+1): # owners start row 7+
        c=ws.cell(r,3).value
        if isinstance(c,(int,float)):
            tot+=c
            if abs(c)>0.005: nz+=1
    bal = "OK" if abs(tot-(d5 or 0))<0.02 else "*** OFF ***"
    print(f"Tract {t}: D5={d5}  owner-sum(rows7+)={tot:.4f}  nonzero_owners={nz}  {bal}")
