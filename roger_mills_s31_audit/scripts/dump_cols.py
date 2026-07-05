import openpyxl, sys
fn, sheet = sys.argv[1], sys.argv[2]
maxr = int(sys.argv[3]) if len(sys.argv)>3 else 60
maxc = int(sys.argv[4]) if len(sys.argv)>4 else 12
wb = openpyxl.load_workbook(fn, data_only=True)
ws = wb[sheet]
for r in range(1, min(maxr, ws.max_row)+1):
    cells=[]
    for cix in range(1, maxc+1):
        c=ws.cell(r,cix)
        if c.value is not None:
            v=str(c.value).replace("\n","\\n")
            if len(v)>45: v=v[:45]+"…"
            cells.append(f"{c.coordinate}={v}")
    if cells:
        print(" | ".join(cells))
