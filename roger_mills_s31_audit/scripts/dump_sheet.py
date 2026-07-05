import openpyxl, sys
fn, sheet = sys.argv[1], sys.argv[2]
maxr = int(sys.argv[3]) if len(sys.argv)>3 else 9999
wb = openpyxl.load_workbook(fn, data_only=False)
ws = wb[sheet]
for row in ws.iter_rows(min_row=1, max_row=min(maxr, ws.max_row)):
    cells=[]
    for c in row:
        if c.value is not None:
            v=str(c.value).replace("\n","\\n")
            cells.append(f"{c.coordinate}={v}")
    if cells:
        print(" | ".join(cells))
