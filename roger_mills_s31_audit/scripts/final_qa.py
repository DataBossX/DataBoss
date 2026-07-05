import openpyxl, zipfile
from openpyxl.workbook.properties import CalcProperties
FN="SECTION31_12N_24W_ROGER_MILLS_FINAL_OWNERSHIP_TITLE_REPORT_CLAUDE_FIXED.xlsx"
wb=openpyxl.load_workbook(FN)
# set full recalc on load properly
wb.calculation=CalcProperties(calcId=0, fullCalcOnLoad=True)
wb.save(FN)

# ---- reopen for QA ----
wb=openpyxl.load_workbook(FN); wbv=openpyxl.load_workbook(FN, data_only=True)
res=[]
def ck(name, ok, detail=""):
    res.append((ok,name,detail)); print(("PASS" if ok else "FAIL"), "-", name, ("| "+detail) if detail else "")

# 1 opens / zip valid
z=zipfile.ZipFile(FN); ck("Workbook zip integrity (no repair)", z.testzip() is None)
# 2 overview first
ck("Overview Map is first tab", wb.sheetnames[0]=="Overview", wb.sheetnames[0])
# 3 sheet count preserved (21 orig + 2 new)
ck("All original sheets preserved (+Audit/Review)", len(wb.sheetnames)==23, f"{len(wb.sheetnames)} sheets")
# 4 tract balances
t=wb["Title"]
BLOCKS=[(1,8,13,80.0),(2,22,95,160.0),(3,104,110,40.0),(4,118,124,80.0),(5,133,134,38.28),
        (6,142,143,51.0),(7,151,154,40.0),(8,163,163,32.8),(9,171,174,75.34),(10,182,184,40.0)]
allbal=True; details=[]
for tno,r0,r1,ac in BLOCKS:
    s=sum(t.cell(r,3).value for r in range(r0,r1+1) if isinstance(t.cell(r,3).value,(int,float)))
    ok=abs(round(s,2)-ac)<0.005; allbal=allbal and ok
    details.append(f"T{tno}={round(s,2)}/{ac}{'' if ok else ' OFF'}")
ck("Every tract balances to its acreage", allbal, "  ".join(details))
# 5 no negatives in any tract net-acre col
neg=[(t.cell(r,2).value, t.cell(r,3).value) for tno,r0,r1,ac in BLOCKS for r in range(r0,r1+1) if isinstance(t.cell(r,3).value,(int,float)) and t.cell(r,3).value<0]
ck("No negative net acres", len(neg)==0, str(neg[:3]))
# 6 no formula errors (cached)
ERR=["#REF!","#VALUE!","#DIV/0!","#NAME?","#NUM!","#NULL!"]
ecount=sum(1 for ws in wbv.worksheets for row in ws.iter_rows() for c in row if isinstance(c.value,str) and c.value in ERR)
ck("No cached formula errors", ecount==0, f"{ecount} found")
# 7 no formula-string errors
scount=sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row if isinstance(c.value,str) and any(e in c.value for e in ERR))
ck("No formula-string errors", scount==0, f"{scount} found")
# 8 OGL columns: numeric OGL or accepted status only
import re
odd=set()
for tno,r0,r1,ac in BLOCKS:
    for r in range(r0,r1+1):
        d=t.cell(r,4).value
        if d is None: continue
        s=str(d).strip()
        if s in ("—","-","","HBP (base)"): continue
        if not re.fullmatch(r'[0-9,\s]+', s): odd.add(s)
ck("OGL columns contain OGL numbers only (or flagged status)", len(odd)==0, str(sorted(odd)))
# 9 no assignment book/page in WI OGL col
wibp=[t.cell(r,4).value for r in range(191,203) if isinstance(t.cell(r,4).value,str) and "/" in t.cell(r,4).value]
ck("No assignment book/page in OGL/Base-OGL column", len(wibp)==0, str(wibp))
# 10 float artifacts gone
msy=0
for tno,r0,r1,ac in BLOCKS:
    for r in range(r0,r1+1):
        v=t.cell(r,3).value
        if isinstance(v,float) and ('e-' in repr(v) or len(repr(v).split('.')[-1])>2): msy+=1
ck("No floating-point artifacts in net acres", msy==0, f"{msy} remain")
# 11 yellow highlight only on intended cells
YEL=set(["C59","C63","C124","C134","C143","C154","C174"])
yfound=set()
for row in t.iter_rows():
    for c in row:
        f=c.fill
        if f and f.fgColor and f.fgColor.rgb and str(f.fgColor.rgb).endswith("FFFF00"):
            yfound.add(c.coordinate)
ck("Yellow fill only on critical review cells", yfound.issubset(YEL) or yfound==YEL, f"yellow={sorted(yfound)}")
# 12 recalc flag
ck("fullCalcOnLoad set (Excel will recompute)", bool(wb.calculation and wb.calculation.fullCalcOnLoad))

print("\n===== SUMMARY:", sum(1 for ok,_,_ in res if ok), "PASS /", sum(1 for ok,_,_ in res if not ok), "FAIL =====")
