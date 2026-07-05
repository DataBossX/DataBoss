import openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
FN="SECTION31_12N_24W_ROGER_MILLS_FINAL_OWNERSHIP_TITLE_REPORT_CLAUDE_FIXED.xlsx"
wb=openpyxl.load_workbook(FN)
audit=json.load(open("audit_edits.json"))

YELLOW=PatternFill("solid", fgColor="FFFF00")
HDR=Font(bold=True, color="FFFFFF"); HDRFILL=PatternFill("solid", fgColor="1F4E78")
thin=Side(style="thin", color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True, vertical="top")

# remove pre-existing versions if re-run
for nm in ["Audit Log","Review Flags"]:
    if nm in wb.sheetnames: del wb[nm]

# ---------- AUDIT LOG ----------
al=wb.create_sheet("Audit Log")
cols=["#","Sheet","Cell","Before","After","Reason / Basis"]
widths=[5,12,8,26,26,80]
for i,(h,w) in enumerate(zip(cols,widths),1):
    cc=al.cell(1,i,h); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD; cc.alignment=Alignment(wrap_text=True,vertical="center")
    al.column_dimensions[chr(64+i)].width=w
al.cell(2,1,"").value=None
r=2
al.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
note=al.cell(r,1,"Audit performed 2026-07-05. Base file: SECTION31_12N_24W_ROGER_MILLS_FINAL_OWNERSHIP_TITLE_REPORT.xlsx. "
     "Controlling ledgers = Tract 1-10 sheets (each already balances exactly to its tract acreage). "
     "Title sheet reconciled to those ledgers; net acres rounded to 2 decimals. No owners, acres, leases or OGL numbers invented.")
note.alignment=wrap; note.font=Font(italic=True)
al.row_dimensions[r].height=48
r=3
for i,(sheet,cell,before,after,reason) in enumerate(audit,1):
    al.cell(r,1,i)
    al.cell(r,2,sheet); al.cell(r,3,cell)
    al.cell(r,4,str(before)); al.cell(r,5,str(after)); al.cell(r,6,reason)
    for cix in range(1,7):
        al.cell(r,cix).border=BORD; al.cell(r,cix).alignment=wrap
    r+=1
al.freeze_panes="A3"

# ---------- REVIEW FLAGS ----------
rf=wb.create_sheet("Review Flags")
cols=["#","Severity","Location","Item","Net Acres / Amount","Why flagged / What is needed"]
widths=[5,10,16,34,16,70]
for i,(h,w) in enumerate(zip(cols,widths),1):
    cc=rf.cell(1,i,h); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD; cc.alignment=Alignment(wrap_text=True,vertical="center")
    rf.column_dimensions[chr(64+i)].width=w
flags=[
 ("HIGH","Title!C124","Tract 4 balance — E/2 NW/4","61.60 NMA",
   "Ownership of 61.60 of 80 NMA undetermined of record; chain over-conveyes / vesting instruments not located. Needs full mineral chain from source instruments (see Curative Requirements)."),
 ("HIGH","Title!C134","Tract 5 balance","37.54 NMA",
   "37.54 of 38.28 NMA undetermined of record; unresolved conveyances in chain. Needs source instruments."),
 ("HIGH","Title!C143","Tract 6 — Ella Pearl Kirk estate","48.80 NMA",
   "Estate/succession not of record. Carried pending probate/heirship (Kirk PB-2022-33). Certified decree/heir determination required."),
 ("HIGH","Title!C154","Tract 7 balance — NE/4 SW/4","36.85 NMA",
   "36.85 of 40 NMA undetermined of record. Needs full mineral chain from source instruments."),
 ("HIGH","Title!C174","Tract 9 balance","18.82 NMA",
   "18.82 of 75.34 NMA undetermined of record; unresolved conveyances. Mitchell->Dold heirship/probate chain must be completed."),
 ("MED","Title!C59 / C63","Tract 2 rollup reconciliation","+/- ~63.70 NMA",
   "Title rollup previously summed to 221.70 vs 160.00 tract acreage. Reconciled to the balanced Tract 2 ledger: Dow C. Bain & Edris O. Bain, JTWROS corrected 63.70->0.00 (intermediate capacity; Trustees capacity retains 9.74), Tim E. Jensen corrected 7.27->9.27. Examiner should confirm against source instruments."),
 ("MED","Title!D22..D163 (multiple)","'HBP (base)' in OGL column","n/a",
   "Several owners carry the lease-status token 'HBP (base)' in the OGL No. column instead of a numeric OGL. Held-by-production carried per client direction; per-owner base OGL numbers are listed on the OGL sheet but were not individually re-keyed to avoid guessing. Populate numeric base OGL per owner when images are pulled."),
 ("MED","Title!B191..B202 (WI 1)","Wellbore assignment chain","WI/NRI not computed",
   "Leasehold/wellbore assignments; exact WI/NRI not determinable without assignment exhibit percentages / OCC review. Book/Page moved out of OGL column (labeled ASSN) and retained in Comments."),
 ("MED","Title!C213","Mortgage/lien burden chain","open",
   "Recorded mortgage/lien burdens (Bk 2356/0001, 2356/0200, 2357/0001, 2357/0474) not yet matched to releases of record."),
 ("LOW","Title!B9 / OGL","First City NBT ETA-Flag Royalty Trust","20.00 NMA",
   "Carried in Tract 1 mineral column but noted as a 1/8 perpetual NPRI (NOT mineral fee). Retained in tract total per prior work product; examiner should confirm whether it belongs in the mineral-acre column or a separate NPRI schedule."),
 ("INFO","Whole workbook","Cursory basis / HBP","n/a",
   "Report is cursory (not certified title). Base leases shown HBP per client direction & OGL sheet; production/spacing/HBP status not independently confirmed; top leases subordinate to prior leases of record. Image/OCC/OTC verification not required for this cursory turn-in per client direction."),
]
r=2
for i,(sev,loc,item,amt,why) in enumerate(flags,1):
    rf.cell(r,1,i); rf.cell(r,2,sev); rf.cell(r,3,loc); rf.cell(r,4,item); rf.cell(r,5,amt); rf.cell(r,6,why)
    for cix in range(1,7):
        rf.cell(r,cix).border=BORD; rf.cell(r,cix).alignment=wrap
    if sev=="HIGH": rf.cell(r,2).fill=YELLOW
    r+=1
rf.freeze_panes="A2"

# ---------- yellow-highlight the genuine review cells on Title ----------
t=wb["Title"]
for cell in ["C59","C63","C124","C134","C143","C154","C174"]:
    t[cell].fill=YELLOW

# ---------- force full recalc on load ----------
try:
    wb.calculation.fullCalcOnLoad=True
except Exception as e:
    print("calc set warn:",e)

wb.save(FN)
print("Finalized. Sheets:", wb.sheetnames)
print("First tab:", wb.sheetnames[0])
