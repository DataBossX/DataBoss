import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

SRC="sources/CURRENT_BROKEN.xlsx"
OUT="SECTION31_12N_24W_ROGER_MILLS_FINAL_OWNERSHIP_TITLE_REPORT_CLAUDE_FIXED.xlsx"
audit=[]  # (sheet, cell, before, after, reason)

wb=openpyxl.load_workbook(SRC)  # keep formulas
tv=wb["Title"]

# ---- Tract block definitions on Title sheet ----
# (tract_no, owner_first_row, owner_last_row, tract_acres, total_cell, balance_row_or_None, total_is_formula)
BLOCKS=[
 (1,   8,  13, 80.0,  "C14",  None, True),
 (2,  22,  95, 160.0, "C96",  None, True),   # SUM formula; fix corrupted cells below
 (3, 104, 110, 40.0,  "C111", None, False),
 (4, 118, 124, 80.0,  "C125", 124, True),
 (5, 133, 134, 38.28, "C135", 134, False),
 (6, 142, 143, 51.0,  "C144", None, True),
 (7, 151, 154, 40.0,  "C155", 154, True),   # C155 refs sheet; owners must still sum to 40
 (8, 163, 163, 32.8,  "C164", None, True),
 (9, 171, 174, 75.34, "C175", 174, True),
 (10,182, 184, 40.0,  "C185", None, False),
]

# ---- STEP A: Tract 2 targeted corrections (reconcile to balanced Tract 2 ledger) ----
# C63 Dow C. Bain & Edris O. Bain, JTWROS: 63.70 -> 0.00 (intermediate owner; nets to 0 per ledger; Trustees capacity C66 holds 9.74)
# C59 Tim E. Jensen: 7.27 -> 9.27 (ledger has two Jensen lines 7.27 + 2.00)
for cell,newv,reason in [
    ("C63",0.0,"Reconcile to balanced Tract 2 ledger: JTWROS capacity nets to 0.00 (Trustees capacity retains 9.74); Title literal 63.70 was a transcription error breaking the 160-ac balance."),
    ("C59",9.27,"Reconcile to Tract 2 ledger: Tim E. Jensen total 9.27 NMA (ledger lines 7.27 + 2.00); Title showed only 7.27."),
]:
    old=tv[cell].value
    tv[cell].value=newv
    audit.append(("Title",cell,old,newv,reason))

# ---- STEP B: round all owner net-acre literals to 2 decimals within tract blocks ----
def round_block(r0,r1):
    for r in range(r0,r1+1):
        c=tv.cell(r,3)  # col C
        if isinstance(c.value,(int,float)) and not isinstance(c.value,bool):
            old=c.value
            new=round(float(old),2)
            if abs(new)==0.0: new=0.0
            if repr(old)!=repr(new):
                c.value=new
                audit.append(("Title",f"C{r}",old,new,"Round Net Acres to 2 decimals (remove floating-point artifact)."))

for (tno,r0,r1,acres,tot,bal,isf) in BLOCKS:
    round_block(r0,r1)

# ---- STEP C: recompute balance-plug rows so each tract equals acreage exactly ----
for (tno,r0,r1,acres,tot,bal,isf) in BLOCKS:
    if bal is None: 
        continue
    others=0.0
    for r in range(r0,r1+1):
        if r==bal: continue
        v=tv.cell(r,3).value
        if isinstance(v,(int,float)): others+=v
    newbal=round(acres-others,2)
    old=tv.cell(bal,3).value
    if isinstance(old,(int,float)) and abs(old-newbal)>0.0001:
        tv.cell(bal,3).value=newbal
        audit.append(("Title",f"C{bal}",old,newbal,f"Recompute balance-of-tract residual so Tract {tno} totals exactly {acres} ac."))

# ---- STEP D: verify/repair literal totals that are not SUM formulas ----
for (tno,r0,r1,acres,tot,bal,isf) in BLOCKS:
    if isf: 
        continue  # formula totals recompute in Excel
    # literal total cell -> set to owner sum (== acres)
    s=0.0
    for r in range(r0,r1+1):
        v=tv.cell(r,3).value
        if isinstance(v,(int,float)): s+=v
    s=round(s,2)
    col=int(tot[1:])
    old=tv[tot].value
    if isinstance(old,(int,float)) and abs(old-s)>0.0001:
        tv[tot].value=s
        audit.append(("Title",tot,old,s,f"Set Tract {tno} REPORT TOTAL to reconciled owner sum {s}."))

# ---- STEP E: WI section - clear book/page values wrongly placed under 'Base OGL's' -> 'ASSN' ----
# rows 191-202 are wellbore/leasehold ASSIGNMENT chain; col D holds Bk/Pg. Book/page preserved in comments col G.
for r in range(191,203):
    d=tv.cell(r,4)
    if isinstance(d.value,str) and "/" in d.value:
        old=d.value
        d.value="ASSN"
        audit.append(("Title",f"D{r}",old,"ASSN","Assignment book/page removed from OGL column per OGL-purity rule; labeled ASSN (conveyance type). Book/Page retained in Comments."))

wb.save(OUT)
print("Saved", OUT)
print("Total audit edits:", len(audit))
import json
json.dump(audit, open("audit_edits.json","w"), indent=0, default=str)
# quick verify: reopen data_only? formulas not recalced by openpyxl; verify literal owner sums
wb2=openpyxl.load_workbook(OUT)
t=wb2["Title"]
def block_sum(r0,r1):
    s=0.0
    for r in range(r0,r1+1):
        v=t.cell(r,3).value
        if isinstance(v,(int,float)): s+=v
    return round(s,2)
print("\nOwner-sum per tract block (must equal acreage):")
for (tno,r0,r1,acres,tot,bal,isf) in BLOCKS:
    s=block_sum(r0,r1)
    print(f"  Tract {tno}: sum={s}  acreage={acres}  {'OK' if abs(s-acres)<0.005 else '*** OFF ***'}")
