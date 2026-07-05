import openpyxl
FN="sources/CURRENT_BROKEN.xlsx"
wbf = openpyxl.load_workbook(FN, data_only=False)   # formulas
wbv = openpyxl.load_workbook(FN, data_only=True)    # cached values

ERR_STRINGS=["#REF!","#VALUE!","#DIV/0!","#NAME?","#NULL!","#NUM!","#N/A"]
print("===== FORMULA ERRORS (cached) =====")
errcount=0
for ws in wbv.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value in ERR_STRINGS:
                print(f"  {ws.title}!{c.coordinate} = {c.value}"); errcount+=1
print("  total:",errcount)

print("\n===== TITLE: tract report totals (cached values) =====")
tv=wbv["Title"]
for r in range(1, tv.max_row+1):
    b=tv.cell(r,2).value
    if isinstance(b,str) and ("TOTAL" in b.upper() or b.strip().startswith("TRACT")):
        c=tv.cell(r,3).value
        print(f"  row{r}: {b[:40]!r} -> C={c!r}")

print("\n===== FLOAT ARTIFACTS in Title col C (Net Acres) =====")
n=0
for r in range(1, tv.max_row+1):
    c=tv.cell(r,3).value
    if isinstance(c,float):
        s=repr(c)
        if 'e-' in s or len(s.split('.')[-1])>4:
            n+=1
            if n<=60: print(f"  C{r} = {c!r}")
print("  total messy floats:",n)

print("\n===== NEGATIVE numeric values in Title col C =====")
for r in range(1, tv.max_row+1):
    c=tv.cell(r,3).value
    if isinstance(c,(int,float)) and c<0:
        print(f"  C{r} = {c}")
