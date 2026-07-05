import openpyxl, re
wb = openpyxl.load_workbook("sources/CURRENT_BROKEN.xlsx", data_only=True)
def norm(s):
    if not s: return ""
    s=str(s).split("\n")[0].lower()
    s=re.sub(r'[^a-z0-9]',' ',s)
    s=re.sub(r'\b(llc|inc|trust|trustee|trustees|company|co|the|a|of|and|revocable|living|family|gst|tax|exempt)\b',' ',s)
    return re.sub(r'\s+',' ',s).strip()
ws=wb["Tract 2"]; ledger={}
for r in range(7, ws.max_row+1):
    d=ws.cell(r,4).value; c=ws.cell(r,3).value
    if d and isinstance(c,(int,float)):
        k=norm(d); ledger[k]=ledger.get(k,0)+c
tv=wb["Title"]; title_sum=0
for r in range(22,96):
    b=tv.cell(r,2).value; c=tv.cell(r,3).value
    if b and isinstance(c,(int,float)):
        k=norm(b); lnet=ledger.get(k); title_sum+=c
        lstr = "NONE" if lnet is None else str(round(lnet,2))
        if lnet is None:
            print(f"  C{r}: title={round(c,2):>8}  ledger={lstr:>8}  NO-MATCH   {str(b).splitlines()[0][:40]}")
        elif abs(round(c,2)-round(lnet,2))>0.02:
            print(f"  C{r}: title={round(c,2):>8}  ledger={lstr:>8}  MISMATCH Δ={round(c-lnet,2):>7}  {str(b).splitlines()[0][:40]}")
print(f"\nTitle Tract2 sum={round(title_sum,2)} (target 160.00)  ledger owner-total={round(sum(ledger.values()),2)}")
