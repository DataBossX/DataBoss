import openpyxl, zipfile
src="sources/CURRENT_BROKEN.xlsx"; dst="test_roundtrip.xlsx"
wb=openpyxl.load_workbook(src)  # formulas
print("loaded sheets:", len(wb.sheetnames), "| first:", wb.sheetnames[0])
wb.save(dst)
wb2=openpyxl.load_workbook(dst)
print("resaved sheets:", len(wb2.sheetnames), "| first:", wb2.sheetnames[0])
print("order preserved:", wb.sheetnames==wb2.sheetnames)
# validate zip integrity
z=zipfile.ZipFile(dst); bad=z.testzip()
print("zip testzip bad:", bad)
# check a formula survived
print("Title C14 formula:", wb2["Title"]["C14"].value)
print("Tract7 total ref  :", wb2["Title"]["C155"].value)
# check merged cells count on a couple sheets
for s in ["Overview","Title","Tract 1"]:
    print(f"  {s}: merged={len(wb2[s].merged_cells.ranges)} dims={wb2[s].dimensions}")
