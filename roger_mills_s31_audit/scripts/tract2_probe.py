import openpyxl
wb = openpyxl.load_workbook("sources/CURRENT_BROKEN.xlsx", data_only=True)
for tname in ["Tract 2"]:
    ws=wb[tname]
    d5=ws.cell(5,4).value
    print(f"{tname}: D5 (tract acres) = {d5}")
    # column C net acres, column D owners
    tot=0.0; rows=[]
    for r in range(1, ws.max_row+1):
        c=ws.cell(r,3).value   # C net acres
        d=ws.cell(r,4).value   # D owner
        if isinstance(c,(int,float)) and c!=0:
            rows.append((r,c,str(d)[:40] if d else ""))
            tot+=c
    print(f"  nonzero-net owners: {len(rows)}  sum(C)={tot:.4f}")
    for r,c,d in rows:
        print(f"    C{r}={c:.4f}  {d}")
