import openpyxl, re
wb=openpyxl.load_workbook("SECTION31_12N_24W_ROGER_MILLS_FINAL_OWNERSHIP_TITLE_REPORT_CLAUDE_FIXED.xlsx")
# 1) formula-string errors anywhere
errs=0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and any(e in c.value for e in ["#REF!","#VALUE!","#DIV/0!","#NAME?","#NUM!","#NULL!"]):
                print("ERR",ws.title,c.coordinate,c.value); errs+=1
print("formula-string errors:",errs)
# 2) WI edit check
t=wb["Title"]
print("\nWI D191-202 after fix:", [t.cell(r,4).value for r in range(191,203)])
# 3) OGL column purity in tract blocks: list distinct non-OGL-number-ish D values
blocks=[(8,13),(22,95),(104,110),(118,124),(133,134),(142,143),(151,154),(163,163),(171,174),(182,184)]
odd=set()
for (r0,r1) in blocks:
    for r in range(r0,r1+1):
        d=t.cell(r,4).value
        if d is None: continue
        s=str(d).strip()
        if s in ("—","-",""): continue
        # OGL numbers: digits, commas, spaces
        if not re.fullmatch(r'[0-9,\s]+', s):
            odd.add(s)
print("\nNon-numeric OGL-column values in tract blocks (need review):")
for o in sorted(odd): print("   ",repr(o))
# 4) open/verify balance items for Review Flags
print("\nOPEN/VERIFY balance rows:")
for r in range(1,t.max_row+1):
    f=t.cell(r,6).value
    if isinstance(f,str) and "OPEN" in f.upper():
        print(f"  Title row {r}: {str(t.cell(r,2).value).splitlines()[0][:45]} | NMA={t.cell(r,3).value}")
