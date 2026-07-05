import openpyxl, sys
fn = sys.argv[1]
wb = openpyxl.load_workbook(fn, data_only=False)
print("FILE:", fn)
print("SHEETS:", len(wb.sheetnames))
for i, ws in enumerate(wb.worksheets):
    print(f"  [{i}] '{ws.title}' dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column} state={ws.sheet_state}")
# images
try:
    for ws in wb.worksheets:
        imgs = getattr(ws, '_images', [])
        if imgs:
            print(f"  IMAGES in '{ws.title}': {len(imgs)}")
except Exception as e:
    print("img err", e)
