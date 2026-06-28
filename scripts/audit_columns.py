#!/usr/bin/env python3
"""audit_columns.py — adversarial column-balance + over-conveyance sweep.

Independent verification pass for the title-report grids. For every instrument
column on every Tract/WI sheet it sums the owner-row body cells (from a recalc'd
copy) and reports any column that does not net to zero, plus any owner whose net
interest exceeds 1.0.

NOTE on interpretation: the workbook's own integrity check is the row-6 RECHECK
formula, which uses SUBTOTAL(9, ...) and therefore deliberately excludes nested
sub-ledger aggregate cells (e.g. Tract 7 rows 161-194) and applies no check to
the working-interest participation columns (WI sheets). A raw column sum will
flag those intentional structures as "imbalanced"; cross-check any hit against
the live RECHECK state from recalc.py before treating it as a defect.

Usage: python scripts/audit_columns.py <recalc'd.xlsx> [<formulas.xlsx>]
The first file should already be recalculated (run recalc.py on a copy first).
The optional second file supplies formulas/headers (defaults to the first).
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter

GRIDS = [f"Tract {i}" for i in range(1, 9)] + ["WI 1", "WI 2"]


def main(values_path, formulas_path=None):
    wbv = openpyxl.load_workbook(values_path, data_only=True)
    wb = openpyxl.load_workbook(formulas_path or values_path, data_only=False)
    cols_checked = 0
    imbalanced = []
    overs = []
    for sn in GRIDS:
        ws, wsv = wb[sn], wbv[sn]
        hcols = [c for c in range(8, ws.max_column + 1)
                 if ws.cell(1, c).value not in (None, "")]
        orows = [r for r in range(8, ws.max_row + 1)
                 if isinstance(ws.cell(r, 5).value, str)
                 and ws.cell(r, 5).value.startswith("=SUM(")]
        if not orows:
            continue
        first, last = min(orows), max(orows)
        for c in hcols:
            cols_checked += 1
            s = 0.0
            nz = 0
            for r in range(first, last + 1):
                v = wsv.cell(r, c).value
                if isinstance(v, (int, float)):
                    s += v
                    if abs(v) > 1e-12:
                        nz += 1
            if abs(s) > 1e-6 and nz > 0:
                imbalanced.append((sn, get_column_letter(c), round(s, 8),
                                   ws.cell(3, c).value))
        for r in range(first, last + 1):
            e = wsv.cell(r, 5).value
            nm = ws.cell(r, 4).value
            if isinstance(e, (int, float)) and e > 1.0001 and isinstance(nm, str) and nm.strip():
                overs.append((sn, "E%d" % r, round(e, 4), nm[:36]))

    print("instrument columns checked:", cols_checked)
    print("\ncolumns not netting to zero (raw sum; verify vs live RECHECK):")
    for sn, col, s, bp in imbalanced:
        print("   %-8s col %-3s sum=%-14s  %s" % (sn, col, s, bp))
    print("\nowners netting > 1.0 (over-conveyance candidates):")
    for sn, cell, e, nm in overs:
        print("   %-8s %-5s net=%-8s  %s" % (sn, cell, e, nm))


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
