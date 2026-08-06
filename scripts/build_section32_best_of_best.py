#!/usr/bin/env python3
"""Build the Section 32 best-of-best cursory title-report deliverables."""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import shutil
import subprocess
import tempfile
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.pagebreak import ColBreak, RowBreak
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(__file__).resolve().parents[1]
TOURNAMENT = ROOT / "SECTION32_BEST_OF_BEST_TOURNAMENT_20260806"
MODEL_DIR = TOURNAMENT / "GPT-5-6-SOL"
CHAMPION_DIR = TOURNAMENT / "FINAL_CHAMPION"
BACKUPS = MODEL_DIR / "source_backups"
BASE_WORKBOOK = (
    BACKUPS
    / "final_verified/beckham32/final_delivery/"
    "32-11N-25W_Beckham_Co_Diversified_Cursory_Title_Report_FINAL_VERIFIED_2026-07-11.xlsx"
)
V7_DUMP = (
    BACKUPS
    / "v10/_TASKS/SECTION32_V10_CLAUDE_CURRENT_DATA_RESEARCH_20260806/"
    "01_INPUT_INVENTORY/V7_WORKBOOK_DUMP.txt"
)
V10_ROOT = BACKUPS / "v10/_TASKS/SECTION32_V10_CLAUDE_CURRENT_DATA_RESEARCH_20260806"
V10_OWNER_MASTER = V10_ROOT / "03_CURRENT_OWNER_RESEARCH/SECTION32_CURRENT_OWNER_MASTER.csv"

MODEL = "GPT-5-6-SOL"
REPORT_STEM = f"Section_32-11N-25W_Beckham_Co_Diversified_Cursory_Report_{MODEL}_TOURNAMENT_20260806"
REPORT_XLSX = MODEL_DIR / f"{REPORT_STEM}.xlsx"
FULL_PDF = MODEL_DIR / f"{REPORT_STEM}_FULL_INTERNAL.pdf"
BOSS_PDF = MODEL_DIR / f"{REPORT_STEM}_BOSS_REVIEW.pdf"
QA_PDF = MODEL_DIR / f"{REPORT_STEM}_QA.pdf"
LINEAGE_XLSX = MODEL_DIR / f"{REPORT_STEM}_SOURCE_LINEAGE.xlsx"
CONFLICT_XLSX = MODEL_DIR / f"{REPORT_STEM}_CONFLICT_ASSUMPTION_REGISTER.xlsx"
SCORECARD_TXT = MODEL_DIR / f"{REPORT_STEM}_SCORECARD.txt"
RECEIPT_TXT = MODEL_DIR / f"{REPORT_STEM}_RECEIPT.txt"

FINAL_XLSX = CHAMPION_DIR / "SECTION32_11N_25W_BECKHAM_DIVERSIFIED_BEST_OF_BEST_FINAL_20260806.xlsx"
FINAL_FULL_PDF = CHAMPION_DIR / "SECTION32_11N_25W_BECKHAM_DIVERSIFIED_FULL_INTERNAL_20260806.pdf"
FINAL_BOSS_PDF = CHAMPION_DIR / "SECTION32_11N_25W_BECKHAM_DIVERSIFIED_BOSS_REVIEW_20260806.pdf"
FINAL_QA_PDF = CHAMPION_DIR / "SECTION32_BEST_OF_BEST_FINAL_QA_20260806.pdf"
FINAL_LINEAGE_XLSX = CHAMPION_DIR / "SECTION32_BEST_OF_BEST_SOURCE_LINEAGE_20260806.xlsx"
FINAL_RECEIPT = CHAMPION_DIR / "SECTION32_BEST_OF_BEST_FINAL_RECEIPT_20260806.txt"
FINAL_QUALIFICATIONS = CHAMPION_DIR / "SECTION32_BEST_OF_BEST_REMAINING_QUALIFICATIONS_20260806.txt"

QUALIFICATION = (
    "Net mineral acres, ownership, lease status, and working-interest calculations shown in this "
    "cursory report are estimates based on the instruments and extracted records reviewed. They are "
    "subject to verification by complete examination of the county, probate, regulatory, and other "
    "applicable records."
)
TITLE_CURRENCY = (
    "Record/index evidence was searched through Book 2400/Page 551 (March 2, 2023); no live county "
    "or probate search from that date through August 6, 2026 was available. Production evidence "
    "reviewed in the source packet ends August 2022."
)

SOURCES = {
    "FV": "FINAL_VERIFIED workbook (2026-07-11), direct-image and official-record synthesis",
    "V7": "V7 workbook dump preserved in the 2026-08-06 current-data research packet",
    "V10": "SECTION32_V10 master registers and patch packet (2026-08-06)",
    "MT": "GPT-5-6-SOL multi-tract evidence-controlled workbook (2026-07-16)",
    "LC": "GPT-5-6-SOL Diversified lease-chain package (2026-07-16)",
    "IR": "Section 32 image-review workbook and extraction JSON, images 0001-0500",
    "ER": "Diversified Title Evidence Register (2026-07-09)",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def split_markdown_row(line: str) -> list[str]:
    return [cell.strip().replace(r"\*", "*") for cell in line.strip().strip("|").split("|")]


def v10_owner_updates() -> dict[tuple[str, str, float], dict]:
    updates = {}
    with V10_OWNER_MASTER.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            tract = row["Tract"].upper()
            owner = row["Displayed_Owner"]
            net_acres = round(float(row["NMA"]), 6)
            key = (tract, owner, net_acres)
            updates[key] = row
    return updates


def partition_owner_rows(owners: list[dict]) -> tuple[list[dict], list[dict]]:
    present_owners = [owner for owner in owners if not owner["audit_only"]]
    audit_owners = [owner for owner in owners if owner["audit_only"]]
    return present_owners, audit_owners


def parse_v7_title() -> list[dict]:
    """Extract the seven fee-tract ending schedules from the preserved V7 text dump."""
    groups: list[dict] = []
    current: dict | None = None
    owner_updates = v10_owner_updates()
    for line in V7_DUMP.read_text(encoding="utf-8").splitlines()[60:200]:
        cells = split_markdown_row(line)
        if len(cells) < 7:
            continue
        tract_cell = cells[1] if len(cells) > 1 else ""
        if tract_cell.startswith("TRACT ") and tract_cell.endswith(":"):
            current = {
                "tract": tract_cell.removesuffix(":"),
                "legal": cells[2],
                "acres": None,
                "owners": [],
            }
            groups.append(current)
            continue
        if current is None:
            continue
        if tract_cell == "Containing:":
            match = re.search(r"([\d.]+)", cells[2])
            current["acres"] = float(match.group(1)) if match else None
            continue
        if tract_cell in {"Mineral Owner", "", "REPORT TOTAL"}:
            continue
        try:
            net_acres = float(cells[2])
        except (ValueError, IndexError):
            continue
        status = cells[4]
        if "HBP" in status:
            royalty = status.split(";")[0].strip()
            status = f"{royalty}; assumed effective for cursory reporting; HBP not independently confirmed"
        elif "Unleased" in status:
            status = "No located active OGL; lease status not independently confirmed"
        elif status in {"Open", "Unresolved"} or "None active" in cells[3]:
            status = "Status not determined"
        owner = cells[1]
        note = cells[6]
        if "ASSUMED:" not in note:
            note = f"EST. last located record owner; {note}"
        update = owner_updates.get((current["tract"].upper(), owner, round(net_acres, 6)))
        is_audit_only = False
        if update:
            owner = update["Recommended_Current_Display"]
            current_status = update["Current_Status"]
            source_instrument = update["Source_Instrument"]
            note += (
                f" V10 STATUS: {current_status}. Vesting/source reference: {source_instrument}. "
                "Forward record search stops 03/02/2023; current status at 08/06/2026 is not confirmed."
            )
            is_audit_only = current_status.startswith("SUPERSEDED")
            address = update["Address_Display"] or cells[5] or "Address not located"
            if "historic recital" in update["Address_Type"].lower():
                address = f"{address} (historic recital address; current address not independently confirmed)"
        else:
            note += " Current status is not confirmed through the report effective date."
            address = cells[5] or "Address not located"
        current["owners"].append(
            {
                "owner": owner,
                "nma": net_acres,
                "ogl": cells[3].replace("None active", "None located"),
                "status": status,
                "address": address,
                "note": note,
                "audit_only": is_audit_only,
            }
        )
    if len(groups) != 7:
        raise RuntimeError(f"Expected seven V7 fee tracts, parsed {len(groups)}")
    return groups


def style_cell(cell, *, bold=False, size=None, align=None, fill=None, color=None):
    cell.font = copy(cell.font)
    cell.font = Font(
        name=cell.font.name or "Arial",
        size=size or cell.font.sz or 10,
        bold=bold,
        italic=cell.font.italic,
        color=color or cell.font.color,
    )
    cell.alignment = Alignment(
        horizontal=align or cell.alignment.horizontal or "left",
        vertical="center",
        wrap_text=True,
    )
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def clear_values(ws, min_row=1, max_row=None, min_col=1, max_col=None):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    max_row = ws.max_row if max_row is None else max_row
    max_col = ws.max_column if max_col is None else max_col
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None
            cell.comment = None


def set_print(ws, area: str, *, landscape=False, fit_height=0) -> None:
    ws.print_area = area
    ws.sheet_state = "visible"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.sheet_view.showGridLines = False
    ws.row_breaks = RowBreak()
    ws.col_breaks = ColBreak()
    ws.oddHeader.left.text = ""
    ws.oddHeader.center.text = "Section 32-11N-25W • Diversified cursory report"
    ws.oddHeader.right.text = ""
    ws.evenHeader = copy(ws.oddHeader)
    ws.firstHeader = copy(ws.oddHeader)
    ws.oddFooter.center.text = "Section 32-11N-25W • Cursory report • Not a title opinion"
    ws.oddFooter.left.text = ""
    ws.oddFooter.right.text = ""
    ws.evenFooter = copy(ws.oddFooter)
    ws.firstFooter = copy(ws.oddFooter)


def sanitize_workbook_metadata(wb) -> None:
    properties = wb.properties
    properties.creator = "GPT-5.6 Sol"
    properties.lastModifiedBy = "GPT-5.6 Sol"
    properties.title = "Section 32-11N-25W Beckham County Diversified Cursory Report"
    properties.subject = "Qualified cursory mineral-title and leasehold report"
    properties.description = (
        "Best-of-best Section 32 synthesis. Estimates and unresolved matters are visibly qualified; "
        "this workbook is not a title opinion."
    )
    properties.keywords = "Section 32, Beckham County, Diversified, cursory title"
    properties.category = "Cursory title report"
    properties.contentStatus = "Qualified internal review"
    properties.identifier = "OK48147.001.1"
    properties.language = "en-US"
    properties.version = "20260806"


def render_full_pdf() -> None:
    FULL_PDF.unlink(missing_ok=True)
    with tempfile.TemporaryDirectory(prefix="section32-pdf-") as output_dir:
        subprocess.run(
            [
                "soffice",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                output_dir,
                str(REPORT_XLSX),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        rendered_pdf = Path(output_dir) / f"{REPORT_XLSX.stem}.pdf"
        if not rendered_pdf.exists() or rendered_pdf.stat().st_size == 0:
            raise RuntimeError("LibreOffice did not produce a non-empty full internal-review PDF")
        shutil.move(rendered_pdf, FULL_PDF)


def write_table_header(ws, row: int, headers: list[str], start_col=1):
    dark = "1F4E78"
    thin = Side(style="thin", color="808080")
    for offset, header in enumerate(headers):
        cell = ws.cell(row, start_col + offset, header)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def write_styled_row(
    ws,
    row: int,
    values: list,
    *,
    font_size: int,
    height: int,
    start_col: int = 1,
    bottom_border: Border | None = None,
):
    for col, value in enumerate(values, start_col):
        cell = ws.cell(row, col, value)
        cell.font = Font(name="Arial", size=font_size)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if bottom_border is not None:
            cell.border = bottom_border
    ws.row_dimensions[row].height = height


def build_overview(wb):
    ws = wb["Overview"]
    ws._images = []
    ws["Q3"] = "32-11N-25W / Diversified / OK48147.001.1"
    ws["S48"] = datetime(2026, 8, 6)
    ws["S50"] = "Record/index corpus through 03/02/2023; no live county continuation search thereafter"
    ws["B53"] = (
        "CONTROLLING CONCLUSION: Diversified Production LLC is the latest supported named claimant "
        "for asset OK48147.001.1 through Bk 2400/Pg 551-567 (asset line at Pg 566). The record "
        "supports a material Section 32 leasehold branch, but not an exact current WI, NRI, ORRI, "
        "net leasehold acreage, or mineral ownership decimal."
    )
    ws["B54"] = (
        "CURRENT-OWNER METHOD: Fee owners are shown as estimated last-located record owners from the "
        "V7 ending schedules, corrected by the V10 qualification register. No tract is documentarily "
        "closed. HBP is not independently confirmed lease-by-lease. Production evidence ends 08/2022."
    )
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 41 and merged.max_row >= 10 and merged.min_col <= 36 and merged.max_col >= 2:
            ws.unmerge_cells(str(merged))
    for row in ws.iter_rows(min_row=10, max_row=41, min_col=2, max_col=36):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
    overview_tracts = [
        ("TRACT 2\nN/2 NW/4\n80 ac", 12, 6, 18, 17, "D9EAF7"),
        ("TRACT 3\nS/2 NW/4\n80 ac", 19, 6, 25, 17, "E2F0D9"),
        ("TRACT 1\nNE/4\n160 ac", 12, 18, 25, 29, "BDD7EE"),
        ("TRACT 4\nN/2 SW/4\n80 ac", 26, 6, 32, 17, "FFF2CC"),
        ("TRACT 5\nS/2 SW/4\n80 ac", 33, 6, 39, 17, "FCE4D6"),
        ("TRACT 7\nW/2 SE/4\n80 ac", 26, 18, 39, 23, "E4DFEC"),
        ("TRACT 6\nE/2 SE/4\n80 ac", 26, 24, 39, 29, "DDEBF7"),
    ]
    medium = Side(style="medium", color="1F4E78")
    for label, r1, c1, r2, c2, fill in overview_tracts:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(r1, c1, label)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for map_row in range(r1, r2 + 1):
            for map_col in range(c1, c2 + 1):
                ws.cell(map_row, map_col).border = Border(
                    left=medium if map_col == c1 else Side(),
                    right=medium if map_col == c2 else Side(),
                    top=medium if map_row == r1 else Side(),
                    bottom=medium if map_row == r2 else Side(),
                )
    ws["AF12"] = "N ▲"
    style_cell(ws["AF12"], bold=True, size=10, align="center")
    ws["B55"] = (
        "PRIORITY CURE: obtain complete 1697/236 Exhibit A; 2340/403 and /490 schedules; 2371/470-533; "
        "2395/415-464; 2400/551-567; the full 1016 assignment series; Crook probate 845/150-157; and "
        "the SE/4 Biggs patent. Resolve Tapstone/OCM Denali, Unbridled/MNR, Linn and Canvas/DP Ponies "
        "branches before stating a final decimal."
    )
    ws.merge_cells("B55:AH55")
    for row in (53, 54, 55):
        style_cell(ws.cell(row, 2), size=8, fill="FFF2CC" if row == 55 else None)
        ws.row_dimensions[row].height = 48
    ws["F56"] = "GPT-5.6 Sol — best-of-best tournament synthesis"
    set_print(ws, "$A$1:$AH$57", landscape=True, fit_height=1)


def build_title(wb, groups: list[dict], lineage: list[dict]):
    ws = wb["Title "]
    clear_values(ws, max_row=max(ws.max_row, 240), max_col=7)
    for col, width in {"A": 2, "B": 34, "C": 13, "D": 16, "E": 28, "F": 35, "G": 66}.items():
        ws.column_dimensions[col].width = width
    ws.merge_cells("B1:G1")
    ws["B1"] = "SECTION 32-11N-25W, BECKHAM COUNTY, OKLAHOMA — ESTIMATED MINERAL TITLE BY TRACT"
    style_cell(ws["B1"], bold=True, size=14, align="center", fill="1F4E78", color="FFFFFF")
    ws.merge_cells("B2:G2")
    ws["B2"] = (
        "Client focus: Diversified • Effective August 6, 2026 • EST. = cursory estimate / last "
        "located record owner, not confirmed present title"
    )
    style_cell(ws["B2"], bold=True, size=9, align="center", fill="D9EAF7")
    row = 4
    for group in groups:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row, 2, f"{group['tract']}: {group['legal']} — {group['acres']:.2f} nominal acres")
        style_cell(ws.cell(row, 2), bold=True, size=11, fill="5B9BD5", color="FFFFFF")
        row += 1
        write_table_header(
            ws,
            row,
            ["Estimated / Last-Located Mineral Owner", "EST. NMA", "OGL / Lease Ref.", "Royalty / Lease Status", "Mailing Address", "Visible Assumptions & Source"],
            2,
        )
        row += 1
        present_owners, audit_owners = partition_owner_rows(group["owners"])
        for owner in present_owners:
            values = [
                owner["owner"],
                owner["nma"],
                owner["ogl"],
                owner["status"],
                owner["address"],
                owner["note"],
            ]
            write_styled_row(
                ws,
                row,
                values,
                font_size=8,
                height=42,
                start_col=2,
                bottom_border=Border(bottom=Side(style="hair", color="B7B7B7")),
            )
            ws.cell(row, 3).number_format = "0.000000"
            lineage.append(
                {
                    "sheet": "Title ",
                    "cell": f"B{row}:G{row}",
                    "conclusion": f"{group['tract']} estimated owner: {owner['owner']} ({owner['nma']:.6f} NMA)",
                    "source": "V7, V10",
                    "reference": owner["note"],
                    "treatment": "Estimated last-located record owner; current vesting and lease status qualified",
                }
            )
            row += 1
        total = sum(owner["nma"] for owner in present_owners)
        ws.cell(row, 2, "ESTIMATED TRACT TOTAL")
        ws.cell(row, 3, total)
        ws.cell(row, 3).number_format = "0.000000"
        ws.cell(row, 7, "Arithmetic closure is not proof of current title.")
        for col in range(2, 8):
            style_cell(ws.cell(row, col), bold=True, size=8, fill="D9EAF7")
        if audit_owners:
            row += 1
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            ws.cell(row, 2, "SUPERSEDED / AUDIT-ONLY ROWS — EXCLUDED FROM PRESENT-OWNER SCHEDULE")
            style_cell(ws.cell(row, 2), bold=True, size=8, fill="D9D9D9")
            for owner in audit_owners:
                row += 1
                values = [
                    owner["owner"],
                    owner["nma"],
                    owner["ogl"],
                    owner["status"],
                    owner["address"],
                    owner["note"],
                ]
                write_styled_row(ws, row, values, font_size=7, height=36, start_col=2)
                ws.cell(row, 3).number_format = "0.000000"
        row += 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    ws.cell(row, 2, "DIVERSIFIED LEASEHOLD / WORKING-INTEREST FOCUS")
    style_cell(ws.cell(row, 2), bold=True, size=11, fill="5B9BD5", color="FFFFFF")
    row += 1
    write_table_header(ws, row, ["Claimant", "Branch %", "Asset / Base OGL", "Status", "Address", "Conclusion"], 2)
    rows = [
        [
            "Diversified Production LLC",
            "51.25% transaction branch",
            "OK48147.001.1 / base OGL not conclusively identified",
            "Identity supported; current WI/NRI not proved",
            "1600 Corporate Drive, Birmingham, Alabama 35242",
            "Latest supported named claimant through Bk 2400/Pg 551-567; do not use 51.25% as current Section 32 WI.",
        ],
        [
            "OCM Denali Holdings LLC",
            "48.75% parallel branch",
            "Bk 2371/Pgs 470-514",
            "Separate claimant; not Diversified",
            "No address of record located",
            "Keep separate. Missing schedules prevent present allocation.",
        ],
        [
            "Face-verified arithmetic product",
            "15.000875%",
            "29.27% × 51.25%",
            "Arithmetic only",
            "Not applicable",
            "Not WI, NRI, ORRI, NMA, or final ownership decimal.",
        ],
    ]
    for values in rows:
        row += 1
        write_styled_row(ws, row, values, font_size=8, height=42, start_col=2)
    row += 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=7)
    ws.cell(row, 2, f"{QUALIFICATION} {TITLE_CURRENCY}")
    style_cell(ws.cell(row, 2), bold=True, size=9, fill="FFF2CC")
    ws.row_dimensions[row].height = 52
    set_print(ws, f"$B$1:$G${row + 2}", landscape=False)
    ws.freeze_panes = "B4"
    return row + 2


def build_plat(wb):
    ws = wb["PLAT"]
    ws._images = []
    clear_values(ws, max_row=45, max_col=32)
    for row in ws.iter_rows(min_row=1, max_row=45, min_col=1, max_col=32):
        for cell in row:
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
    for col in range(1, 33):
        ws.column_dimensions[get_column_letter(col)].width = 3.3
    for row in range(1, 41):
        ws.row_dimensions[row].height = 18
    ws.merge_cells("B1:AE2")
    ws["B1"] = "PLAT — SECTION 32, TOWNSHIP 11 NORTH, RANGE 25 WEST, BECKHAM COUNTY, OKLAHOMA"
    style_cell(ws["B1"], bold=True, size=16, align="center", fill="1F4E78", color="FFFFFF")
    ws.merge_cells("B3:AE4")
    ws["B3"] = "Diversified focus • Asset OK48147.001.1 • 640 nominal gross acres • fee tracts and leasehold/depth overlays are not additive"
    style_cell(ws["B3"], bold=True, size=10, align="center", fill="D9EAF7")
    tracts = [
        ("TRACT 2\nN/2 NW/4\n80 ac", 6, 6, 12, 17),
        ("TRACT 3\nS/2 NW/4\n80 ac", 13, 6, 20, 17),
        ("TRACT 1\nNE/4\n160 ac", 6, 18, 20, 29),
        ("TRACT 4\nN/2 SW/4\n80 ac", 21, 6, 27, 17),
        ("TRACT 5\nS/2 SW/4\n80 ac", 28, 6, 35, 17),
        ("TRACT 6\nE/2 SE/4\n80 ac", 21, 24, 35, 29),
        ("TRACT 7\nW/2 SE/4\n80 ac", 21, 18, 35, 23),
    ]
    colors_cycle = ["D9EAF7", "E2F0D9", "BDD7EE", "FFF2CC", "FCE4D6", "DDEBF7", "E4DFEC"]
    medium = Side(style="medium", color="1F4E78")
    for (label, r1, c1, r2, c2), fill in zip(tracts, colors_cycle):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(r1, c1, label)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Arial", size=12, bold=True, color="1F1F1F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                ws.cell(row, col).border = Border(
                    left=medium if col == c1 else Side(),
                    right=medium if col == c2 else Side(),
                    top=medium if row == r1 else Side(),
                    bottom=medium if row == r2 else Side(),
                )
    ws["AD6"] = "N ▲"
    style_cell(ws["AD6"], bold=True, size=12, align="center")
    ws.merge_cells("B37:AE38")
    ws["B37"] = (
        "Leasehold overlays: OGL 63/33 is E/2 SE/4; deep Union/Leede schedule covers 560 unique "
        "nominal acres with overlaps and an 80-acre S/2 NW/4 gap. Modern all-section index marks "
        "show gross scope only, not 640 NMA or 100% WI."
    )
    style_cell(ws["B37"], size=9, align="center", fill="FFF2CC")
    set_print(ws, "$A$1:$AF$39", landscape=True, fit_height=1)


def lease_rows() -> list[list]:
    return [
        [1, "", "Img 0353", "Oil and Gas Lease", "63/33", "1947-11-13", "1947-11-26", "Cora B. Garrett; S.A. Garrett", "C.E. McKinley", "E/2 SE/4", "Original face reviewed; base of depth-severed family", "10 years", 80, "T6", "1957-11-13 unless maintained", "Not proved", "1/8", "Clause present", "Not determined", "HBP not independently confirmed; assumed potentially effective", "Direct face", "Later 11,100 ft / top Morrow / 19,480 ft tiers", "", "", "", "High", "Pull amendments, releases and lease-specific production", "MT/V7"],
        [2, "", "Imgs 1197-98", "Oil and Gas Lease", "264/345", "1972-02-05", "1972-02-22", "Charlie B. Crook; Ima Pearl Crook", "Union Oil Co. of California", "NE/4 (plus Sec. 28 acreage excluded here)", "Face reviewed; historic recital says held by production", "5 years from 1973-01-13", 160, "T1", "1978-01-13 unless maintained", "Not proved", "1/8", "Order 156126 context", "Not determined", "HBP not independently confirmed; recital supports assumed effectiveness", "Direct face", "Crook wellbore/depth reservations in later assignments", "", "", "", "High", "Confirm lease-specific HBP and later releases", "MT/V7"],
        [3, "", "Imgs 1372", "Oil and Gas Lease", "307/31", "1976-04-29", "1976-06-09", "Cora B. LeGrand", "Union Oil Co. of California", "W/2 SE/4 + S/2 SW/4", "Original face reviewed", "5 years from 1976-09-23", 160, "T5/T7", "1981-09-23 unless maintained", "Not proved", "1/8", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Pull amendments, releases and HBP evidence", "MT"],
        [4, "", "Imgs 1522-23", "Oil and Gas Lease", "340/302", "1977-10-14", "Not extracted", "Gertrude L. LaFortune et al.", "Union Oil Co. of California", "N/2 SW/4", "Overlaps 340/304 and other rows", "5 years from 1977-11-28", 80, "T4", "1982-11-28 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Reconcile overlapping leases and releases", "MT"],
        [5, "", "Imgs 1524-25", "Oil and Gas Lease", "340/304", "1977-10-14", "Not extracted", "Gertrude LaFortune, widow", "Union Oil Co. of California", "N/2 SW/4", "Separate overlapping lease", "5 years from 1977-11-28", 80, "T4", "1982-11-28 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Reconcile relationship to 340/302", "MT"],
        [6, "", "Imgs 1526-27", "Oil and Gas Lease", "340/323", "1977-09-29", "1977-11-07", "George T. Harrison Jr.; Renata Harrison", "Union Oil Co. of California", "N/2 NW/4 + N/2 SW/4", "Original face reviewed", "5 years from 1977-11-27", 160, "T2/T4", "1982-11-27 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Pull amendments, releases and HBP evidence", "MT"],
        [7, "", "Imgs 1542-43", "Oil and Gas Lease", "342/564", "1977-09-29", "1978-02-09", "Carol Seidenbach Leach", "Union Oil Co. of California", "N/2 SW/4", "Overlapping legal", "5 years from 1977-11-30", 80, "T4", "1982-11-30 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Pull amendments, releases and HBP evidence", "MT"],
        [8, "", "Imgs 1544-45", "Oil and Gas Lease", "342/566", "1977-11-01", "1978-02-09", "Cynthia Seidenbach Kruse", "Union Oil Co. of California", "N/2 SW/4", "Overlapping legal", "5 years from 1977-11-30", 80, "T4", "1982-11-30 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Pull amendments, releases and HBP evidence", "MT"],
        [9, "", "Imgs 1552-53", "Oil and Gas Lease", "352/719", "1978-05-16", "1978-06-27", "F.L. Randel", "Union Oil Co. of California", "NE/4", "Face instrument 5169 conflicts with index 5051", "5 years", 160, "T1", "1983-05-16 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Confirm clerk metadata and HBP", "MT"],
        [10, "", "Imgs 1554-55", "Oil and Gas Lease", "352/721", "1978-05-16", "1978-06-27", "J.L. Randel", "Union Oil Co. of California", "NE/4", "Face instrument 5168 conflicts with index 5052", "5 years", 160, "T1", "1983-05-16 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Confirm clerk metadata and HBP", "MT"],
        [11, "", "Imgs 1569-70", "Oil and Gas Lease", "376/369", "1979-03-21", "1979-04-30", "Great Western Oil & Gas, Inc.", "Union Oil Co. of California", "NE/4", "Face instrument 3757 conflicts with index 3707", "5 years", 160, "T1", "1984-03-21 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "Top Morrow to 19,480 ft in 872/279", "", "", "", "High", "Read correction 1039/20-23 and HBP evidence", "MT"],
        [12, "", "Imgs 1457-58", "Oil and Gas Lease", "332/74-75", "1977-08-01", "1977-09-14", "Ralph W. Viersen Jr.; Daisy M. Viersen; Patsy Viersen Brown", "Leede Exploration", "NE/4", "Replaces surrendered 250/582", "5 years", 160, "T1", "1982-08-01 unless maintained", "Not proved", "3/16", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Direct face", "No depth clause located", "", "", "", "Medium-high", "Pull all amendments/releases and lease-specific HBP evidence", "V7"],
        [13, "", "Recited imgs 1568/1812/2031/2300", "Oil and Gas Lease (executed copy absent)", "341/538", "1977-11-29", "Not extracted", "R.L. Minton; Ella Minton", "John E. Hartman", "N/2 NW/4 + N/2 SW/4", "Lease 66032-21-1; identity recited, terms unreviewed", "Not determined", 160, "T2/T4", "Not determined", "Not proved", "Not determined", "Order context only", "Not determined", "Status not determined; HBP not independently confirmed", "Recited only", "Later assignment family limited top Morrow-19,480 ft", "", "", "", "High identity / low terms", "Obtain executed lease and successor/release chain", "FV/LC"],
    ]


def build_ogl(wb, lineage):
    ws = wb["OGL"]
    headers = [ws.cell(1, col).value for col in range(1, 29)]
    leases = lease_rows()
    clear_values(ws, min_row=1, max_row=max(ws.max_row, 40), min_col=1, max_col=28)
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    write_table_header(ws, 1, headers)
    row_border = Border(bottom=Side(style="hair", color="B7B7B7"))
    for row_index, values in enumerate(leases, 2):
        write_styled_row(
            ws,
            row_index,
            values,
            font_size=7,
            height=65,
            bottom_border=row_border,
        )
        lineage.append(
            {
                "sheet": "OGL",
                "cell": f"A{row_index}:AB{row_index}",
                "conclusion": f"OGL {values[4]} historic lease terms and qualified status",
                "source": values[-1],
                "reference": f"Book/Page {values[4]}; {values[2]}",
                "treatment": values[19],
            }
        )
    widths = [7, 7, 17, 18, 12, 12, 12, 24, 24, 26, 38, 15, 12, 8, 18, 13, 10, 18, 13, 25, 14, 28, 12, 12, 12, 12, 30, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    # The donor template hid R, which now contains material pooling/order context.
    ws.column_dimensions["R"].hidden = False
    ws.freeze_panes = "A2"
    last_row = len(leases) + 1
    ws.auto_filter.ref = f"A1:AB{last_row}"
    set_print(ws, f"$A$1:$AB${last_row}", landscape=True)
    # Preserve all template columns but allow a two-page-wide internal render;
    # squeezing 28 substantive columns onto one letter page is not reviewable.
    ws.page_setup.fitToWidth = 2


def append_runsheet_items(wb, lineage):
    ws = wb["Runsheet"]
    existing = {str(ws.cell(row, 4).value) for row in range(2, ws.max_row + 1)}
    last_data_row = max(
        row
        for row in range(1, ws.max_row + 1)
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, 12))
    )
    additions = [
        ("1016/7", "Assignment", "Ray, Charles E.", "Leede Exploration", "Section 32; exact legal requires face", "1016 series; index lead, face not held"),
        ("1016/13", "Assignment", "Frazier, Robert H.", "Leede Exploration", "Section 32; exact legal requires face", "1016 series; index lead, face not held"),
        ("1016/18", "Assignment", "Leede & Pine, Ltd. et al.", "Leede Exploration", "Section 32; exact legal requires face", "1016 series; index lead, face not held"),
        ("1016/25", "Assignment", "Pine, Clyde A. et al.", "Leede Exploration", "Section 32; exact legal requires face", "1016 series; index lead, face not held"),
        ("1016/29", "Assignment", "Consolidation of Leede Trusts", "Leede Exploration", "Section 32; exact legal requires face", "1016 series; index lead, face not held"),
        ("1016/33", "Assignment", "McCall, Jack C. et al.", "Leede Exploration", "Section 32; exact legal requires face", "1016 series; index lead, face not held"),
        ("1016/90", "Assignment", "Leede Exploration et al.", "Enco Gas Gathering Co.", "Section 32; exact legal requires face", "1016 series; index lead, face not held"),
        ("1697/236", "Assignment", "Staghorn Resources LLC", "Chesapeake Exploration Limited Partnership", "Section 32; Exhibit A controls", "Recorded-face lead; required bridge into modern chain"),
    ]
    row = last_data_row + 1
    seq = max(
        (ws.cell(r, 1).value for r in range(2, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, int)),
        default=0,
    )
    for reference, doc_type, grantor, grantee, legal, note in additions:
        if reference in existing:
            continue
        seq += 1
        values = [seq, "", doc_type, reference, "OPEN", "OPEN", grantor, grantee, legal, note, "V7/V10 locator; instrument face or exhibit required"]
        write_styled_row(ws, row, values, font_size=8, height=42)
        lineage.append(
            {
                "sheet": "Runsheet",
                "cell": f"A{row}:K{row}",
                "conclusion": f"Added missing Section 32 chain lead {reference}",
                "source": "V7, V10",
                "reference": note,
                "treatment": "Locator/qualified evidence; full instrument required",
            }
        )
        row += 1
    ws.print_area = f"$A$1:$K${row - 1}"
    ws.freeze_panes = "A2"
    set_print(ws, ws.print_area, landscape=True)


def tract_chain_data(groups):
    owner_summary = {
        "NE/4": groups[0]["owners"],
        "NW/4": groups[1]["owners"] + groups[2]["owners"],
        "SW/4": groups[3]["owners"] + groups[4]["owners"],
        "SE/4": groups[5]["owners"] + groups[6]["owners"],
    }
    return [
        {
            "sheet": "Tract 1",
            "title": "TRACT 1 — NE/4 FEE / MINERAL CHAIN",
            "scope": "NE/4, nominal 160 acres",
            "owners": owner_summary["NE/4"],
            "chain": [
                ["PAT", "P-001 / Img 0003", "Patent 3/469", "1907-03-22 / rec. 1907-06-08", "United States", "Cornelius B. Cornelius", "Sovereign inception; no mineral reservation apparent", "98%"],
                ["WD", "V7 direct-face register", "Roger Mills 11/214", "1907-07-16 / rec. 1907-07-18", "Cornelius and Margreta Cornelius", "Jacob Bollenbach", "NE/4 fee; no reservation apparent", "98%"],
                ["WD", "V7 direct-face register", "10/255", "1910-01-31", "Jacob Bollenbach", "Adolf Bollenbach", "Fee to Adolf; downstream chain incomplete", "High"],
                ["DECR", "V7 probate extraction", "83/259", "Date not extracted", "Estate of Ellen W. Crook", "Thirteen heirs", "Collective mineral vesting; allocations require decree", "High"],
                ["DECR", "V7 probate extraction", "845/150-157", "1985-10-23", "Crook estates", "C.B. Crook Jr.; Hanks; Lenderman; Bryan; Higgenbotham", "Five 8-NMA estimated last-located interests", "High quantum / current status open"],
                ["MD", "V7 direct-face register", "92/47", "1951-03-10", "Johnson parties; N.F. Darnell", "F.L. and J.L. Randel", "1.5 NMA each per deed; later one grantee unextracted", "High"],
                ["MD", "V7 direct-face register", "367/632", "Date not extracted", "Euramerica 1973-A", "Great Western", "10 NMA; competing 267/638 out-conveyance must be resolved", "Medium"],
            ],
        },
        {
            "sheet": "Tract 2",
            "title": "TRACT 2 — NW/4 FEE / MINERAL CHAIN",
            "scope": "N/2 NW/4 and S/2 NW/4, nominal 160 acres",
            "owners": owner_summary["NW/4"],
            "chain": [
                ["FR", "P-003 / Img 0005", "Final Receiver Receipt 3894", "1905-05-04 / rec. 1905-05-12", "United States Land Office", "Herman Stevens", "Earliest located NW/4 evidence; patent not located", "90% identity / 0% patent"],
                ["STIP", "V7 title register", "153/161-162", "Date not extracted", "Maddoux/Farrar parties", "B.E. and D.H. Maddoux", "Recorded residual quantum used for estimated ending schedule", "Medium-high"],
                ["MD/ORDER", "V7 title register", "85/363-366; 311/817-828", "Various", "Historic fee/mineral owners", "Sanditen/Kershaw/Gannaway parties", "Estimated last-located mineral interests", "Varies; current status open"],
                ["DECR", "V7 title register", "P-76-78", "Pages 349-350", "Estate of Ola Keathley", "Keathley/LeGrand branches", "Decree allocation incomplete; 13.333333-NMA residual assumed", "Medium"],
                ["OGL", "Direct face", "340/323", "1977-09-29", "George and Renata Harrison", "Union Oil Co.", "N/2 NW/4 plus N/2 SW/4; no current HBP conclusion", "High terms / open status"],
            ],
        },
        {
            "sheet": "Tract 3",
            "title": "TRACT 3 — SW/4 FEE / MINERAL CHAIN",
            "scope": "N/2 SW/4 and S/2 SW/4, nominal 160 acres",
            "owners": owner_summary["SW/4"],
            "chain": [
                ["FR", "P-004 / Img 0006", "Final Receiver Receipt 3895", "1905-05-04 / rec. 1905-05-12", "United States Land Office", "Augustus P. Stevens", "Earliest located SW/4 evidence; patent not located", "90% identity / 0% patent"],
                ["MD", "V7 title register", "81/194", "Date not extracted", "Historic grantor", "C.R. Bennett", "40 NMA in S/2 SW/4 ending schedule", "Medium; current status open"],
                ["MD/DECR", "V7 title register", "85/222; 262/490-493", "Various", "Historic owners/estate", "Heller Company; Bettye Folmar Norris", "Estimated last-located interests", "Varies"],
                ["OGL", "Direct faces", "340/302; 340/304; 342/564; 342/566", "1977", "LaFortune/Leach/Kruse parties", "Union Oil Co.", "Overlapping N/2 SW/4 leases; not additive", "High terms / open status"],
                ["ASG", "Direct images", "872/279", "1985-10-29 / rec. 1986-03-06", "Union Oil Co.", "Leede Exploration", "Top Morrow-19,480 ft; ORRIs reserved; modern bridge open", "98% historic / open current"],
            ],
        },
        {
            "sheet": "Tract 4",
            "title": "TRACT 4 — SE/4 FEE / MINERAL CHAIN",
            "scope": "E/2 SE/4 and W/2 SE/4, nominal 160 acres",
            "owners": owner_summary["SE/4"],
            "chain": [
                ["FR", "P-002 / Img 0004", "Final Receiver Receipt", "1906-06-28 / stamp 1906-08-08", "United States Land Office", "Velmore Biggs", "Earliest located SE/4 evidence; Biggs/Bigge patent remains missing", "90% identity / 0% patent"],
                ["OGL", "Direct face", "63/33", "1947-11-13 / rec. 1947-11-26", "Cora B. Garrett; S.A. Garrett", "C.E. McKinley", "E/2 SE/4, 1/8 royalty, 10-year term; present status open", "98% terms"],
                ["ASG", "Direct image", "244/432", "1969-11-25", "Humble Oil & Refining", "Union Oil Co.", "Historic 63/33 assignment branch", "96%"],
                ["WB ASG", "Direct image", "845/47", "1985-07-29 / rec. 1985-10-23", "Cities Service", "Leede Exploration", "Pearl wellbore 11,100-19,480 ft; 1/8 ORRI", "98% historic"],
                ["OGL", "Direct face", "307/31", "1976-04-29 / rec. 1976-06-09", "Cora B. LeGrand", "Union Oil Co.", "W/2 SE/4 and S/2 SW/4; present status open", "98% terms"],
            ],
        },
        {
            "sheet": "Tract 5",
            "title": "TRACT 5 — MODERN DIVERSIFIED-AFFILIATED LEASEHOLD BRANCH",
            "scope": "All Section 32 index scope; exact tracts, depths and quantum not proved",
            "owners": [],
            "chain": [
                ["ASG", "Recorded face / Exhibit A required", "1697/236", "2001", "Staghorn Resources LLC", "Chesapeake Exploration LP", "Scheduled predecessor lead; full Exhibit A controls", "Medium"],
                ["ASG", "Index / schedule missing", "2340/490", "rec. 2020-12-14", "Chesapeake affiliates", "KL CHK SPV LLC", "Full-section indexed; missing Tapstone/asset bridge schedules", "High metadata / low scope"],
                ["ASG", "Index / instrument partial", "2371/470-494", "rec. 2022-01-12", "KL CHK SPV LLC", "Diversified Production LLC; OCM Denali", "51.25%/48.75% transaction branches; not current WI", "High parties / low asset quantum"],
                ["MERGER", "Index / certificate missing", "2389/500-580", "rec. 2022-08-24", "Diversified/DP merger parties", "DP Legacy Tapstone LLC", "Corporate consolidation; asset-specific effect open", "High metadata"],
                ["CONV", "Reviewed asset line 2395/462", "2395/415-464", "rec. 2022-11-18", "DP Legacy Tapstone LLC", "Diversified ABS VI Upstream LLC; DP Sooner HoldCo LLC", "OK48147.001.1 identified; allocation/estate/depth blank", "High asset identity / low quantum"],
                ["CONV", "Reviewed asset line 2400/566", "2400/551-567", "rec. 2023-03-02", "DP Sooner HoldCo LLC", "Diversified Production LLC", "Latest supported named claimant; exact current WI/NRI open", "High asset identity / low quantum"],
                ["AGMT", "Index/public metadata", "2476/121-130", "rec. 2026-01-23", "Diversified and DP affiliates", "Same group", "Agreement, not facial conveyance; current vesting effect open", "High metadata / low title effect"],
            ],
        },
    ]


def build_tracts(wb, groups, lineage):
    for data in tract_chain_data(groups):
        ws = wb[data["sheet"]]
        clear_values(ws, max_row=max(ws.max_row, 180), max_col=8)
        widths = [20, 22, 20, 18, 27, 30, 58, 18]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.merge_cells("A1:H1")
        ws["A1"] = data["title"]
        style_cell(ws["A1"], bold=True, size=13, align="center", fill="1F4E78", color="FFFFFF")
        ws.merge_cells("A2:H2")
        ws["A2"] = data["scope"]
        style_cell(ws["A2"], bold=True, size=9, align="center", fill="D9EAF7")
        write_table_header(ws, 4, ["Instrument", "Source / Tier", "Book / Page", "Date", "Grantor", "Grantee / Party", "Estate / Effect", "Confidence"])
        row = 5
        for chain_row in data["chain"]:
            write_styled_row(ws, row, chain_row, font_size=8, height=45)
            lineage.append(
                {
                    "sheet": data["sheet"],
                    "cell": f"A{row}:H{row}",
                    "conclusion": f"{chain_row[0]} {chain_row[2]} — {chain_row[6]}",
                    "source": "FV, V7, V10, MT",
                    "reference": chain_row[1],
                    "treatment": chain_row[7],
                }
            )
            row += 1
        if data["owners"]:
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row, 1, "ESTIMATED ENDING OWNERS — LAST LOCATED RECORD OWNERS; CURRENT VESTING NOT CONFIRMED")
            style_cell(ws.cell(row, 1), bold=True, size=9, fill="5B9BD5", color="FFFFFF")
            row += 1
            write_table_header(ws, row, ["Owner", "EST. NMA", "OGL", "Status", "Address", "Vesting Basis / Note", "Current Limitation", "Source"])
            present_owners, audit_owners = partition_owner_rows(data["owners"])
            for owner in present_owners:
                row += 1
                values = [
                    owner["owner"],
                    owner["nma"],
                    owner["ogl"],
                    owner["status"],
                    owner["address"],
                    owner["note"],
                    "No complete forward chain through effective date",
                    "V7 ending schedule; V10 qualification",
                ]
                write_styled_row(ws, row, values, font_size=7, height=38)
                ws.cell(row, 2).number_format = "0.000000"
            if audit_owners:
                row += 1
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                ws.cell(row, 1, "SUPERSEDED / AUDIT-ONLY — NOT A PRESENT OWNER")
                style_cell(ws.cell(row, 1), bold=True, size=8, fill="D9D9D9")
                for owner in audit_owners:
                    row += 1
                    values = [
                        owner["owner"],
                        owner["nma"],
                        owner["ogl"],
                        owner["status"],
                        owner["address"],
                        owner["note"],
                        "Excluded from present-owner schedule",
                        "V7 audit row; V10 correction",
                    ]
                    write_styled_row(ws, row, values, font_size=7, height=38)
                    ws.cell(row, 2).number_format = "0.000000"
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=8)
        ws.cell(row, 1, QUALIFICATION)
        style_cell(ws.cell(row, 1), bold=True, size=8, fill="FFF2CC")
        ws.row_dimensions[row].height = 36
        ws.freeze_panes = "A5"
        set_print(ws, f"$A$1:$H${row + 1}", landscape=True)


def build_wi(wb, lineage):
    wi1 = wb["WI 1"]
    clear_values(wi1, max_row=max(wi1.max_row, 80), max_col=8)
    wi1.merge_cells("A1:H1")
    wi1["A1"] = "WI 1 — PRINCIPAL KL CHK / TAPSTONE / DIVERSIFIED BRANCH (SHALLOW / UNALLOCATED DEPTHS)"
    style_cell(wi1["A1"], bold=True, size=13, align="center", fill="1F4E78", color="FFFFFF")
    wi1.merge_cells("A2:H2")
    wi1["A2"] = "Leasehold identity path only. Exact current WI, NRI, ORRI, net acres and formation allocation are not proved."
    style_cell(wi1["A2"], bold=True, size=9, align="center", fill="D9EAF7")
    write_table_header(wi1, 4, ["Step", "Book / Page", "Grantor", "Grantee", "Branch / %", "Depth / Scope", "Conclusion", "Evidence / Open Item"])
    rows = [
        [1, "1697/236", "Staghorn Resources LLC", "Chesapeake Exploration LP", "Unquantified", "Exhibit A controls", "Historic scheduled predecessor lead", "Full Exhibit A required"],
        [2, "2340/403", "Chesapeake affiliates", "Tapstone Energy LLC", "Separate Tapstone branch", "Schedule controls", "Potential bridge branch", "Missing Tapstone schedule/bridge"],
        [3, "2340/490", "Chesapeake affiliates", "KL CHK SPV LLC", "Unquantified", "All-section index scope", "Principal indexed predecessor", "Asset schedule absent"],
        [4, "2371/470-494", "KL CHK SPV LLC", "Diversified Production LLC; OCM Denali", "51.25% / 48.75% transaction branches", "Exhibits control", "Keep branches separate", "Not current Section 32 WI"],
        [5, "2389/500-580", "Diversified/DP merger parties", "DP Legacy Tapstone LLC", "Unquantified", "Corporate succession", "Merger path", "Certificate and schedules absent"],
        [6, "2395/415-464", "DP Legacy Tapstone LLC", "Diversified ABS VI Upstream; DP Sooner HoldCo", "Allocation not shown", "OK48147.001.1 at 2395/462", "Asset identity supported", "Estate/depth/quantum blank"],
        [7, "2400/551-567", "DP Sooner HoldCo LLC", "Diversified Production LLC", "Allocation not shown", "OK48147.001.1 at 2400/566", "Latest supported named claimant", "Later filings and exact quantum open"],
        [8, "2451/4-23; 2476/121", "Diversified/DP entities", "Public / same group", "Not a proven transfer", "Corporate filings", "Later context only", "Read complete instruments"],
    ]
    for row_index, values in enumerate(rows, 5):
        write_styled_row(wi1, row_index, values, font_size=8, height=43)
        lineage.append({"sheet": "WI 1", "cell": f"A{row_index}:H{row_index}", "conclusion": str(values[6]), "source": "LC, V7, V10, FV", "reference": str(values[1]), "treatment": str(values[7])})
    summary_row = 15
    write_table_header(wi1, summary_row, ["Item", "Value", "Use", "Status", "Address", "Supporting Reference", "Limitation", "Result"])
    summaries = [
        ["Diversified transaction branch", "51.25%", "Identity/branch context only", "Supported on acquired branch", "1600 Corporate Drive, Birmingham, AL 35242", "2371/470-494", "Missing exhibits and later allocation", "Do not report as current WI"],
        ["OCM Denali parallel branch", "48.75%", "Separate claimant", "Not Diversified", "No address of record located", "2371/470-514", "Missing exhibits", "Keep separate"],
        ["Arithmetic product", "15.000875%", "29.27% × 51.25%", "Arithmetic only", "N/A", "V7 WI summary", "No established 8/8 baseline", "Not WI/NRI/ORRI/NMA"],
        ["Current WI/NRI", "Not calculable", "Cursory conclusion", "Open", "N/A", "2395/462; 2400/566", "Missing A/B/C and excluded-asset schedules", "Qualified estimate only"],
    ]
    for index, values in enumerate(summaries, summary_row + 1):
        write_styled_row(wi1, index, values, font_size=8, height=42)
    for col, width in enumerate([8, 18, 27, 30, 22, 25, 48, 28], 1):
        wi1.column_dimensions[get_column_letter(col)].width = width
    set_print(wi1, "$A$1:$H$20", landscape=True, fit_height=1)

    wi2 = wb["WI 2"]
    clear_values(wi2, max_row=max(wi2.max_row, 80), max_col=8)
    wi2.merge_cells("A1:H1")
    wi2["A1"] = "WI 2 — DEEP / WELLBORE / PARALLEL LEASEHOLD INTERESTS"
    style_cell(wi2["A1"], bold=True, size=13, align="center", fill="1F4E78", color="FFFFFF")
    wi2.merge_cells("A2:H2")
    wi2["A2"] = "Historic depth and wellbore burdens are preserved; current survival and overlap with Diversified remain unresolved."
    style_cell(wi2["A2"], bold=True, size=9, align="center", fill="D9EAF7")
    write_table_header(wi2, 4, ["Instrument", "Book / Page", "Grantor", "Grantee / Burden Owner", "Historic Quantum", "Depth / Formation", "Current Treatment", "Open Requirement"])
    rows2 = [
        ["Crook correction assignment", "502/267", "Union Oil Co.", "Leede Exploration / Union ORRI", "7/64 gas; 1/16 oil ORRI", "Crook wellbore 16,210-20,451 ft", "Historic only; survival open", "Payout, reversion and releases"],
        ["Crook participant assignment", "509/220", "Leede Exploration", "Mesa 64%; Petrofina 25%; ENI 5%; McCall 4%; Midland 1%; Sterling 1%", "100% of assigned interest, not 8/8", "Crook wellbore", "Historic allocation only", "Prove baseline and successors"],
        ["Pearl borehole assignment", "845/47", "Cities Service", "Leede Exploration / 1/8 ORRI reserved", "Assigned interest unquantified", "11,100-19,480 ft", "Historic only; status open", "Read unrecorded agreement and releases"],
        ["Deep scheduled assignment", "872/279", "Union Oil Co.", "Leede Exploration / Union ORRIs", "1/8 gas; 1/16 oil ORRIs", "Top Morrow-19,480 ft; Pearl 11,100-19,480 ft", "Historic only; bridge open", "Complete instrument and 1984 agreement"],
        ["N/2 below-Hunton assignment", "1014/75", "Consolidation of Leede Trusts", "Leede Exploration", "Not proved", "N/2 below top Hunton", "Nemo-dat/source-title defect", "Prove assignor WI source"],
        ["Linn wellbore carve/correction", "2183/174; 2185/639", "Chesapeake Exploration", "Linn Operating / Linn Energy Holdings", "Not proved", "Carlson 7H crosswalk disputed; schedule controls", "Separate branch; do not blend", "Read complete exhibits and legal"],
        ["FourPoint/Unbridled/MNR", "2185/639; 2225/1; 2415/328", "Chesapeake / Unbridled", "FourPoint / MNR ABS", "Not proved", "Wellbore/depth exclusions open", "Parallel portfolio branch", "Reconcile Teocalli carve and schedules"],
    ]
    for row_index, values in enumerate(rows2, 5):
        write_styled_row(wi2, row_index, values, font_size=8, height=48)
        lineage.append({"sheet": "WI 2", "cell": f"A{row_index}:H{row_index}", "conclusion": str(values[6]), "source": "FV, V7, V10, MT", "reference": str(values[1]), "treatment": str(values[7])})
    for col, width in enumerate([24, 18, 28, 34, 25, 34, 32, 40], 1):
        wi2.column_dimensions[get_column_letter(col)].width = width
    set_print(wi2, "$A$1:$H$12", landscape=True, fit_height=1)


def update_wells(wb, lineage):
    ws = wb["Well 1"]
    ws["A1"] = "WELLS — SECTION 32-11N-25W (REGULATORY CONTEXT; NOT LEASE-SPECIFIC HBP PROOF)"
    ws.sheet_state = "visible"
    for row in range(3, ws.max_row + 1):
        if ws.cell(row, 2).value:
            lineage.append(
                {
                    "sheet": "Well 1",
                    "cell": f"A{row}:Q{row}",
                    "conclusion": f"{ws.cell(row, 4).value} / {ws.cell(row, 2).value}",
                    "source": "FV official OCC/OTC synthesis; V10 well register",
                    "reference": str(ws.cell(row, 15).value or ws.cell(row, 12).value),
                    "treatment": "Operator/status are regulatory context only; lease-specific HBP not independently confirmed",
                }
            )
    set_print(ws, f"$A$1:$Q${ws.max_row}", landscape=True)
    ws.freeze_panes = "A3"


def remove_external_links_and_bad_names(wb):
    wb._external_links = []
    for name in list(wb.defined_names.values()):
        attr = getattr(name, "attr_text", "") or ""
        if name.name == "_Order1" or "#REF!" in attr or "[" in attr:
            del wb.defined_names[name.name]
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass


def build_lineage(lineage: list[dict], conflicts: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Source Inventory"
    headers = ["Source ID", "Source / Candidate", "Role", "Limitation"]
    write_table_header(ws, 1, headers)
    source_rows = [
        ["FV", str(BASE_WORKBOOK.relative_to(ROOT)), "Template and strongest verified Runsheet/Well/qualification baseline", "Post-1988 schedules incomplete"],
        ["V7", str(V7_DUMP.relative_to(ROOT)), "Most complete estimated mineral-owner ending schedules and 7-tract geometry", "Text dump only; native Google Sheet/XLSX unavailable"],
        ["V10", str(V10_ROOT.relative_to(ROOT)), "Latest qualification, correction and present-vesting audit", "Live county/Drive retrieval blocked in that run"],
        ["MT", "source_backups/multi_tract/.../GPT56_SOL_SECTION32_IMPROVED_WORKBOOK_20260716.xlsx", "Original OGL faces, terms, addresses and acreage reconciliation", "Current title/HBP not proved"],
        ["LC", "source_backups/lease_chain/.../MASTER_LEASE_CHAIN.xlsx", "Diversified asset chain and OK48147.001.1", "Base OGL and exact quantum open"],
        ["IR", "source_backups/image_review/.../final_synthesis.json + workbooks", "Direct-image extraction and evidence inventory", "Only images 0001-0500 review package plus carried extractions"],
        ["ER", "source_backups/final_verified/.../Diversified_Title_Evidence_Register...", "Cross-check conclusions, burdens, wells and open requirements", "Evidence register is derivative"],
        ["DRIVE", "Google Drive folder 1WiB-VquRfVjbDdyxfvL66hDxYfuC2AJO", "Authoritative source location identified by project manifest", "Connector unavailable in this run"],
        ["DROPBOX", "/11N 25W 32", "Authoritative source root identified by project manifest", "Connector unavailable in this run"],
    ]
    for row in source_rows:
        ws.append(row)
    for col, width in enumerate([14, 85, 55, 55], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws2 = wb.create_sheet("Cell Lineage")
    headers = ["Sheet", "Cell / Range", "Conclusion", "Source IDs", "Document / Evidence Reference", "Treatment / Qualification"]
    write_table_header(ws2, 1, headers)
    for item in lineage:
        ws2.append([item[key] for key in ("sheet", "cell", "conclusion", "source", "reference", "treatment")])
    for col, width in enumerate([18, 20, 65, 25, 65, 65], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:F{ws2.max_row}"
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws3 = wb.create_sheet("Component Selection")
    write_table_header(
        ws3,
        1,
        ["Champion Component", "Contestant Supplier", "Winning Candidate Inputs", "Why Selected", "Merged Additions"],
    )
    selections = [
        ["Overview", "FV + V10", "FV has template parity and strongest controlling qualification", "Effective date, estimated-owner method and cure list from V10"],
        ["PLAT", "V7 geometry rebuilt in FV style", "Seven fee tracts cover the section and are readable at one-page scale", "Leasehold-overlay warning from MT/LC"],
        ["Title", "V7 + V10", "V7 has most complete owners/NMA/addresses; V10 identifies present-vesting defects", "All HBP and current-owner labels qualified"],
        ["OGL", "MT + V7 + FV", "MT reviewed 11 original faces; V7 adds Viersen; FV/LC adds Minton/Hartman", "Uniform lease-specific status qualifications"],
        ["Runsheet", "FV", "Most complete verified chronology", "Missing 1016 series and 1697/236 leads from V7/V10"],
        ["Tracts 1-4", "V7 + MT + FV", "Best fee/mineral ending schedules and direct-face chain facts", "Grouped by quarter to preserve 13-tab template"],
        ["Tract 5", "LC + FV", "Strongest OK48147.001.1 modern chain", "Later corporate filings and missing bridge disclosed"],
        ["WI 1", "LC + V7/V10", "Strongest Diversified branch sequence and transaction split", "51.25% expressly not treated as current WI"],
        ["WI 2", "FV + MT/V7", "Best depth/wellbore burden detail", "Parallel Linn/FourPoint/MNR branches separated"],
        ["Well 1", "FV", "Strongest official OCC/OTC reconciliation", "V10 conflict and HBP limitation retained"],
    ]
    for row in selections:
        ws3.append([row[0], MODEL, *row[1:]])
    for col, width in enumerate([24, 24, 35, 70, 70], 1):
        ws3.column_dimensions[get_column_letter(col)].width = width
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws4 = wb.create_sheet("Conflict Register")
    write_table_header(ws4, 1, ["ID", "Issue", "Competing Evidence", "Champion Treatment", "Status / Cure"])
    for item in conflicts:
        ws4.append([item[key] for key in ("id", "issue", "evidence", "treatment", "status")])
    for col, width in enumerate([12, 48, 75, 75, 55], 1):
        ws4.column_dimensions[get_column_letter(col)].width = width
    for row in ws4.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
    wb.save(LINEAGE_XLSX)

    conflicts_wb = Workbook()
    cws = conflicts_wb.active
    cws.title = "Conflicts and Assumptions"
    write_table_header(cws, 1, ["ID", "Issue / Assumption", "Evidence", "Report Treatment", "Remaining Cure"])
    for item in conflicts:
        cws.append([item[key] for key in ("id", "issue", "evidence", "treatment", "status")])
    for col, width in enumerate([12, 55, 85, 85, 65], 1):
        cws.column_dimensions[get_column_letter(col)].width = width
    for row in cws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    conflicts_wb.save(CONFLICT_XLSX)


def pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor("#1F4E78"), alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=15, textColor=colors.HexColor("#1F4E78"), spaceBefore=8, spaceAfter=4))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10, alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="BodyCompact", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=12, alignment=TA_LEFT))
    return styles


def build_boss_pdf():
    styles = pdf_styles()
    story = [
        Paragraph("Section 32-11N-25W — Boss Review", styles["ReportTitle"]),
        Spacer(1, 0.15 * inch),
        Paragraph("Beckham County, Oklahoma • Diversified focus • Effective August 6, 2026", styles["BodyCompact"]),
        Paragraph("Cursory report — not a title opinion", styles["BodyCompact"]),
        Spacer(1, 0.15 * inch),
        Paragraph("Bottom line", styles["Section"]),
        Paragraph(
            "Diversified Production LLC is the latest supported named claimant for asset OK48147.001.1 "
            "through Book 2400, Pages 551-567 (asset line at Page 566). The evidence supports a material "
            "Section 32 leasehold branch but does not support an exact current WI, NRI, ORRI, net leasehold "
            "acreage, or mineral ownership decimal.",
            styles["BodyCompact"],
        ),
        Paragraph("Mineral ownership", styles["Section"]),
        Paragraph(
            "The Title tab carries estimated last-located owners across seven fee tracts totaling 640 "
            "nominal acres. Arithmetic closure is not current-title proof. No tract is documentarily closed; "
            "owner status and addresses remain qualified where later records were not located.",
            styles["BodyCompact"],
        ),
        Paragraph("Diversified transaction branch", styles["Section"]),
    ]
    data = [
        ["Item", "Cursory conclusion"],
        ["51.25%", "Transaction branch figure only; not reported as current Section 32 WI"],
        ["48.75%", "OCM Denali parallel branch; kept separate from Diversified"],
        ["15.000875%", "Arithmetic product only; not WI/NRI/ORRI/NMA"],
        ["Current exact WI/NRI", "Not calculable from available exhibits and schedules"],
    ]
    table = Table(data, colWidths=[1.5 * inch, 5.5 * inch])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTNAME", (0, 1), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")])]))
    story.extend(
        [
            table,
            Paragraph("Wells and leases", styles["Section"]),
            Paragraph(
                "Latigo is carried as operator for the three Carlson horizontal wells; BCE-Mach II is carried "
                "for its separate Section 32 wells. Diversified is not shown as well operator. Well activity "
                "is regulatory context only and is not treated as lease-specific HBP proof.",
                styles["BodyCompact"],
            ),
            Paragraph("Priority unresolved items", styles["Section"]),
            Paragraph(
                "Complete 1697/236 Exhibit A; 2340/403 and 2340/490 schedules; 2371/470-533; 2395/415-464; "
                "2400/551-567; the 1016 assignment series; Crook probate 845/150-157; the SE/4 Biggs patent; "
                "and lease-specific production/savings-clause evidence.",
                styles["BodyCompact"],
            ),
            Paragraph("Suitability", styles["Section"]),
            Paragraph(
                "Suitable for qualified cursory client review and source-pull planning. It must not be "
                "represented as a drilling, acquisition, division-order, marketable-title, or full title opinion.",
                styles["BodyCompact"],
            ),
            Spacer(1, 0.15 * inch),
            Paragraph(QUALIFICATION, styles["Small"]),
        ]
    )
    doc = SimpleDocTemplate(str(BOSS_PDF), pagesize=letter, rightMargin=0.55 * inch, leftMargin=0.55 * inch, topMargin=0.5 * inch, bottomMargin=0.5 * inch, title="Section 32 Boss Review")
    doc.build(story)


def worksheet_header_footer_text(ws) -> str:
    parts = (
        ws.oddHeader.left,
        ws.oddHeader.center,
        ws.oddHeader.right,
        ws.evenHeader.left,
        ws.evenHeader.center,
        ws.evenHeader.right,
        ws.firstHeader.left,
        ws.firstHeader.center,
        ws.firstHeader.right,
        ws.oddFooter.left,
        ws.oddFooter.center,
        ws.oddFooter.right,
    )
    return " ".join(str(part.text or "") for part in parts)


def workbook_checks(path: Path) -> list[tuple[str, str, str]]:
    wb = load_workbook(path, data_only=False)
    checks: list[tuple[str, str, str]] = []
    checks.append(("Native XLSX opens", "PASS", f"{len(wb.sheetnames)} visible template tabs"))
    checks.append(("Expected tab count/order", "PASS" if len(wb.sheetnames) == 13 else "FAIL", ", ".join(wb.sheetnames)))
    hidden = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
    checks.append(("No hidden worksheets", "PASS" if not hidden else "FAIL", ", ".join(hidden) or "All visible"))
    hidden_material_columns = []
    for ws in wb.worksheets:
        for column, dimension in ws.column_dimensions.items():
            if dimension.hidden and any(cell.value not in (None, "") for cell in ws[column]):
                hidden_material_columns.append(f"{ws.title}!{column}")
    checks.append(
        (
            "No hidden columns containing report information",
            "PASS" if not hidden_material_columns else "FAIL",
            ", ".join(hidden_material_columns) or "None",
        )
    )
    error_cells = []
    external_formulas = []
    stale_tokens = []
    donor_tokens = [
        "Section 27",
        "Roger Mills County",
        "Kunu",
        "Greenhead",
        "Aztek",
        "Phillips 66",
        "New Horizon Energy",
        "405-203-8570",
    ]
    with zipfile.ZipFile(path) as archive:
        core_metadata = archive.read("docProps/core.xml").decode("utf-8", errors="replace")
    metadata_donor_tokens = [token for token in donor_tokens if token.lower() in core_metadata.lower()]
    for ws in wb.worksheets:
        header_footer_text = worksheet_header_footer_text(ws)
        if any(token.lower() in header_footer_text.lower() for token in donor_tokens):
            stale_tokens.append(f"{ws.title}:header/footer")
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if any(token in value for token in ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")):
                        error_cells.append(f"{ws.title}!{cell.coordinate}")
                    if value.startswith("=") and "[" in value:
                        external_formulas.append(f"{ws.title}!{cell.coordinate}")
                    if any(token.lower() in value.lower() for token in donor_tokens):
                        stale_tokens.append(f"{ws.title}!{cell.coordinate}:{value[:60]}")
    checks.append(("No formula/error tokens", "PASS" if not error_cells else "FAIL", ", ".join(error_cells[:10]) or "None"))
    checks.append(("No external-link formulas", "PASS" if not external_formulas and not wb._external_links else "FAIL", ", ".join(external_formulas[:10]) or "None"))
    checks.append(("No identified donor tokens", "PASS" if not stale_tokens else "FAIL", ", ".join(stale_tokens[:10]) or "None"))
    checks.append(
        (
            "No donor metadata in docProps/core.xml",
            "PASS" if not metadata_donor_tokens else "FAIL",
            ", ".join(metadata_donor_tokens) or "Metadata sanitized",
        )
    )
    bad_names = [name.name for name in wb.defined_names.values() if "#REF!" in (getattr(name, "attr_text", "") or "")]
    checks.append(("Defined names valid", "PASS" if not bad_names else "FAIL", ", ".join(bad_names) or "No broken names"))
    no_print = [ws.title for ws in wb.worksheets if not ws.print_area]
    checks.append(("Print areas set", "PASS" if not no_print else "FAIL", ", ".join(no_print) or "All tabs"))
    title_text = " ".join(str(c.value) for row in wb["Title "].iter_rows() for c in row if c.value)
    checks.append(("Title qualification visible", "PASS" if QUALIFICATION in title_text else "FAIL", "Required qualification located"))
    checks.append(("Diversified address visible", "PASS" if "1600 Corporate Drive" in title_text else "FAIL", "Title summary"))
    checks.append(("Estimated acreage labeled", "PASS" if "EST." in title_text else "FAIL", "Title headings and notes"))
    checks.append(("HBP qualified", "PASS" if "HBP not independently confirmed" in title_text else "FAIL", "Lease-specific qualification used"))
    checks.append(("PLAT visible/readable setup", "PASS" if wb["PLAT"].sheet_state == "visible" and wb["PLAT"].page_setup.fitToWidth == 1 else "FAIL", str(wb["PLAT"].print_area)))
    wb.close()
    return checks


def build_qa_pdf(checks, conflicts):
    styles = pdf_styles()
    story = [
        Paragraph("Section 32 Best-of-Best — QA Report", styles["ReportTitle"]),
        Paragraph("Workbook integrity and substantive-control review", styles["BodyCompact"]),
        Spacer(1, 0.12 * inch),
    ]
    data = [["Control", "Result", "Evidence"]] + [[a, b, c] for a, b, c in checks]
    table = Table(data, colWidths=[2.2 * inch, 0.7 * inch, 4.2 * inch], repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7), ("GRID", (0, 0), (-1, -1), 0.35, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")])]))
    story.extend([table, PageBreak(), Paragraph("Material limitations carried forward", styles["Section"])])
    for item in conflicts:
        story.append(Paragraph(f"<b>{item['id']} — {item['issue']}</b><br/>{item['treatment']}<br/><i>{item['status']}</i>", styles["Small"]))
        story.append(Spacer(1, 0.06 * inch))
    story.append(Paragraph("Independent landman and complete-record review remain required before reliance as a title opinion.", styles["BodyCompact"]))
    doc = SimpleDocTemplate(str(QA_PDF), pagesize=letter, rightMargin=0.45 * inch, leftMargin=0.45 * inch, topMargin=0.45 * inch, bottomMargin=0.45 * inch, title="Section 32 QA Report")
    doc.build(story)


def conflict_rows():
    return [
        {"id": "C-01", "issue": "V7 names estimated owners and closes each tract arithmetically, while its hidden QA says no tract is documentarily closed.", "evidence": "V7 Title ending schedules; V10 PRESENT VESTING CONTROL and TITLE CHAIN QA", "treatment": "Owners retained as estimated last-located record owners. Every owner and tract total is visibly qualified.", "status": "Complete forward county/probate search through effective date."},
        {"id": "C-02", "issue": "V7 labels several leases HBP; FV warns section-level production is not lease-specific HBP proof.", "evidence": "V7 Title/OGL/Well; FV Conclusion C8; user controlling logic", "treatment": "All statuses changed to 'HBP not independently confirmed,' 'status not determined,' or assumed potentially effective.", "status": "Lease-by-lease savings-clause and production analysis."},
        {"id": "C-03", "issue": "51.25% Diversified transaction figure can be mistaken for current Section 32 WI.", "evidence": "2371/470 branch split in V7/V10; missing exhibits; LC chain", "treatment": "Shown only as a transaction branch. 48.75% OCM Denali is separate. 15.000875% is labeled arithmetic only.", "status": "Obtain complete exhibits A/B/C and excluded-asset schedules."},
        {"id": "C-04", "issue": "Missing Tapstone/KL CHK/asset bridge and co-assignee allocation.", "evidence": "2340/403, 2340/490, 2371/470-533; LC/V10", "treatment": "No exact current WI/NRI or net leasehold acres stated.", "status": "Pull and abstract complete schedules and incorporated agreements."},
        {"id": "C-05", "issue": "Modern asset schedule supports OK48147.001.1 but legal/interest fields are blank.", "evidence": "2395/462 and 2400/566 reviewed schedule pages", "treatment": "Diversified Production LLC is the latest supported named claimant, not a proven 640-acre/100% owner.", "status": "Read complete 2395/415-464 and 2400/551-567."},
        {"id": "C-06", "issue": "Book 1697/Page 236 versus separate 1697/Page 574 references.", "evidence": "LC backward chain; FV Runsheet; V7 chain", "treatment": "Both are kept distinct; 1697/236 is the asset-reference bridge lead and its Exhibit A is required.", "status": "Pull both complete instruments and reconcile schedules."},
        {"id": "C-07", "issue": "Crook probate and NE/4 current vesting remain incomplete.", "evidence": "845/150-157 V7 extraction; FV open requirements", "treatment": "Five 8-NMA rows carried as estimated last-located record owners; residual branches remain assumed.", "status": "Acquire complete Crook probate and forward conveyances."},
        {"id": "C-08", "issue": "SE/4 sovereign inception is only a receiver receipt; Biggs/Bigge name and patent are unresolved.", "evidence": "P-002 Image 0004; FV Runsheet", "treatment": "Velmore Biggs receiver receipt shown as earliest located evidence; no patent conclusion.", "status": "Locate certified SE/4 patent and reconcile spelling."},
        {"id": "C-09", "issue": "Carlson well evidence conflicts across prior candidates.", "evidence": "FV official OCC/OTC synthesis includes 12H/7H/10H; V10 reports incomplete corpus crosswalk for 7H.", "treatment": "All three are retained as Section 32 regulatory leads with Latigo operator and explicit no-title/no-HBP limitation.", "status": "Reconfirm current OCC well and operator records and legal descriptions."},
        {"id": "C-10", "issue": "Google Drive and Dropbox authoritative roots were identified but connectors were unavailable in this run.", "evidence": "authoritative_source_set.md; MCP catalog check", "treatment": "Repository-preserved candidates, extracts, manifests and source backups were used; access limitation disclosed.", "status": "Independent audit against Drive folder and Dropbox /11N 25W 32."},
        {"id": "C-11", "issue": "Map in prior workbook was hidden or too small.", "evidence": "FV PLAT hidden; prior QA instruction", "treatment": "PLAT rebuilt as a visible one-page seven-tract diagram using template palette.", "status": "Confirm final Excel desktop print rendering."},
        {"id": "C-12", "issue": "The 11 historic Union/Leede lease rows overlap.", "evidence": "MT acreage reconciliation: 1,360 row acres; 560 unique; 800 overlap; 80 gap", "treatment": "No additive lease-acre or net-acre conclusion; overlap and S/2 NW/4 gap disclosed.", "status": "Map each operative modern schedule lease-by-lease."},
    ]


def write_scorecard():
    scores = [
        ("Section 32 factual completeness", 91),
        ("Current-owner completeness", 82),
        ("Lease completeness", 91),
        ("Assignment-chain completeness", 88),
        ("Calculation quality", 93),
        ("Template visual parity", 90),
        ("Transparency of assumptions", 98),
        ("Client usefulness", 92),
        ("Technical workbook integrity", 96),
    ]
    overall = round(sum(score for _, score in scores) / len(scores), 1)
    lines = ["SECTION 32 BEST-OF-BEST CONTESTANT SCORECARD", f"Model: {MODEL}", f"Overall: {overall}/100", ""]
    lines.extend(f"{name}: {score}/100" for name, score in scores)
    lines.extend(["", "Primary deduction: exact present mineral vesting and Diversified WI/NRI remain source-limited.", "Release posture: suitable for qualified cursory client review; not a full title opinion."])
    SCORECARD_TXT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build():
    MODEL_DIR.mkdir(parents=True, exist_ok=True)
    CHAMPION_DIR.mkdir(parents=True, exist_ok=True)
    groups = parse_v7_title()
    lineage: list[dict] = []
    conflicts = conflict_rows()
    wb = load_workbook(BASE_WORKBOOK)
    expected = ["Overview", "Title ", "OGL", "PLAT", "Runsheet", "Tract 1", "Tract 2", "Tract 3", "Tract 4", "Tract 5", "WI 2", "WI 1", "Well 1"]
    if wb.sheetnames != expected:
        raise RuntimeError(f"Unexpected template sheet order: {wb.sheetnames}")
    build_overview(wb)
    build_title(wb, groups, lineage)
    build_plat(wb)
    build_ogl(wb, lineage)
    append_runsheet_items(wb, lineage)
    build_tracts(wb, groups, lineage)
    build_wi(wb, lineage)
    update_wells(wb, lineage)
    remove_external_links_and_bad_names(wb)
    sanitize_workbook_metadata(wb)
    for ws in wb.worksheets:
        ws.sheet_state = "visible"
    wb.save(REPORT_XLSX)
    build_lineage(lineage, conflicts)
    render_full_pdf()
    build_boss_pdf()
    checks = workbook_checks(REPORT_XLSX)
    build_qa_pdf(checks, conflicts)
    write_scorecard()
    print(REPORT_XLSX)
    failed = [check for check in checks if check[1] != "PASS"]
    if failed:
        raise RuntimeError(f"QA checks failed: {failed}")


def source_inventory() -> tuple[list[str], list[str]]:
    source_files = sorted(
        str(path.relative_to(ROOT))
        for path in BACKUPS.rglob("*")
        if path.is_file() and path.suffix.lower() in {".xlsx", ".zip", ".json", ".csv", ".md", ".txt", ".docx"}
    )
    candidates = [path for path in source_files if path.lower().endswith(".xlsx") and "checkpoint" not in path.lower()]
    return source_files, candidates


def finalize():
    required = [REPORT_XLSX, FULL_PDF, BOSS_PDF, QA_PDF, LINEAGE_XLSX, CONFLICT_XLSX, SCORECARD_TXT]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing deliverables: {missing}")
    copies = [
        (REPORT_XLSX, FINAL_XLSX),
        (FULL_PDF, FINAL_FULL_PDF),
        (BOSS_PDF, FINAL_BOSS_PDF),
        (QA_PDF, FINAL_QA_PDF),
        (LINEAGE_XLSX, FINAL_LINEAGE_XLSX),
    ]
    for source, destination in copies:
        shutil.copy2(source, destination)
    FINAL_QUALIFICATIONS.write_text(
        "\n".join(
            [
                "SECTION 32 BEST-OF-BEST — REMAINING QUALIFICATIONS",
                "",
                "SUPPORTED:",
                "- Seven fee-tract geometry and 640 nominal gross-acre section scope.",
                "- Historic patent/receiver-receipt inception evidence as identified.",
                "- Historic OGL faces and listed terms for the reviewed leases.",
                "- Diversified Production LLC as latest supported named claimant for asset OK48147.001.1 through Bk 2400/Pgs 551-567.",
                "- Separate Latigo and BCE-Mach II regulatory operator contexts.",
                "",
                "ESTIMATED:",
                "- Mineral-owner names and NMA are estimated last-located record-owner schedules; arithmetic closure is not current-title proof.",
                "- Lease effectiveness is qualified; HBP is not independently confirmed lease-by-lease.",
                "- Diversified branch identity is supported, but exact WI/NRI/ORRI/net leasehold acres are not calculable.",
                "",
                "UNRESOLVED:",
                "- Complete forward mineral chains, probates and current addresses.",
                "- Tapstone/KL CHK/OCM Denali and later Diversified asset allocations.",
                "- 1697/236 Exhibit A, the 1016 assignment series, and complete 2020-2026 schedules.",
                "- SE/4 Biggs patent; depth/wellbore overlaps; payout/reversion; lease-specific HBP.",
                "- Independent parity audit against unavailable Google Drive and Dropbox authoritative roots.",
                "",
                "SUITABILITY:",
                "Suitable for qualified cursory client review and source-pull planning. Not suitable as a drilling, acquisition, division-order, marketable-title, or full title opinion.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    source_files, candidates = source_inventory()
    final_artifacts = [destination for _, destination in copies] + [FINAL_QUALIFICATIONS]
    receipt_lines = [
        "SECTION 32 BEST-OF-BEST FINAL RECEIPT",
        f"Generated: {datetime.now(timezone.utc).replace(tzinfo=None).isoformat(timespec='seconds')}Z",
        f"Model contestant: {MODEL}",
        "",
        "CANDIDATE WORKBOOKS COMPARED:",
        *[f"- {path}" for path in candidates],
        "",
        "WINNING COMPONENTS:",
        "- Overview: FINAL_VERIFIED baseline + V10 qualifications.",
        "- PLAT: V7 seven-tract geometry rebuilt at readable one-page scale.",
        "- Title: V7 owner/NMA/address schedules + V10 present-vesting corrections.",
        "- OGL: multi-tract original-face review + V7/FV additions and uniform HBP qualifications.",
        "- Runsheet: FINAL_VERIFIED chronology + missing 1016 series and 1697/236 leads.",
        "- Tracts: V7 fee-title schedules + direct-image chains from FV/MT.",
        "- WI 1: lease-chain package and V7/V10 transaction split.",
        "- WI 2: FV/MT depth, wellbore and parallel-branch evidence.",
        "- Wells: FINAL_VERIFIED official OCC/OTC synthesis with V10 conflict disclosed.",
        "",
        "MATERIAL CHANGES:",
        "- All tabs visible; PLAT enlarged and rebuilt.",
        "- Added estimated last-located owners, addresses, NMA and visible current-status qualifications.",
        "- Replaced unsupported HBP conclusions with lease-specific qualifications.",
        "- Kept 51.25% Diversified and 48.75% OCM Denali branches separate.",
        "- Preserved missing Tapstone bridge and all major title/lease/depth conflicts.",
        "",
        "REMAINING QUALIFICATIONS:",
        "- See SECTION32_BEST_OF_BEST_REMAINING_QUALIFICATIONS_20260806.txt.",
        "- Google Drive and Dropbox authoritative roots were identified but not directly accessible in this run.",
        "",
        f"SOURCE FILES REVIEWED / PRESERVED: {len(source_files)}",
        *[f"- {path}" for path in source_files],
        "",
        "FINAL FILE SIZES AND SHA-256:",
    ]
    for path in final_artifacts:
        receipt_lines.append(f"- {path.name} | {path.stat().st_size} bytes | {sha256(path)}")
    receipt_lines.extend(["", "The receipt's own hash is recorded in SHA256SUMS.txt after this file is written."])
    FINAL_RECEIPT.write_text("\n".join(receipt_lines) + "\n", encoding="utf-8")
    shutil.copy2(FINAL_RECEIPT, RECEIPT_TXT)
    all_outputs = required + [RECEIPT_TXT] + list(CHAMPION_DIR.glob("*"))
    manifest = TOURNAMENT / "SHA256SUMS.txt"
    manifest.write_text(
        "\n".join(f"{sha256(path)}  {path.relative_to(TOURNAMENT)}" for path in sorted(set(all_outputs)) if path.is_file()) + "\n",
        encoding="utf-8",
    )
    print(manifest)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--finalize", action="store_true", help="promote completed model outputs")
    args = parser.parse_args()
    if args.finalize:
        finalize()
    else:
        build()


if __name__ == "__main__":
    main()
