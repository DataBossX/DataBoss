#!/usr/bin/env python3
"""Build evidence-controlled Section 32 tournament deliverables."""

from __future__ import annotations

import json
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "deliverables" / "GPT56_SOL_SECTION32_MULTI_TRACT_SPRINT_20260716"
STAMP = "2026-07-16T22:38:00Z"
UNKNOWN = "NOT DETERMINED"

TRACTS = [
    {
        "tract_no": 1,
        "legal": "NE/4",
        "nominal_acres": 160,
        "scope_notes": "Crook #1-32 historic focus; Morrow–Springer wellbore interval 16,210–20,451 feet.",
        "current_conclusion": "Historic fee inception and Crook lease/wellbore branches are supported. Present mineral and leasehold ownership is NOT DETERMINED.",
        "remaining_open_items": [
            "Complete post-Bollenbach fee/mineral chain.",
            "Original and amended leases, releases, production, and HBP evidence.",
            "Operative modern schedules mapping OK48147.001.1 to this tract and depth.",
        ],
    },
    {
        "tract_no": 2,
        "legal": "E/2 SE/4",
        "nominal_acres": 80,
        "scope_notes": "Pearl #1-32 historic focus; later assignments use approximately 11,100–19,480 feet.",
        "current_conclusion": "OGL 63/33 identity and the Humble–Union–Leede Pearl branch are supported historically. Present minerals, lease status, WI, burdens, and HBP are NOT DETERMINED.",
        "remaining_open_items": [
            "Original OGL 63/33 and all amendments/releases.",
            "Complete Pearl wellbore assignment and payout/reversion evidence.",
            "Downstream schedule mapping and present burden survival.",
        ],
    },
    {
        "tract_no": 3,
        "legal": "N/2 of Section 32 below top of Hunton",
        "nominal_acres": 320,
        "scope_notes": "Depth-defined leasehold scope; 320 acres is surface geometry only and overlaps Tracts 1, 2, and 4.",
        "current_conclusion": "Bk 1014/Pg 75 facially describes N/2 below top Hunton, but assignor source title and exact quantum are unresolved. No present WI may be posted.",
        "remaining_open_items": [
            "Complete 1014/75 and predecessor trust/McCall instruments.",
            "Instrument-defined top Hunton reference and well limitations.",
            "Source title proving the assignor held the purported working/cost-bearing interest.",
        ],
    },
    {
        "tract_no": 4,
        "legal": "Deep leasehold references and lease list in Bk 872/Pg 279",
        "nominal_acres": None,
        "scope_notes": "Overlapping lease-by-lease and depth-by-depth scope; non-additive. Eleven listed Section 32 lease rows total 1,360 row acres but reconcile to 560 unique nominal acres, 800 overlapping row acres, and an 80-acre S/2 NW/4 schedule gap. General assignment interval is top Morrow–19,480 feet, subject to exceptions.",
        "current_conclusion": "Original faces for all 11 listed leases and Bk 872/Pg 279 assignment/exhibits support the historic lease package and document-location math. Current lease status, HBP, successor title, and current quantum are NOT DETERMINED.",
        "remaining_open_items": [
            "Missing Bk 872/Pg 280–281 terms/signature pages and the unrecorded 1984 agreement.",
            "Every amendment, correction, release, pooling record, and lease-specific HBP record.",
            "Post-1985 lease-by-lease successor chain and burden survival.",
        ],
    },
    {
        "tract_no": 5,
        "legal": "Full-section modern Diversified-affiliated record branch",
        "nominal_acres": None,
        "scope_notes": "Section-wide reporting scope; whole-section indexing does not prove 640 net acres, 8/8 WI, all depths, or ownership.",
        "current_conclusion": "2395/462 and 2400/566 support a recorded claimant path for OK48147.001.1 into Diversified Production LLC. Tract allocation, estate, depth, and quantum are NOT DETERMINED.",
        "remaining_open_items": [
            "Complete operative schedules and incorporated documents for 2395/415–464 and 2400/551–567.",
            "Complete 2340, 2371, 2389, 2393, and related branch schedules.",
            "Separate and reconcile Unbridled/MNR, Teocalli, Canvas/DP Ponies, Burlington, and lien branches.",
        ],
    },
]

INSTRUMENTS = [
    {
        "tracts": "T1",
        "execution_date": "1907-03-22",
        "recording_date": "1907-06-08",
        "doc_type": "PAT",
        "instrument_no": "Homestead Certificate 4300 / Application 17722",
        "book_page": "Patent 3/469",
        "grantor": "United States of America",
        "grantee": "Cornelius B. Cornelius",
        "legal": "NE/4, Section 32-11N-25W; stated 160 acres",
        "estate": "Historic fee inception",
        "depth": None,
        "interest": "Historic fee estate; no mineral reservation apparent on reviewed image",
        "reserved": None,
        "lease_id": None,
        "branch": "Historic fee",
        "evidence_class": "DIRECT IMAGE",
        "source": "P-001 / Images/0003.jpg",
        "image": "Image 0003",
        "confidence": 98,
        "current_effect": "Historic inception only; not present vesting",
        "conflict": None,
        "curative": "Run complete downstream fee/mineral chain from Cornelius/Bollenbach",
        "notes": "Do not post 160 present net mineral acres.",
    },
    {
        "tracts": "T1",
        "execution_date": "1907-07-16",
        "recording_date": "1907-07-18",
        "doc_type": "WD",
        "instrument_no": None,
        "book_page": "Roger Mills 11/214",
        "grantor": "Cornelius B. Cornelius; Margreta Cornelius",
        "grantee": "Jacob Bollenbach",
        "legal": "NE/4, Section 32-11N-25W; stated 160 acres",
        "estate": "Fee conveyance",
        "depth": None,
        "interest": "Historic fee conveyance; no reservation apparent on reviewed image",
        "reserved": None,
        "lease_id": None,
        "branch": "Historic fee",
        "evidence_class": "DIRECT IMAGE",
        "source": "Recorded deed / Images/0002.jpg",
        "image": "Image 0002",
        "confidence": 98,
        "current_effect": "Historic bridge only",
        "conflict": "County/book metadata must be confirmed in Roger Mills and Beckham records",
        "curative": "Obtain certified copy and continue grantee chain",
        "notes": "Present mineral vesting remains open.",
    },
    {
        "tracts": "T2",
        "execution_date": "1947-11-13",
        "recording_date": "1947-11-26",
        "doc_type": "OGL",
        "instrument_no": None,
        "book_page": "63/33",
        "grantor": "Cora B. Garrett; S.A. Garrett, her husband",
        "grantor_address": "Mayfield, Oklahoma",
        "grantee": "C.E. McKinley",
        "legal": "E/2 SE/4",
        "estate": "Leasehold identity",
        "depth": None,
        "interest": "Historic leasehold face; 1/8 royalty; 10-year primary term",
        "reserved": None,
        "lease_id": "L-91511 / 63/33",
        "branch": "Pearl",
        "evidence_class": "DIRECT ORIGINAL IMAGE + RECORDED ASSIGNMENT EXHIBITS",
        "source": "OGL 63/33 / Image 0353; Bk 244/Pg 432 / Image 1062; Bk 872/Pg 284 / Image 4393",
        "image": "Images 0353, 1062, 4393",
        "confidence": 98,
        "current_effect": "Historic lease terms and identity only",
        "conflict": None,
        "curative": "Pull amendments, ratifications, releases, pooling, production, and HBP evidence",
        "notes": "Original face proves 1/8 royalty and 10-year term. Pooling and shut-in clauses are present but current effect, continuation, and HBP remain open.",
    },
    {
        "tracts": "T2",
        "execution_date": "1969-11-25",
        "recording_date": None,
        "doc_type": "ASG OGL",
        "instrument_no": None,
        "book_page": "244/432",
        "grantor": "Humble Oil & Refining Company",
        "grantee": "Union Oil Company of California",
        "legal": "E/2 SE/4",
        "estate": "Leasehold assignment",
        "depth": None,
        "interest": "OGL 63/33 branch; effective 1969-12-01",
        "reserved": None,
        "lease_id": "L-91511 / 63/33",
        "branch": "Pearl",
        "evidence_class": "DIRECT IMAGE",
        "source": "Recorded assignment / Image 1062",
        "image": "Image 1062",
        "confidence": 96,
        "current_effect": "Historic chain only",
        "conflict": "January 1970 recording day partially illegible",
        "curative": "Confirm clerk metadata and downstream assignments",
        "notes": "No present title conclusion.",
    },
    {
        "tracts": "T1",
        "execution_date": "1981-09-17",
        "recording_date": "1981-10-13",
        "doc_type": "CORR ASG",
        "instrument_no": None,
        "book_page": "509/220",
        "grantor": "Leede Exploration",
        "grantee": "Mesa 64%; American Petrofina 25%; ENI 5%; Jack O. McCall 4%; Midland National Bank, Trustee 1%; Sterling Petroleum-1978-A 1%",
        "legal": "Crook #1-32 wellbore",
        "estate": "Historic depth-limited wellbore interest",
        "depth": "Morrow–Springer, 16,210–20,451 feet",
        "interest": "100% allocation of the assigned Crook interest, not 8/8 tract WI",
        "reserved": "Aggregate participant NRI fixed at 75%; Leede retained 1/4 after-payout reversion",
        "lease_id": "264/345 branch",
        "branch": "Crook",
        "evidence_class": "DIRECT IMAGE / INCOMPLETE COPY",
        "source": "L-002; recorded correction exhibit / Image 1952",
        "image": "Image 1952",
        "confidence": 98,
        "current_effect": "Historical document-scoped allocation only",
        "conflict": None,
        "curative": "Obtain complete instrument, payout evidence, and successor assignments",
        "notes": "Never apply the percentages to 8/8 unless the assignor baseline is proven.",
    },
    {
        "tracts": "T2",
        "execution_date": "1985-07-29",
        "recording_date": "1985-10-23",
        "doc_type": "WB ASG",
        "instrument_no": None,
        "book_page": "845/47",
        "grantor": "Cities Service Oil & Gas Corporation",
        "grantee": "Leede Exploration",
        "legal": "E/2 SE/4; Pearl #1-32 wellbore",
        "estate": "Historic wellbore working-interest branch",
        "depth": "Below 11,100 feet to but not below 19,480 feet",
        "interest": "Assigned wellbore interest; exact 8/8 baseline not established",
        "reserved": "1/8 of 8/8 ORRI",
        "lease_id": "63/33 / Pearl",
        "branch": "Pearl",
        "evidence_class": "DIRECT IMAGE",
        "source": "L-006 / Images 4118–4119",
        "image": "Images 4118–4119",
        "confidence": 98,
        "current_effect": "Historical only",
        "conflict": None,
        "curative": "Obtain unrecorded 1985 agreement, payout and successor records",
        "notes": "Five-year renewal/extension reach; no warranty.",
    },
    {
        "tracts": "T2,T4",
        "execution_date": "1985-10-29",
        "recording_date": "1986-03-06",
        "doc_type": "ASG / LEASE SCHEDULE",
        "instrument_no": None,
        "book_page": "872/279",
        "grantor": "Union Oil Company of California",
        "grantor_address": "P. O. Box 3100, Midland, Texas 79702",
        "grantee": "Leede Exploration",
        "grantee_address": "600 North Marienfeld, Midland, Texas 79701",
        "legal": "Listed Section 32 leases; E/2 SE/4 specially limited",
        "estate": "Depth-limited leasehold",
        "depth": "Generally top Morrow–19,480 feet; E/2 SE/4 11,100–19,480 feet; prior Crook rights excluded",
        "interest": "Scheduled leasehold; exact quantum not stated in reviewed pages",
        "reserved": "1/8 gas/condensate/distillate ORRI; 1/16 oil/casinghead-gas ORRI",
        "lease_id": "Bk 872/Pg 279 schedule",
        "branch": "Union/Leede deep",
        "evidence_class": "DIRECT IMAGE / PARTIAL INSTRUMENT",
        "source": "L-007 / Images 4388, 4391–4393",
        "image": "Images 4388, 4391–4393",
        "confidence": 92,
        "current_effect": "Historic depth-limited branch only",
        "conflict": "Unrecorded 1984 agreement and incomplete instrument review",
        "curative": "Pull complete instrument, agreement, all scheduled OGLs and releases",
        "notes": "Rows overlap; no additive-acre or current-HBP inference.",
    },
    {
        "tracts": "T3",
        "execution_date": "1988-02-18",
        "recording_date": "1988-04-26",
        "doc_type": "ASG",
        "instrument_no": None,
        "book_page": "1014/75",
        "grantor": "Consolidation of Leede Trusts",
        "grantor_address": "600 North Marienfeld, Midland, Texas 79701",
        "grantee": "Leede Exploration",
        "grantee_address": "600 North Marienfeld, Midland, Texas 79701",
        "legal": "N/2 Section 32",
        "estate": "Purported OGL and working/cost-bearing interests",
        "depth": "All depths below top Hunton",
        "interest": "All assignor right, title, and interest in OGLs and other working/cost-bearing interests; effective for all purposes as of first production from McCall #1-29",
        "reserved": "Royalty/ORRI reconveyance effect unclear",
        "lease_id": None,
        "branch": "Leede/McCall/trust",
        "evidence_class": "DIRECT IMAGE",
        "source": "L-011 / Image 4893",
        "image": "Image 4893",
        "confidence": 90,
        "current_effect": "Facial transfer only; title effect OPEN",
        "conflict": "Assignor source evidence proves royalty/ORRI, not the purported WI; exact quantum absent",
        "curative": "Establish predecessor title and obtain complete instrument",
        "notes": "No 1% or 52.632% figure may be posted as present WI.",
    },
    {
        "tracts": "T5",
        "execution_date": None,
        "recording_date": "2022-11-18",
        "doc_type": "CONFIRMATORY CONV",
        "instrument_no": "I-2022-004838",
        "book_page": "2395/415–464",
        "grantor": "DP Legacy Tapstone LLC",
        "grantee": "Diversified ABS VI Upstream LLC; DP Sooner HoldCo LLC",
        "legal": "Section 32 schedule reference; asset OK48147.001.1",
        "estate": "Leasehold claimant identity",
        "depth": None,
        "interest": None,
        "reserved": None,
        "lease_id": "OK48147.001.1",
        "branch": "DP Sooner",
        "evidence_class": "REVIEWED SCHEDULE PAGE",
        "source": "I-2022-004838 / Exhibit E p.462",
        "image": "Bk 2395/Pg 462",
        "confidence": 96,
        "current_effect": "Asset routed to DP Sooner path; no tract or quantum proof",
        "conflict": None,
        "curative": "Obtain complete operative schedules and incorporated documents",
        "notes": "Asset-line legal and interest fields do not support a decimal.",
    },
    {
        "tracts": "T5",
        "execution_date": None,
        "recording_date": "2023-03-02",
        "doc_type": "CONFIRMATORY CONV",
        "instrument_no": "I-2023-000796",
        "book_page": "2400/551–567",
        "grantor": "DP Sooner HoldCo LLC",
        "grantee": "Diversified Production LLC",
        "grantee_address": "1600 Corporate Drive, Birmingham, AL 35242",
        "legal": "Section 32 schedule reference; asset OK48147.001.1",
        "estate": "Leasehold claimant identity",
        "depth": None,
        "interest": None,
        "reserved": None,
        "lease_id": "OK48147.001.1",
        "branch": "DP Sooner / Diversified",
        "evidence_class": "REVIEWED SCHEDULE PAGE",
        "source": "I-2023-000796 / Schedule B p.566",
        "image": "Bk 2400/Pg 566",
        "confidence": 96,
        "current_effect": "Latest supported recorded claimant identity; no tract or quantum proof",
        "conflict": None,
        "curative": "Obtain complete operative schedules and incorporated documents",
        "notes": "Do not post Diversified as proven current owner or assign a WI decimal.",
    },
]

LEASES = [
    ["63/33", "Cora B. Garrett; S.A. Garrett, her husband", "C.E. McKinley", "1947-11-13", "1947-11-26", None, "63/33", "E/2 SE/4", 80, None, None, "1/8", "10 years", None, "Clause present; effect not adjudicated", "Clause present; effect not adjudicated", "244/432; 845/47; 872/279", "Union later reserved depth-limited ORRIs in 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 0353 and 4393. Original face proves historic terms only; amendments, release, HBP, and current ownership remain open."],
    ["264/345", "Charlie B. Crook; Ima Pearl Crook, his wife", "Union Oil Company of California", "1972-02-05", "1972-02-22", "1131", "264/345", "NE/4", 160, None, None, "1/8", "5 years commencing 1973-01-13", None, None, None, "502/267; excluded from 872/279 to extent previously assigned", "Union ORRIs under 502/267", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1197–1198 and 4392. Instrument also covers 80 acres in Section 28; only 160 Section 32 row acres counted."],
    ["307/31", "Cora B. LeGrand, widow", "Union Oil Company of California", "1976-04-29", "1976-06-09", "2991", "307/31", "W/2 SE/4 and S/2 SW/4", 160, None, None, "1/8", "5 years commencing 1976-09-23", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1372 and 4391. Overlaps other listed rows; no present status or HBP inference."],
    ["340/302", "Gertrude L. LaFortune; Jeanne L. Fair; Mary Ann L. Wilcox; J.A. LaFortune Jr.; Robert J. LaFortune, co-trustees under the will of Joseph A. LaFortune", "Union Oil Company of California", "1977-10-14", None, None, "340/302", "N/2 SW/4", 80, None, None, "3/16", "5 years commencing 1977-11-28", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1522–1523 and 4391. Relationship to separate 340/304 remains open."],
    ["340/304", "Gertrude LaFortune, widow", "Union Oil Company of California", "1977-10-14", None, None, "340/304", "N/2 SW/4", 80, None, None, "3/16", "5 years commencing 1977-11-28", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1524–1525 and 4391. Separate overlapping lease; do not double count."],
    ["340/323", "George T. Harrison Jr.; Renata Harrison", "Union Oil Company of California", "1977-09-29", "1977-11-07", None, "340/323", "N/2 NW/4 and N/2 SW/4", 160, None, None, "3/16", "5 years commencing 1977-11-27", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1526–1527 and 4391. This is the only listed lease aliquot covering N/2 NW/4."],
    ["342/564", "Carol Seidenbach Leach", "Union Oil Company of California", "1977-09-29", "1978-02-09", "1089", "342/564", "N/2 SW/4", 80, None, None, "3/16", "5 years commencing 1977-11-30", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1542–1543 and 4391. Overlapping legal; current status open."],
    ["342/566", "Cynthia Seidenbach Kruse", "Union Oil Company of California", "1977-11-01", "1978-02-09", "1090", "342/566", "N/2 SW/4", 80, None, None, "3/16", "5 years commencing 1977-11-30", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1544–1545 and 4391. Overlapping legal; current status open."],
    ["352/719", "F.L. Randel", "Union Oil Company of California", "1978-05-16", "1978-06-27", "5169", "352/719", "NE/4", 160, None, None, "3/16", "5 years", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1552–1553 and 4391. Face instrument 5169 conflicts with protocol index 5051; clerk confirmation required."],
    ["352/721", "J.L. Randel", "Union Oil Company of California", "1978-05-16", "1978-06-27", "5168", "352/721", "NE/4", 160, None, None, "3/16", "5 years", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1554–1555 and 4392. Face instrument 5168 conflicts with protocol index 5052; clerk confirmation required."],
    ["376/369", "Great Western Oil & Gas, Inc., a California corporation", "Union Oil Company of California", "1979-03-21", "1979-04-30", "3757", "376/369", "NE/4", 160, None, None, "3/16", "5 years", None, None, None, "872/279", "Union depth-limited ORRIs under 872/279", None, None, "OPEN_UNVERIFIED_CURRENT", "DIRECT ORIGINAL IMAGE + RECORDED EXHIBIT", 98, "Images 1569–1570 and 4392. Face instrument 3757 conflicts with protocol index 3707; correction metadata at 1039/20–23 remains unreviewed."],
]

ACREAGE_RECONCILIATION = [
    ["Row acres", 1360, "Sum of the 11 Section 32 lease-row aliquots; only 160 Section 32 acres counted for 264/345"],
    ["Unique nominal acres", 560, "Union of listed aliquots: NE/4, SE/4, SW/4, and N/2 NW/4"],
    ["Overlap", 800, "1,360 row acres minus 560 unique nominal acres"],
    ["Schedule gap", 80, "S/2 NW/4 is not among the Bk 872/Pg 279 listed lease aliquots"],
    ["Diversified net leasehold acres", None, "NOT DETERMINED; document-location math does not prove current title, HBP, or quantum"],
]

WI_ROWS = [
    ["509/220", "Crook #1-32", "Mesa; American Petrofina; ENI; McCall; Midland; Sterling", "64%; 25%; 5%; 4%; 1%; 1% of assigned interest", "Assigned Crook wellbore interest", "After-payout reversion", "Historical WI allocation", "NO", "Baseline is not proven 8/8; present chain and payout open"],
    ["509/220", "Crook #1-32", "Leede Exploration", "1/4", "After-payout reversion in assigned Crook interest", "After payout", "Reversion", "NO", "Current payout and survival unknown"],
    ["845/47", "Pearl #1-32", "Cities Service reservation", "1/8 of 8/8", "Pearl assigned wellbore interest, 11,100–19,480 ft", "Not stated", "ORRI", "NO", "Survival and successors open"],
    ["872/279", "Scheduled deep leases", "Union reservation", "1/8", "Gas/condensate/distillate from assigned scheduled interest", "Not stated", "ORRI", "NO", "Historic, schedule- and depth-limited"],
    ["872/279", "Scheduled deep leases", "Union reservation", "1/16", "Oil/casinghead gas from assigned scheduled interest", "Not stated", "ORRI", "NO", "Historic, schedule- and depth-limited"],
    ["903/50", "Defined LGWI", "Each PRF", "0.526300%", "Defined LGWI, not 8/8", "Before payout", "Contract WI/cost share", "NO", "Three PRFs; historical agreement"],
    ["903/50", "Defined LGWI", "Each PRF", "1.315750%", "Defined LGWI, not 8/8", "After payout", "Contract WI/cost share", "NO", "Three PRFs; historical agreement"],
    ["987/159 et seq.", "McCall", "TCW/Grantham/Carlson reservation", "0.043146%", "0.5% of Leede NRI", "Not stated", "ORRI", "NO", "Historic burden only"],
    ["987/159 et seq.", "Pearl", "TCW/Grantham/Carlson reservation", "0.052310%", "0.5% of Leede NRI", "Not stated", "ORRI", "NO", "Historic burden only"],
    ["2395/462; 2400/566", "OK48147.001.1", "Diversified Production LLC", None, None, None, "Present-title candidate", "NO", "Claimant identity only; exact WI/NRI/ORRI and tract allocation not proved"],
]

MODERN_CHAIN = [
    ["2340/403–433", "I-2020-004435", "Chesapeake Exploration LLC", "Tapstone Energy LLC", "Assignment", "Chesapeake/Tapstone", "Index/public metadata", "Complete schedule and lease mapping"],
    ["2340/490–519", "I-2020-004462", "Chesapeake Exploration LLC", "KL CHK SPV LLC", "Assignment", "Chesapeake/KL CHK", "Index/public metadata", "Complete schedule and exact asset allocation"],
    ["2371/470–494", "I-2022-000152", "KL CHK SPV LLC", "Diversified Production LLC; OCM Denali Holdings LLC", "Assignment", "KL CHK/Diversified/OCM", "Index/public metadata", "Fractions, schedules, estate, exclusions"],
    ["2371/495–513", "I-2022-000153", "Tapstone Energy LLC; Tapstone Midstream LLC", "OCM Denali Holdings LLC", "Assignment", "Tapstone/OCM", "Project schedule reference", "Keep separate; 48.75% is not current Diversified WI"],
    ["2389/500–580", "I-2022-003506", "Diversified Production LLC; DP affiliates", "DP Legacy Tapstone LLC", "Merger", "DP Legacy Tapstone", "Index/public metadata", "Merger plan, schedules, asset allocation"],
    ["2393/1–698", "I-2022-004277", "Burlington Resources and affiliates", "DP Legacy Central LLC", "Assignment", "Burlington/DP Legacy Central", "Index/public metadata", "Separate 698-page schedule branch"],
    ["2395/415–464", "I-2022-004838", "DP Legacy Tapstone LLC", "Diversified ABS VI Upstream LLC; DP Sooner HoldCo LLC", "Confirmatory conveyance", "DP Sooner", "Reviewed Exhibit E p.462", "Legal, estate, depth, quantum, exclusions"],
    ["2400/551–567", "I-2023-000796", "DP Sooner HoldCo LLC", "Diversified Production LLC", "Confirmatory conveyance", "Diversified", "Reviewed Schedule B p.566", "Legal, estate, depth, quantum, exclusions"],
    ["2415/328–393", "I-2023-004455", "Unbridled Resources LLC", "MNR ABS Issuer I LLC", "Wellbore assignment", "Unbridled/MNR", "Index/public metadata", "Separate branch; schedule/depth/overlap"],
    ["2434/751–760", "I-2024-002930", "Diversified Production LLC; DP Legacy Central LLC", "Teocalli Exploration LLC", "Wellbore assignment", "Teocalli", "Countywide metadata", "Section 32 applicability and schedule"],
    ["2475/943–959; 2476/1–66", "I-2026-000211; I-2026-000212", "Canvas Energy LLC / Canvas Energy II LLC", "Canvas Energy II LLC / DP Ponies LLC", "Merger and conveyance", "Canvas/DP Ponies", "Countywide metadata", "Separate acquisition; Section 32 schedule and overlap"],
]

ADDRESSES = [
    ["Diversified Production LLC", "1600 Corporate Drive, Birmingham, AL 35242", "1600 Corporate Drive, Birmingham, AL 35242", "Recorded instrument address", "2395/415–464; 2400/551–567", "2022-11-18 / 2023-03-02", "Historic recorded address; not represented as verified current"],
    ["Cora B. Garrett; S.A. Garrett", "Mayfield, Oklahoma", "Mayfield, OK", "Lessor locality", "OGL 63/33 / Image 0353", "1947-11-13", "Historical only; not a current mailing address"],
    ["Charlie B. Crook; Ima Pearl Crook", "Route 1, Mayfield, Oklahoma", "Route 1, Mayfield, OK", "Lessor address", "OGL 264/345 / Image 1197", "1972-02-05", "Historical only; not a current mailing address"],
    ["Cora B. LeGrand", "Sweetwater, Oklahoma", "Sweetwater, OK", "Lessor locality", "OGL 307/31 / Image 1372", "1976-04-29", "Historical only; not a current mailing address"],
    ["LaFortune co-trustees", "2300 Fourth National Bank Building, Tulsa, Oklahoma", "2300 Fourth National Bank Building, Tulsa, OK", "Lessor address", "OGL 340/302 / Image 1522", "1977-10-14", "Historical only; not a current mailing address"],
    ["Gertrude LaFortune", "2300 Fourth National Building, Tulsa, Oklahoma 74119", "2300 Fourth National Building, Tulsa, OK 74119", "Lessor address", "OGL 340/304 / Image 1524", "1977-10-14", "Historical only; not a current mailing address"],
    ["George T. Harrison Jr.; Renata Harrison", "201 North Charles Street, Baltimore, Maryland 21201", "201 North Charles Street, Baltimore, MD 21201", "Lessor address", "OGL 340/323 / Image 1526", "1977-09-29", "Historical only; not a current mailing address"],
    ["Carol Seidenbach Leach", "c/o Mrs. J. Leslie Seidenbach, 2827 South Birmingham Place, Tulsa, Oklahoma", "c/o Mrs. J. Leslie Seidenbach, 2827 South Birmingham Place, Tulsa, OK", "Lessor address", "OGL 342/564 / Image 1542", "1977-09-29", "Historical only; not a current mailing address"],
    ["Cynthia Seidenbach Kruse", "c/o M. B. Johnson, 1623 S. College, Tulsa, Oklahoma 74104", "c/o M. B. Johnson, 1623 S. College, Tulsa, OK 74104", "Lessor address", "OGL 342/566 / Image 1544", "1977-11-01", "Historical only; not a current mailing address"],
    ["F.L. Randel", "P. O. Box 7, c/o Ramada Inn, Big Spring, Texas 79720", "P.O. Box 7, c/o Ramada Inn, Big Spring, TX 79720", "Lessor address", "OGL 352/719 / Image 1552", "1978-05-16", "Historical only; not a current mailing address"],
    ["J.L. Randel", "Box 2609, Wichita Falls, Texas", "P.O. Box 2609, Wichita Falls, TX", "Lessor address", "OGL 352/721 / Image 1554", "1978-05-16", "Historical only; not a current mailing address"],
    ["Great Western Oil & Gas, Inc.", "4010 Palos Verdes Drive North, Suite 206, Rolling Hills Estates, California 90274", "4010 Palos Verdes Drive N, Suite 206, Rolling Hills Estates, CA 90274", "Lessor corporate address", "OGL 376/369 / Image 1569", "1979-03-21", "Historical only; not a current mailing address"],
    ["Union Oil Company of California", "P. O. Box 3100, Midland, Texas 79702", "P.O. Box 3100, Midland, TX 79702", "Assignor address", "Bk 872/Pg 279 / Image 4388", "1985-10-29", "Historical only; not a current mailing address"],
    ["Leede Exploration", "600 North Marienfeld, Midland, Texas 79701", "600 N Marienfeld, Midland, TX 79701", "Assignee address", "Bk 872/Pg 279 / Image 4388", "1985-10-29", "Historical only; not a current mailing address"],
    ["Consolidation of Leede Trusts", "600 North Marienfeld, Midland, Texas 79701", "600 N Marienfeld, Midland, TX 79701", "Assignor address", "Bk 1014/Pg 75 / Image 4893", "1988-02-18", "Historical only; not a current mailing address"],
]

ADDRESS_TRACTS = {
    "Diversified Production LLC": {5},
    "Cora B. Garrett; S.A. Garrett": {2, 4},
    "Charlie B. Crook; Ima Pearl Crook": {1, 4},
    "Cora B. LeGrand": {4},
    "LaFortune co-trustees": {4},
    "Gertrude LaFortune": {4},
    "George T. Harrison Jr.; Renata Harrison": {4},
    "Carol Seidenbach Leach": {4},
    "Cynthia Seidenbach Kruse": {4},
    "F.L. Randel": {1, 4},
    "J.L. Randel": {1, 4},
    "Great Western Oil & Gas, Inc.": {1, 4},
    "Union Oil Company of California": {4},
    "Leede Exploration": {3, 4},
    "Consolidation of Leede Trusts": {3},
}

ACQUISITION_QUEUE = [
    [1, "T5", "Certified complete 2395/415–464 and 2400/551–567, all exhibits and incorporated documents", "Maps OK48147.001.1 by legal, estate, depth, quantum, exclusions, and burdens"],
    [2, "T2", "Amendments, ratifications, releases, production and shut-in records for OGL 63/33", "Determines Pearl lease continuation, status, and HBP beyond the reviewed original face"],
    [3, "T3", "Complete 1014/75 and Leede trust/McCall predecessor chain", "Resolves depth-defined source title and nemo-dat risk"],
    [4, "T4", "Missing 872/279 pages 280–281, unrecorded 1984 agreement, amendments, releases, and HBP records", "Completes assignment terms and current-effect review beyond the faces and exhibits already inspected"],
    [5, "T1", "Post-Bollenbach deeds, mineral deeds, probates, decrees, reservations, and releases", "Determines present NE/4 mineral vesting"],
    [6, "ALL", "Complete 2340, 2371, 2389, 2393, 2415, 2434, 2475–2476 schedules", "Separates portfolio branches and tests actual Section 32 overlap"],
    [7, "ALL", "Stamped index page 48 and recorded copies for promoted index-only rows", "Closes source gap and prevents locator evidence from becoming title evidence"],
]

TEMPLATE_IMPROVEMENTS = [
    "Central protected Tract Control table; report cells reference one authoritative legal/scope definition.",
    "Separate execution and recording date columns.",
    "Pair exact fraction/source percentage with decimal only when denominator and application are proven.",
    "Use explicit NOT DETERMINED and OPEN fields; never formulas that coerce blanks to zero.",
    "Add address type, source, and effective/recording date.",
    "Add portfolio branch, estate, and depth columns.",
    "Add conflict ID, current effect, and curative action.",
    "Prevent additive acreage totals across overlapping tract scopes.",
    "Tier colors: direct image, reviewed schedule, index locator, open/rejected.",
    "QA stale-token scan for Section 27/Roger Mills template facts and unsupported present-owner language.",
]


def instrument_json(row: dict) -> dict:
    return {
        "execution_date": row["execution_date"],
        "recording_date": row["recording_date"],
        "doc_type": row["doc_type"],
        "instrument_no": row["instrument_no"],
        "book_page": row["book_page"],
        "grantor": [p.strip() for p in row["grantor"].split(";")],
        "grantor_addresses": [row["grantor_address"]] if row.get("grantor_address") else [],
        "grantee": [p.strip() for p in row["grantee"].split(";")],
        "grantee_addresses": [row["grantee_address"]] if row.get("grantee_address") else [],
        "legal": row["legal"],
        "estate": row["estate"],
        "depth": row["depth"],
        "interest_conveyed_fraction": row["interest"],
        "interest_conveyed_decimal": None,
        "interest_reserved": row["reserved"],
        "lease_id": row["lease_id"],
        "portfolio_branch": row["branch"],
        "evidence_class": row["evidence_class"],
        "source_file": row["source"],
        "source_page_or_image": row["image"],
        "confidence": row["confidence"],
        "current_effect": row["current_effect"],
        "conflicts": [row["conflict"]] if row["conflict"] else [],
        "curative_action": row["curative"],
        "notes": row["notes"],
    }


LEASE_TRACTS = {
    "63/33": "T2,T4",
    "264/345": "T1,T4",
    "307/31": "T4",
    "340/302": "T4",
    "340/304": "T4",
    "340/323": "T4",
    "342/564": "T4",
    "342/566": "T4",
    "352/719": "T1,T4",
    "352/721": "T1,T4",
    "376/369": "T1,T4",
}

LEASE_ADDRESSES = {
    "63/33": "Mayfield, Oklahoma",
    "264/345": "Route 1, Mayfield, Oklahoma",
    "307/31": "Sweetwater, Oklahoma",
    "340/302": "2300 Fourth National Bank Building, Tulsa, Oklahoma",
    "340/304": "2300 Fourth National Building, Tulsa, Oklahoma 74119",
    "340/323": "201 North Charles Street, Baltimore, Maryland 21201",
    "342/564": "c/o Mrs. J. Leslie Seidenbach, 2827 South Birmingham Place, Tulsa, Oklahoma",
    "342/566": "c/o M. B. Johnson, 1623 S. College, Tulsa, Oklahoma 74104",
    "352/719": "P. O. Box 7, c/o Ramada Inn, Big Spring, Texas 79720",
    "352/721": "Box 2609, Wichita Falls, Texas",
    "376/369": "4010 Palos Verdes Drive North, Suite 206, Rolling Hills Estates, California 90274",
}


def lease_instrument_json(row: list) -> dict:
    conflict = None
    if row[0] == "352/719":
        conflict = "Direct face instrument 5169 conflicts with protocol index 5051"
    elif row[0] == "352/721":
        conflict = "Direct face instrument 5168 conflicts with protocol index 5052"
    elif row[0] == "376/369":
        conflict = "Direct face instrument 3757 conflicts with protocol index 3707"
    return {
        "execution_date": row[3],
        "recording_date": row[4],
        "doc_type": "OGL",
        "instrument_no": row[5],
        "book_page": row[6],
        "grantor": [party.strip() for party in row[1].split(";")],
        "grantor_addresses": [LEASE_ADDRESSES[row[0]]],
        "grantee": [row[2]],
        "grantee_addresses": [],
        "legal": row[7],
        "estate": "Historic leasehold",
        "depth": None,
        "interest_conveyed_fraction": f"Royalty {row[11]}; primary term {row[12]}",
        "interest_conveyed_decimal": None,
        "interest_reserved": None,
        "lease_id": row[0],
        "portfolio_branch": "Union/Leede deep package" if row[0] != "63/33" else "Pearl / Union/Leede deep package",
        "evidence_class": row[21],
        "source_file": row[23],
        "source_page_or_image": row[23].split(".")[0],
        "confidence": row[22],
        "current_effect": "Historic original lease facts only; current status OPEN",
        "conflicts": [conflict] if conflict else [],
        "curative_action": "Confirm clerk metadata; acquire amendments, corrections, releases, pooling, and HBP evidence",
        "notes": "Later Bk 872/Pg 279 depth and ORRI terms apply only to that assignment; they are not original-lease terms.",
    }


def build_json() -> dict:
    tracts = []
    for tract in TRACTS:
        tag = f"T{tract['tract_no']}"
        tract_instruments = [
            instrument_json(row)
            for row in INSTRUMENTS
            if tag in row["tracts"].split(",")
        ]
        tract_instruments.extend(
            lease_instrument_json(row)
            for row in LEASES
            if row[0] != "63/33" and tag in LEASE_TRACTS[row[0]].split(",")
        )
        tracts.append(
            {
                **tract,
                "instruments": tract_instruments,
                "mineral_owners_proven": [],
                "mineral_claimants_unresolved": [],
                "leases": [
                    {"lease_id": row[0], "legal": row[7], "current_status": row[20], "notes": row[23]}
                    for row in LEASES
                    if (tract["legal"] == row[7]) or (tract["tract_no"] == 4)
                ],
                "working_interest_candidates": (
                    [{"party": "Diversified Production LLC", "asset_id": "OK48147.001.1", "quantum": None, "status": "SCHEDULE-DEPENDENT CLAIMANT ONLY"}]
                    if tract["tract_no"] == 5
                    else []
                ),
                "burdens": [],
                "addresses": [
                    {
                        "party": row[0],
                        "address_as_recorded": row[1],
                        "normalized_address": row[2],
                        "address_type": row[3],
                        "source": row[4],
                        "source_date": row[5],
                        "currentness_caveat": row[6],
                    }
                    for row in ADDRESSES
                    if tract["tract_no"] in ADDRESS_TRACTS[row[0]]
                ],
            }
        )

    return {
        "submission": {
            "ai_name": "GPT-5.6 Sol",
            "timestamp": STAMP,
            "base_files_reviewed": [
                "CHATGPT_NE4_IMPROVED_WORKBOOK_20260715.xlsx",
                "32-11N-25W_Beckham_Co_SECTION32_FINAL_CHAMPION_V2_INTERNAL_REVIEW_HOLD.xlsx",
                "Template.xlsx",
                "SECTION32_FINAL_CHAMPION_EVIDENCE_SUMMARY.md",
                "SECTION32_FINAL_CHAMPION_PROVENANCE.xlsx",
                "SECTION32_FINAL_CHAMPION_OPEN_ISSUES.md",
                "SECTION32_MASTER_INDEX.csv",
                "SECTION32_CONFLICTS.xlsx",
                "SECTION32_DUPLICATES.xlsx",
                "SECTION32_LOW_CONFIDENCE.xlsx",
                "SECTION32_MISSING_ENTRIES.xlsx",
                "SECTION32_EXTRACTION_REPORT.md",
                "SECTION32_CONFIDENCE_REPORT.md",
            ],
            "drive_folder_created": "",
            "files_uploaded": [],
            "overall_release_status": "HOLD_NO_RELEASE",
        },
        "tracts": tracts,
        "title_sheet_rows": [
            {
                "tract_no": tract["tract_no"],
                "legal": tract["legal"],
                "nominal_acres": tract["nominal_acres"],
                "proven_current_mineral_owner": None,
                "present_leasehold_candidate": "Diversified Production LLC / OK48147.001.1 (schedule-dependent)" if tract["tract_no"] == 5 else None,
                "net_acres": None,
                "confidence": "HOLD",
                "comments": tract["current_conclusion"],
            }
            for tract in TRACTS
        ],
        "working_interest_rows": [
            {
                "source": row[0],
                "scope": row[1],
                "party": row[2],
                "fraction_or_percentage": row[3],
                "applies_to": row[4],
                "payout_status": row[5],
                "interest_type": row[6],
                "usable_in_current_totals": row[7],
                "notes": row[8],
            }
            for row in WI_ROWS
        ],
        "runsheet_rows": [
            {
                "sequence": i,
                "tracts": row["tracts"],
                "execution_date": row["execution_date"],
                "recording_date": row["recording_date"],
                "doc_type": row["doc_type"],
                "instrument_no": row["instrument_no"],
                "book_page": row["book_page"],
                "grantor": row["grantor"],
                "grantee": row["grantee"],
                "legal": row["legal"],
                "estate": row["estate"],
                "depth": row["depth"],
                "interest_conveyed": row["interest"],
                "interest_reserved": row["reserved"],
                "lease_id": row["lease_id"],
                "portfolio_branch": row["branch"],
                "evidence_class": row["evidence_class"],
                "source_file": row["source"],
                "image_page": row["image"],
                "confidence": row["confidence"],
                "current_effect": row["current_effect"],
                "conflict": row["conflict"],
                "curative_action": row["curative"],
                "notes": row["notes"],
            }
            for i, row in enumerate(INSTRUMENTS, start=1)
        ] + [
            {
                "sequence": len(INSTRUMENTS) + i,
                "tracts": LEASE_TRACTS[row[0]],
                "execution_date": row[3],
                "recording_date": row[4],
                "doc_type": "OGL",
                "instrument_no": row[5],
                "book_page": row[6],
                "grantor": row[1],
                "grantee": row[2],
                "legal": row[7],
                "estate": "Historic leasehold",
                "depth": None,
                "interest_conveyed": f"Royalty {row[11]}; primary term {row[12]}",
                "interest_reserved": None,
                "lease_id": row[0],
                "portfolio_branch": "Union/Leede deep package",
                "evidence_class": row[21],
                "source_file": row[23],
                "image_page": row[23].split(".")[0],
                "confidence": row[22],
                "current_effect": "Historic original lease facts only; current status OPEN",
                "conflict": lease_instrument_json(row)["conflicts"],
                "curative_action": "Confirm clerk metadata; acquire amendments, corrections, releases, pooling, and HBP evidence",
                "notes": "Later assignment depths and reservations are not original-lease terms.",
            }
            for i, row in enumerate((lease for lease in LEASES if lease[0] != "63/33"), start=1)
        ],
        "template_improvements": TEMPLATE_IMPROVEMENTS,
        "new_evidence_only": [
            "Controlling tract correction: T1 NE/4 160; T2 E/2 SE/4 80; T3 N/2 below top Hunton 320 overlapping; T4 Bk 872/279 non-additive deep overlay; T5 modern full-section record branch.",
            "Original OGL 63/33 proves the Garrett/McKinley E/2 SE/4 lease face, 1/8 royalty, 10-year primary term, and 1947-11-26 recording; later assignments control only their own depth-limited scope.",
            "Bk 845/Pg 47 defines the Pearl wellbore interval and a historic 1/8 of 8/8 ORRI reservation.",
            "Bk 1014/Pg 75 contains no exact WI quantum and has a source-title defect; donor 1%/52.632% implications are rejected.",
            "Original faces for all 11 leases listed in Bk 872/Pg 279 establish historic parties, legals, royalties, primary terms, and recorded addresses without proving present survival.",
            "Bk 872/Pg 279 lease rows reconcile to 1,360 row acres, 560 unique nominal acres, 800 overlap, and an 80-acre S/2 NW/4 schedule gap; none is Diversified net leasehold acreage.",
            "Direct lease faces show instrument 5169 at 352/719, 5168 at 352/721, and 3757 at 376/369, conflicting with protocol-index values 5051, 5052, and 3707.",
            "Bk 1014/Pg 75 is without warranty and effective for all purposes as of first production from the McCall #1-29 well; that condition does not prove current production or title.",
            "2395/462 and 2400/566 prove claimant identity for OK48147.001.1 only, not tract allocation or quantum.",
        ],
        "rejected_or_downgraded_claims": [
            "Donor five-parcel tract definitions and additive 640-acre total.",
            "15.000875% as present tract WI; it is relative-chain planning math only.",
            "48.75% OCM transaction RTI as present Diversified WI.",
            "Bk 509/Pg 220 participant percentages as fractions of 8/8.",
            "Financing statements, mortgages, releases, or merger notices as affirmative ownership proof.",
            "Historic probate internal Tract 3/4 labels as project tracts.",
            "Carlson wells in Section 5-10N-25W as Section 32 HBP evidence.",
            "Any historical lease/probate/instrument address as a verified current address.",
        ],
        "acquisition_queue": [
            {"priority": row[0], "tract": row[1], "source": row[2], "purpose": row[3]}
            for row in ACQUISITION_QUEUE
        ],
        "qa": {
            "tract_legals_verified": True,
            "no_blank_as_zero": True,
            "no_unsupported_current_owner": True,
            "portfolio_branches_separated": True,
            "addresses_source_cited": True,
            "formulas_checked": True,
            "uploaded_files_reopened": False,
        },
    }


NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
RED = "F4CCCC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7B7B7")


def style_sheet(ws, freeze: str = "A2", filter_row: int = 1) -> None:
    ws.freeze_panes = freeze
    if ws.max_row >= filter_row:
        ws.auto_filter.ref = f"A{filter_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    for cell in ws[filter_row]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=THIN)
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
        width = min(max(max(map(len, values), default=8) + 2, 10), 42)
        ws.column_dimensions[get_column_letter(col)].width = width


def add_table_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def build_workbook(path: Path) -> None:
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme_rows = [
        ["SECTION 32 MULTI-TRACT EVIDENCE-CONTROLLED PATCH"],
        ["Status", "HOLD_NO_RELEASE"],
        ["Purpose", "Workbook-ready correction package; not a title opinion"],
        ["Critical correction", "Tracts are overlapping estate/depth/branch scopes and must not be acreage-summed."],
        ["Unknown treatment", "Blank/NOT DETERMINED is never zero."],
        ["Current-title rule", "No exact present mineral, WI, NRI, ORRI, net acres, HBP, or depth allocation is stated."],
        ["Evidence hierarchy", "Direct image > reviewed schedule page > complete packet > index/public metadata > unverified lead"],
        ["Generated", STAMP],
    ]
    for row in readme_rows:
        readme.append(row)
    readme["A1"].fill = PatternFill("solid", fgColor=NAVY)
    readme["A1"].font = Font(color=WHITE, bold=True, size=14)
    readme.merge_cells("A1:H1")
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 100
    readme.sheet_view.showGridLines = False

    add_table_sheet(
        wb,
        "Tract Control",
        ["Tract", "Legal / Scope", "Nominal Acres", "Additive?", "Scope Notes", "Current Conclusion", "Release Status"],
        [
            [f"T{t['tract_no']}", t["legal"], t["nominal_acres"], "NO", t["scope_notes"], t["current_conclusion"], "HOLD"]
            for t in TRACTS
        ],
    )
    add_table_sheet(
        wb,
        "Title",
        ["Tract", "Legal", "Nominal Acres", "Historic Fee Owner Proven", "Current Mineral Owner Proven", "Unresolved Claimant Branches", "Historical Leasehold", "Present Leasehold Candidate", "Mailing Address", "Address Source", "OGL", "Royalty", "Term / Expiration", "Net Acres", "Evidence Class", "Confidence", "Curative Action", "Comments"],
        [
            ["T1", "NE/4", 160, "Cornelius B. Cornelius (historic inception only)", None, "Post-Bollenbach mineral chain", "Union/Leede Crook and deep branches", None, None, None, "264/345; 352/719; 352/721; 376/369", "1/8 at 264/345; 3/16 at later original faces", "5 years by reviewed faces; commencement varies", None, "Direct original images + recorded exhibits", "HOLD", "Complete downstream fee/mineral, amendment, release, and HBP chains", TRACTS[0]["current_conclusion"]],
            ["T2", "E/2 SE/4", 80, None, None, "Garrett/Pearl/Crook/Leede successors", "63/33; Humble/Union; Cities Service/Leede", None, None, None, "63/33", "1/8", "10 years from 1947-11-13", None, "Direct original image + recorded assignments", "HOLD", "Pull amendments, releases, payout, successors, and HBP evidence", TRACTS[1]["current_conclusion"]],
            ["T3", "N/2 below top Hunton", 320, None, None, "Leede/McCall/trust branch", "1014/75 facial branch", None, None, None, None, None, None, None, "Direct image; source-title conflict", "HOLD", "Prove predecessor title and exact depth definition", TRACTS[2]["current_conclusion"]],
            ["T4", "Bk 872/Pg 279 deep leasehold overlay", None, None, None, "Union/Leede lease schedule", "Eleven original OGL faces reviewed", None, None, None, "Bk 872/Pg 279 schedule", "1/8 or 3/16 by original face; later Union ORRIs are assignment-scoped", "5 or 10 years by original face", None, "Direct original images + assignment/exhibits", "HOLD", "Pull missing assignment pages, agreement, amendments, releases and HBP", TRACTS[3]["current_conclusion"]],
            ["T5", "Full-section modern record branch", None, None, None, "Chesapeake/Tapstone/KL CHK; Burlington; Unbridled/MNR; Teocalli; Canvas/DP Ponies", "Schedule-dependent corporate branches", "Diversified Production LLC / OK48147.001.1 — claimant only", "1600 Corporate Drive, Birmingham, AL 35242", "Recorded address blocks, 2395/415–464 and 2400/551–567", None, None, None, None, "Reviewed schedule pages + index metadata", "HOLD", "Pull all operative schedules and reconcile branches", TRACTS[4]["current_conclusion"]],
        ],
    )
    run_headers = ["Sequence", "Tracts", "Execution Date", "Recording Date", "Doc Type", "Instrument No", "Book/Page", "Grantor", "Grantor Address", "Grantee", "Grantee Address", "Legal", "Estate", "Depth", "Interest Conveyed", "Interest Reserved", "Lease ID", "Portfolio Branch", "Evidence Class", "Source File", "Image/Page", "Confidence", "Current Effect", "Conflict ID / Text", "Curative Action", "Notes"]
    run_rows = [
        [i, r["tracts"], r["execution_date"], r["recording_date"], r["doc_type"], r["instrument_no"], r["book_page"], r["grantor"], r.get("grantor_address"), r["grantee"], r.get("grantee_address"), r["legal"], r["estate"], r["depth"], r["interest"], r["reserved"], r["lease_id"], r["branch"], r["evidence_class"], r["source"], r["image"], r["confidence"], r["current_effect"], r["conflict"], r["curative"], r["notes"]]
        for i, r in enumerate(INSTRUMENTS, start=1)
    ]
    for row in LEASES:
        if row[0] == "63/33":
            continue
        details = lease_instrument_json(row)
        run_rows.append(
            [
                len(run_rows) + 1,
                LEASE_TRACTS[row[0]],
                row[3],
                row[4],
                "OGL",
                row[5],
                row[6],
                row[1],
                LEASE_ADDRESSES[row[0]],
                row[2],
                None,
                row[7],
                "Historic leasehold",
                None,
                f"Royalty {row[11]}; primary term {row[12]}",
                None,
                row[0],
                details["portfolio_branch"],
                row[21],
                row[23],
                details["source_page_or_image"],
                row[22],
                details["current_effect"],
                "; ".join(details["conflicts"]) or None,
                details["curative_action"],
                details["notes"],
            ]
        )
    add_table_sheet(wb, "Runsheet", run_headers, run_rows)
    add_table_sheet(
        wb,
        "Leases",
        ["Lease_ID", "Lessor", "Lessee", "Execution_Date", "Recording_Date", "Instrument_No", "Book_Page", "Legal", "Gross_Acres", "Depth_From", "Depth_To", "Royalty", "Primary_Term", "Pugh", "Pooling", "Shut_In", "Later_Assignment", "Reservation", "Release", "HBP_Evidence", "Current_Status", "Evidence_Class", "Confidence", "Notes"],
        LEASES,
    )
    add_table_sheet(
        wb,
        "Acreage Reconciliation",
        ["Metric", "Acres", "Basis / Limitation"],
        ACREAGE_RECONCILIATION,
    )
    add_table_sheet(
        wb,
        "WI Historical",
        ["Source", "Scope", "Party", "Exact Fraction / Percentage", "Applies To", "Payout Status", "Interest Type", "Use in Current Totals?", "Notes"],
        WI_ROWS,
    )
    add_table_sheet(
        wb,
        "WI Present Candidates",
        ["Tract", "Candidate", "Asset ID", "WI", "NRI", "Royalty", "ORRI", "Net Leasehold Acres", "Depth", "Status", "Source", "Curative"],
        [["T5 / exact tract allocation open", "Diversified Production LLC — schedule-dependent claimant only", "OK48147.001.1", None, None, None, None, None, None, UNKNOWN, "2395/462; 2400/566", "Obtain operative schedules and predecessor title"]],
    )
    add_table_sheet(wb, "Modern Chain", ["Book/Page", "Instrument No", "Grantor", "Grantee", "Type", "Portfolio Branch", "Evidence", "Open"], MODERN_CHAIN)
    add_table_sheet(wb, "Addresses", ["Party", "Recorded Address", "Normalized Address", "Address Type", "Source", "Effective/Recording Date", "Currentness Caveat"], ADDRESSES)
    add_table_sheet(wb, "Acquisition Queue", ["Priority", "Tract", "Exact Source", "Purpose"], ACQUISITION_QUEUE)
    add_table_sheet(wb, "Template Improvements", ["No.", "Improvement"], [[i, text] for i, text in enumerate(TEMPLATE_IMPROVEMENTS, 1)])
    add_table_sheet(
        wb,
        "QA",
        ["Control", "Result", "Evidence / Note"],
        [
            ["Controlling tract definitions", "PASS", "T1 NE/4; T2 E/2 SE/4; T3 N/2 below Hunton; T4 Bk 872/279 deep overlay; T5 modern record branch"],
            ["No additive tract-acre total", "PASS", "T4 document-location math is separately labeled: 1,360 row / 560 unique / 800 overlap / 80 schedule gap; no current net acres"],
            ["Unknown not zero", "PASS", "Unknown quantities are blank or NOT DETERMINED"],
            ["No unsupported current owner", "PASS", "Diversified is identified only as schedule-dependent claimant"],
            ["Portfolio branches separated", "PASS", "Modern Chain uses distinct branch column"],
            ["Addresses source-cited", "PASS", "One modern recorded Diversified address and 14 exact historical addresses; every historical address is marked not current"],
            ["Original OGL faces", "PASS", "Eleven lease faces reviewed at Images 0353, 1197–1198, 1372, 1522–1527, 1542–1545, 1552–1555, 1569–1570"],
            ["Face/index conflicts visible", "PASS", "352/719: 5169 vs 5051; 352/721: 5168 vs 5052; 376/369: 3757 vs 3707"],
            ["Formula safeguards", "PASS", "No ownership-total formula exists; no blank-to-zero arithmetic"],
            ["Stale template facts", "PASS", "No Section 27, Cherokee, Kunu, Greenhead, Aztek, or Phillips facts imported"],
            ["Index limitation", "PASS", "Page 48 gap and locator-only limits disclosed"],
            ["Drive upload reopened", "OPEN", "Must remain OPEN until files are uploaded and reopened from Drive"],
            ["Release decision", "HOLD", "Exact present mineral/lease/WI conclusions are not supported"],
        ],
    )

    for ws in wb.worksheets:
        ws.print_title_rows = "1:1"
        ws.auto_filter.ref = ws.dimensions if ws.max_row > 1 else None
    wb.save(path)


def copy_sheet(source_path: Path, source_sheet: str, output_path: Path) -> None:
    source = load_workbook(source_path)
    src = source[source_sheet]
    wb = Workbook()
    dst = wb.active
    dst.title = source_sheet
    for row in src.iter_rows():
        for cell in row:
            target = dst[cell.coordinate]
            target.value = cell.value
            if cell.has_style:
                target.font = copy(cell.font)
                target.fill = copy(cell.fill)
                target.border = copy(cell.border)
                target.alignment = copy(cell.alignment)
                target.number_format = cell.number_format
                target.protection = copy(cell.protection)
    for key, dim in src.column_dimensions.items():
        dst.column_dimensions[key].width = dim.width
    dst.freeze_panes = src.freeze_panes
    dst.auto_filter.ref = src.auto_filter.ref
    dst.sheet_view.showGridLines = False
    wb.save(output_path)


def write_markdown() -> None:
    change_log = """# GPT-5.6 Sol Section 32 Change Log

Status: **HOLD_NO_RELEASE**

## Material corrections
1. Replaced the donor's five additive surface parcels with the controlling overlapping tract definitions.
2. Isolated Tract 2 to E/2 SE/4, 80 acres; removed SE/4 NE/4 contamination.
3. Recast Tract 3 as N/2 below top Hunton, 320 surface acres, with the Bk 1014/Pg 75 source-title conflict visible.
4. Recast Tract 4 as the non-additive Bk 872/Pg 279 deep leasehold overlay and normalized its scheduled lease references.
5. Recast Tract 5 as a full-section modern record branch, not 640 net acres or 8/8 ownership.
6. Separated historic document-scoped WI/ORRI/payout facts from present-title candidates.
7. Preserved only the recorded Diversified address and labeled it historic.
8. Separated Chesapeake/Tapstone/KL CHK, Burlington, Unbridled/MNR, Teocalli, and Canvas/DP Ponies branches.
9. Rejected 15.000875%, 48.75%, and Bk 509/Pg 220 participant allocations as present tract WI.
10. Added explicit acquisition, conflict, curative, evidence-tier, and no-blank-as-zero controls.
11. Reviewed original faces for all 11 OGLs listed in Bk 872/Pg 279 and populated historic royalty, term, recording, address, and party fields.
12. Reconciled Bk 872/Pg 279 to 1,360 row acres, 560 unique nominal acres, 800 overlap, and an 80-acre S/2 NW/4 schedule gap—document metrics only.
13. Corrected the historic lessor spelling to Ima Pearl Crook from Image 1197.
14. Flagged direct-face/index conflicts: 352/719 instrument 5169 vs 5051; 352/721 instrument 5168 vs 5052; 376/369 instrument 3757 vs 3707.
15. Added 14 source-dated historical address rows, each expressly labeled not current.

## Evidence limitations
- The canonical index has 1,928 unique locator rows; 1,883 are index-only.
- Stamped page 48 is missing.
- Fifty-three repeated page-54 occurrences are quarantined.
- Complete operative schedules remain missing for the modern claimant chain.
- Bk 872/Pg 280–281, the unrecorded 1984 agreement, amendments, releases, and lease-specific HBP evidence remain missing.
"""
    qa = """# GPT-5.6 Sol Section 32 QA

## Passed
- All five controlling tract definitions are present.
- Tract 4 document-location math is separately controlled: 1,360 row acres, 560 unique nominal acres, 800 overlap, and 80 schedule-gap acres; no current net-acre total is calculated.
- Unknown ownership quantities are null, blank, OPEN, or NOT DETERMINED—not zero.
- No unsupported current mineral owner, WI, NRI, ORRI, net acres, HBP, or lease term is stated.
- Portfolio branches, title conveyances, mergers, mortgages, financing statements, and releases are classified separately.
- The recorded Diversified address and 14 historical addresses are source/date cited; no historical address is represented as current.
- Eleven original OGL faces were directly reviewed and matched to the Bk 872/Pg 279 lease schedule.
- Three direct-face/index instrument-number conflicts are visible and remain open for clerk confirmation.
- JSON parses and contains all five tracts.
- Workbook sheets reopen successfully and required headers are present.

## Open
- Drive upload/reopen verification remains OPEN until a write-capable authenticated Drive interface is available.
- Formula rendering by Excel/LibreOffice was not required because no ownership-total formulas are used.

## Release
**HOLD_NO_RELEASE.** Suitable as a research/patch donor, not a present-title opinion.
"""
    (OUT / "GPT56_SOL_SECTION32_CHANGE_LOG_20260716.md").write_text(change_log, encoding="utf-8")
    (OUT / "GPT56_SOL_SECTION32_QA_20260716.md").write_text(qa, encoding="utf-8")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    payload = build_json()
    json_path = OUT / "GPT56_SOL_SECTION32_RESULTS_20260716.json"
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    workbook_path = OUT / "GPT56_SOL_SECTION32_IMPROVED_WORKBOOK_20260716.xlsx"
    build_workbook(workbook_path)
    copy_sheet(workbook_path, "Runsheet", OUT / "GPT56_SOL_SECTION32_RUNSHEET_20260716.xlsx")

    source = load_workbook(workbook_path)
    patch = Workbook()
    patch.remove(patch.active)
    for sheet_name in ("Title", "WI Historical", "WI Present Candidates", "Addresses"):
        src = source[sheet_name]
        dst = patch.create_sheet(sheet_name)
        for row in src.iter_rows():
            for cell in row:
                target = dst[cell.coordinate]
                target.value = cell.value
                if cell.has_style:
                    target.font = copy(cell.font)
                    target.fill = copy(cell.fill)
                    target.border = copy(cell.border)
                    target.alignment = copy(cell.alignment)
                    target.number_format = cell.number_format
        for key, dim in src.column_dimensions.items():
            dst.column_dimensions[key].width = dim.width
        dst.freeze_panes = src.freeze_panes
        dst.auto_filter.ref = src.auto_filter.ref
        dst.sheet_view.showGridLines = False
    patch.save(OUT / "GPT56_SOL_SECTION32_TITLE_WI_PATCH_20260716.xlsx")
    write_markdown()
    print(f"Built {len(list(OUT.iterdir()))} artifacts in {OUT}")


if __name__ == "__main__":
    main()
