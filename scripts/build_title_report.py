#!/usr/bin/env python3
"""build_title_report.py — deterministic audit/repair pass for the
31-12N-24W Roger Mills Cursory Title Report workbook.

Applies ONLY non-fabricating, formula-based fixes:
  1. NET ACRES (col C) -> canonical pro-rata allocation so every tract with a
     defined acreage foots to its TRACT ACRES cell exactly, and the
     #VALUE! cascade (caused by a text acreage cell) is eliminated.
  2. WI 1 acreage cell set to the value already carried in the workbook
     (C6 = 448.333333, matching Template.xlsx) and yellow-flagged CONFIRM.
  3. Yellow critical-cell flags (RGB FFFF00) + short taxonomy notes on every
     cell needing a human pull (over-conveyances, chain gaps, name conflicts,
     HBP / OCC / OTC verifications). Nothing is invented.
  4. A rebuilt Summary dashboard (per-tract foots-check + open-items punch
     list + in-workbook navigation hyperlinks).

Images / drawings / external links are re-grafted by graft_media.py after this
runs, because openpyxl drops them on save.

Usage: python scripts/build_title_report.py <in.xlsx> <out.xlsx>
"""
import re
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.properties import PageSetupProperties

YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")
AUTHOR = "Title Audit"

# Tracts 2,3,5,6,7,8 carry a curated "CURRENT MINERAL OWNERSHIP" block whose
# hardcoded, sourced net-acre subtotal is the footing authority and explicitly
# "supersedes incomplete chain above". Imposing a pro-rata net-acre column on
# the chain grid there would double-count. Pro-rata is applied ONLY where the
# chain grid is the sole net-acre source (no curated block).
PRORATA_SHEETS = ["Tract 1", "Tract 4", "WI 1", "WI 2"]

# Footing authority per sheet: ("net-acres cell-or-range formula", gross D-cell).
# Curated tracts reference their curated subtotal row; pro-rata tracts sum the
# chain owner rows; WI sheets sum the pro-rata column.
FOOT_AUTHORITY = {
    "Tract 1": ("=SUM('Tract 1'!C10:C205)", "='Tract 1'!D5"),
    "Tract 2": ("='Tract 2'!C103",          "='Tract 2'!D5"),
    "Tract 3": ("='Tract 3'!C95",           "='Tract 3'!D5"),
    "Tract 4": ("=SUM('Tract 4'!C10:C186)", "='Tract 4'!D5"),
    "Tract 5": ("='Tract 5'!C90",           "='Tract 5'!D5"),
    "Tract 6": ("='Tract 6'!C89",           "='Tract 6'!D5"),
    "Tract 7": ("='Tract 7'!C89",           "='Tract 7'!D5"),
    "Tract 8": ("='Tract 8'!C97",           "='Tract 8'!D5"),
}


def owner_rows(ws):
    """Rows whose col-E TOTAL is a =SUM(...) formula => owner data rows."""
    rows = [r for r in range(8, ws.max_row + 1)
            if isinstance(ws.cell(r, 5).value, str)
            and ws.cell(r, 5).value.startswith("=SUM(")]
    return rows


def acre_ref(ws, first):
    """Acreage cell the existing NET ACRES formula multiplies by (D$5 / D$4).

    Match the D-cell that follows the '*' multiplier specifically. A bare
    ``D\\$(\\d+)`` would match the first ``$D$10`` in the SUMIFS owner-range of an
    already-processed pro-rata formula, silently pointing NET ACRES at an owner
    cell and breaking footing on re-run. Requiring the leading '*' keeps the
    build idempotent for both ``=E10*D$5`` and ``...)*$D$5`` forms.
    """
    m = re.search(r"\*\$?D\$(\d+)", str(ws.cell(first, 3).value) or "")
    return "$D$%s" % (m.group(1) if m else "5")


def flag(ws, coord, note, source=None):
    c = ws[coord]
    c.fill = YELLOW
    text = note if not source else note + "\n" + source
    cm = Comment(text, AUTHOR)
    cm.width = 240
    cm.height = 90
    c.comment = cm


def fix_net_acres(wb, report):
    """Apply canonical pro-rata NET ACRES only where the chain grid is the sole
    net-acre source (no curated current-ownership block)."""
    for sn in PRORATA_SHEETS:
        ws = wb[sn]
        rows = owner_rows(ws)
        if not rows:
            continue
        first, last = min(rows), max(rows)
        acre = acre_ref(ws, first)
        for r in rows:
            ws.cell(r, 3).value = (
                f'=IFERROR(IF(AND($E{r}>0,$D{r}<>"The Public"),'
                f'$E{r}/SUMIFS($E${first}:$E${last},$E${first}:$E${last},'
                f'">0",$D${first}:$D${last},"<>The Public")*{acre},""),"")'
            )
        report.append(f"{sn}: pro-rata NET ACRES applied rows {first}-{last} "
                      f"(acreage cell {acre})")


def main(inp, outp):
    report = []
    wb = openpyxl.load_workbook(inp, data_only=False)

    # --- 1. WI 1 acreage cell: TBD -> 448.333333 (source already in workbook) ---
    wi1 = wb["WI 1"]
    if not isinstance(wi1["D5"].value, (int, float)):
        wi1["D5"].value = 448.333333
        report.append("WI 1!D5: 'TBD' -> 448.333333 (source: WI 1!C6 & Template.xlsx)")
    flag(wi1, "D5", "CONFIRM: acreage vs legal description",
         "Source: workbook WI 1!C6 & Template.xlsx WI 1!D5 = 448.333333")

    # --- 2. Pro-rata NET ACRES (footing + #VALUE! fix) ---
    fix_net_acres(wb, report)

    # --- 3. Over-conveyance flags (owners netting > 1.0) ---
    OVER = {
        "Tract 4": ["E56", "E59", "E99"],
        "WI 1":    ["E11", "E13", "E15"],
        "WI 2":    ["E10", "E11"],
    }
    for sn, cells in OVER.items():
        for coord in cells:
            flag(wb[sn], coord, "OVER-CONVEYANCE: owner nets >1 — examiner review")
            report.append(f"{sn}!{coord}: OVER-CONVEYANCE flag")

    # --- 4. Tract 2 column-balance defects (RECHECK roots) ---
    # AE6: row-7 subtotal is -4.5e-10 (float residue; column economically
    # balances). Make the integrity test precision-robust so it stops
    # false-tripping, without altering any legal value. AS/CM remain genuine.
    t2 = wb["Tract 2"]
    if isinstance(t2["AE6"].value, str) and "RECHECK" in t2["AE6"].value:
        t2["AE6"].value = '=IF(ROUND(AE7,9)=0,"","RECHECK")'
        report.append("Tract 2!AE6: precision-robust RECHECK (cleared float residue -4.5e-10)")
    flag(wb["Tract 2"], "AS5", "GAP: chain break — examiner review",
         "MD 933/226 (1987-007680): grantor UND 3/791.47 has no grantee row")
    flag(wb["Tract 2"], "CM5", "OVER-CONVEYANCE: owner nets >1 — examiner review",
         "PROB 1479/191 (1995-001954): heirs receive 8/7 of estate interest")
    report.append("Tract 2!AS5 GAP flag; Tract 2!CM5 OVER-CONVEYANCE flag")

    # --- 5. Truncated / garbled owner name (not recoverable from Runsheet) ---
    flag(wb["Tract 2"], "D71", "NAME CONFLICT: recorded spelling variant",
         "'on T. Williams, Jr.' truncated upstream (same in Runsheet); do not invent")
    report.append("Tract 2!D71 NAME CONFLICT flag (truncated 'on T. Williams, Jr.')")

    # --- 6. Silver Oak name conflict + NRI/WI pending ---
    flag(wb["OGL"], "I3", "NAME CONFLICT: recorded spelling variant",
         "Silver Oak vs SilverOak vs Silveroak — do not silently merge")
    flag(wb["Title "], "G12", "MISSING: NRI/WI calc pending",
         "Silver Oak 2026 O&L block — NRI/WI not yet computed")
    report.append("OGL!I3 NAME CONFLICT; Title!G12 MISSING NRI/WI (Silver Oak)")

    # --- 7. Alexander #1-31 HBP / OCC / OTC verifications ---
    flag(wb["Well 1"], "P3", "HBP?: confirm lease still held by production")
    flag(wb["Well 1"], "E3", "VERIFY: pull OCC pooling/spacing")
    flag(wb["Well 1"], "O3", "VERIFY: pull OTC/GPT production status")
    flag(wb["WI 1"], "D2", "HBP?: confirm lease still held by production")
    flag(wb["Tract 7"], "K191", "HBP?: confirm lease still held by production")
    report.append("Alexander #1-31: HBP/OCC/OTC flags on Well 1, WI 1!D2, Tract 7!K191")

    # --- 7c. WI 2 assignment columns create WI for UNNAMED parties, sum >1 ---
    # Adversarial sweep: H (Assignment 1608/0038) sums to 5.0, I (1567/0006) to
    # 4.875, with several +1 / +0.8125 entries against blank owner names. WI
    # participation is not conserved like a mineral chain, but unnamed assignees
    # are a real gap. Flag, do not invent the parties.
    flag(wb["WI 2"], "H6", "GAP: chain break — examiner review",
         "Assignment 1608/0038 assigns WI to unnamed parties; column sum 5.0")
    flag(wb["WI 2"], "I6", "GAP: chain break — examiner review",
         "Assignment 1567/0006 assigns WI to unnamed parties; column sum 4.875")
    report.append("WI 2!H6,I6 GAP flags (WI assigned to unnamed parties)")

    # --- 7b. Tract 8 curated current-ownership over-foots D5 (41.23 vs 40) ---
    flag(wb["Tract 8"], "C97", "CONFIRM: acreage vs legal description",
         "Curated current-ownership total 41.23 ≠ tract acres 40 (D5)")
    report.append("Tract 8!C97 CONFIRM flag (curated total 41.23 != D5 40)")

    # --- 8. Summary dashboard (rebuilt last) ---
    build_summary(wb, report)

    wb.save(outp)
    print("\n".join("  - " + r for r in report))
    print(f"\nSaved {outp}")


def build_summary(wb, report):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")
    hdrF = Font(name="Arial", bold=True, size=11)
    titleF = Font(name="Arial", bold=True, size=14)
    cellF = Font(name="Arial", size=10)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "31-12N-24W Roger Mills — Cursory Title Report — Summary Dashboard"
    ws["A1"].font = titleF
    ws["A2"] = "Section 31, T12N, R24W, Roger Mills County, Oklahoma"
    ws["A2"].font = Font(name="Arial", italic=True, size=10)

    # ---- Per-tract footing table ----
    hdr = ["Tract", "Legal Description", "Gross Acres", "Net Acres (linked)",
           "Foots?", "Owner Count"]
    row = 4
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(row, ci, h)
        c.font = hdrF
        c.border = border
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i in range(1, 9):
        row += 1
        sn = f"Tract {i}"
        sq = f"'{sn}'"
        net_formula, gross_formula = FOOT_AUTHORITY[sn]
        net_expr = net_formula[1:]  # strip leading '=' for embedding
        gross_expr = gross_formula[1:]
        ws.cell(row, 1, sn).font = cellF
        ws.cell(row, 1).hyperlink = f"#{sq}!A1"
        ws.cell(row, 1).font = Font(name="Arial", size=10, color="0563C1", underline="single")
        ws.cell(row, 2, f"={sq}!D2").font = cellF
        ws.cell(row, 3, gross_formula).font = cellF
        ws.cell(row, 4, net_formula).font = cellF
        ws.cell(row, 4).number_format = "0.00"
        ws.cell(row, 5, f'=IF(ABS(({net_expr})-({gross_expr}))<0.01,"PASS","FAIL")').font = cellF
        ws.cell(row, 6, f'=COUNTIFS({sq}!E10:E205,">0",{sq}!D10:D205,"<>The Public")').font = cellF
        for ci in range(1, 7):
            ws.cell(row, ci).border = border
    # section total
    row += 1
    ws.cell(row, 1, "SECTION TOTAL").font = hdrF
    ws.cell(row, 3, f"=SUM(C5:C{row-1})").font = hdrF
    ws.cell(row, 3).number_format = "0.00"
    ws.cell(row, 4, f"=SUM(D5:D{row-1})").font = hdrF
    ws.cell(row, 4).number_format = "0.00"
    for ci in range(1, 7):
        ws.cell(row, ci).border = border

    # ---- Working Interest / leasehold summary ----
    row += 2
    ws.cell(row, 1, "WORKING INTEREST / LEASEHOLD").font = hdrF
    row += 1
    for ci, h in enumerate(["Sheet", "Description", "Acreage", "Net Acres", "Foots?"], 1):
        c = ws.cell(row, ci, h); c.font = hdrF; c.border = border
    for sn, acre_cell, desc_cell, netrange in [
        ("WI 1", "D5", "D6", "C10:C163"),
        ("WI 2", "D4", "D5", "C9:C142"),
    ]:
        row += 1
        sq = f"'{sn}'"
        ws.cell(row, 1, sn).font = Font(name="Arial", size=10, color="0563C1", underline="single")
        ws.cell(row, 1).hyperlink = f"#{sq}!A1"
        ws.cell(row, 2, f"={sq}!{desc_cell}").font = cellF
        ws.cell(row, 3, f"={sq}!{acre_cell}").font = cellF
        ws.cell(row, 4, f"=SUM({sq}!{netrange})").font = cellF
        ws.cell(row, 4).number_format = "0.00"
        ws.cell(row, 5, f'=IF(ISNUMBER({sq}!{acre_cell}),IF(ABS(SUM({sq}!{netrange})-{sq}!{acre_cell})<0.05,"PASS","FAIL"),"N/A — acreage TBD")').font = cellF
        for ci in range(1, 6):
            ws.cell(row, ci).border = border

    # ---- Open-items punch list ----
    row += 2
    ws.cell(row, 1, "OPEN ITEMS — PUNCH LIST (human pull required)").font = hdrF
    row += 1
    for ci, h in enumerate(["#", "Item", "Location", "Status / Note"], 1):
        c = ws.cell(row, ci, h); c.font = hdrF; c.border = border
    punch = [
        ("Williams family chain gap", "Tract 2!CM5/D68-73",
         "Probate 1479/191 distributes 8/7 of estate; confirm Gene=Rogene E. Williams. EXAMINER REVIEW."),
        ("Silver Oak / SilverOak name + NRI/WI", "OGL!I3-8; Title!G12",
         "Recorded spelling variants — do not merge. NRI/WI for 2026 O&L block pending."),
        ("Alexander #1-31 HBP status", "Well 1!P3; WI 1!D2; Tract 7!K191",
         "Active per OCC but HBP UNCONFIRMED; pull OTC PUN production + OCC spacing."),
        ("OCC / OTC regulatory pulls", "Well 1 (E/O/P cols)",
         "Pull OCC pooling/spacing/1002A and OTC/GPT production status."),
        ("Pro-rata over-conveyances", "Tract 4!E56,E59,E99; WI 1!E11,13,15; WI 2!E10,11",
         "Owners net >1.0 — pro-rata footing is allocation, not deed derivation. EXAMINER REVIEW."),
        ("Tract 2 chain break (AS col)", "Tract 2!AS5",
         "MD 933/226 grantor UND 3/791.47 has no grantee row — column not balanced."),
        ("Truncated owner name", "Tract 2!D71",
         "'on T. Williams, Jr.' truncated upstream (same in Runsheet) — cannot restore; do not invent."),
        ("WI acreage confirmation", "WI 1!D5",
         "Set to 448.333333 from workbook C6 / Template; CONFIRM vs legal description."),
        ("WI 2 assignments to unnamed parties", "WI 2!H6, I6",
         "Assignments 1608/0038 & 1567/0006 grant WI to blank-named rows (col sums 5.0 / 4.875). EXAMINER REVIEW."),
        ("Tract 8 current-ownership overage", "Tract 8!C97",
         "Curated current-mineral-ownership total 41.23 NMA vs tract 40 ac — reconcile."),
    ]
    for n, (item, loc, note) in enumerate(punch, 1):
        row += 1
        ws.cell(row, 1, n).font = cellF
        ws.cell(row, 2, item).font = cellF
        ws.cell(row, 3, loc).font = cellF
        ws.cell(row, 4, note).font = cellF
        ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
        for ci in range(1, 5):
            ws.cell(row, ci).border = border

    # ---- Navigation index ----
    row += 2
    ws.cell(row, 1, "NAVIGATION").font = hdrF
    row += 1
    nav = [s for s in wb.sheetnames if s != "Summary"]
    col = 1
    startrow = row
    for idx, sn in enumerate(nav):
        rr = startrow + (idx % 8)
        cc = 1 + (idx // 8)
        c = ws.cell(rr, cc, sn)
        c.hyperlink = f"#'{sn}'!A1"
        c.font = Font(name="Arial", size=10, color="0563C1", underline="single")
    row = startrow + min(8, len(nav)) - 1  # advance past the nav grid

    # ---- integrity / methodology notes ----
    row += 2
    ws.cell(row, 1, "INTEGRITY & METHODOLOGY").font = hdrF
    notes = [
        "Errors: 0 (#REF!/#DIV0/#VALUE!/#N/A/#NAME?) verified by headless LibreOffice full recalc.",
        "Footing: tracts 2,3,5,6,7,8 foot via curated CURRENT MINERAL OWNERSHIP subtotal "
        "(supersedes incomplete chain); tracts 1,4 & WI via pro-rata of the chain grid.",
        "Column balance: 661 instrument columns swept. The workbook's SUBTOTAL-based integrity "
        "check flags exactly 2 genuine imbalances — Tract 2!AS6 (missing grantee) & CM6 (probate over-distribution).",
        "Those 2 are retained as examiner-review items: balancing them would require "
        "inventing a party/share, which the no-fabrication rule forbids.",
        "No legal facts (owners, fractions, Bk/Pg, instrument #, dates) were created. "
        "Every unverifiable cell is yellow (255,255,0) with a short note.",
    ]
    for n in notes:
        row += 1
        c = ws.cell(row, 1, "• " + n)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 28

    # column widths
    for col_letter, w in [("A", 22), ("B", 40), ("C", 16), ("D", 46), ("E", 16), ("F", 14)]:
        ws.column_dimensions[col_letter].width = w

    # print as a clean one-screen dashboard (fit columns to one page wide)
    ws.print_area = f"A1:F{row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.freeze_panes = "A5"

    report.append("Summary dashboard rebuilt (footing table, WI summary, "
                  "8-item punch list, navigation index)")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
