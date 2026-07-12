#!/usr/bin/env python3
"""Structural conformance comparison of a populated report workbook against its template.

Usage: compare_workbook_to_template.py TEMPLATE.xlsx FINAL.xlsx [REPORT.txt]

Checks sheet names/order, visibility, merged-range deltas, print areas, formula counts,
defined names, and #REF! tokens. Intentional variances should be reviewed against the
project's disclosure list; anything undisclosed is drift.
"""
import sys

import openpyxl


def count_formulas(wb):
    return sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    )


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    tpath, fpath = sys.argv[1], sys.argv[2]
    out_path = sys.argv[3] if len(sys.argv) > 3 else None

    T = openpyxl.load_workbook(tpath)
    F = openpyxl.load_workbook(fpath)

    out = [f"TEMPLATE-CONFORMANCE COMPARISON — {fpath} vs {tpath}"]
    ok = T.sheetnames == F.sheetnames
    out.append(f"Sheet names/order match: {ok}")
    out.append(f"  template: {T.sheetnames}")
    out.append(f"  final   : {F.sheetnames}")

    for name in T.sheetnames:
        if name not in F.sheetnames:
            continue
        tw, fw = T[name], F[name]
        tm = {str(m) for m in tw.merged_cells.ranges}
        fm = {str(m) for m in fw.merged_cells.ranges}
        vis = "OK" if tw.sheet_state == fw.sheet_state else f"DIFF ({tw.sheet_state}->{fw.sheet_state})"
        out.append(
            f"[{name}] visibility {vis}; merges tmpl={len(tm)} final={len(fm)} "
            f"(missing {len(tm - fm)}, added {len(fm - tm)}); "
            f"printarea tmpl={tw.print_area} final={fw.print_area}"
        )

    out.append(f"Formula cells: template={count_formulas(T)} final={count_formulas(F)}")
    out.append(f"Defined names: template={len(T.defined_names)} final={len(F.defined_names)}")
    refs = sum(
        1
        for ws in F.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and "#REF!" in c.value
    )
    out.append(f"#REF! tokens in final: {refs}")

    report = "\n".join(out)
    print(report)
    if out_path:
        with open(out_path, "w") as fh:
            fh.write(report)


if __name__ == "__main__":
    main()
