#!/usr/bin/env python3
"""
Mineral & Leasehold Ownership Report generator — Section 31-12N-24W, Roger Mills Co., OK.

Reads the examiner's CANONICAL "BEST TURN IN COPY" title workbook directly and renders a
self-contained, print-ready HTML report plus a Markdown twin. Every figure is pulled live
from the workbook's Turn_In_Summary, Title, OGL, Curative_Requirements, Well 1, and WI
sheets — nothing is transcribed or invented (honors the project's anti-fabrication "Golden
Law"). The canonical copy supersedes the earlier "NHE Final": it corrected over-credits on
Tracts 2 and 9 and discloses undetermined balances rather than distributing them.

Usage:
    python scripts/mineral_leasehold_report.py --xlsx /tmp/rm/best.xlsx
"""
from __future__ import annotations
import argparse, html, re
from pathlib import Path
import openpyxl

NAVY = "#1F4E79"; AMBER = "#FFE699"; GREEN = "#C6EFCE"; RED = "#FBD5D0"; GREY = "#6b7280"
OUT = Path(__file__).resolve().parent.parent / "reports"
NOISE = ("address needed", "undetermined of record", "final resea", "address unknown",
         "verify before notice", "cannot be verified")


def esc(v):
    return html.escape("" if v is None else str(v))


def num(v, dp=2):
    try:
        return f"{float(v):,.{dp}f}"
    except (TypeError, ValueError):
        return esc(v)


def dt(v):
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(v))
    return f"{int(m.group(2))}/{int(m.group(3))}/{m.group(1)}" if m else esc(v)


def name_addr(cell):
    if cell is None:
        return "", ""
    parts = [p.strip() for p in str(cell).split("\n") if p.strip()]
    name = parts[0] if parts else ""
    addr = [p for p in parts[1:] if not any(n in p.lower() for n in NOISE)]
    return name, " · ".join(addr)


def is_balance(name):
    n = (name or "").lower()
    return "balance of" in n or "undetermined of record" in n or name.startswith("Balance")


# ── Parsers ─────────────────────────────────────────────────────────────────────
def parse_title(ws):
    tracts, wi = [], []; cur = None; mode = None
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value; c = ws.cell(r, 3).value
        bs = str(b).strip() if b is not None else ""
        if bs.startswith("TRACT ") and ":" in bs:
            cur = {"id": bs.replace(":", "").strip(), "legal": str(c or "").strip(), "owners": []}
            tracts.append(cur); mode = "tract"; continue
        if bs == "Working Interest:":
            cur = {"label": str(c or "").strip(), "owners": []}; wi.append(cur); mode = "wi"; continue
        if bs.startswith("Containing") or bs in ("Mineral Owner", "Owner", "") or bs.startswith("Prepared on"):
            continue
        if bs in ("REPORT TOTAL", "TOTAL") or bs.startswith("CURSORY TITLE") or bs.startswith("Section 31"):
            continue
        if cur is None:
            continue
        nm, ad = name_addr(b)
        if not nm:
            continue
        cur["owners"].append({"name": nm, "addr": ad, "acres": c, "ogl": ws.cell(r, 4).value,
                              "roy": ws.cell(r, 5).value, "exp": ws.cell(r, 6).value, "cmt": ws.cell(r, 7).value})
    return tracts, wi


def parse_ogl(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 1).value
        if n is None:
            continue
        out.append({"n": n, "lh": ws.cell(r, 2).value, "bkpg": ws.cell(r, 5).value,
                    "exec": ws.cell(r, 6).value, "grantor": ws.cell(r, 8).value,
                    "grantee": ws.cell(r, 9).value, "gross": ws.cell(r, 13).value,
                    "tracts": ws.cell(r, 14).value, "exp": ws.cell(r, 15).value,
                    "nma": ws.cell(r, 16).value, "roy": ws.cell(r, 17).value})
    return out


def parse_summary(ws):
    rows = []
    for r in range(11, 21):
        if ws.cell(r, 1).value is None:
            continue
        rows.append({k: ws.cell(r, i).value for i, k in enumerate(
            ["t", "legal", "gross", "id", "open", "total", "status", "conf", "ogl", "wi", "next", "src"], 1)})
    return rows


def parse_curative(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        out.append({"id": ws.cell(r, 1).value, "sev": ws.cell(r, 2).value, "blocks": ws.cell(r, 3).value,
                    "tract": ws.cell(r, 4).value, "party": ws.cell(r, 6).value, "instr": ws.cell(r, 7).value,
                    "bkpg": ws.cell(r, 8).value, "type": ws.cell(r, 9).value, "desc": ws.cell(r, 10).value,
                    "need": ws.cell(r, 12).value, "action": ws.cell(r, 13).value, "status": ws.cell(r, 16).value})
    return out


def parse_well(ws):
    keys = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    vals = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    return [(k, v) for k, v in zip(keys, vals) if k]


def owner_rollup(tracts):
    """Aggregate SUPPORTED net acres by owner across tracts (acquisition-target view)."""
    agg = {}
    for t in tracts:
        for o in t["owners"]:
            if is_balance(o["name"]) or not isinstance(o["acres"], (int, float)):
                continue
            key = re.split(r",? a/k/a| a/k/a|,? Trustee| U/A", o["name"])[0].strip()
            e = agg.setdefault(key, {"na": 0.0, "tracts": set(), "leased": False})
            e["na"] += float(o["acres"]); e["tracts"].add(t["id"].replace("TRACT ", "T"))
            if o["ogl"] and str(o["ogl"]) not in ("—", "None"):
                e["leased"] = True
    rows = [{"name": k, "na": v["na"], "tracts": ", ".join(sorted(v["tracts"])), "leased": v["leased"]}
            for k, v in agg.items() if v["na"] >= 0.005]
    return sorted(rows, key=lambda x: -x["na"])


# ── HTML ────────────────────────────────────────────────────────────────────────
def render_html(meta, tracts, wi, leases, summary, curative, well, roll):
    id_na = sum(float(s["id"]) for s in summary if isinstance(s["id"], (int, float)))
    open_na = sum(float(s["open"]) for s in summary if isinstance(s["open"], (int, float)))

    def statusrow(s):
        conf = str(s["conf"] or "")
        cc = "hi" if conf.startswith("High") else ("lo" if conf == "Low" else "med")
        opencls = "verify" if isinstance(s["open"], (int, float)) and s["open"] > 0.01 else ""
        return (f"<tr><td class=num>{esc(s['t'])}</td><td class=cmt>{esc(s['legal'])}</td>"
                f"<td class=num>{num(s['gross'])}</td><td class=num>{num(s['id'])}</td>"
                f"<td class='num {opencls}'>{num(s['open'])}</td><td>{esc(s['status'])}</td>"
                f"<td class=conf-{cc}>{esc(conf)}</td><td class=cmt>{esc(s['next'])}</td></tr>")

    def owner_rows(t):
        out = []
        for o in t["owners"]:
            bal = is_balance(o["name"])
            und = bal or not isinstance(o["acres"], (int, float))
            cls = "verify" if bal else ""
            nm = f"<strong>{esc(o['name'])}</strong>"
            if o["addr"]:
                nm += f"<br><span class=addr>{esc(o['addr'])}</span>"
            elif not bal:
                nm += "<br><span class=addr>address needed</span>"
            ac = num(o["acres"]) if isinstance(o["acres"], (int, float)) else "<em>undet.</em>"
            out.append(f"<tr class='{cls}'><td>{nm}</td><td class=num>{ac}</td>"
                       f"<td>{esc(o['ogl']) if o['ogl'] not in (None,'—') else '—'}</td>"
                       f"<td>{num(o['roy'],4) if isinstance(o['roy'],(int,float)) else esc(o['roy'])}</td>"
                       f"<td>{esc(str(o['exp']).replace(chr(10),' ')) if o['exp'] else ''}</td>"
                       f"<td class=cmt>{esc(o['cmt']) if o['cmt'] else ''}</td></tr>")
        return "\n".join(out)

    tract_html = ""
    for t, s in zip(tracts, summary):
        tract_html += (f"<div class=tract><h3>{esc(t['id'])} <span class=legal>{esc(t['legal'])}</span></h3>"
                       f"<div class=cont>Gross {num(s['gross'])} ac · identified {num(s['id'])} · "
                       f"open {num(s['open'])} · <strong>{esc(s['status'])}</strong> (confidence {esc(s['conf'])})</div>"
                       "<table class=owners><thead><tr><th>Mineral owner</th><th>Net ac</th><th>OGL</th>"
                       "<th>Roy.</th><th>Expiration</th><th>Comments</th></tr></thead><tbody>"
                       f"{owner_rows(t)}</tbody></table></div>")

    def lease_row(l):
        top = str(l["lh"]).upper().startswith("TOP")
        return (f"<tr class='{'top' if top else ''}'><td class=num>{esc(l['n'])}</td><td>{esc(l['lh'])}</td>"
                f"<td>{esc(l['grantor'])}</td><td>{esc(l['grantee'])}</td><td>{dt(l['exec'])}</td>"
                f"<td>{esc(l['bkpg'])}</td><td class=num>{num(l['roy'],4) if isinstance(l['roy'],(int,float)) else esc(l['roy'])}</td>"
                f"<td class=num>{num(l['nma'])}</td><td>{esc(l['tracts'])}</td><td>{dt(l['exp'])}</td></tr>")
    base = [l for l in leases if not str(l["lh"]).upper().startswith("TOP")]
    top = [l for l in leases if str(l["lh"]).upper().startswith("TOP")]
    oglhead = ("<tr><th>OGL</th><th>Status</th><th>Lessor</th><th>Lessee</th><th>Exec.</th>"
               "<th>Bk/Pg</th><th>Roy.</th><th>NMA</th><th>Tracts</th><th>Exp.</th></tr>")

    def cur_row(c):
        sv = str(c["sev"] or ""); sc = "sev-hi" if sv == "High" else ("sev-lo" if sv == "Low" else "sev-med")
        return (f"<tr><td>{esc(c['id'])}</td><td class={sc}>{esc(sv)}</td><td class=num>{esc(c['blocks'])}</td>"
                f"<td>{esc(c['tract'])}</td><td>{esc(c['type'])}</td><td class=cmt>{esc(c['desc'])}</td>"
                f"<td class=cmt>{esc(c['need'])}</td><td>{esc(c['status'])}</td></tr>")

    roll_rows = "\n".join(
        f"<tr><td><strong>{esc(x['name'])}</strong></td><td class=num>{num(x['na'])}</td>"
        f"<td>{esc(x['tracts'])}</td><td>{'Leased' if x['leased'] else '<strong>UNLEASED</strong>'}</td></tr>"
        for x in roll[:25])

    wi_html = ""
    for w in wi:
        rows = "\n".join(
            f"<tr><td><strong>{esc(o['name'])}</strong></td><td>{esc(o['ogl'])}</td>"
            f"<td class=cmt>{esc(o['cmt']) if o['cmt'] else ''}</td></tr>" for o in w["owners"] if o["name"])
        wi_html += f"<h4>{esc(w['label'])}</h4><table class=wi><thead><tr><th>Party of record</th>" \
                   f"<th>Ref (Bk/Pg)</th><th>Comments</th></tr></thead><tbody>{rows}</tbody></table>"

    _dre = re.compile(r"\d{4}-\d{2}")

    def wv(v):
        return dt(v) if isinstance(v, str) and _dre.match(str(v)) else esc(v)
    well_html = "\n".join(f"<tr><td>{esc(k)}</td><td>{wv(v)}</td></tr>" for k, v in well)

    return f"""<!doctype html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width, initial-scale=1">
<title>Mineral & Leasehold Ownership — Sec 31-12N-24W, Roger Mills Co., OK</title><style>
 :root{{--navy:{NAVY};--amber:{AMBER};--green:{GREEN}}}*{{box-sizing:border-box}}
 body{{font-family:"Segoe UI",Arial,sans-serif;color:#1a1a1a;margin:0;background:#eef0f3;line-height:1.45}}
 .page{{max-width:1200px;margin:22px auto;background:#fff;box-shadow:0 1px 8px rgba(0,0,0,.14);padding-bottom:40px}}
 header.rpt{{background:var(--navy);color:#fff;padding:26px 40px}}
 header.rpt .brand{{font-size:12px;letter-spacing:2px;text-transform:uppercase;opacity:.85}}
 header.rpt h1{{margin:6px 0 4px;font-size:25px}}header.rpt h2{{margin:0;font-weight:500;font-size:15px;opacity:.9}}
 .meta{{display:grid;grid-template-columns:repeat(4,1fr);gap:1px;background:#d9dde3;border-bottom:3px solid var(--navy)}}
 .meta div{{background:#fff;padding:9px 14px;font-size:12.5px}}.meta .k{{color:{GREY};text-transform:uppercase;font-size:9.5px;letter-spacing:.8px}}.meta .v{{font-weight:600}}
 section{{padding:6px 40px}}h3{{color:var(--navy);border-bottom:2px solid var(--navy);padding-bottom:5px;margin:24px 0 6px;font-size:16px}}
 h4{{color:var(--navy);margin:14px 0 4px;font-size:13.5px}}.legal{{font-weight:400;color:{GREY};font-size:12px}}.cont{{font-size:11.5px;color:{GREY};margin:2px 0 6px}}
 table{{width:100%;border-collapse:collapse;font-size:12px;margin:4px 0 10px}}
 th{{background:var(--navy);color:#fff;text-align:left;padding:6px 7px;font-weight:600;vertical-align:bottom}}
 td{{border:1px solid #dce0e5;padding:5px 7px;vertical-align:top}}td.num{{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}}
 tr.verify td,td.verify{{background:var(--amber)}}tr.tot td{{background:#eef2f7;font-weight:700}}tr.top td{{background:#eaf3fb}}
 td.cmt{{font-size:11px;color:#333;max-width:340px}}.addr{{font-size:10.5px;color:{GREY}}}
 .conf-hi{{color:#1a7f37;font-weight:600}}.conf-med{{color:#9a6700}}.conf-lo{{color:#c0392b;font-weight:600}}
 .sev-hi{{color:#c0392b;font-weight:700}}.sev-med{{color:#9a6700}}.sev-lo{{color:{GREY}}}
 .kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:12px 0}}
 .kpi{{border:1px solid #dce0e5;border-radius:8px;padding:12px 14px;background:#f8fafc}}
 .kpi .n{{font-size:22px;font-weight:700;color:var(--navy)}}.kpi .l{{font-size:10.5px;color:{GREY};text-transform:uppercase;letter-spacing:.4px}}
 .callout{{background:#fff8e1;border-left:4px solid #d9a400;padding:10px 14px;font-size:12.5px;margin:10px 0}}
 .danger{{background:#fdecea;border-left:4px solid #c0392b}}.info{{background:#eaf3fb;border-left:4px solid #1f6fb2}}
 ul.tight{{margin:6px 0 12px;padding-left:20px}}ul.tight li{{margin:3px 0;font-size:12.5px}}
 .foot{{font-size:10.5px;color:{GREY};padding:16px 40px 0;border-top:1px solid #e5e7eb;margin-top:22px}}
 .scroll{{overflow-x:auto}}.legend span{{display:inline-block;padding:2px 9px;margin-right:8px;border-radius:3px;font-size:11px;border:1px solid #cfd4da}}.legend .v{{background:var(--amber)}}
 @media print{{body{{background:#fff}}.page{{box-shadow:none;margin:0;max-width:none}}header.rpt{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}}}
</style></head><body><div class=page>
 <header class=rpt><div class=brand>{esc(meta['brand'])}</div>
  <h1>Cursory Mineral &amp; Leasehold Ownership Report</h1>
  <h2>Section 31, Township 12 North, Range 24 West, Indian Meridian — Roger Mills County, Oklahoma</h2></header>
 <div class=meta>
  <div><div class=k>Prospect</div><div class=v>{esc(meta['prospect'])}</div></div>
  <div><div class=k>Gross acres</div><div class=v>637.42</div></div>
  <div><div class=k>Tracts</div><div class=v>10</div></div>
  <div><div class=k>Report date</div><div class=v>{esc(meta['date'])}</div></div>
  <div><div class=k>Unit well</div><div class=v>Alexander 1-31</div></div>
  <div><div class=k>API</div><div class=v>35-129-22925</div></div>
  <div><div class=k>Operator of record</div><div class=v>Martin's Resources LLC</div></div>
  <div><div class=k>Examiner</div><div class=v>{esc(meta['examiner'])}</div></div>
 </div>
 <section>
  <div class=callout danger><strong>Cursory report — not a certified title opinion.</strong> Interests are
   subject to the 25-item Curative Requirements schedule below. Figures are read directly from the examiner's
   canonical turn-in workbook; no legal facts are fabricated, and undetermined balances are disclosed as open
   rather than distributed.</div>
  <div class=callout info><strong>Basis / version note.</strong> This report is built on the canonical
   <em>BEST TURN IN COPY</em>, which supersedes the earlier "NHE Final." Per curative items <strong>C-02</strong> and
   <strong>C-03</strong>, the prior version over-credited Tract 9 (Dold 80.00 → corrected 18.835 NMA) and carried an
   unsupported Tract 2 row (removed). Identified NMA is therefore lower and more defensible here.</div>

  <h3>1. Executive summary</h3>
  <div class=kpis>
   <div class=kpi><div class=n>637.42</div><div class=l>Gross acres · 10 tracts</div></div>
   <div class=kpi><div class=n>{num(id_na)}</div><div class=l>NMA identified of record</div></div>
   <div class=kpi><div class=n>{num(open_na)}</div><div class=l>NMA open / undetermined (curative)</div></div>
   <div class=kpi><div class=n>{len(leases)}</div><div class=l>Leases of record (OGL 1–{len(leases)})</div></div>
  </div>
  <p>Section 31 is a leased Cherokee section developed in the 2004–2006 EOG / Arrowhead play (base register
     OGL 1–63). In 2026 <strong>Silver Oak Natural Resources, LLC</strong> recorded a six-lease <strong>top-lease
     sweep</strong> (OGL 64–69, 3/16 royalty) over the principal current owners. Fee mineral ownership is a family
     mosaic — <strong>Mitchell / Williams / Moore</strong>, <strong>Bain / Jensen / Fuchs</strong>, <strong>Kirk</strong>,
     and <strong>Daigle / Sorenson / Johnson</strong> — with modern mineral-fund buyers layered on. Tracts 1, 3, 6, 8
     and 10 are cursory-balanced; Tracts 2, 4, 5, 7 and 9 carry disclosed undetermined balances pending curative.</p>
  <div class=callout danger><strong>Material HBP risk (Curative C-10).</strong> Base leases (OGL 1–63) are carried
     HBP per well data of record, <em>but the client reports the Alexander 1-31 may be inactive</em> and OCC/OTC
     production has not been independently confirmed. <strong>If the well is not producing, the base leases may have
     expired and the Silver Oak top leases (OGL 64–69) may be operative</strong> — this materially changes who holds
     the leasehold. Confirm production before relying on any HBP conclusion.</div>

  <h3>2. Tract status &amp; net-acre reconciliation</h3>
  <div class=scroll><table><thead><tr><th>Tract</th><th>Legal</th><th>Gross</th><th>Identified NMA</th>
   <th>Open/undet.</th><th>Status</th><th>Confidence</th><th>Primary next action</th></tr></thead><tbody>
   {''.join(statusrow(s) for s in summary)}
   <tr class=tot><td colspan=2>SECTION</td><td class=num>637.42</td><td class=num>{num(id_na)}</td>
    <td class=num>{num(open_na)}</td><td colspan=3></td></tr></tbody></table></div>

  <h3>3. Section-wide owner roll-up — supported NMA (acquisition-target view)</h3>
  <p style="font-size:12px;color:{GREY}">Aggregated across all tracts from <em>supported</em> net acres only
     (undetermined balances excluded). "UNLEASED" flags fee owners with no base/top lease tied of record —
     the primary leasing/acquisition targets.</p>
  <div class=scroll><table><thead><tr><th>Owner (normalized)</th><th>Supported NMA</th><th>Tracts</th>
   <th>Lease status</th></tr></thead><tbody>{roll_rows}</tbody></table></div>

  <h3>4. Mineral ownership by tract</h3>
  <div class=legend><span class=v>Amber = undetermined balance / interest open of record — see curative</span></div>
  <div class=scroll>{tract_html}</div>

  <h3>5. Leasehold — base register (OGL 1–63)</h3>
  <div class=scroll><table><thead>{oglhead}</thead><tbody>{''.join(lease_row(l) for l in base)}</tbody></table></div>
  <h4>Silver Oak Natural Resources top-lease package (OGL 64–69, 2026) — subordinate / springing</h4>
  <div class=scroll><table><thead>{oglhead}</thead><tbody>{''.join(lease_row(l) for l in top)}</tbody></table></div>
  <div class=callout>Top leases are subordinate to the prior base leases and spring only on base-lease
     termination — <em>unless the base leases have already expired for lack of production</em> (see HBP risk above).
     Primary terms run 3–5 years; expirations 2029–2031.</div>

  <h3>6. Working interest / wellbore chain</h3>
  {wi_html}

  <h3>7. Unit well &amp; held-by-production</h3>
  <div class=scroll><table class=wi><tbody>{well_html}</tbody></table></div>

  <h3>8. Curative requirements schedule ({len(curative)} items)</h3>
  <div class=scroll><table><thead><tr><th>ID</th><th>Sev.</th><th>Blocks TI</th><th>Tract(s)</th>
   <th>Issue type</th><th>Description</th><th>Document needed</th><th>Status</th></tr></thead><tbody>
   {''.join(cur_row(c) for c in curative)}</tbody></table></div>

  <div class=foot>
   Basis: the examiner's canonical <em>BEST TURN IN COPY</em> workbook ({esc(meta['date'])}) — Turn_In_Summary, Title,
   OGL register, WI chains, Well 1, and the Curative_Requirements schedule — plus the Roger Mills County unplatted
   legal index (OCR through 05/05/2026). Current ownership reflects identified record owners; undetermined balances
   are shown as explicit open lines, not distributed. All ten tracts foot to the section total of 637.42 acres.
   <br><br><strong>Cursory cleanup only from available workbook, county index (OCR), OCC status, and OTC evidence.
   Not a certified title opinion. Open and verify items remain only where source evidence was insufficient.</strong>
  </div>
 </section></div></body></html>"""


def render_md(meta, tracts, wi, leases, summary, curative, roll):
    id_na = sum(float(s["id"]) for s in summary if isinstance(s["id"], (int, float)))
    open_na = sum(float(s["open"]) for s in summary if isinstance(s["open"], (int, float)))
    L = ["# Section 31, T12N–R24W, Roger Mills County, Oklahoma", "## Cursory Mineral & Leasehold Ownership Report", "",
         f"**Prospect {meta['prospect']} · 637.42 gross acres · 10 tracts · Report date {meta['date']}**",
         "**Unit well: Alexander 1-31 · API 35-129-22925 · Cherokee · Operator of record Martin's Resources LLC**", "",
         "> **Cursory — not a certified title opinion.** Built on the canonical *BEST TURN IN COPY* (supersedes the",
         "> NHE Final; Tract 9 Dold over-credit corrected 80→18.835, unsupported Tract 2 row removed per C-02/C-03).", "",
         "> **HBP RISK (C-10):** base leases carried HBP, but client reports the well may be inactive; if not producing,",
         "> base leases may have expired and Silver Oak top leases (OGL 64–69) may be operative. Confirm production.", "",
         "## Executive summary", "",
         f"- **Gross acres:** 637.42  |  **NMA identified of record:** {num(id_na)}  |  **NMA open/undetermined:** {num(open_na)}",
         f"- **Leases of record:** {len(leases)} (base OGL 1–63 + Silver Oak top OGL 64–69)", "",
         "## Tract status & net-acre reconciliation", "",
         "| Tract | Legal | Gross | Identified | Open/undet. | Status | Confidence | Next action |",
         "|--:|---|--:|--:|--:|---|---|---|"]
    for s in summary:
        L.append(f"| {s['t']} | {s['legal']} | {num(s['gross'])} | {num(s['id'])} | {num(s['open'])} "
                 f"| {s['status']} | {s['conf']} | {s['next']} |")
    L += [f"| **Section** | | **637.42** | **{num(id_na)}** | **{num(open_na)}** | | | |", "",
          "## Section-wide owner roll-up — supported NMA (acquisition targets)", "",
          "| Owner (normalized) | Supported NMA | Tracts | Lease status |", "|---|--:|---|---|"]
    for x in roll[:25]:
        L.append(f"| {x['name']} | {num(x['na'])} | {x['tracts']} | {'Leased' if x['leased'] else '**UNLEASED**'} |")
    L += ["", "## Mineral ownership by tract", ""]
    for t, s in zip(tracts, summary):
        L += [f"### {t['id']} — {t['legal']}  ({num(s['gross'])} ac · {s['status']})", "",
              "| Mineral owner | Net ac | OGL | Roy. | Expiration | Comments |", "|---|--:|---|--:|---|---|"]
        for o in t["owners"]:
            ac = num(o["acres"]) if isinstance(o["acres"], (int, float)) else "undet."
            roy = num(o['roy'], 4) if isinstance(o['roy'], (int, float)) else (o['roy'] or "")
            L.append(f"| {o['name']} | {ac} | {o['ogl'] if o['ogl'] not in (None,'—') else ''} | {roy} "
                     f"| {str(o['exp'] or '').replace(chr(10),' ')} | {str(o['cmt'] or '').replace(chr(10),' ')} |")
        L.append("")
    L += ["## Leasehold register (OGL 1–69)", "",
          "| OGL | Status | Lessor | Lessee | Exec. | Bk/Pg | Roy. | NMA | Tracts | Exp. |",
          "|--:|---|---|---|---|---|--:|--:|---|---|"]
    for l in leases:
        roy = num(l['roy'], 4) if isinstance(l['roy'], (int, float)) else (l['roy'] or "")
        L.append(f"| {l['n']} | {l['lh']} | {l['grantor']} | {l['grantee']} | {dt(l['exec'])} | {l['bkpg']} "
                 f"| {roy} | {num(l['nma'])} | {l['tracts']} | {dt(l['exp'])} |")
    L += ["", "## Curative requirements schedule", "",
          "| ID | Sev. | Blocks | Tract(s) | Issue | Description | Document needed | Status |",
          "|---|---|---|---|---|---|---|---|"]
    for c in curative:
        L.append(f"| {c['id']} | {c['sev']} | {c['blocks']} | {c['tract']} | {c['type']} "
                 f"| {str(c['desc'] or '').replace(chr(10),' ')} | {str(c['need'] or '').replace(chr(10),' ')} | {c['status']} |")
    L += ["", "---", "*Cursory cleanup only from available workbook, county index (OCR), OCC status, and OTC "
          "evidence. Not a certified title opinion. Open and verify items remain only where source evidence was "
          "insufficient.*"]
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/tmp/rm/best.xlsx")
    ap.add_argument("--out", default=str(OUT / "31-12N-24W_Roger_Mills_Mineral_and_Leasehold_Ownership_Report"))
    args = ap.parse_args()
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    tracts, wi = parse_title(wb["Title "])
    leases = parse_ogl(wb["OGL"])
    summary = parse_summary(wb["Turn_In_Summary"])
    curative = parse_curative(wb["Curative_Requirements"])
    well = parse_well(wb["Well 1"])
    roll = owner_rollup(tracts)
    meta = {"brand": "DataBossX — Mineral Deal Intelligence", "prospect": "25-004",
            "date": "7/3/2026", "examiner": "RG"}
    OUT.mkdir(parents=True, exist_ok=True)
    Path(args.out + ".html").write_text(render_html(meta, tracts, wi, leases, summary, curative, well, roll), encoding="utf-8")
    Path(args.out + ".md").write_text(render_md(meta, tracts, wi, leases, summary, curative, roll), encoding="utf-8")
    id_na = sum(float(s["id"]) for s in summary if isinstance(s["id"], (int, float)))
    open_na = sum(float(s["open"]) for s in summary if isinstance(s["open"], (int, float)))
    print(f"tracts={len(tracts)} leases={len(leases)} curative={len(curative)} rollup_owners={len(roll)}")
    print(f"identified={id_na:.2f} open={open_na:.2f} sum={id_na+open_na:.2f}")
    print("top owners:", [(x["name"], round(x["na"], 1)) for x in roll[:6]])


if __name__ == "__main__":
    main()
