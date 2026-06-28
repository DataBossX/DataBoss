#!/usr/bin/env python3
"""recalc.py — headless LibreOffice recalculation + error scanner for xlsx workbooks.

Usage:
    python scripts/recalc.py <workbook.xlsx> [timeout_seconds]

Drives a headless LibreOffice instance to fully recalculate every formula in the
workbook (forces a hard recalc, which openpyxl's cached values cannot do), then
re-reads the recalculated values and scans every sheet for Excel error literals
and RECHECK flags.

Prints a JSON report to stdout:
    {
      "status": "ok" | "error",
      "workbook": "...",
      "total_errors": <int>,
      "error_summary": { "#VALUE!": [ "WI 1!C10", ... ], ... },
      "recheck_cells": [ "Tract 2!H7", ... ],
      "sheets": <int>
    }

The workbook is recalculated in place (saved back) so downstream tooling reads
fresh cached values.
"""
import json
import os
import shutil
import subprocess
import sys
import tempfile

ERROR_LITERALS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


# registrymodifications.xcu forcing "Always recalculate" on load for both ODF
# and OOXML (xlsx) documents. Without this LibreOffice leaves xlsx formulas at
# their stale cached values and convert-to performs no real recalculation.
RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry"
 xmlns:xs="http://www.w3.org/2001/XMLSchema"
 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
</oor:items>
"""


def _seed_profile(profile):
    """Write registrymodifications.xcu that forces full recalc on load."""
    reg_dir = os.path.join(profile, "user")
    os.makedirs(reg_dir, exist_ok=True)
    with open(os.path.join(reg_dir, "registrymodifications.xcu"), "w") as fh:
        fh.write(RECALC_XCU)


def recalc_with_libreoffice(path, timeout):
    """Open in headless LibreOffice with full recalc enabled, save back as xlsx."""
    src = os.path.abspath(path)
    workdir = tempfile.mkdtemp(prefix="recalc_")
    # A user profile that forces "always recalculate" on load.
    profile = os.path.join(workdir, "profile")
    os.makedirs(profile, exist_ok=True)
    _seed_profile(profile)
    out_dir = os.path.join(workdir, "out")
    os.makedirs(out_dir, exist_ok=True)
    env = dict(os.environ)
    env["HOME"] = workdir
    cmd = [
        "soffice", "--headless", "--norestore", "--invisible",
        "--nodefault", "--nofirststartwizard", "--nologo",
        "-env:UserInstallation=file://" + profile,
        "--calc", "--convert-to", "xlsx:Calc MS Excel 2007 XML",
        "--outdir", out_dir, src,
    ]
    try:
        subprocess.run(cmd, env=env, timeout=timeout, check=True,
                       capture_output=True)
    except subprocess.TimeoutExpired:
        shutil.rmtree(workdir, ignore_errors=True)
        return None, "libreoffice recalc timed out"
    except subprocess.CalledProcessError as e:
        shutil.rmtree(workdir, ignore_errors=True)
        return None, "libreoffice failed: " + e.stderr.decode("utf-8", "replace")[:500]
    produced = os.path.join(out_dir, os.path.basename(src))
    if not os.path.exists(produced):
        # convert-to may have changed the name; grab the single xlsx in out_dir
        xs = [f for f in os.listdir(out_dir) if f.lower().endswith(".xlsx")]
        if not xs:
            shutil.rmtree(workdir, ignore_errors=True)
            return None, "no recalculated workbook produced"
        produced = os.path.join(out_dir, xs[0])
    shutil.copyfile(produced, src)
    shutil.rmtree(workdir, ignore_errors=True)
    return src, None


def scan(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    error_summary = {}
    recheck_cells = []
    total = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str):
                    if v in ERROR_LITERALS:
                        coord = "%s!%s" % (ws.title, c.coordinate)
                        error_summary.setdefault(v, []).append(coord)
                        total += 1
                    elif "RECHECK" in v.upper():
                        recheck_cells.append("%s!%s" % (ws.title, c.coordinate))
    wb.close()
    return total, error_summary, recheck_cells, len(wb.sheetnames)


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "usage: recalc.py <file> [timeout]"}))
        return 1
    path = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 120
    if not os.path.exists(path):
        print(json.dumps({"status": "error", "message": "file not found: " + path}))
        return 1
    _, err = recalc_with_libreoffice(path, timeout)
    if err:
        print(json.dumps({"status": "error", "message": err}))
        return 1
    total, summary, rechecks, sheets = scan(path)
    # Trim very long location lists for readability but keep counts accurate.
    trimmed = {k: (v if len(v) <= 50 else v[:50] + ["...(%d more)" % (len(v) - 50)])
               for k, v in summary.items()}
    print(json.dumps({
        "status": "ok",
        "workbook": os.path.basename(path),
        "total_errors": total,
        "error_summary": trimmed,
        "recheck_cells": rechecks,
        "sheets": sheets,
    }, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
