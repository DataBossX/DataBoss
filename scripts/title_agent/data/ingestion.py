"""Workbook ingestion — Phase 4.

``WorkbookMapper`` reads an .xlsx **non-destructively** (openpyxl read-only,
values only — it never writes, and the caller keeps the original untouched)
and returns a structured map of the report's topology: the 10 tracts, the OGL
register, the runsheets, the Working-Interest sheets, and the Alexander 1-31
well tab. The validators in Phase 5 consume this map rather than re-parsing the
workbook themselves.

Classification is keyword-first on the sheet title so that, e.g., the
"Alexander 1-31" well tab is never mistaken for a tract just because it
contains digits. Header rows are found heuristically (the densest text row in
the first several rows), which tolerates the title/blank rows real title
workbooks carry above their tables.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Optional

import openpyxl

from ..core.config import SECTION, TRACT_COUNT, WELL_NAME


class SheetKind(str, Enum):
    TRACT = "tract"
    OGL_REGISTER = "ogl_register"
    RUNSHEET = "runsheet"
    WORKING_INTEREST = "working_interest"
    WELL = "well"
    OTHER = "other"


@dataclass(frozen=True)
class SheetProfile:
    """Structural summary of one worksheet."""

    name: str
    kind: SheetKind
    n_rows: int
    n_cols: int
    header_row: Optional[int]
    headers: tuple[str, ...]
    tract_number: Optional[int] = None


@dataclass
class WorkbookTopology:
    """The mapped structure of a title report workbook."""

    path: Path
    section: str
    sheets: list[SheetProfile] = field(default_factory=list)

    def by_kind(self, kind: SheetKind) -> list[SheetProfile]:
        return [s for s in self.sheets if s.kind == kind]

    def tracts(self) -> list[SheetProfile]:
        """Tract sheets ordered by tract number."""
        return sorted(
            self.by_kind(SheetKind.TRACT),
            key=lambda s: s.tract_number if s.tract_number is not None else 1_000,
        )

    def ogl_register(self) -> Optional[SheetProfile]:
        found = self.by_kind(SheetKind.OGL_REGISTER)
        return found[0] if found else None

    def runsheets(self) -> list[SheetProfile]:
        return self.by_kind(SheetKind.RUNSHEET)

    def working_interest_sheets(self) -> list[SheetProfile]:
        return self.by_kind(SheetKind.WORKING_INTEREST)

    def well_tab(self) -> Optional[SheetProfile]:
        found = self.by_kind(SheetKind.WELL)
        return found[0] if found else None

    def found_tract_numbers(self) -> list[int]:
        return sorted(
            s.tract_number for s in self.by_kind(SheetKind.TRACT) if s.tract_number is not None
        )

    def missing_tracts(self, expected: int = TRACT_COUNT) -> list[int]:
        """Which of tracts 1..expected are absent — a completeness signal the
        loop uses to decide the workbook is not yet certifiable."""
        present = set(self.found_tract_numbers())
        return [n for n in range(1, expected + 1) if n not in present]


class WorkbookMapper:
    """Maps and reads title report workbooks without modifying them."""

    # Keyword patterns, checked in priority order. The first match wins, so
    # the well and register tabs are classified before the generic tract rule.
    _WELL_PATTERNS = (
        re.compile(re.escape(WELL_NAME), re.I),
        re.compile(r"\balexander\b", re.I),
        re.compile(r"\b1-31\b"),
        re.compile(r"\bwell\b", re.I),
    )
    _OGL_PATTERNS = (
        re.compile(r"\bogl\b", re.I),
        re.compile(r"lease\s*register", re.I),
        re.compile(r"oil\s*&?\s*and?\s*gas\s*lease", re.I),
    )
    _WI_PATTERNS = (
        re.compile(r"working\s*interest", re.I),
        re.compile(r"\bw\.?i\.?\b", re.I),
    )
    _RUNSHEET_PATTERNS = (
        re.compile(r"run\s*sheet", re.I),
        re.compile(r"\bchain\b", re.I),
    )
    _TRACT_PATTERN = re.compile(r"\btr(?:act)?\.?\s*[-#]?\s*0*(\d{1,2})\b", re.I)

    _HEADER_SCAN_ROWS = 15  # look this far down for the header row

    def map(self, path: Path) -> WorkbookTopology:
        """Return the topology of ``path``. Read-only; the file is not touched."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            topology = WorkbookTopology(path=Path(path), section=SECTION)
            for ws in wb.worksheets:
                topology.sheets.append(self._profile_sheet(ws))
            return topology
        finally:
            wb.close()

    def read_grid(self, path: Path, sheet_name: str) -> list[list[Any]]:
        """Return every cell value of ``sheet_name`` as a row-major grid.
        Read-only. Raises KeyError if the sheet is absent."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                raise KeyError(f"sheet not found: {sheet_name!r}")
            ws = wb[sheet_name]
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

    # ---- internals ------------------------------------------------------- #

    def _profile_sheet(self, ws: Any) -> SheetProfile:
        # read_only worksheets expose max_row/max_column after a pass; iterate
        # a bounded window to find the header without loading the whole grid.
        window: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            window.append(list(row))
            if i + 1 >= self._HEADER_SCAN_ROWS:
                break

        kind, tract_number = self._classify(ws.title)
        header_row, headers = self._detect_header(window)
        n_rows = int(ws.max_row or 0)
        n_cols = int(ws.max_column or 0)
        return SheetProfile(
            name=ws.title,
            kind=kind,
            n_rows=n_rows,
            n_cols=n_cols,
            header_row=header_row,
            headers=tuple(headers),
            tract_number=tract_number,
        )

    def _classify(self, title: str) -> tuple[SheetKind, Optional[int]]:
        for pat in self._WELL_PATTERNS:
            if pat.search(title):
                return SheetKind.WELL, None
        for pat in self._OGL_PATTERNS:
            if pat.search(title):
                return SheetKind.OGL_REGISTER, None
        for pat in self._WI_PATTERNS:
            if pat.search(title):
                return SheetKind.WORKING_INTEREST, None
        for pat in self._RUNSHEET_PATTERNS:
            if pat.search(title):
                return SheetKind.RUNSHEET, None
        m = self._TRACT_PATTERN.search(title)
        if m:
            return SheetKind.TRACT, int(m.group(1))
        return SheetKind.OTHER, None

    def _detect_header(self, window: list[list[Any]]) -> tuple[Optional[int], list[str]]:
        """Pick the densest text row in the window as the header. Returns a
        1-based row index and the trimmed header labels."""
        best_idx: Optional[int] = None
        best_score = 0
        best_row: list[Any] = []
        for idx, row in enumerate(window):
            score = sum(1 for c in row if isinstance(c, str) and c.strip())
            if score > best_score:
                best_score = score
                best_idx = idx
                best_row = row
        if best_idx is None or best_score < 2:
            return None, []
        headers = [str(c).strip() for c in best_row if c is not None and str(c).strip()]
        return best_idx + 1, headers
