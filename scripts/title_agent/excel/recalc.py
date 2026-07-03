"""LibreOffice headless recalculation and error scanning — Phase 3 (recalc).

Gate 6 (Execution) requires that a real spreadsheet engine load the workbook,
recompute every formula, and produce zero ``#REF!`` / ``#N/A`` / OOM errors.
openpyxl does not evaluate formulas, so we drive LibreOffice headless.

``LibreOfficeEngine.run_headless_recalc`` converts the workbook to a fresh
xlsx via ``soffice --headless --convert-to``. Because xml_surgeon set
``fullCalcOnLoad="1"``, LibreOffice recomputes the whole book on load; the
converted copy therefore carries freshly evaluated values (and any error
literals) which ``scan_for_errors`` then reads back with openpyxl.

The "single-sheet isolated mode" from the blueprint is honored by an optional
``sheet_name``: when given, we first extract that one sheet into a throwaway
workbook and recalc only it, so a multi-hundred-MB grid book cannot OOM the
converter. When absent, the whole book is recalculated.

If ``soffice`` is not installed the engine raises ``LibreOfficeUnavailable``
rather than silently passing the gate — a gate can never be cleared by
pretending the check ran.
"""

from __future__ import annotations

import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

# Error literals a spreadsheet engine emits into cells.
_ERROR_LITERALS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


class LibreOfficeUnavailable(RuntimeError):
    """soffice/libreoffice is not on PATH — Gate 6 cannot be evaluated."""


class RecalcError(RuntimeError):
    """The headless conversion failed (non-zero exit, timeout, or no output)."""


@dataclass(frozen=True)
class CellError:
    sheet: str
    cell: str
    value: str


@dataclass
class RecalcResult:
    """Outcome of a recalculation pass."""

    recalculated_path: Path
    errors: list[CellError] = field(default_factory=list)
    oom: bool = False

    @property
    def clean(self) -> bool:
        return not self.errors and not self.oom


class LibreOfficeEngine:
    """Headless LibreOffice recalculation with error detection."""

    _CANDIDATES = ("soffice", "libreoffice")

    def __init__(self, timeout_seconds: int = 180) -> None:
        self._timeout = timeout_seconds

    @classmethod
    def is_available(cls) -> bool:
        """True if the soffice/libreoffice binary is on PATH.

        Note this only checks the binary exists; some sandboxed containers ship
        a soffice that runs ``--version`` but cannot load documents. Use
        ``conversion_works`` for a functional check before relying on Gate 6.
        """
        return cls._locate() is not None

    @classmethod
    def conversion_works(cls) -> bool:
        """Probe whether headless conversion actually functions here.

        Does a trivial CSV -> xlsx round-trip. Returns False if soffice is
        absent OR present-but-non-functional (as in some locked-down CI
        sandboxes). This is what the loop should gate on: a broken engine must
        surface as "Gate 6 not evaluable", never as a silent pass.
        """
        soffice = cls._locate()
        if soffice is None:
            return False
        try:
            with tempfile.TemporaryDirectory(prefix="title_agent_probe_") as d:
                work = Path(d)
                csv = work / "probe.csv"
                csv.write_text("a,b\n1,2\n", encoding="utf-8")
                out = work / "out"
                out.mkdir()
                proc = subprocess.run(
                    [
                        soffice, "--headless",
                        f"-env:UserInstallation=file://{work / 'profile'}",
                        "--convert-to", "xlsx", "--outdir", str(out), str(csv),
                    ],
                    capture_output=True, text=True, timeout=60,
                    env=cls._probe_env(work),
                )
                return proc.returncode == 0 and (out / "probe.xlsx").exists()
        except (subprocess.SubprocessError, OSError):
            return False

    @staticmethod
    def _probe_env(work_dir: Path) -> dict[str, str]:
        import os

        env = dict(os.environ)
        env["HOME"] = str(work_dir)
        return env

    @classmethod
    def _locate(cls) -> Optional[str]:
        for name in cls._CANDIDATES:
            path = shutil.which(name)
            if path:
                return path
        return None

    def run_headless_recalc(
        self, file_path: Path, sheet_name: Optional[str] = None
    ) -> RecalcResult:
        """Recalculate ``file_path`` and return the errors found.

        With ``sheet_name`` the recalc is isolated to that one sheet in a
        throwaway workbook (OOM-safe for huge books). The original file is
        never modified; output lands in a temp directory the caller need not
        manage.
        """
        soffice = self._locate()
        if soffice is None:
            raise LibreOfficeUnavailable(
                "LibreOffice (soffice) not found on PATH; cannot evaluate Gate 6."
            )
        if not file_path.exists():
            raise RecalcError(f"workbook does not exist: {file_path}")

        work_dir = Path(tempfile.mkdtemp(prefix="title_agent_recalc_"))
        source = file_path
        if sheet_name is not None:
            source = self._isolate_sheet(file_path, sheet_name, work_dir)

        out_dir = work_dir / "out"
        out_dir.mkdir(parents=True, exist_ok=True)
        try:
            proc = subprocess.run(
                [
                    soffice,
                    "--headless",
                    "--calc",
                    "--convert-to",
                    "xlsx:Calc MS Excel 2007 XML",
                    "--outdir",
                    str(out_dir),
                    str(source),
                ],
                capture_output=True,
                text=True,
                timeout=self._timeout,
                # A dedicated profile dir avoids clashing with a desktop
                # LibreOffice instance and keeps runs reproducible.
                env=self._headless_env(work_dir),
            )
        except subprocess.TimeoutExpired:
            # A timeout on a large grid is our OOM/too-big signal.
            return RecalcResult(recalculated_path=source, errors=[], oom=True)

        produced = out_dir / (source.stem + ".xlsx")
        if proc.returncode != 0 or not produced.exists():
            raise RecalcError(
                f"headless recalc failed (rc={proc.returncode}): "
                f"{proc.stderr.strip()[:300]}"
            )

        errors = self.scan_for_errors(produced)
        return RecalcResult(recalculated_path=produced, errors=errors, oom=False)

    def scan_for_errors(self, recalculated_path: Path) -> list[CellError]:
        """Read every cell of a recalculated workbook and collect error
        literals. Uses ``data_only`` so cached values (post-recalc) are read,
        not the formulas."""
        errors: list[CellError] = []
        wb = openpyxl.load_workbook(recalculated_path, data_only=True, read_only=True)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if isinstance(val, str) and val in _ERROR_LITERALS:
                            errors.append(
                                CellError(sheet=ws.title, cell=cell.coordinate, value=val)
                            )
        finally:
            wb.close()
        return errors

    def _isolate_sheet(self, file_path: Path, sheet_name: str, work_dir: Path) -> Path:
        """Copy a single sheet into a minimal workbook to bound memory use."""
        src = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
        try:
            if sheet_name not in src.sheetnames:
                raise RecalcError(f"sheet not found: {sheet_name!r}")
            keep = set(src.sheetnames) - {sheet_name}
            for name in keep:
                del src[name]
            isolated = work_dir / f"{file_path.stem}__{sheet_name}.xlsx"
            src.save(isolated)
            return isolated
        finally:
            src.close()

    @staticmethod
    def _headless_env(work_dir: Path) -> dict[str, str]:
        import os

        env = dict(os.environ)
        profile = (work_dir / "lo_profile").as_uri()
        env["HOME"] = str(work_dir)  # keep LO from writing to the real HOME
        env["UserInstallation"] = profile
        return env
