"""Tests for config-file ColumnMap loading (core/wiring.py).

Lets an operator retarget the workbook layout via a JSON file instead of
editing Python — the fix for what the `map` preflight reveals.
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import openpyxl

from ..core.wiring import ColumnMap, WorkbookGateSuite


def test_from_mapping_overrides_only_named_fields(tmp: Path) -> None:
    cm = ColumnMap.from_mapping({"grantor": ["from party", "grantor"]})
    assert cm.grantor == ("from party", "grantor")
    # Unspecified fields keep their defaults.
    assert cm.grantee == ("grantee",)


def test_from_mapping_rejects_unknown_field(tmp: Path) -> None:
    cases: list[dict[str, object]] = [
        {"grantorr": ["x"]}, {"grantor": "not-a-list"}, {"grantor": []}]
    for bad in cases:
        try:
            ColumnMap.from_mapping(bad)  # type: ignore[arg-type]
        except ValueError:
            continue
        raise AssertionError(f"expected ValueError for {bad!r}")


def test_from_json_roundtrip(tmp: Path) -> None:
    f = tmp / "columnmap.json"
    f.write_text(json.dumps({"grantor": ["from"], "grantee": ["to"]}), encoding="utf-8")
    cm = ColumnMap.from_json(f)
    assert cm.grantor == ("from",) and cm.grantee == ("to",)


def test_custom_map_makes_a_nonstandard_layout_certifiable(tmp: Path) -> None:
    # A workbook whose chain columns are "From"/"To" (not Grantor/Grantee) is
    # blocked under defaults but fully mapped once the operator supplies aliases.
    p = tmp / "wb.xlsx"
    _build_from_to(p)

    default_report = WorkbookGateSuite().coverage(p)
    assert not default_report.ok  # Gate 3 columns unresolved under defaults

    cm = ColumnMap.from_mapping({"grantor": ["from", "grantor"], "grantee": ["to", "grantee"]})
    mapped_report = WorkbookGateSuite(column_map=cm).coverage(p)
    # Chain columns now resolve; no Gate 3 blocks remain.
    assert not any("Gate 3" in b for b in mapped_report.blocking)


def _build_from_to(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    recap = wb.create_sheet("Tract Recap")
    recap.append(["Tract", "Gross Acres", "Interest Fraction", "Net Acres"])
    for t in range(1, 11):
        recap.append([t, 63.742, 1.0, 63.742])
    for t in range(1, 11):
        ws = wb.create_sheet(f"Tract {t}")
        ws.append(["From", "To", "Date"])
        ws.append(["USA", f"Owner {t}A", "1980-01-01"])
        ws.append([f"Owner {t}A", f"Owner {t}B", "1990-01-01"])
    wb.save(path)


def _run_all() -> None:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in tests:
        with tempfile.TemporaryDirectory() as d:
            fn(Path(d))
        print(f"  PASS  {fn.__name__}")
    print(f"\n{len(tests)} passed.")


if __name__ == "__main__":
    _run_all()
