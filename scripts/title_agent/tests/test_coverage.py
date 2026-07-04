"""Tests for the mapping coverage preflight (core/wiring.py).

The preflight exists to prevent a false CERTIFIED: if the workbook's columns
don't map to the validators' inputs, the affected gates would run on no data
and silently pass. Coverage must flag that as blocking.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl

from ..core.wiring import WorkbookGateSuite

ROYALTY = 0.1875
GROSS = 63.742


def _build(path: Path, *, recap_headers=None, tract_headers=None) -> None:
    recap_headers = recap_headers or ["Tract", "Gross Acres", "Interest Fraction", "Net Acres"]
    tract_headers = tract_headers or [
        "Grantor", "Grantee", "Date", "Interest Change", "Book", "Page", "Instrument"]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    recap = wb.create_sheet("Tract Recap")
    recap.append(recap_headers)
    for t in range(1, 11):
        recap.append([t, GROSS, 1.0, GROSS])
    ogl = wb.create_sheet("OGL Register")
    ogl.append(["Tract", "Lessor", "Royalty"])
    wi = wb.create_sheet("Working Interest")
    wi.append(["Tract", "Owner", "WI", "NRI"])
    nri = 0.5 * (1 - ROYALTY)
    for t in range(1, 11):
        ogl.append([t, f"Lessor {t}", ROYALTY])
        wi.append([t, f"Owner {t}A", 0.5, nri])
        wi.append([t, f"Owner {t}B", 0.5, nri])
    width = len(tract_headers)
    for t in range(1, 11):
        ws = wb.create_sheet(f"Tract {t}")
        ws.append(tract_headers)
        # Size each data row to the header width so header detection is not
        # confused by rows wider than the header.
        ws.append(["USA", f"Owner {t}A", "1980-01-01", 0.5, 100 + t, 200 + t, "WD"][:width])
        ws.append([f"Owner {t}A", f"Owner {t}B", "1990-01-01", -0.5, 300 + t, 400 + t, "WD"][:width])
    wb.save(path)


def test_clean_layout_is_fully_covered(tmp: Path) -> None:
    p = tmp / "wb.xlsx"
    _build(p)
    report = WorkbookGateSuite().coverage(p)
    assert report.ok
    assert not report.blocking
    # Every tract + recap + OGL mapped.
    assert all(i.status == "mapped" for i in report.items)


def test_unmapped_recap_columns_block(tmp: Path) -> None:
    p = tmp / "wb.xlsx"
    # Recap with headers that ColumnMap can't resolve (no gross/fraction/net).
    _build(p, recap_headers=["Parcel", "Acreage", "Decimal", "Result"])
    report = WorkbookGateSuite().coverage(p)
    assert not report.ok
    assert any("Gate 2" in b for b in report.blocking)


def test_unmapped_chain_columns_block(tmp: Path) -> None:
    p = tmp / "wb.xlsx"
    # Tract sheets with no Grantor/Grantee columns -> Gate 3 can't run.
    _build(p, tract_headers=["From", "To", "Date", "Change", "Bk", "Pg", "Type"])
    report = WorkbookGateSuite().coverage(p)
    assert not report.ok
    assert any("Gate 3" in b for b in report.blocking)
    # 10 tract sheets each flagged.
    assert sum(1 for i in report.items if i.status == "columns_unresolved") >= 10


def test_missing_optional_columns_are_warnings_not_blocks(tmp: Path) -> None:
    p = tmp / "wb.xlsx"
    # Grantor/Grantee present (Gate 3 ok) but no interest-change or book/page.
    _build(p, tract_headers=["Grantor", "Grantee", "Date"])
    report = WorkbookGateSuite().coverage(p)
    assert report.ok  # not blocking
    # But the skipped gates are surfaced as notes/warnings.
    assert report.warnings
    assert any("Gate 1 skipped" in (i.note or "") for i in report.items)


def _run_all() -> None:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in tests:
        with tempfile.TemporaryDirectory() as d:
            fn(Path(d))
        print(f"  PASS  {fn.__name__}")
    print(f"\n{len(tests)} passed.")


if __name__ == "__main__":
    _run_all()
