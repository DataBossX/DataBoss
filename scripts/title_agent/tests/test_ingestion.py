"""Phase 4 tests for data/ingestion.py.

Builds a realistic multi-tab title workbook and verifies the topology map:
tract detection and numbering, OGL/WI/runsheet/well classification, header
detection, and the missing-tract completeness signal.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl

from ..data.ingestion import SheetKind, WorkbookMapper


def _build(path: Path, *, tracts: int = 10) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Tracts 1..n, each with a couple of title rows above the header.
    for t in range(1, tracts + 1):
        ws = wb.create_sheet(f"Tract {t}")
        ws["A1"] = f"Section {31}-12N-24W  Tract {t}"
        ws["A2"] = None
        ws.append(["Instrument", "Grantor", "Grantee", "Gross Acres", "Net Acres"])
        ws.append(["WD", "USA", "John Doe", 63.742, 63.742])
    # Non-tract tabs.
    ogl = wb.create_sheet("OGL Register")
    ogl.append(["Lessor", "Lessee", "Royalty", "Book", "Page"])
    wi = wb.create_sheet("Working Interest")
    wi.append(["Owner", "WI", "NRI"])
    rs = wb.create_sheet("Runsheet Tract 1")
    rs.append(["Seq", "Grantor", "Grantee", "Book", "Page"])
    well = wb.create_sheet("Alexander 1-31")
    well.append(["Tract", "WI", "NRI", "Royalty"])
    wb.save(path)


def test_maps_ten_tracts(tmp: Path) -> None:
    p = tmp / "report.xlsx"
    _build(p, tracts=10)
    topo = WorkbookMapper().map(p)
    assert topo.found_tract_numbers() == list(range(1, 11))
    assert topo.missing_tracts() == []
    # Ordered by tract number.
    assert [s.tract_number for s in topo.tracts()] == list(range(1, 11))


def test_classifies_special_tabs(tmp: Path) -> None:
    p = tmp / "report.xlsx"
    _build(p)
    topo = WorkbookMapper().map(p)
    ogl = topo.ogl_register()
    assert ogl is not None and ogl.kind == SheetKind.OGL_REGISTER
    assert len(topo.working_interest_sheets()) == 1
    well = topo.well_tab()
    assert well is not None and well.name == "Alexander 1-31"
    # "Runsheet Tract 1" must classify as runsheet, not tract (keyword-first).
    assert any(s.name == "Runsheet Tract 1" for s in topo.runsheets())
    assert all(s.kind != SheetKind.TRACT for s in topo.runsheets())


def test_header_detection_skips_title_rows(tmp: Path) -> None:
    p = tmp / "report.xlsx"
    _build(p)
    topo = WorkbookMapper().map(p)
    tract1 = topo.tracts()[0]
    # Header is row 3 (two title/blank rows above), 5 labelled columns.
    assert tract1.header_row == 3
    assert "Grantor" in tract1.headers and "Net Acres" in tract1.headers


def test_missing_tracts_flagged(tmp: Path) -> None:
    p = tmp / "report.xlsx"
    _build(p, tracts=8)  # tracts 9 and 10 absent
    topo = WorkbookMapper().map(p)
    assert topo.missing_tracts() == [9, 10]


def test_read_grid_is_readonly_and_complete(tmp: Path) -> None:
    p = tmp / "report.xlsx"
    _build(p)
    before = p.read_bytes()
    grid = WorkbookMapper().read_grid(p, "Tract 1")
    assert grid[2][:3] == ["Instrument", "Grantor", "Grantee"]
    assert grid[3][0] == "WD"
    # Mapping/reading did not modify the file on disk.
    assert p.read_bytes() == before


def _run_all() -> None:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in tests:
        with tempfile.TemporaryDirectory() as d:
            fn(Path(d))
        print(f"  PASS  {fn.__name__}")
    print(f"\n{len(tests)} passed.")


if __name__ == "__main__":
    _run_all()
