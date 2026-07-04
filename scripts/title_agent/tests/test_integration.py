"""End-to-end integration for the wired agent (core/wiring.py + core/loop.py).

Builds a synthetic 10-tract workbook to the documented convention and drives
the real PerfectionLoop with the real WorkbookGateSuite and SurgeonRepairer:

* a clean book certifies (gates 1-5),
* a chain gap escalates as TITLE_GAP,
* a missing pro-rata formula auto-repairs and then certifies — using a
  functional recalc stub in place of LibreOffice (non-functional in this
  sandbox) so the closed mint -> repair -> recalc -> re-evaluate loop is proven.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl

from ..core.loop import LoopState, PerfectionLoop
from ..core.memory import SQLiteManager, VersionController
from ..core.wiring import SurgeonRepairer, WorkbookGateSuite
from ..validators.rules import InstrumentLine, SourceProbeResult

ROYALTY = 0.1875
GROSS_PER_TRACT = 63.742  # * 10 == 637.42


class _AllFoundProbe:
    def verify(self, line: InstrumentLine) -> SourceProbeResult:
        return SourceProbeResult(found=True, free_to_view=True, detail="ok")


def _build_workbook(
    path: Path, *, chain_gap_tract: int | None = None, blank_net_tract: int | None = None
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    recap = wb.create_sheet("Tract Recap")
    recap.append(["Tract", "Gross Acres", "Interest Fraction", "Net Acres"])
    for t in range(1, 11):
        net = None if t == blank_net_tract else GROSS_PER_TRACT * 1.0
        recap.append([t, GROSS_PER_TRACT, 1.0, net])

    ogl = wb.create_sheet("OGL Register")
    ogl.append(["Tract", "Lessor", "Royalty"])
    wi = wb.create_sheet("Working Interest")
    wi.append(["Tract", "Owner", "WI", "NRI"])
    nri = 0.5 * (1 - ROYALTY)
    for t in range(1, 11):
        ogl.append([t, f"Lessor {t}", ROYALTY])
        wi.append([t, f"Owner {t}A", 0.5, nri])
        wi.append([t, f"Owner {t}B", 0.5, nri])

    for t in range(1, 11):
        ws = wb.create_sheet(f"Tract {t}")
        ws.append(["Grantor", "Grantee", "Date", "Interest Change", "Book", "Page", "Instrument"])
        second_grantor = "Stranger" if t == chain_gap_tract else f"Owner {t}A"
        ws.append([f"USA", f"Owner {t}A", "1980-01-01", 0.5, 100 + t, 200 + t, "WD"])
        ws.append([second_grantor, f"Owner {t}B", "1990-01-01", -0.5, 300 + t, 400 + t, "WD"])
    wb.save(path)


def _seed(tmp: Path, **build_kwargs):
    mgr = SQLiteManager(tmp / "s.db")
    vc = VersionController(mgr, workbook_dir=tmp / "wb")
    v1 = vc.mint_new_version("TitleReport", reason="ingest initial workbook")
    _build_workbook(v1.file_path, **build_kwargs)
    return mgr, vc


def _stub_recalc(path: Path) -> None:
    """Functional stand-in for LibreOffice: materialize Net = Gross * Fraction
    on the recap sheet (columns B*C -> D), simulating formula evaluation."""
    wb = openpyxl.load_workbook(path)  # formulas load as strings; we overwrite
    ws = wb["Tract Recap"]
    for row in range(2, ws.max_row + 1):
        gross, frac = ws.cell(row, 2).value, ws.cell(row, 3).value
        if isinstance(gross, (int, float)) and isinstance(frac, (int, float)):
            ws.cell(row, 4).value = gross * frac
    wb.save(path)


def test_clean_workbook_certifies(tmp: Path) -> None:
    mgr, vc = _seed(tmp)
    suite = WorkbookGateSuite(source_probe=_AllFoundProbe())
    loop = PerfectionLoop(mgr, "TitleReport", suite, SurgeonRepairer(), versions=vc)
    outcome = loop.run()
    assert outcome.state == LoopState.CERTIFIED, outcome.unresolved_failures
    assert outcome.iterations == 1


def test_chain_gap_escalates(tmp: Path) -> None:
    mgr, vc = _seed(tmp, chain_gap_tract=3)
    suite = WorkbookGateSuite(source_probe=_AllFoundProbe())
    loop = PerfectionLoop(mgr, "TitleReport", suite, SurgeonRepairer(), versions=vc)
    outcome = loop.run()
    assert outcome.state == LoopState.ESCALATED
    cats = {e.category for e in outcome.escalations}
    assert "TITLE_GAP" in cats
    refs = {e.reference for e in outcome.escalations}
    assert any("Tract 3" in r for r in refs)


def test_missing_pro_rata_auto_repairs_then_certifies(tmp: Path) -> None:
    mgr, vc = _seed(tmp, blank_net_tract=1)
    suite = WorkbookGateSuite(source_probe=_AllFoundProbe())
    repairer = SurgeonRepairer(recalc=_stub_recalc)  # functional recalc stub
    loop = PerfectionLoop(mgr, "TitleReport", suite, repairer, versions=vc)
    outcome = loop.run()
    assert outcome.state == LoopState.CERTIFIED, outcome.unresolved_failures
    # v1 had the gap; repair minted v2, recalc filled it, v2 certified.
    assert outcome.final_version is not None
    assert outcome.final_version.version_number == 2
    # The repaired file really carries the value now.
    wb = openpyxl.load_workbook(outcome.final_version.file_path, data_only=True)
    assert abs(wb["Tract Recap"].cell(2, 4).value - GROSS_PER_TRACT) < 1e-9


def _run_all() -> None:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in tests:
        with tempfile.TemporaryDirectory() as d:
            fn(Path(d))
        print(f"  PASS  {fn.__name__}")
    print(f"\n{len(tests)} passed.")


if __name__ == "__main__":
    _run_all()
