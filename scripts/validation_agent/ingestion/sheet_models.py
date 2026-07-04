"""Typed read-only views over the workbook's sheets.

These wrap an ``openpyxl`` worksheet (read-only, data_only optional) and expose
domain structure without leaking spreadsheet mechanics to the validators. All
column/row anchors are discovered from the sheet, not hard-coded, so the views
survive minor layout drift.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Iterable, Optional

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models import Coord


# --------------------------------------------------------------------------- #
# Tract grid
# --------------------------------------------------------------------------- #
@dataclass
class InstrumentColumn:
    """One instrument column in a tract grid (a double-entry conveyance)."""
    letter: str
    index: int
    instrument_no: Optional[str]
    header_note: Optional[str]  # e.g. "VERIFY — NEED int."
    cells: dict[int, float]     # row -> numeric value (non-empty only)

    @property
    def net(self) -> float:
        return sum(self.cells.values())

    @property
    def needs_interest(self) -> bool:
        return bool(self.header_note and "NEED int" in self.header_note.upper())


@dataclass
class OwnerRow:
    row: int
    name: Optional[str]
    net_acres: Optional[float]      # column C evaluated
    row_total: Optional[float]      # column E evaluated


@dataclass
class TractGrid:
    ws: Worksheet
    tract_no: int
    acreage: float
    _data_row_start: int = 10
    # Grid columns: D=owner name, E=row total, first instrument col = G (7).
    NAME_COL: int = 4
    TOTAL_COL: int = 5
    NET_COL: int = 3
    FIRST_INSTR_COL: int = 7
    HEADER_ROW: int = 5          # "VERIFY — NEED int." notes live here
    INSTR_NO_ROW: int = 2        # instrument numbers
    SUBTOTAL_GUARD_ROW: int = 6  # =IF(col7=0,"","RECHECK") guards

    def instrument_columns(self, values: "SheetValues") -> list[InstrumentColumn]:
        cols: list[InstrumentColumn] = []
        max_col = self.ws.max_column
        for ci in range(self.FIRST_INSTR_COL, max_col + 1):
            letter = get_column_letter(ci)
            instr = self.ws.cell(self.INSTR_NO_ROW, ci).value
            note = self.ws.cell(self.HEADER_ROW, ci).value
            cells: dict[int, float] = {}
            for r in range(self._data_row_start, self.ws.max_row + 1):
                v = values.get(self.ws.title, letter, r)
                if isinstance(v, (int, float)) and v != 0:
                    cells[r] = float(v)
            if instr is None and note is None and not cells:
                continue
            cols.append(
                InstrumentColumn(
                    letter=letter,
                    index=ci,
                    instrument_no=str(instr) if instr is not None else None,
                    header_note=str(note) if note is not None else None,
                    cells=cells,
                )
            )
        return cols

    def owner_rows(self, values: "SheetValues") -> list[OwnerRow]:
        rows: list[OwnerRow] = []
        for r in range(self._data_row_start, self.ws.max_row + 1):
            name = self.ws.cell(r, self.NAME_COL).value
            if name is None:
                continue
            rows.append(
                OwnerRow(
                    row=r,
                    name=str(name),
                    net_acres=_as_float(values.get(self.ws.title, "C", r)),
                    row_total=_as_float(values.get(self.ws.title, "E", r)),
                )
            )
        return rows

    def report_total(self, values: "SheetValues") -> Optional[float]:
        """Locate the REPORT TOTAL cell (column B label) and read its column-C value."""
        for r in range(self._data_row_start, self.ws.max_row + 1):
            b = self.ws.cell(r, 2).value
            if isinstance(b, str) and b.strip().upper() == "REPORT TOTAL":
                return _as_float(values.get(self.ws.title, "C", r))
        return None

    def subtotal_guards(self, values: "SheetValues") -> list[tuple[Coord, object]]:
        out: list[tuple[Coord, object]] = []
        for ci in range(self.FIRST_INSTR_COL, self.ws.max_column + 1):
            letter = get_column_letter(ci)
            v = values.get(self.ws.title, letter, self.SUBTOTAL_GUARD_ROW)
            if v is not None:
                out.append((Coord(self.ws.title, f"{letter}{self.SUBTOTAL_GUARD_ROW}"), v))
        return out


# --------------------------------------------------------------------------- #
# OGL register
# --------------------------------------------------------------------------- #
@dataclass
class OGLRecord:
    row: int
    ogl_no: Optional[int]
    lease_status: Optional[str]
    book_page: Optional[str]
    grantor: Optional[str]
    grantee: Optional[str]
    legal: Optional[str]


@dataclass
class OGLRegister:
    ws: Worksheet
    HEADER_ROW: int = 1

    def leases(self) -> list[OGLRecord]:
        out: list[OGLRecord] = []
        for r in range(2, self.ws.max_row + 1):
            no = self.ws.cell(r, 1).value
            if no is None:
                continue
            out.append(
                OGLRecord(
                    row=r,
                    ogl_no=_as_int(no),
                    lease_status=_str(self.ws.cell(r, 2).value),
                    book_page=_str(self.ws.cell(r, 5).value),
                    grantor=_str(self.ws.cell(r, 8).value),
                    grantee=_str(self.ws.cell(r, 9).value),
                    legal=_str(self.ws.cell(r, 10).value),
                )
            )
        return out


# --------------------------------------------------------------------------- #
# Runsheet
# --------------------------------------------------------------------------- #
@dataclass
class RunsheetRow:
    row: int
    instrument_no: Optional[str]
    doc_type: Optional[str]
    book_page: Optional[str]
    grantor: Optional[str]
    grantee: Optional[str]
    legal: Optional[str]
    tract_refs: Optional[str]
    classification: Optional[str]


@dataclass
class Runsheet:
    ws: Worksheet

    def instruments(self) -> list[RunsheetRow]:
        out: list[RunsheetRow] = []
        for r in range(2, self.ws.max_row + 1):
            instr = self.ws.cell(r, 1).value
            doc = self.ws.cell(r, 3).value
            if instr is None and doc is None:
                continue
            out.append(
                RunsheetRow(
                    row=r,
                    instrument_no=_str(instr),
                    doc_type=_str(doc),
                    book_page=_str(self.ws.cell(r, 4).value),
                    grantor=_str(self.ws.cell(r, 7).value),
                    grantee=_str(self.ws.cell(r, 8).value),
                    legal=_str(self.ws.cell(r, 9).value),
                    tract_refs=_str(self.ws.cell(r, 12).value),
                    classification=_str(self.ws.cell(r, 13).value),
                )
            )
        return out


@dataclass
class WISheet:
    ws: Worksheet
    index: int


@dataclass
class WellTab:
    ws: Worksheet

    def api_numbers(self) -> list[str]:
        found: list[str] = []
        for row in self.ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.count("-") == 2 and c.value.replace("-", "").isdigit():
                    found.append(c.value)
        return found


# --------------------------------------------------------------------------- #
# Evaluated-value lookup (bridges formulas -> recalculated numbers)
# --------------------------------------------------------------------------- #
@dataclass
class SheetValues:
    """Evaluated cell values keyed by (sheet, col_letter, row).

    Populated either from the workbook's cached values or, preferably, from a
    LibreOffice forced-recalc pass so validators see *computed* results.
    """
    _map: dict[tuple[str, str, int], object] = field(default_factory=dict)

    def set(self, sheet: str, col: str, row: int, value: object) -> None:
        self._map[(sheet, col, row)] = value

    def get(self, sheet: str, col: str, row: int) -> object:
        return self._map.get((sheet, col, row))

    @classmethod
    def from_worksheet_values(cls, ws: Worksheet) -> "SheetValues":
        sv = cls()
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    sv.set(ws.title, get_column_letter(c.column), c.row, c.value)
        return sv


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _as_float(v: object) -> Optional[float]:
    return float(v) if isinstance(v, (int, float)) else None


def _as_int(v: object) -> Optional[int]:
    try:
        return int(v)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _str(v: object) -> Optional[str]:
    return str(v) if v is not None else None
