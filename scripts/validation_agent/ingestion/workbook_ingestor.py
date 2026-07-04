"""Nondestructive workbook ingestion.

Reads an XLSX/XLSM without ever mutating it or executing macros. Produces a
rich, immutable structural view (sheets, headers, used ranges, formulas,
hardcoded-in-formula-region cells, relationships, drawings/media/VBA presence)
that downstream validators consume via the manifest.
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.hashing import sha256_file


class IngestionError(RuntimeError):
    pass


@dataclass
class CellRecord:
    coordinate: str
    row: int
    column: int
    value: Any
    formula: Optional[str] = None
    is_formula: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "coordinate": self.coordinate,
            "row": self.row,
            "column": self.column,
            "value": None if self.value is None else str(self.value),
            "formula": self.formula,
            "is_formula": self.is_formula,
        }


@dataclass
class SheetView:
    name: str
    index: int
    max_row: int
    max_col: int
    used_range: str
    header_row_index: Optional[int]
    headers: List[str]
    rows: List[List[Any]]  # values keyed positionally, header-aligned
    formula_cells: List[CellRecord]
    hardcoded_in_formula_regions: List[CellRecord]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "index": self.index,
            "max_row": self.max_row,
            "max_col": self.max_col,
            "used_range": self.used_range,
            "header_row_index": self.header_row_index,
            "headers": self.headers,
            "rows": [[None if v is None else str(v) for v in r] for r in self.rows],
            "formula_cells": [c.to_dict() for c in self.formula_cells],
            "hardcoded_in_formula_regions": [
                c.to_dict() for c in self.hardcoded_in_formula_regions
            ],
        }


@dataclass
class WorkbookStructure:
    path: str
    filename: str
    extension: str
    is_xlsm: bool
    sha256: str
    zip_valid: bool
    zip_entries: List[str]
    has_vba: bool
    has_drawings: bool
    has_media: bool
    relationships: List[str]
    sheets: List[SheetView] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "path": self.path,
            "filename": self.filename,
            "extension": self.extension,
            "is_xlsm": self.is_xlsm,
            "sha256": self.sha256,
            "zip_valid": self.zip_valid,
            "zip_entries": self.zip_entries,
            "has_vba": self.has_vba,
            "has_drawings": self.has_drawings,
            "has_media": self.has_media,
            "relationships": self.relationships,
            "sheets": [s.to_dict() for s in self.sheets],
            "warnings": self.warnings,
        }


def _validate_zip(path: Path) -> tuple[bool, List[str], List[str]]:
    """Return (is_valid_zip, entry_names, relationship_entries)."""
    try:
        with zipfile.ZipFile(path) as zf:
            bad = zf.testzip()
            if bad is not None:
                return False, [], []
            names = zf.namelist()
    except zipfile.BadZipFile:
        return False, [], []
    rels = [n for n in names if n.endswith(".rels")]
    return True, names, rels


def _first_header_row(ws) -> tuple[Optional[int], List[str]]:
    """Find the first row that looks like headers (>=2 non-empty text cells)."""
    limit = min(ws.max_row or 0, 25)
    max_col = ws.max_column or 0
    for r in range(1, limit + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        text_cells = [str(v).strip() for v in values if v is not None and str(v).strip()]
        if len(text_cells) >= 2:
            return r, [("" if v is None else str(v).strip()) for v in values]
    return None, []


def ingest_workbook(path: Path | str, max_rows_capture: int = 2000) -> WorkbookStructure:
    path = Path(path)
    if not path.exists():
        raise IngestionError(f"workbook not found: {path}")
    ext = path.suffix.lower()
    if ext not in {".xlsx", ".xlsm"}:
        raise IngestionError(f"unsupported workbook type: {ext}")

    zip_valid, entries, rels = _validate_zip(path)
    if not zip_valid:
        raise IngestionError(f"workbook is not a valid XLSX/XLSM ZIP: {path}")

    has_vba = any(n.lower() == "xl/vbaproject.bin" for n in entries)
    has_drawings = any(n.lower().startswith("xl/drawings/") for n in entries)
    has_media = any(n.lower().startswith("xl/media/") for n in entries)
    is_xlsm = ext == ".xlsm"

    structure = WorkbookStructure(
        path=str(path),
        filename=path.name,
        extension=ext,
        is_xlsm=is_xlsm,
        sha256=sha256_file(path),
        zip_valid=zip_valid,
        zip_entries=entries,
        has_vba=has_vba,
        has_drawings=has_drawings,
        has_media=has_media,
        relationships=rels,
    )

    # Load formulas (data_only=False) and cached values (data_only=True).
    # Loading never mutates the file and never executes macros.
    try:
        wb_f = load_workbook(path, data_only=False, keep_vba=is_xlsm, read_only=False)
    except Exception as exc:  # noqa: BLE001
        raise IngestionError(f"failed to open workbook (formulas): {exc}") from exc
    try:
        wb_v = load_workbook(path, data_only=True, keep_vba=is_xlsm, read_only=False)
    except Exception:  # noqa: BLE001
        wb_v = None  # cached values simply unavailable

    for idx, name in enumerate(wb_f.sheetnames):
        ws = wb_f[name]
        ws_v = wb_v[name] if wb_v is not None and name in wb_v.sheetnames else None
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        used_range = (
            f"A1:{get_column_letter(max_col)}{max_row}" if max_row and max_col else "A1"
        )
        header_row, headers = _first_header_row(ws)

        formula_cells: List[CellRecord] = []
        # Track which columns contain formulas (for hardcoded detection).
        formula_columns: Dict[int, int] = {}
        rows_captured: List[List[Any]] = []

        capture_start = (header_row + 1) if header_row else 1
        for r in range(1, min(max_row, max_rows_capture) + 1):
            row_values: List[Any] = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cached = ws_v.cell(row=r, column=c).value if ws_v else None
                is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                if is_formula:
                    formula_cells.append(
                        CellRecord(
                            coordinate=cell.coordinate,
                            row=r,
                            column=c,
                            value=cached,
                            formula=cell.value,
                            is_formula=True,
                        )
                    )
                    formula_columns[c] = formula_columns.get(c, 0) + 1
                # value used by validators: cached computed value if present else literal
                row_values.append(cached if cached is not None else cell.value)
            if r >= capture_start:
                rows_captured.append(row_values)

        # Hardcoded-in-formula-region: a numeric literal in a column where the
        # majority of populated cells are formulas.
        hardcoded: List[CellRecord] = []
        for c, fcount in formula_columns.items():
            if fcount < 2:
                continue
            for r in range(capture_start, min(max_row, max_rows_capture) + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is None:
                    continue
                is_formula = isinstance(v, str) and v.startswith("=")
                if not is_formula and isinstance(v, (int, float)):
                    hardcoded.append(
                        CellRecord(
                            coordinate=cell.coordinate,
                            row=r,
                            column=c,
                            value=v,
                            formula=None,
                            is_formula=False,
                        )
                    )

        structure.sheets.append(
            SheetView(
                name=name,
                index=idx,
                max_row=max_row,
                max_col=max_col,
                used_range=used_range,
                header_row_index=header_row,
                headers=headers,
                rows=rows_captured,
                formula_cells=formula_cells,
                hardcoded_in_formula_regions=hardcoded,
            )
        )

    wb_f.close()
    if wb_v is not None:
        wb_v.close()
    return structure
