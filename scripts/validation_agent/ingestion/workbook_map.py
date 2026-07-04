"""Workbook ingestion & sheet categorisation (read-only)."""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.workbook.workbook import Workbook

from .. import config
from ..models import SheetKind
from .sheet_models import (
    OGLRegister,
    Runsheet,
    SheetValues,
    TractGrid,
    WellTab,
    WISheet,
)

_TRACT_RE = re.compile(r"^\s*tract\s*(\d+)\s*$", re.IGNORECASE)


@dataclass
class WorkbookModel:
    path: Path
    wb: Workbook
    kinds: dict[str, SheetKind] = field(default_factory=dict)

    # -- categorised accessors ------------------------------------------- #
    def sheets_of(self, kind: SheetKind) -> list[str]:
        return [name for name, k in self.kinds.items() if k == kind]

    def tracts(self) -> list[TractGrid]:
        grids: list[TractGrid] = []
        for name, kind in self.kinds.items():
            if kind != SheetKind.TRACT_GRID:
                continue
            m = _TRACT_RE.match(name)
            if not m:
                continue
            no = int(m.group(1))
            grids.append(
                TractGrid(
                    ws=self.wb[name],
                    tract_no=no,
                    acreage=config.TRACT_ACREAGE.get(no, 0.0),
                )
            )
        return sorted(grids, key=lambda g: g.tract_no)

    def ogl(self) -> Optional[OGLRegister]:
        names = self.sheets_of(SheetKind.OGL_REGISTER)
        return OGLRegister(self.wb[names[0]]) if names else None

    def runsheet(self) -> Optional[Runsheet]:
        names = self.sheets_of(SheetKind.RUNSHEET)
        return Runsheet(self.wb[names[0]]) if names else None

    def wi_sheets(self) -> list[WISheet]:
        out: list[WISheet] = []
        for i, name in enumerate(self.sheets_of(SheetKind.WI), start=1):
            out.append(WISheet(self.wb[name], index=i))
        return out

    def wells(self) -> list[WellTab]:
        return [WellTab(self.wb[n]) for n in self.sheets_of(SheetKind.WELL)]

    def cached_values(self) -> SheetValues:
        """Evaluated values from the workbook's stored cache (fast fallback)."""
        data_wb = openpyxl.load_workbook(self.path, data_only=True)
        sv = SheetValues()
        for ws in data_wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is not None:
                        from openpyxl.utils import get_column_letter
                        sv.set(ws.title, get_column_letter(c.column), c.row, c.value)
        return sv

    def checksum(self) -> str:
        import hashlib
        return hashlib.sha256(self.path.read_bytes()).hexdigest()


class WorkbookMapper:
    """Loads a workbook and classifies every sheet by name + shape heuristics."""

    @staticmethod
    def load(path: str | Path) -> WorkbookModel:
        p = Path(path)
        wb = openpyxl.load_workbook(p, data_only=False)
        kinds: dict[str, SheetKind] = {}
        for name in wb.sheetnames:
            kinds[name] = WorkbookMapper._classify(name)
        return WorkbookModel(path=p, wb=wb, kinds=kinds)

    @staticmethod
    def _classify(name: str) -> SheetKind:
        n = name.strip().lower()
        if _TRACT_RE.match(name):
            return SheetKind.TRACT_GRID
        if n == "ogl":
            return SheetKind.OGL_REGISTER
        if n == "runsheet":
            return SheetKind.RUNSHEET
        if n.startswith("wi"):
            return SheetKind.WI
        if n.startswith("well"):
            return SheetKind.WELL
        if n == "overview":
            return SheetKind.OVERVIEW
        if n == "plat":
            return SheetKind.PLAT
        if n == "rawdata":
            return SheetKind.RAWDATA
        return SheetKind.OTHER
