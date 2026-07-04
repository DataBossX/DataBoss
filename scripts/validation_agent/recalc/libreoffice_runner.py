"""Headless LibreOffice recalculation.

Recalculates a copied workbook version by converting it through headless
LibreOffice, producing a NEW version file. If LibreOffice is missing, the
failure is logged and reported honestly — success is never faked. The original
good version is never overwritten with corrupt output.
"""

from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Optional

_COMMON_PATHS = [
    "/usr/bin/libreoffice",
    "/usr/bin/soffice",
    "/opt/libreoffice/program/soffice",
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]


def detect_libreoffice(configured_path: str = "") -> Optional[str]:
    """Return a usable LibreOffice executable path, or None."""
    if configured_path and Path(configured_path).exists():
        return configured_path
    which = shutil.which("libreoffice") or shutil.which("soffice")
    if which:
        return which
    for p in _COMMON_PATHS:
        if Path(p).exists():
            return p
    return None


class LibreOfficeResult:
    def __init__(self, ok: bool, output_path: Optional[Path] = None,
                 error: str = "", available: bool = True,
                 comparison: Optional[dict] = None):
        self.ok = ok
        self.output_path = output_path
        self.error = error
        self.available = available
        self.comparison = comparison

    def to_dict(self) -> dict:
        return {
            "ok": self.ok,
            "available": self.available,
            "output_path": str(self.output_path) if self.output_path else None,
            "error": self.error,
            "comparison": self.comparison,
        }


def _populated_census(path) -> dict[str, int]:
    """Return {sheet_name: populated_cell_count} for a workbook (read-only)."""
    import openpyxl
    census: dict[str, int] = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        for ws in wb.worksheets:
            count = 0
            for row in ws.iter_rows(values_only=True):
                count += sum(1 for c in row if c is not None)
            census[ws.title] = count
    finally:
        wb.close()
    return census


def compare_key_cells(before_path, after_path,
                      min_retained_fraction: float = 0.5) -> dict:
    """Compare structure/population of two workbook versions to detect
    corruption or gross data loss from a recalculation.

    Corruption is flagged when a sheet disappears or when total populated cells
    fall below ``min_retained_fraction`` of the source. Recalc changes computed
    values (not structure), so we compare structure + population, not values.
    """
    before = _populated_census(before_path)
    after = _populated_census(after_path)
    missing_sheets = sorted(set(before) - set(after))
    total_before = sum(before.values())
    total_after = sum(after.values())
    retained = (total_after / total_before) if total_before else 1.0
    corruption = bool(missing_sheets) or retained < min_retained_fraction
    reason = ""
    if missing_sheets:
        reason = f"sheets lost: {', '.join(missing_sheets)}"
    elif corruption:
        reason = (f"populated cells dropped to {retained:.0%} of source "
                  f"({total_after}/{total_before})")
    return {
        "sheets_before": len(before), "sheets_after": len(after),
        "populated_before": total_before, "populated_after": total_after,
        "retained_fraction": round(retained, 3),
        "missing_sheets": missing_sheets,
        "corruption": corruption, "reason": reason,
    }


def recalculate(source_version: str | Path, dest_version: str | Path,
                configured_path: str = "", timeout: int = 120
                ) -> LibreOfficeResult:
    """Recalculate ``source_version`` -> ``dest_version`` (never overwrites)."""
    source = Path(source_version)
    dest = Path(dest_version)
    if dest.exists():
        return LibreOfficeResult(False, error=f"Refusing to overwrite {dest}")

    exe = detect_libreoffice(configured_path)
    if not exe:
        return LibreOfficeResult(
            False, available=False,
            error="LibreOffice not found; recalculation skipped (dependency "
                  "failure logged, not faked).",
        )

    with tempfile.TemporaryDirectory(prefix="dbx_lo_") as tmp:
        tmp_dir = Path(tmp)
        # Convert to xlsx in an isolated profile so we don't touch anything.
        profile = tmp_dir / "profile"
        cmd = [
            exe, "--headless", "--calc",
            f"-env:UserInstallation=file://{profile}",
            "--convert-to", "xlsx:Calc MS Excel 2007 XML",
            "--outdir", str(tmp_dir), str(source),
        ]
        try:
            proc = subprocess.run(cmd, capture_output=True, timeout=timeout)
        except subprocess.TimeoutExpired:
            return LibreOfficeResult(False, error="LibreOffice timed out.")
        except Exception as e:
            return LibreOfficeResult(False, error=f"LibreOffice error: {e}")
        if proc.returncode != 0:
            return LibreOfficeResult(
                False, error=f"LibreOffice exit {proc.returncode}: "
                             f"{proc.stderr.decode(errors='ignore')[:400]}",
            )
        produced = tmp_dir / (source.stem + ".xlsx")
        if not produced.exists():
            # Fall back to any produced xlsx.
            candidates = list(tmp_dir.glob("*.xlsx"))
            if not candidates:
                return LibreOfficeResult(
                    False, error="LibreOffice produced no output file.")
            produced = candidates[0]

        # Verify the produced file opens before accepting it.
        try:
            import openpyxl
            wb = openpyxl.load_workbook(produced, read_only=True)
            wb.close()
        except Exception as e:
            return LibreOfficeResult(
                False, error=f"Recalc output failed to open (corrupt): {e}")

        # Compare key structure/population before vs after; refuse to accept a
        # recalc output that lost sheets or data (never overwrite good with
        # corrupt output).
        try:
            comparison = compare_key_cells(source, produced)
        except Exception as e:
            return LibreOfficeResult(
                False, error=f"Could not compare recalc output: {e}")
        if comparison["corruption"]:
            return LibreOfficeResult(
                False, comparison=comparison,
                error=f"Recalc output rejected as corrupt: {comparison['reason']}")

        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(produced, dest)
        return LibreOfficeResult(True, output_path=dest, comparison=comparison)
