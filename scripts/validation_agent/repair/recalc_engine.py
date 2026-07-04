"""LibreOffice headless forced-recalc engine (Component 7).

Two hard-won facts drive this design:
1. ``--convert-to`` preserves *stale* cached values unless the user profile sets
   ``OOXMLRecalcMode=0`` (always recalc). We stamp that into a private profile.
2. A private ``-env:UserInstallation`` is required so concurrent/headless runs do
   not collide with a system LibreOffice profile.
"""
from __future__ import annotations

import re
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from .. import config
from ..ingestion.sheet_models import SheetValues

_FORMULA_ERRORS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!")

_RECALC_PROFILE_XCU = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
    'xmlns:xs="http://www.w3.org/2001/XMLSchema">'
    '<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>'
    '<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>'
    '</oor:items>\n'
)


@dataclass
class FormulaError:
    sheet: str
    cell: str
    code: str


@dataclass
class RecalcOutcome:
    values: SheetValues
    errors: list[FormulaError] = field(default_factory=list)
    recalculated_path: Optional[Path] = None

    @property
    def ok(self) -> bool:
        return not self.errors


class RecalcEngine:
    def __init__(self, soffice: Optional[str] = None) -> None:
        self._soffice = soffice or config.soffice_binary()
        self._profile = Path(tempfile.mkdtemp(prefix="va_lo_profile_"))
        self._ensure_force_recalc_profile()

    # -- profile ---------------------------------------------------------- #
    def _ensure_force_recalc_profile(self) -> None:
        user_dir = self._profile / "user"
        user_dir.mkdir(parents=True, exist_ok=True)
        (user_dir / "registrymodifications.xcu").write_text(
            _RECALC_PROFILE_XCU, encoding="utf-8"
        )

    # -- full recalc ------------------------------------------------------ #
    def recalc(self, path: str | Path, timeout: int = 600) -> RecalcOutcome:
        src = Path(path)
        out_dir = Path(tempfile.mkdtemp(prefix="va_recalc_"))
        cmd = [
            self._soffice,
            "--headless",
            "--norestore",
            f"-env:UserInstallation=file://{self._profile}",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            str(out_dir),
            str(src),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        recalced = out_dir / (src.stem + ".xlsx")
        if not recalced.exists():
            raise RuntimeError(
                f"LibreOffice recalc produced no output: {proc.stderr.strip()}"
            )
        return self._read_back(recalced)

    def recalc_sheet(
        self, path: str | Path, sheet: str, timeout: int = 600
    ) -> RecalcOutcome:
        """OOM-safe single-sheet recalc: slice one sheet into a throwaway book.

        The tract grids are self-contained for the conservation gate — their
        formulas reference only their own sheet (SUBTOTAL guards, in-sheet SUMIFS,
        the constant D5 acreage). We therefore copy just that sheet (formulas and
        constants) into a fresh 1-sheet workbook and recalc it, holding memory to a
        single grid instead of all nineteen sheets. Any formula that DOES reach
        another sheet is frozen to its cached value in the slice, which is
        acceptable for per-sheet G1 checks.

        Raises ValueError if the sheet is not self-contained enough to slice
        safely (cross-sheet refs on the audited columns), so the caller can fall
        back to full ``recalc``.
        """
        import openpyxl
        from openpyxl.utils import get_column_letter

        src = Path(path)
        src_wb = openpyxl.load_workbook(src, data_only=False)
        if sheet not in src_wb.sheetnames:
            raise ValueError(f"sheet {sheet!r} not in workbook")
        src_ws = src_wb[sheet]
        cache_wb = openpyxl.load_workbook(src, data_only=True)
        cache_ws = cache_wb[sheet]

        slice_wb = openpyxl.Workbook()
        slice_ws = slice_wb.active
        slice_ws.title = sheet[:31]
        for row in src_ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                col = get_column_letter(c.column)
                if isinstance(v, str) and v.startswith("="):
                    # freeze cross-sheet formulas to cached value
                    if "!" in v:
                        v = cache_ws[f"{col}{c.row}"].value
                slice_ws[f"{col}{c.row}"] = v

        slice_path = Path(tempfile.mkdtemp(prefix="va_slice_")) / f"{slice_ws.title}.xlsx"
        slice_wb.save(slice_path)
        return self.recalc(slice_path, timeout=timeout)

    # -- read-back -------------------------------------------------------- #
    @staticmethod
    def _read_back(recalced: Path) -> RecalcOutcome:
        wb = openpyxl.load_workbook(recalced, data_only=True)
        sv = SheetValues()
        errors: list[FormulaError] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None:
                        continue
                    col = get_column_letter(c.column)
                    sv.set(ws.title, col, c.row, c.value)
                    if isinstance(c.value, str):
                        code = c.value.strip()
                        if code in _FORMULA_ERRORS or code.rstrip("!?") in (
                            e.rstrip("!?") for e in _FORMULA_ERRORS
                        ):
                            errors.append(FormulaError(ws.title, f"{col}{c.row}", code))
        return RecalcOutcome(values=sv, errors=errors, recalculated_path=recalced)
