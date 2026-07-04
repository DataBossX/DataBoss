"""Non-destructive XLSX/XLSM XML repair.

Repairs operate only on a *copied* workbook version and always produce the
next version — the input file is never modified. Every unrelated ZIP entry
(drawings, media, styles, shared strings, relationships, and the VBA project
for XLSM) is preserved byte-for-byte. After a formula edit, ``xl/calcChain.xml``
is dropped and ``fullCalcOnLoad`` is set so the recalc engine rebuilds values.

If the repackaged workbook fails validation, the edit is rolled back by simply
not registering the new version — the prior version remains the latest good
copy.
"""

from __future__ import annotations

import shutil
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from lxml import etree

from core.hashing import sha256_file

# Hardened parser: never resolve external entities, load DTDs, or hit the
# network. XLSX XML can come from an untrusted workbook, so this closes XXE /
# billion-laughs / SSRF-via-entity vectors while parsing.
_SAFE_PARSER = etree.XMLParser(
    resolve_entities=False,
    no_network=True,
    load_dtd=False,
    dtd_validation=False,
    huge_tree=False,
)


def _parse(path: Path) -> "etree._ElementTree":
    return etree.parse(str(path), _SAFE_PARSER)


_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"


class XmlRepairError(RuntimeError):
    pass


@dataclass
class RepairOp:
    sheet_name: str
    coordinate: str
    formula: str  # without leading '='


@dataclass
class RepairOutcome:
    ok: bool
    output_path: Optional[Path]
    before_sha256: str
    after_sha256: Optional[str]
    applied: List[str] = field(default_factory=list)
    detail: str = ""


class XmlWorkbookEditor:
    """Edits worksheet XML in a copied workbook and repackages the next version."""

    def __init__(self, source_version_path: Path) -> None:
        self.source = Path(source_version_path)
        if not self.source.exists():
            raise XmlRepairError(f"source version not found: {self.source}")
        self._tmp: Optional[Path] = None
        self._sheet_targets: Dict[str, str] = {}

    # ------------------------------------------------------------------ #
    def _extract(self) -> Path:
        self._tmp = Path(tempfile.mkdtemp(prefix="dbx_xml_"))
        with zipfile.ZipFile(self.source) as zf:
            bad = zf.testzip()
            if bad is not None:
                raise XmlRepairError(f"corrupt entry in source zip: {bad}")
            zf.extractall(self._tmp)
        self._map_sheets()
        return self._tmp

    def _map_sheets(self) -> None:
        assert self._tmp is not None
        wb_xml = self._tmp / "xl" / "workbook.xml"
        rels_xml = self._tmp / "xl" / "_rels" / "workbook.xml.rels"
        if not wb_xml.exists() or not rels_xml.exists():
            raise XmlRepairError("workbook.xml or its rels missing")
        wb_tree = _parse(wb_xml)
        rid_to_target: Dict[str, str] = {}
        rels_tree = _parse(rels_xml)
        for rel in rels_tree.getroot():
            rid = rel.get("Id")
            target = rel.get("Target")
            if rid and target:
                rid_to_target[rid] = target
        for sheet in wb_tree.getroot().iter(f"{{{_MAIN_NS}}}sheet"):
            name = sheet.get("name")
            rid = sheet.get(f"{{{_REL_NS}}}id")
            if name and rid and rid in rid_to_target:
                safe = self._safe_member(rid_to_target[rid])
                if safe is not None:
                    self._sheet_targets[name] = safe

    def _safe_member(self, target: str) -> Optional[str]:
        """Resolve an OOXML relationship Target to a package-relative path that
        is provably contained under the extraction root.

        The Target attribute is attacker-controlled content inside an untrusted
        workbook. ``lstrip`` alone does NOT neutralize embedded ``..`` segments
        (``a/../../etc/x`` survives it), so we normalize and enforce
        containment, rejecting anything that escapes the temp directory.
        """
        assert self._tmp is not None
        t = (target or "").strip().replace("\\", "/")
        if not t:
            return None
        # Absolute package paths ("/xl/...") are relative to the package root;
        # relative paths ("worksheets/sheet1.xml") are relative to xl/.
        rel = t.lstrip("/") if t.startswith("/") else f"xl/{t}"
        parts = [p for p in rel.split("/") if p not in ("", ".")]
        if any(p == ".." for p in parts):
            return None
        rel = "/".join(parts)
        root = self._tmp.resolve()
        resolved = (self._tmp / rel).resolve()
        if resolved != root and root not in resolved.parents:
            return None
        return rel

    # ------------------------------------------------------------------ #
    def _edit_cell(self, worksheet_path: Path, coordinate: str, formula: str) -> bool:
        tree = _parse(worksheet_path)
        root = tree.getroot()
        for cell in root.iter(f"{{{_MAIN_NS}}}c"):
            if cell.get("r") == coordinate:
                # remove any cached value and existing formula, keep style (s)
                for child in list(cell):
                    tag = etree.QName(child).localname
                    if tag in ("f", "v"):
                        cell.remove(child)
                # a formula cell must not carry a t="s" (shared string) type
                if cell.get("t") in ("s", "str", "e"):
                    del cell.attrib["t"]
                f_el = etree.SubElement(cell, f"{{{_MAIN_NS}}}f")
                f_el.text = formula
                tree.write(
                    str(worksheet_path),
                    xml_declaration=True,
                    encoding="UTF-8",
                    standalone=True,
                )
                return True
        return False

    def _drop_calc_chain(self) -> None:
        assert self._tmp is not None
        calc = self._tmp / "xl" / "calcChain.xml"
        if calc.exists():
            calc.unlink()
        # remove content-type override
        ct = self._tmp / "[Content_Types].xml"
        if ct.exists():
            tree = _parse(ct)
            root = tree.getroot()
            for override in list(root):
                if override.get("PartName") == "/xl/calcChain.xml":
                    root.remove(override)
            tree.write(str(ct), xml_declaration=True, encoding="UTF-8", standalone=True)
        # remove workbook rels entry pointing at calcChain
        rels = self._tmp / "xl" / "_rels" / "workbook.xml.rels"
        if rels.exists():
            tree = _parse(rels)
            root = tree.getroot()
            for rel in list(root):
                if (rel.get("Target") or "").endswith("calcChain.xml"):
                    root.remove(rel)
            tree.write(str(rels), xml_declaration=True, encoding="UTF-8", standalone=True)

    def _set_full_calc_on_load(self) -> None:
        assert self._tmp is not None
        wb_xml = self._tmp / "xl" / "workbook.xml"
        tree = _parse(wb_xml)
        root = tree.getroot()
        calc_pr = root.find(f"{{{_MAIN_NS}}}calcPr")
        if calc_pr is None:
            calc_pr = etree.SubElement(root, f"{{{_MAIN_NS}}}calcPr")
        calc_pr.set("fullCalcOnLoad", "1")
        calc_pr.set("calcId", "0")
        tree.write(str(wb_xml), xml_declaration=True, encoding="UTF-8", standalone=True)

    def _repackage(self, dest: Path) -> None:
        assert self._tmp is not None
        if dest.exists():
            raise XmlRepairError(f"refusing to overwrite existing version: {dest}")
        # Write [Content_Types].xml first is not strictly required, but keep order stable.
        files: List[Path] = sorted(
            [p for p in self._tmp.rglob("*") if p.is_file()],
            key=lambda p: str(p.relative_to(self._tmp)),
        )
        with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zf:
            for path in files:
                arcname = str(path.relative_to(self._tmp)).replace("\\", "/")
                zf.write(path, arcname)

    @staticmethod
    def _validate_workbook(path: Path) -> Tuple[bool, str]:
        try:
            with zipfile.ZipFile(path) as zf:
                if zf.testzip() is not None:
                    return False, "zip testzip reported a bad entry"
                names = zf.namelist()
            if "xl/workbook.xml" not in names or "[Content_Types].xml" not in names:
                return False, "required OOXML parts missing after repackage"
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, keep_vba=path.suffix.lower() == ".xlsm")
            _ = wb.sheetnames
            wb.close()
            return True, "ok"
        except Exception as exc:  # noqa: BLE001
            return False, f"validation failed: {exc}"

    # ------------------------------------------------------------------ #
    def apply_formula_repairs(self, ops: List[RepairOp], dest_path: Path) -> RepairOutcome:
        """Apply formula repairs and write the next version. Rolls back on failure."""
        before = sha256_file(self.source)
        applied: List[str] = []
        try:
            try:
                self._extract()
                return self._apply_inner(ops, dest_path, before, applied)
            except XmlRepairError as exc:
                if dest_path.exists():
                    try:
                        dest_path.unlink()
                    except OSError:
                        pass
                return RepairOutcome(
                    ok=False,
                    output_path=None,
                    before_sha256=before,
                    after_sha256=None,
                    applied=applied,
                    detail=f"repair aborted: {exc}",
                )
        finally:
            if self._tmp and self._tmp.exists():
                shutil.rmtree(self._tmp, ignore_errors=True)
                self._tmp = None

    def _apply_inner(
        self, ops: List[RepairOp], dest_path: Path, before: str, applied: List[str]
    ) -> RepairOutcome:
        assert self._tmp is not None
        for op in ops:
            target = self._sheet_targets.get(op.sheet_name)
            if not target:
                # unknown sheet: skip safely, record nothing applied for it
                continue
            ws_path = self._tmp / target
            # Defense-in-depth: never read/write outside the extraction root,
            # even if a target somehow slipped past _safe_member.
            root = self._tmp.resolve()
            resolved = ws_path.resolve()
            if resolved != root and root not in resolved.parents:
                continue
            if not ws_path.exists():
                continue
            if self._edit_cell(ws_path, op.coordinate, op.formula):
                applied.append(f"{op.sheet_name}!{op.coordinate}={op.formula}")
        if applied:
            self._drop_calc_chain()
            self._set_full_calc_on_load()
        self._repackage(dest_path)
        ok, detail = self._validate_workbook(dest_path)
        if not ok:
            # rollback: discard the corrupt output, keep prior version
            if dest_path.exists():
                dest_path.unlink()
            return RepairOutcome(
                ok=False,
                output_path=None,
                before_sha256=before,
                after_sha256=None,
                applied=applied,
                detail=f"rolled back: {detail}",
            )
        return RepairOutcome(
            ok=True,
            output_path=dest_path,
            before_sha256=before,
            after_sha256=sha256_file(dest_path),
            applied=applied,
            detail="ok",
        )
