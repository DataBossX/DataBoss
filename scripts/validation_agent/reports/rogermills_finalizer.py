"""Roger Mills multi-folder title-report finalizer.

Given three source folders (e.g. Roger Mills, Roger Mills 2, Roger Mills 3),
this engine:

  1. Inventories + backs up (timestamped, read-only originals) every workbook.
  2. Classifies sheets (tract / title / OGL / runsheet / sources) and scores
     each candidate report workbook.
  3. Runs a TOURNAMENT (top-K bases) and refinement LOOPS to pick the best
     final per section.
  4. Uses a TEMPLATE workbook as the formatting authority (its formatting is
     preserved; data is written into it, matching your template layout).
  5. Applies safe fixes:
        * TRACT SHEET is authoritative scope — output never adds a tract that
          is not on the tract sheet ("don't go off the tract sheets"); footing
          and formatting are fixed, legal facts are never invented.
        * TITLE SHEET data (section, county, dates, prepared-for) normalized
          into the template's title block.
        * OGL NUMBERS are taken from the OGL sheet and reconciled onto each
          tract by key.
  6. Writes best finals into the output folder (versioned, never overwrites),
     validates each by reopening it, and emits a per-report REVIEW file listing
     everything an examiner must confirm (nothing unverifiable is silently
     trusted).

Every action is recorded to the append-only audit DB. Any paid source
retrieval (browser/API) is gated by the $100 SpendGuard — no silent spend.

IT NEVER INVENTS TITLE FACTS. Missing/ambiguous data is flagged for the human.
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

from ingestion.workbook_ingestor import WorkbookIngestor
from ingestion.sheet_classifier import classify_workbook
from ingestion.data_extractor import _best_column
from core.run_manager import sha256_file

# Column alias tables (header-driven, fuzzy). Override via config if needed.
_TRACT_ID_ALIASES = ["tract", "tract no", "tract number", "parcel", "no", "#"]
_LEGAL_ALIASES = ["legal description", "legal", "description", "land", "lands"]
_ACREAGE_ALIASES = ["acreage", "acres", "gross acres", "net acres", "acre"]
_OGL_TRACT_ALIASES = _TRACT_ID_ALIASES
_OGL_NUM_ALIASES = ["ogl", "ogl number", "ogl no", "ogl #", "lease number",
                    "lease no", "lease", "og lease"]
_LESSOR_ALIASES = ["lessor", "landowner", "mineral owner"]
_LESSEE_ALIASES = ["lessee", "operator", "company"]

# Title-block labels to normalize into the template's title sheet.
_TITLE_LABELS = {
    "section": ["section", "s/t/r", "str", "section township range", "location"],
    "county": ["county"],
    "state": ["state"],
    "prepared_for": ["prepared for", "client", "prepared_for"],
    "effective_date": ["effective date", "effective"],
    "report_date": ["date", "report date", "as of"],
}

_WORKBOOK_EXTS = {".xlsx", ".xlsm"}
_TITLE_SHEET_KEYWORDS = ("title", "cover", "summary", "report")


def _now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


@dataclass
class Candidate:
    path: Path
    sheets: int = 0
    populated: int = 0
    has_tract: bool = False
    has_ogl: bool = False
    has_title: bool = False
    formulas: int = 0
    blanks_ratio: float = 0.0
    mtime: float = 0.0
    score: float = 0.0
    classification: dict = field(default_factory=dict)
    is_template: bool = False
    is_final_copy: bool = False


@dataclass
class ReviewItem:
    kind: str
    subject: str
    detail: str


@dataclass
class FolderResult:
    folder: Path
    candidates: list[Candidate] = field(default_factory=list)
    best: Optional[Candidate] = None
    template: Optional[Candidate] = None
    final_copy: Optional[Candidate] = None
    final_path: Optional[Path] = None
    review: list[ReviewItem] = field(default_factory=list)


@dataclass
class FinalizeConfig:
    horizon_root: Path
    output_dir: Path
    folders: list[Path]
    template_path: Optional[Path] = None
    dry_run: bool = True
    top_k: int = 3           # tournament breadth
    loops: int = 2           # refinement rounds
    tract_sheet: Optional[str] = None
    title_sheet: Optional[str] = None
    ogl_sheet: Optional[str] = None


class RogerMillsFinalizer:
    def __init__(self, config: FinalizeConfig, db=None, audit=None,
                 spend_guard=None):
        self.cfg = config
        self.db = db
        self.audit = audit
        self.spend_guard = spend_guard
        self.stamp = _now_stamp()
        self.backup_root = config.output_dir / "_backups" / self.stamp

    # ------------------------------------------------------------------ #
    def _log(self, event: str, desc: str, data: Any = None) -> None:
        if self.audit:
            self.audit.log(event, desc, data=data)

    def _sheet_rows(self, path: Path) -> dict[str, list[list[Any]]]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        out: dict[str, list[list[Any]]] = {}
        try:
            for ws in wb.worksheets:
                out[ws.title] = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        return out

    @staticmethod
    def _find_header(rows: list[list[Any]]) -> tuple[int, list[str]]:
        for i, row in enumerate(rows[:15]):
            strings = [c for c in row if isinstance(c, str) and c.strip()]
            if len(strings) >= 2:
                return i, [str(c) if c is not None else "" for c in row]
        return -1, []

    # ------------------------------------------------------------------ #
    def inventory_and_backup(self, folder: Path) -> list[Path]:
        """List workbooks and back up EVERY original (read-only), timestamped."""
        workbooks = [p for p in folder.rglob("*")
                     if p.is_file() and p.suffix.lower() in _WORKBOOK_EXTS
                     and not p.name.startswith("~$")]
        if not self.cfg.dry_run:
            backup_dir = self.backup_root / folder.name
            backup_dir.mkdir(parents=True, exist_ok=True)
            for p in workbooks:
                dest = backup_dir / p.name
                if not dest.exists():
                    shutil.copy2(p, dest)  # copy2 opens original read-only
        self._log("inventory", f"{folder.name}: {len(workbooks)} workbook(s)",
                  {"folder": str(folder)})
        return workbooks

    def analyze(self, path: Path) -> Candidate:
        cand = Candidate(path=path)
        try:
            cand.mtime = path.stat().st_mtime
            manifest = WorkbookIngestor(path).ingest()
            cand.sheets = len(manifest.sheets)
            cand.classification = classify_workbook(manifest)
            found = cand.classification.get("found_categories", {})
            cand.has_tract = "tracts" in found
            cand.has_ogl = "ogl_register" in found
            populated = 0
            formulas = 0
            filled = 0
            total_cells = 0
            for s in manifest.sheets:
                populated += s.numeric_cells + s.text_cells + s.formula_cells
                formulas += s.formula_cells
                total_cells += max(s.max_row * s.max_col, 1)
                filled += s.numeric_cells + s.text_cells + s.formula_cells
            cand.populated = populated
            cand.formulas = formulas
            cand.blanks_ratio = 1.0 - (filled / total_cells) if total_cells else 1.0
            # Title sheet detection by name keyword.
            cand.has_title = any(
                any(k in s.name.lower() for k in _TITLE_SHEET_KEYWORDS)
                for s in manifest.sheets
            )
            name = path.name.lower()
            cand.is_template = "template" in name
            cand.is_final_copy = "final" in name
        except Exception as e:
            self._log("analyze_error", f"{path.name}: {e}")
        return cand

    def score(self, cand: Candidate) -> float:
        """Weighted quality score for the tournament."""
        s = 0.0
        s += 25 if cand.has_tract else 0
        s += 20 if cand.has_ogl else 0
        s += 10 if cand.has_title else 0
        s += min(cand.populated / 50.0, 30.0)         # completeness
        s += min(cand.formulas / 5.0, 10.0)           # live formulas
        s += (1.0 - cand.blanks_ratio) * 10.0         # cleanliness
        s += min(cand.mtime / 1e9, 3.0)               # recency
        # A final copy is a strong base to fix; a template is the format source.
        if cand.is_final_copy:
            s += 8
        if cand.is_template:
            s -= 100   # never pick the blank template as the DATA base
        cand.score = round(s, 2)
        return cand.score

    # ------------------------------------------------------------------ #
    def _role_sheet(self, rows_by_sheet: dict, classification: dict,
                    category: str, override: Optional[str],
                    name_keywords: tuple = ()) -> Optional[str]:
        """Resolve the sheet playing a role.

        Order of trust: explicit override > SHEET NAME keyword (the strongest
        signal for these domain sheets, which the user names 'tract sheet',
        'OGL sheet', 'title sheet') > fuzzy classification. Name-first avoids
        the classifier tie where an OGL sheet with a 'Tract' column is
        mislabeled as a tract sheet.
        """
        if override and override in rows_by_sheet:
            return override
        if name_keywords:
            for name in rows_by_sheet:
                if any(k in name.lower() for k in name_keywords):
                    return name
        found = classification.get("found_categories", {})
        names = found.get(category, [])
        if names:
            return names[0]
        return None

    def extract_tracts(self, rows_by_sheet, classification) -> dict[str, dict]:
        sheet = self._role_sheet(rows_by_sheet, classification, "tracts",
                                 self.cfg.tract_sheet, ("tract",))
        if not sheet:
            return {}
        rows = rows_by_sheet[sheet]
        hidx, headers = self._find_header(rows)
        if hidx < 0:
            return {}
        norm = [str(h or "").strip().lower() for h in headers]
        c_id = _best_column(norm, _TRACT_ID_ALIASES)
        c_legal = _best_column(norm, _LEGAL_ALIASES)
        c_ac = _best_column(norm, _ACREAGE_ALIASES)
        tracts: dict[str, dict] = {}
        for row in rows[hidx + 1:]:
            def cell(i):
                return row[i] if 0 <= i < len(row) else None
            tid = cell(c_id)
            legal = cell(c_legal)
            if tid is None and legal is None:
                continue
            key = str(tid).strip() if tid is not None else str(legal)[:24]
            if not key:
                continue
            tracts[key] = {
                "tract": tid, "legal_description": legal,
                "acreage": cell(c_ac), "_sheet": sheet,
            }
        return tracts

    def extract_ogl(self, rows_by_sheet, classification) -> dict[str, dict]:
        sheet = self._role_sheet(rows_by_sheet, classification, "ogl_register",
                                 self.cfg.ogl_sheet, ("ogl",))
        if not sheet:
            return {}
        rows = rows_by_sheet[sheet]
        hidx, headers = self._find_header(rows)
        if hidx < 0:
            return {}
        norm = [str(h or "").strip().lower() for h in headers]
        c_tr = _best_column(norm, _OGL_TRACT_ALIASES)
        c_num = _best_column(norm, _OGL_NUM_ALIASES)
        c_lessor = _best_column(norm, _LESSOR_ALIASES)
        c_lessee = _best_column(norm, _LESSEE_ALIASES)
        ogl: dict[str, dict] = {}
        for row in rows[hidx + 1:]:
            def cell(i):
                return row[i] if 0 <= i < len(row) else None
            tr = cell(c_tr)
            num = cell(c_num)
            if num is None and tr is None:
                continue
            key = str(tr).strip() if tr is not None else str(num)
            ogl[key] = {"ogl_number": num, "lessor": cell(c_lessor),
                        "lessee": cell(c_lessee)}
        return ogl

    def extract_title(self, rows_by_sheet, classification) -> dict[str, Any]:
        sheet = self._role_sheet(rows_by_sheet, classification, "qa",
                                 self.cfg.title_sheet, _TITLE_SHEET_KEYWORDS)
        fields: dict[str, Any] = {}
        if not sheet:
            # Fall back to scanning the first sheet's labeled cells.
            sheet = next(iter(rows_by_sheet), None)
            if not sheet:
                return fields
        rows = rows_by_sheet[sheet]
        # Labeled-cell scan: for each cell equal-ish to a label, take the cell
        # to its right (or below) as the value.
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                if not isinstance(val, str):
                    continue
                label = val.strip().lower().rstrip(":")
                for field_name, aliases in _TITLE_LABELS.items():
                    if field_name in fields:
                        continue
                    if any(label == a or _best_column([label], [a]) >= 90
                           for a in aliases):
                        right = row[c + 1] if c + 1 < len(row) else None
                        below = (rows[r + 1][c] if r + 1 < len(rows)
                                 and c < len(rows[r + 1]) else None)
                        value = right if right not in (None, "") else below
                        if value not in (None, ""):
                            fields[field_name] = value
        return fields

    # ------------------------------------------------------------------ #
    def build_final(self, base: Candidate, folder: FolderResult,
                    all_folder_rows: list[dict]) -> Optional[Path]:
        """Build one template-formatted, fixed final from a base workbook.

        The template (if provided) is the formatting authority: it is copied
        verbatim and data is written into its report sheet. Otherwise the base
        workbook itself is the skeleton.
        """
        review = folder.review
        base_rows = self._sheet_rows(base.path)
        classification = base.classification

        # Authoritative scope + data from the base; merge non-conflicting fills
        # from other workbooks in the same folder (never adds new tracts).
        tracts = self.extract_tracts(base_rows, classification)
        ogl = self.extract_ogl(base_rows, classification)
        title = self.extract_title(base_rows, classification)

        for other_rows, other_cls in all_folder_rows:
            for k, v in self.extract_ogl(other_rows, other_cls).items():
                if k in tracts and (k not in ogl or not ogl[k].get("ogl_number")):
                    if v.get("ogl_number"):
                        ogl[k] = v  # fill a missing OGL number from a sibling

        if not tracts:
            review.append(ReviewItem("blocker", "tract sheet",
                          f"No tract sheet/rows found in {base.path.name}; "
                          f"cannot scope report. Examiner must confirm."))
            return None

        # Reconcile: only tracts on the tract sheet; attach OGL numbers.
        rows_out: list[dict] = []
        for key, t in tracts.items():
            og = ogl.get(key, {})
            if not og.get("ogl_number"):
                review.append(ReviewItem("missing_ogl", f"tract {key}",
                              "No OGL number on the OGL sheet for this tract."))
            if t.get("legal_description") in (None, ""):
                review.append(ReviewItem("missing_legal", f"tract {key}",
                              "Legal description blank; examiner must supply."))
            rows_out.append({
                "tract": t.get("tract", key),
                "legal_description": t.get("legal_description"),
                "acreage": t.get("acreage"),
                "ogl_number": og.get("ogl_number"),
                "lessor": og.get("lessor"),
                "lessee": og.get("lessee"),
            })

        if self.cfg.dry_run:
            self._log("build_final_dryrun",
                      f"{folder.folder.name}: would write {len(rows_out)} tract "
                      f"row(s) from base {base.path.name}",
                      {"tracts": len(rows_out), "title_fields": list(title)})
            return None

        # Skeleton = template copy if available, else the base workbook copy.
        skeleton = (self.cfg.template_path if self.cfg.template_path
                    and self.cfg.template_path.exists() else base.path)
        out_dir = self.cfg.output_dir
        out_dir.mkdir(parents=True, exist_ok=True)
        section = title.get("section") or folder.folder.name
        safe_section = "".join(ch if ch.isalnum() else "_" for ch in str(section))
        out_path = self._unique(out_dir /
                                f"{safe_section}_{folder.folder.name}_FINAL.xlsx")
        shutil.copy2(skeleton, out_path)

        wb = openpyxl.load_workbook(out_path)  # preserves template formatting
        try:
            self._write_into_report_sheet(wb, rows_out, title, review)
            wb.save(out_path)
        finally:
            wb.close()

        # Validate by reopening.
        try:
            chk = openpyxl.load_workbook(out_path, read_only=True)
            chk.close()
        except Exception as e:
            review.append(ReviewItem("blocker", "final workbook",
                          f"Output failed to reopen (corrupt): {e}"))
            return None

        if self.db:
            self.db.record_report_export(
                f"rogermills_{self.stamp}", "workbook", out_path.name,
                str(out_path), sha256_file(out_path),
                note=f"final from {base.path.name}")
        self._log("build_final", f"{folder.folder.name}: wrote {out_path.name}",
                  {"rows": len(rows_out)})
        return out_path

    def _write_into_report_sheet(self, wb, rows_out, title, review) -> None:
        """Write reconciled tract rows + title block into the skeleton's report
        sheet, mapping to the template's own column headers (format preserved)."""
        # Choose the report sheet: a tract-classified sheet, else the first.
        target = None
        for ws in wb.worksheets:
            low = ws.title.lower()
            if any(k in low for k in ("tract", "title report", "report")):
                target = ws
                break
        if target is None:
            target = wb.worksheets[0]

        # Find the header row and map our fields to the template's columns.
        header_row_idx = None
        headers: list[str] = []
        for r in range(1, min(target.max_row, 20) + 1):
            vals = [target.cell(r, c).value for c in range(1, target.max_column + 1)]
            strs = [v for v in vals if isinstance(v, str) and v.strip()]
            if len(strs) >= 2:
                header_row_idx = r
                headers = [str(v) if v is not None else "" for v in vals]
                break
        if header_row_idx is None:
            review.append(ReviewItem("blocker", "template",
                          "Could not locate a header row in the report sheet."))
            return

        norm = [h.strip().lower() for h in headers]
        colmap = {
            "tract": _best_column(norm, _TRACT_ID_ALIASES),
            "legal_description": _best_column(norm, _LEGAL_ALIASES),
            "acreage": _best_column(norm, _ACREAGE_ALIASES),
            "ogl_number": _best_column(norm, _OGL_NUM_ALIASES),
            "lessor": _best_column(norm, _LESSOR_ALIASES),
            "lessee": _best_column(norm, _LESSEE_ALIASES),
        }
        for field_name, idx in colmap.items():
            if idx < 0:
                review.append(ReviewItem("unmapped_column", field_name,
                              "Template has no matching column; value skipped "
                              "(examiner may add the column)."))

        # Write data rows starting just below the header, preserving styles by
        # only setting values (not overwriting cell styles/merges above).
        start = header_row_idx + 1
        for i, row in enumerate(rows_out):
            r = start + i
            for field_name, idx in colmap.items():
                if idx < 0:
                    continue
                target.cell(r, idx + 1).value = row.get(field_name)

        # Title block: write labeled values where labels exist in the sheet.
        self._write_title_block(target, title)

    def _write_title_block(self, ws, title: dict) -> None:
        if not title:
            return
        for r in range(1, min(ws.max_row, 30) + 1):
            for c in range(1, min(ws.max_column, 20) + 1):
                val = ws.cell(r, c).value
                if not isinstance(val, str):
                    continue
                label = val.strip().lower().rstrip(":")
                for field_name, aliases in _TITLE_LABELS.items():
                    if field_name in title and any(
                            label == a or _best_column([label], [a]) >= 90
                            for a in aliases):
                        # Prefer writing to the cell on the right if it's empty.
                        right = ws.cell(r, c + 1)
                        if right.value in (None, ""):
                            right.value = title[field_name]

    @staticmethod
    def _unique(path: Path) -> Path:
        if not path.exists():
            return path
        stem, suffix = path.stem, path.suffix
        n = 1
        while path.with_name(f"{stem}_{n:02d}{suffix}").exists():
            n += 1
        return path.with_name(f"{stem}_{n:02d}{suffix}")

    # ------------------------------------------------------------------ #
    def process_folder(self, folder: Path) -> FolderResult:
        result = FolderResult(folder=folder)
        workbooks = self.inventory_and_backup(folder)
        rows_cache: list[dict] = []
        for wb in workbooks:
            cand = self.analyze(wb)
            self.score(cand)
            if cand.is_template and result.template is None:
                result.template = cand
            if cand.is_final_copy and result.final_copy is None:
                result.final_copy = cand
            result.candidates.append(cand)

        # Tournament: rank data-bearing candidates (exclude blank template).
        ranked = sorted(
            [c for c in result.candidates if not c.is_template],
            key=lambda c: c.score, reverse=True)
        if not ranked:
            result.review.append(ReviewItem("blocker", "folder",
                                  f"No data workbook found in {folder}."))
            return result

        top_k = ranked[:max(1, self.cfg.top_k)]
        # Preload rows for cross-fill.
        all_rows = []
        for c in ranked:
            try:
                all_rows.append((self._sheet_rows(c.path), c.classification))
            except Exception:
                pass

        # Refinement loops: build a final from each of the top-K bases, keep the
        # one that yields the fewest unresolved review items + most tract rows.
        best_final = None
        best_metric = None
        for _loop in range(max(1, self.cfg.loops)):
            for base in top_k:
                trial_review_before = len(result.review)
                path = self.build_final(base, result, all_rows)
                added = len(result.review) - trial_review_before
                if self.cfg.dry_run:
                    # In dry-run nothing is written; the tournament is by score.
                    if best_metric is None or base.score > best_metric:
                        best_metric, result.best = base.score, base
                    continue
                if path is not None:
                    metric = -added  # fewer new review items is better
                    if best_metric is None or metric > best_metric:
                        best_metric, best_final, result.best = metric, path, base
            if not self.cfg.dry_run and best_final is not None:
                break  # a good final was produced; stop looping
        result.final_path = best_final
        return result

    def run(self) -> dict:
        summary = {"stamp": self.stamp, "dry_run": self.cfg.dry_run,
                   "folders": [], "finals": [], "review_total": 0}
        for folder in self.cfg.folders:
            if not folder.exists():
                self._log("folder_missing", f"Folder not found: {folder}")
                summary["folders"].append(
                    {"folder": str(folder), "status": "missing"})
                continue
            res = self.process_folder(folder)
            summary["folders"].append({
                "folder": str(folder),
                "candidates": len(res.candidates),
                "best": res.best.path.name if res.best else None,
                "final": str(res.final_path) if res.final_path else None,
                "review_items": len(res.review),
            })
            if res.final_path:
                summary["finals"].append(str(res.final_path))
            summary["review_total"] += len(res.review)
            self._write_review_file(res)
        self._log("finalize_done",
                  f"{len(summary['finals'])} final(s); "
                  f"{summary['review_total']} review item(s).")
        return summary

    def _write_review_file(self, res: FolderResult) -> None:
        if self.cfg.dry_run or not res.review:
            return
        out = self.cfg.output_dir / f"REVIEW_{res.folder.name}_{self.stamp}.md"
        lines = [f"# Examiner Review — {res.folder.name}", "",
                 f"_Generated {self.stamp}. These items are NOT auto-resolved; "
                 f"a human must confirm each._", "",
                 "| # | Kind | Subject | Detail |",
                 "|---|------|---------|--------|"]
        for i, r in enumerate(res.review, 1):
            lines.append(f"| {i} | {r.kind} | {r.subject} | "
                         f"{r.detail.replace('|', '/')} |")
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
