"""
Title <- OGL reconciler.

Encodes the examiner's rules for improving a cursory-title workbook WITHOUT
inventing facts:

  * The OGL register sheet is the source of truth for lease data. Build an
    OGL-number index (grantor/lessor -> OGL #, royalty, expiration, tracts).
  * On the Title sheet, for each mineral-owner row whose OGL No / Royalty /
    Expiration is blank, fill it ONLY when the owner name confidently matches an
    OGL grantor. Existing values are verified, not overwritten; a conflict is
    reported (never silently changed).
  * DO NOT alter tract legal descriptions or tract structure ("don't go off the
    tract sheets"). Only Title-sheet lease columns and hard-coded totals are
    touched.
  * Hard-coded tract totals are converted to SUM formulas over their owner rows.
  * Everything unmatched / conflicting is reported for examiner review — nothing
    is fabricated.

Output is written to a NEW file; the source is never overwritten.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

try:
    from rapidfuzz import fuzz

    def _sim(a: str, b: str) -> float:
        return fuzz.token_sort_ratio(a, b) / 100.0
except Exception:  # pragma: no cover
    from difflib import SequenceMatcher

    def _sim(a: str, b: str) -> float:
        return SequenceMatcher(None, a, b).ratio()


_SUFFIX = re.compile(r"\b(a/k/a|aka|trustee|trust|llc|l\.l\.c\.|inc|inc\.|company|co\.|"
                     r"revocable|living|family|the|of|and|&|lp|l\.p\.|fund|partners|"
                     r"resources|energy|royalty)\b", re.I)


def _norm(name: str) -> str:
    first = (name or "").split("\n")[0]
    first = first.split(",")[0]  # drop trailing ", Trustee of ..." and addresses
    first = _SUFFIX.sub(" ", first)
    first = re.sub(r"[^a-z0-9 ]", " ", first.lower())
    return re.sub(r"\s+", " ", first).strip()


@dataclass
class OGLEntry:
    ogl_no: str
    grantor: str
    grantor_norm: str
    royalty: object
    expiration: object
    tracts: str


@dataclass
class ReconcileReport:
    filled: List[Dict] = field(default_factory=list)
    verified: List[Dict] = field(default_factory=list)
    conflicts: List[Dict] = field(default_factory=list)
    unmatched: List[Dict] = field(default_factory=list)
    totals_fixed: List[str] = field(default_factory=list)

    def summary(self) -> Dict[str, int]:
        return {"filled": len(self.filled), "verified": len(self.verified),
                "conflicts": len(self.conflicts), "unmatched": len(self.unmatched),
                "totals_fixed": len(self.totals_fixed)}


def build_ogl_index(wb, ogl_sheet: str = "OGL") -> List[OGLEntry]:
    if ogl_sheet not in wb.sheetnames:
        return []
    ws = wb[ogl_sheet]
    # locate columns by header
    header = {str(c.value).strip().lower(): c.column for c in ws[1] if c.value}
    def col(*names):
        for n in names:
            for h, idx in header.items():
                if n in h:
                    return idx
        return None
    c_no = col("ogl #", "ogl#", "ogl no", "ogl")
    c_gr = col("grantor")
    c_roy = col("royalty")
    c_exp = col("exp")
    c_tr = col("tracts")
    entries = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, c_no).value if c_no else None
        gr = ws.cell(r, c_gr).value if c_gr else None
        if no is None or gr is None:
            continue
        entries.append(OGLEntry(
            ogl_no=str(no).strip(),
            grantor=str(gr),
            grantor_norm=_norm(str(gr)),
            royalty=ws.cell(r, c_roy).value if c_roy else None,
            expiration=ws.cell(r, c_exp).value if c_exp else None,
            tracts=str(ws.cell(r, c_tr).value) if c_tr and ws.cell(r, c_tr).value else "",
        ))
    return entries


def parse_tract_set(text: str) -> set:
    """Parse an OGL 'Tracts' cell like 'All (1-8)', '1, 3, 5', '1-3, 5' -> {ints}."""
    if not text:
        return set()
    t = text.lower()
    out = set()
    # 'all (1-8)' or 'all'
    m = re.search(r"all\s*\((\d+)\s*-\s*(\d+)\)", t)
    if m:
        return set(range(int(m.group(1)), int(m.group(2)) + 1))
    if t.strip() == "all":
        return set(range(1, 11))
    for part in re.split(r"[,\s]+", t):
        part = part.strip()
        mm = re.match(r"(\d+)\s*-\s*(\d+)$", part)
        if mm:
            out.update(range(int(mm.group(1)), int(mm.group(2)) + 1))
        elif part.isdigit():
            out.add(int(part))
    return out


def _tract_no(label: str) -> Optional[int]:
    m = re.search(r"tract\s*(\d+)", label, re.I)
    return int(m.group(1)) if m else None


def _best_ogl(owner_norm: str, index: List[OGLEntry], threshold: float,
              tract_no: Optional[int] = None) -> Optional[Tuple[OGLEntry, float]]:
    best, best_s = None, 0.0
    for e in index:
        if not e.grantor_norm:
            continue
        # If we know the tract, only consider OGLs whose Tracts column covers it.
        if tract_no is not None and e.tracts:
            cov = parse_tract_set(e.tracts)
            if cov and tract_no not in cov:
                continue
        s = _sim(owner_norm, e.grantor_norm)
        if s > best_s:
            best, best_s = e, s
    if best and best_s >= threshold:
        return best, best_s
    return None


TRACT_RE = re.compile(r"^\s*tract\s*\d+", re.I)


def reconcile(source_path: str, dest_path: str, *, match_threshold: float = 0.9,
              fill_blanks: bool = True, fix_totals: bool = True) -> ReconcileReport:
    """Produce an improved workbook at dest_path. Returns a reconciliation report."""
    src = Path(source_path)
    dst = Path(dest_path)
    if dst.exists():
        raise FileExistsError(f"Refusing to overwrite {dst}")
    wb = openpyxl.load_workbook(src, data_only=False)
    rep = ReconcileReport()
    index = build_ogl_index(wb)

    title = next((s for s in wb.sheetnames if s.strip().lower().startswith("title")), None)
    if title is None:
        wb.save(dst)
        return rep
    ws = wb[title]

    # Walk tract blocks; identify owner rows and total rows.
    rows = list(ws.iter_rows())
    headers = [(i, str(r[1].value).strip()) for i, r in enumerate(rows)
               if len(r) > 1 and isinstance(r[1].value, str) and TRACT_RE.match(r[1].value)]
    for bi, (i, label) in enumerate(headers):
        end = headers[bi + 1][0] if bi + 1 < len(headers) else len(rows)
        owner_first = owner_last = None
        total_row = None
        for j in range(i + 1, end):
            excel_row = j + 1
            b = rows[j][1].value if len(rows[j]) > 1 else None
            c = rows[j][2].value if len(rows[j]) > 2 else None
            d = rows[j][3].value if len(rows[j]) > 3 else None
            bu = b.strip().upper() if isinstance(b, str) else ""
            if bu in ("REPORT TOTAL", "TOTAL"):
                if total_row is None and isinstance(c, (int, float)):
                    total_row = excel_row
                continue
            if not isinstance(b, str) or bu.startswith("MINERAL OWNER") or bu.startswith("CONTAINING"):
                continue
            # numeric net-acre cell marks an owner row
            is_num = isinstance(c, (int, float))
            if is_num:
                if owner_first is None:
                    owner_first = excel_row
                owner_last = excel_row
            if b.strip().lower().startswith("balance"):
                continue
            # OGL reconciliation for this owner (tract-coverage aware)
            owner_norm = _norm(b)
            hit = _best_ogl(owner_norm, index, match_threshold, _tract_no(label)) if owner_norm else None
            existing = str(d).strip() if d not in (None, "", "—") else ""
            if hit:
                e, score = hit
                if existing:
                    # verify: does the existing OGL# list include the matched one?
                    if e.ogl_no not in re.split(r"[,\s]+", existing):
                        rep.conflicts.append({"tract": label, "row": excel_row,
                                              "owner": b.split("\n")[0], "existing_ogl": existing,
                                              "ogl_match": e.ogl_no, "grantor": e.grantor,
                                              "score": round(score, 2)})
                    else:
                        rep.verified.append({"tract": label, "owner": b.split("\n")[0],
                                             "ogl": existing})
                elif fill_blanks:
                    ws.cell(excel_row, 4).value = e.ogl_no
                    if ws.cell(excel_row, 5).value in (None, ""):
                        ws.cell(excel_row, 5).value = e.royalty
                    if ws.cell(excel_row, 6).value in (None, ""):
                        ws.cell(excel_row, 6).value = e.expiration
                    rep.filled.append({"tract": label, "row": excel_row,
                                       "owner": b.split("\n")[0], "ogl": e.ogl_no,
                                       "grantor": e.grantor, "score": round(score, 2)})
            else:
                if not existing and is_num and c and float(c) > 0:
                    rep.unmatched.append({"tract": label, "owner": b.split("\n")[0]})

        # fix hard-coded total -> SUM formula
        if fix_totals and total_row and owner_first and owner_last:
            cell = ws.cell(total_row, 3)
            if isinstance(cell.value, (int, float)):
                cell.value = f"=SUM(C{owner_first}:C{owner_last})"
                rep.totals_fixed.append(f"{label} C{total_row}")

    wb.save(dst)
    wb.close()
    return rep
