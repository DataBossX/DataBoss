"""Shared pytest fixtures.

Every fixture here is TEST DATA ONLY and must never be treated as a live,
verified legal/title/source fact.
"""
from __future__ import annotations

import sys
import zipfile
from pathlib import Path

import pytest

_AGENT_ROOT = Path(__file__).resolve().parent.parent
if str(_AGENT_ROOT) not in sys.path:
    sys.path.insert(0, str(_AGENT_ROOT))

import openpyxl  # noqa: E402

from db.db_client import DatabaseManager  # noqa: E402
from db.audit_logger import AuditLogger  # noqa: E402

SCHEMA = _AGENT_ROOT / "db" / "schema.sql"


@pytest.fixture
def db(tmp_path):
    mgr = DatabaseManager(tmp_path / "test.sqlite3", SCHEMA)
    yield mgr
    mgr.close()


@pytest.fixture
def audit(db):
    return AuditLogger(db)


@pytest.fixture
def sample_workbook(tmp_path) -> Path:
    """A small .xlsx FIXTURE with formulas, an error literal, and a fake media
    entry so media-preservation can be asserted. NOT real title data."""
    path = tmp_path / "sample_[FIXTURE].xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tract 1"
    ws["A1"] = "FIXTURE - not real title data"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["A4"] = "=A2+A3"        # good formula
    ws["B4"] = "#REF!"          # error literal to be detected/repaired
    wb.create_sheet("OGL")
    wb.create_sheet("WI 1")
    wb.save(path)

    # Inject a fake media part so the XML editor's preservation can be tested.
    tmp2 = tmp_path / "with_media_[FIXTURE].xlsx"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp2, "w") as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        zout.writestr("xl/media/image1.png", b"\x89PNG\r\n\x1a\nFIXTURE")
        zout.writestr("xl/drawings/drawing1.xml", b"<xdr/>")
    return tmp2
