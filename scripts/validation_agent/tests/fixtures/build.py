"""
Test fixture builders.

IMPORTANT: everything here is SYNTHETIC TEST DATA. None of these names, book/page
numbers, or acreages are real title facts and must never be treated as verified
source facts.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl


def build_workbook(path: str | Path) -> Path:
    """A compact but structurally complete fixture workbook (SYNTHETIC)."""
    path = Path(path)
    wb = openpyxl.Workbook()
    # ---- Title sheet with two tract blocks ----
    ws = wb.active
    ws.title = "Title "
    ws["B4"] = "TRACT 1: FIXTURE"
    ws["C4"] = "N/2 NE/4"
    ws["B5"] = "Containing:"
    ws["C5"] = "40 acres, more or less"
    ws["B7"] = "Mineral Owner"; ws["C7"] = "Net Acres"; ws["D7"] = "OGL No."; ws["E7"] = "Royalty"
    # 'Fixture Lessor' matches the OGL register grantor below (blank OGL -> should fill).
    ws["B8"] = "Fixture Lessor"; ws["C8"] = 25
    ws["B9"] = "Fixture Owner B"; ws["C9"] = 15
    ws["B10"] = "REPORT TOTAL"; ws["C10"] = 40   # HARD-CODED total -> formula defect (repairable)
    ws["B11"] = "TOTAL"; ws["C11"] = 40

    ws["B14"] = "TRACT 2: FIXTURE"
    ws["C14"] = "SW/4 NE/4"
    ws["B15"] = "Containing:"
    ws["C15"] = "40 acres, more or less"
    ws["B17"] = "Mineral Owner"; ws["C17"] = "Net Acres"
    ws["B18"] = "Fixture Owner C"; ws["C18"] = 30
    ws["B19"] = "Fixture Owner D"; ws["C19"] = 25   # 30+25 = 55 > 40 -> OVER-conveyance
    ws["B20"] = "Balance undetermined of record"; ws["C20"] = 0
    ws["B21"] = "REPORT TOTAL"; ws["C21"] = "=SUM(C18:C20)"
    ws["B22"] = "TOTAL"; ws["C22"] = "=C21"
    # A chain-gap marker for the chain validator.
    ws["G18"] = "vesting instrument not located of record"

    # ---- 10 tract sheets so the core gate is satisfiable ----
    for i in range(1, 11):
        s = wb.create_sheet(f"Tract {i}")
        s["A1"] = "Mineral Owner"; s["B1"] = "Net Acres"

    # ---- OGL register with expected columns ----
    ogl = wb.create_sheet("OGL")
    ogl.append(["OGL", "Instrument", "Bk/Pg", "Grantor", "Grantee", "Legal", "Royalty"])
    ogl.append([1, "OGL", "1736/0592", "Fixture Lessor", "Fixture Lessee",
                "N/2 NE/4", "3/16"])

    # ---- WI sheets ----
    wb.create_sheet("WI 1")["A1"] = "Working Interest"
    wb.create_sheet("WI 2")["A1"] = "Working Interest"

    # ---- Well sheet ----
    well = wb.create_sheet("Well 1")
    well["A2"] = "Well #"; well["B2"] = "API #"
    well["A3"] = "1-31"; well["B3"] = "35-129-99999"
    well["C3"] = "HBP per recital only; not confirmed"

    wb.save(path)
    return path


def build_workbook_with_media(src: str | Path, dst: str | Path) -> Path:
    """Copy a workbook and inject a media + drawing entry (SYNTHETIC)."""
    src, dst = Path(src), Path(dst)
    png = (b"\x89PNG\r\n\x1a\n" + b"\x00" * 32)  # minimal fake PNG bytes
    drawing = b'<?xml version="1.0"?><xdr:wsDr xmlns:xdr="x"/>'
    with zipfile.ZipFile(src) as zin:
        names = zin.namelist()
        data = {n: zin.read(n) for n in names}
    data["xl/media/image1.png"] = png
    data["xl/drawings/drawing1.xml"] = drawing
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, b in data.items():
            zout.writestr(n, b)
    return dst
