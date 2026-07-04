"""Builders for TEST FIXTURES ONLY.

!!! These workbooks are synthetic test fixtures. They are NEVER to be used as
live, verified source facts. Every value here is fabricated for testing the
validation machinery and carries no legal meaning. !!!
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl


FIXTURE_BANNER = "SYNTHETIC TEST FIXTURE - NOT A VERIFIED SOURCE"


def make_basic_workbook(path: Path, *, with_hardcoded_total: bool = False) -> Path:
    """Create a small workbook with tracts / OGL / working-interest sheets."""
    wb = openpyxl.Workbook()
    tracts = wb.active
    tracts.title = "Tracts"
    tracts.append(["Tract", "Legal Description", "Acreage", FIXTURE_BANNER])
    # 10 tracts summing to 637.42 (fixture math only).
    acreages = [63.742] * 10
    for i, ac in enumerate(acreages, 1):
        tracts.append([f"T{i}", f"Sec {i} T1N R1W", ac, ""])
    if with_hardcoded_total:
        # A footing cell that should be =SUM(C2:C11) but is hardcoded.
        tracts["C13"] = 600.00  # wrong hardcoded literal (formula defect)
        tracts["C12"] = "=SUM(C2:C11)"  # neighbor formula (template evidence)
    else:
        tracts["C12"] = "=SUM(C2:C11)"

    ogl = wb.create_sheet("OGL Register")
    ogl.append(["Lessor", "Lessee", "Date", "Book/Page", "Royalty"])
    ogl.append(["A Owner", "B Operator", "2020-01-01", "bk 100/pg 200", "1/8"])

    wi = wb.create_sheet("Working Interest")
    wi.append(["Party", "WI", "Depth Top", "Depth Base"])
    wi.append(["B Operator", "1.0", "0", "10000"])

    wb.save(path)
    return path


def make_full_workbook(path: Path) -> Path:
    """A complete, internally-consistent fixture that auto-extraction can drive
    all the way to CERTIFIED. Every value is synthetic (see banner)."""
    wb = openpyxl.Workbook()

    tracts = wb.active
    tracts.title = "Tracts"
    tracts.append(["Tract", "Legal Description", "Acreage", FIXTURE_BANNER])
    for i in range(1, 11):
        # Non-S/T/R descriptions so no spurious instrument references appear.
        tracts.append([f"T{i}", f"Lot {i}, Block A", 63.742, ""])
    tracts["C12"] = "=SUM(C2:C11)"

    run = wb.create_sheet("Runsheet")
    run.append(["Tract", "Grantor", "Grantee", "Grantor Interest",
                "Conveyed", "Retained", "Vested Root"])
    # Root conveyance: US -> Alice (whole), then Alice -> Bob (whole).
    run.append(["T1", "US", "Alice", "1/1", "1/1", "0", "yes"])
    run.append(["T1", "Alice", "Bob", "1/1", "1/2", "1/2", ""])

    ogl = wb.create_sheet("OGL Register")
    ogl.append(["Lessor", "Lessee", "Date", "Book/Page", "Royalty"])
    ogl.append(["A Owner", "B Operator", "2020-01-01", "bk 100/pg 200", "1/8"])

    wi = wb.create_sheet("Working Interest")
    wi.append(["Party", "WI", "Depth Top", "Depth Base"])
    wi.append(["B Operator", "1.0", 0, 10000])

    # Sources sheet lists the referenced instrument so the source audit passes.
    src = wb.create_sheet("Sources")
    src.append(["Reference", "Note"])
    src.append(["bk 100/pg 200", "recorded lease (fixture)"])

    wb.save(path)
    return path


def inject_media(path: Path) -> Path:
    """Add fake drawing + media entries to an existing xlsx (fixture only).

    We inject raw ZIP entries so the XML editor's preservation guarantee can be
    tested without requiring Pillow.
    """
    # Read existing entries.
    with zipfile.ZipFile(path) as zf:
        entries = {n: zf.read(n) for n in zf.namelist()}
    entries["xl/drawings/drawing1.xml"] = (
        b'<?xml version="1.0"?><xdr:wsDr '
        b'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/'
        b'spreadsheetDrawing"/>'
    )
    entries["xl/media/image1.png"] = b"\x89PNG\r\n\x1a\n-FAKE-FIXTURE-IMAGE"
    # Rewrite the archive.
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in entries.items():
            zf.writestr(name, data)
    return path


def make_report(path: Path) -> Path:
    """Create a candidate markdown report fixture."""
    path.write_text(
        f"# Prospect Title Report ({FIXTURE_BANNER})\n\n"
        "## Tracts\nSee attached. Book 100 Page 200 recorded.\n\n"
        "## OGL Register\nLease recorded, instrument 12345.\n\n"
        "## Working Interest\nWI schedule attached.\n\n"
        "## Chain of Title\nGrantor to grantee conveyance.\n\n"
        "Certification pending examiner review.\n",
        encoding="utf-8",
    )
    return path
