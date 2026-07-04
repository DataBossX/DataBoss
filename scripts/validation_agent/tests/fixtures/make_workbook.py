"""Deterministic workbook fixtures for tests.

Builds small but structurally realistic workbooks with openpyxl. No fake
production data — these are synthetic test tracts/leases used only by the test
suite.
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook


# 10 tracts summing to 637.42 acres.
_TRACT_ACRES: List[str] = [
    "80.00", "64.00", "100.00", "160.00", "20.00",
    "37.42", "56.00", "48.00", "52.00", "20.00",
]


def build_valid_workbook(path: Path, total_override: Optional[str] = None,
                         break_tract: Optional[float] = None) -> Path:
    wb = Workbook()

    # Tracts sheet
    ws = wb.active
    ws.title = "Tracts"
    ws.append(["Tract", "Legal Description", "OGL", "Net Acres"])
    acres = list(_TRACT_ACRES)
    if break_tract is not None:
        acres[0] = str(break_tract)
    for i, a in enumerate(acres, start=1):
        ws.append([f"Tract {i}", f"SEC {i} T7N R63W", f"OGL-{i:03d}", float(a)])
    total = total_override if total_override is not None else str(
        sum(float(x) for x in acres)
    )
    ws.append(["Total", "", "", float(total)])

    # OGL register
    ogl = wb.create_sheet("OGL Register")
    ogl.append(["OGL", "Lessor", "Lessee", "Book", "Page"])
    for i in range(1, 11):
        ogl.append([f"OGL-{i:03d}", f"Lessor {i}", "DataBossX Energy", 100 + i, 200 + i])

    # Runsheet / chain
    run = wb.create_sheet("Runsheet")
    run.append(["Grantor", "Grantee", "Instrument Type", "Book", "Page",
                "Grantor Interest", "Conveyed", "Retained"])
    run.append(["State of Oklahoma", "Alice Root", "Patent", 1, 1, "1", "1", "0"])
    run.append(["Alice Root", "Bob Buyer", "Warranty Deed", 101, 201,
                "1", "1/2", "1/2"])
    run.append(["Bob Buyer", "Carol Corp", "Warranty Deed", 102, 202,
                "1/2", "1/2", "0"])

    # Working interest
    wi = wb.create_sheet("Working Interest")
    wi.append(["OGL", "Owner", "Working Interest", "Depth", "Grantor Interest", "Conveyed"])
    wi.append(["OGL-001", "Carol Corp", "0.5", "Surface to 8000 ft", "1", "1"])

    # Wells
    wells = wb.create_sheet("Wells")
    wells.append(["Well Name", "API", "Status"])
    wells.append(["Test Well 1", "3500112345", "Plugged"])

    wb.save(path)
    return Path(path)


def build_overconveyance_workbook(path: Path) -> Path:
    """A runsheet where a grantor conveys more than they own."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracts"
    ws.append(["Tract", "OGL", "Net Acres"])
    for i, a in enumerate(_TRACT_ACRES, start=1):
        ws.append([f"Tract {i}", f"OGL-{i:03d}", float(a)])

    run = wb.create_sheet("Runsheet")
    run.append(["Grantor", "Grantee", "Book", "Page",
                "Grantor Interest", "Conveyed", "Retained"])
    run.append(["State of Oklahoma", "Alice Root", 1, 1, "1", "1", "0"])
    # Alice owns 1/2 but conveys 3/4 -> over-conveyance
    run.append(["Alice Root", "Greedy Grantee", 101, 201, "1/2", "3/4", "0"])
    wb.save(path)
    return Path(path)


def build_certifiable_workbook(
    path: Path, defect: str = "none"
) -> Path:
    """A fully clean workbook that certifies once its sources are on disk.

    ``defect`` injects exactly one *safe, repairable* defect:
      * ``"acreage_total"`` — hardcoded wrong total (formula-defect, SUM restore)
      * ``"formula_column"`` — hardcoded literal in a formula-driven column
      * ``"none"`` — no defect (should certify with no repair)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracts"
    ws.append(["Tract", "OGL", "Net Acres", "Check"])
    for i, a in enumerate(_TRACT_ACRES, start=1):
        row_num = i + 1
        check = f"=C{row_num}"
        if defect == "formula_column" and i == 3:
            check = 100.0  # hardcoded literal in the formula-driven "Check" column
        ws.append([f"Tract {i}", f"OGL-{i:03d}", float(a), check])
    correct_total = sum(float(x) for x in _TRACT_ACRES)
    total = "1000.00" if defect == "acreage_total" else str(correct_total)
    ws.append(["Total", "", float(total), ""])

    ogl = wb.create_sheet("OGL Register")
    ogl.append(["OGL", "Lessor", "Lessee", "Book", "Page"])
    for i in range(1, 11):
        ogl.append([f"OGL-{i:03d}", f"Lessor {i}", "DataBossX Energy", 100 + i, 200 + i])

    run = wb.create_sheet("Runsheet")
    run.append(["Grantor", "Grantee", "Instrument Type", "Book", "Page",
                "Grantor Interest", "Conveyed", "Retained"])
    run.append(["State of Oklahoma", "Alice Root", "Patent", 1, 1, "1", "1", "0"])
    run.append(["Alice Root", "Bob Buyer", "Warranty Deed", 101, 201, "1", "1", "0"])

    wi = wb.create_sheet("Working Interest")
    wi.append(["OGL", "Owner", "Working Interest", "Depth", "Grantor Interest", "Conveyed"])
    wi.append(["OGL-001", "Bob Buyer", "0.5", "Surface to 8000 ft", "1", "1"])

    wb.save(path)
    return Path(path)


def materialize_sources(workbook: Path, dest_dir: Path) -> Path:
    """Create one placeholder source file per reference the workbook cites."""
    from ingestion.sheet_classifier import classify_workbook
    from ingestion.workbook_ingestor import ingest_workbook
    from sources.source_finder import build_reference_queue

    dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)
    structure = ingest_workbook(workbook)
    classification = classify_workbook(structure)
    for ref in build_reference_queue(structure, classification):
        (dest_dir / f"{ref.reference_key}_src.pdf").write_bytes(b"%PDF-1.4 placeholder")
    return dest_dir


def build_missing_vesting_workbook(path: Path) -> Path:
    """A grantor with no prior vesting and no root marker -> chain gap."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracts"
    ws.append(["Tract", "OGL", "Net Acres"])
    for i, a in enumerate(_TRACT_ACRES, start=1):
        ws.append([f"Tract {i}", f"OGL-{i:03d}", float(a)])

    run = wb.create_sheet("Runsheet")
    run.append(["Grantor", "Grantee", "Instrument Type", "Book", "Page"])
    run.append(["Mystery Grantor", "New Owner", "Warranty Deed", 500, 600])
    wb.save(path)
    return Path(path)
