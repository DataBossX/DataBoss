from pathlib import Path
import openpyxl
from ingestion.domain_extractor import extract_references, extract_domain, extract_well
from sources.source_finder import SourceFinder, parse_references
from sources.okcounty_client import OKCountyClient
from sources.spend_guard import SpendGuard
from config.settings import Settings
from decimal import Decimal

def _wb_with_refs(tmp_path):
    p = tmp_path/"refs_[FIXTURE].xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Title"
    ws["A1"]="TRACT 1: N/2 NE/4"
    ws["A2"]="Containing: 80 acres, more or less"
    ws["A3"]="Base OGL 1736/0594; instrument 2021-001068; well API 35-129-22925 active"
    wb.save(p); return p

def test_extract_references(tmp_path):
    refs = extract_references(_wb_with_refs(tmp_path))
    assert "1736/0594" in refs["book_page"]
    assert "2021-001068" in refs["instrument_number"]

def test_extract_domain_no_owner_fabrication(tmp_path):
    d = extract_domain(_wb_with_refs(tmp_path))
    assert d["reference_count"] >= 2
    # tract detected with acreage but NO owners invented
    t = d["tracts"][0]
    assert t["tract"] == 1 and t["gross_acres"] == "80"
    assert t["owners"] is None and t["conveyances"] is None

def test_extract_well_no_hbp_assertion(tmp_path):
    w = extract_well(_wb_with_refs(tmp_path))
    assert w["api"] == "35-129-22925"
    assert w["hbp_supported"] is None  # never asserted by extractor

def test_source_finder_queues_and_escalates_in_dryrun(tmp_path, audit, monkeypatch):
    monkeypatch.setenv("DATABOSSX_DRY_RUN","true")
    s = Settings()
    guard = SpendGuard(audit, Decimal("100.00"), "r")
    client = OKCountyClient(s, audit, guard, "r")
    finder = SourceFinder(s, audit, client, "r", search_dirs=[tmp_path/"none"])
    out = finder.build_and_resolve_queue({"book_page":["1736/0594"],"instrument_number":[]})
    assert out[0]["status"] in ("queued","blocked")           # never 'retrieved' in dry-run
    # ONE aggregate escalation logged; full per-doc queue in source_documents
    esc = audit.db.query("SELECT subject, reason FROM escalations WHERE run_id='r'")
    assert len(esc) == 1 and "1736/0594" in esc[0]["reason"]
    q = audit.db.query("SELECT book_page FROM source_documents WHERE run_id='r'")
    assert any(r["book_page"]=="1736/0594" for r in q)
    # no paid spend occurred
    assert guard.cumulative() == Decimal("0.00")

def test_okcounty_dry_run_makes_no_paid_call(audit, monkeypatch):
    monkeypatch.setenv("DATABOSSX_DRY_RUN","true")
    s = Settings()
    guard = SpendGuard(audit, Decimal("100.00"), "r")
    client = OKCountyClient(s, audit, guard, "r")
    res = client.retrieve_document("1736/0594", Decimal("5.00"))
    assert res["status"] == "blocked" and res["reason"] == "dry_run"
    assert guard.cumulative() == Decimal("0.00")

def test_local_source_found_marks_found(tmp_path, audit, monkeypatch):
    # a local file whose name contains the reference token is 'found_local'
    monkeypatch.setenv("DATABOSSX_DRY_RUN","true")
    srcdir = tmp_path/"srcs"; srcdir.mkdir()
    (srcdir/"deed_1736_0594.pdf").write_bytes(b"FIXTURE")
    s = Settings(); guard = SpendGuard(audit, Decimal("100.00"), "r")
    client = OKCountyClient(s, audit, guard, "r")
    finder = SourceFinder(s, audit, client, "r", search_dirs=[srcdir])
    out = finder.build_and_resolve_queue({"book_page":["1736/0594"],"instrument_number":[]})
    assert out[0]["status"] == "found_local"
