"""Phase 16: end-to-end perfection-loop runs.

Two scenarios prove the two terminal behaviours the blueprint requires:

    * a repairable workbook is fixed at the XML level and CERTIFIED, and
    * a workbook with a missing probate HALTS and ESCALATES without inventing
      the missing legal fact.

Both assert the source workbook is never mutated.
"""

from __future__ import annotations

import pytest

pytest.importorskip("openpyxl")
pytest.importorskip("lxml")

from validation_agent.config.settings import config  # noqa: E402
from validation_agent.core.orchestrator import (PerfectionLoopOrchestrator,  # noqa: E402
                                                STATE_CERTIFY, STATE_ESCALATE)
from validation_agent.core.run_manager import RunManager, sha256_of  # noqa: E402
from validation_agent.db.audit_logger import AuditLogger  # noqa: E402
from validation_agent.db.db_client import DatabaseManager  # noqa: E402
from validation_agent.reports.output_generator import OutputGenerator  # noqa: E402
from validation_agent.tests.make_fixture import (build_certifiable_workbook,  # noqa: E402
                                                 build_escalation_workbook)


def _orchestrator(tmp_path):
    dbm = DatabaseManager(db_path=tmp_path / "run.db",
                          schema_path=config.schema_path)
    dbm.init()
    audit = AuditLogger(dbm)
    return PerfectionLoopOrchestrator(audit=audit, run_manager=RunManager(audit))


def test_repairable_workbook_certifies(tmp_path):
    src = build_certifiable_workbook(tmp_path / "clean.xlsx")
    src_sha = sha256_of(src)

    orch = _orchestrator(tmp_path)
    outcome = orch.run(src, timestamp="20260101_000001")

    assert outcome.final_state == STATE_CERTIFY
    assert outcome.certified is True
    assert not outcome.escalations
    # At least one iteration applied a safe XML repair.
    assert any(rec.plan.safe_repairs for rec in outcome.iterations)
    # Source workbook is byte-for-byte unchanged (Golden Law #5).
    assert sha256_of(src) == src_sha

    paths = OutputGenerator().generate(outcome)
    assert "certification" in paths
    assert (outcome.ctx.run_folder / "certification.md").is_file()


def test_missing_probate_halts_and_escalates(tmp_path):
    src = build_escalation_workbook(tmp_path / "esc.xlsx")
    src_sha = sha256_of(src)

    orch = _orchestrator(tmp_path)
    outcome = orch.run(src, timestamp="20260101_000002")

    assert outcome.final_state == STATE_ESCALATE
    assert outcome.certified is False
    categories = {e.category.value for e in outcome.escalations}
    assert "title_chain_gap" in categories
    # No fabrication: source unchanged and nothing certified.
    assert sha256_of(src) == src_sha

    paths = OutputGenerator().generate(outcome)
    assert "certification" not in paths
    assert not (outcome.ctx.run_folder / "certification.md").exists()
    written = paths.get("escalations", "[]")
    assert "ESC-" in written


def test_escalate_always_produces_a_packet(tmp_path):
    """Escalating on structural ERROR gates (e.g. a missing runsheet) must still
    hand the examiner at least one full packet -- never an empty escalation."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for i in range(1, 11):
        ws = wb.create_sheet(f"Tract {i}")
        ws.append(["Tract", "Net Acres"])
        ws.append([i, 63.742])
    ws = wb.create_sheet("OGL Register")
    ws.append(["OGL", "Lessor", "Depth", "Book", "Page"])
    ws.append(["L-1", "ALICE", "X", "5", "100"])
    ws = wb.create_sheet("Working Interest")
    ws.append(["OGL", "Depth"])
    ws.append(["L-1", "X"])
    src = tmp_path / "norunsheet.xlsx"
    wb.save(src)

    orch = _orchestrator(tmp_path)
    outcome = orch.run(src, timestamp="20260101_000004")
    assert outcome.final_state == STATE_ESCALATE
    assert outcome.escalations, "escalation must never be empty"
    assert "0 item(s)" not in outcome.message


def test_escalation_payload_has_all_nine_fields(tmp_path):
    src = build_escalation_workbook(tmp_path / "esc.xlsx")
    orch = _orchestrator(tmp_path)
    outcome = orch.run(src, timestamp="20260101_000003")
    assert outcome.escalations
    esc = outcome.escalations[0]
    d = esc.to_dict()
    for field in ("failed_gate", "location", "subject", "sources_checked",
                  "sources_missing", "automated_checks_attempted",
                  "repair_blocked_reason", "recommended_examiner_action",
                  "estimated_urgency"):
        assert field in d
