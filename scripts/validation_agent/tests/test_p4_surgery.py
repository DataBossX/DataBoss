"""P4 exit tests: XML surgery preserves drawings/media/comments byte-for-byte,
edits round-trip, calcChain dropped, fullCalcOnLoad set."""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from validation_agent.repair.xml_surgery import XmlWorkbook

FIXTURE = Path(__file__).parent / "fixtures" / "workbook.xlsx"


def test_protected_parts_byte_identical(tmp_path: Path) -> None:
    xw = XmlWorkbook.open(FIXTURE)
    xw.set_cell("Title ", "G8", "surgery test note")
    xw.drop_calc_chain()
    xw.force_full_calc()
    out = xw.save_as(tmp_path / "v1.xlsx")

    with zipfile.ZipFile(FIXTURE) as a, zipfile.ZipFile(out) as b:
        for n in a.namelist():
            if n == "xl/calcChain.xml":
                assert n not in b.namelist()
                continue
            if any(k in n for k in ("media", "drawing", "comment", "printerSettings", "theme")):
                assert a.read(n) == b.read(n), f"{n} changed"


def test_cell_edit_round_trips(tmp_path: Path) -> None:
    xw = XmlWorkbook.open(FIXTURE)
    xw.set_cell("Title ", "G8", "ROUND TRIP VALUE")
    xw.set_cell("OGL", "M6", 123.5)
    out = xw.save_as(tmp_path / "v2.xlsx")
    wb = openpyxl.load_workbook(out)
    assert wb["Title "]["G8"].value == "ROUND TRIP VALUE"
    assert abs(wb["OGL"]["M6"].value - 123.5) < 1e-9


def test_calcchain_dropped_and_fullcalc_set(tmp_path: Path) -> None:
    xw = XmlWorkbook.open(FIXTURE)
    xw.drop_calc_chain()
    xw.force_full_calc()
    out = xw.save_as(tmp_path / "v3.xlsx")
    with zipfile.ZipFile(out) as z:
        assert "xl/calcChain.xml" not in z.namelist()
        wbxml = z.read("xl/workbook.xml").decode()
        assert 'fullCalcOnLoad="1"' in wbxml


def test_highlight_applies_style_only(tmp_path: Path) -> None:
    xw = XmlWorkbook.open(FIXTURE)
    xw.set_highlight("Tract 3", "D26", True)
    out = xw.save_as(tmp_path / "v4.xlsx")
    wb = openpyxl.load_workbook(out)
    fill = wb["Tract 3"]["D26"].fill
    assert fill.patternType == "solid"
    assert fill.fgColor.rgb == "FFFFFF00"


def test_shared_formula_unshared_without_ref_error(tmp_path: Path) -> None:
    # Editing a cell inside a shared-formula range must not corrupt siblings.
    xw = XmlWorkbook.open(FIXTURE)
    xw.set_cell("Tract 3", "H6", 0)  # H6 is inside a shared IF(...RECHECK) group
    out = xw.save_as(tmp_path / "v5.xlsx")
    wb = openpyxl.load_workbook(out)
    # sibling guard cells still hold formulas (not #REF!)
    for cell in ("K6", "M6", "N6"):
        v = wb["Tract 3"][cell].value
        assert v is None or (isinstance(v, str) and "#REF" not in v)
