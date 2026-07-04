"""LibreOffice recalc: a missing dependency escalates cleanly (no fake success)
and the runner never overwrites an existing version file."""

import openpyxl

from recalc import libreoffice_runner as lo
from recalc.libreoffice_runner import compare_key_cells
from tests.fixtures.make_fixtures import make_basic_workbook


def test_missing_libreoffice_escalates_cleanly(tmp_path, monkeypatch):
    monkeypatch.setattr(lo, "detect_libreoffice", lambda *a, **k: None)
    src = make_basic_workbook(tmp_path / "wb.xlsx")
    res = lo.recalculate(src, tmp_path / "out.xlsx")
    assert res.ok is False
    assert res.available is False
    assert "not found" in res.error.lower()
    assert res.output_path is None
    assert not (tmp_path / "out.xlsx").exists()   # no fake output


def test_compare_key_cells_detects_intact(tmp_path):
    src = make_basic_workbook(tmp_path / "wb.xlsx")
    # A byte-identical copy is trivially intact.
    import shutil
    dest = tmp_path / "copy.xlsx"
    shutil.copy2(src, dest)
    cmp = compare_key_cells(src, dest)
    assert cmp["corruption"] is False
    assert cmp["missing_sheets"] == []
    assert cmp["retained_fraction"] == 1.0


def test_compare_key_cells_detects_lost_sheet(tmp_path):
    src = make_basic_workbook(tmp_path / "wb.xlsx")
    # Build a "recalc output" that dropped the OGL Register + Working Interest.
    wb = openpyxl.load_workbook(src)
    for name in ("OGL Register", "Working Interest"):
        if name in wb.sheetnames:
            del wb[name]
    bad = tmp_path / "bad.xlsx"
    wb.save(bad)
    cmp = compare_key_cells(src, bad)
    assert cmp["corruption"] is True
    assert "OGL Register" in cmp["missing_sheets"]


def test_compare_key_cells_detects_data_loss(tmp_path):
    src = make_basic_workbook(tmp_path / "wb.xlsx")
    # Same sheets, but nearly all cells wiped -> gross data loss.
    wb = openpyxl.load_workbook(src)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    bad = tmp_path / "empty.xlsx"
    wb.save(bad)
    cmp = compare_key_cells(src, bad)
    assert cmp["corruption"] is True
    assert cmp["retained_fraction"] < 0.5


def test_recalc_refuses_to_overwrite(tmp_path, monkeypatch):
    # Even with a "present" LibreOffice, an existing dest is never clobbered.
    monkeypatch.setattr(lo, "detect_libreoffice",
                        lambda *a, **k: "/usr/bin/soffice")
    src = make_basic_workbook(tmp_path / "wb.xlsx")
    dest = tmp_path / "out.xlsx"
    dest.write_text("existing good version")
    res = lo.recalculate(src, dest)
    assert res.ok is False
    assert "overwrite" in res.error.lower()
    assert dest.read_text() == "existing good version"
