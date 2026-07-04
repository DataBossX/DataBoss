"""Integration test: the full REPAIR -> RECALC -> CERTIFY happy path.

Builds a 10-tract fixture workbook with one formula error, supplies a clean
domain context and the known intended formula, and asserts the orchestrator
performs a real safe repair, a headless LibreOffice recalc (skipped cleanly if
LibreOffice is unavailable), and certifies. FIXTURE data only.
"""
import shutil
import pytest
import openpyxl
from pathlib import Path

from config.settings import get_settings
from core.orchestrator import Orchestrator
from recalc.libreoffice_runner import detect_libreoffice


def _clean_domain():
    # One tract footing exactly to expected acres; owner has source-in; well
    # supported by production; OGL/WI computable; instrument present.
    return {
        "tracts": [{"tract": 1, "gross_acres": "637.42",
                    "owners": [{"name": "FIXTURE Owner", "net_acres": "637.42", "source_in": "deed 1/1"}],
                    "open_balance": "0",
                    "conveyances": [{"grantor": "G", "interest_in": "1/1", "conveyed": "1/1", "retained": "0"}]}],
        "ogls": [{"ogl_no": 1, "tracts": [1]}],
        "working_interests": [{"owner": "WI", "wi": "1.0", "nri": "0.8125"}],
        "well": {"name": "w", "hbp_supported": True, "production_evidence": "OTC PUN (FIXTURE)"},
        "instruments": [{"book_page": "1/1"}],
        "source_documents": [{"book_page": "1/1", "status": "found_local"}],
    }


@pytest.fixture
def ten_tract_workbook(tmp_path):
    p = tmp_path / "cert_[FIXTURE].xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Tract 1"
    ws["A2"] = 2; ws["A3"] = 3
    ws["A4"] = "#DIV/0!"        # error literal to repair -> intended =A2+A3
    for i in range(2, 11):
        wb.create_sheet(f"Tract {i}")
    wb.create_sheet("OGL"); wb.create_sheet("WI 1"); wb.create_sheet("Well")
    wb.save(p)
    return p


def test_repair_recalc_certify(db, audit, ten_tract_workbook, tmp_path, monkeypatch):
    monkeypatch.setenv("DATABOSSX_AGENT_ROOT", str(tmp_path / "agent"))
    (tmp_path / "agent" / "db").mkdir(parents=True, exist_ok=True)
    s = get_settings(reload=True)
    s.outputs_dir = tmp_path / "out"
    orch = Orchestrator(s, db, audit)
    result = orch.run(
        ten_tract_workbook, prospect="FIXTURE-CERT",
        known_formulas={"Tract 1!A4": "=A2+A3"},
        domain_context=_clean_domain())

    # A real safe repair was recorded.
    repairs = db.query("SELECT action, target_range FROM automated_repairs WHERE run_id=?",
                       (result["run_id"],))
    assert any(r["target_range"] == "A4" for r in repairs)

    if detect_libreoffice(s.libreoffice_path):
        # Full happy path: recalc ran and the workbook certified.
        assert result["final_status"] == "CERTIFY", \
            f"expected CERTIFY, got {result['final_status']}"
        assert "certified_workbook" in result["exports"]
    else:
        # No LibreOffice -> recalc escalates cleanly; still terminates.
        assert result["final_status"] in ("ESCALATE", "MAX_ITERATIONS")
