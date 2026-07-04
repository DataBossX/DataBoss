"""End-to-end proof of the Roger Mills finalizer on synthetic fixtures.

Every value here is a SYNTHETIC TEST FIXTURE — never a verified source fact.
The test proves: template formatting is used, tract-sheet scope is respected
(no tract outside the tract sheet appears), OGL numbers come from the OGL sheet,
the title block is filled, missing data is flagged for review (not invented),
and nothing is overwritten.
"""

import openpyxl
import pytest

from db.db_client import DatabaseManager
from db.audit_logger import AuditLogger
from reports.rogermills_finalizer import RogerMillsFinalizer, FinalizeConfig

BANNER = "SYNTHETIC TEST FIXTURE - NOT A VERIFIED SOURCE"


def _make_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Title Report"
    # Title block (single-label rows so they aren't mistaken for the header).
    ws["A1"] = "Section"
    ws["A2"] = "County"
    ws["A3"] = "Effective Date"
    ws["A5"] = BANNER
    headers = ["Tract", "Legal Description", "Acreage", "OGL Number",
               "Lessor", "Lessee"]
    for c, h in enumerate(headers, 1):
        ws.cell(7, c, h)
    wb.save(path)


def _make_report(path, section, tracts, ogl):
    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "Tract Sheet"
    t.append(["Tract", "Legal Description", "Acreage", BANNER])
    for tid, legal, ac in tracts:
        t.append([tid, legal, ac, ""])
    o = wb.create_sheet("OGL")
    o.append(["Tract", "OGL Number", "Lessor", "Lessee"])
    for tid, num, lr, le in ogl:
        o.append([tid, num, lr, le])
    ti = wb.create_sheet("Title")
    ti["A1"], ti["B1"] = "Section", section
    ti["A2"], ti["B2"] = "County", "Roger Mills"
    ti["A3"], ti["B3"] = "Effective Date", "2026-06-27"
    wb.save(path)


@pytest.fixture()
def horizon(tmp_path):
    root = tmp_path / "Horizon"
    root.mkdir()
    (root / ".env").write_text("OKCOUNTY_USERNAME=\nOKCOUNTY_PASSWORD=\n",
                               encoding="utf-8")
    template = root / "Template(30).xlsx"
    _make_template(template)
    # Three folders; tract 3 deliberately has NO OGL number (must be flagged).
    for name, section in [("Roger Mills", "31-12N-24W"),
                          ("Roger Mills 2", "32-12N-24W"),
                          ("Roger Mills 3", "33-12N-24W")]:
        folder = root / name
        folder.mkdir()
        _make_report(
            folder / f"{name} report.xlsx", section,
            tracts=[(1, "Lot 1", 40.0), (2, "Lot 2", 40.0), (3, "Lot 3", 40.0)],
            ogl=[(1, "OGL-1001", "Owner A", "Op A"),
                 (2, "OGL-1002", "Owner B", "Op B")],  # tract 3 missing
        )
    out = root / "rogermillsfinalreports"
    return root, template, out


def _run(root, template, out, dry_run=False):
    db = DatabaseManager(out / "_audit" / "a.db")
    audit = AuditLogger(out / "_audit" / "a.log", db=db)
    cfg = FinalizeConfig(
        horizon_root=root, output_dir=out,
        folders=[root / "Roger Mills", root / "Roger Mills 2",
                 root / "Roger Mills 3"],
        template_path=template, dry_run=dry_run, top_k=2, loops=1,
    )
    summary = RogerMillsFinalizer(cfg, db=db, audit=audit).run()
    db.close()
    return summary


def test_dry_run_writes_nothing(horizon):
    root, template, out = horizon
    summary = _run(root, template, out, dry_run=True)
    assert summary["dry_run"] is True
    assert summary["finals"] == []
    finals = list(out.glob("*_FINAL*.xlsx"))
    assert finals == []


def test_writes_one_final_per_folder(horizon):
    root, template, out = horizon
    summary = _run(root, template, out)
    assert len(summary["finals"]) == 3
    finals = sorted(out.glob("*_FINAL*.xlsx"))
    assert len(finals) == 3


def test_template_format_and_tract_scope_and_ogl(horizon):
    root, template, out = horizon
    _run(root, template, out)
    final = sorted(out.glob("*Roger Mills_FINAL*.xlsx"))[0]
    wb = openpyxl.load_workbook(final)
    assert "Title Report" in wb.sheetnames          # template formatting used
    ws = wb["Title Report"]
    # Header row 7 preserved from the template.
    assert ws.cell(7, 1).value == "Tract"
    # Exactly 3 tract rows written (scope = tract sheet; no phantom tract 4).
    tract_vals = [ws.cell(7 + i, 1).value for i in range(1, 5)]
    assert tract_vals[0] == 1 and tract_vals[1] == 2 and tract_vals[2] == 3
    assert tract_vals[3] in (None, "")              # no 4th tract invented
    # OGL numbers pulled from the OGL sheet for tracts 1 & 2.
    ogl_col = 4
    assert ws.cell(8, ogl_col).value == "OGL-1001"
    assert ws.cell(9, ogl_col).value == "OGL-1002"
    # Tract 3 had no OGL number -> cell left blank (never fabricated).
    assert ws.cell(10, ogl_col).value in (None, "")
    # Title block filled from the title sheet.
    assert ws["B1"].value == "31-12N-24W"


def test_missing_ogl_is_flagged_for_review(horizon):
    root, template, out = horizon
    _run(root, template, out)
    reviews = list(out.glob("REVIEW_Roger Mills_*.md"))
    assert reviews, "a review file must be produced"
    text = reviews[0].read_text(encoding="utf-8")
    assert "missing_ogl" in text          # tract 3's missing OGL is escalated
    assert "tract 3" in text


def test_never_overwrites_finals(horizon):
    root, template, out = horizon
    _run(root, template, out)
    _run(root, template, out)             # run again
    finals = list(out.glob("*Roger Mills_FINAL*.xlsx"))
    assert len(finals) >= 2               # second run numbered, not overwritten


def test_originals_backed_up_and_unmodified(horizon):
    root, template, out = horizon
    src = root / "Roger Mills" / "Roger Mills report.xlsx"
    before = src.read_bytes()
    _run(root, template, out)
    assert src.read_bytes() == before     # original untouched
    backups = list((out / "_backups").rglob("Roger Mills report.xlsx"))
    assert backups                         # a timestamped backup exists
