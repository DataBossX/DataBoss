"""Coverage for instrument, OGL/WI, well/HBP, formula, and audit gates."""

from __future__ import annotations

from openpyxl import Workbook

from core.enums import FailureCategory, GateStatus
from tests._ctx import make_context
from validators.val_audit import AuditCompletenessGate
from validators.val_instrument import InstrumentSourceAuditGate
from validators.val_ogl_wi import (
    OGLRegisterAuditGate,
    WellHBPSupportGate,
    WIDepthConsistencyGate,
)


def _wb_instrument_missing(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Runsheet"
    ws.append(["Grantor", "Grantee", "Book", "Page"])
    ws.append(["Alice", "Bob", "", ""])  # material row, no source
    wb.save(path)
    return path


def _wb_ogl_orphan(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracts"
    ws.append(["Tract", "OGL", "Net Acres"])
    ws.append(["Tract 1", "OGL-999", 10.0])  # not in register
    reg = wb.create_sheet("OGL Register")
    reg.append(["OGL", "Lessor", "Book", "Page"])
    reg.append(["OGL-001", "Lessor 1", 1, 2])
    wb.save(path)
    return path


def _wb_wells_hbp(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Wells"
    ws.append(["Well Name", "API", "Status"])
    ws.append(["Producer 1", "3500199999", "Producing"])  # HBP claim
    wb.save(path)
    return path


def _wb_wi_ambiguous_depth(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Working Interest"
    ws.append(["OGL", "Owner", "Working Interest", "Depth"])
    ws.append(["OGL-001", "Owner 1", "0.5", "TBD"])
    wb.save(path)
    return path


def test_instrument_missing_source_escalates(settings, tmp_path):
    wb = _wb_instrument_missing(tmp_path / "instr.xlsx")
    results = InstrumentSourceAuditGate().validate(make_context(wb, settings))
    esc = [r for r in results if r.status == GateStatus.ESCALATE]
    assert esc
    assert any(r.failure_category == FailureCategory.MISSING_SOURCE for r in esc)


def test_ogl_orphan_escalates(settings, tmp_path):
    wb = _wb_ogl_orphan(tmp_path / "ogl.xlsx")
    results = OGLRegisterAuditGate().validate(make_context(wb, settings))
    esc = [r for r in results if r.status == GateStatus.ESCALATE]
    assert esc
    assert any(r.failure_category == FailureCategory.OGL_MISMATCH_WI_GAP for r in esc)


def test_hbp_unsupported_escalates(settings, tmp_path):
    wb = _wb_wells_hbp(tmp_path / "wells.xlsx")
    ctx = make_context(wb, settings, source_index={"found_keys": []})
    results = WellHBPSupportGate().validate(ctx)
    esc = [r for r in results if r.status == GateStatus.ESCALATE]
    assert esc
    assert any(r.failure_category == FailureCategory.HBP_UNSUPPORTED for r in esc)


def test_hbp_supported_when_source_present(settings, tmp_path):
    wb = _wb_wells_hbp(tmp_path / "wells2.xlsx")
    # well reference present in found sources -> supported
    ctx = make_context(wb, settings, source_index={"found_keys": ["producer 1"]})
    results = WellHBPSupportGate().validate(ctx)
    assert any(r.status == GateStatus.PASS for r in results)


def test_wi_ambiguous_depth_escalates(settings, tmp_path):
    wb = _wb_wi_ambiguous_depth(tmp_path / "wi.xlsx")
    results = WIDepthConsistencyGate().validate(make_context(wb, settings))
    esc = [r for r in results if r.status == GateStatus.ESCALATE]
    assert esc
    assert any(
        r.failure_category == FailureCategory.UNSAFE_LEGAL_INFERENCE for r in esc
    )


def test_audit_gate_fatal_without_db(settings, valid_workbook):
    ctx = make_context(valid_workbook, settings)  # no db/run_id in context
    results = AuditCompletenessGate().validate(ctx)
    assert results[0].status == GateStatus.FAIL
    assert results[0].severity.value == "FATAL"


def test_audit_gate_passes_with_trail(settings, db, audit, valid_workbook):
    audit.log_run("run_a", "folder", "STARTED")
    audit.log_workbook_version("run_a", 0, "v000", "p", "h", "source_copy")
    audit.log_event("X", "run_a")
    ctx = make_context(valid_workbook, settings)
    ctx["db"] = db
    ctx["run_id"] = "run_a"
    results = AuditCompletenessGate().validate(ctx)
    assert results[0].status == GateStatus.PASS
