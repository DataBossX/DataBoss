"""XML editor preserves drawings/media/VBA, drops calcChain, sets fullCalcOnLoad."""

from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from repair.xml_editor import RepairOp, XmlWorkbookEditor


def _make_calc_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"], ws["A2"], ws["A3"] = 1, 2, 3
    ws["B1"] = "=A1*2"
    ws["B2"] = "=A2*2"
    ws["B3"] = 999  # hardcoded literal in the formula column
    wb.save(path)
    return path


def _inject_extra_parts(path: Path) -> Path:
    """Add media, drawing, and calcChain parts to simulate a rich workbook."""
    tmp = path.with_suffix(".injected.xlsx")
    with zipfile.ZipFile(path) as zin:
        names = zin.namelist()
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in names:
                zout.writestr(n, zin.read(n))
            zout.writestr("xl/media/image1.png", b"\x89PNG\r\n\x1a\n-fake-image-bytes")
            zout.writestr("xl/drawings/drawing1.xml",
                          b"<xdr:wsDr xmlns:xdr='x'/>")
            zout.writestr("xl/calcChain.xml", b"<calcChain xmlns='x'/>")
    shutil.move(tmp, path)
    return path


def test_repair_preserves_media_and_drawings(tmp_path):
    src = _make_calc_workbook(tmp_path / "calc.xlsx")
    _inject_extra_parts(src)

    dest = tmp_path / "calc_v001.xlsx"
    editor = XmlWorkbookEditor(src)
    outcome = editor.apply_formula_repairs(
        [RepairOp(sheet_name="Calc", coordinate="B3", formula="A3*2")], dest
    )
    assert outcome.ok, outcome.detail
    assert outcome.output_path == dest

    with zipfile.ZipFile(dest) as zf:
        names = zf.namelist()
        # drawings + media preserved
        assert "xl/media/image1.png" in names
        assert "xl/drawings/drawing1.xml" in names
        # calcChain dropped after formula edit
        assert "xl/calcChain.xml" not in names
        # media bytes intact
        assert zf.read("xl/media/image1.png").endswith(b"-fake-image-bytes")


def test_repair_sets_formula_and_fullcalc(tmp_path):
    src = _make_calc_workbook(tmp_path / "calc2.xlsx")
    dest = tmp_path / "calc2_v001.xlsx"
    editor = XmlWorkbookEditor(src)
    outcome = editor.apply_formula_repairs(
        [RepairOp(sheet_name="Calc", coordinate="B3", formula="A3*2")], dest
    )
    assert outcome.ok, outcome.detail

    wb = load_workbook(dest, data_only=False)
    assert str(wb["Calc"]["B3"].value).startswith("=")
    wb.close()

    with zipfile.ZipFile(dest) as zf:
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
    assert "fullCalcOnLoad" in wb_xml


def test_hashes_before_and_after_differ(tmp_path):
    src = _make_calc_workbook(tmp_path / "calc3.xlsx")
    dest = tmp_path / "calc3_v001.xlsx"
    outcome = XmlWorkbookEditor(src).apply_formula_repairs(
        [RepairOp(sheet_name="Calc", coordinate="B3", formula="A3*2")], dest
    )
    assert outcome.ok
    assert outcome.before_sha256 != outcome.after_sha256
    assert len(outcome.before_sha256) == 64 and len(outcome.after_sha256) == 64


def test_safe_parser_does_not_resolve_external_entities(tmp_path):
    # An XXE payload must not be expanded into file contents.
    secret = tmp_path / "secret.txt"
    secret.write_text("TOP-SECRET-TITLE-DATA", encoding="utf-8")
    payload = (
        '<?xml version="1.0"?>'
        f'<!DOCTYPE r [<!ENTITY xxe SYSTEM "file://{secret}">]>'
        "<r>&xxe;</r>"
    )
    xml_file = tmp_path / "evil.xml"
    xml_file.write_text(payload, encoding="utf-8")

    from repair.xml_editor import _parse

    tree = _parse(xml_file)
    text = tree.getroot().text or ""
    assert "TOP-SECRET-TITLE-DATA" not in text


def test_malicious_rels_target_cannot_escape_temp_dir(tmp_path):
    # A crafted workbook whose sheet relationship Target points outside the
    # package must not cause a write outside the extraction root.
    src = _make_calc_workbook(tmp_path / "evil.xlsx")

    # A would-be victim file outside the temp extraction dir, with a cell the
    # repair op would otherwise match.
    victim = tmp_path / "victim.xml"
    victim.write_text(
        '<?xml version="1.0"?><root><c r="B3"><v>1</v></c></root>', encoding="utf-8"
    )
    victim_before = victim.read_text(encoding="utf-8")

    # Rewrite the workbook's rels so sheet1 Target traverses up to the victim.
    rebuilt = tmp_path / "evil2.xlsx"
    traversal = f"../../../../../../..{victim.resolve()}"
    with zipfile.ZipFile(src) as zin:
        with zipfile.ZipFile(rebuilt, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in zin.namelist():
                data = zin.read(n)
                if n == "xl/_rels/workbook.xml.rels":
                    text = data.decode("utf-8")
                    # point the first worksheet relationship at the traversal path
                    text = re.sub(
                        r'Target="worksheets/sheet1.xml"',
                        f'Target="{traversal}"',
                        text,
                    )
                    data = text.encode("utf-8")
                zout.writestr(n, data)

    dest = tmp_path / "evil_v001.xlsx"
    editor = XmlWorkbookEditor(rebuilt)
    editor.apply_formula_repairs(
        [RepairOp(sheet_name="Calc", coordinate="B3", formula="A3*2")], dest
    )
    # The victim file outside the temp dir must be untouched.
    assert victim.read_text(encoding="utf-8") == victim_before


def test_refuses_to_overwrite_existing_version(tmp_path):
    src = _make_calc_workbook(tmp_path / "calc4.xlsx")
    dest = tmp_path / "calc4_v001.xlsx"
    dest.write_bytes(b"already here")
    outcome = XmlWorkbookEditor(src).apply_formula_repairs(
        [RepairOp(sheet_name="Calc", coordinate="B3", formula="A3*2")], dest
    )
    # rolled back / not ok because dest already exists
    assert outcome.ok is False
