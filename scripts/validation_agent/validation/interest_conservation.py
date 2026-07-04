"""Gate G1 — Interest Conservation.

The tract grid is a double-entry system whose *own* per-column SUBTOTAL guard
cells (row 6: ``=IF(colN7=0,"","RECHECK")``) encode the correct conservation
rule — accounting for the patent root column (sovereign origin, legitimately
sums to 1) and ARTI/fractional conventions that a naive re-summation would
misjudge. This gate therefore treats the guard cells as authoritative:

* a guard reading ``RECHECK`` under a fresh recalc == that instrument column does
  not conserve;
* such columns are almost always ARTI / "NEED int." instruments whose exact
  fraction is not in the workbook — so they are surfaced with the column's header
  note attached, to be routed to source-verification / escalation, never
  auto-invented.
"""
from __future__ import annotations

from openpyxl.utils import get_column_letter

from ..ingestion.sheet_models import SheetValues
from ..ingestion.workbook_map import WorkbookModel
from ..models import Coord, GateId, Severity, ValidationResult
from .base import Validator

_GUARD_ROW = 6
_HEADER_ROW = 5
_FIRST_INSTR_COL = 7


class InterestConservationValidator(Validator):
    gate = GateId.G1_CONSERVATION

    def validate(
        self, model: WorkbookModel, values: SheetValues
    ) -> list[ValidationResult]:
        results: list[ValidationResult] = []
        flagged = 0
        for grid in model.tracts():
            ws = grid.ws
            for ci in range(_FIRST_INSTR_COL, ws.max_column + 1):
                letter = get_column_letter(ci)
                guard = values.get(ws.title, letter, _GUARD_ROW)
                if isinstance(guard, str) and "RECHECK" in guard.upper():
                    flagged += 1
                    note = ws.cell(_HEADER_ROW, ci).value
                    instr = ws.cell(2, ci).value
                    results.append(
                        ValidationResult(
                            gate=self.gate,
                            passed=False,
                            severity=Severity.WARN,  # not auto-fixable -> escalate
                            message=(
                                f"Tract {grid.tract_no} column {letter} does not "
                                f"conserve (guard RECHECK; instr={instr}; "
                                f"note={note})"
                            ),
                            locations=[Coord(ws.title, f"{letter}{_GUARD_ROW}")],
                            metric=None,
                            expected=0.0,
                        )
                    )
        if flagged == 0:
            results.append(
                ValidationResult(
                    gate=self.gate,
                    passed=True,
                    severity=Severity.INFO,
                    message="All instrument columns conserve (no RECHECK guards).",
                )
            )
        return results
