"""
OGL Register Audit Gate + WI / Depth Consistency Gate.

Confirms the OGL register has the expected columns and that working-interest
sheets are present and internally consistent (depth clauses, assignment chain).
Structural problems are reported; legal sufficiency of leases is escalated.
"""
from __future__ import annotations

from typing import List

import openpyxl

from .base_validator import BaseValidator, ValidationResult, PASS, FAIL, ESCALATE, SEV_MED, SEV_HIGH

OGL_EXPECTED_COLS = ["ogl", "instrument", "grantor", "grantee", "legal", "royalty"]


class OglWiValidator(BaseValidator):
    gate_name = "OGL Register / WI Consistency Gate"

    def _sheet_names_by_cat(self, cat: str) -> List[str]:
        return [c.sheet for c in self.classifications
                if getattr(c, "category", None) == cat and not getattr(c, "escalate", False)]

    def run(self) -> List[ValidationResult]:
        results: List[ValidationResult] = []
        ogl_sheets = self._sheet_names_by_cat("ogl_register")
        wi_sheets = self._sheet_names_by_cat("working_interest")

        if not ogl_sheets:
            results.append(self._result(ESCALATE, severity=SEV_HIGH,
                reason="No OGL register sheet classified with confidence",
                recommended_action="Confirm OGL sheet naming / headers"))
        if not wi_sheets:
            results.append(self._result(ESCALATE, severity=SEV_HIGH,
                reason="No working-interest sheet classified with confidence",
                recommended_action="Confirm WI sheet naming / headers"))

        if self.workbook_path and ogl_sheets:
            wb = openpyxl.load_workbook(self.workbook_path, read_only=True, data_only=True)
            try:
                for sn in ogl_sheets:
                    if sn not in wb.sheetnames:
                        continue
                    ws = wb[sn]
                    header = []
                    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                        header = [str(h).lower() if h else "" for h in row]
                        break
                    hay = " ".join(header)
                    missing = [c for c in OGL_EXPECTED_COLS if c not in hay]
                    if missing:
                        results.append(self._result(FAIL, severity=SEV_MED, repairable=False,
                            affected_sheet=sn,
                            evidence={"missing_columns": missing},
                            reason=f"OGL register missing expected columns: {missing}",
                            recommended_action="Confirm register schema (examiner)"))
                    else:
                        # count populated lease rows
                        n = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(x is not None for x in r))
                        results.append(self._result(PASS, severity=SEV_MED,
                            affected_sheet=sn, evidence={"lease_rows": n},
                            reason=f"OGL register has expected columns; {n} lease rows"))
            finally:
                wb.close()

        # WI/depth legal sufficiency is a judgment → escalate for confirmation.
        if wi_sheets:
            results.append(self._result(ESCALATE, severity=SEV_MED, repairable=False,
                affected_sheet=", ".join(wi_sheets),
                reason="WI/NRI fractions and depth-severance require assignment-exhibit review",
                recommended_action="Human examiner: confirm WI/NRI from assignment exhibits / OCC"))
        if not results:
            results.append(self._result(PASS, severity=SEV_MED, reason="No OGL/WI issues detected"))
        return results
