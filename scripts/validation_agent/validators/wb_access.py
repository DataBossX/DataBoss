"""
Read-only workbook access helpers shared by validators.

Locates tract blocks on a Title-style sheet, extracts gross acreage and owner
net-mineral-acre columns, and reports formula vs. literal cells. Everything is
read-only; Decimal is used for money/acreage math (never float equality).
"""
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

TRACT_RE = re.compile(r"^\s*tract\s*\d+", re.I)
CONTAINING_RE = re.compile(r"([\d,]+\.?\d*)\s*acres", re.I)


def to_decimal(v: Any) -> Optional[Decimal]:
    if v is None:
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", "").strip()
            if v == "":
                return None
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def find_title_sheet(wb) -> Optional[str]:
    for ws in wb.worksheets:
        if ws.title.strip().lower().startswith("title"):
            return ws.title
    return None


def read_tract_blocks(workbook_path: str, sheet_name: Optional[str] = None) -> List[Dict]:
    """Return one dict per tract: label, gross (Decimal), owner_nma_sum, has_balance, rows."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        sn = sheet_name or find_title_sheet(wb)
        if sn is None:
            return []
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=False))
        # find tract header rows in column B (index 1)
        headers = []
        for i, row in enumerate(rows):
            b = row[1].value if len(row) > 1 else None
            if isinstance(b, str) and TRACT_RE.match(b):
                headers.append((i, b.strip()))
        blocks = []
        for idx, (i, label) in enumerate(headers):
            end = headers[idx + 1][0] if idx + 1 < len(headers) else len(rows)
            gross = None
            # Prefer the explicit 'Containing:' label row (col B) and read its
            # acreage from col C — a tract's long legal description can itself
            # contain other '… N acres' phrases (e.g. 'the South 51 acres'),
            # which must NOT be mistaken for the tract's gross.
            for j in range(i, min(i + 4, end)):
                b = rows[j][1].value if len(rows[j]) > 1 else None
                if isinstance(b, str) and "containing" in b.strip().lower():
                    c = rows[j][2].value if len(rows[j]) > 2 else None
                    if isinstance(c, str):
                        m = CONTAINING_RE.search(c)
                        if m:
                            gross = to_decimal(m.group(1))
                    elif isinstance(c, (int, float)):
                        gross = to_decimal(c)
                    break
            # Fallback: scan for an 'acres' phrase only if no Containing row found.
            if gross is None:
                for j in range(i, min(i + 3, end)):
                    for cell in rows[j]:
                        if isinstance(cell.value, str):
                            m = CONTAINING_RE.search(cell.value)
                            if m:
                                gross = to_decimal(m.group(1))
                                break
                    if gross is not None:
                        break
            owner_sum = Decimal("0")
            has_balance = False
            owner_rows = 0
            total_is_formula = None
            total_cell = None
            owner_first = None
            owner_last = None
            for j in range(i + 1, end):
                excel_row = j + 1  # openpyxl rows are 1-based; enumerate index j is 0-based
                b = rows[j][1].value if len(rows[j]) > 1 else None
                c = rows[j][2].value if len(rows[j]) > 2 else None
                if isinstance(b, str) and b.strip().upper() in ("REPORT TOTAL", "TOTAL"):
                    if total_cell is None:
                        total_cell = f"C{excel_row}"
                        if isinstance(c, str) and c.startswith("="):
                            total_is_formula = True
                        elif isinstance(c, (int, float)):
                            total_is_formula = False
                    continue
                if isinstance(b, str) and b.strip().lower().startswith("mineral owner"):
                    continue
                if isinstance(b, str) and b.strip().lower().startswith("balance"):
                    has_balance = True
                d = to_decimal(c)
                if d is not None:
                    owner_sum += d
                    owner_rows += 1
                    if owner_first is None:
                        owner_first = excel_row
                    owner_last = excel_row
            sum_range = None
            if owner_first and owner_last:
                sum_range = f"C{owner_first}:C{owner_last}"
            blocks.append({
                "label": label, "gross": gross, "owner_nma_sum": owner_sum,
                "has_balance": has_balance, "owner_rows": owner_rows,
                "total_is_formula": total_is_formula,
                "total_cell": total_cell, "sum_range": sum_range,
                "sheet": sn, "row_index": i,
            })
        return blocks
    finally:
        wb.close()


def scan_for_markers(workbook_path: str, markers: List[str], max_rows: int = 600) -> List[Dict]:
    """Find cells containing any marker string (case-insensitive)."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    hits = []
    low = [m.lower() for m in markers]
    try:
        for ws in wb.worksheets:
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, max_rows)), 1):
                for c in row:
                    if isinstance(c.value, str):
                        cl = c.value.lower()
                        for m in low:
                            if m in cl:
                                hits.append({"sheet": ws.title, "cell": c.coordinate,
                                             "marker": m, "text": c.value[:160]})
                                break
    finally:
        wb.close()
    return hits


def find_book_page_refs(workbook_path: str, max_rows: int = 800) -> List[Dict]:
    """Extract Book/Page (e.g. 1736/0592) and instrument (2019-001855) references."""
    bp = re.compile(r"\b(\d{3,4})\s*/\s*(\d{2,4})\b")
    instr = re.compile(r"\b(19|20)\d{2}-\d{5,6}\b")
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    refs: Dict[str, Dict] = {}
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, max_rows)):
                for c in row:
                    if isinstance(c.value, str):
                        for m in bp.finditer(c.value):
                            key = f"{m.group(1)}/{m.group(2)}"
                            refs.setdefault(key, {"reference": key, "kind": "book_page",
                                                  "sheet": ws.title, "cell": c.coordinate})
                        for m in instr.finditer(c.value):
                            key = m.group(0)
                            refs.setdefault(key, {"reference": key, "kind": "instrument",
                                                  "sheet": ws.title, "cell": c.coordinate})
    finally:
        wb.close()
    return list(refs.values())
