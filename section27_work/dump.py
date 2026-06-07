import openpyxl, sys
from openpyxl.utils import get_column_letter

path = "11N 25W 27 Diversified Cursory Report 6-6-2026.xlsx"
sheets = sys.argv[1:]
wb = openpyxl.load_workbook(path, data_only=False)
for name in sheets:
    ws = wb[name]
    print("="*90)
    print(f"SHEET: '{name}'  max_row={ws.max_row} max_col={ws.max_column}")
    print("="*90)
    for r in range(1, ws.max_row+1):
        cells=[]
        for c in range(1, ws.max_column+1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip()!="":
                cells.append(f"{get_column_letter(c)}{r}={repr(v)}")
        if cells:
            print(" | ".join(cells))
