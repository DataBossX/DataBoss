import openpyxl, sys
from openpyxl.utils import get_column_letter
path="11N 25W 27 Diversified Cursory Report 6-6-2026.xlsx"
name=sys.argv[1]; r1=int(sys.argv[2]); r2=int(sys.argv[3]); c2=int(sys.argv[4]) if len(sys.argv)>4 else None
wb=openpyxl.load_workbook(path)
ws=wb[name]
mc = c2 or ws.max_column
for r in range(r1,r2+1):
    cells=[]
    for c in range(1,mc+1):
        v=ws.cell(row=r,column=c).value
        if v is not None and str(v).strip()!="":
            cells.append(f"{get_column_letter(c)}{r}={repr(v)}")
    if cells: print(" | ".join(cells))
