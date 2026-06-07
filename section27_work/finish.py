import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

SRC="11N 25W 27 Diversified Cursory Report 6-6-2026.xlsx"
OUT="11N 25W 27 Diversified Cursory Report FINAL.xlsx"

wb=openpyxl.load_workbook(SRC)  # keep formulas & formatting
HL=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
changes=[]  # (sheet, cell, old, new, source, confidence, reason)

def setval(sheet, coord, new, source, conf, reason, highlight=True):
    ws=wb[sheet]; c=ws[coord]; old=c.value
    c.value=new
    if highlight: c.fill=HL
    changes.append((sheet, coord, "" if old is None else str(old), str(new), source, conf, reason))

# ---- A. Runsheet rows 1-3: correct truncated Book numbers; add instrument numbers ----
# Supported by 'diversified' tab (2400/551=2023-000796; 2320/438=2019-004809; 2480/824=2026-001031)
setval("Runsheet","F1",2400,"diversified tab (Bk 2400/Pg 551); Runsheet K1 note states B2400/P551","High","Corrected truncated book number 24 -> 2400 to match recorded citation")
setval("Runsheet","F2",2320,"diversified tab (Bk 2320/Pg 438); 2019-004809","High","Corrected truncated book number 232 -> 2320")
setval("Runsheet","F3",2480,"diversified tab (Bk 2480/Pg 824); 2026-001031","High","Corrected truncated book number 248 -> 2480")
setval("Runsheet","B1","2023-000796","diversified tab row (DP Sooner Holdco -> Diversified Production, 03/02/2023, 2400/551)","High","Added missing instrument number for conveyance row")
setval("Runsheet","B2","2019-004809","diversified tab row (Kimbell Royalty Group -> Diversified Rox Minerals, 11/19/2019, 2320/438)","High","Added missing instrument number for deed row")
setval("Runsheet","B3","2026-001031","diversified/diversifiedint tab (DP Legacy Central affidavit, 03/10/2026, 2480/824)","High","Added missing instrument number for affidavit row")

# ---- B. from kellpro rows 2-4: correct truncated Book numbers + hyperlinks; add instrument numbers ----
setval("from kellpro","F2",2400,"diversified tab (2023-000796, Bk 2400/Pg 551)","High","Corrected truncated book number 24 -> 2400 in source companion sheet")
setval("from kellpro","F3",2320,"diversified tab (2019-004809, Bk 2320/Pg 438)","High","Corrected truncated book number 232 -> 2320")
setval("from kellpro","F4",2480,"diversified tab (2026-001031, Bk 2480/Pg 824)","High","Corrected truncated book number 248 -> 2480")
setval("from kellpro","B2","2023-000796","diversified tab","High","Added missing instrument number")
setval("from kellpro","B3","2019-004809","diversified tab","High","Added missing instrument number")
setval("from kellpro","B4","2026-001031","diversified/diversifiedint tab","High","Added missing instrument number")
setval("from kellpro","L2",'=HYPERLINK("https://okcountyrecords.com/results/book%3D002400%3Apage%3D0551%3Asite%3Dbeckham/page-1","Search")',"Corrected book code 000024->002400","High","Fixed hyperlink book code to match corrected book 2400")
setval("from kellpro","L3",'=HYPERLINK("https://okcountyrecords.com/results/book%3D002320%3Apage%3D0438%3Asite%3Dbeckham/page-1","Search")',"Corrected book code 000232->002320","High","Fixed hyperlink book code to match corrected book 2320")
setval("from kellpro","L4",'=HYPERLINK("https://okcountyrecords.com/results/book%3D002480%3Apage%3D0824%3Asite%3Dbeckham/page-1","Search")',"Corrected book code 000248->002480","High","Fixed hyperlink book code to match corrected book 2480")

# ---- C. Full Chain Notes rows 2-4: correct Book/Page + hyperlink; add instrument numbers ----
setval("Full Chain Notes","D2","2400/551","diversified tab (2023-000796)","High","Corrected truncated book number 24 -> 2400")
setval("Full Chain Notes","D3","2320/438","diversified tab (2019-004809)","High","Corrected truncated book number 232 -> 2320")
setval("Full Chain Notes","D4","2480/824","diversified tab (2026-001031)","High","Corrected truncated book number 248 -> 2480")
setval("Full Chain Notes","B2","2023-000796","diversified tab","High","Added missing instrument number")
setval("Full Chain Notes","B3","2019-004809","diversified tab","High","Added missing instrument number")
setval("Full Chain Notes","B4","2026-001031","diversified/diversifiedint tab","High","Added missing instrument number")
setval("Full Chain Notes","K2",'=HYPERLINK("https://okcountyrecords.com/results/book%3D002400%3Apage%3D0551%3Asite%3Dbeckham/page-1","Search")',"Corrected book code","High","Fixed hyperlink book code 000024->002400")
setval("Full Chain Notes","K3",'=HYPERLINK("https://okcountyrecords.com/results/book%3D002320%3Apage%3D0438%3Asite%3Dbeckham/page-1","Search")',"Corrected book code","High","Fixed hyperlink book code 000232->002320")
setval("Full Chain Notes","K4",'=HYPERLINK("https://okcountyrecords.com/results/book%3D002480%3Apage%3D0824%3Asite%3Dbeckham/page-1","Search")',"Corrected book code","High","Fixed hyperlink book code 000248->002480")

# ---- D. Title sheet: cursory annotations (NO invented ownership/royalty/net-acre/HBP) ----
setval("Title ","G8",
    "CURSORY / VERIFY: Per-owner mineral net acres, royalty, expiration and HBP for Section 27-11N-25W are NOT independently verified. See OGL tab (64 base leases), LH and diversifiedint/chain tabs for the leasehold chain, and 'needed' tab for required lease images. Requires landman/title-attorney verification.",
    "OGL, LH, diversifiedint, needed tabs","Cursory","Added cursory pointer/verification note; owner table intentionally left unpopulated to avoid unsupported ownership conclusions")
setval("Title ","B118","Diversified Production, LLC (apparent — VERIFY)",
    "LH tab; diversifiedint; chainfourpoint; chainlegacy","Cursory/apparent","Added apparent working-interest party per workbook chain analysis (report purpose: Diversified interest)")
setval("Title ","G118",
    "APPARENT ONLY: Diversified appears as successor via (A) FourPoint/Maverick and (B) DP Legacy/Tapstone tracks. Net acres, royalty, depths, exclusions (Anadarko Basin Holdings / Teocalli), and HBP are NOT verified from recorded images. Cursory only — requires landman/title-attorney verification.",
    "LH, diversifiedint, chainfourpoint, chainlegacy, inst, needed tabs","Cursory/apparent","Added working-interest caveat tying report to Diversified chain without asserting verified WI/NRI")
setval("Title ","B132",
    "NOTES (CURSORY): This is a leasehold cursory limited to the Cherokee formation, prepared to identify the apparent interest of Diversified Production, LLC in Section 27-11N-25W, Beckham County, OK. It is based on the provided county index/runsheet data and document index PDF; most recorded lease/assignment images were NOT reviewed. Mineral ownership, net acres, royalty, expirations, HBP/production, pooling, depth/Pugh, warranty, option and top-lease clauses are NOT verified. See OGL, LH, diversifiedint, inst and 'needed' tabs. NOT a title opinion; do not rely for drilling, acquisition close, division order, or pay status.",
    "Workbook OGL/LH/needed tabs; project scope notes","Cursory","Added explicit cursory limitation note to Title Notes section")

# ---- E. Overview: add scope/client remark (supported) ----
setval("Overview","B54",
    "Leasehold cursory covering the apparent interest of Diversified Production, LLC (Diversified). Cursory only; not a title opinion. See QA Summary tab.",
    "Project scope; LH tab","Cursory","Added client/scope remark under REMARKS")

wb.save(OUT)
print("Saved interim:", OUT)
print("Total changes so far:", len(changes))
import json
with open("/tmp/changes.json","w") as f: json.dump(changes,f)
