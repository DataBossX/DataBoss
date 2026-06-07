import openpyxl, json
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

OUT="11N 25W 27 Diversified Cursory Report FINAL.xlsx"
wb=openpyxl.load_workbook(OUT)
HL=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HDR=PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
bold=Font(bold=True)
wrap=Alignment(wrap_text=True, vertical="top")
changes=json.load(open("/tmp/changes.json"))
today=datetime(2026,6,7)

# ---- Append to Change Log ----
cl=wb["Change Log"]
# header is row1: Sheet|Cell/Row|Old Value|New Value|Source|Confidence|Reason
start=cl.max_row+1
# add a separator/section label row
cl.cell(row=start, column=1, value=f"--- FINAL QA PASS {today.strftime('%m/%d/%Y')} (29 cells) ---").font=bold
start+=1
for i,(sheet,coord,old,new,source,conf,reason) in enumerate(changes):
    r=start+i
    cl.cell(row=r,column=1,value=sheet)
    cl.cell(row=r,column=2,value=coord)
    cl.cell(row=r,column=3,value=(old[:300] if old else ""))
    cl.cell(row=r,column=4,value=(new[:300]))
    cl.cell(row=r,column=5,value=source)
    cl.cell(row=r,column=6,value=conf)
    cl.cell(row=r,column=7,value=reason)
    for cc in range(1,8):
        cl.cell(row=r,column=cc).fill=HL
print("Change Log now max_row:", cl.max_row)

# ---- QA Summary sheet ----
if "QA Summary" in wb.sheetnames: del wb["QA Summary"]
qa=wb.create_sheet("QA Summary")
qa.sheet_properties.tabColor="00B050"
rows=[
 ("SECTION 27 CURSORY REPORT — FINAL QA SUMMARY", ""),
 ("Prepared", today.strftime("%m/%d/%Y")),
 ("Subject", "Section 27, Township 11 North, Range 25 West, Beckham County, Oklahoma (640 ac, m/l)"),
 ("Prospect", "26-005"),
 ("Client", "Diversified / Diversified Production, LLC"),
 ("Report type", "Leasehold CURSORY title report — limited to the Cherokee formation. NOT a title opinion."),
 ("", ""),
 ("REPORT NATURE / RELIANCE", ""),
 ("", "Cursory only. Based on county index/runsheet data and the provided document index PDF. Most recorded lease/assignment IMAGES were NOT reviewed. Ownership, net acres, royalty, expirations, HBP/production, pooling, depth/Pugh, warranty, option and top-lease clauses are NOT verified. Do not rely for drilling, acquisition close, division order, or pay status."),
 ("", ""),
 ("SHEETS EDITED IN FINAL PASS", ""),
 ("Runsheet", "Corrected 3 truncated Book numbers (24->2400, 232->2320, 248->2480) and added 3 missing instrument numbers for the most-recent Diversified-related rows. Verified against 'diversified' tab."),
 ("from kellpro", "Corrected the same 3 truncated Book numbers + their hyperlink book codes; added 3 instrument numbers (source companion sheet)."),
 ("Full Chain Notes", "Corrected the same 3 Book/Page values + hyperlink book codes; added 3 instrument numbers."),
 ("Title", "Added cursory/verify annotations: Tract 1 comment, an APPARENT-only working-interest entry for Diversified Production, LLC with caveat, and a cursory limitation note. Owner tables intentionally left unpopulated (per-owner net acres/royalty/HBP unverified)."),
 ("Overview", "Added client/scope remark under REMARKS."),
 ("Change Log", "Appended all 29 changed cells (sheet, cell, old, new, source, confidence, reason)."),
 ("QA Summary / Remaining Missing Items", "Added these two summary tabs."),
 ("", ""),
 ("WHAT WAS COMPLETED / CONFIRMED", ""),
 ("OGL tab", "64 base oil & gas lease index entries for Section 27 captured with grantor/grantee, instrument, book/page, dates, source file, and cursory caveats (HBP/full clauses not verified). Leases 1-25 have first-page terms (term/royalty/depth); leases 26-64 flagged 'Needs full lease abstraction'."),
 ("Diversified chain", "Cross-checked across diversified, diversifiedint, DP Legacy, chainfourpoint, chainlegacy, inst, LH. Two apparent tracks documented: (A) Nadel & Gussman / EnerVest -> FourPoint -> Maverick/Unbridled -> Diversified; (B) Chesapeake -> Tapstone / KL CHK SPV -> Diversified Production / DP Legacy / DP Sooner Holdco; ConocoPhillips/Burlington/Louisiana Land -> DP Legacy Central. Potential exclusions to Anadarko Basin Holdings and Teocalli flagged for image review."),
 ("Internal consistency", "No spreadsheet formula error tokens (#REF!, #VALUE!, #NAME?, #DIV/0!, #N/A) present. All expected key sheets retained."),
 ("", ""),
 ("SOURCE GAPS REMAINING (see 'needed' tab for full detail)", ""),
 ("P1 - Lease images", "All 64 base OGL recorded images need to be pulled and abstracted to confirm primary term, royalty, depth/Pugh, option/extension, warranty, top lease, expiration, and HBP effect."),
 ("P1 - Assignment schedules", "FourPoint/EnerVest/Chesapeake/Tapstone/KL CHK and DP Legacy/ConocoPhillips assignment images + lease schedules must be reviewed to confirm Section 27 / Cherokee inclusion, depths, net acres, burdens, and exclusions."),
 ("P1 - Releases/pooling/HBP", "Release (1995-002127), Affidavit of Pooling (2000-000146), and Partial Release (1719/0304) must be matched to OCC orders and production to confirm/deny HBP."),
 ("P1 - Divestiture review", "DP Legacy Central -> Anadarko Basin Holdings (2023-003027, 2023-003028) and DP Legacy/Diversified -> Teocalli (2024-002930) require image review to confirm whether Section 27 / target depths were conveyed out."),
 ("Well tab", "Well 1 tab is a header template only; OCC/RBDMS well, completion, production, spacing and HBP data still required."),
 ("", ""),
 ("ASSUMPTIONS MADE", ""),
 ("Book number corrections", "Book values 24/232/248 were treated as truncations of 2400/2320/2480 because the same instruments appear with full book numbers on the 'diversified' tab and in the rows' own notes. Page numbers were already correct."),
 ("Diversified as WI party", "Diversified Production, LLC listed on Title only as the APPARENT working-interest successor (report purpose), expressly caveated as unverified. No net acres, royalty, NRI, or HBP asserted."),
 ("Scope", "Stale 27-15N-24W references in project notes treated as template/format references only and ignored; final scope is 27-11N-25W."),
 ("", ""),
 ("ITEMS REQUIRING HUMAN / DOCUMENT VERIFICATION", ""),
 ("", "All lease terms, royalty, expirations, HBP/production status, net acres, mineral ownership, working-interest %, NRI, depth/Pugh clauses, warranty, options, top leases, the Diversified leasehold chain (both tracks), and all potential exclusions/divestitures. This cursory must not be relied upon as a title opinion."),
]
qa.column_dimensions['A'].width=34
qa.column_dimensions['B'].width=120
for i,(a,b) in enumerate(rows, start=1):
    ca=qa.cell(row=i,column=1,value=a); cb=qa.cell(row=i,column=2,value=b)
    cb.alignment=wrap
    if b=="" and a and (a.isupper() or a.startswith("SECTION 27")):
        ca.font=bold; ca.fill=HDR
    elif a in ("Prepared","Subject","Prospect","Client","Report type"):
        ca.font=bold

# ---- Remaining Missing Items sheet (concise rollup; 'needed' tab holds full detail) ----
if "Remaining Missing Items" in wb.sheetnames: del wb["Remaining Missing Items"]
rm=wb.create_sheet("Remaining Missing Items")
rm.sheet_properties.tabColor="FF0000"
hdr=["Priority","Category","Item / Target","Count / Refs","Why Needed","Next Action","Status"]
rmrows=[
 ("P1","Base OGL Lease Images","Recorded images for all 64 Section 27 base leases","OGL #1-64 (see 'needed' N-001..N-064)","Confirm term, royalty, depth/Pugh, option, warranty, top lease, expiration, HBP","Pull images; abstract into OGL/baseogls","Open"),
 ("P1","Leasehold Assignment Schedules","FourPoint/EnerVest/Chesapeake/Tapstone/KL CHK + DP Legacy/ConocoPhillips assignments","2017-004597; 2019-002937; 2019-002988; 2020-004435; 2020-004462; 2022-000152; 2022-003506; 2022-004277; 2022-004838; 2022-004844; 2023-000796","Confirm Section 27 / Cherokee inclusion, depths, net acres, burdens, exclusions","Pull images + schedules; map to OGL #","Open"),
 ("P1","Releases / Pooling / HBP","Release, Affidavit of Pooling, Partial Release","1995-002127 (1415/0054); 2000-000146 (1626/0107); 1719/0304","Determine whether base leases were released or are held by production","Match to OCC orders + production","Open"),
 ("P1","Divestiture / Exclusion Review","DP Legacy -> Anadarko Basin Holdings; DP Legacy/Diversified -> Teocalli","2023-003027; 2023-003028; 2024-002930","Confirm whether Section 27 / target depths were conveyed OUT of Diversified","Image/schedule review","Open"),
 ("P2","Well / Production Data","Well 1 tab (OCC/RBDMS)","Well 1 (template only)","Identify section wells, completions, formation, production, spacing, HBP","Pull OCC RBDMS + OTC/production","Open"),
 ("P2","Mineral Ownership / Net Acres","Title owner tables (Tracts 1-5)","Title tab","Per-owner net acres, royalty, expiration, HBP not determined","Full mineral run + lease abstraction","Open"),
 ("P3","Affidavit / Corporate Support","2026 affidavit + corporate succession records","2026-001031; 2025-001808; 2022-003507","Confirm entity succession / identity (not assumed as conveyance)","Image review","Open"),
]
rm.append(hdr)
for c in rm[1]:
    c.font=bold; c.fill=HDR; c.alignment=wrap
for row in rmrows: rm.append(row)
widths=[10,28,46,52,46,34,10]
for i,w in enumerate(widths, start=1):
    rm.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
for r in range(2, rm.max_row+1):
    for c in range(1, 8):
        rm.cell(row=r,column=c).alignment=wrap

wb.save(OUT)
print("Saved with QA Summary + Remaining Missing Items. Sheets:", len(wb.sheetnames))
print(wb.sheetnames)
