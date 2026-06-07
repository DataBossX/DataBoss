import openpyxl
from openpyxl.utils import get_column_letter
wb=openpyxl.load_workbook("Shanwee_Reference.xlsx")
print("SHANWEE SHEETS:", wb.sheetnames)
for nm in wb.sheetnames:
    if 'title' in nm.lower():
        ws=wb[nm]
        print(f"\n--- {nm} (max_row {ws.max_row}) ---")
        for r in range(1,min(ws.max_row,40)+1):
            cells=[]
            for c in range(1,ws.max_column+1):
                v=ws.cell(row=r,column=c).value
                if v is not None and str(v).strip()!="":
                    cells.append(f"{get_column_letter(c)}{r}={repr(v)[:60]}")
            if cells: print(" | ".join(cells))
