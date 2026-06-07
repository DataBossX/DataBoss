import openpyxl
ORIG="11N 25W 27 Diversified Cursory Report 6-6-2026.BACKUP.xlsx"
FIN="11N 25W 27 Diversified Cursory Report FINAL.xlsx"
wo=openpyxl.load_workbook(ORIG); wf=openpyxl.load_workbook(FIN)
print("=== 1. FINAL opens OK ===  sheets:", len(wf.sheetnames))
expected=['Overview','Title ','OGL','Runsheet','runsheet2','runsheet3','Tract 1','Tract 2','Tract 3','WI 1','Well 1','from kellpro','diversified','DP Legacy','Index','LH','diversifiedint','needed','RS','inst','todo','raw','Full Chain Notes','Change Log']
miss=[s for s in expected if s not in wf.sheetnames]
print("=== 2. Missing expected sheets:", miss or "NONE")
print("    New tabs added:", [s for s in wf.sheetnames if s not in wo.sheetnames])
# 3 error scan
errs=['#REF!','#VALUE!','#NAME?','#DIV/0!','#N/A','#NULL!','#NUM!']
cnt=0
for ws in wf.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and any(e in c.value for e in errs): cnt+=1
print("=== 3. Error-token cells:", cnt)
# 4 no sheet wiped: compare max_row*max_col for retained sheets
print("=== 4. Sheet integrity (orig -> final) ===")
wiped=[]
for s in wo.sheetnames:
    a=wo[s]; b=wf[s]
    oa=a.max_row*a.max_column; ob=b.max_row*b.max_column
    flag="" if ob>=oa*0.9 else "  <-- SHRANK!"
    if flag: wiped.append(s)
    print(f"   {s:20s} {a.max_row}x{a.max_column} -> {b.max_row}x{b.max_column}{flag}")
print("   WIPED:", wiped or "NONE")
# 5 verify edits applied
print("=== 5. Spot-check edits ===")
print("   Runsheet F1,B1:", wf['Runsheet']['F1'].value, wf['Runsheet']['B1'].value)
print("   Full Chain Notes D2,B2:", wf['Full Chain Notes']['D2'].value, wf['Full Chain Notes']['B2'].value)
print("   from kellpro F2:", wf['from kellpro']['F2'].value)
print("   Title B118:", wf['Title ']['B118'].value)
print("   Change Log max_row:", wf['Change Log'].max_row)
print("   QA Summary present:", 'QA Summary' in wf.sheetnames, "| Remaining Missing Items:", 'Remaining Missing Items' in wf.sheetnames)
