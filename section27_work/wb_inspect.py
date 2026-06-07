import openpyxl
from openpyxl.utils import get_column_letter

def overview(path, label):
    print("="*80)
    print(f"WORKBOOK: {label}  ({path})")
    print("="*80)
    wb = openpyxl.load_workbook(path, data_only=False)
    print("Sheets:", wb.sheetnames)
    for ws in wb.worksheets:
        print(f"\n--- SHEET: '{ws.title}'  dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column} state={ws.sheet_state}")
        if ws.merged_cells.ranges:
            print("   merged:", len(ws.merged_cells.ranges), "ranges (first 5):", list(ws.merged_cells.ranges)[:5])

overview("11N 25W 27 Diversified Cursory Report 6-6-2026.xlsx", "SECTION 27 MAIN")
