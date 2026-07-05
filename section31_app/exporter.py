"""Build and export the final Section 31 XLSX -- the product.

Writes every canonical sheet in schema order, a Cover sheet with run metadata
and roll-up totals, and applies light styling (bold frozen headers, autosize,
filters). The same function seeds an empty template when there is no data yet,
so the main folder always contains a valid, well-structured workbook.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from .config import COUNTY, SECTION_LABEL, STATE
from .schema import COVER_SHEET, SHEETS

_HEADER_FILL = "1F4E79"
_COVER_FILL = "203864"


def _style_header(ws, ncols: int):
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.freeze_panes = "A2"
    if ncols:
        from openpyxl.utils import get_column_letter

        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _autosize(ws, headers: List[str], rows: List[Dict]):
    from openpyxl.utils import get_column_letter

    for i, h in enumerate(headers, start=1):
        width = len(str(h))
        for r in rows[:200]:
            width = max(width, len(str(r.get(h, ""))))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 48)


def _write_sheet(wb, name: str, headers: List[str], rows: List[Dict]):
    ws = wb.create_sheet(title=name[:31])
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    _style_header(ws, len(headers))
    _autosize(ws, headers, rows)
    return ws


def _write_cover(wb, meta: Dict):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title=COVER_SHEET)
    ws["A1"] = f"{SECTION_LABEL}  |  {COUNTY} County, {STATE}"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=_COVER_FILL)
    ws.merge_cells("A1:B1")
    ws["A2"] = "Final Ownership & Title Report"
    ws["A2"].font = Font(bold=True, size=12, italic=True)
    ws.merge_cells("A2:B2")

    r = 4
    for label, value in meta.items():
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(value)).alignment = Alignment(
            wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70
    # Move cover to the front.
    wb.move_sheet(ws, -(len(wb.sheetnames) - 1))
    return ws


def build_cover_meta(data: Dict[str, List[Dict]], extras: Dict) -> Dict:
    meta = {
        "Report": SECTION_LABEL,
        "County / State": f"{COUNTY}, {STATE}",
        "Generated (UTC)": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ"),
        "Owners (title)": len(data.get("Tract & Title Ownership", [])),
        "WI owners (leasehold)": len(data.get("Leasehold WI Ownership", [])),
        "OGLs registered": len(data.get("OGL Register", [])),
        "Runsheet entries": len(data.get("Runsheet", [])),
        "Wells": len(data.get("Wells", [])),
    }
    meta.update(extras)
    return meta


def export_workbook(
    dest: Path,
    data: Dict[str, List[Dict]],
    cover_extras: Optional[Dict] = None,
) -> Path:
    """Write the final workbook to ``dest``. Returns the path written."""
    from openpyxl import Workbook

    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Drop the default sheet; we add our own in schema order.
    wb.remove(wb.active)

    for name, headers in SHEETS.items():
        _write_sheet(wb, name, headers, data.get(name, []))

    meta = build_cover_meta(data, cover_extras or {})
    _write_cover(wb, meta)

    wb.save(str(dest))
    return dest


def seed_template(dest: Path) -> Path:
    """Create an empty-but-valid final workbook (all sheets, no rows)."""
    empty = {name: [] for name in SHEETS}
    return export_workbook(dest, empty, {"Status": "TEMPLATE -- awaiting first run"})
