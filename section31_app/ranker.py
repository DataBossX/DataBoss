"""Rank candidate XLSX workbooks to choose the best base.

Scoring is transparent and additive so the operator can see *why* a workbook
won. Signals: filename relevance to the section, tab richness, presence of the
ownership/OGL/runsheet vocabulary, row volume, recency, and a small penalty for
obvious drafts/temp files.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from .scanner import Candidate

# Sheet-name vocabulary that signals a real title/ownership workbook.
_VALUE_TABS = [
    "owner", "ownership", "tract", "title", "ogl", "lease", "runsheet",
    "run sheet", "working interest", "wi", "royalty", "net acre", "well",
]
_DRAFT_HINT = ["draft", "temp", "tmp", "copy of", "backup", "old", "~$"]


@dataclass
class RankedWorkbook:
    candidate: Candidate
    score: float = 0.0
    sheet_names: List[str] = field(default_factory=list)
    total_rows: int = 0
    reasons: List[str] = field(default_factory=list)

    @property
    def name(self) -> str:
        return self.candidate.name

    @property
    def path(self) -> str:
        return self.candidate.path


def _inspect_workbook(path: str) -> Dict:
    """Open a local xlsx read-only and summarise tabs + row counts.

    Returns an empty summary for Drive-only or unreadable files (they can still
    be ranked on filename signals alone).
    """
    info = {"sheets": [], "rows": 0, "readable": False}
    p = Path(path)
    if not p.exists() or p.suffix.lower() not in (".xlsx", ".xlsm"):
        return info
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(p), read_only=True, data_only=True)
        info["sheets"] = list(wb.sheetnames)
        total = 0
        for ws in wb.worksheets:
            total += max((ws.max_row or 1) - 1, 0)
        info["rows"] = total
        info["readable"] = True
        wb.close()
    except Exception:
        pass
    return info


def rank_workbooks(candidates: List[Candidate]) -> List[RankedWorkbook]:
    """Score and sort candidate workbooks, best first."""
    ranked: List[RankedWorkbook] = []
    for cand in candidates:
        if cand.kind not in ("workbook", "runsheet"):
            continue
        rw = RankedWorkbook(candidate=cand)
        score = 0.0

        if cand.section_match:
            score += 30
            rw.reasons.append("+30 filename matches Section 31 / 12N-24W")
        lname = cand.name.lower()
        if "final" in lname:
            score += 8
            rw.reasons.append("+8 name says 'final'")
        if any(h in lname for h in _DRAFT_HINT):
            score -= 12
            rw.reasons.append("-12 looks like a draft/temp/backup")

        summary = _inspect_workbook(cand.path)
        rw.sheet_names = summary["sheets"]
        rw.total_rows = summary["rows"]
        if summary["readable"]:
            score += 5
            rw.reasons.append("+5 workbook opened cleanly")
        joined = " ".join(summary["sheets"]).lower()
        hits = sorted({v for v in _VALUE_TABS if v in joined})
        if hits:
            score += min(len(hits) * 6, 36)
            rw.reasons.append(f"+{min(len(hits)*6,36)} value tabs: {', '.join(hits)}")
        if summary["rows"]:
            row_pts = min(summary["rows"] / 25.0, 20)
            score += row_pts
            rw.reasons.append(f"+{row_pts:.0f} data volume ({summary['rows']} rows)")

        if cand.source == "drive":
            score -= 3
            rw.reasons.append("-3 Google Drive fallback (prefer local)")

        rw.score = round(score, 1)
        ranked.append(rw)

    ranked.sort(key=lambda r: r.score, reverse=True)
    return ranked


def select_base(ranked: List[RankedWorkbook]) -> Optional[RankedWorkbook]:
    """The highest-scoring workbook, or None when there are no candidates."""
    return ranked[0] if ranked else None
