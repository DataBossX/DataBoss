"""Read runsheet notes and legal columns -- the controlling fallback.

When workbook tabs disagree, the runsheet's own legal descriptions and notes
win. This module locates a runsheet tab inside any candidate workbook, maps its
loosely-named columns onto the canonical Runsheet schema, and returns clean
rows the pipeline can treat as authoritative.
"""

from __future__ import annotations

import re
from typing import Dict, List, Optional

from .schema import SHEETS

# Loose header -> canonical column. Matching is case/space/punct-insensitive.
_HEADER_MAP = {
    "entry": "entry_no", "entryno": "entry_no", "no": "entry_no", "seq": "entry_no",
    "instrumentdate": "instrument_date", "instdate": "instrument_date",
    "dateofinstrument": "instrument_date", "date": "instrument_date",
    "recorded": "recorded_date", "recordeddate": "recorded_date",
    "filed": "recorded_date", "recording": "recorded_date",
    "doctype": "doc_type", "documenttype": "doc_type", "type": "doc_type",
    "instrumenttype": "doc_type", "kind": "doc_type",
    "grantor": "grantor", "from": "grantor", "lessor": "grantor",
    "grantee": "grantee", "to": "grantee", "lessee": "grantee",
    "instrumentnumber": "instrument_number", "instrumentno": "instrument_number",
    "instno": "instrument_number", "document": "instrument_number",
    "book": "book", "vol": "book", "volume": "book",
    "page": "page", "pg": "page",
    "legal": "legal_description", "legaldescription": "legal_description",
    "description": "legal_description", "landdescription": "legal_description",
    "tract": "legal_description",
    "ogl": "ogl_number", "oglnumber": "ogl_number", "oglno": "ogl_number",
    "lease": "ogl_number", "leaseno": "ogl_number",
    "notes": "notes", "note": "notes", "remarks": "notes", "comment": "notes",
    "comments": "notes",
}

_RUNSHEET_TAB = re.compile(r"run[\s._-]*sheet|abstract|chain", re.I)


def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def find_runsheet_tab(sheet_names: List[str]) -> Optional[str]:
    for name in sheet_names:
        if _RUNSHEET_TAB.search(name):
            return name
    return None


def _map_headers(headers: List[str]) -> Dict[int, str]:
    """Column index -> canonical field, best-effort."""
    mapping: Dict[int, str] = {}
    for i, h in enumerate(headers):
        key = _norm(h)
        if key in _HEADER_MAP:
            mapping[i] = _HEADER_MAP[key]
            continue
        # partial contains match (e.g. "recorded date (county)")
        for token, field in _HEADER_MAP.items():
            if token in key:
                mapping[i] = field
                break
    return mapping


def read_runsheet(path: str, prefer_tab: Optional[str] = None) -> List[Dict]:
    """Extract runsheet rows from a local workbook.

    Returns a list of dicts keyed to the canonical Runsheet columns. Missing
    columns are left blank -- never guessed.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    from pathlib import Path

    p = Path(path)
    if not p.exists():
        return []
    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
    except Exception:
        return []

    tab = prefer_tab or find_runsheet_tab(wb.sheetnames)
    if tab is None:
        wb.close()
        return []

    ws = wb[tab]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []

    mapping = _map_headers(header)
    cols = SHEETS["Runsheet"]
    out: List[Dict] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        record = {c: "" for c in cols}
        for idx, value in enumerate(raw):
            field = mapping.get(idx)
            if field and value not in (None, ""):
                record[field] = str(value).strip()
        record["source"] = f"runsheet::{p.name}::{tab}"
        if any(record[c] for c in ("grantor", "grantee", "legal_description",
                                   "instrument_number")):
            out.append(record)
    wb.close()
    return out
