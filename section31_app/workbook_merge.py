"""Compare workbook tabs across candidates and merge the best data.

The base workbook (highest rank) supplies the spine. Every other candidate is
mined for tabs that map to a canonical sheet; rows are merged in, preferring the
base's values but back-filling blank cells from richer sources. Provenance is
preserved so QA can see where every row came from.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

from .schema import SHEETS

# Which canonical sheet a source tab most likely feeds, by name.
_TAB_TARGETS = [
    (re.compile(r"tract|title|owner|mineral", re.I), "Tract & Title Ownership"),
    (re.compile(r"work.*interest|\bwi\b|leasehold", re.I), "Leasehold WI Ownership"),
    (re.compile(r"ogl|lease.*register|lease.*schedule", re.I), "OGL Register"),
    (re.compile(r"well", re.I), "Wells"),
    (re.compile(r"run.*sheet|abstract", re.I), "Runsheet"),
]

# Per-target header synonyms -> canonical column.
_SYNONYMS: Dict[str, Dict[str, str]] = {
    "Tract & Title Ownership": {
        "tract": "tract", "legal": "legal_description",
        "legaldescription": "legal_description", "description": "legal_description",
        "ogl": "ogl_number", "oglnumber": "ogl_number",
        "owner": "owner_name", "ownername": "owner_name", "name": "owner_name",
        "address": "owner_address", "owneraddress": "owner_address",
        "type": "owner_type", "ownertype": "owner_type",
        "grossacres": "gross_acres", "gross": "gross_acres",
        "netacres": "net_mineral_acres", "netmineralacres": "net_mineral_acres",
        "nma": "net_mineral_acres", "net": "net_mineral_acres",
        "interest": "mineral_interest", "mineralinterest": "mineral_interest",
        "royalty": "royalty_rate", "royaltyrate": "royalty_rate",
        "leasestatus": "lease_status", "status": "lease_status",
        "lessee": "lessee", "source": "source", "remarks": "remarks",
    },
    "Leasehold WI Ownership": {
        "ogl": "ogl_number", "oglnumber": "ogl_number",
        "tract": "tract", "legal": "legal_description",
        "owner": "wi_owner", "wiowner": "wi_owner", "name": "wi_owner",
        "workinginterest": "working_interest", "wi": "working_interest",
        "nri": "net_revenue_interest", "netrevenue": "net_revenue_interest",
        "netrevenueinterest": "net_revenue_interest",
        "netacres": "net_acres", "assignment": "assignment_chain",
        "assignmentchain": "assignment_chain", "chain": "assignment_chain",
        "effective": "effective_date", "effectivedate": "effective_date",
        "status": "lease_status", "source": "source", "remarks": "remarks",
    },
    "OGL Register": {
        "ogl": "ogl_number", "oglnumber": "ogl_number",
        "lessor": "lessor", "lessee": "lessee",
        "date": "instrument_date", "instrumentdate": "instrument_date",
        "recorded": "recorded_date", "recordeddate": "recorded_date",
        "book": "book", "page": "page",
        "legal": "legal_description", "description": "legal_description",
        "grossacres": "gross_acres", "royalty": "royalty_rate",
        "term": "primary_term", "primaryterm": "primary_term",
        "status": "status", "source": "source", "remarks": "remarks",
    },
    "Wells": {
        "api": "api_number", "apinumber": "api_number",
        "well": "well_name", "wellname": "well_name", "name": "well_name",
        "operator": "operator", "status": "well_status",
        "wellstatus": "well_status", "spud": "spud_date",
        "spuddate": "spud_date", "completion": "completion_date",
        "completiondate": "completion_date", "formation": "formation",
        "section": "section", "township": "township", "range": "range",
        "county": "county", "source": "source", "remarks": "remarks",
    },
}


def _norm(text) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def _target_for_tab(tab: str) -> Optional[str]:
    for rx, target in _TAB_TARGETS:
        if rx.search(tab):
            return target
    return None


def _map_headers(headers: List[str], target: str) -> Dict[int, str]:
    syn = _SYNONYMS.get(target, {})
    canonical = set(SHEETS[target])
    mapping: Dict[int, str] = {}
    for i, h in enumerate(headers):
        key = _norm(h)
        if key in canonical:
            mapping[i] = key
        elif key in syn:
            mapping[i] = syn[key]
        else:
            for token, field in syn.items():
                if token and token in key:
                    mapping[i] = field
                    break
    return mapping


def extract_tabs(path: str, base: bool = False) -> Dict[str, List[Dict]]:
    """Map a workbook's tabs onto canonical sheets and pull rows.

    Returns ``{canonical_sheet: [row_dict, ...]}``. Only local .xlsx/.xlsm are
    readable; Drive-only paths yield nothing (they must be downloaded first).
    """
    out: Dict[str, List[Dict]] = {}
    p = Path(path)
    if not p.exists() or p.suffix.lower() not in (".xlsx", ".xlsm"):
        return out
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(p), read_only=True, data_only=True)
    except Exception:
        return out

    for tab in wb.sheetnames:
        target = _target_for_tab(tab)
        if not target or target not in _SYNONYMS:
            continue
        ws = wb[tab]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            continue
        mapping = _map_headers(header, target)
        if not mapping:
            continue
        cols = SHEETS[target]
        bucket = out.setdefault(target, [])
        for raw in rows_iter:
            if raw is None or all(c is None for c in raw):
                continue
            rec = {c: "" for c in cols}
            for idx, value in enumerate(raw):
                field = mapping.get(idx)
                if field and value not in (None, ""):
                    rec[field] = str(value).strip()
            if not rec.get("source"):
                rec["source"] = f"{'BASE::' if base else ''}{p.name}::{tab}"
            if any(v for k, v in rec.items() if k not in ("source",)):
                bucket.append(rec)
    wb.close()
    return out


def _legal_key(text) -> str:
    """Aliquot-aware key so 'SW/4 of SW/4' and 'SW/4 SW/4' collapse together."""
    from .legal_match import normalize_legal

    return re.sub(r"\s", "", normalize_legal(text))


def _row_key(sheet: str, row: Dict) -> str:
    """A dedup key per sheet -- the natural identifier for that record type."""
    if sheet == "Tract & Title Ownership":
        return _norm(row.get("owner_name", "")) + "|" + _legal_key(row.get("legal_description", ""))
    if sheet == "Leasehold WI Ownership":
        return _norm(row.get("wi_owner", "")) + "|" + _norm(row.get("ogl_number", ""))
    if sheet == "OGL Register":
        return _norm(row.get("ogl_number", "")) or _norm(row.get("lessor", "")) + _norm(row.get("lessee", ""))
    if sheet == "Wells":
        return _norm(row.get("api_number", "")) or _norm(row.get("well_name", ""))
    return _norm(str(sorted(row.items())))


def merge_sheet(base_rows: List[Dict], other_rows: List[Dict], sheet: str) -> List[Dict]:
    """Merge ``other_rows`` into ``base_rows``, base values winning on conflict.

    New keys are appended; existing keys back-fill only blank base cells.
    """
    index: Dict[str, Dict] = {}
    ordered: List[Dict] = []
    for row in base_rows:
        key = _row_key(sheet, row)
        if key in index:
            continue
        index[key] = row
        ordered.append(row)
    for row in other_rows:
        key = _row_key(sheet, row)
        if key in index:
            existing = index[key]
            for col, val in row.items():
                if val and not existing.get(col):
                    existing[col] = val
                    existing["remarks"] = (existing.get("remarks", "") +
                                           " | back-filled").strip(" |")
        else:
            index[key] = row
            ordered.append(row)
    return ordered


def merge_workbooks(ranked, base_path: Optional[str]) -> Dict[str, List[Dict]]:
    """Extract + merge every candidate workbook into canonical sheets."""
    merged: Dict[str, List[Dict]] = {}
    if base_path:
        for sheet, rows in extract_tabs(base_path, base=True).items():
            merged[sheet] = rows
    for rw in ranked:
        if rw.path == base_path:
            continue
        for sheet, rows in extract_tabs(rw.path).items():
            merged[sheet] = merge_sheet(merged.get(sheet, []), rows, sheet)
    return merged
