#!/usr/bin/env python3
"""
SECTION 31 QA AUDITOR
=====================
Section 31-12N-24W, Roger Mills County, Oklahoma - title report QA program.

What it does, in order:
  1. Indexes every file under the configured folders            -> file_index.xlsx
  2. Scores every Excel workbook to find the best report        -> workbook_candidates.xlsx
  3. Compares the top candidate workbooks against each other    -> workbook_comparison.xlsx
  4. Runs a full QA audit on the best workbook (read-only)      -> qa_defects.xlsx
  5. Writes mechanical, value-level fix suggestions             -> suggested_corrections.xlsx
  6. Optionally makes a byte-for-byte SAFE WORKING COPY and
     (only if explicitly asked) applies whitelisted fixes to
     THE COPY, never the original.
  7. Writes run_log.txt and SECTION31_QA_SUMMARY.txt

ABSOLUTE RULES (enforced in code):
  * Originals are NEVER opened for writing, never overwritten, never deleted.
  * No title facts are invented: the program only reports what is in the files.
  * HBP is never auto-confirmed; HBP claims without cited evidence are flagged
    for a human examiner.
  * No WI/NRI values are created; corrections are limited to mechanical fixes
    (trim stray spaces, convert numbers stored as text) and are OFF by default,
    applied only to the safe copy, and each one is logged.
  * No tabs/notes are added to the client report; all QA output goes to
    separate files in a new output folder.
  * No silent failures: every exception is logged, and check failures become
    visible defect rows instead of being swallowed.
"""

from __future__ import annotations

import argparse
import hashlib
import logging
import os
import re
import shutil
import sys
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Dependency check with a friendly, non-coder-readable message.
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.stderr.write(
        "\nERROR: the 'openpyxl' package is not installed.\n"
        "Fix: open a command prompt and run:\n\n"
        "    py -m pip install -r requirements.txt\n\n"
        "(or double-click run_section31_qa.bat, which installs it for you)\n"
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DEFAULT_ROOTS = [
    r"D:\Desktop\Horizon",
    r"D:\Desktop\Horizon\Roger Mills 2",
    r"D:\Desktop\Horizon\Roger Mills 2_All Best 7-7-26",
]

SECTION = "31"
TOWNSHIP = "12N"
RANGE = "24W"
COUNTY = "Roger Mills"
SECTION_LABEL = f"{SECTION}-{TOWNSHIP}-{RANGE}"

EXCEL_EXTS = {".xlsx", ".xlsm"}
LEGACY_EXCEL_EXTS = {".xls", ".xlsb"}  # cannot be audited by openpyxl - flagged
SKIP_FILE_PREFIXES = ("~$",)          # Excel lock/temp files
OUTPUT_DIR_PREFIX = "Section31_QA_Output_"

# Filename/sheet keywords recorded in the file index.
INDEX_KEYWORDS = [
    "31-12n-24w", "12n-24w", "section 31", "sec 31", "sec. 31", "sec31",
    "roger mills", "runsheet", "run sheet", "ogl", "lease", "ownership",
    "title", "report", "tract", "hbp", "nhe", "best", "final", "plat",
    "assumption", "notes", "7-8-26", "7-8-2026", "07-08-26", "7-7-26",
]

# Filename scoring: (compiled regex, points, label)
NAME_SCORE_RULES = [
    (re.compile(r"31[\s\-_]*12\s*n[\s\-_]*24\s*w", re.I), 40, "section 31-12N-24W in name"),
    (re.compile(r"\bsec(?:tion|\.)?\s*31\b", re.I), 15, "'Section 31' in name"),
    (re.compile(r"roger\s*mills", re.I), 10, "'Roger Mills' in name"),
    (re.compile(r"7[\-_. ]0?8[\-_. ](?:20)?26", re.I), 30, "target date 7-8-26 in name"),
    (re.compile(r"7[\-_. ]0?7[\-_. ](?:20)?26", re.I), 20, "date 7-7-26 in name"),
    (re.compile(r"\bbest\b", re.I), 10, "'best' in name"),
    (re.compile(r"\bfinal\b", re.I), 8, "'final' in name"),
    (re.compile(r"\bnhe\b", re.I), 5, "'NHE' in name"),
    (re.compile(r"\b(copy|copy\s*\(\d+\))\b", re.I), -5, "'copy' in name"),
    (re.compile(r"\b(old|backup|bak|archive)\b", re.I), -8, "old/backup marker in name"),
    (re.compile(r"\b(draft|wip|working)\b", re.I), -3, "draft marker in name"),
]

# Sheet-name scoring: (compiled regex, points, label)
SHEET_SCORE_RULES = [
    (re.compile(r"owner", re.I), 15, "ownership sheet"),
    (re.compile(r"run\s*sheet|runsheet", re.I), 12, "runsheet sheet"),
    (re.compile(r"\bogl\b|lease", re.I), 10, "OGL/lease sheet"),
    (re.compile(r"tract", re.I), 8, "tract sheet"),
    (re.compile(r"assumption", re.I), 6, "assumptions sheet"),
    (re.compile(r"note", re.I), 4, "notes sheet"),
]

# Column-header vocabulary used to find data tables inside sheets.
OWNER_HEADER_RE = re.compile(r"owner|lessor|grantor|name", re.I)
WI_HEADER_RE = re.compile(r"\bw\.?\s*i\.?\b|working\s*int", re.I)
NRI_HEADER_RE = re.compile(r"\bn\.?\s*r\.?\s*i\.?\b|net\s*rev", re.I)
INTEREST_HEADER_RE = re.compile(r"interest|decimal|\bw\.?i\.?\b|\bn\.?r\.?i\.?\b|\bri\b|royalty", re.I)
ACRES_HEADER_RE = re.compile(r"acre", re.I)
TRACT_HEADER_RE = re.compile(r"tract", re.I)
BOOK_HEADER_RE = re.compile(r"\bbook\b|\bbk\b", re.I)
PAGE_HEADER_RE = re.compile(r"\bpage\b|\bpg\b", re.I)
BOOK_PAGE_HEADER_RE = re.compile(r"book\s*/?\s*page|bk\s*/?\s*pg|recording", re.I)
DATE_HEADER_RE = re.compile(r"date", re.I)
NOTES_HEADER_RE = re.compile(r"note|remark|comment", re.I)

FORMULA_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!"}

TOTAL_ROW_RE = re.compile(r"^\s*(grand\s+)?totals?\s*:?\s*$", re.I)
HBP_RE = re.compile(r"\bhbp\b|held\s+by\s+production", re.I)
HBP_CONFIRM_RE = re.compile(r"confirm|verified|\bheld\b|\byes\b", re.I)
HBP_EVIDENCE_RE = re.compile(
    r"\bwell\b|\bapi\s*#?\s*\d|prod(uction|ucing)|occ\b|completion|form\s*100[23]|allowable|per\s+\w+\s+records",
    re.I,
)
TRACT_LABEL_RE = re.compile(r"\btract\s*#?\s*(\d+[A-Za-z]?)\b", re.I)
STR_LABEL_RE = re.compile(r"\b(\d{1,2})\s*-\s*(\d{1,2})\s*N\s*-\s*(\d{1,2})\s*W\b", re.I)
BOOK_PAGE_VALUE_RE = re.compile(r"^\s*\d{1,5}\s*[/\-:]\s*\d{1,5}\s*$")

FOOTING_TOLERANCE = 0.000001   # 1e-6 of an 8/8ths unit
EIGHT_EIGHTHS_TOLERANCE = 0.0001
MIN_PLAUSIBLE_YEAR = 1850
MAX_HASH_BYTES = 200 * 1024 * 1024  # skip hashing files larger than 200 MB

LOG = logging.getLogger("section31_qa")


# ---------------------------------------------------------------------------
# Data containers
# ---------------------------------------------------------------------------
@dataclass
class FileRecord:
    path: str
    name: str
    ext: str
    size: int
    mtime: str
    sha1: str
    keywords: str
    error: str = ""


@dataclass
class Candidate:
    path: str
    name: str
    mtime: str
    score: float = 0.0
    reasons: list = field(default_factory=list)
    sheets: list = field(default_factory=list)
    hidden_sheets: list = field(default_factory=list)
    data_rows: int = 0
    owner_names: set = field(default_factory=set)
    interest_sums: dict = field(default_factory=dict)  # sheet -> {header: sum}
    error: str = ""


@dataclass
class Defect:
    severity: str      # HIGH / MEDIUM / LOW
    category: str
    sheet: str
    cell: str
    value: str
    message: str
    human_action: str


@dataclass
class Correction:
    sheet: str
    cell: str
    current: str
    suggested: str
    rationale: str
    fix_kind: str      # 'trim_text' | 'text_number' | 'manual'
    applied: str = "NO"


@dataclass
class TableInfo:
    """A detected data table inside a worksheet."""
    sheet: str
    header_row: int
    headers: dict          # column index (1-based) -> header text
    first_data_row: int
    last_data_row: int
    total_rows: list       # row indices whose owner/label cell says "Total"


# ---------------------------------------------------------------------------
# Small utilities
# ---------------------------------------------------------------------------
def norm_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_owner(value) -> str:
    """Normalize an owner name for duplicate detection."""
    text = norm_text(value).casefold()
    text = re.sub(r"[.,;:'\"()]", " ", text)
    text = re.sub(r"\b(llc|l\.l\.c|inc|co|corp|ltd|lp|l\.p|trust|trustee|et\s*ux|et\s*vir|et\s*al)\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def as_number(value):
    """Return value as float if it is numeric (including numbers stored as text)."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100.0
        except ValueError:
            return None
    # simple fraction like 1/8
    m = re.fullmatch(r"(-?\d+)\s*/\s*(\d+)", text)
    if m and int(m.group(2)) != 0:
        return int(m.group(1)) / int(m.group(2))
    try:
        return float(text)
    except ValueError:
        return None


def is_text_number(value) -> bool:
    return isinstance(value, str) and as_number(value) is not None


def sha1_of(path: Path) -> str:
    try:
        if path.stat().st_size > MAX_HASH_BYTES:
            return "(skipped: file > 200 MB)"
        digest = hashlib.sha1()
        with open(path, "rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
    except OSError as exc:
        return f"(hash failed: {exc})"


def fmt_ts(timestamp: float) -> str:
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")


def cell_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


# ---------------------------------------------------------------------------
# Output workbook writer (QA outputs only - never the client report)
# ---------------------------------------------------------------------------
def write_report_xlsx(path: Path, sheet_specs, dry_run: bool) -> None:
    """sheet_specs: list of (sheet_name, headers, rows-of-lists)."""
    if dry_run:
        total = sum(len(rows) for _, _, rows in sheet_specs)
        LOG.info("[DRY-RUN] would write %s (%d data rows)", path.name, total)
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers, rows in sheet_specs:
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = Font(bold=True)
        widths = [len(str(h)) for h in headers]
        for row in rows:
            safe_row = ["" if v is None else v for v in row]
            ws.append(safe_row)
            for i, v in enumerate(safe_row[: len(widths)]):
                widths[i] = max(widths[i], min(len(str(v)), 80))
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 82)
        ws.freeze_panes = "A2"
    wb.save(path)
    LOG.info("Wrote %s", path)


# ---------------------------------------------------------------------------
# Stage 1 - file indexing
# ---------------------------------------------------------------------------
def index_files(roots: list) -> list:
    records: list = []
    seen: set = set()
    for root in roots:
        root_path = Path(root)
        if not root_path.is_dir():
            LOG.warning("Folder not found (skipping): %s", root)
            records.append(FileRecord(str(root_path), root_path.name, "", 0, "", "",
                                      "", error="FOLDER NOT FOUND"))
            continue
        LOG.info("Scanning folder: %s", root_path)
        for dirpath, dirnames, filenames in os.walk(root_path, onerror=lambda e: LOG.warning("Walk error: %s", e)):
            # never descend into our own previous output folders
            dirnames[:] = [d for d in dirnames if not d.startswith(OUTPUT_DIR_PREFIX)]
            for filename in sorted(filenames):
                if filename.startswith(SKIP_FILE_PREFIXES):
                    continue
                full = Path(dirpath) / filename
                try:
                    resolved = str(full.resolve())
                except OSError:
                    resolved = str(full)
                if resolved in seen:      # roots overlap; index each file once
                    continue
                seen.add(resolved)
                try:
                    stat = full.stat()
                    lower = filename.lower()
                    hits = ",".join(k for k in INDEX_KEYWORDS if k in lower)
                    records.append(FileRecord(
                        path=str(full), name=filename, ext=full.suffix.lower(),
                        size=stat.st_size, mtime=fmt_ts(stat.st_mtime),
                        sha1=sha1_of(full), keywords=hits,
                    ))
                except OSError as exc:
                    LOG.error("Cannot stat %s: %s", full, exc)
                    records.append(FileRecord(str(full), filename, full.suffix.lower(),
                                              0, "", "", "", error=str(exc)))
    LOG.info("Indexed %d unique files", len(records))
    return records


# ---------------------------------------------------------------------------
# Table detection inside a worksheet
# ---------------------------------------------------------------------------
def find_tables(ws) -> list:
    """Find header rows (a row containing an owner-ish column plus an
    interest-ish or acres column) and their data extents."""
    tables = []
    max_scan = min(ws.max_row or 0, 30)
    for row_idx in range(1, max_scan + 1):
        headers = {}
        for col_idx in range(1, min(ws.max_column or 0, 40) + 1):
            text = norm_text(ws.cell(row=row_idx, column=col_idx).value)
            if text:
                headers[col_idx] = text
        if len(headers) < 2:
            continue
        header_texts = list(headers.values())
        has_owner = any(OWNER_HEADER_RE.search(h) for h in header_texts)
        has_data_col = any(INTEREST_HEADER_RE.search(h) or ACRES_HEADER_RE.search(h)
                           for h in header_texts)
        # runsheet-style record tables: parties + recording refs, no interest column
        is_record_table = has_owner and any(
            BOOK_HEADER_RE.search(h) or BOOK_PAGE_HEADER_RE.search(h)
            or re.search(r"instrument", h, re.I) for h in header_texts)
        if not (has_owner and (has_data_col or is_record_table)):
            continue
        owner_col = next(c for c, h in headers.items() if OWNER_HEADER_RE.search(h))
        last_data = row_idx
        total_rows = []
        blank_streak = 0
        for r in range(row_idx + 1, (ws.max_row or row_idx) + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in headers]
            if all(v is None or norm_text(v) == "" for v in row_vals):
                blank_streak += 1
                if blank_streak >= 3:
                    break
                continue
            blank_streak = 0
            label = norm_text(ws.cell(row=r, column=owner_col).value)
            if TOTAL_ROW_RE.match(label):
                total_rows.append(r)
            last_data = r
        tables.append(TableInfo(sheet=ws.title, header_row=row_idx, headers=headers,
                                first_data_row=row_idx + 1, last_data_row=last_data,
                                total_rows=total_rows))
        break  # one table per sheet is enough for scoring/QA purposes
    return tables


def data_rows_of(table: TableInfo):
    for r in range(table.first_data_row, table.last_data_row + 1):
        if r not in table.total_rows:
            yield r


# ---------------------------------------------------------------------------
# Stage 2 - workbook scoring
# ---------------------------------------------------------------------------
def score_workbook(record: FileRecord, newest_mtime: float, oldest_mtime: float) -> Candidate:
    cand = Candidate(path=record.path, name=record.name, mtime=record.mtime)
    for rule, points, label in NAME_SCORE_RULES:
        if rule.search(record.name):
            cand.score += points
            cand.reasons.append(f"{points:+d} {label}")
    # folder bonus: living in the "All Best" folder is a strong signal
    if "all best" in record.path.lower():
        cand.score += 12
        cand.reasons.append("+12 located in 'All Best' folder")

    try:
        wb = openpyxl.load_workbook(record.path, read_only=True, data_only=True)
    except Exception as exc:
        cand.error = f"could not open: {exc}"
        cand.score -= 100
        cand.reasons.append("-100 workbook could not be opened")
        LOG.error("Cannot open workbook %s: %s", record.path, exc)
        return cand

    try:
        cand.sheets = list(wb.sheetnames)
        for name in cand.sheets:
            for rule, points, label in SHEET_SCORE_RULES:
                if rule.search(name):
                    cand.score += points
                    cand.reasons.append(f"{points:+d} {label} ('{name}')")
                    break
        try:
            cand.hidden_sheets = [ws.title for ws in wb.worksheets
                                  if ws.sheet_state != "visible"]
        except Exception:
            pass

        # content probe: look at each sheet, find a data table, collect owners
        # and interest-column sums (used for scoring and later for comparison)
        for ws in wb.worksheets:
            try:
                tables = find_tables(ws)
            except Exception as exc:
                LOG.warning("Table scan failed on %s!%s: %s", record.name, ws.title, exc)
                continue
            for table in tables:
                n_rows = sum(1 for _ in data_rows_of(table))
                cand.data_rows += n_rows
                owner_col = next((c for c, h in table.headers.items()
                                  if OWNER_HEADER_RE.search(h)), None)
                if owner_col:
                    for r in data_rows_of(table):
                        owner = norm_owner(ws.cell(row=r, column=owner_col).value)
                        if owner:
                            cand.owner_names.add(owner)
                sums = {}
                for col, header in table.headers.items():
                    if INTEREST_HEADER_RE.search(header) or ACRES_HEADER_RE.search(header):
                        total = 0.0
                        found = False
                        for r in data_rows_of(table):
                            num = as_number(ws.cell(row=r, column=col).value)
                            if num is not None:
                                total += num
                                found = True
                        if found:
                            sums[header] = round(total, 8)
                if sums:
                    cand.interest_sums[ws.title] = sums
        if cand.data_rows:
            bonus = min(15, 3 + cand.data_rows // 10)
            cand.score += bonus
            cand.reasons.append(f"+{bonus} contains {cand.data_rows} table data rows")
        if cand.owner_names:
            cand.score += 5
            cand.reasons.append(f"+5 has {len(cand.owner_names)} distinct owners")
    finally:
        wb.close()

    # recency bonus 0..10 across the candidate pool
    try:
        mtime = datetime.strptime(record.mtime, "%Y-%m-%d %H:%M:%S").timestamp()
        if newest_mtime > oldest_mtime:
            recency = 10 * (mtime - oldest_mtime) / (newest_mtime - oldest_mtime)
            cand.score += round(recency, 1)
            cand.reasons.append(f"+{recency:.1f} recency")
    except ValueError:
        pass
    return cand


def score_all_workbooks(records: list) -> list:
    excel_records = [r for r in records if r.ext in EXCEL_EXTS and not r.error]
    if not excel_records:
        return []
    mtimes = []
    for r in excel_records:
        try:
            mtimes.append(datetime.strptime(r.mtime, "%Y-%m-%d %H:%M:%S").timestamp())
        except ValueError:
            mtimes.append(0.0)
    newest, oldest = max(mtimes), min(mtimes)
    candidates = [score_workbook(r, newest, oldest) for r in excel_records]
    candidates.sort(key=lambda c: c.score, reverse=True)
    return candidates


# ---------------------------------------------------------------------------
# Stage 3 - workbook comparison
# ---------------------------------------------------------------------------
def compare_workbooks(top: list) -> list:
    """Rows: metric, then one column per workbook."""
    rows = []
    names = [c.name for c in top]
    rows.append(["Score"] + [round(c.score, 1) for c in top])
    rows.append(["Last modified"] + [c.mtime for c in top])
    rows.append(["Sheets"] + ["; ".join(c.sheets) for c in top])
    rows.append(["Hidden sheets"] + ["; ".join(c.hidden_sheets) or "(none)" for c in top])
    rows.append(["Table data rows"] + [c.data_rows for c in top])
    rows.append(["Distinct owners"] + [len(c.owner_names) for c in top])
    if len(top) > 1:
        base = top[0].owner_names
        overlaps = ["(baseline)"]
        for c in top[1:]:
            if base or c.owner_names:
                union = base | c.owner_names
                pct = 100.0 * len(base & c.owner_names) / len(union) if union else 0.0
                only_base = len(base - c.owner_names)
                only_c = len(c.owner_names - base)
                overlaps.append(f"{pct:.0f}% overlap; {only_base} only in best; {only_c} only here")
            else:
                overlaps.append("no owner tables found")
        rows.append(["Owner overlap vs best"] + overlaps)
    all_sheet_keys = sorted({(s, h) for c in top for s, sums in c.interest_sums.items()
                             for h in sums})
    for sheet, header in all_sheet_keys:
        rows.append([f"Sum of '{header}' on '{sheet}'"]
                    + [c.interest_sums.get(sheet, {}).get(header, "") for c in top])
    return [["Metric"] + names] + rows


# ---------------------------------------------------------------------------
# Stage 4 - QA audit
# ---------------------------------------------------------------------------
class Auditor:
    """Runs every QA check against the best workbook. Read-only: the workbook
    is loaded into memory and never saved."""

    def __init__(self, path: str):
        self.path = Path(path)
        self.defects: list = []
        self.corrections: list = []
        # values view (formula results as cached by Excel)
        self.wb_vals = openpyxl.load_workbook(path, data_only=True)
        # formulas/formatting view
        self.wb_form = openpyxl.load_workbook(path, data_only=False)
        self.tables: dict = {}
        for ws in self.wb_vals.worksheets:
            found = find_tables(ws)
            if found:
                self.tables[ws.title] = found[0]

    def close(self):
        self.wb_vals.close()
        self.wb_form.close()

    def add(self, severity, category, sheet, cell, value, message, action):
        self.defects.append(Defect(severity, category, sheet, cell,
                                   norm_text(value)[:200], message, action))

    def run_all(self) -> list:
        checks = [
            self.check_section_identity,
            self.check_hidden_sheets,
            self.check_formula_errors,
            self.check_blank_required,
            self.check_bad_negatives,
            self.check_interest_bounds,
            self.check_footing,
            self.check_duplicate_owners,
            self.check_missing_notes_assumptions,
            self.check_hbp,
            self.check_ogl_wi_gaps,
            self.check_tract_mismatch,
            self.check_dates_book_page,
            self.check_format_drift,
            self.check_text_numbers_and_whitespace,
        ]
        for check in checks:
            try:
                check()
            except Exception as exc:
                LOG.error("QA check %s failed: %s\n%s", check.__name__, exc,
                          traceback.format_exc())
                self.add("HIGH", "CHECK_FAILURE", "(program)", "",
                         type(exc).__name__,
                         f"QA check '{check.__name__}' crashed: {exc}",
                         "Report this to the developer; the check did not run.")
        order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
        self.defects.sort(key=lambda d: (order.get(d.severity, 9), d.category, d.sheet))
        return self.defects

    # -- individual checks --------------------------------------------------
    def iter_string_cells(self, wb):
        for ws in wb.worksheets:
            max_row = min(ws.max_row or 0, 5000)
            max_col = min(ws.max_column or 0, 60)
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        yield ws, cell

    def check_section_identity(self):
        """The report should be about Section 31-12N-24W; a prominent other
        section means we may have scored the wrong workbook."""
        wanted = (SECTION, TOWNSHIP.rstrip("Nn"), RANGE.rstrip("Ww"))
        found_wanted = False
        other_sections = {}
        for ws, cell in self.iter_string_cells(self.wb_vals):
            for m in STR_LABEL_RE.finditer(cell.value):
                key = (m.group(1), m.group(2), m.group(3))
                if key == wanted:
                    found_wanted = True
                else:
                    other_sections.setdefault(
                        f"{m.group(1)}-{m.group(2)}N-{m.group(3)}W",
                        f"{ws.title}!{cell.coordinate}")
        if not found_wanted:
            self.add("HIGH", "TRACT_MISMATCH", "(workbook)", "",
                     "", f"No cell mentions {SECTION_LABEL} - is this the right workbook?",
                     "Confirm this file really is the Section 31 report.")
        for label, where in other_sections.items():
            self.add("MEDIUM", "TRACT_MISMATCH", where.split("!")[0], where.split("!")[1],
                     label, f"Workbook references a different section: {label}",
                     "Verify this reference is intentional (e.g. an offset well).")

    def check_hidden_sheets(self):
        for ws in self.wb_form.worksheets:
            if ws.sheet_state != "visible":
                self.add("MEDIUM", "HIDDEN_SHEET", ws.title, "",
                         ws.sheet_state,
                         f"Sheet '{ws.title}' is {ws.sheet_state}",
                         "Unhide and review; hidden data may contradict the report.")

    def check_formula_errors(self):
        for wb in (self.wb_vals, self.wb_form):
            for ws, cell in self.iter_string_cells(wb):
                if cell.value.strip().upper() in FORMULA_ERRORS:
                    if any(d.sheet == ws.title and d.cell == cell.coordinate
                           and d.category == "FORMULA_ERROR" for d in self.defects):
                        continue
                    self.add("HIGH", "FORMULA_ERROR", ws.title, cell.coordinate,
                             cell.value, f"Formula error {cell.value.strip()}",
                             "Repair the formula; downstream totals may be wrong.")

    def _table_ws(self, sheet_name):
        return self.wb_vals[sheet_name]

    def check_blank_required(self):
        for sheet_name, table in self.tables.items():
            ws = self._table_ws(sheet_name)
            required_cols = {c: h for c, h in table.headers.items()
                             if OWNER_HEADER_RE.search(h) or INTEREST_HEADER_RE.search(h)
                             or TRACT_HEADER_RE.search(h)}
            for r in data_rows_of(table):
                row_has_content = any(
                    norm_text(ws.cell(row=r, column=c).value) for c in table.headers)
                if not row_has_content:
                    continue
                for c, header in required_cols.items():
                    if norm_text(ws.cell(row=r, column=c).value) == "":
                        self.add("MEDIUM", "BLANK_REQUIRED", sheet_name, cell_ref(r, c),
                                 "", f"Required cell blank (column '{header}')",
                                 "Fill from source documents or mark 'None of record'.")

    def _interest_cols(self, table):
        return {c: h for c, h in table.headers.items()
                if INTEREST_HEADER_RE.search(h) or ACRES_HEADER_RE.search(h)}

    def check_bad_negatives(self):
        for sheet_name, table in self.tables.items():
            ws = self._table_ws(sheet_name)
            for c, header in self._interest_cols(table).items():
                for r in data_rows_of(table):
                    num = as_number(ws.cell(row=r, column=c).value)
                    if num is not None and num < 0:
                        self.add("HIGH", "BAD_NEGATIVE", sheet_name, cell_ref(r, c),
                                 num, f"Negative value in '{header}'",
                                 "Interests/acres cannot be negative; fix the source entry.")

    def check_interest_bounds(self):
        for sheet_name, table in self.tables.items():
            ws = self._table_ws(sheet_name)
            wi_col = next((c for c, h in table.headers.items() if WI_HEADER_RE.search(h)), None)
            nri_col = next((c for c, h in table.headers.items() if NRI_HEADER_RE.search(h)), None)
            for c, header in self._interest_cols(table).items():
                if ACRES_HEADER_RE.search(header):
                    continue
                for r in data_rows_of(table):
                    num = as_number(ws.cell(row=r, column=c).value)
                    if num is not None and num > 1.0 + EIGHT_EIGHTHS_TOLERANCE:
                        self.add("HIGH", "INTEREST_BOUNDS", sheet_name, cell_ref(r, c),
                                 num, f"'{header}' exceeds 1.0 (8/8ths)",
                                 "An interest greater than 100% is impossible; verify decimal placement.")
            if wi_col and nri_col:
                for r in data_rows_of(table):
                    wi = as_number(ws.cell(row=r, column=wi_col).value)
                    nri = as_number(ws.cell(row=r, column=nri_col).value)
                    if wi is not None and nri is not None and nri > wi + EIGHT_EIGHTHS_TOLERANCE:
                        self.add("HIGH", "INTEREST_BOUNDS", sheet_name, cell_ref(r, nri_col),
                                 f"NRI={nri} WI={wi}", "NRI is greater than WI on the same row",
                                 "NRI can never exceed WI; check burden math.")

    def check_footing(self):
        for sheet_name, table in self.tables.items():
            ws = self._table_ws(sheet_name)
            for c, header in self._interest_cols(table).items():
                body_sum = 0.0
                count = 0
                for r in data_rows_of(table):
                    num = as_number(ws.cell(row=r, column=c).value)
                    if num is not None:
                        body_sum += num
                        count += 1
                if count == 0:
                    continue
                # vs explicit total row
                for tr in table.total_rows:
                    stated = as_number(ws.cell(row=tr, column=c).value)
                    if stated is None:
                        continue
                    diff = body_sum - stated
                    if abs(diff) > FOOTING_TOLERANCE:
                        self.add("HIGH", "NON_FOOTING_TOTAL", sheet_name, cell_ref(tr, c),
                                 stated,
                                 f"Total row says {stated} but column '{header}' sums to "
                                 f"{round(body_sum, 8)} (off by {diff:+.8f})",
                                 "Re-foot the column; find the row(s) causing the variance.")
                        self.corrections.append(Correction(
                            sheet_name, cell_ref(tr, c), str(stated), str(round(body_sum, 8)),
                            f"Column '{header}' actually sums to {round(body_sum, 8)}",
                            "manual"))
                # WI/NRI columns should foot to 1.0 (8/8ths) if they appear to
                # represent a full unit (sum in the neighborhood of 1)
                if not ACRES_HEADER_RE.search(header) and 0.5 < body_sum < 1.5 \
                        and abs(body_sum - 1.0) > EIGHT_EIGHTHS_TOLERANCE:
                    self.add("MEDIUM", "NON_FOOTING_TOTAL", sheet_name,
                             f"col {get_column_letter(c)}", round(body_sum, 8),
                             f"Column '{header}' sums to {round(body_sum, 8)}, not 1.00000000 (8/8ths)",
                             "If this column represents a full unit it must foot to 1.0; "
                             "identify the missing/excess interest.")

    def check_duplicate_owners(self):
        for sheet_name, table in self.tables.items():
            ws = self._table_ws(sheet_name)
            owner_col = next((c for c, h in table.headers.items()
                              if OWNER_HEADER_RE.search(h)), None)
            tract_col = next((c for c, h in table.headers.items()
                              if TRACT_HEADER_RE.search(h)), None)
            if not owner_col:
                continue
            seen = {}
            for r in data_rows_of(table):
                owner = norm_owner(ws.cell(row=r, column=owner_col).value)
                if not owner:
                    continue
                tract = norm_text(ws.cell(row=r, column=tract_col).value) if tract_col else ""
                key = (owner, tract)
                if key in seen:
                    self.add("MEDIUM", "DUPLICATE_OWNER", sheet_name,
                             cell_ref(r, owner_col),
                             ws.cell(row=r, column=owner_col).value,
                             f"Owner appears more than once{' in tract ' + tract if tract else ''} "
                             f"(first at row {seen[key]})",
                             "Confirm whether rows should be combined or represent "
                             "genuinely separate interests (e.g. individual vs trustee).")
                else:
                    seen[key] = r

    def check_missing_notes_assumptions(self):
        joined_sheets = " | ".join(self.wb_vals.sheetnames).lower()
        has_notes = "note" in joined_sheets
        has_assumptions = "assumption" in joined_sheets
        if not (has_notes and has_assumptions):
            for ws, cell in self.iter_string_cells(self.wb_vals):
                low = cell.value.lower()
                if "assumption" in low:
                    has_assumptions = True
                if re.search(r"\bnotes?\b", low):
                    has_notes = True
                if has_notes and has_assumptions:
                    break
        if not has_notes:
            self.add("MEDIUM", "MISSING_NOTES", "(workbook)", "", "",
                     "No notes sheet or notes section found",
                     "Client reports should carry examiner notes; add them in the "
                     "source workbook (not via this tool).")
        if not has_assumptions:
            self.add("MEDIUM", "MISSING_ASSUMPTIONS", "(workbook)", "", "",
                     "No assumptions sheet/section found",
                     "State the assumptions (e.g. HBP basis, unleased treatment) in the "
                     "source workbook.")

    def check_hbp(self):
        for ws, cell in self.iter_string_cells(self.wb_vals):
            text = cell.value
            if not HBP_RE.search(text):
                continue
            claims_confirmed = HBP_CONFIRM_RE.search(text)
            has_evidence = HBP_EVIDENCE_RE.search(text)
            if claims_confirmed and not has_evidence:
                self.add("HIGH", "HBP_UNSUPPORTED", ws.title, cell.coordinate, text,
                         "HBP stated as confirmed/held with no cited evidence "
                         "(no well name, API, OCC/production reference)",
                         "A human must verify HBP against production records before "
                         "this can stand; this tool never confirms HBP.")
            elif not has_evidence:
                self.add("MEDIUM", "HBP_UNSUPPORTED", ws.title, cell.coordinate, text,
                         "HBP mentioned without supporting evidence in the same cell",
                         "Verify HBP status against production/OCC records.")

    def check_ogl_wi_gaps(self):
        for sheet_name, table in self.tables.items():
            if not re.search(r"ogl|lease", sheet_name, re.I):
                continue
            ws = self._table_ws(sheet_name)
            wi_col = next((c for c, h in table.headers.items() if WI_HEADER_RE.search(h)), None)
            exp_col = next((c for c, h in table.headers.items()
                            if re.search(r"expir|primary\s*term", h, re.I)), None)
            notes_col = next((c for c, h in table.headers.items()
                              if NOTES_HEADER_RE.search(h)), None)
            for r in data_rows_of(table):
                if wi_col and as_number(ws.cell(row=r, column=wi_col).value) is None:
                    self.add("MEDIUM", "OGL_WI_GAP", sheet_name, cell_ref(r, wi_col),
                             ws.cell(row=r, column=wi_col).value,
                             "Lease row has no numeric WI",
                             "Trace the lease into the WI schedule or explain the gap.")
                if exp_col:
                    exp_val = ws.cell(row=r, column=exp_col).value
                    exp_date = exp_val if isinstance(exp_val, datetime) else None
                    if exp_date and exp_date.date() < date.today():
                        row_text = " ".join(norm_text(ws.cell(row=r, column=c).value)
                                            for c in table.headers)
                        if not HBP_RE.search(row_text):
                            self.add("HIGH", "OGL_WI_GAP", sheet_name, cell_ref(r, exp_col),
                                     exp_date.date().isoformat(),
                                     "Lease primary term is expired and row carries no HBP notation",
                                     "Determine whether the lease is held by production "
                                     "or has terminated.")
                if notes_col and norm_text(ws.cell(row=r, column=notes_col).value) == "":
                    row_has_content = any(norm_text(ws.cell(row=r, column=c).value)
                                          for c in table.headers)
                    if row_has_content:
                        self.add("LOW", "MISSING_NOTES", sheet_name, cell_ref(r, notes_col),
                                 "", "Lease row has an empty notes/remarks cell",
                                 "Add a remark or leave intentionally blank after review.")

    def check_tract_mismatch(self):
        tracts_by_sheet = {}
        for ws in self.wb_vals.worksheets:
            labels = set()
            max_row = min(ws.max_row or 0, 5000)
            max_col = min(ws.max_column or 0, 60)
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    if isinstance(cell.value, str):
                        for m in TRACT_LABEL_RE.finditer(cell.value):
                            labels.add(m.group(1).upper())
            # values sitting in a 'Tract' column of a detected table count too
            table = self.tables.get(ws.title)
            if table:
                tract_col = next((c for c, h in table.headers.items()
                                  if TRACT_HEADER_RE.search(h)), None)
                if tract_col:
                    for r in data_rows_of(table):
                        val = norm_text(ws.cell(row=r, column=tract_col).value)
                        if re.fullmatch(r"\d+[A-Za-z]?", val):
                            labels.add(val.upper())
            if labels:
                tracts_by_sheet[ws.title] = labels
        if len(tracts_by_sheet) < 2:
            return
        universe = set().union(*tracts_by_sheet.values())
        for sheet, labels in tracts_by_sheet.items():
            missing = universe - labels
            if missing and len(missing) < len(universe):
                self.add("MEDIUM", "TRACT_MISMATCH", sheet, "",
                         ", ".join(sorted(missing)),
                         f"Tract(s) {', '.join(sorted(missing))} appear elsewhere in the "
                         f"workbook but not on sheet '{sheet}'",
                         "Confirm every tract is carried through ownership, OGL and runsheet.")

    def check_dates_book_page(self):
        today = date.today()
        for sheet_name, table in self.tables.items():
            ws = self._table_ws(sheet_name)
            for c, header in table.headers.items():
                is_date_col = bool(DATE_HEADER_RE.search(header))
                is_book_page = bool(BOOK_PAGE_HEADER_RE.search(header))
                is_book = bool(BOOK_HEADER_RE.search(header)) and not is_book_page
                is_page = bool(PAGE_HEADER_RE.search(header)) and not is_book_page
                if not (is_date_col or is_book_page or is_book or is_page):
                    continue
                for r in data_rows_of(table):
                    val = ws.cell(row=r, column=c).value
                    if val is None or norm_text(val) == "":
                        continue
                    if is_date_col:
                        if isinstance(val, datetime):
                            d = val.date()
                            if d > today:
                                self.add("HIGH", "DATE_ISSUE", sheet_name, cell_ref(r, c),
                                         d.isoformat(), f"'{header}' date is in the future",
                                         "A recorded instrument cannot bear a future date.")
                            elif d.year < MIN_PLAUSIBLE_YEAR:
                                self.add("MEDIUM", "DATE_ISSUE", sheet_name, cell_ref(r, c),
                                         d.isoformat(), f"'{header}' date is before {MIN_PLAUSIBLE_YEAR}",
                                         "Verify against the recorded instrument.")
                        elif isinstance(val, str) and not re.search(r"\d", val):
                            self.add("LOW", "DATE_ISSUE", sheet_name, cell_ref(r, c),
                                     val, f"'{header}' cell is text with no digits",
                                     "Enter the instrument date or 'undated'.")
                    elif is_book_page:
                        if isinstance(val, str) and not BOOK_PAGE_VALUE_RE.match(val):
                            self.add("MEDIUM", "BOOK_PAGE_ISSUE", sheet_name, cell_ref(r, c),
                                     val, f"'{header}' not in Book/Page form (e.g. 1234/567)",
                                     "Verify the recording reference.")
                    elif is_book or is_page:
                        num = as_number(val)
                        if num is None or num <= 0 or num != int(num):
                            self.add("MEDIUM", "BOOK_PAGE_ISSUE", sheet_name, cell_ref(r, c),
                                     val, f"'{header}' should be a positive whole number",
                                     "Verify the recording reference.")

    def check_format_drift(self):
        for sheet_name, table in self.tables.items():
            ws_fmt = self.wb_form[sheet_name]
            for c, header in self._interest_cols(table).items():
                formats = {}
                for r in data_rows_of(table):
                    cell = ws_fmt.cell(row=r, column=c)
                    if cell.value is not None:
                        formats.setdefault(cell.number_format, []).append(r)
                if len(formats) > 2:
                    detail = "; ".join(f"'{f}' x{len(rs)}" for f, rs in
                                       sorted(formats.items(), key=lambda kv: -len(kv[1])))
                    self.add("LOW", "FORMAT_DRIFT", sheet_name,
                             f"col {get_column_letter(c)}", detail,
                             f"Column '{header}' mixes {len(formats)} number formats",
                             "Normalize formats so interests display consistently "
                             "(fix in the source workbook, not here).")
            # merged cells overlapping the data region make sums unreliable
            for merged in ws_fmt.merged_cells.ranges:
                if merged.min_row >= table.first_data_row and merged.min_row <= table.last_data_row:
                    self.add("LOW", "FORMAT_DRIFT", sheet_name, str(merged),
                             "", "Merged cells inside the data table",
                             "Merged cells inside data ranges break sorting and footing.")

    def check_text_numbers_and_whitespace(self):
        for sheet_name, table in self.tables.items():
            ws = self._table_ws(sheet_name)
            owner_col = next((c for c, h in table.headers.items()
                              if OWNER_HEADER_RE.search(h)), None)
            for c, header in self._interest_cols(table).items():
                for r in data_rows_of(table):
                    val = ws.cell(row=r, column=c).value
                    if is_text_number(val):
                        self.add("MEDIUM", "TEXT_NUMBER", sheet_name, cell_ref(r, c),
                                 val, f"'{header}' value is a number stored as text",
                                 "Text numbers are excluded from Excel SUM totals.")
                        self.corrections.append(Correction(
                            sheet_name, cell_ref(r, c), repr(val),
                            str(as_number(val)),
                            "Number stored as text; Excel formulas ignore it",
                            "text_number"))
            if owner_col:
                for r in data_rows_of(table):
                    val = ws.cell(row=r, column=owner_col).value
                    if isinstance(val, str) and val != val.strip():
                        self.add("LOW", "FORMAT_DRIFT", sheet_name, cell_ref(r, owner_col),
                                 repr(val), "Owner name has leading/trailing spaces",
                                 "Stray spaces defeat duplicate detection and sorting.")
                        self.corrections.append(Correction(
                            sheet_name, cell_ref(r, owner_col), repr(val),
                            repr(val.strip()), "Trim stray whitespace", "trim_text"))


# ---------------------------------------------------------------------------
# Stage 6 - safe copy (+ optional whitelisted corrections applied to the copy)
# ---------------------------------------------------------------------------
def make_safe_copy(best_path: Path, out_dir: Path, dry_run: bool) -> Path | None:
    copy_name = f"{best_path.stem}__SAFE_WORKING_COPY{best_path.suffix}"
    copy_path = out_dir / copy_name
    if dry_run:
        LOG.info("[DRY-RUN] would create safe working copy: %s", copy_path)
        return None
    shutil.copy2(best_path, copy_path)  # byte-for-byte; all formatting intact
    LOG.info("Safe working copy created (original untouched): %s", copy_path)
    return copy_path


def apply_corrections_to_copy(copy_path: Path, corrections: list) -> int:
    """Apply ONLY whitelisted mechanical fixes (trim whitespace, text->number)
    to the SAFE COPY. The original is never touched. Every change is logged.

    Note: openpyxl re-saves the file, which preserves cell styles, number
    formats, merged cells, column widths and images, but strips charts and
    some exotic drawing objects. That risk applies to the COPY only; the
    summary states it so a human can diff against the untouched original."""
    keep_vba = copy_path.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(copy_path, keep_vba=keep_vba)
    applied = 0
    for corr in corrections:
        if corr.fix_kind not in ("trim_text", "text_number"):
            continue  # anything judgment-based stays human-only
        try:
            ws = wb[corr.sheet]
            cell = ws[corr.cell]
            before = cell.value
            if corr.fix_kind == "trim_text" and isinstance(before, str):
                cell.value = before.strip()
            elif corr.fix_kind == "text_number":
                num = as_number(before)
                if num is None:
                    continue
                cell.value = num
            else:
                continue
            corr.applied = "YES (to safe copy)"
            applied += 1
            LOG.info("Applied fix to COPY %s!%s: %r -> %r",
                     corr.sheet, corr.cell, before, ws[corr.cell].value)
        except Exception as exc:
            corr.applied = f"FAILED: {exc}"
            LOG.error("Fix failed for %s!%s: %s", corr.sheet, corr.cell, exc)
    if applied:
        wb.save(copy_path)
        LOG.info("Saved %d whitelisted fixes to the safe copy (original untouched)", applied)
    wb.close()
    return applied


# ---------------------------------------------------------------------------
# Stage 7 - summary
# ---------------------------------------------------------------------------
def build_summary(roots, records, candidates, best, defects, corrections,
                  outputs, safe_copy, applied_count, dry_run) -> str:
    lines = []
    push = lines.append
    push("=" * 72)
    push(f"SECTION 31 QA SUMMARY - {SECTION_LABEL}, {COUNTY} County, Oklahoma")
    push(f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
         + ("   *** DRY-RUN (no files written) ***" if dry_run else ""))
    push("=" * 72)
    push("")
    push("FOLDERS SCANNED:")
    for r in roots:
        push(f"  - {r}" + ("" if Path(r).is_dir() else "   [NOT FOUND]"))
    push(f"Files indexed: {len(records)}")
    legacy = [r for r in records if r.ext in LEGACY_EXCEL_EXTS]
    if legacy:
        push(f"NOTE: {len(legacy)} legacy Excel file(s) (.xls/.xlsb) could not be "
             "audited - convert to .xlsx to include them:")
        for r in legacy[:10]:
            push(f"  - {r.path}")
    push("")
    push(f"WORKBOOK CANDIDATES SCORED: {len(candidates)}")
    if best:
        push(f"BEST WORKBOOK: {best.path}")
        push(f"  score {best.score:.1f} | modified {best.mtime} | "
             f"{best.data_rows} data rows | {len(best.owner_names)} owners")
        for reason in best.reasons[:8]:
            push(f"    {reason}")
    else:
        push("BEST WORKBOOK: none found - no auditable .xlsx/.xlsm matched.")
    push("")
    sev = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
    by_cat = {}
    for d in defects:
        sev[d.severity] = sev.get(d.severity, 0) + 1
        by_cat[d.category] = by_cat.get(d.category, 0) + 1
    push(f"QA DEFECTS: {len(defects)} total "
         f"(HIGH {sev.get('HIGH', 0)} / MEDIUM {sev.get('MEDIUM', 0)} / LOW {sev.get('LOW', 0)})")
    for cat, n in sorted(by_cat.items(), key=lambda kv: -kv[1]):
        push(f"  {cat:<24} {n}")
    high = [d for d in defects if d.severity == "HIGH"]
    if high:
        push("")
        push("TOP HIGH-SEVERITY ITEMS (fix these first):")
        for d in high[:10]:
            push(f"  [{d.category}] {d.sheet}!{d.cell}: {d.message}")
    push("")
    push(f"SUGGESTED CORRECTIONS: {len(corrections)} "
         f"({applied_count} applied to the SAFE COPY, 0 applied to any original)")
    if safe_copy:
        push(f"SAFE WORKING COPY: {safe_copy}")
        if applied_count:
            push("  NOTE: openpyxl re-saved the copy; styles/widths/images are kept "
                 "but charts, if any, are not. The original is untouched - diff "
                 "against it before adopting the copy.")
    push("")
    push("OUTPUT FILES:")
    for out in outputs:
        push(f"  - {out}")
    push("")
    push("ABSOLUTE-RULES COMPLIANCE:")
    push("  * No original file was opened for writing, modified or deleted.")
    push("  * No title facts were invented; every defect cites its cell.")
    push("  * HBP was never auto-confirmed; unsupported claims were flagged.")
    push("  * No WI/NRI values were created or changed in any original.")
    push("  * No tabs or notes were added to the client report.")
    push("")
    push("HUMAN REVIEW REQUIRED:")
    push("  1. Verify every HIGH defect against recorded instruments.")
    push("  2. Confirm HBP items against production/OCC records - the tool cannot.")
    push("  3. Confirm the chosen 'best workbook' is the intended client report.")
    push("  4. Review suggested_corrections.xlsx before adopting the safe copy.")
    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
def setup_logging(out_dir: Path | None) -> None:
    LOG.setLevel(logging.DEBUG)
    LOG.handlers.clear()
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter("%(levelname)-7s %(message)s"))
    LOG.addHandler(console)
    if out_dir is not None:
        file_handler = logging.FileHandler(out_dir / "run_log.txt", encoding="utf-8")
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(
            logging.Formatter("%(asctime)s %(levelname)-7s %(message)s"))
        LOG.addHandler(file_handler)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        prog="section31_qa",
        description=f"QA auditor for the {SECTION_LABEL} {COUNTY} County title report. "
                    "Read-only by design: originals are never modified.")
    parser.add_argument("--roots", nargs="*", default=DEFAULT_ROOTS,
                        help="Folders to scan (default: the three Horizon folders).")
    parser.add_argument("--out", default=None,
                        help="Output folder (default: a new timestamped folder "
                             "next to the first root).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Scan and audit but write NO files; print what would happen.")
    parser.add_argument("--make-safe-copy", action="store_true",
                        help="Also create a byte-for-byte working copy of the best workbook "
                             "in the output folder.")
    parser.add_argument("--apply-corrections", action="store_true",
                        help="Apply whitelisted mechanical fixes (trim spaces, "
                             "text-number conversion) to the SAFE COPY only. "
                             "Implies --make-safe-copy. Never touches originals.")
    parser.add_argument("--top", type=int, default=4,
                        help="How many top candidates to compare (default 4).")
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    existing_roots = [r for r in args.roots if Path(r).is_dir()]

    # output folder
    out_dir = None
    if not args.dry_run:
        base = Path(args.out) if args.out else (
            Path(existing_roots[0]) if existing_roots else Path.cwd())
        if args.out:
            out_dir = base
        else:
            out_dir = base / f"{OUTPUT_DIR_PREFIX}{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        out_dir.mkdir(parents=True, exist_ok=True)
    setup_logging(out_dir)

    LOG.info("SECTION 31 QA AUDITOR starting (%s mode)",
             "DRY-RUN" if args.dry_run else "normal")
    if not existing_roots:
        LOG.error("None of the configured folders exist: %s", args.roots)
        LOG.error("Pass the right folder(s) with:  --roots \"D:\\path\\to\\files\"")
        return 1

    outputs = []
    try:
        # Stage 1 - index
        records = index_files(args.roots)
        idx_path = (out_dir / "file_index.xlsx") if out_dir else Path("file_index.xlsx")
        write_report_xlsx(idx_path, [(
            "File Index",
            ["Path", "Name", "Ext", "Size (bytes)", "Modified", "SHA1", "Keywords", "Error"],
            [[r.path, r.name, r.ext, r.size, r.mtime, r.sha1, r.keywords, r.error]
             for r in records],
        )], args.dry_run)
        if not args.dry_run:
            outputs.append(idx_path)

        # legacy Excel warning is a visible defect, not a silent skip
        for r in records:
            if r.ext in LEGACY_EXCEL_EXTS:
                LOG.warning("Legacy Excel file cannot be audited (convert to .xlsx): %s", r.path)

        # Stage 2 - score
        candidates = score_all_workbooks(records)
        cand_path = (out_dir / "workbook_candidates.xlsx") if out_dir else Path("workbook_candidates.xlsx")
        write_report_xlsx(cand_path, [(
            "Candidates",
            ["Rank", "Score", "Path", "Modified", "Data rows", "Owners",
             "Sheets", "Hidden sheets", "Scoring reasons", "Error"],
            [[i + 1, round(c.score, 1), c.path, c.mtime, c.data_rows,
              len(c.owner_names), "; ".join(c.sheets), "; ".join(c.hidden_sheets),
              " | ".join(c.reasons), c.error]
             for i, c in enumerate(candidates)],
        )], args.dry_run)
        if not args.dry_run:
            outputs.append(cand_path)

        best = next((c for c in candidates if not c.error), None)
        if best:
            LOG.info("Best workbook: %s (score %.1f)", best.path, best.score)
        else:
            LOG.warning("No auditable workbook found under the scanned folders.")

        # Stage 3 - compare top N
        top = [c for c in candidates if not c.error][: max(args.top, 1)]
        if len(top) >= 1:
            comparison = compare_workbooks(top)
            cmp_path = (out_dir / "workbook_comparison.xlsx") if out_dir else Path("workbook_comparison.xlsx")
            write_report_xlsx(cmp_path, [(
                "Comparison", comparison[0], comparison[1:],
            )], args.dry_run)
            if not args.dry_run:
                outputs.append(cmp_path)

        # Stage 4 - QA audit of the best workbook
        defects: list = []
        corrections: list = []
        if best:
            LOG.info("Auditing best workbook (read-only): %s", best.name)
            auditor = Auditor(best.path)
            try:
                defects = auditor.run_all()
                corrections = auditor.corrections
            finally:
                auditor.close()
            LOG.info("QA audit complete: %d defects, %d suggested corrections",
                     len(defects), len(corrections))

        qa_path = (out_dir / "qa_defects.xlsx") if out_dir else Path("qa_defects.xlsx")
        write_report_xlsx(qa_path, [(
            "QA Defects",
            ["Severity", "Category", "Sheet", "Cell", "Value", "Finding",
             "Required human action"],
            [[d.severity, d.category, d.sheet, d.cell, d.value, d.message, d.human_action]
             for d in defects],
        )], args.dry_run)
        if not args.dry_run:
            outputs.append(qa_path)

        # Stage 5/6 - safe copy and (optional) whitelisted fixes to the copy
        safe_copy = None
        applied = 0
        if best and (args.make_safe_copy or args.apply_corrections):
            safe_copy = make_safe_copy(Path(best.path), out_dir or Path.cwd(), args.dry_run)
            if safe_copy and args.apply_corrections:
                applied = apply_corrections_to_copy(safe_copy, corrections)
            if safe_copy:
                outputs.append(safe_copy)

        corr_path = (out_dir / "suggested_corrections.xlsx") if out_dir else Path("suggested_corrections.xlsx")
        write_report_xlsx(corr_path, [(
            "Suggested Corrections",
            ["Sheet", "Cell", "Current value", "Suggested value", "Why", "Fix type",
             "Applied?"],
            [[c.sheet, c.cell, c.current, c.suggested, c.rationale, c.fix_kind, c.applied]
             for c in corrections],
        )], args.dry_run)
        if not args.dry_run:
            outputs.append(corr_path)

        # Stage 7 - summary
        summary = build_summary(args.roots, records, candidates, best, defects,
                                corrections, outputs
                                + ([out_dir / "run_log.txt", out_dir / "SECTION31_QA_SUMMARY.txt"]
                                   if out_dir else []),
                                safe_copy, applied, args.dry_run)
        print("\n" + summary)
        if out_dir:
            (out_dir / "SECTION31_QA_SUMMARY.txt").write_text(summary, encoding="utf-8")
            LOG.info("Wrote %s", out_dir / "SECTION31_QA_SUMMARY.txt")
        return 0
    except KeyboardInterrupt:
        LOG.error("Interrupted by user.")
        return 130
    except Exception as exc:
        LOG.error("FATAL: %s\n%s", exc, traceback.format_exc())
        LOG.error("The run did NOT complete. No original file was modified.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
