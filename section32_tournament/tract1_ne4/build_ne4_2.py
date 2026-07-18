#!/usr/bin/env python3
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
S="/tmp/claude-0/-home-user-DataBoss/14de7032-a11e-5ecb-9b29-220574151d44/scratchpad"
wb=load_workbook(f"{S}/NE4_build_stage1.xlsx")
thin=Side(style="thin",color="999999"); box=Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill=PatternFill("solid",fgColor="1F3864"); hdr_font=Font(color="FFFFFF",bold=True,size=10)
open_fill=PatternFill("solid",fgColor="FCE4D6")
wrap=Alignment(wrap_text=True,vertical="top"); ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
D=datetime.datetime
def clear_sheet(ws):
    for rng in list(ws.merged_cells.ranges): ws.unmerge_cells(str(rng))
    for row in ws.iter_rows():
        for c in row: c.value=None
def setrow(ws,r,vals,start=1):
    for i,v in enumerate(vals): ws.cell(row=r,column=start+i,value=v)

# ============================ OGL (NE/4 leases) ============================
ws=wb["OGL"]; clear_sheet(ws)
setrow(ws,1,["OGL #","LH","Image #","Instrument","Bk./Pg.","Execution Date","Filing Date","Grantor (Lessor)","Grantee (Lessee)","Legal Description","Misc. Notes","Term","Gross Acres","Tracts","Exp Date","Net Mineral Acres","Royalty"])
for c in range(1,18):
    cell=ws.cell(row=1,column=c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=ctr; cell.border=box
OGL=[
 (1,"264/271","INDEX","OGL","264/271",None,None,"Jean Jenkins Shearer","Union Oil Company of California","NE/4","Union Oil NE/4 leasing program. Original lease not imaged; term/royalty OPEN.","OPEN",160,1,None,"NOT DETERMINED","OPEN"),
 (2,"264/345","IMG 1575","OGL","264/345",D(1972,2,5),None,"Charles R. Cook (Crook) et ux","Union Oil Company of California","NE/4 Sec 32 (+ Sec 28 QQs)","THE CROOK LEASE. HBP by No. 1 Bruner Unit per 395/618. Released as to Leede 1979 (395/618). Base-lease candidate for NE/4 modern leasehold.","OPEN (HBP)",160,1,None,"NOT DETERMINED","OPEN"),
 (3,"270/72","INDEX","OGL","270/72",None,None,"L.S. & J.L. Randell","Union Oil Company of California","NE/4 (all); SE/SE","Union Oil lease; ROL to Randell at 270/553.","OPEN",160,1,None,"NOT DETERMINED","OPEN"),
 (4,"341/714","INDEX","OGL","341/714",None,None,"Jean Watkins Shearer","Leede Exploration","NE/4 (all)","Leede NE/4 leasing.","OPEN",160,1,None,"NOT DETERMINED","OPEN"),
 (5,"352/719","IMG (exh)","OGL","352/719",D(1978,5,16),None,"F.L. Randel","Union Oil Company of California","NE/4 (all)","Union Oil base-lease candidate; on 872/279 deep-rights exhibit.","OPEN",160,1,None,"NOT DETERMINED","OPEN"),
 (6,"352/721","IMG (exh)","OGL","352/721",D(1978,5,16),None,"J.L. Randel","Union Oil Company of California","NE/4 (all)","Union Oil base-lease candidate; on 872/279 exhibit.","OPEN",160,1,None,"NOT DETERMINED","OPEN"),
 (7,"376/369","INDEX/OKC","OGL","376/369",D(1979,3,21),None,"Great Western Oil & Gas, Inc.","Union Oil Company of California","NE/4 (all)","Corrected/re-recorded 1039/20-23 (1988). Union Oil NE/4.","OPEN",160,1,None,"NOT DETERMINED","OPEN"),
 ("OK48147.001.1","(modern lease ID)","2395/462; 2400/566","OGL (base UNPROVEN)","1697/236 cited (Assignment)",None,None,"UNPROVEN base lessor","Diversified Production LLC (record leasehold claimant)","NE/4 within Sec 32","MODERN LEASE ID carried on Diversified schedules. 1697/236 is an ASSIGNMENT (Staghorn->CHK), NOT the base OGL. True base OGL OPEN — likely one of Union Oil leases above HBP. Cherokee formation focus.","OPEN",None,1,None,"NOT DETERMINED","OPEN"),
]
r=2
for row in OGL:
    setrow(ws,r,list(row))
    for c in range(1,18):
        cell=ws.cell(row=r,column=c); cell.border=box; cell.alignment=wrap; cell.font=Font(size=9)
    ws.cell(row=r,column=16).fill=open_fill; ws.cell(row=r,column=17).fill=open_fill
    r+=1
# open items
ws.cell(row=r+1,column=1,value="OPEN"); ws.cell(row=r+1,column=2,value="True base OGL behind lease OK48147.001.1 is not page-proven. 1697/236 is an Assignment, not the base lease. Resolve which Union Oil lease (264/345 / 352/719 / 352/721 / 376/369) is the surviving HBP base for the NE/4.")
ws.cell(row=r+2,column=1,value="OPEN"); ws.cell(row=r+2,column=2,value="Carlson 32-11-25 7H wellbore carve (2183/174; 2185/639) — intersection with NE/4 leasehold and depths unmapped; do not compute full-leasehold WI until resolved.")
for rr in (r+1,r+2):
    ws.cell(row=rr,column=2).alignment=wrap; ws.cell(row=rr,column=2).fill=open_fill
for col,w in {'A':14,'B':16,'C':14,'D':18,'E':16,'F':13,'G':12,'H':26,'I':30,'J':20,'K':46,'L':10,'M':8,'N':7,'O':10,'P':16,'Q':10}.items():
    ws.column_dimensions[col].width=w

# ============================ TITLE (NE/4) ============================
ws=wb["Title "]; clear_sheet(ws)
ws["B2"]="TRACT 1 — NE/4, Section 32-11N-25W, Beckham County, Oklahoma"
ws["B2"].font=Font(bold=True,size=12)
ws["B4"]="TRACT 1:"; ws["C4"]="NE/4 (the Northeast Quarter)"
ws["B5"]="Containing:"; ws["C5"]="160.00 acres, more or less"
ws["B7"]="Mineral Owner"; ws["C7"]="Net Acres"; ws["D7"]="OGL No."; ws["E7"]="Royalty"; ws["F7"]="Status"; ws["G7"]="Comments"
for col in "BCDEFG":
    cell=ws[f"{col}7"]; cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=ctr; cell.border=box
mins=[
 ("FEE MINERAL ESTATE — OPEN","NOT DETERMINED","—","—","OPEN","The NE/4 mineral estate cannot be reduced to a proven current owner/decimal from the supplied Patent-1988 image run plus post-1988 index metadata. It fragmented before 1930 (see clusters below). Blank net acres != zero."),
 ("Garrett fractional-mineral cluster","NOT DETERMINED","—","—","OPEN","Cora B. & S.A. Garrett 1/80 mineral deeds (Bk 89/190-195) to Fowler, Andrews, Kuykendall, Gertrude Garrett, Pauline Roberts, Carl Garrett; later Garrett->Stewart/Fowler MDs (1063/164-167; 1283/4-5). Heirship OPEN."),
 ("Heller / La Fortune / Seidenbach / Sanditen cluster","NOT DETERMINED","—","—","OPEN","The Heller Company fractional MDs pre-1930 (Bk 88-98) seeded numerous small (1/8-1/80) NE/4 mineral interests to Heller heirs, La Fortune, Seidenbach, Sanditen, Jacobs, Lockwood, Murray, Holbrook, Smith, Harrison. Fragmented; OPEN."),
 ("Crook family (surface + partial minerals)","NOT DETERMINED","—","—","OPEN","Charlie B. Crook consolidated NE/4 surface + partial minerals (Bk 66,89,90,106). Probate 845/149-150 devised NE/4 fractions. Current surface lead: Crook, Ola Mae (Bk 2111/585)."),
 ("Randel / Viersen -> Asher Land & Exploration","NOT DETERMINED","—","—","OPEN","F.L./J.L. Randel & Ralph W. Viersen Jr. NE/4 interests conveyed to Asher Land & Exploration (839/329; 842/93-94); Walne/Asher exhibit chain (623/149; 506/205)."),
 ("Hanks, Darrell Lee (surface lead)","NOT DETERMINED","—","—","SURFACE LEAD","SURFACE/CURRENT-OWNER LEAD ONLY for N/2 NE/4 area. TODD 2290/494; 2291/474. Assessor Acct. 050005625. Mineral ownership OPEN."),
 ("Crook, Ola Mae (surface lead)","NOT DETERMINED","—","—","SURFACE LEAD","SURFACE/CURRENT-OWNER LEAD ONLY for SE/4 NE/4 area. Bk 2111/585. Assessor 050005631/050005630. Mineral ownership OPEN."),
]
r=8
for m in mins:
    setrow(ws,r,["",*m])
    for col in "BCDEFG":
        cell=ws[f"{col}{r}"]; cell.border=box; cell.alignment=wrap; cell.font=Font(size=9)
    ws[f"F{r}"].fill=open_fill
    r+=1
ws[f"B{r}"]="REPORT TOTAL"; ws[f"C{r}"]=f"=SUM(C8:C{r-1})"; ws[f"G{r}"]="OPEN — formula preserved; blank owner NMA is NOT zero proof. NE/4 fee minerals unresolved."
ws[f"B{r}"].font=Font(bold=True)
r+=2
# WI block
ws[f"B{r}"]="WORKING INTEREST (Leasehold) — NE/4"; ws[f"B{r}"].font=Font(bold=True,size=11); wr0=r
r+=1
ws[f"B{r}"]="Owner"; ws[f"C{r}"]="Net Acres"; ws[f"D{r}"]="Base OGL"; ws[f"E{r}"]="Royalty"; ws[f"F{r}"]="Status"; ws[f"G{r}"]="Comments"
for col in "BCDEFG":
    cell=ws[f"{col}{r}"]; cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=ctr; cell.border=box
r+=1
wis=[
 ("Diversified Production LLC","NOT DETERMINED","OK48147.001.1","OPEN","CLAIMANT","Record-derived leasehold claimant via DP Sooner HoldCo (2395/415-464 sch. p.2395/462; 2400/551-567 sch. p.2400/566). Addr of record 1600 Corporate Dr, Birmingham AL 35242. Exact WI/NRI/ORRI/net acres NOT DETERMINABLE — schedule interest columns blank. 1697/236 is an Assignment, not base OGL."),
 ("OCM Denali Holdings LLC","NOT DETERMINED","—","—","SEPARATE","Co-assignee with Diversified at 2371/470-494 and Tapstone->OCM Denali 2371/495-513. The 48.75% figure is NOT Diversified's WI and must not be blended."),
 ("Historic operators (chain only)","NOT DETERMINED","264/345 etc.","—","HISTORIC","Union Oil Co. of California -> Leede Exploration (872/279); Chesapeake -> Tapstone -> DP Legacy Tapstone -> Diversified. Latigo Petroleum operates Carlson wells (OCC). Not additive WI."),
]
for w in wis:
    setrow(ws,r,["",*w])
    for col in "BCDEFG":
        cell=ws[f"{col}{r}"]; cell.border=box; cell.alignment=wrap; cell.font=Font(size=9)
    ws[f"F{r}"].fill=open_fill
    r+=1
ws[f"B{r}"]="NOTES"; ws[f"C{r}"]=("HOLD_NO_RELEASE — disclosed cursory. NE/4 fee minerals OPEN; assessor names surface-only. "
 "Diversified WI supported on identity path only; quantum NOT DETERMINED. Base OGL unproven. Carlson 7H carve unresolved. "
 "Index Page 48 missing; Page 54 duplicate. This is a cursory report, not a title opinion.")
ws[f"C{r}"].alignment=wrap
for col,w in {'B':30,'C':16,'D':16,'E':12,'F':14,'G':70}.items(): ws.column_dimensions[col].width=w

wb.save(f"{S}/NE4_build_stage2.xlsx")
print("OGL + Title done")
