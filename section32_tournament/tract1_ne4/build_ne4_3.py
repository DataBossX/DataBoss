#!/usr/bin/env python3
"""Rebuild the Tract 1 sheet as NE/4 (160 ac), following the template's first-
tract conventions: instrument chain across the top header rows, owners down
column D, net-acre formula in C, running total in E, SUBTOTAL audit row,
Notes in column A. Decimal cells are populated only where a fraction is
expressly recited; otherwise left blank and disclosed OPEN (no forced totals).
"""
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
S="/tmp/claude-0/-home-user-DataBoss/14de7032-a11e-5ecb-9b29-220574151d44/scratchpad"
wb=load_workbook(f"{S}/NE4_build_stage2.xlsx")
ws=wb["Tract 1"]
D=datetime.datetime
open_fill=PatternFill("solid",fgColor="FCE4D6")
wrap=Alignment(wrap_text=True,vertical="top")

# --- Update tract identity (keep template skeleton/formulas intact) ---
ws["A1"]="TRACT 1 — NE/4, Sec 32-11N-25W, Beckham Co., OK. 160 ac. Fee mineral chain OPEN; leasehold claimant Diversified Production LLC (OK48147.001.1), quantum NOT DETERMINED."
ws["A1"].alignment=wrap
ws["A1"].font=Font(bold=True,size=9,color="9C0006")
ws["D2"]="NE/4"                     # tract typed exactly as instructed
ws["D5"]=160                        # TRACT ACRES
ws["C6"]=160                        # BASE OGL gross

# The template's instrument header columns (H.. ) carried Roger Mills books
# (PAT006/275, D00038/233 ...). Overwrite them with the documented Section 32
# NE/4 chain-of-title spine: row1=Instr type, row3=Bk/Pg, row4=Date, row5=Conveyance.
import openpyxl.utils as U
chain=[
 ("PAT","6/275",D(1910,5,23),"ARTI"),
 ("MD","89/190",None,"UND 1/80"),
 ("MD","89/193",None,"UND 1/80"),
 ("MD","89/195",None,"UND 1/80"),
 ("WD","90/545",None,"ARTI"),
 ("OGL","264/271",None,"LEASE"),
 ("OGL","264/345",D(1972,2,5),"LEASE"),
 ("OGL","352/719",D(1978,5,16),"LEASE"),
 ("OGL","352/721",D(1978,5,16),"LEASE"),
 ("OGL","376/369",D(1979,3,21),"LEASE"),
 ("REL","395/618",D(1979,11,21),"REL"),
 ("ASG","872/279",D(1985,10,29),"DEEP RTS"),
 ("PA","845/149",D(1985,10,24),"DECREE"),
 ("MD","839/329",D(1985,9,30),"NE MIN"),
 ("ASG","1697/574",D(2001,6,14),"LH"),
 ("ASG","2340/403",D(2020,12,14),"LH"),
 ("ASG","2371/470",D(2022,1,12),"LH"),
 ("MERGER","2389/500",D(2022,8,24),"LH"),
 ("CONV","2395/415",D(2022,11,18),"LH"),
 ("CONV","2400/551",D(2023,3,2),"LH"),
 ("ASG","2183/174",D(2015,2,18),"7H CARVE"),
]
# clear old header cells across H.. up to ~column 80
for col in range(8,90):
    for rr in (1,3,4,5):
        ws.cell(row=rr,column=col).value=None
c=8
for typ,bp,dt,conv in chain:
    ws.cell(row=1,column=c,value=typ)
    ws.cell(row=3,column=c,value=bp)
    if dt: ws.cell(row=4,column=c,value=dt)
    ws.cell(row=5,column=c,value=conv)
    for rr in (1,3,4,5):
        ws.cell(row=rr,column=c).alignment=Alignment(wrap_text=True,vertical="top",horizontal="center")
        ws.cell(row=rr,column=c).font=Font(size=8)
    c+=1

# --- Owners column D (rows 10+): documented NE/4 record parties (chain) ---
# Clear existing template owner rows 10..68 in D and interior decimals H..
for rr in range(10,69):
    ws.cell(row=rr,column=4).value=None
    for col in range(7,90):
        ws.cell(row=rr,column=col).value=None
owners=[
 "United States of America (Patent)",
 "The Heller Company (fractional NE/4 minerals)",
 "Cora B. Garrett & S.A. Garrett (grantors, 1/80 MDs)",
 "Garrett/Heller fractional mineral heirs (Fowler; Andrews; Kuykendall; Roberts; Carl & Gertrude Garrett; La Fortune; Seidenbach; Sanditen) — heirship OPEN",
 "Charlie B. Crook & Ima Pearl Crook (surface + partial minerals)",
 "Estate of Charlie B. Crook / devisees (PA 845/149-150) — fractions OPEN",
 "F.L. & J.L. Randel; Ralph W. Viersen Jr. (NE/4 minerals/lessors)",
 "Asher Land & Exploration Co. (839/329; 842/93-94)",
 "Union Oil Company of California (base leasehold, 264/345 / 352/719-721 / 376/369)",
 "Leede Exploration (deep-rights assignee, 872/279)",
 "Staghorn Resources -> Chesapeake Exploration (1697/574; 1697/236)",
 "Chesapeake -> Tapstone Energy LLC (2340/403)",
 "KL CHK SPV LLC -> Diversified Production LLC + OCM Denali Holdings LLC (2371/470)",
 "DP Legacy Tapstone LLC (2389/500)",
 "Diversified ABS VI Upstream LLC; DP Sooner HoldCo LLC (2395/415)",
 "Diversified Production LLC (2400/551) — RECORD LEASEHOLD CLAIMANT, OK48147.001.1",
 "OCM Denali Holdings LLC (separate co-assignee — NOT Diversified WI)",
 "Linn Operating / FourPoint (Carlson 7H wellbore carve 2183/174; 2185/639)",
 "Current surface leads: Hanks, Darrell Lee (N/2 NE/4); Crook, Ola Mae (SE/4 NE/4) — SURFACE ONLY",
]
r=10
for o in owners:
    ws.cell(row=r,column=4,value=o).alignment=wrap
    ws.cell(row=r,column=4).font=Font(size=9)
    r+=1
# Notes column A anchor
ws["A8"]="NOTES"
ws.cell(row=r+1,column=1,value="OPEN")
ws.cell(row=r+1,column=4,value=("FEE MINERAL DECIMALS NOT COMPUTED — the NE/4 mineral estate is fragmented and the "
 "controlling post-1930 fractional deeds/heirship are not provable from the supplied evidence. "
 "Instrument-chain spine above is factual (county index + direct images). Decimal grid intentionally "
 "left blank rather than forced to 160.00. Leasehold claimant = Diversified Production LLC (OK48147.001.1), "
 "WI/NRI NOT DETERMINED. Do not treat OCM Denali 48.75% as Diversified WI. Carlson 7H carve unresolved."))
ws.cell(row=r+1,column=4).alignment=wrap
ws.cell(row=r+1,column=4).font=Font(size=9,color="9C0006")
ws.cell(row=r+1,column=4).fill=open_fill

wb.save(f"{S}/NE4_build_stage3.xlsx")
print("Tract 1 NE/4 rebuilt; owner rows:", len(owners))
# quick verify sheet order + key cells
wb2=load_workbook(f"{S}/NE4_build_stage3.xlsx")
print("sheets:",wb2.sheetnames)
t=wb2["Tract 1"]
print("D2=",t["D2"].value,"D5=",t["D5"].value)
