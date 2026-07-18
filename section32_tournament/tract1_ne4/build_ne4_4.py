#!/usr/bin/env python3
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
S="/tmp/claude-0/-home-user-DataBoss/14de7032-a11e-5ecb-9b29-220574151d44/scratchpad"
wb=load_workbook(f"{S}/NE4_build_stage3.xlsx")
D=datetime.datetime
thin=Side(style="thin",color="999999"); box=Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill=PatternFill("solid",fgColor="1F3864"); hdr_font=Font(color="FFFFFF",bold=True,size=10)
wrap=Alignment(wrap_text=True,vertical="top"); ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
def unmerge(ws):
    for rng in list(ws.merged_cells.ranges): ws.unmerge_cells(str(rng))
def setrow(ws,r,vals,start=1):
    for i,v in enumerate(vals): ws.cell(row=r,column=start+i,value=v)

# ---- Well 1 : Section 32 NE/4-relevant wells (OCC facts from findings brief) ----
ws=wb["Well 1"]; unmerge(ws)
for row in ws.iter_rows():
    for c in row: c.value=None
ws["A1"]="Wells — Section 32-11N-25W (NE/4 relevant); OCC facts. AC = drilled/not plugged, not proof of production or HBP."
ws["A1"].font=Font(bold=True,size=9)
setrow(ws,2,["Well #","API #","Current Operator","Well Name","Formation","Status","Spacing","Order","Producing?","Notes"])
for c in range(1,11):
    cell=ws.cell(row=2,column=c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=ctr; cell.border=box
wells=[
 (1,"35-009-20312","Complete Oil & Gas Services LLC","Crook #1-32","Morrow (Mesa-Leede unit)","AC/DRY","640","156126 / CD 59995","No 12-mo rows","Mesa->Samson->BP->Latigo->Complete (Form 1073MW appr. 2025-12-03). OTC PUN 009-066620 (BP) Active, no rolling rows. AC != HBP. NE/4 unit well."),
 (2,"35-009-21072","BCE-Mach II LLC","JGL #1-32","Atoka (gas)","AC/gas","640","(formation order stack)","Yes (thru 4/2026)","OTC PUN 009-101800 reports gas Jul-2025..Apr-2026; 2,367 MCF Apr-2026 (NextEra mktg)."),
 (3,"35-009-21085","BCE-Mach II LLC","Hanks Crook #1-32","gas","AC/gas","640","—","No 12-mo rows","Active PUN, no rolling rows 2026-07-09."),
 (4,"35-009-21893","Latigo Petroleum, LLC","Carlson 32-11-25 12H","Granite Wash (Missourian)","AC/producer","640","613064 etc.","Yes (gas thru 5/2026)","Linn->EnerVest->BP->Latigo. 2,113 MCF May-2026. 7H/10H carve at 2183/174; 2185/639."),
 (5,"(permit 9H)","Riviera Operating LLC","Carlson 9H (permit only)","—","EX/NE","—","—","No","Permit-only, expired 2014-04-02; kept separate from drilled 12H."),
]
r=3
for w in wells:
    setrow(ws,r,list(w))
    for c in range(1,11):
        cell=ws.cell(row=r,column=c); cell.border=box; cell.alignment=wrap; cell.font=Font(size=9)
    r+=1
for col,wd in {'A':7,'B':16,'C':24,'D':20,'E':20,'F':11,'G':9,'H':16,'I':14,'J':50}.items(): ws.column_dimensions[col].width=wd

# ---- PLAT scope note ----
ws=wb["PLAT"]; 
try: unmerge(ws)
except: pass
ws["B22"]="SECTION 32-11N-25W — TRACT 1 = NE/4 (160 ac). Gross aliquot; NOT ownership."
ws["B23"]="Tract 1 (this report)"; ws["C23"]="NE/4"; ws["G23"]="160 ac"
ws["B30"]="NOTE: Surface assessor names are NOT proven mineral ownership. Plat not independently examined."

# ---- Neutralize out-of-scope tracts / WI carrying Roger Mills example data ----
for sn in ["Tract 2","Tract 3","Tract 4","Tract 5","WI 1","WI 2"]:
    ws=wb[sn]
    ws["A1"]=(f"OUT OF SCOPE for this deliverable. This tournament report covers TRACT 1 = NE/4 only. "
              f"Any pre-existing cells on this tab are inherited template scaffolding (Section 27-15N-24W example) "
              f"and are NOT Section 32 conclusions.")
    ws["A1"].font=Font(bold=True,size=9,color="9C0006")
    ws["A1"].alignment=wrap

wb.save(f"{S}/32-11N-25W_Beckham_Co_NE4_Tract1_Ownership_Reconstruction.xlsx")
print("FINAL workbook saved")
# verification
wb2=load_workbook(f"{S}/32-11N-25W_Beckham_Co_NE4_Tract1_Ownership_Reconstruction.xlsx")
print("sheets:",wb2.sheetnames)
# formula error scan
errs=0
for sn in wb2.sheetnames:
    for row in wb2[sn].iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.startswith("#") and c.value.endswith("!"): errs+=1
print("literal #REF-style errors:",errs)
