"""The DataBossX Command Center -- one end-to-end run for one project.

Orchestrates the existing engines into a single, non-destructive, evidence-first
pipeline:

    resolve project  ->  ingest & inventory (sha256 evidence trail)
                     ->  extract text / OCR notes
                     ->  chain the OGL <-> runsheet (Horizon intelligence layer)
                     ->  mineral / leasehold / WI / NRI math (databossx.ownership)
                     ->  aggregate defects & missing evidence
                     ->  client Excel + client PDF + HTML dashboard
                     ->  examiner worklist + run manifest + audit log

Every stage degrades gracefully: a missing optional dependency or an absent
input is logged in plain language and the run continues with what it has. The
source tree is only ever read; all output goes to a managed folder and prior
output is backed up first (zero destruction).
"""

from __future__ import annotations

import json
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from . import __version__
from .backup import backup_prior_output
from .ownership import MineralOwner, TractOwnership, division_of_interest
from .projects import GOLDEN_DEMO_KEY, get_project, resolve_root

OUTPUT_DIRNAME = "databossx_output"


@dataclass
class RunResult:
    project: str
    section: str
    source_root: str
    output_dir: str
    generated: str
    ok: bool = True
    rag: str = "GREEN"
    message: str = ""
    counts: Dict[str, int] = field(default_factory=dict)
    deliverables: Dict[str, str] = field(default_factory=dict)
    defects: List[Dict[str, object]] = field(default_factory=list)
    evidence: List[Dict[str, object]] = field(default_factory=list)
    tracts: List[TractOwnership] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


class _Audit:
    """A tiny append-only audit log written next to the output."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self.lines: List[str] = []
        path.parent.mkdir(parents=True, exist_ok=True)

    def __call__(self, event: str, detail: str = "", level: str = "INFO") -> None:
        line = f"{_now()}\t{level}\t{event}\t{detail}"
        self.lines.append(line)
        try:
            with self.path.open("a", encoding="utf-8") as fh:
                fh.write(line + "\n")
        except OSError:
            pass
        print(f"[{level}] {event} {('- ' + detail) if detail else ''}")


# ---------------------------------------------------------------------------
# Ownership-sheet reader
# ---------------------------------------------------------------------------
_OWNER_HEADERS = {
    "owner": "name", "name": "name", "party": "name",
    "mineral_interest": "mineral_interest", "mineral": "mineral_interest",
    "interest": "mineral_interest", "mi": "mineral_interest",
    "royalty": "royalty", "royalty_rate": "royalty", "rr": "royalty",
    "orri": "orri", "override": "orri", "overriding_royalty": "orri",
    "lessee": "lessee", "operator": "lessee",
    "tract": "tract", "legal": "tract", "legal_description": "tract",
    "gross_acres": "gross_acres", "acres": "gross_acres", "acreage": "gross_acres",
    "source": "source", "citation": "source", "evidence": "source",
}


def _canon_owner(h) -> Optional[str]:
    return _OWNER_HEADERS.get(str(h or "").strip().lower().replace(" ", "_"))


def read_ownership_sheet(path: Path) -> List[Dict[str, str]]:
    """Read an 'Ownership' sheet (if present) into raw owner dicts.

    Returns ``[]`` when the workbook has no ownership sheet -- the DOI stage is
    simply skipped rather than fabricated.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if "owner" in sn.lower() or "ownership" in sn.lower():
            ws = wb[sn]
            break
    if ws is None:
        wb.close()
        return []

    header_row, header_map = -1, {}
    for r_idx, values in enumerate(ws.iter_rows(values_only=True)):
        mapped = {i: _canon_owner(v) for i, v in enumerate(values)}
        mapped = {i: f for i, f in mapped.items() if f}
        if len(mapped) > len(header_map):
            header_row, header_map = r_idx, mapped
        if r_idx > 20:
            break

    rows: List[Dict[str, str]] = []
    if header_row >= 0:
        for r_idx, values in enumerate(ws.iter_rows(values_only=True)):
            if r_idx <= header_row:
                continue
            data = {f: (str(values[i]).strip()
                        if i < len(values) and values[i] is not None else "")
                    for i, f in header_map.items()}
            if data.get("name") or data.get("mineral_interest"):
                rows.append(data)
    wb.close()
    return rows


def ownership_to_tracts(rows: List[Dict[str, str]]) -> List[TractOwnership]:
    """Group raw owner dicts by tract and run the division of interest."""
    tracts: Dict[str, Dict[str, object]] = {}
    order: List[str] = []
    for row in rows:
        tract = row.get("tract", "") or "UNSPECIFIED"
        if tract not in tracts:
            tracts[tract] = {"gross": row.get("gross_acres", ""), "owners": []}
            order.append(tract)
        if row.get("gross_acres"):
            tracts[tract]["gross"] = row["gross_acres"]
        tracts[tract]["owners"].append(MineralOwner(
            name=row.get("name", ""),
            mineral_interest=row.get("mineral_interest", ""),
            royalty=row.get("royalty", ""),
            orri=row.get("orri", ""),
            lessee=row.get("lessee", ""),
            source=row.get("source", ""),
        ))
    results: List[TractOwnership] = []
    for tract in order:
        info = tracts[tract]
        results.append(division_of_interest(info["owners"], info["gross"], tract=tract))
    return results


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run_project(project_key: str, *, root: Optional[str] = None,
                output_dir: Optional[str] = None, make_pdf: bool = True,
                base: Optional[str] = None) -> RunResult:
    """Run the full slice for one project and return a :class:`RunResult`.

    Never raises for an expected operational problem (missing folder, missing
    optional dependency); those become a plain-language message on the result.
    """
    proj = get_project(project_key)
    generated = _now()
    if proj is None:
        return RunResult(project=project_key, section="", source_root="",
                         output_dir="", generated=generated, ok=False, rag="RED",
                         message=(f"Unknown project '{project_key}'. Run "
                                  "'databossx list' to see the registered projects."))

    src = resolve_root(proj.key, override=root, base=base)
    assert src is not None
    if not src.exists():
        return RunResult(
            project=proj.name, section=proj.section, source_root=str(src),
            output_dir="", generated=generated, ok=False, rag="RED",
            message=(f"Source folder for '{proj.name}' was not found:\n  {src}\n"
                     "Run DataBossX on the machine where the title files live, or "
                     f"set DATABOSSX_{proj.key.upper()}_ROOT / pass --root."))

    out = Path(output_dir) if output_dir else (src / OUTPUT_DIRNAME)
    out.mkdir(parents=True, exist_ok=True)
    backup_note = backup_prior_output(out)
    audit = _Audit(out / "databossx_audit.log")
    audit("run_start", f"v{__version__} project={proj.name} root={src}")
    if backup_note:
        audit("backup", backup_note)

    result = RunResult(project=proj.name, section=proj.section,
                       source_root=str(src), output_dir=str(out),
                       generated=generated)

    # -- ingest & evidence ---------------------------------------------------
    recs, texts = _ingest(src, out, audit, result)

    # -- chain (Horizon intelligence layer) ---------------------------------
    runsheet_rows, chained = _chain(recs, proj.section, audit, result)

    # -- ownership / WI / NRI ------------------------------------------------
    tracts = _ownership(recs, audit, result)
    result.tracts = tracts

    # -- defects & missing evidence -----------------------------------------
    _collect_defects(result, runsheet_rows, chained, tracts, texts)

    # -- counts & RAG --------------------------------------------------------
    result.counts = {
        "runsheet_rows": len(runsheet_rows),
        "tracts": len(tracts),
        "mineral_owners": sum(len(t.mineral_rows) for t in tracts),
        "doi_rows": sum(len(t.doi_rows) for t in tracts),
        "chain_breaks": len(chained.chain_breaks) if chained else 0,
        "review_items": sum(1 for r in runsheet_rows
                            if str(getattr(r, "status", "")).lower() not in ("", "ok")),
        "defects": len(result.defects),
        "evidence": len(result.evidence),
    }
    result.rag = _rag(result)

    # -- deliverables --------------------------------------------------------
    _write_deliverables(result, out, runsheet_rows, tracts, chained, audit,
                        make_pdf=make_pdf, synthetic=(proj.key == GOLDEN_DEMO_KEY))

    audit("run_done", f"rag={result.rag} deliverables={len(result.deliverables)} "
                      f"defects={len(result.defects)}")
    result.message = f"Run complete -- status {result.rag}. Outputs in {out}"
    return result


def _ingest(src: Path, out: Path, audit, result: RunResult):
    """Inventory + text extraction (reuses the grocery pipeline stages)."""
    recs, texts = [], {}
    try:
        # The root-level engines are already put on sys.path by databossx's
        # package __init__; keep this belt-and-suspenders insert for direct use.
        sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
        import grocery_report_pipeline as g

        log = g.BuildLog()
        recs = g.inventory(src, out, log)
        texts = g.extract_text(recs, out, log)
        for r in recs:
            tr = texts.get(r.path)
            result.evidence.append({
                "file": r.rel_path, "type": r.likely_type, "sha256": r.sha256,
                "bytes": r.size_bytes,
                "method": tr.method if tr else "",
                "note": tr.note if tr else "",
            })
        audit("ingest", f"files={len(recs)} extracted="
                        f"{sum(1 for t in texts.values() if t.status == 'extracted')}")
    except Exception as exc:  # pragma: no cover - defensive
        result.warnings.append(f"Ingest stage limited: {exc}")
        audit("ingest_warn", str(exc), level="WARN")
        # Fall back to a minimal inventory so evidence isn't empty.
        if not recs:
            for p in sorted(src.rglob("*")):
                if p.is_file() and OUTPUT_DIRNAME not in p.parts:
                    result.evidence.append({
                        "file": str(p.relative_to(src)), "type": p.suffix,
                        "sha256": "", "bytes": p.stat().st_size,
                        "method": "", "note": "minimal inventory"})
    return recs, texts


def _chain(recs, section: str, audit, result: RunResult):
    """Build the reconciled runsheet from the best OGL+runsheet workbook."""
    from horizon.pipeline import build_from_workbook, find_reference_workbook

    candidates = [Path(r.path) for r in recs] if recs else []
    wb = find_reference_workbook(candidates) if candidates else None
    if wb is None:
        result.warnings.append(
            "No OGL+runsheet workbook found -- runsheet/chain stage skipped.")
        audit("chain_skip", "no reference workbook", level="WARN")
        return [], None
    try:
        build = build_from_workbook(wb, section=section)
        audit("chain", f"workbook={wb.name} rows={len(build.report.rows)} "
                       f"breaks={len(build.chain_breaks)}")
        return build.report.rows, build
    except Exception as exc:
        result.errors.append(f"Chain stage failed on {wb.name}: {exc}")
        audit("chain_error", str(exc), level="ERROR")
        return [], None


def _ownership(recs, audit, result: RunResult) -> List[TractOwnership]:
    """Read an ownership sheet (if any) and run the WI/NRI math."""
    xlsx = [Path(r.path) for r in recs
            if r.ext in (".xlsx", ".xlsm")] if recs else []
    for path in xlsx:
        try:
            rows = read_ownership_sheet(path)
        except Exception as exc:
            result.warnings.append(f"Could not read ownership from {path.name}: {exc}")
            continue
        if rows:
            tracts = ownership_to_tracts(rows)
            audit("ownership", f"workbook={path.name} tracts={len(tracts)} "
                               f"owners={sum(len(t.mineral_rows) for t in tracts)}")
            return tracts
    result.warnings.append(
        "No ownership sheet found -- mineral/WI/NRI stage skipped.")
    audit("ownership_skip", "no ownership sheet", level="WARN")
    return []


def _collect_defects(result: RunResult, runsheet_rows, chained, tracts, texts):
    for r in runsheet_rows:
        status = str(getattr(r, "status", "")).lower()
        if status and status != "ok":
            result.defects.append({
                "severity": "red" if "escalated" in (getattr(r, "remarks", "") or "").lower()
                else "yellow",
                "category": "examiner-review",
                "subject": getattr(r, "instrument_number", ""),
                "detail": getattr(r, "remarks", ""),
                "source": "runsheet chain",
            })
    if chained:
        for c in chained.chain_breaks:
            result.defects.append({
                "severity": "yellow", "category": "chain-break",
                "subject": c.key, "detail": c.status, "source": "OGL/runsheet cross-ref"})
    for t in tracts:
        for d in t.defects:
            result.defects.append({
                "severity": "red", "category": "ownership-8/8",
                "subject": t.tract, "detail": d, "source": "division of interest"})
        for row in t.mineral_rows + t.doi_rows:
            if row.status == "review" and row.remarks:
                result.defects.append({
                    "severity": "yellow", "category": "ownership-review",
                    "subject": f"{t.tract}: {row.name}", "detail": row.remarks,
                    "source": "division of interest"})
    # Missing evidence: source docs that yielded no text (likely un-OCR'd scans).
    for tr in (texts or {}).values():
        if getattr(tr, "status", "") == "no-text" and tr.method in ("pdf", "image", "ocr"):
            result.defects.append({
                "severity": "yellow", "category": "missing-evidence",
                "subject": tr.rel_path, "detail": tr.note or "no text extracted",
                "source": "text extraction"})


def _rag(result: RunResult) -> str:
    if result.errors:
        return "RED"
    if any(d.get("severity") == "red" for d in result.defects):
        return "RED"
    if result.defects or result.warnings:
        return "YELLOW"
    return "GREEN"


def _write_deliverables(result, out: Path, runsheet_rows, tracts, chained,
                        audit, *, make_pdf: bool, synthetic: bool):
    meta = {
        "project": result.project, "section": result.section,
        "generated": result.generated, "source_root": result.source_root,
    }
    # Excel (required deliverable).
    try:
        from .excel_report import build_client_workbook
        xlsx = build_client_workbook(
            out / "DataBossX_Report.xlsx", meta=meta, counts=result.counts,
            rag=result.rag, runsheet_rows=runsheet_rows, tracts=tracts,
            defects=result.defects, evidence=result.evidence)
        result.deliverables["excel"] = str(xlsx)
        audit("deliverable_excel", xlsx.name)
    except Exception as exc:
        result.errors.append(f"Excel report failed: {exc}")
        audit("excel_error", str(exc), level="ERROR")

    # PDF (required deliverable).
    if make_pdf:
        try:
            from .pdf_report import build_client_pdf
            pdf = build_client_pdf(
                out / "DataBossX_Report.pdf", meta=meta, counts=result.counts,
                rag=result.rag, runsheet_rows=runsheet_rows, tracts=tracts,
                defects=result.defects, evidence=result.evidence,
                synthetic=synthetic)
            result.deliverables["pdf"] = str(pdf)
            audit("deliverable_pdf", pdf.name)
        except Exception as exc:
            result.warnings.append(
                f"PDF report skipped ({exc}). Install 'reportlab' to enable it.")
            audit("pdf_warn", str(exc), level="WARN")

    # Dashboard.
    try:
        from .dashboard import write_dashboard
        dash = write_dashboard(out / "dashboard.html", result)
        result.deliverables["dashboard"] = str(dash)
        audit("deliverable_dashboard", dash.name)
    except Exception as exc:
        result.warnings.append(f"Dashboard skipped: {exc}")
        audit("dashboard_warn", str(exc), level="WARN")

    # Examiner worklist + chain-breaks (reuse Horizon's versioned artifacts).
    if chained is not None:
        try:
            from horizon.artifacts import write_all
            worklists: List[str] = []
            for name, n in write_all(chained, out):
                audit("worklist", f"{name}: {n} row(s)")
                worklists.append(name)
            if worklists:
                result.deliverables["worklists"] = worklists
        except Exception as exc:
            result.warnings.append(f"Examiner worklist skipped: {exc}")

    # Machine-readable run manifest. Deliverables are recorded as portable
    # basenames (relative to the output folder) so the manifest is machine
    # independent; the full paths remain on the RunResult for the CLI.
    def _basename(v):
        if isinstance(v, list):
            return [Path(x).name for x in v]
        return Path(v).name
    try:
        manifest = {
            "schema_id": "dbx.run_manifest", "version": __version__,
            "project": result.project, "section": result.section,
            "generated": result.generated,
            "source_root": result.source_root, "output_dir": str(out),
            "rag": result.rag, "counts": result.counts,
            "deliverables": {k: _basename(v) for k, v in result.deliverables.items()},
            "defect_count": len(result.defects),
            "warnings": result.warnings, "errors": result.errors,
            "release_policy": {"technical_verification_is_not_release": True,
                               "human_approval_required": True},
        }
        (out / "run_manifest.json").write_text(
            json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
        result.deliverables["manifest"] = str(out / "run_manifest.json")
    except OSError as exc:
        result.warnings.append(f"Run manifest not written: {exc}")
