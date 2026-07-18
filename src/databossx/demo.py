"""Generate the bundled **synthetic** golden project.

This corpus is 100%% invented -- fictional owners, made-up instrument numbers,
and a survey label used only as a format placeholder. It exists so the entire
DataBossX slice (ingest -> chain -> ownership/WI/NRI -> defects -> Excel + PDF +
dashboard) can be demonstrated, tested, and self-checked **without any client
data**, honoring the repository's public boundary and the no-fabrication rule
(nothing here is presented as real title).

The corpus is deliberately imperfect so the defect machinery has something to
catch: a broken chain link, an over-conveyance, and a mineral estate that does
not sum to 8/8.
"""

from __future__ import annotations

from pathlib import Path

SYNTHETIC_BANNER = "SYNTHETIC DEMONSTRATION DATA -- NOT A REAL TITLE RECORD"

# (instrument, grantor, grantee, conveyed, legal, doc_type, date, acres)
_OGL_ROWS = [
    ("2018-0001", "UNITED STATES OF AMERICA", "AVERY MINERALS LLC", "1/2",
     "Sec 31 T12N R24W", "Mineral Deed", "2018-01-15", "160"),
    ("2019-0002", "AVERY MINERALS LLC", "BRAND ROYALTY LP", "1/4",
     "Sec 31 T12N R24W", "Mineral Deed", "2019-03-02", "160"),
    # Over-conveyance: Brand holds 1/4 but purports to convey 1/2 -> flagged.
    ("2020-0003", "BRAND ROYALTY LP", "CEDAR OPERATING INC", "1/2",
     "Sec 31 T12N R24W", "Assignment", "2020-06-10", "160"),
    # Second tract, clean chain.
    ("2017-0100", "STATE OF SYNTHETICA", "DELTA FAMILY TRUST", "3/4",
     "Sec 32 T12N R24W", "Patent", "2017-09-01", "80"),
    ("2021-0101", "DELTA FAMILY TRUST", "ECHO ENERGY LLC", "1/4",
     "Sec 32 T12N R24W", "Assignment", "2021-11-20", "80"),
]

# Runsheet notes -- note the missing entry for 2020-0003 (a chain break) and an
# orphan note referencing an instrument that never appears in the OGL register.
_RUNSHEET_ROWS = [
    ("2018-0001", "Recorded Bk 100 Pg 1; patent chain origin"),
    ("2019-0002", "Recorded Bk 210 Pg 45"),
    ("2017-0100", "Recorded Bk 90 Pg 12"),
    ("2021-0101", "Recorded Bk 305 Pg 88"),
    ("2099-9999", "Runsheet note with no matching OGL instrument (orphan)"),
]

# Ownership / division-of-interest inputs (structured mineral ownership).
# Tract 1 mineral estate is intentionally short of 8/8 (7/8) -> a defect.
# (owner, mineral_interest, royalty, orri, lessee, tract, gross_acres, source)
_OWNERSHIP_ROWS = [
    ("Avery Minerals LLC", "1/2", "3/16", "", "Cedar Operating Inc",
     "Sec 31 T12N R24W", "160", "Mineral Deed 2018-0001"),
    ("Brand Royalty LP", "1/4", "3/16", "1/32", "Cedar Operating Inc",
     "Sec 31 T12N R24W", "160", "Mineral Deed 2019-0002"),
    ("Foster Heirs (unleased)", "1/8", "", "", "",
     "Sec 31 T12N R24W", "160", "Probate 2016-0007"),
    # Tract 2 balances to 8/8 and is fully leased -> should be defect-free.
    ("Delta Family Trust", "1/2", "1/8", "", "Echo Energy LLC",
     "Sec 32 T12N R24W", "80", "Patent 2017-0100"),
    ("Echo Energy LLC", "1/2", "1/8", "", "Echo Energy LLC",
     "Sec 32 T12N R24W", "80", "Assignment 2021-0101"),
]


def _write_workbook(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ogl = wb.active
    ogl.title = "OGL"
    ogl.append(["Instrument", "Grantor", "Grantee", "Conveyed",
                "Legal", "Type", "Date", "Acres"])
    for row in _OGL_ROWS:
        ogl.append(list(row))

    run = wb.create_sheet("Runsheet")
    run.append(["Instrument", "Note"])
    for row in _RUNSHEET_ROWS:
        run.append(list(row))

    own = wb.create_sheet("Ownership")
    own.append(["Owner", "Mineral_Interest", "Royalty", "ORRI", "Lessee",
                "Tract", "Gross_Acres", "Source"])
    for row in _OWNERSHIP_ROWS:
        own.append(list(row))

    wb.save(path)
    wb.close()


def _write_source_docs(folder: Path) -> None:
    """A couple of plain-text 'deeds' so the ingest/extract stages have files."""
    (folder / "deed_2018-0001.txt").write_text(
        f"{SYNTHETIC_BANNER}\n\n"
        "MINERAL DEED\n"
        "Grantor: UNITED STATES OF AMERICA\n"
        "Grantee: AVERY MINERALS LLC\n"
        "Instrument No: 2018-0001   Recorded: Book 100 Page 1\n"
        "Conveys an undivided 1/2 mineral interest in and under:\n"
        "Section 31, Township 12 North, Range 24 West, containing 160 acres.\n",
        encoding="utf-8",
    )
    (folder / "assignment_2020-0003.txt").write_text(
        f"{SYNTHETIC_BANNER}\n\n"
        "ASSIGNMENT OF INTEREST\n"
        "Assignor: BRAND ROYALTY LP\n"
        "Assignee: CEDAR OPERATING INC\n"
        "Instrument No: 2020-0003\n"
        "Purports to assign an undivided 1/2 interest in Sec 31 T12N R24W.\n"
        "(Note: assignor's record interest is only 1/4 -- examiner review.)\n",
        encoding="utf-8",
    )


def _write_manifest(folder: Path) -> None:
    import json

    manifest = {
        "schema_id": "dbx.synthetic_project_manifest",
        "schema_version": "1.0",
        "project_id": "GOLDEN-DEMO-001",
        "client": "FICTIONAL CLIENT -- SYNTHETIC ONLY",
        "legal_description": "SYNTHETIC LAND ONLY (Sec 31 & 32, T12N R24W placeholder)",
        "source_policy": "IMMUTABLE_READ_ONLY",
        "release_policy": {
            "technical_verification_is_not_release": True,
            "human_approval_required": True,
        },
        "status": "SYNTHETIC_EXAMPLE",
        "notes": SYNTHETIC_BANNER,
    }
    (folder / "project_manifest.json").write_text(
        json.dumps(manifest, indent=2) + "\n", encoding="utf-8")


def build_golden_demo(dest: Path) -> Path:
    """(Re)generate the synthetic golden project under ``dest``. Idempotent.

    Returns the path to the generated OGL+Runsheet+Ownership workbook.
    """
    dest = Path(dest)
    sources = dest / "sources"
    sources.mkdir(parents=True, exist_ok=True)

    (dest / "README.txt").write_text(
        f"{SYNTHETIC_BANNER}\n\n"
        "This folder is a self-contained, invented land-title project used to\n"
        "demonstrate and test the DataBossX Command Center end to end. No file\n"
        "here represents a real owner, instrument, or tract.\n",
        encoding="utf-8",
    )
    _write_manifest(dest)
    _write_source_docs(sources)
    workbook = sources / "Golden_Demo_OGL_Runsheet.xlsx"
    _write_workbook(workbook)
    return workbook


if __name__ == "__main__":
    from .projects import golden_demo_root

    out = build_golden_demo(golden_demo_root())
    print(f"Synthetic golden project written under {out.parent.parent}")
