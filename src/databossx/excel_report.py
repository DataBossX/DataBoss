"""Client-ready Excel report builder.

Assembles one multi-sheet workbook from a completed run:

* **Summary**             -- run metadata, counts, and a RAG status line.
* **Runsheet**            -- the reconciled OGL chain (canonical title columns).
* **Mineral Ownership**   -- per-owner mineral interest + net mineral acres.
* **Division of Interest**-- WI / royalty / ORRI / NRI, exact.
* **Defects**             -- every examiner-review item and 8/8 discrepancy.
* **Evidence**            -- the source/traceability index.

Structure and column order are stable so a run is format-aligned with the
approved report layout. Blank cells stay blank -- the builder never zero-fills a
value the engines left undetermined.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Sequence

from .ownership import TractOwnership

# Spreadsheet formula-injection guard. openpyxl writes a leading "=" string as a
# live formula, and Excel/LibreOffice treat "= + - @" (and leading tab/CR) as
# formula initiators on open or CSV re-export. Document-derived values (a grantor
# literally named "=HYPERLINK(...)") must never become an executable formula in a
# client's workbook, so any such value is prefixed with a text-forcing apostrophe.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r", "\n")


def _sanitize_workbook(wb) -> None:
    """Neutralize formula-injection in every cell of every sheet before save."""
    from openpyxl.cell.cell import MergedCell

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                v = cell.value
                if isinstance(v, str) and v and v[0] in _FORMULA_TRIGGERS:
                    cell.value = "'" + v

# Canonical runsheet columns (mirror horizon.models.CANONICAL_COLUMNS order).
_RUNSHEET_COLS = [
    ("entry_no", "Entry No"),
    ("instrument_date", "Instrument Date"),
    ("recorded_date", "Recorded Date"),
    ("doc_type", "Doc Type"),
    ("grantor", "Grantor"),
    ("grantee", "Grantee"),
    ("instrument_number", "Instrument Number"),
    ("legal_description", "Legal Description"),
    ("gross_acres", "Gross Acres"),
    ("conveyed_interest", "Conveyed Interest"),
    ("retained_interest", "Retained Interest"),
    ("net_mineral_acres", "Net Mineral Acres"),
    ("status", "Status"),
    ("remarks", "Remarks"),
]

_HEADER_FILL = "1F3A5F"
_BANNER_FILL = "13233B"
_REVIEW_FILL = "FBE4D5"


def _style_header(ws, row_idx: int, ncols: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autosize(ws, max_width: int = 60) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 10),
                                      min(max_width, len(str(cell.value)) + 2))
    from openpyxl.utils import get_column_letter
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _banner(ws, text: str, ncols: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.append([text])
    ws.merge_cells(start_row=ws.max_row, start_column=1,
                   end_row=ws.max_row, end_column=max(ncols, 1))
    cell = ws.cell(row=ws.max_row, column=1)
    cell.fill = PatternFill("solid", fgColor=_BANNER_FILL)
    cell.font = Font(bold=True, color="FFFFFF", size=13)
    cell.alignment = Alignment(vertical="center")


def _mark_review_rows(ws, header_row: int, status_col: int) -> None:
    """Shade any data row whose Status cell is not a clean 'ok'."""
    from openpyxl.styles import PatternFill

    fill = PatternFill("solid", fgColor=_REVIEW_FILL)
    for r in range(header_row + 1, ws.max_row + 1):
        val = str(ws.cell(row=r, column=status_col).value or "").lower()
        if val and val != "ok":
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill


def _summary_sheet(wb, meta: Dict[str, object], counts: Dict[str, int],
                   rag: str) -> None:
    ws = wb.active
    ws.title = "Summary"
    _banner(ws, "DataBossX -- Land & Title Intelligence Report", 2)
    ws.append([])
    rows = [
        ("Project", meta.get("project", "")),
        ("Section", meta.get("section", "")),
        ("Generated (UTC)", meta.get("generated", "")),
        ("Source root", meta.get("source_root", "")),
        ("Overall status", rag),
        ("", ""),
        ("Runsheet rows", counts.get("runsheet_rows", 0)),
        ("Tracts", counts.get("tracts", 0)),
        ("Mineral owners", counts.get("mineral_owners", 0)),
        ("Division-of-interest rows", counts.get("doi_rows", 0)),
        ("Chain breaks", counts.get("chain_breaks", 0)),
        ("Examiner-review items", counts.get("review_items", 0)),
        ("Defects", counts.get("defects", 0)),
        ("Evidence records", counts.get("evidence", 0)),
    ]
    label_row = ws.max_row + 1
    for label, value in rows:
        ws.append([label, value])
    from openpyxl.styles import Font
    for r in range(label_row, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.append([])
    _banner(ws,
             "Unreviewed output is DRAFT work product -- not a certified abstract "
             "or title opinion. Human release gate required.", 2)
    _autosize(ws)


def _runsheet_sheet(wb, rows: Sequence[object]) -> None:
    ws = wb.create_sheet("Runsheet")
    headers = [h for _, h in _RUNSHEET_COLS]
    _banner(ws, "Runsheet -- reconciled chain of title", len(headers))
    ws.append(headers)
    header_row = ws.max_row
    _style_header(ws, header_row, len(headers))
    for row in rows:
        ws.append([_get(row, key) for key, _ in _RUNSHEET_COLS])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    status_idx = [i for i, (k, _) in enumerate(_RUNSHEET_COLS, start=1)
                  if k == "status"][0]
    _mark_review_rows(ws, header_row, status_idx)
    _autosize(ws)


def _mineral_sheet(wb, tracts: List[TractOwnership]) -> None:
    ws = wb.create_sheet("Mineral Ownership")
    headers = ["Tract", "Owner", "Mineral Interest", "Mineral Decimal",
               "Net Mineral Acres", "Lessee", "Status", "Remarks", "Source"]
    _banner(ws, "Mineral ownership", len(headers))
    ws.append(headers)
    header_row = ws.max_row
    _style_header(ws, header_row, len(headers))
    for t in tracts:
        for r in t.mineral_rows:
            d = r.display()
            ws.append([t.tract, d["name"], d["mineral_interest"],
                       d["mineral_decimal"], d["net_mineral_acres"], d["lessee"],
                       d["status"], d["remarks"], d["source"]])
        # Tract total row.
        from openpyxl.styles import Font
        ws.append([t.tract, "TOTAL", _fmt_frac(t.total_mineral), "", "", "",
                   "", "8/8 check", ""])
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _mark_review_rows(ws, header_row, 7)
    _autosize(ws)


def _doi_sheet(wb, tracts: List[TractOwnership]) -> None:
    ws = wb.create_sheet("Division of Interest")
    headers = ["Tract", "Party", "Role", "Working Interest", "Royalty Rate",
               "NRI (decimal)", "Lessee", "Status", "Remarks", "Source"]
    _banner(ws, "Division of interest -- working interest & net revenue", len(headers))
    ws.append(headers)
    header_row = ws.max_row
    _style_header(ws, header_row, len(headers))
    for t in tracts:
        for r in t.doi_rows:
            d = r.display()
            ws.append([t.tract, d["name"], d["role"], d["working_interest"],
                       d["royalty_rate"], d["nri"], d["lessee"], d["status"],
                       d["remarks"], d["source"]])
        from openpyxl.styles import Font
        ws.append([t.tract, "TOTAL", "",
                   _fmt_dec(t.total_working_interest),
                   "", _fmt_dec(t.total_nri), "", "WI & NRI totals", "", ""])
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _mark_review_rows(ws, header_row, 8)
    _autosize(ws)


def _defects_sheet(wb, defects: Sequence[Dict[str, object]]) -> None:
    ws = wb.create_sheet("Defects")
    headers = ["Severity", "Category", "Subject", "Detail", "Source"]
    _banner(ws, "Defects & missing evidence -- resolve before release", len(headers))
    ws.append(headers)
    header_row = ws.max_row
    _style_header(ws, header_row, len(headers))
    if not defects:
        ws.append(["info", "none", "", "No defects detected in this run.", ""])
    for d in defects:
        ws.append([d.get("severity", ""), d.get("category", ""),
                   d.get("subject", ""), d.get("detail", ""), d.get("source", "")])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _mark_review_rows(ws, header_row, 1)  # severity in col 1
    _autosize(ws)


def _evidence_sheet(wb, evidence: Sequence[Dict[str, object]]) -> None:
    ws = wb.create_sheet("Evidence")
    headers = ["File", "Type", "SHA-256", "Bytes", "Extract Method", "Note"]
    _banner(ws, "Evidence & audit trail -- source traceability", len(headers))
    ws.append(headers)
    header_row = ws.max_row
    _style_header(ws, header_row, len(headers))
    for e in evidence:
        ws.append([e.get("file", ""), e.get("type", ""), e.get("sha256", ""),
                   e.get("bytes", ""), e.get("method", ""), e.get("note", "")])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws)


def _abstract_sheet(wb, abstracts) -> None:
    ws = wb.create_sheet("Abstract")
    headers = ["Tract", "Seq", "Instrument Date", "Doc Type", "Grantor",
               "Grantee", "Instrument Number", "Conveyed", "Retained",
               "Net Mineral Acres", "Status", "Remarks"]
    _banner(ws, "Chain-of-title abstract -- ownership chain per tract", len(headers))
    ws.append(headers)
    header_row = ws.max_row
    _style_header(ws, header_row, len(headers))
    from openpyxl.styles import Font
    for a in abstracts or []:
        # A tract sub-header row for readability.
        ws.append([f"TRACT: {a.legal_description or a.tract}"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True)
        for e in a.entries:
            ws.append([a.tract, e.seq, e.instrument_date, e.doc_type, e.grantor,
                       e.grantee, e.instrument_number, e.conveyed_interest,
                       e.retained_interest, e.net_mineral_acres, e.status,
                       e.remarks])
    if not abstracts:
        ws.append(["", "", "", "", "", "", "", "", "", "",
                   "info", "No chain to abstract in this run."])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _mark_review_rows(ws, header_row, 11)  # status in col 11
    _autosize(ws)


def build_client_workbook(
    path: Path,
    *,
    meta: Dict[str, object],
    counts: Dict[str, int],
    rag: str,
    runsheet_rows: Sequence[object],
    tracts: List[TractOwnership],
    defects: Sequence[Dict[str, object]],
    evidence: Sequence[Dict[str, object]],
    abstracts: Optional[Sequence[object]] = None,
) -> Path:
    """Write the full client workbook to ``path`` and return it."""
    import openpyxl  # noqa: F401  (ensures a clear error if openpyxl is absent)

    wb = openpyxl.Workbook()
    _summary_sheet(wb, meta, counts, rag)
    _runsheet_sheet(wb, runsheet_rows)
    _abstract_sheet(wb, abstracts)
    _mineral_sheet(wb, tracts)
    _doi_sheet(wb, tracts)
    _defects_sheet(wb, defects)
    _evidence_sheet(wb, evidence)
    _sanitize_workbook(wb)  # formula-injection guard (client-facing deliverable)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path


# -- small helpers ----------------------------------------------------------
def _get(row: object, key: str) -> object:
    if isinstance(row, dict):
        return row.get(key, "") or ""
    return getattr(row, key, "") or ""


def _fmt_frac(frac) -> str:
    from horizon.interest import format_fraction
    return format_fraction(frac) if frac is not None else ""


def _fmt_dec(frac, places: int = 8) -> str:
    from horizon.interest import format_acres
    return format_acres(frac, places) if frac is not None else ""
