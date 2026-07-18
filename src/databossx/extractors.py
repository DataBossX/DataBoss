from __future__ import annotations

import csv
import hashlib
import io
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree


@dataclass(frozen=True)
class ExtractionResult:
    status: str
    extractor: str
    text: str
    note: str = ""

    @property
    def text_sha256(self) -> str | None:
        return hashlib.sha256(self.text.encode("utf-8")).hexdigest() if self.text else None


def _plain_text(path: Path) -> ExtractionResult:
    raw = path.read_bytes()
    text = raw.decode("utf-8", errors="replace")
    replacements = text.count("\ufffd")
    note = f"{replacements} invalid UTF-8 sequence(s) replaced" if replacements else ""
    return ExtractionResult("COMPLETE", "stdlib-text", text, note)


def _delimited(path: Path) -> ExtractionResult:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
    output = io.StringIO()
    for row in csv.reader(io.StringIO(text), dialect):
        output.write("\t".join(row) + "\n")
    return ExtractionResult("COMPLETE", "stdlib-delimited", output.getvalue())


def _docx(path: Path) -> ExtractionResult:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    words = [node.text or "" for node in root.iter() if node.tag.endswith("}t")]
    return ExtractionResult("COMPLETE", "stdlib-docx-xml", "\n".join(words))


def _xlsx(path: Path) -> ExtractionResult:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ExtractionResult(
            "UNAVAILABLE", "openpyxl", "", "Install openpyxl to extract spreadsheets"
        )
    workbook = load_workbook(path, read_only=True, data_only=False)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            if any(value is not None for value in row):
                lines.append("\t".join("" if value is None else str(value) for value in row))
    workbook.close()
    return ExtractionResult("COMPLETE", "openpyxl", "\n".join(lines))


def _pdf(path: Path) -> ExtractionResult:
    try:
        import fitz
    except ImportError:
        return ExtractionResult(
            "UNAVAILABLE", "pymupdf", "", "Install PyMuPDF for PDF text extraction"
        )
    document = fitz.open(path)
    text = "\n\n".join(page.get_text("text") for page in document)
    document.close()
    if not text.strip():
        return ExtractionResult(
            "NEEDS_OCR", "pymupdf", "", "PDF contains no embedded searchable text"
        )
    return ExtractionResult("COMPLETE", "pymupdf", text)


def _image(path: Path) -> ExtractionResult:
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return ExtractionResult(
            "UNAVAILABLE", "tesseract", "", "Install Pillow and pytesseract for image OCR"
        )
    try:
        with Image.open(path) as image:
            text = pytesseract.image_to_string(image)
    except Exception as exc:
        return ExtractionResult("FAILED", "tesseract", "", f"OCR failed: {exc}")
    return ExtractionResult("COMPLETE", "tesseract", text)


def extract(path: Path, extension: str) -> ExtractionResult:
    suffix = extension.lower()
    try:
        if suffix in {".txt", ".md", ".json", ".xml", ".html", ".htm", ".log"}:
            return _plain_text(path)
        if suffix in {".csv", ".tsv"}:
            return _delimited(path)
        if suffix == ".docx":
            return _docx(path)
        if suffix in {".xlsx", ".xlsm"}:
            return _xlsx(path)
        if suffix == ".pdf":
            return _pdf(path)
        if suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}:
            return _image(path)
        return ExtractionResult(
            "UNSUPPORTED", "none", "", f"No deterministic extractor for {suffix or 'no extension'}"
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        return ExtractionResult("FAILED", "deterministic", "", str(exc))


def extract_vault_object(vault_path: Path, extension: str) -> ExtractionResult:
    with tempfile.TemporaryDirectory(prefix="databossx-extract-") as temp:
        copy = Path(temp) / f"source{extension}"
        copy.write_bytes(vault_path.read_bytes())
        return extract(copy, extension)
