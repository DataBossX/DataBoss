from __future__ import annotations

from fractions import Fraction
from pathlib import Path
from xml.sax.saxutils import escape

from .economics import EconomicResult
from .ledger import Defect, LedgerResult, fraction_text

DRAFT_NOTICE = "DRAFT — NOT A CERTIFIED ABSTRACT OR TITLE OPINION"


def safe_cell(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _all_defects(
    ledger: LedgerResult, economics: list[EconomicResult]
) -> tuple[Defect, ...]:
    return (
        *ledger.defects,
        *(defect for result in economics for defect in result.defects),
    )


def _write_rows(sheet, headers: list[str], rows: list[list[object]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="183C33")
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        sheet.append([safe_cell(value) for value in row])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column]
        sheet.column_dimensions[column[0].column_letter].width = min(
            max(max(map(len, values)) + 2, 12), 48
        )


def render_xlsx(
    destination: Path,
    title_case: dict,
    opening: list[dict],
    instruments: list[dict],
    ledger: LedgerResult,
    economics: list[EconomicResult],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Case Summary"
    cover.append([DRAFT_NOTICE])
    cover["A1"].font = Font(bold=True, color="9C3F18", size=14)
    cover.append(["Title case", safe_cell(title_case["name"])])
    cover.append(["Legal description", safe_cell(title_case["legal_description"])])
    gross = Fraction(title_case["gross_acres_num"], title_case["gross_acres_den"])
    cover.append(["Gross acres (exact)", fraction_text(gross)])
    cover.append(
        [
            "Blocking defects",
            sum(
                defect.severity == "BLOCKING"
                for defect in _all_defects(ledger, economics)
            ),
        ]
    )
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 70

    runsheet = workbook.create_sheet("Runsheet")
    _write_rows(
        runsheet,
        [
            "Sequence", "Instrument Type", "Recording Reference", "Effective Date",
            "Grantor", "Grantee", "Stated Interest", "Basis", "Review Status",
            "Evidence Asset ID", "Character Range", "Notes",
        ],
        [
            [
                item["sequence_no"], item["instrument_type"],
                item["recording_reference"], item["effective_date"] or "",
                item["grantor_name"], item["grantee_name"],
                fraction_text(Fraction(item["conveyed_num"], item["conveyed_den"])),
                item["interest_basis"], item["review_status"],
                item["evidence_asset_version_id"] or "",
                (
                    f"{item['evidence_char_start']}:{item['evidence_char_end']}"
                    if item["evidence_char_start"] is not None
                    else ""
                ),
                item["notes"],
            ]
            for item in instruments
        ],
    )

    ownership = workbook.create_sheet("Current Ownership")
    _write_rows(
        ownership,
        [
            "Owner", "Interest", "Numerator", "Denominator",
            "Gross Acres", "Net Mineral Acres",
        ],
        [
            [
                owner, fraction_text(interest), interest.numerator,
                interest.denominator, fraction_text(gross),
                fraction_text(interest * gross),
            ]
            for owner, interest in sorted(ledger.ownership.items())
        ],
    )

    ledger_sheet = workbook.create_sheet("Ownership Ledger")
    _write_rows(
        ledger_sheet,
        [
            "Sequence", "Recording Reference", "Grantor", "Grantee", "Conveyed",
            "Grantor Before", "Grantor After", "Grantee After", "Estate Total",
        ],
        [
            [
                entry.sequence_no, entry.recording_reference, entry.grantor_name,
                entry.grantee_name, fraction_text(entry.conveyed),
                fraction_text(entry.grantor_before), fraction_text(entry.grantor_after),
                fraction_text(entry.grantee_after), fraction_text(entry.estate_total),
            ]
            for entry in ledger.entries
        ],
    )

    defects = workbook.create_sheet("Defects and Curative")
    _write_rows(
        defects,
        ["Severity", "Code", "Sequence", "Recording Reference", "Detail"],
        [
            [
                defect.severity, defect.code, defect.sequence_no or "",
                defect.recording_reference or "", defect.detail,
            ]
            for defect in _all_defects(ledger, economics)
        ],
    )

    evidence = workbook.create_sheet("Evidence Index")
    if economics:
        evidence_records = [
            (
                "Mineral Instrument",
                item["sequence_no"],
                item["recording_reference"],
                item,
            )
            for item in instruments
        ]
        for unit in title_case["economic_units"]:
            evidence_records.append(
                ("Economic Unit", "", unit["lease_recording_reference"], unit)
            )
            evidence_records.extend(
                (
                    f"Economic {event['event_kind']}",
                    event["sequence_no"],
                    event["recording_reference"],
                    event,
                )
                for event in unit["events"]
            )
        _write_rows(
            evidence,
            [
                "Record Type", "Sequence", "Recording Reference",
                "Asset Version ID", "Source SHA-256", "Extraction SHA-256",
                "Span SHA-256", "Character Start", "Character End",
                "Cited Source Text",
            ],
            [
                [
                    record_type, sequence, recording,
                    item["evidence_asset_version_id"] or "",
                    item.get("evidence_source_sha256") or "",
                    item.get("evidence_extraction_sha256") or "",
                    item.get("evidence_span_sha256") or "",
                    (
                        item["evidence_char_start"]
                        if item["evidence_char_start"] is not None
                        else ""
                    ),
                    (
                        item["evidence_char_end"]
                        if item["evidence_char_end"] is not None
                        else ""
                    ),
                    item.get("evidence_span_text") or "",
                ]
                for record_type, sequence, recording, item in evidence_records
            ],
        )
        units_sheet = workbook.create_sheet("Calculation Units")
        _write_rows(
            units_sheet,
            [
                "Unit", "Legal Description", "Lease Reference", "As Of",
                "Depth Scope", "Product Scope", "Gross Acres",
                "Leased Mineral Fraction", "Base Royalty", "WI Basis",
                "Review Status", "Unsupported Terms",
            ],
            [
                [
                    unit["unit_label"], unit["legal_description"],
                    unit["lease_recording_reference"], unit["effective_as_of"] or "",
                    unit["depth_scope"], unit["product_scope"],
                    fraction_text(
                        Fraction(unit["gross_acres_num"], unit["gross_acres_den"])
                    ),
                    fraction_text(
                        Fraction(
                            unit["leased_mineral_num"],
                            unit["leased_mineral_den"],
                        )
                    ),
                    fraction_text(
                        Fraction(
                            unit["base_royalty_num"], unit["base_royalty_den"]
                        )
                    ),
                    unit["wi_basis"], unit["review_status"],
                    unit["unsupported_terms"],
                ]
                for unit in title_case["economic_units"]
            ],
        )
        facts_sheet = workbook.create_sheet("Reviewed Lease Facts")
        _write_rows(
            facts_sheet,
            [
                "Unit", "Evidence Asset", "Source SHA-256",
                "Span SHA-256", "Cited Source Text",
            ],
            [
                [
                    unit["unit_label"], unit["evidence_asset_version_id"] or "",
                    unit["evidence_source_sha256"] or "",
                    unit["evidence_span_sha256"] or "",
                    unit["evidence_span_text"] or "",
                ]
                for unit in title_case["economic_units"]
            ],
        )
        leasehold_sheet = workbook.create_sheet("Leasehold Ledger")
        _write_rows(
            leasehold_sheet,
            [
                "Unit", "Sequence", "Recording Reference", "Assignor",
                "Assignee", "Conveyed", "Assignor Before", "Assignor After",
                "Assignee After",
            ],
            [
                [
                    result.unit_label, entry.sequence_no,
                    entry.recording_reference, entry.from_party, entry.to_party,
                    fraction_text(entry.conveyed),
                    fraction_text(entry.from_before),
                    fraction_text(entry.from_after),
                    fraction_text(entry.to_after),
                ]
                for result in economics
                for entry in result.entries
            ],
        )
        wi_sheet = workbook.create_sheet("Leasehold and WI")
        _write_rows(
            wi_sheet,
            [
                "Unit", "Party", "Leasehold", "Working Interest",
                "Net Leasehold Acres", "Tract-Weighted WI",
            ],
            [
                [
                    result.unit_label, party,
                    fraction_text(result.leasehold.get(party, Fraction())),
                    fraction_text(result.working_interest.get(party, Fraction())),
                    fraction_text(
                        result.net_leasehold_acres.get(party, Fraction())
                    ),
                    fraction_text(result.tract_weighted_wi.get(party, Fraction())),
                ]
                for result in economics
                for party in sorted(
                    set(result.leasehold) | set(result.working_interest)
                )
            ],
        )
        nri_sheet = workbook.create_sheet("Net Revenue Detail")
        _write_rows(
            nri_sheet,
            [
                "Unit", "Party", "Working Interest", "NRI",
                "Burden Received", "Tract-Weighted NRI",
            ],
            [
                [
                    result.unit_label, party,
                    fraction_text(result.working_interest.get(party, Fraction())),
                    fraction_text(result.nri.get(party, Fraction())),
                    fraction_text(result.burdens_received.get(party, Fraction())),
                    fraction_text(result.tract_weighted_nri.get(party, Fraction())),
                ]
                for result in economics
                for party in sorted(set(result.working_interest) | set(result.nri))
            ],
        )
        conservation = workbook.create_sheet("Revenue Conservation")
        _write_rows(
            conservation,
            [
                "Unit", "Leased Mineral Fraction", "Royalty Share",
                "Tract-Weighted NRI Total", "Conserved Total",
                "Blocking Defects",
            ],
            [
                [
                    result.unit_label,
                    fraction_text(result.leased_mineral_fraction),
                    fraction_text(result.royalty_share),
                    fraction_text(sum(result.tract_weighted_nri.values(), Fraction())),
                    fraction_text(
                        result.royalty_share
                        + sum(result.tract_weighted_nri.values(), Fraction())
                    ),
                    len(result.blocking_defects),
                ]
                for result in economics
            ],
        )
    else:
        _write_rows(
            evidence,
            [
                "Sequence", "Recording Reference", "Asset Version ID",
                "Source SHA-256", "Extraction SHA-256", "Span SHA-256",
                "Character Start", "Character End", "Cited Source Text",
            ],
            [
                [
                    item["sequence_no"], item["recording_reference"],
                    item["evidence_asset_version_id"] or "",
                    item.get("evidence_source_sha256") or "",
                    item.get("evidence_extraction_sha256") or "",
                    item.get("evidence_span_sha256") or "",
                    (
                        item["evidence_char_start"]
                        if item["evidence_char_start"] is not None
                        else ""
                    ),
                    (
                        item["evidence_char_end"]
                        if item["evidence_char_end"] is not None
                        else ""
                    ),
                    item.get("evidence_span_text") or "",
                ]
                for item in instruments
            ],
        )
    workbook.save(destination)


def render_pdf(
    destination: Path,
    title_case: dict,
    ledger: LedgerResult,
    economics: list[EconomicResult],
) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table

    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(destination),
        pagesize=letter,
        rightMargin=0.55 * inch,
        leftMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
        title=f"Draft Abstract Aid — {title_case['name']}",
    )
    story = [
        Paragraph(DRAFT_NOTICE, styles["Title"]),
        Spacer(1, 12),
        Paragraph(
            f"<b>Title case:</b> {escape(str(title_case['name']))}",
            styles["BodyText"],
        ),
        Paragraph(
            f"<b>Legal description:</b> "
            f"{escape(str(title_case['legal_description']))}",
            styles["BodyText"],
        ),
        Paragraph(
            "This examiner aid organizes reviewed evidence and exact arithmetic. "
            "It requires qualified human review and is not legal advice.",
            styles["BodyText"],
        ),
        Spacer(1, 14),
        Paragraph("Current mineral ownership", styles["Heading2"]),
    ]
    ownership_rows = [["Owner", "Exact interest", "Numerator", "Denominator"]]
    ownership_rows.extend(
        [
            owner,
            fraction_text(interest),
            str(interest.numerator),
            str(interest.denominator),
        ]
        for owner, interest in sorted(ledger.ownership.items())
    )
    ownership_table = Table(
        ownership_rows,
        repeatRows=1,
        colWidths=[3 * inch, 1.2 * inch, 0.9 * inch, 1 * inch],
    )
    ownership_table.setStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#183C33")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
    )
    story.append(ownership_table)
    if economics:
        story.extend(
            [Spacer(1, 14), Paragraph("Leasehold, WI and NRI", styles["Heading2"])]
        )
        economic_rows = [
            ["Unit", "Party", "Leasehold", "WI", "NRI", "Tract NRI"]
        ]
        for result in economics:
            for party in sorted(
                set(result.leasehold)
                | set(result.working_interest)
                | set(result.nri)
            ):
                economic_rows.append(
                    [
                        result.unit_label,
                        party,
                        fraction_text(result.leasehold.get(party, Fraction())),
                        fraction_text(
                            result.working_interest.get(party, Fraction())
                        ),
                        fraction_text(result.nri.get(party, Fraction())),
                        fraction_text(
                            result.tract_weighted_nri.get(party, Fraction())
                        ),
                    ]
                )
        economic_table = Table(
            economic_rows,
            repeatRows=1,
            colWidths=[1.2 * inch, 1.8 * inch, 0.8 * inch, 0.7 * inch, 0.7 * inch, 0.9 * inch],
        )
        economic_table.setStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#183C33")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
        story.append(economic_table)
    story.extend(
        [PageBreak(), Paragraph("Defects and curative", styles["Heading2"])]
    )
    defect_rows = [["Severity", "Code", "Recording reference", "Detail"]]
    defect_rows.extend(
        [
            defect.severity,
            defect.code,
            defect.recording_reference or "",
            defect.detail,
        ]
        for defect in _all_defects(ledger, economics)
    )
    if len(defect_rows) == 1:
        defect_rows.append(
            ["—", "NONE", "", "No deterministic blocking defect detected"]
        )
    defect_table = Table(
        defect_rows,
        repeatRows=1,
        colWidths=[0.8 * inch, 1.25 * inch, 1.3 * inch, 3.6 * inch],
    )
    defect_table.setStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#183C33")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]
    )
    story.append(defect_table)
    document.build(story)
