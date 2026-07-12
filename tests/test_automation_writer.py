"""Workbook writer tests: safety, atomicity, locking, and OOXML preservation.

All data is synthetic and fictional.
"""

from __future__ import annotations

import io
import threading
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from automation.writer import (
    WorkbookSecurityError,
    _workbook_lock,
    create_staging_copy,
    sanitize_cell_value,
    update_workbook,
)

HEADERS = [
    "Name (Owner)",
    "Verified Address",
    "Status",
    "Last Affecting Doc No",
    "Last Doc Type",
    "Last Doc Date",
    "Source URL",
    "Notes",
    "Confidence%",
]

OWNER_A = "Example Owner A"
OWNER_B = "Example Owner B"


def make_workbook(path: Path, owners=(OWNER_A,)) -> None:
    workbook = Workbook()
    target = workbook.active
    target.title = "DSU NOTICE LIST"
    target.append(HEADERS)
    for owner in owners:
        target.append([owner])
    supporting = workbook.create_sheet("Supporting Evidence")
    supporting["A1"] = "=1+1"
    supporting["B2"] = "must survive"
    workbook.save(path)


def _make_signed_xlsx(path: Path) -> None:
    make_workbook(path)
    with zipfile.ZipFile(path, "a") as package:
        package.writestr("_xmlsignatures/sig1.xml", "<Signature/>")


def _make_xlsm(path: Path, sentinel: bytes) -> None:
    """Build a representative macro-enabled workbook with a VBA project part."""
    buffer = io.BytesIO()
    make_workbook(buffer)
    buffer.seek(0)
    with zipfile.ZipFile(buffer) as original:
        parts = {name: original.read(name) for name in original.namelist()}

    content_types = parts["[Content_Types].xml"].decode("utf-8")
    content_types = content_types.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    content_types = content_types.replace(
        "</Types>",
        '<Override PartName="/xl/vbaProject.bin" '
        'ContentType="application/vnd.ms-office.vbaProject"/></Types>',
    )
    parts["[Content_Types].xml"] = content_types.encode("utf-8")

    rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
    rels = rels.replace(
        "</Relationships>",
        '<Relationship Id="rIdVbaProject" '
        'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
        'Target="vbaProject.bin"/></Relationships>',
    )
    parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")
    parts["xl/vbaProject.bin"] = sentinel

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as package:
        for name, data in parts.items():
            package.writestr(name, data)


def test_writer_preserves_source_and_unrelated_sheets(tmp_path):
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    source_before = source.read_bytes()
    staging = create_staging_copy(source, tmp_path / "staging")

    changed = update_workbook(
        staging,
        "DSU NOTICE LIST",
        OWNER_A.lower(),
        {"status": "Verified", "doc_no": "SYN-0001", "confidence": 0.97},
    )

    assert changed is True
    assert source.read_bytes() == source_before
    workbook = load_workbook(staging, data_only=False)
    assert workbook.sheetnames == ["DSU NOTICE LIST", "Supporting Evidence"]
    assert workbook["Supporting Evidence"]["A1"].value == "=1+1"
    assert workbook["Supporting Evidence"]["B2"].value == "must survive"
    assert workbook["DSU NOTICE LIST"]["C2"].value == "Verified"
    assert workbook["DSU NOTICE LIST"]["I2"].value == 97
    workbook.close()


def test_every_written_field_round_trips(tmp_path):
    path = tmp_path / "fields.xlsx"
    make_workbook(path)
    row = {
        "verified_address": "1 Example Way, Fictional County",
        "status": "Verified",
        "doc_no": "SYN-2026-0001",
        "instrument": "Warranty Deed",
        "recording_date": "2026-01-02",
        "source_url": "https://example.test/records/SYN-0001",
        "notes": "synthetic note",
        "confidence": 0.5,
    }
    assert update_workbook(path, "DSU NOTICE LIST", OWNER_A, row)
    workbook = load_workbook(path)
    sheet = workbook["DSU NOTICE LIST"]
    assert sheet["B2"].value == row["verified_address"]
    assert sheet["C2"].value == "Verified"
    assert sheet["D2"].value == "SYN-2026-0001"
    assert sheet["E2"].value == "Warranty Deed"
    assert sheet["F2"].value == "2026-01-02"
    assert sheet["G2"].value == "https://example.test/records/SYN-0001"
    assert sheet["H2"].value == "synthetic note"
    assert sheet["I2"].value == 50
    workbook.close()


@pytest.mark.parametrize(
    "payload",
    [
        "=1+1",
        "+1+1",
        "-1+1",
        "@SUM(A1:A9)",
        "=cmd|' /C calc'!A0",
        "\t=1+1",
    ],
)
def test_formula_injection_is_neutralized(tmp_path, payload):
    path = tmp_path / "inject.xlsx"
    make_workbook(path)
    assert update_workbook(path, "DSU NOTICE LIST", OWNER_A, {"notes": payload})
    workbook = load_workbook(path)
    value = workbook["DSU NOTICE LIST"]["H2"].value
    workbook.close()
    # Stored as literal text (apostrophe-prefixed), never as an active formula.
    assert isinstance(value, str)
    assert value.startswith("'")
    assert not value.startswith(("=", "+", "-", "@"))


def test_sanitize_allowlisted_formula_passes_through():
    assert sanitize_cell_value("=SUM(A1:A2)", allowlisted_formulas=["=SUM(A1:A2)"]) == "=SUM(A1:A2)"
    assert sanitize_cell_value("plain text") == "plain text"
    assert sanitize_cell_value(97) == 97


def test_writer_does_not_save_when_owner_is_missing(tmp_path):
    workbook_path = tmp_path / "staging.xlsx"
    make_workbook(workbook_path)
    before = workbook_path.read_bytes()
    assert update_workbook(workbook_path, "DSU NOTICE LIST", "Unknown Owner", {}) is False
    assert workbook_path.read_bytes() == before


def test_stale_lock_file_does_not_block_recovery(tmp_path):
    workbook_path = tmp_path / "staging.xlsx"
    make_workbook(workbook_path)
    lock_path = tmp_path / ".staging.xlsx.lock"
    lock_path.write_text("left by terminated process", encoding="utf-8")

    assert update_workbook(
        workbook_path, "DSU NOTICE LIST", OWNER_A, {"status": "Recovered", "confidence": 1}
    )
    workbook = load_workbook(workbook_path)
    assert workbook["DSU NOTICE LIST"]["C2"].value == "Recovered"
    workbook.close()


def test_lock_is_exclusive_while_held(tmp_path):
    workbook_path = tmp_path / "locked.xlsx"
    make_workbook(workbook_path)
    with _workbook_lock(workbook_path):
        with pytest.raises(TimeoutError):
            with _workbook_lock(workbook_path, timeout=0.5):
                pass


def test_concurrent_updates_serialize_without_data_loss(tmp_path):
    workbook_path = tmp_path / "concurrent.xlsx"
    make_workbook(workbook_path, owners=(OWNER_A, OWNER_B))

    errors: list[Exception] = []

    def worker(owner: str, status: str) -> None:
        try:
            update_workbook(workbook_path, "DSU NOTICE LIST", owner, {"status": status})
        except Exception as exc:  # pragma: no cover - surfaced via assertion
            errors.append(exc)

    threads = [
        threading.Thread(target=worker, args=(OWNER_A, "A-Verified")),
        threading.Thread(target=worker, args=(OWNER_B, "B-Verified")),
    ]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()

    assert not errors
    workbook = load_workbook(workbook_path)
    sheet = workbook["DSU NOTICE LIST"]
    assert sheet["C2"].value == "A-Verified"
    assert sheet["C3"].value == "B-Verified"
    workbook.close()
    # Atomic replacement leaves no partial temp files behind.
    assert not list(tmp_path.glob(".concurrent-*"))


def test_unsupported_workbook_type_is_refused(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")
    with pytest.raises(WorkbookSecurityError, match="unsupported workbook type"):
        update_workbook(path, "Sheet", OWNER_A, {})


def test_signed_workbook_is_refused_fail_closed(tmp_path):
    path = tmp_path / "signed.xlsx"
    _make_signed_xlsx(path)
    with pytest.raises(WorkbookSecurityError, match="signed"):
        update_workbook(path, "DSU NOTICE LIST", OWNER_A, {"status": "x"})


def test_xlsm_vba_project_is_preserved(tmp_path):
    sentinel = b"SYNTHETIC_VBA_PROJECT_BINARY_v1"
    path = tmp_path / "macro.xlsm"
    _make_xlsm(path, sentinel)

    changed = update_workbook(path, "DSU NOTICE LIST", OWNER_A, {"status": "Verified"})
    assert changed is True

    with zipfile.ZipFile(path) as package:
        names = package.namelist()
        assert "xl/vbaProject.bin" in names
        assert package.read("xl/vbaProject.bin") == sentinel
    workbook = load_workbook(path, keep_vba=True)
    assert workbook["DSU NOTICE LIST"]["C2"].value == "Verified"
    workbook.close()
