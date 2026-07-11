import json
from pathlib import Path

import openpyxl
from PIL import Image, ImageDraw

import databoss_title_factory.core as factory_core
from databoss_title_factory.core import (
    OUTPUT_DIR_NAME,
    build_inventory,
    build_runsheet,
    export_safe_xlsx,
    extract_and_reconcile,
    latest_run,
    preprocess_images,
    resume_pipeline,
    run_ocr,
    start_run,
    tournament_reconcile,
)
from databoss_title_factory.workbook_audit import audit_workbook, compare_workbooks
from databoss_title_factory.project_db import (
    is_pause_requested,
    request_pause,
    stage_statuses,
)


def _source_text() -> str:
    return """
Instrument Number: 2026-001234
Warranty Deed
Instrument Date: 01/02/2026
Recorded Date: 01/05/2026
Grantor: Ada Owner
Grantee: Beacon Minerals LLC
Book: 44
Page: 219
Legal Description: NE/4 Section 31, Township 12 North, Range 24 West
Interest Conveyed: 1/2 mineral interest
"""


def test_inventory_ocr_extract_and_runsheet_are_cited_and_versioned(tmp_path: Path):
    project = tmp_path / "project"
    project.mkdir()
    (project / "deed.txt").write_text(_source_text(), encoding="utf-8")

    ctx = start_run(project)
    inventory = build_inventory(ctx)
    assert len(inventory) == 1
    assert inventory[0]["source_path"] == "deed.txt"

    ocr = run_ocr(ctx)
    assert ocr[0]["citation"] == "deed.txt#file"
    assert ocr[0]["method"] == "native_text"

    instruments = extract_and_reconcile(ctx, weak_threshold=0.60)
    assert instruments[0]["instrument_number"] == "2026-001234"
    assert "deed.txt#file" in instruments[0]["citation"]
    assert instruments[0]["status"] == "PROVISIONAL - SAMPLE QA"
    assert instruments[0]["field_provenance"]["instrument_number"]["source_file_hash"]

    bundle = build_runsheet(ctx)
    assert len(bundle["runsheet"]) == 1
    assert (ctx.run_dir / "instruments.json").exists()
    assert (ctx.run_dir / "instruments.csv").exists()
    archive = [
        json.loads(line)
        for line in (ctx.run_dir / "candidates" / "candidate_archive.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
    ]
    assert len(archive) == 2
    assert all(row["status"] == "CANDIDATE - RETAINED" and row["sha256"] for row in archive)
    assert (project / OUTPUT_DIR_NAME / "latest_run.txt").read_text().strip() == ctx.run_id

    second = start_run(project)
    assert second.run_id != ctx.run_id
    assert ctx.run_dir.exists()
    assert latest_run(project).run_id == second.run_id
    request_pause(second.output_dir, second.run_id)
    assert is_pause_requested(second.output_dir, second.run_id) is True
    request_pause(second.output_dir, second.run_id, False)
    assert is_pause_requested(second.output_dir, second.run_id) is False
    statuses = stage_statuses(ctx.output_dir, ctx.run_id)
    assert {row["stage"] for row in statuses} >= {"INVENTORY", "OCR", "RECONCILE", "RUNSHEET"}


def test_preprocess_images_creates_copy_and_keeps_original(tmp_path: Path):
    project = tmp_path / "project"
    project.mkdir()
    source = project / "scan.png"
    image = Image.new("RGB", (500, 300), "white")
    ImageDraw.Draw(image).text((20, 20), "Instrument Number: 12345", fill="black")
    image.save(source)
    original = source.read_bytes()

    ctx = start_run(project)
    build_inventory(ctx)
    pages = preprocess_images(ctx)

    assert len(pages) == 1
    assert (ctx.run_dir / pages[0]["preprocessed_path"]).exists()
    assert {
        "orientation_corrected", "grayscale", "deskewed",
        "contrast_enhanced", "thresholded", "sharpened",
    }.issubset({variant["recipe"] for variant in pages[0]["variants"]})
    assert source.read_bytes() == original


def test_tournament_quarantines_unsupported_candidate():
    cursor = [
        {
            "instrument_number": "",
            "grantor": "Ada Owner",
            "confidence": 0.9,
            "citation": "",
            "source_path": "scan.pdf",
            "source_locator": "page=1",
        }
    ]
    reconciled = tournament_reconcile(cursor, [], weak_threshold=0.4)
    assert reconciled[0]["status"] == "QUARANTINED - REVIEW REQUIRED"
    assert "missing instrument number" in reconciled[0]["reconciliation_notes"]
    assert "missing source citation" in reconciled[0]["reconciliation_notes"]


def test_tournament_reconciles_instrument_number_disagreement_by_citation():
    common = {
        "grantor": "Ada Owner",
        "grantee": "Beacon LLC",
        "confidence": 0.95,
        "citation": "deed.pdf#page=1",
        "source_path": "deed.pdf",
        "source_locator": "page=1",
    }
    cursor = [{**common, "instrument_number": "2026-100"}]
    codex = [{**common, "instrument_number": "2026-101"}]

    reconciled = tournament_reconcile(cursor, codex, weak_threshold=0.40)

    assert len(reconciled) == 1
    assert "instrument_number" in reconciled[0]["reconciliation_notes"]
    assert reconciled[0]["instrument_number"] == ""
    assert reconciled[0]["status"] == "QUARANTINED - REVIEW REQUIRED"


def test_image_ocr_retains_blocks_and_field_bounding_boxes(tmp_path: Path, monkeypatch):
    project = tmp_path / "project"
    project.mkdir()
    image = Image.new("RGB", (800, 400), "white")
    image.save(project / "instrument.png")
    words = [
        ("Instrument", 10, 10, 90, 30),
        ("Number:", 95, 10, 150, 30),
        ("2026-55", 155, 10, 220, 30),
        ("Warranty", 10, 40, 80, 60),
        ("Deed", 85, 40, 125, 60),
    ]

    def fake_ocr(_path, config="--oem 3 --psm 6"):
        return {
            "text": "Instrument Number: 2026-55\nWarranty Deed",
            "confidence": 0.98,
            "blocks": [
                {
                    "text": text,
                    "normalized_text": text,
                    "confidence": 0.98,
                    "bounding_box": {
                        "x1": x1, "y1": y1, "x2": x2, "y2": y2,
                        "coordinate_space": "pixels",
                    },
                    "block_number": 1,
                    "paragraph_number": 1,
                    "line_number": 1,
                    "word_number": index,
                }
                for index, (text, x1, y1, x2, y2) in enumerate(words, start=1)
            ],
            "ocr_config": config,
            "engine_version": "test-engine",
        }

    monkeypatch.setattr(factory_core, "_ocr_image", fake_ocr)
    ctx = start_run(project)
    build_inventory(ctx)
    ocr = run_ocr(ctx)
    rows = extract_and_reconcile(ctx, weak_threshold=0.4)

    assert ocr[0]["blocks"][0]["bounding_box"]["x1"] == 10
    provenance = rows[0]["field_provenance"]["instrument_number"]
    assert provenance["source_bounding_box"] == {
        "x1": 155,
        "y1": 10,
        "x2": 220,
        "y2": 30,
        "coordinate_space": "pixels",
    }
    assert provenance["review_status"] == "DIRECT_SOURCE_SUPPORT"


def test_runsheet_sorts_dates_chronologically(tmp_path: Path):
    project = tmp_path / "project"
    project.mkdir()
    ctx = start_run(project)
    records = [
        {"instrument_number": "2", "recorded_date": "12/01/2025", "status": "ACCEPTED"},
        {"instrument_number": "1", "recorded_date": "02/01/2025", "status": "ACCEPTED"},
    ]
    (ctx.run_dir / "instruments.json").write_text(json.dumps(records), encoding="utf-8")

    bundle = build_runsheet(ctx)

    assert [row["instrument_number"] for row in bundle["runsheet"]] == ["1", "2"]


def test_external_candidate_without_current_run_citation_is_quarantined(tmp_path: Path):
    project = tmp_path / "project"
    project.mkdir()
    (project / "source.txt").write_text(_source_text(), encoding="utf-8")
    candidate = {
        "instrument_number": "2026-001234",
        "instrument_type": "Warranty Deed",
        "grantor": "Ada Owner",
        "grantee": "Beacon Minerals LLC",
        "legal_description": "NE/4 Section 31",
        "confidence": 0.99,
        "citation": "different-file.pdf#page=99",
    }
    candidate_path = project / "external.json"
    candidate_path.write_text(json.dumps([candidate]), encoding="utf-8")
    ctx = start_run(project)
    build_inventory(ctx)
    run_ocr(ctx)

    rows = extract_and_reconcile(
        ctx,
        weak_threshold=0.4,
        cursor_json=candidate_path,
        codex_json=candidate_path,
    )

    assert rows[0]["status"] == "QUARANTINED - REVIEW REQUIRED"
    assert "current OCR evidence ledger" in rows[0]["reconciliation_notes"]


def test_external_candidate_cannot_forge_source_support(tmp_path: Path):
    project = tmp_path / "project"
    project.mkdir()
    source = project / "source.txt"
    source.write_text("Instrument Number: REAL-123\nWarranty Deed", encoding="utf-8")
    ctx = start_run(project)
    build_inventory(ctx)
    evidence = run_ocr(ctx)[0]
    forged = {
        "instrument_number": "FORGED-999",
        "instrument_type": "Warranty Deed",
        "confidence": 1.0,
        "citation": evidence["citation"],
        "source_path": evidence["source_path"],
        "source_file_hash": evidence["source_file_hash"],
        "unrecognized_but_must_be_archived": {"raw": "retain me"},
        "field_provenance": {
            "instrument_number": {
                "source_support_score": 1.0,
                "source_bounding_box": {
                    "x1": 1, "y1": 1, "x2": 2, "y2": 2,
                    "coordinate_space": "pixels",
                },
            }
        },
    }
    candidate_path = project / "forged.json"
    candidate_path.write_text(json.dumps([forged, "malformed candidate"]), encoding="utf-8")

    rows = extract_and_reconcile(
        ctx,
        weak_threshold=0.4,
        cursor_json=candidate_path,
        codex_json=candidate_path,
    )

    assert rows[0]["instrument_number"] == ""
    assert rows[0]["status"] == "QUARANTINED - REVIEW REQUIRED"
    archive = (ctx.run_dir / "candidates" / "candidate_archive.jsonl").read_text(
        encoding="utf-8"
    )
    assert '"FORGED-999"' in archive
    assert '"retain me"' in archive
    assert '"malformed candidate"' in archive
    assert (ctx.run_dir / "candidates" / "cursor_raw_input.json").read_bytes() == (
        candidate_path.read_bytes()
    )


def test_source_hash_change_after_inventory_blocks_ocr(tmp_path: Path):
    project = tmp_path / "project"
    project.mkdir()
    source = project / "source.txt"
    source.write_text("Original source", encoding="utf-8")
    ctx = start_run(project)
    inventory = build_inventory(ctx)
    original_hash = inventory[0]["sha256"]
    source.write_text("Changed after inventory", encoding="utf-8")

    records = run_ocr(ctx)

    assert records == []
    failures = json.loads(
        (ctx.quarantine_dir / "ocr_failures.json").read_text(encoding="utf-8")
    )
    assert "SOURCE HASH MISMATCH" in failures[0]["error"]
    assert original_hash not in {
        row.get("source_file_hash") for row in records
    }


def test_missing_documents_includes_unrepresented_source_and_clears_stale_failures(
    tmp_path: Path,
):
    project = tmp_path / "project"
    project.mkdir()
    (project / "deed.txt").write_text(_source_text(), encoding="utf-8")
    (project / "unclassified.txt").write_text("No instrument facts here.", encoding="utf-8")
    ctx = start_run(project)
    build_inventory(ctx)
    (ctx.quarantine_dir / "ocr_failures.json").write_text(
        json.dumps([{"source_path": "old.pdf", "error": "stale"}]),
        encoding="utf-8",
    )

    run_ocr(ctx)
    assert json.loads(
        (ctx.quarantine_dir / "ocr_failures.json").read_text(encoding="utf-8")
    ) == []
    extract_and_reconcile(ctx, weak_threshold=0.6)
    bundle = build_runsheet(ctx)

    assert any(
        row["citation"] == "unclassified.txt"
        and "No instrument candidate" in row["missing_or_issue"]
        for row in bundle["missing_documents"]
    )
    assert all(row["citation"] != "old.pdf" for row in bundle["missing_documents"])


def test_resume_continues_same_run_after_ocr(tmp_path: Path):
    project = tmp_path / "project"
    project.mkdir()
    (project / "deed.txt").write_text(_source_text(), encoding="utf-8")
    ctx = start_run(project)
    build_inventory(ctx)
    run_ocr(ctx)

    result = resume_pipeline(ctx, weak_threshold=0.6)

    assert result["run_id"] == ctx.run_id
    assert result["completed"] == ["RECONCILE", "RUNSHEET"]
    assert result["skipped"] == ["INVENTORY", "OCR"]
    assert latest_run(project).run_id == ctx.run_id
    assert (ctx.run_dir / "runsheet_bundle.json").is_file()


def test_multiframe_tiff_accounts_for_every_frame(tmp_path: Path, monkeypatch):
    project = tmp_path / "project"
    project.mkdir()
    first = Image.new("L", (100, 60), "white")
    second = Image.new("L", (100, 60), "black")
    source = project / "index.tiff"
    first.save(source, save_all=True, append_images=[second])

    def fake_ocr(path, config="--oem 3 --psm 6"):
        return {
            "text": f"Page evidence {Path(path).name}",
            "confidence": 0.9,
            "blocks": [],
            "ocr_config": config,
            "engine_version": "test-engine",
        }

    monkeypatch.setattr(factory_core, "_ocr_image", fake_ocr)
    ctx = start_run(project)
    inventory = build_inventory(ctx)
    assert inventory[0]["page_count"] == 2

    records = run_ocr(ctx)

    assert [record["citation"] for record in records] == [
        "index.tiff#page=1",
        "index.tiff#page=2",
    ]


def test_export_preserves_template_and_uses_separate_control_workbook(tmp_path: Path):
    project = tmp_path / "project"
    project.mkdir()
    (project / "source.txt").write_text("Source evidence", encoding="utf-8")
    template = project / "template.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Original Template"
    sheet["A1"] = "DO NOT CHANGE"
    sheet["B2"] = "=SUM(1,2)"
    collision = workbook.create_sheet("DBTF Runsheet")
    collision["A1"] = "TEMPLATE CONTENT"
    workbook.save(template)
    original_template = template.read_bytes()

    candidate = {
        "instrument_number": "2026-77",
        "instrument_type": "Oil and Gas Lease",
        "instrument_date": "01/02/2026",
        "recorded_date": "01/05/2026",
        "book": "10",
        "page": "20",
        "grantor": "=HYPERLINK(\"https://example.invalid\",\"Owner\")",
        "grantee": "Beacon LLC",
        "legal_description": "NE/4 Section 31",
        "interest_conveyed": "1/2",
        "lease_royalty_terms": "3/16",
        "confidence": 0.95,
        "citation": "source.txt#file",
        "source_path": "source.txt",
        "source_locator": "file",
    }
    cursor_json = project / "cursor.json"
    codex_json = project / "codex.json"
    cursor_json.write_text(json.dumps([candidate]), encoding="utf-8")
    codex_json.write_text(json.dumps([candidate]), encoding="utf-8")

    ctx = start_run(project)
    build_inventory(ctx)
    run_ocr(ctx)
    extract_and_reconcile(ctx, cursor_json=cursor_json, codex_json=codex_json)
    candidate_csv = (ctx.run_dir / "candidates" / "cursor_output.csv").read_text(
        encoding="utf-8-sig"
    )
    assert "'=HYPERLINK" in candidate_csv
    build_runsheet(ctx)
    exported = export_safe_xlsx(ctx, template, section="32-11N-25W")

    assert exported != template
    assert template.read_bytes() == original_template
    assert exported.read_bytes() == original_template
    result = openpyxl.load_workbook(exported, data_only=False)
    assert result["Original Template"]["A1"].value == "DO NOT CHANGE"
    assert result["Original Template"]["B2"].value == "=SUM(1,2)"
    assert result["DBTF Runsheet"]["A1"].value == "TEMPLATE CONTENT"
    assert result.sheetnames == ["Original Template", "DBTF Runsheet"]
    result.close()
    manifest = json.loads(
        (ctx.run_dir / "export_manifest.json").read_text(encoding="utf-8")
    )
    assert manifest["client_worksheets_added"] == []
    assert manifest["preservation_audit_passed"] is True
    control = Path(manifest["control_workbook_path"])
    assert control.is_file()
    control_book = openpyxl.load_workbook(control, read_only=True)
    assert {
        "Run Control", "File Manifest", "OCR Provenance", "Candidate Archive",
        "Reconciled Instruments", "Field Conflicts", "Review Queue",
        "Missing Documents", "Workbook Audit",
    }.issubset(control_book.sheetnames)
    control_book.close()
    backup = Path(manifest["backup_record"]["backup_path"])
    assert backup.read_bytes() == original_template
    readiness = json.loads(
        (exported.parent / "FINAL_READINESS_STATEMENT.json").read_text(encoding="utf-8")
    )
    assert readiness["ready_to_submit"] is False
    assert readiness["conclusion"] == "NOT READY TO SUBMIT"


def test_workbook_audit_detects_formula_mutation(tmp_path: Path):
    before_path = tmp_path / "before.xlsx"
    after_path = tmp_path / "after.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "=SUM(1,2)"
    workbook.active.merge_cells("B2:C2")
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    workbook.save(before_path)
    workbook.close()
    after_path.write_bytes(before_path.read_bytes())

    exact = compare_workbooks(audit_workbook(before_path), audit_workbook(after_path))
    assert exact["preservation_passed"] is True

    changed = openpyxl.load_workbook(after_path)
    changed.active["A1"] = "=SUM(1,3)"
    changed.save(after_path)
    changed.close()
    comparison = compare_workbooks(
        audit_workbook(before_path),
        audit_workbook(after_path),
    )
    assert comparison["preservation_passed"] is False
    assert any(item["feature"].endswith(".formulas") for item in comparison["differences"])


def test_workbook_audit_rejects_plain_cell_and_style_mutation(tmp_path: Path):
    before_path = tmp_path / "before.xlsx"
    after_path = tmp_path / "after.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "Original"
    workbook.save(before_path)
    workbook.save(after_path)
    workbook.close()
    changed = openpyxl.load_workbook(after_path)
    changed.active["A1"] = "Altered"
    changed.active["A1"].font = openpyxl.styles.Font(bold=True)
    changed.save(after_path)
    changed.close()

    comparison = compare_workbooks(
        audit_workbook(before_path),
        audit_workbook(after_path),
    )

    assert comparison["binary_identical"] is False
    assert comparison["preservation_passed"] is False
    assert any(
        item["feature"].endswith(".cells_or_styles")
        for item in comparison["differences"]
    )
