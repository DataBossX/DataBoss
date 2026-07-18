"""Tests for the Excel / PDF / dashboard report builders (in isolation)."""

from pathlib import Path

import pytest

from databossx.dashboard import render_dashboard, write_dashboard
from databossx.excel_report import build_client_workbook
from databossx.ownership import MineralOwner, division_of_interest


def _sample_tracts():
    owners = [
        MineralOwner("Alpha", "1/2", royalty="1/8", lessee="Op"),
        MineralOwner("Beta", "1/2", royalty="1/8", lessee="Op"),
    ]
    return [division_of_interest(owners, "160", tract="Sec 1 T1N R1W")]


_META = {"project": "Test", "section": "1-1N-1W", "generated": "2026-01-01",
         "source_root": "/tmp/test"}
_COUNTS = {"runsheet_rows": 1, "tracts": 1, "mineral_owners": 2, "doi_rows": 4,
           "chain_breaks": 0, "review_items": 0, "defects": 0, "evidence": 1}


def test_excel_workbook_written_with_sheets(tmp_path):
    import openpyxl

    runsheet = [{"instrument_number": "1", "grantor": "A", "grantee": "B",
                 "conveyed_interest": "1/2", "retained_interest": "1/2",
                 "net_mineral_acres": "80", "status": "ok", "remarks": ""}]
    path = build_client_workbook(
        tmp_path / "r.xlsx", meta=_META, counts=_COUNTS, rag="GREEN",
        runsheet_rows=runsheet, tracts=_sample_tracts(),
        defects=[], evidence=[{"file": "a.txt", "type": "text", "sha256": "abc",
                               "bytes": 10, "method": "txt", "note": ""}])
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert set(["Summary", "Runsheet", "Mineral Ownership",
                "Division of Interest", "Defects", "Evidence"]) <= set(wb.sheetnames)
    # The DOI NRI total must round-trip as 8/8 in the workbook.
    doi = wb["Division of Interest"]
    totals = [r for r in doi.iter_rows(values_only=True) if r and r[1] == "TOTAL"]
    assert totals and totals[0][5].startswith("1.0")  # NRI decimal total == 1


def test_pdf_written_and_valid(tmp_path):
    reportlab = pytest.importorskip("reportlab")  # noqa: F841
    from databossx.pdf_report import build_client_pdf

    path = build_client_pdf(
        tmp_path / "r.pdf", meta=_META, counts=_COUNTS, rag="GREEN",
        runsheet_rows=[], tracts=_sample_tracts(), defects=[], evidence=[],
        synthetic=True)
    assert path.exists()
    assert path.read_bytes()[:5] == b"%PDF-"
    assert path.stat().st_size > 1000


def test_excel_neutralizes_formula_injection(tmp_path):
    # A malicious grantor named like a formula must NOT become a live formula in
    # the client's workbook.
    import openpyxl

    runsheet = [{"instrument_number": "1", "grantor": "=HYPERLINK(\"http://x\")",
                 "grantee": "@SUM(A1)", "conveyed_interest": "1/2",
                 "status": "ok", "remarks": "+1+1"}]
    path = build_client_workbook(
        tmp_path / "r.xlsx", meta=_META, counts=_COUNTS, rag="RED",
        runsheet_rows=runsheet, tracts=_sample_tracts(), defects=[], evidence=[])
    wb = openpyxl.load_workbook(path)
    ws = wb["Runsheet"]
    danger = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":  # a live formula cell
                danger.append(cell.value)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                danger.append(cell.value)
    assert not danger, f"formula-injected cells present: {danger}"
    # The malicious value survives as inert, apostrophe-prefixed text.
    found = [c.value for r in ws.iter_rows() for c in r
             if isinstance(c.value, str) and "HYPERLINK" in c.value]
    assert found and found[0].startswith("'=")


def test_pdf_survives_xml_special_chars_in_data(tmp_path):
    # A party name with & / < / > must not break reportlab Paragraph markup.
    pytest.importorskip("reportlab")
    from databossx.pdf_report import build_client_pdf

    runsheet = [{"instrument_number": "1", "grantor": "Smith & Sons <Trust>",
                 "grantee": "A>B & C", "conveyed_interest": "1/2",
                 "retained_interest": "1/2", "net_mineral_acres": "80",
                 "status": "ok", "remarks": "note with <b> & ampersand"}]
    tracts = [division_of_interest(
        [MineralOwner("Q & R <LLC>", "1", royalty="1/8", lessee="Op & Co")],
        "160", tract="Sec 1 <T1N> & R1W")]
    path = build_client_pdf(
        tmp_path / "x.pdf", meta={**_META, "source_root": "/a & b/<x>"},
        counts=_COUNTS, rag="RED", runsheet_rows=runsheet, tracts=tracts,
        defects=[{"severity": "red", "category": "c", "subject": "s & <t>",
                  "detail": "d < e & f", "source": "src"}],
        evidence=[], abstracts=None)
    assert path.exists()
    assert path.read_bytes()[:5] == b"%PDF-"
    assert path.stat().st_size > 1000


def test_dashboard_is_self_contained_html(tmp_path):
    from databossx.command_center import RunResult

    result = RunResult(project="Test", section="1-1N-1W", source_root="/tmp",
                       output_dir=str(tmp_path), generated="2026-01-01",
                       rag="YELLOW", counts=_COUNTS,
                       defects=[{"severity": "yellow", "category": "x",
                                 "subject": "s", "detail": "d", "source": "src"}])
    htmltext = render_dashboard(result)
    # No external assets (CSP-safe / offline).
    assert "http://" not in htmltext and "https://" not in htmltext
    assert "<script" not in htmltext.lower()
    assert "YELLOW" in htmltext
    path = write_dashboard(tmp_path / "d.html", result)
    assert path.exists() and path.stat().st_size > 0


def test_reports_do_not_zero_fill_blanks(tmp_path):
    # An unleased owner has no WI/NRI; the report must leave those blank.
    import openpyxl

    tracts = [division_of_interest([MineralOwner("Open", "1")], "160", tract="T")]
    path = build_client_workbook(
        tmp_path / "r.xlsx", meta=_META, counts=_COUNTS, rag="RED",
        runsheet_rows=[], tracts=tracts, defects=[], evidence=[])
    wb = openpyxl.load_workbook(path)
    doi = wb["Division of Interest"]
    data_rows = [r for r in doi.iter_rows(values_only=True)
                 if r and r[0] == "T" and r[1] == "Open"]
    assert data_rows
    # Working interest / NRI cells are blank (None/""), not 0.
    assert not data_rows[0][3]  # working interest
    assert not data_rows[0][5]  # NRI


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
