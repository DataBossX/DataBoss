"""Safety-critical tests for the DataBossX baseline tooling.

Covers the prompt's required checks:
  * backup excludes secrets
  * secret scan detects fake keys
  * workbook review copy does not alter source hash
  * hyperlink export works
  * dry-run copy does not write
  * diagnostics bundle excludes .env / .auth
  * CLI starts / first-task pass runs
"""
from __future__ import annotations

from pathlib import Path

import pytest

from databossx.core import backup, diagnostics, file_guard, paths, secret_scan
from databossx.excel import (
    mock_workbook,
    review_workbook,
    workbook_fingerprint,
    workbook_inspector,
)


def test_secret_scan_detects_fake_keys(tmp_path: Path):
    env = tmp_path / ".env"
    env.write_text("OPENAI_API_KEY=sk-abcdef0123456789ABCDEF0123456789ABCDEF\n")  # pragma: allowlist secret
    result = secret_scan.scan(root=tmp_path)
    assert result.findings, "expected the fake key to be detected"
    assert any("OpenAI" in f.label or "key" in f.label.lower() for f in result.findings)


def test_secret_scan_detects_keys_in_dotenv_variants(tmp_path: Path):
    # .env.local / .env.production must be scanned too (not just exact .env).
    (tmp_path / ".env.local").write_text("OPENAI_API_KEY=sk-abcdef0123456789ABCDEF0123456789\n")  # pragma: allowlist secret
    (tmp_path / ".env.production").write_text("GROK_API_KEY=xai-abcdef0123456789ABCDEF\n")  # pragma: allowlist secret
    result = secret_scan.scan(root=tmp_path)
    found_files = {f.path for f in result.findings}
    assert ".env.local" in found_files
    assert ".env.production" in found_files


def test_secret_scan_ignores_clean_files(tmp_path: Path):
    (tmp_path / "ok.py").write_text("x = 1\nname = 'hello world'\n")
    result = secret_scan.scan(root=tmp_path)
    assert result.findings == []


def test_secret_scan_respects_allowlist_pragma(tmp_path: Path):
    # A deliberate fake key marked with the allowlist pragma must be suppressed,
    # so `scan --fail-on-tracked` stays usable as a CI gate.
    (tmp_path / "fixture.py").write_text(
        'KEY = "sk-abcdef0123456789ABCDEF0123456789"  # pragma: allowlist secret\n'
    )
    assert secret_scan.scan(root=tmp_path).findings == []

    # Without the pragma the same value is still reported. (The pragma comment
    # here suppresses the finding on *this* source line only; the string written
    # to real.py has no pragma, so the scan below still detects it.)
    (tmp_path / "real.py").write_text('KEY = "sk-abcdef0123456789ABCDEF0123456789"\n')  # pragma: allowlist secret
    assert secret_scan.scan(root=tmp_path).findings


def test_backup_excludes_secrets(tmp_path: Path, monkeypatch):
    # Build a fake project root with a secret + a normal file.
    (tmp_path / "keep.py").write_text("print('hi')\n")
    (tmp_path / ".env").write_text("OPENAI_API_KEY=sk-secret\n")
    (tmp_path / "logs").mkdir()
    (tmp_path / "logs" / "app.log").write_text("noise\n")

    # Redirect artifact dirs into tmp.
    monkeypatch.setattr(paths, "ROOT", tmp_path)
    monkeypatch.setattr(paths, "BACKUPS", tmp_path / "backups")
    monkeypatch.setattr(paths, "LOGS", tmp_path / "logs")
    monkeypatch.setattr(paths, "ALL_ARTIFACT_DIRS", (tmp_path / "backups", tmp_path / "logs"))

    result = backup.create_backup(root=tmp_path)
    import zipfile

    with zipfile.ZipFile(result.zip_path) as zf:
        names = zf.namelist()
    assert "keep.py" in names
    assert ".env" not in names
    assert not any(n.endswith(".log") for n in names)
    assert len(result.sha256) == 64


def test_review_copy_does_not_alter_source_hash(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(paths, "REVIEW_OUTPUTS", tmp_path / "review_outputs")
    monkeypatch.setattr(paths, "ALL_ARTIFACT_DIRS", (tmp_path / "review_outputs",))
    src = mock_workbook.create_mock_workbook(out_path=tmp_path / "src.xlsx")
    before = workbook_fingerprint.fingerprint(src)
    review_workbook.create_review_copy(src)
    after = workbook_fingerprint.fingerprint(src)
    assert workbook_fingerprint.unchanged(before, after)


def test_review_copy_has_ai_columns(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(paths, "REVIEW_OUTPUTS", tmp_path / "review_outputs")
    monkeypatch.setattr(paths, "ALL_ARTIFACT_DIRS", (tmp_path / "review_outputs",))
    src = mock_workbook.create_mock_workbook(out_path=tmp_path / "src.xlsx")
    copy = review_workbook.create_review_copy(src)
    from openpyxl import load_workbook

    ws = load_workbook(copy).active
    header_vals = {c.value for row in ws.iter_rows(min_row=2, max_row=2) for c in row}
    for col in review_workbook.REVIEW_COLUMNS:
        assert col in header_vals


def test_hyperlink_export_works(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(paths, "REPORTS", tmp_path / "reports")
    monkeypatch.setattr(paths, "ALL_ARTIFACT_DIRS", (tmp_path / "reports",))
    src = mock_workbook.create_mock_workbook(out_path=tmp_path / "src.xlsx")
    result = workbook_inspector.inspect(src)
    assert result.total_hyperlinks >= 5
    csv_path = workbook_inspector.export_hyperlinks_csv(result)
    rows = csv_path.read_text().strip().splitlines()
    assert rows[0] == "sheet,cell,target"
    assert len(rows) >= 6  # header + 5 links


def test_dry_run_copy_does_not_write(tmp_path: Path):
    src = tmp_path / "a.txt"
    src.write_text("hi")
    dst = tmp_path / "b.txt"
    plan = file_guard.guarded_copy(src, dst, dry_run=True)
    assert plan.performed is False
    assert not dst.exists()


def test_guard_refuses_overwrite(tmp_path: Path):
    src = tmp_path / "a.txt"
    src.write_text("hi")
    dst = tmp_path / "b.txt"
    dst.write_text("existing")
    with pytest.raises(file_guard.GuardError):
        file_guard.guarded_copy(src, dst)


def test_guard_refuses_protected_area(tmp_path: Path):
    src = tmp_path / "a.txt"
    src.write_text("hi")
    dst = tmp_path / "Horizon" / "b.txt"
    with pytest.raises(file_guard.GuardError):
        file_guard.guarded_copy(src, dst)


def test_diagnostics_excludes_env(tmp_path: Path, monkeypatch):
    reports = tmp_path / "reports"
    logs = tmp_path / "logs"
    reviews = tmp_path / "review_outputs"
    for d in (reports, logs, reviews):
        d.mkdir()
    (reports / "r.md").write_text("ok")
    (logs / ".env").write_text("OPENAI_API_KEY=sk-secret")  # should be excluded
    monkeypatch.setattr(paths, "ROOT", tmp_path)
    monkeypatch.setattr(paths, "REPORTS", reports)
    monkeypatch.setattr(paths, "LOGS", logs)
    monkeypatch.setattr(paths, "REVIEW_OUTPUTS", reviews)
    monkeypatch.setattr(paths, "DIAGNOSTICS", tmp_path / "diagnostics")
    monkeypatch.setattr(paths, "ALL_ARTIFACT_DIRS", (tmp_path / "diagnostics",))
    result = diagnostics.build()
    import zipfile

    with zipfile.ZipFile(result.zip_path) as zf:
        names = zf.namelist()
    assert not any(n.endswith(".env") for n in names)
    assert any(n.endswith("r.md") for n in names)


def test_diagnostics_excludes_review_workbooks(tmp_path: Path, monkeypatch):
    reports = tmp_path / "reports"
    logs = tmp_path / "logs"
    reviews = tmp_path / "review_outputs"
    for d in (reports, logs, reviews):
        d.mkdir()
    (reports / "r.md").write_text("ok")
    (reviews / "client_REVIEW.xlsx").write_text("PRETEND CLIENT DATA")
    monkeypatch.setattr(paths, "ROOT", tmp_path)
    monkeypatch.setattr(paths, "REPORTS", reports)
    monkeypatch.setattr(paths, "LOGS", logs)
    monkeypatch.setattr(paths, "REVIEW_OUTPUTS", reviews)
    monkeypatch.setattr(paths, "DIAGNOSTICS", tmp_path / "diagnostics")
    monkeypatch.setattr(paths, "ALL_ARTIFACT_DIRS", (tmp_path / "diagnostics",))
    result = diagnostics.build()
    import zipfile

    with zipfile.ZipFile(result.zip_path) as zf:
        names = zf.namelist()
    assert not any("review_outputs" in n or n.endswith(".xlsx") for n in names)


def test_cli_first_task_runs(tmp_path: Path, monkeypatch):
    # Point all artifact dirs into tmp so the test is hermetic.
    for attr in ("REPORTS", "LOGS", "BACKUPS", "DIAGNOSTICS", "ROLLBACK", "MEMORY", "REVIEW_OUTPUTS"):
        monkeypatch.setattr(paths, attr, tmp_path / attr.lower())
    monkeypatch.setattr(
        paths, "ALL_ARTIFACT_DIRS",
        tuple(tmp_path / a for a in ("reports", "logs", "backups", "diagnostics", "rollback", "memory", "review_outputs")),
    )
    monkeypatch.setattr(paths, "ROOT", tmp_path)
    (tmp_path / "sample.py").write_text("x=1\n")
    from databossx.ui import command_center

    artifacts = command_center.run_first_task()
    assert Path(artifacts["source_change_proof"]).exists()
    assert Path(artifacts["review_copy"]).exists()
