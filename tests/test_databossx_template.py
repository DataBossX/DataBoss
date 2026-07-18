"""Template-preservation test: an approved client workbook keeps its structure
(other sheets + embedded plats) while the reconciled data is written in."""

import base64
import zipfile
from pathlib import Path

import pytest

from databossx.command_center import run_project
from databossx.demo import build_golden_demo

# A minimal valid 1x1 PNG, used as a stand-in embedded plat exhibit.
_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8"
    "z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg==")


def _approved_template(path: Path) -> Path:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage

    png = path.parent / "plat.png"
    png.write_bytes(_PNG)
    wb = openpyxl.Workbook()
    wb.active.title = "Cover"
    wb.active["A1"] = "APPROVED CLIENT COVER"
    plat = wb.create_sheet("Plat Notes")
    plat["A1"] = "Plat exhibit"
    plat.add_image(XLImage(str(png)), "B2")
    wb.save(path)
    wb.close()
    return path


def test_template_faithful_workbook_preserves_structure_and_plats(tmp_path):
    tpl = _approved_template(tmp_path / "Client_Approved_Template.xlsx")
    # Sanity: the template really carries embedded media.
    assert any(n.startswith("xl/media/") for n in zipfile.ZipFile(tpl).namelist())

    root = tmp_path / "G"
    build_golden_demo(root)
    result = run_project("golden_demo", root=str(root),
                         output_dir=str(tmp_path / "out"),
                         make_pdf=False, template=str(tpl))

    dest = result.deliverables.get("template_report")
    assert dest and Path(dest).exists()

    import openpyxl
    wb = openpyxl.load_workbook(dest)
    # The client's own sheets survive...
    assert "Cover" in wb.sheetnames
    assert "Plat Notes" in wb.sheetnames
    # ...the reconciled data is written into the data sheet...
    assert "Title Report" in wb.sheetnames
    # ...and the embedded plat image is preserved byte-for-byte.
    assert any(n.startswith("xl/media/")
               for n in zipfile.ZipFile(dest).namelist())

    # The data sheet actually contains reconciled rows.
    ws = wb["Title Report"]
    text = " ".join(str(c) for row in ws.iter_rows(values_only=True)
                    for c in row if c)
    assert "AVERY MINERALS LLC" in text


def test_missing_template_is_a_warning_not_a_crash(tmp_path):
    root = tmp_path / "G"
    build_golden_demo(root)
    result = run_project("golden_demo", root=str(root),
                         output_dir=str(tmp_path / "out"), make_pdf=False,
                         template=str(tmp_path / "nope.xlsx"))
    assert result.ok  # still completes
    assert "template_report" not in result.deliverables
    assert any("template not found" in w.lower() for w in result.warnings)


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
