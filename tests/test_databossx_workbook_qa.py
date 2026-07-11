"""Tests for the examiner-level workbook QA engine."""

from __future__ import annotations

import openpyxl
import pytest

from databossx.workbook_qa import DIVERSIFIED_32_SPEC, validate_workbook

SMALL_SPEC = {
    "name": "test-spec",
    "expected_sheets": ["Overview", "Runsheet", "Tract 1"],
    "max_defined_names": 0,
    "expected_formula_cells": 1,
    "runsheet_sheet": "Runsheet",
    "min_runsheet_hyperlinks": 1,
    "contamination_tokens": ["Roger Mills"],
    "required_tokens": ["2371"],
    "evidence_only_sheets": ["Tract 1"],
}


def build_workbook(path, contaminate=False, fabricate=False, drop_token=False):
    wb = openpyxl.Workbook()
    wb.active.title = "Overview"
    rs = wb.create_sheet("Runsheet")
    tract = wb.create_sheet("Tract 1")
    wb["Overview"]["A1"] = "=SUM(B1:B2)"
    rs["A1"] = "Instrument"
    rs["A2"] = "stale Roger Mills entry" if contaminate else "clean entry"
    rs["A3"] = "" if drop_token else "Bk 2371/P533-553"
    rs["B2"] = "link"
    rs["B2"].hyperlink = "https://example.test/detail/1"
    tract["A1"] = "evidence matrix"
    if fabricate:
        tract["B2"] = 0.125  # an unsupported ownership decimal
    wb.save(path)
    return path


def test_clean_workbook_passes(tmp_path):
    p = build_workbook(tmp_path / "clean.xlsx")
    report = validate_workbook(p, SMALL_SPEC)
    assert report["verdict"] == "PASS", report["failed_checks"]


def test_contamination_fails(tmp_path):
    p = build_workbook(tmp_path / "dirty.xlsx", contaminate=True)
    report = validate_workbook(p, SMALL_SPEC)
    assert "contamination_scan" in report["failed_checks"]


def test_fabricated_decimal_fails(tmp_path):
    p = build_workbook(tmp_path / "fab.xlsx", fabricate=True)
    report = validate_workbook(p, SMALL_SPEC)
    assert "fabricated_decimal_guard" in report["failed_checks"]


def test_missing_required_token_fails(tmp_path):
    p = build_workbook(tmp_path / "missing.xlsx", drop_token=True)
    report = validate_workbook(p, SMALL_SPEC)
    assert "required_tokens" in report["failed_checks"]


def test_wrong_sheet_order_fails(tmp_path):
    p = build_workbook(tmp_path / "order.xlsx")
    wb = openpyxl.load_workbook(p)
    wb.move_sheet("Tract 1", offset=-1)
    wb.save(p)
    report = validate_workbook(p, SMALL_SPEC)
    assert "sheet_names_order" in report["failed_checks"]


def test_error_literal_fails(tmp_path):
    p = build_workbook(tmp_path / "err.xlsx")
    wb = openpyxl.load_workbook(p)
    wb["Overview"]["C1"] = "#REF!"
    wb.save(p)
    report = validate_workbook(p, SMALL_SPEC)
    assert "formula_error_cells" in report["failed_checks"]


def test_certified_spec_shape():
    """The certified Section 32 spec matches the audit's verified facts."""
    assert len(DIVERSIFIED_32_SPEC["expected_sheets"]) == 13
    assert DIVERSIFIED_32_SPEC["expected_formula_cells"] == 12
    assert DIVERSIFIED_32_SPEC["min_runsheet_hyperlinks"] == 34
    assert DIVERSIFIED_32_SPEC["max_defined_names"] == 0
