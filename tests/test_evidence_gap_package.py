import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.build_evidence_gap_package import (
    OUTPUT_NAMES,
    RELEASE_STATUS,
    UNABLE_TO_VERIFY,
    build,
)

GENERATED = "2026-07-15T00:00:00+00:00"
METADATA_NAMES = ("SHA256SUMS.txt", "INVOCATION_MANIFEST.json")


class EvidenceGapPackageTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporary_directory.cleanup)
        self.output = Path(self.temporary_directory.name) / "package"
        self.project = {
            "client": "Diversified",
            "legal": "Section 32, T11N-R25W",
            "county": "Beckham County, Oklahoma",
            "source_note": "Official API returned HTTP 401; no credential configured.",
        }

    def test_builds_complete_validated_package(self) -> None:
        build(self.output, self.project, GENERATED)

        for name in (*OUTPUT_NAMES, *METADATA_NAMES):
            self.assertTrue((self.output / name).is_file(), name)

        manifest = json.loads(
            (self.output / "INVOCATION_MANIFEST.json").read_text(encoding="utf-8")
        )
        self.assertEqual(RELEASE_STATUS, "BLOCKED — DO NOT RELEASE")
        self.assertEqual(
            UNABLE_TO_VERIFY,
            "Unable to verify from official county records.",
        )
        self.assertEqual(manifest["release_status"], RELEASE_STATUS)
        self.assertEqual(manifest["official_verification"], UNABLE_TO_VERIFY)
        self.assertEqual(manifest["input_counts"]["verified_instruments"], 0)
        self.assertEqual(set(manifest["outputs"]), set(OUTPUT_NAMES))

        workbook = load_workbook(self.output / "VERIFIED_MASTER_WORKBOOK.xlsx")
        self.addCleanup(workbook.close)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Master Verification",
                "Error Report",
                "Missing Documents",
                "Title Gaps",
                "Diversified Chain",
                "Source Citations",
                "QA Summary",
            ],
        )
        verification = workbook["Master Verification"]
        self.assertEqual(verification["V2"].value, "NOT FOUND")
        self.assertIn(UNABLE_TO_VERIFY, verification["Y2"].value)

    def test_escapes_formula_like_client_input(self) -> None:
        project = dict(self.project, client="=HYPERLINK(\"https://example.test\")")
        build(self.output, project, GENERATED)

        workbook = load_workbook(self.output / "DIVERSIFIED_OWNERSHIP_CHAIN.xlsx")
        self.addCleanup(workbook.close)
        destination = workbook["Diversified Chain"]["C2"]
        self.assertEqual(destination.data_type, "s")
        self.assertTrue(destination.value.startswith("'="))

    def test_refuses_repository_public_output(self) -> None:
        repository_output = (
            Path(__file__).resolve().parents[1] / "public-title-package-forbidden"
        )
        with self.assertRaisesRegex(ValueError, "ignored private directory"):
            build(repository_output, self.project, GENERATED)

    def test_refuses_to_overwrite_package(self) -> None:
        build(self.output, self.project, GENERATED)
        with self.assertRaises(FileExistsError):
            build(self.output, self.project, GENERATED)


if __name__ == "__main__":
    unittest.main()
