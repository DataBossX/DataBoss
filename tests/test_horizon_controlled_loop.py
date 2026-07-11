"""Tests for hash-bound, deterministic workbook QA and repair runs."""

import hashlib
import json
import zipfile
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
import pytest
from lxml import etree

from horizon.controlled_loop import ControlledWorkbookLoop
from horizon.project_manifest import (
    ControlFileError,
    load_project_manifest,
    load_work_order,
)
from horizon.workbook_qa import inspect_workbook


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _set_cached_value(path: Path, cell_reference: str, value: float) -> None:
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    temporary = path.with_suffix(".cached.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temporary, "w", zipfile.ZIP_DEFLATED
    ) as output:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                root = etree.fromstring(data)
                cell = root.find(f".//{{{namespace}}}c[@r={cell_reference!r}]")
                cached = cell.find(f"{{{namespace}}}v")
                cached.text = str(value)
                data = etree.tostring(
                    root,
                    xml_declaration=True,
                    encoding="UTF-8",
                    standalone=True,
                )
            output.writestr(item, data)
    temporary.replace(path)


def _make_workbook(
    path: Path, formula: str = "=SUM(B2:B3)", *, cache_formula: bool = True
) -> None:
    workbook = openpyxl.Workbook()
    title = workbook.active
    title.title = "Title"
    title.append(["Owner", "Current NMA"])
    title.append(["Alpha LLC", 20])
    title.append(["Beta LLC", 20])
    title["B4"] = formula
    title["D2"] = 40
    workbook.create_sheet("Runsheet")
    workbook.save(path)
    workbook.close()
    if cache_formula:
        _set_cached_value(path, "B4", 40)


def _write_controls(
    root: Path,
    candidate: Path,
    template: Path,
    *,
    candidate_hash: str = "",
    allowed_repairs: Optional[List[str]] = None,
) -> Tuple[Path, Path, Path]:
    project = root / "project"
    project.mkdir()
    profile = project / "workbook_profile.json"
    profile.write_text(
        json.dumps(
            {
                "required_sheet_order": ["Title", "Runsheet"],
                "preserve_sheet_order": True,
                "allow_new_sheets": False,
                "current_ownership_ranges": [
                    {"sheet": "Title", "range": "B2:B3"}
                ],
                "current_owner_columns": [
                    {
                        "sheet": "Title",
                        "column": "A",
                        "start_row": 2,
                        "end_row": 3,
                    }
                ],
                "total_assertions": [
                    {
                        "check_id": "ownership_totals",
                        "sheet": "Title",
                        "cell": "D2",
                        "expected": 40,
                        "tolerance": 0.01,
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    reported_hash = candidate_hash or _sha256(candidate)
    manifest_path = project / "project_manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "schema_id": "dbx.project_manifest",
                "schema_version": "1.0",
                "project_id": "DBX-TEST",
                "source_policy": "IMMUTABLE_READ_ONLY",
                "candidate_deliverables": [
                    {
                        "path": "candidate.xlsx",
                        "reported_sha256": reported_hash,
                        "status": "CANDIDATE",
                    }
                ],
                "required_checks": [
                    "negative_current_ownership",
                    "ownership_totals",
                    "no_duplicate_owners",
                    "no_broken_formulas",
                    "template_compliance",
                    "human_landman_release",
                ],
                "release_policy": {
                    "technical_verification_is_not_release": True,
                    "approved_hash_required": True,
                    "human_gate": "G7",
                },
            }
        ),
        encoding="utf-8",
    )
    work_order_path = project / "work_order.json"
    work_order_path.write_text(
        json.dumps(
            {
                "schema_id": "dbx.work_order",
                "schema_version": "1.0",
                "work_order_id": "WO-TEST-001",
                "project_id": "DBX-TEST",
                "objective": "Repair and verify one workbook without releasing it",
                "candidate_path": "candidate.xlsx",
                "candidate_local_path": str(candidate),
                "template_path": str(template),
                "template_expected_sha256": _sha256(template),
                "profile_path": str(profile),
                "profile_expected_sha256": _sha256(profile),
                "staging_root": str(root / "runs"),
                "acceptance_tests": [
                    "negative_current_ownership",
                    "ownership_totals",
                    "no_duplicate_owners",
                    "no_broken_formulas",
                    "template_compliance",
                    "human_landman_release",
                ],
                "allowed_repairs": allowed_repairs or [],
                "constraints": {"edit_originals": False},
                "retry_policy": {
                    "max_attempts_per_defect": 3,
                    "stop_on_regression": True,
                },
                "promotion": {"require_human_approval": True},
            }
        ),
        encoding="utf-8",
    )
    return manifest_path, work_order_path, profile


def test_control_files_bind_work_order_to_manifest_candidate(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate)
    _make_workbook(template)
    manifest_path, work_order_path, _ = _write_controls(
        tmp_path, candidate, template
    )

    manifest = load_project_manifest(manifest_path)
    order = load_work_order(work_order_path, manifest)

    assert order.project_id == manifest.project_id
    assert order.expected_sha256 == _sha256(candidate)
    assert order.require_human_approval


def test_work_order_cannot_disable_human_approval(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate)
    _make_workbook(template)
    manifest_path, work_order_path, _ = _write_controls(
        tmp_path, candidate, template
    )
    data = json.loads(work_order_path.read_text(encoding="utf-8"))
    data["promotion"]["require_human_approval"] = False
    work_order_path.write_text(json.dumps(data), encoding="utf-8")

    manifest = load_project_manifest(manifest_path)
    with pytest.raises(
        ControlFileError, match="human approval cannot be disabled"
    ):
        load_work_order(work_order_path, manifest)


def test_work_order_cannot_omit_manifest_checks(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate)
    _make_workbook(template)
    manifest_path, work_order_path, _ = _write_controls(
        tmp_path, candidate, template
    )
    data = json.loads(work_order_path.read_text(encoding="utf-8"))
    data["acceptance_tests"].remove("ownership_totals")
    work_order_path.write_text(json.dumps(data), encoding="utf-8")

    manifest = load_project_manifest(manifest_path)
    with pytest.raises(ControlFileError, match="must exactly match"):
        load_work_order(work_order_path, manifest)


def test_deterministic_qa_reports_rule_based_failures(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate, formula="=#REF!")
    _make_workbook(template)
    manifest_path, work_order_path, profile_path = _write_controls(
        tmp_path, candidate, template
    )
    manifest = load_project_manifest(manifest_path)
    order = load_work_order(work_order_path, manifest)
    profile = json.loads(profile_path.read_text(encoding="utf-8"))

    report = inspect_workbook(
        candidate,
        order.acceptance_tests,
        template_path=template,
        profile=profile,
    )

    formula_check = next(
        check for check in report.checks if check.check_id == "no_broken_formulas"
    )
    assert not formula_check.passed
    assert formula_check.findings[0].cell == "B4"
    assert formula_check.findings[0].repairable
    assert report.score.mathematical_accuracy == 100
    assert not report.score.technical_pass
    assert not report.score.promotion_ready


def test_uncalculated_formula_is_blocking(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate, cache_formula=False)
    _make_workbook(template)
    manifest_path, work_order_path, profile_path = _write_controls(
        tmp_path, candidate, template
    )
    manifest = load_project_manifest(manifest_path)
    order = load_work_order(work_order_path, manifest)
    profile = json.loads(profile_path.read_text(encoding="utf-8"))

    report = inspect_workbook(
        candidate,
        order.acceptance_tests,
        template_path=template,
        profile=profile,
    )

    formula_check = next(
        check for check in report.checks if check.check_id == "no_broken_formulas"
    )
    assert formula_check.status == "failed"
    assert formula_check.findings[0].code == "formula_not_calculated"
    assert not report.score.technical_pass


def test_template_formula_replaced_by_constant_is_blocking(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate)
    _make_workbook(template)
    workbook = openpyxl.load_workbook(candidate)
    workbook["Title"]["B4"] = 40
    workbook.save(candidate)
    workbook.close()
    manifest_path, work_order_path, profile_path = _write_controls(
        tmp_path, candidate, template
    )
    manifest = load_project_manifest(manifest_path)
    order = load_work_order(work_order_path, manifest)
    profile = json.loads(profile_path.read_text(encoding="utf-8"))

    report = inspect_workbook(
        candidate,
        order.acceptance_tests,
        template_path=template,
        profile=profile,
    )

    formula_check = next(
        check for check in report.checks if check.check_id == "no_broken_formulas"
    )
    assert formula_check.findings[0].code == "authoritative_formula_mismatch"
    assert not report.score.technical_pass


def test_nan_total_cannot_pass_assertion(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate)
    _make_workbook(template)
    workbook = openpyxl.load_workbook(candidate)
    workbook["Title"]["D2"] = "NaN"
    workbook.save(candidate)
    workbook.close()
    _set_cached_value(candidate, "B4", 40)
    manifest_path, work_order_path, profile_path = _write_controls(
        tmp_path, candidate, template
    )
    manifest = load_project_manifest(manifest_path)
    order = load_work_order(work_order_path, manifest)
    profile = json.loads(profile_path.read_text(encoding="utf-8"))

    report = inspect_workbook(
        candidate,
        order.acceptance_tests,
        template_path=template,
        profile=profile,
    )

    totals = next(
        check for check in report.checks if check.check_id == "ownership_totals"
    )
    assert totals.status == "failed"
    assert totals.findings[0].code == "total_out_of_tolerance"


def test_loop_repairs_staging_only_then_requires_recalculation(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate, formula="=#REF!")
    _make_workbook(template)
    original_bytes = candidate.read_bytes()
    manifest_path, work_order_path, _ = _write_controls(
        tmp_path,
        candidate,
        template,
        allowed_repairs=["restore_formula_from_template"],
    )

    result = ControlledWorkbookLoop.from_files(
        manifest_path, work_order_path
    ).run()

    assert result.status == "blocked"
    assert result.attempts == 1
    assert candidate.read_bytes() == original_bytes
    assert result.staged_workbook is not None
    repaired = openpyxl.load_workbook(result.staged_workbook, data_only=False)
    assert repaired["Title"]["B4"].value == "=SUM(B2:B3)"
    repaired.close()

    receipt = json.loads(result.receipt_path.read_text(encoding="utf-8"))
    assert receipt["input"]["actual_sha256"] == _sha256(candidate)
    assert receipt["output"]["sha256"] == _sha256(result.staged_workbook)
    assert receipt["human_approval_required"] is True
    assert receipt["promotion_executed"] is False
    assert result.promotion_package is None
    assert receipt["authorities"]["template"]["sha256"] == _sha256(template)
    assert receipt["authorities"]["profile"]["sha256"] == _sha256(
        tmp_path / "project" / "workbook_profile.json"
    )


def test_clean_loop_stops_at_hash_bound_human_gate(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate)
    _make_workbook(template)
    manifest_path, work_order_path, _ = _write_controls(
        tmp_path, candidate, template
    )

    result = ControlledWorkbookLoop.from_files(
        manifest_path, work_order_path
    ).run()

    assert result.status == "technically_verified_pending_approval"
    assert result.promotion_package is not None
    receipt = json.loads(result.receipt_path.read_text(encoding="utf-8"))
    package = json.loads(result.promotion_package.read_text(encoding="utf-8"))
    assert package["status"] == "PENDING_HUMAN_APPROVAL"
    assert package["approval_must_name_sha256"] == receipt["output"]["sha256"]
    assert package["promotion_executed"] is False


def test_hash_mismatch_fails_before_copy_and_writes_receipt(tmp_path):
    candidate = tmp_path / "candidate.xlsx"
    template = tmp_path / "template.xlsx"
    _make_workbook(candidate)
    _make_workbook(template)
    manifest_path, work_order_path, _ = _write_controls(
        tmp_path,
        candidate,
        template,
        candidate_hash="0" * 64,
    )

    result = ControlledWorkbookLoop.from_files(
        manifest_path, work_order_path
    ).run()

    assert result.status == "failed"
    assert result.staged_workbook is None
    receipt = json.loads(result.receipt_path.read_text(encoding="utf-8"))
    assert receipt["error_type"] == "ControlFileError"
    assert "hash mismatch" in receipt["error"].lower()
    assert receipt["promotion_executed"] is False
