"""Tests for the operationalized intelligence layer (pipeline.py)."""

from pathlib import Path

import pytest

from horizon.chaining import OGLRecord, RunsheetNote
from horizon.models import REVIEW_TAG
from horizon.pipeline import (
    build_from_workbook,
    chain_to_report,
    find_reference_workbook,
    read_ogl_records,
    read_runsheet_notes,
    score_reference_workbook,
)


def test_chain_to_report_balanced_tract():
    ogl = [
        OGLRecord(instrument_number="1", grantor="ROOT", grantee="X",
                  conveyed_interest="1/2", legal_description="NE/4 Sec 31"),
        OGLRecord(instrument_number="2", grantor="X", grantee="Y",
                  conveyed_interest="1/4", legal_description="NE/4 Sec 31"),
    ]
    notes = [RunsheetNote(instrument_number="1", note="warranty deed"),
             RunsheetNote(instrument_number="2", note="mineral deed")]
    build = chain_to_report(ogl, notes, section="31-12N-24W")
    assert len(build.report.rows) == 2
    assert build.chain_breaks == []
    # first row: grantor held 8/8, conveyed 1/2, retained 1/2
    assert build.report.rows[0].retained_interest == "1/2"
    assert build.report.rows[0].status == "ok"
    # runsheet note carried into the verification trail
    assert "warranty deed" in build.report.rows[0].remarks


def test_chain_to_report_flags_chain_break():
    ogl = [OGLRecord(instrument_number="100", grantor="A", grantee="B",
                     conveyed_interest="1/2", legal_description="Sec 31")]
    notes = [RunsheetNote(instrument_number="999", note="orphan")]  # no OGL match
    build = chain_to_report(ogl, notes)
    # instrument 100 has no runsheet note -> unreferenced_ogl break
    assert any(c.key == "100" for c in build.chain_breaks)
    assert REVIEW_TAG in build.report.rows[0].status
    assert "Chain break" in build.report.rows[0].remarks


def test_chain_to_report_over_conveyance_review():
    ogl = [
        OGLRecord(instrument_number="1", grantor="ROOT", grantee="X",
                  conveyed_interest="1/2", legal_description="Sec 31"),
        OGLRecord(instrument_number="2", grantor="X", grantee="Y",
                  conveyed_interest="3/4", legal_description="Sec 31"),  # >held
    ]
    notes = [RunsheetNote(instrument_number="1"), RunsheetNote(instrument_number="2")]
    build = chain_to_report(ogl, notes)
    assert "Sec31" in build.tracts_reviewed or any(build.tracts_reviewed)
    assert build.report.rows[1].status == REVIEW_TAG
    assert "over-conveyance" in build.report.rows[1].remarks.lower()


def _write_reference_workbook(path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ogl = wb.active
    ogl.title = "OGL"
    ogl.append(["Instrument", "Grantor", "Grantee", "Conveyed", "Legal"])
    ogl.append(["2019-1", "ROOT", "X", "1/2", "NE/4 Sec 31"])
    ogl.append(["2019-2", "X", "Y", "1/4", "NE/4 Sec 31"])
    run = wb.create_sheet("Runsheet")
    run.append(["Instrument", "Note"])
    run.append(["2019-1", "warranty deed"])
    run.append(["2019-2", "mineral deed"])
    wb.save(path)
    wb.close()


def test_read_reference_sheets(tmp_path):
    wb = tmp_path / "ref.xlsx"
    _write_reference_workbook(wb)
    ogl = read_ogl_records(wb)
    notes = read_runsheet_notes(wb)
    assert len(ogl) == 2
    assert ogl[0].grantor == "ROOT"
    assert ogl[0].conveyed_interest == "1/2"
    assert len(notes) == 2
    assert notes[0].note == "warranty deed"


def test_build_from_workbook_end_to_end(tmp_path):
    wb = tmp_path / "ref.xlsx"
    _write_reference_workbook(wb)
    build = build_from_workbook(wb, section="31-12N-24W")
    assert len(build.report.rows) == 2
    assert build.chain_breaks == []          # both instruments matched
    assert build.report.rows[0].retained_interest == "1/2"
    assert build.report.rows[1].retained_interest == "1/4"


def test_score_and_find_reference_workbook(tmp_path):
    ref = tmp_path / "31-12N-24W Roger Mills Cursory Title Report NHE.xlsx"
    _write_reference_workbook(ref)
    # a decoy workbook with no OGL/runsheet sheets
    import openpyxl
    decoy = tmp_path / "random.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(decoy)
    wb.close()

    assert score_reference_workbook(ref) > score_reference_workbook(decoy)
    picked = find_reference_workbook([decoy, ref])
    assert picked == ref


def test_find_reference_workbook_none_when_no_candidates(tmp_path):
    import openpyxl
    decoy = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.save(decoy)
    wb.close()
    assert find_reference_workbook([decoy]) is None
