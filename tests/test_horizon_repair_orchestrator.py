"""End-to-end tests for lxml repair, report I/O, and the improvement loop."""

import zipfile
from pathlib import Path

import pytest

from horizon.audit import AuditLog
from horizon.config import HorizonConfig
from horizon.models import ReportModel, TitleRow
from horizon.orchestrator import Orchestrator
from horizon.repair import repair_workbook
from horizon.report_io import read_report, write_report
from horizon.repair import _HAVE_LXML
from horizon.validation import Requirements


def _make_xlsx_with_error_formula(path: Path):
    """Build a minimal .xlsx containing a cell with an errored formula and a
    media part, so we can prove repair fixes the formula but preserves media."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "ok"
    ws["A2"] = "=#REF!"          # broken formula
    wb.save(path)
    # inject a fake media part to prove it survives repair untouched
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as zin, \
            zipfile.ZipFile(tmp, "w") as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        zout.writestr("xl/media/plat1.png", b"\x89PNG\r\n\x1a\nFAKEPLATDATA")
    tmp.replace(path)


@pytest.mark.skipif(not _HAVE_LXML, reason="lxml required for XML repair")
def test_repair_removes_error_formula_and_preserves_media(tmp_path):
    src = tmp_path / "report.xlsx"
    _make_xlsx_with_error_formula(src)
    dest = tmp_path / "report_v002.xlsx"
    result = repair_workbook(src, dest)
    assert result.repaired
    assert result.media_preserved == 1
    assert any("errored formula" in f for f in result.fixes)
    # media survived byte-for-byte
    with zipfile.ZipFile(dest) as zf:
        assert zf.read("xl/media/plat1.png") == b"\x89PNG\r\n\x1a\nFAKEPLATDATA"
    # source workbook was never modified
    assert src.exists()


def test_report_io_roundtrip(tmp_path):
    report = ReportModel(section="31-12N-24W", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="2019-001",
                 conveyed_interest="1/2", legal_description="NE/4 Sec 31"),
    ])
    out = tmp_path / "rpt_v001.xlsx"
    write_report(report, out)
    loaded = read_report(out, section="31-12N-24W")
    assert len(loaded.rows) == 1
    assert loaded.rows[0].grantor == "A"
    assert loaded.rows[0].instrument_number == "2019-001"


@pytest.fixture()
def cfg_audit(tmp_path):
    cfg = HorizonConfig(root=tmp_path / "Horizon")
    cfg.ensure_workspace()
    return cfg, AuditLog(path=None, echo=False)


def test_orchestrator_converges_on_clean_report(cfg_audit):
    cfg, audit = cfg_audit
    base = "31-12N-24W_report"
    # seed a v001 so ingest has something to read
    seed = ReportModel(section="31-12N-24W", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="100",
                 conveyed_interest="1/2", retained_interest="1/2"),
    ])
    write_report(seed, cfg.final_reports / f"{base}_v001.xlsx")

    orch = Orchestrator(cfg, audit)
    result = orch.run(seed, base, Requirements())
    assert result.converged
    assert result.loops_run == 1
    assert result.final_version.name == f"{base}_v002.xlsx"
    assert result.final_version.exists()


def test_orchestrator_escalates_when_unrepairable(cfg_audit):
    cfg, audit = cfg_audit
    base = "31-12N-24W_report"
    # A required instrument that is absent -> validation error with no workbook
    # to repair (no seed version) -> escalate and stop.
    report = ReportModel(section="31-12N-24W", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="100"),
    ])
    reqs = Requirements(required_instruments={"999"})
    orch = Orchestrator(cfg, audit)
    result = orch.run(report, base, reqs)
    assert not result.converged
    assert result.exhausted
    assert audit.count("ESCALATED") >= 1


def test_orchestrator_respects_max_loops(cfg_audit):
    cfg, audit = cfg_audit
    cfg.max_loops = 3
    base = "31-12N-24W_report"
    # seed a workbook so repair() has a source, but repair makes no fixes ->
    # repaired_path stays None -> loop escalates after first iteration.
    seed = ReportModel(section="31-12N-24W", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="100"),
    ])
    write_report(seed, cfg.final_reports / f"{base}_v001.xlsx")
    reqs = Requirements(required_instruments={"999"})
    orch = Orchestrator(cfg, audit)
    result = orch.run(seed, base, reqs)
    assert result.loops_run <= cfg.max_loops
    assert not result.converged
