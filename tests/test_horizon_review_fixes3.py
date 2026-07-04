"""Regression tests for the third round of PR#18 review fixes."""

from fractions import Fraction
from pathlib import Path

import openpyxl

from horizon.chaining import OGLRecord, RunsheetNote, reconcile_chain
from horizon.config import HorizonConfig
from horizon.main import main
from horizon.models import REVIEW_TAG, ReportModel, TitleRow
from horizon.pipeline import (
    _is_ogl_sheet,
    _is_runsheet_sheet,
    chain_to_report,
    read_runsheet_notes,
    score_reference_workbook,
)
from horizon.report_io import write_report
from horizon.validation import Requirements, load_requirements


# -- #1 ambiguous single tab not scored as two registers -------------------
def test_ambiguous_tab_not_double_counted(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "OGL Notes"   # matches both hint sets
    wb.active.append(["Instrument", "Grantor"])
    p = tmp_path / "amb.xlsx"
    wb.save(p)
    wb.close()
    assert not _is_ogl_sheet("OGL Notes")
    assert not _is_runsheet_sheet("OGL Notes")
    # neither register unambiguously present -> no runsheet credit
    assert score_reference_workbook(p) < 6


# -- #3 runsheet reader never reuses the OGL tab ---------------------------
def test_runsheet_reader_returns_empty_without_runsheet_tab(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OGL"
    ws.append(["Instrument", "Grantor", "Grantee"])
    ws.append(["100", "A", "B"])
    p = tmp_path / "ogl_only.xlsx"
    wb.save(p)
    wb.close()
    assert read_runsheet_notes(p) == []   # no fabricated notes from the OGL tab


# -- #4 integrity break poisons the downstream ledger ----------------------
def test_continuity_break_poisons_downstream():
    records = [
        OGLRecord(instrument_number="1", grantor="ROOT", grantee="X", conveyed_interest="1/2"),
        # grantor Z != prior grantee X -> continuity break (balanced math)
        OGLRecord(instrument_number="2", grantor="Z", grantee="Y", conveyed_interest="1/4"),
        OGLRecord(instrument_number="3", grantor="Y", grantee="W", conveyed_interest="1/8"),
    ]
    result = reconcile_chain("T", records, starting_interest=Fraction(1))
    # link 2 broke continuity -> holder poisoned -> link 3 can't reconcile
    assert result.links[1].holder_after is None
    assert result.links[2].reconciliation.status == "examiner_review"


# -- #2 instrument spanning multiple tracts is flagged ---------------------
def test_cross_tract_instrument_flagged():
    ogl = [
        OGLRecord(instrument_number="1", grantor="A", grantee="B",
                  conveyed_interest="1/2", legal_description="NE/4 Sec 31"),
        OGLRecord(instrument_number="1", grantor="A", grantee="B",
                  conveyed_interest="1/2", legal_description="SW/4 Sec 31"),
    ]
    build = chain_to_report(ogl, [RunsheetNote(instrument_number="1")])
    flagged = [r for r in build.report.rows if "spans multiple tracts" in r.remarks]
    assert flagged and all(r.status == REVIEW_TAG for r in flagged)


# -- #5 Golden Source read failure is surfaced -----------------------------
def test_golden_source_read_failure_surfaced(tmp_path):
    bad = tmp_path / "project_notes_updated.xlsx"
    bad.write_bytes(b"this is not a valid xlsx")   # exists but unreadable
    reqs = load_requirements(bad)
    assert "ERROR reading Golden Source" in reqs.source


# -- #6 bare Instrument column of non-numeric types not loaded -------------
def test_non_numeric_instrument_values_not_required(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Instrument"])          # bare header
    ws.append(["Warranty Deed"])       # a *type*, not a number
    ws.append(["Mineral Deed"])
    gs = tmp_path / "project_notes_updated.xlsx"
    wb.save(gs)
    wb.close()
    reqs = load_requirements(gs)
    assert reqs.required_instruments == set()   # nothing numeric -> no gates


# -- #7 review artifacts are versioned -------------------------------------
def test_review_artifacts_versioned(tmp_path):
    from horizon.artifacts import write_all
    build = chain_to_report(
        [OGLRecord(instrument_number="1", grantor="A", grantee="B", conveyed_interest="1/2")],
        [RunsheetNote(instrument_number="1")])
    first = dict(write_all(build, tmp_path))
    second = dict(write_all(build, tmp_path))
    names = set(first) | set(second)
    assert any("_v001.csv" in n for n in names)
    assert any("_v002.csv" in n for n in names)   # second run did not overwrite


# -- #8 dry-run returns non-zero on validation failure ---------------------
def test_dry_run_nonzero_on_validation_failure(tmp_path):
    root = tmp_path / "Horizon"
    root.mkdir()
    cfg = HorizonConfig(root=root)
    cfg.ensure_workspace()
    base = "31-12N-24W_Roger_Mills_Cursory_Title_Report"
    write_report(ReportModel(section="31-12N-24W", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="100")]),
        cfg.final_reports / f"{base}_v001.xlsx")
    # Golden Source requiring an instrument that is absent -> validation error
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Instrument Number"])
    ws.append(["999"])
    wb.save(root / "project_notes_updated.xlsx")
    wb.close()
    rc = main(["--root", str(root), "--dry-run"])
    assert rc == 1
