"""Regression tests for the fifth round of PR#18 review fixes."""

import io
import zipfile
from pathlib import Path

import openpyxl
import pytest

from horizon.audit import AuditLog
from horizon.chaining import OGLRecord, RunsheetNote
from horizon.config import HorizonConfig
from horizon.models import REVIEW_TAG, ReportModel, TitleRow
from horizon.orchestrator import Orchestrator
from horizon.pipeline import chain_to_report, find_reference_workbook, score_reference_workbook
from horizon.report_io import read_report, write_report
from horizon.validation import Requirements

_HAS_PIL = False
try:
    from PIL import Image as _PILImage  # noqa: F401
    _HAS_PIL = True
except ImportError:
    pass


# -- #1 a workbook with no OGL sheet scores 0 despite a matching filename ----
def test_no_ogl_sheet_scores_zero(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Summary"          # not an OGL register
    p = tmp_path / "Roger Mills Cursory NHE.xlsx"   # tempting filename
    wb.save(p)
    wb.close()
    assert score_reference_workbook(p) == 0
    assert find_reference_workbook([p]) is None


# -- #4 rows with no legal description are flagged --------------------------
def test_missing_legal_flagged():
    ogl = [OGLRecord(instrument_number="1", grantor="A", grantee="B",
                     conveyed_interest="1/2", legal_description="")]
    build = chain_to_report(ogl, [RunsheetNote(instrument_number="1")])
    row = build.report.rows[0]
    assert "No legal description" in row.remarks
    assert row.status == REVIEW_TAG


# -- #3 emit preserves embedded plats/media --------------------------------
@pytest.mark.skipif(not _HAS_PIL, reason="Pillow needed to embed a test image")
def test_write_report_preserves_media(tmp_path):
    # build a seed report workbook that carries an embedded plat image
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    seed = tmp_path / "seed.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Title Report"
    plats = wb.create_sheet("Plats")
    b = io.BytesIO(); PILImage.new("RGB", (8, 8), (0, 128, 255)).save(b, format="PNG"); b.seek(0)
    plats.add_image(XLImage(b), "A1")
    wb.save(seed)
    wb.close()
    assert any("media" in n for n in zipfile.ZipFile(seed).namelist())

    out = tmp_path / "out.xlsx"
    report = ReportModel(section="31-12N-24W", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="1")])
    write_report(report, out, template=seed)
    # media survived, and the data sheet holds the new report
    assert any("media" in n for n in zipfile.ZipFile(out).namelist())
    loaded = read_report(out)
    assert loaded.rows and loaded.rows[0].grantor == "A"


def test_write_report_fresh_when_no_template(tmp_path):
    out = tmp_path / "fresh.xlsx"
    write_report(ReportModel(section="X", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="1")]), out)
    assert out.exists()
    assert not any("media" in n for n in zipfile.ZipFile(out).namelist())


# -- #2 a repaired report that passes converges on the final iteration ------
def test_final_iteration_revalidates(tmp_path, monkeypatch):
    cfg = HorizonConfig(root=tmp_path / "H")
    cfg.ensure_workspace()
    cfg.max_loops = 1
    audit = AuditLog(path=None, echo=False)
    base = "31-12N-24W_Roger_Mills_Cursory_Title_Report"
    # seed a v001 so ingest/repair have a file
    write_report(ReportModel(section="31-12N-24W", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="1")]),
        cfg.final_reports / f"{base}_v001.xlsx")

    orch = Orchestrator(cfg, audit)

    # First validate() fails, forcing a repair; after reload the report passes.
    calls = {"n": 0}
    real_validate = orch.validate

    def fake_validate(report, reqs):
        calls["n"] += 1
        vr = real_validate(report, reqs)
        if calls["n"] == 1:
            vr.passed = False  # force one repair cycle
        return vr

    monkeypatch.setattr(orch, "validate", fake_validate)
    # make repair produce a new version so the loop reloads and revalidates
    monkeypatch.setattr(orch, "repair",
                        lambda src, stem: write_report(
                            read_report(src, section=cfg.section),
                            __import__("horizon.versioning", fromlist=["next_version_path"])
                            .next_version_path(cfg.final_reports, stem)))

    report = read_report(cfg.final_reports / f"{base}_v001.xlsx", section=cfg.section)
    result = orch.run(report, base, Requirements())
    assert result.converged  # final revalidation caught the now-passing report
