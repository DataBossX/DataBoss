"""Regression tests for the sixth round of PR#18 review fixes."""

import io
import zipfile
from pathlib import Path

import openpyxl
import pytest

from horizon.chaining import OGLRecord, RunsheetNote
from horizon.models import REVIEW_TAG, ReportModel, TitleRow
from horizon.pipeline import chain_to_report
from horizon.report_io import read_report, write_report
from horizon.validation import Requirements, validate_report

_HAS_PIL = False
try:
    from PIL import Image as _PILImage  # noqa: F401
    _HAS_PIL = True
except ImportError:
    pass


def _media_seed(path: Path, active_title: str):
    """A workbook whose ACTIVE sheet is a plat tab carrying an image."""
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = active_title
    b = io.BytesIO(); PILImage.new("RGB", (8, 8), (0, 200, 100)).save(b, format="PNG"); b.seek(0)
    ws.add_image(XLImage(b), "A1")
    wb.save(path)
    wb.close()


# -- #1 read targets the data sheet by name, not whatever is active --------
@pytest.mark.skipif(not _HAS_PIL, reason="Pillow needed")
def test_read_targets_data_sheet_not_active(tmp_path):
    seed = tmp_path / "seed.xlsx"
    _media_seed(seed, active_title="Plats")   # active sheet is a plat, not data
    out = tmp_path / "out.xlsx"
    report = ReportModel(section="31-12N-24W", rows=[
        TitleRow(grantor="ROOT", grantee="X", instrument_number="1",
                 conveyed_interest="1/2")])
    write_report(report, out, template=seed)
    loaded = read_report(out)   # must find "Title Report", not the Plats tab
    assert loaded.rows and loaded.rows[0].grantor == "ROOT"


# -- #3 a media template without a Title Report sheet keeps the plat tab ----
@pytest.mark.skipif(not _HAS_PIL, reason="Pillow needed")
def test_media_template_preserves_plat_sheet(tmp_path):
    seed = tmp_path / "seed.xlsx"
    _media_seed(seed, active_title="Plats")
    out = tmp_path / "out.xlsx"
    write_report(ReportModel(section="X", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="1")]), out, template=seed)
    wb = openpyxl.load_workbook(out)
    assert "Plats" in wb.sheetnames        # plat tab NOT clobbered
    assert "Title Report" in wb.sheetnames  # data sheet added alongside
    assert any("media" in n for n in zipfile.ZipFile(out).namelist())
    wb.close()


# -- #4 over-conveyance: no negative retained, no validation stall ----------
def test_over_conveyance_no_negative_retained():
    ogl = [
        OGLRecord(instrument_number="1", grantor="ROOT", grantee="X",
                  conveyed_interest="1/2", legal_description="Sec 31"),
        OGLRecord(instrument_number="2", grantor="X", grantee="Y",
                  conveyed_interest="3/4", legal_description="Sec 31"),  # over-conveys
    ]
    build = chain_to_report(ogl, [RunsheetNote(instrument_number="1"),
                                  RunsheetNote(instrument_number="2")])
    over = build.report.rows[1]
    assert over.status == REVIEW_TAG
    assert over.retained_interest == ""          # no negative retained emitted
    # the built report must not stall validation with a redundant error
    vr = validate_report(build.report, Requirements())
    assert not any(i.severity == "error" for i in vr.issues)


def test_flagged_row_skips_interest_gate():
    report = ReportModel(section="X", rows=[
        TitleRow(grantor="A", grantee="B", instrument_number="1",
                 conveyed_interest="1/2", retained_interest="-1/4",
                 status="Needs Examiner Review")])
    vr = validate_report(report, Requirements())
    assert not any(i.severity == "error" for i in vr.errors)


# -- #5 pandas is not a horizon dependency ---------------------------------
def test_horizon_requirements_has_no_pandas():
    reqs = Path(__file__).resolve().parent.parent / "horizon" / "requirements.txt"
    body = reqs.read_text()
    # a real dependency line for pandas (ignore the explanatory comment)
    dep_lines = [ln.strip() for ln in body.splitlines()
                 if ln.strip() and not ln.strip().startswith("#")]
    assert not any(ln.lower().startswith("pandas") for ln in dep_lines)
