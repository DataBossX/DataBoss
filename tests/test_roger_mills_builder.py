#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Regression tests for automation/roger_mills_title_report_builder.py (codexv2).

Run with:  python -m unittest tests.test_roger_mills_builder
These tests need only openpyxl (the tool's one hard dependency). Tests that
require PyMuPDF (PDF runsheet harvesting) are skipped automatically if it is
not installed, so the suite stays green on a minimal install.
"""
from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
import zipfile
from fractions import Fraction
from pathlib import Path

try:
    import openpyxl
except Exception:  # pragma: no cover - openpyxl is required to run these tests
    openpyxl = None

# Import the builder module by path (it lives in automation/, not a package).
_MOD_PATH = Path(__file__).resolve().parent.parent / "automation" / "roger_mills_title_report_builder.py"
_spec = importlib.util.spec_from_file_location("rm_builder", _MOD_PATH)
rm = importlib.util.module_from_spec(_spec)
sys.modules["rm_builder"] = rm  # let dataclasses resolve annotations by module name
_spec.loader.exec_module(rm)  # type: ignore

HEADERS = ["Entry", "Instrument Date", "Doc Type", "Grantor", "Grantee",
           "Book", "Page", "Instrument No", "Legal Description", "Interest", "Remarks"]


def _write_report(path: Path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Title Report"
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    wb.save(path)


def _title_rows(rows):
    """Build rm.TitleRow objects straight from (grantor, grantee, interest,
    instrument_no, date) tuples for chain-only tests."""
    out = []
    for i, (grantor, grantee, interest, inst, date) in enumerate(rows, start=1):
        out.append(rm.TitleRow(data={
            "grantor": grantor, "grantee": grantee, "interest": interest,
            "instrument_no": inst, "instrument_date": date,
        }, source="test", source_row=i))
    return out


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class ParseInterestTests(unittest.TestCase):
    def test_fraction(self):
        self.assertEqual(rm.parse_interest("1/2")[0], Fraction(1, 2))
        self.assertEqual(rm.parse_interest("an undivided 3/16")[0], Fraction(3, 16))

    def test_fraction_product(self):
        self.assertEqual(rm.parse_interest("1/2 of 1/8")[0], Fraction(1, 16))
        self.assertEqual(rm.parse_interest("1/2 x 1/4")[0], Fraction(1, 8))

    def test_percent(self):
        self.assertEqual(rm.parse_interest("6.25%")[0], Fraction(1, 16))
        self.assertEqual(rm.parse_interest("50%")[0], Fraction(1, 2))

    def test_decimal(self):
        self.assertEqual(rm.parse_interest("0.25")[0], Fraction(1, 4))

    def test_nma_needs_gross(self):
        val, kind, _ = rm.parse_interest("20 NMA")
        self.assertIsNone(val)
        self.assertEqual(kind, "nma_no_gross")
        self.assertEqual(rm.parse_interest("20 NMA", gross_acres=160)[0], Fraction(1, 8))

    def test_unparseable(self):
        val, kind, _ = rm.parse_interest("see deed")
        self.assertIsNone(val)
        self.assertEqual(kind, "unknown")

    def test_empty(self):
        self.assertIsNone(rm.parse_interest("")[0])


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class ChainTests(unittest.TestCase):
    def test_running_and_ledger_balance(self):
        rows = _title_rows([
            ("USA", "ALPHA CORP", "1", "1", "1/1/1980"),
            ("ALPHA CORP", "BRAVO LLC", "1/2", "2", "1/1/1990"),
            ("BRAVO LLC", "CHARLIE JONES", "1/4", "3", "1/1/2000"),
        ])
        links, ledger = rm.build_interest_chain(rows)
        # conveyed fractions
        self.assertEqual([lk.fraction for lk in links],
                         [Fraction(1), Fraction(1, 2), Fraction(1, 4)])
        # ledger nets to zero (every conveyance out equals a conveyance in)
        self.assertEqual(sum(ledger.values(), Fraction(0)), Fraction(0))
        # BRAVO received 1/2, then conveyed 1/4 -> nets 1/4
        self.assertEqual(ledger[rm.norm_name("BRAVO LLC")], Fraction(1, 4))

    def test_over_conveyance_flag(self):
        rows = _title_rows([
            ("ALPHA CORP", "BRAVO LLC", "1/4", "1", "1/1/1990"),   # ALPHA untracked -> no flag
            ("BRAVO LLC", "CHARLIE JONES", "1/2", "2", "1/1/1995"),  # BRAVO holds 1/4 -> over
        ])
        links, _ = rm.build_interest_chain(rows)
        self.assertIn("OVER_CONVEYANCE", links[1].flags)

    def test_unparsed_interest_flag(self):
        links, _ = rm.build_interest_chain(
            _title_rows([("ALPHA CORP", "BRAVO LLC", "see deed", "1", "1/1/1990")]))
        self.assertIn("UNPARSED_INTEREST", links[0].flags)
        self.assertIsNone(links[0].fraction)

    def test_proportional_of_grantor(self):
        # ALPHA ends holding 1/2; conveys "1/2 of grantor's interest" -> 1/4 conveyed
        rows = _title_rows([
            ("USA", "ALPHA CORP", "1/2", "1", "1/1/1980"),
            ("ALPHA CORP", "BRAVO LLC", "1/2 of grantor's interest", "2", "1/1/1990"),
        ])
        links, ledger = rm.build_interest_chain(rows)
        self.assertEqual(links[1].fraction, Fraction(1, 4))
        self.assertIn("PROPORTIONAL", links[1].flags)
        self.assertEqual(ledger[rm.norm_name("BRAVO LLC")], Fraction(1, 4))

    def test_proportional_no_basis(self):
        links, _ = rm.build_interest_chain(_title_rows(
            [("ALPHA CORP", "BRAVO LLC", "1/2 of grantor's interest", "1", "1/1/1990")]))
        self.assertIn("PROPORTIONAL_NO_BASIS", links[0].flags)
        self.assertIsNone(links[0].fraction)


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class TidyAndZipTests(unittest.TestCase):
    def test_zip_extraction_is_safe_and_idempotent(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            zpath = root / "a.zip"
            with zipfile.ZipFile(zpath, "w") as zf:
                zf.writestr("inner/file.txt", "hello")
                zf.writestr("../evil.txt", "should be skipped")  # traversal attempt
            _, n = rm.extract_all_zips(root)
            self.assertGreaterEqual(n, 1)
            self.assertTrue((root / rm.EXTRACT_DIRNAME / "a" / "inner" / "file.txt").exists())
            self.assertFalse((root.parent / "evil.txt").exists())  # traversal blocked
            # second run extracts nothing new
            _, n2 = rm.extract_all_zips(root)
            self.assertEqual(n2, 0)

    def test_duplicate_and_trash_detection(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            (root / "a.txt").write_text("same content")
            (root / "b.txt").write_text("same content")   # exact duplicate
            (root / "c.txt").write_text("different")
            (root / "empty.txt").write_bytes(b"")           # trash (empty)
            (root / "~$lock.xlsx").write_bytes(b"x")         # trash (office lock)
            recs = rm.inventory(root)
            dups, trash = rm.find_duplicates_and_trash(recs)
            dup_names = {r.path.name for g in dups for r in g}
            self.assertEqual(dup_names, {"a.txt", "b.txt"})
            self.assertEqual({r.path.name for r in trash}, {"empty.txt", "~$lock.xlsx"})


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class DelimitedLoaderTests(unittest.TestCase):
    def test_csv_report_rows(self):
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "extra_report.csv"
            p.write_text("Grantor,Grantee,Instrument No,Interest\nX,Y,2024-1,1/32\n")
            rec = rm.FileRec(path=p, rel=p.name, ext=".csv", size=p.stat().st_size,
                             mtime=rm._dt.datetime.now())
            rows = rm.title_rows_from_table(rec)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].data["interest"], "1/32")
            self.assertEqual(rows[0].data["grantee"], "Y")


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class DeedExtractionTests(unittest.TestCase):
    def test_parse_deed_text(self):
        text = ("MINERAL DEED made this 7th day of July, 2018, by and between "
                "A. RANCHER, Grantor, and E. WHITE, Grantee. Grantor does grant, "
                "bargain, sell and convey unto E. WHITE an undivided 1/4 interest "
                "in the NW/4. Recorded in Book 455 Page 9. Instrument No. 2018-345.")
        d = rm.parse_deed_text(text)
        self.assertEqual(d["doc_type"], "Mineral Deed")
        self.assertEqual(d["grantor"], "A. RANCHER")
        self.assertEqual(d["grantee"], "E. WHITE")
        self.assertEqual((d["book"], d["page"]), ("455", "9"))
        self.assertEqual(d["instrument_no"], "2018-345")
        self.assertEqual(d["confidence"], "5/5")

    def test_parse_deed_blank_when_absent(self):
        d = rm.parse_deed_text("Some unrelated text with no title fields at all.")
        self.assertEqual(d["grantor"], "")
        self.assertEqual(d["instrument_no"], "")
        self.assertEqual(d["confidence"], "0/5")


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class TitleSummaryTests(unittest.TestCase):
    def test_current_owners_positive_only_with_nma(self):
        _, ledger = rm.build_interest_chain(_title_rows([
            ("USA", "ALPHA CORP", "1", "1", "1/1/1980"),
            ("ALPHA CORP", "BRAVO LLC", "1/2", "2", "1/1/1990"),
        ]), gross_acres=160)
        owners = rm.current_owners(ledger, gross_acres=160)
        parties = {o["party"]: o for o in owners}
        # USA is negative (-1) and must be excluded; positives keep their NMA
        self.assertNotIn(rm.norm_name("USA"), parties)
        self.assertEqual(parties[rm.norm_name("BRAVO LLC")]["fraction"], Fraction(1, 2))
        self.assertEqual(parties[rm.norm_name("BRAVO LLC")]["nma"], 80.0)

    def test_current_owners_no_gross_leaves_nma_none(self):
        _, ledger = rm.build_interest_chain(_title_rows([
            ("USA", "ALPHA CORP", "1", "1", "1/1/1980"),
        ]))
        owners = rm.current_owners(ledger)
        self.assertTrue(all(o["nma"] is None for o in owners))


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class SelfTestModeTests(unittest.TestCase):
    def test_self_test_passes(self):
        self.assertEqual(rm.run_self_test(), 0)


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class HtmlReportTests(unittest.TestCase):
    def test_html_report_is_self_contained(self):
        links, ledger = rm.build_interest_chain(_title_rows([
            ("USA", "ALPHA CORP", "1", "1", "1/1/1980"),
            ("ALPHA CORP", "BRAVO LLC", "see deed", "2", "1/1/1990"),  # flagged
        ]))
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "report.html"
            rm.write_html_report(p, "31-12N-24W", links, ledger, [], [],
                                 {"Records": 2}, ["do a spot-check"])
            html = p.read_text(encoding="utf-8")
        self.assertIn("<table", html)
        self.assertIn('class="flag"', html)          # the flagged row is marked
        self.assertNotIn("http://", html)             # no external assets
        self.assertNotIn("https://", html)
        self.assertIn("do a spot-check", html)        # checklist rendered


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class FullBuildTests(unittest.TestCase):
    def test_end_to_end_build(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td) / "Roger Mills"
            root.mkdir()
            # template (formatting authority) + one report
            _write_report(root / "Template(30).xlsx", [])
            _write_report(root / "Roger_Mills_Title_Report.xlsx", [
                [1, "1/1/1980", "Patent", "USA", "A", 10, 1, "1980-1", "NW/4 31-12N-24W", "1", ""],
                [2, "1/1/1990", "Mineral Deed", "A", "B", 20, 2, "1990-2", "NW/4 31-12N-24W", "1/2", ""],
            ])
            out = root / "out.xlsx"
            rc = rm.main([
                "--root", str(root), "--output", str(out),
                "--support-dir", str(root / "files"),
                "--section", "31-12N-24W", "--gross-acres", "160",
            ])
            self.assertEqual(rc, 0)
            self.assertTrue(out.exists())
            wb = openpyxl.load_workbook(out)
            for sheet in ("Interest Chain", "Ownership Ledger", "Review Flags"):
                self.assertIn(sheet, wb.sheetnames)
            # chain sheet has header + 2 conveyances
            self.assertEqual(wb["Interest Chain"].max_row, 3)


if __name__ == "__main__":
    unittest.main()
