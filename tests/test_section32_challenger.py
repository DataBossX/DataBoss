"""Automated tests for Section 32 Challenger Package deliverables and workbook integrity."""

import hashlib
import json
from pathlib import Path
import openpyxl
import pytest

from horizon.workbook_qa import inspect_workbook, EXCEL_ERROR_VALUES


def test_section32_challenger_workbook_integrity():
    wb_path = Path("/workspace/SECTION32_GEMINI37_CHALLENGER_20260830.xlsx")
    assert wb_path.exists(), "Section 32 Challenger workbook must exist on disk"

    wb = openpyxl.load_workbook(wb_path, data_only=False)
    
    # 1. Exact 13 sheets contract
    expected_sheets = [
        "Overview", "Title ", "OGL", "PLAT", "Runsheet",
        "Tract 1", "Tract 2", "Tract 3", "Tract 4", "Tract 5",
        "WI 2", "WI 1", "Well 1"
    ]
    assert wb.sheetnames == expected_sheets, f"Sheet order/names mismatch: {wb.sheetnames}"

    # 2. No formula errors in any cell
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None:
                    val_str = str(cell.value)
                    for err in EXCEL_ERROR_VALUES:
                        assert err not in val_str, f"Found formula error {err} in {sheet_name}!{cell.coordinate}"

    # 3. HOLD marker verification
    overview = wb["Overview"]
    found_hold = False
    for row in overview.iter_rows(values_only=True):
        for cell_val in row:
            if cell_val and "FOR REVIEW — HOLD NO EXTERNAL RELEASE" in str(cell_val):
                found_hold = True
                break
    assert found_hold, "Mandatory HOLD marker missing from Overview sheet"

    wb.close()


def test_section32_challenger_pdf_deliverables():
    full_pdf = Path("/workspace/SECTION32_GEMINI37_FULL_INTERNAL_20260830.pdf")
    boss_pdf = Path("/workspace/SECTION32_GEMINI37_BOSS_REVIEW_20260830.pdf")

    assert full_pdf.exists() and full_pdf.stat().st_size > 5000, "Full Internal PDF must exist and be populated"
    assert boss_pdf.exists() and boss_pdf.stat().st_size > 3000, "Boss Review PDF must exist and be populated"


def test_section32_handoff_and_hashes():
    handoff_path = Path("/workspace/MACHINE_READABLE_HANDOFF.json")
    assert handoff_path.exists(), "MACHINE_READABLE_HANDOFF.json must exist"

    with handoff_path.open(encoding="utf-8") as f:
        data = json.load(f)

    assert data["reviewer_model"] == "Gemini 3.7 Flash"
    assert data["hold_preserved"] is True
    assert data["protected_files_unchanged"] is True
    assert len(data["files_created"]) >= 15

    # Verify SHA256 matches for created files
    for file_entry in data["files_created"]:
        fpath = Path(file_entry["path_or_drive_id"])
        if fpath.exists():
            computed_sha = hashlib.sha256(fpath.read_bytes()).hexdigest()
            assert computed_sha == file_entry["sha256"], f"Hash mismatch for {file_entry['name']}"


def test_qa_inspection_score():
    wb_path = Path("/workspace/SECTION32_GEMINI37_CHALLENGER_20260830.xlsx")
    profile = {
        "current_owner_columns": [
            {"sheet": "Title ", "column": "B", "start_row": 5, "end_row": 9}
        ]
    }
    report = inspect_workbook(wb_path, ["no_duplicate_owners"], profile=profile)
    assert report.score.technical_pass is True
    assert report.score.unresolved_conflicts == 0
    assert report.score.mathematical_accuracy == 100.0
