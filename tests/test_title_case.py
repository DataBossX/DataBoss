from __future__ import annotations

import hashlib
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from databossx.config import KernelConfig  # noqa: E402
from databossx.products.title.ledger import calculate_ownership  # noqa: E402
from databossx.products.title.manager import _positive_fraction  # noqa: E402
from databossx.products.title.render import safe_cell  # noqa: E402
from databossx.service import KernelService  # noqa: E402


def config_for(tmp_path: Path) -> KernelConfig:
    runtime = tmp_path / "runtime"
    return KernelConfig(
        repo_root=ROOT,
        runtime_root=runtime,
        database_path=runtime / "kernel.sqlite3",
        vault_root=runtime / "vault",
        projects_root=runtime / "projects",
        migrations_root=ROOT / "migrations",
    )


def reviewed_instrument(
    sequence: int,
    recording: str,
    grantor: str,
    grantee: str,
    numerator: int,
    denominator: int,
    asset_id: str,
    text_length: int,
    basis: str = "ABSOLUTE_ESTATE",
) -> dict:
    return {
        "sequence_no": sequence,
        "instrument_type": "Mineral Deed",
        "recording_reference": recording,
        "effective_date": f"202{sequence}-01-01",
        "grantor_name": grantor,
        "grantee_name": grantee,
        "conveyed_interest": {
            "numerator": numerator,
            "denominator": denominator,
        },
        "interest_basis": basis,
        "evidence_asset_version_id": asset_id,
        "evidence_char_start": 0,
        "evidence_char_end": text_length,
        "review_status": "REVIEWED",
    }


def test_branching_exact_ownership_preserves_all_owner_balances() -> None:
    opening = [{"owner_name": "Alpha", "interest_num": 1, "interest_den": 1}]
    evidence = "evidence"
    instruments = [
        {
            **reviewed_instrument(1, "BK1-P1", "Alpha", "Beta", 1, 2, evidence, 1),
            "conveyed_num": 1,
            "conveyed_den": 2,
        },
        {
            **reviewed_instrument(
                2, "BK1-P2", "Beta", "Gamma", 1, 2, evidence, 1, "OF_GRANTOR"
            ),
            "conveyed_num": 1,
            "conveyed_den": 2,
        },
    ]
    result = calculate_ownership(opening, instruments)
    assert str(result.ownership["Alpha"]) == "1/2"
    assert str(result.ownership["Beta"]) == "1/4"
    assert str(result.ownership["Gamma"]) == "1/4"
    assert sum(result.ownership.values()) == 1
    assert not result.blocking_defects
    assert all(entry.estate_total == 1 for entry in result.entries)


def test_duplicate_and_overconveyance_are_blocking_and_not_applied() -> None:
    opening = [{"owner_name": "Alpha", "interest_num": 1, "interest_den": 1}]
    base = {
        **reviewed_instrument(1, "DUP-1", "Alpha", "Beta", 3, 2, "asset", 1),
        "conveyed_num": 3,
        "conveyed_den": 2,
    }
    duplicate = {
        **reviewed_instrument(2, " dup 1 ", "Alpha", "Gamma", 1, 2, "asset", 1),
        "conveyed_num": 1,
        "conveyed_den": 2,
    }
    result = calculate_ownership(opening, [base, duplicate])
    assert result.ownership == {"Alpha": 1}
    assert {defect.code for defect in result.blocking_defects} == {
        "OVER_CONVEYANCE",
        "DUPLICATE_INSTRUMENT",
    }
    assert safe_cell("=HYPERLINK(\"https://evil.example\")").startswith("'=")
    with pytest.raises(ValueError, match="exact numerator"):
        _positive_fraction({"numerator": 1.9, "denominator": 2}, "interest")
    with pytest.raises(ValueError, match="exact numerator"):
        _positive_fraction({"numerator": True, "denominator": 2}, "interest")


def test_title_package_xlsx_pdf_hash_review_end_to_end(tmp_path: Path) -> None:
    source = tmp_path / "source"
    source.mkdir()
    first_text = "Mineral Deed Alpha conveys an undivided 1/2 to Beta. Book 1 Page 1."
    second_text = "Mineral Deed Beta conveys 1/2 of grantor interest to Gamma. Book 1 Page 2."
    (source / "deed-1.txt").write_text(first_text, encoding="utf-8")
    (source / "deed-2.txt").write_text(second_text, encoding="utf-8")
    service = KernelService(config_for(tmp_path))
    project = service.create_project("Synthetic Title Case", source)
    service.ingest_project(project["id"])
    assets = {asset["rel_path"]: asset for asset in service.list_assets(project["id"])}
    payload = {
        "name": "Synthetic <b> Branching & Ownership",
        "legal_description": "SYNTHETIC Section 1, T1N, R1W",
        "gross_acres": {"numerator": 160, "denominator": 1},
        "opening_ownership": [
            {
                "owner_name": "Alpha",
                "interest": {"numerator": 1, "denominator": 1},
            }
        ],
        "instruments": [
            reviewed_instrument(
                1,
                "BK1-P1",
                "Alpha",
                "Beta",
                1,
                2,
                assets["deed-1.txt"]["id"],
                len(first_text),
            ),
            reviewed_instrument(
                2,
                "BK1-P2",
                "Beta",
                "Gamma",
                1,
                2,
                assets["deed-2.txt"]["id"],
                len(second_text),
                "OF_GRANTOR",
            ),
        ],
    }
    title_case = service.create_title_case(project["id"], payload)
    assert all(
        len(instrument["evidence_source_sha256"]) == 64
        and len(instrument["evidence_extraction_sha256"]) == 64
        and len(instrument["evidence_span_sha256"]) == 64
        and instrument["evidence_span_text"]
        for instrument in title_case["instruments"]
    )
    package = service.build_title_package(title_case["id"])
    assert (source / "deed-1.txt").read_text(encoding="utf-8") == first_text
    assert (source / "deed-2.txt").read_text(encoding="utf-8") == second_text
    assert package["run_status"] == "WAITING_HUMAN"
    assert package["blocking_defect_count"] == 0
    artifacts = {item["rel_path"]: item for item in package["artifacts"]}
    assert {
        "Title_Examiner_Packet.xlsx",
        "Draft_Abstract_Aid.pdf",
        "title_case_summary.json",
        "package_manifest.json",
    } <= artifacts.keys()
    with pytest.raises(ValueError, match="export is blocked"):
        service.export_artifact(artifacts["Title_Examiner_Packet.xlsx"]["id"])

    _, workbook_path = service.artifact(
        artifacts["Title_Examiner_Packet.xlsx"]["id"]
    )
    with workbook_path.open("rb") as workbook_file:
        workbook = load_workbook(workbook_file, read_only=True, data_only=False)
        assert set(workbook.sheetnames) == {
            "Case Summary",
            "Runsheet",
            "Current Ownership",
            "Ownership Ledger",
            "Defects and Curative",
            "Evidence Index",
        }
        ownership_rows = list(
            workbook["Current Ownership"].iter_rows(values_only=True)
        )
        workbook.close()
    assert ("Alpha", "1/2", 1, 2, "160", "80") in ownership_rows
    assert ("Beta", "1/4", 1, 4, "160", "40") in ownership_rows
    assert ("Gamma", "1/4", 1, 4, "160", "40") in ownership_rows

    reviewer = service.register_title_reviewer(
        "Qualified Examiner", "QUALIFIED_EXAMINER", "629104"
    )
    with pytest.raises(ValueError, match="does not match"):
        service.review_title_package(
            package["pipeline_run_id"],
            "0" * 64,
            reviewer["id"],
            "629104",
            "APPROVE",
        )
    with pytest.raises(ValueError, match="credential is invalid"):
        service.review_title_package(
            package["pipeline_run_id"],
            package["package_manifest_sha256"],
            reviewer["id"],
            "wrong-pin",
            "APPROVE",
        )
    approved = service.review_title_package(
        package["pipeline_run_id"],
        package["package_manifest_sha256"],
        reviewer["id"],
        "629104",
        "APPROVE",
        "Synthetic test approval only",
    )
    assert approved["review_status"] == "APPROVED"
    exported, exported_path = service.export_artifact(
        artifacts["Title_Examiner_Packet.xlsx"]["id"]
    )
    assert exported["title_review_status"] == "APPROVED"
    assert exported_path.is_file()
    assert service.get_title_case(title_case["id"])["status"] == "APPROVED_FOR_EXPORT"
    rebuilt = service.build_title_package(title_case["id"])
    assert rebuilt["review_status"] == "AWAITING_REVIEW"
    assert rebuilt["approved_manifest_sha256"] is None
    assert service.health()["audit_chain_valid"] is True


def test_unreviewed_instrument_blocks_approval(tmp_path: Path) -> None:
    source = tmp_path / "source"
    source.mkdir()
    (source / "notice.txt").write_text("Unresolved title evidence", encoding="utf-8")
    service = KernelService(config_for(tmp_path))
    project = service.create_project("Defect Case", source)
    service.ingest_project(project["id"])
    payload = {
        "name": "Needs Examiner",
        "legal_description": "SYNTHETIC TRACT",
        "gross_acres": {"numerator": 40, "denominator": 1},
        "opening_ownership": [
            {"owner_name": "Alpha", "interest": {"numerator": 1, "denominator": 1}}
        ],
        "instruments": [
            {
                "sequence_no": 1,
                "instrument_type": "Mineral Deed",
                "recording_reference": "UNKNOWN-1",
                "grantor_name": "Alpha",
                "grantee_name": "Beta",
                "conveyed_interest": {"numerator": 1, "denominator": 2},
                "interest_basis": "ABSOLUTE_ESTATE",
                "review_status": "NEEDS_REVIEW",
            }
        ],
    }
    title_case = service.create_title_case(project["id"], payload)
    package = service.build_title_package(title_case["id"])
    assert package["blocking_defect_count"] == 1
    assert package["defects"][0]["code"] == "INSTRUMENT_NEEDS_REVIEW"
    with pytest.raises(ValueError, match="Blocking defects"):
        reviewer = service.register_title_reviewer(
            "Defect Examiner", "QUALIFIED_EXAMINER", "183746"
        )
        service.review_title_package(
            package["pipeline_run_id"],
            package["package_manifest_sha256"],
            reviewer["id"],
            "183746",
            "APPROVE",
        )


def test_migration_invalidates_pre_provenance_pending_package(tmp_path: Path) -> None:
    evidence_text = "Recorded mineral deed"
    evidence_sha256 = hashlib.sha256(evidence_text.encode("utf-8")).hexdigest()
    migrations = tmp_path / "migrations"
    migrations.mkdir()
    for number in range(1, 5):
        source = next((ROOT / "migrations").glob(f"00{number}_*.sql"))
        shutil.copy2(source, migrations / source.name)
    config = config_for(tmp_path)
    config = KernelConfig(
        repo_root=config.repo_root,
        runtime_root=config.runtime_root,
        database_path=config.database_path,
        vault_root=config.vault_root,
        projects_root=config.projects_root,
        migrations_root=migrations,
    )
    service = KernelService(config)
    with service.db.transaction(immediate=True) as connection:
        connection.execute(
            """
            INSERT INTO projects(
                id, name, source_root, source_kind, status, created_at
            ) VALUES ('p', 'Legacy', '/tmp', 'synthetic', 'DRAFT', 'now')
            """
        )
        connection.execute(
            """
            INSERT INTO ingest_runs(
                id, project_id, status, started_at
            ) VALUES ('i', 'p', 'SUCCEEDED', 'now')
            """
        )
        connection.execute(
            """
            INSERT INTO pipeline_runs(
                id, project_id, ingest_run_id, pipeline, pipeline_version,
                input_manifest_sha256, status, run_dir, started_at
            ) VALUES (
                'r', 'p', 'i', 'title-examiner-packet', '1',
                'hash', 'WAITING_HUMAN', '/tmp', 'now'
            )
            """
        )
        connection.execute(
            """
            INSERT INTO title_cases(
                id, project_id, name, legal_description,
                gross_acres_num, gross_acres_den, created_at
            ) VALUES ('c', 'p', 'Legacy', 'Synthetic', 1, 1, 'now')
            """
        )
        connection.execute(
            """
            INSERT INTO title_package_details(
                pipeline_run_id, title_case_id, package_manifest_sha256,
                blocking_defect_count, review_status
            ) VALUES ('r', 'c', 'old-hash', 0, 'AWAITING_REVIEW')
            """
        )
        connection.execute(
            """
            INSERT INTO blobs(sha256, size_bytes, vault_relpath, created_at)
            VALUES (?, ?, ?, 'now')
            """,
            (evidence_sha256, len(evidence_text), evidence_sha256),
        )
        connection.execute(
            """
            INSERT INTO asset_versions(
                id, project_id, ingest_run_id, rel_path, extension, size_bytes,
                source_mtime_ns, blob_sha256, custody_at
            ) VALUES ('a2', 'p', 'i', 'deed.txt', '.txt', ?, 0, ?, 'now')
            """,
            (len(evidence_text), evidence_sha256),
        )
        connection.execute(
            """
            INSERT INTO text_extractions(
                id, asset_version_id, status, extractor, extractor_version,
                text_sha256, text_content, char_count, created_at
            ) VALUES ('e2', 'a2', 'SUCCEEDED', 'plain_text', '1', ?, ?, ?, 'now')
            """,
            (evidence_sha256, evidence_text, len(evidence_text)),
        )
        connection.execute(
            """
            INSERT INTO pipeline_runs(
                id, project_id, ingest_run_id, pipeline, pipeline_version,
                input_manifest_sha256, status, run_dir, started_at
            ) VALUES (
                'r2', 'p', 'i', 'title-examiner-packet', '1',
                'hash', 'WAITING_HUMAN', '/tmp', 'now'
            )
            """
        )
        connection.execute(
            """
            INSERT INTO title_cases(
                id, project_id, name, legal_description,
                gross_acres_num, gross_acres_den, created_at
            ) VALUES ('c2', 'p', 'Bound', 'Synthetic', 1, 1, 'now')
            """
        )
        connection.execute(
            """
            INSERT INTO title_instruments(
                id, title_case_id, sequence_no, instrument_type,
                recording_reference, grantor_name, grantee_name,
                conveyed_num, conveyed_den, interest_basis,
                evidence_asset_version_id, evidence_char_start, evidence_char_end,
                review_status, recording_reference_key, evidence_ingest_run_id,
                evidence_source_sha256, evidence_extraction_sha256,
                evidence_span_sha256, evidence_span_text
            ) VALUES (
                'ti2', 'c2', 1, 'Mineral Deed', 'BK1-P1', 'Grantor', 'Grantee',
                1, 1, 'ABSOLUTE_ESTATE', 'a2', 0, ?, 'REVIEWED', 'bk1p1', 'i',
                ?, ?, ?, ?
            )
            """,
            (
                len(evidence_text),
                evidence_sha256,
                evidence_sha256,
                evidence_sha256,
                evidence_text,
            ),
        )
        connection.execute(
            """
            INSERT INTO title_package_details(
                pipeline_run_id, title_case_id, package_manifest_sha256,
                blocking_defect_count, review_status
            ) VALUES ('r2', 'c2', 'bound-hash', 0, 'AWAITING_REVIEW')
            """
        )
    for number in (5, 6):
        migration = next((ROOT / "migrations").glob(f"00{number}_*.sql"))
        shutil.copy2(migration, migrations / migration.name)
    upgraded = KernelService(config)
    with upgraded.db.connect() as connection:
        package = connection.execute(
            "SELECT * FROM title_package_details WHERE pipeline_run_id='r'"
        ).fetchone()
        run = connection.execute(
            "SELECT * FROM pipeline_runs WHERE id='r'"
        ).fetchone()
        bound_package = connection.execute(
            "SELECT * FROM title_package_details WHERE pipeline_run_id='r2'"
        ).fetchone()
        bound_run = connection.execute(
            "SELECT * FROM pipeline_runs WHERE id='r2'"
        ).fetchone()
    assert package["review_status"] == "INVALIDATED_PRE_PROVENANCE"
    assert run["status"] == "FAILED_TERMINAL"
    assert bound_package["review_status"] == "AWAITING_REVIEW"
    assert bound_run["status"] == "WAITING_HUMAN"
