from __future__ import annotations

import json
import sys
from fractions import Fraction
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from databossx.config import KernelConfig  # noqa: E402
from databossx.products.title.economics import calculate_economics  # noqa: E402
from databossx.service import KernelService  # noqa: E402


def config_for(tmp_path: Path) -> KernelConfig:
    runtime = tmp_path / "runtime"
    return KernelConfig(
        repo_root=ROOT,
        runtime_root=runtime,
        database_path=runtime / "kernel.sqlite3",
        vault_root=runtime / "vault",
        projects_root=runtime / "projects",
        migrations_root=ROOT / "migrations",
    )


def unit_row(mineral: Fraction = Fraction(1), royalty: Fraction = Fraction(3, 16)) -> dict:
    return {
        "id": "unit-1",
        "unit_label": "Synthetic Lease Unit",
        "gross_acres_num": 160,
        "gross_acres_den": 1,
        "leased_mineral_num": mineral.numerator,
        "leased_mineral_den": mineral.denominator,
        "base_royalty_num": royalty.numerator,
        "base_royalty_den": royalty.denominator,
        "wi_basis": "WI_EQUALS_LEASEHOLD",
        "review_status": "REVIEWED",
        "unsupported_terms": "",
        "evidence_span_sha256": "a" * 64,
    }


def event(
    sequence: int,
    kind: str,
    recording: str,
    to_party: str,
    interest: Fraction,
    *,
    from_party: str | None = None,
    basis: str = "ABSOLUTE_UNIT",
    burden_kind: str | None = None,
) -> dict:
    return {
        "sequence_no": sequence,
        "event_kind": kind,
        "recording_reference": recording,
        "from_party": from_party,
        "to_party": to_party,
        "interest_num": interest.numerator,
        "interest_den": interest.denominator,
        "interest_basis": basis,
        "burden_kind": burden_kind,
        "review_status": "REVIEWED",
        "evidence_span_sha256": "b" * 64,
    }


def test_exact_full_lease_wi_nri_and_royalty() -> None:
    result = calculate_economics(
        unit_row(),
        [event(1, "OPENING_LEASEHOLD", "L-1", "Operator", Fraction(1))],
    )
    assert result.leasehold == {"Operator": Fraction(1)}
    assert result.working_interest == {"Operator": Fraction(1)}
    assert result.nri == {"Operator": Fraction(13, 16)}
    assert result.royalty_share == Fraction(3, 16)
    assert not result.blocking_defects


def test_partial_assignment_burden_and_tract_conservation() -> None:
    events = [
        event(1, "OPENING_LEASEHOLD", "L-1", "Operator", Fraction(1)),
        event(
            2,
            "ASSIGNMENT",
            "A-1",
            "Assignee",
            Fraction(1, 2),
            from_party="Operator",
            basis="OF_ASSIGNOR",
        ),
        event(
            3,
            "REVENUE_BURDEN",
            "ORRI-1",
            "ORRI Owner",
            Fraction(1, 16),
            from_party="Operator",
            basis="LEASE_8_8",
            burden_kind="ORRI",
        ),
    ]
    result = calculate_economics(
        unit_row(Fraction(1, 2), Fraction(1, 4)), events
    )
    assert result.leasehold == {
        "Operator": Fraction(1, 2),
        "Assignee": Fraction(1, 2),
    }
    assert result.nri == {
        "Operator": Fraction(5, 16),
        "Assignee": Fraction(3, 8),
        "ORRI Owner": Fraction(1, 16),
    }
    assert result.net_leasehold_acres["Operator"] == 40
    assert sum(result.tract_weighted_nri.values()) + result.royalty_share == Fraction(
        1, 2
    )
    assert not result.blocking_defects


def evidence_fields(asset_id: str, text: str) -> dict:
    return {
        "evidence_asset_version_id": asset_id,
        "evidence_char_start": 0,
        "evidence_char_end": len(text),
        "review_status": "REVIEWED",
    }


def economic_payload(asset_id: str, text: str) -> dict:
    evidence = evidence_fields(asset_id, text)
    return {
        "name": "Synthetic Economic Case",
        "legal_description": "SYNTHETIC TRACT",
        "gross_acres": {"numerator": 160, "denominator": 1},
        "opening_ownership": [
            {"owner_name": "Mineral Owner", "interest": {"numerator": 1, "denominator": 1}}
        ],
        "instruments": [],
        "economic_units": [
            {
                "unit_label": "Lease Unit A",
                "legal_description": "SYNTHETIC TRACT — ALL DEPTHS",
                "lease_recording_reference": "LEASE-100",
                "effective_as_of": "2026-01-01",
                "depth_scope": "ALL DEPTHS — OPERATOR REVIEWED",
                "product_scope": "OIL AND GAS",
                "gross_acres": {"numerator": 160, "denominator": 1},
                "leased_mineral_interest": {"numerator": 1, "denominator": 2},
                "base_royalty": {"numerator": 1, "denominator": 4},
                "wi_basis": "WI_EQUALS_LEASEHOLD",
                "unsupported_terms": [],
                **evidence,
                "events": [
                    {
                        "sequence_no": 1,
                        "event_kind": "OPENING_LEASEHOLD",
                        "recording_reference": "LEASE-100-OPEN",
                        "to_party": "Operator",
                        "interest": {"numerator": 1, "denominator": 1},
                        "interest_basis": "ABSOLUTE_UNIT",
                        **evidence,
                    },
                    {
                        "sequence_no": 2,
                        "event_kind": "ASSIGNMENT",
                        "recording_reference": "ASSIGN-200",
                        "from_party": "Operator",
                        "to_party": "Assignee",
                        "interest": {"numerator": 1, "denominator": 2},
                        "interest_basis": "OF_ASSIGNOR",
                        **evidence,
                    },
                ],
            }
        ],
    }


def test_economic_package_v2_outputs_exact_sheets_and_manifest(tmp_path: Path) -> None:
    source = tmp_path / "source"
    source.mkdir()
    text = (
        "SYNTHETIC REVIEWED LEASE FACTS. 160 gross acres; one-half leased "
        "mineral interest; one-fourth royalty; Operator assigns one-half to Assignee."
    )
    (source / "lease-facts.txt").write_text(text, encoding="utf-8")
    service = KernelService(config_for(tmp_path))
    project = service.create_project("Economic Package", source)
    service.ingest_project(project["id"])
    asset = service.list_assets(project["id"])[0]
    title_case = service.create_title_case(
        project["id"], economic_payload(asset["id"], text)
    )
    package = service.build_title_package(title_case["id"])
    assert package["blocking_defect_count"] == 0
    artifacts = {item["rel_path"]: item for item in package["artifacts"]}

    _, workbook_path = service.artifact(artifacts["Title_Examiner_Packet.xlsx"]["id"])
    with workbook_path.open("rb") as workbook_file:
        workbook = load_workbook(workbook_file, read_only=True, data_only=False)
        assert {
            "Calculation Units",
            "Reviewed Lease Facts",
            "Leasehold Ledger",
            "Leasehold and WI",
            "Net Revenue Detail",
            "Revenue Conservation",
        } <= set(workbook.sheetnames)
        revenue_rows = list(
            workbook["Revenue Conservation"].iter_rows(values_only=True)
        )
        workbook.close()
    assert ("Lease Unit A", "1/2", "1/8", "3/8", "1/2", 0) in revenue_rows

    _, manifest_path = service.artifact(artifacts["package_manifest.json"]["id"])
    manifest = json.loads(manifest_path.read_text("utf-8"))
    assert manifest["schema"] == "databossx.title-package.v2"
    assert manifest["blocking_defect_count"] == 0

    with service.db.transaction(immediate=True) as connection:
        connection.execute(
            """
            UPDATE text_extractions
               SET text_content='tampered economic evidence'
             WHERE asset_version_id=?
            """,
            (asset["id"],),
        )
    with pytest.raises(ValueError, match="hashes or source span changed"):
        service.build_title_package(title_case["id"])


def test_unsupported_term_and_over_assignment_block_package() -> None:
    unit = unit_row()
    unit["unsupported_terms"] = "Payout reversion requires examiner interpretation"
    result = calculate_economics(
        unit,
        [
            event(1, "OPENING_LEASEHOLD", "L-1", "Operator", Fraction(1)),
            event(
                2,
                "ASSIGNMENT",
                "A-1",
                "Assignee",
                Fraction(3, 2),
                from_party="Operator",
            ),
        ],
    )
    assert {defect.code for defect in result.blocking_defects} >= {
        "UNSUPPORTED_ECONOMIC_TERM",
        "LEASEHOLD_OVER_ASSIGNMENT",
    }


def test_mixed_evidence_snapshots_are_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source"
    source.mkdir()
    text = "SYNTHETIC REVIEWED LEASE FACTS one-half minerals and one-quarter royalty."
    (source / "lease.txt").write_text(text, encoding="utf-8")
    service = KernelService(config_for(tmp_path))
    project = service.create_project("Snapshot Gate", source)
    first_ingest = service.ingest_project(project["id"])
    first_asset = next(
        asset
        for asset in service.list_assets(project["id"])
        if asset["ingest_run_id"] == first_ingest["id"]
    )
    second_ingest = service.ingest_project(project["id"])
    second_asset = next(
        asset
        for asset in service.list_assets(project["id"])
        if asset["ingest_run_id"] == second_ingest["id"]
    )
    payload = economic_payload(second_asset["id"], text)
    payload["economic_units"][0]["events"][0].update(
        evidence_fields(first_asset["id"], text)
    )
    with pytest.raises(ValueError, match="one evidence snapshot"):
        service.create_title_case(project["id"], payload)
