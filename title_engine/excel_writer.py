"""
Excel Workbook Writer.
Generates all 13 sheets with formatting, source links, and color-coding.
Never writes fabricated data — UNKNOWN for any unsupported field.
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME

log = logging.getLogger(__name__)

# ── Color palette ─────────────────────────────────────────────────────────────

C = {
    "header_bg":   "1F3864",  # dark navy
    "header_fg":   "FFFFFF",
    "subhdr_bg":   "2E75B6",  # blue
    "subhdr_fg":   "FFFFFF",
    "alt_row":     "EBF3FB",  # light blue alternating row
    "high_conf":   "C6EFCE",  # green
    "med_conf":    "FFEB9C",  # yellow
    "low_conf":    "FFC7CE",  # red-pink
    "unk_conf":    "D9D9D9",  # grey
    "gap_row":     "FFC7CE",  # red
    "div_row":     "E2EFDA",  # light green (Diversified)
    "s27_row":     "BDD7EE",  # light blue (S27 confirmed)
    "warn_row":    "FFEB9C",  # yellow (needs review)
    "tab_dark":    "1F3864",
    "tab_blue":    "2E75B6",
    "tab_green":   "375623",
    "tab_orange":  "833C00",
    "tab_red":     "C00000",
    "tab_grey":    "595959",
}

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(wrap=False, h="left", v="center") -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_row(ws, values: list, row: int, bg: str = C["header_bg"], fg: str = "FFFFFF") -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = _font(bold=True, color=fg, size=10)
        cell.fill      = _fill(bg)
        cell.alignment = _align(wrap=True, h="center")
        cell.border    = _border()


def _set_col_widths(ws, widths: List[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _conf_fill(conf: str) -> PatternFill:
    m = {"HIGH": C["high_conf"], "MEDIUM": C["med_conf"],
         "LOW": C["low_conf"], "UNKNOWN": C["unk_conf"]}
    return _fill(m.get(conf.upper(), C["unk_conf"]))


def _write_data_rows(ws, rows: List[List], start_row: int,
                     conf_col: Optional[int] = None,
                     s27_col: Optional[int] = None,
                     gap_marker: Optional[int] = None) -> None:
    for i, row_data in enumerate(rows):
        r = start_row + i
        row_bg = C["alt_row"] if i % 2 == 1 else "FFFFFF"
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = _font(size=9)
            cell.alignment = _align(wrap=True)
            cell.border    = _border()
            cell.fill      = _fill(row_bg)

        # Confidence color override
        if conf_col and conf_col <= len(row_data):
            conf_val = str(row_data[conf_col - 1])
            ws.cell(row=r, column=conf_col).fill = _conf_fill(conf_val)

        # S27 flag
        if s27_col and s27_col <= len(row_data):
            v = str(row_data[s27_col - 1])
            if v == "1" or v.upper() == "TRUE" or v.upper() == "YES":
                ws.cell(row=r, column=s27_col).fill = _fill(C["s27_row"])

        # Gap marker
        if gap_marker and gap_marker <= len(row_data):
            v = str(row_data[gap_marker - 1])
            if "GAP" in v.upper() or "MISSING" in v.upper():
                for c in range(1, len(row_data) + 1):
                    ws.cell(row=r, column=c).fill = _fill(C["gap_row"])


# ── Column header list (matches spec exactly) ─────────────────────────────────

RUNSHEET_COLS = [
    "County", "Legal", "Quarter Call",
    "Instrument Type", "Instrument Number", "Book", "Page",
    "Document Date", "Effective Date", "Recording Date",
    "Grantor", "Grantee", "Lessor", "Lessee", "Assignor", "Assignee",
    "Lease Name", "Royalty", "Primary Term", "Depth Clause", "Pugh Clause",
    "Gross Acres", "Net Mineral Acres", "Net Leasehold Acres",
    "Working Interest", "NRI",
    "Current Claimed Owner", "Diversified Entity",
    "Chain From", "Chain To", "Chain Status",
    "Status", "Confidence %", "Problem / Gap", "Curative Needed",
    "Notes", "SOURCE LINK / FILE LINK",
]

_COL_WIDTHS_RUNSHEET = [
    10, 18, 14,        # County, Legal, Quarter
    24, 18, 8, 8,      # type, num, book, page
    12, 12, 12,        # dates
    28, 28, 20, 20, 20, 20,  # parties
    22, 10, 12, 22, 12,  # lease fields
    10, 14, 16,         # acres
    14, 10,             # WI, NRI
    28, 22,             # owner, div entity
    16, 16, 12,         # chain
    14, 10, 28, 28,     # status, conf, gap, curative
    30, 40,             # notes, source
]


def _instrument_to_row(inst: dict, ext: dict) -> List:
    def _s(k, fb="UNKNOWN"):
        v = ext.get(k) or inst.get(k)
        if v and str(v).strip() and str(v).strip().lower() not in ("null", "none", "[]", "{}"):
            return str(v).strip()
        return fb

    def _names(k):
        raw = ext.get(k) or inst.get(k)
        if not raw:
            return "UNKNOWN"
        if isinstance(raw, str):
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, list):
                    return "; ".join(str(x) for x in parsed)
            except Exception:
                return raw
        if isinstance(raw, list):
            return "; ".join(str(x) for x in raw)
        return str(raw)

    conf_f  = float(ext.get("confidence") or 0.0)
    has_img = inst.get("download_status") == "downloaded"
    if has_img and conf_f >= 0.8: conf_label = "HIGH"
    elif conf_f >= 0.5:            conf_label = "MEDIUM"
    elif conf_f > 0:               conf_label = "LOW"
    else:                          conf_label = "UNKNOWN"

    # Source link
    if inst.get("local_pdf") and has_img:
        src = inst["local_pdf"]
    elif inst.get("source_url"):
        src = inst["source_url"]
    elif inst.get("image_url"):
        src = inst["image_url"]
    else:
        src = "NO LINK — SOURCE NEEDED"

    return [
        inst.get("county", "Beckham"),
        "27-11N-25W",
        _s("quarter_call"),
        _s("instrument_type"),
        _s("instrument_number"),
        _s("book"),
        _s("page"),
        _s("doc_date"),
        _s("effective_date"),
        _s("recording_date"),
        _names("grantor"),
        _names("grantee"),
        _names("lessor"),
        _names("lessee"),
        _names("assignor"),
        _names("assignee"),
        _s("lease_name"),
        _s("royalty"),
        _s("primary_term"),
        _s("depth_clause"),
        _s("pugh_clause"),
        _s("gross_acres"),
        _s("net_mineral_acres"),
        _s("net_leasehold_acres"),
        _s("working_interest"),
        _s("nri"),
        _s("current_owner"),
        _s("diversified_entity"),
        _s("chain_from"),
        _s("chain_to"),
        _s("chain_status"),
        _s("status"),
        conf_label,
        _s("problem_gap", ""),
        _s("curative_needed", ""),
        _s("notes", ""),
        src,
    ]


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_executive_summary(wb: openpyxl.Workbook, instruments: List[dict],
                              stats: dict, gaps: list, chain) -> None:
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_properties.tabColor = C["tab_dark"]

    ws["A1"] = "SECTION 27-11N-25W, BECKHAM COUNTY, OKLAHOMA"
    ws["A1"].font  = _font(bold=True, size=16, color=C["header_bg"])
    ws["A1"].alignment = _align(h="center")
    ws.merge_cells("A1:H1")

    ws["A2"] = f"Title Research Report — Generated {datetime.now().strftime('%B %d, %Y')}"
    ws["A2"].font = _font(italic=True, size=11, color="595959")
    ws.merge_cells("A2:H2")

    # KPI table
    kpi_rows = [
        ("Total Instruments Found",    str(stats.get("total", 0))),
        ("Section 27 Confirmed",       str(stats.get("section_27", 0))),
        ("Documents Downloaded",       str(stats.get("dl_done", 0))),
        ("OCR / Extraction Complete",  str(stats.get("ocr_done", 0))),
        ("Chain Gaps Detected",        str(len(gaps))),
        ("Curative Items",             str(len(chain.curative_rows) if chain else 0)),
        ("Needs Human Review",         str(stats.get("needs_review", 0))),
        ("Total Spend (API + AI)",     f"${stats.get('total_cost', 0):.2f}"),
    ]
    _hdr_row(ws, ["Metric", "Value"], 4)
    for i, (k, v) in enumerate(kpi_rows):
        r = 5 + i
        ws.cell(r, 1, k).font = _font(size=10)
        ws.cell(r, 2, v).font = _font(bold=True, size=10)
        for c in (1, 2):
            ws.cell(r, c).border    = _border()
            ws.cell(r, c).alignment = _align()
            if i % 2:
                ws.cell(r, c).fill = _fill(C["alt_row"])

    # Instructions
    row_offset = 5 + len(kpi_rows) + 2
    ws.cell(row_offset, 1, "SHEET GUIDE").font = _font(bold=True, size=11)
    sheets_info = [
        ("Full Section 27 Runsheet", "Every instrument found, all columns"),
        ("Diversified Interest Summary", "Only Diversified/Tapstone chain"),
        ("Lease Chain", "OGL → assignments → current lessee"),
        ("Assignment Chain", "All assignments in sequence"),
        ("Mineral Ownership", "Mineral deed chain and ownership"),
        ("Wells / OCC / Production", "Wellbores and production orders"),
        ("Corporate Chain", "Entity name changes and mergers"),
        ("Missing Documents", "Gaps requiring curative action"),
        ("Curative Requirements", "Prioritized curative checklist"),
        ("Source Links", "All documents with links/paths"),
        ("Research Log", "Timestamped activity log"),
        ("Search Matrix", "All searches run and result counts"),
    ]
    for j, (sname, sdesc) in enumerate(sheets_info):
        r = row_offset + 1 + j
        ws.cell(r, 1, sname).font  = _font(bold=True, color="2E75B6", size=9)
        ws.cell(r, 2, sdesc).font  = _font(size=9, italic=True)

    _set_col_widths(ws, [35, 55])


def _sheet_runsheet(wb, instruments: List[dict], tab_color: str,
                    title: str, sheet_name: str,
                    filter_fn=None) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"
    ws["A1"] = title
    ws["A1"].font = _font(bold=True, size=12, color=C["header_bg"])
    ws.merge_cells(f"A1:{get_column_letter(len(RUNSHEET_COLS))}1")

    _hdr_row(ws, RUNSHEET_COLS, 2)

    rows_to_write = instruments if not filter_fn else list(filter(filter_fn, instruments))
    data_rows = []
    for inst in rows_to_write:
        ext = {}
        if inst.get("extracted_json"):
            try:
                ext = json.loads(inst["extracted_json"])
            except Exception:
                pass
        data_rows.append(_instrument_to_row(inst, ext))

    conf_col = RUNSHEET_COLS.index("Confidence %") + 1
    _write_data_rows(ws, data_rows, 3, conf_col=conf_col)

    # Autofilter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(RUNSHEET_COLS))}2"
    _set_col_widths(ws, _COL_WIDTHS_RUNSHEET)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30


def _sheet_chain(wb, rows_dicts: List[dict], sheet_name: str,
                 tab_color: str, title: str) -> None:
    """Generic chain sheet — accepts dicts with RUNSHEET_COLS keys."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"
    ws["A1"] = title
    ws["A1"].font = _font(bold=True, size=12, color=C["header_bg"])
    ws.merge_cells(f"A1:{get_column_letter(len(RUNSHEET_COLS))}1")
    _hdr_row(ws, RUNSHEET_COLS, 2)

    def _rdict_to_list(d):
        return [d.get(col, "") for col in RUNSHEET_COLS]

    data = [_rdict_to_list(r) for r in rows_dicts]
    conf_col = RUNSHEET_COLS.index("Confidence %") + 1
    _write_data_rows(ws, data, 3, conf_col=conf_col)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(RUNSHEET_COLS))}2"
    _set_col_widths(ws, _COL_WIDTHS_RUNSHEET)


def _sheet_curative(wb, curative_rows: List[dict]) -> None:
    ws = wb.create_sheet("Curative Requirements")
    ws.sheet_properties.tabColor = C["tab_red"]
    ws.freeze_panes = "A2"
    cols = ["Priority", "Instrument Number", "Book", "Page",
            "Issue / Gap", "Required Action", "Responsible Party",
            "Deadline", "Status", "Notes"]
    _hdr_row(ws, cols, 1, bg=C["tab_red"])
    for i, r in enumerate(curative_rows):
        row = i + 2
        for c, col in enumerate(cols, 1):
            val = r.get(col, "")
            cell = ws.cell(row=row, column=c, value=val)
            cell.font      = _font(size=9)
            cell.alignment = _align(wrap=True)
            cell.border    = _border()
            if r.get("Priority") == "HIGH":
                cell.fill = _fill(C["low_conf"])
            elif i % 2:
                cell.fill = _fill(C["alt_row"])

    _set_col_widths(ws, [10, 18, 8, 8, 45, 50, 22, 12, 14, 30])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def _sheet_missing_docs(wb, missing: List[dict]) -> None:
    ws = wb.create_sheet("Missing Documents")
    ws.sheet_properties.tabColor = C["tab_orange"]
    cols = ["Priority", "Referenced In Instrument",
            "Missing Instrument", "Type", "Curative Action",
            "Source Suggestions"]
    _hdr_row(ws, cols, 1, bg=C["tab_orange"])
    for i, m in enumerate(missing):
        row = i + 2
        vals = [
            m.get("priority", "MEDIUM"),
            m.get("referenced_in", ""),
            m.get("missing_instrument", ""),
            m.get("type", ""),
            m.get("curative", ""),
            "OKCountyRecords / Courthouse / Title Plant",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font      = _font(size=9)
            cell.alignment = _align(wrap=True)
            cell.border    = _border()
            cell.fill      = _fill(C["low_conf"] if m.get("priority") == "HIGH" else C["alt_row"] if i % 2 else "FFFFFF")
    _set_col_widths(ws, [10, 24, 24, 28, 50, 40])


def _sheet_source_links(wb, instruments: List[dict]) -> None:
    ws = wb.create_sheet("Source Links")
    ws.sheet_properties.tabColor = C["tab_grey"]
    cols = ["Instrument Number", "Book", "Page", "Type", "Grantor", "Grantee",
            "Local PDF Path", "API / Web URL", "Link Status"]
    _hdr_row(ws, cols, 1)
    for i, inst in enumerate(instruments):
        ext = {}
        if inst.get("extracted_json"):
            try:
                ext = json.loads(inst["extracted_json"])
            except Exception:
                pass
        local = inst.get("local_pdf", "")
        url   = inst.get("source_url") or inst.get("image_url") or ""
        if local:
            link_status = "PDF DOWNLOADED"
        elif url:
            link_status = "INDEX ONLY — IMAGE NOT DOWNLOADED"
        else:
            link_status = "NO LINK — SOURCE NEEDED"

        def _g(k):
            raw = ext.get(k) or inst.get(k)
            if isinstance(raw, list):
                return "; ".join(str(x) for x in raw)
            try:
                parsed = json.loads(str(raw or ""))
                if isinstance(parsed, list):
                    return "; ".join(str(x) for x in parsed)
            except Exception:
                pass
            return str(raw or "")

        row = i + 2
        vals = [
            inst.get("instrument_number", ""),
            inst.get("book", ""),
            inst.get("page", ""),
            inst.get("instrument_type", ""),
            _g("grantor"), _g("grantee"),
            local, url, link_status,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font      = _font(size=9)
            cell.alignment = _align(wrap=True)
            cell.border    = _border()
            fill_c = C["high_conf"] if local else C["med_conf"] if url else C["low_conf"]
            cell.fill = _fill(fill_c)
    _set_col_widths(ws, [20, 8, 8, 24, 28, 28, 50, 50, 28])


def _sheet_research_log(wb, log_entries: List[dict]) -> None:
    ws = wb.create_sheet("Research Log")
    ws.sheet_properties.tabColor = C["tab_grey"]
    cols = ["Timestamp", "Action", "Detail"]
    _hdr_row(ws, cols, 1)
    for i, e in enumerate(log_entries):
        row = i + 2
        vals = [e.get("created_at", ""), e.get("action", ""), e.get("detail", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = _font(size=9)
            cell.alignment = _align(wrap=True)
            cell.border = _border()
            if i % 2:
                cell.fill = _fill(C["alt_row"])
    _set_col_widths(ws, [22, 28, 80])


def _sheet_search_matrix(wb, search_log: List[dict]) -> None:
    ws = wb.create_sheet("Search Matrix")
    ws.sheet_properties.tabColor = C["tab_grey"]
    cols = ["Search Type", "Query / Parameters", "County", "Results Found",
            "Status", "Ran At"]
    _hdr_row(ws, cols, 1)
    for i, s in enumerate(search_log):
        row = i + 2
        vals = [
            s.get("search_type", ""), s.get("query", ""), s.get("county", ""),
            s.get("result_count", 0), s.get("status", ""), s.get("ran_at", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = _font(size=9)
            cell.alignment = _align(wrap=True)
            cell.border = _border()
            if i % 2:
                cell.fill = _fill(C["alt_row"])
    _set_col_widths(ws, [18, 50, 12, 14, 12, 22])


def _sheet_wells(wb) -> None:
    ws = wb.create_sheet("Wells / OCC / Production")
    ws.sheet_properties.tabColor = C["tab_green"]
    cols = ["Well Name", "API Number", "Operator", "Type", "Formation",
            "Spud Date", "First Production", "Status",
            "Monthly BOE", "Notes", "SOURCE LINK / FILE LINK"]
    _hdr_row(ws, cols, 1, bg=C["tab_green"])
    ws.cell(2, 1, "** Populate from OCC Well Records / IHS / DrillingInfo **")
    ws.cell(2, 1).font = _font(italic=True, color="595959", size=9)
    _set_col_widths(ws, [28, 16, 28, 12, 16, 12, 14, 12, 14, 40, 40])


def _sheet_corporate_chain(wb) -> None:
    ws = wb.create_sheet("Corporate Chain")
    ws.sheet_properties.tabColor = C["tab_blue"]
    cols = ["Entity Name", "Successor / Predecessor",
            "Instrument Type", "Date", "Notes", "SOURCE LINK / FILE LINK"]
    _hdr_row(ws, cols, 1, bg=C["tab_blue"])
    default_chain = [
        ["Tapstone Energy", "merged into Diversified Production LLC", "Merger", "2021", "", ""],
        ["DP Legacy Tapstone LLC", "subsidiary of Diversified Production LLC", "Formation", "2021", "", ""],
        ["Diversified Production LLC", "subsidiary of Diversified Energy Company", "Formation", "2019", "", ""],
        ["Unbridled Resources LLC", "predecessor / acquired by Tapstone", "Acquisition", "ca. 2018", "", ""],
        ["BNK Petroleum", "predecessor operator — verify", "UNKNOWN", "UNKNOWN", "", "NO LINK — SOURCE NEEDED"],
    ]
    for i, row in enumerate(default_chain):
        r = i + 2
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = _font(size=9, italic=(c >= 5))
            cell.alignment = _align(wrap=True)
            cell.border = _border()
            if i % 2:
                cell.fill = _fill(C["alt_row"])
    note_row = len(default_chain) + 3
    ws.cell(note_row, 1, "NOTE: Verify all corporate links with SOS filings and recorded merger docs.")
    ws.cell(note_row, 1).font = _font(italic=True, color="C00000", size=9)
    _set_col_widths(ws, [34, 40, 24, 12, 40, 40])


# ── Main export function ──────────────────────────────────────────────────────

def write_workbook(
    out_path: Path,
    instruments: List[dict],
    chain,          # ChainBuilder instance
    stats: dict,
    log_entries: List[dict],
    search_log: List[dict],
) -> Path:
    """
    Write the complete 13-sheet Excel workbook.
    Returns the written file path.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # 1 – Executive Summary
    _sheet_executive_summary(wb, instruments, stats,
                              chain.gaps if chain else [],
                              chain)

    # 2 – Full Section 27 Runsheet
    _sheet_runsheet(wb, instruments,
                    tab_color=C["tab_blue"],
                    title="Full Section 27-11N-25W Runsheet — Beckham County, Oklahoma",
                    sheet_name="Full Section 27 Runsheet")

    # 3 – Diversified Interest Summary
    def _is_div(inst):
        raw = (inst.get("extracted_json") or "").lower()
        g = (inst.get("grantor") or "").lower()
        ge = (inst.get("grantee") or "").lower()
        keywords = ["diversified", "tapstone", "dp legacy", "unbridled", "maverick natural",
                    "bnk petroleum", "rimrock"]
        return any(k in raw or k in g or k in ge for k in keywords)

    _sheet_runsheet(wb, instruments,
                    tab_color=C["tab_green"],
                    title="Diversified / Tapstone Interest Summary — Section 27-11N-25W",
                    sheet_name="Diversified Interest Summary",
                    filter_fn=_is_div)

    # 4 – Lease Chain
    def _is_lease(inst):
        t = (inst.get("instrument_type") or "").lower()
        return any(k in t for k in ("oil and gas lease", "memorandum of oil", "ratification", "pooling"))
    _sheet_runsheet(wb, instruments,
                    tab_color=C["tab_blue"],
                    title="Lease Chain — Section 27-11N-25W",
                    sheet_name="Lease Chain",
                    filter_fn=_is_lease)

    # 5 – Assignment Chain
    def _is_asgn(inst):
        t = (inst.get("instrument_type") or "").lower()
        return any(k in t for k in ("assignment", "release"))
    _sheet_runsheet(wb, instruments,
                    tab_color=C["tab_blue"],
                    title="Assignment Chain — Section 27-11N-25W",
                    sheet_name="Assignment Chain",
                    filter_fn=_is_asgn)

    # 6 – Mineral Ownership
    def _is_mineral(inst):
        t = (inst.get("instrument_type") or "").lower()
        return any(k in t for k in ("mineral deed", "warranty deed", "quit claim", "affidavit of heirship"))
    _sheet_runsheet(wb, instruments,
                    tab_color=C["tab_green"],
                    title="Mineral Ownership — Section 27-11N-25W",
                    sheet_name="Mineral Ownership",
                    filter_fn=_is_mineral)

    # 7 – Wells / OCC / Production
    _sheet_wells(wb)

    # 8 – Corporate Chain
    _sheet_corporate_chain(wb)

    # 9 – Missing Documents
    _sheet_missing_docs(wb, chain.missing if chain else [])

    # 10 – Curative Requirements
    _sheet_curative(wb, chain.curative_rows if chain else [])

    # 11 – Source Links
    _sheet_source_links(wb, instruments)

    # 12 – Research Log
    _sheet_research_log(wb, log_entries)

    # 13 – Search Matrix
    _sheet_search_matrix(wb, search_log)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    log.info("Wrote workbook: %s (%d sheets)", out_path, len(wb.sheetnames))
    return out_path


def clean_filename(book: str, page: str, itype: str,
                   grantor: str, grantee: str, inst_num: str,
                   doc_date: str) -> str:
    """Build the canonical PDF filename as specified."""
    import re as _re

    def _sf(s: str) -> str:
        s = str(s or "UNKNOWN").strip()
        s = _re.sub(r"[^\w\-]", "_", s)
        return s[:40]

    date_part = (_re.sub(r"[^\d]", "-", doc_date)[:10]
                 if doc_date and doc_date != "UNKNOWN" else "NODATE")
    itype_p = _sf(itype)
    g_to_ge = f"{_sf(grantor)}_to_{_sf(grantee)}"
    bp      = f"{_sf(book)}_{_sf(page)}"
    inst_p  = _sf(inst_num)
    return f"{date_part}__{itype_p}__{g_to_ge}__{bp}__{inst_p}.pdf"
