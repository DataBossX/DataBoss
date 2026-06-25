"""Orchestrator / CLI for the cursory title report generator.

Runs the full pipeline against a :class:`TitleReport` and writes every
deliverable into an output directory:

  * <Tract>_Cursory_Title_Report_<date>.xlsx
  * SOURCE_LOG.csv, CURATIVE_LIST.csv, MISSING_OR_UNVERIFIED_INSTRUMENTS.csv
  * QA_RESULTS.md, CHANGE_SUMMARY.md

Usage:
    python -m title_report.generate [--out DIR] [--date YYYY-MM-DD]

With no arguments it runs the bundled placeholder dataset for 31-11N-24W so the
pipeline is demonstrable without live source access.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from .models import TitleReport
from .sample_data import build_sample_report
from .specimen_31_12N_24W import build_report as build_31_12n_24w
from . import workbook as wbmod
from . import deliverables as dl
from . import qa as qamod

# Tract registry: slug -> builder(report_date) -> TitleReport.
TRACTS = {
    "31-12N-24W": build_31_12n_24w,   # fully-populated specimen (default)
    "31-11N-24W": build_sample_report,
}
DEFAULT_TRACT = "31-12N-24W"


def _slug(rpt: TitleReport) -> str:
    cfg = rpt.config
    county = cfg.county.replace(" ", "_")
    return f"{cfg.section}-{cfg.township}-{cfg.range_}_{county}"


_FOOTER_MARKERS = ("TOTALS", "Live formulas", "Derived columns", "(")


def _workbook_counts(path: str) -> dict:
    """Read back per-sheet data-row counts for reconciliation.

    Workbook layout: row 1 = nav, row 2 = header, data from row 3. Footer/notes
    (e.g. the matrix TOTALS row) are excluded via marker tokens in column A.
    """
    wb = load_workbook(path, read_only=True)
    counts = {}
    for name in ("Source Log", "Runsheet", "Abstractions", "Wells & HBP",
                 "Curative", "NRI-WI Matrix"):
        if name not in wb.sheetnames:
            counts[name] = None
            continue
        ws = wb[name]
        n = 0
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r <= 2:                        # nav + header
                continue
            first = row[0]
            if first is None or str(first).strip() == "":
                continue
            if str(first).startswith(_FOOTER_MARKERS):
                continue
            n += 1
        counts[name] = n
    wb.close()
    return counts


def build_tract(tract: str = DEFAULT_TRACT, report_date: Optional[str] = None) -> TitleReport:
    """Build the TitleReport for a known tract slug."""
    if tract not in TRACTS:
        raise KeyError(f"unknown tract {tract!r}; known: {', '.join(TRACTS)}")
    builder = TRACTS[tract]
    return builder(report_date) if report_date else builder()


def generate(rpt: Optional[TitleReport] = None, out_dir: str = "title_report_output",
             run_date: Optional[str] = None) -> dict:
    rpt = rpt or build_tract()
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    run_date = run_date or rpt.config.report_date

    # 1. CSV / MD deliverables (written first so the secrets scan can cover them)
    src_csv = dl.write_source_log(rpt, out)
    cur_csv = dl.write_curative_list(rpt, out)
    mis_csv = dl.write_missing(rpt, out)
    missing = dl.compute_missing(rpt)
    change_md = dl.write_change_summary(
        rpt, out, run_date, prior_exists=(out / "CHANGE_SUMMARY.md").exists())

    # 2. Workbook (first pass, so we can read back counts for reconciliation)
    xlsx_path = str(out / f"{_slug(rpt)}_Cursory_Title_Report_{run_date}.xlsx")
    wbmod.build_workbook(rpt, xlsx_path, qa_summary=None, missing=missing)
    counts = _workbook_counts(xlsx_path)

    # 3. QA over the model + written deliverables
    deliverable_paths = [Path(p) for p in (src_csv, cur_csv, mis_csv, change_md)]
    qa_summary = qamod.run_qa(rpt, counts, deliverable_paths, out)

    # 4. Rebuild workbook with the QA dashboard populated
    wbmod.build_workbook(rpt, xlsx_path, qa_summary=qa_summary, missing=missing)

    return {
        "workbook": xlsx_path,
        "source_log": src_csv,
        "curative_list": cur_csv,
        "missing": mis_csv,
        "change_summary": change_md,
        "qa_results": str(out / "QA_RESULTS.md"),
        "qa_passed": qa_summary["passed"],
        "is_placeholder": rpt.config.is_placeholder,
        "data_status": rpt.config.data_status,
        "counts": counts,
    }


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Generate a cursory title report.")
    ap.add_argument("--out", default="title_report_output", help="output directory")
    ap.add_argument("--date", default=None, help="report/run date YYYY-MM-DD")
    ap.add_argument("--tract", default=DEFAULT_TRACT, choices=list(TRACTS),
                    help="tract to generate")
    args = ap.parse_args(argv)

    rpt = build_tract(args.tract, args.date)
    result = generate(rpt, out_dir=args.out, run_date=args.date or rpt.config.report_date)

    print(f"Workbook:       {result['workbook']}")
    print(f"Source log:     {result['source_log']}")
    print(f"Curative list:  {result['curative_list']}")
    print(f"Missing list:   {result['missing']}")
    print(f"Change summary: {result['change_summary']}")
    print(f"QA results:     {result['qa_results']}")
    status = "PASS" if result["qa_passed"] else "FAIL"
    final = (f"NOT FINAL ({result['data_status']} data)"
             if result["data_status"] != "EXAMINED" else
             ("FINAL-eligible" if result["qa_passed"] else "NOT FINAL (QA failed)"))
    print(f"QA: {status}  |  {final}  |  counts={result['counts']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
