"""End-to-end and unit tests for the title report generator."""

from __future__ import annotations

from fractions import Fraction
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fractions import Fraction
from title_report import fractions as fr
from title_report.models import UNKNOWN, OwnershipEntry, ProjectConfig, TitleReport
from title_report.sample_data import build_sample_report
from title_report.specimen_31_12N_24W import build_report as build_specimen
from title_report.deliverables import compute_missing
from title_report.generate import generate, build_tract, TRACTS

REQUIRED_SHEETS = {
    "Cover", "Contents", "Summary", "Source Log", "Runsheet", "Abstractions",
    "Chain - Mineral & Surface", "Chain - Leasehold & WI", "Wells & HBP",
    "Curative", "NRI-WI Matrix", "QA Dashboard", "Missing-Unverified",
}


def _generate_specimen(tmp_path):
    return generate(build_specimen(), out_dir=str(tmp_path))


# ── fraction math ────────────────────────────────────────────────────────────

def test_royalty_nri_exact():
    assert fr.net_revenue_interest(Fraction(1), Fraction(3, 16)) == Fraction(3, 16)
    assert fr.net_revenue_interest(Fraction(1, 2), Fraction(3, 16)) == Fraction(3, 32)


def test_wi_nri_exact():
    assert fr.wi_net_revenue_interest(Fraction(1), Fraction(3, 16)) == Fraction(13, 16)


def test_unknown_inputs_stay_unknown():
    assert fr.net_revenue_interest(UNKNOWN, Fraction(3, 16)) == UNKNOWN
    assert fr.wi_net_revenue_interest(Fraction(1), UNKNOWN) == UNKNOWN
    assert fr.net_mineral_acres(UNKNOWN, Fraction(640)) == UNKNOWN


def test_parse_and_render():
    assert fr.parse_fraction("3/16") == Fraction(3, 16)
    assert fr.parse_fraction("0.1875") == Fraction(3, 16)
    assert fr.parse_fraction("") is None
    assert fr.frac_str(Fraction(3, 16)) == "3/16"
    assert fr.frac_str(UNKNOWN) == UNKNOWN


# ── pipeline ─────────────────────────────────────────────────────────────────

def test_generate_writes_all_deliverables(tmp_path):
    result = generate(out_dir=str(tmp_path))
    for key in ("workbook", "source_log", "curative_list", "missing",
                "change_summary", "qa_results"):
        assert Path(result[key]).exists(), f"missing deliverable: {key}"


def test_workbook_has_all_sheets(tmp_path):
    result = generate(out_dir=str(tmp_path))
    wb = load_workbook(result["workbook"])
    assert REQUIRED_SHEETS.issubset(set(wb.sheetnames))


def test_nri_matrix_uses_live_formulas(tmp_path):
    result = generate(out_dir=str(tmp_path))
    wb = load_workbook(result["workbook"])
    ws = wb["NRI-WI Matrix"]
    formulas = [c.value for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    assert formulas, "expected at least one live Excel formula in NRI-WI matrix"


def test_row_reconciliation_matches_model(tmp_path):
    rpt = build_specimen()
    result = generate(rpt, out_dir=str(tmp_path))
    counts = result["counts"]
    assert counts["Runsheet"] == len(rpt.instruments)
    assert counts["Curative"] == len(rpt.curative)
    assert counts["NRI-WI Matrix"] == len(rpt.ownership)
    assert counts["Source Log"] == len(rpt.sources)


def test_qa_runs_and_reports(tmp_path):
    result = generate(out_dir=str(tmp_path))
    qa_text = Path(result["qa_results"]).read_text()
    assert "Citation coverage" in qa_text
    assert "Fraction math" in qa_text
    assert "Secrets scan" in qa_text


def test_placeholder_blocks_final(tmp_path):
    result = generate(out_dir=str(tmp_path))
    assert result["is_placeholder"] is True
    qa_text = Path(result["qa_results"]).read_text()
    assert "NOT FINAL" in qa_text


def test_computed_nri_requires_basis_language():
    """A computed interest without basis language must be caught by QA."""
    cfg = ProjectConfig(section="1", township="1N", range_="1W", county="Test")
    bad = OwnershipEntry(owner="X", role="royalty",
                         mineral_interest=Fraction(1), lease_royalty=Fraction(1, 8),
                         basis_language=UNKNOWN, citation="SRC-1")
    rpt = TitleReport(config=cfg, ownership=[bad])
    from title_report.qa import _fraction_math
    chk = _fraction_math(rpt)
    assert chk["status"] == "FAIL"
    assert "basis language" in chk["detail"]


def test_secrets_scan_flags_key(tmp_path):
    from title_report.qa import _secrets_scan
    p = tmp_path / "leak.txt"
    p.write_text("api_key = sk-ABCDEFGHIJKLMNOPQRSTUVWX")
    chk = _secrets_scan([p])
    assert chk["status"] == "FAIL"


def test_missing_flags_uncited_and_placeholder():
    rpt = build_sample_report()
    missing = compute_missing(rpt)
    types = {m["type"] for m in missing}
    assert "Placeholder run" in types


# ── 31-12N-24W specimen ──────────────────────────────────────────────────────

def test_orri_nri_dispatch():
    assert fr.entry_nri("overriding_royalty", UNKNOWN, UNKNOWN, UNKNOWN, UNKNOWN,
                        Fraction(1, 32)) == Fraction(1, 32)
    assert fr.entry_nri("working_interest", UNKNOWN, UNKNOWN, Fraction(1),
                        Fraction(7, 32), UNKNOWN) == Fraction(25, 32)


def test_specimen_minerals_and_nri_reconcile_to_one():
    rpt = build_specimen()
    minerals = [o.mineral_interest for o in rpt.ownership if o.role == "royalty"]
    assert fr.total_known(minerals) == Fraction(1)
    nris = [fr.entry_nri(o.role, o.mineral_interest, o.lease_royalty,
                         o.working_interest, o.burdens, o.override_royalty)
            for o in rpt.ownership]
    assert fr.total_known(nris) == Fraction(1)


def test_specimen_qa_all_pass(tmp_path):
    result = _generate_specimen(tmp_path)
    assert result["qa_passed"] is True
    assert result["data_status"] == "SPECIMEN"


def test_specimen_every_runsheet_cell_filled(tmp_path):
    """The specimen leaves no UNKNOWN in the runsheet (fully populated)."""
    result = _generate_specimen(tmp_path)
    wb = load_workbook(result["workbook"])
    ws = wb["Runsheet"]
    blanks = [(c.row, c.column) for row in ws.iter_rows(min_row=3) for c in row
              if c.value == UNKNOWN]
    assert not blanks, f"unexpected UNKNOWN cells in specimen runsheet: {blanks}"


def test_specimen_workbook_has_working_links(tmp_path):
    result = _generate_specimen(tmp_path)
    wb = load_workbook(result["workbook"])
    ext = internal = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.hyperlink:
                    tgt = c.hyperlink.target or c.hyperlink.location or ""
                    if str(tgt).startswith("http"):
                        ext += 1
                    else:
                        internal += 1
    assert ext >= 10, f"expected external portal links, got {ext}"
    assert internal >= 20, f"expected internal cross-links, got {internal}"


def test_contents_links_resolve_to_real_sheets(tmp_path):
    result = _generate_specimen(tmp_path)
    wb = load_workbook(result["workbook"])
    names = set(wb.sheetnames)
    ws = wb["Contents"]
    targets = []
    for row in ws.iter_rows():
        for c in row:
            link = c.hyperlink
            if not link:
                continue
            loc = link.location or link.target or ""
            # internal link form: "#'Sheet'!A1" or "'Sheet'!A1"
            if "!" in loc:
                sheet = loc.split("!")[0].lstrip("#").strip("'")
                targets.append(sheet)
    assert targets, "Contents has no internal links"
    for t in targets:
        assert t in names, f"Contents links to missing sheet {t!r}"


def test_build_tract_registry():
    assert "31-12N-24W" in TRACTS and "31-11N-24W" in TRACTS
    rpt = build_tract("31-12N-24W", "2026-06-25")
    assert rpt.config.report_date == "2026-06-25"
    assert rpt.config.section == "31" and rpt.config.township == "12N"


def test_html_dashboard_generated(tmp_path):
    result = _generate_specimen(tmp_path)
    html_path = Path(result["html"])
    assert html_path.exists()
    text = html_path.read_text(encoding="utf-8")
    assert "<html" in text.lower()
    assert "RECONCILES 8/8" in text          # specimen reconciles
    assert "31-12N-24W" in text


# ── DOTO adapter ─────────────────────────────────────────────────────────────

def _make_doto_db(path):
    import sqlite3
    con = sqlite3.connect(path)
    con.executescript("""
      CREATE TABLE image_queue (id INTEGER PRIMARY KEY AUTOINCREMENT, dedup_key TEXT,
        county TEXT, book TEXT, page TEXT, instrument_number TEXT, section TEXT,
        township TEXT, range_val TEXT, status TEXT);
      CREATE TABLE downloads (id INTEGER PRIMARY KEY AUTOINCREMENT, queue_item_id INTEGER,
        file_path TEXT, status TEXT);
      CREATE TABLE analyses (id INTEGER PRIMARY KEY AUTOINCREMENT, download_id INTEGER,
        queue_item_id INTEGER, instrument_type TEXT, instrument_number TEXT, book TEXT,
        page TEXT, recording_date TEXT, instrument_date TEXT, grantors TEXT, grantees TEXT,
        legal_description TEXT, section TEXT, township TEXT, range_val TEXT,
        interest_conveyed TEXT, lease_royalty_terms TEXT, title_requirement_affected TEXT,
        confidence_score REAL, needs_review INTEGER);
    """)
    con.execute("INSERT INTO image_queue (dedup_key,county,book,page,instrument_number,"
                "section,township,range_val,status) VALUES "
                "('k1','Roger Mills','100','5','I-100','31','12N','24W','analyzed')")
    qid = con.execute("SELECT id FROM image_queue").fetchone()[0]
    con.execute("INSERT INTO downloads (queue_item_id,file_path,status) VALUES (?,?,?)",
                (qid, "/x/1.pdf", "success"))
    did = con.execute("SELECT id FROM downloads").fetchone()[0]
    con.execute(
        "INSERT INTO analyses (download_id,queue_item_id,instrument_type,instrument_number,"
        "book,page,recording_date,grantors,grantees,legal_description,section,township,"
        "range_val,interest_conveyed,lease_royalty_terms,confidence_score,needs_review) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (did, qid, "Warranty Deed", "I-100", "100", "5", "1990-02-01",
         '["A Smith"]', '["B Jones"]', "All of Sec 31-12N-24W", "31", "12N", "24W",
         "Surface + minerals", "3/16", 0.92, 0))
    con.commit()
    con.close()


def test_doto_available_tracts(tmp_path):
    from title_report.adapters.doto import available_tracts
    db = tmp_path / "doto.db"
    _make_doto_db(str(db))
    tracts = available_tracts(str(db))
    assert len(tracts) == 1
    assert tracts[0]["slug"] == "31-12N-24W"
    assert tracts[0]["county"] == "Roger Mills"
    assert available_tracts(str(tmp_path / "missing.db")) == []


def test_doto_build_report_examined(tmp_path):
    from title_report.adapters.doto import build_report_from_db
    db = tmp_path / "doto.db"
    _make_doto_db(str(db))
    rpt = build_report_from_db(str(db), "31", "12N", "24W",
                               report_date="2026-06-25", source_cutoff_date="2026-06-24")
    assert rpt.config.data_status == "EXAMINED"
    assert len(rpt.instruments) == 1
    assert len(rpt.chain_links) == 1
    assert rpt.instruments[0].grantors == ["A Smith"]
    # no invented ownership / wells
    assert rpt.ownership == [] and rpt.wells == []


def test_doto_report_passes_qa_end_to_end(tmp_path):
    from title_report.adapters.doto import build_report_from_db
    db = tmp_path / "doto.db"
    _make_doto_db(str(db))
    rpt = build_report_from_db(str(db), "31", "12N", "24W",
                               report_date="2026-06-25", source_cutoff_date="2026-06-24")
    result = generate(rpt, out_dir=str(tmp_path / "out"))
    assert result["qa_passed"] is True
    assert result["data_status"] == "EXAMINED"
    assert result["counts"]["Runsheet"] == 1
