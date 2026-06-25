"""Excel workbook generation for the cursory title report.

Produces one polished, fully-linked ``.xlsx``:

  * Cover + a Contents tab whose rows hyperlink to every sheet, and a Summary
    dashboard with live ownership reconciliation.
  * Every data tab carries an auto-filter, frozen header, and a "Contents"
    back-link.
  * Working hyperlinks throughout: citations link to their Source Log row;
    instrument references link to their Runsheet row; source locations, recorded
    images, and well records link out to the live public portals.
  * The NRI/WI matrix uses live, fraction-formatted Excel formulas (royalty,
    working-interest, and overriding-royalty roles) plus a reconciliation row
    that proves the interests sum to 8/8.
"""

from __future__ import annotations

import re
from fractions import Fraction
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import TitleReport, UNKNOWN
from . import fractions as fr

# ── external portals for out-links ───────────────────────────────────────────
OCC_WELL_URL = "https://gisdata-occokc.opendata.arcgis.com/"
OCC_IMG_URL = "http://imaging.occ.ok.gov/"

# ── sheet names (ASCII for reliable internal hyperlinks) ─────────────────────
S_COVER = "Cover"
S_CONTENTS = "Contents"
S_SUMMARY = "Summary"
S_SOURCES = "Source Log"
S_RUNSHEET = "Runsheet"
S_ABSTRACT = "Abstractions"
S_CHAIN_MIN = "Chain - Mineral & Surface"
S_CHAIN_LEASE = "Chain - Leasehold & WI"
S_WELLS = "Wells & HBP"
S_CURATIVE = "Curative"
S_MATRIX = "NRI-WI Matrix"
S_QA = "QA Dashboard"
S_MISSING = "Missing-Unverified"

SHEET_GUIDE = [
    (S_SUMMARY, "Dashboard: tract facts, schedule counts, ownership reconciliation, HBP."),
    (S_SOURCES, "Every source consulted, with live portal links and page refs."),
    (S_RUNSHEET, "Chronological instrument index with linked recorded images."),
    (S_ABSTRACT, "Clause-level abstraction of each material instrument."),
    (S_CHAIN_MIN, "Mineral & surface chain of title, root patent to present."),
    (S_CHAIN_LEASE, "Leasehold / working-interest / burdens chain."),
    (S_WELLS, "Well & held-by-production schedule (spacing, pooling, status)."),
    (S_CURATIVE, "Encumbrances & curative items by materiality and priority."),
    (S_MATRIX, "Net revenue / working-interest matrix with live formulas."),
    (S_QA, "QA gate results for this generation."),
    (S_MISSING, "Missing or unverified items flagged for follow-up."),
]

# ── styling ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=18, bold=True, color="1F4E79")
SUB_FONT = Font(size=11, italic=True, color="595959")
BANNER_FILL = PatternFill("solid", fgColor="FFF2CC")
BANNER_FONT = Font(bold=True, color="7F6000")
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FONT = Font(bold=True, color="9C0006")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
REVIEW_FILL = PatternFill("solid", fgColor="FFE699")
HIGH_FILL = PatternFill("solid", fgColor="FF9999")
MED_FILL = PatternFill("solid", fgColor="FFD580")
LOW_FILL = PatternFill("solid", fgColor="C6EFCE")
LINK_FONT = Font(color="0563C1", underline="single")
UNKNOWN_FONT = Font(italic=True, color="808080")
LABEL_FONT = Font(bold=True)
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FRACTION_FMT = "# ??/??"          # displays 0.1875 as "3/16", 0.78125 as "25/32"
ACRE_FMT = "#,##0.00"

_HEADER_ROW = 2                    # row 1 holds nav; headers on row 2; data row 3+
_SRC_RE = re.compile(r"(SRC-\d+)")


# ── helpers ──────────────────────────────────────────────────────────────────

def _join(values) -> str:
    if isinstance(values, list):
        return "; ".join(str(v) for v in values)
    return "" if values is None else str(values)


def _num_or_blank(value):
    """Fraction -> float for formulas; UNKNOWN -> '' (blank, not 0)."""
    if isinstance(value, Fraction):
        return float(value)
    if value == UNKNOWN or value is None:
        return ""
    return value


def _nav(ws: Worksheet) -> None:
    c = ws.cell(row=1, column=1, value="↩ Contents")
    c.hyperlink = f"#'{S_CONTENTS}'!A1"
    c.font = LINK_FONT


def _ext(cell, url: str, text: str) -> None:
    cell.value = text
    if url and url != UNKNOWN:
        cell.hyperlink = url
        cell.font = LINK_FONT


def _internal(cell, sheet: str, row: int, text: str) -> None:
    cell.value = text
    cell.hyperlink = f"#'{sheet}'!A{row}"
    cell.font = LINK_FONT


def _link_citation(cell, citation: str, src_rows: Dict[str, int]) -> None:
    """Render a citation, hyperlinked to its Source Log row when resolvable."""
    cell.value = citation
    m = _SRC_RE.search(citation or "")
    if m and m.group(1) in src_rows:
        cell.hyperlink = f"#'{S_SOURCES}'!A{src_rows[m.group(1)]}"
        cell.font = LINK_FONT


def _link_instrument(cell, instrument_no: str, instr_rows: Dict[str, int]) -> None:
    cell.value = instrument_no
    if instrument_no in instr_rows:
        cell.hyperlink = f"#'{S_RUNSHEET}'!A{instr_rows[instrument_no]}"
        cell.font = LINK_FONT


def _headers(ws: Worksheet, headers: List[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _finalize_table(ws: Worksheet, ncols: int) -> None:
    last = ws.max_row
    if last >= _HEADER_ROW:
        ws.auto_filter.ref = f"A{_HEADER_ROW}:{get_column_letter(ncols)}{last}"
    ws.freeze_panes = ws.cell(row=_HEADER_ROW + 1, column=1)
    # grey-italic any UNKNOWN cells for at-a-glance scanning
    for row in ws.iter_rows(min_row=_HEADER_ROW + 1):
        for cell in row:
            if cell.value == UNKNOWN:
                cell.font = UNKNOWN_FONT
    _autofit(ws, ncols)


def _autofit(ws: Worksheet, ncols: int) -> None:
    for i in range(1, ncols + 1):
        letter = get_column_letter(i)
        max_len = 8
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 60)


# ── sheets ───────────────────────────────────────────────────────────────────

def _sheet_cover(ws: Worksheet, rpt: TitleReport) -> None:
    cfg = rpt.config
    ws.cell(row=1, column=1, value=f"{cfg.report_type}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=cfg.tract_label).font = SUB_FONT
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")

    if cfg.data_status != "EXAMINED":
        b = ws.cell(row=4, column=1,
                    value=(f"DATA STATUS: {cfg.data_status} — illustrative, internally-consistent "
                           "title model. NOT an examination of record title; no live records pulled."))
        b.fill = BANNER_FILL
        b.font = BANNER_FONT
        ws.merge_cells("A4:E4")

    rows = [
        ("Section-Township-Range", f"{cfg.section}-{cfg.township}-{cfg.range_}"),
        ("County / State", f"{cfg.county} / {cfg.state}"),
        ("Report Type", cfg.report_type),
        ("Report Date", cfg.report_date),
        ("Source Cutoff Date", cfg.source_cutoff_date),
        ("Gross Acres (assumed)", fr.frac_str(cfg.gross_acres)),
        ("Net Mineral Acres (assumed)", fr.frac_str(cfg.net_mineral_acres)),
        ("Acreage Basis", cfg.acreage_basis),
        ("Prepared For", cfg.prepared_for),
        ("Preparer", cfg.preparer),
    ]
    r = 6
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Assumptions & Scope").font = SECTION_FONT
    r += 1
    for a in cfg.assumptions:
        ws.cell(row=r, column=1, value=f"• {a}")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    r += 1
    link = ws.cell(row=r, column=1, value="▶ Go to Contents")
    link.hyperlink = f"#'{S_CONTENTS}'!A1"
    link.font = LINK_FONT
    ws.column_dimensions["A"].width = 30
    for col in "BCDE":
        ws.column_dimensions[col].width = 22


def _sheet_contents(ws: Worksheet, rpt: TitleReport) -> None:
    ws.cell(row=1, column=1, value="Contents").font = TITLE_FONT
    ws.cell(row=3, column=1, value="Sheet").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=2, value="Description").font = HEADER_FONT
    ws.cell(row=3, column=2).fill = HEADER_FILL
    r = 4
    for name, desc in SHEET_GUIDE:
        link = ws.cell(row=r, column=1, value=name)
        link.hyperlink = f"#'{name}'!A1"
        link.font = LINK_FONT
        ws.cell(row=r, column=2, value=desc)
        r += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70


def _sheet_sources(ws: Worksheet, rpt: TitleReport) -> Dict[str, int]:
    _nav(ws)
    headers = ["Source ID", "Category", "Name", "Open", "Accessed", "Doc Count",
               "Credential Req'd", "Page Ref", "Notes"]
    _headers(ws, headers)
    src_rows: Dict[str, int] = {}
    r = _HEADER_ROW + 1
    for s in rpt.sources:
        src_rows[s.source_id] = r
        ws.cell(row=r, column=1, value=s.source_id)
        ws.cell(row=r, column=2, value=s.category)
        ws.cell(row=r, column=3, value=s.name)
        _ext(ws.cell(row=r, column=4), s.location, "Open portal ↗")
        ws.cell(row=r, column=5, value=s.accessed_date)
        ws.cell(row=r, column=6, value=fr.frac_str(s.doc_count))
        ws.cell(row=r, column=7, value="YES" if s.credential_required else "no")
        ws.cell(row=r, column=8, value=s.page_ref)
        ws.cell(row=r, column=9, value=s.notes)
        r += 1
    _finalize_table(ws, len(headers))
    return src_rows


def _sheet_runsheet(ws: Worksheet, rpt: TitleReport, src_rows) -> Dict[str, int]:
    _nav(ws)
    headers = ["Instrument #", "Type", "Book", "Page", "Instrument Date",
               "Recording Date", "Grantors", "Grantees", "Legal (raw)",
               "Legal (normalized)", "Consideration", "Classification",
               "Image", "Citation", "Notes"]
    _headers(ws, headers)
    instr_rows: Dict[str, int] = {}
    r = _HEADER_ROW + 1
    for i in rpt.instruments:
        instr_rows[i.instrument_no] = r
        ws.cell(row=r, column=1, value=i.instrument_no)
        ws.cell(row=r, column=2, value=i.instrument_type)
        ws.cell(row=r, column=3, value=i.book)
        ws.cell(row=r, column=4, value=i.page)
        ws.cell(row=r, column=5, value=i.instrument_date)
        ws.cell(row=r, column=6, value=i.recording_date)
        ws.cell(row=r, column=7, value=_join(i.grantors))
        ws.cell(row=r, column=8, value=_join(i.grantees))
        ws.cell(row=r, column=9, value=i.legal_raw)
        ws.cell(row=r, column=10, value=i.legal_normalized)
        ws.cell(row=r, column=11, value=i.consideration)
        ws.cell(row=r, column=12, value=i.classification)
        _ext(ws.cell(row=r, column=13), i.image_url, "View image ↗")
        _link_citation(ws.cell(row=r, column=14), i.citation, src_rows)
        ws.cell(row=r, column=15, value=i.notes)
        r += 1
    _finalize_table(ws, len(headers))
    return instr_rows


def _sheet_abstractions(ws: Worksheet, rpt: TitleReport, src_rows, instr_rows) -> None:
    _nav(ws)
    headers = ["Instrument #", "Grant Clause", "Reservations", "Royalty/WI Language",
               "Exceptions", "Pugh/Shut-in", "Depth Limits", "Pooling Refs",
               "Liens", "Probate", "Tags", "Confidence", "Needs Review", "Citation"]
    _headers(ws, headers)
    r = _HEADER_ROW + 1
    for a in rpt.abstractions:
        _link_instrument(ws.cell(row=r, column=1), a.instrument_no, instr_rows)
        ws.cell(row=r, column=2, value=a.grant_clause)
        ws.cell(row=r, column=3, value=a.reservations)
        ws.cell(row=r, column=4, value=a.royalty_wi_language)
        ws.cell(row=r, column=5, value=a.exceptions)
        ws.cell(row=r, column=6, value=a.pugh_shutin)
        ws.cell(row=r, column=7, value=a.depth_limits)
        ws.cell(row=r, column=8, value=a.pooling_refs)
        ws.cell(row=r, column=9, value=a.liens)
        ws.cell(row=r, column=10, value=a.probate)
        ws.cell(row=r, column=11, value=_join(a.tags))
        ws.cell(row=r, column=12, value=fr.decimal_str(a.confidence, 2)
                if isinstance(a.confidence, Fraction) else a.confidence)
        ws.cell(row=r, column=13, value="YES" if a.needs_review else "no")
        _link_citation(ws.cell(row=r, column=14), a.citation, src_rows)
        if a.needs_review:
            ws.cell(row=r, column=13).fill = REVIEW_FILL
        r += 1
    _finalize_table(ws, len(headers))


def _sheet_chain(ws: Worksheet, rpt: TitleReport, chain_type: str, src_rows, instr_rows) -> None:
    _nav(ws)
    headers = ["Seq", "Instrument #", "Date", "Conveyance", "From", "To",
               "Interest", "Gap Note", "Citation"]
    _headers(ws, headers)
    links = sorted([c for c in rpt.chain_links if c.chain_type == chain_type],
                   key=lambda x: x.seq)
    r = _HEADER_ROW + 1
    for c in links:
        ws.cell(row=r, column=1, value=c.seq)
        _link_instrument(ws.cell(row=r, column=2), c.instrument_no, instr_rows)
        ws.cell(row=r, column=3, value=c.date)
        ws.cell(row=r, column=4, value=c.conveyance)
        ws.cell(row=r, column=5, value=c.from_party)
        ws.cell(row=r, column=6, value=c.to_party)
        ws.cell(row=r, column=7, value=c.interest)
        gap = ws.cell(row=r, column=8, value=c.gap_note or "— continuous —")
        if c.gap_note:
            gap.fill = WARN_FILL
        else:
            gap.fill = OK_FILL
        _link_citation(ws.cell(row=r, column=9), c.citation, src_rows)
        r += 1
    _finalize_table(ws, len(headers))


def _sheet_wells(ws: Worksheet, rpt: TitleReport, src_rows) -> None:
    _nav(ws)
    headers = ["API #", "Well Name", "Operator", "Formation", "Spacing Order",
               "Pooling Order", "First Production", "Status", "HBP Support", "OCC", "Citation"]
    _headers(ws, headers)
    r = _HEADER_ROW + 1
    for w in rpt.wells:
        _ext(ws.cell(row=r, column=1), OCC_WELL_URL, w.api_number)
        ws.cell(row=r, column=2, value=w.well_name)
        ws.cell(row=r, column=3, value=w.operator)
        ws.cell(row=r, column=4, value=w.formation)
        ws.cell(row=r, column=5, value=w.spacing_order)
        ws.cell(row=r, column=6, value=w.pooling_order)
        ws.cell(row=r, column=7, value=w.first_production)
        ws.cell(row=r, column=8, value=w.status)
        hbp = ws.cell(row=r, column=9, value=w.hbp_support)
        hbp.fill = OK_FILL if w.hbp_support == "Supported" else (
            REVIEW_FILL if w.hbp_support in ("Unknown", UNKNOWN) else WARN_FILL)
        _ext(ws.cell(row=r, column=10), OCC_IMG_URL, "OCC imaging ↗")
        _link_citation(ws.cell(row=r, column=11), w.citation, src_rows)
        r += 1
    _finalize_table(ws, len(headers))


def _sheet_curative(ws: Worksheet, rpt: TitleReport, src_rows, instr_rows) -> None:
    _nav(ws)
    headers = ["Issue ID", "Type", "Instrument #", "Description", "Materiality",
               "Priority", "Status", "Recommended Action", "Citation"]
    _headers(ws, headers)
    r = _HEADER_ROW + 1
    for c in rpt.curative:
        ws.cell(row=r, column=1, value=c.issue_id)
        ws.cell(row=r, column=2, value=c.issue_type)
        _link_instrument(ws.cell(row=r, column=3), c.instrument_no, instr_rows)
        ws.cell(row=r, column=4, value=c.description)
        mat = ws.cell(row=r, column=5, value=c.materiality)
        mat.fill = {"High": HIGH_FILL, "Medium": MED_FILL, "Low": LOW_FILL}.get(
            c.materiality, PatternFill())
        ws.cell(row=r, column=6, value=c.priority)
        stat = ws.cell(row=r, column=7, value=c.status)
        stat.fill = OK_FILL if c.status == "Cleared" else REVIEW_FILL
        ws.cell(row=r, column=8, value=c.recommended_action)
        _link_citation(ws.cell(row=r, column=9), c.citation, src_rows)
        r += 1
    _finalize_table(ws, len(headers))


def _sheet_matrix(ws: Worksheet, rpt: TitleReport, src_rows) -> Dict[str, int]:
    """NRI/WI matrix with live, fraction-formatted formulas and a reconciliation row."""
    _nav(ws)
    headers = ["Owner", "Role", "Mineral Interest", "Lease Royalty", "Working Interest",
               "Burdens", "ORRI", "NRI (computed)", "NMA (ac)", "Basis Language", "Citation"]
    _headers(ws, headers)
    gross = rpt.config.gross_acres
    gross_dec = float(gross) if isinstance(gross, Fraction) else None

    first = _HEADER_ROW + 1
    r = first
    for o in rpt.ownership:
        ws.cell(row=r, column=1, value=o.owner)
        ws.cell(row=r, column=2, value=o.role)
        for col, val in ((3, o.mineral_interest), (4, o.lease_royalty),
                         (5, o.working_interest), (6, o.burdens), (7, o.override_royalty)):
            c = ws.cell(row=r, column=col, value=_num_or_blank(val))
            c.number_format = FRACTION_FMT
        # NRI formula by role
        if o.role == "working_interest":
            f = f'=IF(OR(E{r}="",F{r}=""),"UNKNOWN",E{r}*(1-F{r}))'
        elif o.role == "overriding_royalty":
            f = f'=IF(G{r}="","UNKNOWN",G{r})'
        else:
            f = f'=IF(OR(C{r}="",D{r}=""),"UNKNOWN",C{r}*D{r})'
        nri = ws.cell(row=r, column=8, value=f)
        nri.number_format = FRACTION_FMT
        if gross_dec is not None:
            nma = ws.cell(row=r, column=9, value=f'=IF(C{r}="","",C{r}*{gross_dec})')
            nma.number_format = ACRE_FMT
        else:
            ws.cell(row=r, column=9, value=UNKNOWN)
        ws.cell(row=r, column=10, value=o.basis_language)
        _link_citation(ws.cell(row=r, column=11), o.citation, src_rows)
        r += 1

    last = r - 1
    total_row = r + 1
    ws.cell(row=total_row, column=1, value="TOTALS").font = LABEL_FONT
    mineral_total = ws.cell(row=total_row, column=3, value=f"=SUM(C{first}:C{last})")
    mineral_total.number_format = FRACTION_FMT
    mineral_total.font = LABEL_FONT
    nri_total = ws.cell(row=total_row, column=8, value=f"=SUM(H{first}:H{last})")
    nri_total.number_format = FRACTION_FMT
    nri_total.font = LABEL_FONT
    check = ws.cell(
        row=total_row, column=9,
        value=(f'=IF(AND(ROUND(C{total_row},6)=1,ROUND(H{total_row},6)=1),'
               f'"RECONCILES 8/8 ✓","CHECK TOTALS")'))
    check.font = LABEL_FONT

    note = ws.cell(
        row=total_row + 2, column=1,
        value=("Live formulas: royalty NRI = Mineral×Royalty; WI NRI = WI×(1−Burdens); "
               "ORRI NRI = ORRI. NMA = Mineral×gross acres. Interests shown as fractions; "
               "blank inputs render UNKNOWN, never 0. Minerals and total NRI must each equal 1 (8/8)."))
    note.font = Font(italic=True, size=9, color="595959")
    ws.merge_cells(start_row=total_row + 2, start_column=1, end_row=total_row + 2, end_column=11)
    _finalize_table(ws, len(headers))
    return {"total_row": total_row}


def _sheet_qa(ws: Worksheet, rpt: TitleReport, qa_summary: Optional[dict]) -> None:
    _nav(ws)
    _headers(ws, ["Check", "Result", "Detail"])
    r = _HEADER_ROW + 1
    if qa_summary:
        for chk in qa_summary.get("checks", []):
            ws.cell(row=r, column=1, value=chk["name"])
            res = ws.cell(row=r, column=2, value=chk["status"])
            res.fill = OK_FILL if chk["status"] == "PASS" else (
                WARN_FILL if chk["status"] == "FAIL" else MED_FILL)
            ws.cell(row=r, column=3, value=chk["detail"])
            r += 1
    else:
        ws.cell(row=r, column=1, value="(QA not yet run)")
    _finalize_table(ws, 3)


def _sheet_missing(ws: Worksheet, rpt: TitleReport, missing: List[dict]) -> None:
    _nav(ws)
    _headers(ws, ["Item", "Type", "Reason", "Citation/Source"])
    r = _HEADER_ROW + 1
    for m in missing:
        ws.cell(row=r, column=1, value=m.get("item", UNKNOWN))
        ws.cell(row=r, column=2, value=m.get("type", UNKNOWN))
        ws.cell(row=r, column=3, value=m.get("reason", UNKNOWN))
        ws.cell(row=r, column=4, value=m.get("source", UNKNOWN))
        r += 1
    if not missing:
        ws.cell(row=r, column=1, value="(none flagged)")
    _finalize_table(ws, 4)


def _sheet_summary(ws: Worksheet, rpt: TitleReport, matrix_ctx: dict,
                   qa_summary: Optional[dict]) -> None:
    _nav(ws)
    cfg = rpt.config
    ws.cell(row=2, column=1, value="Title Summary Dashboard").font = TITLE_FONT
    ws.merge_cells("A2:D2")
    if cfg.data_status != "EXAMINED":
        b = ws.cell(row=3, column=1, value=f"DATA STATUS: {cfg.data_status} (illustrative)")
        b.fill = BANNER_FILL
        b.font = BANNER_FONT
        ws.merge_cells("A3:D3")

    r = 5
    ws.cell(row=r, column=1, value="Tract Facts").font = SECTION_FONT
    r += 1
    facts = [
        ("Tract", cfg.tract_label),
        ("Report Date", cfg.report_date),
        ("Source Cutoff", cfg.source_cutoff_date),
        ("Gross Acres", fr.frac_str(cfg.gross_acres)),
    ]
    for label, val in facts:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Schedule Counts").font = SECTION_FONT
    r += 1
    counts = [
        (S_SOURCES, len(rpt.sources)), (S_RUNSHEET, len(rpt.instruments)),
        (S_ABSTRACT, len(rpt.abstractions)), ("Chain links", len(rpt.chain_links)),
        (S_WELLS, len(rpt.wells)), (S_CURATIVE, len(rpt.curative)),
        (S_MATRIX, len(rpt.ownership)),
    ]
    for label, n in counts:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=n)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Ownership Reconciliation (live)").font = SECTION_FONT
    r += 1
    tr = matrix_ctx["total_row"]
    ws.cell(row=r, column=1, value="Total Mineral Interest").font = LABEL_FONT
    mt = ws.cell(row=r, column=2, value=f"='{S_MATRIX}'!C{tr}")
    mt.number_format = FRACTION_FMT
    r += 1
    ws.cell(row=r, column=1, value="Total NRI").font = LABEL_FONT
    nt = ws.cell(row=r, column=2, value=f"='{S_MATRIX}'!H{tr}")
    nt.number_format = FRACTION_FMT
    r += 1
    ws.cell(row=r, column=1, value="Reconciliation").font = LABEL_FONT
    rc = ws.cell(row=r, column=2, value=f"='{S_MATRIX}'!I{tr}")
    rc.font = LABEL_FONT
    r += 1

    r += 1
    ws.cell(row=r, column=1, value="HBP / Wells").font = SECTION_FONT
    r += 1
    hbp = "; ".join(f"{w.well_name} ({w.status}; HBP {w.hbp_support})" for w in rpt.wells) or "none"
    ws.cell(row=r, column=1, value="Wells").font = LABEL_FONT
    ws.cell(row=r, column=2, value=hbp)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

    r += 1
    ws.cell(row=r, column=1, value="Curative by Materiality").font = SECTION_FONT
    r += 1
    by_mat = {"High": 0, "Medium": 0, "Low": 0}
    open_n = 0
    for c in rpt.curative:
        by_mat[c.materiality] = by_mat.get(c.materiality, 0) + 1
        if c.status != "Cleared":
            open_n += 1
    for label in ("High", "Medium", "Low"):
        cell = ws.cell(row=r, column=1, value=f"{label} items")
        cell.font = LABEL_FONT
        cell.fill = {"High": HIGH_FILL, "Medium": MED_FILL, "Low": LOW_FILL}[label]
        ws.cell(row=r, column=2, value=by_mat.get(label, 0))
        r += 1
    ws.cell(row=r, column=1, value="Open items").font = LABEL_FONT
    ws.cell(row=r, column=2, value=open_n)
    r += 1

    if qa_summary:
        r += 1
        ws.cell(row=r, column=1, value="QA Overall").font = SECTION_FONT
        res = ws.cell(row=r, column=2, value="PASS" if qa_summary["passed"] else "FAIL")
        res.fill = OK_FILL if qa_summary["passed"] else WARN_FILL

    ws.column_dimensions["A"].width = 28
    for col in "BCD":
        ws.column_dimensions[col].width = 26


# ── public entry point ───────────────────────────────────────────────────────

def build_workbook(
    rpt: TitleReport,
    output_path: str,
    qa_summary: Optional[dict] = None,
    missing: Optional[List[dict]] = None,
) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    # Create sheets up-front so display order is fixed and internal links resolve.
    ws = {name: wb.create_sheet(name) for name, _ in
          [(S_COVER, None), (S_CONTENTS, None), (S_SUMMARY, None), (S_SOURCES, None),
           (S_RUNSHEET, None), (S_ABSTRACT, None), (S_CHAIN_MIN, None),
           (S_CHAIN_LEASE, None), (S_WELLS, None), (S_CURATIVE, None),
           (S_MATRIX, None), (S_QA, None), (S_MISSING, None)]}

    # Populate in dependency order (Source Log + Runsheet first to build link maps).
    src_rows = _sheet_sources(ws[S_SOURCES], rpt)
    instr_rows = _sheet_runsheet(ws[S_RUNSHEET], rpt, src_rows)
    _sheet_abstractions(ws[S_ABSTRACT], rpt, src_rows, instr_rows)
    _sheet_chain(ws[S_CHAIN_MIN], rpt, "mineral_surface", src_rows, instr_rows)
    _sheet_chain(ws[S_CHAIN_LEASE], rpt, "leasehold_wi_burdens", src_rows, instr_rows)
    _sheet_wells(ws[S_WELLS], rpt, src_rows)
    _sheet_curative(ws[S_CURATIVE], rpt, src_rows, instr_rows)
    matrix_ctx = _sheet_matrix(ws[S_MATRIX], rpt, src_rows)
    _sheet_qa(ws[S_QA], rpt, qa_summary)
    _sheet_missing(ws[S_MISSING], rpt, missing or [])

    _sheet_cover(ws[S_COVER], rpt)
    _sheet_contents(ws[S_CONTENTS], rpt)
    _sheet_summary(ws[S_SUMMARY], rpt, matrix_ctx, qa_summary)

    wb.active = wb.sheetnames.index(S_COVER)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
