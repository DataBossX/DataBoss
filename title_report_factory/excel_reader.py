"""Read prior title workbooks (the Section 10 example, an existing Section 27
workbook, a KellPro export, an index) and surface their rows as Documents.

This is best-effort: it maps loosely-named columns to the Document fields and
preserves the original cell text. It never invents values for missing columns.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from .models import Document, UNKNOWN

# Loose header aliases -> Document attribute.
COLUMN_ALIASES: Dict[str, List[str]] = {
    "doc_type": ["document type", "doc type", "instrument", "instrument type", "type"],
    "grantor": ["grantor", "assignor", "lessor", "grantor/assignor/lessor",
                "grantor/lessor/assignor", "from"],
    "grantee": ["grantee", "assignee", "lessee", "grantee/assignee/lessee",
                "grantee/lessee/assignee", "to"],
    "execution_date": ["execution date", "executed", "date executed", "dated"],
    "effective_date": ["effective date", "effective"],
    "recording_date": ["recording date", "recorded", "file date", "filed"],
    "instrument_no": ["instrument no.", "instrument no", "instrument number",
                      "inst no", "doc no", "document no", "document number", "reception"],
    "book_page": ["book/page", "book page", "book/pg", "bk/pg", "book", "vol/page"],
    "legal_description": ["legal description", "legal", "lands", "description"],
    "interest": ["interest/lands affected", "interest", "interest conveyed",
                 "lands affected", "interest conveyed/reserved/affected"],
    "title_effect": ["title effect", "effect"],
    "notes": ["notes", "important notes", "remarks", "comments"],
}


def _norm(s) -> str:
    return str(s).strip().lower() if s is not None else ""


def _build_header_map(header_row) -> Dict[int, str]:
    """Map column index -> Document attribute using COLUMN_ALIASES."""
    idx_to_attr: Dict[int, str] = {}
    for col_idx, cell in enumerate(header_row):
        h = _norm(cell)
        if not h:
            continue
        for attr, aliases in COLUMN_ALIASES.items():
            if h in aliases:
                idx_to_attr[col_idx] = attr
                break
    return idx_to_attr


def read_workbook_documents(path: str, tract_filter: Optional[str] = None) -> List[Document]:
    """Extract Document rows from a workbook's data sheets.

    Returns [] if openpyxl is unavailable or the file cannot be opened — the
    caller logs that as a limitation rather than failing the whole run.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    p = Path(path)
    if not p.exists():
        return []

    try:
        wb = load_workbook(p, data_only=True, read_only=True)
    except Exception:
        return []

    docs: List[Document] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # Find the header row: first row that maps >= 3 known columns.
        header_idx = None
        header_map: Dict[int, str] = {}
        for i, row in enumerate(rows[: min(10, len(rows))]):
            hm = _build_header_map(row)
            if len(hm) >= 3:
                header_idx, header_map = i, hm
                break
        if header_idx is None:
            continue

        for row in rows[header_idx + 1:]:
            if not any(c not in (None, "") for c in row):
                continue
            doc = Document(source_ref=f"{p.name}::{ws.title}")
            populated = False
            for col_idx, attr in header_map.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val in (None, ""):
                    continue
                setattr(doc, attr, str(val).strip())
                populated = True
            if not populated:
                continue
            doc.match_confidence = 60  # came from a prior workbook, needs verification
            doc.relevance_confidence = 50
            if tract_filter and tract_filter.lower() not in (doc.legal_description or "").lower():
                doc.notes = (doc.notes + " | Tract match not confirmed from legal text").strip(" |")
            docs.append(doc)

    wb.close()
    return docs
