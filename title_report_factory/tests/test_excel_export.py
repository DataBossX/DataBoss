"""End-to-end test: the pipeline produces a valid workbook with all sheets."""
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from title_report_factory.pipeline import run, _planned_sheets

CONFIG = os.path.join(os.path.dirname(__file__), "..", "configs", "section_27_diversified.json")


def test_pipeline_produces_all_sheets(tmp_path=None):
    summary = run(CONFIG, log=lambda *a, **k: None)
    wb_path = summary["workbook_path"]
    assert os.path.exists(wb_path), "workbook not written"

    wb = load_workbook(wb_path, read_only=True)
    try:
        for s in _planned_sheets():
            assert s in wb.sheetnames, f"missing sheet {s}"
    finally:
        wb.close()

    # truth checks
    assert summary["overall_confidence_score"] <= 68, "cursory run must stay honestly capped"
    assert summary["leases_count"] == 64, "expected 64 base OGLs"
    assert summary["wells_count"] == 4, "expected 4 OCC wells"
    assert summary["index_rows_created_count"] >= 64
    assert any("Cherokee" in f for f in summary["major_title_findings"])
    assert summary["curative_issues"], "curative issues must be present"


def test_no_blank_required_cells():
    summary = run(CONFIG, log=lambda *a, **k: None)
    wb = load_workbook(summary["workbook_path"])
    ws = wb["Index Text"]
    # header is row 2; data starts row 3. Instrument No. column = 5.
    blanks = 0
    for row in ws.iter_rows(min_row=3, min_col=5, max_col=5, values_only=True):
        if row[0] in (None, ""):
            blanks += 1
    wb.close()
    assert blanks == 0, "Instrument No. column should never be blank (blank_cell_rule)"
