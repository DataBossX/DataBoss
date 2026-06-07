"""Automated tests for document parsing and Excel export.

Run with:  python -m pytest title_report_factory/tests -q
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from title_report_factory import document_classifier
from title_report_factory.excel_template_writer import SHEET_ORDER, write_workbook
from title_report_factory.extractors import interest as interest_x
from title_report_factory.extractors import legal_description as legal_x
from title_report_factory.extractors import party_and_date as party_x
from title_report_factory.utils import parse_date, short_date

SAMPLE_LEASE = """
OIL AND GAS LEASE

THIS AGREEMENT made and entered this 5th day of March, 2019, by and between
John Q. Smith, a single person, hereinafter called Lessor, and
Acme Operating LLC, hereinafter called Lessee.

Lessor hereby grants the following described land in Beckham County, Oklahoma:
Section 27, Township 11 North, Range 25 West, the NE/4 and the S/2,
containing 480 acres, more or less.

Royalty: 3/16. Primary term: 3 years.
Recorded 3/12/2019 as Instrument No. 2019-001234, Book 512, Page 88.
Effective date: 3/1/2019.
"""

SAMPLE_DEED = """
MINERAL DEED
This Mineral Deed dated January 2, 2020, between Mary Roe (Grantor) and
John Q. Smith (Grantee), conveys an undivided 1/2 interest in all oil, gas and
other minerals in Section 27, Township 11 North, Range 25 West, Beckham County,
Oklahoma. Grantor reserves 1/4. Recorded as Document No. 2020-009876.
"""


# --------------------------- classification --------------------------------
def test_classify_lease():
    label, conf, _ = document_classifier.classify(SAMPLE_LEASE)
    assert label == "Oil & Gas Lease"
    assert conf >= 60


def test_classify_mineral_deed():
    label, conf, _ = document_classifier.classify(SAMPLE_DEED)
    assert label == "Mineral Deed"
    assert conf >= 60


def test_classify_empty_is_unknown_low_conf():
    label, conf, _ = document_classifier.classify("")
    assert label == "Unknown"
    assert conf == 0.0


# --------------------------- legal description -----------------------------
def test_legal_extraction_matches_subject():
    legal = legal_x.extract_legal(SAMPLE_LEASE, "Beckham County", "Oklahoma")
    assert legal.section == "27"
    assert legal.township == "11N"
    assert legal.range == "25W"
    assert legal.confidence >= 80
    assert legal_x.matches_subject(legal, "27", "11N", "25W")


def test_legal_abbreviated_form():
    legal = legal_x.extract_legal("lands in Sec 27 T11N R25W, Beckham County")
    assert (legal.section, legal.township, legal.range) == ("27", "11N", "25W")


def test_legal_no_match_is_low_confidence():
    legal = legal_x.extract_legal("no legal here")
    assert legal.section is None
    assert legal.confidence == 0.0


# --------------------------- parties & dates -------------------------------
def test_party_extraction_lease():
    parties = party_x.extract_parties(SAMPLE_LEASE)
    assert any("Smith" in n for n in parties.get("lessor", []))
    assert any("Acme" in n for n in parties.get("lessee", []))


def test_recording_data():
    rec = party_x.extract_recording_data(SAMPLE_LEASE)
    assert rec["instrument_no"].startswith("2019")
    assert rec["book_page"] == "512/88"


def test_dates():
    dates = party_x.extract_dates(SAMPLE_LEASE)
    assert dates.get("recording_date") == "3/12/2019"
    assert dates.get("effective_date") == "3/1/2019"


def test_interest_extraction():
    money = interest_x.extract_interest(SAMPLE_DEED, "Mineral Deed")
    assert "1/2" in money.get("interest", "") or "1/4" in money.get("interest", "")


# --------------------------- date utilities --------------------------------
@pytest.mark.parametrize("raw,expected", [
    ("3/12/2019", "3/12/2019"),
    ("March 5, 2019", "3/5/2019"),
    ("2019-03-12", "3/12/2019"),
])
def test_short_date(raw, expected):
    assert short_date(raw) == expected


def test_parse_date_invalid():
    assert parse_date("not a date") is None


# --------------------------- excel export ----------------------------------
def test_write_workbook_has_all_sheets(tmp_path):
    sheets = {
        name: {"title": name, "subtitle": "test",
               "headers": ["A", "Date", "Confidence"],
               "rows": [{"A": "x", "Date": "3/12/2019", "Confidence": 88}]}
        for name in SHEET_ORDER
    }
    out = tmp_path / "out.xlsx"
    write_workbook(out, sheets)
    assert out.exists()
    wb = load_workbook(out)
    for name in SHEET_ORDER:
        assert name in wb.sheetnames
    # Frozen panes set, header present a few rows down.
    ws = wb["Overview"]
    assert ws.freeze_panes is not None


def test_write_workbook_dates_are_real_dates(tmp_path):
    sheets = {name: {"headers": ["Recording Date"], "rows":
                     [{"Recording Date": "3/12/2019"}]} for name in SHEET_ORDER}
    out = tmp_path / "out.xlsx"
    write_workbook(out, sheets)
    wb = load_workbook(out)
    ws = wb["Run Sheet"]
    # find the date cell (header row is row 4, data row 5)
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.number_format and "yyyy" in str(cell.number_format).lower():
                found = True
    assert found


# --------------------------- end-to-end (empty inputs) ---------------------
def test_pipeline_runs_with_empty_inputs(tmp_path):
    from title_report_factory.config_loader import RunConfig
    from title_report_factory.pipeline import run_pipeline

    (tmp_path / "inputs").mkdir()
    cfg = RunConfig(
        project_name="t", input_dirs=["inputs"], output_dir="outputs",
        output_file_name="t.xlsx",
    )
    cfg.base_dir = str(tmp_path)
    result = run_pipeline(cfg)
    assert Path(result.workbook_path).exists()
    wb = load_workbook(result.workbook_path)
    for name in SHEET_ORDER:
        assert name in wb.sheetnames
    # Honest empty-state: overall score should be modest, limitations recorded.
    assert result.errors_or_limitations
    # Side-car artifacts exist.
    out = tmp_path / "outputs"
    assert (out / "completion_summary.json").exists()
    assert (out / "source_inventory.json").exists()
    assert (out / "confidence_summary.json").exists()
    assert (out / "curative_issues.json").exists()
    summary = json.loads((out / "completion_summary.json").read_text())
    assert "overall_confidence_score" in summary
