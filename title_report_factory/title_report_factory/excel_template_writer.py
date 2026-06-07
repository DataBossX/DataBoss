"""Write the finished Excel workbook.

If a template workbook is supplied and present, we sample its header style
(font, fill) so the output matches the house style. Otherwise we apply a clean,
professional default: bold white header on a dark blue band, frozen header row,
auto-sized columns, wrapped text, and M/D/YYYY short dates.

The workbook is always written to the outputs folder -- original files are never
overwritten.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .utils import LOG, parse_date

# Sheet order required by the spec.
SHEET_ORDER = [
    "Overview", "Title", "PLAT", "Run Sheet", "OGL", "Well Data",
    "Index Text", "Source Notes", "Confidence Summary", "Curative Issues",
]

_DEFAULT_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_DEFAULT_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
_TITLE_FONT = Font(bold=True, size=16, color="1F3864", name="Calibri")
_SUB_FONT = Font(italic=True, size=10, color="595959", name="Calibri")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_DATE_FMT = "m/d/yyyy"


class StyleProfile:
    def __init__(self, header_fill: PatternFill, header_font: Font):
        self.header_fill = header_fill
        self.header_font = header_font


def _sample_template_style(template_path: Optional[str]) -> StyleProfile:
    """Best-effort: copy the first header cell style from a template workbook."""
    if template_path and Path(template_path).exists():
        try:
            wb = load_workbook(template_path)
            ws = wb.worksheets[0]
            cell = ws.cell(row=1, column=1)
            fill = cell.fill if cell.fill and cell.fill.fill_type == "solid" else _DEFAULT_HEADER_FILL
            font = Font(bold=True,
                        color=(cell.font.color.rgb if cell.font and cell.font.color else "FFFFFF"),
                        size=cell.font.size or 11,
                        name=cell.font.name or "Calibri")
            LOG.info("Sampled header style from template %s", Path(template_path).name)
            return StyleProfile(fill, font)
        except Exception as exc:
            LOG.warning("Could not sample template style: %s", exc)
    return StyleProfile(_DEFAULT_HEADER_FILL, _DEFAULT_HEADER_FONT)


def _looks_like_date_header(header: str) -> bool:
    return "date" in (header or "").lower()


def _write_table(ws, title: str, subtitle: str,
                 headers: List[str], rows: List[Dict],
                 style: StyleProfile, start_row: int = 1) -> None:
    # Title band.
    ws.cell(row=start_row, column=1, value=title).font = _TITLE_FONT
    ws.cell(row=start_row + 1, column=1, value=subtitle).font = _SUB_FONT
    header_row = start_row + 3

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = _BORDER

    for r, row in enumerate(rows, start=header_row + 1):
        for c, h in enumerate(headers, start=1):
            val = row.get(h, "")
            cell = ws.cell(row=r, column=c)
            if _looks_like_date_header(h):
                d = parse_date(val)
                if d is not None:
                    cell.value = d
                    cell.number_format = _DATE_FMT
                else:
                    cell.value = "" if val in (None,) else str(val)
            else:
                cell.value = val if isinstance(val, (int, float)) else (
                    "" if val is None else str(val))
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws, headers, header_row)


def _autosize(ws, headers: List[str], header_row: int) -> None:
    for c, h in enumerate(headers, start=1):
        letter = get_column_letter(c)
        max_len = len(str(h))
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)


def write_workbook(out_path: Path,
                   sheets: Dict[str, dict],
                   template_path: Optional[str] = None) -> Path:
    """`sheets` maps sheet name -> {headers: [...], rows: [ {col: val} ], title, subtitle}."""
    style = _sample_template_style(template_path)
    wb = Workbook()
    wb.remove(wb.active)

    for name in SHEET_ORDER:
        payload = sheets.get(name, {"headers": ["Note"], "rows": [], "title": name, "subtitle": ""})
        ws = wb.create_sheet(title=name[:31])
        _write_table(
            ws,
            title=payload.get("title", name),
            subtitle=payload.get("subtitle", ""),
            headers=payload.get("headers", ["Note"]),
            rows=payload.get("rows", []),
            style=style,
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    LOG.info("Workbook written: %s", out_path)
    return out_path
