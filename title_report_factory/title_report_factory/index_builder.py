"""Build the verified full-text index.

Two inputs feed the index:
  1. Documents we processed ourselves (OCR -> classify -> extract).
  2. Existing index/runsheet rows already present in uploaded Excel workbooks
     (Overview, Title, PLAT, OGLs, Runsheet, "from kellpro", source data, index).

We read those sheets, detect highlighted/flagged rows, and attempt to match
each existing index row to a source document. Unmatched rows are kept and
explicitly marked "Unmatched" rather than dropped.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from .models import Document
from .utils import LOG, NEEDS_REVIEW, NOT_FOUND

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill  # noqa: F401
    _HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAVE_OPENPYXL = False

# Sheet names we especially care about (case-insensitive contains match).
_INTERESTING_SHEETS = [
    "overview", "title", "plat", "ogl", "run", "kellpro",
    "source", "index", "runsheet",
]


def _cell_is_highlighted(cell) -> bool:
    """Detect a non-default (e.g. yellow) solid fill."""
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type not in ("solid",):
        return False
    rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return False
    rgb = rgb.upper()
    # Ignore white/none/black defaults.
    return rgb not in ("00000000", "FFFFFFFF", "FFFFFF", "000000", "")


def read_existing_workbook(path: Path) -> Dict[str, dict]:
    """Return {sheet_name: {headers, rows, highlighted_row_indices}}."""
    if not _HAVE_OPENPYXL:
        return {}
    out: Dict[str, dict] = {}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        LOG.error("Could not open workbook %s: %s", path.name, exc)
        return {}

    for ws in wb.worksheets:
        lname = ws.title.lower()
        if not any(key in lname for key in _INTERESTING_SHEETS):
            # Still capture, but mark as generic.
            pass
        headers: List[str] = []
        rows: List[List] = []
        highlighted: List[int] = []
        for r_idx, row in enumerate(ws.iter_rows()):
            values = [c.value for c in row]
            if r_idx == 0:
                headers = [str(v) if v is not None else "" for v in values]
                continue
            if all(v is None for v in values):
                continue
            if any(_cell_is_highlighted(c) for c in row):
                highlighted.append(len(rows))
            rows.append(values)
        out[ws.title] = {
            "headers": headers,
            "rows": rows,
            "highlighted_row_indices": highlighted,
        }
    LOG.info("Read %d sheet(s) from %s", len(out), path.name)
    return out


def _match_existing_row(headers: List[str], row: List,
                        docs: List[Document]) -> Optional[str]:
    """Try to match an existing index row to one of our documents."""
    joined = " ".join(str(v) for v in row if v is not None).lower()
    for d in docs:
        if d.instrument_no and d.instrument_no.lower() in joined:
            return d.doc_id
        if d.book_page and d.book_page.lower() in joined:
            return d.doc_id
    return None


def build_index(docs: List[Document],
                existing: Dict[str, Dict[str, dict]]) -> List[dict]:
    """Produce the Index Text rows.

    `existing` maps source-file-name -> sheets dict from read_existing_workbook.
    """
    rows: List[dict] = []
    idx = 0

    # 1) Rows for documents we processed ourselves.
    for d in docs:
        idx += 1
        rows.append({
            "Index No.": idx,
            "Source File": Path(d.source_file).name,
            "Matched Document": d.matched_document or d.document_type,
            "Document Type": d.document_type,
            "Instrument No.": d.instrument_no or NOT_FOUND,
            "Book/Page": d.book_page or NOT_FOUND,
            "Recording Date": d.recording_date or NOT_FOUND,
            "Execution Date": d.execution_date or NOT_FOUND,
            "Effective Date": d.effective_date or NOT_FOUND,
            "Grantor/Lessor/Assignor": "; ".join(d.grantors) or NOT_FOUND,
            "Grantee/Lessee/Assignee": "; ".join(d.grantees) or NOT_FOUND,
            "Legal Description": d.legal.raw or NOT_FOUND,
            "Interest Conveyed/Reserved/Affected": d.interest or NEEDS_REVIEW,
            "Oil & Gas Lease Terms": d.lease_terms or "",
            "Royalty": d.royalty or "",
            "Gross Acres": d.gross_acres or "",
            "Net Acres": d.net_acres or "",
            "Important Notes": "; ".join(d.issues) if d.issues else "",
            "Title Effect": d.title_effect or NEEDS_REVIEW,
            "Source Reference": d.source_reference or Path(d.source_file).name,
            "OCR Confidence": _fmt_conf(d.ocr_confidence),
            "Document Match Confidence": _fmt_conf(d.match_confidence),
            "Title Relevance Confidence": _fmt_conf(d.relevance_confidence),
        })

    # 2) Rows carried over from existing workbook index/runsheet sheets.
    for src_name, sheets in existing.items():
        for sheet_name, payload in sheets.items():
            lname = sheet_name.lower()
            if not any(k in lname for k in ("index", "run", "kellpro")):
                continue
            headers = payload["headers"]
            for r_i, row in enumerate(payload["rows"]):
                matched = _match_existing_row(headers, row, docs)
                idx += 1
                highlighted = r_i in payload["highlighted_row_indices"]
                note = "Carried from existing workbook"
                if highlighted:
                    note += "; FLAGGED/HIGHLIGHTED row in source"
                if not matched:
                    note += "; Unmatched to a source document"
                rows.append({
                    "Index No.": idx,
                    "Source File": f"{src_name} [{sheet_name}]",
                    "Matched Document": matched or "Unmatched",
                    "Document Type": _guess_col(headers, row, ("type", "instrument")) or NEEDS_REVIEW,
                    "Instrument No.": _guess_col(headers, row, ("instrument", "doc no", "document")) or NOT_FOUND,
                    "Book/Page": _guess_col(headers, row, ("book", "page")) or NOT_FOUND,
                    "Recording Date": _guess_col(headers, row, ("record",)) or NOT_FOUND,
                    "Execution Date": _guess_col(headers, row, ("execut", "dated")) or NOT_FOUND,
                    "Effective Date": _guess_col(headers, row, ("effective",)) or NOT_FOUND,
                    "Grantor/Lessor/Assignor": _guess_col(headers, row, ("grantor", "lessor", "assignor")) or NOT_FOUND,
                    "Grantee/Lessee/Assignee": _guess_col(headers, row, ("grantee", "lessee", "assignee")) or NOT_FOUND,
                    "Legal Description": _guess_col(headers, row, ("legal", "description", "lands")) or NOT_FOUND,
                    "Interest Conveyed/Reserved/Affected": _guess_col(headers, row, ("interest", "affect")) or NEEDS_REVIEW,
                    "Oil & Gas Lease Terms": _guess_col(headers, row, ("lease term", "term")) or "",
                    "Royalty": _guess_col(headers, row, ("royalty",)) or "",
                    "Gross Acres": _guess_col(headers, row, ("gross",)) or "",
                    "Net Acres": _guess_col(headers, row, ("net",)) or "",
                    "Important Notes": note,
                    "Title Effect": _guess_col(headers, row, ("title effect", "effect", "notes")) or NEEDS_REVIEW,
                    "Source Reference": f"{src_name}:{sheet_name}:row{r_i + 2}",
                    "OCR Confidence": "",
                    "Document Match Confidence": "Matched" if matched else "Unmatched",
                    "Title Relevance Confidence": "",
                })
    LOG.info("Built %d index row(s)", len(rows))
    return rows


def _guess_col(headers: List[str], row: List, keys) -> str:
    for i, h in enumerate(headers):
        hl = (h or "").lower()
        if any(k in hl for k in keys):
            if i < len(row) and row[i] not in (None, ""):
                return str(row[i])
    return ""


def _fmt_conf(value) -> str:
    if value is None:
        return ""
    try:
        return str(int(round(float(value))))
    except (TypeError, ValueError):
        return ""
