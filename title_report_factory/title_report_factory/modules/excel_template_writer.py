"""excel_template_writer -- render the finished cursory workbook.

Builds the 10 required sheets in the style of the Section 10 (Shanwee) template:
Calibri 9pt body, gray bold headers, wrapped text, frozen header panes,
short (M/D/YYYY) dates, and the cursory tract/title layout. No important cell is
left blank (blank_cell_rule).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..model import ReportContext, blank_safe

# --- shared styles -----------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="D8D8D8")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUBJECT_FILL = PatternFill("solid", fgColor="DDEBF7")
TRACT_FILL = PatternFill("solid", fgColor="F2F2F2")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BODY_FONT = Font(name="Calibri", size=9)
BOLD = Font(name="Calibri", size=9, bold=True)
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
WHITE_BOLD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_L = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_BOLD = Alignment(horizontal="right", vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title_row(ws, text, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def _header(ws, headers, row, widths=None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _row(ws, row, values, fills=None, wrap=True):
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP_L if wrap else Alignment(vertical="top")
        c.border = BORDER
        if fills and fills.get(j):
            c.fill = fills[j]


# ---------------------------------------------------------------------------
def write_workbook(ctx: ReportContext, qc: dict, out_path: str) -> str:
    cfg = ctx.config
    subj = cfg["subject_property"]
    wb = Workbook()
    wb.remove(wb.active)

    _overview(wb, ctx, qc, subj)
    _title(wb, ctx, subj)
    _plat(wb, ctx, subj)
    _runsheet(wb, ctx)
    _ogl(wb, ctx)
    _well_data(wb, ctx)
    _index_text(wb, ctx)
    _source_notes(wb, ctx)
    _confidence_summary(wb, ctx, qc)
    _curative(wb, ctx)

    wb.save(out_path)
    return out_path


def _str(subj):
    return f"Section {subj['section']}, Township {subj['township']}, Range {subj['range']}"


# ---- 1. Overview -----------------------------------------------------------
def _overview(wb, ctx, qc, subj):
    ws = wb.create_sheet("Overview")
    for col, w in zip("ABCDEFGH", [22, 26, 14, 14, 14, 14, 16, 16]):
        ws.column_dimensions[col].width = w
    _title_row(ws, "CURSORY TITLE REPORT", 8)
    rows = [
        ("Prospect / Client:", subj["entity"], "Examiner:", "RG"),
        ("Subject Lands:", _str(subj), "Report Type:", "Cursory (index-level)"),
        ("County / State:", f"{subj['county']} County, {subj['state']}", "Cursory Report Dated:", "6/7/2026"),
        ("Meridian:", subj.get("meridian", "Indian Meridian"), "Gross Acres:", f"{subj['gross_acres']:.2f}"),
        ("Formation Limit:", subj.get("formation_limit", "Cherokee"), "Through Index Entry:", "5/22/2026 (Bk/Pg 2486/0716)"),
    ]
    r = 3
    for label, val, label2, val2 in rows:
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=1).alignment = RIGHT_BOLD
        c = ws.cell(row=r, column=2, value=val); c.font = BODY_FONT; c.alignment = WRAP_L
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=5, value=label2).font = BOLD
        ws.cell(row=r, column=5).alignment = RIGHT_BOLD
        c2 = ws.cell(row=r, column=6, value=val2); c2.font = BODY_FONT; c2.alignment = WRAP_L
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="TRACTS (640.00 acres, more or less):").font = BOLD
    r += 1
    tracts = [("Tract 1", "NE/4", "160.00"), ("Tract 2", "NW/4", "160.00"),
              ("Tract 3", "SE/4", "160.00"), ("Tract 4", "SW/4", "160.00 (less 1.00 ac Buffalo Church tract)")]
    _header(ws, ["Tract", "Description", "Acres", "Notes"], r, None)
    ws.freeze_panes = None
    for t, d, a in [(x[0], x[1], x[2]) for x in tracts]:
        r += 1
        note = "Quarter section" if "less" not in a else "SW/4 less 1.00-ac church tract"
        _row(ws, r, [t, f"{d} {_str(subj)}", a, note])

    r += 2
    ws.cell(row=r, column=1, value="REMARKS:").font = BOLD
    r += 1
    remarks = (
        "Cursoried only insofar as the CHEROKEE formation. This is an index/record-level cursory "
        "report: conclusions rest on the county index, the provided source workbook, and official "
        "OCC/BLM public-record context. No recorded instrument images, lease schedules, or OTC "
        "production were independently reviewed; Diversified's interest is APPARENT only. "
        "Do not rely for drilling, acquisition close, division order, or pay status."
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=8)
    c = ws.cell(row=r, column=1, value=remarks); c.font = BODY_FONT; c.alignment = WRAP_L
    c.fill = WARN_FILL
    r += 5
    ws.cell(row=r, column=1, value="Overall Confidence:").font = BOLD
    ws.cell(row=r, column=2, value=f"{qc['overall_confidence']}/100 — {qc['overall_band']}").font = BODY_FONT


# ---- 2. Title --------------------------------------------------------------
def _title(wb, ctx, subj):
    ws = wb.create_sheet("Title")
    widths = [40, 12, 18, 10, 16, 46]
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    _title_row(ws, f"TITLE — {_str(subj)} (CURSORIED ONLY INSOFAR AS THE CHEROKEE FORMATION)", 6)

    # group leases by tract (NE/NW/SE/SW)
    from .legal_description_extractor import tract_for_quarter
    tract_names = {"NE/4": "TRACT 1: NE/4", "NW/4": "TRACT 2: NW/4",
                   "SE/4": "TRACT 3: SE/4", "SW/4": "TRACT 4: SW/4"}
    groups = {k: [] for k in tract_names}
    section_wide = []
    for lz in ctx.leases:
        t = tract_for_quarter(lz.legal_description)
        if t in groups:
            groups[t].append(lz)
        else:
            section_wide.append(lz)

    def render_block(start_row, header_text, header_fill, leases, empty_msg=None):
        r = start_row + 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=header_text)
        c.font = BOLD; c.fill = header_fill
        r += 1
        _header(ws, ["Mineral Owner / Lessor", "Net Acres", "OGL No. / Bk-Pg / Inst.",
                     "Royalty", "Expiration", "Comments"], r, widths)
        ws.freeze_panes = None
        seen = set()
        wrote = False
        for lz in leases:
            owner = lz.grantor or "Unknown"
            if owner in seen:
                continue
            seen.add(owner)
            wrote = True
            r += 1
            comment = (f"Lessee: {lz.grantee or 'Unknown'}. Term: {lz.term or 'Unknown'}. "
                       f"{lz.notes[:160] if lz.notes else ''} Src: {lz.source_file or 'Not Found'}.")
            _row(ws, r, [owner, blank_safe(lz.net_acres, "Not Verified"),
                         f"OGL {lz.ogl_no} / {lz.book_page} / {lz.instrument_no}",
                         blank_safe(lz.royalty, "Unknown"),
                         blank_safe(lz.exp_date, "Unknown") + ("\nHBP unverified" if lz.exp_date else ""),
                         comment])
        if not wrote and empty_msg:
            r += 1
            _row(ws, r, [empty_msg, "Needs Review", "Not Found", "Unknown", "Unknown",
                         "Verify mineral ownership for this tract."])
        return r

    r = 3
    for key, label in tract_names.items():
        r = render_block(
            r, f"{label} {_str(subj)}  —  Containing 160.00 acres, more or less",
            TRACT_FILL, groups[key],
            empty_msg="No quarter-specific lease matched to this tract in the index")

    # Section-wide leases (legal description shows the whole section, no quarter).
    if section_wide:
        r = render_block(
            r,
            "SECTION-WIDE LESSORS — lease legal shows all of Section 27-11N-25W (quarter not "
            "specified in index; interest spans all four tracts; verify allocation by image).",
            WARN_FILL, section_wide)


# ---- 3. PLAT ---------------------------------------------------------------
def _plat(wb, ctx, subj):
    ws = wb.create_sheet("PLAT")
    _title_row(ws, f"PLAT — {_str(subj)}", 6)
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 18
    grid = {
        (3, 1): "TRACT 2 — NW/4\n160 ac", (3, 4): "TRACT 1 — NE/4\n160 ac",
        (8, 1): "TRACT 4 — SW/4\n~159 ac\n(less church tract)", (8, 4): "TRACT 3 — SE/4\n160 ac",
    }
    for (r, c), label in grid.items():
        ws.merge_cells(start_row=r, start_column=c, end_row=r + 3, end_column=c + 2)
        cell = ws.cell(row=r, column=c, value=label)
        cell.font = BOLD; cell.alignment = CENTER; cell.fill = TRACT_FILL
        for rr in range(r, r + 4):
            for cc in range(c, c + 3):
                ws.cell(row=rr, column=cc).border = BORDER
    r = 13
    ws.cell(row=r, column=1, value="Wells located in Section 27 (OCC RBDMS):").font = BOLD
    for w in ctx.wells:
        r += 1
        ws.cell(row=r, column=1, value=f"• {w.name} ({w.api}) — {w.status}; {w.location}").font = BODY_FONT
        ws.cell(row=r, column=1).alignment = WRAP_L


# ---- 4. Run Sheet ----------------------------------------------------------
def _runsheet(wb, ctx):
    ws = wb.create_sheet("Run Sheet")
    headers = ["No.", "Document Type", "Grantor", "Grantee", "Execution Date", "Effective Date",
               "Recording Date", "Instrument No.", "Book/Page", "Legal Description",
               "Interest/Lands Affected", "Title Notes", "Source", "Confidence"]
    widths = [5, 20, 26, 26, 12, 12, 12, 14, 12, 18, 26, 40, 22, 11]
    _title_row(ws, "RUN SHEET — Apparent Diversified Chain (chronological)", len(headers))
    _header(ws, headers, 2, widths)
    r = 2
    for i, cl in enumerate(ctx.chain, 1):
        r += 1
        fill = {12: WARN_FILL} if cl.confidence < 50 else None
        _row(ws, r, [i, cl.doc_type, blank_safe(cl.grantor, "Unknown"), blank_safe(cl.grantee, "Unknown"),
                     "Unknown", "Unknown", blank_safe(cl.date, "Unknown"),
                     blank_safe(cl.instrument_no, "Not Found"), blank_safe(cl.book_page, "Not Found"),
                     cl.legal_scope, cl.interest_affected,
                     f"{cl.chain_function} | {cl.limitations}",
                     cl.source[:80] if cl.source else "diversifiedint",
                     f"{cl.confidence} ({cl.confidence_label})"], fills=fill)


# ---- 5. OGL ----------------------------------------------------------------
def _ogl(wb, ctx):
    ws = wb.create_sheet("OGL")
    headers = ["Lease No.", "Lessor", "Lessee", "Lease Date", "Effective Date", "Recording Date",
               "Instrument No.", "Book/Page", "Legal Description", "Gross Acres", "Net Acres",
               "Royalty", "Primary Term", "Extension/Option", "Assignments",
               "Current Lessee/Operator if Determinable", "HBP Notes", "Source", "Confidence"]
    widths = [8, 30, 22, 12, 12, 12, 14, 12, 26, 10, 10, 9, 14, 14, 14, 24, 28, 18, 10]
    _title_row(ws, "OIL & GAS LEASES — Section 27-11N-25W (index/first-page level)", len(headers))
    _header(ws, headers, 2, widths)
    r = 2
    for lz in ctx.leases:
        r += 1
        hbp = "Primary term expired on face; HBP status unverified (tie to OCC production)."
        _row(ws, r, [lz.ogl_no, blank_safe(lz.grantor, "Unknown"), blank_safe(lz.grantee, "Unknown"),
                     blank_safe(lz.execution_date, "Unknown"), blank_safe(lz.execution_date, "Unknown"),
                     blank_safe(lz.filing_date, "Unknown"), lz.instrument_no, lz.book_page,
                     lz.legal_description, blank_safe(lz.gross_acres, "Unknown"),
                     blank_safe(lz.net_acres, "Not Verified"), blank_safe(lz.royalty, "Unknown"),
                     blank_safe(lz.term, "Unknown"), "Needs Review", "Needs Review (see Run Sheet/Index)",
                     "Sanguine/French/Arrowhead/Todco era — current holder via assignments unverified",
                     hbp, blank_safe(lz.source_file, "OGL sheet"), lz.confidence])


# ---- 6. Well Data ----------------------------------------------------------
def _well_data(wb, ctx):
    ws = wb.create_sheet("Well Data")
    headers = ["Well Name", "API", "Operator", "Location", "Section-Township-Range", "Formation",
               "Status", "Spud Date", "Completion Date", "First Production", "Last Production Found",
               "Lease/Unit", "HBP Relevance", "Source", "Notes", "Confidence"]
    widths = [18, 14, 22, 30, 20, 12, 18, 12, 14, 14, 22, 22, 40, 30, 26, 10]
    _title_row(ws, "WELL DATA — OCC RBDMS / Completion context", len(headers))
    _header(ws, headers, 2, widths)
    r = 2
    for w in ctx.wells:
        r += 1
        _row(ws, r, [w.name, w.api, w.operator, w.location, w.str_label, w.formation, w.status,
                     blank_safe(w.spud_date, "Unknown"), blank_safe(w.completion_date, "Unknown"),
                     blank_safe(w.first_production, "Unknown"), w.last_production, w.lease_unit,
                     w.hbp_relevance, w.source, blank_safe(w.notes, "—"), w.confidence])


# ---- 7. Index Text ---------------------------------------------------------
def _index_text(wb, ctx):
    ws = wb.create_sheet("Index Text")
    headers = ["Index No.", "Source File", "Matched Document", "Document Type", "Instrument No.",
               "Book/Page", "Recording Date", "Execution Date", "Effective Date",
               "Grantor/Lessor/Assignor", "Grantee/Lessee/Assignee", "Legal Description",
               "Interest Conveyed/Reserved/Affected", "Oil & Gas Lease Terms", "Royalty",
               "Gross Acres", "Net Acres", "Important Notes", "Title Effect", "Source Reference",
               "OCR Confidence", "Document Match Confidence", "Title Relevance Confidence"]
    widths = [8, 16, 16, 16, 14, 12, 12, 12, 12, 26, 26, 22, 26, 22, 9, 10, 10, 40, 22, 24, 10, 12, 12]
    _title_row(ws, "INDEX TEXT — consolidated full instrument index", len(headers))
    _header(ws, headers, 2, widths)
    r = 2
    for it in ctx.instruments:
        r += 1
        fill = {18: WARN_FILL} if "TRAP" in (it.notes or "") else None
        _row(ws, r, [it.index_no, it.source_file, it.matched_document, it.doc_type,
                     blank_safe(it.instrument_no, "Not Found"), blank_safe(it.book_page, "Not Found"),
                     blank_safe(it.recording_date, "Unknown"), blank_safe(it.execution_date, "Unknown"),
                     blank_safe(it.effective_date, "Unknown"), blank_safe(it.grantor, "Unknown"),
                     blank_safe(it.grantee, "Unknown"), it.legal_description, it.interest,
                     blank_safe(it.lease_terms, "Not Applicable"), blank_safe(it.royalty, "Not Applicable"),
                     it.gross_acres, it.net_acres, it.notes, it.title_effect,
                     it.source_reference, it.ocr_confidence, it.match_confidence,
                     it.relevance_confidence], fills=fill)


# ---- 8. Source Notes -------------------------------------------------------
def _source_notes(wb, ctx):
    ws = wb.create_sheet("Source Notes")
    for col, w in zip("ABCDEF", [26, 40, 14, 14, 18, 50]):
        ws.column_dimensions[col].width = w
    _title_row(ws, "SOURCE NOTES — inventory, sources used & limitations", 6)
    r = 3
    ws.cell(row=r, column=1, value="Local Source Files (logged source-of-record)").font = BOLD
    r += 1
    _header(ws, ["File", "SHA-256 (first 16)", "Type", "Size (KB)", "Pages", "Note"], r,
            [26, 40, 14, 14, 18, 50])
    ws.freeze_panes = None
    for sf in ctx.source_files:
        r += 1
        _row(ws, r, [sf.name, sf.sha256[:16], sf.kind, round(sf.size_bytes / 1024),
                     sf.pages if sf.pages else "—", sf.note])

    r += 2
    ws.cell(row=r, column=1, value="External Sources Used").font = BOLD
    r += 1
    _header(ws, ["Source", "URL/Reference", "Used For", "Reliability", "Accessed", "Limitation"], r,
            [26, 40, 14, 14, 18, 50])
    ws.freeze_panes = None
    for s in ctx.sources_used:
        r += 1
        _row(ws, r, [s.get("name", ""), s.get("url", ""), s.get("used_for", ""),
                     s.get("reliability", ""), s.get("accessed", ""), s.get("limitation", "")])

    r += 2
    ws.cell(row=r, column=1, value="Limitations & Missing Capabilities").font = BOLD
    for lim in ctx.limitations:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value="• " + lim); c.font = BODY_FONT; c.alignment = WRAP_L

    r += 2
    ws.cell(row=r, column=1, value="Executive Summary").font = BOLD
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 10, end_column=6)
    c = ws.cell(row=r, column=1, value=ctx.executive_summary); c.font = BODY_FONT; c.alignment = WRAP_L


# ---- 9. Confidence Summary -------------------------------------------------
def _confidence_summary(wb, ctx, qc):
    ws = wb.create_sheet("Confidence Summary")
    for col, w in zip("ABCD", [30, 18, 18, 60]):
        ws.column_dimensions[col].width = w
    _title_row(ws, "CONFIDENCE SUMMARY", 4)
    r = 3
    ws.cell(row=r, column=1, value="Overall Confidence").font = BOLD
    ws.cell(row=r, column=2, value=f"{qc['overall_confidence']}/100").font = BODY_FONT
    ws.cell(row=r, column=4, value=qc["overall_band"]).font = BODY_FONT
    r += 1
    ws.cell(row=r, column=1, value="QC Checks Passed").font = BOLD
    ws.cell(row=r, column=2, value=f"{qc['checks_passed']}/{qc['checks_total']}").font = BODY_FONT
    r += 2
    _header(ws, ["Sheet", "Score", "Band", "Meaning"], r, [30, 18, 18, 60])
    ws.freeze_panes = None
    from ..model import confidence_band
    for sheet, score in qc["sheet_scores"].items():
        r += 1
        _row(ws, r, [sheet, f"{score}/100", confidence_band(score), ""])

    r += 2
    ws.cell(row=r, column=1, value="QC Check Detail").font = BOLD
    r += 1
    _header(ws, ["Check", "Result", "Detail", ""], r, [40, 14, 50, 10])
    ws.freeze_panes = None
    for ch in qc["checks"]:
        r += 1
        _row(ws, r, [ch["check"], "PASS" if ch["pass"] else "REVIEW", ch["detail"], ""])

    r += 2
    ws.cell(row=r, column=1, value="Score Legend").font = BOLD
    for band in ["90-100 Strong, clear, source-backed", "70-89 Good, minor uncertainty",
                 "50-69 Usable but needs review", "25-49 Weak or incomplete",
                 "0-24 Unreliable or speculative"]:
        r += 1
        ws.cell(row=r, column=1, value=band).font = BODY_FONT


# ---- 10. Curative Issues ---------------------------------------------------
def _curative(wb, ctx):
    ws = wb.create_sheet("Curative Issues")
    headers = ["Issue ID", "Severity", "Category", "Description", "Related Instruments",
               "Recommended Action", "Source"]
    widths = [12, 10, 16, 55, 24, 50, 24]
    _title_row(ws, "CURATIVE ISSUES & MISSING DOCUMENTS", len(headers))
    _header(ws, headers, 2, widths)
    r = 2
    sev_fill = {"High": PatternFill("solid", fgColor="F8CBAD"),
                "Medium": WARN_FILL, "Low": PatternFill("solid", fgColor="E2EFDA")}
    for iss in ctx.curative:
        r += 1
        _row(ws, r, [iss.issue_id, iss.severity, iss.category, iss.description,
                     iss.related_instruments, iss.recommended_action, iss.source],
             fills={2: sev_fill.get(iss.severity)})
