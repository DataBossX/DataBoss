"""source_reader -- read all sheets of an .xlsx into plain Python rows."""
from __future__ import annotations

from openpyxl import load_workbook


def read_sheet(path: str, sheet: str) -> list[list]:
    """Return rows (list of cell-value lists) for one sheet; [] if absent."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        return rows
    finally:
        wb.close()


def read_all_sheets(path: str) -> dict[str, list[list]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        out = {}
        for name in wb.sheetnames:
            ws = wb[name]
            out[name] = [list(r) for r in ws.iter_rows(values_only=True)]
        return out
    finally:
        wb.close()


def as_dicts(rows: list[list], header_row: int = 0) -> list[dict]:
    """Turn rows into dicts keyed by the header row's column names."""
    if not rows or header_row >= len(rows):
        return []
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[header_row])]
    out = []
    for r in rows[header_row + 1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = r[i] if i < len(r) else None
        out.append(d)
    return out
