"""pipeline -- orchestrate the full cursory title report workflow.

Stages (with parallel-friendly, independent extraction passes):
  1. config + file inventory
  2. records gathering (local + optional county download)
  3. OCR (capability-aware)
  4. classify + extract (OGL / wells / chain / index)
  5. multi-pass tournament review
  6. write workbook
  7. quality control + second pass
  8. export artifacts + final JSON summary
"""
from __future__ import annotations

import os
from concurrent.futures import ThreadPoolExecutor

from .config_loader import load_config, input_path
from .model import ReportContext
from .modules import (
    file_inventory, ocr_engine, okcounty_browser_downloader,
    ogl_analyzer, well_data_collector, chain_of_title_builder,
    index_builder, multi_ai_tournament_reviewer, quality_control,
    excel_template_writer, final_report_exporter, source_reader,
)


def _gather_sources_used(cfg: dict, pub_path: str | None) -> list[dict]:
    used = []
    if pub_path:
        for d in source_reader.as_dicts(source_reader.read_sheet(pub_path, "Source Register")):
            used.append({
                "name": str(d.get("Exact Source Name", "") or ""),
                "url": str(d.get("Source URL", "") or ""),
                "used_for": str(d.get("Used For", "") or ""),
                "reliability": str(d.get("Reliability", "") or ""),
                "accessed": str(d.get("Date Accessed", "") or ""),
                "limitation": str(d.get("Limitation", "") or ""),
            })
    return used


def _executive_summary(path: str, sheet: str) -> str:
    rows = source_reader.read_sheet(path, sheet)
    lines = []
    for r in rows:
        for cell in r:
            if cell and str(cell).strip():
                lines.append(str(cell).strip())
    return "\n".join(lines[:18])


def run(config_path: str, log=print) -> dict:
    cfg = load_config(config_path)
    subj = cfg["subject_property"]
    ctx = ReportContext(config=cfg)
    ctx.tools_used = ["openpyxl", "python-stdlib (hashlib/re/concurrent.futures)"]

    # 1. Inventory --------------------------------------------------------
    log("[1/8] Inventorying local input files...")
    ctx.source_files = file_inventory.build_inventory(cfg["_input_dir"])
    log(f"      {len(ctx.source_files)} files inventoried.")

    src = input_path(cfg, "source_workbook")
    pub = input_path(cfg, "public_records_workbook")
    pdf = input_path(cfg, "index_pdf")
    if not src:
        raise FileNotFoundError("source_workbook not found in inputs")

    # 2. Records gathering (optional county download) ---------------------
    log("[2/8] Gathering records (local + optional county download)...")
    dl = okcounty_browser_downloader.download_records(
        cfg, subj.get("search_variations", []), os.path.join(cfg["_output_dir"], "downloads"))
    ctx.tools_used.append("okcounty_browser_downloader (Playwright, optional)")
    ctx.limitations.append(
        f"OKCountyRecords live download: {dl.status} — {dl.detail}")
    ctx.sources_used = _gather_sources_used(cfg, pub)

    # 3. OCR (capability-aware) ------------------------------------------
    log("[3/8] OCR pass over scanned PDF index...")
    caps = ocr_engine.capabilities()
    ctx.tools_used.append("ocr_engine (pytesseract/PyMuPDF, optional)")
    if pdf:
        res = ocr_engine.ocr_pdf(pdf)
        ctx.ocr_cache[os.path.basename(pdf)] = {
            "status": res.ocr_status, "pages_processed": res.pages_processed,
            "mean_confidence": res.mean_confidence, "detail": res.detail,
            "text_chars": len(res.text),
        }
        if res.ocr_status != "ok":
            ctx.limitations.append(
                f"Index PDF '{os.path.basename(pdf)}' ({_pdf_pages(ctx, pdf)} pp) not OCR'd: {res.detail} "
                "Structured index data from the source workbook was used instead.")
    log(f"      OCR capable: {caps['can_ocr']}")

    # 4. Parallel extraction passes --------------------------------------
    log("[4/8] Running parallel extraction passes (OGL / wells / chain)...")
    sm = cfg["source_sheet_map"]
    psm = cfg["public_records_sheet_map"]
    with ThreadPoolExecutor(max_workers=4) as ex:
        f_ogl = ex.submit(ogl_analyzer.build_leases, src, sm["ogl"], subj)
        f_chain = ex.submit(chain_of_title_builder.build_chain, src, sm, subj)
        f_wells = (ex.submit(well_data_collector.build_wells, pub, psm["wells"], psm["completions"])
                   if pub else None)
        ctx.leases = f_ogl.result()
        ctx.chain = f_chain.result()
        ctx.wells = f_wells.result() if f_wells else []
    log(f"      {len(ctx.leases)} leases, {len(ctx.chain)} chain links, {len(ctx.wells)} wells.")

    ctx.instruments = index_builder.build_index(
        src, sm, ctx.leases, ctx.chain, subj, cfg.get("name_match_traps", []))
    log(f"      {len(ctx.instruments)} consolidated index rows.")

    ctx.executive_summary = _executive_summary(src, sm.get("executive_summary", "notes 3"))

    # 5. Tournament review -----------------------------------------------
    log("[5/8] Multi-pass tournament review (facts / effect / skeptic)...")
    multi_ai_tournament_reviewer.run_tournament(ctx)
    ctx.tools_used.append("multi_ai_tournament_reviewer (heuristic passes)")
    if not multi_ai_tournament_reviewer.llm_review("ping"):
        ctx.limitations.append("Multi-LLM tournament backend not configured (no API keys); "
                               "deterministic heuristic review passes were used.")
    log(f"      {len(ctx.curative)} curative issues raised.")

    # 6. Write workbook (first pass) -------------------------------------
    log("[6/8] Writing Excel workbook...")
    out_wb = os.path.join(cfg["_output_dir"], cfg["output"]["workbook_name"])
    qc = quality_control.run_qc(ctx, _planned_sheets())
    excel_template_writer.write_workbook(ctx, qc, out_wb)

    # 7. Second QC pass (verify the file actually opens) -----------------
    log("[7/8] Second QC pass (verify workbook integrity)...")
    produced = _verify_workbook(out_wb)
    qc = quality_control.run_qc(ctx, produced)
    excel_template_writer.write_workbook(ctx, qc, out_wb)  # rewrite with final scores
    produced = _verify_workbook(out_wb)
    log(f"      Workbook opens; sheets: {produced}")

    # 8. Export artifacts -------------------------------------------------
    log("[8/8] Exporting artifacts + final summary...")
    paths = final_report_exporter.export_artifacts(ctx, qc, cfg["_output_dir"], out_wb)

    summary = _final_summary(ctx, qc, paths, dl, caps)
    import json
    with open(os.path.join(cfg["_output_dir"], "completion_summary.json"), "w", encoding="utf-8") as fh:
        json.dump(summary, fh, indent=2, ensure_ascii=False, default=str)
    return summary


def _pdf_pages(ctx, pdf):
    for sf in ctx.source_files:
        if sf.path == pdf:
            return sf.pages or "?"
    return "?"


def _planned_sheets():
    return ["Overview", "Title", "PLAT", "Run Sheet", "OGL", "Well Data",
            "Index Text", "Source Notes", "Confidence Summary", "Curative Issues"]


def _verify_workbook(path: str) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _final_summary(ctx, qc, paths, dl, caps) -> dict:
    high = [c.description for c in ctx.curative if c.severity == "High"]
    findings = [
        "Apparent Diversified leasehold via two unverified tracks: Track A "
        "(Nadel&Gussman/EnerVest -> FourPoint -> Maverick -> Diversified, partly corporate-level) "
        "and Track B (ConocoPhillips/Burlington/LL&E -> DP Legacy -> Diversified entities).",
        f"{len(ctx.leases)} base oil & gas leases (1994-2001) cover the section; all primary terms "
        "expired on their face — current validity depends on unconfirmed HBP.",
        "Cherokee-formation HBP NOT confirmed: active wells exist but the matched completion "
        "formation is Hogshooter, not Cherokee.",
        "Possible divestitures/exclusions (DP Legacy -> Anadarko Basin Holdings 2023; "
        "DP Legacy/Diversified -> Teocalli wellbore assignment 2024) require schedule review.",
        "Name-match trap excluded: 'Diversified Financial' judgment/lien records are a different "
        "entity and were not treated as subject mineral ownership.",
    ]
    return {
        "workbook_path": paths["workbook"],
        "overall_confidence_score": qc["overall_confidence"],
        "overall_confidence_band": qc["overall_band"],
        "documents_reviewed_count": len(ctx.instruments),
        "documents_downloaded_count": len(dl.downloaded),
        "documents_ocr_count": sum(1 for v in ctx.ocr_cache.values() if v.get("status") == "ok"),
        "index_rows_created_count": len(ctx.instruments),
        "leases_count": len(ctx.leases),
        "wells_count": len(ctx.wells),
        "chain_links_count": len(ctx.chain),
        "qc_checks_passed": f"{qc['checks_passed']}/{qc['checks_total']}",
        "major_title_findings": findings,
        "missing_documents": [c.description for c in ctx.curative
                              if c.category in ("Leasehold", "Chain of title")][:8],
        "curative_issues": [{"id": c.issue_id, "severity": c.severity, "desc": c.description}
                            for c in ctx.curative],
        "tools_used": sorted(set(ctx.tools_used)),
        "artifacts": {k: v for k, v in paths.items()},
        "errors_or_limitations": ctx.limitations,
    }
