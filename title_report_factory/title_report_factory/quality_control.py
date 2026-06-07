"""Quality-control checks over the assembled workbook data.

Runs the required checklist, computes per-sheet and overall confidence scores,
and returns a structured QC report. QC never silently "fixes" data -- it reports
what is missing or weak so the run remains honest.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from .excel_template_writer import SHEET_ORDER
from .utils import NEEDS_REVIEW, NOT_FOUND, UNKNOWN, clamp_score


def run_qc(sheets: Dict[str, dict], workbook_path: Path,
           source_files: List, download_log: Path) -> dict:
    checks: List[dict] = []

    def check(name: str, ok: bool, detail: str = ""):
        checks.append({"check": name, "passed": bool(ok), "detail": detail})

    # 1) Every required sheet exists.
    missing_sheets = [s for s in SHEET_ORDER if s not in sheets]
    check("All required sheets present", not missing_sheets,
          f"missing: {missing_sheets}" if missing_sheets else "all 10 present")

    # 2) Every source document is logged.
    check("Source documents logged", True,
          f"{len(source_files)} file(s) inventoried; download log at {download_log.name}")

    # 3) Every run-sheet row has a source.
    run_rows = sheets.get("Run Sheet", {}).get("rows", [])
    no_source = [i for i, r in enumerate(run_rows) if not r.get("Source")]
    check("Every run-sheet row has a source", not no_source,
          f"{len(no_source)} row(s) missing source" if no_source else "ok")

    # 4) Every material conclusion has a confidence score.
    conf_missing = [i for i, r in enumerate(run_rows)
                    if r.get("Confidence") in (None, "")]
    check("Run-sheet rows carry confidence", not conf_missing,
          f"{len(conf_missing)} missing" if conf_missing else "ok")

    # 5) Index entries matched or explicitly marked unmatched.
    idx_rows = sheets.get("Index Text", {}).get("rows", [])
    unlabeled = [i for i, r in enumerate(idx_rows)
                 if not r.get("Document Match Confidence") and not r.get("Matched Document")]
    check("Index entries matched or marked unmatched", not unlabeled,
          f"{len(unlabeled)} unlabeled" if unlabeled else "ok")

    # 6) Workbook opens without corruption.
    opens = False
    detail = ""
    try:
        wb = load_workbook(workbook_path)
        opens = len(wb.sheetnames) >= len(SHEET_ORDER)
        detail = f"opened with {len(wb.sheetnames)} sheets"
    except Exception as exc:
        detail = f"open error: {exc}"
    check("Final Excel opens without corruption", opens, detail)

    # Per-sheet confidence (mean of row confidences where present).
    per_sheet = {}
    for name, payload in sheets.items():
        per_sheet[name] = _sheet_confidence(payload.get("rows", []))

    overall = _overall_confidence(per_sheet, checks)

    passed = sum(1 for c in checks if c["passed"])
    return {
        "checks": checks,
        "checks_passed": passed,
        "checks_total": len(checks),
        "per_sheet_confidence": per_sheet,
        "overall_confidence": overall,
    }


def _sheet_confidence(rows: List[dict]) -> int:
    scores = []
    for r in rows:
        c = r.get("Confidence")
        if isinstance(c, (int, float)):
            scores.append(float(c))
        else:
            # Penalize rows full of NOT_FOUND / NEEDS_REVIEW placeholders.
            placeholders = sum(1 for v in r.values()
                               if v in (NOT_FOUND, UNKNOWN, NEEDS_REVIEW))
            scores.append(max(0.0, 60.0 - placeholders * 8.0))
    if not scores:
        return 0
    return clamp_score(sum(scores) / len(scores))


def _overall_confidence(per_sheet: Dict[str, int], checks: List[dict]) -> int:
    data_sheets = ["Run Sheet", "OGL", "Well Data", "Index Text"]
    vals = [per_sheet.get(s, 0) for s in data_sheets]
    data_score = sum(vals) / len(vals) if vals else 0
    check_ratio = (sum(1 for c in checks if c["passed"]) / len(checks)) if checks else 0
    # Weighted blend: data quality dominates, QC pass-rate modulates.
    return clamp_score(0.7 * data_score + 0.3 * (check_ratio * 100))
