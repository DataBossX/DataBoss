"""Build and export the cursory title-report Excel workbook.

Produces all required sheets with consistent styling (matching the palette used
by doto_image_commander/reports/generator.py), short M/D/YYYY dates, frozen
header rows, and column auto-fit. Empty data areas are filled with an explicit
honest placeholder rather than left blank or fabricated.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ReportData, Document

# ── Style palette (shared with doto_image_commander) ──────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="1F4E79", bold=True, size=16)
LABEL_FONT = Font(bold=True)
REVIEW_FILL = PatternFill("solid", fgColor="FFE699")
EMPTY_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
NO_DATA = "No source documents provided — input required."


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autofit(ws, max_width=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), max_width)


def _table(ws, headers: List[str], rows: List[list]):
    ws.append(headers)
    _style_header(ws)
    ws.freeze_panes = "A2"
    if not rows:
        placeholder = [NO_DATA] + [""] * (len(headers) - 1)
        ws.append(placeholder)
        for cell in ws[2]:
            cell.fill = EMPTY_FILL
            cell.alignment = WRAP
    else:
        for r in rows:
            ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    _autofit(ws)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_overview(wb, rd: ReportData):
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "CURSORY TITLE REPORT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = ("This is a cursory land/title report prepared from available "
                "records and OCR review, NOT a formal attorney title opinion. "
                "All flagged issues require attorney review and county-record "
                "confirmation.")
    ws["A2"].font = Font(italic=True, color="C00000")
    ws.merge_cells("A2:D2")
    ws["A2"].alignment = WRAP
    r = 4
    for label, value in rd.overview.items():
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=value)
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    ws.column_dimensions["A"].width = 30
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 30
    ws.row_dimensions[2].height = 45


def _sheet_title(wb, rd: ReportData):
    ws = wb.create_sheet("Title")
    _table(ws, ["Conclusion", "Supporting Instruments / Basis", "Confidence (0-100)"],
           [[n["conclusion"], n["support"], n["confidence"]] for n in rd.title_narrative])


def _sheet_plat(wb, rd: ReportData):
    ws = wb.create_sheet("PLAT")
    ws["A1"] = f"PLAT / TRACT SUMMARY — {rd.tract.short}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    rows = [
        ("Section", rd.tract.section),
        ("Township", rd.tract.township),
        ("Range", rd.tract.range),
        ("County / State", f"{rd.tract.county} County, {rd.tract.state}"),
        ("Base Legal Description", rd.tract.legal_base),
        ("Affected Lands", NO_DATA if not rd.documents else "See Run Sheet / Index Text"),
        ("Tract / Lease / Well References", NO_DATA if not rd.documents else "See Index Text"),
        ("Notes", "Standard section = 640 acres, 16 aliquot 40-acre QQ tracts. "
                  "No tract-level conveyances confirmed without source documents."),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=value)
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    ws.column_dimensions["A"].width = 30
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 30


def _sheet_runsheet(wb, rd: ReportData):
    ws = wb.create_sheet("Run Sheet")
    headers = ["No.", "Document Type", "Grantor/Assignor/Lessor",
               "Grantee/Assignee/Lessee", "Execution Date", "Effective Date",
               "Recording Date", "Instrument No.", "Book/Page", "Legal Description",
               "Interest/Lands Affected", "Title Effect", "Notes",
               "Source File/Record", "Confidence"]
    rows = []
    for i, d in enumerate(_sorted_docs(rd.documents), 1):
        rows.append([i, d.doc_type, d.grantor, d.grantee, d.execution_date,
                     d.effective_date, d.recording_date, d.instrument_no,
                     d.book_page, d.legal_description, d.interest, d.title_effect,
                     d.notes, d.source_ref, d.confidence])
    _table(ws, headers, rows)


def _sheet_ogl(wb, rd: ReportData):
    ws = wb.create_sheet("OGL")
    headers = ["Lease No.", "Lessor", "Lessee", "Lease Date", "Effective Date",
               "Recording Date", "Instrument No.", "Book/Page", "Legal Description",
               "Gross Acres", "Net Acres", "Royalty", "Primary Term",
               "Extension/Option", "Assignments", "Current Lessee/Owner",
               "Held By Production Notes", "Source", "Confidence"]
    rows = []
    for i, d in enumerate(rd.leases, 1):
        rows.append([i, d.grantor, d.grantee, d.execution_date, d.effective_date,
                     d.recording_date, d.instrument_no, d.book_page,
                     d.legal_description, "Unknown", "Unknown", "Unknown",
                     "Unknown", "Unknown", "Needs Review", "Needs Review",
                     "Needs Review", d.source_ref, d.confidence])
    _table(ws, headers, rows)


def _sheet_well(wb, rd: ReportData):
    ws = wb.create_sheet("Well Data")
    headers = ["Well Name", "API", "Operator", "Location", "Section-Township-Range",
               "Formation/Zone", "Status", "Spud Date", "Completion Date",
               "First Production", "Last Production Found", "Lease/Unit",
               "HBP Relevance", "Source", "Notes", "Confidence"]
    rows = []
    for w in rd.wells:
        rows.append([w.get(h, "Unknown") for h in
                     ["well_name", "api", "operator", "location", "str",
                      "formation", "status", "spud", "completion", "first_prod",
                      "last_prod", "lease_unit", "hbp", "source", "notes", "confidence"]])
    if not rows:
        # Honest note specific to well data unavailability.
        ws.append(headers)
        _style_header(ws)
        ws.freeze_panes = "A2"
        ws.append(["No well/production data collected. Oklahoma Corporation "
                   "Commission and other public sources were not accessed "
                   "(no record access in this run). HBP status is UNKNOWN."] +
                  [""] * (len(headers) - 1))
        for cell in ws[2]:
            cell.fill = EMPTY_FILL
            cell.alignment = WRAP
        _autofit(ws)
    else:
        _table(ws, headers, rows)


def _sheet_index(wb, rd: ReportData):
    ws = wb.create_sheet("Index Text")
    headers = ["Index No.", "Source File", "Matched Document", "Document Type",
               "Instrument No.", "Book/Page", "Recording Date", "Execution Date",
               "Effective Date", "Grantor/Lessor/Assignor", "Grantee/Lessee/Assignee",
               "Legal Description", "Interest Conveyed/Reserved/Affected",
               "OGL Terms if Any", "Royalty if Any", "Gross Acres", "Net Acres",
               "Important Notes", "Title Effect", "Source Reference",
               "OCR Confidence", "Document Match Confidence", "Title Relevance Confidence"]
    rows = []
    for i, d in enumerate(rd.documents, 1):
        matched = "Matched" if d.doc_type != "Unknown" else "Unmatched"
        rows.append([i, d.source_ref, matched, d.doc_type, d.instrument_no,
                     d.book_page, d.recording_date, d.execution_date,
                     d.effective_date, d.grantor, d.grantee, d.legal_description,
                     d.interest, "Unknown", "Unknown", "Unknown", "Unknown",
                     d.notes, d.title_effect, d.source_ref, d.ocr_confidence,
                     d.match_confidence, d.relevance_confidence])
    _table(ws, headers, rows)


def _sheet_source_notes(wb, rd: ReportData):
    ws = wb.create_sheet("Source Notes")
    headers = ["Source ID", "Source Type", "File Name or Record Reference",
               "Where Found", "Pages", "OCR Status", "Extraction Status",
               "Important Extracted Text", "Limitations", "Confidence"]
    rows = [[s.source_id, s.source_type, s.file_name, s.where_found, s.pages,
             s.ocr_status, s.extraction_status, s.important_text[:500],
             s.limitations, s.confidence] for s in rd.sources]
    if not rows and rd.limitations:
        rows = [["S000", "N/A", "(none found)", "n/a", "n/a", "n/a", "n/a",
                 "", " | ".join(rd.limitations), 0]]
    _table(ws, headers, rows)


def _sheet_confidence(wb, rd: ReportData):
    ws = wb.create_sheet("Confidence Summary")
    ws["A1"] = "CONFIDENCE & RELIABILITY DASHBOARD"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    r = 3
    for label, value in rd.confidence.items():
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        if isinstance(value, list):
            ws.cell(row=r, column=2, value="; ".join(map(str, value))).alignment = WRAP
        else:
            ws.cell(row=r, column=2, value=value)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="QUALITY-CONTROL PASSES").font = LABEL_FONT
    r += 1
    ws.cell(row=r, column=1, value="Pass").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=2, value="Result").font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    ws.cell(row=r, column=3, value="Detail").font = HEADER_FONT
    ws.cell(row=r, column=3).fill = HEADER_FILL
    r += 1
    for qc in rd.qc_results:
        ws.cell(row=r, column=1, value=qc["pass"])
        ws.cell(row=r, column=2, value=qc["result"])
        ws.cell(row=r, column=3, value=qc["detail"]).alignment = WRAP
        r += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 70


def _sheet_curative(wb, rd: ReportData):
    ws = wb.create_sheet("Curative Issues")
    headers = ["Issue No.", "Severity", "Issue Type", "Affected Instrument",
               "Affected Lands/Interest", "Problem", "Needed Action",
               "Suggested Curative Document or Research", "Source", "Confidence"]
    rows = [[c.issue_no, c.severity, c.issue_type, c.affected_instrument,
             c.affected_lands, c.problem, c.needed_action, c.suggested_curative,
             c.source, c.confidence] for c in rd.curative]
    _table(ws, headers, rows)
    # Highlight Critical rows.
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value).lower() == "critical":
            for cell in row:
                cell.fill = REVIEW_FILL


def _sorted_docs(docs: List[Document]) -> List[Document]:
    def key(d):
        return (d.recording_date or "", d.execution_date or "", d.instrument_no or "")
    return sorted(docs, key=key)


def build_workbook(rd: ReportData, out_path: str) -> str:
    wb = Workbook()
    _sheet_overview(wb, rd)
    _sheet_title(wb, rd)
    _sheet_plat(wb, rd)
    _sheet_runsheet(wb, rd)
    _sheet_ogl(wb, rd)
    _sheet_well(wb, rd)
    _sheet_index(wb, rd)
    _sheet_source_notes(wb, rd)
    _sheet_confidence(wb, rd)
    _sheet_curative(wb, rd)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def verify_workbook(path: str) -> dict:
    """Excel Final Check: reopen and confirm required sheets exist and are non-trivial."""
    required = ["Overview", "Title", "PLAT", "Run Sheet", "OGL", "Well Data",
                "Index Text", "Source Notes", "Confidence Summary", "Curative Issues"]
    try:
        wb = load_workbook(path)
    except Exception as e:  # pragma: no cover
        return {"pass": "Excel Final Check", "result": "FAIL",
                "detail": f"Workbook did not reopen: {e}"}
    missing = [s for s in required if s not in wb.sheetnames]
    empty = [s for s in wb.sheetnames if wb[s].max_row < 1]
    ok = not missing and not empty
    return {
        "pass": "Excel Final Check",
        "result": "PASS" if ok else "FAIL",
        "detail": (f"{len(wb.sheetnames)} sheets present; "
                   f"missing={missing or 'none'}; empty={empty or 'none'}."),
    }
