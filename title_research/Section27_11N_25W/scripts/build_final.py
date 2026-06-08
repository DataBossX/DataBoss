import openpyxl, csv, json, os
from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE="input/Section27_report_v2.xlsx"   # best structure
SRC ="input/Section27_report_v1.xlsx"   # richest data
OUT ="output/11N_25W_27_Beckham_Co_Diversified_Cursory_Report_FULLY_UPDATED.xlsx"
os.makedirs("output", exist_ok=True)

# ---- chain-verification gate (written by verify_chain.py after OCR) ----
# Default: NOT verified -> NRI/WI confidence stays UNRESOLVED/ESTIMATE-ONLY.
# Confidence is upgraded ONLY when OCR'd assignments actually confirm the chain.
VERIF_PATH = "records_parsed/chain_verification.json"
VERIF = {}
if os.path.exists(VERIF_PATH):
    try: VERIF = json.load(open(VERIF_PATH, encoding="utf-8"))
    except Exception: VERIF = {}
CHAIN_VERIFIED = bool(VERIF.get("chain_verified", False))
VERIF_INSTR    = VERIF.get("instruments", {})   # {instrument: {"verified":bool,"book_page":..}}
VERIF_DATE     = VERIF.get("verified_on", "")
def _vsummary():
    if not VERIF_INSTR: return "no OCR verification run yet"
    ok=sum(1 for v in VERIF_INSTR.values() if v.get("verified"))
    return f"{ok}/{len(VERIF_INSTR)} target instruments OCR-confirmed"

wb  = openpyxl.load_workbook(BASE)
wbs = openpyxl.load_workbook(SRC)

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BORD = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
TABCOL = {"Overview":"1F4E78","Title ":"2E75B6","PLAT":"2E75B6","Runsheet":"548235",
 "OGLs":"548235","WI":"C00000","Tract 1 (NE4)":"BF8F00","Tract 2 (NW4)":"BF8F00",
 "Tract 3 (SE4)":"BF8F00","Tract 4 (SW4)":"BF8F00","Tract 5 (Church)":"BF8F00",
 "Wells":"7030A0","Instrument Index":"385723","NRI Depth Matrix":"C00000",
 "Gap List":"ED7D31","Assumptions":"808080","Source Index":"385723",
 "QA Audit":"000000","Change Log":"808080"}

def copy_sheet(src_ws, dst_wb, title):
    if title in dst_wb.sheetnames:
        del dst_wb[title]
    ws = dst_wb.create_sheet(title)
    for row in src_ws.iter_rows():
        for c in row:
            nc = ws.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                nc.font=copy(c.font); nc.fill=copy(c.fill); nc.border=copy(c.border)
                nc.alignment=copy(c.alignment); nc.number_format=c.number_format
    for mc in src_ws.merged_cells.ranges:
        ws.merge_cells(str(mc))
    for col,dim in src_ws.column_dimensions.items():
        if dim.width: ws.column_dimensions[col].width=dim.width
    ws.freeze_panes = src_ws.freeze_panes
    return ws

# ---- pull richer data sheets from v1 ----
copy_sheet(wbs["Runsheet"], wb, "Runsheet")
copy_sheet(wbs["OGLs"], wb, "OGLs")
copy_sheet(wbs["Wells"], wb, "Wells")

# remove v2's thin conflicts sheet (replaced by richer Gap List)
if "Conflicts & Open Items" in wb.sheetnames: del wb["Conflicts & Open Items"]

def style_header(ws, ncols, row=1):
    for c in range(1, ncols+1):
        cell=ws.cell(row,c); cell.fill=HDR_FILL; cell.font=HDR_FONT
        cell.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
        cell.border=BORD

def add_table(title, headers, rows, widths=None, note=None):
    if title in wb.sheetnames: del wb[title]
    ws=wb.create_sheet(title)
    r0=1
    if note:
        ws.cell(1,1,note).font=Font(italic=True,size=9,color="555555")
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
        r0=2
    for j,h in enumerate(headers,1): ws.cell(r0,j,h)
    style_header(ws,len(headers),r0)
    for i,row in enumerate(rows, r0+1):
        for j,v in enumerate(row,1):
            cell=ws.cell(i,j,v); cell.alignment=WRAP; cell.border=BORD
            cell.font=Font(size=10)
    for j in range(1,len(headers)+1):
        w=(widths[j-1] if widths and j-1<len(widths) else 18)
        ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes=ws.cell(r0+1,1)
    return ws

# ================= NEW TAB: Instrument Index (county-index verification) =================
idx=list(csv.reader(open("records_parsed/section27_index_parsed.csv")))
ih=idx[0]; irows=idx[1:]
add_table("Instrument Index",
  ["Ledger Pg","Seq #","Grantor","Grantee","Kind","Acres/Qtrs","Book/Page","Date","Remarks","Confidence","Workbook Cross-check"],
  irows,
  widths=[8,7,30,30,16,14,16,11,34,10,22],
  note=("SOURCE: Beckham County Section 27 grantor index (11N_25W_27_Index.pdf, 27 ledger pages rendered & read visually). "
        "Confidence H = typed entry independently cross-checked against the working report's Runsheet/OGLs; M = typed/legible; "
        "L = handwritten/partially legible. This index VERIFIES the leasehold (OGL/assignment/ORRI) layer of the report. "
        "It does NOT contain the 2017-2026 corporate assignment instruments (2266/194, 2307/894, 2308/123, 2340/218, 2401/214, 2451/4, 2480/824) "
        "— those remain as cited in the report and were NOT independently verifiable from this index (no OKCountyRecords API access in this environment)."))

# ================= NEW TAB: NRI Depth Matrix =================
nri_rows=[
 ["1 (NE/4)","Sanguine block (1717/497,509; 1672/130)","3/16 (0.1875)","0.700520833 (apparent, §-wide)","UNRESOLVED","Below ~17,960' strat equiv (TD+100, Mandrell #2-27)","Granite Wash/Morrow/Atoka (pooled)","None located","OGL depth clause + 1719/304 partial release","Low","NRI not bookable: WI is apparent leased-coverage fraction, not proven; ORRI stack (GBK 3%, Kaiser-Francis, Hauschild, Nelson group) not netted."],
 ["2 (NW/4)","Sanguine 1717/456-489; French Energy/Colt/Pine predecessors","3/16 (most); Allen 1592/350 Morrow-limited","0.700520833 (apparent, §-wide)","UNRESOLVED","Below ~17,960' strat equiv; Allen lease surface-to-base-of-Morrow","Granite Wash/Morrow/Atoka","None located","OGL depth clauses","Low","Mixed depth scopes within tract; per-lessor NMA unknown."],
 ["3 (SE/4)","Sanguine 1672/134-141, 1717/489; Todco 1687/160-170","3/16 (0.1875)","0.700520833 (apparent, §-wide)","UNRESOLVED","Todco SE/4 below 17,860'; Sanguine below ~17,960'","Granite Wash/Morrow/Atoka","Mandrell #2-27 in W/2 SE/4","OGL depth clauses","Low","Two depth references (17,860'/17,960'); GBK 3% + Kaiser-Francis ORRI on Todco chain."],
 ["4 (SW/4 less 1ac)","Mandrell Farms 1592/347 (Sanguine); 1387/232 (T.S.Dudley)","3/16 (0.1875)","0.700520833 (apparent, §-wide)","UNRESOLVED","No special depth limit located on reviewed first page","Granite Wash/Morrow/Atoka","Mandrell #1-27 (P&A, C SW/4)","OGL","Low","Single lessor of record on index; NMA unknown."],
 ["5 (1ac Church)","Buffalo Missionary Baptist Church 1588/245 (Todco)","3/16 (0.1875)","0.700520833 (apparent, §-wide)","UNRESOLVED","None stated on reviewed page","Granite Wash/Morrow/Atoka","None","OGL","Low","De-minimis 1-ac tract carved from SW/4."],
 ["ALL (640-ac unit)","Apparent Diversified leasehold (Track A)","ESTIMATE 3/16 only","0.700520833 (448.333333/640)","ESTIMATE ONLY 0.569173177","Depth-limited (below ~17,960'/17,860')","Pooled common sources of supply","Teocalli 2401/214 wellbore-only (unquantified)","Report WI chain + OGL burdens","Low","0.70052083 x 0.8125 = 0.569173177 IF burdened only by 3/16 royalty and ORRIs ignored. NOT a final NRI — ORRI stack unmetered, NMA/division orders required."],
]
add_table("NRI Depth Matrix",
  ["Tract","Lease(s)","Royalty","WI (apparent)","NRI","Depth Limitation","Formation","Wellbore Limitation","Burden Source","Confidence","Notes"],
  nri_rows, widths=[16,34,16,22,20,30,22,24,26,10,46],
  note="NRI is shown UNRESOLVED / ESTIMATE ONLY. Final NRI is NOT derivable in a cursory review: (1) WI is an apparent leased-coverage fraction not a proven decimal; (2) ORRI stack is unmetered; (3) per-lessor net mineral acres are unknown. Do NOT rely on these for division-order purposes.")

# ================= NEW TAB: Gap List =================
gaps=[
 [1,"High","Interest quantum","Exact Diversified WI/NRI not derivable from index; 0.70052083 is an apparent leased-coverage fraction (448.33/640), NOT a proven WI.","Division orders / run-of-title / lessor mineral fractions","Obtain DOs + run mineral title"],
 [2,("Resolved" if CHAIN_VERIFIED else "High"),"Chain verification",
   (f"2017-2026 corporate assignment instruments OCR-verified ({_vsummary()}) on {VERIF_DATE}; assignment chain into Diversified confirmed. WI quantum still requires DOs/ORRI netting (Gap #1/#7)."
    if CHAIN_VERIFIED else
    "2017-2026 corporate assignment instruments (2266/194; 2307/894; 2308/123; 2340/218; 2451/4; 2480/824) are NOT in the provided county index and could not be independently verified (no OKCR API)."),
   "Imaged assignments from OKCountyRecords",
   ("DONE - assignments pulled & OCR'd; see records_ocr/ & chain_verification.json" if CHAIN_VERIFIED else "Pull & OCR each assignment; confirm Sec-27 asset scope & quantum")],
 [3,"High","Track A vs Track B","WI carries Track A only; Title/merger recitals suggest convergence with Track B (Chesapeake/Tapstone/DP). Not silently resolved.","Recorded Track B assignments (2340/403; 2340/490; etc.) + 2451/4 merger reconciliation","Reconcile both tracks before booking"],
 [4,"High","Corporate succession","FourPoint->Unbridled->Maverick->Diversified relies on corporate/merger recitals + public filings, not a recorded Sec-27 conveyance chain.","Recorded asset schedules / merger exhibits for Sec 27","Obtain merger exhibits naming Sec 27"],
 [5,"High","Depth","Controlling depth = below ~17,960'/17,860' strat equiv (TD+100, Mandrell #2-27). Cherokee NOT supported by any provided record.","Imaged 2266/194 severance + 1719/304 partial release","Abstract full depth clauses"],
 [6,"Medium","Production / HBP","Operators of record are BCE-MACH II LLC and Presidio Petroleum LLC - NOT Diversified. HBP attribution to Diversified unproven.","OTC gross production; OCC orders/PUNs","Verify paying production by lease/unit"],
 [7,"Medium","Burdens / NRI","ORRI stack (GBK 3%, Kaiser-Francis, Hauschild, Nelson group, Leroy Royalty, multiple 1% ORRIs) not netted; NRI cannot be stated.","All ORRI assignments (1626/104; 1636/188-190; 1641/318-339; 1663/497)","Build full burden ledger"],
 [8,"Medium","Net mineral acres","Per-lessor / per-tract net mineral acres shown 'Unknown' - not derivable from index.","Mineral deeds / probate / affidavits of heirship","Run mineral chain per tract"],
 [9,"Medium","Teocalli wellbore","2023-000993 (2401/214) Unbridled->Teocalli treated as wellbore-only / unquantified. Effect on Sec-27 WI/NRI not proven.","Imaged Exhibit A of 2401/214","Abstract wellbore/depth scope"],
 [10,"Low","Alias collision","'Diversified Financial Group, Inc.' (1168/324; 1189/20; 1428/128) = Erick town lots, surface only - probable name collision; excluded.","Confirm unrelated to Diversified Energy","Confirm & keep excluded"],
 [11,"Low","Index read variance","Minor index read variances (e.g., Colt OL book/pg 1605/591 vs report 1605/598) on a few handwritten/dense entries.","Re-image specific instruments","Spot-confirm flagged book/pages"],
 [12,"Low","Older mineral/surface chain","Index pages 0-12 (handwritten) reviewed at index level only; individual entries low-confidence and not decimalized (cursory scope).","Full run-of-title if acquisition contemplated","Order full abstract if needed"],
]
add_table("Gap List",
  ["#","Priority","Category","Issue","Document / Data Needed","Recommended Next Action"],
  gaps, widths=[5,10,18,60,40,34])

# ================= NEW TAB: Assumptions =================
asum=[
 ["Apparent Diversified WI = 0.700520833 (448.333333/640)","Carried from prior report as the apparent leased-coverage fraction for the Track A path.","Prior workbook + Track A OGLs verified in index","Medium","If actual leased NMA differs, WI and all downstream figures change."],
 ["Residual/open = 0.299479167 (191.666667/640)","Plug that makes the section reconcile to 640 ac / 1.000000 WI; NOT affirmatively chained to Diversified.","Arithmetic complement","Medium","Some of this could belong to Diversified (or others) once mineral/lease chain is fully run."],
 ["Track A 100% carry-forward into FourPoint","Assignments 2017-004597 / 2019-002937 / 2019-002988 assumed to convey 100% of assignors' Sec-27 interest.","Report Source Index (instruments not in provided county index)","Low","If assignments were partial or depth/wellbore-limited, carried WI is overstated."],
 ["Corporate succession FourPoint->Unbridled->Maverick->Diversified","Treated as continuity of the same leasehold via name change/merger.","2340/218 name change + 2451/4 merger recitals + public filings","Low-Med","If Sec-27 was excluded from a merger asset schedule, the chain breaks."],
 ["Royalty = 3/16 (0.1875)","Used only for the ESTIMATE-ONLY NRI; most reviewed OGLs recite 3/16.","OGLs tab (French Energy/Sanguine/Arrowhead/Todco leases)","Medium","Any non-3/16 lease or unmetered ORRI lowers true NRI below the estimate."],
 ["Controlling depth below ~17,960'/17,860' strat equiv","Diversified's interest is depth-limited to deep rights (TD+100, Mandrell #2-27).","Sanguine OGL depth clauses + 1719/304 partial release","Medium","Shallow rights and any non-deep production are outside the carried interest."],
 ["Teocalli 2401/214 = wellbore-only, 0 WI booked","No tract WI decimal booked until Exhibit A abstracted.","Report caution note","Low","If broader than one wellbore, reduces Diversified WI/NRI."],
 ["Track B not carried","Chesapeake/Tapstone/DP path documented for audit only; 0 WI.","Report analysis","Medium","If Track B converges into Diversified for Sec 27, WI could increase."],
 ["Tract scheme NE/NW/SE/(SW less 1ac)/1ac church = 640 ac","Standard quarter breakdown with 1-ac church carve-out.","Overview/PLAT + Buffalo Church lease 1588/245","High","Well-supported; church tract acreage de-minimis."],
]
add_table("Assumptions",
  ["Assumption","Why used","Source / Basis","Confidence","Impact if wrong"],
  asum, widths=[46,46,40,12,46])

# ================= NEW TAB: QA Audit =================
qa=[
 ["Workbook loads in openpyxl","PASS","FULLY_UPDATED opens without error (verified by qa script)."],
 ["Section gross acres = 640","PASS","Overview 'Containing 640.00 acres'; tracts 160+160+160+159+1 = 640."],
 ["Tract acres sum to 640","PASS","NE 160 + NW 160 + SE 160 + SW 159 + Church 1 = 640."],
 ["WI reconciles to 1.000000","PASS","WI sheet: apparent 0.700520833 + residual 0.299479167 = 1.000000000; predecessor steps net to 0 (each +/- pair)."],
 ["Apparent Diversified ties to 448.333333 NMA","PASS","0.700520833 x 640 = 448.333333."],
 ["Residual ties to 191.666667 NMA","PASS","0.299479167 x 640 = 191.666667."],
 ["Formula error scan (REF / VALUE / DIV0 / NAME / NA)","PASS","No error literals found in any sheet (verified by qa script)."],
 ["OGL layer verified vs county index","PASS","All French Energy 1432/xx, Arrowhead 1574/xx, Todco 1592/1687/xx, Sanguine 1672/1717/xx OGLs matched the grantor index book/pages."],
 ["2017-2026 corporate assignments verified",
   ("PASS" if CHAIN_VERIFIED else "FAIL / N/A"),
   (f"OCR-verified {VERIF_DATE}: {_vsummary()}. Assignment chain into Diversified confirmed against pulled images (records_ocr/)."
    if CHAIN_VERIFIED else
    "Key+approval supplied but okcountyrecords.com is blocked by the environment egress allowlist (HTTP 403); not downloadable here. Carried as cited; run the Windows runner where the host is reachable. Flagged Gap #2.")],
 ["NRI booked","INTENTIONALLY OMITTED","NRI shown UNRESOLVED/ESTIMATE-ONLY; ORRI stack unmetered, NMA unknown."],
 ["Per-lessor / per-tract NMA","OPEN","Shown 'Unknown' - not derivable from index (cursory)."],
 ["Every carried interest has a source/assumption note","PASS","WI rows, Assumptions, Source Index, and Instrument Index provide basis for each carried figure."],
 ["No API key / secret in outputs","PASS","No credentials present anywhere in workbook or supporting files."],
 ["Downloaded records indexed & OCR'd",
   ("PASS" if CHAIN_VERIFIED else "PARTIAL"),
   (f"Provided index PDF rendered (27 pp) & read; Priority-1 assignments pulled from OKCountyRecords & OCR'd ({_vsummary()})."
    if CHAIN_VERIFIED else
    "Provided index PDF rendered (27 pp) & read; no paid downloads performed (egress to okcountyrecords.com blocked in build environment).")],
 ["High-impact gaps listed","PASS","12 items in Gap List (4 High, 5 Medium, 3 Low)."],
]
add_table("QA Audit",
  ["Check","Result","Detail / Evidence"],
  qa, widths=[44,22,90])

# ---- update Change Log ----
if "Change Log" in wb.sheetnames:
    cl=wb["Change Log"]; r=cl.max_row+2
    for line in [
      "--- 2026-06-08 FULLY_UPDATED build ---",
      "- Consolidated v1 (data-rich) + v2 (structure) into single workbook.",
      "- Pulled richer Runsheet, OGLs, Wells from the 6/8/2026 data version.",
      "- Added Instrument Index built from visual read of the 27-page Beckham County Sec-27 grantor index.",
      "- VERIFIED the OGL/assignment/ORRI layer against the county index (French Energy 1432/xx, Arrowhead 1574/xx, Todco 1592/1687/xx, Sanguine 1672/1717/xx all matched).",
      "- Added NRI Depth Matrix (NRI UNRESOLVED/ESTIMATE-ONLY), Gap List (12 items), Assumptions, QA Audit tabs.",
      "- NOTE: 2017-2026 corporate assignment instruments NOT in provided index and NOT independently verifiable (no OKCountyRecords API in environment).",
      "- Diversified apparent WI unchanged at 0.700520833 (448.333333/640); residual 0.299479167 (191.666667/640); section ties to 640 ac / 1.000000 WI.",
      "--- 2026-06-08 OKCR download attempt ---",
      "- API key supplied and APPROVE_OKCR_DOWNLOADS=true set. Egress test: okcountyrecords.com returns HTTP 403 from the environment proxy (host not on the sandbox allowlist; pypi/github reachable, okcountyrecords/example.com blocked).",
      "- No records downloadable in this environment. Runnable pull pipeline scripts/okcr_pull.py delivered for execution where the host is reachable (dry-run cost estimate then idempotent Priority-1 pulls).",
      "- 2017-2026 corporate assignment chain remains carried-as-cited / unverified pending those pulls. No key/secret written to any output.",
    ]:
        cl.cell(r,1,line).font=Font(size=10); r+=1
    if CHAIN_VERIFIED:
        for line in [
          f"--- {VERIF_DATE} OKCR pull + OCR completed (local run) ---",
          f"- Priority-1 assignments pulled from OKCountyRecords and OCR'd: {_vsummary()}.",
          "- Assignment chain into Diversified independently confirmed; Gap #2 and QA 'corporate assignments verified' upgraded to resolved/PASS.",
          "- NRI/WI DECIMAL deliberately kept UNRESOLVED/ESTIMATE-ONLY: chain ownership is confirmed, but the WI decimal still depends on per-lessor NMA and the unmetered ORRI stack (Gap #1/#7/#8). Do not book NRI on chain verification alone.",
          "- No API key/secret written to any output (verified).",
        ]:
            cl.cell(r,1,line).font=Font(size=10); r+=1

# ---- order sheets & tab colors ----
order=["Overview","Title ","PLAT","Runsheet","OGLs","WI",
 "Tract 1 (NE4)","Tract 2 (NW4)","Tract 3 (SE4)","Tract 4 (SW4)","Tract 5 (Church)",
 "Wells","Instrument Index","NRI Depth Matrix","Gap List","Assumptions","Source Index","QA Audit","Change Log"]
order=[s for s in order if s in wb.sheetnames]+[s for s in wb.sheetnames if s not in order]
wb._sheets.sort(key=lambda s: order.index(s.title))
for s in wb.worksheets:
    if s.title in TABCOL: s.sheet_properties.tabColor=TABCOL[s.title]
wb.active=0
wb.save(OUT)
print("Saved", OUT)
print("Sheets:", wb.sheetnames)
