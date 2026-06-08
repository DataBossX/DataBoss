import openpyxl, sys
def dump(path, label):
    print("="*70); print(label, path); print("="*70)
    wb = openpyxl.load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        print(f"\n### SHEET: '{ws.title}'  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column} color={ws.sheet_properties.tabColor}")
for p,l in [("input/Section27_report_v1.xlsx","REPORT v1"),
            ("input/Section27_report_v2.xlsx","REPORT v2"),
            ("input/project_notes.xlsx","NOTES")]:
    dump(p,l)
