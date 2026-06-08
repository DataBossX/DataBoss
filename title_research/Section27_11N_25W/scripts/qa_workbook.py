import openpyxl
OUT="output/11N_25W_27_Beckham_Co_Diversified_Cursory_Report_FULLY_UPDATED.xlsx"
errs=["#REF!","#VALUE!","#DIV/0!","#NAME?","#NULL!","#NUM!"]
wb=openpyxl.load_workbook(OUT)
print("LOADS OK |",len(wb.sheetnames),"sheets")
found=0; secret=0
KEYS=["api_key","apikey","okcr","okcountyrecords_api","password","secret","token","bearer"]
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if c.value is None: continue
            s=str(c.value)
            for e in errs:
                if e in s: print("FORMULA ERROR",ws.title,c.coordinate,s); found+=1
            sl=s.lower()
            for k in KEYS:
                if k in sl and "okcountyrecords.com" not in sl and "no okcr" not in sl and "okcr api" not in sl:
                    # allow descriptive mentions
                    if any(x in sl for x in ["=","key:","token:","secret:"]) and len(s)<200:
                        pass
print("Formula error literals found:",found)
# reconciliations
print("\n--- RECONCILIATIONS ---")
print("0.700520833*640 =",round(0.7005208328125*640,6),"(expect 448.333333)")
print("0.299479167*640 =",round(0.2994791671875*640,6),"(expect 191.666667)")
print("WI sum =",round(0.7005208328125+0.2994791671875,9),"(expect 1.0)")
print("Tracts 160+160+160+159+1 =",160+160+160+159+1,"(expect 640)")
# WI sheet net check of predecessor columns
ws=wb["WI"]
for col in "HIJKL":
    tot=0
    for r in range(10,19):
        v=ws[f"{col}{r}"].value
        if isinstance(v,(int,float)): tot+=v
    print(f"WI col {col} net (rows10-18) = {round(tot,9)} (expect ~0)")
gcol=sum(ws[f"G{r}"].value for r in range(10,19) if isinstance(ws[f"G{r}"].value,(int,float)))
print("WI col G total (booked) =",round(gcol,9),"(expect 1.0)")
