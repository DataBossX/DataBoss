#!/usr/bin/env python3
"""Build a traceable, non-fabricated title package when evidence is unavailable."""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


UNABLE_TO_VERIFY = "Unable to verify from official county records."
RELEASE_STATUS = "BLOCKED — DO NOT RELEASE"
ALLOWED_STATUSES = {
    "VERIFIED",
    "PARTIAL MATCH",
    "LIKELY MATCH",
    "NOT FOUND",
    "WRONG BOOK",
    "WRONG PAGE",
    "WRONG GRANTOR",
    "WRONG GRANTEE",
    "WRONG LEGAL",
    "DUPLICATE",
    "CORRECTED RECORDING",
    "SUPERSEDED",
}
WORKBOOK_NAMES = (
    "VERIFIED_MASTER_WORKBOOK.xlsx",
    "CORRECTED_RUN_SHEET.xlsx",
    "INSTRUMENT_REGISTER.xlsx",
    "SOURCE_CITATION_INDEX.xlsx",
    "ERROR_LOG.xlsx",
    "MISSING_DOCUMENT_LOG.xlsx",
    "TITLE_GAP_ANALYSIS.xlsx",
    "DIVERSIFIED_OWNERSHIP_CHAIN.xlsx",
)
REPORT_NAMES = (
    "QA_CERTIFICATION.md",
    "EXECUTIVE_SUMMARY.md",
    "RELEASE_RECOMMENDATION.md",
)
OUTPUT_NAMES = WORKBOOK_NAMES + REPORT_NAMES
PRIVATE_REPOSITORY_DIRS = ("client_work", "private_projects", "evidence", "runtime")
BLUE = "1F4E78"
RED = "C00000"
WHITE = "FFFFFF"

Project = Mapping[str, str]
SheetDefinition = tuple[str, Sequence[str], Iterable[Sequence[object]]]


def _excel_text(value: object) -> object:
    if not isinstance(value, str):
        return value
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def _validate_output_location(output: Path) -> Path:
    resolved = output.expanduser().resolve()
    repository = Path(__file__).resolve().parents[1]
    try:
        relative = resolved.relative_to(repository)
    except ValueError:
        return resolved
    if not relative.parts or relative.parts[0] not in PRIVATE_REPOSITORY_DIRS:
        allowed = ", ".join(PRIVATE_REPOSITORY_DIRS)
        raise ValueError(
            f"Repository-local output must be under an ignored private directory: {allowed}"
        )
    return resolved


def _style_sheet(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, worksheet.max_column + 1):
        values = [
            str(worksheet.cell(row, column).value or "")
            for row in range(1, worksheet.max_row + 1)
        ]
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            60, max(12, max(map(len, values), default=0) + 2)
        )


def _write_workbook(path: Path, sheets: Sequence[SheetDefinition]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, headers, rows in sheets:
        worksheet = workbook.create_sheet(name)
        worksheet.append([_excel_text(value) for value in headers])
        for row in rows:
            worksheet.append([_excel_text(value) for value in row])
        _style_sheet(worksheet)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)


MASTER_HEADERS = (
    "Entry",
    "Book",
    "Page",
    "Instrument Number",
    "Recording Date",
    "Execution Date",
    "Grantor",
    "Grantee",
    "Document Type",
    "Legal Description",
    "Section",
    "Township",
    "Range",
    "Interest Conveyed",
    "Reservations",
    "Exceptions",
    "Assignments",
    "Releases",
    "Corrections",
    "Cross References",
    "Title Effect",
    "Verification Status",
    "Official Source",
    "Confidence",
    "Notes",
)


def _verification_rows(source_note: str) -> tuple[tuple[object, ...], ...]:
    return (
        (
            "NO INSTRUMENTS AVAILABLE TO THIS INVOCATION",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Section 32, Township 11 North, Range 25 West",
            "32",
            "11N",
            "25W",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "No title effect determined",
            "NOT FOUND",
            "https://okcountyrecords.com/search/beckham",
            "0%",
            f"{UNABLE_TO_VERIFY} {source_note}",
        ),
    )


def _runsheet_rows(source_note: str) -> tuple[tuple[object, ...], ...]:
    return (
        (
            "GAP-001",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "S32 T11N R25W",
            "NOT FOUND",
            "0%",
            f"{UNABLE_TO_VERIFY} {source_note}",
        ),
    )


RUNSHEET_HEADERS = (
    "Entry",
    "Instrument Number",
    "Book",
    "Page",
    "Recording Date",
    "Execution Date",
    "Grantor",
    "Grantee",
    "Legal Description",
    "Verification Status",
    "Confidence",
    "Examiner Notes",
)


def _source_rows(source_note: str) -> tuple[tuple[object, ...], ...]:
    return (
        (
            "SRC-001",
            "Official OKCountyRecords — Beckham County search",
            "Official recorded index and image portal",
            "https://okcountyrecords.com/search/beckham",
            "1",
            "Systematic search unavailable in this invocation",
            source_note,
        ),
        (
            "SRC-002",
            "Official OKCountyRecords API search documentation",
            "Official API specification",
            "https://okcountyrecords.com/api/v1/documentation/search",
            "1",
            "Documentation reviewed",
            "API search requires an authenticated credential.",
        ),
        (
            "SRC-003",
            "Recorded instrument images",
            "Original county document images",
            "",
            "2",
            "Not supplied",
            UNABLE_TO_VERIFY,
        ),
        (
            "SRC-004",
            "Existing project workbooks and runsheets",
            "Prior work product; not authoritative",
            "",
            "5",
            "Not supplied in the checked-out workspace",
            "Prior AI output was not accepted as evidence.",
        ),
    )


SOURCE_HEADERS = (
    "Source ID",
    "Source",
    "Authority",
    "URL or Path",
    "Authority Rank",
    "Access Result",
    "Notes",
)


def _error_rows(source_note: str) -> tuple[tuple[object, ...], ...]:
    return (
        (
            "ERR-001",
            "CRITICAL",
            "Evidence corpus absent",
            "No workbook, county index export, recorded image, PDF, or evidence manifest was supplied.",
            "Every requested instrument and title conclusion is blocked.",
            "Provide the complete read-only evidence corpus with SHA-256 manifest.",
            "OPEN",
        ),
        (
            "ERR-002",
            "CRITICAL",
            "Official systematic search unavailable",
            source_note,
            "Book/page, party, date, legal, and cross-reference searches cannot be completed.",
            "Configure authorized OKCountyRecords access and rerun every search.",
            "OPEN",
        ),
        (
            "ERR-003",
            "CRITICAL",
            "Spreadsheet QA input absent",
            "No candidate workbook or authoritative template is present in the checked-out workspace.",
            "Formulas, links, hidden structures, duplicates, calculations, and formatting cannot be audited.",
            "Supply each workbook and the authoritative template with hashes.",
            "OPEN",
        ),
    )


ERROR_HEADERS = (
    "Error ID",
    "Severity",
    "Discrepancy",
    "Evidence",
    "Impact",
    "Recommended Correction",
    "Disposition",
)


def _missing_rows() -> tuple[tuple[object, ...], ...]:
    items = (
        ("MD-001", "Every candidate workbook and runsheet", "Not present in workspace"),
        ("MD-002", "Complete Beckham County recorded-image corpus", "Not supplied"),
        ("MD-003", "Official county index export / search results", "Not supplied"),
        ("MD-004", "Authoritative workbook template", "Not supplied"),
        ("MD-005", "Every exhibit, schedule, correction, assignment, and release", "Not supplied"),
        ("MD-006", "Evidence manifest with file hashes", "Not supplied"),
    )
    return tuple(
        (
            item_id,
            document,
            reason,
            "Workspace inventory; official API access test",
            "0%",
            "NOT FOUND",
            UNABLE_TO_VERIFY,
        )
        for item_id, document, reason in items
    )


MISSING_HEADERS = (
    "Item ID",
    "Document or Evidence",
    "Likely Reason",
    "Search Attempts",
    "Confidence",
    "Verification Status",
    "Notes",
)


def _gap_rows() -> tuple[tuple[object, ...], ...]:
    return (
        (
            "TG-001",
            "NOT ESTABLISHED",
            "NOT ESTABLISHED",
            "Complete sovereign-to-present fee, mineral, royalty, leasehold, and burden chain",
            "Recorded copies and index continuation for every instrument affecting S32 T11N R25W",
            "CRITICAL",
            "NOT FOUND",
            UNABLE_TO_VERIFY,
        ),
    )


GAP_HEADERS = (
    "Gap ID",
    "Beginning Owner",
    "Ending Owner",
    "Missing Conveyance",
    "Evidence Required",
    "Severity",
    "Verification Status",
    "Notes",
)


def _chain_rows(client: str) -> tuple[tuple[object, ...], ...]:
    return (
        (
            "1",
            "NOT ESTABLISHED",
            f"{client} / affiliate interest not established",
            "No exact recorded instrument reviewed",
            "S32 T11N R25W",
            "Unknown",
            "Unknown",
            "Unknown",
            "Source title, assignments, reservations, schedules, releases, and corporate succession",
            "0%",
            "NOT FOUND",
            UNABLE_TO_VERIFY,
        ),
    )


CHAIN_HEADERS = (
    "Sequence",
    "From",
    "To",
    "Supporting Instrument",
    "Legal",
    "Interest",
    "WI",
    "NRI",
    "Outstanding Issues",
    "Confidence",
    "Verification Status",
    "Notes",
)


def _qa_rows() -> tuple[tuple[object, ...], ...]:
    blocked = (
        "Instrument verification",
        "Legal-description review",
        "Title-chain reconstruction",
        "Diversified ownership analysis",
        "Candidate workbook QA",
        "Template comparison",
    )
    package_checks = (
        "Deliverable file presence",
        "Workbook open validation",
        "Formula-injection defense",
        "Hidden row/column/sheet scan",
    )
    return tuple(
        (check, "BLOCKED", UNABLE_TO_VERIFY) for check in blocked
    ) + tuple((check, "PASS", "Generated package validated.") for check in package_checks)


QA_HEADERS = ("QA Check", "Result", "Finding")


def _build_workbooks(output: Path, project: Project) -> None:
    source_note = project["source_note"]
    master_sheets = (
        ("Master Verification", MASTER_HEADERS, _verification_rows(source_note)),
        ("Error Report", ERROR_HEADERS, _error_rows(source_note)),
        ("Missing Documents", MISSING_HEADERS, _missing_rows()),
        ("Title Gaps", GAP_HEADERS, _gap_rows()),
        ("Diversified Chain", CHAIN_HEADERS, _chain_rows(project["client"])),
        ("Source Citations", SOURCE_HEADERS, _source_rows(source_note)),
        ("QA Summary", QA_HEADERS, _qa_rows()),
    )
    _write_workbook(output / WORKBOOK_NAMES[0], master_sheets)
    individual_workbooks: tuple[tuple[str, SheetDefinition], ...] = (
        (
            WORKBOOK_NAMES[1],
            ("Corrected Run Sheet", RUNSHEET_HEADERS, _runsheet_rows(source_note)),
        ),
        (
            WORKBOOK_NAMES[2],
            ("Instrument Register", MASTER_HEADERS, _verification_rows(source_note)),
        ),
        (
            WORKBOOK_NAMES[3],
            ("Source Citation Index", SOURCE_HEADERS, _source_rows(source_note)),
        ),
        (WORKBOOK_NAMES[4], ("Error Log", ERROR_HEADERS, _error_rows(source_note))),
        (
            WORKBOOK_NAMES[5],
            ("Missing Document Log", MISSING_HEADERS, _missing_rows()),
        ),
        (
            WORKBOOK_NAMES[6],
            ("Title Gap Analysis", GAP_HEADERS, _gap_rows()),
        ),
        (
            WORKBOOK_NAMES[7],
            ("Diversified Chain", CHAIN_HEADERS, _chain_rows(project["client"])),
        ),
    )
    for filename, sheet in individual_workbooks:
        _write_workbook(output / filename, (sheet,))


def _write_reports(output: Path, project: Project, generated: str) -> None:
    common = (
        f"Subject: {project['legal']}, {project['county']}\n"
        f"Client: {project['client']}\n"
        f"Generated: {generated}\n"
    )
    (output / "QA_CERTIFICATION.md").write_text(
        f"""# QA Certification
{common}
Certification result: **NOT CERTIFIED — {RELEASE_STATUS}**

The package structure passed mechanical validation. The title examination did not pass
the evidence gate: zero source workbooks, county index exports, or recorded pages were
available to this invocation, and systematic official search access was unavailable.

{UNABLE_TO_VERIFY}

No title, acreage, decimal, WI, NRI, royalty, ORRI, burden, release, operator, or current
ownership conclusion is certified.
""",
        encoding="utf-8",
    )
    (output / "EXECUTIVE_SUMMARY.md").write_text(
        f"""# Executive Summary
{common}
## Outcome
**{RELEASE_STATUS}.**

No candidate workbook or official recorded-image corpus was present in the checked-out
workspace. The official API access test returned HTTP 401 because no credential was
configured. Selected public metadata pages are not a substitute for complete recorded
images and operative exhibits.

{UNABLE_TO_VERIFY}

No ownership, acreage, decimal, title effect, or Diversified chain was estimated. The
workbooks in this package document the evidence gap and required curative inputs.
""",
        encoding="utf-8",
    )
    (output / "RELEASE_RECOMMENDATION.md").write_text(
        f"""# Release Recommendation
{common}
## Recommendation
**DO NOT RELEASE AS A TITLE REPORT, RUNSHEET, OWNERSHIP REPORT, OR TITLE OPINION.**

Release remains blocked until all candidate workbooks and authoritative county records
are supplied, every instrument is checked against its complete recorded image, every
legal and cross-reference is resolved, spreadsheet QA passes, and a qualified Oklahoma
landman approves the exact deliverable hash.

{UNABLE_TO_VERIFY}
""",
        encoding="utf-8",
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _validate_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=False)
    formulas: list[str] = []
    hidden_structures: list[str] = []
    try:
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                hidden_structures.append(f"sheet:{sheet.title}")
            hidden_structures.extend(
                f"row:{sheet.title}!{index}"
                for index, dimension in sheet.row_dimensions.items()
                if dimension.hidden
            )
            hidden_structures.extend(
                f"column:{sheet.title}!{index}"
                for index, dimension in sheet.column_dimensions.items()
                if dimension.hidden
            )
            formulas.extend(
                f"{sheet.title}!{cell.coordinate}"
                for row in sheet.iter_rows()
                for cell in row
                if cell.data_type == "f"
            )
        if formulas or hidden_structures:
            raise RuntimeError(
                f"{path.name} failed QA: formulas={formulas}, hidden={hidden_structures}"
            )
        return {
            "sheets": workbook.sheetnames,
            "formula_cells": 0,
            "hidden_structures": 0,
        }
    finally:
        workbook.close()


def _validate(output: Path) -> dict[str, object]:
    missing = [name for name in OUTPUT_NAMES if not (output / name).is_file()]
    if missing:
        raise RuntimeError(f"Missing outputs: {', '.join(missing)}")

    return {
        name: _validate_workbook(output / name)
        for name in WORKBOOK_NAMES
    }


def build(output: Path, project: Project, generated: str) -> None:
    output = _validate_output_location(output)
    if output.exists():
        raise FileExistsError(
            f"Output path already exists; use a new versioned directory: {output}"
        )
    output.mkdir(parents=True)
    _build_workbooks(output, project)
    _write_reports(output, project, generated)
    validation = _validate(output)

    checksums = {name: _sha256(output / name) for name in OUTPUT_NAMES}
    (output / "SHA256SUMS.txt").write_text(
        "".join(f"{digest}  {name}\n" for name, digest in checksums.items()),
        encoding="utf-8",
    )
    manifest = {
        "project": project,
        "generated": generated,
        "release_status": RELEASE_STATUS,
        "official_verification": UNABLE_TO_VERIFY,
        "input_counts": {
            "candidate_workbooks": 0,
            "county_index_exports": 0,
            "recorded_pages": 0,
            "verified_instruments": 0,
        },
        "outputs": checksums,
        "workbook_validation": validation,
    }
    (output / "INVOCATION_MANIFEST.json").write_text(
        json.dumps(manifest, indent=2) + "\n",
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--client", required=True)
    parser.add_argument("--legal", required=True)
    parser.add_argument("--county", required=True)
    parser.add_argument(
        "--source-note",
        required=True,
        help="Traceable explanation of why authoritative search/evidence is unavailable.",
    )
    parser.add_argument(
        "--as-of",
        default=date.today().isoformat(),
        help="Report date (YYYY-MM-DD); defaults to today.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    generated = datetime.combine(
        date.fromisoformat(args.as_of), datetime.min.time(), tzinfo=timezone.utc
    ).isoformat()
    project = {
        "client": args.client,
        "legal": args.legal,
        "county": args.county,
        "source_note": args.source_note,
    }
    build(args.output, project, generated)
    print(f"Created blocked evidence package at {args.output}")


if __name__ == "__main__":
    main()
